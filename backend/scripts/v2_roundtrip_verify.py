#!/usr/bin/env python3
"""V2 工作簿往返（roundtrip）验证脚本（2026-08-18）。

覆盖各 tab 的 下载→修改→validate→apply→查库 闭环，重点验证只读哈希：
1. 哈希正确性：原样上传过 / 改只读列拒 / 改可编辑列过
2. 03_备件明细：改人工成本+原因+备注 → 落库（ManualCostOverride + line_note）
3. 05_实收回款：空表 CREATE → 落库（CollectionSnapshot）
4. 02_回款计划：空表 CREATE → 落库（CollectionMilestone）
5. 04_费用报销：新增行 → 拒（报销只能改已有行）
6. 06_领用返还：无实体ID行 → 拒

用法：DATABASE_URL=... python3 scripts/v2_roundtrip_verify.py
退出码：0=全过 1=有失败（会打印失败详情）
"""
import io
import sys
import uuid
from datetime import date

import openpyxl
from sqlalchemy import text

from app.db import SessionLocal
from app.services import maintenance_project_master_workbook as master

PASS = 0
FAIL = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  ✅ {name}")
    else:
        FAIL += 1
        print(f"  ❌ {name} {detail}")


def build(db, project_id: str, sheets=None) -> bytes:
    return master.build_project_master_v2(db, project_id=project_id)


def validate(db, project_id: str, content: bytes):
    return master.validate_project_master_v2(db, project_id=project_id, data=content)


def apply(db, project_id: str, content: bytes, operator: str = "roundtrip-verify"):
    plan = master.validate_project_master_v2(db, project_id=project_id, data=content)
    return master.apply_project_master_v2(db, plan, operated_by=operator,
                                          import_batch_id=f"rt-{uuid.uuid4().hex[:12]}")


def main() -> int:
    db = SessionLocal()
    try:
        # ---- 准备：找 03 有数据的项目 ----
        big_project = db.execute(text("""
            SELECT a.project_id FROM maintenance_source_order_assignment a
            JOIN maintenance_project p ON p.project_id = a.project_id
            WHERE a.is_active GROUP BY a.project_id
            ORDER BY count(*) DESC LIMIT 1
        """)).scalar()
        print(f"测试项目（备件最全）: {big_project}")

        # ================= 1. 哈希正确性（核心） =================
        print("\n== 1. 只读哈希正确性 ==")
        content = build(db, big_project)
        # 1a. 原样上传应通过
        plan = validate(db, big_project, content)
        check("原样上传 validate 通过", plan is not None)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["03_备件明细"]
        headers = [c.value for c in ws[1]]
        pn_col = headers.index("PN") + 1
        old_pn = ws.cell(2, pn_col).value
        ws.cell(2, pn_col).value = f"{old_pn}MODIFIED"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        try:
            validate(db, big_project, buf.read())
            check("改只读列(PN) 应拒绝", False, "竟然通过了！")
        except master.WorkbookError as exc:
            check("改只读列(PN) 被拒", exc.code == "readonly_cell_modified",
                  f"code={exc.code}")
        # 1c. 改可编辑列（人工未税单位成本）→ 应通过
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["03_备件明细"]
        manual_col = headers.index("人工未税单位成本") + 1
        reason_col = headers.index("人工成本原因") + 1
        note_col = headers.index("备注") + 1
        ws.cell(2, manual_col).value = 1234.56
        ws.cell(2, reason_col).value = "哈希测试补价"
        ws.cell(2, note_col).value = "哈希测试备注"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        plan = validate(db, big_project, buf.read())
        check("改可编辑列 validate 通过", plan is not None and len(plan.cost_refills) >= 1,
              f"refills={len(plan.cost_refills) if plan else 0}")
        buf.seek(0)
        editable_content = buf.read()

        # ================= 2. 03 备件明细：改成本+备注落库 =================
        print("\n== 2. 03_备件明细 roundtrip ==")
        # 用 Excel 第 2 行的实体ID（与 1c 修改的行对齐），而非 SQL 第一条
        wb_diag = openpyxl.load_workbook(io.BytesIO(content))
        ws_diag = wb_diag["03_备件明细"]
        h_diag = [c.value for c in ws_diag[1]]
        line_id = int(ws_diag.cell(2, h_diag.index("实体ID") + 1).value)
        before_note = db.execute(text(
            "SELECT line_note FROM f_maintenance_line WHERE id=:i"), {"i": line_id}).scalar()
        apply(db, big_project, editable_content)
        db.commit()
        row = db.execute(text(
            "SELECT unit_cost_ex_tax, reason FROM maintenance_manual_cost_override WHERE line_id=:i"),
            {"i": line_id}).one_or_none()
        check("成本覆盖落库", row is not None and float(row[0]) == 1234.56, f"row={row}")
        check("原因落库", row is not None and row[1] == "哈希测试补价", f"reason={row[1] if row else None}")
        note_after = db.execute(text(
            "SELECT line_note FROM f_maintenance_line WHERE id=:i"), {"i": line_id}).scalar()
        check("备注(line_note)落库", note_after == "哈希测试备注", f"note={note_after!r}")

        # ================= 3. 05_实收回款：空表 CREATE =================
        print("\n== 3. 05_实收回款 CREATE ==")
        content = build(db, big_project)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["05_实收回款"]
        h = {c.value: i for i, c in enumerate(ws[1])}
        contract_no = db.execute(text(
            "SELECT contract_no FROM maintenance_project_contract WHERE project_id=:p LIMIT 1"),
            {"p": big_project}).scalar()
        if contract_no:
            ws.cell(2, h["合同编号"] + 1).value = contract_no
            ws.cell(2, h["报告月份"] + 1).value = "2026-07"
            ws.cell(2, h["累计实收金额（含税）"] + 1).value = 50000.00
            ws.cell(2, h["回款凭证号"] + 1).value = "RT-2026-07"
            ws.cell(2, h["备注"] + 1).value = "roundtrip 测试"
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            try:
                apply(db, big_project, buf.read())
                db.commit()
                snap = db.execute(text(
                    "SELECT count(*) FROM maintenance_collection_snapshot "
                    "WHERE project_id=:p AND report_month='2026-07-01' AND status='confirmed'"),
                    {"p": big_project}).scalar()
                check("实收回款 CREATE 落库", snap == 1, f"count={snap}")
            except Exception as exc:
                db.rollback()
                check("实收回款 CREATE 落库", False, str(exc)[:200])
        else:
            print("  ⏭ 项目无合同，跳过 05（不影响核心验证）")

        # ================= 4. 02_回款计划 CREATE =================
        print("\n== 4. 02_回款计划 CREATE ==")
        content = build(db, big_project)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["02_回款计划"]
        h = {c.value: i for i, c in enumerate(ws[1])}
        if contract_no:
            ws.cell(2, h["操作"] + 1).value = "CREATE"
            ws.cell(2, h["合同编号"] + 1).value = contract_no
            ws.cell(2, h["期次"] + 1).value = 1
            ws.cell(2, h["计划回款日期"] + 1).value = "2026-09-30"
            ws.cell(2, h["计划回款金额（含税）"] + 1).value = 30000.00
            ws.cell(2, h["备注"] + 1).value = "roundtrip 计划测试"
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            try:
                apply(db, big_project, buf.read())
                db.commit()
                ms = db.execute(text(
                    "SELECT count(*) FROM maintenance_collection_milestone "
                    "WHERE project_id=:p AND is_active AND planned_date='2026-09-30'"),
                    {"p": big_project}).scalar()
                check("回款计划 CREATE 落库", ms == 1, f"count={ms}")
            except Exception as exc:
                db.rollback()
                check("回款计划 CREATE 落库", False, str(exc)[:200])
        else:
            print("  ⏭ 项目无合同，跳过 02")

        # ================= 5. 04_费用报销：新增行应拒 =================
        print("\n== 5. 04_费用报销：新增行守卫 ==")
        content = build(db, big_project)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["04_费用报销"]
        ws.cell(2, 1).value = "NEW-BXD-001"
        ws.cell(2, 3).value = "2026-08-01"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        try:
            validate(db, big_project, buf.read())
            check("报销新增行应拒", False, "竟然通过了")
        except master.WorkbookError as exc:
            check("报销新增行被拒", "not_recognized" in exc.code or "not_found" in exc.code,
                  f"code={exc.code}")

        # ================= 6. 06_领用返还：无实体ID应拒 =================
        print("\n== 6. 06_领用返还：无实体ID守卫 ==")
        content = build(db, big_project)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["06_领用返还"]
        ws.cell(2, 1).value = "SITE-001"
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        try:
            validate(db, big_project, buf.read())
            check("领用无实体ID应拒", False, "竟然通过了")
        except master.WorkbookError as exc:
            check("领用无实体ID被拒", "not_recognized" in exc.code, f"code={exc.code}")

        print(f"\n========== 结果：PASS={PASS} FAIL={FAIL} ==========")
        return 0 if FAIL == 0 else 1
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
