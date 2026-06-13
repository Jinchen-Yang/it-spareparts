"""agent 全部批次 合并+替代 总回执 xlsx —— 直接从库的 merge_log/substitute/audit 出（权威）。
  uv run python scripts/pn_eval_receipt.py [--out pn_eval_receipt_all.xlsx]
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import create_engine, text
from app.config import get_settings

ap = argparse.ArgumentParser()
ap.add_argument("--out", default="pn_eval_receipt_all.xlsx")
args = ap.parse_args()

eng = create_engine(get_settings().database_url)
wb = Workbook()
ws1 = wb.active; ws1.title = "合并"
ws1.append(["log_id", "源型号(墓碑)", "存活目标", "采购行", "销售行", "理由"])
ws2 = wb.create_sheet("替代")
ws2.append(["型号A", "型号B", "类型", "说明"])

with eng.connect() as conn:
    merges = conn.execute(text("""
        SELECT ml.id, sp.pn_std src, tp.pn_std tgt, ml.merge_reason rsn, ml.affected_records aff
        FROM product_merge_logs ml
        JOIN dim_part sp ON sp.id=ml.source_part_id
        JOIN dim_part tp ON tp.id=ml.target_part_id
        WHERE ml.created_by='agent:pn-match-eval' AND sp.status='merged'
          AND ml.id NOT IN (SELECT (before_json->>'merge_log_id')::int FROM sys_audit_log
                            WHERE action='unmerge' AND before_json ? 'merge_log_id')
        ORDER BY ml.id""")).fetchall()
    for m in merges:
        aff = m.aff or {}
        ws1.append([m.id, m.src, m.tgt,
                    len(aff.get("f_purchase_line_ids", [])),
                    len(aff.get("f_sales_line_ids", [])), m.rsn])
    subs = conn.execute(text("""
        SELECT pa.pn_std sa, pb.pn_std sb, s.substitute_type stype, s.note snote
        FROM sys_audit_log al
        JOIN part_substitute s ON s.id = al.entity_id
        JOIN dim_part pa ON pa.id=s.part_id_a
        JOIN dim_part pb ON pb.id=s.part_id_b
        WHERE al.entity_type='substitute' AND al.action='create'
          AND al.operated_by='agent:pn-match-eval'
        ORDER BY s.id""")).fetchall()
    for s in subs:
        ws2.append([s.sa, s.sb, s.stype, s.snote])

for ws, widths in ((ws1, [7, 34, 34, 7, 7, 56]), (ws2, [34, 36, 12, 56])):
    for c, w in enumerate(widths, 1):
        cell = ws.cell(1, c); cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.freeze_panes = "A2"
wb.save(args.out)
print(f"合并 {ws1.max_row - 1} 行, 替代 {ws2.max_row - 1} 行 → {args.out}")
