"""批量落地：读判定好的 xlsx，对 AI=same 按"安全配方"自动合并，其余给残差清单。

安全配方(自动合并)= AI判same + (①一方无采购 或 ②近期采/销价差≤15%) + 采购价无背离>50% + 无形态冲突词。
残差(不自动合,留联网/人工)= 价格背离>50% | 中等价差(需联网) | 文本形态冲突词命中。
形态冲突(平台/整机/扩展柜/license/清洗带…)主要由改进后的判定 prompt 在判定层拦掉，此处再做一层文本兜底。
方向=体量大者存活。合并走可回滚的 merge_parts。
  uv run python scripts/pn_eval_scale_apply.py plan --in pn_eval_b2.xlsx
  uv run python scripts/pn_eval_scale_apply.py go   --in pn_eval_b2.xlsx
"""
import argparse, statistics, time
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from app.config import get_settings
from app.db import SessionLocal
from app.services import merge, profit, inventory

PRICE_TOL, NEAR_DAYS, DIVERGE = 0.15, 180, 0.50
FORM_WORDS = ["平台", "准系统", "裸机", "刀框", "机框", "扩展柜", "控制柜",
              "license", "许可", "清洗", "主控板", "接口板"]


def med(xs): return statistics.median(xs) if xs else None
def drange(rows):
    ds = [d for _, d in rows if d]; return (min(ds), max(ds)) if ds else None
def near(a, b):
    if not a or not b: return None
    lo, hi = max(a[0], b[0]), min(a[1], b[1]); return 0 if lo <= hi else (lo - hi).days
def gap(pa, pb):
    if pa is None or pb is None: return None
    return abs(pa - pb) / max(min(pa, pb), 1e-9)


def info(conn, pn):
    row = conn.execute(text("SELECT id,status,pn_compact FROM dim_part WHERE pn_std=:p"), {"p": pn}).fetchone()
    if not row: return None
    pid, stt, compact = row
    pur = conn.execute(text("SELECT pl.unit_price,po.order_date FROM f_purchase_line pl "
        "JOIN f_purchase_order po ON pl.order_id=po.id WHERE pl.part_id=:i"), {"i": pid}).fetchall()
    sal = conn.execute(text("SELECT sl.unit_price,so.order_date FROM f_sales_line sl "
        "JOIN f_sales_order so ON sl.order_id=so.id WHERE sl.part_id=:i"), {"i": pid}).fetchall()
    cl = lambda rows: [(float(p), d) for p, d in rows if p is not None]
    return {"id": pid, "status": stt, "compact": compact, "pur": cl(pur), "sal": cl(sal)}


def form_conflict(a_text, b_text):
    a, b = (a_text or "").lower(), (b_text or "").lower()
    for w in FORM_WORDS:
        wl = w.lower()
        if (wl in a) != (wl in b):
            return w
    return None


def classify(A, B):
    ap, bp, as_, bs = len(A["pur"]), len(B["pur"]), len(A["sal"]), len(B["sal"])
    if min(ap, bp) == 0:
        return "merge①空壳"
    pm = gap(med([p for p,_ in A["pur"]]), med([p for p,_ in B["pur"]])); np_ = near(drange(A["pur"]), drange(B["pur"]))
    sm = gap(med([p for p,_ in A["sal"]]), med([p for p,_ in B["sal"]])) if as_ and bs else None
    ns = near(drange(A["sal"]), drange(B["sal"])) if as_ and bs else None
    okp = pm is not None and pm <= PRICE_TOL and np_ is not None and np_ <= NEAR_DAYS
    oks = sm is not None and sm <= PRICE_TOL and ns is not None and ns <= NEAR_DAYS
    if okp or oks:
        return "merge②价格佐证"
    if pm is not None and pm > DIVERGE:
        return "残差·背离>50%"
    return "残差·需联网"


def direction(A, B, apn, bpn):
    va, vb = len(A["pur"])+len(A["sal"]), len(B["pur"])+len(B["sal"])
    if va != vb: return (bpn, apn) if va > vb else (apn, bpn)
    if len(apn) != len(bpn): return (bpn, apn) if len(apn) > len(bpn) else (apn, bpn)
    return (bpn, apn)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["plan", "go"])
    ap.add_argument("--in", dest="inp", default="pn_eval_b2.xlsx")
    args = ap.parse_args()

    eng = create_engine(get_settings().database_url)
    wb = load_workbook(args.inp, data_only=True); ws = wb.active
    H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
    def g(r, n): return r[ix[n]]
    sames = [r for r in ws.iter_rows(min_row=2, values_only=True) if str(g(r,"AI判定")).strip()=="same"]

    plan, residual = [], []
    from collections import Counter
    cc = Counter()
    with eng.connect() as conn:
        for r in sames:
            apn, bpn = str(g(r,"源型号")), str(g(r,"候选型号"))
            fc = form_conflict(f"{apn} {g(r,'源描述')}", f"{bpn} {g(r,'候选描述')}")
            A, B = info(conn, apn), info(conn, bpn)
            if not A or not B or A["status"] != "active" or B["status"] != "active":
                cc["跳过·非active"] += 1; continue
            if fc:
                cc[f"残差·形态词({fc})"] += 1
                residual.append((g(r,"序号"), apn, bpn, f"形态词:{fc}")); continue
            if A["compact"] and A["compact"] == B["compact"]:
                cat = "merge⓪同PN写法差异"          # 归一化PN相同=同一商品,价差为波动,照合
            else:
                cat = classify(A, B)
            cc[cat] += 1
            if cat.startswith("merge"):
                src, tgt = direction(A, B, apn, bpn)
                plan.append({"#": g(r,"序号"), "src": src, "tgt": tgt, "cat": cat, "conf": g(r,"AI置信度")})
            else:
                residual.append((g(r,"序号"), apn, bpn, cat))

    print(f"AI=same 共 {len(sames)}；分类：")
    for k, v in cc.most_common():
        print(f"  {v:3}  {k}")
    print(f"\n→ 自动合并 {len(plan)}；残差(留联网/人工) {len(residual)}")
    if args.cmd == "plan":
        print("\n自动合并清单(墓碑→存活):")
        for p in plan: print(f"  #{p['#']:>3} [{p['cat']}] {p['src'][:30]:<31} → {p['tgt'][:30]}")
        print("\n残差清单:")
        for s, a, b, why in residual: print(f"  #{s:>3} {a[:28]:<29} ⇄ {b[:28]:<29} | {why}")
        return

    # go
    db = SessionLocal(); done, failed = [], []
    for p in plan:
        try:
            res = merge.merge_parts(db, p["src"], p["tgt"],
                f"AI同款(conf {p['conf']})+{p['cat']}+无形态冲突; pn_eval_scale #{p['#']}", "agent:pn-match-eval")
            done.append(res["merge_log_id"])
            print(f"  ✓ #{p['#']:>3} log#{res['merge_log_id']} {p['src'][:20]:<21}→ {p['tgt'][:20]}")
        except Exception as e:
            db.rollback(); failed.append((p["#"], f"{type(e).__name__}: {e}"))
            print(f"  ✗ #{p['#']} {type(e).__name__}: {e}")
    if done:
        print("重算利润+库存成本…"); t0 = time.time()
        profit.recompute(db); inventory.backfill_costs(db); db.commit()
        print(f" 重算 {time.time()-t0:.0f}s")
    db.close()
    print(f"\n成功 {len(done)} / 失败 {len(failed)}；log_ids={done}")
    if failed:
        for s, e in failed: print(f"  ✗ #{s}: {e}")
    print(f"\n残差 {len(residual)} 对待联网/人工：")
    for s, a, b, why in residual: print(f"  #{s:>3} {a[:28]:<29} ⇄ {b[:28]:<29} | {why}")


if __name__ == "__main__":
    main()
