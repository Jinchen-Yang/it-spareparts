"""规模化：采 needs_review 里"新的"候选对(排除批1已判的)，用改进后 prompt 判定，出 xlsx。
判定逻辑/prompt 复用 pn_match_eval(已含完整度规则)。判完由 pn_eval_scale_apply 按配方落地。
  uv run python scripts/pn_eval_scale.py sample --limit 300 --out pn_eval_b2.xlsx --done pn_eval.xlsx
"""
import argparse, sys, time
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent))
import pn_match_eval as ev
from openpyxl import load_workbook


def load_done_keys(paths):
    """paths 可逗号分隔多个 xlsx，全部已判对都排除。"""
    keys = set()
    for path in [p.strip() for p in str(paths).split(",") if p.strip()]:
        if not Path(path).exists():
            continue
        wb = load_workbook(path, data_only=True); ws = wb.active
        H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            a, b = r[ix["源型号"]], r[ix["候选型号"]]
            if a and b:
                keys.add(ev.pair_key(str(a), str(b)))
    return keys


def collect_fresh(db, done_keys, min_score, per_source, limit):
    from sqlalchemy import select
    from app.models.dimensions import DimPart
    from app.services import part_resolver
    sources = db.execute(select(DimPart).where(
        DimPart.status == "active", DimPart.needs_review.is_(True))).scalars().all()
    print(f"  needs_review 源 {len(sources)} 个；已判对 {len(done_keys)} 个排除", flush=True)
    buckets = {"high": [], "mid": [], "low": []}
    seen = set(); by_pn = {p.pn_std: p for p in sources}
    for i, src in enumerate(sources):
        if i % 200 == 0 and i:
            print(f"  …已扫 {i}", flush=True)
        res = part_resolver.resolve(db, src.pn_std, limit=per_source + 3)
        taken = 0
        for it in res["items"]:
            if it["pn_std"] == src.pn_std or it["score"] < min_score:
                continue
            key = ev.pair_key(src.pn_std, it["pn_std"])
            if key in seen or key in done_keys:
                continue
            seen.add(key)
            cand = by_pn.get(it["pn_std"]) or db.execute(
                select(DimPart).where(DimPart.pn_std == it["pn_std"])).scalar_one_or_none()
            if cand is None or cand.status != "active":
                continue
            buckets[ev.band(it["score"])].append({"a": src, "b": cand, "score": round(float(it["score"]), 3)})
            taken += 1
            if taken >= per_source:
                break
    out = []; per_band = max(1, limit // 3)
    for name in ("high", "mid", "low"):
        out.extend(buckets[name][:per_band])
    if len(out) < limit:
        rest = [p for name in buckets for p in buckets[name][per_band:]]
        out.extend(rest[:limit - len(out)])
    print(f"  采到新对 {len(out)}（high/mid/low="
          f"{len(buckets['high'])}/{len(buckets['mid'])}/{len(buckets['low'])}）", flush=True)
    return out[:limit]


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser("sample")
    s.add_argument("--out", default="pn_eval_b2.xlsx")
    s.add_argument("--done", default="pn_eval.xlsx")
    s.add_argument("--limit", type=int, default=300)
    s.add_argument("--per-source", type=int, default=2)
    s.add_argument("--min-score", type=float, default=0.45)
    args = ap.parse_args()

    from app.db import SessionLocal
    from app.config import get_settings
    db = SessionLocal()
    try:
        done = load_done_keys(args.done)
        print("· 采新候选对…", flush=True)
        pairs = collect_fresh(db, done, args.min_score, args.per_source, args.limit)
        for p in pairs:
            p["a"]._specs = ev._specs_text(db, p["a"].id)
            p["b"]._specs = ev._specs_text(db, p["b"].id)
        st = get_settings()
        if not st.llm_api_key:
            sys.exit("✗ 未配置 LLM_API_KEY")
        from openai import OpenAI
        client = OpenAI(api_key=st.llm_api_key, base_url=st.llm_base_url, timeout=st.llm_timeout_seconds)
        print(f"· 调用 {st.llm_model}（思考模式·含完整度规则）判 {len(pairs)} 对…", flush=True)
        results = []; t0 = time.time()
        for idx, p in enumerate(pairs, 1):
            d = ev._judge(client, st.llm_model, p["a"], p["b"], p["score"])
            tag = d.get("verdict") or d.get("_error") or "?"
            print(f"  [{idx}/{len(pairs)}] {p['a'].pn_std} ⇄ {p['b'].pn_std} → {tag}", flush=True)
            results.append((p, d))
        print(f"· 判定完成 {time.time()-t0:.0f}s", flush=True)
        ev._write_xlsx(Path(args.out), results)
        print(f"✓ 已写出 {args.out}（{len(results)} 行）")
    finally:
        db.close()


if __name__ == "__main__":
    main()
