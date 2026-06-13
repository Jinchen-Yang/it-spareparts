"""从库的合并日志/替代表生成 45 个 AI=same 的处置回执 xlsx。"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import create_engine, text
from app.config import get_settings

eng = create_engine(get_settings().database_url)
wb = load_workbook("pn_eval.xlsx", data_only=True); ws = wb.active
H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
def g(r, n): return r[ix[n]]
sames = [r for r in ws.iter_rows(min_row=2, values_only=True) if str(g(r,"AI判定")).strip()=="same"]

SUBST = {24, 25, 282, 249, 31}
KEEP_CONFIG = {18, 248}
KEEP_PLATFORM = {29, 221, 251}
FAILED = {19}

with eng.connect() as conn:
    logs = conn.execute(text("""
        SELECT ml.id AS log, sp.pn_std AS src, tp.pn_std AS tgt
        FROM product_merge_logs ml
        JOIN dim_part sp ON sp.id=ml.source_part_id
        JOIN dim_part tp ON tp.id=ml.target_part_id
        WHERE ml.created_by='agent:pn-match-eval' AND sp.status='merged'""")).fetchall()
    by_src = {r.src: (r.log, r.tgt) for r in logs}

out = Workbook(); o = out.active; o.title = "处置回执"
o.append(["序号", "源型号", "候选型号", "AI置信度", "处置", "详情/理由"])
for r in sames:
    s = g(r,"序号"); a, b = str(g(r,"源型号")), str(g(r,"候选型号")); conf = g(r,"AI置信度")
    if s in FAILED:
        o.append([s, a, b, conf, "未执行(并入同簇)", "华为S7706三重复:本对的S7706已并入#17目标,核心交换机留独立见#18"])
    elif s in SUBST:
        o.append([s, a, b, conf, "记替代·不合并", "同物多码;part_substitute互替;联网确认同一实物;价差大不并成本"])
    elif s in KEEP_PLATFORM:
        o.append([s, a, b, conf, "保持独立", "平台≠整机(用户裁定),两个SKU"])
    elif s in KEEP_CONFIG:
        o.append([s, a, b, conf, "保持独立·待复核", "同型号不同配置(价差>200%),倾向不同款"])
    elif a in by_src:
        log, tgt = by_src[a]; o.append([s, a, b, conf, "已合并", f"log#{log} → 存活『{tgt}』(可unmerge回滚)"])
    elif b in by_src:
        log, tgt = by_src[b]; o.append([s, a, b, conf, "已合并", f"log#{log} → 存活『{tgt}』(可unmerge回滚)"])
    else:
        o.append([s, a, b, conf, "?", "未分类"])

hdr = PatternFill("solid", fgColor="305496")
for c in range(1, 7):
    cell = o.cell(1, c); cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = hdr; cell.alignment = Alignment(horizontal="center")
for c, w in zip(range(1, 7), [5, 26, 26, 9, 16, 52]):
    o.column_dimensions[o.cell(1, c).column_letter].width = w
o.freeze_panes = "A2"
out.save("pn_eval_disposition.xlsx")
# 汇总
from collections import Counter
cc = Counter(o.cell(r, 5).value for r in range(2, o.max_row + 1))
print("处置汇总:", dict(cc))
print("written pn_eval_disposition.xlsx")
