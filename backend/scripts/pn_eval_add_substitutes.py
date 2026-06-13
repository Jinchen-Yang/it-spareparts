"""给 5 对同物多码建 part_substitute(互替)，证据入 note。用户裁定：不合并、记替代。"""
from openpyxl import load_workbook
from app.db import SessionLocal
from app.services import substitute

NOTES = {
 24:  "同一实物·华为OEM贴牌=东芝AL14SEB060N(600GB 10K SAS);双方各有交易,价差50%,按设计§9记替代不合并",
 25:  "同一实物·华为02312RBV=东芝AL15SEB120N(1.2TB 10K SAS);亚马逊/cmicomputer确认OEM贴牌,价差22%,记替代",
 282: "同一实物·华为02311JDA=希捷ST2000NX0433(2TB 7.2K SAS);联网确认,价差44%,记替代",
 249: "同一实物·澜起M88JTMC5220R=Jintide C5220R 24核(重封装Intel Xeon);澜起官网确认,记替代",
 31:  "同型号·希捷ST1200MM0009浪潮存储专用变体;价差167%/500%疑渠道/锁定差异,不合并,记替代",
}
wb = load_workbook("pn_eval.xlsx", data_only=True); ws = wb.active
H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
def g(r, n): return r[ix[n]]
rows = {g(r,"序号"): r for r in ws.iter_rows(min_row=2, values_only=True) if g(r,"序号") in NOTES}

db = SessionLocal()
for serial, note in NOTES.items():
    r = rows.get(serial)
    if not r:
        print(f"  #{serial} 未找到"); continue
    a, b = str(g(r,"源型号")), str(g(r,"候选型号"))
    try:
        res = substitute.add_substitute(db, a, b, note, "agent:pn-match-eval",
                                        direction="both", substitute_type="same_spec")
        print(f"  ✓ #{serial} {'已建' if res['created'] else '已存在'}: {a}  ⇄  {b}")
    except Exception as e:
        print(f"  ✗ #{serial} {type(e).__name__}: {e}")
db.close()
