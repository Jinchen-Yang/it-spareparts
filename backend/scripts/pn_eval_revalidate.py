"""用改后的 SYSTEM_PROMPT 重判：已知错合应翻转，控制组(写法差异/OEM)应仍 same。"""
import sys
from pathlib import Path
from types import SimpleNamespace
from openpyxl import load_workbook
sys.path.insert(0, str(Path("scripts").resolve()))
import pn_match_eval as ev
from app.config import get_settings
from openai import OpenAI

s = get_settings()
client = OpenAI(api_key=s.llm_api_key, base_url=s.llm_base_url, timeout=s.llm_timeout_seconds)
model = s.llm_model

ERR = [29, 221, 251, 18, 248]          # 已知错合(平台/整机·配置)，应翻转
CTRL = [17, 88, 51, 119, 40, 25]       # 控制：写法差异/同码/OEM，应仍 same
want = {**{x: "错合·应翻转" for x in ERR}, **{x: "控制·应 same" for x in CTRL}}

wb = load_workbook("pn_eval.xlsx", data_only=True); ws = wb.active
H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
def g(r, n): return r[ix[n]]
rows = {g(r,"序号"): r for r in ws.iter_rows(min_row=2, values_only=True) if g(r,"序号") in want}

def mk(r, which):
    p = "源" if which == "a" else "候选"
    return SimpleNamespace(pn_std=str(g(r, p+"型号")), description=g(r, p+"描述"),
        brand=g(r, p+"品牌"), category_major=None, category_minor=None,
        _specs=g(r, p+"规格") or "（无）")

print(f"{'序号':>4} | {'期望':<8} | 旧→新      | 理由")
flip, ctrl = 0, 0
for serial in ERR + CTRL:
    r = rows.get(serial)
    if not r:
        print(f"#{serial} 缺"); continue
    a, b = mk(r, "a"), mk(r, "b")
    d = ev._judge(client, model, a, b, float(g(r,"初筛相似度")))
    v = d.get("verdict") or d.get("_error", "?")
    if serial in ERR and v in ("different", "uncertain", "not_a_single_part"): flip += 1
    if serial in CTRL and v == "same": ctrl += 1
    why = (d.get("conflict") or d.get("reason") or "")[:58]
    print(f"#{serial:>3} | {want[serial]:<8} | same→{v:<14} | {why}")
print(f"\n错合翻转 {flip}/{len(ERR)}   控制保持 same {ctrl}/{len(CTRL)}")
print("→ prompt 修正" + ("有效 ✓" if flip >= len(ERR)-1 and ctrl >= len(CTRL)-1 else "需再调 ✗"))
