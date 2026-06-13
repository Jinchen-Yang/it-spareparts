"""型号「是否同款」判定——AI 准确率测试脚本（治理 Agent 落地前的地基验证）。

为什么有这个脚本
----------------
PN 标准化/合并 Agent（见 docs/PN标准化Agent设计方案.md、PR #7）能不能自动化，
整个压在一个假设上：**AI 到底能不能把"两个长得像的型号是不是同一个实物"判准。**
判得准 → 才敢让它自动合并/标准化；判不准 → 架构再漂亮也是空中楼阁。
所以在写整套 Agent 之前，先用几百对真实样本量一个准确率出来——尤其是
"错合率"（AI 说是同款、其实不是的比例），这个数直接决定自动化能放权到哪一步。

它做什么
--------
1) sample 模式：从真库里用现成的 part_resolver 捞出"疑似同款"的候选对，
   逐对喂给 DeepSeek v4（思考模式）判 same/different/uncertain/not_a_single_part，
   连同 AI 理由写进一个 Excel，留两列空白给人工核对。
2) score 模式：读人工填好的 Excel，算出准确率/错合率/召回率等指标。
3) selftest：纯函数自测（不连库、不花 token），验证脚本本身没写坏。

怎么跑（在 backend/ 目录下）
---------------------------
  uv run python scripts/pn_match_eval.py selftest                      # 可选，不花钱
  uv run python scripts/pn_match_eval.py sample --dry-run --limit 300  # 只出样本，不调 AI
  uv run python scripts/pn_match_eval.py sample --limit 300 --out pn_eval.xlsx   # 真调 DS
  # —— 人工在 xlsx「人工判定」列填 same/different/uncertain/not_a_single_part ——
  uv run python scripts/pn_match_eval.py score --in pn_eval.xlsx

前置：本地 .env 的 DATABASE_URL 指向真库；LLM_API_KEY / LLM_BASE_URL / LLM_MODEL
配好 DS v4。思考模式由本脚本强制开启，不依赖 .env 的 LLM_EXTRA_BODY。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path

# 让 `python scripts/xxx.py` 与 `uv run` 都能 import app.*
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

# ============================================================
# 判定 prompt —— 本脚本的核心；要调就改这里
# ============================================================

SYSTEM_PROMPT = """\
你是 IT 服务器备件领域的资深主数据专家。下面给你两条来自 ERP 的型号记录（A 和 B），
它们由不同员工手工录入，写法可能很乱（大小写、空格、连字符、中英混写、品牌前后缀、
描述繁简不一）。你的唯一任务：判断 A 和 B 是不是**同一个可下单的实物商品**——
即同一厂商、同一规格、可互相替代发货，从而可以在系统里合并成一条。

【判定铁律（按优先级）】
1) 纯写法差异不影响判定：连字符/空格/大小写不同、品牌中英写法不同（超微=Supermicro、
   希捷=Seagate）、描述详略不同——这些都算同款。
2) 任一关键规格不同 = 不同款（different），哪怕型号字符串极像：
   - 容量（480G/960G/1.92T）、显存档（V100 16G vs 32G、A100 40G vs 80G）
   - 接口/协议（SATA/SAS/NVMe；PCIe vs SXM/NVLink）
   - 转速（7.2K/10K/15K）、尺寸封装（2.5"/3.5"、HHHL/FHFL）
   - 代际/类型（DDR4 vs DDR5、Gen3 vs Gen4）、电压功率
3) 厂商料号 vs 经销商/整机厂 SKU，若指向同一实物 = 同款（这是别名关系，不是不同商品）。
4) 下列情况绝不可判同款：
   - 一方标"兼容/compatible/OEM/拆机/翻新/二手"，另一方是原厂全新 → different
   - "型号"是纯中文、没有厂商字母数字料号，或含"一批/一套/若干/配件/组件"等字样 →
     这通常不是单个商品而是一批货或一个分类，判 not_a_single_part
   - 型号尾缀不同（-00005 / 固件码 / 地区码）且无法确认可互换 → uncertain
5) 信息不足以下定论时，判 uncertain，不要猜。

【成本意识——最重要】
把两个**不同**的东西错合成一条，会静默污染成本、库存与报价，后果严重且难发现；
而漏合（本是同款却没判出）只是少省点事，代价小得多。
所以**拿不准时一律 uncertain 或 different，绝不轻易 same。**

【输出格式】可以先思考，最后只输出一个 JSON 对象（除 JSON 外不要多余文字）：
{
  "verdict": "same | different | uncertain | not_a_single_part",
  "confidence": 0.0,
  "canonical_pn": "若 same 给出最规范的厂商料号写法，否则 null",
  "key_attrs": {"brand": "", "capacity": "", "interface": "", "form_factor": ""},
  "conflict": "若 different 写明哪个规格冲突，否则 null",
  "reason": "一句中文，点到具体字段说明为什么"
}"""

USER_TEMPLATE = """\
A 记录
- 型号PN: {a_pn}
- 描述: {a_desc}
- 品牌: {a_brand}
- 品类: {a_cat}
- 已抽取规格: {a_specs}

B 记录
- 型号PN: {b_pn}
- 描述: {b_desc}
- 品牌: {b_brand}
- 品类: {b_cat}
- 已抽取规格: {b_specs}

（系统初筛文本相似度: {score}）
请判断 A 与 B 是否同款，按要求输出 JSON。"""

VERDICTS = ("same", "different", "uncertain", "not_a_single_part")

# 输出表列（顺序即列序）；带★的两列留给人工
COLUMNS = [
    "序号", "源型号", "源描述", "源品牌", "源规格",
    "候选型号", "候选描述", "候选品牌", "候选规格", "初筛相似度",
    "AI判定", "AI置信度", "AI规范PN", "AI冲突点", "AI理由", "AI思考(节选)",
    "★人工判定", "★AI对不对", "备注",
]


# ============================================================
# 纯函数（selftest 覆盖这些，无需库/AI）
# ============================================================

def extract_json(text: str | None) -> dict | None:
    """从模型输出里抠出 JSON（容忍 ```json 代码块或前后多余文字）。"""
    if not text:
        return None
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.S)
    if m:
        try:
            return json.loads(m.group(1))
        except json.JSONDecodeError:
            pass
    i, j = text.find("{"), text.rfind("}")
    if i != -1 and j > i:
        try:
            return json.loads(text[i:j + 1])
        except json.JSONDecodeError:
            return None
    return None


def pair_key(a: str, b: str) -> tuple[str, str]:
    """无向对去重键（A↔B 与 B↔A 视为同一对）。"""
    return (a, b) if a <= b else (b, a)


def band(score: float) -> str:
    """按初筛相似度分档，采样时各档均衡，避免只测到容易的对。"""
    if score >= 0.85:
        return "high"
    if score >= 0.65:
        return "mid"
    return "low"


def compute_metrics(rows: list[dict]) -> dict:
    """rows: [{"ai": verdict, "human": verdict}]，只统计 human 已填的行。"""
    labeled = [r for r in rows if r.get("human") in VERDICTS and r.get("ai") in VERDICTS]
    n = len(labeled)
    out: dict = {"已标注对数": n}
    if n == 0:
        return out
    agree = sum(1 for r in labeled if r["ai"] == r["human"])
    ai_same = [r for r in labeled if r["ai"] == "same"]
    human_same = [r for r in labeled if r["human"] == "same"]
    out["总体一致率"] = round(agree / n, 4)
    # 最关键：AI 说 same 里，其实不是 same 的比例 = 错合风险
    if ai_same:
        wrong = sum(1 for r in ai_same if r["human"] != "same")
        out["★错合率(AI判same却不是)"] = round(wrong / len(ai_same), 4)
        out["same精确率"] = round(1 - wrong / len(ai_same), 4)
    out["AI判same对数"] = len(ai_same)
    if human_same:
        caught = sum(1 for r in human_same if r["ai"] == "same")
        out["same召回率"] = round(caught / len(human_same), 4)
        out["真同款被AI判different数"] = sum(1 for r in human_same if r["ai"] == "different")
    out["真同款对数"] = len(human_same)
    out["AI判uncertain占比"] = round(
        sum(1 for r in labeled if r["ai"] == "uncertain") / n, 4)
    # 混淆矩阵
    cm: dict[str, dict[str, int]] = {a: {b: 0 for b in VERDICTS} for a in VERDICTS}
    for r in labeled:
        cm[r["human"]][r["ai"]] += 1
    out["混淆矩阵(行=人工,列=AI)"] = cm
    return out


# ============================================================
# 库 + AI（仅 sample/score 用，selftest 不触发）
# ============================================================

def _specs_text(db, part_id: int) -> str:
    from app.models.master_data import ProductSpec
    from sqlalchemy import select
    rows = db.execute(select(ProductSpec).where(ProductSpec.part_id == part_id)).scalars().all()
    if not rows:
        return "（无）"
    return "; ".join(
        f"{s.spec_key}={s.spec_value}{('(' + s.spec_unit + ')') if s.spec_unit else ''}"
        for s in rows)


def _collect_pairs(db, source_kind: str, min_score: float, per_source: int,
                   limit: int) -> list[dict]:
    """用 resolver 捞候选对，按相似度分档均衡采样到 limit 对。"""
    from sqlalchemy import select
    from app.models.dimensions import DimPart
    from app.services import part_resolver

    q = select(DimPart).where(DimPart.status == "active")
    if source_kind == "needs_review":
        q = q.where(DimPart.needs_review.is_(True))
    sources = db.execute(q).scalars().all()
    print(f"  源型号候选 {len(sources)} 个（source-kind={source_kind}）", flush=True)

    buckets: dict[str, list[dict]] = {"high": [], "mid": [], "low": []}
    seen: set[tuple[str, str]] = set()
    by_pn = {p.pn_std: p for p in sources}
    for i, src in enumerate(sources):
        if i % 200 == 0 and i:
            print(f"  …已扫 {i} 个源", flush=True)
        res = part_resolver.resolve(db, src.pn_std, limit=per_source + 3)
        taken = 0
        for it in res["items"]:
            if it["pn_std"] == src.pn_std or it["score"] < min_score:
                continue
            key = pair_key(src.pn_std, it["pn_std"])
            if key in seen:
                continue
            seen.add(key)
            cand = by_pn.get(it["pn_std"]) or db.execute(
                select(DimPart).where(DimPart.pn_std == it["pn_std"])).scalar_one_or_none()
            if cand is None:
                continue
            buckets[band(it["score"])].append({
                "a": src, "b": cand, "score": round(float(it["score"]), 3),
            })
            taken += 1
            if taken >= per_source:
                break

    # 各档均衡取样
    out: list[dict] = []
    per_band = max(1, limit // 3)
    for name in ("high", "mid", "low"):
        out.extend(buckets[name][:per_band])
    # 不足则用剩余补满
    if len(out) < limit:
        rest = [p for name in buckets for p in buckets[name][per_band:]]
        out.extend(rest[:limit - len(out)])
    print(f"  采到候选对 {len(out)} 对（high/mid/low="
          f"{len(buckets['high'])}/{len(buckets['mid'])}/{len(buckets['low'])}）", flush=True)
    return out[:limit]


def _judge(client, model: str, a, b, score: float) -> dict:
    """调 DS v4（思考模式）判一对。返回解析后的 dict（含 _raw/_reasoning/_error）。"""
    user = USER_TEMPLATE.format(
        a_pn=a.pn_std, a_desc=a.description or "（无）", a_brand=a.brand or "（无）",
        a_cat=" / ".join(x for x in (a.category_major, a.category_minor) if x) or "（无）",
        a_specs=a._specs, b_pn=b.pn_std, b_desc=b.description or "（无）",
        b_brand=b.brand or "（无）",
        b_cat=" / ".join(x for x in (b.category_major, b.category_minor) if x) or "（无）",
        b_specs=b._specs, score=score)
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "system", "content": SYSTEM_PROMPT},
                      {"role": "user", "content": user}],
            extra_body={"thinking": {"type": "enabled"}},   # 思考模式开
            temperature=0,
        )
        msg = resp.choices[0].message
        data = extract_json(msg.content) or {}
        data["_raw"] = msg.content
        data["_reasoning"] = getattr(msg, "reasoning_content", None)
        if not data.get("verdict"):
            data["_error"] = "未解析出 verdict"
        return data
    except Exception as e:  # noqa: BLE001 —— 单对失败不打断整批
        return {"verdict": "", "_error": f"{type(e).__name__}: {e}"}


def cmd_sample(args) -> None:
    from app.db import SessionLocal
    from app.config import get_settings

    out_path = Path(args.out)
    db = SessionLocal()
    try:
        print("· 采样候选对…", flush=True)
        pairs = _collect_pairs(db, args.source_kind, args.min_score,
                               args.per_source, args.limit)
        for p in pairs:
            p["a"]._specs = _specs_text(db, p["a"].id)
            p["b"]._specs = _specs_text(db, p["b"].id)

        client = model = None
        if not args.dry_run:
            s = get_settings()
            if not s.llm_api_key:
                sys.exit("✗ 未配置 LLM_API_KEY，无法调 AI；或加 --dry-run 只出样本表。")
            from openai import OpenAI
            client = OpenAI(api_key=s.llm_api_key, base_url=s.llm_base_url,
                            timeout=s.llm_timeout_seconds)
            model = s.llm_model
            print(f"· 调用 {model}（思考模式）判 {len(pairs)} 对…", flush=True)

        results = []
        t0 = time.time()
        for idx, p in enumerate(pairs, 1):
            a, b, score = p["a"], p["b"], p["score"]
            if args.dry_run:
                d = {}
            else:
                d = _judge(client, model, a, b, score)
                tag = d.get("verdict") or d.get("_error") or "?"
                print(f"  [{idx}/{len(pairs)}] {a.pn_std} ⇄ {b.pn_std} → {tag}", flush=True)
            results.append((p, d))
        if not args.dry_run:
            print(f"· 判定完成，耗时 {time.time() - t0:.0f}s", flush=True)

        _write_xlsx(out_path, results)
        print(f"\n✓ 已写出 {out_path}（{len(results)} 行）")
        print("  下一步：在「★人工判定」列填 same/different/uncertain/not_a_single_part，")
        print("  重点抽查 AI 判 same 的那些对不对；填完跑：")
        print(f"    uv run python scripts/pn_match_eval.py score --in {out_path}")
    finally:
        db.close()


def _write_xlsx(path: Path, results: list[tuple]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "型号同款判定"
    ws.append(COLUMNS)
    hdr_fill = PatternFill("solid", fgColor="305496")
    human_fill = PatternFill("solid", fgColor="FFF2CC")
    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF" if not name.startswith("★") else "9C5700")
        cell.fill = hdr_fill if not name.startswith("★") else human_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (p, d) in enumerate(results, 1):
        a, b = p["a"], p["b"]
        reasoning = (d.get("_reasoning") or "")[:400]
        ka = d.get("key_attrs")
        ws.append([
            i, a.pn_std, a.description, a.brand, a._specs,
            b.pn_std, b.description, b.brand, b._specs, p["score"],
            d.get("verdict") or d.get("_error", ""), d.get("confidence"),
            d.get("canonical_pn"), d.get("conflict"),
            d.get("reason") or (json.dumps(ka, ensure_ascii=False) if ka else ""),
            reasoning, "", "", d.get("_error", ""),
        ])
    # 人工判定列下拉
    dv = DataValidation(type="list", formula1='"same,different,uncertain,not_a_single_part"',
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"Q2:Q{len(results) + 1}")   # 第17列=Q「★人工判定」
    widths = [5, 22, 30, 12, 26, 22, 30, 12, 26, 8, 12, 8, 20, 18, 34, 30, 12, 9, 16]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def cmd_score(args) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(args.in_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    ai_i, human_i = header.index("AI判定"), header.index("★人工判定")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        ai = (str(r[ai_i]).strip() if r[ai_i] else "")
        human = (str(r[human_i]).strip() if r[human_i] else "")
        rows.append({"ai": ai, "human": human})
    metrics = compute_metrics(rows)
    print("\n==== 准确率报告 ====")
    print(json.dumps(metrics, ensure_ascii=False, indent=2))
    if metrics.get("已标注对数", 0) == 0:
        print("\n⚠ 还没有人工标注（「★人工判定」列为空），先填了再跑 score。")
    else:
        print("\n看点：★错合率 = AI 说同款但其实不是的比例，这是能不能让 AI 自动合并的命门。")


def cmd_selftest(_args) -> None:
    ok = True

    def check(name, cond):
        nonlocal ok
        print(f"  {'✓' if cond else '✗'} {name}")
        ok = ok and cond

    check("extract_json 纯JSON", extract_json('{"verdict":"same"}') == {"verdict": "same"})
    check("extract_json 代码块", extract_json('```json\n{"verdict":"different"}\n```')
          == {"verdict": "different"})
    check("extract_json 带前后文",
          extract_json('思考...\n最终：{"verdict":"uncertain","confidence":0.4} 完')
          == {"verdict": "uncertain", "confidence": 0.4})
    check("extract_json 坏输入", extract_json("没有json") is None)
    check("pair_key 无向", pair_key("B", "A") == pair_key("A", "B") == ("A", "B"))
    check("band 分档", (band(0.9), band(0.7), band(0.5)) == ("high", "mid", "low"))
    m = compute_metrics([
        {"ai": "same", "human": "same"},
        {"ai": "same", "human": "different"},   # 错合
        {"ai": "different", "human": "different"},
        {"ai": "uncertain", "human": "same"},   # 漏判→uncertain
    ])
    check("错合率=0.5", m["★错合率(AI判same却不是)"] == 0.5)
    check("same精确率=0.5", m["same精确率"] == 0.5)
    check("same召回率=0.5", m["same召回率"] == 0.5)
    check("总体一致率=0.5", m["总体一致率"] == 0.5)
    print("自测通过 ✓" if ok else "自测有失败 ✗")
    sys.exit(0 if ok else 1)


def main() -> None:
    ap = argparse.ArgumentParser(description="型号同款判定 AI 准确率测试")
    sub = ap.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("sample", help="采样候选对并调 AI 判定，输出待人工核对的 Excel")
    s.add_argument("--out", default="pn_eval.xlsx")
    s.add_argument("--limit", type=int, default=300, help="总候选对上限")
    s.add_argument("--per-source", type=int, default=2, help="每个源型号最多取几个候选")
    s.add_argument("--min-score", type=float, default=0.45, help="初筛相似度下限")
    s.add_argument("--source-kind", choices=["needs_review", "all"], default="needs_review")
    s.add_argument("--dry-run", action="store_true", help="只出样本表，不调 AI（不花 token）")
    s.set_defaults(func=cmd_sample)

    sc = sub.add_parser("score", help="读人工标注好的 Excel，算准确率指标")
    sc.add_argument("--in", dest="in_path", default="pn_eval.xlsx")
    sc.set_defaults(func=cmd_score)

    st = sub.add_parser("selftest", help="纯函数自测（不连库、不花 token）")
    st.set_defaults(func=cmd_selftest)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
