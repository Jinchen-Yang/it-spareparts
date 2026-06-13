"""PN 标准化发现层 —— 设计§5 standardize 任务的 L0(advisory)·不联网 MVP。

孤儿待审型号(无合并候选) → AI 凭字段 + 自身知识(不联网)提"厂商规范 PN"建议
→ 人工在 Excel 审 → apply 走已建的 part_rename.rename_part 落库。

  uv run python scripts/pn_standardize.py selftest                              # 纯函数自测,不花钱
  uv run python scripts/pn_standardize.py suggest --limit 50 --out pn_std.xlsx  # 出建议(调 AI)
  # 人工在「★采纳」填 yes、「★最终PN」可改写,然后:
  uv run python scripts/pn_standardize.py apply --in pn_std.xlsx                # 采纳的走 rename
"""
from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))           # 复用 pn_match_eval
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))       # import app.*
import pn_match_eval as ev   # extract_json / _specs_text

# ============================================================
# 标准化建议 prompt —— 核心；要调改这里
# ============================================================
SYSTEM_PROMPT = """\
你是 IT 服务器备件领域的资深主数据专家。下面给你一条**录入很乱**的型号记录（中英混写、
夹中文描述、空格大小写不一、可能缺品牌）。只凭这些字段 + 你**自身知识**（不联网），
给出它最可能的**厂商规范 PN 写法**和品牌，用于把它"正名"成标准型号。

【判定规则】
1) 能认出是某厂商的标准型号 → canonical_pn 给厂商官方料号/型号的规范写法
   （去掉多余中文描述、整理空格大小写连字符；如"HP惠普 DL380 G9 整机"→"HP ProLiant DL380 Gen9"，
   "ST1200MM0009 浪潮存储专用"→"ST1200MM0009"），brand 给规范品牌名。
2) 只是写法脏（型号本身对，混了中文/空格）→ 给清洗后的规范写法即可。
3) **绝不编造、绝不替换料号**：canonical_pn 只能是输入里**已出现料号**的规范写法、或其厂商官方
   等价全称；**严禁输出输入中没有的另一个料号**（哪怕你觉得真品是别的）。想不到合规写法、或会引入
   新料号 → canonical_pn=null、uncertain=true。（如"511-1231 PN重复"——'PN重复'是录入批注不是料号，
   规范名就是 511-1231，绝不能改成别的号。）
4) 纯中文无厂商字母数字料号，或含"一批/一套/若干/配件/组件/通用项"等 → 不是单个可下单商品，
   canonical_pn=null、not_a_part=true。
5) 若记录里有多个料号（厂商料号 + 原厂型号），canonical_pn 取**最权威的原厂型号**写法。

【输出格式】可先思考，最后只输出一个 JSON（除 JSON 外不要多余文字）：
{
  "canonical_pn": "规范PN写法，或 null",
  "brand": "规范品牌名，或 null",
  "confidence": 0.0,
  "uncertain": false,
  "not_a_part": false,
  "reason": "一句中文，说明依据/为什么拿不准"
}"""

USER_TEMPLATE = """\
待标准化记录：
- 当前 pn_std: {pn}
- 描述: {desc}
- 品牌: {brand}
- 品类: {cat}
- 已抽取规格: {specs}

请给出它的厂商规范 PN 写法与品牌，按要求输出 JSON。"""

COLUMNS = [
    "part_id", "当前pn_std", "描述", "品牌", "品类", "规格",
    "AI建议规范PN", "AI品牌", "AI置信度", "AI不确定", "AI非单品", "AI理由",
    "★采纳(yes)", "★最终PN(空=用建议)", "备注",
]


# ============================================================
# 纯函数（selftest 覆盖）
# ============================================================
def adopt_target(adopt, final_pn, ai_pn) -> str | None:
    """据人工列决定改名目标：采纳=yes 时取「★最终PN」(空则用 AI 建议)，否则 None(不改)。"""
    if str(adopt or "").strip().lower() not in ("yes", "y", "采纳", "是", "true", "1"):
        return None
    target = (str(final_pn).strip() if final_pn else "") or (str(ai_pn).strip() if ai_pn else "")
    return target or None


# ============================================================
# AI（suggest 用）
# ============================================================
def _suggest(client, model, part) -> dict:
    user = USER_TEMPLATE.format(
        pn=part.pn_std, desc=part.description or "（无）", brand=part.brand or "（无）",
        cat=" / ".join(x for x in (part.category_major, part.category_minor) if x) or "（无）",
        specs=part._specs)
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "system", "content": SYSTEM_PROMPT},
                      {"role": "user", "content": user}],
            extra_body={"thinking": {"type": "enabled"}}, temperature=0)
        msg = resp.choices[0].message
        data = ev.extract_json(msg.content) or {}
        data["_reasoning"] = getattr(msg, "reasoning_content", None)
        if "canonical_pn" not in data and not data.get("not_a_part"):
            data["_error"] = "未解析出 canonical_pn"
        return data
    except Exception as e:  # noqa: BLE001
        return {"_error": f"{type(e).__name__}: {e}"}


def cmd_suggest(args) -> None:
    from sqlalchemy import select

    from app.config import get_settings
    from app.db import SessionLocal
    from app.models.dimensions import DimPart
    from app.services import match_candidates

    db = SessionLocal()
    try:
        ids = match_candidates.parts_without_candidates(db)
        ids = ids[args.offset: args.offset + args.limit]
        parts = db.execute(select(DimPart).where(DimPart.id.in_(ids))).scalars().all()
        print(f"· 孤儿待标准化型号 {len(parts)} 个（offset={args.offset} limit={args.limit}）", flush=True)
        for p in parts:
            p._specs = ev._specs_text(db, p.id)
        s = get_settings()
        if not s.llm_api_key:
            sys.exit("✗ 未配置 LLM_API_KEY")
        from openai import OpenAI
        client = OpenAI(api_key=s.llm_api_key, base_url=s.llm_base_url, timeout=s.llm_timeout_seconds)
        print(f"· 调 {s.llm_model}（思考模式·不联网）提规范名 {len(parts)} 个…", flush=True)
        results, t0 = [], time.time()
        for i, p in enumerate(parts, 1):
            d = _suggest(client, s.llm_model, p)
            tag = d.get("canonical_pn") or ("非单品" if d.get("not_a_part") else
                                            "不确定" if d.get("uncertain") else d.get("_error", "?"))
            print(f"  [{i}/{len(parts)}] {p.pn_std[:28]} → {tag}", flush=True)
            results.append((p, d))
        print(f"· 完成 {time.time() - t0:.0f}s", flush=True)
        _write_xlsx(Path(args.out), results)
        print(f"\n✓ 已写出 {args.out}（{len(results)} 行）")
        print("  人工在「★采纳」填 yes、「★最终PN」可改写；然后：")
        print(f"    uv run python scripts/pn_standardize.py apply --in {args.out}")
    finally:
        db.close()


def _write_xlsx(path: Path, results: list[tuple]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook(); ws = wb.active; ws.title = "PN标准化建议"
    ws.append(COLUMNS)
    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(1, c)
        star = name.startswith("★")
        cell.font = Font(bold=True, color="9C5700" if star else "FFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFF2CC" if star else "305496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for p, d in results:
        ws.append([
            p.id, p.pn_std, p.description, p.brand,
            " / ".join(x for x in (p.category_major, p.category_minor) if x) or None, p._specs,
            d.get("canonical_pn"), d.get("brand"), d.get("confidence"),
            bool(d.get("uncertain")), bool(d.get("not_a_part")),
            d.get("reason") or d.get("_error"), "", "", d.get("_error", ""),
        ])
    dv = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"M2:M{len(results) + 1}")   # 第13列=M「★采纳」
    for c, w in zip(range(1, len(COLUMNS) + 1),
                    [8, 26, 30, 12, 16, 24, 26, 12, 8, 8, 8, 34, 10, 22, 14]):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.freeze_panes = "A2"; wb.save(path)


def cmd_apply(args) -> None:
    from openpyxl import load_workbook

    from app.db import SessionLocal
    from app.services import part_rename

    wb = load_workbook(args.in_path, data_only=True); ws = wb.active
    H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
    db = SessionLocal()
    renamed, to_candidate, skipped, failed = 0, 0, 0, []
    try:
        for r in ws.iter_rows(min_row=2, values_only=True):
            pid = r[ix["part_id"]]
            target = adopt_target(r[ix["★采纳(yes)"]], r[ix["★最终PN(空=用建议)"]], r[ix["AI建议规范PN"]])
            if pid is None or target is None:
                skipped += 1
                continue
            try:
                res = part_rename.rename_part(db, int(pid), target,
                                              "标准化:AI建议+人工采纳", "agent:pn-standardize")
                if res.get("renamed"):
                    renamed += 1
                    print(f"  ✓ #{pid} {res['old_pn'][:24]} → {target}")
                else:
                    to_candidate += 1
                    print(f"  ↦ #{pid} 目标名被占用 → 转合并候选 #{res.get('merge_candidate_id')}")
            except part_rename.RenameError as e:
                failed.append((pid, str(e)))
                print(f"  ✗ #{pid} {e}")
                db.rollback()
        print(f"\n改名 {renamed} / 转候选 {to_candidate} / 跳过(未采纳) {skipped} / 失败 {len(failed)}")
    finally:
        db.close()


def cmd_selftest(_args) -> None:
    ok = True

    def check(name, cond):
        nonlocal ok
        print(f"  {'✓' if cond else '✗'} {name}")
        ok = ok and cond

    check("extract_json 复用", ev.extract_json('{"canonical_pn":"X"}') == {"canonical_pn": "X"})
    check("采纳=yes 用建议", adopt_target("yes", "", "AI-PN") == "AI-PN")
    check("采纳=yes 用最终PN覆盖", adopt_target("yes", "FINAL", "AI-PN") == "FINAL")
    check("采纳=no 不改", adopt_target("no", "X", "Y") is None)
    check("采纳空 不改", adopt_target("", "X", "Y") is None)
    check("采纳yes但都空 → None", adopt_target("yes", "", "") is None)
    print("自测通过 ✓" if ok else "自测有失败 ✗")
    sys.exit(0 if ok else 1)


def main() -> None:
    ap = argparse.ArgumentParser(description="PN 标准化发现层（不联网 MVP）")
    sub = ap.add_subparsers(dest="cmd", required=True)
    sg = sub.add_parser("suggest")
    sg.add_argument("--out", default="pn_std_suggest.xlsx")
    sg.add_argument("--limit", type=int, default=50)
    sg.add_argument("--offset", type=int, default=0)
    sg.set_defaults(func=cmd_suggest)
    ap_ = sub.add_parser("apply")
    ap_.add_argument("--in", dest="in_path", default="pn_std_suggest.xlsx")
    ap_.set_defaults(func=cmd_apply)
    st = sub.add_parser("selftest")
    st.set_defaults(func=cmd_selftest)
    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
