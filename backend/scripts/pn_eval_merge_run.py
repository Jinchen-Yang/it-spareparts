"""执行 AI=same 中规则①②够格的合并（用户授权·直接执行）。
模式: plan(只列清单) | smoke(对一空壳对做合并+回滚round-trip) | go(批量合并+重算)
方向: 留有业务历史/体量大的一方为目标(存活)，墓碑掉空壳/小的一方。
平台≠整机(#29/#221/#251)按用户裁定排除。"""
import sys, statistics, time
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from app.config import get_settings
from app.db import SessionLocal
from app.services import merge, profit, inventory

PRICE_TOL, NEAR_DAYS = 0.15, 180
EXCLUDE_SERIAL = {29, 221, 251}   # 平台/整机，保持独立

eng = create_engine(get_settings().database_url)
wb = load_workbook("pn_eval.xlsx", data_only=True); ws = wb.active
H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
def g(r, n): return r[ix[n]]
sames = [r for r in ws.iter_rows(min_row=2, values_only=True) if str(g(r,"AI判定")).strip()=="same"]

def info(conn, pn):
    row = conn.execute(text("SELECT id,status FROM dim_part WHERE pn_std=:p"), {"p": pn}).fetchone()
    if not row: return None
    pid, stt = row
    pur = conn.execute(text("SELECT pl.unit_price,po.order_date FROM f_purchase_line pl "
        "JOIN f_purchase_order po ON pl.order_id=po.id WHERE pl.part_id=:i"), {"i": pid}).fetchall()
    sal = conn.execute(text("SELECT sl.unit_price,so.order_date FROM f_sales_line sl "
        "JOIN f_sales_order so ON sl.order_id=so.id WHERE sl.part_id=:i"), {"i": pid}).fetchall()
    cl = lambda rows: [(float(p), d) for p, d in rows if p is not None]
    return {"id": pid, "status": stt, "pur": cl(pur), "sal": cl(sal)}

def med(xs): return statistics.median(xs) if xs else None
def drange(rows):
    ds = [d for _, d in rows if d]; return (min(ds), max(ds)) if ds else None
def near(a, b):
    if not a or not b: return None
    lo, hi = max(a[0], b[0]), min(a[1], b[1]); return 0 if lo <= hi else (lo - hi).days
def gap(pa, pb):
    if pa is None or pb is None: return None
    return abs(pa - pb) / max(min(pa, pb), 1e-9)

def classify(A, B):
    ap, bp, as_, bs = len(A["pur"]), len(B["pur"]), len(A["sal"]), len(B["sal"])
    if min(ap, bp) == 0: return "①空壳"
    pm = gap(med([p for p,_ in A["pur"]]), med([p for p,_ in B["pur"]])); np_ = near(drange(A["pur"]), drange(B["pur"]))
    sm = gap(med([p for p,_ in A["sal"]]), med([p for p,_ in B["sal"]])) if as_ and bs else None
    ns = near(drange(A["sal"]), drange(B["sal"])) if as_ and bs else None
    okp = pm is not None and pm <= PRICE_TOL and np_ is not None and np_ <= NEAR_DAYS
    oks = sm is not None and sm <= PRICE_TOL and ns is not None and ns <= NEAR_DAYS
    return "②价格佐证" if (okp or oks) else "other"

def direction(A, B, apn, bpn):
    va, vb = len(A["pur"])+len(A["sal"]), len(B["pur"])+len(B["sal"])
    if va != vb: return (bpn, apn) if va > vb else (apn, bpn)      # 体量大者为目标
    if len(apn) != len(bpn): return (bpn, apn) if len(apn) > len(bpn) else (apn, bpn)
    return (bpn, apn)

plan = []
with eng.connect() as conn:
    for r in sames:
        serial = g(r,"序号")
        if serial in EXCLUDE_SERIAL: continue
        apn, bpn = str(g(r,"源型号")), str(g(r,"候选型号"))
        A, B = info(conn, apn), info(conn, bpn)
        if not A or not B or A["status"] != "active" or B["status"] != "active": continue
        cat = classify(A, B)
        if cat not in ("①空壳", "②价格佐证"): continue
        src, tgt = direction(A, B, apn, bpn)
        plan.append({"#": serial, "src": src, "tgt": tgt, "cat": cat, "conf": g(r,"AI置信度")})

mode = sys.argv[1] if len(sys.argv) > 1 else "plan"
print(f"待合并清单 {len(plan)} 对（墓碑 → 存活）：")
for p in plan:
    print(f"  #{p['#']:>3} [{p['cat']}] {p['src'][:32]:<33} → {p['tgt'][:32]}")

if mode == "plan":
    sys.exit(0)

if mode == "smoke":
    db = SessionLocal()
    try:
        p = next((x for x in plan if x["#"] == 93), plan[0])
        print(f"\n=== SMOKE 合并+回滚 round-trip: #{p['#']} {p['src']} → {p['tgt']} ===")
        res = merge.merge_parts(db, p["src"], p["tgt"], f"smoke #{p['#']}", "agent:pn-match-eval")
        print(" 合并 ok  log_id =", res["merge_log_id"], " counts =", res["affected_counts"])
        s = db.execute(text("SELECT status,merged_into_id FROM dim_part WHERE pn_std=:p"), {"p": p["src"]}).fetchone()
        print(" 源状态 =", s, "(应 merged)")
        un = merge.unmerge(db, res["merge_log_id"], "agent:pn-match-eval")
        print(" 回滚 ok =", un)
        s2 = db.execute(text("SELECT status FROM dim_part WHERE pn_std=:p"), {"p": p["src"]}).fetchone()
        print(" 回滚后源状态 =", s2, "(应 active) → round-trip", "通过 ✓" if s2 and s2[0]=="active" else "失败 ✗")
    finally:
        db.close()
    sys.exit(0)

if mode == "go":
    db = SessionLocal()
    done, failed = [], []
    try:
        for p in plan:
            try:
                res = merge.merge_parts(db, p["src"], p["tgt"],
                    f"AI同款(conf {p['conf']})+规则{p['cat']}; pn_match_eval #{p['#']}", "agent:pn-match-eval")
                done.append({"#": p["#"], "log": res["merge_log_id"], "src": p["src"], "tgt": p["tgt"],
                             "counts": res["affected_counts"]})
                print(f"  ✓ #{p['#']:>3} log#{res['merge_log_id']} {p['src'][:22]:<23}→ {p['tgt'][:22]}")
            except Exception as e:
                db.rollback()
                failed.append({"#": p["#"], "err": f"{type(e).__name__}: {e}", "src": p["src"], "tgt": p["tgt"]})
                print(f"  ✗ #{p['#']} {type(e).__name__}: {e}")
        print(f"\n合并完成: 成功 {len(done)} / 失败 {len(failed)}")
        print("重算利润 + 库存成本…")
        t0 = time.time()
        rc = profit.recompute(db); ic = inventory.backfill_costs(db); db.commit()
        print(f" recompute = {rc}")
        print(f" inventory_costs = {ic}")
        print(f" 重算耗时 {time.time()-t0:.0f}s")
    finally:
        db.close()
    if failed:
        print("\n失败明细:")
        for f in failed: print(f"  #{f['#']} {f['src']} → {f['tgt']}: {f['err']}")
    print("\nlog_ids(可 unmerge):", [d["log"] for d in done])
    sys.exit(0)
