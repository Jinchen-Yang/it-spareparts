"""按 序号 强制合并指定 AI=same 对（用于写法差异/联网确认后的补合）。方向=体量大者存活。"""
import sys, time
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from app.config import get_settings
from app.db import SessionLocal
from app.services import merge, profit, inventory

import argparse
_ap = argparse.ArgumentParser()
_ap.add_argument("--in", dest="inp", default="pn_eval.xlsx")
_ap.add_argument("--reason", default="归一化近似同款(写法/品牌同义/token序/spec隐含)")
_ap.add_argument("serials", nargs="+", type=int)
_args = _ap.parse_args()
serials = set(_args.serials)
reason_tag = _args.reason
eng = create_engine(get_settings().database_url)
wb = load_workbook(_args.inp, data_only=True); ws = wb.active
H = [c.value for c in ws[1]]; ix = {n: i for i, n in enumerate(H)}
def g(r, n): return r[ix[n]]
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if g(r,"序号") in serials]

def vol(conn, pn):
    row = conn.execute(text("SELECT id,status FROM dim_part WHERE pn_std=:p"), {"p": pn}).fetchone()
    if not row: return None
    pid, stt = row
    pc = conn.execute(text("SELECT count(*) FROM f_purchase_line WHERE part_id=:i"), {"i": pid}).scalar()
    sc = conn.execute(text("SELECT count(*) FROM f_sales_line WHERE part_id=:i"), {"i": pid}).scalar()
    return {"id": pid, "status": stt, "v": pc + sc}

plan = []
with eng.connect() as conn:
    for r in rows:
        apn, bpn = str(g(r,"源型号")), str(g(r,"候选型号"))
        A, B = vol(conn, apn), vol(conn, bpn)
        if not A or not B or A["status"] != "active" or B["status"] != "active":
            print(f"  跳过 #{g(r,'序号')}: A={A and A['status']} B={B and B['status']}"); continue
        if A["v"] != B["v"]:
            src, tgt = (bpn, apn) if A["v"] > B["v"] else (apn, bpn)
        else:
            src, tgt = (bpn, apn) if len(apn) >= len(bpn) else (apn, bpn)
        plan.append((g(r,"序号"), src, tgt, g(r,"AI置信度")))

db = SessionLocal(); done = []
for serial, src, tgt, conf in plan:
    try:
        res = merge.merge_parts(db, src, tgt, f"AI同款(conf {conf})+{reason_tag}; pn_match_eval #{serial}", "agent:pn-match-eval")
        done.append(res["merge_log_id"])
        print(f"  ✓ #{serial} log#{res['merge_log_id']} {src} → {tgt}")
    except Exception as e:
        db.rollback(); print(f"  ✗ #{serial} {type(e).__name__}: {e}")
if done:
    profit.recompute(db); inventory.backfill_costs(db); db.commit(); print("重算完成")
db.close()
print("log_ids:", done)
