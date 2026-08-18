#!/usr/bin/env python3
"""V2 工作簿全量往返测试（2026-08-18）——多项目 × 全 sheet。

策略：主动造数据保证每个 sheet 有内容可测（生产库领用/回款为空），
每个 sheet 走「下载 → 填不同信息 → validate → apply → 查库断言」闭环，
并验证只读列修改被哈希守卫拒绝。测试结束清理造的数据。

用法：DATABASE_URL=... python3 scripts/v2_full_roundtrip.py
"""
import io
import sys
import uuid
from datetime import date, timedelta

import openpyxl
from sqlalchemy import text

from app.db import SessionLocal
from app.services import maintenance_project_master_workbook as master
from app.models.maintenance_project_operations import (
    MaintenanceSiteIssue,
    MaintenanceSiteIssueLine,
)

PASS = 0
FAIL = 0


def check(name: str, condition: bool, detail: str = "") -> None:
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"    ✅ {name}")
    else:
        FAIL += 1
        print(f"    ❌ {name} {detail}")


def v_validate(db, project_id: str, content: bytes):
    return master.validate_project_master_v2(db, project_id=project_id, data=content)


def v_apply(db, project_id: str, content: bytes, operator: str = "full-roundtrip"):
    plan = master.validate_project_master_v2(db, project_id=project_id, data=content)
    return master.apply_project_master_v2(db, plan, operated_by=operator,
                                          import_batch_id=f"rt-{uuid.uuid4().hex[:10]}")


def save_wb(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def prep_data(db, project_id: str):
    """给测试项目造合同/报销/领用/回款计划（返回可清理的 id 列表）。"""
    created = {"contracts": [], "expenses": [], "site_lines": [], "milestones": [], "snapshots": []}
    proj = db.execute(text(
        "SELECT project_code, display_name, salesperson, cmo_name FROM maintenance_project WHERE project_id=:p"),
        {"p": project_id}).one()
    # 合同（含金额，供 02/05/概览）
    cid = str(uuid.uuid4())
    ccid = str(uuid.uuid4())
    db.execute(text("""
        INSERT INTO maintenance_project_contract
        (project_contract_id, project_id, contract_id, contract_no, contract_amount,
         amount_inc_tax, contract_status, status_mapping_state, status_mapping_version,
         included_in_total, effective_from, effective_to, source, version)
        VALUES (:ccid, :pid, :cid, :cno, 100000.00, 113000.00, '正常', 'mapped', 'v1',
                true, :dfrom, :dto, 'test-seed', 1)
    """), {"ccid": ccid, "pid": project_id, "cid": cid,
           "cno": f"TEST-CONTRACT-{uuid.uuid4().hex[:6]}",
           "dfrom": date.today() - timedelta(days=30),
           "dto": date.today() + timedelta(days=330)})
    contract_no = db.execute(text(
        "SELECT contract_no FROM maintenance_project_contract WHERE project_contract_id=:c"),
        {"c": ccid}).scalar()
    created["contracts"].append(ccid)
    # 报销（BXD 导入形态）
    raw_line = f"test-bxd-{uuid.uuid4().hex[:8]}"
    db.execute(text("""
        INSERT INTO f_project_expense
        (raw_line_id, bxd_no, expense_date, person, expense_type, fee_category, reason,
         linked_sales_order_no, amount_ex_tax, amount_inc_tax, amount, tax_basis,
         data_status, remark, import_batch_id)
        VALUES (:r, :bxd, :d, '测试报销员', '交通', '差旅', '测试报销', :cno,
                100.00, 113.00, 100.00, 'ex', 'approved', '测试备注原始', 168)
    """), {"r": raw_line, "bxd": f"TEST-BXD-{uuid.uuid4().hex[:6]}", "d": date.today(),
           "cno": contract_no})
    created["expenses"].append(raw_line)
    # 领用单 + 行（ORM 插入，自动处理默认值）
    issue_id = str(uuid.uuid4())
    line_id = str(uuid.uuid4())
    db.add(MaintenanceSiteIssue(
        issue_id=issue_id, project_id=project_id,
        issue_no=f"TEST-SITE-{uuid.uuid4().hex[:6]}", issue_date=date.today(),
        raw_status="已确认", status_mapping_state="mapped",
        normalized_status="confirmed", status_mapping_version="v1",
    ))
    db.add(MaintenanceSiteIssueLine(
        issue_line_id=line_id, issue_id=issue_id, line_no=1, part_id=1,
        pn="TEST-PN-001", quantity=2, serial_number="SN123", no_return=False,
        remark="测试领用备注", algorithm_version="synthetic-v1",
    ))
    created["site_lines"].append(line_id)
    db.commit()
    return created, contract_no


def cleanup_data(db, project_id: str) -> None:
    # 先删引用合同/项目的子表，再删合同
    db.execute(text("DELETE FROM maintenance_collection_milestone WHERE project_id=:p"),
               {"p": project_id})
    db.execute(text("DELETE FROM maintenance_collection_snapshot WHERE project_id=:p "
                    "AND remark LIKE 'roundtrip-%'"), {"p": project_id})
    db.execute(text("DELETE FROM f_project_expense WHERE raw_line_id LIKE 'test-bxd-%'"))
    db.execute(text("DELETE FROM maintenance_site_issue_line WHERE issue_id IN "
                    "(SELECT issue_id FROM maintenance_site_issue WHERE issue_no LIKE 'TEST-SITE-%')"))
    db.execute(text("DELETE FROM maintenance_site_issue WHERE project_id=:p "
                    "AND issue_no LIKE 'TEST-SITE-%'"), {"p": project_id})
    db.execute(text("DELETE FROM maintenance_project_contract WHERE project_id=:p "
                    "AND source='test-seed'"), {"p": project_id})
    db.commit()


def test_sheet_02_plan(db, project_id: str, contract_no: str) -> None:
    print("  [02_回款计划]")
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["02_回款计划"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    # CREATE
    ws.cell(2, h["操作"] + 1).value = "CREATE"
    ws.cell(2, h["合同编号"] + 1).value = contract_no
    ws.cell(2, h["期次"] + 1).value = 1
    ws.cell(2, h["计划回款日期"] + 1).value = "2026-10-31"
    ws.cell(2, h["计划回款金额（含税）"] + 1).value = 30000.00
    ws.cell(2, h["备注"] + 1).value = "roundtrip-CREATE"
    v_apply(db, project_id, save_wb(wb))
    db.commit()
    ms = db.execute(text(
        "SELECT milestone_id, planned_amount FROM maintenance_collection_milestone "
        "WHERE project_id=:p AND planned_date='2026-10-31' AND follow_up_note='roundtrip-CREATE'"),
        {"p": project_id}).one_or_none()
    check("CREATE 落库", ms is not None and float(ms[1]) == 30000.00, f"ms={ms}")
    # UPDATE（改金额）
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["02_回款计划"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    found = False
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, h["备注"] + 1).value == "roundtrip-CREATE":
            ws.cell(r, h["操作"] + 1).value = "UPDATE"
            ws.cell(r, h["计划回款金额（含税）"] + 1).value = 35000.00
            ws.cell(r, h["备注"] + 1).value = "roundtrip-UPDATE"
            found = True
            break
    if found:
        v_apply(db, project_id, save_wb(wb))
        db.commit()
        ms = db.execute(text(
            "SELECT planned_amount FROM maintenance_collection_milestone "
            "WHERE project_id=:p AND follow_up_note='roundtrip-UPDATE'"),
            {"p": project_id}).one_or_none()
        check("UPDATE 落库", ms is not None and float(ms[0]) == 35000.00, f"ms={ms}")
    # VOID
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["02_回款计划"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, h["备注"] + 1).value == "roundtrip-UPDATE":
            ws.cell(r, h["操作"] + 1).value = "VOID"
            break
    v_apply(db, project_id, save_wb(wb))
    db.commit()
    active = db.execute(text(
        "SELECT count(*) FROM maintenance_collection_milestone WHERE project_id=:p "
        "AND planned_date='2026-10-31' AND is_active"), {"p": project_id}).scalar()
    check("VOID 后不再 active", active == 0, f"active={active}")


def test_sheet_05_receipts(db, project_id: str, contract_no: str) -> None:
    print("  [05_实收回款]")
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["05_实收回款"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    ws.cell(2, h["合同编号"] + 1).value = contract_no
    ws.cell(2, h["报告月份"] + 1).value = "2026-08"
    ws.cell(2, h["累计实收金额（含税）"] + 1).value = 40000.00
    ws.cell(2, h["回款凭证号"] + 1).value = "RT-RECEIPT-001"
    ws.cell(2, h["备注"] + 1).value = "roundtrip-receipt"
    v_apply(db, project_id, save_wb(wb))
    db.commit()
    snap = db.execute(text(
        "SELECT cumulative_amount FROM maintenance_collection_snapshot "
        "WHERE project_id=:p AND report_month='2026-08-01' AND remark='roundtrip-receipt'"),
        {"p": project_id}).one_or_none()
    check("CREATE 落库", snap is not None and float(snap[0]) == 40000.00, f"snap={snap}")
    # UPDATE
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["05_实收回款"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, h["备注"] + 1).value == "roundtrip-receipt":
            ws.cell(r, h["累计实收金额（含税）"] + 1).value = 45000.00
            ws.cell(r, h["备注"] + 1).value = "roundtrip-receipt-upd"
            break
    v_apply(db, project_id, save_wb(wb))
    db.commit()
    snap = db.execute(text(
        "SELECT cumulative_amount FROM maintenance_collection_snapshot "
        "WHERE project_id=:p AND remark='roundtrip-receipt-upd'"),
        {"p": project_id}).one_or_none()
    check("UPDATE 落库", snap is not None and float(snap[0]) == 45000.00, f"snap={snap}")


def test_sheet_04_expense(db, project_id: str) -> None:
    print("  [04_费用报销]")
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["04_费用报销"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    # 找到造好的报销行（第 2 行）
    row2 = [c.value for c in ws[2]]
    entity = row2[h["实体ID"]]
    check("报销行导出有实体ID", entity is not None, f"entity={entity}")
    if entity:
        # 改未税金额 + 备注
        ws.cell(2, h["未税金额"] + 1).value = 222.00
        ws.cell(2, h["备注"] + 1).value = "roundtrip-expense"
        v_apply(db, project_id, save_wb(wb))
        db.commit()
        exp = db.execute(text(
            "SELECT amount_ex_tax, remark FROM f_project_expense WHERE raw_line_id=:r"),
            {"r": entity}).one_or_none()
        check("金额/备注落库", exp is not None and float(exp[0]) == 222.00
              and exp[1] == "roundtrip-expense", f"exp={exp}")


def test_sheet_06_site(db, project_id: str) -> None:
    print("  [06_领用返还]")
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["06_领用返还"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    row2 = [c.value for c in ws[2]]
    entity = row2[h["实体ID"]]
    check("领用行导出有实体ID", entity is not None, f"entity={entity}")
    if entity:
        # 改是否应返还 + 备注 + 数量
        ws.cell(2, h["是否应返还"] + 1).value = "是"
        ws.cell(2, h["备注"] + 1).value = "roundtrip-site"
        ws.cell(2, h["领用数量"] + 1).value = 5
        v_apply(db, project_id, save_wb(wb))
        db.commit()
        site = db.execute(text(
            "SELECT no_return, remark, quantity FROM maintenance_site_issue_line WHERE issue_line_id=:i"),
            {"i": entity}).one_or_none()
        # 导出语义：no_return=True 显示"否"，no_return=False 显示"是"（#43 口径）
        check("领用落库", site is not None and site[0] is False and site[1] == "roundtrip-site"
              and int(site[2]) == 5, f"site={site}")


def test_sheet_03_parts(db, project_id: str) -> None:
    print("  [03_备件明细]")
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb["03_备件明细"]
    h = {c.value: i for i, c in enumerate(ws[1])}
    row2 = [c.value for c in ws[2]]
    entity = row2[h["实体ID"]]
    check("备件行导出有实体ID", entity is not None, f"entity={entity}")
    if entity:
        ws.cell(2, h["人工未税单位成本"] + 1).value = 456.78
        ws.cell(2, h["人工成本原因"] + 1).value = "roundtrip-parts"
        ws.cell(2, h["备注"] + 1).value = "roundtrip-parts-note"
        v_apply(db, project_id, save_wb(wb))
        db.commit()
        ov = db.execute(text(
            "SELECT unit_cost_ex_tax, reason FROM maintenance_manual_cost_override WHERE line_id=:i"),
            {"i": entity}).one_or_none()
        check("成本覆盖落库", ov is not None and float(ov[0]) == 456.78 and ov[1] == "roundtrip-parts",
              f"ov={ov}")
        note = db.execute(text("SELECT line_note FROM f_maintenance_line WHERE id=:i"),
                          {"i": entity}).scalar()
        check("备注落库", note == "roundtrip-parts-note", f"note={note!r}")
        # 哈希守卫：改只读列（PN）应拒
        wb2 = openpyxl.load_workbook(io.BytesIO(content))
        ws2 = wb2["03_备件明细"]
        ws2.cell(2, h["PN"] + 1).value = "MODIFIED-PN"
        try:
            v_validate(db, project_id, save_wb(wb2))
            check("改只读列(PN) 应拒", False, "竟然通过")
        except master.WorkbookError as exc:
            check("改只读列(PN) 被拒", exc.code == "readonly_cell_modified", f"code={exc.code}")


def main() -> int:
    db = SessionLocal()
    try:
        projects = db.execute(text("""
            SELECT a.project_id, p.display_name, count(*) AS parts
            FROM maintenance_source_order_assignment a
            JOIN maintenance_project p ON p.project_id = a.project_id
            WHERE a.is_active
            GROUP BY a.project_id, p.display_name
            ORDER BY parts DESC LIMIT 2
        """)).all()
        for pid, name, parts in projects:
            print(f"\n===== 项目 {name} (parts={parts}) =====")
            created, contract_no = prep_data(db, pid)
            try:
                test_sheet_03_parts(db, pid)
                test_sheet_02_plan(db, pid, contract_no)
                test_sheet_05_receipts(db, pid, contract_no)
                test_sheet_04_expense(db, pid)
                test_sheet_06_site(db, pid)
            finally:
                cleanup_data(db, pid)
        print(f"\n========== 结果：PASS={PASS} FAIL={FAIL} ==========")
        return 0 if FAIL == 0 else 1
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
