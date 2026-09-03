"""2.7.0 项目总表行级三路合并：基线令牌 + rebase + 冲突 + 强制接管。

矩阵对应 2026-09-02 维修计划的验收口径：
- 未触碰行在服务端变更后原样重传 → 不再整本 409（自动 rebase、零回写）；
- 不同行并发编辑互不阻塞；
- 同行同字段并发 → 冲突三值明细 + 零写入；force_takeover 覆盖留痕；
- 删行=作废但服务端该行已变 → 整行冲突；接管后作废生效；
- 基线令牌防篡改。
"""
import io
import uuid
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.models.maintenance import FMaintenanceLine
from app.models.maintenance_project_operations import (
    MaintenanceProjectOperationAudit,
)
from app.services import maintenance_project_master_workbook as master
from app.services import maintenance_project_operations as operations

from tests.test_maintenance_project_master_v2_editable import (
    _batch,
    _make_project_with_line,
    _save,
)


def _add_line(db, order, part, *, qty=Decimal("5")) -> FMaintenanceLine:
    line = FMaintenanceLine(
        raw_line_id=f"raw-line-{uuid.uuid4()}", order_id=order.id,
        line_no=order.id % 100 + 10, part_id=part.id,
        pn_std=part.pn_std, pn_raw=part.pn_std, description=part.description,
        qty=qty, return_qty=Decimal("0"), cost_source="direct",
        cost_tax_basis="ex", confidence="high",
        import_batch_id=_batch(db),
    )
    db.add(line)
    db.commit()
    return line


def _download_parts(db, project_id: str):
    content = master.build_project_master_v2(
        db, project_id=project_id, sheets=(master.V2_SHEET_PARTS,))
    return content, load_workbook(io.BytesIO(content))


def _row_of(ws, entity_id) -> int:
    """按隐藏实体ID列定位行号（不假设导出排序）。"""
    headers = {c.value: c.column for c in ws[1]}
    id_col = headers["实体ID"]
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, id_col).value or "").strip() == str(entity_id):
            return row
    raise AssertionError(f"未找到实体ID {entity_id} 所在行")


def _apply(db, project_id, wb, *, force=False, operated_by="merge-tester"):
    plan = master.validate_project_master_v2(
        db, project_id=project_id, data=_save(wb), force_takeover=force)
    return plan, master.apply_project_master_v2(
        db, plan, operated_by=operated_by, import_batch_id=str(uuid.uuid4()))


def test_untouched_rows_rebase_after_foreign_write(db):
    """旧 409 场景回归：他人写入推进 revision 后，未触碰行原样重传应成功零回写。"""
    project, part, order, line = _make_project_with_line(db)
    second = _add_line(db, order, part)
    content, _ = _download_parts(db, project.project_id)
    wb = load_workbook(io.BytesIO(content))

    # 模拟他人写入：直接改第二行并推进 revision（等同一次导入/他人应用）
    second.qty = Decimal("9")
    operations.bump_workbook_revision(db, project_id=project.project_id)
    db.commit()

    # 原样重传（未触碰任何行）
    plan, result = _apply(db, project.project_id, wb)
    assert not plan.conflicts
    assert result["revision_drift"] is True
    assert result["changes"] == []
    db.refresh(second)
    assert second.qty == Decimal("9.00"), "未触碰行不得回写导出旧值"
    db.refresh(line)
    assert line.qty == Decimal("2.00")


def test_stale_export_value_does_not_overwrite_server_change(db):
    """反方向保护：文件里携带旧值但用户没改 → 不产生任何 UPDATE（防静默覆盖）。"""
    project, _part, _order, line = _make_project_with_line(db)
    content, _ = _download_parts(db, project.project_id)
    wb = load_workbook(io.BytesIO(content))

    line.qty = Decimal("7")
    operations.bump_workbook_revision(db, project_id=project.project_id)
    db.commit()

    plan, result = _apply(db, project.project_id, wb)
    assert result["changes"] == []
    db.refresh(line)
    assert line.qty == Decimal("7.00")


def test_disjoint_rows_concurrent_edits_both_succeed(db):
    project, part, order, line = _make_project_with_line(db)
    second = _add_line(db, order, part)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))

    first_wb = load_workbook(io.BytesIO(content))
    second_wb = load_workbook(io.BytesIO(content))
    first_ws = first_wb[master.V2_SHEET_PARTS]
    second_ws = second_wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in first_ws[1]}
    first_ws.cell(_row_of(first_ws, line.id), headers["需求数量"], 3)
    second_ws.cell(_row_of(second_ws, second.id), headers["需求数量"], 8)

    _, first_result = _apply(db, project.project_id, first_wb, operated_by="a")
    assert first_result["changes"], first_result
    _, second_result = _apply(db, project.project_id, second_wb, operated_by="b")
    assert not second_result["conflicts"]
    db.refresh(line)
    db.refresh(second)
    assert line.qty == Decimal("3.00")
    assert second.qty == Decimal("8.00")


def test_row_conflict_lists_three_values_and_zero_write(db):
    project, _part, _order, line = _make_project_with_line(db)
    content, _ = _download_parts(db, project.project_id)
    ws_row = load_workbook(io.BytesIO(content))[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws_row[1]}
    ws_row.cell(_row_of(ws_row, line.id), headers["需求数量"], 4)
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    ws.cell(_row_of(ws, line.id), headers["需求数量"], 6)

    # 他人先把 qty 改成 4（即 ws_row 的值）并提交
    _, _ = _apply(db, project.project_id, ws_row.parent, operated_by="other")

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert plan.conflicts
    conflict = plan.conflicts[0]
    assert conflict["field"] == "需求数量"
    assert conflict["entity_id"] == str(line.id)
    audits_before = int(db.scalar(
        select(func.count(MaintenanceProjectOperationAudit.id))) or 0)
    with pytest.raises(master.WorkbookError) as raised:
        master.apply_project_master_v2(
            db, plan, operated_by="loser", import_batch_id=str(uuid.uuid4()))
    assert raised.value.code == "row_conflicts"
    db.rollback()
    db.refresh(line)
    assert line.qty == Decimal("4.00")
    assert int(db.scalar(
        select(func.count(MaintenanceProjectOperationAudit.id)))) == audits_before


def test_force_takeover_overrides_and_records(db):
    project, _part, _order, line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    other_wb = load_workbook(io.BytesIO(content))
    my_wb = load_workbook(io.BytesIO(content))
    headers = {c.value: c.column for c in other_wb[master.V2_SHEET_PARTS][1]}
    other_wb[master.V2_SHEET_PARTS].cell(
        _row_of(other_wb[master.V2_SHEET_PARTS], line.id), headers["需求数量"], 4)
    my_wb[master.V2_SHEET_PARTS].cell(
        _row_of(my_wb[master.V2_SHEET_PARTS], line.id), headers["需求数量"], 6)
    _, _ = _apply(db, project.project_id, other_wb, operated_by="other")

    plan, result = _apply(db, project.project_id, my_wb, force=True,
                          operated_by="taker")
    assert not plan.conflicts
    assert result["force_takeover"] is True
    assert result["overridden"] and result["overridden"][0]["field"] == "需求数量"
    db.refresh(line)
    assert line.qty == Decimal("6.00")


def test_deleted_row_conflicts_when_server_changed(db):
    project, part, order, line = _make_project_with_line(db)
    second = _add_line(db, order, part)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]

    # 他人修改第二行后，用户删除该行（清空整行 → 缺行=作废）
    second.qty = Decimal("9")
    operations.bump_workbook_revision(db, project_id=project.project_id)
    db.commit()
    blank_row = _row_of(ws, second.id)
    for col in range(1, len(master.V2_PART_HEADERS) + 1):
        ws.cell(blank_row, col).value = None

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert plan.conflicts, "删行但服务端该行已变必须整行冲突"
    assert plan.conflicts[0]["sheet"] == "03_备件明细"
    db.refresh(second)
    assert second.is_active is True, "未接管时零写入"

    takeover_plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb), force_takeover=True)
    assert not takeover_plan.conflicts
    master.apply_project_master_v2(
        db, takeover_plan, operated_by="taker",
        import_batch_id=str(uuid.uuid4()))
    db.refresh(second)
    assert second.is_active is False


def test_baseline_token_tamper_rejected(db):
    project, _part, _order, line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {c.value: c.column for c in ws[1]}
    # 篡改基线令牌（伪造成另一行的基线）
    ws.cell(_row_of(ws, line.id), headers["需求数量"], 3)
    ws.cell(_row_of(ws, line.id), headers[master.V2_BASE_COLUMN],
            '{"key_id":"dev-v1","payload":{"sheet":"03_备件明细",'
            '"entity":"999999","base":{}},"signature":"deadbeef"}')
    with pytest.raises(master.WorkbookError) as raised:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert raised.value.code == "invalid_concurrency_token"


def test_apply_result_carries_field_changes(db):
    project, _part, _order, line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    headers = {c.value: c.column for c in wb[master.V2_SHEET_PARTS][1]}
    row_no = _row_of(wb[master.V2_SHEET_PARTS], line.id)
    wb[master.V2_SHEET_PARTS].cell(row_no, headers["需求数量"], 3)
    wb[master.V2_SHEET_PARTS].cell(row_no, headers["备注"], "回执应记录字段改动")

    _, result = _apply(db, project.project_id, wb)
    fields = {(c["sheet"], c["entity_id"], c["field"]) for c in result["changes"]}
    assert ("03_备件明细", str(line.id), "需求数量") in fields
    assert ("03_备件明细", str(line.id), "备注") in fields
    db.refresh(line)
    assert line.qty == Decimal("3.00")


def test_site_pn_flexible_resolution_and_warnings(db):
    """2026-09-03：PN 柔性解析——粘连取首段/补前导零自动命中并告警；
    未匹配批量一次性报全量（不再首错即断）。"""
    from app.models.dimensions import DimPart

    project, part, order, _line = _make_project_with_line(db)
    glue = DimPart(pn_std="AL15SEB120N", description="粘连PN测试")
    zero = DimPart(pn_std="06200288", description="前导零货号测试")
    db.add_all([glue, zero])
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    headers = {c.value: c.column for c in ws[1]}
    # 写到示例行之后的干净行：示例行带【示例】备注，直接覆盖其部分
    # 单元格会被 _is_example_row 整行跳过（数据静默丢失）。
    base = ws.max_row + 1
    ws.cell(base, headers["领用单号"], "LY-001")
    ws.cell(base, headers["领用日期"], "2026-09-01")
    ws.cell(base, headers["PN"], "AL15SEB120N V0231E4000000000")
    ws.cell(base, headers["领用数量"], 1)
    ws.cell(base + 1, headers["领用单号"], "LY-002")
    ws.cell(base + 1, headers["领用日期"], "2026-09-01")
    ws.cell(base + 1, headers["PN"], "6200288")
    ws.cell(base + 1, headers["领用数量"], 1)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert not plan.conflicts
    assert any("粘连" in w for w in plan.warnings), plan.warnings
    assert any("前导零" in w for w in plan.warnings), plan.warnings
    result = master.apply_project_master_v2(
        db, plan, operated_by="pn-flex", import_batch_id=str(uuid.uuid4()))
    assert result["site_creates"] == 2
    assert any("粘连" in w for w in result["warnings"])


def test_unmatched_pns_reported_in_batch(db):
    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    headers = {c.value: c.column for c in ws[1]}
    base = ws.max_row + 1
    for offset, no in ((0, "LY-A"), (1, "LY-B")):
        ws.cell(base + offset, headers["领用单号"], no)
        ws.cell(base + offset, headers["领用日期"], "2026-09-01")
        ws.cell(base + offset, headers["PN"], f"TOTALLY-UNKNOWN-{no}")
        ws.cell(base + offset, headers["领用数量"], 1)
    with pytest.raises(master.WorkbookError) as raised:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert raised.value.code == "part_not_found"
    msg = raised.value.message
    assert "TOTALLY-UNKNOWN-LY-A" in msg and "TOTALLY-UNKNOWN-LY-B" in msg, msg
    assert "共 2 个" in msg


def test_duplicate_manual_rows_rejected_not_500(db):
    """同文件重复手工行（06 同单号/PN/SN；04 同单号+序号）解析期 422，
    不得走到 apply 期撞主键变 500（本地 E2E 实测 UniqueViolation）。"""
    from openpyxl import load_workbook as _lw
    from tests.test_maintenance_project_master_v2_editable import (
        _make_project_with_line as _mk,
    )
    from app.models.dimensions import DimPart

    project, part, order, _line = _mk(db)
    # 06 重复行
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    wb = _lw(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    headers = {c.value: c.column for c in ws[1]}
    base = ws.max_row + 1
    for r in (base, base + 1):
        ws.cell(r, headers["领用单号"], "CKD-DUP-001")
        ws.cell(r, headers["领用日期"], "2026-09-01")
        ws.cell(r, headers["PN"], part.pn_std)
        ws.cell(r, headers["SN"], "SN-SAME")
        ws.cell(r, headers["领用数量"], 1)
    with pytest.raises(master.WorkbookError) as raised:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert raised.value.code == "duplicate_site_row"

    # 04 重复行（同费用单号+序号）
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,))
    wb = _lw(io.BytesIO(content))
    ws = wb[master.V2_SHEET_EXPENSE]
    headers = {c.value: c.column for c in ws[1]}
    base = ws.max_row + 1
    from app.models.maintenance_project import MaintenanceProjectContract
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id))
    for r in (base, base + 1):
        ws.cell(r, headers["操作"], "CREATE")
        ws.cell(r, headers["费用单号"], "BXD-DUP-001")
        ws.cell(r, headers["明细序号"], 1)
        ws.cell(r, headers["报销日期"], "2026-09-01")
        ws.cell(r, headers["维保销售订单（归集键）"], contract.contract_no)
        ws.cell(r, headers["未税金额"], 100)
    with pytest.raises(master.WorkbookError) as raised:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert raised.value.code == "duplicate_expense_row"

def test_manual_row_claiming_existing_line_accepts_normalized_pn(db):
    """2026-09-03 契约：柔性解析命中的 PN，在「认领既有明细」路径上不得被二次严格匹配打死。

    真实场景（中国电信云 1028 行领用实录）：客户手工补一行明细，PN 带 WWN 粘连；
    对应明细在下载之后才由 WBDD 导入落库，因此该行走「认领已存在全局明细」分支。
    _resolve_part_flexible 已把粘连 PN 解析成标准 PN，但
    _v2_refill_for_existing_line 会重新读原始单元格做 _exact_part_for_pn——
    必然 part_not_found 且当场中断整本上传，正好废掉柔性解析要解决的场景。
    """
    from app.models.dimensions import DimPart

    project, _part, order, _line = _make_project_with_line(db)
    _content, wb = _download_parts(db, project.project_id)
    ws = wb[master.V2_SHEET_PARTS]
    headers = {c.value: c.column for c in ws[1]}

    # 下载之后才落库的一条明细——不在本次导出里，只能走认领分支
    glued_part = DimPart(pn_std="AL15SEB120N", description="粘连PN备件")
    db.add(glued_part)
    db.commit()
    late_line = _add_line(db, order, glued_part, qty=Decimal("7"))

    row = ws.max_row + 1
    ws.cell(row, headers["维保单号"], order.order_no)
    ws.cell(row, headers["PN"], "AL15SEB120N V0231E4000000000")
    ws.cell(row, headers["需求数量"], 7)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert any("粘连" in w for w in plan.warnings), plan.warnings
    master.apply_project_master_v2(
        db, plan, operated_by="claim-flex", import_batch_id=str(uuid.uuid4()))
    db.refresh(late_line)
    # 认领的是既有明细，不得凭空再造一条
    assert late_line.pn_std == "AL15SEB120N"
    assert db.scalar(select(func.count()).select_from(FMaintenanceLine).where(
        FMaintenanceLine.order_id == order.id,
        FMaintenanceLine.part_id == glued_part.id,
        FMaintenanceLine.is_active.is_(True),
    )) == 1


def test_untouched_field_not_reverted_when_server_changed_it(db):
    """D-02 契约：用户只碰字段 A，他人在导出后改了同行字段 B —— 上传不得把 B
    静默回滚成导出时的旧值。

    根因（2026-09-03 评审 P1）：_v2_merge_row 只做分类，不限制下游写什么；
    _v2_refill_for_existing_line 是拿整行单元格与服务端现值逐一比对生成 update，
    于是用户表里「没碰但已过期」的 B 会盖掉他人改动，且因 B 未触碰而不报冲突。
    修法是合并时把未触碰字段 rebase 成服务端现值再进 refill。
    """
    project, _part, _order, line = _make_project_with_line(db, qty=Decimal("5"))
    line.description = "原始描述"
    db.commit()

    _content, wb = _download_parts(db, project.project_id)
    ws = wb[master.V2_SHEET_PARTS]
    headers = {c.value: c.column for c in ws[1]}
    target = _row_of(ws, line.id)

    # 导出之后，另一个写入方改了「描述」——用户全程没碰这个字段
    line.description = "他人改的描述"
    db.commit()

    # 用户只改「需求数量」
    ws.cell(target, headers["需求数量"], 9)

    plan, _result = _apply(db, project.project_id, wb)
    db.refresh(line)
    assert line.qty == Decimal("9.000"), line.qty
    assert line.description == "他人改的描述", (
        f"静默回滚了他人的改动：{line.description!r}")
    # 未触碰字段自动 rebase，不应制造冲突
    assert not plan.conflicts, plan.conflicts
