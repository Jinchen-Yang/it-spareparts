"""2.7.0 项目总表编辑权下放：项目负责人/销售（2026-09-02 拍板全量放开）。

- primary_manager 挂靠或 canonical 销售 → 下载/校验/应用全量可用（无
  action 键、无 data_profit 也放行——当日拍板成本列对编辑者可见）；
- 仅 viewer 挂靠或无关账号 → 403 fail-closed；
- 合同额编辑对负责人/销售放开（此前需管理员双键）。
"""
import io
import uuid
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app import permissions as permissions_mod
from app.auth import hash_password
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectUserAssignment,
)
from app.models.system import SysUser
from app.security import UserContext
from app.services import maintenance_project_assignments as assignments
from app.services import maintenance_project_master_workbook as master

from tests.test_maintenance_project_master_v2_editable import (
    _make_project_with_line,
    _save,
)

_PASSWORD = "pw123456"


def _user(db, username, *, salesperson_name=None, display_name=None):
    base = permissions_mod.effective("readonly", None)
    user = SysUser(
        username=username, role="readonly", display_name=display_name or username,
        salesperson_name=salesperson_name,
        password_hash=hash_password(_PASSWORD), is_active=True,
        template_code="readonly", template_version=1, template_perms=base,
        perm_overrides={"page_maintenance": True},
        permissions=permissions_mod.effective_from_snapshot(base, {"page_maintenance": True}),
    )
    db.add(user)
    db.flush()
    return user


def _client_for(db, user) -> TestClient:
    db.commit()
    client = TestClient(app)
    login = client.post("/api/auth/login",
                        json={"username": user.username, "password": _PASSWORD})
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def _manager_client(db, project):
    user = _user(db, f"mgr-{uuid.uuid4().hex[:6]}")
    db.add(MaintenanceProjectUserAssignment(
        assignment_id=str(uuid.uuid4()), project_id=project.project_id,
        responsibility_type="primary_manager", user_id=user.id, version=1,
        assigned_by="test", assignment_reason="负责人编辑权测试",
    ))
    return _client_for(db, user)


def _sales_client(db, project, *, match=True):
    name = project.salesperson if match and project.salesperson else "销售甲"
    if match and not project.salesperson:
        project.salesperson = name
    user = _user(db, f"sales-{uuid.uuid4().hex[:6]}",
                 salesperson_name=name if match else "无关销售",
                 display_name=name if match else "无关销售")
    return _client_for(db, user)


def _viewer_client(db, project):
    user = _user(db, f"viewer-{uuid.uuid4().hex[:6]}")
    db.add(MaintenanceProjectUserAssignment(
        assignment_id=str(uuid.uuid4()), project_id=project.project_id,
        responsibility_type="viewer", user_id=user.id, version=1,
        assigned_by="test", assignment_reason="仅可见不应可编",
    ))
    return _client_for(db, user)


def _download(client, project_id):
    return client.get(f"/api/maintenance/projects/stable/{project_id}"
                      f"/master-workbook.xlsx")


def _upload(client, project_id, wb, *, force=False):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    files = {"file": ("m.xlsx", buf, "application/vnd.openxmlformats-"
                                     "officedocument.spreadsheetml.sheet")}
    data = {"force_takeover": "true"} if force else None
    return client.post(
        f"/api/maintenance/projects/stable/{project_id}/master-workbook/apply",
        files=files, data=data)


def test_is_project_workbook_editor_matrix(db):
    project, _part, _order, _line = _make_project_with_line(db)
    manager = _user(db, f"mgr-{uuid.uuid4().hex[:6]}")
    db.add(MaintenanceProjectUserAssignment(
        assignment_id=str(uuid.uuid4()), project_id=project.project_id,
        responsibility_type="primary_manager", user_id=manager.id, version=1,
        assigned_by="test", assignment_reason="负责人",
    ))
    db.commit()

    yes = UserContext(user_id=manager.username, role="readonly",
                      is_authenticated=True)
    assert assignments.is_project_workbook_editor(
        db, project_id=project.project_id, user_ctx=yes)

    project.salesperson = "销售甲"
    sales = UserContext(user_id="whatever", role="readonly",
                        salesperson_name="销售甲", is_authenticated=True)
    assert assignments.is_project_workbook_editor(
        db, project_id=project.project_id, user_ctx=sales)

    stranger = UserContext(user_id="nobody", role="readonly",
                           salesperson_name="别的销售", is_authenticated=True)
    assert not assignments.is_project_workbook_editor(
        db, project_id=project.project_id, user_ctx=stranger)
    anonymous = UserContext(user_id=None, role="readonly", is_authenticated=False)
    assert not assignments.is_project_workbook_editor(
        db, project_id=project.project_id, user_ctx=anonymous)


def test_project_manager_without_action_key_can_download_and_apply(
    db, monkeypatch,
):
    from app.config import get_settings
    monkeypatch.setattr(get_settings(), "maintenance_project_master_v2_enabled",
                        True)
    project, _part, _order, line = _make_project_with_line(db)
    client = _manager_client(db, project)

    resp = _download(client, project.project_id)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {c.value: c.column for c in ws[1]}
    target_row = next(
        r for r in range(2, ws.max_row + 1)
        if str(ws.cell(r, headers["实体ID"]).value or "") == str(line.id))
    ws.cell(target_row, headers["需求数量"], 3)

    applied = _upload(client, project.project_id, wb)
    assert applied.status_code == 200, applied.text
    assert applied.json()["line_updates"] >= 1
    db.refresh(line)
    assert line.qty == Decimal("3.00")


def test_salesperson_of_project_can_download(db, monkeypatch):
    from app.config import get_settings
    monkeypatch.setattr(get_settings(), "maintenance_project_master_v2_enabled",
                        True)
    project, _part, _order, _line = _make_project_with_line(db)
    client = _sales_client(db, project, match=True)
    resp = _download(client, project.project_id)
    assert resp.status_code == 200, resp.text


def test_viewer_or_outside_salesperson_denied(db, monkeypatch):
    from app.config import get_settings
    monkeypatch.setattr(get_settings(), "maintenance_project_master_v2_enabled",
                        True)
    project, _part, _order, _line = _make_project_with_line(db)

    viewer = _viewer_client(db, project)
    assert _download(viewer, project.project_id).status_code == 403

    outsider = _sales_client(db, project, match=False)
    assert _download(outsider, project.project_id).status_code == 403


def test_contract_amount_editable_by_manager(db, monkeypatch):
    from app.config import get_settings
    monkeypatch.setattr(get_settings(), "maintenance_project_master_v2_enabled",
                        True)
    project, _part, _order, _line = _make_project_with_line(db)
    client = _manager_client(db, project)

    resp = _download(client, project.project_id)
    assert resp.status_code == 200, resp.text
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb[master.V2_SHEET_OVERVIEW]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == "合同总额（含税）":
            ws.cell(row, 2, "20000.00")
            break
    else:
        pytest.fail("概览缺少合同总额行")
    applied = _upload(client, project.project_id, wb)
    assert applied.status_code == 200, applied.text
