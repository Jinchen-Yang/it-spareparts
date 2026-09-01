"""Database adapter contract for manager monthly workbook v3 (#206)."""

from __future__ import annotations

import io
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from sqlalchemy import select

from app.auth import hash_password
from app.models.maintenance_manager import (
    MaintenanceCollectionMilestone,
    MaintenanceManagerUploadBatch,
    MaintenanceManagerUploadBatchProject,
)
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
    MaintenanceProjectUserAssignment,
)
from app.models.maintenance_project_operations import MaintenanceCollectionSnapshot
from app.models.system import SysUser
from app.security import UserContext
from app.services.maintenance_manager_workbook_adapter import (
    MaintenanceManagerWorkbookAdapter,
    ManagerWorkbookConflict,
)
from app.services.maintenance_manager_workbook_v3 import (
    OVERVIEW_SHEET,
    PLAN_SHEET,
    PLAN_TABLE,
    ManagerWorkbookV3Error,
)
from app.services.maintenance_project_operations import project_operations


HMAC_KEY = b"synthetic-manager-workbook-adapter-key"


def _seed(db, *, username: str = "synthetic_manager") -> tuple[SysUser, str, str]:
    user = SysUser(
        username=username,
        role="purchaser",
        display_name="合成项目经理",
        password_hash=hash_password("synthetic-password-123"),
    )
    project = MaintenanceProject(
        project_id=f"project-{username}",
        project_code=f"PM-{username}",
        display_name="合成项目经理工作簿项目",
        lifecycle_status="ongoing",
    )
    db.add_all([user, project])
    db.flush()
    assignment = MaintenanceProjectUserAssignment(
        assignment_id=f"assignment-{username}",
        project_id=project.project_id,
        responsibility_type="primary_manager",
        user_id=user.id,
        version=1,
        assigned_at=datetime.now(UTC),
        assigned_by="synthetic-admin",
        assignment_reason="合成测试负责人映射",
    )
    contract = MaintenanceProjectContract(
        project_contract_id=f"pc-{username}",
        project_id=project.project_id,
        contract_id=f"contract-{username}",
        contract_no=f"XS-{username}",
        contract_amount=100000,
        amount_inc_tax=113000,
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=date(2026, 1, 1),
        source="synthetic-test",
    )
    actual = MaintenanceCollectionSnapshot(
        collection_id=f"collection-{username}",
        project_id=project.project_id,
        project_contract_id=contract.project_contract_id,
        report_month=date(2026, 8, 1),
        cumulative_amount=30000,
        status="confirmed",
    )
    db.add_all([assignment, contract, actual])
    db.commit()
    return user, project.project_id, contract.project_contract_id


def _adapter(db, user: SysUser) -> MaintenanceManagerWorkbookAdapter:
    return MaintenanceManagerWorkbookAdapter(
        db,
        user_ctx=_ctx(user),
        operator=user.username,
        as_of=date(2026, 8, 9),
    )


def _ctx(user: SysUser) -> UserContext:
    return UserContext(
        role=user.role,
        user_id=user.username,
        is_authenticated=True,
        permissions={
            "page_maintenance": True,
            "data_profit": True,
            "data_purchase_cost": True,
        },
    )


def _set_plan(content: bytes, relation_id: str) -> bytes:
    return _set_plan_values(content, relation_id, date(2026, 9, 15), 25000)


def _set_plan_values(
    content: bytes,
    relation_id: str,
    planned_date: date,
    planned_amount: int,
) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        sheet = book[PLAN_SHEET]
        table = sheet.tables[PLAN_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("项目合同关系ID") + 1).value == relation_id
            and sheet.cell(row, headers.index("计划期次") + 1).value == 1
        )
        sheet.cell(target, headers.index("计划回款日期") + 1, planned_date)
        sheet.cell(target, headers.index("计划回款金额（含税）") + 1, planned_amount)
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def test_validate_then_apply_is_atomic_and_replay_safe(db):
    user, project_id, relation_id = _seed(db)
    adapter = _adapter(db, user)
    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    validation, batch = adapter.validate(
        date(2026, 8, 1),
        _set_plan(artifact.content, relation_id),
        hmac_key=HMAC_KEY,
    )
    assert validation.can_apply is True
    assert len(validation.milestone_changes) == 1
    db.commit()

    result = adapter.apply(batch.batch_id)
    db.commit()
    milestone = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 1,
        )
    )
    assert milestone is not None
    assert milestone.planned_amount == 25000
    assert result["changed_rows"] == 1
    assert db.get(MaintenanceManagerUploadBatch, batch.batch_id).status == "applied"
    assert db.get(MaintenanceManagerUploadBatchProject, (batch.batch_id, project_id)) is not None

    replay = adapter.apply(batch.batch_id)
    db.commit()
    assert replay == result
    assert db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 1,
        )
    ).version == 1

    directory = project_operations(
        db,
        as_of=date(2026, 8, 9),
        user_ctx=_ctx(user),
        owner_scope="me",
    )
    # 2026-08-25：「项目经理月度更新」任务退役（工作簿入口不存在，死胡同提示）
    assert not [
        task for task in directory["rows"][0]["task_summary"]["rows"]
        if task["task_type"] == "项目经理月度更新"
    ]
    retired = project_operations(
        db,
        as_of=date(2026, 8, 9),
        user_ctx=_ctx(user),
        owner_scope="me",
        task_type="项目经理月度更新",
        task_status="pending",
    )
    assert retired["total"] == 0


def test_manager_snapshot_and_workbook_use_inc_tax_contract_amount(db):
    user, _project_id, relation_id = _seed(
        db, username="inc_tax_contract_manager")
    adapter = _adapter(db, user)
    artifact, snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)

    contract = snapshot["projects"][0]["contracts"][0]
    assert contract["project_contract_id"] == relation_id
    assert contract["contract_amount"] == 113000

    book = load_workbook(io.BytesIO(artifact.content), data_only=True)
    try:
        overview = book[OVERVIEW_SHEET]
        values = list(overview.values)
        assert any(Decimal(str(cell)) == Decimal("113000")
                   for row in values for cell in row if isinstance(cell, (int, float)))
    finally:
        book.close()


def test_manager_workbook_fails_closed_for_unmapped_contract_fact(db):
    """经理月表不得绕过项目卡的 canonical 合同完整性门禁计算回款率。"""
    user, project_id, relation_id = _seed(
        db, username="unmapped_contract_manager")
    contract = db.get(MaintenanceProjectContract, relation_id)
    contract.status_mapping_state = "unmapped"
    contract.included_in_total = False
    db.commit()

    artifact, snapshot = _adapter(db, user).export(
        date(2026, 8, 1), hmac_key=HMAC_KEY)

    project = snapshot["projects"][0]
    assert project["project_id"] == project_id
    assert project["contract_facts_complete"] is False
    book = load_workbook(io.BytesIO(artifact.content), data_only=True)
    try:
        overview = book[OVERVIEW_SHEET]
        row = list(overview.iter_rows(min_row=6, max_row=6, values_only=True))[0]
        assert row[6] is None                 # 全部合同额（含税）
        assert row[7] == "missing"           # 合同额完整性
        assert row[9] is None                 # 实收/合同额
        assert row[11] is None                # 计划/合同额
    finally:
        book.close()


@pytest.mark.parametrize("conflict_kind", ["duplicate", "shared"])
def test_manager_workbook_fails_closed_for_duplicate_or_shared_contract(
    db, conflict_kind,
):
    username = "dup_mgr" if conflict_kind == "duplicate" else "shared_mgr"
    user, project_id, relation_id = _seed(db, username=username)
    original = db.get(MaintenanceProjectContract, relation_id)
    other_project_id = project_id
    if conflict_kind == "shared":
        other = MaintenanceProject(
            project_id=f"project-{conflict_kind}-other",
            project_code=f"PM-{conflict_kind}-other",
            display_name="共享合同的另一个项目",
            lifecycle_status="ongoing",
        )
        db.add(other)
        db.flush()
        other_project_id = other.project_id
    db.add(MaintenanceProjectContract(
        project_contract_id=f"pc-{conflict_kind}-other",
        project_id=other_project_id,
        contract_id=original.contract_id,
        contract_no=f"XS-{conflict_kind}-other",
        contract_amount=Decimal("100000.00"),
        amount_inc_tax=Decimal("113000.00"),
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=(date(2026, 2, 1)
                        if conflict_kind == "duplicate" else date(2026, 1, 1)),
        source="synthetic-test",
    ))
    db.commit()

    artifact, snapshot = _adapter(db, user).export(
        date(2026, 8, 1), hmac_key=HMAC_KEY)

    assert snapshot["projects"][0]["contract_facts_complete"] is False
    book = load_workbook(io.BytesIO(artifact.content), data_only=True)
    try:
        row = list(book[OVERVIEW_SHEET].iter_rows(
            min_row=6, max_row=6, values_only=True))[0]
        assert row[6] is None
        assert row[7] == "missing"
        assert row[9] is None
        assert row[11] is None
    finally:
        book.close()


def test_adapter_writes_day_precision_and_pending_state_for_new_nodes(db):
    """manager workbook 写路径统一写 date_precision=day，新节点 pending/false（设计 §4.1）。"""
    user, _project_id, relation_id = _seed(db, username="day_precision_manager")
    adapter = _adapter(db, user)
    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    _validation, batch = adapter.validate(
        date(2026, 8, 1),
        _set_plan(artifact.content, relation_id),
        hmac_key=HMAC_KEY,
    )
    db.commit()
    adapter.apply(batch.batch_id)
    db.commit()
    milestone = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 1,
        )
    )
    assert milestone.date_precision == "day"
    assert milestone.follow_up_status == "pending"
    assert milestone.follow_up_review_required is False
    assert milestone.followed_up_by is None
    assert milestone.followed_up_at is None


def test_adapter_keeps_handled_state_and_marks_review_when_planned_facts_change(db):
    """已处理节点被新批次改日期/金额：保留 handled 与处理人/时间，置 review_required=true。"""
    user, _project_id, relation_id = _seed(db, username="handled_plan_manager")
    adapter = _adapter(db, user)
    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    _validation, batch = adapter.validate(
        date(2026, 8, 1),
        _set_plan(artifact.content, relation_id),
        hmac_key=HMAC_KEY,
    )
    db.commit()
    adapter.apply(batch.batch_id)
    db.commit()
    milestone = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 1,
        )
    )
    assert milestone.version == 1
    followed_up_at = datetime(2026, 8, 10, 9, 30, tzinfo=UTC)
    milestone.follow_up_status = "handled"
    milestone.followed_up_by = user.id
    milestone.followed_up_at = followed_up_at
    milestone.follow_up_note = "合成已跟进备注"
    db.commit()

    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    changed = _set_plan_values(artifact.content, relation_id, date(2026, 10, 15), 30000)
    _validation, second_batch = adapter.validate(
        date(2026, 8, 1),
        changed,
        hmac_key=HMAC_KEY,
    )
    db.commit()
    adapter.apply(second_batch.batch_id)
    db.commit()

    db.expire_all()
    milestone = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 1,
        )
    )
    assert milestone.follow_up_status == "handled"
    assert milestone.followed_up_by == user.id
    assert milestone.followed_up_at == followed_up_at
    assert milestone.follow_up_note == "合成已跟进备注"
    assert milestone.follow_up_review_required is True
    assert milestone.planned_date == date(2026, 10, 15)
    assert milestone.planned_amount == 30000
    assert milestone.version == 2


def test_scope_change_after_validation_fails_before_any_write(db):
    user, _project_id, relation_id = _seed(db, username="scope_conflict_manager")
    adapter = _adapter(db, user)
    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    _validation, batch = adapter.validate(
        date(2026, 8, 1),
        _set_plan(artifact.content, relation_id),
        hmac_key=HMAC_KEY,
    )
    db.commit()
    assignment = db.scalar(
        select(MaintenanceProjectUserAssignment).where(
            MaintenanceProjectUserAssignment.user_id == user.id,
            MaintenanceProjectUserAssignment.archived_at.is_(None),
        )
    )
    assignment.archived_at = datetime.now(UTC)
    assignment.archived_by = "synthetic-admin"
    assignment.archive_reason = "合成范围冲突"
    assignment.version += 1
    db.commit()

    with pytest.raises(ManagerWorkbookConflict, match="范围"):
        adapter.apply(batch.batch_id)
    db.rollback()
    assert db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id
        )
    ) is None


def test_expired_preview_revalidates_into_same_idempotency_ledger(db):
    user, project_id, relation_id = _seed(
        db,
        username="expired_preview_manager",
    )
    adapter = _adapter(db, user)
    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    edited = _set_plan(artifact.content, relation_id)
    first_validation, batch = adapter.validate(
        date(2026, 8, 1),
        edited,
        hmac_key=HMAC_KEY,
    )
    batch.expires_at = datetime.now(UTC) - timedelta(seconds=1)
    db.commit()

    second_validation, refreshed = adapter.validate(
        date(2026, 8, 1),
        edited,
        hmac_key=HMAC_KEY,
    )
    db.commit()

    assert second_validation.validation_id == first_validation.validation_id
    assert refreshed.batch_id == batch.batch_id
    assert refreshed.status == "valid"
    assert refreshed.expires_at > datetime.now(UTC)
    assert len(second_validation.milestone_changes) == 1

    refreshed.expires_at = datetime.now(UTC) - timedelta(seconds=1)
    project = db.get(MaintenanceProject, project_id)
    project.version += 1
    db.commit()

    with pytest.raises(ManagerWorkbookV3Error, match="数据版本已变化"):
        adapter.validate(
            date(2026, 8, 1),
            edited,
            hmac_key=HMAC_KEY,
        )
    db.rollback()
    unchanged_batch = db.get(MaintenanceManagerUploadBatch, batch.batch_id)
    assert unchanged_batch.expires_at <= datetime.now(UTC)


def test_applied_status_becomes_stale_when_current_scope_changes(db):
    user, _project_id, _relation_id = _seed(
        db,
        username="status_scope_manager",
    )
    adapter = _adapter(db, user)
    artifact, _snapshot = adapter.export(date(2026, 8, 1), hmac_key=HMAC_KEY)
    _validation, batch = adapter.validate(
        date(2026, 8, 1),
        artifact.content,
        hmac_key=HMAC_KEY,
    )
    db.commit()
    adapter.apply(batch.batch_id)
    db.commit()
    assert adapter.status(date(2026, 8, 1))["latest_batch"]["status"] == "applied"

    project = MaintenanceProject(
        project_id="status-scope-new-project",
        project_code="PM-STATUS-SCOPE-NEW",
        display_name="新增本人范围项目",
        lifecycle_status="ongoing",
    )
    db.add(project)
    db.flush()
    db.add_all([
        MaintenanceProjectUserAssignment(
            assignment_id="status-scope-new-assignment",
            project_id=project.project_id,
            responsibility_type="primary_manager",
            user_id=user.id,
            assigned_at=datetime.now(UTC),
            assigned_by="synthetic-admin",
            assignment_reason="新增范围验证",
        ),
        MaintenanceProjectContract(
            project_contract_id="status-scope-new-relation",
            project_id=project.project_id,
            contract_id="status-scope-new-contract",
            contract_no="XS-STATUS-SCOPE-NEW",
            contract_amount=1000,
            contract_status="active",
            status_mapping_state="mapped",
            status_mapping_version="synthetic-v1",
            included_in_total=True,
            effective_from=date(2026, 1, 1),
            source="synthetic-test",
        ),
    ])
    db.commit()

    current = adapter.status(date(2026, 8, 1))

    assert current["latest_batch"]["status"] == "stale_scope"
    assert current["latest_batch"]["scope_matches_current"] is False
    assert current["project_count"] == 2

    # 2026-08-25：「项目经理月度更新」任务退役——pending 过滤恒为空
    directory = project_operations(
        db,
        as_of=date(2026, 8, 9),
        user_ctx=_ctx(user),
        owner_scope="me",
        task_type="项目经理月度更新",
        task_status="pending",
    )
    assert directory["total"] == 0
