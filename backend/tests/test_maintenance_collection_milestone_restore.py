"""实名管理员受控恢复已作废回款计划。"""

from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.config import get_settings
from app.main import app
from app.models.maintenance_manager import MaintenanceCollectionMilestone
from app.models.maintenance_project import MaintenanceProjectContract
from app.models.maintenance_project_operations import (
    MaintenanceProjectOperationAudit,
    MaintenanceProjectWorkbookState,
)
from app.services import maintenance_collection_milestone_restore as restore
from app.services import maintenance_project_master_workbook as master
from app.services import maintenance_project_operations as operations
from tests.boss_board_helpers import client_for, make_project


_BASE = "/api/maintenance/projects/stable"


def _seed(db):
    project = make_project(db, code=f"恢复测试-{uuid.uuid4().hex[:8]}")
    contract = MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project.project_id,
        contract_id=f"C-{uuid.uuid4().hex[:8]}",
        contract_no="XSDD-RESTORE-001",
        amount_inc_tax=Decimal("151250.00"),
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="test",
        effective_from=date(2025, 1, 1),
        source="test",
        version=1,
    )
    db.add(contract)
    db.flush()
    milestones = []
    for sequence, planned_date in ((1, date(2025, 7, 1)), (2, date(2025, 10, 1))):
        milestone = MaintenanceCollectionMilestone(
            milestone_id=str(uuid.uuid4()),
            project_id=project.project_id,
            project_contract_id=contract.project_contract_id,
            sequence=sequence,
            planned_date=planned_date,
            planned_amount=Decimal("75625.00"),
            completeness_state="complete",
            source="project_master_v2",
            date_precision="month",
            is_active=False,
            version=2,
        )
        db.add(milestone)
        milestones.append(milestone)
    operations.get_or_create_workbook_state(db, project_id=project.project_id)
    db.commit()
    return project, contract, milestones


def _specs(contract, milestones):
    return [
        restore.MilestoneRestoreSpec(
            entity_id=milestone.milestone_id,
            expected_version=2,
            contract_no=contract.contract_no,
            sequence=milestone.sequence,
            planned_date=milestone.planned_date,
            planned_amount=milestone.planned_amount,
            date_precision=milestone.date_precision,
        )
        for milestone in milestones
    ]


def _payload(contract, milestones):
    return {
        "reason": "按用户确认的 Excel 恢复误作废回款计划",
        "items": [
            {
                "entity_id": milestone.milestone_id,
                "expected_version": 2,
                "contract_no": contract.contract_no,
                "sequence": milestone.sequence,
                "planned_date": milestone.planned_date.isoformat(),
                "planned_amount": str(milestone.planned_amount),
                "date_precision": milestone.date_precision,
            }
            for milestone in milestones
        ],
    }


def _admin_client(db, username="milestone-restore-admin"):
    return client_for(db, username=username, role="admin", overrides={
        "page_maintenance": True,
        "data_profit": True,
        "action_maintenance_expense_collection_upload": True,
    })


def test_restore_is_atomic_audited_revisioned_and_idempotent(db):
    project, contract, milestones = _seed(db)
    before_state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    before_revision = before_state.revision

    result = restore.restore_collection_milestones(
        db,
        project_id=project.project_id,
        specs=_specs(contract, milestones),
        reason="按 Excel 恢复两条测试计划",
        operated_by="restore-admin",
    )
    db.commit()
    assert result["restored_count"] == 2
    assert result["idempotent_replay_count"] == 0
    for milestone in milestones:
        db.refresh(milestone)
        assert milestone.is_active is True
        assert milestone.version == 3
        assert milestone.planned_amount == Decimal("75625.00")
    audits = list(db.scalars(select(MaintenanceProjectOperationAudit).where(
        MaintenanceProjectOperationAudit.project_id == project.project_id,
        MaintenanceProjectOperationAudit.action == "RESTORE",
    )))
    assert len(audits) == 2
    assert {audit.entity_id for audit in audits} == {
        milestone.milestone_id for milestone in milestones
    }
    assert all(audit.before_json["is_active"] is False for audit in audits)
    assert all(audit.after_json["is_active"] is True for audit in audits)
    db.refresh(before_state)
    assert before_state.revision == before_revision + 1

    replay = restore.restore_collection_milestones(
        db,
        project_id=project.project_id,
        specs=_specs(contract, milestones),
        reason="按 Excel 恢复两条测试计划",
        operated_by="restore-admin",
    )
    db.commit()
    assert replay["restored_count"] == 0
    assert replay["idempotent_replay_count"] == 2
    assert db.scalar(select(func.count()).select_from(
        MaintenanceProjectOperationAudit).where(
            MaintenanceProjectOperationAudit.project_id == project.project_id,
            MaintenanceProjectOperationAudit.action == "RESTORE",
        )) == 2
    db.refresh(before_state)
    assert before_state.revision == before_revision + 1

    exported = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_PLAN,),
    )
    workbook = load_workbook(io.BytesIO(exported), data_only=True)
    worksheet = workbook[master.V2_SHEET_PLAN]
    entity_col = next(cell.column for cell in worksheet[1] if cell.value == "实体ID")
    version_col = next(cell.column for cell in worksheet[1] if cell.value == "基础版本")
    exported_rows = {
        worksheet.cell(row, entity_col).value: worksheet.cell(row, version_col).value
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, entity_col).value
    }
    assert exported_rows == {
        milestone.milestone_id: 3 for milestone in milestones
    }


def test_api_rolls_back_entire_batch_when_one_excel_fact_mismatches(db):
    project, contract, milestones = _seed(db)
    payload = _payload(contract, milestones)
    payload["items"][1]["planned_amount"] = "75626.00"
    response = _admin_client(db).post(
        f"{_BASE}/{project.project_id}/collection-milestones/restore",
        json=payload,
    )
    assert response.status_code == 409, response.text
    assert response.json()["detail"]["code"] == "milestone_restore_conflict"
    db.expire_all()
    rows = list(db.scalars(select(MaintenanceCollectionMilestone).where(
        MaintenanceCollectionMilestone.project_id == project.project_id,
    )))
    assert all(row.is_active is False and row.version == 2 for row in rows)
    assert db.scalar(select(func.count()).select_from(
        MaintenanceProjectOperationAudit).where(
            MaintenanceProjectOperationAudit.project_id == project.project_id,
    )) == 0


def test_active_row_without_restore_audit_is_not_accepted_as_replay(db):
    project, contract, milestones = _seed(db)
    milestones[0].is_active = True
    milestones[0].version = 3
    db.commit()

    with pytest.raises(restore.MilestoneRestoreConflict, match="没有可验证的恢复审计"):
        restore.restore_collection_milestones(
            db,
            project_id=project.project_id,
            specs=_specs(contract, [milestones[0]]),
            reason="尝试重放一条没有恢复回执的计划",
            operated_by="restore-admin",
        )
    db.rollback()
    assert db.scalar(select(func.count()).select_from(
        MaintenanceProjectOperationAudit).where(
            MaintenanceProjectOperationAudit.project_id == project.project_id,
            MaintenanceProjectOperationAudit.action == "RESTORE",
        )) == 0


def test_api_requires_real_admin_not_uploader_or_shared_password(db):
    project, contract, milestones = _seed(db)
    payload = _payload(contract, milestones)
    uploader = client_for(db, username="milestone-restore-uploader", role="readonly",
                          overrides={
                              "page_maintenance": True,
                              "data_profit": True,
                              "action_maintenance_expense_collection_upload": True,
                          })
    denied = uploader.post(
        f"{_BASE}/{project.project_id}/collection-milestones/restore",
        json=payload,
    )
    assert denied.status_code == 403

    shared = TestClient(app)
    login = shared.post(
        "/api/auth/login",
        json={"username": "admin", "password": get_settings().admin_password},
    )
    assert login.status_code == 200, login.text
    shared.headers["Authorization"] = f"Bearer {login.json()['token']}"
    shared_denied = shared.post(
        f"{_BASE}/{project.project_id}/collection-milestones/restore",
        json=payload,
    )
    assert shared_denied.status_code == 403
    db.expire_all()
    assert all(
        row.is_active is False
        for row in db.scalars(select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_id == project.project_id,
        ))
    )
