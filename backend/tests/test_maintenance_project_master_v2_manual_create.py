import io
import uuid
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import select

from app.models.dimensions import DimPart
from app.models.maintenance import FProjectExpense
from app.models.maintenance_project import MaintenanceProject, MaintenanceProjectContract
from app.models.maintenance_project_operations import MaintenanceSiteIssueLine
from app.services import maintenance_project_master_workbook as master


def _project(db):
    project = MaintenanceProject(
        project_id=str(uuid.uuid4()),
        project_code="MANUAL-CREATE",
        display_name="手工新增兼容测试",
        lifecycle_status="ongoing",
    )
    part = DimPart(pn_std="MANUAL-PN-001", description="手工领用测试备件")
    db.add_all([project, part])
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project.project_id,
        contract_id="MANUAL-CONTRACT",
        contract_no="XSDD-MANUAL-001",
        amount_inc_tax=Decimal("10000.00"),
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="v1",
        effective_from=date(2026, 1, 1),
        source="ledger",
        version=1,
    ))
    db.commit()
    return project, part


def _manual_workbook(db, project_id: str) -> bytes:
    content = master.build_project_master_v2(db, project_id=project_id)
    wb = load_workbook(io.BytesIO(content))
    expense = wb[master.V2_SHEET_EXPENSE]
    # V2.1：第 1 列为「操作」，留空 = 空白实体手工新增（CREATE 语义）
    expense.append([
        "", "BXD-20260818-0001", 1, "2026-08-18 10:00:00", "测试人员",
        "维保费用", "交通费", "人工新增报销", "XSDD-MANUAL-001",
        "手工新增兼容测试", "测试销售", "已生效", 100, "ex", 100,
        113, "人工新增", "",
    ])
    site = wb[master.V2_SHEET_SITE]
    site.append([
        "CKD-20260818-0001", "2026-08-18", "MANUAL-PN-001", "SN-001",
        1, "是", "人工新增领用", 1, "", "", "",
    ])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def test_v2_blank_entity_ids_create_expense_and_site_rows(db):
    project, part = _project(db)
    content = _manual_workbook(db, project.project_id)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=content,
    )
    assert plan.summary["expense_creates"] == 1
    assert plan.summary["site_creates"] == 1

    result = master.apply_project_master_v2(
        db,
        plan,
        operated_by="manual-test",
        import_batch_id=str(uuid.uuid4()),
    )
    assert result["expense_creates"] == 1
    assert result["site_creates"] == 1

    expense = db.scalar(select(FProjectExpense))
    assert expense is not None
    assert expense.raw_line_id
    assert expense.bxd_no == "BXD-20260818-0001"
    assert expense.line_no == 1
    assert expense.amount_ex_tax == Decimal("100.00")
    assert expense.amount_inc_tax == Decimal("113.00")
    assert expense.data_status == "已结束"

    site_line = db.scalar(select(MaintenanceSiteIssueLine))
    assert site_line is not None
    assert site_line.issue_line_id
    assert site_line.part_id == part.id
    assert site_line.pn == "MANUAL-PN-001"

    exported = load_workbook(io.BytesIO(
        master.build_project_master_v2(db, project_id=project.project_id)
    ), data_only=True)
    assert exported[master.V2_SHEET_EXPENSE]["R2"].value  # V2.1：实体ID 移至第 18 列（操作列插入） == expense.raw_line_id
    assert exported[master.V2_SHEET_SITE]["K2"].value == site_line.issue_line_id


def test_v2_reupload_of_blank_ids_is_idempotent(db):
    project, _part = _project(db)
    content = _manual_workbook(db, project.project_id)

    for _ in range(2):
        plan = master.validate_project_master_v2(
            db, project_id=project.project_id, data=content,
        )
        master.apply_project_master_v2(
            db,
            plan,
            operated_by="manual-test",
            import_batch_id=str(uuid.uuid4()),
        )

    assert len(db.scalars(select(FProjectExpense)).all()) == 1
    assert len(db.scalars(select(MaintenanceSiteIssueLine)).all()) == 1


def test_v2_single_sheet_export_only_validates_that_sheet(db):
    project, _part = _project(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_EXPENSE,),
    )
    wb = load_workbook(io.BytesIO(content), data_only=True)
    assert wb.sheetnames == [
        master.V2_SHEET_EXPENSE,
        master.V2_SHEET_DICTIONARY,
        master.V2_SHEET_USAGE,
        master.V2_SHEET_META,
    ]
    assert wb[master.V2_SHEET_META]["B6"].value == master.V2_SHEET_EXPENSE
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=content,
    )
    assert plan.sheets == (master.V2_SHEET_EXPENSE,)


def test_v2_manual_site_create_prices_line_immediately(db):
    """2026-08-24：工作簿建领用行后立即取价（此前不调取价，新行成本为空，
    要等全局回填才恢复——8-24 两项目新增 38 行无价的根因）。"""
    from app.etl import loader
    from tests import factories as f
    from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder
    from app.models.maintenance_source_assignment import (
        MaintenanceSourceOrderAssignment,
    )
    from app.models.system import SysImportBatch

    project, part = _project(db)
    # 种一条挂靠本项目的 WBDD 需求价格（未税 300）作为最强证据层
    batch = SysImportBatch(
        filename="manual-dem.xlsx", file_type="maintenance",
        file_hash=uuid.uuid4().hex.ljust(64, "0"), status="success")
    db.add(batch)
    db.flush()
    rid = f"WBDD-MANUAL-{uuid.uuid4().hex[:6]}"
    loader.load(
        db,
        f.maintenance_result(
            {rid: f.maintenance_head(rid, order_no=rid, on=date(2026, 8, 10),
                                     project=project.display_name)},
            [f.maintenance_line(rid, f"{rid}-L1", part.pn_std, qty="1")],
        ),
        batch.id, date(2026, 8, 10), mode="upsert",
    )
    order = db.scalar(select(FMaintenanceOrder).where(
        FMaintenanceOrder.raw_order_id == rid))
    db.add(MaintenanceSourceOrderAssignment(
        assignment_id=str(uuid.uuid4()), source_order_id=rid,
        project_id=project.project_id, is_active=True, version=1,
        created_by="tester"))
    dem_line = db.scalar(select(FMaintenanceLine).where(
        FMaintenanceLine.order_id == order.id))
    dem_line.cost_source = "direct"
    dem_line.cost_amount_ex_tax = Decimal("300")
    dem_line.cost_amount_inc_tax = (Decimal("300") * Decimal("1.13")).quantize(
        Decimal("0.01"))
    db.commit()

    content = _manual_workbook(db, project.project_id)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=content)
    result = master.apply_project_master_v2(
        db, plan, operated_by="manual-test", import_batch_id=str(uuid.uuid4()))
    assert result["site_creates"] == 1

    site_line = db.scalar(select(MaintenanceSiteIssueLine))
    assert site_line is not None
    assert site_line.cost_source == "maint_demand"
    assert site_line.unit_cost_ex_tax == Decimal("300.00")
    assert site_line.cost_amount_inc_tax == Decimal("339.00")
