"""按时间范围批量导出项目工作簿 ZIP 的公共接口契约。"""
import asyncio
import csv
import io
import struct
from datetime import date
from decimal import Decimal
from zipfile import ZIP_STORED, ZipFile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import text
from starlette.requests import ClientDisconnect

from app import permissions
from app.api import maintenance as maintenance_api
from app.auth import hash_password
from app.db import SessionLocal
from app.main import app
from app.models.dimensions import DimPart
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceContractWorkbookState,
)
from app.models.system import SysImportBatch, SysUser
from app.security import UserContext
from app.services import maintenance_workbook_export


def _admin_client(db) -> TestClient:
    db.add(SysUser(
        username="maintenance_workbook_export_admin",
        role="admin",
        password_hash=hash_password("pw123456"),
        is_active=True,
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_workbook_export_admin", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _custom_client(
    db,
    *,
    username: str,
    role: str = "readonly",
    overrides: dict[str, bool],
) -> TestClient:
    base = permissions.effective(role, None)
    db.add(SysUser(
        username=username,
        role=role,
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code=role,
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": username, "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _batch(db) -> SysImportBatch:
    batch = SysImportBatch(
        filename="maintenance-workbook-export.xlsx",
        file_type="maintenance",
        file_hash="maintenance-workbook-export",
    )
    db.add(batch)
    db.flush()
    return batch


def _order(
    db,
    batch: SysImportBatch,
    *,
    raw_id: str,
    contract: str | None,
    on: date | None = date(2026, 7, 15),
    status: str = "已生效",
) -> FMaintenanceOrder:
    order = FMaintenanceOrder(
        raw_order_id=raw_id,
        order_no=f"WBDD-{raw_id}",
        order_date=on,
        linked_sales_order_no=contract,
        project_raw=f"项目-{raw_id}",
        project_std=f"项目-{raw_id}",
        data_status=status,
        import_batch_id=batch.id,
    )
    db.add(order)
    return order


def _manifest(archive: ZipFile) -> list[dict[str, str]]:
    raw = archive.read("导出清单.csv").decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(raw)))


def _only_workbook(archive: ZipFile):
    member = next(
        name for name in archive.namelist()
        if name.startswith("项目工作簿/")
    )
    return load_workbook(io.BytesIO(archive.read(member)), data_only=False)


def _extra_field_ids(extra: bytes) -> list[int]:
    field_ids: list[int] = []
    offset = 0
    while offset < len(extra):
        assert len(extra) - offset >= 4
        field_id, size = struct.unpack_from("<HH", extra, offset)
        offset += 4
        assert size <= len(extra) - offset
        field_ids.append(field_id)
        offset += size
    return field_ids


def _assert_zip_uses_only_zip32_member_headers(payload: bytes) -> None:
    local_header = struct.Struct("<IHHHHHIIIHH")
    with ZipFile(io.BytesIO(payload)) as archive:
        assert archive.testzip() is None
        for info in archive.infolist():
            fields = local_header.unpack_from(payload, info.header_offset)
            assert fields[0] == 0x04034B50
            assert fields[1] < 45
            assert fields[7] != 0xFFFFFFFF
            assert fields[8] != 0xFFFFFFFF
            name_length, extra_length = fields[9], fields[10]
            extra_offset = info.header_offset + local_header.size + name_length
            local_extra = payload[extra_offset:extra_offset + extra_length]
            assert 0x0001 not in _extra_field_ids(local_extra)
            assert 0x0001 not in _extra_field_ids(info.extra)
            assert archive.read(info.filename)


def test_bulk_export_returns_one_zip_with_one_contract_workbook(db):
    batch = _batch(db)
    _order(db, batch, raw_id="ONE", contract="XSDD-ONE")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "application/zip"
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["content-disposition"] == (
        'attachment; filename="maintenance_project_workbooks_all.zip"; '
        "filename*=UTF-8''maintenance_project_workbooks_all.zip"
    )
    assert response.content[:2] == b"PK"
    with ZipFile(io.BytesIO(response.content)) as archive:
        assert archive.testzip() is None
        members = archive.namelist()
        assert all(info.compress_type == ZIP_STORED for info in archive.infolist())
        workbooks = [name for name in members if name.startswith("项目工作簿/")]
        assert len(workbooks) == 1
        assert workbooks[0].endswith(".xlsx")
        assert archive.read(workbooks[0])[:2] == b"PK"
        assert _manifest(archive) == [{
            "记录类型": "已生成",
            "合同号": "XSDD-ONE",
            "文件名": workbooks[0],
            "命中订单数": "1",
            "命中最早日期": "2026-07-15",
            "命中最晚日期": "2026-07-15",
            "跳过维保单号": "",
            "跳过原始订单ID": "",
            "跳过制单日期": "",
            "说明": "",
        }]


def test_bulk_export_uses_zip32_headers_accepted_by_browser_validator(db):
    batch = _batch(db)
    _order(db, batch, raw_id="ZIP32", contract="XSDD-ZIP32")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 200, response.text
    _assert_zip_uses_only_zip32_member_headers(response.content)


@pytest.mark.parametrize(
    "params",
    [
        {"date_from": "2026-07-01"},
        {"date_to": "2026-07-31"},
        {"date_from": "2026-08-01", "date_to": "2026-07-31"},
    ],
)
def test_bulk_export_rejects_incomplete_or_reversed_date_range(db, params):
    response = _admin_client(db).get(
        "/api/maintenance/export-workbooks",
        params=params,
    )

    assert response.status_code == 422


def test_bulk_export_rejects_range_without_active_orders(db):
    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 422
    assert response.json()["detail"] == "所选范围内没有已生效维保订单"


def test_bulk_export_rejects_when_all_selected_orders_lack_contract(db):
    batch = _batch(db)
    _order(db, batch, raw_id="UNLINKED", contract=None)
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 422
    assert response.json()["detail"] == "所选范围内的已生效维保订单均未关联合同"


def test_workbook_selection_excludes_contracts_before_maintenance_cost_start_date(db):
    batch = _batch(db)
    _order(
        db,
        batch,
        raw_id="PRE-COST-START",
        contract="XSDD-PRE-COST-START",
        on=date(2023, 12, 31),
    )
    db.commit()
    client = _admin_client(db)

    single = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-PRE-COST-START"},
    )
    bulk = client.get("/api/maintenance/export-workbooks")

    assert single.status_code == 422
    assert "起算日" in single.json()["detail"]
    assert bulk.status_code == 422
    assert "起算日" in bulk.json()["detail"]


def test_contract_workbook_audits_expense_inc_ex_and_evidence_status(db):
    batch = _batch(db)
    _order(db, batch, raw_id="EXPENSE-AUDIT", contract="XSDD-EXPENSE-AUDIT")
    db.add_all([
        FProjectExpense(
            raw_line_id="EXPENSE-AUDIT-ROW",
            bxd_no="BXD-EXPENSE-AUDIT",
            line_no=1,
            data_status="已结束",
            expense_date=date(2026, 7, 16),
            linked_sales_order_no="XSDD-EXPENSE-AUDIT",
            amount=Decimal("100"),
            amount_ex_tax=Decimal("100"),
            amount_inc_tax=Decimal("113"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XSDD-EXPENSE-AUDIT",
            revision=1,
            expense_complete_through=date.max,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-EXPENSE-AUDIT"},
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    try:
        summary = {}
        for row in workbook["项目预算"].iter_rows(min_row=3, values_only=True):
            if row[0]:
                summary[str(row[0])] = row[1]
            if len(row) > 2 and row[2]:
                summary[str(row[2])] = row[3]
        assert summary["报销费用（含税）"] == 113
        assert summary["报销费用（未税）"] == 100
        assert summary["费用证据状态"] == "完整"

        details = workbook["报销明细"]
        headers = [cell.value for cell in details[2]]
        row = dict(zip(headers, [cell.value for cell in details[3]], strict=True))
        assert row["报销金额（含税）"] == 113
        assert row["报销金额（未税）"] == 100
        assert row["费用证据状态"] == "双口径已确认"
    finally:
        workbook.close()


def test_contract_workbook_keeps_unready_expense_tax_amounts_blank(db):
    batch = _batch(db)
    _order(db, batch, raw_id="EXPENSE-UNREADY", contract="XSDD-EXPENSE-UNREADY")
    db.add(FProjectExpense(
        raw_line_id="EXPENSE-UNREADY-ROW",
        bxd_no="BXD-EXPENSE-UNREADY",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 16),
        linked_sales_order_no="XSDD-EXPENSE-UNREADY",
        amount=Decimal("100"),
        amount_ex_tax=Decimal("100"),
        amount_inc_tax=Decimal("113"),
        tax_basis="ex",
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-EXPENSE-UNREADY"},
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    try:
        summary = {}
        for row in workbook["项目预算"].iter_rows(min_row=3, values_only=True):
            if row[0]:
                summary[str(row[0])] = row[1]
            if len(row) > 2 and row[2]:
                summary[str(row[2])] = row[3]
        assert summary["报销费用（含税）"] == "—"
        assert summary["报销费用（未税）"] == "—"
        assert summary["费用证据状态"] == "未就绪（无记录不等于0）"

        details = workbook["报销明细"]
        headers = [cell.value for cell in details[2]]
        row = dict(zip(headers, [cell.value for cell in details[3]], strict=True))
        assert row["报销金额（含税）"] is None
        assert row["报销金额（未税）"] is None
        assert row["费用证据状态"] == "费用快照未就绪"
    finally:
        workbook.close()


def test_contract_workbook_excludes_expenses_before_maintenance_cost_start_date(db):
    batch = _batch(db)
    _order(db, batch, raw_id="POST-START", contract="XSDD-EXPENSE-START")
    db.add_all([
        FProjectExpense(
            raw_line_id="EXPENSE-BEFORE-START",
            bxd_no="BXD-BEFORE-START",
            line_no=1,
            data_status="已结束",
            expense_date=date(2023, 12, 31),
            linked_sales_order_no="XSDD-EXPENSE-START",
            amount=Decimal("100"),
            amount_ex_tax=Decimal("100"),
            amount_inc_tax=Decimal("113"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XSDD-EXPENSE-START",
            revision=1,
            expense_complete_through=date.max,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-EXPENSE-START"},
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    try:
        summary = {}
        for row in workbook["项目预算"].iter_rows(min_row=3, values_only=True):
            if row[0]:
                summary[str(row[0])] = row[1]
            if len(row) > 2 and row[2]:
                summary[str(row[2])] = row[3]
        assert summary["报销费用（含税）"] == 0
        assert summary["报销费用（未税）"] == 0
        assert workbook["报销明细"].max_row == 2
    finally:
        workbook.close()


def test_bulk_export_lists_each_unlinked_order_without_skipping_linked_contracts(db):
    batch = _batch(db)
    _order(db, batch, raw_id="LINKED", contract="XSDD-LINKED", on=date(2026, 7, 14))
    _order(db, batch, raw_id="UNLINKED", contract=None, on=date(2026, 7, 16))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 200, response.text
    with ZipFile(io.BytesIO(response.content)) as archive:
        rows = _manifest(archive)
        assert len([row for row in rows if row["记录类型"] == "已生成"]) == 1
        assert [row for row in rows if row["记录类型"] == "已跳过"] == [{
            "记录类型": "已跳过",
            "合同号": "",
            "文件名": "",
            "命中订单数": "",
            "命中最早日期": "",
            "命中最晚日期": "",
            "跳过维保单号": "WBDD-UNLINKED",
            "跳过原始订单ID": "UNLINKED",
            "跳过制单日期": "2026-07-16",
            "说明": "未关联合同",
        }]


def test_bulk_export_uses_inclusive_dates_and_all_excludes_null_dates_and_inactive(db):
    batch = _batch(db)
    _order(db, batch, raw_id="A-FIRST", contract="XSDD-A", on=date(2026, 7, 1))
    _order(db, batch, raw_id="A-LAST", contract="XSDD-A", on=date(2026, 7, 31))
    _order(db, batch, raw_id="OUTSIDE", contract="XSDD-OUTSIDE", on=date(2026, 8, 1))
    _order(db, batch, raw_id="NULL-DATE", contract="XSDD-NULL", on=None)
    _order(
        db,
        batch,
        raw_id="INACTIVE",
        contract="XSDD-INACTIVE",
        on=date(2026, 7, 15),
        status="已取消",
    )
    db.commit()
    client = _admin_client(db)

    ranged = client.get(
        "/api/maintenance/export-workbooks",
        params={"date_from": "2026-07-01", "date_to": "2026-07-31"},
    )
    all_dates = client.get("/api/maintenance/export-workbooks")

    assert ranged.status_code == 200, ranged.text
    with ZipFile(io.BytesIO(ranged.content)) as archive:
        generated = [row for row in _manifest(archive) if row["记录类型"] == "已生成"]
        assert [(row["合同号"], row["命中订单数"]) for row in generated] == [
            ("XSDD-A", "2"),
        ]
        assert generated[0]["命中最早日期"] == "2026-07-01"
        assert generated[0]["命中最晚日期"] == "2026-07-31"

    assert all_dates.status_code == 200, all_dates.text
    with ZipFile(io.BytesIO(all_dates.content)) as archive:
        generated = [row for row in _manifest(archive) if row["记录类型"] == "已生成"]
        assert [row["合同号"] for row in generated] == [
            "XSDD-A", "XSDD-OUTSIDE",
        ]


def test_date_range_limits_each_generated_workbook_to_the_closed_interval(db):
    batch = _batch(db)
    inside = _order(
        db, batch, raw_id="COMPLETE-IN", contract="XSDD-COMPLETE", on=date(2026, 7, 15),
    )
    outside = _order(
        db, batch, raw_id="COMPLETE-OUT", contract="XSDD-COMPLETE", on=date(2026, 6, 15),
    )
    part = DimPart(pn_std="PART-COMPLETE")
    db.add(part)
    db.flush()
    db.add_all([
        FMaintenanceLine(
            raw_line_id="LINE-COMPLETE-IN",
            order_id=inside.id,
            line_no=1,
            part_id=part.id,
            pn_std="PART-COMPLETE",
            qty=Decimal("1"),
            anomaly_flags=[],
            import_batch_id=batch.id,
        ),
        FMaintenanceLine(
            raw_line_id="LINE-COMPLETE-OUT",
            order_id=outside.id,
            line_no=1,
            part_id=part.id,
            pn_std="PART-COMPLETE",
            qty=Decimal("1"),
            anomaly_flags=[],
            import_batch_id=batch.id,
        ),
    ])
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-COMPLETE-IN",
            data_status="已结束",
            expense_date=date(2026, 7, 20),
            linked_sales_order_no="XSDD-COMPLETE",
            amount=Decimal("10"),
            amount_ex_tax=Decimal("10"),
            amount_inc_tax=Decimal("11.30"),
            import_batch_id=batch.id,
        ),
        FProjectExpense(
            raw_line_id="EXP-COMPLETE-OUT",
            data_status="已结束",
            expense_date=date(2026, 6, 20),
            linked_sales_order_no="XSDD-COMPLETE",
            amount=Decimal("20"),
            amount_ex_tax=Decimal("20"),
            amount_inc_tax=Decimal("22.60"),
            import_batch_id=batch.id,
        ),
    ])
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbooks",
        params={"date_from": "2026-07-01", "date_to": "2026-07-31"},
    )

    assert response.status_code == 200, response.text
    with ZipFile(io.BytesIO(response.content)) as archive:
        workbook = _only_workbook(archive)
    exported_orders = {
        cell.value for cell in workbook["备件明细-氚云"]["A"][1:]
    }
    exported_expense_dates = {
        cell.value
        for cell in workbook["报销明细"]["A"][2:]
        if cell.value is not None and cell.row < workbook["报销明细"].max_row
    }
    assert exported_orders == {"WBDD-COMPLETE-IN"}
    assert exported_expense_dates == {"2026-07-20"}


def test_bulk_export_requires_profit_permission_before_reading_selected_orders(
    db,
    monkeypatch,
):
    batch = _batch(db)
    _order(db, batch, raw_id="PROFIT", contract="XSDD-PROFIT")
    db.commit()
    builder_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbooks_zip",
        lambda *_args, **_kwargs: builder_calls.append("bulk"),
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        lambda *_args, **_kwargs: builder_calls.append("single"),
    )
    client = _custom_client(
        db,
        username="maintenance_workbook_no_profit",
        overrides={
            "page_maintenance": True,
            "data_purchase_cost": True,
            "data_profit": False,
        },
    )

    responses = [
        client.get("/api/maintenance/export-workbooks"),
        client.get(
            "/api/maintenance/export-workbook",
            params={"contract": "XSDD-PROFIT"},
        ),
    ]

    for response in responses:
        assert response.status_code == 403
        assert response.json()["detail"] == "无成本及利润查看权限，不能导出项目成本工作簿"
    assert builder_calls == []


def test_bulk_export_keeps_auth_page_cost_and_scoped_sales_boundaries(db):
    assert TestClient(app).get("/api/maintenance/export-workbooks").status_code == 401
    no_page = _custom_client(
        db,
        username="maintenance_workbook_no_page",
        overrides={},
    )
    assert no_page.get("/api/maintenance/export-workbooks").status_code == 403

    batch = _batch(db)
    _order(db, batch, raw_id="PERMISSIONS", contract="XSDD-PERMISSIONS")
    db.commit()
    no_cost = _custom_client(
        db,
        username="maintenance_workbook_no_cost",
        overrides={
            "page_maintenance": True,
            "data_purchase_cost": False,
            "data_profit": False,
        },
    )
    scoped = _custom_client(
        db,
        username="maintenance_workbook_scoped_sales",
        role="sales",
        overrides={
            "page_maintenance": True,
            "data_purchase_cost": True,
            "data_profit": True,
            "own_customers_only": True,
        },
    )

    no_cost_response = no_cost.get("/api/maintenance/export-workbooks")
    scoped_response = scoped.get("/api/maintenance/export-workbooks")

    assert no_cost_response.status_code == 403
    assert scoped_response.status_code == 403
    assert scoped_response.json()["detail"] == "受限销售账号不能导出项目成本工作簿"


def test_bulk_export_sanitizes_traversal_and_resolves_member_name_collisions(db):
    batch = _batch(db)
    _order(db, batch, raw_id="PATH-A", contract="=../CON")
    _order(db, batch, raw_id="PATH-B", contract="=..\\CON")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 200, response.text
    with ZipFile(io.BytesIO(response.content)) as archive:
        members = [
            name for name in archive.namelist()
            if name.startswith("项目工作簿/")
        ]
        assert len(members) == 2
        assert len(set(members)) == 2
        assert all(
            not name.startswith(("/", "\\"))
            and ".." not in name.split("/")
            and "\\" not in name
            for name in members
        )
        rows = _manifest(archive)
        assert {row["文件名"] for row in rows} == set(members)
        assert all(row["合同号"].startswith("'") for row in rows)


def test_single_and_bulk_workbooks_force_formula_like_contracts_to_text(db):
    batch = _batch(db)
    _order(db, batch, raw_id="FORMULA", contract="=2+2")
    db.commit()
    client = _admin_client(db)

    single = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "=2+2"},
    )
    bulk = client.get("/api/maintenance/export-workbooks")

    assert single.status_code == 200, single.text
    assert bulk.status_code == 200, bulk.text
    single_book = load_workbook(io.BytesIO(single.content), data_only=False)
    with ZipFile(io.BytesIO(bulk.content)) as archive:
        bulk_book = _only_workbook(archive)
    for workbook in (single_book, bulk_book):
        anchor = workbook["报销明细"]["B1"]
        assert anchor.data_type == "s"
        assert anchor.value == "'=2+2"


def test_workbook_renderer_forces_all_dynamic_formula_like_text_to_strings(db):
    batch = _batch(db)
    order = _order(db, batch, raw_id="TEXT", contract="XSDD-TEXT")
    order.order_no = "=ORDER"
    order.project_raw = "+PROJECT"
    part = DimPart(pn_std="PART-TEXT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-TEXT",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="-PN",
        pn_raw="@RAW",
        description="\x01=DESCRIPTION",
        qty=Decimal("1"),
        serial_numbers="+SERIAL",
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.add(FProjectExpense(
        raw_line_id="EXP-TEXT",
        bxd_no="=EXPENSE",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 16),
        person="+PERSON",
        expense_type="-TYPE",
        fee_category="@CATEGORY",
        reason="  =REASON",
        linked_sales_order_no="XSDD-TEXT",
        amount=Decimal("10"),
        amount_ex_tax=Decimal("10"),
        amount_inc_tax=Decimal("11.30"),
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-TEXT"},
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    parts = workbook["备件明细-氚云"]
    expenses = workbook["报销明细"]
    assert [parts.cell(2, column).data_type for column in (1, 4, 10, 11, 16)] == ["s"] * 5
    assert [parts.cell(2, column).value for column in (1, 4, 10, 11, 16)] == [
        "'=ORDER", "'+PROJECT", "'-PN", "'=DESCRIPTION", "'+SERIAL",
    ]
    assert [expenses.cell(3, column).data_type for column in (2, 3, 4, 5, 11)] == ["s"] * 5
    assert [expenses.cell(3, column).value for column in (2, 3, 4, 5, 11)] == [
        "'+PERSON", "'-TYPE", "'@CATEGORY", "'  =REASON", "'=EXPENSE",
    ]


def test_single_and_bulk_reject_text_over_excel_cell_limit_without_truncation(db):
    batch = _batch(db)
    order = _order(db, batch, raw_id="TEXT-LIMIT", contract="XSDD-TEXT-LIMIT")
    part = DimPart(pn_std="PART-TEXT-LIMIT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-TEXT-LIMIT",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="PART-TEXT-LIMIT",
        description="X" * 32768,
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()
    client = _admin_client(db)

    single = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-TEXT-LIMIT"},
    )
    bulk = client.get("/api/maintenance/export-workbooks")

    for response in (single, bulk):
        assert response.status_code == 413
        assert "32767" in response.json()["detail"]


def test_bulk_export_rejects_contract_limit_before_rendering_any_workbook(db, monkeypatch):
    monkeypatch.setattr(maintenance_workbook_export, "MAX_WORKBOOKS", 1, raising=False)
    batch = _batch(db)
    _order(db, batch, raw_id="LIMIT-A", contract="XSDD-LIMIT-A")
    _order(db, batch, raw_id="LIMIT-B", contract="XSDD-LIMIT-B")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert response.json()["detail"] == "命中合同超过批量上限 1 本"


def test_bulk_export_rejects_selected_order_limit_before_contract_selection(db, monkeypatch):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_SELECTED_ORDERS",
        1,
        raising=False,
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "_contract_matches",
        lambda *_args: (_ for _ in ()).throw(
            AssertionError("订单总数超限后不得继续选择合同"),
        ),
    )
    batch = _batch(db)
    _order(db, batch, raw_id="ORDER-LIMIT-A", contract="XSDD-ORDER-LIMIT-A")
    _order(db, batch, raw_id="ORDER-LIMIT-B", contract="XSDD-ORDER-LIMIT-B")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert response.json()["detail"] == "命中维保订单超过批量上限 1 条"


def test_bulk_export_rejects_total_part_line_limit_before_build(db, monkeypatch):
    monkeypatch.setattr(maintenance_workbook_export, "MAX_PART_LINES", 0, raising=False)
    batch = _batch(db)
    order = _order(db, batch, raw_id="PART-LIMIT", contract="XSDD-PART-LIMIT")
    part = DimPart(pn_std="PART-LIMIT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-PART-LIMIT",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="PART-LIMIT",
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert response.json()["detail"] == "项目工作簿备件明细超过批量上限 0 行"


def test_bulk_export_rejects_single_workbook_part_line_limit_before_build(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES_PER_WORKBOOK",
        0,
        raising=False,
    )
    builder_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        lambda *_args: builder_calls.append("built"),
    )
    batch = _batch(db)
    order = _order(db, batch, raw_id="PART-SINGLE-LIMIT", contract="XSDD-PART-SINGLE")
    part = DimPart(pn_std="PART-SINGLE-LIMIT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-PART-SINGLE-LIMIT",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="PART-SINGLE-LIMIT",
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert "单个项目工作簿备件明细超过上限 0 行" in response.json()["detail"]
    assert builder_calls == []


def test_bulk_export_counts_zero_detail_order_placeholder_before_build(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES_PER_WORKBOOK",
        0,
        raising=False,
    )
    builder_calls = []

    def fake_builder(*_args, **_kwargs):
        builder_calls.append("built")
        return io.BytesIO(b"not-a-real-workbook")

    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        fake_builder,
    )
    batch = _batch(db)
    _order(
        db,
        batch,
        raw_id="PART-ZERO-DETAIL-LIMIT",
        contract="XSDD-PART-ZERO-DETAIL",
    )
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert (
        "单个项目工作簿备件明细超过上限 0 行"
        in response.json()["detail"]
    )
    assert builder_calls == []


def test_bulk_export_counts_mixed_detail_and_placeholder_rows_before_build(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES_PER_WORKBOOK",
        1,
        raising=False,
    )
    builder_calls = []

    def fake_builder(*_args, **_kwargs):
        builder_calls.append("built")
        return io.BytesIO(b"not-a-real-workbook")

    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        fake_builder,
    )
    batch = _batch(db)
    order_with_detail = _order(
        db,
        batch,
        raw_id="PART-MIXED-WITH-DETAIL",
        contract="XSDD-PART-MIXED",
    )
    _order(
        db,
        batch,
        raw_id="PART-MIXED-WITHOUT-DETAIL",
        contract="XSDD-PART-MIXED",
    )
    part = DimPart(pn_std="PART-MIXED-LIMIT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-PART-MIXED-LIMIT",
        order_id=order_with_detail.id,
        line_no=1,
        part_id=part.id,
        pn_std="PART-MIXED-LIMIT",
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert (
        "单个项目工作簿备件明细超过上限 1 行"
        in response.json()["detail"]
    )
    assert builder_calls == []


def test_bulk_export_allows_mixed_rendered_part_rows_exactly_at_limits(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES",
        2,
        raising=False,
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES_PER_WORKBOOK",
        2,
        raising=False,
    )
    batch = _batch(db)
    order_with_detail = _order(
        db,
        batch,
        raw_id="PART-MIXED-BOUNDARY-WITH-DETAIL",
        contract="XSDD-PART-MIXED-BOUNDARY",
    )
    _order(
        db,
        batch,
        raw_id="PART-MIXED-BOUNDARY-WITHOUT-DETAIL",
        contract="XSDD-PART-MIXED-BOUNDARY",
    )
    part = DimPart(pn_std="PART-MIXED-BOUNDARY")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-PART-MIXED-BOUNDARY",
        order_id=order_with_detail.id,
        line_no=1,
        part_id=part.id,
        pn_std="PART-MIXED-BOUNDARY",
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 200, response.text
    with ZipFile(io.BytesIO(response.content)) as archive:
        workbook = _only_workbook(archive)
        try:
            assert workbook["备件明细-氚云"].max_row == 3
        finally:
            workbook.close()


def test_bulk_export_rejects_total_expense_line_limit_before_build(db, monkeypatch):
    monkeypatch.setattr(maintenance_workbook_export, "MAX_EXPENSE_LINES", 0)
    batch = _batch(db)
    _order(db, batch, raw_id="EXPENSE-LIMIT", contract="XSDD-EXPENSE-LIMIT")
    db.add(FProjectExpense(
        raw_line_id="EXP-LIMIT",
        bxd_no="BXD-LIMIT",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 16),
        linked_sales_order_no="XSDD-EXPENSE-LIMIT",
        amount=Decimal("10"),
        amount_ex_tax=Decimal("10"),
        amount_inc_tax=Decimal("11.30"),
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert response.json()["detail"] == "项目工作簿报销明细超过批量上限 0 行"


def test_bulk_export_rejects_single_workbook_expense_line_limit_before_build(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_EXPENSE_LINES_PER_WORKBOOK",
        0,
        raising=False,
    )
    builder_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        lambda *_args: builder_calls.append("built"),
    )
    batch = _batch(db)
    _order(db, batch, raw_id="EXP-SINGLE-LIMIT", contract="XSDD-EXP-SINGLE")
    db.add(FProjectExpense(
        raw_line_id="EXP-SINGLE-LIMIT",
        bxd_no="BXD-SINGLE-LIMIT",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 16),
        linked_sales_order_no="XSDD-EXP-SINGLE",
        amount=Decimal("10"),
        amount_ex_tax=Decimal("10"),
        amount_inc_tax=Decimal("11.30"),
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert "单个项目工作簿报销明细超过上限 0 行" in response.json()["detail"]
    assert builder_calls == []


def test_bulk_export_rejects_dynamic_text_budget_before_rendering(db, monkeypatch):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        512,
        raising=False,
    )
    builder_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        lambda *_args: builder_calls.append("built"),
    )
    batch = _batch(db)
    order = _order(db, batch, raw_id="TEXT-BUDGET", contract="XSDD-TEXT-BUDGET")
    part = DimPart(pn_std="PART-TEXT-BUDGET")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-TEXT-BUDGET",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="PART-TEXT-BUDGET",
        description="X" * 600,
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert "单个项目工作簿动态文本超过安全上限" in response.json()["detail"]
    assert builder_calls == []


def test_single_export_rejects_part_line_limit_before_materializing_data(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES_PER_WORKBOOK",
        0,
        raising=False,
    )
    materialize_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: materialize_calls.append("materialized"),
    )
    batch = _batch(db)
    order = _order(db, batch, raw_id="SINGLE-PART-LIMIT", contract="XSDD-SINGLE-PART")
    part = DimPart(pn_std="SINGLE-PART-LIMIT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-SINGLE-PART-LIMIT",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="SINGLE-PART-LIMIT",
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-SINGLE-PART"},
    )

    assert response.status_code == 413
    assert "单个项目工作簿备件明细超过上限 0 行" in response.json()["detail"]
    assert materialize_calls == []


def test_single_export_counts_zero_detail_order_placeholder_before_materializing(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_PART_LINES_PER_WORKBOOK",
        0,
        raising=False,
    )
    materialize_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: materialize_calls.append("materialized"),
    )
    batch = _batch(db)
    _order(
        db,
        batch,
        raw_id="SINGLE-ZERO-DETAIL-LIMIT",
        contract="XSDD-SINGLE-ZERO-DETAIL",
    )
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-SINGLE-ZERO-DETAIL"},
    )

    assert response.status_code == 413
    assert (
        "单个项目工作簿备件明细超过上限 0 行"
        in response.json()["detail"]
    )
    assert materialize_calls == []


def test_single_export_rejects_expense_line_limit_before_materializing_data(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_EXPENSE_LINES_PER_WORKBOOK",
        0,
        raising=False,
    )
    materialize_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: materialize_calls.append("materialized"),
    )
    batch = _batch(db)
    _order(db, batch, raw_id="SINGLE-EXP-LIMIT", contract="XSDD-SINGLE-EXP")
    db.add(FProjectExpense(
        raw_line_id="SINGLE-EXP-LIMIT",
        bxd_no="BXD-SINGLE-EXP-LIMIT",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 16),
        linked_sales_order_no="XSDD-SINGLE-EXP",
        amount=Decimal("10"),
        amount_ex_tax=Decimal("10"),
        amount_inc_tax=Decimal("11.30"),
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-SINGLE-EXP"},
    )

    assert response.status_code == 413
    assert "单个项目工作簿报销明细超过上限 0 行" in response.json()["detail"]
    assert materialize_calls == []


def test_single_export_rejects_dynamic_text_budget_before_materializing_data(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        512,
        raising=False,
    )
    materialize_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: materialize_calls.append("materialized"),
    )
    batch = _batch(db)
    order = _order(db, batch, raw_id="SINGLE-TEXT-LIMIT", contract="XSDD-SINGLE-TEXT")
    part = DimPart(pn_std="SINGLE-TEXT-LIMIT")
    db.add(part)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="LINE-SINGLE-TEXT-LIMIT",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="SINGLE-TEXT-LIMIT",
        description="X" * 600,
        qty=Decimal("1"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-SINGLE-TEXT"},
    )

    assert response.status_code == 413
    assert "单个项目工作簿动态文本超过安全上限" in response.json()["detail"]
    assert materialize_calls == []


def test_single_export_counts_zero_detail_placeholder_dynamic_text_preflight(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        128,
        raising=False,
    )
    materialize_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: materialize_calls.append("materialized"),
    )
    batch = _batch(db)
    order = _order(
        db,
        batch,
        raw_id="R",
        contract="C",
    )
    order.project_raw = "X" * 80
    order.project_std = None
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "C"},
    )

    assert response.status_code == 413
    assert "单个项目工作簿动态文本超过安全上限" in response.json()["detail"]
    assert materialize_calls == []


def test_dynamic_text_budget_counts_generated_date_and_label_text_before_materializing(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        30,
        raising=False,
    )
    materialize_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: materialize_calls.append("materialized"),
    )
    batch = _batch(db)
    part = DimPart(pn_std="P")
    db.add(part)
    db.flush()
    order = FMaintenanceOrder(
        raw_order_id="R",
        order_no="O",
        order_date=date(2026, 7, 15),
        linked_sales_order_no="C",
        data_status="已生效",
        import_batch_id=batch.id,
    )
    db.add(order)
    db.flush()
    db.add(FMaintenanceLine(
        raw_line_id="L",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std="P",
        qty=Decimal("1"),
        cost_source="window",
        confidence="high",
        anomaly_flags=[],
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "C"},
    )

    assert response.status_code == 413
    assert "单个项目工作簿动态文本超过安全上限" in response.json()["detail"]
    assert materialize_calls == []


def test_dynamic_text_budget_counts_rendered_fee_category_prefix_before_materializing(
    db,
    monkeypatch,
):
    # 该最小数据在旧口径下为 50 bytes；55 落在加入 15-byte“费用分类：”前缀前后之间。
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        55,
        raising=False,
    )
    materialize_calls = []
    real_materialize = (
        maintenance_workbook_export.maintenance_cost.contract_workbook_data
    )

    def capture_materialize(*args):
        materialize_calls.append("materialized")
        return real_materialize(*args)

    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        capture_materialize,
    )
    batch = _batch(db)
    _order(db, batch, raw_id="FEE-PREFIX", contract="C")
    db.add(FProjectExpense(
        raw_line_id="EXP-FEE-PREFIX",
        bxd_no="B",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 16),
        fee_category="X",
        linked_sales_order_no="C",
        amount=Decimal("10"),
        amount_ex_tax=Decimal("10"),
        amount_inc_tax=Decimal("11.30"),
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "C"},
    )

    assert response.status_code == 413
    assert "单个项目工作簿动态文本超过安全上限" in response.json()["detail"]
    assert materialize_calls == []


def test_excel_column_limit_uses_max_categories_per_contract_not_global_distinct(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_EXCEL_COLUMNS",
        4,
        raising=False,
    )
    batch = _batch(db)
    for index, contract in enumerate(("XSDD-CATEGORY-A", "XSDD-CATEGORY-B"), 1):
        _order(db, batch, raw_id=f"CATEGORY-{index}", contract=contract)
        db.add(FProjectExpense(
            raw_line_id=f"EXP-CATEGORY-{index}",
            bxd_no=f"BXD-CATEGORY-{index}",
            line_no=1,
            data_status="已结束",
            expense_date=date(2026, 7, 16),
            fee_category=f"仅合同{index}的分类",
            linked_sales_order_no=contract,
            amount=Decimal("10"),
            amount_ex_tax=Decimal("10"),
            amount_inc_tax=Decimal("11.30"),
            import_batch_id=batch.id,
        ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 200, response.text
    with ZipFile(io.BytesIO(response.content)) as archive:
        assert len([
            name for name in archive.namelist()
            if name.startswith("项目工作簿/")
        ]) == 2


def test_excel_column_preflight_counts_normalized_and_part_named_fee_categories(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_EXCEL_COLUMNS",
        4,
        raising=False,
    )
    batch = _batch(db)
    contract = "XSDD-CATEGORY-NORMALIZATION"
    _order(db, batch, raw_id="CATEGORY-NORMALIZATION", contract=contract)
    for index, category in enumerate((None, "", "备件消耗"), 1):
        db.add(FProjectExpense(
            raw_line_id=f"EXP-CATEGORY-NORMALIZATION-{index}",
            bxd_no=f"BXD-CATEGORY-NORMALIZATION-{index}",
            line_no=index,
            data_status="已结束",
            expense_date=date(2026, 7, 16),
            fee_category=category,
            linked_sales_order_no=contract,
            amount=Decimal("10"),
            amount_ex_tax=Decimal("10"),
            amount_inc_tax=Decimal("11.30"),
            import_batch_id=batch.id,
        ))
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert response.json()["detail"] == "项目预算 Sheet 超过 Excel 列数上限"


def test_fee_categories_named_like_cost_tiers_remain_visible_in_monthly_table(db):
    batch = _batch(db)
    contract = "XSDD-COST-TIER-CATEGORY"
    _order(db, batch, raw_id="COST-TIER-CATEGORY", contract=contract)
    for index, (category, amount) in enumerate((
        ("备件实际参考", Decimal("11")),
        ("备件估算参考", Decimal("22")),
    ), 1):
        db.add(FProjectExpense(
            raw_line_id=f"EXP-COST-TIER-CATEGORY-{index}",
            bxd_no=f"BXD-COST-TIER-CATEGORY-{index}",
            line_no=index,
            data_status="已结束",
            expense_date=date(2026, 7, 16),
            fee_category=category,
            linked_sales_order_no=contract,
            amount=amount,
            amount_ex_tax=amount,
            amount_inc_tax=(amount * Decimal("1.13")).quantize(Decimal("0.01")),
            import_batch_id=batch.id,
        ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": contract},
    )

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    try:
        sheet = workbook["项目预算"]
        heading_index = next(
            index
            for index, cells in enumerate(sheet.iter_rows(), 1)
            if cells[0].value == "月份"
        )
        headings = [
            cell.value for cell in sheet[heading_index]
            if cell.value is not None
        ]
        values = [cell.value for cell in sheet[heading_index + 1]][:len(headings)]
        row = dict(zip(headings, values, strict=True))
        assert row["当前已导入报销（非全量）·备件实际参考"] == 11
        assert row["当前已导入报销（非全量）·备件估算参考"] == 22
    finally:
        workbook.close()


@pytest.mark.parametrize("mode", ("single", "bulk"))
def test_workbook_build_holds_shared_source_lock_against_concurrent_import(
    db,
    mode,
):
    batch = _batch(db)
    contract = f"XSDD-SNAPSHOT-{mode}"
    _order(db, batch, raw_id=f"SNAPSHOT-{mode}", contract=contract)
    db.commit()

    resource = (
        maintenance_workbook_export.build_contract_workbook_file(db, contract)
        if mode == "single"
        else maintenance_workbook_export.build_contract_workbooks_zip(db)
    )
    try:
        with SessionLocal() as importer:
            acquired_during_export = importer.scalar(
                text("SELECT pg_try_advisory_xact_lock(:k)"),
                {"k": 0x5350_4152},
            )
            assert acquired_during_export is False

            db.rollback()
            acquired_after_export = importer.scalar(
                text("SELECT pg_try_advisory_xact_lock(:k)"),
                {"k": 0x5350_4152},
            )
            assert acquired_after_export is True
    finally:
        resource.close()
        db.rollback()


def test_bulk_export_rejects_final_zip_size_without_returning_partial_archive(db, monkeypatch):
    monkeypatch.setattr(maintenance_workbook_export, "MAX_ZIP_BYTES", 1, raising=False)
    batch = _batch(db)
    _order(db, batch, raw_id="ZIP-LIMIT", contract="XSDD-ZIP-LIMIT")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 413
    assert response.json()["detail"] == "批量工作簿 ZIP 超过 512 MiB 上限"


def test_zip_budget_rejects_before_writing_beyond_hard_limit():
    output = io.BytesIO()
    limited = maintenance_workbook_export._SizeLimitedFile(output, max_size=3)
    assert limited.write(b"12") == 2

    with pytest.raises(
        maintenance_workbook_export.WorkbookExportRejected,
        match="512 MiB",
    ):
        limited.write(b"34")

    assert output.getvalue() == b"12"


def test_second_bulk_export_is_rejected_before_selection_with_retry_after(db, monkeypatch):
    class BusyLock:
        def acquire(self, blocking=True):
            assert blocking is False
            return False

        def release(self):
            raise AssertionError("未取得锁时不得 release")

    monkeypatch.setattr(
        maintenance_workbook_export,
        "_BULK_EXPORT_LOCK",
        BusyLock(),
        raising=False,
    )

    response = _admin_client(db).get("/api/maintenance/export-workbooks")

    assert response.status_code == 429
    assert response.headers["retry-after"] == "5"
    assert response.json()["detail"] == "已有批量工作簿导出正在执行，请稍后重试"


def test_bulk_export_audits_exact_date_scope(db, monkeypatch):
    output = io.BytesIO(b"zip bytes")
    access_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbooks_zip",
        lambda *_args, **_kwargs: output,
    )
    monkeypatch.setattr(
        maintenance_api,
        "record_access_log",
        lambda *args: access_calls.append(args),
    )

    response = maintenance_api.export_workbooks(
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
        db=db,
        _auth="admin",
        _page=None,
        ctx=UserContext(user_id="admin", role="admin"),
    )
    response._resource.close()

    assert access_calls == [(
        UserContext(user_id="admin", role="admin"),
        "export_workbooks",
        "maintenance",
        {"date_from": "2026-07-01", "date_to": "2026-07-31"},
    )]


def test_bulk_export_releases_database_transaction_before_streaming(db, monkeypatch):
    output = io.BytesIO(b"zip bytes")
    rollback_calls = []
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbooks_zip",
        lambda *_args, **_kwargs: output,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args: None)
    monkeypatch.setattr(db, "rollback", lambda: rollback_calls.append("rollback"))

    response = maintenance_api.export_workbooks(
        date_from=None,
        date_to=None,
        db=db,
        _auth="admin",
        _page=None,
        ctx=UserContext(user_id="admin", role="admin"),
    )

    assert rollback_calls == ["rollback"]
    assert not output.closed
    response._resource.close()


def test_single_builder_closes_output_and_workbook_when_rendering_fails(db, monkeypatch):
    output = io.BytesIO()

    class TrackedWorkbook:
        closed = False

        def close(self):
            self.closed = True

    workbook = TrackedWorkbook()
    monkeypatch.setattr(
        maintenance_workbook_export,
        "SpooledTemporaryFile",
        lambda **_kwargs: output,
    )
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: {},
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "Workbook",
        lambda: workbook,
    )
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_workbook_renderer,
        "render_contract_workbook",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("render failed")),
    )

    with pytest.raises(RuntimeError, match="render failed"):
        maintenance_workbook_export.build_contract_workbook_file(
            db,
            "XSDD-FAIL",
            resource_limits_preflighted=True,
        )

    assert output.closed
    assert workbook.closed


def test_inner_workbook_size_limit_closes_all_streams_and_releases_lock(
    db,
    monkeypatch,
):
    monkeypatch.setattr(
        maintenance_workbook_export,
        "MAX_WORKBOOK_BYTES",
        1,
        raising=False,
    )
    batch = _batch(db)
    _order(db, batch, raw_id="INNER-SIZE", contract="XSDD-INNER-SIZE")
    db.commit()
    outer = io.BytesIO()
    inner = io.BytesIO()
    outputs = iter((outer, inner))

    class TrackedWorkbook:
        closed = False

        def save(self, target):
            target.write(b"12")

        def close(self):
            self.closed = True

    workbook = TrackedWorkbook()
    monkeypatch.setattr(
        maintenance_workbook_export,
        "SpooledTemporaryFile",
        lambda **_kwargs: next(outputs),
    )
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_cost,
        "contract_workbook_data",
        lambda *_args: {},
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "Workbook",
        lambda: workbook,
    )
    monkeypatch.setattr(
        maintenance_workbook_export.maintenance_workbook_renderer,
        "render_contract_workbook",
        lambda *_args, **_kwargs: workbook,
    )

    with pytest.raises(
        maintenance_workbook_export.WorkbookExportRejected,
        match="单本项目工作簿超过 256 MiB 上限",
    ):
        maintenance_workbook_export.build_contract_workbooks_zip(db)

    assert inner.closed
    assert outer.closed
    assert workbook.closed
    assert maintenance_workbook_export._BULK_EXPORT_LOCK.acquire(blocking=False)
    maintenance_workbook_export._BULK_EXPORT_LOCK.release()


def test_bulk_builder_closes_outer_stream_and_releases_lock_on_inner_failure(
    db,
    monkeypatch,
):
    batch = _batch(db)
    _order(db, batch, raw_id="INNER-FAIL", contract="XSDD-INNER-FAIL")
    db.commit()
    output = io.BytesIO()
    monkeypatch.setattr(
        maintenance_workbook_export,
        "SpooledTemporaryFile",
        lambda **_kwargs: output,
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        lambda *_args: (_ for _ in ()).throw(RuntimeError("inner failed")),
    )

    with pytest.raises(RuntimeError, match="inner failed"):
        maintenance_workbook_export.build_contract_workbooks_zip(db)

    assert output.closed
    assert maintenance_workbook_export._BULK_EXPORT_LOCK.acquire(blocking=False)
    maintenance_workbook_export._BULK_EXPORT_LOCK.release()


def test_bulk_export_closes_zip_stream_when_asgi_send_disconnects(db, monkeypatch):
    output = io.BytesIO(b"zip bytes")
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbooks_zip",
        lambda *_args, **_kwargs: output,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args: None)
    response = maintenance_api.export_workbooks(
        date_from=None,
        date_to=None,
        db=db,
        _auth="admin",
        _page=None,
        ctx=UserContext(user_id="admin", role="admin"),
    )

    async def receive():
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(_message):
        raise OSError("client disconnected")

    with pytest.raises(ClientDisconnect):
        asyncio.run(response(
            {
                "type": "http",
                "method": "GET",
                "path": "/api/maintenance/export-workbooks",
                "headers": [],
                "asgi": {"version": "3.0", "spec_version": "2.4"},
            },
            receive,
            send,
        ))

    assert output.closed


def test_bulk_export_streams_zip_in_bounded_binary_chunks(db, monkeypatch):
    chunk_size = 1024 * 1024
    output = io.BytesIO(b"X" * (chunk_size * 2 + 17))
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbooks_zip",
        lambda *_args, **_kwargs: output,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args: None)
    response = maintenance_api.export_workbooks(
        date_from=None,
        date_to=None,
        db=db,
        _auth="admin",
        _page=None,
        ctx=UserContext(user_id="admin", role="admin"),
    )
    sent = []

    async def receive():
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(message):
        sent.append(message)

    asyncio.run(response(
        {
            "type": "http",
            "method": "GET",
            "path": "/api/maintenance/export-workbooks",
            "headers": [],
            "asgi": {"version": "3.0", "spec_version": "2.4"},
        },
        receive,
        send,
    ))

    body_sizes = [
        len(message.get("body", b""))
        for message in sent
        if message["type"] == "http.response.body" and message.get("body")
    ]
    assert body_sizes == [chunk_size, chunk_size, 17]
    assert output.closed


def test_member_name_filters_unicode_controls_and_obeys_utf8_leaf_limit():
    used_names = set()
    first = maintenance_workbook_export._member_name(
        "\u202e\u2066\u0085" + "😀" * 64,
        used_names,
    )
    second = maintenance_workbook_export._member_name(
        "\u202e\u2066\u0085" + "😀" * 63 + "😁",
        used_names,
    )

    for member_name in (first, second):
        leaf = member_name.rsplit("/", 1)[-1]
        assert "\u202e" not in leaf
        assert "\u2066" not in leaf
        assert "\u0085" not in leaf
        assert len(leaf.encode("utf-8")) <= 240
        assert leaf.endswith(".xlsx")
    assert first != second
