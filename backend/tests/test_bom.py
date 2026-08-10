"""整机拆解 / 多格式文件 / 美化报表 / 15天采购价 单测（不依赖 LLM）。"""
import io

import pytest
from openpyxl import load_workbook

from app import permissions, security
from app.agent import tools
from app.auth import hash_password
from app.db import SessionLocal
from app.models.system import SysUser
from app.services import agent_files as af


@pytest.fixture(scope="module")
def db():
    s = SessionLocal()
    yield s
    s.close()


@pytest.fixture()
def ctx():
    return security.UserContext(
        user_id="bom-admin", role="admin",
        permissions=permissions.effective("admin", None), is_authenticated=True,
        authn="sys_user", has_stable_subject=True,
        token_version=0,
    )


def _docx_bytes(lines: list[str]) -> bytes:
    from docx import Document
    doc = Document()
    for ln in lines:
        doc.add_paragraph(ln)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _owner(db, username: str):
    user = db.query(SysUser).filter_by(username=username).one_or_none()
    if user is None:
        db.add(SysUser(
            username=username, role="admin", password_hash=hash_password("pw123456"),
            permissions=permissions.effective("admin", None),
        ))
        db.commit()
    return af.verified_artifact_owner(db, security.UserContext(
        user_id=username,
        role="admin",
        permissions=permissions.effective("admin", None),
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
        token_version=0,
    ))


def test_upload_docx_and_read(db, ctx):
    owner = _owner(db, "bom-admin")
    up = af.save_upload(_docx_bytes(["1台 Dell R750", "2× Xeon Gold 6330", "8× 32GB DDR4"]),
                        "整机配置.docx", owner)
    assert up["ext"] == "docx" and up["file_kind"] == "Word"
    rd = af.read_document(up["file_id"], owner)
    assert "Xeon Gold 6330" in rd["content"] and rd["vision_used"] is False


def test_upload_txt_and_read(db, ctx):
    owner = _owner(db, "bom-admin")
    up = af.save_upload(
        "服务器配置:\nCPU x2 6330\n内存 32G x8".encode(),
        "config.txt",
        owner,
    )
    rd = af.read_document(up["file_id"], owner)
    assert "6330" in rd["content"]


def test_image_degrades_without_vision(db, ctx):
    from PIL import Image
    buf = io.BytesIO()
    Image.new("RGB", (60, 30), "white").save(buf, "PNG")
    owner = _owner(db, "bom-admin")
    up = af.save_upload(buf.getvalue(), "cfg.png", owner)
    rd = af.read_document(up["file_id"], owner)
    assert "未配置视觉模型" in rd["content"]  # 无 VISION_API_KEY 时优雅降级


def test_reject_executable_ext(db, ctx):
    with pytest.raises(af.FileError):
        af.save_upload(b"MZ", "evil.exe", _owner(db, "bom-admin"))


def test_read_document_tool(db, ctx):
    owner = _owner(db, "bom-admin")
    up = af.save_upload(
        _docx_bytes(["RTX 4090 显卡 x2"]),
        "x.docx",
        owner,
    )
    r = tools.dispatch(db, "read_document", {"file_id": up["file_id"]}, ctx)
    assert "RTX 4090" in r["content"]


def test_write_report_styled(db, ctx):
    owner = _owner(db, "bom-admin")
    headers = ["部件", "数量", "采购价", "备注"]
    rows = [
        ["CPU", 2, 3500.0, "ok"],
        ["内存", 8, 800.0, "规格需确认"],
        ["X", 1, None, "库内未找到"],
    ]
    evidence = af._mint_report_provenance(
        owner,
        title="报价单",
        headers=headers,
        rows=rows,
        output_name=None,
        money_cols=[2],
        contained_resources=set(),
        contained_fields=set(),
        required_positive_keys=set(),
    )
    r = af.write_report(
        "报价单", headers, rows, None, owner, money_cols=[2], provenance=evidence
    )
    assert "download_url" in r and r["rows_written"] == 3
    download = af.get_download_info(r["file_id"], owner)
    workbook = load_workbook(io.BytesIO(download.content))
    ws = workbook.active
    # 标题行1、表头行2、数据行3-5；表头靛蓝填充
    assert ws.cell(2, 1).fill.fgColor.rgb.endswith("4F46E5")
    assert ws.cell(4, 1).fill.fgColor.rgb.endswith("FFF3E0")  # 需确认→橙
    assert ws.cell(5, 1).fill.fgColor.rgb.endswith("FDECEA")  # 未找到→红
    assert ws.cell(3, 3).number_format == "#,##0.00"
    workbook.close()


def test_lookup_prices_bulk_has_15day(db, ctx):
    r = tools.dispatch(db, "lookup_prices_bulk", {"queries": ["ST8000NM000A"]}, ctx)
    item = r["results"][0]
    if item["status"] == "ok":  # 依赖真库
        assert "recent_purchase_days" in item and item["recent_purchase_days"] == 15
        assert "recent_purchase_avg" in item
