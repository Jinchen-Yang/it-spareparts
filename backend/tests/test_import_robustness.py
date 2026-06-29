"""导入鲁棒性修复回归测试（审计 2026-06-28 第一波 P1）。

覆盖：
- I-1 空表头单元格不再被误判「重复列名」整文件拒收
- I-2 超大文件被行数上限拦截，不进入 pandas 全量物化（防 OOM）
- I-3 Excel 日期序列号按 1899 纪元正确转换，不再静默变 1970-01-01
- I-4 金额/数量越列限抛 ValueError → 走坏行隔离，不毒化整批
"""
from datetime import date

import openpyxl
import pytest

from app import config
from app.etl import cleaner, mapping, reader, transform


# ---------- I-3 Excel 日期序列号 ----------

def test_parse_date_excel_serial_not_1970():
    """纯数字序列号按 Excel 纪元(1899-12-30)转换，而非被当纳秒塌缩到 1970。"""
    serial = (date(2025, 1, 1) - date(1899, 12, 30)).days   # = 45658
    assert cleaner.parse_date(serial) == date(2025, 1, 1)
    assert cleaner.parse_date(float(serial)) == date(2025, 1, 1)
    # 另一已知值
    assert cleaner.parse_date((date(2026, 6, 28) - date(1899, 12, 30)).days) == date(2026, 6, 28)


def test_parse_date_normal_inputs_unaffected():
    """正常 datetime / 字符串日期 / 空值不受序列号修复影响。"""
    import pandas as pd
    assert cleaner.parse_date(pd.Timestamp("2025-03-15")) == date(2025, 3, 15)
    assert cleaner.parse_date("2025-03-15") == date(2025, 3, 15)
    assert cleaner.parse_date("2025/03/15") == date(2025, 3, 15)
    assert cleaner.parse_date(None) is None
    assert cleaner.parse_date("") is None


def test_parse_date_out_of_range_serial_raises():
    """越界数字（非合理 Excel 序列号）抛 ValueError → 走 bad_date 隔离，而非静默 1970。"""
    with pytest.raises(ValueError):
        cleaner.parse_date(0)
    with pytest.raises(ValueError):
        cleaner.parse_date(-5)
    with pytest.raises(ValueError):
        cleaner.parse_date(10_000_000)


# ---------- I-4 金额/数量越列限隔离 ----------

def test_parse_money_within_limit_ok():
    # Money = Numeric(14,2)：整数位上限 12
    assert cleaner.parse_money("999999999999.99") is not None      # 12 位整数，合法


def test_parse_money_overflow_raises():
    with pytest.raises(ValueError):
        cleaner.parse_money("1000000000000")     # 13 位整数 → 越界
    with pytest.raises(ValueError):
        cleaner.parse_money("-1000000000000")


def test_parse_qty_overflow_raises():
    # Qty = Numeric(14,3)：整数位上限 11
    assert cleaner.parse_qty("99999999999") is not None            # 11 位，合法
    with pytest.raises(ValueError):
        cleaner.parse_qty("100000000000")        # 12 位整数 → 越界


def test_overflow_amount_isolated_as_bad_row(tmp_path):
    """整批里一行金额越界 → 该行进 errors(bad_number)，其余好行照常解析（坏行隔离）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号(必填)", "数据ID(不可修改)", "明细.数据ID(不可修改)",
               "明细.产品名称(必填)", "明细.单价(必填)"])
    ws.append(["CGDD-1", "O1", "L1", "ST8000NM000A", "100"])              # 好行
    ws.append(["CGDD-2", "O2", "L2", "ST8000NM000A", "1000000000000"])    # 金额越界坏行
    p = tmp_path / "purchase_overflow.xlsx"
    wb.save(p)
    df, ft = reader.read_excel(str(p))
    assert ft == mapping.PURCHASE
    res = transform.transform(df, mapping.PURCHASE)
    assert len(res.lines) == 1                                            # 好行照常入库
    assert any(e.error_type == "bad_number" for e in res.errors)          # 坏行隔离


# ---------- I-1 空表头单元格不误判重复列 ----------

def _purchase_xlsx_blank_headers(tmp_path):
    """单表头采购文件，但表头行有 2 个空单元格（合并单元格/字段视图未导满的真实形态）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号(必填)", None, "明细.数据ID(不可修改)",
               "明细.产品名称(必填)", None, "明细.单价(必填)"])   # 2 个空表头
    ws.append(["CGDD-1", "x", "L1", "ST8000NM000A", "y", "100"])
    p = tmp_path / "purchase_blank_headers.xlsx"
    wb.save(p)
    return str(p)


def test_blank_header_cells_not_rejected(tmp_path):
    """≥2 个空表头单元格不再被误判「重复列名」整文件拒收。"""
    df, ft = reader.read_excel(_purchase_xlsx_blank_headers(tmp_path))
    assert ft == mapping.PURCHASE
    assert "明细.产品名称(必填)" in df.columns
    assert df["明细.产品名称(必填)"].iloc[0] == "ST8000NM000A"


def test_real_duplicate_columns_still_rejected(tmp_path):
    """真实重复的非空列名仍要报错（不能因修复把真重复也放过）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号(必填)", "明细.产品名称(必填)", "明细.产品名称(必填)"])
    ws.append(["CGDD-1", "ST8000NM000A", "DUP"])
    p = tmp_path / "purchase_dup.xlsx"
    wb.save(p)
    with pytest.raises(reader.ReaderError):
        reader.read_excel(str(p))


# ---------- I-2 超大文件行数上限 ----------

def test_oversized_rowcount_rejected(tmp_path, monkeypatch):
    """超过行数上限的文件被干净拒绝，不进入 pandas 全量物化。"""
    monkeypatch.setattr(config, "IMPORT_MAX_ROWS", 50)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["采购单号(必填)", "明细.数据ID(不可修改)", "明细.产品名称(必填)"])
    for i in range(60):     # 60 行 > 上限 50
        ws.append([f"CGDD-{i}", f"L{i}", "ST8000NM000A"])
    p = tmp_path / "huge.xlsx"
    wb.save(p)
    with pytest.raises(reader.ReaderError) as ei:
        reader.read_excel(str(p))
    assert "行数" in str(ei.value)


# ---------- 无价格列检测（导入前预检 + 报告留痕） ----------

def _purchase_xlsx(tmp_path, name, header, row):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append(row)
    p = tmp_path / name
    wb.save(p)
    return str(p)


def test_has_price_columns_helper():
    assert mapping.has_price_columns(["明细.单价(必填)", "明细.采购数量(必填)"], mapping.PURCHASE) is True
    # 只有数量/最近采购价(参考价，非成交价)→ 无价格列
    assert mapping.has_price_columns(["明细.采购数量(必填)", "明细.最近采购价"], mapping.PURCHASE) is False
    assert mapping.has_price_columns(["库存数量"], mapping.INVENTORY) is True   # 库存不涉及
    assert mapping.has_price_columns(["乱"], None) is True                      # 未识别不拦


def test_peek_columns_missing_price(tmp_path):
    p = _purchase_xlsx(
        tmp_path, "no_price.xlsx",
        ["采购单号(必填)", "数据ID(不可修改)", "明细.数据ID(不可修改)",
         "明细.产品名称(必填)", "明细.采购数量(必填)"],
        ["CGDD-1", "O1", "L1", "ST8000NM000A", "2"])
    cols, ft = reader.peek_columns(p)
    assert ft == mapping.PURCHASE
    assert mapping.has_price_columns(cols, ft) is False


def test_report_flags_missing_price(db, tmp_path):
    from app.etl import pipeline
    no_price = _purchase_xlsx(
        tmp_path, "np.xlsx",
        ["采购单号(必填)", "数据ID(不可修改)", "明细.数据ID(不可修改)",
         "明细.产品名称(必填)", "明细.采购数量(必填)"],
        ["CGDD-1", "O1", "L1", "ST8000NM000A", "2"])
    b1 = pipeline.run_import(db, no_price, "np.xlsx")
    db.commit()
    assert b1.report_json["missing_price_columns"] is True

    with_price = _purchase_xlsx(
        tmp_path, "wp.xlsx",
        ["采购单号(必填)", "数据ID(不可修改)", "明细.数据ID(不可修改)",
         "明细.产品名称(必填)", "明细.采购数量(必填)", "明细.单价(必填)"],
        ["CGDD-2", "O2", "L2", "ST8000NM000A", "2", "100"])
    b2 = pipeline.run_import(db, with_price, "wp.xlsx")
    db.commit()
    assert b2.report_json["missing_price_columns"] is False
