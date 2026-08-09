"""Public API contract for the own-scope manager workbook v3 (#206)."""

from __future__ import annotations

import io
from datetime import UTC, date, datetime

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from sqlalchemy import select

from app import auth
from app.api import maintenance_manager_workbooks
from app.auth import hash_password
from app.models.maintenance_manager import MaintenanceCollectionMilestone
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
    MaintenanceProjectUserAssignment,
)
from app.models.system import SysUser
from app.services.maintenance_manager_workbook_v3 import (
    METADATA_SHEET,
    OVERVIEW_SHEET,
    OVERVIEW_TABLE,
    PLAN_SHEET,
    PLAN_TABLE,
)


def _client(db, *, username: str) -> tuple[TestClient, SysUser]:
    user = SysUser(
        username=username,
        role="admin",
        display_name="合成项目经理",
        password_hash=hash_password("synthetic-password-123"),
    )
    db.add(user)
    db.commit()
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_manager_workbooks.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": username, "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client, user


def _project(db, *, suffix: str, manager: SysUser | None) -> tuple[str, str]:
    project = MaintenanceProject(
        project_id=f"pm-api-project-{suffix}",
        project_code=f"PM-API-{suffix}",
        display_name=f"合成项目 {suffix}",
        lifecycle_status="ongoing",
    )
    db.add(project)
    db.flush()
    if manager is not None:
        db.add(
            MaintenanceProjectUserAssignment(
                assignment_id=f"pm-api-assignment-{suffix}",
                project_id=project.project_id,
                responsibility_type="primary_manager",
                user_id=manager.id,
                assigned_at=datetime.now(UTC),
                assigned_by="synthetic-admin",
                assignment_reason="合成 API 负责人映射",
            )
        )
    contract = MaintenanceProjectContract(
        project_contract_id=f"pm-api-pc-{suffix}",
        project_id=project.project_id,
        contract_id=f"pm-api-contract-{suffix}",
        contract_no=f"XS-PM-API-{suffix}",
        contract_amount=100000,
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=date(2026, 1, 1),
        source="synthetic-test",
    )
    db.add(contract)
    db.commit()
    return project.project_id, contract.project_contract_id


def _edit_plan(
    content: bytes,
    relation_id: str,
    *,
    sequence: int = 1,
    planned_date=...,
    planned_amount=...,
) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        sheet = book[PLAN_SHEET]
        table = sheet.tables[PLAN_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("项目合同关系ID") + 1).value == relation_id
            and sheet.cell(row, headers.index("计划期次") + 1).value == sequence
        )
        if planned_date is not ...:
            sheet.cell(target, headers.index("计划回款日期") + 1, planned_date)
        if planned_amount is not ...:
            sheet.cell(target, headers.index("计划回款金额（含税）") + 1, planned_amount)
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def test_export_validate_apply_uses_only_explicit_own_scope(db):
    client, manager = _client(db, username="pm_api_manager")
    owned_project_id, relation_id = _project(db, suffix="owned", manager=manager)
    _other_project_id, _other_relation = _project(db, suffix="other", manager=None)

    download = client.get(
        "/api/maintenance/project-manager/workbooks/v3",
        params={"report_month": "2026-08"},
    )
    assert download.status_code == 200, download.text
    assert download.headers["cache-control"] == "no-store"
    book = load_workbook(io.BytesIO(download.content), read_only=False, data_only=False)
    try:
        overview = book[OVERVIEW_SHEET]
        table = overview.tables[OVERVIEW_TABLE]
        _min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
        assert max_row - min_row == 1
        assert overview.cell(min_row + 1, 1).value == owned_project_id
    finally:
        book.close()

    edited = _edit_plan(
        download.content,
        relation_id,
        planned_date=date(2026, 9, 20),
        planned_amount=18000,
    )
    validation = client.post(
        "/api/maintenance/project-manager/workbooks/v3/validate",
        params={"report_month": "2026-08"},
        files={
            "file": (
                "synthetic-manager-v3.xlsx",
                edited,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validation.status_code == 200, validation.text
    preview = validation.json()
    assert preview["can_apply"] is True
    assert preview["changes"] == {
        "service_periods": 0,
        "planned_collection_milestones": 1,
        "acceptance_due_dates": 0,
        "total": 1,
    }
    assert preview["items"] == [{
        "kind": "planned_collection_milestone",
        "project_id": owned_project_id,
        "project_code": "PM-API-owned",
        "project_name": "合成项目 owned",
        "project_contract_id": relation_id,
        "contract_no": "XS-PM-API-owned",
        "sequence": 1,
        "before": {
            "planned_date": None,
            "planned_amount": None,
            "completeness_state": None,
        },
        "after": {
            "planned_date": "2026-09-20",
            "planned_amount": "18000.00",
            "completeness_state": "complete",
        },
    }]

    mismatch = client.post(
        "/api/maintenance/project-manager/workbooks/v3/apply",
        json={
            "validation_token": preview["validation_token"],
            "data_version": "wrong-version",
        },
    )
    assert mismatch.status_code == 409
    assert db.scalar(select(MaintenanceCollectionMilestone)) is None

    applied = client.post(
        "/api/maintenance/project-manager/workbooks/v3/apply",
        json={
            "validation_token": preview["validation_token"],
            "data_version": preview["data_version"],
        },
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["changed_rows"] == 1
    milestone = db.scalar(select(MaintenanceCollectionMilestone))
    assert milestone is not None
    assert milestone.project_id == owned_project_id
    assert milestone.planned_amount == 18000

    repeated = client.post(
        "/api/maintenance/project-manager/workbooks/v3/apply",
        json={
            "validation_token": preview["validation_token"],
            "data_version": preview["data_version"],
        },
    )
    assert repeated.status_code == 200
    assert repeated.json() == applied.json()

    revalidated = client.post(
        "/api/maintenance/project-manager/workbooks/v3/validate",
        params={"report_month": "2026-08"},
        files={
            "file": (
                "synthetic-manager-v3.xlsx",
                edited,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert revalidated.status_code == 200, revalidated.text
    replay_preview = revalidated.json()
    assert replay_preview["validation_token"] == preview["validation_token"]
    assert replay_preview["status"] == "applied"
    assert replay_preview["already_applied"] is True
    assert replay_preview["can_apply"] is False

    reapplied = client.post(
        "/api/maintenance/project-manager/workbooks/v3/apply",
        json={
            "validation_token": replay_preview["validation_token"],
            "data_version": replay_preview["data_version"],
        },
    )
    assert reapplied.status_code == 200
    assert reapplied.json() == applied.json()


def test_status_fails_closed_without_contract_amount_permission(db):
    user = SysUser(
        username="pm_api_restricted",
        role="purchaser",
        display_name="合成受限项目经理",
        password_hash=hash_password("synthetic-password-123"),
        permissions={
            "page_maintenance": True,
            "data_purchase_cost": True,
            "data_profit": False,
        },
    )
    db.add(user)
    db.commit()
    _project(db, suffix="restricted", manager=user)
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_manager_workbooks.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": user.username, "password": "synthetic-password-123"},
    )
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"

    response = client.get(
        "/api/maintenance/project-manager/workbooks/v3/status",
        params={"report_month": "2026-08"},
    )

    assert response.status_code == 403
    assert "全部合同额" in response.json()["detail"]


def test_apply_requires_dedicated_manager_workbook_action(db):
    user = SysUser(
        username="pm_api_apply_denied",
        role="purchaser",
        display_name="合成只读项目经理",
        password_hash=hash_password("synthetic-password-123"),
        permissions={
            "page_maintenance": True,
            "data_purchase_cost": True,
            "data_profit": True,
            "action_maintenance_manager_workbook_apply": False,
        },
    )
    db.add(user)
    db.commit()
    _project(db, suffix="apply-denied", manager=user)
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_manager_workbooks.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": user.username, "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"

    download = client.get(
        "/api/maintenance/project-manager/workbooks/v3",
        params={"report_month": "2026-08"},
    )
    assert download.status_code == 200, download.text
    validation = client.post(
        "/api/maintenance/project-manager/workbooks/v3/validate",
        params={"report_month": "2026-08"},
        files={
            "file": (
                "synthetic-manager-v3.xlsx",
                download.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert validation.status_code == 200, validation.text

    denied = client.post(
        "/api/maintenance/project-manager/workbooks/v3/apply",
        json={
            "validation_token": validation.json()["validation_token"],
            "data_version": validation.json()["data_version"],
        },
    )

    assert denied.status_code == 403
    assert denied.json()["detail"] == "无此操作权限"


def test_status_and_export_fail_closed_without_active_project_assignment(db):
    client, _manager = _client(db, username="pm_api_unassigned")
    _project(db, suffix="unassigned", manager=None)

    status_response = client.get(
        "/api/maintenance/project-manager/workbooks/v3/status",
        params={"report_month": "2026-08"},
    )
    export_response = client.get(
        "/api/maintenance/project-manager/workbooks/v3",
        params={"report_month": "2026-08"},
    )

    assert status_response.status_code == 403
    assert "未分配任何有效维保项目" in status_response.json()["detail"]
    assert export_response.status_code == 403
    assert "未分配任何有效维保项目" in export_response.json()["detail"]


def test_api_rejects_signed_metadata_tampering_before_creating_batch(db):
    client, manager = _client(db, username="pm_api_metadata_tamper")
    _project(db, suffix="metadata-tamper", manager=manager)
    download = client.get(
        "/api/maintenance/project-manager/workbooks/v3",
        params={"report_month": "2026-08"},
    )
    assert download.status_code == 200, download.text
    book = load_workbook(io.BytesIO(download.content), data_only=False)
    try:
        book[METADATA_SHEET].cell(2, 2, "different-owner")
        output = io.BytesIO()
        book.save(output)
    finally:
        book.close()

    validation = client.post(
        "/api/maintenance/project-manager/workbooks/v3/validate",
        params={"report_month": "2026-08"},
        files={
            "file": (
                "metadata-tampered.xlsx",
                output.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert validation.status_code == 422, validation.text
    detail = validation.json()["detail"]
    assert detail["issues"][0]["code"] == "metadata_tampered"


def test_api_preserves_blank_existing_node_and_reports_partial_new_node(db):
    client, manager = _client(db, username="pm_api_partial")
    _project_id, relation_id = _project(db, suffix="partial", manager=manager)
    existing = MaintenanceCollectionMilestone(
        milestone_id="pm-api-existing-node",
        project_id=_project_id,
        project_contract_id=relation_id,
        sequence=1,
        planned_date=date(2026, 9, 1),
        planned_amount=12000,
        completeness_state="complete",
        source="direct_api",
        source_batch_id=None,
    )
    db.add(existing)
    db.commit()
    download = client.get(
        "/api/maintenance/project-manager/workbooks/v3",
        params={"report_month": "2026-08"},
    )
    assert download.status_code == 200, download.text
    blanked = _edit_plan(
        download.content,
        relation_id,
        sequence=1,
        planned_date=None,
        planned_amount=None,
    )
    partial = _edit_plan(
        blanked,
        relation_id,
        sequence=2,
        planned_date=date(2026, 10, 1),
        planned_amount=None,
    )

    validation = client.post(
        "/api/maintenance/project-manager/workbooks/v3/validate",
        params={"report_month": "2026-08"},
        files={"file": ("partial.xlsx", partial, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert validation.status_code == 200, validation.text
    preview = validation.json()
    assert preview["can_apply"] is True
    assert preview["changes"]["planned_collection_milestones"] == 1
    assert [warning["code"] for warning in preview["warnings"]] == [
        "partial_plan_node"
    ]
    applied = client.post(
        "/api/maintenance/project-manager/workbooks/v3/apply",
        json={
            "validation_token": preview["validation_token"],
            "data_version": preview["data_version"],
        },
    )
    assert applied.status_code == 200, applied.text
    preserved = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 1,
        )
    )
    assert preserved.planned_date == date(2026, 9, 1)
    assert preserved.planned_amount == 12000
    assert preserved.version == 1
    partial_row = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id == relation_id,
            MaintenanceCollectionMilestone.sequence == 2,
        )
    )
    assert partial_row.completeness_state == "date_only"


def test_api_rejects_editing_financially_confirmed_actual_column(db):
    client, manager = _client(db, username="pm_api_actual_readonly")
    _project_id, _relation_id = _project(db, suffix="actual-readonly", manager=manager)
    download = client.get(
        "/api/maintenance/project-manager/workbooks/v3",
        params={"report_month": "2026-08"},
    )
    book = load_workbook(io.BytesIO(download.content), data_only=False)
    try:
        sheet = book[OVERVIEW_SHEET]
        table = sheet.tables[OVERVIEW_TABLE]
        min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        sheet.cell(min_row + 1, headers.index("财务确认实收（只读）") + 1, 99999)
        output = io.BytesIO()
        book.save(output)
    finally:
        book.close()

    validation = client.post(
        "/api/maintenance/project-manager/workbooks/v3/validate",
        params={"report_month": "2026-08"},
        files={"file": ("tampered.xlsx", output.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert validation.status_code == 200, validation.text
    preview = validation.json()
    assert preview["can_apply"] is False
    assert [error["code"] for error in preview["errors"]] == [
        "readonly_actual_changed"
    ]
