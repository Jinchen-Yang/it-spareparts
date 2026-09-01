"""Public contracts for project-bound atomic replenishment (#260)."""

from concurrent.futures import ThreadPoolExecutor
from decimal import Decimal
from io import BytesIO
from threading import Barrier, Event
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.exc import DBAPIError

from app import permissions
from app.auth import hash_password
from app.config import get_settings
from app.db import engine
from app.main import app
from app.models.dimensions import DimPart
from app.models.inventory import PartPool, PartPoolMember, PartPoolPricePolicy
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectUserAssignment,
)
from app.models.replenishment import (
    ReplenishmentApplication,
    ReplenishmentApplicationLine,
    ReplenishmentApplicationVersion,
    ReplenishmentAuditEvent,
    ReplenishmentCartDraft,
)
from app.models.system import SysUser
from app.security import UserContext
from app.services import (
    maintenance_project_assignments,
    pool_price_analysis,
    replenishment,
    replenishment_cart,
    replenishment_screening,
)


_PASSWORD = "synthetic-replenishment-atomic-password"


def _admin_client(db, *, username: str) -> TestClient:
    template = permissions.effective("admin", None)
    template["page_replenishment_beta"] = False
    db.add(
        SysUser(
            username=username,
            password_hash=hash_password(_PASSWORD),
            role="admin",
            display_name="补库原子提交管理员",
            salesperson_name="销售甲",
            is_active=True,
            template_code="admin",
            template_version=1,
            template_perms=template,
            perm_overrides={"page_replenishment_beta": True},
        )
    )
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login", json={"username": username, "password": _PASSWORD}
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def _sales_client(
    db, *, username: str, salesperson_name: str | None
) -> TestClient:
    template = permissions.effective("sales", None)
    # a9e2f7c4d1b8 upgrades the production sales template/account snapshots to
    # the unified maintenance row scope; code templates intentionally stay
    # frozen at the older fallback shape.
    template["own_maintenance_projects_only"] = True
    template["page_replenishment_beta"] = False
    db.add(
        SysUser(
            username=username,
            password_hash=hash_password(_PASSWORD),
            role="sales",
            display_name="补库原子提交销售",
            salesperson_name=salesperson_name,
            is_active=True,
            template_code="sales",
            template_version=1,
            template_perms=template,
            perm_overrides={
                "page_replenishment_beta": True,
                "action_replenishment_create": True,
                "data_pool_price_governance": True,
            },
        )
    )
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login", json={"username": username, "password": _PASSWORD}
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def _explicit_replenishment_client(
    db,
    *,
    username: str,
    role: str,
    salesperson_name: str | None = None,
) -> tuple[SysUser, TestClient]:
    template = permissions.effective(role, None)
    template["page_replenishment_beta"] = False
    user = SysUser(
        username=username,
        password_hash=hash_password(_PASSWORD),
        role=role,
        display_name=f"补库显式授权-{role}",
        salesperson_name=salesperson_name,
        is_active=True,
        template_code=role,
        template_version=1,
        template_perms=template,
        perm_overrides={
            "page_replenishment_beta": True,
            "action_replenishment_create": True,
            "data_pool_price_governance": True,
        },
    )
    db.add(user)
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login", json={"username": username, "password": _PASSWORD}
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return user, client


def _assign_manager(db, *, project: MaintenanceProject, user: SysUser) -> None:
    db.add(
        MaintenanceProjectUserAssignment(
            assignment_id=str(uuid4()),
            project_id=project.project_id,
            responsibility_type="primary_manager",
            user_id=user.id,
            source_manager_text=user.display_name,
            version=1,
            assigned_by="replenishment-tdd",
            assignment_reason="补库项目范围测试",
        )
    )
    db.commit()


def _project(db, *, salesperson: str = "销售甲") -> MaintenanceProject:
    project = MaintenanceProject(
        project_id=str(uuid4()),
        project_code=f"REPL-{uuid4().hex[:8].upper()}",
        display_name="补库原子提交合成项目",
        salesperson=salesperson,
        lifecycle_status="ongoing",
        is_active=True,
    )
    db.add(project)
    db.commit()
    return project


def _get(client: TestClient, path: str):
    settings = get_settings()
    original = settings.replenishment_beta_enabled
    try:
        settings.replenishment_beta_enabled = True
        return client.get(path)
    finally:
        settings.replenishment_beta_enabled = original


def _post(client: TestClient, payload: dict):
    settings = get_settings()
    original = settings.replenishment_beta_enabled
    try:
        settings.replenishment_beta_enabled = True
        return client.post("/api/replenishment-beta/applications", json=payload)
    finally:
        settings.replenishment_beta_enabled = original


def _beta_request(client: TestClient, method: str, path: str, **kwargs):
    settings = get_settings()
    original = settings.replenishment_beta_enabled
    try:
        settings.replenishment_beta_enabled = True
        return client.request(method, path, **kwargs)
    finally:
        settings.replenishment_beta_enabled = original


def _post_with_auto_review(client: TestClient, payload: dict):
    settings = get_settings()
    original = settings.replenishment_auto_review_enabled
    try:
        settings.replenishment_auto_review_enabled = True
        return _post(client, payload)
    finally:
        settings.replenishment_auto_review_enabled = original


def _archive_manager(
    db, *, project: MaintenanceProject, user: SysUser
) -> None:
    assignment = db.scalar(
        select(MaintenanceProjectUserAssignment).where(
            MaintenanceProjectUserAssignment.project_id == project.project_id,
            MaintenanceProjectUserAssignment.user_id == user.id,
            MaintenanceProjectUserAssignment.archived_at.is_(None),
        )
    )
    assert assignment is not None
    archived = maintenance_project_assignments.archive_primary_manager(
        db,
        assignment_id=assignment.assignment_id,
        version=assignment.version,
        reason="验证撤权后的最小写权限",
        operated_by="replenishment-tdd",
    )
    assert archived is not None
    db.commit()


def test_named_admin_atomically_submits_one_project_cart(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-ATOMIC-PN-001", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_atomic_admin")

    response = _post(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": "仅用于合成公开接口测试",
            "lines": [
                {"part_id": part.id, "quantity": 2, "special_note": None}
            ],
        },
    )

    assert response.status_code == 201, response.text
    payload = response.json()
    assert payload["project"]["project_id"] == project.project_id
    assert payload["status"] == "submitted"
    assert payload["workflow_mode"] == "system_screening"
    assert payload["stage"] == "screening_complete"
    assert payload["versions"][0]["status"] == "submitted"
    assert payload["versions"][0]["lines"][0]["quantity"] == 2
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationLine)) == 1
    audit = db.scalar(select(ReplenishmentAuditEvent))
    assert audit is not None
    assert audit.after_json["project_id"] == project.project_id


def test_atomic_submission_reuses_one_price_fact_snapshot(db, monkeypatch):
    project = _project(db)
    part = DimPart(pn_std="REPL-ONE-FACT-SNAPSHOT", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_one_fact_admin")
    calls = []
    snapshots = [
        {
            part.id: {
                "purchase": {
                    "weighted_avg": 25.0,
                    "total_qty": 2.0,
                    "order_count": 1,
                    "line_count": 1,
                    "latest_date": "2026-08-17",
                },
                "sales": None,
            }
        },
        {part.id: {"purchase": None, "sales": None}},
    ]

    def changing_price_facts(*_args, **_kwargs):
        calls.append(1)
        return snapshots[min(len(calls) - 1, 1)]

    monkeypatch.setattr(
        pool_price_analysis, "aggregate_part_price_facts", changing_price_facts
    )
    response = _post(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )

    assert response.status_code == 201, response.text
    line = response.json()["versions"][0]["lines"][0]
    activity = next(
        check
        for check in line["screening"]["checks"]
        if check["key"] == "recent_activity"
    )
    assert len(calls) == 1
    assert line["purchase"]["order_count"] == 1
    assert activity["detail"]["purchase_samples"] == 1


def test_capabilities_exposes_system_screening_without_review_action(db):
    client = _admin_client(db, username="replenishment_capabilities_admin")

    response = _get(client, "/api/replenishment-beta/capabilities")

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["can_review"] is False
    assert payload["workflow_mode"] == "system_screening"
    assert payload["stage"] == "screening_complete"
    assert "人工审批" not in payload["data_contract"]


def test_admin_project_picker_lists_all_and_only_active_projects(db):
    active = _project(db)
    db.add(
        MaintenanceProject(
            project_id=str(uuid4()),
            project_code="REPL-INACTIVE",
            display_name="停用项目",
            salesperson="销售乙",
            lifecycle_status="completed",
            is_active=False,
        )
    )
    db.commit()
    client = _admin_client(db, username="replenishment_project_admin")

    response = _get(client, "/api/replenishment-beta/projects")

    assert response.status_code == 200, response.text
    assert response.json() == {
        "items": [
            {
                "project_id": active.project_id,
                "project_code": active.project_code,
                "display_name": active.display_name,
            }
        ]
    }


def test_maintenance_manager_project_picker_uses_assigned_scope(db):
    assigned = _project(db, salesperson="其他销售")
    _project(db, salesperson="无关销售")
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_assigned_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=assigned, user=user)

    response = _get(client, "/api/replenishment-beta/projects")

    assert response.status_code == 200, response.text
    assert [item["project_id"] for item in response.json()["items"]] == [
        assigned.project_id
    ]


def test_maintenance_manager_can_replace_draft_only_in_assigned_scope(db):
    assigned = _project(db, salesperson="其他销售")
    unassigned = _project(db, salesperson="无关销售")
    part = DimPart(pn_std="REPL-MANAGER-DRAFT-PN", status="active")
    db.add(part)
    db.commit()
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_draft_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=assigned, user=user)
    payload = {
        "expected_version": None,
        "request_note": "负责人云端草稿",
        "lines": [{"part_id": part.id, "quantity": 2, "special_note": None}],
    }

    allowed = _beta_request(
        client,
        "PUT",
        f"/api/replenishment-beta/cart-drafts/{assigned.project_id}",
        json=payload,
    )
    denied = _beta_request(
        client,
        "PUT",
        f"/api/replenishment-beta/cart-drafts/{unassigned.project_id}",
        json=payload,
    )
    missing = _beta_request(
        client,
        "PUT",
        f"/api/replenishment-beta/cart-drafts/{uuid4()}",
        json=payload,
    )

    assert allowed.status_code == 200, allowed.text
    assert allowed.json()["draft"]["project_id"] == assigned.project_id
    assert denied.status_code == missing.status_code == 404
    assert denied.json() == missing.json()


def test_out_of_scope_write_is_rejected_before_project_row_lock(db, monkeypatch):
    target = _project(db, salesperson="无关销售")
    user, _client = _explicit_replenishment_client(
        db,
        username="replenishment_unassigned_lock_guard",
        role="maintenance_manager",
    )
    user_ctx = UserContext(
        user_id=user.username,
        role=user.role,
        salesperson_name=user.salesperson_name,
        permissions=permissions.effective_for_user(user),
        is_authenticated=True,
    )
    project_locks: list[str] = []
    original_scalar = db.scalar

    def tracked_scalar(statement, *args, **kwargs):
        sql = str(statement)
        if "FROM maintenance_project" in sql and "FOR UPDATE" in sql:
            project_locks.append(sql)
        return original_scalar(statement, *args, **kwargs)

    monkeypatch.setattr(db, "scalar", tracked_scalar)

    try:
        replenishment._authorized_project(
            db,
            project_id=target.project_id,
            username=user.username,
            user_ctx=user_ctx,
        )
    except replenishment.ReplenishmentError as exc:
        assert exc.code == "project_unavailable"
        assert exc.status_code == 404
    else:  # pragma: no cover - failure branch is the assertion itself
        raise AssertionError("越权项目不应通过补库写范围校验")

    assert project_locks == []


def test_authenticated_username_cannot_expand_scope_with_forged_context_role(db):
    target = _project(db, salesperson="无关销售")
    user, _client = _explicit_replenishment_client(
        db,
        username="replenishment_forged_role_guard",
        role="maintenance_manager",
    )

    try:
        replenishment._authorized_project(
            db,
            project_id=target.project_id,
            username=user.username,
            user_ctx=UserContext(
                user_id=user.username,
                role="admin",
                is_authenticated=True,
            ),
        )
    except replenishment.ReplenishmentError as exc:
        assert exc.code == "project_unavailable"
        assert exc.status_code == 404
    else:  # pragma: no cover - assertion branch
        raise AssertionError("调用方伪造的 admin context 不得覆盖数据库实名账号范围")


@pytest.mark.parametrize(
    "revoked_key",
    [
        "page_replenishment_beta",
        "data_pool_price_governance",
        "action_replenishment_create",
    ],
)
def test_stale_request_context_cannot_write_after_permission_revoke(
    db, revoked_key
):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std=f"REPL-REVOKE-{revoked_key}", status="active")
    db.add(part)
    db.commit()
    user, _client = _explicit_replenishment_client(
        db,
        username=f"replenishment_revoke_{revoked_key}",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=user)
    stale_ctx = UserContext(
        user_id=user.username,
        role=user.role,
        salesperson_name=user.salesperson_name,
        permissions=permissions.runtime_safe(permissions.effective_for_user(user)),
        is_authenticated=True,
    )
    user.perm_overrides = {**(user.perm_overrides or {}), revoked_key: False}
    user.token_version = (user.token_version or 0) + 1
    db.commit()

    with pytest.raises(replenishment.ReplenishmentError) as exc:
        replenishment.submit_application_atomic(
            db,
            username=user.username,
            user_ctx=stale_ctx,
            client_request_id=str(uuid4()),
            project_id=project.project_id,
            lines=[
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        )

    assert exc.value.code == "permission_revoked"
    assert exc.value.status_code == 403
    assert db.scalar(
        select(func.count()).select_from(ReplenishmentApplication)
    ) == 0


def test_boss_full_scope_can_draft_and_submit_any_active_project(db):
    target = _project(db, salesperson="其他销售")
    other = _project(db, salesperson="无关销售")
    inactive = MaintenanceProject(
        project_id=str(uuid4()),
        project_code="REPL-BOSS-INACTIVE",
        display_name="Boss 不可选停用项目",
        salesperson="其他销售",
        lifecycle_status="completed",
        is_active=False,
    )
    part = DimPart(pn_std="REPL-BOSS-FULL-SCOPE-PN", status="active")
    db.add_all([inactive, part])
    db.commit()
    _user, client = _explicit_replenishment_client(
        db,
        username="replenishment_full_scope_boss",
        role="boss",
    )

    listing = _get(client, "/api/replenishment-beta/projects")
    draft = _beta_request(
        client,
        "PUT",
        f"/api/replenishment-beta/cart-drafts/{target.project_id}",
        json={
            "expected_version": None,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    submitted = _beta_request(
        client,
        "POST",
        f"/api/replenishment-beta/cart-drafts/{target.project_id}/submit",
        json={
            "expected_version": draft.json()["draft"]["version"],
            "client_request_id": draft.json()["draft"]["client_request_id"],
        },
    )

    assert listing.status_code == 200, listing.text
    assert {item["project_id"] for item in listing.json()["items"]} == {
        target.project_id,
        other.project_id,
    }
    assert inactive.project_id not in {
        item["project_id"] for item in listing.json()["items"]
    }
    assert draft.status_code == 200, draft.text
    assert submitted.status_code == 201, submitted.text
    assert submitted.json()["project"]["project_id"] == target.project_id


def test_maintenance_manager_submit_uses_assignment_salesperson_union(db):
    assigned = _project(db, salesperson="其他销售")
    sold = _project(db, salesperson="销售甲")
    unassigned = _project(db, salesperson="无关销售")
    part = DimPart(pn_std="REPL-MANAGER-UNION-PN", status="active")
    db.add(part)
    db.commit()
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_union_manager",
        role="maintenance_manager",
        salesperson_name="销售甲",
    )
    _assign_manager(db, project=assigned, user=user)

    listing = _get(client, "/api/replenishment-beta/projects")
    allowed = _post(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": sold.project_id,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    denied_payload = {
        "client_request_id": str(uuid4()),
        "project_id": unassigned.project_id,
        "request_note": None,
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }
    denied = _post(client, denied_payload)
    missing = _post(
        client,
        {**denied_payload, "client_request_id": str(uuid4()), "project_id": str(uuid4())},
    )

    assert listing.status_code == 200, listing.text
    assert {item["project_id"] for item in listing.json()["items"]} == {
        assigned.project_id,
        sold.project_id,
    }
    assert allowed.status_code == 201, allowed.text
    assert denied.status_code == missing.status_code == 404
    assert denied.json() == missing.json()


def test_revoked_manager_cannot_read_write_or_delete_retained_old_draft(db):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std="REPL-REVOKED-DRAFT-PN", status="active")
    db.add(part)
    db.commit()
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_revoked_draft_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=user)
    path = f"/api/replenishment-beta/cart-drafts/{project.project_id}"
    payload = {
        "expected_version": None,
        "request_note": "撤权前草稿",
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }
    created = _beta_request(client, "PUT", path, json=payload)
    assert created.status_code == 200, created.text
    draft = created.json()["draft"]

    _archive_manager(db, project=project, user=user)

    historical = _beta_request(client, "GET", path)
    replace_denied = _beta_request(
        client,
        "PUT",
        path,
        json={**payload, "expected_version": 1},
    )
    submit_denied = _beta_request(
        client,
        "POST",
        f"{path}/submit",
        json={
            "expected_version": draft["version"],
            "client_request_id": draft["client_request_id"],
        },
    )
    deleted = _beta_request(
        client,
        "DELETE",
        f"{path}?expected_version=1",
    )

    assert historical.status_code == 404
    assert replace_denied.status_code == 404
    assert submit_denied.status_code == 404
    assert deleted.status_code == 404
    db.expire_all()
    assert db.scalar(
        select(func.count())
        .select_from(ReplenishmentCartDraft)
        .where(ReplenishmentCartDraft.draft_id == draft["draft_id"])
    ) == 1


def test_cart_create_without_expected_version_cannot_overwrite_existing_draft(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-CART-CREATE-ONLY-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_cart_create_only_admin")
    path = f"/api/replenishment-beta/cart-drafts/{project.project_id}"
    first_payload = {
        "expected_version": None,
        "request_note": "第一页保存",
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }

    first = _beta_request(client, "PUT", path, json=first_payload)
    stale_create = _beta_request(
        client,
        "PUT",
        path,
        json={
            **first_payload,
            "request_note": "另一页不得静默覆盖",
            "lines": [
                {"part_id": part.id, "quantity": 2, "special_note": None}
            ],
        },
    )
    current = _beta_request(client, "GET", path)

    assert first.status_code == 200, first.text
    assert stale_create.status_code == 409, stale_create.text
    assert stale_create.json()["detail"]["code"] == "version_conflict"
    assert current.status_code == 200, current.text
    assert current.json()["draft"]["version"] == 1
    assert current.json()["draft"]["request_note"] == "第一页保存"
    assert current.json()["draft"]["lines"][0]["quantity"] == 1


def test_cart_submit_response_loss_retry_recovers_deleted_draft(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-CART-RETRY-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_cart_retry_admin")
    path = f"/api/replenishment-beta/cart-drafts/{project.project_id}"
    draft = _beta_request(
        client,
        "PUT",
        path,
        json={
            "expected_version": None,
            "request_note": "响应丢失重试",
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        },
    ).json()["draft"]
    submit_payload = {
        "expected_version": draft["version"],
        "client_request_id": draft["client_request_id"],
    }

    first = _beta_request(client, "POST", f"{path}/submit", json=submit_payload)
    retry = _beta_request(client, "POST", f"{path}/submit", json=submit_payload)
    cart_after = _beta_request(client, "GET", path)

    assert first.status_code == retry.status_code == 201
    assert first.json()["idempotent"] is False
    assert retry.json()["idempotent"] is True
    assert first.json()["application_id"] == retry.json()["application_id"]
    assert cart_after.status_code == 200
    assert cart_after.json() == {"draft": None}
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1


def test_cart_submit_requires_stable_retry_key_and_keeps_draft(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-CART-REQUIRED-KEY-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_cart_required_key_admin")
    path = f"/api/replenishment-beta/cart-drafts/{project.project_id}"
    draft = _beta_request(
        client,
        "PUT",
        path,
        json={
            "expected_version": None,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        },
    ).json()["draft"]

    response = _beta_request(
        client,
        "POST",
        f"{path}/submit",
        json={"expected_version": draft["version"]},
    )

    assert response.status_code == 422
    assert _beta_request(client, "GET", path).json()["draft"]["draft_id"] == draft["draft_id"]
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 0


def test_maintenance_manager_can_resubmit_owned_application_in_current_scope(db):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std="REPL-MANAGER-REVISION-PN", status="active")
    db.add(part)
    db.commit()
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_revision_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=user)
    created = _post_with_auto_review(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    assert created.status_code == 201, created.text
    assert created.json()["status"] == "needs_revision"

    revised = _beta_request(
        client,
        "POST",
        f"/api/replenishment-beta/applications/{created.json()['application_id']}"
        "/revisions",
        json={
            "expected_application_version": created.json()["version"],
            "client_request_id": str(uuid4()),
            "lines": [
                {"part_id": part.id, "quantity": 2, "special_note": "复提"}
            ],
        },
    )

    assert revised.status_code == 200, revised.text
    assert revised.json()["latest_version_no"] == 2
    assert revised.json()["versions"][0]["lines"][0]["quantity"] == 2


def test_full_line_revision_retry_is_idempotent_and_payload_bound(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-FULL-REVISION-RETRY-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_full_revision_retry_admin")
    created = _post_with_auto_review(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        },
    )
    assert created.status_code == 201, created.text
    assert created.json()["status"] == "needs_revision"
    path = (
        f"/api/replenishment-beta/applications/{created.json()['application_id']}"
        "/revisions"
    )
    request_id = str(uuid4())
    body = {
        "expected_application_version": created.json()["version"],
        "client_request_id": request_id,
        "lines": [
            {"part_id": part.id, "quantity": 2, "special_note": "复提重试"}
        ],
    }

    first = _beta_request(client, "POST", path, json=body)
    retry = _beta_request(client, "POST", path, json=body)
    changed = _beta_request(
        client,
        "POST",
        path,
        json={
            **body,
            "lines": [
                {"part_id": part.id, "quantity": 3, "special_note": "复提重试"}
            ],
        },
    )

    assert first.status_code == retry.status_code == 200
    assert first.json()["idempotent"] is False
    assert retry.json()["idempotent"] is True
    assert first.json()["latest_version_no"] == retry.json()["latest_version_no"] == 2
    assert first.json()["versions"][0]["lines"][0]["quantity"] == 2
    assert changed.status_code == 409, changed.text
    assert changed.json()["detail"]["code"] == "idempotency_conflict"


def test_resolution_revision_retry_is_idempotent_and_payload_bound(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-RESOLUTION-RETRY-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_resolution_retry_admin")
    created = _post_with_auto_review(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        },
    )
    assert created.status_code == 201, created.text
    rejected_line = created.json()["versions"][0]["lines"][0]
    path = (
        f"/api/replenishment-beta/applications/{created.json()['application_id']}"
        "/revisions"
    )
    request_id = str(uuid4())
    body = {
        "expected_application_version": created.json()["version"],
        "client_request_id": request_id,
        "resolutions": [
            {
                "request_line_id": rejected_line["request_line_id"],
                "action": "replace",
                "part_id": part.id,
                "special_note": "逐条复提重试",
            }
        ],
    }

    first = _beta_request(client, "POST", path, json=body)
    retry = _beta_request(client, "POST", path, json=body)
    changed = _beta_request(
        client,
        "POST",
        path,
        json={
            **body,
            "resolutions": [
                {**body["resolutions"][0], "quantity": 3}
            ],
        },
    )

    assert first.status_code == retry.status_code == 200
    assert first.json()["idempotent"] is False
    assert retry.json()["idempotent"] is True
    assert first.json()["latest_version_no"] == retry.json()["latest_version_no"] == 2
    assert first.json()["versions"][0]["lines"][0]["quantity"] == 1
    assert changed.status_code == 409, changed.text
    assert changed.json()["detail"]["code"] == "idempotency_conflict"


def test_revision_rejects_lines_and_resolutions_in_the_same_request(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-REVISION-EXACTLY-ONE-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_revision_exactly_one_admin")
    created = _post_with_auto_review(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        },
    )
    assert created.status_code == 201, created.text
    rejected = created.json()["versions"][0]["lines"][0]
    version_count = db.scalar(
        select(func.count()).select_from(ReplenishmentApplicationVersion)
    )

    response = _beta_request(
        client,
        "POST",
        f"/api/replenishment-beta/applications/{created.json()['application_id']}"
        "/revisions",
        json={
            "expected_application_version": created.json()["version"],
            "client_request_id": str(uuid4()),
            "lines": [
                {"part_id": part.id, "quantity": 2, "special_note": "全量"}
            ],
            "resolutions": [
                {
                    "request_line_id": rejected["request_line_id"],
                    "action": "remove",
                }
            ],
        },
    )

    assert response.status_code == 422
    assert db.scalar(
        select(func.count()).select_from(ReplenishmentApplicationVersion)
    ) == version_count


@pytest.mark.parametrize(
    ("lines", "resolutions"),
    [
        ([], None),
        (None, []),
        ([{}] * (replenishment.MAX_LINES + 1), None),
        (None, [{}] * (replenishment.MAX_LINES + 1)),
    ],
)
def test_revision_service_rejects_empty_or_oversized_mode_before_locking(
    db,
    lines,
    resolutions,
):
    with pytest.raises(replenishment.ReplenishmentError) as exc:
        replenishment.apply_revision_atomic(
            db,
            "not-a-real-application",
            username="revision-service-boundary",
            user_ctx=UserContext(
                user_id="revision-service-boundary",
                role="admin",
                is_authenticated=True,
            ),
            expected_application_version=1,
            client_request_id="revision-service-boundary-key",
            lines=lines,
            resolutions=resolutions,
        )

    assert exc.value.status_code == 422
    assert exc.value.code == "revision_content_required"


def test_revoked_owner_keeps_history_but_cannot_apply_revision(db):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std="REPL-REVOKED-REVISION-PN", status="active")
    db.add(part)
    db.commit()
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_revoked_revision_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=user)
    created = _post_with_auto_review(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    assert created.status_code == 201, created.text
    application_id = created.json()["application_id"]
    version_before = db.scalar(
        select(func.count()).select_from(ReplenishmentApplicationVersion)
    )
    _archive_manager(db, project=project, user=user)

    listing = _get(client, "/api/replenishment-beta/applications")
    detail = _get(
        client,
        f"/api/replenishment-beta/applications/{application_id}",
    )
    revision = _beta_request(
        client,
        "POST",
        f"/api/replenishment-beta/applications/{application_id}/revisions",
        json={
            "expected_application_version": created.json()["version"],
            "client_request_id": str(uuid4()),
            "lines": [
                {"part_id": part.id, "quantity": 2, "special_note": "撤权后复提"}
            ],
        },
    )

    assert listing.status_code == 200, listing.text
    assert [item["application_id"] for item in listing.json()["items"]] == [
        application_id
    ]
    assert detail.status_code == 200, detail.text
    assert revision.status_code == 404, revision.text
    assert revision.json()["detail"]["code"] == "project_unavailable"
    assert (
        db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion))
        == version_before
    )


def test_revoked_owner_cannot_replay_atomic_submit(db):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std="REPL-REVOKED-ATOMIC-REPLAY-PN", status="active")
    db.add(part)
    db.commit()
    user, client = _explicit_replenishment_client(
        db,
        username="replenishment_revoked_atomic_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=user)
    payload = {
        "client_request_id": str(uuid4()),
        "project_id": project.project_id,
        "request_note": None,
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }
    created = _post(client, payload)
    assert created.status_code == 201, created.text
    _archive_manager(db, project=project, user=user)

    replay = _post(client, payload)

    assert replay.status_code == 404, replay.text
    assert replay.json()["detail"]["code"] == "project_unavailable"
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1


def test_boss_does_not_inherit_admin_application_owner_exception(db):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std="REPL-BOSS-OWNER-SCOPE-PN", status="active")
    db.add(part)
    db.commit()
    owner, owner_client = _explicit_replenishment_client(
        db,
        username="replenishment_owner_manager",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=owner)
    created = _post(
        owner_client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    assert created.status_code == 201, created.text
    application_id = created.json()["application_id"]
    _boss, boss_client = _explicit_replenishment_client(
        db,
        username="replenishment_non_owner_boss",
        role="boss",
    )
    admin_client = _admin_client(
        db,
        username="replenishment_non_owner_admin",
    )
    revision_body = {
        "expected_application_version": created.json()["version"],
        "client_request_id": str(uuid4()),
        "lines": [{"part_id": part.id, "quantity": 2, "special_note": None}],
    }

    boss_list = _get(boss_client, "/api/replenishment-beta/applications")
    boss_detail = _get(
        boss_client,
        f"/api/replenishment-beta/applications/{application_id}",
    )
    boss_revision = _beta_request(
        boss_client,
        "POST",
        f"/api/replenishment-beta/applications/{application_id}/revisions",
        json=revision_body,
    )
    admin_list = _get(admin_client, "/api/replenishment-beta/applications")
    admin_detail = _get(
        admin_client,
        f"/api/replenishment-beta/applications/{application_id}",
    )
    admin_revision = _beta_request(
        admin_client,
        "POST",
        f"/api/replenishment-beta/applications/{application_id}/revisions",
        json=revision_body,
    )

    assert boss_list.status_code == 200, boss_list.text
    assert boss_list.json()["total"] == 0
    assert boss_detail.status_code == 404
    assert boss_revision.status_code == 404
    assert admin_list.status_code == 200, admin_list.text
    assert admin_list.json()["total"] == 1
    assert admin_detail.status_code == 200, admin_detail.text
    assert admin_revision.status_code == 409, admin_revision.text


def test_inflight_admin_context_cannot_revise_other_owner_after_role_demotion(db):
    project = _project(db, salesperson="其他销售")
    part = DimPart(pn_std="REPL-DEMOTED-ADMIN-OWNER-PN", status="active")
    db.add(part)
    db.commit()
    owner, owner_client = _explicit_replenishment_client(
        db,
        username="replenishment_demotion_owner",
        role="maintenance_manager",
    )
    _assign_manager(db, project=project, user=owner)
    created = _post_with_auto_review(
        owner_client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None}
            ],
        },
    )
    assert created.status_code == 201, created.text
    assert created.json()["status"] == "needs_revision"
    stale_admin, _client = _explicit_replenishment_client(
        db,
        username="replenishment_demoted_admin",
        role="admin",
    )
    stale_ctx = UserContext(
        user_id=stale_admin.username,
        role="admin",
        salesperson_name=stale_admin.salesperson_name,
        permissions=permissions.runtime_safe(
            permissions.effective_for_user(stale_admin)
        ),
        is_authenticated=True,
    )
    stale_admin.role = "boss"
    stale_admin.token_version = (stale_admin.token_version or 0) + 1
    db.commit()
    version_count = db.scalar(
        select(func.count()).select_from(ReplenishmentApplicationVersion)
    )

    with pytest.raises(replenishment.ReplenishmentError) as exc:
        replenishment.apply_revision_atomic(
            db,
            created.json()["application_id"],
            username=stale_admin.username,
            user_ctx=stale_ctx,
            expected_application_version=created.json()["version"],
            client_request_id=str(uuid4()),
            lines=[
                {"part_id": part.id, "quantity": 2, "special_note": "越权复提"}
            ],
        )

    assert exc.value.code == "not_found"
    assert exc.value.status_code == 404
    assert db.scalar(
        select(func.count()).select_from(ReplenishmentApplicationVersion)
    ) == version_count


def test_sales_project_picker_uses_exact_salesperson_mapping(db):
    exact = _project(db, salesperson="销售甲")
    db.add_all(
        [
            MaintenanceProject(
                project_id=str(uuid4()),
                project_code="REPL-OTHER-SALES",
                display_name="其他销售项目",
                salesperson="销售乙",
                lifecycle_status="ongoing",
                is_active=True,
            ),
            MaintenanceProject(
                project_id=str(uuid4()),
                project_code="REPL-WHITESPACE-MISMATCH",
                display_name="空格不匹配项目",
                salesperson="销售甲 ",
                lifecycle_status="ongoing",
                is_active=True,
            ),
        ]
    )
    db.commit()
    client = _sales_client(
        db, username="replenishment_project_sales", salesperson_name="销售甲"
    )

    response = _get(client, "/api/replenishment-beta/projects")

    assert response.status_code == 200, response.text
    assert [item["project_id"] for item in response.json()["items"]] == [
        exact.project_id
    ]


def test_sales_without_mapping_sees_no_projects_and_submit_fails_closed(db):
    project = _project(db, salesperson="销售甲")
    part = DimPart(pn_std="REPL-UNMAPPED-SALES-PN", status="active")
    db.add(part)
    db.commit()
    client = _sales_client(
        db, username="replenishment_unmapped_sales", salesperson_name=None
    )
    payload = {
        "client_request_id": str(uuid4()),
        "project_id": project.project_id,
        "request_note": None,
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }

    listing = _get(client, "/api/replenishment-beta/projects")
    existing = _post(client, payload)
    missing = _post(client, {**payload, "project_id": str(uuid4())})

    assert listing.status_code == 200
    assert listing.json() == {"items": []}
    assert existing.status_code == missing.status_code == 404
    assert existing.json() == missing.json()
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 0


def test_same_owner_and_request_replays_without_duplicate_business_rows(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-IDEMPOTENT-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_idempotent_admin")
    payload = {
        "client_request_id": str(uuid4()),
        "project_id": project.project_id,
        "request_note": "顺序重试",
        "lines": [{"part_id": part.id, "quantity": 3, "special_note": None}],
    }

    first = _post(client, payload)
    second = _post(client, payload)

    assert first.status_code == second.status_code == 201
    assert first.json()["idempotent"] is False
    assert second.json()["idempotent"] is True
    assert first.json()["application_id"] == second.json()["application_id"]
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationLine)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == 1


def test_direct_submit_locks_project_before_idempotency_advisory(db, monkeypatch):
    project = _project(db)
    part = DimPart(pn_std="REPL-DIRECT-LOCK-ORDER-PN", status="active")
    db.add(part)
    db.commit()
    username = "replenishment_direct_lock_order_admin"
    _admin_client(db, username=username)
    events: list[str] = []
    original_authorized_project = replenishment._authorized_project
    original_execute = db.execute

    def tracked_authorized_project(*args, **kwargs):
        events.append("project")
        return original_authorized_project(*args, **kwargs)

    def tracked_execute(statement, *args, **kwargs):
        if "pg_advisory_xact_lock" in str(statement):
            events.append("advisory")
        return original_execute(statement, *args, **kwargs)

    monkeypatch.setattr(replenishment, "_authorized_project", tracked_authorized_project)
    monkeypatch.setattr(db, "execute", tracked_execute)

    replenishment.submit_application_atomic(
        db,
        username=username,
        user_ctx=UserContext(
            user_id=username,
            role="admin",
            is_authenticated=True,
        ),
        client_request_id=str(uuid4()),
        project_id=project.project_id,
        lines=[{"part_id": part.id, "quantity": 1, "special_note": None}],
    )

    assert events == ["project", "advisory"]


def test_cart_submit_lock_order_is_project_draft_project_advisory(db, monkeypatch):
    project = _project(db)
    part = DimPart(pn_std="REPL-CART-LOCK-ORDER-PN", status="active")
    db.add(part)
    db.commit()
    username = "replenishment_cart_lock_order_admin"
    client = _admin_client(db, username=username)
    draft = _beta_request(
        client,
        "PUT",
        f"/api/replenishment-beta/cart-drafts/{project.project_id}",
        json={
            "expected_version": None,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    assert draft.status_code == 200, draft.text
    events: list[str] = []
    original_authorized_project = replenishment._authorized_project
    original_execute = db.execute
    original_scalar = db.scalar

    def tracked_authorized_project(*args, **kwargs):
        events.append("project")
        return original_authorized_project(*args, **kwargs)

    def tracked_execute(statement, *args, **kwargs):
        if "pg_advisory_xact_lock" in str(statement):
            events.append("advisory")
        return original_execute(statement, *args, **kwargs)

    def tracked_scalar(statement, *args, **kwargs):
        sql = str(statement)
        if "FROM replenishment_cart_draft" in sql and "FOR UPDATE" in sql:
            events.append("draft")
        return original_scalar(statement, *args, **kwargs)

    monkeypatch.setattr(replenishment, "_authorized_project", tracked_authorized_project)
    monkeypatch.setattr(db, "execute", tracked_execute)
    monkeypatch.setattr(db, "scalar", tracked_scalar)

    replenishment_cart.submit_cart_draft_atomic(
        db,
        username=username,
        user_ctx=UserContext(
            user_id=username,
            role="admin",
            is_authenticated=True,
        ),
        project_id=project.project_id,
        expected_version=draft.json()["draft"]["version"],
        client_request_id=draft.json()["draft"]["client_request_id"],
    )

    assert events == ["project", "draft", "project", "advisory"]


def test_same_owner_and_request_with_different_payload_is_conflict(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-IDEMPOTENCY-CONFLICT-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_conflict_admin")
    request_id = str(uuid4())
    payload = {
        "client_request_id": request_id,
        "project_id": project.project_id,
        "request_note": None,
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }

    first = _post(client, payload)
    changed = _post(
        client,
        {
            **payload,
            "lines": [{"part_id": part.id, "quantity": 2, "special_note": None}],
        },
    )

    assert first.status_code == 201
    assert changed.status_code == 409
    assert changed.json()["detail"]["code"] == "idempotency_conflict"
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationLine)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == 1


def test_same_owner_and_request_across_projects_is_conflict(db):
    first_project = _project(db)
    second_project = _project(db)
    part = DimPart(pn_std="REPL-CROSS-PROJECT-IDEMPOTENCY-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_cross_project_idempotency_admin")
    request_id = str(uuid4())
    payload = {
        "client_request_id": request_id,
        "project_id": first_project.project_id,
        "request_note": None,
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }

    first = _post(client, payload)
    cross_project = _post(
        client,
        {**payload, "project_id": second_project.project_id},
    )

    assert first.status_code == 201, first.text
    assert cross_project.status_code == 409, cross_project.text
    assert cross_project.json()["detail"]["code"] == "idempotency_conflict"
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1


def test_validation_failure_leaves_zero_business_rows(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-DUPLICATE-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_atomic_failure_admin")

    response = _post(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1, "special_note": None},
                {"part_id": part.id, "quantity": 2, "special_note": None},
            ],
        },
    )

    assert response.status_code == 409
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationLine)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == 0


def test_fractional_quantity_is_rejected_before_any_business_write(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-FRACTIONAL-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_fractional_admin")

    response = _post(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [
                {"part_id": part.id, "quantity": 1.5, "special_note": None}
            ],
        },
    )

    assert response.status_code == 422
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationLine)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == 0


def test_concurrent_same_owner_and_request_returns_one_application(db):
    project = _project(db)
    part = DimPart(pn_std="REPL-CONCURRENT-PN", status="active")
    db.add(part)
    db.commit()
    authenticated = _admin_client(db, username="replenishment_concurrent_admin")
    clients = [
        TestClient(app, raise_server_exceptions=False),
        TestClient(app, raise_server_exceptions=False),
    ]
    for client in clients:
        client.headers["Authorization"] = authenticated.headers["Authorization"]
    payload = {
        "client_request_id": str(uuid4()),
        "project_id": project.project_id,
        "request_note": "并发重试",
        "lines": [{"part_id": part.id, "quantity": 4, "special_note": None}],
    }
    barrier = Barrier(2)

    def submit(client: TestClient):
        barrier.wait()
        return _post(client, payload)

    try:
        with ThreadPoolExecutor(max_workers=2) as executor:
            responses = list(executor.map(submit, clients))
    finally:
        for client in clients:
            client.close()

    assert [response.status_code for response in responses] == [201, 201]
    assert len({response.json()["application_id"] for response in responses}) == 1
    assert sorted(response.json()["idempotent"] for response in responses) == [False, True]
    db.expire_all()
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationVersion)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplicationLine)) == 1
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == 1


def test_atomic_submit_locks_project_account_scope_and_part_eligibility(db, monkeypatch):
    project = _project(db)
    part = DimPart(pn_std="REPL-ELIGIBILITY-LOCK-PN", status="active")
    db.add(part)
    db.commit()
    client = _admin_client(db, username="replenishment_eligibility_lock_admin")
    entered_screening = Event()
    release_screening = Event()
    original_screen = replenishment_screening.screen

    def blocked_screen(*args, **kwargs):
        entered_screening.set()
        assert release_screening.wait(timeout=5)
        return original_screen(*args, **kwargs)

    def update_times_out(statement: str, params: dict) -> bool:
        try:
            with engine.begin() as connection:
                connection.execute(text("SET LOCAL lock_timeout = '200ms'"))
                connection.execute(text(statement), params)
        except DBAPIError:
            return True
        return False

    monkeypatch.setattr(replenishment_screening, "screen", blocked_screen)
    payload = {
        "client_request_id": str(uuid4()),
        "project_id": project.project_id,
        "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
    }
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(_post, client, payload)
        assert entered_screening.wait(timeout=5)
        try:
            project_locked = update_times_out(
                "UPDATE maintenance_project SET is_active = false "
                "WHERE project_id = :project_id",
                {"project_id": project.project_id},
            )
            part_locked = update_times_out(
                "UPDATE dim_part SET status = 'merged' WHERE id = :part_id",
                {"part_id": part.id},
            )
            account_scope_locked = update_times_out(
                "UPDATE sys_user SET salesperson_name = :salesperson_name "
                "WHERE username = :username",
                {
                    "salesperson_name": "并发撤销后的销售",
                    "username": "replenishment_eligibility_lock_admin",
                },
            )
        finally:
            release_screening.set()
        response = future.result(timeout=10)

    assert project_locked is True
    assert part_locked is True
    assert account_scope_locked is True
    assert response.status_code == 201, response.text


def test_retired_mutation_and_derived_routes_return_410_without_writes(db):
    client = _admin_client(db, username="replenishment_retired_routes_admin")
    application_id = str(uuid4())
    line_id = str(uuid4())
    version_id = str(uuid4())
    requests = [
        (
            "PATCH",
            f"/api/replenishment-beta/applications/{application_id}",
            {"json": {"expected_version": 1, "warehouse": None, "request_note": None}},
        ),
        (
            "POST",
            f"/api/replenishment-beta/applications/{application_id}/lines",
            {"json": {"expected_version": 1, "part_id": 1, "quantity": 1}},
        ),
        (
            "PATCH",
            f"/api/replenishment-beta/applications/{application_id}/lines/{line_id}",
            {"json": {"expected_version": 1, "part_id": 1, "quantity": 1}},
        ),
        (
            "DELETE",
            f"/api/replenishment-beta/applications/{application_id}/lines/{line_id}"
            "?expected_version=1",
            {},
        ),
        (
            "POST",
            f"/api/replenishment-beta/applications/{application_id}/submit",
            {"json": {"expected_version": 1}},
        ),
        (
            "POST",
            f"/api/replenishment-beta/applications/{application_id}/revision",
            {"json": {"expected_version": 1}},
        ),
        (
            "POST",
            f"/api/replenishment-beta/applications/{application_id}/review-results",
            {
                "json": {
                    "version_id": version_id,
                    "content_digest": "a" * 64,
                    "idempotency_key": "retired-review-key",
                    "decisions": [
                        {"line_id": line_id, "decision": "approved", "reason": None}
                    ],
                }
            },
        ),
        (
            "GET",
            f"/api/replenishment-beta/applications/{application_id}"
            "/exports/manual-review.xlsx",
            {},
        ),
        (
            "GET",
            f"/api/replenishment-beta/applications/{application_id}"
            "/exports/wbdd-subset.xlsx",
            {},
        ),
        ("GET", f"/api/replenishment-beta/applications/{application_id}/evidence", {}),
        (
            "GET",
            f"/api/replenishment-beta/applications/{application_id}"
            "/exports/purchase-list.xlsx",
            {},
        ),
    ]

    settings = get_settings()
    original = settings.replenishment_beta_enabled
    try:
        settings.replenishment_beta_enabled = True
        responses = [
            client.request(method, path, **kwargs)
            for method, path, kwargs in requests
        ]
    finally:
        settings.replenishment_beta_enabled = original

    assert [response.status_code for response in responses] == [410] * len(requests)
    assert db.scalar(select(func.count()).select_from(ReplenishmentApplication)) == 0
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == 0


def test_retired_review_route_exposes_no_decision_request_contract(db):
    client = _admin_client(db, username="replenishment_retired_schema_admin")

    schema = client.get("/openapi.json").json()

    operation = schema["paths"][
        "/api/replenishment-beta/applications/{application_id}/review-results"
    ]["post"]
    assert "requestBody" not in operation
    assert "ReviewDecision" not in schema["components"]["schemas"]
    assert "ReviewWrite" not in schema["components"]["schemas"]


def test_retired_routes_do_not_validate_removed_workflow_inputs(db):
    client = _admin_client(db, username="replenishment_retired_inputs_admin")
    application_id = str(uuid4())
    line_id = str(uuid4())
    requests = [
        ("PATCH", f"/api/replenishment-beta/applications/{application_id}"),
        ("POST", f"/api/replenishment-beta/applications/{application_id}/lines"),
        (
            "PATCH",
            f"/api/replenishment-beta/applications/{application_id}/lines/{line_id}",
        ),
        (
            "DELETE",
            f"/api/replenishment-beta/applications/{application_id}/lines/{line_id}",
        ),
        ("POST", f"/api/replenishment-beta/applications/{application_id}/submit"),
        ("POST", f"/api/replenishment-beta/applications/{application_id}/revision"),
        (
            "POST",
            f"/api/replenishment-beta/applications/{application_id}/review-results",
        ),
    ]

    settings = get_settings()
    original = settings.replenishment_beta_enabled
    try:
        settings.replenishment_beta_enabled = True
        responses = [
            client.request(method, path, json={}) for method, path in requests
        ]
    finally:
        settings.replenishment_beta_enabled = original

    assert [response.status_code for response in responses] == [410] * len(requests)


def test_system_screening_export_is_frozen_and_read_only(db, monkeypatch):
    project = _project(db)
    part = DimPart(
        pn_std="REPL-FROZEN-EXPORT-PN", description="提交时描述", status="active"
    )
    db.add(part)
    db.flush()
    pool = PartPool(
        group_id=980000 + part.id,
        name="补库冻结池",
        status="active",
        source="manual",
        member_count=1,
    )
    db.add(pool)
    db.flush()
    db.add(PartPoolMember(group_id=pool.group_id, part_id=part.id))
    policy = PartPoolPricePolicy(
        group_id=pool.group_id,
        sales_floor_ex_tax=Decimal("88.00"),
        sales_input_value=Decimal("88.00"),
        sales_input_basis="ex_tax",
        valid_to=None,
        changed_by="tester",
    )
    db.add(policy)
    db.commit()
    client = _admin_client(db, username="replenishment_frozen_export_admin")
    created = _post(
        client,
        {
            "client_request_id": str(uuid4()),
            "project_id": project.project_id,
            "request_note": None,
            "lines": [{"part_id": part.id, "quantity": 1, "special_note": None}],
        },
    )
    assert created.status_code == 201, created.text
    audit_count = db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent))

    policy.sales_floor_ex_tax = Decimal("777.00")
    part.description = "提交后被修改的实时描述"
    db.commit()

    def reject_live_query(*_args, **_kwargs):
        raise AssertionError("export must not query mutable screening sources")

    monkeypatch.setattr(replenishment_screening, "screen", reject_live_query)
    monkeypatch.setattr(replenishment_screening, "latest_sales_history", reject_live_query)
    monkeypatch.setattr(replenishment_screening, "pool_floor_prices", reject_live_query)
    response = _get(
        client,
        f"/api/replenishment-beta/applications/{created.json()['application_id']}"
        "/exports/system-screening.xlsx",
    )

    assert response.status_code == 200, response.text
    rows = [
        [cell.value for cell in row]
        for row in load_workbook(BytesIO(response.content)).active.iter_rows()
    ]
    headers, body = rows[1], rows[2]
    assert body[headers.index("产品描述")] == "提交时描述"
    assert body[headers.index("池内最低价(未税)")] == 88.0
    assert "批准" not in rows[0][0]
    assert "驳回" not in rows[0][0]
    db.expire_all()
    assert db.scalar(select(func.count()).select_from(ReplenishmentAuditEvent)) == audit_count


def test_legacy_unbound_history_cannot_export_system_screening(db):
    username = "replenishment_legacy_export_admin"
    client = _admin_client(db, username=username)
    db.execute(
        text(
            "ALTER TABLE replenishment_application "
            "DISABLE TRIGGER trg_replenishment_project_binding"
        )
    )
    try:
        db.execute(
            text(
                "INSERT INTO replenishment_application "
                "(application_id, application_no, owner_username, "
                "is_legacy_project_unbound, status) VALUES "
                "('legacy-export-app', 'BLK-LEGACY-EXPORT', :owner, true, 'submitted')"
            ),
            {"owner": username},
        )
    finally:
        db.execute(
            text(
                "ALTER TABLE replenishment_application "
                "ENABLE TRIGGER trg_replenishment_project_binding"
            )
        )
    db.execute(
        text(
            "INSERT INTO replenishment_application_version "
            "(version_id, application_id, version_no, status, warehouse, "
            "content_digest, created_by, submitted_by, submitted_at) VALUES "
            "('legacy-export-version', 'legacy-export-app', 1, 'submitted', "
            "'历史仓', :digest, :owner, :owner, now())"
        ),
        {"digest": "a" * 64, "owner": username},
    )
    db.commit()

    response = _get(
        client,
        "/api/replenishment-beta/applications/legacy-export-app/"
        "exports/system-screening.xlsx",
    )

    assert response.status_code == 409, response.text
    assert response.json()["detail"]["code"] == "legacy_project_unbound"
