"""V2.1 项目总表 03 备件明细全字段可编辑：改/作废/新增/数量重算/审计/读侧过滤。

对应 REQUIREMENTS #55。删除=软作废（is_active=false），不计入计算、不再导出、
06 关联行级联作废；氚云 loader 白名单不含作废列（重传不复活）。
"""
import io
import uuid
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select

from app.models.dimensions import DimPart
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceDemandTombstone,
    MaintenanceManualCostOverride,
)
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectAuditLog,
    MaintenanceProjectContract,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
    MaintenanceProjectOperationAudit,
    MaintenanceProjectWorkbookOperation,
    MaintenanceProjectWorkbookState,
    MaintenanceSiteIssue,
    MaintenanceSiteIssueLine,
)
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.models.sales import FSalesOrder
from app.models.system import SysImportBatch
from app.security import UserContext
from app.services import maintenance_project_master_workbook as master
from app.services import maintenance_project_operations as operations


def _make_project_with_line(db, *, qty=Decimal("2"), unit_cost=Decimal("100.00"),
                            cost_source="direct", return_qty=Decimal("0"),
                            source_line_id=None):
    tag = uuid.uuid4().hex[:8]
    project = MaintenanceProject(
        project_id=str(uuid.uuid4()), project_code=f"EDIT-{tag}",
        display_name="可编辑测试", lifecycle_status="ongoing",
    )
    # PN 查找走 upper() 归一，夹具须用大写 hex 避免大小写错配
    part = DimPart(pn_std=f"EDIT-PN-{tag.upper()}", description="可编辑备件")
    db.add_all([project, part])
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=project.project_id,
        contract_id="EDIT-CONTRACT", contract_no="XSDD-EDIT-001",
        amount_inc_tax=Decimal("10000.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1,
    ))
    order = FMaintenanceOrder(
        raw_order_id=f"raw-order-{uuid.uuid4()}", order_no="WBDD-EDIT-001",
        order_date=date(2026, 8, 1), linked_sales_order_no="XSDD-EDIT-001",
        project_raw="可编辑测试", data_status="已生效",
        import_batch_id=_batch(db),
    )
    db.add(order)
    db.flush()
    line = FMaintenanceLine(
        raw_line_id=source_line_id or f"raw-line-{uuid.uuid4()}",
        order_id=order.id, line_no=1, part_id=part.id,
        pn_std=part.pn_std, pn_raw=part.pn_std, description=part.description,
        qty=qty, return_qty=return_qty,
        unit_cost=unit_cost if unit_cost is not None else None,
        unit_cost_ex_tax=unit_cost if unit_cost is not None else None,
        unit_cost_inc_tax=(unit_cost * Decimal("1.13")).quantize(Decimal("0.01"))
        if unit_cost is not None else None,
        cost_amount_ex_tax=(unit_cost * max(qty - return_qty, Decimal(0))).quantize(Decimal("0.01"))
        if unit_cost is not None else None,
        cost_amount_inc_tax=(unit_cost * Decimal("1.13") * max(qty - return_qty, Decimal(0))).quantize(Decimal("0.01"))
        if unit_cost is not None else None,
        cost_amount=(unit_cost * max(qty - return_qty, Decimal(0))).quantize(Decimal("0.01"))
        if unit_cost is not None else None,
        cost_source=cost_source, cost_tax_basis="ex", confidence="high",
        import_batch_id=_batch(db),
    )
    db.add(line)
    db.add(MaintenanceSourceOrderAssignment(
        assignment_id=str(uuid.uuid4()), project_id=project.project_id,
        source_order_id=order.raw_order_id, is_active=True,
        created_by="test",
    ))
    db.commit()
    return project, part, order, line


def _batch(db) -> int:
    from app.models.system import SysImportBatch
    b = SysImportBatch(
        filename="test.xlsx", file_type="maintenance",
        file_hash=uuid.uuid4().hex, status="success",
    )
    db.add(b)
    db.flush()
    return b.id


def _parts_sheet(db, project_id: str):
    content = master.build_project_master_v2(
        db, project_id=project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    return content, wb, wb[master.V2_SHEET_PARTS]


def _overview_values(db, project_id: str) -> dict:
    content = master.build_project_master_v2(
        db, project_id=project_id, sheets=(master.V2_SHEET_OVERVIEW,))
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    return {
        row[0].value: row[1].value
        for row in workbook[master.V2_SHEET_OVERVIEW].iter_rows(min_col=1, max_col=2)
        if row[0].value
    }


def test_v2_overview_strict_cost_keeps_real_zero_and_rejects_dirty_fact(db):
    project, _part, _order, line = _make_project_with_line(
        db, qty=Decimal("1"), unit_cost=Decimal("0"),
    )
    values = _overview_values(db, project.project_id)
    assert Decimal(values["备件成本（含税）"]) == Decimal("0.00")
    assert values["缺成本行数"] == 0

    line.cost_source = "future_source"
    db.commit()
    values = _overview_values(db, project.project_id)
    assert values["备件成本（含税）"] == "—"
    assert values["缺成本行数"] == 1


def test_v2_overview_active_manual_override_uses_net_quantity(db):
    project, _part, _order, line = _make_project_with_line(
        db, qty=Decimal("3"), return_qty=Decimal("1"), unit_cost=None,
        cost_source=None,
    )
    db.add(MaintenanceManualCostOverride(
        line_id=line.id,
        unit_cost_ex_tax=Decimal("8.00"),
        unit_cost_inc_tax=Decimal("9.04"),
        active=True,
        updated_by="tester",
    ))
    db.commit()

    values = _overview_values(db, project.project_id)
    assert Decimal(values["备件成本（含税）"]) == Decimal("18.08")
    assert values["缺成本行数"] == 0


def _reupload(db, project_id, wb):
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project_id, data=buf.getvalue())
    return master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))


def test_v21_parts_sheet_has_operation_column_and_new_template(db):
    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(db, project_id=project.project_id)
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = [c.value for c in ws[1]]
    assert headers[0] == "操作"
    assert headers.count("实体ID") == 1
    meta = {r[0].value: r[1].value for r in wb[master.V2_SHEET_META].iter_rows(min_col=1, max_col=2)}
    assert meta["template_version"] == "2.6.0"
    assert meta["metadata_hmac_algorithm"] == "HMAC-SHA256"
    assert len(meta["metadata_hmac"]) == 64


def test_v25_non_contract_apply_bumps_workbook_revision_once(db):
    """Parts/site/expense facts invalidate stale readers, not only contracts."""

    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_PARTS,),
    )
    state = operations.get_or_create_workbook_state(
        db, project_id=project.project_id)
    db.commit()
    before = state.revision

    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["需求数量"], 3)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    master.apply_project_master_v2(
        db,
        plan,
        operated_by="revision-v2",
        import_batch_id=str(uuid.uuid4()),
    )

    db.expire_all()
    state = operations.get_or_create_workbook_state(
        db, project_id=project.project_id)
    assert state.revision == before + 1


def test_v26_different_workbook_from_same_revision_is_zero_write_stale(
    db,
):
    """First writer wins; its ACK replay wins before the stale revision gate."""

    project, _part, _order, line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_PARTS,),
    )
    first_wb = load_workbook(io.BytesIO(content))
    second_wb = load_workbook(io.BytesIO(content))
    first_ws = first_wb[master.V2_SHEET_PARTS]
    second_ws = second_wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in first_ws[1]}
    first_ws.cell(2, headers["需求数量"], 3)
    second_ws.cell(2, headers["需求数量"], 4)
    first_plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(first_wb))
    second_plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(second_wb))

    first = master.apply_project_master_v2(
        db,
        first_plan,
        operated_by="occ-writer",
        import_batch_id=str(uuid.uuid4()),
    )
    state_after_first = operations.get_or_create_workbook_state(
        db, project_id=project.project_id).revision
    replay = master.apply_project_master_v2(
        db,
        first_plan,
        operated_by="occ-writer",
        import_batch_id=str(uuid.uuid4()),
    )
    assert first["replayed"] is False
    assert replay["replayed"] is True
    assert operations.get_or_create_workbook_state(
        db, project_id=project.project_id).revision == state_after_first

    audits_before = int(db.scalar(
        select(func.count(MaintenanceProjectOperationAudit.id))) or 0)
    with pytest.raises(master.WorkbookError) as raised:
        master.apply_project_master_v2(
            db,
            second_plan,
            operated_by="occ-writer",
            import_batch_id=str(uuid.uuid4()),
        )
    assert raised.value.code == "stale_workbook"
    db.rollback()
    db.refresh(line)
    assert line.qty == Decimal("3.00")
    assert int(db.scalar(
        select(func.count(MaintenanceProjectOperationAudit.id))) or 0) == audits_before


def test_v1_apply_rejects_foreign_part_and_site_hidden_ids(db):
    project, _part, _order, _line = _make_project_with_line(db)
    other, other_part, _other_order, foreign_line = _make_project_with_line(db)
    _issue, foreign_site_line = _site_issue(db, other, other_part)

    foreign_cost_plan = master.MasterPlan(
        project_id=project.project_id,
        cost_refills=(master.CostRefill(
            line_id=foreign_line.id,
            unit_cost_ex_tax=Decimal("99.00"),
            unit_cost_inc_tax=Decimal("111.87"),
            reason="crafted-foreign-id",
        ),),
        sheets=(master.SHEET_PARTS,),
    )
    with pytest.raises(master.WorkbookError) as cost_error:
        master.apply(
            db,
            foreign_cost_plan,
            operated_by="v1-scope-test",
            import_batch_id=str(uuid.uuid4()),
        )
    assert cost_error.value.code == "project_scope_denied"
    db.rollback()
    db.refresh(foreign_line)
    assert foreign_line.unit_cost_ex_tax == Decimal("100.00")

    foreign_site_plan = master.MasterPlan(
        project_id=project.project_id,
        site_flags=(master.SiteReturnFlag(
            issue_line_id=foreign_site_line.issue_line_id,
            no_return=True,
        ),),
        sheets=(master.SHEET_SITE,),
    )
    with pytest.raises(master.WorkbookError) as site_error:
        master.apply(
            db,
            foreign_site_plan,
            operated_by="v1-scope-test",
            import_batch_id=str(uuid.uuid4()),
        )
    assert site_error.value.code == "project_scope_denied"
    db.rollback()
    db.refresh(foreign_site_line)
    assert foreign_site_line.no_return is not True


def test_v26_contract_fallback_uses_only_active_success_sales_evidence(db):
    project, _part, _order, _line = _make_project_with_line(db)
    project.period_from = date(2026, 1, 1)
    contract_no = f"XSDD-FALLBACK-{uuid.uuid4().hex[:8]}"
    batches = {
        "valid": SysImportBatch(
            filename="valid.xlsx", file_type="sales",
            file_hash=uuid.uuid4().hex, status="success"),
        "failed": SysImportBatch(
            filename="failed.xlsx", file_type="sales",
            file_hash=uuid.uuid4().hex, status="failed"),
        "wrong_type": SysImportBatch(
            filename="wrong.xlsx", file_type="maintenance",
            file_hash=uuid.uuid4().hex, status="success"),
    }
    db.add_all(batches.values())
    db.flush()
    db.add_all([
        FSalesOrder(
            raw_order_id=f"sales-valid-{uuid.uuid4()}",
            order_no=contract_no,
            amount_ex_tax=Decimal("100.00"), tax_rate=Decimal("0.13"),
            data_status="已生效", import_batch_id=batches["valid"].id,
        ),
        FSalesOrder(
            raw_order_id=f"sales-inactive-{uuid.uuid4()}",
            order_no=contract_no,
            amount_ex_tax=Decimal("999.00"), tax_rate=Decimal("0.13"),
            data_status="已作废", import_batch_id=batches["valid"].id,
        ),
        FSalesOrder(
            raw_order_id=f"sales-failed-{uuid.uuid4()}",
            order_no=contract_no,
            amount_ex_tax=Decimal("888.00"), tax_rate=Decimal("0.13"),
            data_status="已生效", import_batch_id=batches["failed"].id,
        ),
        FSalesOrder(
            raw_order_id=f"sales-wrong-type-{uuid.uuid4()}",
            order_no=contract_no,
            amount_ex_tax=Decimal("777.00"), tax_rate=Decimal("0.13"),
            data_status="已生效", import_batch_id=batches["wrong_type"].id,
        ),
    ])
    db.commit()

    contract, created = master._ensure_contract_for_xsdd_apply(
        db,
        project=project,
        contract_no=contract_no,
        operated_by="fallback-test",
        assigned_xsdd_nos={contract_no},
    )
    assert created is True
    assert contract.contract_amount == Decimal("100.00")
    assert contract.amount_inc_tax == Decimal("113.00")


def test_v26_contract_fallback_rejects_conflicting_active_candidates(db):
    project, _part, _order, _line = _make_project_with_line(db)
    project.period_from = date(2026, 1, 1)
    contract_no = f"XSDD-AMBIG-{uuid.uuid4().hex[:8]}"
    first_batch = SysImportBatch(
        filename="first.xlsx", file_type="sales",
        file_hash=uuid.uuid4().hex, status="success")
    second_batch = SysImportBatch(
        filename="second.xlsx", file_type="sales",
        file_hash=uuid.uuid4().hex, status="success")
    db.add_all([first_batch, second_batch])
    db.flush()
    db.add_all([
        FSalesOrder(
            raw_order_id=f"sales-ambig-a-{uuid.uuid4()}",
            order_no=contract_no,
            amount_ex_tax=Decimal("100.00"), tax_rate=Decimal("0.13"),
            data_status="已生效", import_batch_id=first_batch.id,
        ),
        FSalesOrder(
            raw_order_id=f"sales-ambig-b-{uuid.uuid4()}",
            order_no=contract_no,
            amount_ex_tax=Decimal("200.00"), tax_rate=Decimal("0.13"),
            data_status="已生效", import_batch_id=second_batch.id,
        ),
    ])
    db.commit()

    with pytest.raises(master.WorkbookError) as raised:
        master._ensure_contract_for_xsdd_apply(
            db,
            project=project,
            contract_no=contract_no,
            operated_by="fallback-test",
            assigned_xsdd_nos={contract_no},
        )
    assert raised.value.code == "sales_order_ambiguous"
    db.rollback()
    assert db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id,
        MaintenanceProjectContract.contract_no == contract_no,
    )) is None


def test_v24_overview_masks_known_contract_subtotal_when_incomplete(db):
    project, _part, _order, _line = _make_project_with_line(db)
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project.project_id,
        contract_id="EDIT-CONTRACT-MISSING",
        contract_no="XSDD-EDIT-MISSING",
        amount_inc_tax=None,
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="v1",
        effective_from=date(2026, 1, 1),
        source="ledger",
        version=1,
    ))
    db.commit()

    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    values = {
        row[0].value: row[1].value
        for row in workbook[master.V2_SHEET_OVERVIEW].iter_rows(
            min_col=1, max_col=2
        )
        if row[0].value
    }
    meta = {
        row[0].value: row[1].value
        for row in workbook[master.V2_SHEET_META].iter_rows(
            min_col=1, max_col=2
        )
    }

    assert values["合同总额（含税）"] == "—"
    assert values["成本率"] == "—"
    assert values["合同额口径"] == "合同事实不完整/存在共享冲突"
    assert meta["contract_total_exported"] is None
    assert meta["contract_editable"] == "false"


def test_v24_overview_unique_contract_total_roundtrips_to_inc_tax_with_audit(db):
    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_OVERVIEW]
    amount_cell = next(
        row[1] for row in ws.iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    assert str(amount_cell.value) == "10000.00"
    assert amount_cell.fill.fill_type == "solid"
    amount_cell.value = Decimal("12345.67")

    result = _reupload(db, project.project_id, wb)
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id,
    ))
    assert result["contract_updates"] == 1
    assert contract.amount_inc_tax == Decimal("12345.67")
    assert contract.contract_amount is None
    assert contract.version == 2
    assert contract.source == "project_master_workbook"
    audit = db.scalar(select(MaintenanceProjectAuditLog).where(
        MaintenanceProjectAuditLog.project_id == project.project_id,
        MaintenanceProjectAuditLog.entity_type == "project_contract",
        MaintenanceProjectAuditLog.action == "workbook_update_amount",
    ))
    assert audit is not None
    assert audit.before_json["amount_inc_tax"] == "10000.00"
    assert audit.after_json["amount_inc_tax"] == "12345.67"


def test_v24_overview_accepts_zero_as_a_real_contract_total(db):
    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_OVERVIEW]
    amount_cell = next(
        row[1] for row in ws.iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    amount_cell.value = 0

    _reupload(db, project.project_id, wb)
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id,
    ))
    assert contract.amount_inc_tax == Decimal("0.00")


def test_v24_unique_missing_contract_amount_stays_editable(db):
    """Fail-closed display does not remove the unique-contract repair path."""

    project, _part, _order, _line = _make_project_with_line(db)
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id,
    ))
    contract.amount_inc_tax = None
    db.commit()

    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook[master.V2_SHEET_OVERVIEW]
    amount_cell = next(
        row[1]
        for row in worksheet.iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    meta = {
        row[0].value: row[1].value
        for row in workbook[master.V2_SHEET_META].iter_rows(
            min_col=1, max_col=2
        )
    }
    assert amount_cell.value == "—"
    assert amount_cell.fill.fill_type == "solid"
    assert meta["contract_total_exported"] is None
    assert meta["contract_editable"] == "true"

    amount_cell.value = Decimal("12345.67")
    result = _reupload(db, project.project_id, workbook)

    db.refresh(contract)
    assert result["contract_updates"] == 1
    assert contract.amount_inc_tax == Decimal("12345.67")


def test_v24_stale_overview_rejects_contract_added_after_validation(db):
    """validate 后合同集合变化，apply 必须零写入而不是改旧合同。"""
    import pytest

    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_OVERVIEW]
    amount_cell = next(
        row[1] for row in ws.iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    amount_cell.value = Decimal("12345.67")
    data = _save(wb)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=data)

    original = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id,
    ))
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=project.project_id,
        contract_id="EDIT-CONTRACT-LATE", contract_no="XSDD-EDIT-LATE",
        amount_inc_tax=Decimal("5000.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1,
    ))
    db.commit()

    with pytest.raises(master.WorkbookError) as exc:
        master.apply_project_master_v2(
            db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    assert exc.value.code == "contract_total_ambiguous"
    db.rollback()
    db.refresh(original)
    assert original.amount_inc_tax == Decimal("10000.00")
    assert original.version == 1


def test_v24_overview_rejects_project_total_when_multiple_contracts_are_current(db):
    import pytest

    project, _part, _order, _line = _make_project_with_line(db)
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=project.project_id,
        contract_id="EDIT-CONTRACT-2", contract_no="XSDD-EDIT-002",
        amount_inc_tax=Decimal("5000.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1,
    ))
    db.commit()
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_OVERVIEW]
    amount_cell = next(
        row[1] for row in ws.iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    amount_cell.value = Decimal("16000.00")
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=buf.getvalue())
    assert exc.value.code == "contract_total_ambiguous"


def test_v24_overview_rejects_cross_project_shared_contract(db):
    import pytest

    project, _part, _order, _line = _make_project_with_line(db)
    other = MaintenanceProject(
        project_id=str(uuid.uuid4()), project_code=f"OTHER-{uuid.uuid4().hex[:8]}",
        display_name="共享合同的另一个项目", lifecycle_status="ongoing",
    )
    db.add(other)
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=other.project_id,
        # 内部稳定 ID 不同也不能绕过共享 XSDD 合同号门禁。
        contract_id="EDIT-CONTRACT-OTHER", contract_no="XSDD-EDIT-001",
        amount_inc_tax=Decimal("10000.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1,
    ))
    db.commit()

    content = master.build_project_master_v2(
        db, project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,),
    )
    wb = load_workbook(io.BytesIO(content))
    values = {
        row[0].value: row[1].value
        for row in wb[master.V2_SHEET_OVERVIEW].iter_rows(
            min_col=1, max_col=2
        )
        if row[0].value
    }
    meta = {
        row[0].value: row[1].value
        for row in wb[master.V2_SHEET_META].iter_rows(
            min_col=1, max_col=2
        )
    }
    assert values["合同总额（含税）"] == "—"
    assert values["成本率"] == "—"
    assert meta["contract_total_exported"] is None
    assert meta["contract_editable"] == "false"
    amount_cell = next(
        row[1] for row in wb[master.V2_SHEET_OVERVIEW].iter_rows(
            min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    amount_cell.value = Decimal("12000.00")

    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert exc.value.code == "contract_total_ambiguous"


def test_v24_change_pn_retires_old_manual_evidence_and_runs_authoritative_reprice(
    db, monkeypatch,
):
    project, _part, _order, line = _make_project_with_line(
        db, unit_cost=None, cost_source=None)
    replacement = DimPart(
        pn_std=f"REPRICE-{uuid.uuid4().hex[:8].upper()}",
        description="替换后的备件",
    )
    db.add(replacement)
    db.flush()
    override = MaintenanceManualCostOverride(
        line_id=line.id,
        unit_cost_ex_tax=Decimal("10.00"),
        unit_cost_inc_tax=Decimal("11.30"),
        reason="旧 PN 的人工证据",
        active=True,
        version=1,
        updated_by="tester",
    )
    db.add(override)
    db.commit()
    calls: list[tuple[bool, set[int] | None]] = []

    def fake_recompute(_db, *, commit=True, line_ids=None):
        calls.append((commit, line_ids))
        return {"lines_in_scope": 1}

    monkeypatch.setattr("app.services.maintenance_cost.recompute", fake_recompute)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=9, value=replacement.pn_std)
    _reupload(db, project.project_id, wb)

    db.refresh(line)
    db.refresh(override)
    assert line.part_id == replacement.id
    assert line.pn_std == replacement.pn_std
    assert override.active is False
    assert override.version == 2
    assert calls == [(False, {line.id})]


def test_v21_void_line_sets_inactive_and_excludes_from_export(db):
    project, _part, _order, line = _make_project_with_line(db)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    # row 2 = the exported line; set 操作=VOID
    ws.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, wb)

    db.refresh(line)
    assert line.is_active is False
    assert line.voided_by == "tester"
    assert line.voided_at is not None

    # 再导出：该行不再出现
    _content2, _wb2, ws2 = _parts_sheet(db, project.project_id)
    data_rows = [r for r in ws2.iter_rows(min_row=2, values_only=True)
                 if any(v not in (None, "") for v in r)]
    # 只剩空白新增行（无实体ID）——没有真实行
    assert all(r[21] in (None, "") for r in data_rows)


def test_v21_void_excluded_from_assigned_lines_and_bundle(db):
    project, _part, _order, line = _make_project_with_line(db)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, wb)

    rows = master._assigned_lines(db, project_id=project.project_id, window=None)
    assert rows == []
    # 看板成本：作废行不计入
    from app.services import maintenance_boss_board as board
    window = (date(2026, 1, 1), date(2026, 12, 31))
    bundle = board._cost_bundle(db, window=window, project_id=project.project_id,
                                can_cost=True)
    assert bundle["value"]["actual_amount"] == 0
    assert bundle["value"]["missing_lines"] == 0


def test_v21_change_quantity_recomputes_cost_amount(db):
    project, _part, _order, line = _make_project_with_line(
        db, qty=Decimal("2"), unit_cost=Decimal("100.00"))
    original_amount = line.cost_amount_inc_tax
    assert original_amount == Decimal("226.00")  # 2 × 113

    _content, wb, ws = _parts_sheet(db, project.project_id)
    # 需求数量 is column 11 (header index: 操作1, 维保单号2, 制单3, XSDD4, 需求类型5,
    # 仓库6, 销售7, 业务8, PN9, 描述10, 需求数量11)
    ws.cell(row=2, column=11, value=5)
    _reupload(db, project.project_id, wb)

    db.refresh(line)
    assert line.qty == Decimal("5.00")
    assert line.cost_amount_inc_tax == Decimal("565.00")  # 5 × 113
    assert line.cost_amount_ex_tax == Decimal("500.00")


def test_v24_second_manual_cost_edit_updates_the_materialized_line(db):
    project, _part, _order, line = _make_project_with_line(
        db, unit_cost=None, cost_source=None)

    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=18, value=100)
    ws.cell(row=2, column=19, value="首次补价")
    _reupload(db, project.project_id, wb)
    db.refresh(line)
    assert line.cost_source == "manual"
    assert line.unit_cost_ex_tax == Decimal("100.00")
    assert line.cost_amount_inc_tax == Decimal("226.00")

    _content2, wb2, ws2 = _parts_sheet(db, project.project_id)
    ws2.cell(row=2, column=18, value=150)
    ws2.cell(row=2, column=19, value="证据修正")
    _reupload(db, project.project_id, wb2)
    db.refresh(line)
    override = db.scalar(select(MaintenanceManualCostOverride).where(
        MaintenanceManualCostOverride.line_id == line.id,
        MaintenanceManualCostOverride.active.is_(True),
    ))
    assert override.version == 2
    assert override.unit_cost_ex_tax == Decimal("150.00")
    assert line.unit_cost_ex_tax == Decimal("150.00")
    assert line.unit_cost_inc_tax == Decimal("169.50")
    assert line.cost_amount_inc_tax == Decimal("339.00")


def test_v21_void_cascades_to_site_issue_line(db):
    project, part, order, line = _make_project_with_line(
        db, source_line_id="cascade-source-line-1")
    # 06 领用行通过 source_line_id 文本关联到该 03 行
    issue = MaintenanceSiteIssue(
        issue_id=str(uuid.uuid4()), project_id=project.project_id,
        issue_no="CKD-CASC-1", issue_date=date(2026, 8, 2),
        raw_status="已确认", status_mapping_state="mapped",
        normalized_status="confirmed", status_mapping_version="v1",
        source="legacy",
    )
    db.add(issue)
    db.flush()
    site_line = MaintenanceSiteIssueLine(
        issue_line_id=str(uuid.uuid4()), issue_id=issue.issue_id, line_no=1,
        part_id=part.id, pn=part.pn_std, quantity=Decimal("1"),
        source_order_id=order.raw_order_id, source_line_id=line.raw_line_id,
        algorithm_version="v1",
    )
    db.add(site_line)
    db.commit()

    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, wb)

    db.refresh(site_line)
    assert site_line.is_active is False
    # 06 导出不再包含级联作废的行
    rows = db.execute(
        select(MaintenanceSiteIssueLine).join(
            MaintenanceSiteIssue,
            MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id,
        ).where(MaintenanceSiteIssue.project_id == project.project_id,
                MaintenanceSiteIssueLine.is_active.is_(True))
    ).all()
    assert rows == []


def test_v21_void_does_not_cascade_same_source_line_id_in_another_project(db):
    project, _part, _order, line = _make_project_with_line(
        db, source_line_id="cross-project-source-line")
    other, other_part, other_order, _other_line = _make_project_with_line(db)
    foreign_issue = MaintenanceSiteIssue(
        issue_id=str(uuid.uuid4()), project_id=other.project_id,
        issue_no="CKD-FOREIGN", issue_date=date(2026, 8, 2),
        raw_status="已确认", status_mapping_state="mapped",
        normalized_status="confirmed", status_mapping_version="v1",
        source="legacy",
    )
    db.add(foreign_issue)
    db.flush()
    foreign_site_line = MaintenanceSiteIssueLine(
        issue_line_id=str(uuid.uuid4()), issue_id=foreign_issue.issue_id,
        line_no=1, part_id=other_part.id, pn=other_part.pn_std,
        quantity=Decimal("1"), source_order_id=other_order.raw_order_id,
        source_line_id=line.raw_line_id, algorithm_version="v1",
    )
    db.add(foreign_site_line)
    db.commit()

    _content, workbook, sheet = _parts_sheet(db, project.project_id)
    sheet.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, workbook)

    db.refresh(foreign_site_line)
    assert foreign_site_line.is_active is True


def test_v21_create_new_line_under_existing_order(db, monkeypatch):
    project, part, order, _line = _make_project_with_line(db)
    repriced: list[set[int]] = []

    def fake_recompute(_db, *, commit=True, line_ids=None):
        assert commit is False
        repriced.append(set(line_ids or ()))
        return {"lines_in_scope": len(line_ids or ())}

    monkeypatch.setattr("app.services.maintenance_cost.recompute", fake_recompute)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    # 找一个空白新增行（row 3）填入 CREATE
    ws.cell(row=3, column=1, value="CREATE")
    ws.cell(row=3, column=2, value=order.order_no)   # 维保单号
    ws.cell(row=3, column=9, value=part.pn_std)      # PN
    ws.cell(row=3, column=11, value=3)               # 需求数量
    ws.cell(row=3, column=25, value="新增行备注")     # 备注（1-based：第 25 列）
    _reupload(db, project.project_id, wb)

    new_line = db.scalar(
        select(FMaintenanceLine).where(
            FMaintenanceLine.order_id == order.id,
            FMaintenanceLine.edited_source == "workbook_manual",
        ))
    assert new_line is not None
    assert new_line.qty == Decimal("3.00")
    assert new_line.raw_line_id.startswith("manual-line:")
    assert new_line.line_note == "新增行备注"
    assert new_line.is_active is True
    assert repriced == [{new_line.id}]
    # 审计写了 CREATE
    audit = db.scalar(select(MaintenanceProjectOperationAudit).where(
        MaintenanceProjectOperationAudit.entity_id == str(new_line.id),
        MaintenanceProjectOperationAudit.action == "CREATE"))
    assert audit is not None


def test_v23_blank_manual_rows_reassign_existing_wbdd_without_duplicates(db):
    """项目内人工回传是归属确认：复用全局行、保留旧挂靠历史、重复上传幂等。"""
    old_project, part, order, line = _make_project_with_line(db)
    target = MaintenanceProject(
        project_id=str(uuid.uuid4()),
        project_code=f"TARGET-{uuid.uuid4().hex[:8]}",
        display_name="人工认证后的正确项目",
        lifecycle_status="ongoing",
    )
    db.add(target)
    order.project_raw = target.display_name
    order.project_std = target.display_name
    db.commit()

    content = master.build_project_master_v2(
        db, project_id=target.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    row_no = ws.max_row + 1
    ws.cell(row_no, headers["维保单号"], order.order_no)
    ws.cell(row_no, headers["PN"], part.pn_std)
    ws.cell(row_no, headers["描述"], line.description)
    ws.cell(row_no, headers["需求数量"], float(line.qty))
    ws.cell(row_no, headers["退货数量"], float(line.return_qty or 0))

    data = _save(wb)
    ctx = UserContext(user_id="assignment-admin", role="admin", is_authenticated=True)
    plan = master.validate_project_master_v2(
        db, project_id=target.project_id, data=data, user_ctx=ctx)
    assert plan.summary["order_reassignments"] == 1
    assert plan.summary["line_creates"] == 0
    assert plan.summary["line_updates"] == 0
    assert plan.assignment_changes[0].previous_project_id == old_project.project_id

    line_count = int(db.scalar(select(func.count(FMaintenanceLine.id))) or 0)
    master.apply_project_master_v2(
        db, plan, operated_by="assignment-admin",
        import_batch_id=str(uuid.uuid4()), user_ctx=ctx)

    assert int(db.scalar(select(func.count(FMaintenanceLine.id))) or 0) == line_count
    history = list(db.scalars(
        select(MaintenanceSourceOrderAssignment)
        .where(MaintenanceSourceOrderAssignment.source_order_id == order.raw_order_id)
        .order_by(MaintenanceSourceOrderAssignment.created_at)
    ).all())
    assert len(history) == 2
    assert sum(row.is_active for row in history) == 1
    assert next(row for row in history if row.is_active).project_id == target.project_id
    assert next(row for row in history if not row.is_active).project_id == old_project.project_id

    # 重放同一份人工表：既有全局行已在目标项目，既不再改挂，也不重复建行。
    replay = master.validate_project_master_v2(
        db, project_id=target.project_id, data=data, user_ctx=ctx)
    assert replay.summary["order_reassignments"] == 0
    assert replay.summary["line_creates"] == 0
    assert replay.summary["line_updates"] == 0
    master.apply_project_master_v2(
        db, replay, operated_by="assignment-admin",
        import_batch_id=str(uuid.uuid4()), user_ctx=ctx)
    assert int(db.scalar(select(func.count(FMaintenanceLine.id))) or 0) == line_count


def test_v23_multi_order_cascade_writes_one_bounded_audit_per_order(db):
    """多张 WBDD 同时整单作废时，审计 entity_id 不再逗号拼接溢出 varchar(64)。"""
    project, part, first_order, _first_line = _make_project_with_line(db)
    second_order = FMaintenanceOrder(
        raw_order_id=f"raw-order-{uuid.uuid4()}",
        order_no="WBDD-EDIT-002",
        order_date=date(2026, 8, 2),
        linked_sales_order_no="XSDD-EDIT-001",
        project_raw=project.display_name,
        project_std=project.display_name,
        data_status="已生效",
        import_batch_id=_batch(db),
    )
    db.add(second_order)
    db.flush()
    second_line = FMaintenanceLine(
        raw_line_id=f"raw-line-{uuid.uuid4()}",
        order_id=second_order.id,
        line_no=1,
        part_id=part.id,
        pn_std=part.pn_std,
        pn_raw=part.pn_std,
        description=part.description,
        qty=Decimal("1"),
        return_qty=Decimal("0"),
        is_active=True,
        import_batch_id=_batch(db),
    )
    db.add_all([
        second_line,
        MaintenanceSourceOrderAssignment(
            assignment_id=str(uuid.uuid4()),
            project_id=project.project_id,
            source_order_id=second_order.raw_order_id,
            is_active=True,
            created_by="test",
        ),
    ])
    db.commit()

    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    entity_col = next(cell.column for cell in ws[1] if cell.value == "实体ID")
    for row_no in range(ws.max_row, 1, -1):
        if ws.cell(row_no, entity_col).value not in (None, ""):
            ws.delete_rows(row_no)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert plan.summary["line_voids"] == 2
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))

    order_audits = list(db.scalars(
        select(MaintenanceProjectOperationAudit).where(
            MaintenanceProjectOperationAudit.project_id == project.project_id,
            MaintenanceProjectOperationAudit.entity_type == "maintenance_order",
            MaintenanceProjectOperationAudit.action == "VOID",
        )
    ).all())
    assert {row.entity_id for row in order_audits} == {
        first_order.raw_order_id,
        second_order.raw_order_id,
    }
    assert all(len(row.entity_id) <= 64 for row in order_audits)


def test_v21_editing_locked_column_is_rejected(db):
    project, _part, _order, _line = _make_project_with_line(db)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    # XSDD is column 4 — locked
    ws.cell(row=2, column=4, value="XSDD-TAMPERED")
    import pytest
    from app.services.maintenance_expense_collection_workbook import WorkbookError
    with pytest.raises(WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id,
            data=_save(wb))
    assert exc.value.code == "readonly_cell_modified"


def test_v21_audit_written_for_update_and_void(db):
    project, _part, _order, line = _make_project_with_line(db)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=11, value=4)  # 改数量
    _reupload(db, project.project_id, wb)

    audits = db.scalars(select(MaintenanceProjectOperationAudit).where(
        MaintenanceProjectOperationAudit.entity_type == "maintenance_line",
        MaintenanceProjectOperationAudit.entity_id == str(line.id))).all()
    actions = {a.action for a in audits}
    assert "UPDATE" in actions

    # 再作废
    _content2, wb2, ws2 = _parts_sheet(db, project.project_id)
    ws2.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, wb2)
    audits2 = db.scalars(select(MaintenanceProjectOperationAudit).where(
        MaintenanceProjectOperationAudit.entity_type == "maintenance_line",
        MaintenanceProjectOperationAudit.entity_id == str(line.id))).all()
    assert {a.action for a in audits2} >= {"UPDATE", "VOID"}


def test_v21_unchanged_reupload_is_idempotent(db):
    """原样回传下载的工作簿：不应写 UPDATE 审计、不应改 edited_source、不应重算金额。
    回归：导出总是带上数量/人工成本列，早期实现未与现值 diff，导致原样上传把所有
    行标成 workbook_manual 并写假审计。"""
    project, _part, _order, line = _make_project_with_line(db)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    original_amount = line.cost_amount_inc_tax
    _reupload(db, project.project_id, wb)

    db.refresh(line)
    assert line.edited_source == "wbdd"
    assert line.cost_amount_inc_tax == original_amount
    audits = db.scalars(select(MaintenanceProjectOperationAudit).where(
        MaintenanceProjectOperationAudit.entity_type == "maintenance_line",
        MaintenanceProjectOperationAudit.entity_id == str(line.id))).all()
    assert audits == []


def test_v21_cannot_edit_already_voided_line(db):
    """已作废行再次出现在旧工作簿里被改并上传 → 报 line_not_found（哈希/实体失效）。"""
    import pytest
    from app.services.maintenance_expense_collection_workbook import WorkbookError
    project, _part, _order, line = _make_project_with_line(db)
    # 先作废
    _c, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, wb)
    db.refresh(line)
    assert line.is_active is False

    # 用另一份（作废前导出的）工作簿改数量后上传：实体已作废应拒绝
    _c2, wb2, ws2 = _parts_sheet(db, project.project_id)
    # 作废后导出已不含该行；手工在第 2 行塞一条带旧实体ID 的改单不可行（行已消失），
    # 因此直接校验：对已作废行 db.get 走解析会报 line_not_found。
    from app.services import maintenance_project_master_workbook as master
    # 构造一个带已作废行实体ID 的最小工作簿
    wb3 = load_workbook(io.BytesIO(_c2))
    ws3 = wb3[master.V2_SHEET_PARTS]
    ws3.cell(row=2, column=1, value="UPDATE")
    ws3.cell(row=2, column=22, value=line.id)  # 实体ID（openpyxl 1-based；22=实体ID）
    ws3.cell(row=2, column=11, value=9)        # 需求数量
    with pytest.raises(WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb3))
    assert exc.value.code == "line_not_found"


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------
# 2026-08-19（#264/#267）：04 报销作废（显式 VOID + 缺行=作废）、防呆、
# void-fast 一键批量作废、latest/missing 差异清单。
# ----------------------------------------------------------------------

def _attribute_expense(db, project, expense):
    db.add(MaintenanceProjectExpenseAttribution(
        expense_id=f"bxd:{expense.raw_line_id}",
        project_id=project.project_id,
        project_contract_id=db.scalar(select(
            MaintenanceProjectContract.project_contract_id
        ).where(MaintenanceProjectContract.project_id == project.project_id)),
        raw_expense_line_id=expense.raw_line_id,
        expense_ref=f"{expense.bxd_no}#{expense.line_no}",
        expense_date=expense.expense_date,
        applicant=expense.person,
        category=expense.fee_category or expense.expense_type,
        expense_reason=expense.reason,
        tax_basis=expense.tax_basis or "default_ex",
        amount_ex_tax=expense.amount_ex_tax,
        amount_inc_tax=expense.amount_inc_tax,
        tax_rate_used=Decimal("0.13"),
        raw_status=expense.data_status or "已生效",
        status_mapping_state="mapped",
        normalized_status="approved",
        status_mapping_version="synthetic-v2.5",
        ownership_mapping_state="mapped",
        ownership_mapping_version="ownership-v1",
        version=1,
    ))


def _make_project_with_expense(db, *, data_status=None, raw_line_id=None):
    project, _part, order, line = _make_project_with_line(db)
    expense = FProjectExpense(
        raw_line_id=raw_line_id or f"bxd-{uuid.uuid4()}",
        import_batch_id=_batch(db),
        bxd_no="BXD-EDIT-001", line_no=1,
        expense_date=date(2026, 8, 3), person="测试员",
        linked_sales_order_no="XSDD-EDIT-001",
        amount=Decimal("500.00"), tax_basis="ex",
        amount_ex_tax=Decimal("500.00"),
        amount_inc_tax=Decimal("565.00"),
        data_status=data_status,
    )
    db.add(expense)
    _attribute_expense(db, project, expense)
    db.commit()
    return project, order, line, expense


def test_expense_update_syncs_canonical_once_with_one_version_bump(db):
    from app.services import maintenance_expense_collection_workbook as ec

    project, _order, _line, expense = _make_project_with_expense(db)
    attribution = db.get(
        MaintenanceProjectExpenseAttribution, f"bxd:{expense.raw_line_id}"
    )
    before_version = attribution.version
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,)
    )
    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook[master.V2_SHEET_EXPENSE]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    row_no = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, headers["实体ID"]).value == expense.raw_line_id
    )
    worksheet.cell(row_no, headers["操作"], "UPDATE")
    worksheet.cell(row_no, headers["未税金额"], 600)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(workbook)
    )
    master.apply_project_master_v2(
        db, plan, operated_by="expense-once-test",
        import_batch_id=str(uuid.uuid4()),
    )
    db.flush()
    db.refresh(expense)
    db.refresh(attribution)

    assert expense.amount_ex_tax == Decimal("600.00")
    assert attribution.amount_ex_tax == Decimal("600.00")
    assert attribution.amount_inc_tax == Decimal("678.00")
    assert attribution.version == before_version + 1
    assert attribution.status_mapping_version == ec.PROTOCOL_VERSION


def test_v2_expense_negative_reversal_preserves_inc_tax_source_basis(db):
    project, _order, _line, expense = _make_project_with_expense(db)
    attribution = db.get(
        MaintenanceProjectExpenseAttribution, f"bxd:{expense.raw_line_id}"
    )
    expense.tax_basis = "inc"
    expense.amount = Decimal("565.00")
    attribution.tax_basis = "inc"
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,)
    )
    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook[master.V2_SHEET_EXPENSE]
    headers = {cell.value: cell.column for cell in worksheet[1]}
    row_no = next(
        row
        for row in range(2, worksheet.max_row + 1)
        if worksheet.cell(row, headers["实体ID"]).value == expense.raw_line_id
    )
    worksheet.cell(row_no, headers["操作"], "UPDATE")
    worksheet.cell(row_no, headers["未税金额"], -200)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(workbook)
    )
    master.apply_project_master_v2(
        db,
        plan,
        operated_by="negative-reversal-test",
        import_batch_id=str(uuid.uuid4()),
    )

    db.refresh(expense)
    db.refresh(attribution)
    assert expense.tax_basis == "inc"
    assert expense.amount == Decimal("-226.00")
    assert expense.amount_ex_tax == Decimal("-200.00")
    assert expense.amount_inc_tax == Decimal("-226.00")
    assert attribution.tax_basis == "inc"
    assert attribution.amount_ex_tax == Decimal("-200.00")
    assert attribution.amount_inc_tax == Decimal("-226.00")


def test_v21_expense_missing_row_becomes_void(db):
    """缺行=作废（用户核心诉求）：下载→删掉报销行→回传→行从系统消失。

    两行删一行恰好等于 50%（不触发防呆）；删光全部行则被 50% 防呆整本拒绝。
    """
    import pytest
    from app.services.maintenance_expense_collection_workbook import WorkbookError
    project, _order, _line, expense = _make_project_with_expense(db)
    keeper = FProjectExpense(
        raw_line_id=f"bxd-{uuid.uuid4()}",
        bxd_no="BXD-EDIT-001", line_no=2,
        expense_date=date(2026, 8, 4), person="测试员",
        linked_sales_order_no="XSDD-EDIT-001",
        amount=Decimal("300.00"), tax_basis="ex",
        amount_ex_tax=Decimal("300.00"),
        amount_inc_tax=Decimal("339.00"),
        import_batch_id=_batch(db),
    )
    db.add(keeper)
    _attribute_expense(db, project, keeper)
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_EXPENSE]
    # 2026-08-22 撤防呆：删光全部 → 放行为全量作废（用户拍板大批量修改期）
    wb_all_deleted = load_workbook(io.BytesIO(content))
    wb_all_deleted[master.V2_SHEET_EXPENSE].delete_rows(2, 2)
    buf_all = io.BytesIO()
    wb_all_deleted.save(buf_all)
    plan_all = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf_all.getvalue())
    assert all(r.operation == "VOID" for r in plan_all.expense_updates)
    # 正常路径：只删第一行（=50%，放行）
    ws.delete_rows(2)
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert expense.raw_line_id in plan.expense_voids
    assert any(r["sheet"] == "04_费用报销" and r["entity_id"] == expense.raw_line_id
               for r in plan.will_void_rows)
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))

    db.refresh(expense)
    assert expense.data_status == "已作废"
    # 再导出：作废行彻底不出现（用户拍板：干脆不导出）
    content2 = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,))
    ws2 = load_workbook(io.BytesIO(content2))[master.V2_SHEET_EXPENSE]
    rows = [r for r in ws2.iter_rows(min_row=2, values_only=True)
            if any(v not in (None, "") for v in r)]
    assert all(r[17] != expense.raw_line_id for r in rows)  # 实体ID 列（1-based→idx 17）


def test_v21_expense_explicit_void_operation(db):
    project, _order, _line, expense = _make_project_with_expense(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,))
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_EXPENSE].cell(row=2, column=1, value="VOID")
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.expense_voids == (expense.raw_line_id,)
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.refresh(expense)
    assert expense.data_status == "已作废"


def test_v21_row_loss_guard_removed_all_rows_voidable(db):
    """2026-08-22 撤防呆：≥2 行的表删光全部数据行 → 全量 VOID 放行（用户拍板）。"""
    import pytest
    from app.services.maintenance_expense_collection_workbook import WorkbookError

    project, part, _order, _line = _make_project_with_line(db)
    second = FMaintenanceLine(
        raw_line_id=f"raw-line-{uuid.uuid4()}",
        order_id=_order.id, line_no=2, part_id=part.id,
        pn_std=part.pn_std, pn_raw=part.pn_std, description=part.description,
        qty=Decimal("1"), return_qty=Decimal("0"), import_batch_id=_batch(db),
    )
    db.add(second)
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_PARTS].delete_rows(2, 2)  # 删光 2 条数据行 → 全量 VOID
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert [r.operation for r in plan.cost_refills].count("VOID") == 2


def test_void_fast_tombstones_deactivates_assignment(db):
    from app.services import maintenance_demands

    project, _part, order, _line = _make_project_with_line(db)
    result = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="氚云已删除", operated_by="tester")
    assert result["voided"] == 1 and result["already_voided"] == 0

    assignment = db.scalar(select(MaintenanceSourceOrderAssignment).where(
        MaintenanceSourceOrderAssignment.source_order_id == order.raw_order_id))
    assert assignment.is_active is False
    tombstone = db.get(MaintenanceDemandTombstone, order.raw_order_id)
    assert tombstone is not None and tombstone.restored_at is None
    # 总表行随即消失（读侧联动）
    assert master._assigned_lines(db, project_id=project.project_id, window=None) == []


def test_void_fast_unknown_order_rejects_whole_batch(db):
    import pytest
    from app.services import maintenance_demands

    project, _part, order, _line = _make_project_with_line(db)
    with pytest.raises(maintenance_demands.MaintenanceDemandNotFound):
        maintenance_demands.void_fast(
            db, source_order_ids=[order.raw_order_id, "raw-order-does-not-exist"],
            reason="测试", operated_by="tester")
    # 整批零写入：已知单也没有墓碑
    assert db.get(MaintenanceDemandTombstone, order.raw_order_id) is None


def test_void_fast_already_tombstoned_is_idempotent(db):
    from app.services import maintenance_demands

    _project, _part, order, _line = _make_project_with_line(db)
    first = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="第一次", operated_by="tester")
    second = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="重复点击", operated_by="tester")
    assert first["voided"] == 1
    assert second["already_voided"] == 1 and second["voided"] == 0


def test_latest_missing_reports_disappeared_orders(db):
    from app.models.maintenance_wbdd_import import MaintenanceWbddImportReceipt
    from app.services import maintenance_wbdd_import as wbdd

    project, _part, order, _line = _make_project_with_line(db)
    # 用 order 所在批次造一张回执：文件单集 = 该批次全部单（含 order）
    batch_id = order.import_batch_id
    db.add(MaintenanceWbddImportReceipt(
        batch_id=batch_id, idempotency_key="test-key-0001",
        uploaded_by="tester", file_hash=uuid.uuid4().hex,
        report_json={"batch_id": batch_id},
    ))
    db.commit()
    result = wbdd.latest_missing(db)
    assert result["readiness"] == "ready"
    # 文件里存在 order → 不在缺失清单
    assert all(m["source_order_id"] != order.raw_order_id
               for m in result["missing_orders"])

    # 再造一张挂在别的批次的新单（库里有、回执批次文件里没有）→ 进缺失清单
    other_batch = _batch(db)
    ghost = FMaintenanceOrder(
        raw_order_id=f"raw-ghost-{uuid.uuid4()}", order_no="WBDD-GHOST-001",
        order_date=order.order_date, linked_sales_order_no="XSDD-EDIT-001",
        project_raw="可编辑测试", data_status="已生效",
        import_batch_id=other_batch,
    )
    db.add(ghost)
    db.commit()
    result2 = wbdd.latest_missing(db)
    assert any(m["source_order_id"] == ghost.raw_order_id
               for m in result2["missing_orders"])
    entry = next(m for m in result2["missing_orders"]
                 if m["source_order_id"] == ghost.raw_order_id)
    assert entry["order_no"] == "WBDD-GHOST-001"
    assert entry["line_count"] == 0


def test_void_fast_returns_per_order_results(db):
    """#265 冻结契约：逐单 results（幂等命中 already_voided 不算错误）。"""
    from app.services import maintenance_demands

    _project, _part, order, _line = _make_project_with_line(db)
    result = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="契约测试", operated_by="tester")
    assert result["results"] == [
        {"source_order_id": order.raw_order_id,
         "order_no": order.order_no, "status": "voided"},
    ]
    again = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="重复", operated_by="tester")
    assert again["results"][0]["status"] == "already_voided"


def test_search_include_voided_view(db):
    """默认视图不见已作废单；include_voided=True 可见且带 is_voided 标记。"""
    from app.services import maintenance_demands

    _project, _part, order, _line = _make_project_with_line(db)
    maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="测试视图", operated_by="tester")

    active = maintenance_demands.search_demands(
        db, q=None, page=1, page_size=20)
    assert all(item["source_order_id"] != order.raw_order_id
               for item in active["items"])

    with_voided = maintenance_demands.search_demands(
        db, q=None, page=1, page_size=20, include_voided=True)
    entry = next(item for item in with_voided["items"]
                 if item["source_order_id"] == order.raw_order_id)
    assert entry["is_voided"] is True
    assert entry["order_no"] == order.order_no


def test_review_p2_void_fast_idempotent_replay(db):
    """同幂等键同请求重放 → 返回首次结果；同键异请求 → 冲突。"""
    import pytest
    from app.services import maintenance_demands

    _project, _part, order, _line = _make_project_with_line(db)
    first = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="重放测试", operated_by="tester", idempotency_key="replay-key-01")
    db.commit()
    replay = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="重放测试", operated_by="tester", idempotency_key="replay-key-01")
    assert replay["replayed"] is True
    assert replay["voided"] == first["voided"] == 1
    with pytest.raises(maintenance_demands.DeleteIntentConflict):
        maintenance_demands.void_fast(
            db, source_order_ids=[order.raw_order_id],
            reason="不同的请求", operated_by="tester", idempotency_key="replay-key-01")


def test_void_fast_bumps_owner_workbook_revision_once(db):
    """OCC 写者失效：实际作废给归属项目 +1；同键重放与 already_voided 都 +0。"""
    from app.services import maintenance_demands

    project, _part, order, _line = _make_project_with_line(db)
    operations.get_or_create_workbook_state(db, project_id=project.project_id)
    db.commit()
    revision_before = db.get(
        MaintenanceProjectWorkbookState, project.project_id).revision

    result = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="氚云已删除", operated_by="tester",
        idempotency_key="void-occ-key-01")
    assert result["voided"] == 1 and result["already_voided"] == 0
    db.expire_all()
    state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    assert state.revision == revision_before + 1
    data_version_after_void = state.data_version

    # 同键同请求重放：返回首次结果，revision/data_version 不再动（+0）
    replay = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="氚云已删除", operated_by="tester",
        idempotency_key="void-occ-key-01")
    assert replay["replayed"] is True
    db.expire_all()
    state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    assert state.revision == revision_before + 1
    assert state.data_version == data_version_after_void

    # 无键重复点击：already_voided 幂等放行，同样 +0
    again = maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="重复点击", operated_by="tester")
    assert again["already_voided"] == 1 and again["voided"] == 0
    db.expire_all()
    state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    assert state.revision == revision_before + 1
    assert state.data_version == data_version_after_void


def test_cascade_tombstone_validates_prelocked_states_and_bumps_owner(db):
    """级联墓碑：_prelocked_states 缺 owner → Conflict 零写；覆盖时复用并 +1。"""
    from app.services import maintenance_demands

    project, _part, order, line = _make_project_with_line(db)
    line.is_active = False  # 活动行归零 → 满足级联条件
    db.commit()

    # 预锁集合缺 owner → 拒绝，且不允许持锁后补拿新 state（零写）
    with pytest.raises(maintenance_demands.DeleteIntentConflict):
        maintenance_demands.cascade_tombstone_orders(
            db, source_order_ids=[order.raw_order_id],
            operated_by="tester", reason="行全部作废级联",
            _prelocked_states={})
    assert db.get(MaintenanceDemandTombstone, order.raw_order_id) is None

    states = operations.lock_workbook_states(
        db, project_ids={project.project_id})
    revision_before = states[project.project_id].revision
    cascaded = maintenance_demands.cascade_tombstone_orders(
        db, source_order_ids=[order.raw_order_id],
        operated_by="tester", reason="行全部作废级联",
        _prelocked_states=states)
    assert cascaded == [order.raw_order_id]
    db.expire_all()
    assert db.get(
        MaintenanceProjectWorkbookState, project.project_id
    ).revision == revision_before + 1


def test_restore_fails_closed_on_dirty_active_assignment(db):
    """restore 不复活挂靠（正常 +0）；历史脏 active 挂靠 → fail closed 零写。"""
    from app.services import maintenance_demands

    project, _part, order, _line = _make_project_with_line(db)
    maintenance_demands.void_fast(
        db, source_order_ids=[order.raw_order_id],
        reason="作废", operated_by="tester")
    db.expire_all()
    revision_after_void = db.get(
        MaintenanceProjectWorkbookState, project.project_id).revision
    # 构造历史脏数据：已墓碑单上重新出现 active 挂靠
    db.add(MaintenanceSourceOrderAssignment(
        assignment_id=str(uuid.uuid4()), project_id=project.project_id,
        source_order_id=order.raw_order_id, is_active=True,
        created_by="dirty",
    ))
    db.commit()

    with pytest.raises(maintenance_demands.MaintenanceDemandError):
        maintenance_demands.restore_demand(
            db, source_order_id=order.raw_order_id,
            reason="尝试恢复", operated_by="tester")
    tombstone = db.get(MaintenanceDemandTombstone, order.raw_order_id)
    assert tombstone.restored_at is None
    db.expire_all()
    assert db.get(
        MaintenanceProjectWorkbookState, project.project_id
    ).revision == revision_after_void


def test_review_p1_cross_project_line_is_rejected(db):
    """拿别项目的实体ID 在本项目入口上传 → line_not_in_project 拒绝。"""
    import pytest
    from app.services.maintenance_expense_collection_workbook import WorkbookError

    project_a, _p, _o, line_a = _make_project_with_line(db)
    project_b, _p2, _o2, _line_b = _make_project_with_line(db)
    _content, wb, _ws = _parts_sheet(db, project_b.project_id)
    ws = wb[master.V2_SHEET_PARTS]
    # 把 A 项目的实体ID 塞进 B 项目的空白新增行并标 UPDATE
    ws.cell(row=2, column=1, value="UPDATE")
    ws.cell(row=2, column=22, value=line_a.id)
    ws.cell(row=2, column=11, value=9)
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project_b.project_id, data=buf.getvalue())
    assert exc.value.code in ("line_not_in_project", "readonly_cell_modified", "line_not_found")


def test_review_p1_pn_change_updates_part_id(db):
    """改 PN 必须同步换 part_id，否则库存/成本 join 腐化。"""
    project, part_old, _order, line = _make_project_with_line(db)
    part_new = DimPart(pn_std="EDIT-PN-002", description="新备件")
    db.add(part_new)
    db.commit()
    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=9, value=part_new.pn_std)  # PN
    _reupload(db, project.project_id, wb)
    db.refresh(line)
    assert line.pn_std == part_new.pn_std
    assert line.part_id == part_new.id


def test_review_p1_late_imported_expense_is_not_voided(db):
    """下载后新导入的报销行不在导出全集里 → 缺行判定不得误杀；
    用户真删的行（导出时存在）照常作废（P1，Codex review #272）。"""
    project, _order, _line, expense = _make_project_with_expense(db)
    keeper = FProjectExpense(
        raw_line_id=f"bxd-{uuid.uuid4()}",
        bxd_no="BXD-EDIT-001", line_no=2,
        expense_date=date(2026, 8, 4), person="测试员",
        linked_sales_order_no="XSDD-EDIT-001",
        amount=Decimal("300.00"), tax_basis="ex",
        amount_ex_tax=Decimal("300.00"), amount_inc_tax=Decimal("339.00"),
        import_batch_id=_batch(db),
    )
    db.add(keeper)
    _attribute_expense(db, project, keeper)
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,))
    # 下载之后、上传之前：新导入一行报销（不在导出全集）
    late = FProjectExpense(
        raw_line_id=f"bxd-late-{uuid.uuid4()}",
        bxd_no="BXD-LATE-001", line_no=1,
        expense_date=date(2026, 8, 5), person="迟到导入",
        linked_sales_order_no="XSDD-EDIT-001",
        amount=Decimal("100.00"), tax_basis="ex",
        amount_ex_tax=Decimal("100.00"), amount_inc_tax=Decimal("113.00"),
        import_batch_id=_batch(db),
    )
    db.add(late)
    _attribute_expense(db, project, late)
    db.commit()
    # 用户在旧文件里删掉导出时存在的那行
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_EXPENSE].delete_rows(2)
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.expense_voids == (expense.raw_line_id,)
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.refresh(expense)
    db.refresh(late)
    assert expense.data_status == "已作废"
    assert late.data_status != "已作废"


def test_v22_parts_missing_row_becomes_void(db):
    """03 删行=作废（与 04 同语义，2026-08-20 用户拍板）。
    两行删一行＝50%，不触发防呆；唯一行被删会被防呆拦（契约行为）。"""
    project, part, order, line = _make_project_with_line(db)
    second = FMaintenanceLine(
        raw_line_id=f"raw-line-{uuid.uuid4()}",
        order_id=order.id, line_no=2, part_id=part.id,
        pn_std=part.pn_std, pn_raw=part.pn_std, description=part.description,
        qty=Decimal("1"), return_qty=Decimal("0"), import_batch_id=_batch(db),
    )
    db.add(second)
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_PARTS].delete_rows(2)  # 删第 2 行（第一条明细）
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert any(r.operation == "VOID" and r.line_id == line.id for r in plan.cost_refills)
    assert any(r["sheet"] == "03_备件明细" and r.get("reason") == "上传文件缺行"
               for r in plan.will_void_rows)
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.refresh(line)
    assert line.is_active is False


def test_v22_void_all_lines_cascades_order_tombstone(db):
    """整单行全作废 → 级联整单墓碑：搜索消失、挂靠停用、重传不复活。"""
    from app.services import maintenance_demands

    project, _part, order, line = _make_project_with_line(db)
    _content, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=1, value="VOID")
    _reupload(db, project.project_id, wb)

    db.refresh(line)
    assert line.is_active is False
    tombstone = db.get(MaintenanceDemandTombstone, order.raw_order_id)
    assert tombstone is not None and tombstone.restored_at is None
    assignment = db.scalar(select(MaintenanceSourceOrderAssignment).where(
        MaintenanceSourceOrderAssignment.source_order_id == order.raw_order_id))
    assert assignment.is_active is False
    # 默认搜索不见
    result = maintenance_demands.search_demands(db, q=None, page=1, page_size=20)
    assert all(item["source_order_id"] != order.raw_order_id
               for item in result["items"])


def test_v22_template_has_usage_sheet_dropdown_and_yellow_editable(db):
    project, _p, _o, _l = _make_project_with_line(db)
    content = master.build_project_master_v2(db, project_id=project.project_id)
    wb = load_workbook(io.BytesIO(content))
    # 使用说明 sheet
    assert master.V2_SHEET_USAGE in wb.sheetnames
    usage_texts = "\n".join(str(c.value) for row in wb[master.V2_SHEET_USAGE].iter_rows()
                            for c in row if c.value)
    assert "黄底" in usage_texts and "VOID" in usage_texts
    # 操作列下拉
    ws = wb[master.V2_SHEET_PARTS]
    dvs = list(ws.data_validations.dataValidation)
    assert any(dv.type == "list" and "VOID" in (dv.formula1 or "") for dv in dvs)
    # 可编辑数据区黄底（PN 列=9，第 2 行）
    from openpyxl.styles import PatternFill
    fill = ws.cell(row=2, column=9).fill
    assert fill.start_color.rgb in ("00FFE699", "FFFFE699") or fill.fgColor.rgb == "00FFE699"
    # 合同总额可编辑协议升级后的模板版本
    meta = {r[0].value: r[1].value for r in wb[master.V2_SHEET_META].iter_rows(min_col=1, max_col=2)}
    assert meta["template_version"] == "2.6.0"


def test_latest_missing_allows_full_sync_at_any_ratio(db):
    """2026-08-24 用户拍板：缺失占比任意（含 >50%/100%）都不再拦截——
    差异清单与批量作废解锁，可全量跟随修改；占比仅作展示。"""
    from app.models.maintenance_wbdd_import import MaintenanceWbddImportReceipt
    from app.services import maintenance_wbdd_import as wbdd

    project, _part, order, _line = _make_project_with_line(db)
    # 再造 4 张同窗口、不在文件批次里的单 → 库内 5 张活跃，文件只有 1 张（80%）
    others = []
    for i in range(4):
        b = _batch(db)
        o = FMaintenanceOrder(
            raw_order_id=f"raw-ghost-{i}-{uuid.uuid4()}", order_no=f"WBDD-GHOST-{i}",
            order_date=order.order_date, linked_sales_order_no="XSDD-EDIT-001",
            project_raw="可编辑测试", data_status="已生效", import_batch_id=b,
        )
        others.append(o)
        db.add(o)
    db.add(MaintenanceWbddImportReceipt(
        batch_id=order.import_batch_id, idempotency_key="suspicious-key-0001",
        uploaded_by="tester", file_hash=uuid.uuid4().hex,
        report_json={"batch_id": order.import_batch_id},
    ))
    db.commit()
    result = wbdd.latest_missing(db)
    assert result["missing_count"] == 4
    assert result["db_active_in_window"] == 5
    assert "suspicious" not in result
    assert result["missing_ratio"] > 0.5
    assert len(result["missing_orders"]) == 4


def test_e2e_fix_single_row_delete_allowed(db):
    """E2E #1：单行表删除唯一一行不再被防呆拦截（防呆只拦批量损失）。"""
    project, _order, _line, _expense = _make_project_with_expense(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_EXPENSE,))
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_EXPENSE].delete_rows(2)  # 唯一一行
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.expense_voids == (_expense.raw_line_id,)
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.refresh(_expense)
    assert _expense.data_status == "已作废"


def _project_with_receipt(db):
    project, _part, _order, _line = _make_project_with_line(db)
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id
    ))
    snapshot = MaintenanceCollectionSnapshot(
        collection_id=str(uuid.uuid4()),
        project_id=project.project_id,
        project_contract_id=contract.project_contract_id,
        report_month=date(2026, 6, 1),
        cumulative_amount=Decimal("82325.40"),
        status="confirmed",
        receipt_reference="SKD-20260630-0007",
        remark=None,
        source="direct_api",
        import_batch_id=None,
        version=1,
    )
    db.add(snapshot)
    db.commit()
    return project, contract, snapshot


def test_receipt_row_deleted_from_export_is_voided_on_upload(db):
    """05 实收回款删行覆盖：只作废导出签名 envelope 中消失的稳定行。"""

    project, _contract, snapshot = _project_with_receipt(db)

    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_RECEIPTS,),
    )
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook[master.V2_SHEET_RECEIPTS]
    headers = {cell.value: cell.column for cell in sheet[1]}
    row_no = next(
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, headers["实体ID"]).value == snapshot.collection_id
    )
    sheet.delete_rows(row_no, 1)

    plan = master.validate_project_master_v2(
        db,
        project_id=project.project_id,
        data=_save(workbook),
    )
    assert plan.summary["collection_voids"] == 1
    assert plan.will_void_rows == ({
        "sheet": master.V2_SHEET_RECEIPTS,
        "entity_id": snapshot.collection_id,
        "label": "XSDD-EDIT-001 2026-06",
        "reason": "上传文件缺行",
    },)

    master.apply_project_master_v2(
        db,
        plan,
        operated_by="receipt-delete-test",
        import_batch_id=str(uuid.uuid4()),
    )
    db.refresh(snapshot)
    assert snapshot.status == "void"


def test_unchanged_receipt_export_is_not_voided(db):
    project, _contract, snapshot = _project_with_receipt(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_RECEIPTS,),
    )

    plan = master.validate_project_master_v2(
        db,
        project_id=project.project_id,
        data=content,
    )
    assert plan.summary["collection_voids"] == 0
    assert plan.will_void_rows == ()

    master.apply_project_master_v2(
        db,
        plan,
        operated_by="receipt-unchanged-test",
        import_batch_id=str(uuid.uuid4()),
    )
    db.refresh(snapshot)
    assert snapshot.status == "confirmed"


def test_receipt_created_after_export_is_not_voided_by_old_file(db):
    project, contract, snapshot = _project_with_receipt(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_RECEIPTS,),
    )
    later_snapshot = MaintenanceCollectionSnapshot(
        collection_id=str(uuid.uuid4()),
        project_id=project.project_id,
        project_contract_id=contract.project_contract_id,
        report_month=date(2026, 7, 1),
        cumulative_amount=Decimal("100000.00"),
        status="confirmed",
        receipt_reference="SKD-20260731-0008",
        remark="导出后新增",
        source="direct_api",
        import_batch_id=None,
        version=1,
    )
    db.add(later_snapshot)
    db.commit()

    plan = master.validate_project_master_v2(
        db,
        project_id=project.project_id,
        data=content,
    )
    assert plan.summary["collection_voids"] == 0
    master.apply_project_master_v2(
        db,
        plan,
        operated_by="receipt-later-test",
        import_batch_id=str(uuid.uuid4()),
    )
    db.refresh(snapshot)
    db.refresh(later_snapshot)
    assert snapshot.status == "confirmed"
    assert later_snapshot.status == "confirmed"


def test_receipt_row_without_exported_identity_fails_closed(db):
    project, _contract, snapshot = _project_with_receipt(db)
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_RECEIPTS,),
    )
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook[master.V2_SHEET_RECEIPTS]
    headers = {cell.value: cell.column for cell in sheet[1]}
    row_no = next(
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, headers["实体ID"]).value == snapshot.collection_id
    )
    sheet.cell(row_no, headers["实体ID"]).value = None

    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db,
            project_id=project.project_id,
            data=_save(workbook),
        )
    assert exc.value.code == "receipt_identity_lost"
    db.refresh(snapshot)
    assert snapshot.status == "confirmed"


def test_e2e_fix_summary_distinguishes_qty_from_cost(db):
    """E2E #5：改数量报 line_updates/qty_updates，不再误报 cost_overrides。"""
    project, _part, _order, line = _make_project_with_line(db)
    _c, wb, ws = _parts_sheet(db, project.project_id)
    ws.cell(row=2, column=11, value=5)
    buf = io.BytesIO(); wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.summary["qty_updates"] == 1
    assert plan.summary["line_updates"] == 1
    assert plan.summary["cost_overrides"] == 0


def test_v23_example_rows_present_and_ignored_on_upload(db):
    """每个数据 sheet 底部有灰色示例行；回传时被系统忽略（零变更幂等）。"""
    project, _order, _line, _expense = _make_project_with_expense(db)
    content = master.build_project_master_v2(db, project_id=project.project_id)
    wb = load_workbook(io.BytesIO(content))
    for sheet_name in (master.V2_SHEET_PLAN, master.V2_SHEET_PARTS,
                       master.V2_SHEET_EXPENSE, master.V2_SHEET_RECEIPTS,
                       master.V2_SHEET_SITE):
        ws = wb[sheet_name]
        # finalize 会向下多刷 20 行样式（max_row 被撑大），全表扫找示例标记
        found = any(
            str(c.value or "").strip() in ("示例",) or str(c.value or "").startswith("【示例】")
            for row in ws.iter_rows(min_row=2) for c in row
        )
        assert found, sheet_name
    # 原样回传（含示例行）：零变更
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.summary["line_creates"] == 0
    assert plan.summary["line_updates"] == 0
    assert plan.summary["line_voids"] == 0
    assert plan.summary["expense_creates"] == 0
    assert plan.summary["expense_voids"] == 0
    assert plan.summary["plan_creates"] == 0
    assert not plan.will_void_rows


def test_v23_plan_accepts_xsdd_without_ledger_contract(db):
    """台账未导入的项目也能填 02：合同号=挂靠 XSDD → 自动建合同（用户踩坑）。"""
    from app.models.maintenance_project import MaintenanceProjectContract

    project, _part, order, _line = _make_project_with_line(db)
    # 该项目挂靠单的 XSDD = XSDD-EDIT-001（fixture 里 linked_sales_order_no）
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PLAN]
    # 在示例行后追加一条 CREATE（用挂靠 XSDD，不依赖台账）
    ws.append(["CREATE", "XSDD-EDIT-001", 1, "2026-09-30", "day", 40000,
               None, None, None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.summary["plan_creates"] == 1
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id,
        MaintenanceProjectContract.contract_no == "XSDD-EDIT-001"))
    assert contract is not None  # 自动建出来了


def test_v23_plan_rejects_with_helpful_error(db):
    """不属于本项目的合同号 → 报错列出可用合同。"""
    import pytest
    from app.services.maintenance_expense_collection_workbook import WorkbookError

    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PLAN]
    ws.append(["CREATE", "XSDD-别的项目", 1, "2026-09-30", "day", 100, None, None,
               None, None, None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=buf.getvalue())
    assert "可用" in str(exc.value) and "XSDD-EDIT-001" in str(exc.value)


def test_v23_plan_modify_without_operation_applies(db):
    """02 改单元格（不填操作）→ 生效；改了日期/金额才下发。"""
    from app.models.maintenance_manager import MaintenanceCollectionMilestone

    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    # 先 CREATE 一期
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PLAN]
    ws.append(["CREATE", "XSDD-EDIT-001", 1, "2026-09-30", "day", 40000,
               None, None, None, None, None, None, None, None])
    buf = io.BytesIO(); wb.save(buf)
    plan = master.validate_project_master_v2(db, project_id=project.project_id, data=buf.getvalue())
    master.apply_project_master_v2(db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))

    # 重新下载，不填操作直接改金额 → 生效
    content2 = master.build_project_master_v2(db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb2 = load_workbook(io.BytesIO(content2))
    ws2 = wb2[master.V2_SHEET_PLAN]
    found = False
    for r in ws2.iter_rows(min_row=2):
        if r[12].value:  # 实体ID 列（1-based 13 → idx 12）
            r[5].value = 30000  # 计划金额
            found = True
    assert found
    buf2 = io.BytesIO(); wb2.save(buf2)
    plan2 = master.validate_project_master_v2(db, project_id=project.project_id, data=buf2.getvalue())
    assert plan2.summary["plan_updates"] == 1
    master.apply_project_master_v2(db, plan2, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    ms = db.scalar(select(MaintenanceCollectionMilestone).where(
        MaintenanceCollectionMilestone.project_id == project.project_id))
    assert ms.planned_amount == Decimal("30000")

    # 原样再传（什么都不改）→ 零变更
    content3 = master.build_project_master_v2(db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    plan3 = master.validate_project_master_v2(db, project_id=project.project_id, data=content3)
    assert plan3.summary["plan_updates"] == 0 and plan3.summary["plan_creates"] == 0


def test_v23_plan_delete_row_voids(db):
    """02 删行=作废（对齐 03/04）。"""
    from app.models.maintenance_manager import MaintenanceCollectionMilestone

    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_PLAN].append(["CREATE", "XSDD-EDIT-001", 1, "2026-09-30", "day", 40000,
                                     None, None, None, None, None, None, None, None])
    buf = io.BytesIO(); wb.save(buf)
    plan = master.validate_project_master_v2(db, project_id=project.project_id, data=buf.getvalue())
    master.apply_project_master_v2(db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))

    content2 = master.build_project_master_v2(db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb2 = load_workbook(io.BytesIO(content2))
    ws2 = wb2[master.V2_SHEET_PLAN]
    for r in range(ws2.max_row, 1, -1):
        if ws2.cell(r, 13).value:  # 实体ID
            ws2.delete_rows(r)
    buf2 = io.BytesIO(); wb2.save(buf2)
    plan2 = master.validate_project_master_v2(db, project_id=project.project_id, data=buf2.getvalue())
    assert plan2.summary["plan_voids"] == 1
    assert any(x["sheet"] == "02_回款计划" for x in plan2.will_void_rows)
    master.apply_project_master_v2(db, plan2, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    ms = db.scalar(select(MaintenanceCollectionMilestone).where(
        MaintenanceCollectionMilestone.project_id == project.project_id))
    assert ms.is_active is False


# ---------------------------------------------------------------- 06 缺行=作废（2026-08-23 用户口径）

def _site_issue(db, project, part, *, qty="2", line_no=1, issue_no="ISS-V2-1"):
    from app.models.maintenance_project_operations import (
        MaintenanceSiteIssue,
        MaintenanceSiteIssueLine,
    )

    issue = MaintenanceSiteIssue(
        issue_id=str(uuid.uuid4()), project_id=project.project_id,
        issue_no=issue_no, issue_date=date(2026, 8, 10),
        raw_status="已确认", status_mapping_state="mapped",
        normalized_status="confirmed", status_mapping_version="t",
        source="legacy")
    db.add(issue)
    db.flush()
    line = MaintenanceSiteIssueLine(
        issue_line_id=str(uuid.uuid4()), issue_id=issue.issue_id,
        line_no=line_no, part_id=part.id, pn=part.pn_std,
        quantity=Decimal(qty), algorithm_version="t",
        tax_rate_used=Decimal("0.13"), is_active=True)
    db.add(line)
    db.commit()
    return issue, line


def _append_site_line(db, issue, part, *, qty="1", line_no=2):
    line = MaintenanceSiteIssueLine(
        issue_line_id=str(uuid.uuid4()), issue_id=issue.issue_id,
        line_no=line_no, part_id=part.id, pn=part.pn_std,
        quantity=Decimal(qty), algorithm_version="t",
        tax_rate_used=Decimal("0.13"), is_active=True,
    )
    db.add(line)
    db.commit()
    return line


def _site_row_for_entity(ws, entity_id):
    return next(
        row_no
        for row_no in range(2, ws.max_row + 1)
        if ws.cell(row_no, 11).value == entity_id
    )


def test_v2_site_header_change_on_one_repeated_row_propagates_once_and_reprices(
    db, monkeypatch
):
    project, part, _order, _line = _make_project_with_line(db)
    issue, first = _site_issue(db, project, part)
    second = _append_site_line(db, issue, part)
    original_version = issue.version
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,)
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    changed_row = _site_row_for_entity(ws, first.issue_line_id)
    ws.cell(changed_row, 1, "ISS-V2-RENAMED")
    ws.cell(changed_row, 2, date(2026, 8, 12))
    repriced = []

    def _capture_reprice(_db, *, lines):
        repriced.extend((issue_date, line.issue_line_id) for issue_date, line in lines)

    from app.services import maintenance_consumption_cost

    monkeypatch.setattr(maintenance_consumption_cost, "resolve_lines", _capture_reprice)
    _reupload(db, project.project_id, wb)

    db.refresh(issue)
    assert issue.issue_no == "ISS-V2-RENAMED"
    assert issue.issue_date == date(2026, 8, 12)
    assert issue.version == original_version + 1
    assert repriced == [
        (date(2026, 8, 12), first.issue_line_id),
        (date(2026, 8, 12), second.issue_line_id),
    ]
    assert db.scalar(select(func.count()).select_from(
        MaintenanceProjectOperationAudit
    ).where(
        MaintenanceProjectOperationAudit.project_id == project.project_id,
        MaintenanceProjectOperationAudit.entity_type == "site_issue",
        MaintenanceProjectOperationAudit.entity_id == issue.issue_id,
        MaintenanceProjectOperationAudit.action == "UPDATE",
    )) == 1


def test_v2_site_repeated_header_conflict_fails_closed(db):
    project, part, _order, _line = _make_project_with_line(db)
    issue, first = _site_issue(db, project, part)
    second = _append_site_line(db, issue, part)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,)
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    ws.cell(_site_row_for_entity(ws, first.issue_line_id), 1, "ISS-V2-A")
    ws.cell(_site_row_for_entity(ws, second.issue_line_id), 1, "ISS-V2-B")
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=buf.getvalue()
        )

    assert exc.value.code == "conflicting_issue_header"
    db.refresh(issue)
    assert issue.issue_no == "ISS-V2-1"


def test_v2_site_pn_change_updates_part_identity_clears_old_evidence_and_reprices(
    db, monkeypatch
):
    project, old_part, _order, _line = _make_project_with_line(db)
    issue, site_line = _site_issue(db, project, old_part)
    site_line.manual_unit_cost = Decimal("10.00")
    site_line.manual_unit_cost_inc_tax = Decimal("11.30")
    site_line.manual_evidence = "旧 PN 人工证据"
    new_part = DimPart(
        pn_std=f"SITE-NEW-{uuid.uuid4().hex[:8].upper()}",
        description="替换备件",
    )
    db.add(new_part)
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,)
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    ws.cell(_site_row_for_entity(ws, site_line.issue_line_id), 3, new_part.pn_std)
    repriced = []

    def _capture_reprice(_db, *, lines):
        repriced.extend((issue_date, line.issue_line_id) for issue_date, line in lines)

    from app.services import maintenance_consumption_cost

    monkeypatch.setattr(maintenance_consumption_cost, "resolve_lines", _capture_reprice)
    _reupload(db, project.project_id, wb)

    db.refresh(site_line)
    assert site_line.pn == new_part.pn_std
    assert site_line.part_id == new_part.id
    assert site_line.manual_unit_cost is None
    assert site_line.manual_unit_cost_inc_tax is None
    assert site_line.manual_evidence is None
    assert repriced == [(issue.issue_date, site_line.issue_line_id)]


def test_v2_site_missing_row_voids_line_and_issue(db):
    """06 删行覆盖上传 → 领用行作废；单只剩空 → 单据状态置 void。"""
    from app.models.maintenance_project_operations import (
        MaintenanceSiteIssue,
        MaintenanceSiteIssueLine,
    )

    project, part, _order, _line = _make_project_with_line(db)
    issue, sline = _site_issue(db, project, part)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    # 删掉全部数据行（示例行保留与否都行——解析跳过）
    data_rows = ws.max_row - 1
    ws.delete_rows(2, data_rows)
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.summary["site_voids"] == 1
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.refresh(sline)
    db.refresh(issue)
    assert sline.is_active is False
    assert issue.normalized_status == "void"
    # 再导出：作废行不再出现
    content2 = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    wb2 = load_workbook(io.BytesIO(content2))
    body_rows = [r for r in wb2[master.V2_SHEET_SITE].iter_rows(min_row=2, values_only=True)
                 if any(v not in (None, "") for v in r)
                 and not str(r[2] or "").startswith("（示例）")]
    assert all(row[10] != sline.issue_line_id for row in body_rows)


def test_v2_site_roundtrip_unchanged_is_zero_ops(db):
    """原样回传（不删任何行）→ 零作废零更新（幂等）。"""
    project, part, _order, _line = _make_project_with_line(db)
    _site_issue(db, project, part)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=content)
    assert plan.summary["site_voids"] == 0
    assert plan.summary["site_creates"] == 0
    # site_updates 允许 ≥0：解析器对原样行也会记一条「更新」（apply 幂等，
    # 无实际变化）——既有行为，不在本断言范围


def test_v2_site_partial_delete_voids_only_missing(db):
    """两张单各一行，只删一张 → 只作废被删那张。"""
    from app.models.maintenance_project_operations import (
        MaintenanceSiteIssueLine,
    )

    project, part, _order, _line = _make_project_with_line(db)
    issue_a, line_a = _site_issue(db, project, part, issue_no="ISS-A")
    issue_b, line_b = _site_issue(db, project, part, issue_no="ISS-B")
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_SITE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_SITE]
    # 删掉 A 那一行（数据首行）
    ws.delete_rows(2, 1)
    buf = io.BytesIO()
    wb.save(buf)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=buf.getvalue())
    assert plan.summary["site_voids"] == 1
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.refresh(line_a)
    db.refresh(line_b)
    assert line_a.is_active is False
    assert line_b.is_active is True


# ----------------------------------------------------------- V2.5 entity envelope

def _v25_cross_project_facts(db):
    project_a, order_a, line_a, expense_a = _make_project_with_expense(db)
    project_b, order_b, line_b, expense_b = _make_project_with_expense(db)
    contract_b = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project_b.project_id,
    ))
    unique_b_contract = f"XSDD-SEC-{uuid.uuid4().hex[:8]}"
    contract_b.contract_no = unique_b_contract
    contract_b.contract_id = f"CONTRACT-SEC-{uuid.uuid4().hex[:8]}"
    order_b.linked_sales_order_no = unique_b_contract
    expense_b.linked_sales_order_no = unique_b_contract
    part_a = db.get(DimPart, line_a.part_id)
    part_b = db.get(DimPart, line_b.part_id)
    issue_a, site_a = _site_issue(
        db, project_a, part_a, issue_no=f"ISS-SEC-A-{uuid.uuid4().hex[:6]}")
    issue_b, site_b = _site_issue(
        db, project_b, part_b, issue_no=f"ISS-SEC-B-{uuid.uuid4().hex[:6]}")
    db.commit()
    return {
        "project_a": project_a,
        "order_a": order_a,
        "line_a": line_a,
        "expense_a": expense_a,
        "issue_a": issue_a,
        "site_a": site_a,
        "project_b": project_b,
        "order_b": order_b,
        "line_b": line_b,
        "expense_b": expense_b,
        "issue_b": issue_b,
        "site_b": site_b,
    }


@pytest.mark.parametrize(
    ("kind", "operation", "expected_code"),
    [
        ("parts", "VOID", "line_not_in_project"),
        ("site", None, "project_mismatch"),
        ("expense", "UPDATE", "expense_not_in_project"),
        ("expense", "VOID", "expense_not_in_project"),
    ],
)
def test_v25_explicit_cross_project_entity_tamper_is_zero_write(
    db, kind, operation, expected_code
):
    facts = _v25_cross_project_facts(db)
    sheet_by_kind = {
        "parts": master.V2_SHEET_PARTS,
        "site": master.V2_SHEET_SITE,
        "expense": master.V2_SHEET_EXPENSE,
    }
    content = master.build_project_master_v2(
        db,
        project_id=facts["project_a"].project_id,
        sheets=(sheet_by_kind[kind],),
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[sheet_by_kind[kind]]
    if kind == "parts":
        ws.cell(2, 1, operation)
        ws.cell(2, 22, facts["line_b"].id)
    elif kind == "site":
        ws.cell(2, 11, facts["site_b"].issue_line_id)
        ws.cell(2, 5, 99)
    else:
        ws.cell(2, 1, operation)
        ws.cell(2, 18, facts["expense_b"].raw_line_id)
        if operation == "UPDATE":
            ws.cell(2, 15, 999)

    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db,
            project_id=facts["project_a"].project_id,
            data=_save(wb),
        )
    assert exc.value.code == expected_code
    db.refresh(facts["line_b"])
    db.refresh(facts["site_b"])
    db.refresh(facts["expense_b"])
    assert facts["line_b"].is_active is True
    assert facts["site_b"].is_active is True
    assert facts["site_b"].quantity == Decimal("2")
    assert facts["expense_b"].amount_ex_tax == Decimal("500.00")
    assert facts["expense_b"].data_status != "已作废"


@pytest.mark.parametrize(
    ("kind", "meta_key"),
    [
        ("parts", "parts_row_ids"),
        ("site", "site_row_ids"),
        ("expense", "expense_row_ids"),
    ],
)
def test_v25_forged_implicit_void_identity_set_fails_hmac_and_writes_nothing(
    db, kind, meta_key
):
    facts = _v25_cross_project_facts(db)
    sheet_by_kind = {
        "parts": master.V2_SHEET_PARTS,
        "site": master.V2_SHEET_SITE,
        "expense": master.V2_SHEET_EXPENSE,
    }
    foreign_id = {
        "parts": str(facts["line_b"].id),
        "site": facts["site_b"].issue_line_id,
        "expense": facts["expense_b"].raw_line_id,
    }[kind]
    content = master.build_project_master_v2(
        db,
        project_id=facts["project_a"].project_id,
        sheets=(sheet_by_kind[kind],),
    )
    wb = load_workbook(io.BytesIO(content))
    meta = wb[master.V2_SHEET_META]
    target = next(row[1] for row in meta.iter_rows(min_col=1, max_col=2)
                  if row[0].value == meta_key)
    target.value = foreign_id

    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db,
            project_id=facts["project_a"].project_id,
            data=_save(wb),
        )
    assert exc.value.code == "template_signature_invalid"
    db.refresh(facts["line_b"])
    db.refresh(facts["site_b"])
    db.refresh(facts["expense_b"])
    assert facts["line_b"].is_active is True
    assert facts["site_b"].is_active is True
    assert facts["expense_b"].data_status != "已作废"


def test_v25_unsigned_old_template_fails_closed_with_redownload_message(db):
    project, _part, _order, _line = _make_project_with_line(db)
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    meta = wb[master.V2_SHEET_META]
    version = next(row[1] for row in meta.iter_rows(min_col=1, max_col=2)
                   if row[0].value == "template_version")
    version.value = "2.4.0"
    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert exc.value.code == "template_version_mismatch"
    assert "重新下载" in exc.value.message


@pytest.mark.parametrize(
    ("kind", "mode"),
    [
        ("parts", "implicit_void"),
        ("parts", "explicit_void"),
        ("site", "implicit_void"),
        ("expense", "implicit_void"),
        ("expense", "explicit_void"),
        ("expense", "explicit_update"),
    ],
)
def test_v25_apply_rechecks_scope_after_ownership_changes(db, kind, mode):
    facts = _v25_cross_project_facts(db)
    sheet_by_kind = {
        "parts": master.V2_SHEET_PARTS,
        "site": master.V2_SHEET_SITE,
        "expense": master.V2_SHEET_EXPENSE,
    }
    content = master.build_project_master_v2(
        db,
        project_id=facts["project_a"].project_id,
        sheets=(sheet_by_kind[kind],),
    )
    wb = load_workbook(io.BytesIO(content))
    ws = wb[sheet_by_kind[kind]]
    if mode == "implicit_void":
        ws.delete_rows(2, 1)
    elif mode == "explicit_void":
        ws.cell(2, 1, "VOID")
    else:
        ws.cell(2, 1, "UPDATE")
        ws.cell(2, 15, 777)
    plan = master.validate_project_master_v2(
        db,
        project_id=facts["project_a"].project_id,
        data=_save(wb),
    )
    if kind == "parts":
        from app.services import maintenance_source_assignments

        assignment = db.scalar(select(MaintenanceSourceOrderAssignment).where(
            MaintenanceSourceOrderAssignment.source_order_id
            == facts["order_a"].raw_order_id,
            MaintenanceSourceOrderAssignment.is_active.is_(True),
        ))
        maintenance_source_assignments.assign_source_orders(
            db,
            project_id=facts["project_b"].project_id,
            items=[{
                "source_order_id": facts["order_a"].raw_order_id,
                "expected_assignment_id": assignment.assignment_id,
                "expected_version": assignment.version,
            }],
            reason="模拟 validate 后并发改挂",
            operated_by="security-test",
            user_ctx=UserContext(
                user_id="security-admin", role="admin", is_authenticated=True),
        )
    elif kind == "site":
        facts["issue_a"].project_id = facts["project_b"].project_id
    else:
        contract_b = db.scalar(select(MaintenanceProjectContract).where(
            MaintenanceProjectContract.project_id == facts["project_b"].project_id,
        ))
        facts["expense_a"].linked_sales_order_no = contract_b.contract_no
    db.commit()

    audits_before = int(db.scalar(
        select(func.count(MaintenanceProjectOperationAudit.id))
    ) or 0)
    with pytest.raises(master.WorkbookError):
        master.apply_project_master_v2(
            db,
            plan,
            operated_by="security-test",
            import_batch_id=str(uuid.uuid4()),
        )
    audits_after = int(db.scalar(
        select(func.count(MaintenanceProjectOperationAudit.id))
    ) or 0)
    assert audits_after == audits_before
    db.refresh(facts["line_a"])
    db.refresh(facts["site_a"])
    db.refresh(facts["expense_a"])
    assert facts["line_a"].is_active is True
    assert facts["site_a"].is_active is True
    assert facts["expense_a"].data_status != "已作废"


def test_v25_apply_uses_global_parent_before_child_lock_order(db):
    """SQL evidence for the deadlock-safe apply order used by all V2 writes."""

    from sqlalchemy import event

    project, part, _order, _line = _make_project_with_line(db)
    _issue, _site_line = _site_issue(
        db, project, part, issue_no=f"ISS-LOCK-{uuid.uuid4().hex[:6]}")
    content = master.build_project_master_v2(
        db,
        project_id=project.project_id,
        sheets=(master.V2_SHEET_PARTS, master.V2_SHEET_SITE),
    )
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_PARTS].delete_rows(2, 1)
    wb[master.V2_SHEET_SITE].delete_rows(2, 1)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))

    statements: list[str] = []
    engine = db.get_bind()

    def capture(_conn, _cursor, statement, _parameters, _context, _executemany):
        normalized = " ".join(statement.lower().split())
        if "for update" in normalized:
            statements.append(normalized)

    event.listen(engine, "before_cursor_execute", capture)
    try:
        master.apply_project_master_v2(
            db,
            plan,
            operated_by="lock-order-test",
            import_batch_id=str(uuid.uuid4()),
        )
    finally:
        event.remove(engine, "before_cursor_execute", capture)

    def lock_index(table: str) -> int:
        return next(
            index for index, statement in enumerate(statements)
            if f"from {table} " in statement
        )

    state_lock = lock_index("maintenance_project_workbook_state")
    project_lock = lock_index("maintenance_project")
    order_lock = lock_index("f_maintenance_order")
    assignment_lock = lock_index("maintenance_source_order_assignment")
    line_lock = lock_index("f_maintenance_line")
    issue_lock = lock_index("maintenance_site_issue")
    site_line_lock = lock_index("maintenance_site_issue_line")
    assert state_lock < project_lock < order_lock < assignment_lock < line_lock
    assert project_lock < issue_lock < site_line_lock


def test_v25_apply_receipt_replays_manual_create_once(db, monkeypatch):
    """ACK 丢失后同文件重试：回放同结果，手工行只建一次。"""

    from sqlalchemy import event

    project, part, order, _line = _make_project_with_line(db)
    monkeypatch.setattr(
        "app.services.maintenance_cost.recompute",
        lambda *_args, **_kwargs: {"lines_in_scope": 1},
    )
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    row_no = ws.max_row + 1
    ws.cell(row_no, headers["操作"], "CREATE")
    ws.cell(row_no, headers["维保单号"], order.order_no)
    ws.cell(row_no, headers["PN"], part.pn_std)
    ws.cell(row_no, headers["需求数量"], 3)
    data = _save(wb)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=data)
    assert plan.summary["line_creates"] == 1

    advisory_statements: list[str] = []
    engine = db.get_bind()

    def capture(_conn, _cursor, statement, _parameters, _context, _executemany):
        if "pg_advisory_xact_lock" in statement.lower():
            advisory_statements.append(statement.lower())

    event.listen(engine, "before_cursor_execute", capture)
    try:
        first = master.apply_project_master_v2(
            db, plan, operated_by="idempotency-test",
            import_batch_id=str(uuid.uuid4()))
        second = master.apply_project_master_v2(
            db, plan, operated_by="idempotency-test",
            import_batch_id=str(uuid.uuid4()))
    finally:
        event.remove(engine, "before_cursor_execute", capture)

    manual_lines = list(db.scalars(select(FMaintenanceLine).where(
        FMaintenanceLine.order_id == order.id,
        FMaintenanceLine.edited_source == "workbook_manual",
    )).all())
    receipts = list(db.scalars(select(MaintenanceProjectWorkbookOperation).where(
        MaintenanceProjectWorkbookOperation.project_id == project.project_id,
        MaintenanceProjectWorkbookOperation.operation_type == "file_apply",
    )).all())
    assert len(manual_lines) == 1
    assert manual_lines[0].raw_line_id.startswith("manual-line:")
    assert len(receipts) == 1
    assert first["replayed"] is False
    assert second["replayed"] is True
    assert first["import_batch_id"] == second["import_batch_id"]
    assert len(advisory_statements) >= 2


def _v25_two_milestones(db):
    from app.models.maintenance_manager import MaintenanceCollectionMilestone

    project, _part, _order, _line = _make_project_with_line(db)
    contract_a = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id))
    contract_b = MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project.project_id,
        contract_id=f"EDIT-CONTRACT-{uuid.uuid4().hex[:8]}",
        contract_no=f"XSDD-EDIT-{uuid.uuid4().hex[:8]}",
        amount_inc_tax=Decimal("20000.00"),
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="v1",
        effective_from=date(2026, 1, 1),
        source="ledger",
        version=1,
    )
    db.add(contract_b)
    db.flush()
    milestone_a = MaintenanceCollectionMilestone(
        milestone_id=str(uuid.uuid4()), project_id=project.project_id,
        project_contract_id=contract_a.project_contract_id, sequence=1,
        planned_date=date(2026, 9, 1), planned_amount=Decimal("1000"),
        completeness_state="complete", source="direct_api", version=1,
    )
    milestone_b = MaintenanceCollectionMilestone(
        milestone_id=str(uuid.uuid4()), project_id=project.project_id,
        project_contract_id=contract_b.project_contract_id, sequence=1,
        planned_date=date(2026, 10, 1), planned_amount=Decimal("2000"),
        completeness_state="complete", source="direct_api", version=1,
    )
    db.add_all([milestone_a, milestone_b])
    db.commit()
    return project, contract_a, contract_b, milestone_a, milestone_b


def test_v25_plan_void_uses_hidden_entity_not_editable_coordinate(db):
    project, _contract_a, contract_b, milestone_a, milestone_b = (
        _v25_two_milestones(db)
    )
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PLAN]
    headers = {cell.value: cell.column for cell in ws[1]}
    row_no = next(
        row for row in range(2, ws.max_row + 1)
        if ws.cell(row, headers["实体ID"]).value == milestone_a.milestone_id
    )
    ws.cell(row_no, headers["操作"], "VOID")
    ws.cell(row_no, headers["合同编号"], contract_b.contract_no)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert plan.milestone_changes[0].entity_id == milestone_a.milestone_id
    assert plan.milestone_changes[0].contract_no != contract_b.contract_no
    master.apply_project_master_v2(
        db, plan, operated_by="milestone-test",
        import_batch_id=str(uuid.uuid4()))
    db.refresh(milestone_a)
    db.refresh(milestone_b)
    assert milestone_a.is_active is False
    assert milestone_b.is_active is True


def test_v25_plan_update_cannot_redirect_hidden_entity(db):
    project, _contract_a, contract_b, milestone_a, milestone_b = (
        _v25_two_milestones(db)
    )
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PLAN]
    headers = {cell.value: cell.column for cell in ws[1]}
    row_no = next(
        row for row in range(2, ws.max_row + 1)
        if ws.cell(row, headers["实体ID"]).value == milestone_a.milestone_id
    )
    ws.cell(row_no, headers["操作"], "UPDATE")
    ws.cell(row_no, headers["合同编号"], contract_b.contract_no)
    ws.cell(row_no, headers["计划回款金额（含税）"], 9999)
    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert exc.value.code == "milestone_identity_mismatch"
    db.refresh(milestone_a)
    db.refresh(milestone_b)
    assert milestone_a.planned_amount == Decimal("1000")
    assert milestone_b.planned_amount == Decimal("2000")


def test_v25_missing_plan_entity_fails_controlled_without_writes(db):
    project, _contract_a, _contract_b, milestone_a, milestone_b = (
        _v25_two_milestones(db)
    )
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PLAN]
    entity_col = next(cell.column for cell in ws[1] if cell.value == "实体ID")
    for row_no in range(ws.max_row, 1, -1):
        if ws.cell(row_no, entity_col).value == milestone_a.milestone_id:
            ws.delete_rows(row_no)
            break
    db.delete(milestone_a)
    db.commit()
    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project.project_id, data=_save(wb))
    assert exc.value.code == "milestone_not_found"
    db.refresh(milestone_b)
    assert milestone_b.is_active is True


def test_v25_shared_xsdd_expenses_are_read_only_and_forgery_fails(db):
    project_a, _order_a, _line_a, expense_a = _make_project_with_expense(db)
    project_b, _order_b, _line_b, expense_b = _make_project_with_expense(db)
    content = master.build_project_master_v2(
        db, project_id=project_a.project_id,
        sheets=(master.V2_SHEET_EXPENSE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_EXPENSE]
    headers = {cell.value: cell.column for cell in ws[1]}
    entity_values = {
        ws.cell(row_no, headers["实体ID"]).value
        for row_no in range(2, ws.max_row + 1)
    }
    assert expense_a.raw_line_id not in entity_values
    row_no = ws.max_row + 1
    ws.cell(row_no, headers["操作"], "UPDATE")
    ws.cell(row_no, headers["费用单号"], expense_b.bxd_no)
    ws.cell(row_no, headers["明细序号"], expense_b.line_no)
    ws.cell(row_no, headers["报销日期"], expense_b.expense_date)
    ws.cell(row_no, headers["维保销售订单（归集键）"], "XSDD-EDIT-001")
    ws.cell(row_no, headers["未税金额"], 999)
    ws.cell(row_no, headers["实体ID"], expense_b.raw_line_id)
    with pytest.raises(master.WorkbookError) as exc:
        master.validate_project_master_v2(
            db, project_id=project_a.project_id, data=_save(wb))
    assert exc.value.code == "contract_not_found"
    db.refresh(expense_a)
    db.refresh(expense_b)
    assert expense_a.amount_ex_tax == Decimal("500.00")
    assert expense_b.amount_ex_tax == Decimal("500.00")


def test_v25_apply_rechecks_stable_expense_attribution(db):
    project_a, _order_a, _line_a, expense_a = _make_project_with_expense(db)
    content = master.build_project_master_v2(
        db, project_id=project_a.project_id,
        sheets=(master.V2_SHEET_EXPENSE,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_EXPENSE]
    ws.cell(2, 1, "UPDATE")
    ws.cell(2, 15, 777)
    plan = master.validate_project_master_v2(
        db, project_id=project_a.project_id, data=_save(wb))
    project_b, _part_b, _order_b, _line_b = _make_project_with_line(db)
    attribution = db.get(
        MaintenanceProjectExpenseAttribution, f"bxd:{expense_a.raw_line_id}")
    attribution.project_id = project_b.project_id
    attribution.project_contract_id = db.scalar(select(
        MaintenanceProjectContract.project_contract_id
    ).where(MaintenanceProjectContract.project_id == project_b.project_id))
    db.commit()
    with pytest.raises(master.WorkbookError) as exc:
        master.apply_project_master_v2(
            db, plan, operated_by="expense-scope-test",
            import_batch_id=str(uuid.uuid4()))
    assert exc.value.code == "expense_not_in_project"
    db.rollback()
    db.refresh(expense_a)
    assert expense_a.amount_ex_tax == Decimal("500.00")


def test_v25_validate_does_not_create_xsdd_contract_and_null_date_is_422(db):
    """validate 纯读；apply 缺业务日期时返回受控错误而非 NOT NULL 500。"""

    project, _part, _order, _line = _make_project_with_line(db)
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id))
    db.delete(contract)
    db.commit()
    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PLAN,))
    wb = load_workbook(io.BytesIO(content))
    wb[master.V2_SHEET_PLAN].append([
        "CREATE", "XSDD-EDIT-001", 1, "2026-09-30", "day", 1000,
        None, None, None, None, None, None, None, None,
    ])
    data = _save(wb)
    before = int(db.scalar(select(func.count(
        MaintenanceProjectContract.project_contract_id
    )).where(MaintenanceProjectContract.project_id == project.project_id)) or 0)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=data)
    after_validate = int(db.scalar(select(func.count(
        MaintenanceProjectContract.project_contract_id
    )).where(MaintenanceProjectContract.project_id == project.project_id)) or 0)
    assert before == after_validate == 0
    with pytest.raises(master.WorkbookError) as exc:
        master.apply_project_master_v2(
            db, plan, operated_by="contract-date-test",
            import_batch_id=str(uuid.uuid4()))
    assert exc.value.code == "contract_effective_date_missing"
    db.rollback()
    after_apply = int(db.scalar(select(func.count(
        MaintenanceProjectContract.project_contract_id
    )).where(MaintenanceProjectContract.project_id == project.project_id)) or 0)
    assert after_apply == 0
