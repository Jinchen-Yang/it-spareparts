"""WBDD 跨页检索与可恢复逻辑删除的端到端契约。"""

import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select, text
from sqlalchemy.exc import DBAPIError

from app.auth import hash_password
from app.db import SessionLocal
from app.etl import loader
from app.main import app
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    MaintenanceDemandDeleteEvent,
    MaintenanceDemandDeleteIntent,
    MaintenanceDemandTombstone,
)
from app.models.system import SysAccessLog, SysImportBatch, SysUser
from app.services import maintenance_demands
from app.services import inventory, maintenance_cost, maintenance_export
from app.security import UserContext
from tests import factories as f


def _admin_client(db, username: str = "demand-delete-admin") -> TestClient:
    db.add(
        SysUser(
            username=username,
            role="admin",
            display_name="合成实名管理员",
            password_hash=hash_password("synthetic-password-123"),
        )
    )
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": username, "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def _seed_demands(db, *, count: int = 3, lines_each: int = 2) -> list[str]:
    batch = SysImportBatch(
        filename="synthetic-maintenance.xlsx",
        file_type="maintenance",
        file_hash=f"synthetic-maintenance-{count}-{lines_each}",
        status="processing",
    )
    db.add(batch)
    db.flush()
    orders = {}
    lines = []
    ids = []
    for index in range(1, count + 1):
        raw_id = f"RAW-WBDD-{index:04d}"
        ids.append(raw_id)
        orders[raw_id] = f.maintenance_head(
            raw_id,
            order_no=f"WBDD-20260809-{index:04d}",
            on=date(2026, 8, index),
            project=f"合成维保项目-{index}",
            sales_order=f"XSDD-SYNTH-{index}",
        )
        for line_index in range(1, lines_each + 1):
            lines.append(
                f.maintenance_line(
                    raw_id,
                    f"RAW-LINE-{index:04d}-{line_index:02d}",
                    f"PN-SYNTH-{index}-{line_index}",
                    description=f"合成备件 {index}-{line_index}",
                )
            )
    loader.load(
        db,
        f.maintenance_result(orders, lines),
        batch.id,
        date(2026, 8, 9),
    )
    batch.status = "success"
    db.commit()
    return ids


def _exported_first_line(db) -> tuple:
    output = maintenance_export.build_workbook(
        db,
        UserContext(user_id="demand-delete-admin", role="admin"),
    )
    try:
        output.seek(0)
        workbook = load_workbook(
            io.BytesIO(output.read()),
            read_only=True,
            data_only=True,
        )
    finally:
        output.close()
    try:
        return next(
            workbook["订单明细"].iter_rows(min_row=2, values_only=True)
        )
    finally:
        workbook.close()


def test_search_is_post_body_header_paginated_and_supports_pn(db):
    _seed_demands(db, count=3, lines_each=2)
    client = _admin_client(db)

    first = client.post(
        "/api/maintenance/demands/search",
        json={"q": "PN-SYNTH", "page": 1, "page_size": 2},
    )

    assert first.status_code == 200, first.text
    assert first.json()["total"] == 3
    assert len(first.json()["items"]) == 2
    assert all(item["line_count"] == 2 for item in first.json()["items"])
    assert len({item["source_order_id"] for item in first.json()["items"]}) == 2
    assert all(len(item["version_digest"]) == 64 for item in first.json()["items"])
    assert client.get("/api/maintenance/demands/search?q=PN-SYNTH").status_code == 405


def test_search_rejects_overlong_q_without_reflection_log_or_audit(db, caplog):
    client = _admin_client(db, username="demand-overlong-search-admin")
    sentinel = "WBDD-PRIVATE-SEARCH-SENTINEL-" + "x" * 256

    response = client.post(
        "/api/maintenance/demands/search",
        json={"q": sentinel, "page": 1, "page_size": 20},
    )

    assert response.status_code == 422
    assert "PRIVATE-SEARCH-SENTINEL" not in response.text
    assert sentinel not in caplog.text
    db.expire_all()
    assert (
        db.scalar(
            select(SysAccessLog.id).where(
                SysAccessLog.username == "demand-overlong-search-admin",
                SysAccessLog.action == "maintenance_demand_search",
            )
        )
        is None
    )


def test_delete_intent_requires_unique_bounded_selection_and_reason(db):
    raw_ids = _seed_demands(db, count=2, lines_each=2)
    client = _admin_client(db)

    blank = client.post(
        "/api/maintenance/demands/delete-intents",
        json={
            "source_order_ids": raw_ids,
            "reason": "   ",
            "idempotency_key": "intent-blank-reason",
        },
    )
    duplicate = client.post(
        "/api/maintenance/demands/delete-intents",
        json={
            "source_order_ids": [raw_ids[0], raw_ids[0]],
            "reason": "合成重复选择",
            "idempotency_key": "intent-duplicate-selection",
        },
    )

    assert blank.status_code == 422
    assert duplicate.status_code == 409
    assert raw_ids[0] not in duplicate.text
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandTombstone)) == 0


def test_delete_intent_rejects_oversized_selection_without_reflection_or_audit(
    db, caplog
):
    username = "demand-private-selection-admin"
    client = _admin_client(db, username=username)
    sentinel = "PRIVATE-STABLE-ID-SENTINEL"
    source_order_ids = [f"{sentinel}-{index:04d}" for index in range(1_001)]
    access_count_before = db.scalar(
        select(func.count())
        .select_from(SysAccessLog)
        .where(SysAccessLog.username == username)
    )

    response = client.post(
        "/api/maintenance/demands/delete-intents",
        json={
            "source_order_ids": source_order_ids,
            "reason": "合成超限选择",
            "idempotency_key": "safe-delete-private-selection-limit",
        },
    )

    assert response.status_code == 422
    assert sentinel not in response.text
    assert sentinel not in caplog.text
    db.expire_all()
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandDeleteIntent)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandDeleteEvent)) == 0
    assert (
        db.scalar(
            select(func.count())
            .select_from(SysAccessLog)
            .where(SysAccessLog.username == username)
        )
        == access_count_before
    )


def test_delete_intent_rejects_invalid_source_ids_without_reflection(db, caplog):
    client = _admin_client(db, username="demand-invalid-source-id-admin")
    sentinel = "PRIVATE-INVALID-SOURCE-ID-SENTINEL"
    invalid_selections = [
        [""],
        ["   "],
        [sentinel + "x" * 64],
        [{"private_id": sentinel}],
    ]

    for index, source_order_ids in enumerate(invalid_selections):
        response = client.post(
            "/api/maintenance/demands/delete-intents",
            json={
                "source_order_ids": source_order_ids,
                "reason": "合成非法稳定 ID",
                "idempotency_key": f"safe-delete-invalid-source-id-{index}",
            },
        )
        assert response.status_code == 422
        assert sentinel not in response.text

    assert sentinel not in caplog.text
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandDeleteIntent)) == 0


def test_delete_intent_limits_and_idempotency_are_atomic(db, monkeypatch):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    first = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成幂等验证",
        idempotency_key="safe-delete-idempotent-create",
        operated_by="demand-delete-admin",
        now=now,
    )
    db.commit()
    replay = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成幂等验证",
        idempotency_key="safe-delete-idempotent-create",
        operated_by="demand-delete-admin",
        now=now + timedelta(seconds=1),
    )
    assert replay["intent_id"] == first["intent_id"]
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandDeleteIntent)) == 1

    with pytest.raises(maintenance_demands.DeleteIntentConflict):
        maintenance_demands.create_delete_intent(
            db,
            source_order_ids=[raw_id],
            reason="同一幂等键却换了理由",
            idempotency_key="safe-delete-idempotent-create",
            operated_by="demand-delete-admin",
            now=now + timedelta(seconds=2),
        )
    with pytest.raises(maintenance_demands.MaintenanceDemandError):
        maintenance_demands.create_delete_intent(
            db,
            source_order_ids=[f"RAW-{index}" for index in range(1_001)],
            reason="超过订单头上限",
            idempotency_key="safe-delete-too-many-headers",
            operated_by="demand-delete-admin",
            now=now,
        )

    oversized = {
        "RAW-OVERSIZED": {
            "source_order_id": "RAW-OVERSIZED",
            "order_no": "WBDD-OVERSIZED",
            "order_date": None,
            "project": None,
            "project_raw": None,
            "linked_sales_order_no": None,
            "line_count": 20_001,
            "downstream_references": [],
            "version_digest": "f" * 64,
        }
    }
    monkeypatch.setattr(maintenance_demands, "_load_snapshots", lambda *args, **kwargs: oversized)
    with pytest.raises(maintenance_demands.DeleteIntentConflict):
        maintenance_demands.create_delete_intent(
            db,
            source_order_ids=["RAW-OVERSIZED"],
            reason="超过备件行上限",
            idempotency_key="safe-delete-too-many-lines",
            operated_by="demand-delete-admin",
            now=now,
        )
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandTombstone)) == 0


def test_intent_snapshot_holds_shared_source_lock_against_import_writer(db):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]

    maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成快照并发保护",
        idempotency_key="safe-delete-shared-source-lock",
        operated_by="demand-delete-admin",
    )

    with SessionLocal() as competing:
        acquired = competing.scalar(
            text("SELECT pg_try_advisory_xact_lock(:key)"),
            {"key": maintenance_demands.DATA_CHANGE_ADVISORY_LOCK_KEY},
        )
        assert acquired is False
        competing.rollback()

    db.commit()
    with SessionLocal() as competing:
        acquired = competing.scalar(
            text("SELECT pg_try_advisory_xact_lock(:key)"),
            {"key": maintenance_demands.DATA_CHANGE_ADVISORY_LOCK_KEY},
        )
        assert acquired is True
        competing.rollback()


def test_intent_digest_and_review_item_set_are_database_immutable(db):
    raw_id, unselected_raw_id = _seed_demands(db, count=2, lines_each=1)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成不可变复核身份",
        idempotency_key="safe-delete-immutable-intent",
        operated_by="demand-delete-admin",
    )
    db.commit()

    with pytest.raises(DBAPIError, match="intent identity is immutable"):
        db.execute(
            text(
                "UPDATE maintenance_demand_delete_intent "
                "SET selection_digest = :digest WHERE intent_id = :intent_id"
            ),
            {"digest": "0" * 64, "intent_id": intent["intent_id"]},
        )
    db.rollback()
    persisted = db.get(MaintenanceDemandDeleteIntent, intent["intent_id"])
    assert persisted.selection_digest == intent["selection_digest"]

    with pytest.raises(DBAPIError, match="intent item set is immutable"):
        db.execute(
            text(
                "INSERT INTO maintenance_demand_delete_intent_item "
                "(intent_id, source_order_id, ordinal, version_digest, snapshot_json) "
                "VALUES (:intent_id, :source_order_id, 1, :digest, '{}'::jsonb)"
            ),
            {
                "intent_id": intent["intent_id"],
                "source_order_id": unselected_raw_id,
                "digest": "1" * 64,
            },
        )
    db.rollback()
    assert db.scalar(
        text(
            "SELECT count(*) FROM maintenance_demand_delete_intent_item "
            "WHERE intent_id = :intent_id"
        ),
        {"intent_id": intent["intent_id"]},
    ) == 1


def test_digest_mismatch_never_arms_or_deletes(db):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成摘要篡改验证",
        idempotency_key="safe-delete-digest-mismatch",
        operated_by="demand-delete-admin",
        now=now,
    )
    with pytest.raises(maintenance_demands.DeleteIntentConflict):
        maintenance_demands.arm_delete_intent(
            db,
            intent_id=intent["intent_id"],
            digest="0" * 64,
            operated_by="demand-delete-admin",
            now=now,
        )
    db.rollback()
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandTombstone)) == 0


def test_full_review_server_wait_atomic_execute_and_replay(db):
    raw_ids = _seed_demands(db, count=2, lines_each=2)
    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)

    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=raw_ids,
        reason="两张合成测试单重复导入，确认逻辑删除",
        idempotency_key="safe-delete-intent-001",
        operated_by="demand-delete-admin",
        now=now,
    )
    db.commit()
    assert intent["status"] == "reviewed"
    assert [item["source_order_id"] for item in intent["items"]] == raw_ids
    assert all(item["order_no"].startswith("WBDD-") for item in intent["items"])
    assert all(item["line_count"] == 2 for item in intent["items"])
    assert all(item["downstream_references"] for item in intent["items"])

    armed = maintenance_demands.arm_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now,
    )
    db.commit()
    assert armed["not_before"] == now + timedelta(seconds=7)

    try:
        maintenance_demands.execute_delete_intent(
            db,
            intent_id=intent["intent_id"],
            digest=intent["selection_digest"],
            operated_by="demand-delete-admin",
            now=now + timedelta(seconds=6, milliseconds=990),
        )
        raise AssertionError("6.99 秒不应允许执行")
    except maintenance_demands.DeleteIntentTooEarly as exc:
        assert exc.not_before == now + timedelta(seconds=7)
        db.rollback()

    result = maintenance_demands.execute_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now + timedelta(seconds=7),
    )
    db.commit()
    replay = maintenance_demands.execute_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now + timedelta(seconds=8),
    )
    db.commit()

    assert result == replay
    assert result["status"] == "executed"
    assert result["header_count"] == 2
    assert result["line_count"] == 4
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandTombstone)) == 2
    assert db.scalar(
        select(func.count())
        .select_from(MaintenanceDemandDeleteEvent)
        .where(MaintenanceDemandDeleteEvent.event_type == "executed")
    ) == 1
    assert db.scalar(select(func.count()).select_from(FMaintenanceOrder)) == 2
    assert db.scalar(select(func.count()).select_from(FMaintenanceLine)) == 4
    assert maintenance_demands.search_demands(db, q=None, page=1, page_size=20)["total"] == 0


def test_execute_detects_any_item_change_and_leaves_batch_untouched(db):
    raw_ids = _seed_demands(db, count=2, lines_each=1)
    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=raw_ids,
        reason="合成并发冲突验证",
        idempotency_key="safe-delete-conflict-001",
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.arm_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now,
    )
    db.commit()

    changed = db.scalar(
        select(FMaintenanceOrder).where(FMaintenanceOrder.raw_order_id == raw_ids[1])
    )
    changed.project_std = "并发变化后的项目"
    db.commit()

    try:
        maintenance_demands.execute_delete_intent(
            db,
            intent_id=intent["intent_id"],
            digest=intent["selection_digest"],
            operated_by="demand-delete-admin",
            now=now + timedelta(seconds=7),
        )
        raise AssertionError("版本变化必须整批冲突")
    except maintenance_demands.DeleteIntentConflict:
        db.commit()

    assert db.scalar(select(func.count()).select_from(MaintenanceDemandTombstone)) == 0


def test_restore_stays_shadow_only_and_preserves_stable_cost_project_and_export(
    db,
    monkeypatch,
):
    settings = maintenance_demands.get_settings()
    monkeypatch.setattr(settings, "maintenance_cutover_enabled", False)
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    source_order = db.scalar(
        select(FMaintenanceOrder).where(FMaintenanceOrder.raw_order_id == raw_id)
    )
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    source_line.unit_cost = Decimal("100.00")
    source_line.cost_amount = Decimal("100.00")
    source_line.cost_source = "direct"
    source_line.cost_tax_basis = "ex"
    source_line.unit_cost_ex_tax = Decimal("100.00")
    source_line.unit_cost_inc_tax = Decimal("113.00")
    source_line.cost_amount_ex_tax = Decimal("100.00")
    source_line.cost_amount_inc_tax = Decimal("113.00")
    source_line.linked_purchase_order_no = "CGDD-STABLE"
    source_line.confidence = "high"
    db.commit()

    stable_cost_before = (
        source_line.unit_cost,
        source_line.cost_amount,
        source_line.cost_source,
        source_line.cost_tax_basis,
        source_line.unit_cost_inc_tax,
        source_line.unit_cost_ex_tax,
        source_line.cost_amount_inc_tax,
        source_line.cost_amount_ex_tax,
        source_line.linked_purchase_order_no,
        source_line.confidence,
        tuple(source_line.anomaly_flags or ()),
    )
    stable_project_before = maintenance_cost.projects_aggregate(db)["rows"]
    stable_export_before = _exported_first_line(db)

    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="Beta 影子删除恢复验证",
        idempotency_key="safe-delete-shadow-restore",
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.arm_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.execute_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now + timedelta(seconds=7),
    )
    db.commit()

    restored = maintenance_demands.restore_demand(
        db,
        source_order_id=raw_id,
        reason="Beta 恢复不得改写稳定成本",
        operated_by="demand-delete-admin",
        now=now + timedelta(minutes=1),
    )
    db.commit()

    assert restored["cost_state"] == "stable_unchanged"
    assert restored["invalidated_line_count"] == 0
    assert restored["cutover_enabled"] is False
    db.expire_all()
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    assert (
        source_line.unit_cost,
        source_line.cost_amount,
        source_line.cost_source,
        source_line.cost_tax_basis,
        source_line.unit_cost_inc_tax,
        source_line.unit_cost_ex_tax,
        source_line.cost_amount_inc_tax,
        source_line.cost_amount_ex_tax,
        source_line.linked_purchase_order_no,
        source_line.confidence,
        tuple(source_line.anomaly_flags or ()),
    ) == stable_cost_before
    assert maintenance_cost.projects_aggregate(db)["rows"] == stable_project_before
    assert _exported_first_line(db) == stable_export_before


def test_tombstone_survives_same_raw_order_reimport_and_controlled_restore(
    db,
    monkeypatch,
):
    settings = maintenance_demands.get_settings()
    monkeypatch.setattr(settings, "maintenance_cutover_enabled", True)
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    source_order = db.scalar(
        select(FMaintenanceOrder).where(FMaintenanceOrder.raw_order_id == raw_id)
    )
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    source_line.unit_cost = Decimal("100.00")
    source_line.cost_amount = Decimal("100.00")
    source_line.cost_source = "direct"
    source_line.cost_tax_basis = "ex"
    source_line.unit_cost_ex_tax = Decimal("100.00")
    source_line.unit_cost_inc_tax = Decimal("113.00")
    source_line.cost_amount_ex_tax = Decimal("100.00")
    source_line.cost_amount_inc_tax = Decimal("113.00")
    source_line.linked_purchase_order_no = "CGDD-STALE"
    source_line.confidence = "high"
    db.commit()
    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成重导防复活验证",
        idempotency_key="safe-delete-reimport-001",
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.arm_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.execute_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now + timedelta(seconds=7),
    )
    db.commit()

    second_batch = SysImportBatch(
        filename="synthetic-maintenance-reimport.xlsx",
        file_type="maintenance",
        file_hash="synthetic-maintenance-reimport",
        status="processing",
    )
    db.add(second_batch)
    db.flush()
    loader.load(
        db,
        f.maintenance_result(
            {
                raw_id: f.maintenance_head(
                    raw_id,
                    order_no="WBDD-REIMPORTED-0001",
                    on=date(2026, 8, 2),
                    project="重导后的项目名",
                )
            },
            [
                f.maintenance_line(
                    raw_id,
                    "RAW-LINE-0001-01",
                    "PN-SYNTH-1-1",
                    qty="3",
                    description="重导后的备件描述",
                )
            ],
        ),
        second_batch.id,
        date(2026, 8, 9),
        mode="upsert",
    )
    purchase_batch = SysImportBatch(
        filename="synthetic-purchase-evidence-change.xlsx",
        file_type="purchase",
        file_hash="synthetic-purchase-evidence-change",
        status="processing",
    )
    db.add(purchase_batch)
    db.flush()
    loader.load(
        db,
        f.purchase_result(
            {
                "RAW-CGDD-NEW": f.purchase_head(
                    "RAW-CGDD-NEW",
                    order_no="CGDD-NEW",
                    on=date(2026, 8, 2),
                    linked_maintenance_order_no="WBDD-REIMPORTED-0001",
                )
            },
            [
                f.purchase_line(
                    "RAW-CGDD-NEW",
                    "RAW-CGDD-LINE-NEW",
                    "PN-SYNTH-1-1",
                    qty="3",
                    price="250",
                )
            ],
        ),
        purchase_batch.id,
        date(2026, 8, 9),
        mode="upsert",
    )
    db.commit()
    assert maintenance_demands.search_demands(db, q=None, page=1, page_size=20)["total"] == 0
    db.expire_all()
    source_order = db.scalar(
        select(FMaintenanceOrder).where(FMaintenanceOrder.raw_order_id == raw_id)
    )
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    assert source_order.project_std == "重导后的项目名"
    assert source_line.qty == Decimal("3")
    # Reimport intentionally does not own derived cost columns, so the old
    # value remains archived while the tombstone is active.
    assert source_line.unit_cost == Decimal("100.00")

    restored = maintenance_demands.restore_demand(
        db,
        source_order_id=raw_id,
        reason="管理员确认恢复合成需求单",
        operated_by="demand-delete-admin",
        now=now + timedelta(minutes=1),
    )
    db.commit()
    assert restored["status"] == "restored"
    assert restored["cost_state"] == "pending_recompute"
    assert restored["invalidated_line_count"] == 1
    assert restored["cutover_enabled"] is True
    db.expire_all()
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    for field in (
        "unit_cost",
        "cost_amount",
        "cost_source",
        "cost_tax_basis",
        "unit_cost_inc_tax",
        "unit_cost_ex_tax",
        "cost_amount_inc_tax",
        "cost_amount_ex_tax",
        "linked_purchase_order_no",
        "confidence",
        "reference_side",
    ):
        assert getattr(source_line, field) is None
    assert "cost_recompute_pending" in source_line.anomaly_flags
    assert maintenance_demands.search_demands(db, q=None, page=1, page_size=20)["total"] == 1

    maintenance_cost.recompute(db)
    db.expire_all()
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    assert source_line.unit_cost == Decimal("250.00")
    assert source_line.cost_amount == Decimal("750.00")
    assert source_line.linked_purchase_order_no == "CGDD-NEW"
    assert "cost_recompute_pending" not in source_line.anomaly_flags

    # A cutover rollback is safe after the required recompute: the same
    # canonical cost, project aggregate and export remain visible instead of
    # resurrecting the archived pre-reimport cost snapshot.
    cutover_project = maintenance_cost.projects_aggregate(db)["rows"]
    cutover_export = _exported_first_line(db)
    monkeypatch.setattr(settings, "maintenance_cutover_enabled", False)
    db.expire_all()
    source_line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == source_order.id)
    )
    assert source_line.unit_cost == Decimal("250.00")
    assert source_line.cost_amount == Decimal("750.00")
    assert maintenance_cost.projects_aggregate(db)["rows"] == cutover_project
    assert _exported_first_line(db) == cutover_export


def test_restore_rejects_overlong_reason_without_reflection_log_or_audit(db, caplog):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    username = "demand-restore-private-admin"
    client = _admin_client(db, username=username)
    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成恢复理由上限验证",
        idempotency_key="safe-delete-private-restore-reason",
        operated_by=username,
        now=now,
    )
    maintenance_demands.arm_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by=username,
        now=now,
    )
    maintenance_demands.execute_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by=username,
        now=now + timedelta(seconds=7),
    )
    db.commit()
    sentinel = "RESTORE-PRIVATE-REASON-SENTINEL-" + "x" * 1_100

    response = client.post(
        f"/api/maintenance/demands/{raw_id}/restore",
        json={"reason": sentinel},
    )

    assert response.status_code == 422
    assert "PRIVATE-REASON-SENTINEL" not in response.text
    assert sentinel not in caplog.text
    db.expire_all()
    tombstone = db.get(MaintenanceDemandTombstone, raw_id)
    assert tombstone.restored_at is None
    assert (
        db.scalar(
            select(func.count())
            .select_from(MaintenanceDemandDeleteEvent)
            .where(MaintenanceDemandDeleteEvent.event_type == "restored")
        )
        == 0
    )
    access_rows = list(
        db.scalars(select(SysAccessLog).where(SysAccessLog.username == username))
    )
    assert all(sentinel not in str(row.detail) for row in access_rows)


def test_restore_rejects_overlong_source_id_without_reflection_or_audit(db, caplog):
    username = "demand-restore-private-source-admin"
    client = _admin_client(db, username=username)
    sentinel = "RESTORE-PRIVATE-SOURCE-ID-SENTINEL-" + "x" * 64
    access_count_before = db.scalar(
        select(func.count())
        .select_from(SysAccessLog)
        .where(SysAccessLog.username == username)
    )

    response = client.post(
        f"/api/maintenance/demands/{sentinel}/restore",
        json={"reason": "合成恢复路径 ID 上限验证"},
    )

    assert response.status_code == 422
    assert "PRIVATE-SOURCE-ID-SENTINEL" not in response.text
    assert sentinel not in caplog.text
    db.expire_all()
    assert (
        db.scalar(
            select(func.count())
            .select_from(SysAccessLog)
            .where(SysAccessLog.username == username)
        )
        == access_count_before
    )
    assert (
        db.scalar(
            select(func.count())
            .select_from(MaintenanceDemandDeleteEvent)
            .where(MaintenanceDemandDeleteEvent.event_type == "restored")
        )
        == 0
    )


def test_write_endpoints_fail_closed_for_anonymous_and_shared_password(db):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    body = {
        "source_order_ids": [raw_id],
        "reason": "无权限请求不应建立意图",
        "idempotency_key": "unsafe-auth-intent",
    }
    anonymous = TestClient(app)
    assert anonymous.post("/api/maintenance/demands/delete-intents", json=body).status_code == 401

    shared = TestClient(app)
    login = shared.post(
        "/api/auth/login",
        json={"username": "admin", "password": "admin"},
    )
    assert login.status_code == 200, login.text
    shared.headers["Authorization"] = f"Bearer {login.json()['token']}"
    assert shared.post("/api/maintenance/demands/delete-intents", json=body).status_code == 403


def test_beta_tombstone_changes_stable_reads_only_after_explicit_cutover(db, monkeypatch):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    order = db.scalar(
        select(FMaintenanceOrder).where(FMaintenanceOrder.raw_order_id == raw_id)
    )
    line = db.scalar(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == order.id)
    )
    assert inventory.dynamic_stock_map(db)[line.part_id]["out_maint"] == 1
    assert maintenance_cost.projects_aggregate(db)["rows"][0]["order_count"] == 1

    now = datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)
    intent = maintenance_demands.create_delete_intent(
        db,
        source_order_ids=[raw_id],
        reason="合成有效读模型排除验证",
        idempotency_key="safe-delete-effective-reads",
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.arm_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now,
    )
    maintenance_demands.execute_delete_intent(
        db,
        intent_id=intent["intent_id"],
        digest=intent["selection_digest"],
        operated_by="demand-delete-admin",
        now=now + timedelta(seconds=7),
    )
    db.commit()

    # Beta deletion is a shadow fact until the explicit business cutover.  It
    # disappears from Beta demand/search consumers, but the production-stable
    # cost, inventory and export paths keep their pre-Beta truth unchanged.
    assert maintenance_demands.search_demands(
        db,
        q=raw_id,
        page=1,
        page_size=20,
    )["total"] == 0
    assert inventory.dynamic_stock_map(db)[line.part_id]["out_maint"] == 1
    assert maintenance_cost.projects_aggregate(db)["rows"][0]["order_count"] == 1
    workbook = maintenance_export.build_workbook(
        db,
        UserContext(user_id="demand-delete-admin", role="admin"),
    )
    try:
        workbook.seek(0)
        assert workbook.read(2) == b"PK"
    finally:
        workbook.close()

    settings = maintenance_demands.get_settings()
    monkeypatch.setattr(settings, "maintenance_cutover_enabled", True)
    cutover_stock = inventory.dynamic_stock_map(db)
    assert cutover_stock.get(line.part_id, {}).get("out_maint", Decimal(0)) == 0
    assert maintenance_cost.projects_aggregate(db)["rows"] == []
    with pytest.raises(maintenance_export.ExcelExportEmpty):
        maintenance_export.build_workbook(
            db,
            UserContext(user_id="demand-delete-admin", role="admin"),
        )

    monkeypatch.setattr(settings, "maintenance_cutover_enabled", False)
    assert inventory.dynamic_stock_map(db)[line.part_id]["out_maint"] == 1
    assert maintenance_cost.projects_aggregate(db)["rows"][0]["order_count"] == 1


def test_api_uses_server_clock_for_exact_seven_second_gate(db, monkeypatch):
    raw_id = _seed_demands(db, count=1, lines_each=1)[0]
    client = _admin_client(db, "demand-delete-clock-admin")
    clock = {"now": datetime(2026, 8, 9, 12, 0, tzinfo=timezone.utc)}
    monkeypatch.setattr(maintenance_demands, "_utc_now", lambda: clock["now"])

    created = client.post(
        "/api/maintenance/demands/delete-intents",
        json={
            "source_order_ids": [raw_id],
            "reason": "服务端时钟验证",
            "idempotency_key": "safe-delete-server-clock",
        },
    ).json()
    armed = client.post(
        f"/api/maintenance/demands/delete-intents/{created['intent_id']}/arm",
        json={"digest": created["selection_digest"]},
    )
    assert armed.status_code == 200, armed.text

    clock["now"] += timedelta(seconds=6, milliseconds=990)
    too_early = client.post(
        f"/api/maintenance/demands/delete-intents/{created['intent_id']}/execute",
        json={"digest": created["selection_digest"]},
    )
    assert too_early.status_code == 425
    assert too_early.json()["detail"]["not_before"].endswith("+00:00")
    assert db.scalar(select(func.count()).select_from(MaintenanceDemandTombstone)) == 0

    clock["now"] += timedelta(milliseconds=10)
    executed = client.post(
        f"/api/maintenance/demands/delete-intents/{created['intent_id']}/execute",
        json={"digest": created["selection_digest"]},
    )
    assert executed.status_code == 200, executed.text
    assert executed.json()["status"] == "executed"
