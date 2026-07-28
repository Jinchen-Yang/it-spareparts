"""维保页结构性权限：隐藏金额之外，分类、排序、筛选与 Agent 截断也不能泄漏。"""
import io
from datetime import date
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app import permissions, security
from app.agent import tools
from app.auth import hash_password
from app.etl import loader
from app.main import app
from app.models.maintenance import FMaintenanceLine
from app.models.system import SysImportBatch, SysUser
from app.services import maintenance_cost, maintenance_cost_quality
from tests import factories as f


@pytest.fixture()
def maintenance_permission_data(db):
    batch = SysImportBatch(
        filename="maintenance-permission.xlsx", file_type="maintenance",
        file_hash="maintenance-permission-structure",
    )
    db.add(batch)
    db.flush()

    specs = [
        # 项目名字母序 A/M/Z 与成本序 Z/M/A、利润状态序 Z/M/A 相反；
        # 出库日期序 A/M/Z，用来证明受限结果没有沿用隐藏分类/金额排序。
        ("A", "A-low", date(2026, 3, 30), "100"),
        ("M", "M-mid", date(2026, 3, 20), "950"),
        ("Z", "Z-high", date(2026, 3, 10), "1200"),
    ]
    for tag, project, out_date, cost in specs:
        loader.load(db, f.sales_result(
            {f"S-{tag}": f.sales_head(
                f"S-{tag}", order_no=f"XS-{tag}", amount_ex_tax=Decimal("1000"),
            )},
            [f.sales_line(f"S-{tag}", f"SL-{tag}", f"PN-{tag}", qty="1", price="1")],
        ), batch.id, date(2026, 6, 1))
        loader.load(db, f.purchase_result(
            {f"P-{tag}": f.purchase_head(
                f"P-{tag}", on=date(2026, 3, 5), source_type="维保需求",
                linked_maintenance_order_no=f"WBDD-{tag}",
            )},
            [f.purchase_line(
                f"P-{tag}", f"PL-{tag}", f"PN-{tag}", qty="1", price=cost,
            )],
        ), batch.id, date(2026, 6, 1))
        loader.load(db, f.maintenance_result(
            {f"M-{tag}": f.maintenance_head(
                f"M-{tag}", order_no=f"WBDD-{tag}", on=out_date,
                sales_order=f"XS-{tag}", project=project,
                maint_end=date(2027, 12, 31),
            )},
            [f.maintenance_line(
                f"M-{tag}", f"ML-{tag}", f"PN-{tag}", qty="1",
            )],
        ), batch.id, date(2026, 6, 1))
    db.commit()
    maintenance_cost.recompute(db)
    return specs


def _limited_client(db, username: str, *, cost: bool, profit: bool) -> TestClient:
    base = permissions.effective("readonly", None)
    overrides = {
        "page_maintenance": True,
        "data_purchase_cost": cost,
        "data_profit": profit,
    }
    effective = permissions.effective_from_snapshot(base, overrides)
    assert permissions.combo_errors(effective) == []
    db.add(SysUser(
        username=username, role="readonly", display_name=username,
        password_hash=hash_password("pw123456"), is_active=True,
        template_code="readonly", template_version=1, template_perms=base,
        perm_overrides=overrides, permissions=effective,
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login", json={"username": username, "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _ctx(*, cost: bool, profit: bool) -> security.UserContext:
    perms = permissions.effective("readonly", {
        "page_maintenance": True,
        "data_purchase_cost": cost,
        "data_profit": profit,
    })
    return security.UserContext(
        user_id="limited", role="readonly", permissions=perms, is_authenticated=True,
    )


def test_api_profit_blind_board_removes_status_filter_counts_and_ranking(
    db, maintenance_permission_data,
):
    client = _limited_client(db, "maint-profit-blind", cost=True, profit=False)
    response = client.get("/api/maintenance/board", params={"status": "red"})
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["profit_restricted"] is True
    assert data["ranking_restricted"] is True
    assert data["effective_sort"] == "last_out"
    assert data["status_filter_applied"] is False
    assert "status_counts" not in data
    assert "decision_status_counts" not in data
    assert len(data["rows"]) == 3                 # red 筛选被忽略，不泄漏命中集合
    assert all(
        "status" not in row and "decision_status" not in row
        for row in data["rows"]
    )
    assert all(row["cost_quality"] == "actual_only" for row in data["rows"])
    assert all(row["actual_lines"] == 1 for row in data["rows"])
    assert all(row["known_cost_total"] is not None for row in data["rows"])
    assert all(row["spent_expense"] is None for row in data["rows"])
    assert all(row["expense_data_available"] is False for row in data["rows"])
    assert all(row["budget"] is None for row in data["rows"])
    assert all(row["remaining"] is None for row in data["rows"])
    assert all(row["remaining_pct"] is None for row in data["rows"])
    for row in data["rows"]:
        for field in (
            "revenue_inc", "revenue_ex",
            "parts_gross_profit_inc", "parts_gross_profit_ex",
            "parts_gross_margin_inc", "parts_gross_margin_ex",
            "parts_profit_status_inc", "parts_profit_status_ex",
            "contribution_profit_inc", "contribution_profit_ex",
            "contribution_margin_inc", "contribution_margin_ex",
            "contribution_status_inc", "contribution_status_ex",
        ):
            assert row[field] is None
    assert [row["contract"] for row in data["rows"]] == ["XS-A", "XS-M", "XS-Z"]

    projects_response = client.get("/api/maintenance/projects")
    assert projects_response.status_code == 200, projects_response.text
    project_rows = projects_response.json()["rows"]
    assert all(row["known_cost_total"] is not None for row in project_rows)
    assert all(row["contract_amount"] is None for row in project_rows)


def test_api_cost_blind_projects_use_neutral_sort_before_any_consumer_truncation(
    db, maintenance_permission_data,
):
    client = _limited_client(db, "maint-cost-blind", cost=False, profit=False)
    response = client.get("/api/maintenance/projects")
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["ranking_restricted"] is True
    assert data["effective_sort"] == "project"
    assert [row["project"] for row in data["rows"]] == ["A-low", "M-mid", "Z-high"]
    assert all(row["cost_total"] is None for row in data["rows"])
    assert all(row["known_cost_total"] is None for row in data["rows"])
    assert all(row["actual_lines"] is None for row in data["rows"])
    assert all(row["estimated_lines"] is None for row in data["rows"])
    assert all(row["missing_cost_lines"] is None for row in data["rows"])
    assert all(row["cost_quality"] is None for row in data["rows"])
    assert all(row["by_source"] is None for row in data["rows"])
    assert all(row["coverage_pct"] is None for row in data["rows"])
    assert all(row["parts_cost_inc_tax"] is None for row in data["rows"])
    assert all(row["parts_cost_ex_tax"] is None for row in data["rows"])
    assert all(row["parts_cost_inc_tax_complete"] is None for row in data["rows"])
    assert all(row["parts_cost_ex_tax_complete"] is None for row in data["rows"])
    assert all(row["parts_cost_inc_tax_quality"] is None for row in data["rows"])
    assert all(row["parts_cost_ex_tax_quality"] is None for row in data["rows"])


def test_agent_uses_same_board_collapse_and_neutral_project_truncation(
    db, maintenance_permission_data,
):
    profit_blind_ctx = _ctx(cost=True, profit=False)
    profit_blind = tools.dispatch(
        db, "get_maintenance_board", {"status": "red"},
        profit_blind_ctx,
    )
    assert profit_blind["profit_restricted"] is True
    assert "status_counts" not in profit_blind
    assert "decision_status_counts" not in profit_blind
    assert len(profit_blind["rows"]) == 3
    assert all(
        "status" not in row and "decision_status" not in row
        for row in profit_blind["rows"]
    )
    assert all(row["known_cost_total"] is not None for row in profit_blind["rows"])
    assert all(row["budget"] is None for row in profit_blind["rows"])
    assert all(row["remaining"] is None for row in profit_blind["rows"])
    assert all(row["remaining_pct"] is None for row in profit_blind["rows"])
    for row in profit_blind["rows"]:
        for field in (
            "revenue_inc", "revenue_ex",
            "parts_gross_profit_inc", "parts_gross_profit_ex",
            "parts_gross_margin_inc", "parts_gross_margin_ex",
            "parts_profit_status_inc", "parts_profit_status_ex",
            "contribution_profit_inc", "contribution_profit_ex",
            "contribution_margin_inc", "contribution_margin_ex",
            "contribution_status_inc", "contribution_status_ex",
        ):
            assert row[field] is None

    cost_blind = tools.dispatch(
        db, "get_maintenance_projects", {"top": 2},
        _ctx(cost=False, profit=False),
    )
    assert cost_blind["ranking_restricted"] is True
    assert cost_blind["effective_sort"] == "project"
    assert [row["project"] for row in cost_blind["rows"]] == ["A-low", "M-mid"]
    assert all(row["cost_total"] is None for row in cost_blind["rows"])
    assert "按项目名" in cost_blind["note"] and "成本最高" not in cost_blind["note"]

    cost_blind_board = tools.dispatch(
        db, "get_maintenance_board", {"status": "red"},
        _ctx(cost=False, profit=False),
    )
    assert cost_blind_board["ranking_restricted"] is True
    assert cost_blind_board["status_filter_applied"] is False
    assert "status_counts" not in cost_blind_board
    assert "decision_status_counts" not in cost_blind_board
    for row in cost_blind_board["rows"]:
        assert "status" not in row and "decision_status" not in row
        for field in (
            "actual_cost_inc", "actual_cost_ex",
            "estimated_cost_inc", "estimated_cost_ex",
            "actual_lines", "estimated_lines", "missing_cost_lines",
            "known_cost_total", "cost_quality", "coverage_pct",
            "parts_cost_inc_tax", "parts_cost_ex_tax",
            "parts_cost_inc_tax_complete", "parts_cost_ex_tax_complete",
            "parts_cost_inc_tax_quality", "parts_cost_ex_tax_quality",
            "parts_cost_inc_tax_missing_lines", "parts_cost_ex_tax_missing_lines",
        ):
            assert row[field] is None
        for project in row["projects"]:
            assert project["actual_lines"] is None
            assert project["estimated_lines"] is None
            assert project["missing_cost_lines"] is None
            assert project["cost_quality"] is None
            assert project["parts_cost_inc_tax"] is None
            assert project["parts_cost_ex_tax"] is None
            assert project["parts_cost_inc_tax_complete"] is None
            assert project["parts_cost_ex_tax_complete"] is None

    skills = tools.dispatch(db, "list_skills", {}, profit_blind_ctx)["skills"]
    assert "maintenance_health_check" not in {item["skill"] for item in skills}


def test_default_purchaser_template_keeps_cost_facts_but_hides_budget_decisions(
    db,
    maintenance_permission_data,
):
    purchaser_permissions = permissions.effective("purchaser", None)
    assert purchaser_permissions["page_maintenance"] is True
    assert purchaser_permissions["data_purchase_cost"] is True
    assert purchaser_permissions["data_profit"] is False
    ctx = security.UserContext(
        user_id="default-purchaser",
        role="purchaser",
        permissions=purchaser_permissions,
        is_authenticated=True,
    )

    projects = maintenance_cost.projects_aggregate(db, user_ctx=ctx)
    assert all(row["known_cost_total"] is not None for row in projects["rows"])
    assert all(row["cost_quality"] == "actual_only" for row in projects["rows"])
    assert all(row["contract_amount"] is None for row in projects["rows"])

    board = maintenance_cost.board(db, status="red", user_ctx=ctx)
    assert board["profit_restricted"] is True
    assert board["decision_restricted"] is True
    assert board["status_filter_applied"] is False
    assert board["effective_sort"] == "last_out"
    assert "status_counts" not in board
    assert "decision_status_counts" not in board
    assert len(board["rows"]) == 3
    for row in board["rows"]:
        assert row["known_cost_total"] is not None
        assert row["cost_quality"] == "actual_only"
        assert row["budget"] is None
        assert row["remaining"] is None
        assert row["remaining_pct"] is None
        assert "status" not in row
        assert "decision_status" not in row

    agent_board = tools.dispatch(
        db,
        "get_maintenance_board",
        {"status": "red"},
        ctx,
    )
    assert agent_board["status_filter_applied"] is False
    assert agent_board["effective_sort"] == "last_out"
    assert "status_counts" not in agent_board
    assert "decision_status_counts" not in agent_board
    assert all(row["known_cost_total"] is not None for row in agent_board["rows"])
    assert all(row["budget"] is None for row in agent_board["rows"])
    assert all("decision_status" not in row for row in agent_board["rows"])


def test_boss_sees_same_incomplete_cost_gate_across_api_agent_and_workbook(
    db,
    maintenance_permission_data,
):
    boss_permissions = permissions.effective("boss", None)
    assert boss_permissions["page_maintenance"] is True
    assert boss_permissions["data_purchase_cost"] is True
    assert boss_permissions["data_profit"] is True

    missing_line = db.execute(
        select(FMaintenanceLine).where(FMaintenanceLine.raw_line_id == "ML-A"),
    ).scalar_one()
    missing_line.cost_source = "none"
    missing_line.cost_tax_basis = None
    missing_line.unit_cost = None
    missing_line.cost_amount = None
    db.add(SysUser(
        username="maintenance-boss",
        role="boss",
        display_name="维保老板",
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="boss",
        template_version=1,
        template_perms=boss_permissions,
        perm_overrides={},
        permissions=boss_permissions,
    ))
    db.commit()

    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance-boss", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})

    projects_response = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    assert projects_response.status_code == 200, projects_response.text
    project = next(
        row
        for row in projects_response.json()["rows"]
        if row["project"] == "A-low"
    )
    assert project["actual_lines"] == 0
    assert project["estimated_lines"] == 0
    assert project["missing_cost_lines"] == 1
    assert project["known_cost_total"] == 0.0
    assert project["cost_quality"] == "incomplete"
    assert project["contract_amount"] is not None

    board_response = client.get(
        "/api/maintenance/board",
        params={"status": "incomplete_cost", "lifecycle": "all"},
    )
    assert board_response.status_code == 200, board_response.text
    board = board_response.json()
    assert board["status_filter_applied"] is True
    assert board["decision_status_counts"]["incomplete_cost"] == 1
    assert [row["contract"] for row in board["rows"]] == ["XS-A"]
    boss_row = board["rows"][0]
    assert boss_row["decision_status"] == boss_row["status"] == "incomplete_cost"
    assert boss_row["cost_quality"] == "incomplete"
    assert boss_row["missing_cost_lines"] == 1
    assert boss_row["remaining"] is None
    assert boss_row["remaining_pct"] is None

    boss_ctx = security.UserContext(
        user_id="maintenance-boss",
        role="boss",
        permissions=boss_permissions,
        is_authenticated=True,
    )
    agent_board = tools.dispatch(
        db,
        "get_maintenance_board",
        {"status": "incomplete_cost"},
        boss_ctx,
    )
    assert agent_board["decision_status_counts"]["incomplete_cost"] == 1
    assert [row["contract"] for row in agent_board["rows"]] == ["XS-A"]
    assert agent_board["rows"][0]["decision_status"] == "incomplete_cost"
    assert agent_board["rows"][0]["remaining"] is None
    assert agent_board["rows"][0]["remaining_pct"] is None

    workbook_response = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XS-A"},
    )
    assert workbook_response.status_code == 200, workbook_response.text
    workbook = load_workbook(
        io.BytesIO(workbook_response.content),
        read_only=True,
        data_only=True,
    )
    try:
        assert workbook.sheetnames == [
            "项目预算",
            "备件明细-氚云",
            "报销明细",
            "填写说明",
        ]
    finally:
        workbook.close()


def test_agent_board_reuses_service_decision_without_recomputing(monkeypatch):
    captured = {}
    service_result = {
        "rows": [{
            "contract": "XS-SERVICE-TRUTH",
            "cost_quality": "incomplete",
            "known_cost_total": 1.0,
            "budget": 1000.0,
            "decision_status": "incomplete_cost",
            "status": "incomplete_cost",
            "remaining": None,
            "remaining_pct": None,
        }],
        "status_counts": {"red": 0, "yellow": 0, "green": 0, "no_budget": 0},
        "decision_status_counts": {
            "incomplete_cost": 1,
            "red": 0,
            "yellow": 0,
            "green": 0,
            "no_budget": 0,
        },
    }

    def fake_board(db, date_from, date_to, status, *, user_ctx):
        captured.update({
            "db": db,
            "date_from": date_from,
            "date_to": date_to,
            "status": status,
            "user_ctx": user_ctx,
        })
        return service_result

    monkeypatch.setattr(maintenance_cost, "board", fake_board)

    ctx = _ctx(cost=True, profit=True)
    result = tools.dispatch(
        None,
        "get_maintenance_board",
        {"status": "incomplete_cost"},
        ctx,
    )

    assert captured["status"] == "incomplete_cost"
    assert captured["user_ctx"] is ctx
    assert result["rows"][0]["decision_status"] == "incomplete_cost"
    assert result["rows"][0]["remaining"] is None
    assert result["rows"][0]["remaining_pct"] is None
    assert result["decision_status_counts"]["incomplete_cost"] == 1

    schema = next(
        item["function"]
        for item in tools.TOOLS
        if item["function"]["name"] == "get_maintenance_board"
    )
    assert schema["parameters"]["properties"]["status"]["enum"] == [
        "incomplete_cost",
        "expense_data_unavailable",
        "red",
        "yellow",
        "green",
        "no_budget",
    ]


def test_cost_blind_lines_remove_cost_derived_anomaly_flags(
    db, maintenance_permission_data,
):
    line = db.execute(
        select(FMaintenanceLine).where(FMaintenanceLine.raw_line_id == "ML-A"),
    ).scalar_one()
    line.anomaly_flags = [
        *maintenance_cost_quality.COST_DERIVED_ANOMALY_FLAGS,
        "future_date",
    ]
    db.commit()

    client = _limited_client(db, "maint-line-cost-blind", cost=False, profit=False)
    response = client.get("/api/maintenance/lines", params={"project": "A-low"})

    assert response.status_code == 200, response.text
    assert response.json()["rows"][0]["anomaly_flags"] == ["future_date"]
    assert response.json()["rows"][0]["cost_tier"] is None


def test_new_maintenance_fields_are_registered_without_masking_generic_status():
    purchase_fields = set(permissions.config.FIELD_GROUPS["purchase_cost"])
    assert {
        "actual_cost_inc", "actual_cost_ex",
        "estimated_cost_inc", "estimated_cost_ex", "known_cost_total",
        "actual_lines", "estimated_lines", "missing_cost_lines",
        "cost_quality", "cost_tier", "by_source", "coverage_pct",
        "unit_cost_inc_tax", "unit_cost_ex_tax",
        "cost_amount_inc_tax", "cost_amount_ex_tax",
        "parts_cost_inc_tax", "parts_cost_ex_tax",
        "parts_cost_inc_tax_complete", "parts_cost_ex_tax_complete",
        "parts_cost_inc_tax_quality", "parts_cost_ex_tax_quality",
        "parts_cost_inc_tax_missing_lines", "parts_cost_ex_tax_missing_lines",
        "reference_side", "reference_pool_group_id", "reference_pool_version",
        "reference_sample_count", "reference_from_date", "reference_to_date",
        "reference_latest_date",
    } <= purchase_fields
    profit_amount_fields = set(permissions.config.FIELD_GROUPS["profit_amount"])
    profit_rate_fields = set(permissions.config.FIELD_GROUPS["profit_rate"])
    assert {
        "parts_gross_profit_inc", "parts_gross_profit_ex",
        "parts_profit_status_inc", "parts_profit_status_ex",
        "contribution_profit_inc", "contribution_profit_ex",
        "contribution_status_inc", "contribution_status_ex",
    } <= profit_amount_fields
    assert {
        "parts_gross_margin_inc", "parts_gross_margin_ex",
        "contribution_margin_inc", "contribution_margin_ex",
    } <= profit_rate_fields
    assert {
        "gross_profit_inc", "gross_profit_ex",
        "profit_status_inc", "profit_status_ex",
    }.isdisjoint(profit_amount_fields)
    assert {
        "gross_margin_inc", "gross_margin_ex",
    }.isdisjoint(profit_rate_fields)
    assert "decision_status" in profit_amount_fields
    assert {
        "contract_amount", "budget", "remaining", "remaining_pct",
    } <= profit_amount_fields
    assert "status" not in {
        field
        for fields in permissions.config.FIELD_GROUPS.values()
        for field in fields
    }
