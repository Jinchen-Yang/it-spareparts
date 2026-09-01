"""维保期限双源 P1：project.period_* 唯一事实源的写入/读取契约（K3-F）。

覆盖：helper 四态与倒置拒绝、provenance 互斥、no-op/provenance-only 不增版本、
catalog PATCH 双表同步、manager workbook apply 后 project 与 projection 一致、
旧版本冲突与 scope 漂移零写、多变化单项目只 +1、operations/reminders 读
canonical project 期限（即使 projection 冲突）、只读漂移分类 helper。
"""

from __future__ import annotations

import io
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from sqlalchemy import select

from app.auth import hash_password
from app.models.maintenance_ledger import MaintenanceLedgerImportBatch
from app.models.maintenance_manager import (
    MaintenanceCollectionMilestone,
    MaintenanceManagerUploadBatch,
    MaintenanceServicePeriod,
)
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectAuditLog,
    MaintenanceProjectContract,
    MaintenanceProjectUserAssignment,
)
from app.models.system import SysUser
from app.security import UserContext
from app.services import maintenance_boss_board
from app.services import maintenance_project
from app.services import maintenance_periods
from app.services import maintenance_project_catalog as catalog
from app.services import maintenance_project_operations as operations
from app.services.maintenance_collection_reminders import (
    get_project_collection_milestones,
    search_collection_reminders,
)
from app.services.maintenance_manager_workbook_adapter import (
    MaintenanceManagerWorkbookAdapter,
    ManagerWorkbookConflict,
)
from app.services.maintenance_manager_workbook_v3 import (
    OVERVIEW_SHEET,
    OVERVIEW_TABLE,
    PLAN_SHEET,
    PLAN_TABLE,
)

HMAC_KEY = b"synthetic-period-canonical-key"
AS_OF = date(2026, 8, 15)
REPORT_MONTH = date(2026, 8, 1)


def _ctx(username: str) -> UserContext:
    return UserContext(
        role="admin",
        user_id=username,
        is_authenticated=True,
        permissions=None,
    )


def _user(db, *, username: str) -> SysUser:
    user = SysUser(
        username=username,
        role="admin",
        display_name=f"合成{username}",
        password_hash=hash_password("synthetic-password-123"),
    )
    db.add(user)
    db.commit()
    return user


def _project(
    db,
    *,
    suffix: str,
    manager: SysUser | None = None,
    period_from: date | None = None,
    period_to: date | None = None,
    lifecycle_status: str = "missing",
) -> tuple[MaintenanceProject, MaintenanceProjectContract]:
    project = MaintenanceProject(
        project_id=f"period-project-{suffix}",
        project_code=f"PERIOD-{suffix}",
        display_name=f"合成期限项目 {suffix}",
        period_from=period_from,
        period_to=period_to,
        lifecycle_status=lifecycle_status,
        is_active=True,
    )
    db.add(project)
    db.flush()
    if manager is not None:
        db.add(
            MaintenanceProjectUserAssignment(
                assignment_id=f"period-assignment-{suffix}",
                project_id=project.project_id,
                responsibility_type="primary_manager",
                user_id=manager.id,
                assigned_at=datetime.now(UTC),
                assigned_by="synthetic-admin",
                assignment_reason="合成期限负责人映射",
            )
        )
    contract = MaintenanceProjectContract(
        project_contract_id=f"period-pc-{suffix}",
        project_id=project.project_id,
        contract_id=f"period-contract-{suffix}",
        contract_no=f"XS-PERIOD-{suffix}",
        contract_amount=Decimal("100000.00"),
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=date(2026, 1, 1),
        source="synthetic-test",
    )
    db.add(contract)
    db.commit()
    return project, contract


def _ledger_batch(db, *, batch_id: str) -> str:
    db.add(
        MaintenanceLedgerImportBatch(
            batch_id=batch_id,
            file_hash=f"h-{batch_id}",
            filename="台账.xlsx",
            idempotency_key=f"key-{batch_id}",
            source_kind="project_manager_xls_v1",
            uploaded_by="合成管理员",
            status="applied",
            applied_by="合成管理员",
            applied_at=datetime(2026, 8, 1, tzinfo=UTC),
        )
    )
    db.flush()
    return batch_id


def _manager_batch(db, user: SysUser, *, batch_id: str) -> str:
    db.add(
        MaintenanceManagerUploadBatch(
            batch_id=batch_id,
            owner_user_id=user.id,
            report_month=REPORT_MONTH,
            protocol_version="v3",
            template_version="synthetic-tpl",
            export_id=f"export-{batch_id}",
            file_sha256="a" * 64,
            file_size=100,
            operation_key=f"operation-{batch_id}",
            semantic_hash="b" * 64,
            scope_version="c" * 64,
            data_version="d" * 64,
            status="valid",
            plan_json={},
            issues_json=[],
            created_by="synthetic-test",
            created_at=datetime.now(UTC),
            expires_at=datetime.now(UTC) + timedelta(hours=24),
        )
    )
    db.flush()
    return batch_id


def _lock_state_then_project(db, project_id: str) -> MaintenanceProject:
    """调用方锁序契约：先 workbook state，再 project。"""
    operations.get_or_create_workbook_state(db, project_id=project_id, lock=True)
    return db.scalar(
        select(MaintenanceProject)
        .where(MaintenanceProject.project_id == project_id)
        .with_for_update()
    )


def _projection(db, project_id: str) -> MaintenanceServicePeriod | None:
    return db.scalar(
        select(MaintenanceServicePeriod).where(
            MaintenanceServicePeriod.project_id == project_id
        )
    )


def _edit_workbook(
    content: bytes,
    *,
    project_id: str,
    relation_id: str,
    service_start=...,
    service_end=...,
    planned_date=...,
    planned_amount=...,
) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        overview = book[OVERVIEW_SHEET]
        table = overview.tables[OVERVIEW_TABLE]
        min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
        headers = [
            overview.cell(min_row, column).value
            for column in range(min_col, _max_col + 1)
        ]
        overview_row = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if overview.cell(row, headers.index("项目ID") + 1).value == project_id
        )
        if service_start is not ...:
            overview.cell(
                overview_row, headers.index("维保开始日期") + 1, service_start
            )
        if service_end is not ...:
            overview.cell(overview_row, headers.index("维保结束日期") + 1, service_end)
        if planned_date is not ... or planned_amount is not ...:
            plan = book[PLAN_SHEET]
            plan_table = plan.tables[PLAN_TABLE]
            p_min_col, p_min_row, p_max_col, p_max_row = range_boundaries(
                plan_table.ref
            )
            p_headers = [
                plan.cell(p_min_row, column).value
                for column in range(p_min_col, p_max_col + 1)
            ]
            plan_row = next(
                row
                for row in range(p_min_row + 1, p_max_row + 1)
                if plan.cell(row, p_headers.index("项目合同关系ID") + 1).value
                == relation_id
                and plan.cell(row, p_headers.index("计划期次") + 1).value == 1
            )
            if planned_date is not ...:
                plan.cell(plan_row, p_headers.index("计划回款日期") + 1, planned_date)
            if planned_amount is not ...:
                plan.cell(
                    plan_row, p_headers.index("计划回款金额（含税）") + 1, planned_amount
                )
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _adapter(db, user: SysUser) -> MaintenanceManagerWorkbookAdapter:
    return MaintenanceManagerWorkbookAdapter(
        db,
        user_ctx=_ctx(user.username),
        operator=user.username,
        as_of=AS_OF,
    )


# ---------- helper 纯函数与写入契约 ----------


def test_completeness_state_and_lifecycle_status_pure_functions():
    assert maintenance_periods.completeness_state(None, None) == "empty"
    assert maintenance_periods.completeness_state(date(2026, 1, 1), None) == "start_only"
    assert maintenance_periods.completeness_state(None, date(2026, 12, 31)) == "end_only"
    assert (
        maintenance_periods.completeness_state(date(2026, 1, 1), date(2026, 12, 31))
        == "complete"
    )

    as_of = date(2026, 8, 15)
    assert maintenance_periods.lifecycle_status(None, None, as_of) == "missing"
    assert (
        maintenance_periods.lifecycle_status(date(2026, 1, 1), date(2026, 12, 31), as_of)
        == "ongoing"
    )
    assert (
        maintenance_periods.lifecycle_status(date(2025, 1, 1), date(2025, 12, 31), as_of)
        == "ended"
    )
    # 只有起点且未结束 → ongoing；只有终点且已过期 → ended
    assert maintenance_periods.lifecycle_status(date(2026, 1, 1), None, as_of) == "ongoing"
    assert maintenance_periods.lifecycle_status(None, date(2026, 1, 1), as_of) == "ended"
    # 尚未开始：按既有口径仍是 missing
    assert (
        maintenance_periods.lifecycle_status(date(2027, 1, 1), date(2027, 12, 31), as_of)
        == "missing"
    )


def test_helper_rejects_inverted_range_without_writes(db):
    project, _contract = _project(db, suffix="inverted")
    locked = _lock_state_then_project(db, project.project_id)

    with pytest.raises(maintenance_periods.MaintenancePeriodError):
        maintenance_periods.apply_canonical_period_locked(
            db,
            project=locked,
            period_from=date(2026, 12, 31),
            period_to=date(2026, 1, 1),
            source=maintenance_periods.SOURCE_DIRECT_API,
            as_of=AS_OF,
            operated_by="synthetic-admin",
            reason="合成倒置期限",
        )
    db.rollback()
    assert _projection(db, project.project_id) is None
    reloaded = db.get(MaintenanceProject, project.project_id)
    assert reloaded.period_from is None and reloaded.period_to is None
    assert reloaded.version == 1


def test_helper_enforces_mutually_exclusive_provenance(db):
    manager = _user(db, username="period_prov_manager")
    project, _contract = _project(db, suffix="prov")
    ledger_batch = _ledger_batch(db, batch_id="ledger-batch-prov")
    manager_batch = _manager_batch(db, manager, batch_id="manager-batch-prov")
    locked = _lock_state_then_project(db, project.project_id)

    ledger_result = maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
        source=maintenance_periods.SOURCE_LEDGER,
        ledger_batch_id=ledger_batch,
        as_of=AS_OF,
        operated_by="synthetic-admin",
        reason="台账期限写入",
    )
    assert ledger_result["project_changed"] is True
    assert ledger_result["projection_changed"] is True
    period = _projection(db, project.project_id)
    assert period.source == "project_manager_xls_v1"
    assert period.ledger_batch_id == ledger_batch
    assert period.source_batch_id is None
    assert period.version == 1
    assert locked.period_from == date(2026, 1, 1)
    assert locked.period_to == date(2026, 12, 31)
    assert locked.lifecycle_status == "ongoing"

    # 跨来源覆盖：manager workbook 必须清掉 ledger_batch_id（CHECK 互斥）。
    workbook_result = maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 2, 1),
        period_to=date(2026, 11, 30),
        source=maintenance_periods.SOURCE_MANAGER_WORKBOOK,
        source_batch_id=manager_batch,
        as_of=AS_OF,
        operated_by=manager.username,
        reason="月度工作簿期限修订",
    )
    assert workbook_result["before"]["projection"]["ledger_batch_id"] == ledger_batch
    period = _projection(db, project.project_id)
    assert period.source == "manager_workbook_v3"
    assert period.source_batch_id == manager_batch
    assert period.ledger_batch_id is None
    assert period.version == 2

    # direct_api 两者清空。
    maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 3, 1),
        period_to=date(2026, 10, 31),
        source=maintenance_periods.SOURCE_DIRECT_API,
        as_of=AS_OF,
        operated_by="synthetic-admin",
        reason="面板直接编辑期限",
    )
    period = _projection(db, project.project_id)
    assert period.source == "direct_api"
    assert period.source_batch_id is None
    assert period.ledger_batch_id is None
    assert period.version == 3


def test_helper_noop_and_provenance_only_do_not_bump_projection_version(db):
    manager = _user(db, username="period_noop_manager")
    project, _contract = _project(db, suffix="noop")
    manager_batch = _manager_batch(db, manager, batch_id="manager-batch-noop")
    locked = _lock_state_then_project(db, project.project_id)

    maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
        source=maintenance_periods.SOURCE_DIRECT_API,
        as_of=AS_OF,
        operated_by="synthetic-admin",
        reason="初始期限",
    )
    assert _projection(db, project.project_id).version == 1

    noop = maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
        source=maintenance_periods.SOURCE_DIRECT_API,
        as_of=AS_OF,
        operated_by="synthetic-admin",
        reason="完全相同的重复写入",
    )
    assert noop["project_changed"] is False
    assert noop["projection_changed"] is False
    assert _projection(db, project.project_id).version == 1

    # 日期不变、只换来源批次：provenance-only，projection.version 不增。
    provenance_only = maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
        source=maintenance_periods.SOURCE_MANAGER_WORKBOOK,
        source_batch_id=manager_batch,
        as_of=AS_OF,
        operated_by=manager.username,
        reason="同日期跨来源回写",
    )
    assert provenance_only["projection_changed"] is True
    period = _projection(db, project.project_id)
    assert period.version == 1
    assert period.source == "manager_workbook_v3"
    assert period.source_batch_id == manager_batch


# ---------- catalog PATCH ----------


def test_catalog_patch_syncs_projection_lifecycle_version_and_revision(db):
    _user(db, username="period_catalog_admin")
    created = catalog.create_project(
        db,
        project_code="PERIOD-CATALOG",
        display_name="合成 catalog 期限项目",
        project_manager_id=None,
        reason="建立期限 catalog 测试主档",
        operated_by="period_catalog_admin",
    )
    db.commit()
    project_id = created["project_id"]

    # 宽日期区间保证任意业务日期下 lifecycle 都是 ongoing。
    after = catalog.update_project(
        db,
        project_id=project_id,
        version=1,
        updates={
            "display_name": "合成 catalog 期限项目（改名）",
            "period_from": date(2020, 1, 1),
            "period_to": date(2099, 12, 31),
        },
        reason="设置维保期限并改名",
        operated_by="period_catalog_admin",
    )
    db.commit()

    assert after["period_from"] == "2020-01-01"
    assert after["period_to"] == "2099-12-31"
    assert after["lifecycle_status"] == "ongoing"
    # 多个字段合并后版本只 +1
    assert after["version"] == 2

    period = _projection(db, project_id)
    assert period is not None
    assert period.service_start == date(2020, 1, 1)
    assert period.service_end == date(2099, 12, 31)
    assert period.completeness_state == "complete"
    assert period.source == "direct_api"
    assert period.source_batch_id is None
    assert period.ledger_batch_id is None
    assert period.version == 1

    audits = list(
        db.scalars(
            select(MaintenanceProjectAuditLog).where(
                MaintenanceProjectAuditLog.project_id == project_id
            )
        )
    )
    assert [row.action for row in audits] == ["create", "update"]
    state = operations.get_or_create_workbook_state(db, project_id=project_id)
    assert state.revision == 1


def test_catalog_patch_rejects_inverted_period(db):
    _user(db, username="period_catalog_inverted_admin")
    created = catalog.create_project(
        db,
        project_code="PERIOD-CATALOG-INVERTED",
        display_name="合成倒置期限项目",
        project_manager_id=None,
        reason="建立倒置期限测试主档",
        operated_by="period_catalog_inverted_admin",
    )
    db.commit()

    with pytest.raises(catalog.MaintenanceProjectCatalogError):
        catalog.update_project(
            db,
            project_id=created["project_id"],
            version=1,
            updates={
                "period_from": date(2026, 12, 31),
                "period_to": date(2026, 1, 1),
            },
            reason="倒置期限应被拒绝",
            operated_by="period_catalog_inverted_admin",
        )
    db.rollback()
    reloaded = db.get(MaintenanceProject, created["project_id"])
    assert reloaded.period_from is None and reloaded.period_to is None
    assert reloaded.version == 1
    assert _projection(db, created["project_id"]) is None


# ---------- manager workbook apply ----------


def test_manager_workbook_apply_syncs_project_and_projection_with_single_bump(db):
    manager = _user(db, username="period_apply_manager")
    project, contract = _project(db, suffix="apply", manager=manager)
    adapter = _adapter(db, manager)
    artifact, _snapshot = adapter.export(REPORT_MONTH, hmac_key=HMAC_KEY)

    edited = _edit_workbook(
        artifact.content,
        project_id=project.project_id,
        relation_id=contract.project_contract_id,
        service_start=date(2026, 1, 1),
        service_end=date(2026, 12, 31),
        planned_date=date(2026, 9, 20),
        planned_amount=18000,
    )
    validation, batch = adapter.validate(REPORT_MONTH, edited, hmac_key=HMAC_KEY)
    assert validation.can_apply, [issue.message for issue in validation.errors]
    assert len(validation.service_period_changes) == 1
    assert len(validation.milestone_changes) == 1

    result = adapter.apply(batch.batch_id, data_version=validation.data_version)
    db.commit()
    assert result["changed_rows"] == 2

    reloaded = db.get(MaintenanceProject, project.project_id)
    assert reloaded.period_from == date(2026, 1, 1)
    assert reloaded.period_to == date(2026, 12, 31)
    assert reloaded.lifecycle_status == "ongoing"
    # service + milestone 两类变化合并，project.version 只 +1
    assert reloaded.version == 2

    period = _projection(db, project.project_id)
    assert period.service_start == reloaded.period_from
    assert period.service_end == reloaded.period_to
    assert period.completeness_state == "complete"
    assert period.source == "manager_workbook_v3"
    assert period.source_batch_id == batch.batch_id
    assert period.ledger_batch_id is None
    assert period.version == 1

    milestone = db.scalar(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id
            == contract.project_contract_id
        )
    )
    assert milestone is not None
    assert milestone.planned_date == date(2026, 9, 20)

    audits = list(
        db.scalars(
            select(MaintenanceProjectAuditLog).where(
                MaintenanceProjectAuditLog.project_id == project.project_id
            )
        )
    )
    assert len(audits) == 1
    assert audits[0].action == "update"
    assert audits[0].before_json["version"] == 1
    assert audits[0].after_json["version"] == 2

    state = operations.get_or_create_workbook_state(db, project_id=project.project_id)
    assert state.revision == 1


def test_manager_workbook_apply_conflict_on_external_period_change_zero_writes(db):
    manager = _user(db, username="period_conflict_manager")
    project, contract = _project(db, suffix="conflict", manager=manager)
    adapter = _adapter(db, manager)
    artifact, _snapshot = adapter.export(REPORT_MONTH, hmac_key=HMAC_KEY)
    edited = _edit_workbook(
        artifact.content,
        project_id=project.project_id,
        relation_id=contract.project_contract_id,
        service_start=date(2026, 1, 1),
        service_end=date(2026, 12, 31),
        planned_date=date(2026, 9, 20),
        planned_amount=18000,
    )
    validation, batch = adapter.validate(REPORT_MONTH, edited, hmac_key=HMAC_KEY)
    assert validation.can_apply

    # 校验后外部直改期限（旧 service_period_version=0 瞬间过期）。
    catalog.update_project(
        db,
        project_id=project.project_id,
        version=1,
        updates={"period_from": date(2026, 3, 1), "period_to": date(2026, 3, 31)},
        reason="校验后外部修改期限制造版本冲突",
        operated_by="period_conflict_manager",
    )
    db.commit()

    with pytest.raises(ManagerWorkbookConflict):
        adapter.apply(batch.batch_id, data_version=validation.data_version)
    db.rollback()

    # 零写：milestone 未创建、batch 仍可应用、外部期限改动原样保留。
    assert (
        db.scalar(
            select(MaintenanceCollectionMilestone).where(
                MaintenanceCollectionMilestone.project_contract_id
                == contract.project_contract_id
            )
        )
        is None
    )
    assert db.get(MaintenanceManagerUploadBatch, batch.batch_id).status == "valid"
    reloaded = db.get(MaintenanceProject, project.project_id)
    assert reloaded.period_from == date(2026, 3, 1)
    assert reloaded.version == 2


def test_manager_workbook_apply_scope_drift_zero_writes(db):
    manager = _user(db, username="period_drift_manager")
    project, contract = _project(db, suffix="drift", manager=manager)
    adapter = _adapter(db, manager)
    artifact, _snapshot = adapter.export(REPORT_MONTH, hmac_key=HMAC_KEY)
    edited = _edit_workbook(
        artifact.content,
        project_id=project.project_id,
        relation_id=contract.project_contract_id,
        planned_date=date(2026, 9, 20),
        planned_amount=18000,
    )
    validation, batch = adapter.validate(REPORT_MONTH, edited, hmac_key=HMAC_KEY)
    assert validation.can_apply

    # 校验后本人负责范围漂移：新分配一个项目。
    _project(db, suffix="drift-new", manager=manager)

    with pytest.raises(ManagerWorkbookConflict):
        adapter.apply(batch.batch_id, data_version=validation.data_version)
    db.rollback()

    assert (
        db.scalar(
            select(MaintenanceCollectionMilestone).where(
                MaintenanceCollectionMilestone.project_contract_id
                == contract.project_contract_id
            )
        )
        is None
    )
    assert db.get(MaintenanceManagerUploadBatch, batch.batch_id).status == "valid"
    reloaded = db.get(MaintenanceProject, project.project_id)
    assert reloaded.version == 1


# ---------- 读侧：operations / reminders 只看 canonical project 期限 ----------


def test_operations_and_reminders_read_canonical_project_period(db):
    manager = _user(db, username="period_read_manager")
    # project 期限只有起点（start_only）；projection 刻意冲突写成 complete，
    # 模拟历史双源漂移——业务展示/筛选必须只认 project。
    project, contract = _project(
        db,
        suffix="read",
        manager=manager,
        period_from=date(2026, 1, 1),
        period_to=None,
        lifecycle_status="ongoing",
    )
    db.add(
        MaintenanceServicePeriod(
            project_id=project.project_id,
            service_start=date(2020, 1, 1),
            service_end=date(2020, 12, 31),
            completeness_state="complete",
            source="direct_api",
        )
    )
    db.commit()

    start_only = operations.project_operations(
        db,
        as_of=AS_OF,
        user_ctx=_ctx(manager.username),
        owner_scope="all",
        reminder="service_period:start_only",
    )
    assert start_only["total"] == 1
    card = start_only["rows"][0]
    assert card["project_id"] == project.project_id
    assert card["manager_tracking"]["service_period"] == {
        "service_start": "2026-01-01",
        "service_end": None,
        "completeness_state": "start_only",
    }

    empty = operations.project_operations(
        db,
        as_of=AS_OF,
        user_ctx=_ctx(manager.username),
        owner_scope="all",
        reminder="service_period:empty",
    )
    assert empty["total"] == 0
    complete = operations.project_operations(
        db,
        as_of=AS_OF,
        user_ctx=_ctx(manager.username),
        owner_scope="all",
        reminder="service_period:end_only",
    )
    assert complete["total"] == 0

    detail = get_project_collection_milestones(
        db,
        project_id=project.project_id,
        as_of=AS_OF,
        user_ctx=_ctx(manager.username),
    )
    assert detail["project"]["service_period"] == {
        "service_start": "2026-01-01",
        "service_end": None,
        "completeness_state": "start_only",
    }


def test_cross_day_reads_end_period_without_snapshot_write(db, monkeypatch):
    """PostgreSQL 回归：跨过 period_to 次日，无日切写入也必须动态变 ended。

    同时覆盖 SQL 筛选（老板看板、operations 目录）与对应 payload，以及工作区、
    stable project read model、回款提醒目录/详情。存储列保持旧 ``ongoing``，证明
    结果不是某条读取路径偷偷落库刷新快照得到的。
    """
    assert db.get_bind().dialect.name == "postgresql"
    manager = _user(db, username="period_cross_day_manager")
    period_to = date(2026, 8, 15)
    next_business_day = period_to + timedelta(days=1)
    project, _contract = _project(
        db,
        suffix="cross-day",
        manager=manager,
        period_from=date(2026, 1, 1),
        period_to=period_to,
        # 模拟昨天写入的兼容快照；今天没有任何写操作。
        lifecycle_status="ongoing",
    )
    db.commit()
    ctx = _ctx(manager.username)

    # boss_board 没有显式 as_of 参数，冻结其请求级 business_today。
    monkeypatch.setattr(
        maintenance_boss_board,
        "business_today",
        lambda: next_business_day,
    )
    boss_ended = maintenance_boss_board.projects(
        db,
        user_ctx=ctx,
        q_text=project.project_code,
        lifecycle="ended",
    )
    assert boss_ended["total"] == 1
    assert boss_ended["rows"][0]["lifecycle"] == "ended"
    boss_ongoing = maintenance_boss_board.projects(
        db,
        user_ctx=ctx,
        q_text=project.project_code,
        lifecycle="ongoing",
    )
    assert boss_ongoing["total"] == 0

    operations_ended = operations.project_operations(
        db,
        as_of=next_business_day,
        user_ctx=ctx,
        owner_scope="all",
        q_text=project.project_code,
        lifecycle="ended",
    )
    assert operations_ended["total"] == 1
    assert operations_ended["rows"][0]["lifecycle_status"] == "ended"
    operations_ongoing = operations.project_operations(
        db,
        as_of=next_business_day,
        user_ctx=ctx,
        owner_scope="all",
        q_text=project.project_code,
        lifecycle="ongoing",
    )
    assert operations_ongoing["total"] == 0

    workspace = operations.project_workspace(
        db,
        project_id=project.project_id,
        as_of=next_business_day,
        user_ctx=ctx,
    )
    assert workspace is not None
    assert workspace["project"]["lifecycle_status"] == "ended"

    directory = maintenance_project.project_directory(
        db,
        q_text=project.project_code,
        page=1,
        page_size=20,
        include_inactive=False,
        as_of=next_business_day,
        user_ctx=ctx,
    )
    assert directory["rows"][0]["lifecycle_status"] == "ended"

    reminder_directory = search_collection_reminders(
        db,
        as_of=next_business_day,
        user_ctx=ctx,
        q_text=project.project_code,
        owner_scope="all",
        reminder_state=None,
        page=1,
        page_size=20,
    )
    assert reminder_directory["rows"][0]["project"]["lifecycle_status"] == "ended"
    reminder_detail = get_project_collection_milestones(
        db,
        project_id=project.project_id,
        as_of=next_business_day,
        user_ctx=ctx,
    )
    assert reminder_detail is not None
    assert reminder_detail["project"]["lifecycle_status"] == "ended"

    db.expire_all()
    stored = db.get(MaintenanceProject, project.project_id)
    assert stored is not None
    assert stored.lifecycle_status == "ongoing"
    assert stored.version == 1


# ---------- 只读漂移分类（不自动修复历史冲突） ----------


def test_classify_period_divergence_is_readonly(db):
    diverged, _ = _project(
        db,
        suffix="div",
        period_from=date(2026, 1, 1),
        period_to=None,
        lifecycle_status="ongoing",
    )
    db.add(
        MaintenanceServicePeriod(
            project_id=diverged.project_id,
            service_start=date(2020, 1, 1),
            service_end=date(2020, 12, 31),
            completeness_state="complete",
            source="direct_api",
        )
    )
    missing, _ = _project(
        db,
        suffix="miss",
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
        lifecycle_status="ongoing",
    )
    synced, _ = _project(db, suffix="sync")
    locked = _lock_state_then_project(db, synced.project_id)
    maintenance_periods.apply_canonical_period_locked(
        db,
        project=locked,
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
        source=maintenance_periods.SOURCE_DIRECT_API,
        as_of=AS_OF,
        operated_by="synthetic-admin",
        reason="同步项目不应出现在漂移清单",
    )
    db.commit()

    rows = maintenance_periods.classify_period_divergence(db)
    by_id = {row["project_id"]: row for row in rows}
    assert by_id[diverged.project_id]["category"] == "diverged"
    assert by_id[diverged.project_id]["project_period_from"] == "2026-01-01"
    assert by_id[diverged.project_id]["projection_service_start"] == "2020-01-01"
    assert by_id[missing.project_id]["category"] == "missing_projection"
    assert synced.project_id not in by_id

    # 只读：分类后两侧数据与版本完全不变。
    assert db.get(MaintenanceProject, diverged.project_id).version == 1
    period = db.get(MaintenanceServicePeriod, diverged.project_id)
    assert period.version == 1
    assert period.service_start == date(2020, 1, 1)
