"""AB-3：报销/回款往返工作簿（F3 改判并入 v1）。

口径出处：`docs/releases/v1.23-addon-pack.md` AB-3 + REQUIREMENTS #8/#30/#31。
"""
import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app import permissions
from app.auth import hash_password
from app.main import app
from app.models.maintenance import FProjectExpense
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
    MaintenanceProjectUserAssignment,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
)
from app.models.system import SysImportBatch, SysUser
from app.services import maintenance_expense_collection_workbook as wbk
from app.services import maintenance_project_operations as operations

_PASSWORD = "synthetic-wbk-password-1"
_BASE = "/api/maintenance/projects/stable"
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _client(
    db,
    *,
    username,
    overrides,
    assign_existing_projects: bool = True,
) -> TestClient:
    base = permissions.effective("readonly", None)
    user = SysUser(
        username=username, role="readonly", display_name=username,
        password_hash=hash_password(_PASSWORD), is_active=True,
        template_code="readonly", template_version=1, template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides))
    db.add(user)
    db.flush()
    if assign_existing_projects:
        for project_id in db.scalars(select(MaintenanceProject.project_id)):
            db.add(MaintenanceProjectUserAssignment(
                assignment_id=str(uuid.uuid4()),
                project_id=project_id,
                responsibility_type="viewer",
                user_id=user.id,
                version=1,
                assigned_by="synthetic-workbook-test",
                assignment_reason="往返工作簿测试项目可见范围",
            ))
    db.commit()
    client = TestClient(app)
    login = client.post("/api/auth/login",
                        json={"username": username, "password": _PASSWORD})
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def uploader(
    db,
    username="wbk-uploader",
    *,
    assign_existing_projects: bool = True,
) -> TestClient:
    return _client(db, username=username, overrides={
        "page_maintenance": True, "data_purchase_cost": True, "data_profit": True,
        "action_maintenance_expense_collection_upload": True},
        assign_existing_projects=assign_existing_projects)


def reader(
    db,
    username="wbk-reader",
    *,
    assign_existing_projects: bool = True,
) -> TestClient:
    """能看金额但不能传（只读对账）。"""
    return _client(db, username=username, overrides={
        "page_maintenance": True, "data_purchase_cost": True, "data_profit": True},
        assign_existing_projects=assign_existing_projects)


@pytest.fixture()
def project(db):
    proj = MaintenanceProject(project_id=str(uuid.uuid4()), project_code="合成项目A",
                              display_name="合成项目A", lifecycle_status="ongoing")
    db.add(proj)
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=proj.project_id,
        contract_id="C-1", contract_no="HT-001",
        amount_inc_tax=Decimal("100000.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1))
    db.commit()
    return proj


def _expense(db, *, raw_line_id="BXD-1#1", contract_no="HT-001",
             ex_tax="100.00"):
    batch = SysImportBatch(filename="e.xlsx", file_type="maintenance",
                           file_hash=uuid.uuid4().hex * 2, status="success")
    db.add(batch)
    db.flush()
    expense = FProjectExpense(
        raw_line_id=raw_line_id, bxd_no="BXD-20260101-1", line_no=1,
        data_status="已结束", expense_date=date(2026, 7, 1), person="张三",
        expense_type="差旅", fee_category="交通", reason="现场维保",
        linked_sales_order_no=contract_no,
        amount=Decimal(ex_tax), amount_ex_tax=Decimal(ex_tax),
        amount_inc_tax=(Decimal(ex_tax) * Decimal("1.13")).quantize(Decimal("0.01")),
        tax_basis="ex", tax_rate_used=Decimal("0.13"), import_batch_id=batch.id)
    db.add(expense)
    contract = db.scalars(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.contract_no == contract_no
    )).one()
    db.add(MaintenanceProjectExpenseAttribution(
        expense_id=f"bxd:{raw_line_id}",
        project_id=contract.project_id,
        project_contract_id=contract.project_contract_id,
        raw_expense_line_id=expense.raw_line_id,
        expense_ref=f"{expense.bxd_no}#{expense.line_no}",
        expense_date=expense.expense_date,
        applicant=expense.person,
        category=expense.fee_category,
        expense_reason=expense.reason,
        amount_ex_tax=expense.amount_ex_tax,
        amount_inc_tax=expense.amount_inc_tax,
        tax_rate_used=Decimal("0.13"),
        raw_status=expense.data_status,
        status_mapping_state="mapped",
        normalized_status="approved",
        status_mapping_version="synthetic-expense-workbook",
        ownership_mapping_state="mapped",
        ownership_mapping_version="synthetic-expense-workbook",
        version=1,
    ))
    db.commit()


def _download(client, project) -> bytes:
    resp = client.get(f"{_BASE}/{project.project_id}/expense-collection-workbook.xlsx")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(_XLSX)
    return resp.content


def _upload(client, project, content: bytes, *, action="apply"):
    return client.post(
        f"{_BASE}/{project.project_id}/expense-collection-workbook/{action}",
        files={"file": ("wbk.xlsx", io.BytesIO(content), _XLSX)})


# ---------- 导出 ----------

def test_workbook_has_exactly_the_two_signed_sheets(db, project):
    _expense(db)
    wb = load_workbook(io.BytesIO(_download(uploader(db), project)))
    # AB-3「合二为一」：一个文件两张 sheet；不得顺带导出冻结的 v3 其余四表
    assert wb.sheetnames[:2] == [wbk.SHEET_EXPENSE, wbk.SHEET_COLLECTION]
    visible = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"]
    assert visible == [wbk.SHEET_EXPENSE, wbk.SHEET_COLLECTION]
    assert [c.value for c in wb[wbk.SHEET_COLLECTION][1]][:4] == [
        "操作", "合同编号", "报告月份", "累计回款金额(含税)"]


def test_download_requires_profit_visibility_but_not_upload_action(db, project):
    """只读对账的人应当能把表拉下来核对，不必为此获得写权限。"""
    assert reader(db).get(
        f"{_BASE}/{project.project_id}/expense-collection-workbook.xlsx"
    ).status_code == 200
    # data_profit 在 readonly 模板里默认是 True，必须显式关掉才测得到这条门
    no_money = _client(db, username="wbk-nomoney",
                       overrides={"page_maintenance": True, "data_profit": False})
    assert no_money.get(
        f"{_BASE}/{project.project_id}/expense-collection-workbook.xlsx"
    ).status_code == 403


def test_unknown_project_fails_closed_for_scoped_account(db):
    assert uploader(db, "wbk-404").get(
        f"{_BASE}/{uuid.uuid4()}/expense-collection-workbook.xlsx"
    ).status_code == 403


@pytest.mark.parametrize("action", ["download", "validate", "apply"])
def test_unassigned_account_cannot_access_any_workbook_route(db, project, action):
    """下载、预演、应用必须共用同一项目行权限，不能只封写端。"""
    authorized = uploader(db, f"wbk-scope-source-{action}")
    content = _download(authorized, project)
    denied = uploader(
        db,
        f"wbk-scope-denied-{action}",
        assign_existing_projects=False,
    )
    if action == "download":
        response = denied.get(
            f"{_BASE}/{project.project_id}/expense-collection-workbook.xlsx"
        )
    else:
        response = _upload(denied, project, content, action=action)
    assert response.status_code == 403


# ---------- 报销：未税可改，含税由系统算 ----------

def test_expense_inc_tax_is_computed_not_accepted_from_the_sheet(db, project):
    """REQUIREMENTS #8：正式金额列=含税，由未税×1.13 算出，不接受人工直填。"""
    _expense(db, ex_tax="100.00")
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_EXPENSE]
    ws.cell(row=2, column=8, value=200)        # 未税改成 200
    ws.cell(row=2, column=9, value=999999)     # 含税列乱填，应被忽略
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue())
    assert resp.status_code == 200, resp.text
    assert resp.json()["expense_updates"] == 1
    db.expire_all()
    expense = db.execute(select(FProjectExpense)).scalars().one()
    assert expense.amount_ex_tax == Decimal("200.00")
    assert expense.amount_inc_tax == Decimal("226.00")   # 200 × 1.13，不是 999999
    assert expense.tax_basis == "ex"


def test_negative_reversal_preserves_inc_tax_source_basis(db, project):
    """冲销允许负数；编辑展示未税值时仍保留来源的含税金额口径。"""
    _expense(db, ex_tax="100.00")
    expense = db.execute(select(FProjectExpense)).scalars().one()
    attribution = db.get(
        MaintenanceProjectExpenseAttribution, f"bxd:{expense.raw_line_id}"
    )
    expense.tax_basis = "inc"
    expense.amount = Decimal("113.00")
    attribution.tax_basis = "inc"
    db.commit()

    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_EXPENSE]
    ws.cell(row=2, column=8, value=-200)
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue())

    assert resp.status_code == 200, resp.text
    db.expire_all()
    expense = db.execute(select(FProjectExpense)).scalars().one()
    attribution = db.get(
        MaintenanceProjectExpenseAttribution, f"bxd:{expense.raw_line_id}"
    )
    assert expense.tax_basis == "inc"
    assert expense.amount == Decimal("-226.00")
    assert expense.amount_ex_tax == Decimal("-200.00")
    assert expense.amount_inc_tax == Decimal("-226.00")
    assert attribution.tax_basis == "inc"
    assert attribution.amount_ex_tax == Decimal("-200.00")
    assert attribution.amount_inc_tax == Decimal("-226.00")


def test_manual_expense_row_without_entity_id_is_created(db, project):
    """人工回填兼容：空实体 ID 的完整报销行由后端生成稳定键并新增。"""
    _expense(db)
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_EXPENSE]
    ws.append(["BXD-NEW", "2026-07-02", "李四", "差旅", "交通", "编的",
               "HT-001", 50, "", "已结束"])
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue())
    assert resp.status_code == 200, resp.text
    assert resp.json()["expense_creates"] == 1
    rows = db.scalars(select(FProjectExpense).order_by(FProjectExpense.id)).all()
    assert len(rows) == 2
    assert rows[1].bxd_no == "BXD-NEW"
    assert rows[1].amount_ex_tax == Decimal("50.00")
    assert rows[1].raw_line_id.startswith("EXP:")
    attribution = db.get(
        MaintenanceProjectExpenseAttribution,
        f"bxd:{rows[1].raw_line_id}",
    )
    assert attribution is not None
    assert attribution.project_id == project.project_id


def test_shared_current_contract_hides_expense_from_workbook_and_rows_api(db, project):
    """同一 XSDD 被两个项目当前占用时，合同号不能再充当项目授权边界。"""
    _expense(db)
    other = MaintenanceProject(
        project_id=str(uuid.uuid4()), project_code="合成项目B",
        display_name="合成项目B", lifecycle_status="ongoing",
    )
    db.add(other)
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=other.project_id,
        contract_id="C-SHARED-B", contract_no="HT-001",
        amount_inc_tax=Decimal("1.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1,
    ))
    db.commit()

    client = uploader(db, "shared-contract-reader")
    workbook = load_workbook(io.BytesIO(_download(client, project)))
    assert workbook[wbk.SHEET_EXPENSE].max_row == 1
    response = client.get(
        f"{_BASE}/{project.project_id}/expense-rows",
    )
    assert response.status_code == 200, response.text
    assert response.json()["rows"] == []


def test_crafted_foreign_raw_line_id_is_rejected_at_validation(db, project):
    other = MaintenanceProject(
        project_id=str(uuid.uuid4()), project_code="合成项目B",
        display_name="合成项目B", lifecycle_status="ongoing",
    )
    db.add(other)
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=other.project_id,
        contract_id="C-B", contract_no="HT-002",
        amount_inc_tax=Decimal("1.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1,
    ))
    db.commit()
    _expense(db, raw_line_id="FOREIGN-BXD#1", contract_no="HT-002")

    client = uploader(db, "foreign-expense-crafter")
    workbook = load_workbook(io.BytesIO(_download(client, project)))
    row = ["BXD-FOREIGN", "2026-07-01", "外部", "差旅", "交通", "越权",
           "HT-001", 999, "", "已结束", ""]
    workbook[wbk.SHEET_EXPENSE].append(row + ["FOREIGN-BXD#1"])
    payload = io.BytesIO()
    workbook.save(payload)

    response = _upload(client, project, payload.getvalue(), action="validate")
    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "expense_not_in_project"


def test_apply_rechecks_expense_attribution_after_validation(db, project):
    _expense(db, ex_tax="100.00")
    workbook = load_workbook(io.BytesIO(wbk.build_workbook(
        db, project_id=project.project_id)))
    workbook[wbk.SHEET_EXPENSE].cell(row=2, column=8, value=250)
    payload = io.BytesIO()
    workbook.save(payload)
    plan = wbk.validate(db, project_id=project.project_id, data=payload.getvalue())

    other = MaintenanceProject(
        project_id=str(uuid.uuid4()), project_code="归因改派项目",
        display_name="归因改派项目", lifecycle_status="ongoing",
    )
    db.add(other)
    db.flush()
    other_contract = MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=other.project_id,
        contract_id=f"C-{uuid.uuid4().hex[:8]}",
        contract_no="HT-OTHER",
        amount_inc_tax=Decimal("100000.00"),
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="synthetic-scope-recheck",
        effective_from=date(2026, 1, 1),
        source="synthetic-test",
        version=1,
    )
    db.add(other_contract)
    attribution = db.get(MaintenanceProjectExpenseAttribution, "bxd:BXD-1#1")
    attribution.project_id = other.project_id
    attribution.project_contract_id = other_contract.project_contract_id
    raw_expense = db.scalar(select(FProjectExpense).where(
        FProjectExpense.raw_line_id == "BXD-1#1",
    ))
    raw_expense.linked_sales_order_no = other_contract.contract_no
    db.commit()

    with pytest.raises(wbk.WorkbookError) as raised:
        wbk.apply(
            db,
            plan,
            operated_by="scope-recheck",
            import_batch_id=str(uuid.uuid4()),
        )
    assert raised.value.code == "expense_not_in_project"
    db.rollback()
    db.expire_all()
    expense = db.scalar(select(FProjectExpense).where(
        FProjectExpense.raw_line_id == "BXD-1#1"))
    assert expense.amount_ex_tax == Decimal("100.00")


def test_blank_amount_leaves_the_row_untouched(db, project):
    _expense(db, ex_tax="100.00")
    client = uploader(db)
    content = _download(client, project)
    assert _upload(client, project, content).json()["expense_updates"] == 0
    db.expire_all()
    assert db.execute(select(FProjectExpense)).scalars().one().amount_ex_tax \
        == Decimal("100.00")


def test_standalone_apply_bumps_project_workbook_revision_once(db, project):
    """04/05 standalone writes invalidate older project-workbook snapshots."""

    _expense(db, ex_tax="100.00")
    state = operations.get_or_create_workbook_state(
        db, project_id=project.project_id)
    db.commit()
    before = state.revision

    client = uploader(db, "revision-standalone")
    wb = load_workbook(io.BytesIO(_download(client, project)))
    wb[wbk.SHEET_EXPENSE].cell(row=2, column=8, value=125)
    buf = io.BytesIO()
    wb.save(buf)
    response = _upload(client, project, buf.getvalue())
    assert response.status_code == 200, response.text

    db.expire_all()
    state = operations.get_or_create_workbook_state(
        db, project_id=project.project_id)
    assert state.revision == before + 1


# ---------- 回款：月度累计快照 ----------

def test_collection_create_appends_monthly_cumulative_snapshot(db, project):
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_COLLECTION]
    ws.cell(row=2, column=1, value="CREATE")
    ws.cell(row=2, column=2, value="HT-001")
    ws.cell(row=2, column=3, value="2026-07")
    ws.cell(row=2, column=4, value=30000)
    ws.cell(row=2, column=5, value="SK-001")
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue())
    assert resp.status_code == 200, resp.text
    assert resp.json()["collection_creates"] == 1
    snapshot = db.execute(select(MaintenanceCollectionSnapshot)).scalars().one()
    assert snapshot.report_month == date(2026, 7, 1)
    assert snapshot.cumulative_amount == Decimal("30000.00")
    assert snapshot.status == "confirmed" and snapshot.source == "workbook"
    assert snapshot.import_batch_id is not None


def test_reupload_overwrites_the_same_month(db, project):
    """「上传覆盖」：同合同同月份再传一次是改写累计额，不是追加第二条。"""
    client = uploader(db)

    def _send(amount):
        wb = load_workbook(io.BytesIO(_download(client, project)))
        ws = wb[wbk.SHEET_COLLECTION]
        row = ws.max_row + 1 if ws.cell(row=2, column=2).value else 2
        ws.cell(row=row, column=1, value="CREATE")
        ws.cell(row=row, column=2, value="HT-001")
        ws.cell(row=row, column=3, value="2026-07")
        ws.cell(row=row, column=4, value=amount)
        buf = io.BytesIO()
        wb.save(buf)
        return _upload(client, project, buf.getvalue())

    assert _send(30000).status_code == 200
    assert _send(45000).status_code == 200
    snapshots = db.execute(select(MaintenanceCollectionSnapshot)).scalars().all()
    assert len(snapshots) == 1
    db.expire_all()
    assert snapshots[0].cumulative_amount == Decimal("45000.00")


def test_void_marks_snapshot_void_rather_than_deleting(db, project):
    """缺行 ≠ 删除：作废必须显式写 VOID，且留痕不物理删。"""
    contract = db.execute(select(MaintenanceProjectContract)).scalars().one()
    db.add(MaintenanceCollectionSnapshot(
        collection_id=str(uuid.uuid4()), project_id=project.project_id,
        project_contract_id=contract.project_contract_id,
        report_month=date(2026, 6, 1), cumulative_amount=Decimal("10000.00"),
        status="confirmed", source="legacy", version=1))
    db.commit()
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_COLLECTION]
    ws.cell(row=2, column=1, value="VOID")
    buf = io.BytesIO()
    wb.save(buf)
    assert _upload(client, project, buf.getvalue()).status_code == 200
    db.expire_all()
    snapshot = db.execute(select(MaintenanceCollectionSnapshot)).scalars().one()
    assert snapshot.status == "void"          # 仍在库，只是作废


def test_contract_outside_the_project_is_rejected(db, project):
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_COLLECTION]
    ws.cell(row=2, column=1, value="CREATE")
    ws.cell(row=2, column=2, value="HT-别人家的")
    ws.cell(row=2, column=3, value="2026-07")
    ws.cell(row=2, column=4, value=100)
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue())
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "contract_not_found"


def test_one_bad_row_rejects_the_whole_file(db, project):
    """整份拒绝：静默跳过一行 = 操作者以为改了、实际没改，比报错难发现。"""
    _expense(db, ex_tax="100.00")
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    wb[wbk.SHEET_EXPENSE].cell(row=2, column=8, value=250)      # 合法改动
    ws = wb[wbk.SHEET_COLLECTION]
    ws.cell(row=2, column=1, value="CREATE")
    ws.cell(row=2, column=2, value="HT-001")
    ws.cell(row=2, column=3, value="不是月份")                   # 非法
    ws.cell(row=2, column=4, value=100)
    buf = io.BytesIO()
    wb.save(buf)
    assert _upload(client, project, buf.getvalue()).status_code == 422
    db.expire_all()
    # 合法的那一行也必须没有落库
    assert db.execute(select(FProjectExpense)).scalars().one().amount_ex_tax \
        == Decimal("100.00")


def test_validate_is_side_effect_free(db, project):
    _expense(db, ex_tax="100.00")
    client = uploader(db)
    wb = load_workbook(io.BytesIO(_download(client, project)))
    wb[wbk.SHEET_EXPENSE].cell(row=2, column=8, value=777)
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue(), action="validate")
    assert resp.status_code == 200 and resp.json()["expense_updates"] == 1
    db.expire_all()
    assert db.execute(select(FProjectExpense)).scalars().one().amount_ex_tax \
        == Decimal("100.00")


# ---------- 权限 ----------

def test_upload_requires_the_dedicated_action_key(db, project):
    """WBDD-only 账号传不了（AB-3 权限要求）。"""
    content = _download(uploader(db, "wbk-up2"), project)
    assert _upload(reader(db, "wbk-read2"), project, content).status_code == 403
    wbdd_only = _client(db, username="wbk-wbddonly", overrides={
        "page_maintenance": True, "action_maintenance_wbdd_import": True})
    assert _upload(wbdd_only, project, content).status_code == 403


def test_unauthenticated_is_401(db, project):
    anon = TestClient(app)
    assert anon.get(
        f"{_BASE}/{project.project_id}/expense-collection-workbook.xlsx"
    ).status_code == 401


def test_non_xlsx_upload_is_415(db, project):
    client = uploader(db, "wbk-415")
    resp = client.post(
        f"{_BASE}/{project.project_id}/expense-collection-workbook/apply",
        files={"file": ("x.csv", io.BytesIO(b"a,b"), "text/csv")})
    assert resp.status_code == 415


# ---------- #47：报销备注列 ----------

def test_expense_sheet_has_editable_remark_column(db, project):
    _expense(db)
    wb = load_workbook(io.BytesIO(_download(uploader(db, "rmk-1"), project)))
    # 表头行末尾还有一列隐藏的行标识（值为空），所以按名字取位置而不是取 [-1]
    headers = [c.value for c in wb[wbk.SHEET_EXPENSE][1]]
    assert "备注" in headers
    remark_col = headers.index("备注") + 1
    assert remark_col == len(wbk._EXPENSE_HEADERS)
    # 黄底＝可编辑（与未税金额同色），与只读列区分
    ws = wb[wbk.SHEET_EXPENSE]
    assert ws.cell(row=1, column=remark_col).fill.fgColor.rgb \
        == ws.cell(row=1, column=8).fill.fgColor.rgb
    # 2026-08-17 全面放开：除「含税金额(系统计算)」外所有列黄底可改（支持手工新增报销）。
    # 只读仅剩含税金额列（第 9 列），备注黄底应与它不同色。
    assert ws.cell(row=1, column=remark_col).fill.fgColor.rgb \
        != ws.cell(row=1, column=9).fill.fgColor.rgb


def test_remark_roundtrips_into_the_database(db, project):
    _expense(db)
    client = uploader(db, "rmk-2")
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_EXPENSE]
    ws.cell(row=2, column=len(wbk._EXPENSE_HEADERS), value="客户确认可报")
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload(client, project, buf.getvalue())
    assert resp.status_code == 200, resp.text
    assert resp.json()["expense_updates"] == 1
    db.expire_all()
    assert db.execute(select(FProjectExpense)).scalars().one().remark == "客户确认可报"


def test_existing_remark_is_exported_back(db, project):
    _expense(db)
    expense = db.execute(select(FProjectExpense)).scalars().one()
    expense.remark = "上一轮写的备注"
    db.commit()
    wb = load_workbook(io.BytesIO(_download(uploader(db, "rmk-3"), project)))
    ws = wb[wbk.SHEET_EXPENSE]
    assert ws.cell(row=2, column=len(wbk._EXPENSE_HEADERS)).value == "上一轮写的备注"


def test_remark_only_edit_does_not_touch_amounts(db, project):
    """只改备注不填金额也是合法回填：金额三列一个都不能动。"""
    _expense(db, ex_tax="100.00")
    client = uploader(db, "rmk-4")
    wb = load_workbook(io.BytesIO(_download(client, project)))
    ws = wb[wbk.SHEET_EXPENSE]
    ws.cell(row=2, column=8).value = None                      # 未税留空
    ws.cell(row=2, column=len(wbk._EXPENSE_HEADERS), value="只改备注")
    buf = io.BytesIO()
    wb.save(buf)
    assert _upload(client, project, buf.getvalue()).status_code == 200
    db.expire_all()
    expense = db.execute(select(FProjectExpense)).scalars().one()
    assert expense.remark == "只改备注"
    assert expense.amount_ex_tax == Decimal("100.00")
    assert expense.amount_inc_tax == Decimal("113.00")


def test_clearing_remark_writes_null(db, project):
    _expense(db)
    expense = db.execute(select(FProjectExpense)).scalars().one()
    expense.remark = "待清空"
    db.commit()
    client = uploader(db, "rmk-5")
    wb = load_workbook(io.BytesIO(_download(client, project)))
    wb[wbk.SHEET_EXPENSE].cell(row=2, column=len(wbk._EXPENSE_HEADERS)).value = None
    buf = io.BytesIO()
    wb.save(buf)
    assert _upload(client, project, buf.getvalue()).status_code == 200
    db.expire_all()
    assert db.execute(select(FProjectExpense)).scalars().one().remark is None


def test_unchanged_remark_is_not_a_write(db, project):
    _expense(db)
    expense = db.execute(select(FProjectExpense)).scalars().one()
    expense.remark = "没动过"
    db.commit()
    client = uploader(db, "rmk-6")
    content = _download(client, project)
    assert _upload(client, project, content).json()["expense_updates"] == 0
