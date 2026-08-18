"""补库证据与四列导出测试（F2，round-4 修复版）。

反例覆盖（Codex round-4 最小集）：
- draft / 非 owner / other-owner 全部拒绝（404/409）；
- 导出精确四列 + 公式转义 + 跨版本累计批准意向；
- 未来/作废事实不进入 inactivity/high-frequency；CKD 已应用成本为区间来源；
  归档/excluded 替代件被过滤。
"""

import io
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.models.dimensions import DimPart
from app.models.inventory import PartPool, PartPoolMember
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder
from app.models.maintenance_ckd_import import (
    MaintenanceCkdHeadRow,
    MaintenanceCkdImportBatch,
    MaintenanceCkdLineRow,
)
from app.models.maintenance_project import MaintenanceProject
from app.models.purchase import FPurchaseLine, FPurchaseOrder
from app.models.replenishment import (
    ReplenishmentApplication,
    ReplenishmentApplicationLine,
    ReplenishmentApplicationVersion,
    ReplenishmentReview,
    ReplenishmentReviewLine,
)
from app.models.sales import FSalesLine, FSalesOrder
from app.models.system import SysImportBatch, SysUser
from app.services import maintenance_replenishment_evidence as evidence
from app.services import replenishment as replenishment_service

OWNER = "evidence-owner"
HEX64 = "ab" * 32


@pytest.fixture()
def owner_user(db):
    db.add(
        SysUser(
            username=OWNER,
            role="sales",
            display_name="证据测试销售",
            password_hash="x",
        )
    )
    db.commit()


def _project(db, code: str = "EVTEST-0001") -> MaintenanceProject:
    project = MaintenanceProject(
        project_id=str(uuid4()),
        project_code=code,
        display_name=f"证据测试项目-{code}",
        lifecycle_status="ongoing",
        is_active=True,
        version=1,
    )
    db.add(project)
    db.flush()
    return project


def _line(
    db,
    *,
    line_id: str,
    request_line_id: str,
    version_id: str,
    line_no: int,
    part_id: int,
    pn: str,
    quantity: str = "5",
    purchase: dict | None = None,
    sales: dict | None = None,
    pool_group_id: int | None = None,
    pool_name: str | None = None,
) -> ReplenishmentApplicationLine:
    line = ReplenishmentApplicationLine(
        line_id=line_id,
        request_line_id=request_line_id,
        version_id=version_id,
        line_no=line_no,
        part_id=part_id,
        pn_std=pn,
        quantity=Decimal(quantity),
        pool_group_id=pool_group_id,
        pool_name=pool_name,
        pool_version=1 if pool_group_id is not None else None,
        price_window_from=date(2026, 2, 1),
        price_window_to=date(2026, 8, 1),
        price_as_of=date(2026, 8, 1),
        purchase_stats_json=purchase
        or {"weighted_avg": None, "total_qty": 0, "order_count": 0},
        sales_stats_json=sales
        or {"weighted_avg": None, "total_qty": 0, "order_count": 0},
        # 2026-08-18：version 提交 guard（guard_replenishment_atomic_submission）
        # 要求行级 screening_json 完整——建行时补一份符合 validator 的筛查快照
        screening_json={
            "schema_version": 1,
            "as_of": "2026-08-17",
            "lookback_days": 182,
            "checks": [
                {
                    "key": "pool_membership",
                    "passed": pool_group_id is not None,
                    "detail": {
                        "in_pool": pool_group_id is not None,
                        "pool_name": pool_name,
                        "pool_status": "active" if pool_group_id is not None else None,
                    },
                },
                {
                    "key": "recent_activity",
                    "passed": False,
                    "detail": {
                        "window": {"from": "2026-02-17", "to": "2026-08-17"},
                        "purchase_samples": 0,
                        "sales_samples": 0,
                    },
                },
                {
                    "key": "niche_pn",
                    "passed": False,
                    "detail": {
                        "is_niche": False,
                        "purchase_samples": 0,
                        "sales_samples": 0,
                        "rule": "no_purchase_or_sales_in_182_days",
                    },
                },
            ],
            "anomaly_count": 1,
            "latest_sales": sales or {"weighted_avg": None, "total_qty": 0},
            "pool_floor_ex_tax": None,
        },
        evidence_digest=HEX64,
    )
    db.add(line)
    db.flush()
    return line


def _version(db, *, version_id: str, application_id: str, version_no: int) -> None:
    db.add(
        ReplenishmentApplicationVersion(
            version_id=version_id,
            application_id=application_id,
            version_no=version_no,
            status="draft",
            warehouse="北京前置库",
            created_by=OWNER,
        )
    )
    db.flush()


def _submit_version(db, *, version_id: str) -> None:
    """行插入完成后再提交冻结（触发器的 submitted 不可变守卫只允许 draft 建行）。"""
    version = db.get(ReplenishmentApplicationVersion, version_id)
    version.status = "submitted"
    version.content_digest = HEX64
    version.submitted_by = OWNER
    version.submitted_at = datetime(2026, 8, 1, tzinfo=timezone.utc)
    db.flush()


def _review(
    db, *, version_id: str, decisions: list[tuple[str, str, str | None]]
) -> None:
    review_id = f"review-{version_id}"
    approved = sum(1 for _, d, _ in decisions if d == "approved")
    rejected = sum(1 for _, d, _ in decisions if d == "rejected")
    db.add(
        ReplenishmentReview(
            review_id=review_id,
            version_id=version_id,
            idempotency_key=f"key-{version_id}",
            payload_digest=HEX64,
            approved_count=approved,
            rejected_count=rejected,
            reviewed_by="reviewer",
        )
    )
    db.flush()
    for index, (line_id, decision, reason) in enumerate(decisions):
        db.add(
            ReplenishmentReviewLine(
                review_line_id=f"{review_id}-{index}",
                review_id=review_id,
                version_line_id=line_id,
                decision=decision,
                reason=reason,
            )
        )
    db.flush()


@pytest.fixture()
def approved_two_version(db, owner_user):
    """v1：PN-A 打回、PN-B 通过；v2：PN-A 复提通过 → application approved。"""
    part_a = DimPart(pn_std="=WEBSERVICE(1)", description="公式样式PN")
    part_b = DimPart(pn_std="PN-B-001", description="正常PN")
    db.add_all([part_a, part_b])
    db.flush()
    application = ReplenishmentApplication(
        application_id="evidence-app-1",
        application_no="EV-APP-0001",
        owner_username=OWNER,
        project_id=_project(db).project_id,
        project_code_snapshot="EVTEST-0001",
        project_name_snapshot="证据测试项目",
        client_request_id="evidence-manual-crid",
        request_digest="abababababababababababababababababababababababababababababababab",
        status="draft",
        latest_version_no=2,
        version=1,
    )
    db.add(application)
    db.flush()
    _version(db, version_id="ev-ver-1", application_id="evidence-app-1", version_no=1)
    line_a1 = _line(
        db, line_id="ev-a1", request_line_id="ev-req-a", version_id="ev-ver-1",
        line_no=1, part_id=part_a.id, pn="=WEBSERVICE(1)", quantity="5",
    )
    line_b1 = _line(
        db, line_id="ev-b1", request_line_id="ev-req-b", version_id="ev-ver-1",
        line_no=2, part_id=part_b.id, pn="PN-B-001", quantity="3",
        purchase={"weighted_avg": 100.0, "total_qty": 60, "order_count": 6},
        sales={"weighted_avg": 113.0, "total_qty": 40, "order_count": 4},
    )
    _submit_version(db, version_id="ev-ver-1")
    _review(
        db,
        version_id="ev-ver-1",
        decisions=[
            ("ev-a1", "rejected", "请说明特殊原因"),
            ("ev-b1", "approved", None),
        ],
    )
    _version(db, version_id="ev-ver-2", application_id="evidence-app-1", version_no=2)
    line_a2 = _line(
        db, line_id="ev-a2", request_line_id="ev-req-a", version_id="ev-ver-2",
        line_no=1, part_id=part_a.id, pn="=WEBSERVICE(1)", quantity="5",
        purchase={"weighted_avg": 100.0, "total_qty": 60, "order_count": 6},
        sales={"weighted_avg": 113.0, "total_qty": 40, "order_count": 4},
    )
    _submit_version(db, version_id="ev-ver-2")
    _review(db, version_id="ev-ver-2", decisions=[("ev-a2", "approved", None)])
    application.status = "approved"
    application.version += 1
    db.commit()
    return {"application_id": application.application_id, "part_b": part_b.id}


def test_draft_rejected_409_and_other_owner_404(db, owner_user):
    part = DimPart(pn_std="PN-DRAFT-001", description="草稿件")
    db.add(part)
    db.flush()
    application = ReplenishmentApplication(
        application_id="evidence-draft-1",
        application_no="EV-APP-DRAFT",
        owner_username=OWNER,
        project_id=_project(db).project_id,
        project_code_snapshot="EVTEST-0001",
        project_name_snapshot="证据测试项目",
        client_request_id="evidence-manual-crid",
        request_digest="abababababababababababababababababababababababababababababababab",
        status="draft",
        latest_version_no=1,
        version=1,
    )
    db.add(application)
    db.flush()
    _version(db, version_id="ev-draft-ver-1", application_id="evidence-draft-1", version_no=1)
    _line(
        db, line_id="ev-draft-a1", request_line_id="ev-draft-req-a",
        version_id="ev-draft-ver-1", line_no=1, part_id=part.id, pn="PN-DRAFT-001",
    )
    db.commit()

    with pytest.raises(replenishment_service.ReplenishmentError) as exc:
        evidence.application_evidence(
            db, "evidence-draft-1", username=OWNER, role="sales"
        )
    assert exc.value.status_code == 409
    with pytest.raises(replenishment_service.ReplenishmentError) as exc:
        evidence.export_purchase_list(
            db, "evidence-draft-1", username=OWNER, role="sales"
        )
    assert exc.value.status_code == 409

    # 非 owner 与不存在同 404
    for fn in (evidence.application_evidence, evidence.export_purchase_list):
        with pytest.raises(replenishment_service.ReplenishmentError) as exc:
            fn(db, "evidence-draft-1", username="other-sales", role="sales")
        assert exc.value.status_code == 404
        with pytest.raises(replenishment_service.ReplenishmentError) as exc:
            fn(db, "no-such-app", username=OWNER, role="sales")
        assert exc.value.status_code == 404


def test_export_exactly_four_columns_cumulative_and_escaped(db, approved_two_version):
    data = evidence.export_purchase_list(
        db, "evidence-app-1", username=OWNER, role="sales"
    )
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook["补库采购清单"]
    assert sheet.max_column == 4
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0] == ("PN", "数量", "采购金额(参考)", "销售金额(参考)")
    assert len(rows) == 3  # 表头 + 两条累计批准意向
    pns = {row[0] for row in rows[1:]}
    assert pns == {"'=WEBSERVICE(1)", "PN-B-001"}  # 公式型 PN 已转义
    for row in rows[1:]:
        assert len(row) == 4


def test_evidence_uses_effective_facts_only(db, owner_user):
    """未来/作废事实不参与；CKD 已应用成本是区间来源；excluded 替代件被过滤。"""
    part_c = DimPart(pn_std="PN-C-001", description="高频候选")
    part_alt = DimPart(pn_std="PN-ALT-001", description="已排除替代", is_excluded=True)
    db.add_all([part_c, part_alt])
    db.flush()
    pool = PartPool(group_id=9002, name="证据池", status="active", member_count=1)
    db.add(pool)
    db.flush()
    db.add(PartPoolMember(group_id=9002, part_id=part_c.id))
    db.add(PartPoolMember(group_id=9002, part_id=part_alt.id))
    import_batch = SysImportBatch(
        filename="w.xlsx", file_type="maintenance", file_hash="h-ev", status="success"
    )
    db.add(import_batch)
    db.flush()
    today = date.today()
    # 未来 60 件补库供货（已生效）→ 不得制造高频
    future_order = FMaintenanceOrder(
        raw_order_id="ev-wbdd-future",
        order_no="WBDD-20260801-9999",
        order_date=today + timedelta(days=30),
        demand_type="补库供货",
        business_type="整体维保",
        project_raw="EV项目",
        project_std="EV项目",
        warehouse="北京成品仓",
        data_status="已生效",
        import_batch_id=import_batch.id,
    )
    db.add(future_order)
    db.flush()
    db.add(
        FMaintenanceLine(
            raw_line_id="ev-line-future",
            order_id=future_order.id,
            line_no=1,
            part_id=part_c.id,
            pn_std="PN-C-001",
            qty=Decimal("60"),
            unit_cost=Decimal("300"),
            import_batch_id=import_batch.id,
        )
    )
    # 昨天的已作废采购/销售 → 不得消除 365 天无记录提醒
    cancelled_purchase = FPurchaseOrder(
        raw_order_id="ev-cg-cancelled",
        order_no="CG-CANCELLED-1",
        order_date=today - timedelta(days=1),
        data_status="已作废",
        import_batch_id=import_batch.id,
    )
    db.add(cancelled_purchase)
    db.flush()
    db.add(
        FPurchaseLine(
            raw_line_id="ev-cg-line-cancelled",
            order_id=cancelled_purchase.id,
            line_no=1,
            part_id=part_c.id,
            pn_std="PN-C-001",
            qty=Decimal("2"),
            unit_price=Decimal("10"),
            line_amount=Decimal("20"),
            import_batch_id=import_batch.id,
        )
    )
    cancelled_sales = FSalesOrder(
        raw_order_id="ev-xs-cancelled",
        order_no="XS-CANCELLED-1",
        order_date=today - timedelta(days=1),
        data_status="已作废",
        import_batch_id=import_batch.id,
    )
    db.add(cancelled_sales)
    db.flush()
    db.add(
        FSalesLine(
            raw_line_id="ev-xs-line-cancelled",
            order_id=cancelled_sales.id,
            line_no=1,
            part_id=part_c.id,
            pn_std="PN-C-001",
            qty=Decimal("1"),
            unit_price=Decimal("12"),
            line_amount=Decimal("12"),
            import_batch_id=import_batch.id,
        )
    )
    # 已应用 CKD 发货单成本 90/110
    ckd_batch = MaintenanceCkdImportBatch(
        batch_id="ev-ckd-batch",
        file_hash="h-ckd",
        filename="发货单.xlsx",
        idempotency_key="ev-ckd-key",
        uploaded_by=OWNER,
        status="applied",
        applied_by=OWNER,
        applied_at=datetime.now(timezone.utc),
    )
    db.add(ckd_batch)
    db.flush()
    db.add(
        MaintenanceCkdHeadRow(
            row_id="ev-ckd-head-1",
            batch_id="ev-ckd-batch",
            row_no=1,
            order_no="CKD-20260801-0001",
            category="维保供货",
            data_status_raw="已生效",
            order_date=today,
            issues=[],
        )
    )
    db.flush()
    # 销售出库头（已生效但非维保供货）→ 成本不进区间（round-5 Blocker 12 反例）
    db.add(
        MaintenanceCkdHeadRow(
            row_id="ev-ckd-head-sales",
            batch_id="ev-ckd-batch",
            row_no=2,
            order_no="CKD-20260801-0099",
            category="销售出库",
            data_status_raw="已生效",
            order_date=today,
            issues=[],
        )
    )
    db.flush()
    db.add(
        MaintenanceCkdLineRow(
            row_id="ev-ckd-line-sales",
            batch_id="ev-ckd-batch",
            head_row_id="ev-ckd-head-sales",
            row_no=3,
            pn="PN-C-001",
            out_qty=Decimal("1"),
            unit_cost=Decimal("300"),
        )
    )
    for index, cost in enumerate(("90", "110")):
        db.add(
            MaintenanceCkdLineRow(
                row_id=f"ev-ckd-line-{index}",
                batch_id="ev-ckd-batch",
                head_row_id="ev-ckd-head-1",
                row_no=index + 1,
                pn="PN-C-001",
                out_qty=Decimal("1"),
                unit_cost=Decimal(cost),
            )
        )

    application = ReplenishmentApplication(
        application_id="evidence-app-2",
        application_no="EV-APP-0002",
        owner_username=OWNER,
        project_id=_project(db).project_id,
        project_code_snapshot="EVTEST-0001",
        project_name_snapshot="证据测试项目",
        client_request_id="evidence-manual-crid",
        request_digest="abababababababababababababababababababababababababababababababab",
        status="draft",
        latest_version_no=1,
        version=1,
    )
    db.add(application)
    db.flush()
    _version(db, version_id="ev2-ver-1", application_id="evidence-app-2", version_no=1)
    _line(
        db, line_id="ev2-line-c", request_line_id="ev2-req-c", version_id="ev2-ver-1",
        line_no=1, part_id=part_c.id, pn="PN-C-001",
        pool_group_id=9002, pool_name="证据池",
    )
    _submit_version(db, version_id="ev2-ver-1")
    _review(db, version_id="ev2-ver-1", decisions=[("ev2-line-c", "approved", None)])
    application.status = "approved"
    application.version += 1
    db.commit()

    payload = evidence.application_evidence(
        db, "evidence-app-2", username=OWNER, role="sales"
    )
    row = payload["lines"][0]
    assert row["is_high_frequency"] is False  # 未来订单不计
    assert row["recent_supply_qty"] == 0.0
    assert row["inactive_365d"] is True  # 作废采购/销售不计
    assert row["inactive_sides"] == ["purchase", "sales"]
    assert float(row["ckd_unit_cost_min"]) == 90.0
    assert float(row["ckd_unit_cost_max"]) == 110.0
    assert row["ckd_unit_cost_sample_count"] == 2
    # 旧 WBDD 派生成本 300 不是区间来源
    assert 300.0 not in (float(row["ckd_unit_cost_min"]), float(row["ckd_unit_cost_max"]))
    # 排除件不得成为替代建议
    assert all(alt["pn_std"] != "PN-ALT-001" for alt in row["pool_alternatives"])


def test_evidence_and_export_owner_scope_http(db, approved_two_version):
    """F2 两端点 owner/other-owner/admin 的 HTTP 200/404 矩阵（round-5 Blocker 1/12）。"""
    from fastapi.testclient import TestClient

    from app import permissions
    from app.auth import hash_password
    from app.config import get_settings
    from app.main import app
    from app.models.system import SysUser

    password = "safe-test-password"
    custom = permissions.effective("readonly", None)
    custom.update(
        {
            "page_replenishment_beta": True,
            "action_replenishment_create": True,
            "data_pool_price_governance": True,
        }
    )
    owner_user = db.query(SysUser).filter_by(username=OWNER).one()
    owner_user.password_hash = hash_password(password)
    owner_user.permissions = custom
    owner_user.is_active = True
    db.add_all(
        [
            SysUser(
                username="other-sales-http",
                role="sales",
                display_name="其他销售",
                password_hash=hash_password(password),
                permissions=custom,
                is_active=True,
            ),
            SysUser(
                username="evidence_admin_http",
                role="admin",
                display_name="管理员",
                password_hash=hash_password(password),
                permissions=custom,
                is_active=True,
            ),
        ]
    )
    db.commit()

    def client_for(username: str) -> TestClient:
        client = TestClient(app)
        login = client.post(
            "/api/auth/login", json={"username": username, "password": password}
        )
        assert login.status_code == 200, login.text
        client.headers["Authorization"] = f"Bearer {login.json()['token']}"
        return client

    settings = get_settings()
    original = settings.replenishment_beta_enabled
    try:
        settings.replenishment_beta_enabled = True
        evidence_path = "/api/replenishment-beta/applications/evidence-app-1/evidence"
        export_path = (
            "/api/replenishment-beta/applications/evidence-app-1/"
            "exports/purchase-list.xlsx"
        )
        owner = client_for(OWNER)
        other = client_for("other-sales-http")
        admin = client_for("evidence_admin_http")
        # 2026-08-18：旧 evidence/导出 HTTP 端点已停用（_retired → 410），
        # _retired 在权限检查前执行——对所有用户（含非 owner/不存在）统一 410。
        for who in (owner, other, admin):
            assert who.get(evidence_path).status_code == 410
            assert who.get(export_path).status_code == 410
        missing = owner.get(
            "/api/replenishment-beta/applications/no-such-app/evidence"
        )
        assert missing.status_code == 410
    finally:
        settings.replenishment_beta_enabled = original
