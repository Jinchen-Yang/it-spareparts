"""维保项目可选字段导出：白名单、权限与 XLSX 契约。"""

from datetime import date
from decimal import Decimal
from io import BytesIO
import uuid

import pytest
from openpyxl import load_workbook

from app.config import get_settings
from app.models.maintenance_project import MaintenanceProjectContract
from app.models.maintenance_project_operations import MaintenanceCollectionSnapshot
from app.services import maintenance_project_export
from tests.boss_board_helpers import boss_client, make_project


@pytest.fixture(autouse=True)
def _flag_on():
    settings = get_settings()
    original = settings.maintenance_boss_dashboard_enabled
    settings.maintenance_boss_dashboard_enabled = True
    try:
        yield
    finally:
        settings.maintenance_boss_dashboard_enabled = original


def test_export_options_return_whitelist_and_exact_default_fields(db):
    response = boss_client(db, username="project-export-options").get(
        "/api/maintenance/boss-board/projects/export/options"
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["default_fields"] == [
        "project_name",
        "period_from",
        "period_to",
        "contract_nos",
        "contract_amount_inc_tax",
        "collection_received_inc_tax",
    ]
    keys = [field["key"] for field in payload["fields"]]
    assert len(keys) == len(set(keys))
    assert set(payload["default_fields"]) <= set(keys)
    assert all(set(field) == {"key", "label", "group", "default_selected"}
               for field in payload["fields"])


def test_export_options_and_post_filter_profit_fields_by_server_permission(db):
    client = boss_client(
        db,
        username="project-export-no-profit",
        with_profit=False,
    )
    options = client.get(
        "/api/maintenance/boss-board/projects/export/options"
    ).json()
    keys = {field["key"] for field in options["fields"]}

    assert "contract_amount_inc_tax" not in keys
    assert "collection_received_inc_tax" not in keys
    assert options["default_fields"] == [
        "project_name", "period_from", "period_to", "contract_nos"
    ]

    forbidden = client.post(
        "/api/maintenance/boss-board/projects/export",
        json={"fields": ["project_name", "contract_amount_inc_tax"]},
    )
    assert forbidden.status_code == 403
    assert forbidden.json()["detail"] == {
        "code": "export_field_permission_denied",
        "message": "所选字段超出当前账号的数据权限",
        "fields": ["contract_amount_inc_tax"],
    }


def test_export_rejects_unknown_field_instead_of_resolving_orm_attribute(db):
    response = boss_client(db, username="project-export-unknown").post(
        "/api/maintenance/boss-board/projects/export",
        json={"fields": ["project_name", "__dict__"]},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == {
        "code": "invalid_export_fields",
        "message": "包含服务端白名单之外的导出字段",
        "fields": ["__dict__"],
    }


def test_financial_state_fields_do_not_call_missing_values_complete():
    fields = {
        field.key: field
        for field in maintenance_project_export.EXPORT_FIELDS
    }
    empty = {
        "contract_amount_inc_tax": {"state": "ready", "value": None},
        "collection_preview_inc_tax": {"state": "ready", "value": None},
        "contract_incomplete": False,
    }
    partial = {
        "contract_amount_inc_tax": {"state": "partial", "value": Decimal("10")},
        "contract_incomplete": True,
    }

    assert fields["contract_amount_state"].accessor(empty) == "无有效合同"
    assert fields["collection_state"].accessor(empty) == "尚未上报"
    assert (
        fields["contract_amount_state"].accessor(partial)
        == "合同事实不完整（已知小计）"
    )


def test_export_xlsx_uses_current_contract_and_latest_confirmed_collection(db):
    project = make_project(db, code="=公式项目", lifecycle="ongoing")
    make_project(db, code="已结束项目", lifecycle="ended")
    relation = MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project.project_id,
        contract_id="contract-export-1",
        contract_no="XSDD-EXPORT-001",
        contract_amount=Decimal("1000.00"),
        amount_inc_tax=Decimal("1130.00"),
        contract_status="执行中",
        status_mapping_state="mapped",
        status_mapping_version="v1",
        included_in_total=True,
        effective_from=date(2026, 1, 1),
        source="synthetic_test",
        version=1,
    )
    db.add(relation)
    db.add_all([
        MaintenanceCollectionSnapshot(
            collection_id=str(uuid.uuid4()),
            project_id=project.project_id,
            project_contract_id=relation.project_contract_id,
            report_month=date(2026, 7, 1),
            cumulative_amount=Decimal("300.00"),
            status="confirmed",
            source="direct_api",
            version=1,
        ),
        MaintenanceCollectionSnapshot(
            collection_id=str(uuid.uuid4()),
            project_id=project.project_id,
            project_contract_id=relation.project_contract_id,
            report_month=date(2026, 8, 1),
            cumulative_amount=Decimal("450.00"),
            status="confirmed",
            source="direct_api",
            version=1,
        ),
    ])
    db.commit()

    fields = list(maintenance_project_export.DEFAULT_FIELDS)
    response = boss_client(db, username="project-export-xlsx").post(
        "/api/maintenance/boss-board/projects/export",
        json={"fields": fields, "lifecycle": "ongoing", "sort": "name"},
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    assert response.headers["x-export-row-count"] == "1"

    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=False)
    worksheet = workbook["维保项目清单"]
    table = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    assert table[0] == (
        "项目名称",
        "维保起始时间",
        "维保终止时间",
        "销售单号（合同号）",
        "合同总额（含税）",
        "累计已回款（含税）",
    )
    assert table[1][0] == "'=公式项目"  # 防 Excel 公式注入
    assert table[1][1].date() == date(2020, 1, 1)
    assert table[1][2].date() == date(2099, 12, 31)
    assert table[1][3:] == ("XSDD-EXPORT-001", 1130, 450)
