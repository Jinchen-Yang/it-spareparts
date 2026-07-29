"""项目追踪工作簿表单闭环（§17）：宽松报销导入 / 页级锚 / 内容幂等键 /
多 sheet 白名单 / upsert=以本表为准 / 导出↔导入 round-trip。"""
from decimal import Decimal

import pandas as pd
import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from app.etl import mapping, pipeline, reader
from app.etl.transform import transform
from app.models.maintenance import FProjectExpense
from app.models.system import SysImportBatch

_CANON = ["报销日期", "报销人员", "报销类别", "费用分类", "支出事由",
          "报销金额", "流程状态", "单号", "序号"]


@pytest.fixture()
def batch(db):
    b = SysImportBatch(filename="t.xlsx", file_type="expense", file_hash="hwf")
    db.add(b)
    db.flush()
    return b


def _canon_row(d="2026-05-01", person="张三", amount=100, reason="外援",
               status=None, bxd=None, seq=None, fee="外援劳务"):
    return dict(zip(_CANON, [d, person, "维保费用", fee, reason, amount, status, bxd, seq]))


def _workbook_xlsx(tmp_path, name, exp_rows, anchor="XSDD-1", with_parts=True,
                   with_total_row=True, marker=""):
    """构造项目追踪工作簿：预算页(kv) + 备件页(WBDD表头) + 报销页(锚+canonical) + 说明页。

    marker：写进预算页的杂散单元格，用于制造"内容相同但文件 hash 不同"的重传场景。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "项目预算"
    ws.append(["合同（销售订单）", anchor])
    ws.append(["合同金额（含税参考）", 50000])
    if marker:
        ws.append(["备注", marker])
    if with_parts:
        ws2 = wb.create_sheet("备件明细-氚云")
        ws2.append(["数据ID(不可修改)", "需求单号", "制单日期", "需求类型",
                    "需求明细.数据ID(不可修改)", "需求明细.需供货产品",
                    "需求明细.产品描述", "需求明细.需求数量"])
        ws2.append(["MH1", "WBDD-20260501-0001", "2026-05-01", "报修供货",
                    "ML1", "PN-X", "HDD", 1])
    ws3 = wb.create_sheet("报销明细")
    ws3.append(["销售订单", anchor])
    ws3.append(_CANON)
    for r in exp_rows:
        ws3.append([r.get(c) for c in _CANON])
    if with_total_row:
        ws3.append([None, None, None, None, "合计（仅已结束）", 999999, None, None, None])
    ws4 = wb.create_sheet("填写说明")
    ws4["A1"] = "说明"
    p = tmp_path / name
    wb.save(str(p))
    return str(p)


# ---------- 识别（§17.3 宽松签名） ----------

def test_detect_loose_expense_signature():
    assert mapping.detect_file_type(_CANON) == mapping.EXPENSE
    assert mapping.detect_file_type(["报销日期", "报销金额"]) == mapping.EXPENSE
    assert mapping.detect_file_type(["费用分类", "报销金额(必填)"]) == mapping.EXPENSE
    # 裸「金额」不吃：防误伤
    assert mapping.detect_file_type(["日期", "金额"]) is None


# ---------- transform：必需列 / 锚 / 默认状态 / 内容键 ----------

def test_loose_transform_anchor_defaults_and_content_key():
    df = pd.DataFrame([
        _canon_row(amount=100),                      # 无状态 → 默认已结束
        _canon_row(amount=100),                      # 同内容 → dup_idx 区分
        _canon_row(d=None, amount=300, reason="月度合计"),  # 合计行 → 跳过（非合计缺日期=报错另测）
        _canon_row(amount=None),                     # 缺金额 → 跳过
        _canon_row(amount=200, status="流程中", bxd="BXD-20260501-3", seq=2),
    ])
    res = transform(df, mapping.EXPENSE, anchor="XSDD-A")
    assert not res.errors
    assert res.rows_skipped_no_data == 2
    assert len(res.lines) == 3
    k0, k1, k2 = (r["raw_line_id"] for r in res.lines)
    assert k0.startswith("EXP:") and k0.endswith("#0")
    assert k1.startswith("EXP:") and k1.endswith("#1") and k0[:41] == k1[:41]
    assert k2.startswith("BXD-20260501-3#2@")        # 单号+序号 复合键，带合同域后缀
    assert all(r["linked_sales_order_no"] == "XSDD-A" for r in res.lines)
    assert res.lines[0]["data_status"] == "已结束"    # 默认生效
    assert res.lines[2]["data_status"] == "流程中"
    assert res.rows_inactive == 1


def test_loose_transform_row_link_beats_anchor_and_missing_link_errors():
    df = pd.DataFrame([
        {**_canon_row(amount=50), "销售订单": "XSDD-ROW"},
        _canon_row(amount=60),
    ])
    res = transform(df, mapping.EXPENSE, anchor="XSDD-PAGE")
    assert res.lines[0]["linked_sales_order_no"] == "XSDD-ROW"    # 行级优先
    assert res.lines[1]["linked_sales_order_no"] == "XSDD-PAGE"
    res2 = transform(pd.DataFrame([_canon_row(amount=70)]), mapping.EXPENSE)  # 无锚无行级
    assert not res2.lines and res2.errors[0].error_type == "missing_link"


def test_expense_transform_uses_postgresql_money_rounding_for_both_tax_bases():
    df = pd.DataFrame([
        _canon_row(amount=Decimal("0.50"), reason="未税中点"),
        {
            **_canon_row(
                amount=Decimal("1.00"),
                reason="含税反算",
                d="2026-05-02",
            ),
            "金额口径": "含税",
        },
        _canon_row(
            amount=Decimal("0.505"),
            reason="先舍入权威金额再派生",
            d="2026-05-03",
        ),
        {
            **_canon_row(
                amount=Decimal("1.00"),
                reason="显式双列含税权威",
                d="2026-05-04",
            ),
            "未税金额": Decimal("0.88"),
            "含税金额": Decimal("1.00"),
            "金额口径": "含税",
        },
    ])

    result = transform(df, mapping.EXPENSE, anchor="XSDD-ROUND")

    assert not result.errors
    assert result.lines[0]["amount_ex_tax"] == Decimal("0.50")
    assert result.lines[0]["amount_inc_tax"] == Decimal("0.57")
    assert result.lines[1]["amount_inc_tax"] == Decimal("1.00")
    assert result.lines[1]["amount_ex_tax"] == Decimal("0.88")
    assert result.lines[2]["amount_ex_tax"] == Decimal("0.51")
    assert result.lines[2]["amount_inc_tax"] == Decimal("0.58")
    assert result.lines[3]["tax_basis"] == "inc"
    assert result.lines[3]["amount"] == Decimal("1.00")
    assert result.lines[3]["amount_ex_tax"] == Decimal("0.88")
    assert result.lines[3]["amount_inc_tax"] == Decimal("1.00")


def test_expense_transform_rejects_raw_amount_that_disagrees_with_tax_basis():
    row = {
        **_canon_row(amount=Decimal("999.00"), reason="原值冲突"),
        "未税金额": Decimal("0.88"),
        "含税金额": Decimal("1.00"),
        "金额口径": "含税",
    }

    result = transform(
        pd.DataFrame([row]),
        mapping.EXPENSE,
        anchor="XSDD-ROUND",
    )

    assert not result.lines
    assert result.errors[0].error_type == "inconsistent_raw_amount"


# ---------- reader：锚行探测 + 多 sheet ----------

def test_read_workbook_multisheet_anchor(db, tmp_path):
    p = _workbook_xlsx(tmp_path, "wb1.xlsx", [_canon_row()])
    sheets = reader.read_workbook(p)
    by_type = {s.file_type: s for s in sheets}
    assert set(by_type) == {mapping.MAINTENANCE, mapping.EXPENSE}   # 预算/说明页不识别
    exp = by_type[mapping.EXPENSE]
    assert exp.sheet_name == "报销明细" and exp.anchor == "XSDD-1"
    assert list(exp.df.columns)[:6] == _CANON[:6]


# ---------- pipeline：多 sheet 白名单 + 报告 ----------

def test_run_import_workbook_only_ingests_expense(db, tmp_path):
    p = _workbook_xlsx(tmp_path, "wb2.xlsx", [_canon_row(), _canon_row(amount=200, reason="快递")])
    batch = pipeline.run_import(db, p, "wb2.xlsx")
    db.commit()
    assert batch.file_type == "workbook"
    assert batch.rows_inserted == 2
    rep = batch.report_json
    assert len(rep["skipped_sheets"]) == 1 and rep["skipped_sheets"][0].startswith("备件明细-氚云")
    assert rep["rows_skipped_no_data"] == 1                         # 合计行
    # 备件页确实没入库
    from app.models.maintenance import FMaintenanceOrder
    assert db.scalar(select(func.count()).select_from(FMaintenanceOrder)) == 0
    # 锚归集
    assert db.scalar(select(func.count()).select_from(FProjectExpense)
                     .where(FProjectExpense.linked_sales_order_no == "XSDD-1")) == 2


def test_run_import_workbook_reupload_idempotent(db, tmp_path):
    rows = [_canon_row(), _canon_row(amount=200, reason="快递")]
    p1 = _workbook_xlsx(tmp_path, "wb3.xlsx", rows)
    pipeline.run_import(db, p1, "wb3.xlsx")
    db.commit()
    # 同报销内容但文件字节不同（预算页加备注换 hash）→ 不触发文件级去重，考验行级内容键
    p2 = _workbook_xlsx(tmp_path, "wb3b.xlsx", rows, marker="财务月更 v2")
    b2 = pipeline.run_import(db, p2, "wb3b.xlsx")
    db.commit()
    assert b2.rows_inserted == 0 and b2.rows_skipped == 2           # 内容键幂等
    assert db.scalar(select(func.count()).select_from(FProjectExpense)) == 2


def test_run_import_upsert_replaces_contract(db, tmp_path):
    p1 = _workbook_xlsx(tmp_path, "wb4.xlsx",
                        [_canon_row(), _canon_row(amount=200, reason="快递")])
    pipeline.run_import(db, p1, "wb4.xlsx")
    db.commit()
    # 财务改了金额（100→150）并删掉快递行 → 修复模式=以本表为准
    p2 = _workbook_xlsx(tmp_path, "wb4b.xlsx", [_canon_row(amount=150)])
    b2 = pipeline.run_import(db, p2, "wb4b.xlsx", mode="upsert")
    db.commit()
    assert b2.report_json["expense_rows_replaced"] == 2
    amts = db.scalars(select(FProjectExpense.amount)
                      .where(FProjectExpense.linked_sales_order_no == "XSDD-1")).all()
    assert amts == [Decimal("150")]                                 # 无旧行残影


def test_workbook_without_expense_falls_back_to_first_sheet(db, tmp_path):
    """无报销页的多页文件 → 老语义：导第一个可识别页，其余页进 ignored_sheets 提示。

    典型场景=隐藏备份页/副本页（审查 D1）：此前会整批失败，现照常导入。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "库存"
    ws.append(["产品库存ID", "产品名称(PN)", "库存数量", "仓库"])
    ws.append(["INV-8", "PN-Y", 5, "总仓"])
    ws2 = wb.create_sheet("库存 (2)")                                # 副本页
    ws2.append(["产品库存ID", "产品名称(PN)", "库存数量", "仓库"])
    ws2.append(["INV-9", "PN-Y", 3, "总仓"])
    p = tmp_path / "wb5.xlsx"
    wb.save(str(p))
    batch = pipeline.run_import(db, str(p), "wb5.xlsx")
    db.commit()
    assert batch.status == "success" and batch.file_type == "inventory"
    assert batch.rows_inserted == 1                                  # 只导第一页
    assert batch.report_json["ignored_sheets"][0].startswith("库存 (2)")


def test_single_sheet_maintenance_still_imports(db, tmp_path):
    """单表维保出库文件行为不变（维保需求单样式不动，权威源不受白名单影响）。"""
    df = pd.DataFrame([{
        "数据ID(不可修改)": "MH9", "需求单号": "WBDD-20260501-0009",
        "制单日期": "2026-05-01", "需求类型": "报修供货",
        "销售订单": "XSDD-9", "项目名": "某项目",
        "需求明细.数据ID(不可修改)": "ML9", "需求明细.需供货产品": "PN-Z",
        "需求明细.产品描述": "HDD", "需求明细.需求数量": 1,
    }])
    p = tmp_path / "wbdd.xlsx"
    df.to_excel(str(p), index=False)
    batch = pipeline.run_import(db, str(p), "wbdd.xlsx")
    db.commit()
    assert batch.file_type == mapping.MAINTENANCE and batch.rows_inserted == 1


# ---------- round-trip：导出 → 再导入 = 零新增（§17.6 核心不变量）----------

def test_export_workbook_true_roundtrip(db, batch, tmp_path):
    """真实闭环：宽松导入产生内容键 → 导出 → 再导入 → 零新增零错误。"""
    from app.api.maintenance import _build_workbook
    from app.services import maintenance_cost

    p1 = _workbook_xlsx(tmp_path, "rt1.xlsx",
                        [_canon_row(amount=100), _canon_row(amount=100),
                         # 非整数金额：库内 Numeric(…,2) → 导出 float → 再解析，
                         # parse_money 统一量化 0.01 保证内容键稳定
                         _canon_row(amount=123.45, reason="快递", status="流程中"),
                         _canon_row(amount=100.5, reason="打车"),
                         _canon_row(amount=300, bxd="BXD-20260502-1", seq=1)],
                        anchor="XSDD-RT2", with_parts=False, with_total_row=False)
    pipeline.run_import(db, p1, "rt1.xlsx")
    db.commit()
    n0 = db.scalar(select(func.count()).select_from(FProjectExpense)
                   .where(FProjectExpense.linked_sales_order_no == "XSDD-RT2"))
    assert n0 == 5

    data = maintenance_cost.contract_workbook_data(db, "XSDD-RT2")
    wb = _build_workbook("XSDD-RT2", data)
    p2 = tmp_path / "rt2.xlsx"
    wb.save(str(p2))

    b2 = pipeline.run_import(db, str(p2), "rt2.xlsx")
    db.commit()
    assert b2.rows_inserted == 0 and b2.rows_error == 0             # 零新增零错误
    assert db.scalar(select(func.count()).select_from(FProjectExpense)
                     .where(FProjectExpense.linked_sales_order_no == "XSDD-RT2")) == n0


# ---------- 审查修复回归（R2-1/2/3/4 + 空白表单）----------

def test_no_ffill_new_row_keeps_defaults(db, tmp_path):
    """员工在状态=流程中的既有行下面只填 日期+金额 续新行：新行不得继承上一行的
    流程状态/人员/事由（R2-1——否则新报销静默不计已花）。"""
    p = _workbook_xlsx(tmp_path, "ff1.xlsx",
                       [_canon_row(amount=100, status="流程中", reason="旧账"),
                        {"报销日期": "2026-05-02", "报销金额": 88}],
                       with_parts=False, with_total_row=False)
    batch = pipeline.run_import(db, str(p), "ff1.xlsx")
    db.commit()
    assert batch.rows_inserted == 2
    row = db.scalar(select(FProjectExpense).where(FProjectExpense.amount == Decimal("88")))
    assert row.data_status == "已结束"          # 默认生效，而不是继承「流程中」
    assert row.person is None and row.reason is None


def test_cross_contract_same_docno_no_collision(db, tmp_path):
    """两个合同的工作簿都手填 单号=1/序号=1：键带合同域，不互撞不互改（R2-2）。"""
    pa = _workbook_xlsx(tmp_path, "ca.xlsx", [_canon_row(amount=100, bxd="1", seq=1)],
                        anchor="XSDD-CA", with_parts=False, with_total_row=False)
    pb = _workbook_xlsx(tmp_path, "cb.xlsx", [_canon_row(amount=200, bxd="1", seq=1)],
                        anchor="XSDD-CB", with_parts=False, with_total_row=False)
    pipeline.run_import(db, pa, "ca.xlsx")
    pipeline.run_import(db, pb, "cb.xlsx", mode="upsert")   # B 合同修复模式也不得波及 A
    db.commit()
    rows = {r.linked_sales_order_no: r.amount for r in db.scalars(select(FProjectExpense)).all()}
    assert rows == {"XSDD-CA": Decimal("100"), "XSDD-CB": Decimal("200")}


def test_duplicate_docno_in_file_isolated(db, tmp_path):
    """同一文件内 单号+序号 重复：后行成 duplicate_key 错误行，不炸整批（R2-4）。"""
    p = _workbook_xlsx(tmp_path, "dup1.xlsx",
                       [_canon_row(amount=100, bxd="B1", seq=1),
                        _canon_row(amount=200, reason="重复", bxd="B1", seq=1)],
                       with_parts=False, with_total_row=False)
    batch = pipeline.run_import(db, str(p), "dup1.xlsx")
    db.commit()
    assert batch.rows_inserted == 1 and batch.rows_error == 1
    assert batch.report_json["errors_preview"][0]["error_type"] == "duplicate_key"


def test_upsert_requires_clean_sheet(db, tmp_path):
    """修复模式（以本表为准）遇错误行整批拒绝：半截行不许静默丢账（R2-3）。"""
    ok = _workbook_xlsx(tmp_path, "cl0.xlsx", [_canon_row(amount=100)],
                        with_parts=False, with_total_row=False)
    pipeline.run_import(db, ok, "cl0.xlsx")
    db.commit()
    bad = _workbook_xlsx(tmp_path, "cl1.xlsx",
                         [_canon_row(amount=150),
                          {"报销金额": 66, "支出事由": "有金额没日期"}],   # missing_date 错误行
                         with_parts=False, with_total_row=False)
    from app.etl.reader import ReaderError
    with pytest.raises(ReaderError, match="修复模式"):
        pipeline.run_import(db, str(bad), "cl1.xlsx", mode="upsert")
    db.commit()
    # 原数据未被动过（删除未发生）
    amts = db.scalars(select(FProjectExpense.amount)).all()
    assert amts == [Decimal("100")]


def test_amount_without_date_errors_but_total_row_skips():
    df = pd.DataFrame([
        {"报销金额": 66, "支出事由": "半截行"},
        {"报销金额": 999, "支出事由": "合计（仅已结束）"},
    ])
    res = transform(df, mapping.EXPENSE, anchor="XSDD-X")
    assert res.rows_skipped_no_data == 1
    assert len(res.errors) == 1 and res.errors[0].error_type == "missing_date"


def test_blank_form_export_reimports_clean(db, batch, tmp_path):
    """空白表单（新合同导出，报销页只有锚+表头）再导入：成功、零行、零错误。"""
    from app.api.maintenance import _build_workbook
    from app.services import maintenance_cost

    data = maintenance_cost.contract_workbook_data(db, "XSDD-BLANK")
    wb = _build_workbook("XSDD-BLANK", data)
    p = tmp_path / "blank.xlsx"
    wb.save(str(p))
    b = pipeline.run_import(db, str(p), "blank.xlsx")
    db.commit()
    assert b.status == "success" and b.rows_inserted == 0 and b.rows_error == 0
