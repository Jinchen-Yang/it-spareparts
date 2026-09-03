from __future__ import annotations

import hashlib
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from app.etl import loader, pipeline
from app.models.maintenance_project import (
    MaintenanceProjectAuditLog,
    MaintenanceProject,
    MaintenanceProjectAlias,
    MaintenanceProjectContract,
    MaintenanceProjectXsdd,
)
from app.models.sales import FSalesOrder
from app.services import maintenance_bulk_import as bulk
from app.services import maintenance_project_identity


class _ScalarsOnlyDb:
    def __init__(self, rows=()):
        self.rows = list(rows)

    def scalars(self, _statement):
        return list(self.rows)


class _ScalarSequenceDb:
    def __init__(self, *rows):
        self.rows = list(rows)

    def scalar(self, _statement):
        return self.rows.pop(0) if self.rows else None


def _sheet(field_indexes: dict[str, int], rows: list[tuple]) -> bulk.DetectedSheet:
    width = max(field_indexes.values()) + 1
    return bulk.DetectedSheet(
        name="Sheet1",
        header_row=2,
        header_rows=(1, 2),
        headers=tuple("" for _ in range(width)),
        system_headers=tuple("" for _ in range(width)),
        field_indexes=field_indexes,
        field_matches={},
        rows=tuple((row_no, tuple(values)) for row_no, values in rows),
    )


def _ordinary_sales_workbook(
    tmp_path,
    *,
    order_no: str,
    raw_order_id: str,
    project_name: str,
    period_from: date | None = date(2026, 1, 1),
    period_to: date | None = date(2026, 12, 31),
    maintenance_business: str = "是",
    business_type: str = "备件维保",
    data_status: str = "已生效",
    project_manager: str = "",
    order_amount: str = "113",
    tax_rate: str = "13%",
    tax_amount: str = "13",
    amount_ex_tax: str = "100",
) -> str:
    workbook = Workbook()
    sheet = workbook.active
    system_headers = [
        "SeqNo", "ObjectId", "Status", "F0000118", "F0000059",
        "F0000119", "F0000134", "F0000131", "F0000132", "F0000021", "F0000053",
        "F0000054", "F0000055", "F0000056", "D0001F0000001",
        "D0001F0000002", "D0001F0000003", "D0001F0000004",
    ]
    captions = [
        "订单编号(必填)", "数据ID(不可修改)", "数据状态", "维保业务", "业务类型#",
        "项目名称(必填)", "项目经理(必填)", "维保起始日期(必填)", "维保终止日期(必填)",
        "订单金额", "是否含税(必填)",
        "税率(必填)", "税金", "不含税金额", "订单明细.数据ID(不可修改)",
        "订单明细.产品名称", "订单明细.订单数量", "订单明细.单价",
    ]
    sheet.append(system_headers)
    sheet.append(captions)
    sheet.append([
        order_no, raw_order_id, data_status, maintenance_business, business_type,
        project_name, project_manager, period_from, period_to,
        order_amount, "含税", tax_rate,
        tax_amount, amount_ex_tax,
        f"{raw_order_id}-line", "PN-AUTO-1", "1", order_amount,
    ])
    path = tmp_path / f"{raw_order_id}-{abs(hash(project_name))}.xlsx"
    workbook.save(path)
    return str(path)


def test_sales_exact_machine_header_wins_duplicate_caption_at_row_20():
    workbook = Workbook()
    sheet = workbook.active
    systems = (
        "SeqNo",
        "F0000021",
        "F0000053",
        "Purchase.F0000099",
        "Sales.F0000054",
        "F0000055",
        "F0000056",
    )
    captions = (
        "订单编号(必填)",
        "订单金额",
        "是否含税(必填)",
        "税率(必填)",
        "税率(必填)",
        "税金",
        "不含税金额",
    )
    for column, value in enumerate(systems, start=1):
        sheet.cell(19, column, value)
    for column, value in enumerate(captions, start=1):
        sheet.cell(20, column, value)
    values = (
        "XSDD-20240708-0093",
        Decimal("31440000"),
        "含税",
        "6%",  # another document's identically captioned field
        "13%",
        Decimal("3616991.15"),
        Decimal("27823008.85"),
    )
    for column, value in enumerate(values, start=1):
        sheet.cell(21, column, value)
    stream = BytesIO()
    workbook.save(stream)

    adapter, detected = bulk._detect(stream.getvalue())

    assert adapter.key == "sales_contract_amount"
    assert detected.header_row == 20
    assert detected.field_indexes["tax_rate"] == 4
    inc, amount_ex, rate = adapter._amounts(detected, detected.rows[0][1])
    assert inc == Decimal("31440000.00")
    assert amount_ex == Decimal("27823008.85")
    assert rate == Decimal("0.130000")


def test_sales_amount_can_derive_missing_rate_from_ex_tax_and_tax():
    adapter = bulk.SalesContractAmountAdapter()
    detected = _sheet(
        {
            "order_amount": 0,
            "tax_flag": 1,
            "tax_rate": 2,
            "tax_amount": 3,
            "amount_ex_tax": 4,
        },
        [(3, ("31440000", "含税", "", "3616991.15", "27823008.85"))],
    )

    inc, amount_ex, rate = adapter._amounts(detected, detected.rows[0][1])

    assert inc == Decimal("31440000.00")
    assert amount_ex == Decimal("27823008.85")
    assert rate == Decimal("0.130000")


def test_sales_amount_preserves_numeric_zero_values():
    adapter = bulk.SalesContractAmountAdapter()
    detected = _sheet(
        {
            "order_amount": 0,
            "tax_flag": 1,
            "tax_rate": 2,
            "tax_amount": 3,
            "amount_ex_tax": 4,
        },
        [(
            3,
            (
                Decimal("0.00"),
                "含税",
                Decimal("0"),
                Decimal("0.00"),
                Decimal("0.00"),
            ),
        )],
    )

    inc, amount_ex, rate = adapter._amounts(detected, detected.rows[0][1])

    assert inc == Decimal("0.00")
    assert amount_ex == Decimal("0.00")
    assert rate == Decimal("0")


def test_sales_form_without_tax_rate_column_is_recognized_from_ex_and_tax():
    workbook = Workbook()
    sheet = workbook.active
    for column, value in enumerate(
        ("SeqNo", "F0000055", "F0000056"), start=1
    ):
        sheet.cell(1, column, value)
    for column, value in enumerate(
        ("订单编号(必填)", "税金", "不含税金额"), start=1
    ):
        sheet.cell(2, column, value)
    for column, value in enumerate(
        ("XSDD-1", "13", "100"), start=1
    ):
        sheet.cell(3, column, value)
    stream = BytesIO()
    workbook.save(stream)

    adapter, detected = bulk._detect(stream.getvalue())
    inc, amount_ex, rate = adapter._amounts(detected, detected.rows[0][1])

    assert adapter.key == "sales_contract_amount"
    assert inc == Decimal("113.00")
    assert amount_ex == Decimal("100.00")
    assert rate == Decimal("0.130000")


def test_sales_amount_never_treats_missing_rate_as_zero_for_lone_ex_tax():
    adapter = bulk.SalesContractAmountAdapter()
    detected = _sheet(
        {"amount_ex_tax": 0, "tax_rate": 1},
        [(3, ("100", ""))],
    )

    with pytest.raises(bulk.BulkImportInvalid, match="不能从单一未税金额"):
        adapter._amounts(detected, detected.rows[0][1])


def test_sales_conflicting_duplicate_blocks_entire_order(monkeypatch):
    contract = SimpleNamespace(
        project_id="project-1",
        project_contract_id="relation-1",
        amount_inc_tax=Decimal("10.00"),
        version=1,
    )
    monkeypatch.setattr(
        bulk,
        "_contract_maps",
        lambda _db: ({"20240101-0001": [contract]}, {"20240101-0001": contract}),
    )
    monkeypatch.setattr(bulk, "_all_contracts_by_order", lambda _db, _values: {})
    monkeypatch.setattr(bulk, "_assignment_evidence", lambda _db, _values: ({}, {}))
    detected = _sheet(
        {
            "order_no": 0,
            "order_amount": 1,
            "tax_flag": 2,
            "tax_rate": 3,
            "tax_amount": 4,
            "amount_ex_tax": 5,
        },
        [
            (3, ("XSDD-20240101-0001", "113", "含税", "13%", "13", "100")),
            (4, ("XSDD-20240101-0001", "226", "含税", "13%", "26", "200")),
        ],
    )

    plan = bulk.SalesContractAmountAdapter().build_plan(_ScalarsOnlyDb(), detected)

    assert [operation["action"] for operation in plan["operations"]] == ["blocked"]
    assert all(row["action"] == "error" for row in plan["rows"])
    assert any(
        issue["code"] == "order_level_fail_closed" for issue in plan["issues"]
    )


def test_receipt_risk_row_blocks_all_months_without_partial_cumulative(monkeypatch):
    contract = SimpleNamespace(
        project_id="project-1",
        project_contract_id="relation-1",
        contract_no="XSDD-20240101-0001",
        version=1,
    )
    monkeypatch.setattr(
        bulk,
        "_contract_maps",
        lambda _db: ({"20240101-0001": [contract]}, {"20240101-0001": contract}),
    )
    detected = _sheet(
        {"order_no": 0, "receipt_no": 1, "receipt_date": 2, "actual_amount": 3, "remark": 4},
        [
            (3, ("XSDD-20240101-0001", "SK-1", date(2026, 1, 10), "100", "")),
            (4, ("XSDD-20240101-0001", "SK-2", date(2026, 1, 20), "20", "坏账")),
            (5, ("XSDD-20240101-0001", "SK-3", date(2026, 2, 1), "50", "")),
        ],
    )

    plan = bulk.ReceiptCollectionAdapter().build_plan(_ScalarsOnlyDb(), detected)

    assert {operation["action"] for operation in plan["operations"]} == {"conflict"}
    assert {operation["report_month"] for operation in plan["operations"]} == {
        "2026-01-01",
        "2026-02-01",
    }
    assert all(
        operation["new_cumulative_amount"] is None
        for operation in plan["operations"]
    )
    assert not any(
        operation["action"] in {"create", "noop"}
        for operation in plan["operations"]
    )


def test_incomplete_receipt_history_blocks_every_month_for_contract(monkeypatch):
    contract = SimpleNamespace(
        project_id="project-1",
        project_contract_id="relation-1",
        contract_no="XSDD-20240101-0001",
        version=1,
    )
    existing = SimpleNamespace(
        project_contract_id="relation-1",
        report_month=date(2025, 12, 1),
        status="confirmed",
        cumulative_amount=Decimal("10.00"),
        collection_id="collection-1",
        version=1,
        receipt_reference="SK-OLD",
    )
    monkeypatch.setattr(
        bulk,
        "_contract_maps",
        lambda _db: ({"20240101-0001": [contract]}, {"20240101-0001": contract}),
    )
    detected = _sheet(
        {"order_no": 0, "receipt_no": 1, "receipt_date": 2, "actual_amount": 3},
        [
            (3, ("XSDD-20240101-0001", "SK-1", date(2026, 1, 10), "100")),
            (4, ("XSDD-20240101-0001", "SK-2", date(2026, 2, 10), "50")),
        ],
    )

    plan = bulk.ReceiptCollectionAdapter().build_plan(
        _ScalarsOnlyDb([existing]), detected
    )

    assert len(plan["operations"]) == 2
    assert all(operation["action"] == "conflict" for operation in plan["operations"])
    assert all(
        any(
            issue["code"] == "incomplete_receipt_history"
            for issue in operation["issues"]
        )
        for operation in plan["operations"]
    )


def _transfer_batch(*, status: str, selected: list[str]) -> tuple[str, str, str, object]:
    token = f"1.{('x' * 32)}"
    payload_hash = "a" * 64
    data_version = "version-1"
    selection_hash = bulk._canonical_hash(
        {"payload_hash": payload_hash, "row_keys": sorted(selected)}
    )
    report = {
        "token_hash": hashlib.sha256(token.encode("utf-8")).hexdigest(),
        "payload_hash": payload_hash,
        "data_version": data_version,
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        "selected_row_keys": selected,
        "selection_hash": selection_hash,
        "result": {
            "status": "done",
            "rows": [
                {
                    "row_key": key,
                    "action": "update_contract",
                    "project_id": "project-1",
                }
                for key in selected
            ],
        },
    }
    batch = SimpleNamespace(
        id=1,
        file_type=bulk.TRANSFER_BATCH_TYPE,
        uploaded_by="operator",
        status=status,
        report_json=report,
    )
    return token, payload_hash, data_version, batch


def test_success_replay_rejects_different_row_keys():
    token, payload_hash, data_version, batch = _transfer_batch(
        status="success", selected=["row-1"]
    )

    with pytest.raises(bulk.BulkImportConflict, match="row_keys"):
        bulk.apply_transfer(
            _ScalarSequenceDb(batch),
            preview_token=token,
            payload_hash=payload_hash,
            data_version=data_version,
            row_keys=["row-2"],
            operated_by="operator",
        )


def test_success_replay_accepts_same_row_key_set_in_any_order():
    token, payload_hash, data_version, batch = _transfer_batch(
        status="success", selected=["row-1", "row-2"]
    )

    result = bulk.apply_transfer(
        _ScalarSequenceDb(batch),
        preview_token=token,
        payload_hash=payload_hash,
        data_version=data_version,
        row_keys=["row-2", "row-1"],
        operated_by="operator",
    )

    assert result["status"] == "done"


def test_success_replay_rechecks_current_project_scope():
    token, payload_hash, data_version, batch = _transfer_batch(
        status="success", selected=["row-1"]
    )

    with pytest.raises(bulk.BulkImportScopeDenied):
        bulk.apply_transfer(
            _ScalarSequenceDb(batch),
            preview_token=token,
            payload_hash=payload_hash,
            data_version=data_version,
            row_keys=["row-1"],
            operated_by="operator",
            allowed_project_ids={"project-2"},
        )


def test_processing_apply_rejects_out_of_scope_target_before_adapter_write():
    token, payload_hash, data_version, batch = _transfer_batch(
        status="processing", selected=[]
    )
    batch.report_json.update(
        {
            "row_map": {"row-1": {"plan_index": 0, "operation_index": 0}},
            "public": {
                "rows": [{"row_key": "row-1", "row_status": "ready"}]
            },
            "plans": [
                {
                    "form_type": "sales_contract_amount",
                    "plan": {
                        "operations": [
                            {
                                "action": "update_contract",
                                "project_id": "project-2",
                            }
                        ]
                    },
                }
            ],
        }
    )

    with pytest.raises(bulk.BulkImportScopeDenied):
        bulk.apply_transfer(
            _ScalarSequenceDb(batch),
            preview_token=token,
            payload_hash=payload_hash,
            data_version=data_version,
            row_keys=["row-1"],
            operated_by="operator",
            allowed_project_ids={"project-1"},
        )


def test_preview_rejects_out_of_scope_operation_before_persist(monkeypatch):
    artifact = bulk.PreviewArtifact(
        adapter_key="sales_contract_amount",
        file_type="maint_contract",
        file_hash="b" * 64,
        filename="sales.xlsx",
        plan={
            "operations": [
                {"action": "update_contract", "project_id": "project-2"}
            ],
            "summary": {"source_rows": 1},
        },
    )
    monkeypatch.setattr(bulk, "build_preview", lambda *_args: artifact)

    with pytest.raises(bulk.BulkImportScopeDenied):
        bulk.preview_transfer(
            object(),
            [("sales.xlsx", b"xlsx")],
            operated_by="operator",
            allowed_project_ids={"project-1"},
        )


def test_ordinary_sales_upload_creates_one_xsdd_project_and_retains_peer_names(
    db, tmp_path
):
    order_no = "XSDD-20260902-0001"
    raw_order_id = "sales-auto-project-1"
    pre_delivery_name = "预交付-联通云服务器维保"
    formal_name = "联通云平台服务器维保集中项目"
    first_path = _ordinary_sales_workbook(
        tmp_path,
        order_no=order_no,
        raw_order_id=raw_order_id,
        project_name=pre_delivery_name,
    )

    first = pipeline.run_import(
        db,
        first_path,
        "sales-first.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    owner = db.get(MaintenanceProjectXsdd, "20260902-0001")
    assert owner is not None
    project = db.get(MaintenanceProject, owner.project_id)
    assert project is not None
    assert project.display_name == pre_delivery_name
    assert project.period_from == date(2026, 1, 1)
    assert project.period_to == date(2026, 12, 31)
    assert first.report_json["maintenance_sales_project_sync"]["status"] == "applied"
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id
    ))
    assert contract is not None
    assert contract.contract_no == order_no
    assert contract.amount_inc_tax == Decimal("113.00")

    second_path = _ordinary_sales_workbook(
        tmp_path,
        order_no=order_no,
        raw_order_id=raw_order_id,
        project_name=formal_name,
        period_from=None,
        period_to=date(2027, 12, 31),
    )
    second = pipeline.run_import(
        db,
        second_path,
        "sales-formal.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    assert second.report_json["maintenance_sales_project_sync"]["noop"] == 1
    assert project.period_from is None
    assert project.period_to == date(2027, 12, 31)
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 1
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectContract)) == 1
    aliases = set(db.scalars(select(MaintenanceProjectAlias.alias_name).where(
        MaintenanceProjectAlias.project_id == project.project_id
    )))
    assert {pre_delivery_name, formal_name} <= aliases
    peers = maintenance_project_identity.peer_names_by_project(
        db, [project.project_id]
    )
    assert {pre_delivery_name, formal_name} <= set(peers[project.project_id])

    project_version = project.version
    contract_version = contract.version
    alias_count = db.scalar(select(func.count()).select_from(MaintenanceProjectAlias))
    with open(second_path, "rb") as workbook_file:
        replay = bulk.sync_uploaded_sales_workbook(
            db,
            workbook_file.read(),
            "sales-formal-replay.xlsx",
            operated_by="sales-importer",
            import_batch_id=second.id,
        )
    assert replay["noop"] == 1
    assert project.version == project_version
    assert contract.version == contract_version
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectAlias)) == alias_count

    # Name equality alone never merges or blocks different XSDDs.  Only the
    # XSDD owner relation decides identity.
    other_xsdd_path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0003",
        raw_order_id="sales-auto-project-3",
        project_name=formal_name,
    )
    pipeline.run_import(
        db,
        other_xsdd_path,
        "sales-same-name-other-xsdd.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )
    other_owner = db.get(MaintenanceProjectXsdd, "20260902-0003")
    assert other_owner is not None and other_owner.project_id != project.project_id
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 2


def test_ordinary_sales_auto_project_failure_rolls_back_all_facts(db, tmp_path):
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0002",
        raw_order_id="sales-auto-project-invalid",
        project_name="缺期限的维保项目",
        period_from=None,
        period_to=None,
    )

    pipeline.run_import(
        db,
        path,
        "sales-missing-period.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )
    project = db.scalar(select(MaintenanceProject))
    assert project is not None
    assert project.period_from is None and project.period_to is None
    assert project.version == 1
    assert db.scalar(select(func.count()).select_from(FSalesOrder)) == 1
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectContract)) == 1
    assert db.get(MaintenanceProjectXsdd, "20260902-0002") is not None


def test_owner_backed_maintenance_sales_still_requires_project_name(db, tmp_path):
    project = MaintenanceProject(
        project_id="sales-owner-blank-name-project",
        project_code="XSDD-20260902-0021",
        display_name="既有 XSDD owner 项目",
        lifecycle_status="ongoing",
        is_active=True,
        version=1,
    )
    db.add(project)
    db.flush()
    db.add(MaintenanceProjectXsdd(
        xsdd_norm="20260902-0021",
        project_id=project.project_id,
        source="test-map-only-owner",
    ))
    db.commit()

    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0021",
        raw_order_id="sales-owner-blank-name",
        project_name="",
    )
    with pytest.raises(
        loader.ImportIntegrityError,
        match="维保销售订单自动建项失败",
    ):
        pipeline.run_import(
            db,
            path,
            "sales-owner-blank-name.xlsx",
            uploaded_by="sales-importer",
            mode="upsert",
            auto_assign_maintenance_projects=True,
        )
    db.rollback()

    assert db.scalar(select(func.count()).select_from(FSalesOrder)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectContract)) == 0


@pytest.mark.parametrize(
    ("period_from", "period_to"),
    [(date(2026, 1, 1), None), (None, date(2026, 12, 31))],
)
def test_ordinary_sales_auto_project_preserves_one_sided_period(
    db, tmp_path, period_from, period_to
):
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0017",
        raw_order_id="sales-one-sided-period",
        project_name="单侧期限维保项目",
        period_from=period_from,
        period_to=period_to,
    )
    pipeline.run_import(
        db,
        path,
        "sales-one-sided-period.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )
    project = db.scalar(select(MaintenanceProject))
    assert project is not None
    assert project.period_from == period_from
    assert project.period_to == period_to


def test_ordinary_sales_auto_project_never_reopens_the_xlsx(db, tmp_path, monkeypatch):
    """契约：自动建项只吃 loader 已解析的 TransformResult，绝不重开 XLSX。

    重开等于 openpyxl 以 read_only=False 实体化每个 worksheet（真实销售导出
    有 19 个 sheet），抵消 loader 的 load_selected_workbook 内存边界，还可能
    选到与已入库事实不同的那一张表。把服务侧的 load_workbook 换成炸弹即可
    把这条边界钉死。
    """

    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0033",
        raw_order_id="sales-no-reopen",
        project_name="不重开工作簿维保项目",
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
    )

    def _explode(*args, **kwargs):
        raise AssertionError("自动建项重开了 XLSX，内存边界被抵消")

    monkeypatch.setattr(bulk, "load_workbook", _explode)

    pipeline.run_import(
        db,
        path,
        "sales-no-reopen.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    project = db.scalar(select(MaintenanceProject))
    assert project is not None
    assert project.project_code == "XSDD-20260902-0033"


def test_ordinary_sales_inverted_period_rolls_back_all_facts(db, tmp_path):
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0018",
        raw_order_id="sales-inverted-period",
        project_name="倒置期限维保项目",
        period_from=date(2027, 1, 1),
        period_to=date(2026, 1, 1),
    )
    with pytest.raises(loader.ImportIntegrityError, match="自动建项失败"):
        pipeline.run_import(
            db,
            path,
            "sales-invalid.xlsx",
            uploaded_by="sales-importer",
            mode="upsert",
            auto_assign_maintenance_projects=True,
        )
    db.rollback()

    assert db.scalar(select(func.count()).select_from(FSalesOrder)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectContract)) == 0


def test_existing_sales_owner_preserves_period_when_source_period_is_inverted(
    db, tmp_path
):
    order_no = "XSDD-20260902-0019"
    raw_order_id = "sales-inverted-existing-owner"
    original = _ordinary_sales_workbook(
        tmp_path,
        order_no=order_no,
        raw_order_id=raw_order_id,
        project_name="已有期限维保项目",
        period_from=date(2026, 1, 1),
        period_to=date(2026, 12, 31),
    )
    pipeline.run_import(
        db,
        original,
        "sales-valid-period.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    inverted = _ordinary_sales_workbook(
        tmp_path,
        order_no=order_no,
        raw_order_id=raw_order_id,
        project_name="已有期限维保项目",
        period_from=date(2027, 1, 1),
        period_to=date(2026, 1, 1),
        order_amount="226",
        tax_amount="26",
        amount_ex_tax="200",
    )
    pipeline.run_import(
        db,
        inverted,
        "sales-inverted-existing-period.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    owner = db.get(MaintenanceProjectXsdd, "20260902-0019")
    project = db.get(MaintenanceProject, owner.project_id)
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id
    ))
    assert project.period_from == date(2026, 1, 1)
    assert project.period_to == date(2026, 12, 31)
    assert contract.amount_inc_tax == Decimal("226.00")


def test_ordinary_non_maintenance_sales_upload_never_creates_project(db, tmp_path):
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0004",
        raw_order_id="ordinary-hardware-sale",
        project_name="普通备件销售",
        maintenance_business="否",
        business_type="备件销售",
    )

    batch = pipeline.run_import(
        db,
        path,
        "ordinary-hardware-sale.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    assert batch.report_json["maintenance_sales_project_sync"]["status"] == "no_maintenance_rows"
    assert db.scalar(select(func.count()).select_from(FSalesOrder)) == 1
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectContract)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectXsdd)) == 0


def test_skip_mode_existing_sales_fact_does_not_trigger_project_overwrite(db, tmp_path):
    order_no = "XSDD-20260902-0005"
    raw_order_id = "sales-skip-existing"
    original = _ordinary_sales_workbook(
        tmp_path,
        order_no=order_no,
        raw_order_id=raw_order_id,
        project_name="普通销售原始事实",
        maintenance_business="否",
        business_type="备件销售",
    )
    pipeline.run_import(
        db,
        original,
        "sales-skip-original.xlsx",
        uploaded_by="sales-importer",
        mode="upsert",
        auto_assign_maintenance_projects=True,
    )

    maintenance_revision = _ordinary_sales_workbook(
        tmp_path,
        order_no=order_no,
        raw_order_id=raw_order_id,
        project_name="不应在 skip 模式生效的维保项目",
    )
    skipped = pipeline.run_import(
        db,
        maintenance_revision,
        "sales-skip-maintenance.xlsx",
        uploaded_by="sales-importer",
        mode="skip",
        auto_assign_maintenance_projects=True,
    )

    assert skipped.report_json["maintenance_sales_project_sync"]["status"] == "no_maintenance_rows"
    sales = db.scalar(select(FSalesOrder).where(FSalesOrder.raw_order_id == raw_order_id))
    assert sales is not None and sales.business_type == "备件销售"
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 0
    assert db.scalar(select(func.count()).select_from(MaintenanceProjectContract)) == 0


def test_ordinary_sales_auto_project_takes_first_of_multiple_managers(db, tmp_path):
    """销售订单「项目经理(必填)」是多值，建项负责人取首位、原值留审计。

    2026-09-03 负责人拍板。真实导出实测形如「廖晓娟;司珂梓」，且不同订单行
    人名顺序不一致——所以必须逐行取各自的首位，不能全批写同一个人。
    项目表没有备注列，完整原值写进建项审计 reason，不为此加迁移。
    """
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0031",
        raw_order_id="sales-multi-manager",
        project_name="多负责人维保项目",
        project_manager="廖晓娟;司珂梓",
    )
    pipeline.run_import(
        db, path, "sales-multi-manager.xlsx", uploaded_by="sales-importer",
        mode="upsert", auto_assign_maintenance_projects=True,
    )
    project = db.scalar(select(MaintenanceProject))
    assert project is not None
    assert project.project_code == "XSDD-20260902-0031"
    assert project.project_manager_id == "廖晓娟"
    audit = db.scalars(
        select(MaintenanceProjectAuditLog).where(
            MaintenanceProjectAuditLog.project_id == project.project_id)
    ).all()
    assert any("廖晓娟、司珂梓" in (entry.reason or "") for entry in audit), [
        entry.reason for entry in audit
    ]


def test_ordinary_sales_auto_project_ignores_non_maintenance_rows(db, tmp_path):
    """只有明确「维保业务=是」且已生效的行才建项（2026-09-03 拍板口径）。"""
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0032",
        raw_order_id="sales-not-maintenance",
        project_name="非维保销售订单",
        maintenance_business="否",
        business_type="设备销售",
    )
    pipeline.run_import(
        db, path, "sales-not-maintenance.xlsx", uploaded_by="sales-importer",
        mode="upsert", auto_assign_maintenance_projects=True,
    )
    assert db.scalar(select(func.count()).select_from(FSalesOrder)) == 1
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 0


def test_explicit_no_maintenance_flag_wins_over_business_type(db, tmp_path):
    """「维保业务=否」但业务类型里带「维修」的单次维修，不得被自动建项。

    这是收窄门槛前的或逻辑（维保业务=是 或 业务类型含维保/运维/维修）判错的
    那一类：业务类型是分类，维保业务才是这一行是不是维保业务的权威事实，
    源头上人已经明确说「否」就不能反过来靠分类词把项目建出来。
    """
    path = _ordinary_sales_workbook(
        tmp_path,
        order_no="XSDD-20260902-0033",
        raw_order_id="sales-single-repair",
        project_name="单次维修不建项",
        maintenance_business="否",
        business_type="单次维修",
    )
    pipeline.run_import(
        db, path, "sales-single-repair.xlsx", uploaded_by="sales-importer",
        mode="upsert", auto_assign_maintenance_projects=True,
    )
    assert db.scalar(select(func.count()).select_from(FSalesOrder)) == 1
    assert db.scalar(select(func.count()).select_from(MaintenanceProject)) == 0
