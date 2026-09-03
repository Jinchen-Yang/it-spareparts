"""validate→apply 窗口的并发契约。

背景（2026-09-03 实测）：POST .../master-workbook/apply 在同一请求、同一 Session
里先跑 validate_project_master_v2（全程零加锁读服务端值做三路合并），再跑
apply_project_master_v2（第一把锁是 advisory）。两者之间任何一次并发提交都不会
被 plan 感知，apply 会按已过期的判定把用户值写下去——实测：并发方写的 qty=9 被
静默改成 3，plan.conflicts 空、回执 overridden 空、revision_drift 为 False，
三处留痕全沉默。

修法是行指纹哨兵：validate 给**用户真正触碰过的行**记下当时的服务端行指纹，
apply 在拿到行锁之后、任何写之前重算比对，不等就整本 fail-closed。
只给触碰行发指纹是刻意的——未触碰行本来就不写，给它们发指纹会让「项目里任何
一行被人改过就整本 409」，那正是 2.7.0 明确废掉的整本硬拒。

另有一层加固（expire_all + populate_existing）：apply 取锁后清 identity map，
让锁内重读拿到真实行。SQLAlchemy 的 identity map 持弱引用，真实请求里 validate
的局部变量出栈后对象通常已被回收，所以这一层多数时候是「本来就没事」；但只要有
任何东西（关系缓存、引用环、未来的重构）让对象存活下来，锁内的 version CAS 就会
拿陈旧值跟陈旧值比、永远判「没变」。这里把它钉成不变式，而不是依赖回收时机。
"""
import io
import uuid
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import select, text

from app.db import SessionLocal
from app.models.maintenance_project import MaintenanceProjectContract
from app.services import maintenance_project_master_workbook as master

from tests.test_maintenance_project_master_v2_editable import (
    _make_project_with_line,
    _save,
)


def _peer_commit(fn):
    """模拟另一个请求：独立 Session，提交后立即关闭。"""
    other = SessionLocal()
    try:
        fn(other)
        other.commit()
    finally:
        other.close()


def _overview_amount_cell(ws):
    return next(
        row[1] for row in ws.iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )


def test_apply_sees_concurrent_contract_bump_and_refuses_to_overwrite(db):
    """01 合同额：窗口内并发改额并 bump version → apply 必须报 stale_contract。

    修复前：apply 侧 FOR UPDATE 重读命中 identity map 的旧对象，
    svc 的 `contract.version != change.base_version` 恒为假，12345.67 直接盖掉
    并发方的 99999.00，且 version 从 1 算成 2 写回，把并发方已写的 2 抹平。
    """
    project, *_ = _make_project_with_line(db)
    db.commit()

    content = master.build_project_master_v2(
        db, project_id=project.project_id,
        sheets=(master.V2_SHEET_OVERVIEW,))
    wb = load_workbook(io.BytesIO(content))
    _overview_amount_cell(wb[master.V2_SHEET_OVERVIEW]).value = Decimal("12345.67")
    data = _save(wb)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=data)
    assert plan.contract_amount_change is not None

    contract_id = db.scalar(select(MaintenanceProjectContract.project_contract_id)
                            .where(MaintenanceProjectContract.project_id
                                   == project.project_id))

    def _bump(other):
        peer = other.get(MaintenanceProjectContract, contract_id)
        peer.amount_inc_tax = Decimal("99999.00")
        peer.version = peer.version + 1

    _peer_commit(_bump)

    with pytest.raises(master.WorkbookError) as raised:
        master.apply_project_master_v2(
            db, plan, operated_by="racer", import_batch_id=str(uuid.uuid4()))
    assert raised.value.code == "stale_contract"
    db.rollback()

    # 并发方的值必须原样保留，零写入。
    row = db.execute(text(
        "select amount_inc_tax, version from maintenance_project_contract "
        "where project_contract_id = :i"), {"i": contract_id}).first()
    assert row[0] == Decimal("99999.00")
    assert row[1] == 2


def test_validate_leaves_no_pending_writes_so_expire_all_discards_nothing(db):
    """expire_all 的安全前提：validate 必须只读。

    Session.expire_all() 会丢弃未 flush 的改动。validate 一旦开始写，
    这个修法就会静默吞掉改动——这条用例把「validate 只读」钉成契约。
    """
    project, *_ = _make_project_with_line(db)
    db.commit()

    content = master.build_project_master_v2(db, project_id=project.project_id)
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["需求数量"], 3)

    master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))

    assert not db.new, f"validate 新增了对象：{db.new}"
    assert not db.deleted, f"validate 删除了对象：{db.deleted}"
    assert not db.dirty, f"validate 修改了对象：{db.dirty}"


# ---------- validate→apply 窗口：行指纹哨兵 ----------

def _parts_workbook_with_qty(db, project_id, *, row, qty):
    content = master.build_project_master_v2(
        db, project_id=project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(row, headers["需求数量"], qty)
    return _save(wb)


def test_concurrent_commit_between_validate_and_apply_is_refused(db):
    """窗口内并发改同一行 → 整本零写入报 row_conflicts，并发方的值原样保留。

    这是 2026-09-03 实测复现的那条时序：修复前 apply 会把用户值 3 静默写入，
    plan.conflicts 空、overridden 空、revision_drift False，三处留痕全沉默。
    """
    from app.models.maintenance import FMaintenanceLine

    project, _part, _order, line = _make_project_with_line(db)
    line_id = line.id
    db.commit()

    data = _parts_workbook_with_qty(db, project.project_id, row=2, qty=3)
    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=data)
    assert not plan.conflicts
    assert plan.row_guards, "触碰行必须发出指纹"

    _peer_commit(lambda o: setattr(
        o.get(FMaintenanceLine, line_id), "qty", Decimal("9")))

    with pytest.raises(master.WorkbookError) as raised:
        master.apply_project_master_v2(
            db, plan, operated_by="racer", import_batch_id=str(uuid.uuid4()))
    assert raised.value.code == "row_conflicts"
    assert raised.value.issues
    db.rollback()

    assert db.scalar(select(FMaintenanceLine.qty).where(
        FMaintenanceLine.id == line_id)) == Decimal("9.000")


def test_concurrent_change_to_an_untouched_row_does_not_block_the_upload(db):
    """负对照：并发改的是用户**没碰**的那一行 → 必须照常写入，不得整本 409。

    2.7.0 明确废除了「整本 revision 硬拒」：未触碰行服务端已变也不回写、自动
    rebase。哨兵只给触碰行发指纹，正是为了不把那条语义倒退回去。
    """
    from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder

    project, part, order, line = _make_project_with_line(db)
    touched_id = line.id
    untouched = FMaintenanceLine(
        raw_line_id=f"raw-line-{uuid.uuid4()}", order_id=order.id, line_no=2,
        part_id=part.id, pn_std=part.pn_std, pn_raw=part.pn_std,
        description=part.description, qty=Decimal("5"),
        return_qty=Decimal("0"), cost_source="direct", cost_tax_basis="ex",
        confidence="high", import_batch_id=line.import_batch_id,
    )
    db.add(untouched)
    db.commit()
    untouched_id = untouched.id

    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    id_col = headers.get("维保明细ID") or headers.get("明细ID")
    target_row = next(
        r for r in range(2, ws.max_row + 1)
        if str(ws.cell(r, id_col).value or "") == str(touched_id)
    ) if id_col else 2
    ws.cell(target_row, headers["需求数量"], 3)

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    assert {g.entity_id for g in plan.row_guards} == {str(touched_id)}, (
        "未触碰行不得进指纹集合，否则整本硬拒会复活")

    _peer_commit(lambda o: setattr(
        o.get(FMaintenanceLine, untouched_id), "qty", Decimal("42")))

    master.apply_project_master_v2(
        db, plan, operated_by="racer", import_batch_id=str(uuid.uuid4()))
    db.commit()

    assert db.scalar(select(FMaintenanceLine.qty).where(
        FMaintenanceLine.id == touched_id)) == Decimal("3.000")
    assert db.scalar(select(FMaintenanceLine.qty).where(
        FMaintenanceLine.id == untouched_id)) == Decimal("42.000")


# ---------- 写阶段重读不得冲掉未 flush 的改动 ----------

def test_same_row_quantity_and_manual_cost_both_land(db):
    """同一行既改数量、又填人工成本 → 两者都必须落库。

    SessionLocal 是 autoflush=False。写阶段若对同一行做带 populate_existing
    的重读，会把已经赋值但尚未 flush 的 qty/PN/备注/重算金额直接冲回库里的旧
    值，而人工成本照常提交——半截应用，审计还记的是被冲回去的值
    （Codex P1，2026-09-04）。规则：populate_existing 只出现在取锁阶段。
    """
    from app.models.maintenance import FMaintenanceLine

    project, _part, _order, line = _make_project_with_line(
        db, unit_cost=None, cost_source="none")
    line_id = line.id
    db.commit()

    content = master.build_project_master_v2(
        db, project_id=project.project_id, sheets=(master.V2_SHEET_PARTS,))
    wb = load_workbook(io.BytesIO(content))
    ws = wb[master.V2_SHEET_PARTS]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["需求数量"], 7)
    ws.cell(2, headers["人工未税单位成本"], Decimal("88.00"))
    ws.cell(2, headers["人工成本原因"], "本次人工定价")

    plan = master.validate_project_master_v2(
        db, project_id=project.project_id, data=_save(wb))
    master.apply_project_master_v2(
        db, plan, operated_by="tester", import_batch_id=str(uuid.uuid4()))
    db.commit()

    row = db.execute(
        select(FMaintenanceLine.qty, FMaintenanceLine.unit_cost_ex_tax,
               FMaintenanceLine.cost_source)
        .where(FMaintenanceLine.id == line_id)
    ).first()
    assert row[0] == Decimal("7.000"), f"数量改动被写阶段重读冲掉了：{row[0]}"
    assert row[1] == Decimal("88.00")
    assert row[2] == "manual"
