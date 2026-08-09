"""项目维保四表工作簿 v2 的纯服务契约测试。"""

from __future__ import annotations

import hashlib
import io
import zipfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table

from app.services import maintenance_project_workbook_v2 as workbook_v2


HMAC_KEY = b"fixed-test-only-maintenance-workbook-v2-key"
BASE_V2_FIXTURE = (
    Path(__file__).with_name("fixtures")
    / "maintenance_project_workbook_v2_base_7bfe7fc.xlsx"
)
BASE_V2_FIXTURE_SHA256 = (
    "846b31f5dcec7314cd91797da3a32111ee1ab976f46c00cc804513a4469157a9"
)


def _append_collection(
    content: bytes,
    *,
    amount: int = 40000,
    resave_marker: str | None = None,
) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        sheet = book["01_总览"]
        table = sheet.tables["tbl_collections_v2"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if not sheet.cell(row, headers.index("__entity_id") + 1).value
        )
        values = {
            "操作": "CREATE",
            "项目合同关系ID": "project-contract-001",
            "合同编号": "XSDD-20260001",
            "报告月份": "2026-08",
            "累计回款金额（含税）": amount,
            "回款凭证号": "HK-NEW",
            "状态": "已确认",
            "备注": "二期款",
        }
        for header, value in values.items():
            sheet.cell(target, headers.index(header) + 1, value)
        if resave_marker:
            book.properties.subject = resave_marker
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _edit_workbook(content: bytes, edit) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        edit(book)
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _add_zip_member(content: bytes, name: str, payload: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content), "r") as source:
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                target.writestr(item, source.read(item.filename))
            target.writestr(name, payload)
    return output.getvalue()


def _replace_zip_bytes(content: bytes, old: bytes, new: bytes) -> bytes:
    output = io.BytesIO()
    replaced = False
    with zipfile.ZipFile(io.BytesIO(content), "r") as source:
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                payload = source.read(item.filename)
                if old in payload:
                    payload = payload.replace(old, new)
                    replaced = True
                target.writestr(item, payload)
    assert replaced
    return output.getvalue()


class _FakeRepository:
    def __init__(self, *, revision: int = 7):
        self.revision = revision
        self.files: set[str] = set()
        self.operations: dict[str, str] = {}
        self.atomic_calls: list[tuple] = []

    def current_revision(self, project_id: str) -> int:
        assert project_id == "project-001"
        return self.revision

    def applied_file(self, file_sha256: str) -> bool:
        return file_sha256 in self.files

    def applied_operation(self, operation_key: str) -> str | None:
        return self.operations.get(operation_key)

    def apply_collections_atomically(self, validation, creates) -> None:
        self.atomic_calls.append(tuple(creates))
        for create in creates:
            self.operations[create.operation_key] = create.payload_hash
        self.files.add(validation.file_sha256)


class _FakeEndpointAdapter(_FakeRepository):
    def __init__(self, workspace):
        super().__init__(revision=workspace["workbook_revision"])
        self.workspace = workspace
        self.validations = {}
        self.errors = {}

    def load_workspace(self, project_id):
        assert project_id == "project-001"
        return self.workspace

    def save_validation(self, validation):
        self.validations[validation.validation_id] = validation

    def load_validation(self, validation_id):
        return self.validations.get(validation_id)

    def save_validation_error(self, validation_id, project_id, issues, error_workbook):
        self.errors[validation_id] = error_workbook

    def load_validation_error(self, validation_id):
        return self.errors.get(validation_id)


def _workspace(*, revision: int = 7) -> dict:
    return {
        "project": {
            "project_id": "project-001",
            "project_code": "XSDD-20260001",
            "project_name": "医院维保项目",
            "manager_name": "张经理",
            "version": 3,
        },
        "workbook_revision": revision,
        "as_of": date(2026, 8, 8),
        "data_version": "maintenance-snapshot-42",
        "contracts": [
            {
                "project_contract_id": "project-contract-001",
                "contract_no": "XSDD-20260001",
                "contract_amount": Decimal("100000.00"),
                "version": 2,
            }
        ],
        "collections": [
            {
                "collection_id": "collection-001",
                "project_contract_id": "project-contract-001",
                "contract_no": "XSDD-20260001",
                "report_month": "2026-07",
                "cumulative_amount": Decimal("30000.00"),
                "voucher_no": "HK-001",
                "status": "已确认",
                "remark": "一期款",
                "version": 1,
            }
        ],
        "consumptions": [
            {
                "consumption_id": "issue-line-001",
                "issue_no": "CKD-001",
                "issue_date": date(2026, 7, 22),
                "part_no": "PN-001",
                "part_name": "备件一",
                "quantity": Decimal("2"),
                "unit_cost": None,
                "cost_amount": None,
                "unit_cost_ex_tax": None,
                "unit_cost_inc_tax": None,
                "cost_amount_ex_tax": None,
                "cost_amount_inc_tax": None,
                "cost_status": "缺少价格成本",
                "cost_source": "none",
            }
        ],
        "expenses": [
            {
                "expense_id": "expense-001",
                "expense_no": "BX-001",
                "expense_date": date(2026, 7, 25),
                "applicant": "李四",
                "category": "差旅费",
                "amount": Decimal("1130.00"),
                "amount_ex_tax": Decimal("1000.00"),
                "amount_inc_tax": Decimal("1130.00"),
                "approval_status": "已审批",
                "remark": "现场支持",
            }
        ],
        "tasks": [
            {
                "task_id": "task-001",
                "task_type": "项目经理追踪",
                "title": "确认下月巡检",
                "due_date": date(2026, 8, 20),
                "status": "待处理",
                "owner": "张经理",
                "detail": "联系客户确认时间",
            }
        ],
    }


def test_export_is_four_visible_sheets_and_keeps_missing_cost_as_blank():
    workspace = _workspace()
    workspace["consumptions"].append(
        {
            "consumption_id": "issue-line-sales-estimate",
            "issue_no": "CKD-002",
            "issue_date": date(2026, 7, 23),
            "part_no": "PN-SALES-ESTIMATE",
            "part_name": "销售回退估算备件",
            "quantity": Decimal("2"),
            "unit_cost": Decimal("100.00"),
            "cost_amount": Decimal("200.00"),
            "unit_cost_ex_tax": Decimal("100.00"),
            "unit_cost_inc_tax": Decimal("113.00"),
            "cost_amount_ex_tax": Decimal("200.00"),
            "cost_amount_inc_tax": Decimal("226.00"),
            "cost_status": "available",
            "cost_source": "sales_window",
            "cost_evidence_kind": "sales_estimate",
            "cost_is_estimate": True,
            "cost_source_label": "估算（销售前后 7 天数量加权）",
        }
    )
    workspace["tasks"].append(
        {
            "task_id": "task-sales-fallback-estimate",
            "task_type": "cost",
            "title": "核对项目成本中的销售回退估算",
            "due_date": None,
            "status": "open",
            "owner": "张经理",
            "detail": (
                "1 条已确认现场领用按销售前后 7 天数量加权估算；"
                "已计入成本进度，但不等于采购或人工确认单价"
            ),
        }
    )
    artifact = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )

    assert artifact.preview == {
        "contracts": 1,
        "collections": 1,
        "consumptions": 2,
        "expenses": 1,
        "tasks": 2,
        "missing_cost_rows": 1,
    }
    book = load_workbook(io.BytesIO(artifact.content), data_only=False)
    try:
        visible = [sheet.title for sheet in book.worksheets if sheet.sheet_state == "visible"]
        assert visible == list(workbook_v2.VISIBLE_SHEETS)
        assert set(book["01_总览"].tables) == {
            "tbl_project_contracts_v2",
            "tbl_collections_v2",
        }
        assert set(book["02_备件消耗"].tables) == {"tbl_consumptions_v2"}
        assert set(book["03_报销单"].tables) == {"tbl_expenses_v2"}
        assert set(book["04_项目经理追踪与提醒"].tables) == {"tbl_tasks_v2"}
        overview = book["01_总览"]
        collection_table = overview.tables["tbl_collections_v2"]
        _, collection_header, _, collection_end = range_boundaries(collection_table.ref)
        assert overview.protection.sheet is True
        assert overview.cell(collection_header + 1, 1).protection.locked is True
        assert overview.cell(collection_end, 1).protection.locked is False
        assert book["02_备件消耗"].protection.sheet is True

        contract_sheet = book["01_总览"]
        contract_table = contract_sheet.tables["tbl_project_contracts_v2"]
        contract_min_col, contract_min_row, contract_max_col, _ = range_boundaries(
            contract_table.ref
        )
        assert "合同额（含税，全部合同）" in [
            contract_sheet.cell(contract_min_row, col).value
            for col in range(contract_min_col, contract_max_col + 1)
        ]

        expense_sheet = book["03_报销单"]
        expense_headers = [cell.value for cell in expense_sheet[1]]
        assert "已审批金额（未税）" in expense_headers
        assert "已审批金额（含税）" in expense_headers
        assert expense_sheet.cell(
            2, expense_headers.index("已审批金额（未税）") + 1
        ).value == 1000
        assert expense_sheet.cell(
            2, expense_headers.index("已审批金额（含税）") + 1
        ).value == 1130

        sheet = book["02_备件消耗"]
        table = sheet.tables["tbl_consumptions_v2"]
        # Table starts at A1; verify unknown cost stays unknown, never becomes 0.
        header_values = [
            sheet.cell(1, col).value
            for col in range(1, len(table.tableColumns) + 1)
        ]
        unit_cost_col = header_values.index("未税单位成本") + 1
        amount_ex_col = header_values.index("实际消耗成本（未税）") + 1
        amount_inc_col = header_values.index("实际消耗成本（含税）") + 1
        assert sheet.cell(2, unit_cost_col).value is None
        assert sheet.cell(2, header_values.index("含税单位成本") + 1).value is None
        assert sheet.cell(2, amount_ex_col).value is None
        assert sheet.cell(2, amount_inc_col).value is None
        assert sheet.cell(2, header_values.index("成本完整性") + 1).value == "缺少价格成本"
        assert sheet.cell(3, header_values.index("成本完整性") + 1).value == "已计入（销售回退估算）"
        assert (
            sheet.cell(3, header_values.index("成本来源") + 1).value
            == "估算（销售前后 7 天数量加权）"
        )
        summary_labels = [
            sheet_value
            for sheet_value in (
                book["01_总览"].cell(row, 1).value
                for row in range(1, 24)
            )
            if sheet_value
        ]
        assert "现场领用已计成本（未税，含销售回退估算）" in summary_labels
        assert "项目已计成本（含税，含销售回退估算）" in summary_labels
        task_sheet = book["04_项目经理追踪与提醒"]
        task_headers = [cell.value for cell in task_sheet[1]]
        assert task_sheet.cell(3, task_headers.index("标题") + 1).value == (
            "核对项目成本中的销售回退估算"
        )
        assert "不等于采购或人工确认单价" in task_sheet.cell(
            3,
            task_headers.index("详细说明") + 1,
        ).value
    finally:
        book.close()


def test_summary_uses_latest_monthly_cumulative_snapshot_not_month_sum():
    workspace = _workspace()
    workspace["collections"].append({
        "collection_id": "collection-002",
        "project_contract_id": "project-contract-001",
        "contract_no": "XSDD-20260001",
        "report_month": "2026-08",
        "cumulative_amount": Decimal("40000.00"),
        "voucher_no": "HK-002",
        "status": "已确认",
        "remark": "八月累计",
        "version": 1,
    })

    summary = workbook_v2.compute_project_summary(workspace)

    assert summary["total_contract_amount"] == Decimal("100000.00")
    assert summary["confirmed_cumulative_collection_amount"] == Decimal("40000.00")
    assert summary["collection_rate"] == Decimal("0.4")
    assert summary["known_consumption_cost"] == Decimal("0")
    assert summary["approved_expense_ex_tax"] == Decimal("1000.00")
    assert summary["approved_expense_inc_tax"] == Decimal("1130.00")
    assert summary["approved_expense"] == Decimal("1130.00")
    assert summary["known_project_cost"] == Decimal("1130.00")
    assert summary["missing_cost_rows"] == 1
    assert summary["cost_alert"] == "incomplete"


def test_summary_excludes_confirmed_collection_from_historical_or_not_included_contracts():
    workspace = _workspace()
    workspace["contracts"][0].update(
        {"included_in_total": True, "is_effective": True}
    )
    workspace["contracts"].extend(
        [
            {
                "project_contract_id": "project-contract-historical",
                "contract_no": "XSDD-HISTORICAL",
                "contract_amount": Decimal("200000.00"),
                "included_in_total": True,
                "is_effective": False,
                "version": 1,
            },
            {
                "project_contract_id": "project-contract-excluded",
                "contract_no": "XSDD-EXCLUDED",
                "contract_amount": Decimal("300000.00"),
                "included_in_total": False,
                "is_effective": False,
                "version": 1,
            },
        ]
    )
    workspace["collections"].extend(
        [
            {
                "collection_id": "collection-historical",
                "project_contract_id": "project-contract-historical",
                "contract_no": "XSDD-HISTORICAL",
                "report_month": "2026-08",
                "cumulative_amount": Decimal("180000.00"),
                "status": "已确认",
                "version": 1,
            },
            {
                "collection_id": "collection-excluded",
                "project_contract_id": "project-contract-excluded",
                "contract_no": "XSDD-EXCLUDED",
                "report_month": "2026-08",
                "cumulative_amount": Decimal("250000.00"),
                "status": "已确认",
                "version": 1,
            },
        ]
    )

    summary = workbook_v2.compute_project_summary(workspace)

    assert summary["total_contract_amount"] == Decimal("100000.00")
    assert summary["confirmed_cumulative_collection_amount"] == Decimal("30000.00")
    assert summary["collection_rate"] == Decimal("0.3")


@pytest.mark.parametrize(
    ("known_cost", "cost_complete", "canonical_status", "expected"),
    [
        ("79000.00", False, "unknown", "incomplete"),
        ("80000.00", False, "yellow", "incomplete"),
        ("100010.00", False, "red", "red"),
        ("79000.00", True, "normal", "green"),
    ],
)
def test_summary_reuses_canonical_metrics_and_fails_closed_below_safe_thresholds(
    known_cost,
    cost_complete,
    canonical_status,
    expected,
):
    workspace = _workspace()
    workspace["canonical_metrics"] = {
        "contract_amount_basis": "inc_tax",
        "cost_progress_basis": "inc_tax",
        "total_contract_amount": "100000.00",
        "known_contract_amount": "100000.00",
        "contract_amount_complete": True,
        "received_amount": "30000.00",
        "collection_progress_pct": "30.00",
        "site_requisition_known_cost": known_cost,
        "approved_expense": "0.00",
        "actual_project_cost_known": known_cost,
        "cost_rate_lower_bound_pct": str(
            (Decimal(known_cost) / Decimal("100000.00") * Decimal("100"))
        ),
        "cost_status": canonical_status,
        "cost_complete": cost_complete,
        "missing_cost_lines": 1 if not cost_complete else 0,
    }
    workspace["canonical_completeness"] = {
        "status": "incomplete" if not cost_complete else "complete",
        "issues": ([{"code": "expense_data_not_ready"}] if not cost_complete else []),
    }
    # Deliberately contradictory workbook rows prove the summary does not
    # reimplement a second KPI definition beside project_workspace.
    workspace["collections"][0]["cumulative_amount"] = Decimal("99999.00")
    workspace["consumptions"][0]["cost_amount"] = Decimal("1.00")

    summary = workbook_v2.compute_project_summary(workspace)

    assert summary["confirmed_cumulative_collection_amount"] == Decimal("30000.00")
    assert summary["collection_rate"] == Decimal("0.3")
    assert summary["known_project_cost"] == Decimal(known_cost)
    assert summary["cost_alert"] == expected


def test_summary_closes_both_ratios_when_canonical_contract_scope_is_incomplete():
    workspace = _workspace()
    workspace["canonical_metrics"] = {
        "contract_amount_basis": "inc_tax",
        "cost_progress_basis": "inc_tax",
        "total_contract_amount": None,
        "known_contract_amount": "100000.00",
        "contract_amount_complete": False,
        "received_amount": "30000.00",
        "collection_progress_pct": None,
        "site_requisition_known_cost": "90000.00",
        "approved_expense": "0.00",
        "actual_project_cost_known": "90000.00",
        "cost_rate_lower_bound_pct": None,
        "cost_status": "unknown",
        "cost_complete": False,
        "missing_cost_lines": 0,
    }
    workspace["canonical_completeness"] = {
        "status": "incomplete",
        "issues": [{"code": "missing_contract_amount"}],
    }

    summary = workbook_v2.compute_project_summary(workspace)

    assert summary["total_contract_amount"] is None
    assert summary["collection_rate"] is None
    assert summary["cost_rate_lower_bound"] is None
    assert summary["cost_alert"] == "no_contract_amount"


@pytest.mark.parametrize(
    ("contract_basis", "cost_basis"),
    [(None, "inc_tax"), ("ex_tax", "inc_tax"), ("inc_tax", None), ("inc_tax", "ex_tax")],
)
def test_summary_fails_closed_when_canonical_tax_basis_is_missing_or_invalid(
    contract_basis,
    cost_basis,
):
    workspace = _workspace()
    workspace["canonical_metrics"] = {
        "contract_amount_basis": contract_basis,
        "cost_progress_basis": cost_basis,
        "total_contract_amount": "100000.00",
        "known_contract_amount": "100000.00",
        "contract_amount_complete": True,
        "received_amount": "30000.00",
        "collection_progress_pct": "30.00",
        "site_requisition_known_cost_ex_tax": "70000.00",
        "site_requisition_known_cost_inc_tax": "80000.00",
        "approved_expense_ex_tax": "0.00",
        "approved_expense_inc_tax": "0.00",
        "actual_project_cost_known_ex_tax": "70000.00",
        "actual_project_cost_known_inc_tax": "80000.00",
        "cost_rate_lower_bound_pct": "80.00",
        "cost_status": "yellow",
        "cost_complete": True,
        "missing_cost_lines": 0,
    }
    workspace["canonical_completeness"] = {"status": "complete", "issues": []}

    summary = workbook_v2.compute_project_summary(workspace)

    if contract_basis == "inc_tax":
        assert summary["collection_rate"] == Decimal("0.3")
    else:
        assert summary["collection_rate"] is None
    assert summary["cost_rate_lower_bound"] is None
    assert summary["cost_alert"] == "basis_unknown"
    assert summary["completeness_status"] == "incomplete"
    assert "basis_unknown" in summary["completeness_issues"]


@pytest.mark.parametrize(
    ("cost", "expected"),
    [
        ("79990.00", "green"),
        ("80000.00", "green"),
        ("80004.00", "green"),
        ("80005.00", "yellow"),
        ("80010.00", "yellow"),
        ("100000.00", "yellow"),
        ("100004.00", "yellow"),
        ("100005.00", "red"),
        ("100010.00", "red"),
    ],
)
def test_fallback_cost_alert_uses_displayed_strict_thresholds(cost, expected):
    workspace = _workspace()
    workspace["expenses"] = []
    workspace["consumptions"][0]["unit_cost"] = Decimal(cost)
    workspace["consumptions"][0]["cost_amount"] = Decimal(cost)
    workspace["consumptions"][0]["unit_cost_ex_tax"] = Decimal(cost)
    workspace["consumptions"][0]["unit_cost_inc_tax"] = Decimal(cost)
    workspace["consumptions"][0]["cost_amount_ex_tax"] = Decimal(cost)
    workspace["consumptions"][0]["cost_amount_inc_tax"] = Decimal(cost)
    summary = workbook_v2.compute_project_summary(workspace)
    assert summary["cost_alert"] == expected


@pytest.mark.parametrize(
    ("cost_pct", "expected"),
    [
        ("79.99", "green"),
        ("80.00", "green"),
        ("80.004", "green"),
        ("80.005", "yellow"),
        ("80.01", "yellow"),
        ("100.00", "yellow"),
        ("100.004", "yellow"),
        ("100.005", "red"),
        ("100.01", "red"),
    ],
)
def test_canonical_cost_alert_uses_displayed_strict_thresholds(cost_pct, expected):
    workspace = _workspace()
    known_cost = Decimal(cost_pct) * Decimal("1000")
    workspace["canonical_metrics"] = {
        "contract_amount_basis": "inc_tax",
        "cost_progress_basis": "inc_tax",
        "total_contract_amount": "100000.00",
        "known_contract_amount": "100000.00",
        "contract_amount_complete": True,
        "received_amount": "30000.00",
        "collection_progress_pct": "30.00",
        "site_requisition_known_cost": str(known_cost),
        "approved_expense": "0.00",
        "actual_project_cost_known": str(known_cost),
        "cost_rate_lower_bound_pct": cost_pct,
        "cost_status": "normal",
        "cost_complete": True,
        "missing_cost_lines": 0,
    }
    workspace["canonical_completeness"] = {"status": "complete", "issues": []}

    summary = workbook_v2.compute_project_summary(workspace)

    assert summary["cost_alert"] == expected


def test_summary_fails_closed_when_any_contract_amount_is_missing():
    workspace = _workspace()
    workspace["contracts"].append({
        "project_contract_id": "project-contract-missing",
        "contract_no": "XSDD-MISSING",
        "contract_amount": None,
        "version": 1,
    })

    summary = workbook_v2.compute_project_summary(workspace)

    assert summary["known_contract_amount"] == Decimal("100000.00")
    assert summary["contract_amount_complete"] is False
    assert summary["total_contract_amount"] is None
    assert summary["collection_rate"] is None
    assert summary["cost_rate_lower_bound"] is None
    assert summary["cost_alert"] == "no_contract_amount"


def test_contract_sheet_displays_all_relationships_but_denominator_is_labeled():
    workspace = _workspace()
    workspace["contracts"][0].update(
        {
            "contract_status": "履行中",
            "status_mapping_state": "mapped",
            "included_in_total": True,
            "is_effective": True,
            "effective_from": "2026-01-01",
            "effective_to": None,
        }
    )
    workspace["contracts"].append(
        {
            "project_contract_id": "project-contract-historical",
            "contract_no": "XSDD-HISTORICAL",
            "contract_amount": None,
            "contract_status": "已终止",
            "status_mapping_state": "mapped",
            "included_in_total": False,
            "is_effective": False,
            "effective_from": "2025-01-01",
            "effective_to": "2025-12-31",
            "version": 1,
        }
    )

    summary = workbook_v2.compute_project_summary(workspace)
    artifact = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )

    assert summary["total_contract_amount"] == Decimal("100000.00")
    assert summary["contract_amount_complete"] is True
    assert artifact.preview["contracts"] == 2
    book = load_workbook(io.BytesIO(artifact.content), data_only=False)
    try:
        sheet = book["01_总览"]
        table = sheet.tables[workbook_v2.CONTRACT_TABLE]
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        headers = [
            sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        historical = {
            headers[index]: sheet.cell(min_row + 2, min_col + index).value
            for index in range(len(headers))
        }
        assert historical["合同编号"] == "XSDD-HISTORICAL"
        assert historical["原始合同状态"] == "已终止"
        assert historical["是否计入全部合同额"] == "否"
        assert historical["当前是否生效"] == "否"
        assert historical["金额完整性"] == "缺少合同额"
    finally:
        book.close()


def test_signed_metadata_and_unchanged_upload_is_noop():
    workspace = _workspace()
    artifact = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
        export_id="11111111-1111-4111-8111-111111111111",
    )

    book = load_workbook(io.BytesIO(artifact.content), data_only=False)
    try:
        assert book["98_字典"].sheet_state == "hidden"
        assert book["99_实体版本"].sheet_state == "veryHidden"
        assert book["99_元数据"].sheet_state == "veryHidden"
        metadata = {
            book["99_元数据"].cell(row, 1).value:
            book["99_元数据"].cell(row, 2).value
            for row in range(2, book["99_元数据"].max_row + 1)
        }
        assert metadata["protocol_id"] == workbook_v2.PROTOCOL_ID
        assert metadata["schema_version"] == "2.0"
        assert metadata["error_report"] == "false"
        assert len(metadata["metadata_hmac"]) == 64
    finally:
        book.close()

    validation = workbook_v2.validate_project_workbook(
        artifact.content,
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )
    assert validation.project_id == "project-001"
    assert validation.export_id == artifact.export_id
    assert validation.expected_revision == 7
    assert validation.unchanged is True
    assert validation.creates == ()


def test_base_7bfe7fc_workbook_remains_a_valid_unchanged_upload():
    content = BASE_V2_FIXTURE.read_bytes()
    assert hashlib.sha256(content).hexdigest() == BASE_V2_FIXTURE_SHA256

    validation = workbook_v2.validate_project_workbook(
        content,
        workspace=_workspace(),
        hmac_key=HMAC_KEY,
    )

    assert validation.project_id == "project-001"
    assert validation.export_id == "22222222-2222-4222-8222-222222222222"
    assert validation.expected_revision == 7
    assert validation.unchanged is True
    assert validation.creates == ()


def test_appended_collection_is_idempotent_by_file_and_logical_row_key():
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
        export_id="22222222-2222-4222-8222-222222222222",
    )
    uploaded = _append_collection(exported.content)
    validation = workbook_v2.validate_project_workbook(
        uploaded,
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )
    assert validation.unchanged is False
    assert len(validation.creates) == 1
    create = validation.creates[0]
    assert create.operation_key == (
        f"{exported.export_id}|01_总览|{create.client_row_id}"
    )
    assert create.cumulative_amount == Decimal("40000.00")
    assert create.report_month == date(2026, 8, 1)

    repository = _FakeRepository()
    first = workbook_v2.apply_project_workbook(validation, repository=repository)
    assert first.status == "applied"
    assert first.created == 1
    assert len(repository.atomic_calls) == 1

    exact_replay = workbook_v2.apply_project_workbook(validation, repository=repository)
    assert exact_replay.status == "file_replay"
    assert exact_replay.created == 0
    assert len(repository.atomic_calls) == 1

    # Excel re-save changes the ZIP hash, but the signed export/sheet/client id
    # operation key still makes the same logical CREATE a no-op.
    resaved = _append_collection(exported.content, resave_marker="second-save")
    logical = workbook_v2.validate_project_workbook(
        resaved,
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )
    logical_replay = workbook_v2.apply_project_workbook(logical, repository=repository)
    assert logical_replay.status == "logical_replay"
    assert logical_replay.replayed == 1
    assert len(repository.atomic_calls) == 1


@pytest.mark.parametrize("mutation", ["tamper_readonly", "remove_existing"])
def test_readonly_tamper_or_missing_existing_collection_is_rejected(mutation):
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )

    def edit(book):
        if mutation == "tamper_readonly":
            sheet = book["02_备件消耗"]
            headers = [cell.value for cell in sheet[1]]
            sheet.cell(2, headers.index("成本完整性") + 1, "成本完整")
            return
        sheet = book["01_总览"]
        table = sheet.tables["tbl_collections_v2"]
        min_col, min_row, max_col, _ = range_boundaries(table.ref)
        for column in range(min_col, max_col + 1):
            sheet.cell(min_row + 1, column).value = None

    uploaded = _edit_workbook(exported.content, edit)
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            uploaded,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert caught.value.status_code == 422


@pytest.mark.parametrize("cell", ["B1", "B6", "B8", "B14"])
def test_summary_tamper_is_rejected(cell):
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    tampered = _edit_workbook(
        exported.content,
        lambda book: setattr(book["01_总览"][cell], "value", "tampered"),
    )

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error, match="项目摘要被修改"):
        workbook_v2.validate_project_workbook(
            tampered,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )


def test_sparse_far_away_cell_is_rejected_before_rectangular_iteration():
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    unsafe = _edit_workbook(
        exported.content,
        lambda book: setattr(book["02_备件消耗"]["XFD1048576"], "value", "x"),
    )

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error, match="声明范围"):
        workbook_v2.validate_project_workbook(
            unsafe,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )


def test_formula_prefix_business_text_roundtrips_without_digest_drift():
    workspace = _workspace()
    workspace["project"]["project_name"] = "-医院维保项目"
    workspace["collections"][0]["remark"] = "@一期款"
    workspace["consumptions"][0]["part_name"] = "+备件一"
    workspace["tasks"][0]["title"] = "=确认下月巡检"
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )

    validation = workbook_v2.validate_project_workbook(
        exported.content,
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )

    assert validation.unchanged is True


def test_stale_revision_is_409_and_validation_failure_has_zero_writes():
    original = _workspace(revision=7)
    exported = workbook_v2.build_project_workbook(
        original,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    uploaded = _append_collection(exported.content)
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            uploaded,
            workspace=_workspace(revision=8),
            hmac_key=HMAC_KEY,
        )
    assert caught.value.status_code == 409
    assert caught.value.issues[0].code == "stale_workbook"

    validation = workbook_v2.validate_project_workbook(
        uploaded,
        workspace=original,
        hmac_key=HMAC_KEY,
    )
    repository = _FakeRepository(revision=8)
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as apply_error:
        workbook_v2.apply_project_workbook(validation, repository=repository)
    assert apply_error.value.status_code == 409
    assert repository.atomic_calls == []

    adapter = _FakeEndpointAdapter(_workspace(revision=8))
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as wrapper_error:
        workbook_v2.validate_and_store_project_workbook(
            "project-001",
            uploaded,
            adapter=adapter,
            hmac_key=HMAC_KEY,
        )
    assert wrapper_error.value.status_code == 409
    assert adapter.errors == {}

    tampered = _edit_workbook(
        uploaded,
        lambda book: setattr(book["03_报销单"]["F2"], "value", 9999),
    )
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error):
        workbook_v2.validate_project_workbook(
            tampered,
            workspace=original,
            hmac_key=HMAC_KEY,
        )
    assert repository.atomic_calls == []


def test_same_project_contract_and_report_month_is_rejected():
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    duplicate_month = _append_collection(exported.content)

    def change_to_existing_month(book):
        sheet = book["01_总览"]
        table = sheet.tables["tbl_collections_v2"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
        target = next(
            row for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("操作") + 1).value == "CREATE"
        )
        sheet.cell(target, headers.index("报告月份") + 1, "2026-07")

    duplicate_month = _edit_workbook(duplicate_month, change_to_existing_month)
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            duplicate_month,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert caught.value.issues[0].code == "duplicate_contract_month"


@pytest.mark.parametrize(
    ("workspace_change", "month", "expected_code"),
    [
        (lambda workspace: None, "2026-09", "future_report_month"),
        (
            lambda workspace: workspace["contracts"][0].update(
                {"effective_from": "2026-09-01"}
            ),
            "2026-08",
            "contract_inactive_for_report_month",
        ),
        (
            lambda workspace: workspace["contracts"][0].update(
                {"status_mapping_state": "unmapped"}
            ),
            "2026-08",
            "unmapped_project_contract",
        ),
    ],
)
def test_validate_rejects_business_invalid_appends_before_apply(
    workspace_change, month, expected_code
):
    workspace = _workspace()
    workspace_change(workspace)
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    uploaded = _append_collection(exported.content)

    def replace_month(book):
        sheet = book["01_总览"]
        table = sheet.tables[workbook_v2.COLLECTION_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("操作") + 1).value == "CREATE"
        )
        sheet.cell(target, headers.index("报告月份") + 1, month)

    uploaded = _edit_workbook(uploaded, replace_month)
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            uploaded,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert caught.value.issues[0].code == expected_code


def test_hidden_nonconfirmed_snapshot_still_blocks_duplicate_contract_month():
    workspace = _workspace()
    workspace["collections"][0].update(
        {
            "report_month": "2026-08",
            "status": "未确认",
        }
    )
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    uploaded = _append_collection(exported.content)

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            uploaded,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert caught.value.issues[0].code == "duplicate_contract_month"


def test_all_collection_statuses_are_signed_locked_and_covered_by_snapshot_validation():
    workspace = _workspace()
    workspace["collections"].extend(
        [
            {
                "collection_id": "collection-unconfirmed",
                "project_contract_id": "project-contract-001",
                "contract_no": "XSDD-20260001",
                "report_month": "2026-05",
                "cumulative_amount": Decimal("10000.00"),
                "voucher_no": "HK-UNCONFIRMED",
                "status": "未确认",
                "remark": "历史未确认事实",
                "version": 4,
            },
            {
                "collection_id": "collection-void",
                "project_contract_id": "project-contract-001",
                "contract_no": "XSDD-20260001",
                "report_month": "2026-06",
                "cumulative_amount": Decimal("20000.00"),
                "voucher_no": "HK-VOID",
                "status": "已作废",
                "remark": "历史作废事实",
                "version": 5,
            },
        ]
    )
    artifact = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )

    assert artifact.preview["collections"] == 3
    book = load_workbook(io.BytesIO(artifact.content), data_only=False)
    try:
        overview = book["01_总览"]
        table = overview.tables[workbook_v2.COLLECTION_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            overview.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        historical_rows = [
            {
                headers[column - min_col]: overview.cell(row, column).value
                for column in range(min_col, max_col + 1)
            }
            for row in range(min_row + 1, min_row + 4)
        ]
        assert [row["状态"] for row in historical_rows] == [
            "已确认",
            "未确认",
            "已作废",
        ]
        assert all(row["操作"] == "KEEP" for row in historical_rows)
        assert all(row["__row_token"] for row in historical_rows)
        for row in range(min_row + 1, min_row + 4):
            assert all(
                overview.cell(row, column).protection.locked
                for column in range(min_col, max_col + 1)
            )

        versions = book["99_实体版本"]
        version_table = versions.tables[workbook_v2.ENTITY_VERSION_TABLE]
        v_min_col, v_min_row, v_max_col, v_max_row = range_boundaries(version_table.ref)
        version_rows = {
            versions.cell(row, v_min_col + 1).value
            for row in range(v_min_row + 1, v_max_row + 1)
            if versions.cell(row, v_min_col).value == "collection"
        }
        assert version_rows == {
            "collection-001",
            "collection-unconfirmed",
            "collection-void",
        }
    finally:
        book.close()

    unchanged = workbook_v2.validate_project_workbook(
        artifact.content,
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )
    assert unchanged.unchanged is True
    assert unchanged.creates == ()

    def tamper_void_fact(book):
        sheet = book["01_总览"]
        table = sheet.tables[workbook_v2.COLLECTION_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        entity_col = min_col + headers.index("__entity_id")
        amount_col = min_col + headers.index("累计回款金额（含税）")
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, entity_col).value == "collection-void"
        )
        sheet.cell(target, amount_col, Decimal("20001.00"))

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            _edit_workbook(artifact.content, tamper_void_fact),
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert "只读业务数据或既有回款被修改" in str(caught.value)


def test_client_row_id_is_signed_and_cumulative_amount_cannot_decrease():
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    uploaded = _append_collection(exported.content, amount=20000)

    def replace_client_id(book):
        sheet = book["01_总览"]
        table = sheet.tables["tbl_collections_v2"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
        target = next(
            row for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("操作") + 1).value == "CREATE"
        )
        sheet.cell(target, headers.index("__client_row_id") + 1, "33333333-3333-4333-8333-333333333333")

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as client_error:
        workbook_v2.validate_project_workbook(
            _edit_workbook(uploaded, replace_client_id),
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert "client_row_id" in str(client_error.value)

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as amount_error:
        workbook_v2.validate_project_workbook(
            uploaded,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert amount_error.value.issues[0].code == "cumulative_decrease"


@pytest.mark.parametrize(
    ("header", "value", "expected_code"),
    [
        ("回款凭证号", "V" * 129, "receipt_reference_too_long"),
        ("备注", "R" * 32768, "collection_remark_too_long"),
    ],
)
def test_collection_text_limits_are_rejected_during_validate(
    header,
    value,
    expected_code,
):
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    uploaded = _append_collection(exported.content)

    marker = "OVERSIZED_REMARK_MARKER" if header == "备注" else value

    def set_oversized_value(book):
        sheet = book["01_总览"]
        table = sheet.tables[workbook_v2.COLLECTION_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [
            sheet.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        operation_col = min_col + headers.index("操作")
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, operation_col).value == "CREATE"
        )
        sheet.cell(target, min_col + headers.index(header), marker)

    unsafe = _edit_workbook(uploaded, set_oversized_value)
    if header == "备注":
        unsafe = _replace_zip_bytes(
            unsafe,
            marker.encode("utf-8"),
            value.encode("utf-8"),
        )

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            unsafe,
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )

    assert caught.value.issues[0].code == expected_code


def test_export_rejects_more_than_validator_row_limit():
    workspace = _workspace()
    template = dict(workspace["consumptions"][0])
    workspace["consumptions"] = [
        {**template, "consumption_id": f"issue-line-{index}"}
        for index in range(workbook_v2.MAX_ROWS_PER_TABLE + 1)
    ]

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.build_project_workbook(
            workspace,
            hmac_key=HMAC_KEY,
            exported_by="tester",
        )

    assert caught.value.issues[0].code == "export_row_limit"


def test_export_at_validator_row_limit_can_be_validated_unchanged():
    workspace = _workspace()
    template = dict(workspace["consumptions"][0])
    workspace["consumptions"] = [
        {**template, "consumption_id": f"issue-line-{index}"}
        for index in range(workbook_v2.MAX_ROWS_PER_TABLE)
    ]

    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    validation = workbook_v2.validate_project_workbook(
        exported.content,
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )

    assert validation.unchanged is True


def test_error_workbook_is_first_sheet_error_list_and_cannot_be_imported():
    issues = (
        workbook_v2.WorkbookIssue(
            "invalid_cumulative_amount",
            "累计回款金额必须大于 0",
            "01_总览",
            14,
            "累计回款金额（含税）",
        ),
    )
    content = workbook_v2.build_error_workbook(
        issues,
        hmac_key=HMAC_KEY,
        project_id="project-001",
        source_sha256="a" * 64,
    )
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        assert book.sheetnames[0] == "00_错误清单"
        assert book["00_错误清单"]["A2"].value == "invalid_cumulative_amount"
        metadata = {
            book["99_元数据"].cell(row, 1).value:
            book["99_元数据"].cell(row, 2).value
            for row in range(2, book["99_元数据"].max_row + 1)
        }
        assert metadata["error_report"] == "true"
        assert len(metadata["metadata_hmac"]) == 64
    finally:
        book.close()

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            content,
            workspace=_workspace(),
            hmac_key=HMAC_KEY,
        )
    assert "不可再次导入" in str(caught.value)


@pytest.mark.parametrize(
    ("member", "payload", "message"),
    [
        ("xl/vbaProject.bin", b"macro", "宏或外部链接"),
        ("xl/externalLinks/externalLink1.xml", b"<externalLink/>", "宏或外部链接"),
        (
            "xl/worksheets/_rels/synthetic.xml.rels",
            b'<Relationship TargetMode = "External" Target="https://example.invalid"/>',
            "宏或外部链接",
        ),
        ("xl/media/compression-bomb.bin", b"0" * 2_000_000, "ZIP bomb"),
    ],
)
def test_macro_external_link_and_zip_bomb_are_rejected(member, payload, message):
    exported = workbook_v2.build_project_workbook(
        _workspace(),
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    unsafe = _add_zip_member(exported.content, member, payload)
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            unsafe,
            workspace=_workspace(),
            hmac_key=HMAC_KEY,
        )
    assert message in str(caught.value)


def test_huge_cumulative_amount_is_a_controlled_validation_error():
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    uploaded = _append_collection(exported.content)

    def set_huge_amount(book):
        sheet = book["01_总览"]
        table = sheet.tables["tbl_collections_v2"]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
        target = next(
            row for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("操作") + 1).value == "CREATE"
        )
        sheet.cell(target, headers.index("累计回款金额（含税）") + 1, "1e100")

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            _edit_workbook(uploaded, set_huge_amount),
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert caught.value.issues[0].code == "invalid_cumulative_amount"


def test_appended_cumulative_amount_uses_half_up_and_rejects_rounded_overflow():
    workspace = _workspace()
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    rounded = workbook_v2.validate_project_workbook(
        _append_collection(exported.content, amount=Decimal("30000.005")),
        workspace=workspace,
        hmac_key=HMAC_KEY,
    )
    assert rounded.creates[0].cumulative_amount == Decimal("30000.01")

    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            _append_collection(
                exported.content,
                amount=Decimal("999999999999.999"),
            ),
            workspace=workspace,
            hmac_key=HMAC_KEY,
        )
    assert caught.value.issues[0].code == "invalid_cumulative_amount"


def test_cross_version_upload_is_controlled_422_and_hmac_key_has_no_default():
    legacy = Workbook()
    try:
        sheet = legacy.active
        sheet.title = "99_元数据"
        sheet.append(("key", "value"))
        sheet.append(("protocol_id", "ITDATA_MAINT_ROUNDTRIP/1.0"))
        sheet.append(("schema_version", "1.0"))
        table = Table(displayName="tbl_metadata_v1", ref="A1:B3")
        sheet.add_table(table)
        output = io.BytesIO()
        legacy.save(output)
    finally:
        legacy.close()
    with pytest.raises(workbook_v2.ProjectWorkbookV2Error) as caught:
        workbook_v2.validate_project_workbook(
            output.getvalue(),
            workspace=_workspace(),
            hmac_key=HMAC_KEY,
        )
    assert caught.value.status_code == 422

    with pytest.raises(ValueError, match="hmac_key"):
        workbook_v2.build_project_workbook(
            _workspace(),
            hmac_key=b"",
            exported_by="tester",
        )


def test_endpoint_adapter_stores_server_plan_and_retrievable_error_workbook():
    workspace = _workspace()
    adapter = _FakeEndpointAdapter(workspace)
    exported = workbook_v2.build_project_workbook(
        workspace,
        hmac_key=HMAC_KEY,
        exported_by="tester",
    )
    valid_attempt = workbook_v2.validate_and_store_project_workbook(
        "project-001",
        _append_collection(exported.content),
        adapter=adapter,
        hmac_key=HMAC_KEY,
    )
    assert valid_attempt.valid is True
    assert valid_attempt.validation_id in adapter.validations
    result = workbook_v2.apply_stored_project_workbook(
        "project-001",
        valid_attempt.validation_id,
        adapter=adapter,
    )
    assert result.status == "applied"

    tampered = _edit_workbook(
        exported.content,
        lambda book: setattr(book["02_备件消耗"]["I2"], "value", "伪造完整"),
    )
    failed_attempt = workbook_v2.validate_and_store_project_workbook(
        "project-001",
        tampered,
        adapter=adapter,
        hmac_key=HMAC_KEY,
    )
    assert failed_attempt.valid is False
    report = workbook_v2.load_validation_error_workbook(
        failed_attempt.validation_id,
        adapter=adapter,
    )
    report_book = load_workbook(io.BytesIO(report), data_only=False)
    try:
        assert report_book.sheetnames[0] == "00_错误清单"
    finally:
        report_book.close()
