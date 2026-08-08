"""Public API contract for the stable-project four-sheet workbook loop."""

from __future__ import annotations

import asyncio
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from types import SimpleNamespace
import threading

from fastapi import FastAPI, Response
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
import pytest
from sqlalchemy import func, null, select

from app import auth
from app.api import (
    maintenance_project_operations,
    maintenance_project_workbooks,
    maintenance_projects,
)
from app.auth import hash_password
from app.models.dimensions import DimPart
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
    MaintenanceProjectWorkbookOperation,
    MaintenanceProjectWorkbookState,
    MaintenanceProjectWorkbookValidation,
    MaintenanceSiteIssue,
    MaintenanceSiteIssueLine,
)
from app.models.system import SysUser
from app.security import UserContext
from app.services import maintenance_project_workbook_adapter
from app.services.maintenance_project_workbook_v2 import (
    COLLECTION_TABLE,
    PROTOCOL_ID,
    ProjectWorkbookV2Error,
)


def _client(db, *, username: str) -> TestClient:
    db.add(
        SysUser(
            username=username,
            role="admin",
            display_name="合成工作簿管理员",
            password_hash=hash_password("synthetic-password-123"),
        )
    )
    db.commit()
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_project_operations.router, prefix="/api")
    app.include_router(maintenance_project_workbooks.router, prefix="/api")
    app.include_router(maintenance_projects.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": username, "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def _permission_client(db, *, username: str, permissions: dict) -> TestClient:
    db.add(
        SysUser(
            username=username,
            role="boss",
            display_name="合成工作簿权限账号",
            password_hash=hash_password("synthetic-password-123"),
            permissions=permissions,
        )
    )
    db.commit()
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_project_operations.router, prefix="/api")
    app.include_router(maintenance_project_workbooks.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": username, "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    return client


def _project_and_contract(client: TestClient, db, *, suffix: str) -> tuple[str, dict]:
    project_id = f"project-workbook-{suffix}"
    db.add(
        MaintenanceProject(
            project_id=project_id,
            project_code=f"WB-{suffix}",
            display_name=f"合成工作簿项目 {suffix}",
            project_manager_id="manager-synthetic",
            lifecycle_status="ongoing",
        )
    )
    db.commit()
    response = client.post(
        f"/api/maintenance/projects/stable/{project_id}/contracts",
        json={
            "contract_id": f"contract-{suffix}",
            "contract_no": f"XS-{suffix}",
            "contract_amount": "1000.00",
            "contract_status": "active",
            "status_mapping_state": "mapped",
            "status_mapping_version": "synthetic-map-v1",
            "included_in_total": True,
            "effective_from": "2026-01-01",
            "source": "synthetic-test",
            "reason": "建立工作簿接口测试合同",
        },
    )
    assert response.status_code == 201, response.text
    return project_id, response.json()


def _append_collection(
    source: bytes,
    *,
    project_contract_id: str,
    contract_no: str,
    report_month: str = "2026-08",
    amount: str = "320.00",
) -> bytes:
    book = load_workbook(BytesIO(source))
    try:
        sheet = book["01_总览"]
        table = sheet.tables[COLLECTION_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        assert min_col == 1 and max_col == 12
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, 9).value in (None, "")
            and all(
                sheet.cell(row, column).value in (None, "") for column in range(1, 9)
            )
        )
        sheet.cell(target, 1, "CREATE")
        sheet.cell(target, 2, project_contract_id)
        sheet.cell(target, 3, contract_no)
        sheet.cell(target, 4, report_month)
        sheet.cell(target, 5, amount)
        sheet.cell(target, 6, "SYNTHETIC-VOUCHER-001")
        sheet.cell(target, 7, "已确认")
        sheet.cell(target, 8, "合成月度回填")
        output = BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _download(client: TestClient, project_id: str) -> bytes:
    response = client.get(f"/api/maintenance/projects/stable/{project_id}/workbook")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert response.headers["cache-control"] == "no-store"
    return response.content


def _validate(client: TestClient, project_id: str, content: bytes, name="update.xlsx"):
    return client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/validate",
        files={
            "file": (
                name,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_export_preflights_row_limit_before_loading_full_workspace(db, monkeypatch):
    client = _client(db, username="workbook_preflight_admin")
    project_id, _contract = _project_and_contract(client, db, suffix="preflight")
    monkeypatch.setattr(
        maintenance_project_workbook_adapter,
        "MAX_ROWS_PER_TABLE",
        0,
        raising=False,
    )

    def fail_full_workspace_load(*_args, **_kwargs):
        raise AssertionError("full ORM workspace loaded before row-limit preflight")

    monkeypatch.setattr(
        maintenance_project_workbook_adapter.operations,
        "project_workbook_workspace",
        fail_full_workspace_load,
    )

    response = client.get(
        f"/api/maintenance/projects/stable/{project_id}/workbook"
    )

    assert response.status_code == 422, response.text
    assert "合同" in response.json()["detail"]["message"]


def test_workspace_preflight_counts_each_exported_fact_category(db, monkeypatch):
    client = _client(db, username="workbook_preflight_categories_admin")
    project_id, contract = _project_and_contract(
        client,
        db,
        suffix="preflight-cat",
    )
    second_contract = MaintenanceProjectContract(
        project_contract_id="preflight-second-contract",
        project_id=project_id,
        contract_id="preflight-second-contract-source",
        contract_no="XS-PREFLIGHT-SECOND",
        contract_amount=100,
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-map-v1",
        included_in_total=True,
        effective_from=date(2026, 1, 1),
        source="synthetic-test",
    )
    part = DimPart(pn_std="PN-WORKBOOK-PREFLIGHT")
    issue = MaintenanceSiteIssue(
        issue_id="preflight-issue",
        project_id=project_id,
        issue_no="ISSUE-PREFLIGHT",
        issue_date=date(2026, 8, 1),
        raw_status="confirmed",
        status_mapping_state="mapped",
        normalized_status="confirmed",
        status_mapping_version="synthetic-map-v1",
    )
    db.add_all([second_contract, part, issue])
    db.flush()
    db.add_all(
        [
            MaintenanceCollectionSnapshot(
                collection_id=f"preflight-collection-{index}",
                project_id=project_id,
                project_contract_id=contract["project_contract_id"],
                report_month=date(2026, month, 1),
                cumulative_amount=index * 100,
                status="confirmed",
            )
            for index, month in enumerate((7, 8), start=1)
        ]
        + [
            MaintenanceSiteIssueLine(
                issue_line_id=f"preflight-line-{index}",
                issue_id=issue.issue_id,
                line_no=index,
                part_id=part.id,
                pn=part.pn_std,
                quantity=1,
                algorithm_version="synthetic-cost-v1",
            )
            for index in (1, 2)
        ]
        + [
            MaintenanceProjectExpenseAttribution(
                expense_id=f"preflight-expense-{index}",
                project_id=project_id,
                expense_ref=f"BX-PREFLIGHT-{index}",
                expense_date=date(2026, 8, index),
                amount_ex_tax=10,
                raw_status="approved",
                status_mapping_state="mapped",
                normalized_status="approved",
                status_mapping_version="synthetic-map-v1",
            )
            for index in (1, 2)
        ]
    )
    db.commit()
    monkeypatch.setattr(
        maintenance_project_workbook_adapter,
        "MAX_ROWS_PER_TABLE",
        1,
    )
    adapter = maintenance_project_workbook_adapter.MaintenanceProjectWorkbookAdapter(
        db,
        user_ctx=UserContext(
            user_id="workbook_preflight_categories_admin",
            role="admin",
            is_authenticated=True,
        ),
        operator="workbook_preflight_categories_admin",
        as_of=date(2026, 8, 9),
    )

    with pytest.raises(ProjectWorkbookV2Error) as caught:
        adapter.load_workspace(project_id)

    message = str(caught.value)
    assert "合同=2" in message
    assert "回款=2" in message
    assert "已确认现场领用明细=2" in message
    assert "已审批报销=2" in message


def test_workspace_preflight_bounds_ineligible_facts_loaded_for_completeness(
    db,
    monkeypatch,
):
    project_id = "project-workbook-loaded-facts"
    project = MaintenanceProject(
        project_id=project_id,
        project_code="WB-LOADED-FACTS",
        display_name="合成全量事实预检项目",
        lifecycle_status="ongoing",
    )
    part = DimPart(pn_std="PN-WORKBOOK-LOADED-FACTS")
    issue = MaintenanceSiteIssue(
        issue_id="loaded-void-issue",
        project_id=project_id,
        issue_no="ISSUE-LOADED-VOID",
        issue_date=date(2026, 8, 1),
        raw_status="void",
        status_mapping_state="mapped",
        normalized_status="void",
        status_mapping_version="synthetic-map-v1",
    )
    db.add_all([project, part, issue])
    db.flush()
    db.add_all(
        [
            MaintenanceSiteIssueLine(
                issue_line_id=f"loaded-void-line-{index}",
                issue_id=issue.issue_id,
                line_no=index,
                part_id=part.id,
                pn=part.pn_std,
                quantity=1,
                algorithm_version="synthetic-cost-v1",
            )
            for index in (1, 2)
        ]
        + [
            MaintenanceProjectExpenseAttribution(
                expense_id=f"loaded-rejected-expense-{index}",
                project_id=project_id,
                expense_ref=f"BX-LOADED-REJECTED-{index}",
                expense_date=date(2026, 8, index),
                amount_ex_tax=10,
                raw_status="rejected",
                status_mapping_state="mapped",
                normalized_status="rejected",
                status_mapping_version="synthetic-map-v1",
            )
            for index in (1, 2)
        ]
    )
    db.commit()
    monkeypatch.setattr(
        maintenance_project_workbook_adapter,
        "MAX_ROWS_PER_TABLE",
        1,
    )
    adapter = maintenance_project_workbook_adapter.MaintenanceProjectWorkbookAdapter(
        db,
        user_ctx=UserContext(
            user_id="workbook_loaded_facts_admin",
            role="admin",
            is_authenticated=True,
        ),
        operator="workbook_loaded_facts_admin",
        as_of=date(2026, 8, 9),
    )

    with pytest.raises(ProjectWorkbookV2Error) as caught:
        adapter.load_workspace(project_id)

    message = str(caught.value)
    assert "现场领用全量事实=2" in message
    assert "报销全量事实=2" in message
    assert "已确认现场领用明细" not in message
    assert "已审批报销" not in message


def test_validate_opportunistically_expires_compacts_and_retires_records(db):
    client = _client(db, username="workbook_retention_validate_admin")
    project_id, _contract = _project_and_contract(
        client,
        db,
        suffix="retention-validate",
    )
    now = datetime.now(timezone.utc)
    rows = [
        MaintenanceProjectWorkbookValidation(
            validation_id="retention-expired-valid",
            project_id=project_id,
            export_id="retention-export-valid",
            expected_revision=0,
            file_sha256="1" * 64,
            plan_json={"large": "plan"},
            status="valid",
            issues_json=[{"message": "stale"}],
            error_workbook=None,
            created_by="workbook_retention_validate_admin",
            expires_at=now - timedelta(minutes=1),
        ),
        MaintenanceProjectWorkbookValidation(
            validation_id="retention-expired-error",
            project_id=project_id,
            export_id="retention-export-error",
            expected_revision=0,
            file_sha256="2" * 64,
            plan_json=null(),
            status="error",
            issues_json=[{"message": "stale"}],
            error_workbook=b"large-error-workbook",
            created_by="workbook_retention_validate_admin",
            expires_at=now - timedelta(minutes=1),
        ),
        MaintenanceProjectWorkbookValidation(
            validation_id="retention-current-applied",
            project_id=project_id,
            export_id="retention-export-applied",
            expected_revision=0,
            file_sha256="3" * 64,
            plan_json={"large": "applied-plan"},
            status="applied",
            issues_json=[],
            error_workbook=None,
            created_by="workbook_retention_validate_admin",
            expires_at=now + timedelta(days=1),
            applied_at=now,
        ),
        MaintenanceProjectWorkbookValidation(
            validation_id="retention-old-expired",
            project_id=project_id,
            export_id="retention-export-old-expired",
            expected_revision=0,
            file_sha256="4" * 64,
            plan_json=null(),
            status="expired",
            issues_json=[],
            error_workbook=None,
            created_by="workbook_retention_validate_admin",
            expires_at=now - timedelta(days=8),
        ),
        MaintenanceProjectWorkbookValidation(
            validation_id="retention-old-applied",
            project_id=project_id,
            export_id="retention-export-old-applied",
            expected_revision=0,
            file_sha256="5" * 64,
            plan_json=null(),
            status="applied",
            issues_json=[],
            error_workbook=None,
            created_by="workbook_retention_validate_admin",
            expires_at=now - timedelta(days=31),
            applied_at=now - timedelta(days=31),
        ),
    ]
    db.add_all(rows)
    db.commit()

    response = _validate(client, project_id, b"not-an-xlsx")

    assert response.status_code == 200, response.text
    for validation_id in (
        "retention-expired-valid",
        "retention-expired-error",
    ):
        row = db.get(MaintenanceProjectWorkbookValidation, validation_id)
        db.refresh(row)
        assert row.status == "expired"
        assert row.plan_json is None
        assert row.issues_json == []
        assert row.error_workbook is None
    applied = db.get(
        MaintenanceProjectWorkbookValidation,
        "retention-current-applied",
    )
    db.refresh(applied)
    assert applied.status == "applied"
    assert applied.plan_json is None
    assert (
        db.get(MaintenanceProjectWorkbookValidation, "retention-old-expired")
        is None
    )
    assert (
        db.get(MaintenanceProjectWorkbookValidation, "retention-old-applied")
        is None
    )


def test_validation_cleanup_limits_each_phase_to_one_fixed_batch(db, monkeypatch):
    project = MaintenanceProject(
        project_id="project-workbook-cleanup-batch",
        project_code="WB-CLEANUP-BATCH",
        display_name="合成清理批次项目",
        lifecycle_status="ongoing",
    )
    db.add(project)
    db.flush()
    now = datetime.now(timezone.utc)
    db.add_all(
        [
            MaintenanceProjectWorkbookValidation(
                validation_id=f"cleanup-batch-{index:03d}",
                project_id=project.project_id,
                export_id=f"cleanup-export-{index:03d}",
                expected_revision=0,
                file_sha256=f"{index:064x}",
                plan_json={"large": "plan"},
                status="valid",
                issues_json=[{"message": "stale"}],
                error_workbook=None,
                created_by="workbook_cleanup_batch_admin",
                expires_at=now - timedelta(minutes=1),
            )
            for index in range(3)
        ]
    )
    db.commit()
    monkeypatch.setattr(
        maintenance_project_workbook_adapter,
        "VALIDATION_CLEANUP_BATCH_SIZE",
        2,
    )

    maintenance_project_workbook_adapter.cleanup_project_workbook_validations(
        db,
        now=now,
    )
    db.commit()

    statuses = list(
        db.scalars(
            select(MaintenanceProjectWorkbookValidation.status).order_by(
                MaintenanceProjectWorkbookValidation.validation_id
            )
        )
    )
    assert statuses == ["expired", "expired", "valid"]


def test_apply_and_error_download_persist_opportunistic_expiration(db):
    client = _client(db, username="workbook_retention_paths_admin")
    project_id, _contract = _project_and_contract(
        client,
        db,
        suffix="retention-paths",
    )
    state = db.get(MaintenanceProjectWorkbookState, project_id)
    db.refresh(state)
    now = datetime.now(timezone.utc)
    expired_valid = MaintenanceProjectWorkbookValidation(
        validation_id="retention-apply-expired",
        project_id=project_id,
        export_id="retention-apply-export",
        expected_revision=state.revision,
        file_sha256="6" * 64,
        plan_json={"large": "expired-plan"},
        status="valid",
        issues_json=[{"message": "stale"}],
        error_workbook=None,
        created_by="workbook_retention_paths_admin",
        expires_at=now - timedelta(seconds=1),
    )
    db.add(expired_valid)
    db.commit()

    apply_response = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": expired_valid.validation_id,
            "data_version": state.data_version,
        },
    )

    assert apply_response.status_code == 409, apply_response.text
    db.expire_all()
    persisted_valid = db.get(
        MaintenanceProjectWorkbookValidation,
        expired_valid.validation_id,
    )
    assert persisted_valid.status == "expired"
    assert persisted_valid.plan_json is None
    assert persisted_valid.issues_json == []

    expired_error = MaintenanceProjectWorkbookValidation(
        validation_id="retention-download-expired",
        project_id=project_id,
        export_id="retention-download-export",
        expected_revision=state.revision,
        file_sha256="7" * 64,
        plan_json=null(),
        status="error",
        issues_json=[{"message": "stale"}],
        error_workbook=b"large-error-workbook",
        created_by="workbook_retention_paths_admin",
        expires_at=now - timedelta(seconds=1),
    )
    db.add(expired_error)
    db.commit()

    download_response = client.get(
        "/api/maintenance/workbook-validations/"
        f"{expired_error.validation_id}/errors.xlsx"
    )

    assert download_response.status_code == 404, download_response.text
    db.expire_all()
    persisted_error = db.get(
        MaintenanceProjectWorkbookValidation,
        expired_error.validation_id,
    )
    assert persisted_error.status == "expired"
    assert persisted_error.issues_json == []
    assert persisted_error.error_workbook is None


def test_export_validate_apply_is_a_server_owned_atomic_loop(db):
    client = _client(db, username="workbook_loop_admin")
    project_id, contract = _project_and_contract(client, db, suffix="loop")
    workbook = _download(client, project_id)

    export_state = db.get(MaintenanceProjectWorkbookState, project_id)
    db.refresh(export_state)
    assert export_state.last_export_id
    assert export_state.last_exported_at
    export_ledgers = db.scalar(
        select(func.count())
        .select_from(MaintenanceProjectWorkbookOperation)
        .where(MaintenanceProjectWorkbookOperation.operation_type == "file_export")
    )
    assert export_ledgers == 1

    edited = _append_collection(
        workbook,
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
    )
    validated = _validate(client, project_id, edited)
    assert validated.status_code == 200, validated.text
    plan = validated.json()
    assert set(plan) == {
        "validation_token",
        "project_id",
        "data_version",
        "filename",
        "preview",
        "changes",
        "warnings",
        "errors",
        "can_apply",
    }
    assert plan["project_id"] == project_id
    assert plan["filename"] == "update.xlsx"
    assert plan["changes"] == {"collection_append": 1}
    assert plan["errors"] == []
    assert plan["can_apply"] is True
    assert plan["preview"]["protocol_version"] == PROTOCOL_ID
    assert plan["preview"]["latest_tracking_month"] == "2026-08"
    assert validated.headers["cache-control"] == "no-store"
    assert validated.headers["x-content-type-options"] == "nosniff"
    assert [sheet["code"] for sheet in plan["preview"]["sheets"]] == [
        "overview",
        "site_requisitions",
        "approved_expenses",
        "manager_tracking",
    ]
    assert [sheet["name"] for sheet in plan["preview"]["sheets"]] == [
        "01_总览",
        "02_备件消耗",
        "03_报销单",
        "04_项目经理追踪与提醒",
    ]

    validation_row = db.get(
        MaintenanceProjectWorkbookValidation, plan["validation_token"]
    )
    db.refresh(validation_row)
    assert validation_row.status == "valid"
    assert validation_row.plan_json["creates"][0]["cumulative_amount"] == "320.00"

    applied = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": plan["validation_token"],
            "data_version": plan["data_version"],
        },
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["applied"] is True
    assert applied.json()["changed_rows"] == 1
    assert applied.json()["data_version"] != plan["data_version"]
    assert applied.headers["cache-control"] == "no-store"
    assert applied.headers["x-content-type-options"] == "nosniff"

    db.expire_all()
    collection = db.scalar(
        select(MaintenanceCollectionSnapshot).where(
            MaintenanceCollectionSnapshot.project_id == project_id
        )
    )
    assert collection is not None
    assert collection.report_month == date(2026, 8, 1)
    assert str(collection.cumulative_amount) == "320.00"
    assert collection.status == "confirmed"
    assert db.get(MaintenanceProjectWorkbookState, project_id).revision == 2
    assert (
        db.get(MaintenanceProjectWorkbookValidation, plan["validation_token"]).status
        == "applied"
    )
    assert (
        db.get(
            MaintenanceProjectWorkbookValidation,
            plan["validation_token"],
        ).plan_json
        is None
    )
    operation_types = list(
        db.scalars(
            select(MaintenanceProjectWorkbookOperation.operation_type)
            .where(MaintenanceProjectWorkbookOperation.project_id == project_id)
            .order_by(MaintenanceProjectWorkbookOperation.id)
        )
    )
    assert operation_types == ["file_export", "collection_create", "file_apply"]


def test_invalid_file_returns_persisted_error_workbook_without_writes(
    db,
    monkeypatch,
):
    access_logs = []
    monkeypatch.setattr(
        maintenance_project_workbooks,
        "record_access_log",
        lambda *args: access_logs.append(args),
    )
    client = _client(db, username="workbook_error_admin")
    project_id, _contract = _project_and_contract(client, db, suffix="error")
    before_revision = db.get(MaintenanceProjectWorkbookState, project_id).revision

    response = _validate(client, project_id, b"not-an-xlsx")
    assert response.status_code == 200, response.text
    result = response.json()
    assert result["can_apply"] is False
    assert result["errors"]
    assert result["changes"] == {"collection_append": 0}

    error_file = client.get(
        f"/api/maintenance/workbook-validations/{result['validation_token']}/errors.xlsx"
    )
    assert error_file.status_code == 200, error_file.text
    assert error_file.content.startswith(b"PK")
    assert error_file.headers["cache-control"] == "no-store"
    assert error_file.headers["x-content-type-options"] == "nosniff"
    assert len(access_logs) == 1
    assert access_logs[0][1:] == (
        "download_workbook_errors",
        f"maintenance_workbook_validation:{result['validation_token']}",
        {"validation_id": result["validation_token"]},
    )
    other_client = _client(db, username="workbook_error_other_admin")
    wrong_user = other_client.get(
        f"/api/maintenance/workbook-validations/{result['validation_token']}/errors.xlsx"
    )
    assert wrong_user.status_code == 404
    assert len(access_logs) == 1
    row = db.get(MaintenanceProjectWorkbookValidation, result["validation_token"])
    db.refresh(row)
    assert row.status == "error"
    assert row.plan_json is None
    assert (
        db.get(MaintenanceProjectWorkbookState, project_id).revision == before_revision
    )
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )


def test_validate_worker_owns_session_and_runs_outside_event_loop(
    tmp_path,
    monkeypatch,
):
    upload = tmp_path / "update.xlsx"
    upload.write_bytes(b"synthetic")
    events = []
    worker_threads = []

    class FakeSession:
        def commit(self):
            events.append(("commit", threading.get_ident()))

        def rollback(self):
            events.append(("rollback", threading.get_ident()))

        def close(self):
            events.append(("close", threading.get_ident()))

    class FakeAdapter:
        def __init__(self, worker_db, **_kwargs):
            worker_threads.append(threading.get_ident())
            assert isinstance(worker_db, FakeSession)

        def validate(self, project_id, content, *, hmac_key):
            assert project_id == "project-worker"
            assert content == b"synthetic"
            assert hmac_key == b"x" * 16
            return (
                SimpleNamespace(metadata={}, creates=(), unchanged=True),
                (),
                "validation-worker",
            )

        def load_workspace(self, project_id):
            assert project_id == "project-worker"
            return {
                "as_of": "2026-08-09",
                "data_version": "worker-version",
                "contracts": [],
                "collections": [],
                "consumptions": [],
                "expenses": [],
                "tasks": [],
            }

    monkeypatch.setattr(
        maintenance_project_workbooks,
        "SessionLocal",
        FakeSession,
    )
    monkeypatch.setattr(
        maintenance_project_workbooks,
        "MaintenanceProjectWorkbookAdapter",
        FakeAdapter,
    )
    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_real_operator",
        lambda worker_db, ident: ident["sub"],
    )
    ctx = UserContext(
        user_id="worker-user",
        role="admin",
        is_authenticated=True,
    )

    async def exercise():
        loop_thread = threading.get_ident()
        result = await maintenance_project_workbooks._run_project_workbook_validation_worker(
            project_id="project-worker",
            upload_path=str(upload),
            original_name="update.xlsx",
            ident={"sub": "worker-user"},
            ctx=ctx,
            hmac_key=b"x" * 16,
        )
        return loop_thread, result

    loop_thread, result = asyncio.run(exercise())

    assert result["validation_token"] == "validation-worker"
    assert worker_threads and worker_threads[0] != loop_thread
    assert events == [
        ("commit", worker_threads[0]),
        ("close", worker_threads[0]),
    ]


def test_validate_worker_cancellation_waits_for_terminal_thread(monkeypatch):
    entered = threading.Event()
    release = threading.Event()
    finished = threading.Event()

    def blocking_worker(**_kwargs):
        entered.set()
        assert release.wait(timeout=2)
        finished.set()
        return {"can_apply": True}

    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_validate_project_workbook_in_worker",
        blocking_worker,
    )

    async def exercise():
        task = asyncio.create_task(
            maintenance_project_workbooks._run_project_workbook_validation_worker()
        )
        assert await asyncio.to_thread(entered.wait, 1)
        task.cancel()
        await asyncio.sleep(0.05)
        completed_before_release = task.done()
        release.set()
        with pytest.raises(asyncio.CancelledError):
            await task
        return completed_before_release

    assert asyncio.run(exercise()) is False
    assert finished.is_set()


def test_cancelled_validate_keeps_temp_and_limiter_until_worker_finishes(
    tmp_path,
    monkeypatch,
):
    upload = tmp_path / "owned-until-worker-finishes.xlsx"
    upload.write_bytes(b"synthetic")
    entered = threading.Event()
    release = threading.Event()
    finished = threading.Event()

    async def parsed_upload(_request):
        return str(upload), "update.xlsx"

    def blocking_worker(**kwargs):
        assert kwargs["upload_path"] == str(upload)
        assert upload.exists()
        entered.set()
        assert release.wait(timeout=2)
        assert upload.exists()
        finished.set()
        return {"can_apply": True}

    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_parse_and_save_roundtrip_upload",
        parsed_upload,
    )
    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_validate_project_workbook_in_worker",
        blocking_worker,
    )
    ctx = UserContext(
        user_id="cancel-user",
        role="admin",
        is_authenticated=True,
    )
    limiter = maintenance_project_workbooks._PROJECT_WORKBOOK_VALIDATE_LIMITER

    async def exercise():
        task = asyncio.create_task(
            maintenance_project_workbooks.validate_project_workbook_upload(
                request=object(),
                response=Response(),
                project_id="project-cancel",
                ident={"sub": "cancel-user"},
                _auth="admin",
                _page=None,
                _action=None,
                ctx=ctx,
            )
        )
        assert await asyncio.to_thread(entered.wait, 1)
        task.cancel()
        await asyncio.sleep(0.05)
        assert task.done() is False
        assert upload.exists()
        assert limiter.acquire(blocking=False) is False
        release.set()
        with pytest.raises(asyncio.CancelledError):
            await task

    asyncio.run(exercise())

    assert finished.is_set()
    assert upload.exists() is False
    assert limiter.acquire(blocking=False)
    limiter.release()


def test_validate_concurrency_limiter_rejects_before_upload_read(db, monkeypatch):
    client = _client(db, username="workbook_validate_busy_admin")
    project_id, _contract = _project_and_contract(
        client,
        db,
        suffix="validate-busy",
    )
    upload_read = False

    async def fail_if_read(_request):
        nonlocal upload_read
        upload_read = True
        raise AssertionError("busy validation consumed upload")

    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_parse_and_save_roundtrip_upload",
        fail_if_read,
    )
    limiter = maintenance_project_workbooks._PROJECT_WORKBOOK_VALIDATE_LIMITER
    assert limiter.acquire(blocking=False)
    try:
        response = _validate(client, project_id, b"must-not-be-read")
    finally:
        limiter.release()

    assert response.status_code == 429
    assert response.headers["retry-after"] == "5"
    assert upload_read is False


def test_unchanged_monthly_workbook_can_confirm_zero_row_update(db):
    client = _client(db, username="workbook_unchanged_admin")
    project_id, _contract = _project_and_contract(client, db, suffix="unchanged")
    workbook = _download(client, project_id)

    validated = _validate(client, project_id, workbook)
    assert validated.status_code == 200, validated.text
    plan = validated.json()
    assert plan["changes"] == {"collection_append": 0}
    assert plan["can_apply"] is True
    assert plan["warnings"] == ["未检测到新增回款；确认后将记录本月已更新"]
    before = db.get(MaintenanceProjectWorkbookState, project_id)
    db.refresh(before)
    before_revision = before.revision

    applied = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": plan["validation_token"],
            "data_version": plan["data_version"],
        },
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["applied"] is True
    assert applied.json()["changed_rows"] == 0
    assert applied.json()["data_version"] != plan["data_version"]

    db.expire_all()
    state = db.get(MaintenanceProjectWorkbookState, project_id)
    assert state.revision == before_revision + 1
    assert state.last_applied_at is not None
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )
    assert (
        db.scalar(
            select(func.count())
            .select_from(MaintenanceProjectWorkbookOperation)
            .where(MaintenanceProjectWorkbookOperation.operation_type == "file_apply")
        )
        == 1
    )
    assert (
        db.get(MaintenanceProjectWorkbookValidation, plan["validation_token"]).status
        == "applied"
    )

    replay = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": plan["validation_token"],
            "data_version": applied.json()["data_version"],
        },
    )
    assert replay.status_code == 409


def test_export_preserves_approved_expense_business_fields(db):
    client = _client(db, username="workbook_expense_admin")
    project_id, contract = _project_and_contract(client, db, suffix="expense")
    created = client.post(
        f"/api/maintenance/projects/stable/{project_id}/expenses",
        json={
            "expense_id": "expense-workbook-001",
            "project_contract_id": contract["project_contract_id"],
            "expense_ref": "BX-WORKBOOK-001",
            "expense_date": "2026-08-01",
            "applicant": "合成报销人",
            "category": "差旅费",
            "expense_reason": "项目现场支持",
            "amount_ex_tax": "50.00",
            "raw_status": "synthetic-finished",
            "status_mapping_state": "mapped",
            "normalized_status": "approved",
            "status_mapping_version": "synthetic-expense-map-v1",
            "reason": "验证报销字段进入工作簿",
        },
    )
    assert created.status_code == 201, created.text
    excluded = client.post(
        f"/api/maintenance/projects/stable/{project_id}/contracts",
        json={
            "contract_id": "contract-expense-historical",
            "contract_no": "XS-EXPENSE-HISTORICAL",
            "contract_amount": None,
            "contract_status": "已终止",
            "status_mapping_state": "mapped",
            "status_mapping_version": "synthetic-map-v1",
            "included_in_total": False,
            "effective_from": "2025-01-01",
            "effective_to": "2025-12-31",
            "source": "synthetic-test",
            "reason": "验证工作簿展示全部关联合同",
        },
    )
    assert excluded.status_code == 201, excluded.text

    book = load_workbook(BytesIO(_download(client, project_id)), data_only=False)
    try:
        sheet = book["03_报销单"]
        assert [sheet.cell(2, column).value for column in range(1, 9)] == [
            "expense-workbook-001",
            "BX-WORKBOOK-001",
            datetime(2026, 8, 1, 0, 0),
            "合成报销人",
            "差旅费",
            50,
            "已审批",
            "项目现场支持",
        ]
        overview = book["01_总览"]
        contract_table = overview.tables["tbl_project_contracts_v2"]
        min_col, min_row, max_col, max_row = range_boundaries(contract_table.ref)
        headers = [
            overview.cell(min_row, column).value
            for column in range(min_col, max_col + 1)
        ]
        contracts = [
            {
                headers[index]: overview.cell(row, min_col + index).value
                for index in range(len(headers))
            }
            for row in range(min_row + 1, max_row + 1)
        ]
        historical = next(
            row
            for row in contracts
            if row["合同编号"] == "XS-EXPENSE-HISTORICAL"
        )
        assert historical["原始合同状态"] == "已终止"
        assert historical["是否计入全部合同额"] == "否"
        assert historical["当前是否生效"] == "否"
        assert historical["金额完整性"] == "缺少合同额"
    finally:
        book.close()


def test_apply_rejects_expired_wrong_project_replay_and_client_plan(db):
    client = _client(db, username="workbook_fail_close_admin")
    project_id, contract = _project_and_contract(client, db, suffix="source")
    other_project_id, _ = _project_and_contract(client, db, suffix="other")
    content = _append_collection(
        _download(client, project_id),
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
    )
    plan = _validate(client, project_id, content).json()
    body = {
        "validation_token": plan["validation_token"],
        "data_version": plan["data_version"],
    }

    other_client = _client(db, username="workbook_apply_other_admin")
    wrong_user = other_client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json=body,
    )
    assert wrong_user.status_code == 409
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )

    wrong_project = client.post(
        f"/api/maintenance/projects/stable/{other_project_id}/workbook/apply",
        json=body,
    )
    assert wrong_project.status_code == 409
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )

    injected = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={**body, "plan": {"creates": [{"cumulative_amount": "0.01"}]}},
    )
    assert injected.status_code == 422
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )

    row = db.get(MaintenanceProjectWorkbookValidation, plan["validation_token"])
    db.refresh(row)
    row.expires_at = datetime.now(timezone.utc) - timedelta(seconds=1)
    db.commit()
    expired = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json=body,
    )
    assert expired.status_code == 409
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )

    db.expire_all()
    row = db.get(MaintenanceProjectWorkbookValidation, plan["validation_token"])
    assert row.status == "expired"
    assert row.plan_json is None
    fresh_plan = _validate(client, project_id, content).json()
    body = {
        "validation_token": fresh_plan["validation_token"],
        "data_version": fresh_plan["data_version"],
    }
    applied = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json=body,
    )
    assert applied.status_code == 200, applied.text
    replay = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json=body,
    )
    assert replay.status_code == 409
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 1
    )


def test_validate_requires_exactly_one_xlsx_file(db):
    client = _client(db, username="workbook_upload_admin")
    project_id, _contract = _project_and_contract(client, db, suffix="upload")

    wrong_extension = _validate(client, project_id, b"content", name="update.xls")
    assert wrong_extension.status_code == 400
    multiple = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/validate",
        files=[
            ("file", ("one.xlsx", b"one", "application/octet-stream")),
            ("file", ("two.xlsx", b"two", "application/octet-stream")),
        ],
    )
    assert multiple.status_code == 400 or multiple.status_code == 422


@pytest.mark.parametrize(
    ("case_name", "permission_overrides", "path_kind"),
    [
        ("no-page", {"page_maintenance": False}, "export"),
        ("no-cost", {"data_purchase_cost": False}, "export"),
        ("no-profit", {"data_profit": False}, "export"),
        (
            "no-action",
            {"action_maintenance_roundtrip_apply": False},
            "validate",
        ),
    ],
)
def test_workbook_endpoints_fail_closed_by_permission(
    db,
    monkeypatch,
    case_name,
    permission_overrides,
    path_kind,
):
    admin = _client(db, username=f"workbook_permission_seed_{case_name}")
    project_id, _contract = _project_and_contract(
        admin,
        db,
        suffix=f"perm-{case_name}",
    )
    permissions = {
        "page_maintenance": True,
        "data_customer": True,
        "data_purchase_cost": True,
        "data_profit": True,
        "action_maintenance_roundtrip_apply": True,
        **permission_overrides,
    }
    client = _permission_client(
        db,
        username=f"workbook_permission_{case_name}",
        permissions=permissions,
    )
    if path_kind == "export":
        response = client.get(
            f"/api/maintenance/projects/stable/{project_id}/workbook"
        )
        assert response.status_code == 403
        return

    upload_read = False

    async def fail_if_upload_is_read(_request):
        nonlocal upload_read
        upload_read = True
        raise AssertionError("unauthorized upload reached the parser")

    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_parse_and_save_roundtrip_upload",
        fail_if_upload_is_read,
    )
    response = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/validate",
        files={
            "file": (
                "update.xlsx",
                b"not-read",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 403
    assert upload_read is False


def test_project_workbook_export_validate_apply_and_error_download_require_customer_before_side_effects(
    db,
    monkeypatch,
):
    admin = _client(db, username="workbook_customer_permission_seed")
    project_id, _contract = _project_and_contract(
        admin,
        db,
        suffix="customer-permission",
    )
    client = _permission_client(
        db,
        username="workbook_without_customer_permission",
        permissions={
            "page_maintenance": True,
            "data_customer": False,
            "data_purchase_cost": True,
            "data_profit": True,
            "action_maintenance_roundtrip_apply": True,
        },
    )
    state = db.get(MaintenanceProjectWorkbookState, project_id)
    assert state is not None
    before = {
        "revision": state.revision,
        "data_version": state.data_version,
        "last_export_id": state.last_export_id,
        "operations": db.scalar(
            select(func.count()).select_from(MaintenanceProjectWorkbookOperation)
        ),
        "validations": db.scalar(
            select(func.count()).select_from(MaintenanceProjectWorkbookValidation)
        ),
        "collections": db.scalar(
            select(func.count()).select_from(MaintenanceCollectionSnapshot)
        ),
    }
    upload_read = False

    async def fail_if_upload_is_read(_request):
        nonlocal upload_read
        upload_read = True
        raise AssertionError("unauthorized upload reached the parser")

    monkeypatch.setattr(
        maintenance_project_workbooks,
        "_parse_and_save_roundtrip_upload",
        fail_if_upload_is_read,
    )

    responses = [
        client.get(f"/api/maintenance/projects/stable/{project_id}/workbook"),
        client.post(
            f"/api/maintenance/projects/stable/{project_id}/workbook/validate",
            files={
                "file": (
                    "must-not-be-read.xlsx",
                    b"must-not-be-read",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        ),
        client.post(
            f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
            json={
                "validation_token": "must-not-be-read",
                "data_version": state.data_version,
            },
        ),
        client.get(
            "/api/maintenance/workbook-validations/must-not-be-read/errors.xlsx"
        ),
    ]

    assert [response.status_code for response in responses] == [403, 403, 403, 403]
    assert upload_read is False
    db.expire_all()
    unchanged_state = db.get(MaintenanceProjectWorkbookState, project_id)
    after = {
        "revision": unchanged_state.revision,
        "data_version": unchanged_state.data_version,
        "last_export_id": unchanged_state.last_export_id,
        "operations": db.scalar(
            select(func.count()).select_from(MaintenanceProjectWorkbookOperation)
        ),
        "validations": db.scalar(
            select(func.count()).select_from(MaintenanceProjectWorkbookValidation)
        ),
        "collections": db.scalar(
            select(func.count()).select_from(MaintenanceCollectionSnapshot)
        ),
    }
    assert after == before


def test_workbook_export_requires_authentication(db):
    admin = _client(db, username="workbook_anonymous_seed_admin")
    project_id, _contract = _project_and_contract(
        admin,
        db,
        suffix="perm-anonymous",
    )
    app = FastAPI()
    app.include_router(maintenance_project_workbooks.router, prefix="/api")
    response = TestClient(app).get(
        f"/api/maintenance/projects/stable/{project_id}/workbook"
    )
    assert response.status_code == 401


def test_stale_plan_and_same_file_replay_fail_closed(db):
    client = _client(db, username="workbook_stale_admin")
    project_id, contract = _project_and_contract(client, db, suffix="stale")
    content = _append_collection(
        _download(client, project_id),
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
    )
    first_plan = _validate(client, project_id, content).json()
    cloned_plan = _validate(client, project_id, content).json()

    changed = client.post(
        f"/api/maintenance/projects/stable/{project_id}/collections",
        json={
            "project_contract_id": contract["project_contract_id"],
            "report_month": "2026-07-01",
            "cumulative_amount": "100.00",
            "status": "confirmed",
            "reason": "制造服务端 revision 变化",
        },
    )
    assert changed.status_code == 201, changed.text
    stale = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": first_plan["validation_token"],
            "data_version": first_plan["data_version"],
        },
    )
    assert stale.status_code == 409
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 1
    )

    # A fresh export is required after any project revision change.
    fresh_content = _append_collection(
        _download(client, project_id),
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
    )
    applied_plan = _validate(client, project_id, fresh_content).json()
    duplicate_plan = _validate(client, project_id, fresh_content).json()
    applied = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": applied_plan["validation_token"],
            "data_version": applied_plan["data_version"],
        },
    )
    assert applied.status_code == 200, applied.text
    replay = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": duplicate_plan["validation_token"],
            "data_version": applied.json()["data_version"],
        },
    )
    assert replay.status_code == 409
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 2
    )
    assert cloned_plan["validation_token"] != first_plan["validation_token"]


@pytest.mark.parametrize("master_change", ["rename", "manager", "archive"])
def test_project_master_change_invalidates_a_validated_workbook_plan(
    db,
    master_change,
):
    client = _client(db, username=f"workbook_master_{master_change}_admin")
    project_id, contract = _project_and_contract(
        client,
        db,
        suffix=f"master-{master_change}",
    )
    content = _append_collection(
        _download(client, project_id),
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
    )
    plan = _validate(client, project_id, content).json()
    before_revision = db.get(MaintenanceProjectWorkbookState, project_id).revision

    if master_change == "archive":
        changed = client.post(
            f"/api/maintenance/projects/stable/{project_id}/archive",
            json={"version": 1, "reason": "验证归档使旧工作簿计划失效"},
        )
    else:
        field = (
            {"display_name": "校验后更正的项目名称"}
            if master_change == "rename"
            else {"project_manager_id": "manager-after-validation"}
        )
        changed = client.patch(
            f"/api/maintenance/projects/stable/{project_id}",
            json={
                "version": 1,
                **field,
                "reason": "验证主档变化使旧工作簿计划失效",
            },
        )
    assert changed.status_code == 200, changed.text
    db.expire_all()
    assert (
        db.get(MaintenanceProjectWorkbookState, project_id).revision
        == before_revision + 1
    )

    stale = client.post(
        f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
        json={
            "validation_token": plan["validation_token"],
            "data_version": plan["data_version"],
        },
    )
    assert stale.status_code == 409, stale.text
    assert (
        db.scalar(
            select(func.count())
            .select_from(MaintenanceCollectionSnapshot)
            .where(MaintenanceCollectionSnapshot.project_id == project_id)
        )
        == 0
    )


def test_archived_project_rejects_new_collection(db):
    client = _client(db, username="workbook_archived_collection_admin")
    project_id, contract = _project_and_contract(
        client,
        db,
        suffix="archived-collection",
    )
    archived = client.post(
        f"/api/maintenance/projects/stable/{project_id}/archive",
        json={"version": 1, "reason": "项目结束后禁止新增回款"},
    )
    assert archived.status_code == 200, archived.text

    response = client.post(
        f"/api/maintenance/projects/stable/{project_id}/collections",
        json={
            "project_contract_id": contract["project_contract_id"],
            "report_month": "2026-08-01",
            "cumulative_amount": "320.00",
            "status": "confirmed",
            "reason": "归档后误录回款",
        },
    )
    assert response.status_code == 400, response.text
    assert "归档" in response.json()["detail"]
    assert (
        db.scalar(
            select(func.count())
            .select_from(MaintenanceCollectionSnapshot)
            .where(MaintenanceCollectionSnapshot.project_id == project_id)
        )
        == 0
    )


def test_unexpected_second_row_failure_rolls_back_first_row_and_ledgers(
    db, monkeypatch
):
    client = _client(db, username="workbook_atomic_admin")
    project_id, contract = _project_and_contract(client, db, suffix="atomic")
    content = _append_collection(
        _download(client, project_id),
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
        report_month="2026-07",
        amount="100.00",
    )
    content = _append_collection(
        content,
        project_contract_id=contract["project_contract_id"],
        contract_no=contract["contract_no"],
        report_month="2026-08",
        amount="200.00",
    )
    plan = _validate(client, project_id, content).json()
    before_revision = db.get(MaintenanceProjectWorkbookState, project_id).revision
    original_create = maintenance_project_workbook_adapter.operations.create_collection
    calls = 0

    def fail_on_second(*args, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("synthetic second-row failure")
        return original_create(*args, **kwargs)

    monkeypatch.setattr(
        maintenance_project_workbook_adapter.operations,
        "create_collection",
        fail_on_second,
    )
    with pytest.raises(RuntimeError, match="second-row failure"):
        client.post(
            f"/api/maintenance/projects/stable/{project_id}/workbook/apply",
            json={
                "validation_token": plan["validation_token"],
                "data_version": plan["data_version"],
            },
        )

    db.expire_all()
    assert (
        db.scalar(select(func.count()).select_from(MaintenanceCollectionSnapshot)) == 0
    )
    assert (
        db.get(MaintenanceProjectWorkbookState, project_id).revision == before_revision
    )
    assert (
        db.get(MaintenanceProjectWorkbookValidation, plan["validation_token"]).status
        == "valid"
    )
    applied_ledgers = db.scalar(
        select(func.count())
        .select_from(MaintenanceProjectWorkbookOperation)
        .where(
            MaintenanceProjectWorkbookOperation.operation_type.in_(
                ["collection_create", "file_apply"]
            )
        )
    )
    assert applied_ledgers == 0
