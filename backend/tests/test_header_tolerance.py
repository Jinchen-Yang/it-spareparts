"""列名容差归一：氚云非标导出视图让列丢 (必填) 注解（实测 #25 采购订单整文件 empty_pn）。"""
from datetime import date
from decimal import Decimal

import openpyxl
import pandas as pd
import pytest
from sqlalchemy import select

from app.etl import mapping, pipeline, reader, transform
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder
from app.models.purchase import FPurchaseLine, FPurchaseOrder
from app.models.sales import FSalesLine, FSalesOrder


def test_canonicalize_columns():
    # (必填) 容差：裸列名归一到规范键
    assert mapping.canonicalize_columns(["采购单号", "明细.产品名称", "供应商编码#"]) == \
        ["采购单号(必填)", "明细.产品名称(必填)", "供应商编码#"]
    # 标准导出零改动（已是规范键，走精确分支）
    assert mapping.canonicalize_columns(["明细.产品名称(必填)"]) == ["明细.产品名称(必填)"]
    # (PN) 等语义后缀不被剥
    assert mapping.canonicalize_columns(["产品名称(PN)"]) == ["产品名称(PN)"]


def test_canonicalize_no_duplicate_columns():
    # 裸名与规范名并存 → 只认规范名、不造重复列
    out = mapping.canonicalize_columns(["明细.产品名称", "明细.产品名称(必填)"])
    assert out.count("明细.产品名称(必填)") == 1
    # 产品名称 与 产品名称#（#25 实况：两列并存）→ 不冲突，# 列原样保留
    out = mapping.canonicalize_columns(["明细.产品名称", "明细.产品名称#"])
    assert out.count("明细.产品名称(必填)") == 1 and "明细.产品名称#" in out


def test_purchase_tax_marker_alias_is_not_shadowed_by_expense_loose_column():
    columns = mapping.canonicalize_columns(["是否含税(选填)"])
    frame = reader._coalesce_value_aliases(
        pd.DataFrame([["是"]], columns=columns),
        mapping.PURCHASE,
    )

    assert list(frame.columns) == ["是否含税(必填)"]
    assert frame["是否含税(必填)"].iloc[0] == "是"


def test_sales_tax_block_is_resolved_within_sales_document_type(tmp_path):
    """销售税字段不能被采购同名 canonical key 截走并静默落成 NULL。"""

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号(必填)", "数据ID(不可修改)", "订单日期(必填)",
        "销售人员(必填)", "客户名称", "业务类型#", "订单金额",
        "是否含税(必填)", "税率(必填)", "税金", "不含税金额", "数据状态",
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    ws.append([f"F{i:07d}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "XSDD-20240708-0093", "RAW-SALES-TAX", "2024-07-08",
        "王雪菲", "测试客户", "算力运维", 31_440_000,
        "含税", "13.0%", Decimal("3616991.15"), Decimal("27823008.85"),
        "已生效", "RAW-SALES-TAX-L1", "TEST-PN", 1, 31_440_000, 31_440_000,
    ])
    path = tmp_path / "sales-tax-block.xlsx"
    wb.save(path)

    frame, file_type = reader.read_excel(str(path))
    result = transform.transform(frame, file_type)
    order = result.orders["RAW-SALES-TAX"]

    assert file_type == mapping.SALES
    assert "税率" in frame.columns and "税率(必填)" not in frame.columns
    assert "是否含税" in frame.columns and "是否含税(必填)" not in frame.columns
    assert order["amount_inc_tax"] == Decimal("31440000.00")
    assert order["amount_ex_tax"] == Decimal("27823008.85")
    assert order["tax_amount"] == Decimal("3616991.15")
    assert order["tax_rate"] == Decimal("0.1300")
    assert order["is_tax_inclusive"] is True


def test_detect_file_type_tolerant():
    assert mapping.detect_file_type(["采购单号", "明细.产品名称"]) == mapping.PURCHASE
    assert mapping.detect_file_type(["采购单号(必填)"]) == mapping.PURCHASE
    assert mapping.detect_file_type(["订单编号", "业务类型"]) == mapping.SALES        # 无 必填/# 也认
    assert mapping.detect_file_type(["订单编号(必填)", "业务类型#"]) == mapping.SALES
    assert mapping.detect_file_type(["产品库存ID"]) == mapping.INVENTORY
    assert mapping.detect_file_type(["库存数量", "产品名称(PN)"]) == mapping.INVENTORY
    assert mapping.detect_file_type(["乱七八糟", "随便"]) is None


def _purchase_xlsx_bare(tmp_path):
    """双表头采购文件，但中文列名是「裸」的（无 (必填)）——复刻 #25 故障形态。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([f"F000000{i}" for i in range(1, 8)])              # 第0行 F码 → 双表头
    ws.append(["采购单号", "数据ID(不可修改)", "明细.数据ID(不可修改)",
               "明细.产品名称", "明细.采购数量", "明细.单价", "供应商"])  # 第1行 裸中文名
    ws.append(["CGDD-1", "O1", "L1", "ST8000NM000A", "2", "100", "测试供应商"])
    p = tmp_path / "purchase_bare.xlsx"
    wb.save(p)
    return str(p)


def test_reader_canonicalizes_bare_headers(tmp_path):
    df, ft = reader.read_excel(_purchase_xlsx_bare(tmp_path))
    assert ft == mapping.PURCHASE
    # 裸列名已归一为规范键，PN 列有值（不再整列空 → 不会全行 empty_pn）
    assert "明细.产品名称(必填)" in df.columns
    assert "采购单号(必填)" in df.columns
    assert df["明细.产品名称(必填)"].iloc[0] == "ST8000NM000A"


@pytest.mark.parametrize(
    ("identity_column", "identities"),
    [
        ("订单编号(必填)", ["XSDD-A", "XSDD-A", "XSDD-B", None]),
        ("数据ID(不可修改)", ["RAW-A", "RAW-A", "RAW-B", None]),
    ],
)
def test_header_ffill_supports_a_single_identity_column(identity_column, identities):
    df = pd.DataFrame({
        identity_column: identities,
        "业务类型#": ["备件销售", None, None, None],
    })

    filled = reader._ffill_head_columns(df, mapping.SALES)

    assert filled["业务类型#"].iloc[:2].tolist() == ["备件销售", "备件销售"]
    assert filled["业务类型#"].iloc[2:].isna().all()


@pytest.mark.parametrize(
    ("raw_ids", "order_nos"),
    [
        ([None, "RAW-A", None], ["XSDD-A", "XSDD-A", None]),
        (["RAW-A", "RAW-A", None], [None, "XSDD-A", None]),
    ],
)
def test_header_ffill_supplements_identity_within_current_group(raw_ids, order_nos):
    df = pd.DataFrame({
        "数据ID(不可修改)": raw_ids,
        "订单编号(必填)": order_nos,
        "业务类型#": ["备件销售", None, None],
    })

    filled = reader._ffill_head_columns(df, mapping.SALES)

    assert filled["业务类型#"].tolist() == ["备件销售", "备件销售", "备件销售"]


@pytest.mark.parametrize(
    ("raw_ids", "order_nos", "must_stay_empty"),
    [
        ([None, "RAW-B", None], ["XSDD-A", None, None], "订单编号(必填)"),
        (["RAW-A", None, None], [None, "XSDD-B", None], "数据ID(不可修改)"),
    ],
)
def test_header_ffill_does_not_join_complementary_unproven_identities(
    raw_ids, order_nos, must_stay_empty,
):
    df = pd.DataFrame({
        "数据ID(不可修改)": raw_ids,
        "订单编号(必填)": order_nos,
        "业务类型#": ["备件销售", None, None],
    })

    filled = reader._ffill_head_columns(df, mapping.SALES)

    assert filled["业务类型#"].iloc[0] == "备件销售"
    assert filled["业务类型#"].iloc[1:].isna().all()
    assert filled[must_stay_empty].iloc[1:].isna().all()


def _sales_xlsx(tmp_path, business_type_header, double_header, raw_order_id, business_type):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号", "数据ID(不可修改)", business_type_header,
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    if double_header:
        ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        f"XSDD-{raw_order_id}", raw_order_id, business_type,
        f"{raw_order_id}-LINE", "ST8000NM000A", "1", "100", "100",
    ])
    path = tmp_path / f"{raw_order_id}.xlsx"
    wb.save(path)
    return str(path)


def _sales_xlsx_with_maintenance_period_headers(
    tmp_path,
    *,
    period_from_header: str,
    period_to_header: str,
    raw_order_id: str,
) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号", "数据ID(不可修改)", "业务类型#",
        period_from_header, period_to_header,
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        f"XSDD-{raw_order_id}", raw_order_id, "备件维保",
        "2026-01-01", "2026-12-31",
        f"{raw_order_id}-LINE", "ST8000NM000A", "1", "100", "100",
    ])
    path = tmp_path / f"{raw_order_id}-period.xlsx"
    wb.save(path)
    return str(path)


def _sales_xlsx_with_business_type_aliases(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号", "数据ID(不可修改)", "业务类型#", "业务类型",
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "XSDD-SALES-A", "SALES-A", "标准A", "裸A",
        "SALES-A-LINE", "ST8000NM000A", "1", "100", "100",
    ])
    ws.append([
        "XSDD-SALES-B", "SALES-B", None, "裸B",
        "SALES-B-LINE-1", "ST8000NM000B", "1", "200", "200",
    ])
    ws.append([
        None, None, None, None,
        "SALES-B-LINE-2", "ST8000NM000C", "1", "300", "300",
    ])
    path = tmp_path / "sales-business-type-aliases.xlsx"
    wb.save(path)
    return str(path)


def _sales_xlsx_with_empty_new_order_business_type(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号", "数据ID(不可修改)", "业务类型#", "业务类型",
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "XSDD-SALES-A", "SALES-A", "标准A", None,
        "SALES-A-LINE", "ST8000NM000A", "1", "100", "100",
    ])
    ws.append([
        "XSDD-SALES-B", "SALES-B", None, None,
        "SALES-B-LINE-1", "ST8000NM000B", "1", "200", "200",
    ])
    ws.append([
        None, None, None, None,
        "SALES-B-LINE-2", "ST8000NM000C", "1", "300", "300",
    ])
    path = tmp_path / "sales-empty-new-order-business-type.xlsx"
    wb.save(path)
    return str(path)


def _sales_xlsx_with_repeated_order_identity(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号", "数据ID(不可修改)", "业务类型#", "业务类型", "销售人员", "仓库",
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "XSDD-SALES-REPEATED", "SALES-REPEATED", "备件销售", None, "销售A", "销售仓",
        "SALES-REPEATED-LINE-1", "ST8000NM000A", "1", "100", "100",
    ])
    ws.append([
        "XSDD-SALES-REPEATED", "SALES-REPEATED", None, None, None, None,
        "SALES-REPEATED-LINE-2", "ST8000NM000B", "1", "200", "200",
    ])
    path = tmp_path / "sales-repeated-order-identity.xlsx"
    wb.save(path)
    return str(path)


def _sales_xlsx_with_recovered_raw_identity(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "订单编号", "数据ID(不可修改)", "业务类型#",
        "订单明细.数据ID(不可修改)", "订单明细.产品名称",
        "订单明细.订单数量", "订单明细.单价", "订单明细.金额",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "XSDD-SALES-RECOVERED", None, "备件销售",
        None, None, None, None, None,
    ])
    ws.append([
        "XSDD-SALES-RECOVERED", "SALES-RECOVERED", None,
        "SALES-RECOVERED-LINE", "ST8000NM000A", "1", "100", "100",
    ])
    path = tmp_path / "sales-recovered-raw-identity.xlsx"
    wb.save(path)
    return str(path)


def _purchase_xlsx_with_continuation(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "采购单号(必填)", "数据ID(不可修改)", "采购人员(必填)",
        "供应商(必填)", "批量采购(必填)",
        "明细.数据ID(不可修改)", "明细.产品名称(必填)",
        "明细.采购数量(必填)", "明细.单价(必填)", "明细.合计金额",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "CGDD-STATE", "PURCHASE-STATE", "采购A", "供应商A", "销售订单",
        "PURCHASE-STATE-LINE-1", "ST8000NM000A", "1", "100", "100",
    ])
    ws.append([
        None, None, None, None, None,
        "PURCHASE-STATE-LINE-2", "ST8000NM000B", "2", "200", "400",
    ])
    path = tmp_path / "purchase-state-machine.xlsx"
    wb.save(path)
    return str(path)


def _maintenance_xlsx_with_continuation(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "需求单号", "数据ID(不可修改)", "需求类型", "业务类型",
        "销售人员", "出库仓库(必填)", "维保起始日期", "维保终止日期",
        "需求明细.数据ID(不可修改)", "需求明细.需供货产品",
        "需求明细.需求数量", "需求明细.退货数量",
    ]
    ws.append([f"F000000{i}" for i in range(1, len(headers) + 1)])
    ws.append(headers)
    ws.append([
        "WBDD-STATE", "MAINTENANCE-STATE", "报修供货", "备件维保",
        "销售A", "总仓", "2026-01-01", "2026-12-31",
        "MAINTENANCE-STATE-LINE-1", "ST8000NM000A", "1", "0",
    ])
    ws.append([
        None, None, None, None, None, None, None, None,
        "MAINTENANCE-STATE-LINE-2", "ST8000NM000B", "2", "1",
    ])
    path = tmp_path / "maintenance-state-machine.xlsx"
    wb.save(path)
    return str(path)


@pytest.mark.parametrize(
    ("business_type_header", "double_header", "raw_order_id", "business_type"),
    [
        ("业务类型", False, "SALES-BARE", "备件销售"),
        ("业务类型#", True, "SALES-TRITIUM", "整机销售"),
    ],
)
def test_sales_business_type_survives_transform_and_load(
    db, tmp_path, business_type_header, double_header, raw_order_id, business_type,
):
    path = _sales_xlsx(
        tmp_path, business_type_header, double_header, raw_order_id, business_type,
    )
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.SALES
    assert "业务类型#" in df.columns
    assert "业务类型" not in df.columns
    assert df["业务类型#"].tolist() == [business_type]
    assert result.orders[raw_order_id]["business_type"] == business_type

    pipeline.run_import(db, path, f"{raw_order_id}.xlsx")
    loaded = db.scalar(
        select(FSalesOrder).where(FSalesOrder.raw_order_id == raw_order_id)
    )
    assert loaded is not None
    assert loaded.business_type == business_type


@pytest.mark.parametrize(
    ("period_from_header", "period_to_header", "raw_order_id"),
    [
        ("维保起始日期", "维保终止日期", "SALES-PERIOD-BARE"),
        (
            "维保起始日期(必填)",
            "维保终止日期(必填)",
            "SALES-PERIOD-REQUIRED",
        ),
    ],
)
def test_sales_maintenance_period_header_variants_survive_transform(
    tmp_path,
    period_from_header,
    period_to_header,
    raw_order_id,
):
    path = _sales_xlsx_with_maintenance_period_headers(
        tmp_path,
        period_from_header=period_from_header,
        period_to_header=period_to_header,
        raw_order_id=raw_order_id,
    )
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.SALES
    assert "维保起始日期(必填)" in df.columns
    assert "维保终止日期(必填)" in df.columns
    assert "维保起始日期" not in df.columns
    assert "维保终止日期" not in df.columns
    assert result.orders[raw_order_id]["maintenance_period_from"] == "2026-01-01"
    assert result.orders[raw_order_id]["maintenance_period_to"] == "2026-12-31"


def test_sales_business_type_aliases_coalesce_before_ffill(db, tmp_path):
    path = _sales_xlsx_with_business_type_aliases(tmp_path)
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.SALES
    assert "业务类型" not in df.columns
    assert df["业务类型#"].tolist() == ["标准A", "裸B", "裸B"]
    assert result.orders["SALES-A"]["business_type"] == "标准A"
    assert result.orders["SALES-B"]["business_type"] == "裸B"

    pipeline.run_import(db, path, "sales-business-type-aliases.xlsx")
    loaded = dict(db.execute(
        select(FSalesOrder.raw_order_id, FSalesOrder.business_type).where(
            FSalesOrder.raw_order_id.in_(["SALES-A", "SALES-B"])
        )
    ).all())
    assert loaded == {"SALES-A": "标准A", "SALES-B": "裸B"}


def test_sales_business_type_does_not_ffill_across_orders(db, tmp_path):
    path = _sales_xlsx_with_empty_new_order_business_type(tmp_path)
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.SALES
    assert df["业务类型#"].iloc[0] == "标准A"
    assert df["业务类型#"].iloc[1:].isna().all()
    assert result.orders["SALES-A"]["business_type"] == "标准A"
    assert result.orders["SALES-B"]["business_type"] is None

    pipeline.run_import(db, path, "sales-empty-new-order-business-type.xlsx")
    loaded = dict(db.execute(
        select(FSalesOrder.raw_order_id, FSalesOrder.business_type).where(
            FSalesOrder.raw_order_id.in_(["SALES-A", "SALES-B"])
        )
    ).all())
    assert loaded == {"SALES-A": "标准A", "SALES-B": None}


def test_sales_business_type_ffills_when_order_identity_repeats(db, tmp_path):
    path = _sales_xlsx_with_repeated_order_identity(tmp_path)
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.SALES
    assert df["业务类型#"].tolist() == ["备件销售", "备件销售"]
    assert len(result.orders) == 1
    assert len(result.lines) == 2
    assert result.orders["SALES-REPEATED"]["business_type"] == "备件销售"
    assert result.orders["SALES-REPEATED"]["salesperson"] == "销售A"
    assert result.orders["SALES-REPEATED"]["warehouse"] == "销售仓"

    pipeline.run_import(db, path, "sales-repeated-order-identity.xlsx")
    loaded = db.scalar(
        select(FSalesOrder).where(FSalesOrder.raw_order_id == "SALES-REPEATED")
    )
    assert loaded is not None
    assert loaded.business_type == "备件销售"
    assert loaded.salesperson == "销售A"
    assert loaded.warehouse == "销售仓"
    assert len(db.scalars(
        select(FSalesLine).where(FSalesLine.order_id == loaded.id)
    ).all()) == 2


def test_sales_identity_can_recover_without_splitting_order(db, tmp_path):
    path = _sales_xlsx_with_recovered_raw_identity(tmp_path)
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.SALES
    assert df["业务类型#"].tolist() == ["备件销售", "备件销售"]
    assert len(result.lines) == 1
    assert result.orders["SALES-RECOVERED"]["business_type"] == "备件销售"

    pipeline.run_import(db, path, "sales-recovered-raw-identity.xlsx")
    loaded = db.scalar(
        select(FSalesOrder).where(FSalesOrder.raw_order_id == "SALES-RECOVERED")
    )
    assert loaded is not None
    assert loaded.business_type == "备件销售"
    assert len(db.scalars(
        select(FSalesLine).where(FSalesLine.order_id == loaded.id)
    ).all()) == 1


def test_purchase_continuation_preserves_head_and_detail_count(db, tmp_path):
    path = _purchase_xlsx_with_continuation(tmp_path)
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.PURCHASE
    assert len(result.orders) == 1
    assert len(result.lines) == 2
    assert result.orders["PURCHASE-STATE"]["purchaser"] == "采购A"
    assert result.orders["PURCHASE-STATE"]["source_type_raw"] == "销售订单"

    pipeline.run_import(db, path, "purchase-state-machine.xlsx")
    loaded = db.scalar(
        select(FPurchaseOrder).where(
            FPurchaseOrder.raw_order_id == "PURCHASE-STATE"
        )
    )
    assert loaded is not None
    assert loaded.purchaser == "采购A"
    assert loaded.source_type_raw == "销售订单"
    assert len(db.scalars(
        select(FPurchaseLine).where(FPurchaseLine.order_id == loaded.id)
    ).all()) == 2


def test_maintenance_continuation_preserves_head_and_detail_count(db, tmp_path):
    path = _maintenance_xlsx_with_continuation(tmp_path)
    df, file_type = reader.read_excel(path)
    result = transform.transform(df, file_type)

    assert file_type == mapping.MAINTENANCE
    assert len(result.orders) == 1
    assert len(result.lines) == 2
    assert result.orders["MAINTENANCE-STATE"]["demand_type"] == "报修供货"
    assert result.orders["MAINTENANCE-STATE"]["business_type"] == "备件维保"
    assert result.orders["MAINTENANCE-STATE"]["salesperson"] == "销售A"
    assert result.orders["MAINTENANCE-STATE"]["warehouse"] == "总仓"
    assert result.orders["MAINTENANCE-STATE"]["maint_start"] == date(2026, 1, 1)
    assert result.orders["MAINTENANCE-STATE"]["maint_end"] == date(2026, 12, 31)

    pipeline.run_import(db, path, "maintenance-state-machine.xlsx")
    loaded = db.scalar(
        select(FMaintenanceOrder).where(
            FMaintenanceOrder.raw_order_id == "MAINTENANCE-STATE"
        )
    )
    assert loaded is not None
    assert loaded.demand_type == "报修供货"
    assert loaded.business_type == "备件维保"
    assert loaded.salesperson == "销售A"
    assert loaded.warehouse == "总仓"
    assert loaded.maint_start == date(2026, 1, 1)
    assert loaded.maint_end == date(2026, 12, 31)
    assert len(db.scalars(
        select(FMaintenanceLine).where(FMaintenanceLine.order_id == loaded.id)
    ).all()) == 2
