"""单项目/单合同下载必须区分对象不存在和日期范围零命中。"""

from __future__ import annotations

import io
from datetime import date

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app import permissions, security
from app.agent import tools
from app.auth import hash_password
from app.main import app
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder
from app.models.system import SysUser
from app.services import (
    maintenance_cost,
    maintenance_margin_evidence,
    maintenance_roundtrip,
    maintenance_workbook_export,
)
from tests.test_maintenance_export_headers import _admin_client
from tests.test_maintenance_roundtrip import _seed_contract


def _scoped_sales_client(db, *, salesperson_name: str) -> TestClient:
    base = permissions.effective("sales", None)
    overrides = {"page_maintenance": True, "own_customers_only": True}
    effective = permissions.effective_from_snapshot(base, overrides)
    db.add(SysUser(
        username="maintenance_scoped_sales",
        role="sales",
        salesperson_name=salesperson_name,
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="sales",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=effective,
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_scoped_sales", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def test_single_project_csv_returns_404_for_unknown_project(db):
    client = _admin_client(db)

    response = client.get(
        "/api/maintenance/lines/export",
        params={"project": "不存在项目"},
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "项目不存在：不存在项目"


def test_single_project_csv_returns_422_when_project_exists_but_range_is_empty(db):
    _seed_contract(db, suffix="project-range", contract="XSDD-PROJECT-RANGE")
    client = _admin_client(db)

    response = client.get(
        "/api/maintenance/lines/export",
        params={
            "project": "回填项目-project-range",
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
        },
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "项目存在，但所选范围内没有可导出的明细"


@pytest.mark.parametrize("object_state", ["before_start", "inactive", "zero_lines"])
def test_single_project_csv_visible_object_without_current_cost_lines_is_422(
    db,
    object_state,
):
    order_id, line_id = _seed_contract(
        db,
        suffix=f"project-{object_state}",
        contract=f"XSDD-PROJECT-{object_state}",
    )
    order = db.get(FMaintenanceOrder, order_id)
    assert order is not None
    if object_state == "before_start":
        order.order_date = date(2023, 12, 31)
    elif object_state == "inactive":
        order.data_status = "已取消"
    else:
        line = db.get(FMaintenanceLine, line_id)
        assert line is not None
        db.delete(line)
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/lines/export",
        params={"project": f"回填项目-project-{object_state}"},
    )

    assert response.status_code == 422
    assert response.json()["detail"] == "项目存在，但所选范围内没有可导出的明细"


def test_scoped_sales_project_existence_and_count_use_same_visible_row_scope(db):
    _seed_contract(db, suffix="own-project", contract="XSDD-OWN")
    _seed_contract(db, suffix="other-project", contract="XSDD-OTHER")
    own_order = db.scalar(select(FMaintenanceOrder).where(
        FMaintenanceOrder.linked_sales_order_no == "XSDD-OWN",
    ))
    other_order = db.scalar(select(FMaintenanceOrder).where(
        FMaintenanceOrder.linked_sales_order_no == "XSDD-OTHER",
    ))
    own_order.salesperson = "当前销售"
    other_order.salesperson = "其他销售"
    db.commit()
    client = _scoped_sales_client(db, salesperson_name="当前销售")

    own = client.get(
        "/api/maintenance/lines/export",
        params={"project": "回填项目-own-project"},
    )
    invisible = client.get(
        "/api/maintenance/lines/export",
        params={"project": "回填项目-other-project"},
    )
    projects = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    project_csv = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )

    assert own.status_code == 200, own.text
    assert invisible.status_code == 404
    assert invisible.json()["detail"] == "项目不存在：回填项目-other-project"
    assert projects.status_code == 200, projects.text
    assert [row["project"] for row in projects.json()["rows"]] == [
        "回填项目-own-project",
    ]
    assert project_csv.status_code == 200, project_csv.text
    decoded = project_csv.content.decode("utf-8-sig")
    assert "回填项目-own-project" in decoded
    assert "回填项目-other-project" not in decoded

    hidden_contract = client.get(
        "/api/maintenance/board",
        params={"lifecycle": "all", "q": "XSDD-OTHER"},
    )
    unknown_contract = client.get(
        "/api/maintenance/board",
        params={"lifecycle": "all", "q": "XSDD-NOT-FOUND"},
    )
    assert hidden_contract.status_code == unknown_contract.status_code == 403
    assert hidden_contract.json()["detail"] == unknown_contract.json()["detail"]


def test_scoped_sales_shared_contract_projects_never_load_or_return_full_revenue(
    db,
    monkeypatch,
):
    _seed_contract(db, suffix="shared-a", contract="XSDD-SHARED")
    _seed_contract(db, suffix="shared-b", contract="XSDD-SHARED")
    orders = db.scalars(
        select(FMaintenanceOrder)
        .where(FMaintenanceOrder.linked_sales_order_no == "XSDD-SHARED")
        .order_by(FMaintenanceOrder.id),
    ).all()
    assert len(orders) == 2
    orders[0].salesperson = "销售A"
    orders[1].salesperson = "销售B"
    db.commit()

    def forbidden_revenue_query(*_args, **_kwargs):
        raise AssertionError("scoped project facts must not load full contract revenue")

    monkeypatch.setattr(
        maintenance_margin_evidence,
        "load_contract_revenue_evidence",
        forbidden_revenue_query,
    )
    client = _scoped_sales_client(db, salesperson_name="销售A")

    projects = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    project_csv = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    ctx = security.UserContext(
        user_id="maintenance_scoped_sales",
        role="sales",
        salesperson_name="销售A",
        permissions={
            "page_maintenance": True,
            "own_customers_only": True,
            "data_purchase_cost": True,
            "data_profit": True,
        },
        is_authenticated=True,
        authn="sys_user",
        token_version=0,
    )
    agent_projects = tools.dispatch(
        db,
        "get_maintenance_projects",
        {},
        ctx,
    )

    assert projects.status_code == 200, projects.text
    assert [row["project"] for row in projects.json()["rows"]] == [
        "回填项目-shared-a",
    ]
    assert projects.json()["rows"][0]["contract_amount"] is None
    assert project_csv.status_code == 200, project_csv.text
    decoded = project_csv.content.decode("utf-8-sig")
    assert "回填项目-shared-a" in decoded
    assert "回填项目-shared-b" not in decoded
    assert "1130" not in decoded
    assert [row["project"] for row in agent_projects["rows"]] == [
        "回填项目-shared-a",
    ]
    assert agent_projects["rows"][0]["contract_amount"] is None


def test_scoped_sales_agent_board_rejects_before_service_query(
    db,
    monkeypatch,
):
    calls: list[str] = []

    def forbidden_board(*_args, **_kwargs):
        calls.append("called")
        raise AssertionError("scoped agent must reject before contract service")

    monkeypatch.setattr(maintenance_cost, "board", forbidden_board)
    _scoped_sales_client(db, salesperson_name="销售A")
    ctx = security.UserContext(
        user_id="maintenance_scoped_sales",
        role="sales",
        salesperson_name="销售A",
        permissions={
            "page_maintenance": True,
            "own_customers_only": True,
            "data_purchase_cost": True,
            "data_profit": True,
        },
        is_authenticated=True,
        authn="sys_user",
        token_version=0,
    )

    response = tools.dispatch(
        db,
        "get_maintenance_board",
        {"status": "red"},
        ctx,
    )

    # Capability policy rejects before the handler and uses the same non-enumerating response as
    # an unknown tool; the scoped account must not learn that a contract-level capability exists.
    assert response == {
        "error": "未知工具或无权限",
        "kind": "capability_denied",
    }
    assert calls == []


def test_scoped_sales_contract_routes_reject_before_any_object_builder(
    db,
    monkeypatch,
):
    client = _scoped_sales_client(db, salesperson_name="销售A")
    builder_calls: list[str] = []

    def forbidden_builder(*_args, **_kwargs):
        builder_calls.append("called")
        raise AssertionError("scoped contract route must reject before object query")

    monkeypatch.setattr(maintenance_cost, "board", forbidden_builder)
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbook_file",
        forbidden_builder,
    )
    monkeypatch.setattr(
        maintenance_workbook_export,
        "build_contract_workbooks_zip",
        forbidden_builder,
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "build_roundtrip_template",
        forbidden_builder,
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "build_roundtrip_template_bundle",
        forbidden_builder,
    )

    responses = [
        client.get("/api/maintenance/board", params={"q": "B对象"}),
        client.get("/api/maintenance/board/export", params={"q": "B对象"}),
        client.get(
            "/api/maintenance/export-workbook",
            params={"contract": "XSDD-B"},
        ),
        client.get("/api/maintenance/export-workbooks"),
        client.get(
            "/api/maintenance/roundtrip-template",
            params={"contract": "XSDD-B"},
        ),
        client.get("/api/maintenance/roundtrip-templates"),
    ]

    assert all(response.status_code == 403 for response in responses)
    assert len({response.json()["detail"] for response in responses}) <= 2
    assert builder_calls == []


def test_single_contract_downloads_return_404_for_unknown_contract(db):
    client = _admin_client(db)

    workbook = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-NOT-FOUND"},
    )
    roundtrip = client.get(
        "/api/maintenance/roundtrip-template",
        params={"contract": "XSDD-NOT-FOUND"},
    )

    assert workbook.status_code == 404
    assert workbook.json()["detail"] == "合同不存在：XSDD-NOT-FOUND"
    assert roundtrip.status_code == 404
    assert roundtrip.json()["detail"] == "合同不存在：XSDD-NOT-FOUND"


def test_unknown_contract_blank_roundtrip_template_remains_signed_and_structured(db):
    client = _admin_client(db)

    response = client.get(
        "/api/maintenance/roundtrip-template",
        params={"contract": "XSDD-NEW", "blank": True},
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["x-content-type-options"] == "nosniff"
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    try:
        assert workbook.sheetnames == list(maintenance_roundtrip.SHEET_NAMES)
        assert set(workbook["01_项目"].tables) == {"tbl_projects_v1"}
        assert set(workbook["02_维保订单"].tables) == {"tbl_orders_v1"}
        assert set(workbook["03_订单明细"].tables) == {"tbl_order_lines_v1"}
        metadata = {
            str(workbook["99_元数据"].cell(row=row, column=1).value): str(
                workbook["99_元数据"].cell(row=row, column=2).value or "",
            )
            for row in range(2, workbook["99_元数据"].max_row + 1)
        }
        assert metadata["contract_scope"] == "XSDD-NEW"
        assert metadata["template_mode"] == "blank"
        assert len(metadata["metadata_hmac"]) == 64
        assert metadata["metadata_hmac"] == maintenance_roundtrip._metadata_hmac(
            {
                key: value
                for key, value in metadata.items()
                if key != "metadata_hmac"
            },
        )
    finally:
        workbook.close()


def test_single_contract_downloads_return_422_when_range_is_empty(db):
    _seed_contract(db, suffix="contract-range", contract="XSDD-CONTRACT-RANGE")
    client = _admin_client(db)
    params = {
        "contract": "XSDD-CONTRACT-RANGE",
        "date_from": "2026-06-01",
        "date_to": "2026-06-30",
    }

    workbook = client.get("/api/maintenance/export-workbook", params=params)
    roundtrip = client.get("/api/maintenance/roundtrip-template", params=params)

    assert workbook.status_code == 422
    assert workbook.json()["detail"] == "合同存在，但所选范围内没有可导出的维保数据"
    assert roundtrip.status_code == 422
    assert roundtrip.json()["detail"] == "合同存在，但所选范围内没有可导出的维保数据"


def test_global_roundtrip_template_returns_422_when_scope_has_no_data(db):
    client = _admin_client(db)

    response = client.get("/api/maintenance/roundtrip-template")

    assert response.status_code == 422
    assert response.json()["detail"] == "所选范围内没有可导出的维保数据"


def test_project_summary_csv_and_orders_xlsx_return_422_for_empty_scope(db):
    client = _admin_client(db)

    projects = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    orders = client.get("/api/maintenance/orders/export")

    assert projects.status_code == 422
    assert projects.json()["detail"] == "所选范围内没有可导出的项目数据"
    assert orders.status_code == 422
    assert orders.json()["detail"] == "所选范围内没有可导出的维保订单"
