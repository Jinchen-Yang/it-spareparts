"""维保订单 XLSX 导出的公共接口契约。"""
import asyncio
import io
from datetime import date
from decimal import Decimal
from unittest.mock import Mock

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet
from openpyxl.worksheet._writer import ALL_TEMP_FILES
from sqlalchemy import select, text
from starlette.requests import ClientDisconnect

from app import config, permissions
from app.api import maintenance as maintenance_api
from app.auth import hash_password
from app.db import SessionLocal
from app.etl import loader
from app.main import app
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    MaintenanceManualCostOverride,
)
from app.models.system import SysImportBatch, SysUser
from app.security import UserContext
from app.services import maintenance_cost_quality, maintenance_export
from tests import factories as f


def _admin_client(db) -> TestClient:
    db.add(SysUser(
        username="maintenance_order_export_admin",
        role="admin",
        password_hash=hash_password("pw123456"),
        is_active=True,
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_order_export_admin", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _cost_blind_client(db) -> TestClient:
    base = permissions.effective("readonly", None)
    overrides = {"page_maintenance": True, "data_purchase_cost": False}
    db.add(SysUser(
        username="maintenance_order_export_cost_blind",
        role="readonly",
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="readonly",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_order_export_cost_blind", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _customer_blind_purchaser_client(db) -> TestClient:
    base = permissions.effective("purchaser", None)
    overrides = {"page_maintenance": True, "data_customer": False}
    db.add(SysUser(
        username="maintenance_order_export_customer_blind",
        role="purchaser",
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="purchaser",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_order_export_customer_blind", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _readonly_client(db) -> TestClient:
    db.add(SysUser(
        username="maintenance_order_export_readonly",
        role="readonly",
        password_hash=hash_password("pw123456"),
        is_active=True,
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_order_export_readonly", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _scoped_sales_with_maintenance_page_client(db) -> TestClient:
    base = permissions.effective("sales", None)
    overrides = {"page_maintenance": True, "own_customers_only": True}
    db.add(SysUser(
        username="maintenance_order_export_scoped_sales",
        role="sales",
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="sales",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "maintenance_order_export_scoped_sales", "password": "pw123456"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _workbook(response):
    return load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)


def _seed_orders(db, specs):
    batch = SysImportBatch(
        filename="maintenance-order-export.xlsx",
        file_type="maintenance",
        file_hash="maintenance-order-export",
    )
    db.add(batch)
    db.flush()
    orders = {}
    lines = []
    for raw_id, order_date, status, maint_end, line_count in specs:
        order = f.maintenance_head(
            raw_id,
            order_no=f"WBDD-{raw_id}",
            on=order_date or date(2026, 1, 1),
            project=f"项目-{raw_id}",
            data_status=status,
            maint_end=maint_end,
        )
        order["order_date"] = order_date
        orders[raw_id] = order
        lines.extend(
            f.maintenance_line(raw_id, f"LINE-{raw_id}-{index}", f"PN-{raw_id}-{index}")
            for index in range(1, line_count + 1)
        )
    loader.load(db, f.maintenance_result(orders, lines), batch.id, date(2026, 7, 1))
    db.commit()


def test_order_export_merges_unsynced_active_manual_override(db):
    """逐单总导出不得把仍只存在 override 表中的历史人工成本显示为 0/空。"""
    _seed_orders(db, [
        ("MANUAL-EXPORT", date(2026, 7, 1), "已生效", None, 1),
    ])
    line = db.scalar(select(FMaintenanceLine).where(
        FMaintenanceLine.raw_line_id == "LINE-MANUAL-EXPORT-1",
    ))
    assert line is not None
    line.qty = Decimal("5")
    line.return_qty = Decimal("2")
    line.cost_source = "none"
    line.cost_tax_basis = None
    line.unit_cost = None
    line.cost_amount = None
    line.unit_cost_inc_tax = None
    line.unit_cost_ex_tax = None
    line.cost_amount_inc_tax = None
    line.cost_amount_ex_tax = None
    line.confidence = None
    line.price_month = "2026-01"
    line.anomaly_flags = ["no_cost"]
    db.add(MaintenanceManualCostOverride(
        line_id=line.id,
        unit_cost_ex_tax=Decimal("10.00"),
        unit_cost_inc_tax=Decimal("11.30"),
        reason="历史人工成本依据",
        active=True,
        updated_by="test",
    ))
    db.commit()

    output = maintenance_export.build_workbook(
        db,
        UserContext(user_id="admin", role="admin"),
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        sheet = workbook["订单明细"]
        headers = [cell.value for cell in sheet[1]]
        values = [cell.value for cell in sheet[2]]
        exported = dict(zip(headers, values, strict=True))
        assert exported["单价"] == 10
        assert exported["金额"] == 30
        assert exported["含税单位成本"] == 11.3
        assert exported["未税单位成本"] == 10
        assert exported["含税成本金额"] == 33.9
        assert exported["未税成本金额"] == 30
        assert exported["成本事实层级"] == "实际采购参考"
        assert exported["成本来源"] == "manual"
        assert exported["含税口径"] == "ex"
        assert exported["置信度"] == "high"
        assert exported["取价月"] is None
        assert exported["异常标记"] in (None, "")
    finally:
        workbook.close()
        output.close()

    db.refresh(line)
    assert line.cost_source == "none"
    assert line.cost_amount is None


def test_empty_export_returns_clear_422_without_generating_an_xlsx(db):
    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 422
    assert response.json()["detail"] == "所选范围内没有可导出的维保订单"


def test_legacy_project_workbook_returns_404_for_unknown_contract(db):
    response = _admin_client(db).get(
        "/api/maintenance/export-workbook",
        params={"contract": "XSDD-NOT-FOUND"},
    )

    assert response.status_code == 404
    assert response.json()["detail"] == "合同不存在：XSDD-NOT-FOUND"


def test_export_closes_workbook_stream_after_response_is_consumed(db, monkeypatch):
    output = io.BytesIO(b"workbook bytes")
    monkeypatch.setattr(maintenance_export, "build_workbook", lambda *args: output)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    assert response.content == b"workbook bytes"
    assert output.closed


def test_export_closes_workbook_stream_when_asgi_send_disconnects(db, monkeypatch):
    output = io.BytesIO(b"workbook bytes")
    monkeypatch.setattr(maintenance_export, "build_workbook", lambda *args: output)
    response = maintenance_api.orders_export(
        date_from=None, date_to=None, db=db, _auth="admin", _page=None,
        ctx=UserContext(user_id="admin", role="admin"),
    )

    async def receive():
        return {"type": "http.request", "body": b"", "more_body": False}

    async def send(_message):
        raise OSError("client disconnected")

    with pytest.raises(ClientDisconnect):
        asyncio.run(response(
            {
                "type": "http", "method": "GET", "path": "/", "headers": [],
                "asgi": {"version": "3.0", "spec_version": "2.4"},
            },
            receive,
            send,
        ))

    assert output.closed


def test_build_workbook_cleans_writer_temp_files_when_database_stream_fails(db, monkeypatch):
    before = set(ALL_TEMP_FILES)
    monkeypatch.setattr(db, "execute", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("db failed")))

    with pytest.raises(RuntimeError, match="db failed"):
        maintenance_export.build_workbook(db, UserContext(user_id="admin", role="admin"))

    assert set(ALL_TEMP_FILES) == before


def test_build_workbook_closes_output_and_cleans_temp_files_when_save_fails(db, monkeypatch):
    _seed_orders(db, [("SAVE-FAIL", date(2026, 7, 1), "已生效", None, 1)])
    before = set(ALL_TEMP_FILES)
    output = io.BytesIO()
    monkeypatch.setattr(maintenance_export, "SpooledTemporaryFile", lambda **kwargs: output)
    monkeypatch.setattr(
        maintenance_export,
        "_save_workbook",
        lambda *args: (_ for _ in ()).throw(RuntimeError("save failed")),
    )

    with pytest.raises(RuntimeError, match="save failed"):
        maintenance_export.build_workbook(db, UserContext(user_id="admin", role="admin"))

    assert output.closed
    assert set(ALL_TEMP_FILES) == before


def test_export_rejects_final_xlsx_size_and_cleans_all_resources(db, monkeypatch):
    _seed_orders(db, [("SIZE-LIMIT", date(2026, 7, 1), "已生效", None, 1)])
    before = set(ALL_TEMP_FILES)
    output = io.BytesIO()
    monkeypatch.setattr(maintenance_export, "SpooledTemporaryFile", lambda **kwargs: output)
    monkeypatch.setattr(maintenance_export, "MAX_WORKBOOK_BYTES", 1, raising=False)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 413
    assert response.json()["detail"] == "维保订单 XLSX 超过 256 MiB 上限"
    assert output.closed
    assert set(ALL_TEMP_FILES) == before
    assert maintenance_export._ORDER_EXPORT_LOCK.acquire(blocking=False)
    maintenance_export._ORDER_EXPORT_LOCK.release()


def test_xlsx_size_guard_never_writes_beyond_hard_limit():
    output = io.BytesIO()
    limited = maintenance_export._SizeLimitedFile(output, max_size=3)
    assert limited.write(b"12") == 2

    with pytest.raises(
        maintenance_export.ExcelExportTooLarge,
        match="256 MiB",
    ):
        limited.write(b"34")

    assert output.getvalue() == b"12"


def test_build_workbook_cleans_temp_files_when_row_append_fails(db, monkeypatch):
    _seed_orders(db, [("APPEND", date(2026, 7, 15), "已生效", None, 1)])
    before = set(ALL_TEMP_FILES)
    original_append = WriteOnlyWorksheet.append
    calls = 0

    def fail_data_append(worksheet, row):
        nonlocal calls
        calls += 1
        if calls == 3:
            raise RuntimeError("append failed")
        return original_append(worksheet, row)

    monkeypatch.setattr(WriteOnlyWorksheet, "append", fail_data_append)

    with pytest.raises(RuntimeError, match="append failed"):
        maintenance_export.build_workbook(db, UserContext(user_id="admin", role="admin"))

    assert set(ALL_TEMP_FILES) == before


def test_export_rejects_date_from_without_date_to(db):
    response = _admin_client(db).get(
        "/api/maintenance/orders/export",
        params={"date_from": "2026-07-01"},
    )

    assert response.status_code == 422


def test_export_rejects_date_to_without_date_from(db):
    response = _admin_client(db).get(
        "/api/maintenance/orders/export",
        params={"date_to": "2026-07-31"},
    )

    assert response.status_code == 422


def test_export_rejects_reversed_date_range(db):
    response = _admin_client(db).get(
        "/api/maintenance/orders/export",
        params={"date_from": "2026-07-31", "date_to": "2026-07-01"},
    )

    assert response.status_code == 422


def test_export_rejects_detail_rows_over_excel_sheet_limit_and_cleans_temp_files(db, monkeypatch):
    _seed_orders(db, [("ROW-LIMIT", date(2026, 7, 15), "已生效", None, 2)])
    before = set(ALL_TEMP_FILES)
    monkeypatch.setattr(maintenance_export, "MAX_DATA_ROWS_PER_SHEET", 1, raising=False)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 422
    assert "1048575" in response.json()["detail"]
    assert set(ALL_TEMP_FILES) == before


def test_export_preflight_rejects_order_rows_over_limit_before_writing(db, monkeypatch):
    _seed_orders(db, [
        ("ORDER-LIMIT-1", date(2026, 7, 15), "已生效", None, 0),
        ("ORDER-LIMIT-2", date(2026, 7, 16), "已生效", None, 0),
    ])
    monkeypatch.setattr(maintenance_export, "MAX_DATA_ROWS_PER_SHEET", 1)
    append = Mock()
    monkeypatch.setattr(WriteOnlyWorksheet, "append", append)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 422
    assert "维保订单" in response.json()["detail"]
    append.assert_not_called()


def test_export_preflight_rejects_detail_rows_over_limit_before_writing(db, monkeypatch):
    _seed_orders(db, [("DETAIL-LIMIT", date(2026, 7, 15), "已生效", None, 2)])
    monkeypatch.setattr(maintenance_export, "MAX_DATA_ROWS_PER_SHEET", 1)
    append = Mock()
    monkeypatch.setattr(WriteOnlyWorksheet, "append", append)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 422
    assert "订单明细" in response.json()["detail"]
    append.assert_not_called()


def test_export_preflight_rejects_dynamic_text_budget_before_writing(db, monkeypatch):
    _seed_orders(db, [("TEXT-LIMIT", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.description = "超限动态文本"
    db.commit()
    monkeypatch.setattr(maintenance_export, "MAX_DYNAMIC_TEXT_BYTES", 1, raising=False)
    append = Mock()
    monkeypatch.setattr(WriteOnlyWorksheet, "append", append)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 413
    assert response.json()["detail"] == "维保订单 XLSX 动态文本超过 64 MiB 上限"
    append.assert_not_called()


def test_second_order_export_is_rejected_before_preflight_with_retry_after(db, monkeypatch):
    class BusyLock:
        def acquire(self, blocking=True):
            assert blocking is False
            return False

        def release(self):
            raise AssertionError("未取得锁时不得 release")

    monkeypatch.setattr(
        maintenance_export,
        "_ORDER_EXPORT_LOCK",
        BusyLock(),
        raising=False,
    )

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 429
    assert response.headers["retry-after"] == "5"
    assert response.json()["detail"] == "已有逐单维保导出正在执行，请稍后重试"


def test_database_lock_rejects_concurrent_export_across_sessions(db):
    _seed_orders(db, [("LOCK", date(2026, 7, 1), "已生效", None, 1)])
    client = _admin_client(db)
    ctx = UserContext(user_id="holder", role="admin")
    with SessionLocal() as holder:
        held_output = maintenance_export.build_workbook(holder, ctx)
        try:
            response = client.get("/api/maintenance/orders/export")
            assert response.status_code == 429
            assert response.headers["retry-after"] == "5"
            assert response.json()["detail"] == "已有逐单维保导出正在执行，请稍后重试"
        finally:
            held_output.close()
            holder.rollback()

    retry = client.get("/api/maintenance/orders/export")
    assert retry.status_code == 200, retry.text


def test_export_holds_shared_source_lock_from_preflight_through_materialization(
    db,
    monkeypatch,
):
    _seed_orders(db, [("SOURCE-SNAPSHOT", date(2026, 7, 15), "已生效", None, 1)])
    original_preflight = maintenance_export._preflight_row_limits
    original_save = maintenance_export._save_workbook
    protected_stages = []

    def assert_import_lock_blocked(stage):
        with SessionLocal() as importer:
            acquired = importer.scalar(
                text("SELECT pg_try_advisory_xact_lock(:k)"),
                {"k": config.DATA_CHANGE_ADVISORY_LOCK_KEY},
            )
            protected_stages.append((stage, acquired))
            importer.rollback()

    def checked_preflight(*args, **kwargs):
        assert_import_lock_blocked("preflight")
        return original_preflight(*args, **kwargs)

    def checked_save(*args, **kwargs):
        assert_import_lock_blocked("materialization")
        return original_save(*args, **kwargs)

    monkeypatch.setattr(maintenance_export, "_preflight_row_limits", checked_preflight)
    monkeypatch.setattr(maintenance_export, "_save_workbook", checked_save)

    output = maintenance_export.build_workbook(
        db,
        UserContext(user_id="admin", role="admin"),
    )
    output.close()
    assert protected_stages == [
        ("preflight", False),
        ("materialization", False),
    ]

    db.rollback()
    with SessionLocal() as importer:
        assert importer.scalar(
            text("SELECT pg_try_advisory_xact_lock(:k)"),
            {"k": config.DATA_CHANGE_ADVISORY_LOCK_KEY},
        ) is True


def test_all_export_contains_every_active_order_and_all_its_lines(db):
    _seed_orders(db, [
        ("OLD", date(2023, 12, 31), "已生效", date(2024, 1, 1), 2),
        ("NULL-DATE", None, "已生效", None, 1),
        ("CURRENT", date(2026, 7, 15), "已生效", date(2027, 1, 1), 2),
        ("INACTIVE", date(2026, 7, 15), "审批中", date(2027, 1, 1), 1),
    ])

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    workbook = _workbook(response)
    order_rows = list(workbook["维保订单"].values)[1:]
    line_rows = list(workbook["订单明细"].values)[1:]
    assert {row[1] for row in order_rows} == {"OLD", "NULL-DATE", "CURRENT"}
    assert len(order_rows) == 3
    assert {row[1] for row in line_rows} == {
        "LINE-OLD-1", "LINE-OLD-2", "LINE-NULL-DATE-1", "LINE-CURRENT-1", "LINE-CURRENT-2",
    }
    assert len(line_rows) == 5


def test_date_range_export_is_inclusive_and_excludes_null_dates(db):
    _seed_orders(db, [
        ("BEFORE", date(2026, 6, 30), "已生效", None, 1),
        ("FROM", date(2026, 7, 1), "已生效", None, 1),
        ("TO", date(2026, 7, 31), "已生效", None, 1),
        ("AFTER", date(2026, 8, 1), "已生效", None, 1),
        ("NULL-DATE", None, "已生效", None, 1),
    ])

    response = _admin_client(db).get(
        "/api/maintenance/orders/export",
        params={"date_from": "2026-07-01", "date_to": "2026-07-31"},
    )

    assert response.status_code == 200, response.text
    workbook = _workbook(response)
    order_rows = list(workbook["维保订单"].values)[1:]
    line_rows = list(workbook["订单明细"].values)[1:]
    assert [row[1] for row in order_rows] == ["FROM", "TO"]
    assert [row[1] for row in line_rows] == ["LINE-FROM-1", "LINE-TO-1"]


def test_cost_blind_user_exports_facts_with_all_cost_metadata_blank(db):
    _seed_orders(db, [("COSTED", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.unit_cost = Decimal("12.34")
    line.cost_amount = Decimal("24.68")
    line.cost_source = "direct"
    line.cost_tax_basis = "inc"
    line.price_month = "2026-07"
    line.trace_months = 0
    line.linked_purchase_order_no = "CGDD-SECRET"
    line.price_distance_days = 0
    line.confidence = "high"
    line.unit_cost_inc_tax = Decimal("12.34")
    line.unit_cost_ex_tax = Decimal("10.92")
    line.cost_amount_inc_tax = Decimal("24.68")
    line.cost_amount_ex_tax = Decimal("21.84")
    line.reference_side = "purchase"
    line.reference_pool_group_id = 77
    line.reference_pool_version = 3
    line.reference_sample_count = 4
    line.reference_from_date = date(2026, 6, 1)
    line.reference_to_date = date(2026, 6, 30)
    line.reference_latest_date = date(2026, 6, 30)
    db.commit()

    response = _cost_blind_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["订单明细"].values)
    row = dict(zip(values[0], values[1]))
    assert row["维保单号"] == "WBDD-COSTED"
    assert [row[key] for key in (
        "单价", "金额", "成本事实层级", "成本来源", "含税口径", "取价月", "追溯月数", "关联采购单",
        "距采购天数", "置信度",
        "含税单位成本", "未税单位成本", "含税成本金额", "未税成本金额",
        "参考侧", "参考池ID", "参考池版本", "参考样本数",
        "参考起始日", "参考截止日", "最近样本日",
    )] == [None] * 21


def test_order_workbook_fail_closes_invalid_cost_amount_and_marks_missing_tier(db):
    _seed_orders(db, [("INVALID-COST", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.unit_cost = Decimal("999.00")
    line.cost_amount = Decimal("999.00")
    line.cost_source = "future_source"
    line.cost_tax_basis = "inc"
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["订单明细"].values)
    row = dict(zip(values[0], values[1]))
    assert row["单价"] is None
    assert row["金额"] is None
    assert row["成本事实层级"] == "成本缺失"
    assert row["成本来源"] == "future_source"


def test_cost_blind_user_exports_data_quality_flags_without_cost_conclusions(db):
    _seed_orders(db, [("FLAGS", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.anomaly_flags = [
        "future_date",
        *maintenance_cost_quality.COST_DERIVED_ANOMALY_FLAGS,
        "has_return",
        "missing_qty",
    ]
    db.commit()

    response = _cost_blind_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["订单明细"].values)
    row = dict(zip(values[0], values[1]))
    assert row["异常标记"] == "future_date、has_return、missing_qty"
    assert row["维保单号"] == "WBDD-FLAGS"


def test_customer_blind_purchaser_exports_order_facts_without_end_customer(db):
    _seed_orders(db, [("CUSTOMER", date(2026, 7, 15), "已生效", None, 1)])
    order = db.execute(select(FMaintenanceOrder)).scalar_one()
    order.end_customer = "敏感终端客户"
    db.commit()

    response = _customer_blind_purchaser_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["维保订单"].values)
    row = dict(zip(values[0], values[1]))
    assert row["终端客户"] is None
    assert row["维保单号"] == "WBDD-CUSTOMER"
    assert row["项目名"] == "项目-CUSTOMER"


def test_export_escapes_every_excel_formula_prefix_in_text_cells(db):
    _seed_orders(db, [("FORMULA", date(2026, 7, 15), "已生效", None, 1)])
    order = db.execute(select(FMaintenanceOrder)).scalar_one()
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    order.order_no = "=WBDD"
    order.linked_sales_order_no = "+XSDD"
    order.project_raw = "-项目"
    order.end_customer = "@客户"
    order.demand_type = "\t需求"
    order.business_type = "\r业务"
    order.salesperson = "\n销售"
    line.pn_std = "=PN"
    line.pn_raw = "+RAW"
    line.description = "-描述"
    line.serial_numbers = "@SN"
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    workbook = _workbook(response)
    order_values = list(workbook["维保订单"].values)
    order_row = dict(zip(order_values[0], order_values[1]))
    line_values = list(workbook["订单明细"].values)
    line_row = dict(zip(line_values[0], line_values[1]))
    assert [order_row[key] for key in (
        "维保单号", "销售订单", "项目名", "终端客户", "需求类型", "业务类型", "销售人员",
    )] == ["'=WBDD", "'+XSDD", "'-项目", "'@客户", "'\t需求", "'\r业务", "'\n销售"]
    assert [line_row[key] for key in ("维保单号", "PN", "原始PN", "产品描述", "发货SN")] == [
        "'=WBDD", "'=PN", "'+RAW", "'-描述", "'@SN",
    ]


def test_export_removes_invalid_xml_controls_before_formula_escaping(db):
    _seed_orders(db, [("CONTROL", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.description = "\x0b=危险\x0c文本"
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["订单明细"].values)
    row = dict(zip(values[0], values[1]))
    assert row["产品描述"] == "'=危险文本"


def test_export_removes_xml_forbidden_unicode_noncharacters(db):
    _seed_orders(db, [("NONCHAR", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.description = "前\ufffe中\uffff后"
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["订单明细"].values)
    row = dict(zip(values[0], values[1]))
    assert row["产品描述"] == "前中后"


def test_range_export_filename_contains_actual_boundaries(db):
    _seed_orders(db, [("RANGE", date(2026, 7, 15), "已生效", None, 1)])
    response = _admin_client(db).get(
        "/api/maintenance/orders/export",
        params={"date_from": "2026-07-01", "date_to": "2026-07-31"},
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-disposition"] == (
        'attachment; filename="maintenance_orders_2026-07-01_2026-07-31.xlsx"; '
        "filename*=UTF-8''maintenance_orders_2026-07-01_2026-07-31.xlsx"
    )


def test_export_audit_log_records_all_and_date_range_scopes(db, monkeypatch):
    _seed_orders(db, [("AUDIT", date(2026, 7, 15), "已生效", None, 1)])
    audit = Mock()
    monkeypatch.setattr(maintenance_api, "record_access_log", audit)
    client = _admin_client(db)

    assert client.get("/api/maintenance/orders/export").status_code == 200
    assert client.get("/api/maintenance/orders/export", params={
        "date_from": "2026-07-01", "date_to": "2026-07-31",
    }).status_code == 200

    assert audit.call_args_list[0].args[3] == {"scope": "all"}
    assert audit.call_args_list[1].args[3] == {
        "date_from": "2026-07-01", "date_to": "2026-07-31",
    }


def test_export_keeps_anonymous_401_and_no_page_403(db):
    assert TestClient(app).get("/api/maintenance/orders/export").status_code == 401
    assert _readonly_client(db).get("/api/maintenance/orders/export").status_code == 403


def test_scoped_sales_remains_forbidden_after_page_permission_is_granted(db, monkeypatch):
    build_workbook = Mock()
    monkeypatch.setattr(maintenance_export, "build_workbook", build_workbook)

    response = _scoped_sales_with_maintenance_page_client(db).get(
        "/api/maintenance/orders/export",
    )

    assert response.status_code == 403
    build_workbook.assert_not_called()


def test_as_of_returns_authenticated_business_date(db, monkeypatch):
    monkeypatch.setattr(maintenance_api, "business_today", lambda: date(2026, 7, 17))

    response = _admin_client(db).get("/api/maintenance/as-of")

    assert response.status_code == 200
    assert response.headers["cache-control"] == "no-store"
    assert response.json() == {"as_of": "2026-07-17"}


def test_export_rejects_text_that_excel_would_silently_truncate(db):
    _seed_orders(db, [("TOO-LONG", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.description = "X" * 32768
    db.commit()
    before = set(ALL_TEMP_FILES)

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 422
    assert "32767" in response.json()["detail"]
    assert set(ALL_TEMP_FILES) == before


def test_export_keeps_formula_escaped_text_at_excel_cell_limit(db):
    _seed_orders(db, [("ESCAPED-LIMIT", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    description = "=" + "X" * 32765
    line.description = description
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    values = list(_workbook(response)["订单明细"].values)
    row = dict(zip(values[0], values[1]))
    exported_description = row["产品描述"]
    assert isinstance(exported_description, str)
    assert exported_description == "'" + description
    assert len(exported_description) == 32767


def test_export_rejects_formula_escaped_text_over_excel_cell_limit(db):
    _seed_orders(db, [("ESCAPED-TOO-LONG", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.description = "=" + "X" * 32766
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 422
    assert "32767" in response.json()["detail"]


def test_export_uses_excel_date_quantity_and_money_formats(db):
    _seed_orders(db, [("FORMATS", date(2026, 7, 15), "已生效", None, 1)])
    line = db.execute(select(FMaintenanceLine)).scalar_one()
    line.qty = Decimal("1234.5")
    line.return_qty = Decimal("0.5")
    line.unit_cost = Decimal("12.34")
    line.cost_amount = Decimal("15227.30")
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    workbook = load_workbook(io.BytesIO(response.content), data_only=True)
    order_sheet = workbook["维保订单"]
    line_sheet = workbook["订单明细"]
    assert order_sheet["D2"].is_date
    assert line_sheet["F2"].is_date
    assert line_sheet["K2"].number_format == "0.00"
    assert line_sheet["L2"].number_format == "0.00"
    assert line_sheet["N2"].number_format == "#,##0.00"
    assert line_sheet["O2"].number_format == "#,##0.00"


def test_export_keeps_all_50000_detail_rows(db):
    _seed_orders(db, [("STRESS", date(2026, 7, 15), "已生效", None, 1)])
    seed = db.execute(select(FMaintenanceLine)).scalar_one()
    db.execute(text("""
        INSERT INTO f_maintenance_line
            (raw_line_id, order_id, line_no, part_id, pn_std, pn_raw, qty, import_batch_id)
        SELECT
            'LINE-STRESS-' || n, :order_id, n, :part_id, 'PN-STRESS', 'PN-STRESS', 1, :batch_id
        FROM generate_series(2, 50000) AS n
    """), {
        "order_id": seed.order_id,
        "part_id": seed.part_id,
        "batch_id": seed.import_batch_id,
    })
    db.commit()

    response = _admin_client(db).get("/api/maintenance/orders/export")

    assert response.status_code == 200, response.text
    workbook = _workbook(response)
    assert sum(1 for _ in workbook["维保订单"].values) == 2
    assert sum(1 for _ in workbook["订单明细"].values) == 50001
