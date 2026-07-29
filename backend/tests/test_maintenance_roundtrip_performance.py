"""Hard SQL-count gates for DEV-15 roundtrip workbook imports.

These tests use the real PostgreSQL test database. They count cursor executions
only around the import itself, so workbook generation and fixture setup cannot
hide a row-by-row query or flush regression.
"""
from __future__ import annotations

import uuid
from collections import Counter
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import event, func, select

from app.models.dimensions import DimCustomer, DimPart
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceManualCostOverride,
    MaintenanceRoundtripOperation,
)
from app.models.system import SysAuditLog, SysImportBatch
from app.services import maintenance_roundtrip
from tests.test_maintenance_roundtrip import _export_to_path, _seed_contract


@dataclass(slots=True)
class _SqlCount:
    total: int = 0
    selects: int = 0
    verbs: Counter[str] = field(default_factory=Counter)
    statements: list[str] = field(default_factory=list)

    def describe(self) -> str:
        return (
            f"total={self.total}, selects={self.selects}, "
            f"verbs={dict(self.verbs)}, statements={self.statements}"
        )


@contextmanager
def _count_sql(db) -> Iterator[_SqlCount]:
    count = _SqlCount()
    bind = db.get_bind()

    def before_cursor_execute(
        _connection,
        _cursor,
        statement,
        _parameters,
        _context,
        _executemany,
    ):
        normalized = " ".join(statement.strip().split())
        verb = normalized.split(" ", 1)[0].upper() if normalized else "EMPTY"
        count.total += 1
        count.verbs[verb] += 1
        count.statements.append(normalized[:180])
        if verb == "SELECT":
            count.selects += 1

    event.listen(bind, "before_cursor_execute", before_cursor_execute)
    try:
        yield count
    finally:
        event.remove(bind, "before_cursor_execute", before_cursor_execute)


def _seed_lines(db, *, suffix: str, contract: str, row_count: int) -> list[int]:
    batch = SysImportBatch(
        filename=f"roundtrip-perf-seed-{suffix}.xlsx",
        file_type="maintenance",
        file_hash=f"roundtrip-perf-seed-{suffix}",
    )
    part = DimPart(pn_std=f"PN-RT-PERF-{suffix}")
    db.add_all([batch, part])
    db.flush()
    order = FMaintenanceOrder(
        raw_order_id=f"ORDER-RT-PERF-{suffix}",
        order_no=f"WBDD-RT-PERF-{suffix}",
        order_date=date(2026, 7, 15),
        linked_sales_order_no=contract,
        project_raw=f"性能门禁-{suffix}",
        project_std=f"性能门禁-{suffix}",
        data_status="已生效",
        import_batch_id=batch.id,
    )
    db.add(order)
    db.flush()
    lines = [
        FMaintenanceLine(
            raw_line_id=f"LINE-RT-PERF-{suffix}-{index:04d}",
            order_id=order.id,
            line_no=index,
            part_id=part.id,
            pn_std=part.pn_std,
            description=f"性能门禁原始行 {index}",
            qty=Decimal("2.000"),
            return_qty=Decimal("0.000"),
            cost_source="none",
            anomaly_flags=["no_cost"],
            import_batch_id=batch.id,
        )
        for index in range(1, row_count + 1)
    ]
    db.add_all(lines)
    db.commit()
    return [line.id for line in lines]


def _mark_rows(
    path,
    *,
    sheet: str,
    operation: str,
    values: dict[str, object],
) -> None:
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook[sheet]
        headers = {
            str(cell.value): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        table = next(iter(worksheet.tables.values()))
        last_row = int(table.ref.rsplit(":", 1)[1][1:])
        for row in range(2, last_row + 1):
            worksheet.cell(row=row, column=headers["操作"], value=operation)
            for header, value in values.items():
                worksheet.cell(row=row, column=headers[header], value=value)
        workbook.save(path)
    finally:
        workbook.close()


def _mark_expense_creates(path, *, contract: str, row_count: int) -> None:
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook["04_报销明细"]
        headers = {
            str(cell.value): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        for row in range(2, row_count + 2):
            values = {
                "操作": "CREATE",
                "合同号": contract,
                "报销日期": date(2026, 7, 20),
                "未税金额": Decimal("1.00"),
                "流程状态": "已结束",
                "变更原因": "DEV-15 SQL 门禁",
                "__client_row_id": str(uuid.uuid4()),
            }
            for header, value in values.items():
                worksheet.cell(row=row, column=headers[header], value=value)
        table = next(iter(worksheet.tables.values()))
        table.ref = (
            f"A1:{get_column_letter(len(headers))}{row_count + 1}"
        )
        workbook.save(path)
    finally:
        workbook.close()


def _run_line_update(db, tmp_path, row_count: int) -> _SqlCount:
    suffix = f"LINE-{row_count}"
    contract = f"XSDD-RT-PERF-{suffix}"
    _seed_lines(db, suffix=suffix, contract=contract, row_count=row_count)
    path = _export_to_path(
        db,
        tmp_path / f"line-update-{row_count}.xlsx",
        contract=contract,
    )
    _mark_rows(
        path,
        sheet="03_订单明细",
        operation="UPDATE",
        values={
            "产品描述": f"批量更新 {row_count}",
            "需求数量": Decimal("3.000"),
            "退货数量": Decimal("1.000"),
            "变更原因": "DEV-15 SQL 门禁",
        },
    )

    with _count_sql(db) as count:
        result = maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="perf-tester",
        )

    assert result["changed_rows"] == row_count
    assert result["counts"]["update"] == row_count
    return count


def _run_manual_create(db, tmp_path, row_count: int) -> _SqlCount:
    suffix = f"MANUAL-{row_count}"
    contract = f"XSDD-RT-PERF-{suffix}"
    _seed_lines(db, suffix=suffix, contract=contract, row_count=row_count)
    path = _export_to_path(
        db,
        tmp_path / f"manual-create-{row_count}.xlsx",
        contract=contract,
    )
    _mark_rows(
        path,
        sheet="05_人工成本回填",
        operation="CREATE",
        values={
            "人工未税单位成本": Decimal("2.50"),
            "回填原因": "三个月内无采购和销售参考",
            "依据说明": "DEV-15 SQL 门禁",
        },
    )

    with _count_sql(db) as count:
        result = maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="perf-tester",
        )

    assert result["changed_rows"] == row_count
    assert result["counts"]["create"] == row_count
    return count


def _run_expense_create(db, tmp_path, row_count: int) -> _SqlCount:
    suffix = f"EXPENSE-{row_count}"
    contract = f"XSDD-RT-PERF-{suffix}"
    _seed_contract(db, suffix=suffix, contract=contract)
    path = _export_to_path(
        db,
        tmp_path / f"expense-create-{row_count}.xlsx",
        contract=contract,
    )
    _mark_expense_creates(path, contract=contract, row_count=row_count)

    with _count_sql(db) as count:
        result = maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="perf-tester",
        )

    assert result["changed_rows"] == row_count
    assert result["counts"]["create"] == row_count
    return count


def test_line_update_100_and_1000_rows_have_constant_bounded_sql(db, tmp_path):
    counts = {
        row_count: _run_line_update(db, tmp_path, row_count)
        for row_count in (100, 1000)
    }
    hundred = counts[100]
    thousand = counts[1000]

    assert thousand.selects <= 20, thousand.describe()
    assert thousand.total <= 50, thousand.describe()
    assert thousand.selects == hundred.selects, {
        size: count.describe() for size, count in counts.items()
    }
    # The second import recomputes 1100 total lines and may need one extra
    # 1000-row UPDATE chunk, but cursor executions must otherwise be constant.
    assert thousand.total <= hundred.total + 2, {
        size: count.describe() for size, count in counts.items()
    }
    assert db.scalar(
        select(func.count(SysAuditLog.id)).where(
            SysAuditLog.entity_type == "f_maintenance_line",
        )
    ) == 1100
    assert db.scalar(select(func.count(MaintenanceRoundtripOperation.id))) == 1100
    line_ids = set(db.scalars(select(FMaintenanceLine.id)))
    audit_ids = set(db.scalars(
        select(SysAuditLog.entity_id).where(
            SysAuditLog.entity_type == "f_maintenance_line",
        )
    ))
    ledger_ids = {
        int(result["entity_id"])
        for result in db.scalars(
            select(MaintenanceRoundtripOperation.result_json)
        )
    }
    assert audit_ids == line_ids
    assert ledger_ids == line_ids


def test_manual_create_100_and_1000_rows_have_constant_bounded_sql(db, tmp_path):
    counts = {
        row_count: _run_manual_create(db, tmp_path, row_count)
        for row_count in (100, 1000)
    }
    hundred = counts[100]
    thousand = counts[1000]

    assert thousand.selects <= 22, thousand.describe()
    assert thousand.total <= 50, thousand.describe()
    assert thousand.selects == hundred.selects, {
        size: count.describe() for size, count in counts.items()
    }
    assert thousand.total <= hundred.total + 2, {
        size: count.describe() for size, count in counts.items()
    }
    assert db.scalar(select(func.count(MaintenanceManualCostOverride.id))) == 1100
    assert db.scalar(
        select(func.count(SysAuditLog.id)).where(
            SysAuditLog.entity_type == "maintenance_manual_cost_override",
        )
    ) == 1100
    assert db.scalar(select(func.count(MaintenanceRoundtripOperation.id))) == 1100
    override_ids = set(db.scalars(select(MaintenanceManualCostOverride.id)))
    audit_ids = set(db.scalars(
        select(SysAuditLog.entity_id).where(
            SysAuditLog.entity_type == "maintenance_manual_cost_override",
        )
    ))
    ledger_ids = {
        int(result["entity_id"])
        for result in db.scalars(
            select(MaintenanceRoundtripOperation.result_json)
        )
    }
    assert audit_ids == override_ids
    assert ledger_ids == override_ids


def test_expense_create_100_and_1000_rows_have_constant_bounded_sql(db, tmp_path):
    counts = {
        row_count: _run_expense_create(db, tmp_path, row_count)
        for row_count in (100, 1000)
    }
    hundred = counts[100]
    thousand = counts[1000]

    assert thousand.selects <= 22, thousand.describe()
    assert thousand.total <= 50, thousand.describe()
    assert thousand.selects == hundred.selects, {
        size: count.describe() for size, count in counts.items()
    }
    assert thousand.total <= hundred.total + 2, {
        size: count.describe() for size, count in counts.items()
    }
    assert db.scalar(select(func.count(FProjectExpense.id))) == 1100
    assert db.scalar(
        select(func.count(SysAuditLog.id)).where(
            SysAuditLog.entity_type == "f_project_expense",
        )
    ) == 1100
    assert db.scalar(select(func.count(MaintenanceRoundtripOperation.id))) == 1100
    expense_ids = set(db.scalars(select(FProjectExpense.id)))
    audit_ids = set(db.scalars(
        select(SysAuditLog.entity_id).where(
            SysAuditLog.entity_type == "f_project_expense",
        )
    ))
    ledger_ids = {
        int(result["entity_id"])
        for result in db.scalars(
            select(MaintenanceRoundtripOperation.result_json)
        )
    }
    assert audit_ids == expense_ids
    assert ledger_ids == expense_ids


def test_customer_resolution_for_1000_names_is_one_select_one_insert(db):
    changes = [
        maintenance_roundtrip._Change(
            kind="order",
            operation="UPDATE",
            row=maintenance_roundtrip._ParsedRow(
                sheet="02_维保订单",
                excel_row=index + 2,
                values={},
            ),
            entity=None,
            values={"customer_name": f"性能客户-{index:04d}"},
            contracts=frozenset(),
        )
        for index in range(1000)
    ]

    with _count_sql(db) as count:
        customer_ids = maintenance_roundtrip._customer_ids_for_changes(db, changes)

    assert len(customer_ids) == 1000
    assert count.selects == 1, count.describe()
    assert count.total == 2, count.describe()
    assert db.scalar(select(func.count(DimCustomer.id))) == 1000


def test_exact_hash_replay_is_two_selects_without_commit_or_recompute(
    db,
    tmp_path,
    monkeypatch,
):
    _seed_contract(
        db,
        suffix="PERF-EXACT",
        contract="XSDD-RT-PERF-EXACT",
    )
    path = _export_to_path(
        db,
        tmp_path / "exact-hash-replay.xlsx",
        contract="XSDD-RT-PERF-EXACT",
    )
    first = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="perf-tester",
    )
    commits = 0

    def after_commit(_session):
        nonlocal commits
        commits += 1

    def fail_if_recomputed(_db):
        raise AssertionError("exact hash replay reached recompute")

    event.listen(db, "after_commit", after_commit)
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_recompute_in_transaction",
        fail_if_recomputed,
    )
    try:
        with _count_sql(db) as count:
            replay = maintenance_roundtrip.import_roundtrip_workbook(
                db,
                str(path),
                filename=path.name,
                operated_by="perf-tester",
            )
    finally:
        event.remove(db, "after_commit", after_commit)

    assert replay["no_op"] is True
    assert replay["batch_id"] == first["batch_id"]
    assert count.selects == 2, count.describe()
    assert count.total == 2, count.describe()
    assert commits == 0
