"""Agent Capability Kernel security contracts (#219)."""

import asyncio
import json
import os
import subprocess
import sys
import threading
from dataclasses import replace
from datetime import date
from io import BytesIO
from pathlib import Path
from types import MappingProxyType, SimpleNamespace

import pytest
from fastapi import HTTPException
from fastapi.security import HTTPAuthorizationCredentials
from openpyxl import load_workbook
from sqlalchemy import select
from starlette.datastructures import UploadFile

from app import auth, config, security
from app.agent import limits, provider, runtime, tools
from app.api import agent as agent_api, chat_sessions
from app.models.chat import ChatMessage
from app.models.system import SysUser
from app.services import agent_files, chat_store


def _sys_ctx(user_id: str = "alice", role: str = "sales") -> security.UserContext:
    return security.UserContext(
        user_id=user_id,
        role=role,
        is_authenticated=True,
        authn="sys_user",
        token_version=0,
    )


@pytest.fixture(autouse=True)
def _explicit_non_rbac_compatibility(monkeypatch):
    """Most kernel unit tests isolate policy layers without a live request identity."""
    monkeypatch.setattr(config, "ENABLE_RBAC", False)


def _egress_settings(**overrides):
    values = {
        "environment": "test",
        "agent_allow_loopback_http": False,
        "agent_allow_unattested_private_for_development": True,
        "enable_agent": True,
        "llm_provider": "openai_compatible",
        "llm_api_key": "test-primary-key",
        "vision_api_key": "test-vision-key",
        "llm_timeout_seconds": 60,
        "llm_max_retries": 2,
        "llm_max_tokens": None,
        "llm_max_tool_iters": 8,
        "vision_timeout_seconds": 90,
        "vision_max_pages": 8,
        "llm_extra_body_dict": lambda: None,
        "llm_trust_zone": "private",
        "llm_base_url": "https://agent-private.test:8000/v1",
        "llm_model": "primary-model",
        "llm_approved_models": "primary-model,another-primary-model",
        "llm_private_base_urls": "https://agent-private.test:8000",
        "llm_approved_external_base_urls": "",
        "vision_trust_zone": "private",
        "vision_base_url": "https://vision-private.test:8443/v1",
        "vision_model": "vision-model",
        "vision_approved_models": "vision-model,another-vision-model",
        "vision_private_base_urls": "https://vision-private.test:8443",
        "vision_approved_external_base_urls": "",
        "agent_model_context_egress_enabled": True,
        "agent_external_file_egress_enabled": False,
    }
    values.update(overrides)
    return SimpleNamespace(**values)


def test_egress_configuration_defaults_fail_closed():
    fields = config.Settings.model_fields
    assert fields["enable_agent"].default is False
    assert fields["llm_trust_zone"].default == "unknown"
    assert fields["llm_private_base_urls"].default == ""
    assert fields["llm_approved_external_base_urls"].default == ""
    assert fields["vision_trust_zone"].default == "unknown"
    assert fields["vision_private_base_urls"].default == ""
    assert fields["vision_approved_external_base_urls"].default == ""
    assert fields["agent_allow_loopback_http"].default is False
    assert fields["agent_model_context_egress_enabled"].default is False
    assert fields["agent_allow_unattested_private_for_development"].default is False
    assert fields["agent_external_file_egress_enabled"].default is False


def test_current_release_hard_rejects_agent_enablement_in_production():
    secure = config.Settings(
        environment="prod",
        enable_agent=True,
        admin_password="non-default-admin-password",
        secret_key="non-default-secret-key-with-sufficient-entropy",
        database_url="postgresql+psycopg://app:nondefault@db.invalid/app",
    )

    warnings = config.check_security(secure)

    assert warnings == [
        "当前版本未通过 Agent 生产门禁，生产环境必须保持 ENABLE_AGENT=false"
    ]


@pytest.mark.parametrize("environment", ["production", "PROD", "typo"])
def test_unknown_or_case_drifted_environment_is_rejected(environment):
    with pytest.raises(ValueError):
        config.Settings(environment=environment)


def test_main_import_rejects_prod_agent_even_with_nondefault_credentials():
    environment = os.environ.copy()
    environment.update({
        "ENVIRONMENT": "prod",
        "ENABLE_AGENT": "true",
        "ADMIN_PASSWORD": "non-default-admin-password",
        "SECRET_KEY": "non-default-secret-key-with-sufficient-entropy",
        "DATABASE_URL": "postgresql+psycopg://app:nondefault@db.invalid/app",
    })
    for name in (
        "OPENAI_ORG_ID",
        "OPENAI_PROJECT_ID",
        "OPENAI_CUSTOM_HEADERS",
        "OPENAI_LOG",
        "SSLKEYLOGFILE",
    ):
        environment.pop(name, None)

    result = subprocess.run(
        [sys.executable, "-c", "import app.main"],
        cwd=str(Path(__file__).resolve().parents[1]),
        env=environment,
        text=True,
        capture_output=True,
        timeout=20,
        check=False,
    )

    assert result.returncode != 0
    assert "生产环境安全自检失败" in result.stderr
    assert "ENABLE_AGENT=false" in result.stderr


def test_agent_kill_switch_covers_stateless_files_and_session_namespaces(monkeypatch):
    disabled = SimpleNamespace(enable_agent=False)
    monkeypatch.setattr(agent_api, "get_settings", lambda: disabled)
    monkeypatch.setattr(chat_sessions, "get_settings", lambda: disabled)

    for guard in (agent_api._require_agent_enabled, chat_sessions._require_agent_enabled):
        with pytest.raises(HTTPException) as exc_info:
            guard()
        assert exc_info.value.status_code == 404

    assert any(
        dependency.dependency is agent_api._require_agent_enabled
        for dependency in agent_api.router.dependencies
    )
    assert any(
        dependency.dependency is chat_sessions._require_agent_enabled
        for dependency in chat_sessions.router.dependencies
    )


def test_vision_capability_declares_both_required_egress_edges():
    spec = tools._SPEC_BY_NAME["read_document_with_vision"]

    assert spec.egress == tools._vision_egress()
    assert [(edge.purpose, edge.projection_id, edge.max_bytes, edge.policy_version,
             edge.retention_policy)
            for edge in spec.egress] == [
        (
            tools.PURPOSE_DOCUMENT_ASSISTANCE,
            tools.PROJECTION_VISION_INPUT,
            tools.VISION_INPUT_MAX_BYTES,
            tools.EGRESS_POLICY_VERSION,
            tools.RETENTION_NO_ADDITIONAL_EGRESS_ARCHIVE,
        ),
        (
            tools.PURPOSE_DOCUMENT_ASSISTANCE,
            tools.PROJECTION_VISION_OCR_RESULT,
            tools.VISION_OCR_RESULT_MAX_BYTES,
            tools.EGRESS_POLICY_VERSION,
            tools.RETENTION_NO_ADDITIONAL_EGRESS_ARCHIVE,
        ),
    ]


def test_upload_audit_records_only_extension_and_size(monkeypatch):
    secret = "CUSTOMER-SECRET-FILENAME-8472"
    events: list[tuple] = []
    monkeypatch.setattr(
        agent_api,
        "record_access_log",
        lambda ctx, action, resource, filters=None: events.append(
            (action, resource, filters)
        ),
    )
    content = b"safe local text"
    upload = UploadFile(file=BytesIO(content), filename=f"{secret}.txt")

    result = asyncio.run(agent_api.upload(upload, role="sales", ctx=_sys_ctx()))

    assert result["ext"] == "txt"
    assert events == [("upload", "agent_file", {"ext": "txt", "size_bytes": len(content)})]
    assert secret not in str(events)


@pytest.mark.parametrize(
    ("authn", "user_id"),
    [
        ("shared", "legacy-shared-user"),
        (None, None),
        ("sys_user", ""),
    ],
)
def test_upload_rejects_unstable_owner_before_read_audit_or_storage(
    monkeypatch,
    authn,
    user_id,
):
    reads: list[str] = []
    audits: list[str] = []
    saves: list[str] = []

    class DeniedUpload:
        filename = "customer-secret.xlsx"

        async def read(self):
            reads.append("read")
            return b"must-not-be-read"

    monkeypatch.setattr(
        agent_api,
        "record_access_log",
        lambda *_args, **_kwargs: audits.append("audit"),
    )
    monkeypatch.setattr(
        agent_files,
        "save_upload",
        lambda *_args, **_kwargs: saves.append("save"),
    )
    ctx = security.UserContext(
        user_id=user_id,
        role="admin" if authn == "shared" else "guest",
        is_authenticated=authn is not None,
        authn=authn,
        token_version=0 if authn == "sys_user" else None,
    )

    with pytest.raises(HTTPException) as denied:
        asyncio.run(agent_api.upload(DeniedUpload(), role=ctx.role, ctx=ctx))

    assert denied.value.status_code == 404
    assert denied.value.detail == agent_api._FILE_NOT_FOUND
    assert reads == audits == saves == []


def test_chat_access_audit_records_structure_without_prompt(monkeypatch):
    secret = "CUSTOMER-CHAT-PROMPT-SECRET-8472"
    events: list[tuple] = []
    monkeypatch.setattr(
        agent_api,
        "record_access_log",
        lambda ctx, action, resource, filters=None: events.append(
            (action, resource, filters)
        ),
    )
    monkeypatch.setattr(provider, "is_configured", lambda: False)
    req = agent_api.ChatRequest(messages=[
        agent_api.ChatMessage(role="user", content="first"),
        agent_api.ChatMessage(role="assistant", content="answer"),
        agent_api.ChatMessage(role="user", content=secret),
    ])
    ctx = _sys_ctx()

    agent_api.chat(req, db=None, role="sales", ctx=ctx)
    agent_api.chat_stream(req, db=None, role="sales", ctx=ctx)

    assert events == [
        (
            "chat",
            "agent",
            {
                "message_count": 3,
                "last_message_chars": len(secret),
                "endpoint": "chat",
                "stream": False,
            },
        ),
        (
            "chat_stream",
            "agent",
            {
                "message_count": 3,
                "last_message_chars": len(secret),
                "endpoint": "chat_stream",
                "stream": True,
            },
        ),
    ]
    assert secret not in str(events)


def test_persisted_chat_audit_records_structure_without_prompt(monkeypatch):
    secret = "SESSION-CHAT-PROMPT-SECRET-8472"
    events: list[tuple] = []
    monkeypatch.setattr(
        chat_sessions,
        "_owned_or_404",
        lambda db, ident, session_id: SimpleNamespace(id=session_id, title="safe"),
    )
    monkeypatch.setattr(chat_sessions, "acquire_session", lambda session_id: None)
    monkeypatch.setattr(
        chat_sessions,
        "record_access_log",
        lambda ctx, action, resource, filters=None: events.append(
            (action, resource, filters)
        ),
    )

    with pytest.raises(HTTPException) as exc_info:
        chat_sessions.chat_stream(
            42,
            chat_sessions.SendMessageRequest(message=secret),
            db=None,
            ident={"sub": "alice", "authn": "sys_user"},
            ctx=_sys_ctx(),
        )

    assert getattr(exc_info.value, "status_code", None) == 409
    assert events == [(
        "chat_stream",
        "agent",
        {
            "message_count": 1,
            "last_message_chars": len(secret),
            "endpoint": "session_chat_stream",
            "stream": True,
            "session_id": 42,
        },
    )]
    assert secret not in str(events)


def test_stateless_chat_endpoints_preflight_model_egress_before_runtime(monkeypatch):
    request = agent_api.ChatRequest(messages=[
        agent_api.ChatMessage(role="user", content="customer context"),
    ])
    monkeypatch.setattr(agent_api, "record_access_log", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: False)
    monkeypatch.setattr(
        runtime,
        "run",
        lambda *_args, **_kwargs: pytest.fail("denied chat reached runtime"),
    )
    monkeypatch.setattr(
        runtime,
        "run_stream",
        lambda *_args, **_kwargs: pytest.fail("denied stream reached runtime"),
    )

    result = agent_api.chat(request, db=None, role="sales", ctx=_sys_ctx())
    response = agent_api.chat_stream(request, db=None, role="sales", ctx=_sys_ctx())
    body = asyncio.run(_stream_body(response))

    assert result["code"] == "AGENT_MODEL_EGRESS_DENIED"
    assert result["retriable"] is False
    assert "AGENT_MODEL_EGRESS_DENIED" in body


def test_stateless_api_dto_refreshes_identity_before_releasing_artifact_ids(monkeypatch):
    artifact_id = "e" * 12
    tool = {
        "name": "write_report",
        "args": {},
        "args_are_shape": True,
        "artifact_ids": [artifact_id],
    }
    request = agent_api.ChatRequest(messages=[
        agent_api.ChatMessage(role="user", content="safe request"),
    ])

    monkeypatch.setattr(agent_api, "record_access_log", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: True)
    monkeypatch.setattr(tools, "refresh_runtime_context", lambda _db, _ctx: None)
    monkeypatch.setattr(
        runtime,
        "run",
        lambda *_args, **_kwargs: {"answer": "safe", "tool_calls": [tool]},
    )

    def forged_stream(*_args, **_kwargs):
        yield {"type": "tool", **tool}
        yield {
            "type": "tool_done",
            "name": "write_report",
            "ok": True,
            "artifact_ids": [artifact_id],
        }
        yield {"type": "done", "answer": "safe", "tool_calls": [tool]}

    monkeypatch.setattr(runtime, "run_stream", forged_stream)

    result = agent_api.chat(request, db=None, role="sales", ctx=_sys_ctx())
    response = agent_api.chat_stream(request, db=None, role="sales", ctx=_sys_ctx())
    body = asyncio.run(_stream_body(response))

    assert artifact_id not in json.dumps(result, ensure_ascii=False)
    assert artifact_id not in body
    assert result["tool_calls"][0]["name"] == "write_report"
    assert '"type": "done"' in body


def test_session_history_refreshes_identity_for_every_artifact_id(db, monkeypatch):
    own_id = "a" * 12
    revoked_id = "b" * 12
    session = chat_store.create_session(db, "alice")
    db.add(ChatMessage(
        session_id=session.id,
        role="assistant",
        content="safe",
        tools=[{
            "name": "write_report",
            "args": {},
            "args_are_shape": True,
            "artifact_ids": [own_id, revoked_id],
        }],
        stopped=False,
    ))
    db.commit()
    allowed = _sys_ctx()
    decisions = iter((allowed, allowed, None))
    monkeypatch.setattr(
        tools,
        "refresh_runtime_context",
        lambda _db, _ctx: next(decisions),
    )
    monkeypatch.setattr(agent_files, "owner_of", lambda _artifact_id: "alice")

    result = chat_sessions.list_messages(
        session.id,
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=allowed,
    )

    assert result["items"][0]["tools"][0]["artifact_ids"] == [own_id]
    assert revoked_id not in json.dumps(result, ensure_ascii=False, default=str)


def test_checkpoint_store_rechecks_each_artifact_id_with_fresh_authorizer(db, monkeypatch):
    own_id = "c" * 12
    revoked_id = "d" * 12
    session = chat_store.create_session(db, "alice")
    db.commit()
    allowed = _sys_ctx()
    decisions = iter((allowed, None))
    monkeypatch.setattr(
        tools,
        "refresh_runtime_context",
        lambda _db, _ctx: next(decisions),
    )
    monkeypatch.setattr(agent_files, "owner_of", lambda _artifact_id: "alice")

    chat_store.save_assistant_progress(
        session.id,
        None,
        "safe",
        [{
            "name": "write_report",
            "args": {},
            "args_are_shape": True,
            "artifact_ids": [own_id, revoked_id],
        }],
        stopped=False,
        artifact_authorizer=tools.fresh_artifact_authorizer(allowed),
    )

    db.expire_all()
    stored = db.scalar(select(ChatMessage).where(ChatMessage.session_id == session.id))
    assert stored.tools[0]["artifact_ids"] == [own_id]
    assert revoked_id not in json.dumps(stored.tools, ensure_ascii=False)


async def _stream_body(response) -> str:
    chunks: list[str] = []
    async for chunk in response.body_iterator:
        chunks.append(chunk.decode() if isinstance(chunk, bytes) else chunk)
    return "".join(chunks)


def test_session_worker_resanitizes_forged_runtime_trace_before_sse_and_db(
    db,
    monkeypatch,
    caplog,
):
    sentinel = "CUSTOMER-FORGED-RUNTIME-SENTINEL-8472"
    reasoning_canary = "SESSION-REASONING-CANARY-8472"
    error_canary = "SESSION-ERROR-TELEMETRY-CANARY-8472"
    artifact_id = "c" * 12
    session = chat_store.create_session(db, "alice")
    worker_ready = threading.Event()
    worker_release = threading.Event()

    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: True)

    def forged_runtime(*_args, **_kwargs):
        yield {"type": "thinking", "text": reasoning_canary}
        yield {"type": "delta", "text": "visible "}
        yield {"type": "delta", "text": "<th"}
        yield {"type": "delta", "text": f"ink>{reasoning_canary}"}
        yield {"type": "delta", "text": "</thi"}
        yield {"type": "delta", "text": "nk>answer"}
        yield {
            "type": "tool",
            "name": "search_parts",
            "args": {"query": sentinel},
        }
        worker_ready.set()
        assert worker_release.wait(timeout=5)
        yield {
            "type": "error",
            "message": error_canary,
            "kind": error_canary,
            "code": error_canary,
            "retriable": error_canary,
            "args": {"query": error_canary},
            "result": {"raw": error_canary},
            "debug": error_canary,
        }
        # Error is terminal at the session seam: these forged follow-up events must never run.
        yield {
            "type": "tool_done",
            "name": "search_parts",
            "ok": True,
            "artifact_ids": [artifact_id, sentinel],
        }
        yield {
            "type": "done",
            "answer": "ok",
            "tool_calls": [{"name": "search_parts", "args": {"query": sentinel}}],
        }

    monkeypatch.setattr(runtime, "run_stream", forged_runtime)

    response = chat_sessions.chat_stream(
        session.id,
        chat_sessions.SendMessageRequest(message="safe request"),
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    assert worker_ready.wait(timeout=5)
    attach_response = chat_sessions.chat_attach(
        session.id,
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    worker_release.set()
    body = asyncio.run(_stream_body(response))
    attach_body = asyncio.run(_stream_body(attach_response))

    db.expire_all()
    messages = chat_store.list_messages(db, session.id)
    assistant = next(message for message in messages if message["role"] == "assistant")
    serialized_tools = json.dumps(assistant["tools"], ensure_ascii=False)
    assert sentinel not in body
    assert sentinel not in attach_body
    assert sentinel not in serialized_tools
    assert sentinel not in caplog.text
    assert reasoning_canary not in body
    assert reasoning_canary not in attach_body
    assert reasoning_canary not in json.dumps(messages, ensure_ascii=False, default=str)
    assert reasoning_canary not in caplog.text
    assert error_canary not in body
    assert error_canary not in attach_body
    assert error_canary not in json.dumps(messages, ensure_ascii=False, default=str)
    assert error_canary not in caplog.text
    assert runtime.GENERIC_ERROR_CODE in body
    assert runtime.GENERIC_ERROR_CODE in attach_body
    assert "<think>" not in body
    assert "<think>" not in attach_body
    assert assistant["content"] == "visible answer"
    assert assistant["stopped"] is True
    assert assistant["tools"][0]["args"]["arg_keys"] == ["query"]
    assert "artifact_ids" not in assistant["tools"][0]
    assert artifact_id not in body
    assert '"type": "tool_done"' not in body
    assert '"type": "done"' not in body


def test_subscriber_limit_is_explicit_retry_event_on_initial_and_attach_streams(
    db,
    monkeypatch,
):
    session = chat_store.create_session(db, "alice")
    saturated = chat_sessions._RunHub()
    for _ in range(chat_sessions._MAX_SUBSCRIBERS_PER_RUN):
        assert saturated.subscribe()[1] is not None

    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: True)
    monkeypatch.setattr(chat_sessions, "acquire_session", lambda _sid: saturated)
    monkeypatch.setattr(chat_sessions, "get_run", lambda _sid: saturated)
    # The stream-path contract is under test; do not run an unrelated background model worker.
    monkeypatch.setattr(chat_sessions, "_start_agent_worker", lambda *_args, **_kwargs: None)

    initial = chat_sessions.chat_stream(
        session.id,
        chat_sessions.SendMessageRequest(message="safe request"),
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    initial_body = asyncio.run(_stream_body(initial))
    attach = chat_sessions.chat_attach(
        session.id,
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    attach_body = asyncio.run(_stream_body(attach))

    for body in (initial_body, attach_body):
        assert '"type": "subscriber_evicted"' in body
        assert '"retry_attach": true' in body
        assert '"type": "done"' not in body


def test_session_worker_caps_forged_cumulative_visible_buffer(db, monkeypatch):
    session = chat_store.create_session(db, "alice")
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: True)

    def forged_runtime(*_args, **_kwargs):
        yield {"type": "delta", "text": "x" * (limits.MAX_VISIBLE_RUN_BYTES + 1)}
        yield {"type": "done", "answer": "unreachable", "tool_calls": []}

    monkeypatch.setattr(runtime, "run_stream", forged_runtime)

    response = chat_sessions.chat_stream(
        session.id,
        chat_sessions.SendMessageRequest(message="safe request"),
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    body = asyncio.run(_stream_body(response))

    assert runtime.MODEL_OUTPUT_BUDGET_CODE in body
    assert '"type": "delta"' not in body
    assert '"type": "done"' not in body


def test_session_checkpoint_outage_has_bounded_attempts_and_logs(db, monkeypatch):
    session = chat_store.create_session(db, "alice")
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: True)
    monkeypatch.setattr(chat_sessions, "_CHECKPOINT_BYTES", 1)
    monkeypatch.setattr(chat_sessions, "_CHECKPOINT_SECONDS", 10_000.0)
    monkeypatch.setattr(chat_sessions, "_MAX_CHECKPOINT_FAILURES", 3)
    attempts = 0
    log_calls: list[tuple] = []

    def fail_checkpoint(*_args, **_kwargs):
        nonlocal attempts
        attempts += 1
        raise RuntimeError("database unavailable")

    def forged_runtime(*_args, **_kwargs):
        for _ in range(20):
            yield {
                "type": "delta",
                "text": "x" * limits.FIRST_STREAM_DELTA_BATCH_BYTES,
            }
        yield {"type": "done", "answer": "forged", "tool_calls": []}

    monkeypatch.setattr(chat_store, "save_assistant_progress", fail_checkpoint)
    monkeypatch.setattr(runtime, "run_stream", forged_runtime)
    monkeypatch.setattr(
        chat_sessions._log,
        "error",
        lambda *args, **kwargs: log_calls.append((args, kwargs)),
    )

    response = chat_sessions.chat_stream(
        session.id,
        chat_sessions.SendMessageRequest(message="safe request"),
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    asyncio.run(_stream_body(response))

    assert attempts == chat_sessions._MAX_CHECKPOINT_FAILURES + 1  # final attempt is allowed
    assert len(log_calls) == attempts
    assert all("checkpoint assistant message failed" in call[0][0] for call in log_calls)


def test_session_history_is_not_saved_or_sent_to_approved_external_model(
    db,
    monkeypatch,
):
    """Legacy history has no provenance, so every prompt is customer-file sensitivity in v1."""
    session = chat_store.create_session(db, "alice")
    chat_store.append_message(db, session, "user", "legacy customer workbook excerpt")
    chat_store.append_message(db, session, "assistant", "legacy derived answer")
    before = chat_store.list_messages(db, session.id)

    settings = _egress_settings(
        llm_trust_zone="approved_external",
        llm_base_url="https://approved-primary.test/v1",
        llm_private_base_urls="",
        llm_approved_external_base_urls="https://approved-primary.test",
        agent_external_file_egress_enabled=True,
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    # A deployment-wide flag cannot prove consent by the current user. v1 blocks all
    # approved-external conversation/file egress, including legacy history without provenance.
    assert tools.primary_model_call_allowed() is False
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(
        provider,
        "chat_stream",
        lambda *_args, **_kwargs: pytest.fail("external history reached provider"),
    )

    response = chat_sessions.chat_stream(
        session.id,
        chat_sessions.SendMessageRequest(message="new question"),
        db=db,
        ident={"sub": "alice", "authn": "sys_user"},
        ctx=_sys_ctx(),
    )
    body = asyncio.run(_stream_body(response))

    db.expire_all()
    after = chat_store.list_messages(db, session.id)
    assert after == before
    assert "AGENT_MODEL_EGRESS_DENIED" in body
    assert not chat_sessions.is_generating(session.id)


def test_prompt_context_policy_is_customer_file_fail_closed():
    private = _egress_settings(agent_external_file_egress_enabled=False)
    approved = _egress_settings(
        llm_trust_zone="approved_external",
        llm_base_url="https://approved-primary.test/v1",
        llm_private_base_urls="",
        llm_approved_external_base_urls="https://approved-primary.test",
        agent_external_file_egress_enabled=False,
    )

    assert tools.primary_model_call_allowed(private) is True
    assert tools.primary_model_call_allowed(approved) is False
    approved.agent_external_file_egress_enabled = True
    assert tools.primary_model_call_allowed(approved) is False


def test_stateless_stream_projects_forged_runtime_error_before_sse(monkeypatch, caplog):
    canary = "STATELESS-ERROR-TELEMETRY-CANARY-8472"
    request = agent_api.ChatRequest(messages=[
        agent_api.ChatMessage(role="user", content="safe request"),
    ])
    monkeypatch.setattr(agent_api, "record_access_log", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(runtime, "primary_model_call_allowed", lambda: True)
    monkeypatch.setattr(
        runtime,
        "run_stream",
        lambda *_args, **_kwargs: iter([{
            "type": "error",
            "message": canary,
            "kind": canary,
            "code": canary,
            "retriable": canary,
            "args": {"raw": canary},
            "result": canary,
        }]),
    )

    response = agent_api.chat_stream(request, db=None, role="sales", ctx=_sys_ctx())
    body = asyncio.run(_stream_body(response))

    assert canary not in body
    assert canary not in caplog.text
    assert runtime.GENERIC_ERROR_CODE in body
    assert runtime.GENERIC_ERROR_MESSAGE in body


def test_user_context_preserves_verified_authentication_provenance(monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    monkeypatch.setattr(
        auth,
        "verify_token_db",
        lambda token, db: {
            "sub": "alice",
            "role": "sales",
            "name": "Alice",
            "perms": {"page_chat": True},
            "authn": "sys_user",
            "tv": 7,
        },
    )
    creds = HTTPAuthorizationCredentials(scheme="Bearer", credentials="signed-token")

    ctx = security.get_current_user_context(creds, db=None)

    assert ctx.authn == "sys_user"
    assert ctx.user_id == "alice"
    assert ctx.token_version == 7


def test_dispatch_reloads_current_role_permissions_and_salesperson(db, monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    user = SysUser(
        username="fresh-agent-user",
        role="sales",
        salesperson_name="Current Sales Row Subject",
        password_hash=auth.hash_password("pw123456"),
        permissions={"page_inventory": True},
        token_version=4,
    )
    db.add(user)
    db.commit()
    seen: list[security.UserContext] = []

    def capture(_db, _query, **kwargs):
        seen.append(kwargs["user_ctx"])
        return {"items": [], "total": 0}

    monkeypatch.setattr(tools.inventory, "list_dynamic", capture)
    stale_snapshot = security.UserContext(
        user_id=user.username,
        role="admin",
        salesperson_name="Stale",
        permissions={"page_inventory": False},
        is_authenticated=True,
        authn="sys_user",
        token_version=4,
    )

    result = tools.dispatch(db, "get_inventory", {}, stale_snapshot)

    assert result == {"items": [], "total": 0}
    assert len(seen) == 1
    assert seen[0].role == "sales"
    assert seen[0].salesperson_name == "Current Sales Row Subject"
    assert seen[0].permissions["page_inventory"] is True


def test_runtime_stops_on_permission_epoch_change_and_new_run_uses_new_schemas(db, monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    user = SysUser(
        username="runtime-schema-user",
        role="sales",
        password_hash=auth.hash_password("pw123456"),
        permissions={"page_parts": True},
        token_version=2,
    )
    db.add(user)
    db.commit()
    ctx = security.UserContext(
        user_id=user.username,
        role="sales",
        permissions={"page_parts": True},
        is_authenticated=True,
        authn="sys_user",
        token_version=2,
    )
    schema_names: list[set[str]] = []
    provider_calls = 0

    def fake(_messages, schemas=None, **_kwargs):
        nonlocal provider_calls
        provider_calls += 1
        schema_names.append({
            schema["function"]["name"] for schema in (schemas or [])
        })
        if provider_calls == 1:
            yield "result", provider.ChatResult(
                content=None,
                tool_calls=[provider.ToolCall(
                    id="c",
                    name="list_skills",
                    arguments="{}",
                )],
            )
        else:
            yield "result", provider.ChatResult(content="done", tool_calls=[])

    def dispatch(*_args, **_kwargs):
        user.permissions = {"page_parts": False}
        db.commit()
        return {"items": []}

    monkeypatch.setattr(provider, "chat_stream", fake)
    monkeypatch.setattr(tools, "dispatch", dispatch)

    first = runtime.run(db, [{"role": "user", "content": "safe"}], ctx)
    second = runtime.run(db, [{"role": "user", "content": "safe"}], ctx)

    assert first["code"] == runtime.CAPABILITY_REVOKED_CODE
    assert second["answer"] == "done"
    assert provider_calls == 2
    assert "search_parts" in schema_names[0]
    assert "search_parts" not in schema_names[1]


def test_runtime_stops_before_next_provider_when_live_identity_is_disabled(db, monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    user = SysUser(
        username="runtime-revoked-user",
        role="sales",
        password_hash=auth.hash_password("pw123456"),
        permissions={"page_parts": True},
        token_version=3,
    )
    db.add(user)
    db.commit()
    ctx = security.UserContext(
        user_id=user.username,
        role="sales",
        permissions={"page_parts": True},
        is_authenticated=True,
        authn="sys_user",
        token_version=3,
    )
    provider_calls = 0

    def fake(_messages, _schemas=None, **_kwargs):
        nonlocal provider_calls
        provider_calls += 1
        if provider_calls > 1:
            pytest.fail("disabled identity reached a second provider call")
        yield "result", provider.ChatResult(
            content=None,
            tool_calls=[provider.ToolCall(
                id="c",
                name="search_parts",
                arguments='{"query":"PN"}',
            )],
        )

    def dispatch(*_args, **_kwargs):
        user.is_active = False
        db.commit()
        return {"items": []}

    monkeypatch.setattr(provider, "chat_stream", fake)
    monkeypatch.setattr(tools, "dispatch", dispatch)

    events = list(runtime.run_stream(db, [{"role": "user", "content": "safe"}], ctx))

    assert provider_calls == 1
    assert events[-1] == runtime.identity_stale_error_event()


def test_run_policy_lease_rejects_same_origin_model_drift_before_second_provider(monkeypatch):
    settings = _egress_settings()
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    provider_calls = 0

    def fake(_messages, _schemas=None, **_kwargs):
        nonlocal provider_calls
        provider_calls += 1
        if provider_calls > 1:
            pytest.fail("drifted run reached a second provider call")
        yield "result", provider.ChatResult(
            content=None,
            tool_calls=[provider.ToolCall(
                id="c",
                name="search_parts",
                arguments='{"query":"PN"}',
            )],
        )

    def dispatch(*_args, **_kwargs):
        settings.llm_model = "another-primary-model"
        return {"items": []}

    monkeypatch.setattr(provider, "chat_stream", fake)
    monkeypatch.setattr(tools, "dispatch", dispatch)

    events = list(runtime.run_stream(None, [{"role": "user", "content": "safe"}], _sys_ctx()))

    assert provider_calls == 1
    assert events[-1]["code"] == runtime.MODEL_EGRESS_DENIED_CODE


def test_midstream_field_permission_change_blocks_next_public_byte(db, monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    canary = "REVOKED-FIELD-DATA-CANARY-8472"
    user = SysUser(
        username="runtime-midstream-field-user",
        role="purchaser",
        password_hash=auth.hash_password("pw123456"),
        permissions={"page_chat": True, "data_purchase_cost": True},
        token_version=9,
    )
    db.add(user)
    db.commit()
    ctx = security.UserContext(
        user_id=user.username,
        role=user.role,
        permissions=user.permissions,
        is_authenticated=True,
        authn="sys_user",
        token_version=9,
    )

    def fake(_messages, _schemas=None, **_kwargs):
        yield "delta", "A" * limits.FIRST_STREAM_DELTA_BATCH_BYTES
        user.permissions = {"page_chat": True, "data_purchase_cost": False}
        db.commit()
        yield "delta", canary * 20
        yield "result", provider.ChatResult(content="unreachable", tool_calls=[])

    monkeypatch.setattr(provider, "chat_stream", fake)

    events = list(runtime.run_stream(db, [{"role": "user", "content": "safe"}], ctx))
    serialized = json.dumps(events, ensure_ascii=False)

    assert events[0]["type"] == "delta"
    assert events[-1] == runtime.capability_revoked_error_event()
    assert canary not in serialized


def test_handler_permission_revocation_blocks_result_before_message_append(db, monkeypatch):
    canary = "POST-HANDLER-PERMISSION-CANARY-8472"
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    user = SysUser(
        username="runtime-handler-revoke-user",
        role="sales",
        password_hash=auth.hash_password("pw123456"),
        permissions={"page_parts": True},
        token_version=5,
    )
    db.add(user)
    db.commit()
    ctx = security.UserContext(
        user_id=user.username,
        role="sales",
        permissions={"page_parts": True},
        is_authenticated=True,
        authn="sys_user",
        token_version=5,
    )
    provider_messages: list[list[dict]] = []

    def fake(messages, _schemas=None, **_kwargs):
        provider_messages.append(messages)
        yield "result", provider.ChatResult(
            content=None,
            tool_calls=[provider.ToolCall(
                id="c",
                name="search_parts",
                arguments='{"query":"PN"}',
            )],
        )

    def dispatch(*_args, **_kwargs):
        user.permissions = {"page_parts": False}
        db.commit()
        return {"data": canary}

    monkeypatch.setattr(provider, "chat_stream", fake)
    monkeypatch.setattr(tools, "dispatch", dispatch)

    events = list(runtime.run_stream(db, [{"role": "user", "content": "safe"}], ctx))

    assert len(provider_messages) == 1
    assert events[-1] == runtime.capability_revoked_error_event()
    assert canary not in json.dumps(provider_messages, ensure_ascii=False)
    assert canary not in json.dumps(events, ensure_ascii=False)


def test_revoked_result_is_purged_if_permission_changes_after_append(db, monkeypatch):
    canary = "POST-APPEND-PERMISSION-CANARY-8472"
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    user = SysUser(
        username="runtime-append-revoke-user",
        role="sales",
        password_hash=auth.hash_password("pw123456"),
        permissions={"page_parts": True},
        token_version=6,
    )
    db.add(user)
    db.commit()
    ctx = security.UserContext(
        user_id=user.username,
        role="sales",
        permissions={"page_parts": True},
        is_authenticated=True,
        authn="sys_user",
        token_version=6,
    )
    provider_calls = 0
    retained_messages: list[dict] | None = None
    original_append = provider.append_tool_result

    def fake(_messages, _schemas=None, **_kwargs):
        nonlocal provider_calls
        provider_calls += 1
        if provider_calls > 1:
            pytest.fail("revoked appended result reached provider")
        yield "result", provider.ChatResult(
            content=None,
            tool_calls=[provider.ToolCall(
                id="c",
                name="search_parts",
                arguments='{"query":"PN"}',
            )],
        )

    def append_then_revoke(messages, tool_call_id, content):
        nonlocal retained_messages
        retained_messages = messages
        message = original_append(messages, tool_call_id, content)
        user.permissions = {"page_parts": False}
        db.commit()
        return message

    monkeypatch.setattr(provider, "chat_stream", fake)
    monkeypatch.setattr(tools, "dispatch", lambda *_args, **_kwargs: {"data": canary})
    monkeypatch.setattr(provider, "append_tool_result", append_then_revoke)

    events = list(runtime.run_stream(db, [{"role": "user", "content": "safe"}], ctx))

    assert provider_calls == 1
    assert retained_messages is not None
    assert canary not in json.dumps(retained_messages, ensure_ascii=False)
    assert events[-1] == runtime.capability_revoked_error_event()


def test_shared_credentials_receive_no_capabilities_or_stateless_provider_bytes(
    db,
    monkeypatch,
):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    shared = security.UserContext(
        user_id="shared-readonly",
        role="readonly",
        is_authenticated=True,
        authn="shared",
        token_version=0,
    )
    request = agent_api.ChatRequest(messages=[
        agent_api.ChatMessage(role="user", content="safe request"),
    ])
    monkeypatch.setattr(agent_api, "record_access_log", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(provider, "is_configured", lambda: True)
    monkeypatch.setattr(
        provider,
        "chat_stream",
        lambda *_args, **_kwargs: pytest.fail("shared credential reached provider"),
    )

    result = agent_api.chat(request, db=db, role="readonly", ctx=shared)
    response = agent_api.chat_stream(request, db=db, role="readonly", ctx=shared)
    body = asyncio.run(_stream_body(response))

    assert tools.tools_for(shared) == []
    assert result["code"] == runtime.IDENTITY_STALE_CODE
    assert runtime.IDENTITY_STALE_CODE in body


def test_shared_credentials_cannot_open_persistent_session_chat(db, monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    monkeypatch.setattr(
        provider,
        "chat_stream",
        lambda *_args, **_kwargs: pytest.fail("shared session reached provider"),
    )

    with pytest.raises(HTTPException) as exc_info:
        chat_sessions.chat_stream(
            1,
            chat_sessions.SendMessageRequest(message="safe request"),
            db=db,
            ident={"sub": "shared-readonly", "authn": "shared"},
            ctx=security.UserContext(
                user_id="shared-readonly",
                role="readonly",
                is_authenticated=True,
                authn="shared",
            ),
        )

    assert exc_info.value.status_code == 403


@pytest.mark.parametrize("failure", ["missing", "inactive", "token_version", "db_error"])
def test_dispatch_identity_failures_are_stable_and_never_reach_handler(
    db,
    monkeypatch,
    failure,
):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    username = f"identity-{failure}"
    if failure != "missing" and failure != "db_error":
        db.add(SysUser(
            username=username,
            role="sales",
            password_hash=auth.hash_password("pw123456"),
            permissions={"page_parts": True},
            is_active=failure != "inactive",
            token_version=2 if failure == "token_version" else 1,
        ))
        db.commit()
    selected_db = SimpleNamespace(
        scalar=lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("db secret"))
    ) if failure == "db_error" else db
    calls = 0

    def forbidden(*_args, **_kwargs):
        nonlocal calls
        calls += 1
        return {"unexpected": True}

    monkeypatch.setattr(tools.part_resolver, "resolve", forbidden)
    ctx = security.UserContext(
        user_id=username,
        role="admin",
        permissions={"page_parts": True},
        is_authenticated=True,
        authn="sys_user",
        token_version=1,
    )

    result = tools.dispatch(selected_db, "search_parts", {"query": "PN"}, ctx)

    assert result == {
        "error": "身份状态已失效，请重新登录",
        "kind": "identity_denied",
        "code": "AGENT_IDENTITY_STALE",
        "retriable": False,
    }
    assert calls == 0


def test_policy_fingerprint_is_stable_across_registration_order():
    baseline = tools.capability_policy_fingerprint(tools.TOOL_SPECS)

    assert tools.CAPABILITY_POLICY_VERSION == "v1"
    assert tools.CAPABILITY_POLICY_FINGERPRINT == baseline
    assert len(baseline) == 64
    assert tools.capability_policy_fingerprint(tuple(reversed(tools.TOOL_SPECS))) == baseline


def test_canonical_tool_schema_is_deeply_immutable_and_fingerprint_stays_stable():
    spec = tools.TOOL_SPECS[0]
    baseline = tools.capability_policy_fingerprint()

    assert isinstance(spec.schema, MappingProxyType)
    assert isinstance(spec.schema["function"], MappingProxyType)
    assert isinstance(
        spec.schema["function"]["parameters"]["properties"],
        MappingProxyType,
    )
    with pytest.raises(TypeError):
        spec.schema["function"]["name"] = "attacker"
    with pytest.raises(TypeError):
        spec.schema["function"]["parameters"]["properties"]["rogue"] = {
            "type": "string",
        }

    assert tools.capability_policy_fingerprint() == baseline


def test_policy_fingerprint_rejects_unregistered_policy_metadata():
    first = tools.TOOL_SPECS[0]
    changed_edge = replace(
        first.egress[0],
        max_bytes=first.egress[0].max_bytes + 1,
    )
    changed = replace(first, egress=(changed_edge,))
    changed_specs = (changed, *tools.TOOL_SPECS[1:])

    assert tools._valid_egress(changed.egress) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint(changed_specs)


@pytest.mark.parametrize(
    ("field", "value"),
    [
        ("purpose", "alternate_purpose"),
        ("projection_id", "alternate_projection_v1"),
        ("max_bytes", 1),
        ("policy_version", "egress-v2"),
        ("retention_policy", ""),
        ("retention_policy", "unknown_retention"),
    ],
)
def test_unregistered_egress_contract_metadata_fails_closed(
    field,
    value,
):
    first = tools.TOOL_SPECS[0]
    changed_edge = replace(first.egress[0], **{field: value})
    changed = replace(first, egress=(changed_edge,))
    changed_specs = (changed, *tools.TOOL_SPECS[1:])

    assert tools._valid_egress(changed.egress) is False
    assert tools._edge_egress_allowed(changed_edge, settings=_egress_settings()) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint(changed_specs)


def test_top_level_sensitivity_must_equal_maximum_edge_sensitivity():
    first = tools.TOOL_SPECS[0]
    poisoned = replace(first, sensitivity=tools.DataSensitivity.INTERNAL)

    assert tools._allowed(poisoned, _sys_ctx()) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint((poisoned,))


def test_policy_fingerprint_changes_when_permission_policy_changes():
    first = tools.TOOL_SPECS[0]
    permission = tools._PagePermission("page_inventory")
    changed = replace(
        first,
        permission=permission,
        permission_id=permission.policy_id,
    )

    assert (
        tools.capability_policy_fingerprint((changed, *tools.TOOL_SPECS[1:]))
        != tools.CAPABILITY_POLICY_FINGERPRINT
    )


def test_policy_fingerprint_changes_when_function_parameters_change():
    first = tools.TOOL_SPECS[0]
    changed_schema = tools._thaw_json(first.schema)
    changed_schema["function"]["parameters"]["properties"]["new_filter"] = {
        "type": "string",
    }
    changed = replace(first, schema=changed_schema)

    assert (
        tools.capability_policy_fingerprint((changed, *tools.TOOL_SPECS[1:]))
        != tools.CAPABILITY_POLICY_FINGERPRINT
    )


def test_policy_fingerprint_covers_budget_handler_version_and_stable_subjects(monkeypatch):
    first = tools.TOOL_SPECS[0]
    changed_budget = replace(
        first,
        budget=replace(first.budget, max_payload_bytes=first.budget.max_payload_bytes + 1),
    )
    changed_version = replace(first, implementation_version="audit-test-v2")

    def replacement_handler(db, args, ctx):
        return {"ok": True}

    changed_handler = replace(first, handler=replacement_handler)

    assert (
        tools.capability_policy_fingerprint((changed_budget, *tools.TOOL_SPECS[1:]))
        != tools.CAPABILITY_POLICY_FINGERPRINT
    )
    assert (
        tools.capability_policy_fingerprint((changed_version, *tools.TOOL_SPECS[1:]))
        != tools.CAPABILITY_POLICY_FINGERPRINT
    )
    assert (
        tools.capability_policy_fingerprint((changed_handler, *tools.TOOL_SPECS[1:]))
        != tools.CAPABILITY_POLICY_FINGERPRINT
    )

    monkeypatch.setattr(
        tools,
        "STABLE_SUBJECT_EFFECTS",
        frozenset({tools.ToolEffect.FILE_READ}),
    )
    assert tools.capability_policy_fingerprint() != tools.CAPABILITY_POLICY_FINGERPRINT


def test_runtime_policy_fingerprint_is_normalized_and_secret_free():
    base = _egress_settings(
        llm_base_url="HTTPS://GPU0.TAILNET:443/v1",
        llm_private_base_urls="https://gpu0.tailnet/, https://backup.tailnet:8443",
        llm_api_key="SECRET-A",
        vision_api_key="VISION-SECRET-A",
    )
    equivalent = SimpleNamespace(
        **{
            **base.__dict__,
            "llm_base_url": "https://gpu0.tailnet:443/v1/",
            "llm_private_base_urls": "https://backup.tailnet:8443 https://GPU0.TAILNET:443",
            "llm_api_key": "SECRET-B",
            "vision_api_key": "VISION-SECRET-B",
        }
    )
    changed = SimpleNamespace(**{**base.__dict__, "agent_external_file_egress_enabled": True})
    changed_zone = SimpleNamespace(**{**base.__dict__, "llm_trust_zone": "approved_external"})
    changed_origin = SimpleNamespace(
        **{**base.__dict__, "llm_base_url": "https://another.tailnet:8443/v1"}
    )
    changed_path = SimpleNamespace(
        **{**base.__dict__, "llm_base_url": "https://gpu0.tailnet/another/path"}
    )
    changed_allowlist = SimpleNamespace(
        **{**base.__dict__, "llm_private_base_urls": "https://gpu0.tailnet"}
    )
    changed_vision_model = SimpleNamespace(
        **{**base.__dict__, "vision_model": "another-vision-model"}
    )

    fingerprint = tools.runtime_policy_fingerprint(base)

    assert len(fingerprint) == 64
    assert tools.runtime_policy_fingerprint(equivalent) == fingerprint
    assert tools.runtime_policy_fingerprint(changed) != fingerprint
    assert tools.runtime_policy_fingerprint(changed_zone) != fingerprint
    assert tools.runtime_policy_fingerprint(changed_origin) != fingerprint
    assert tools.runtime_policy_fingerprint(changed_path) != fingerprint
    assert tools.runtime_policy_fingerprint(changed_allowlist) != fingerprint
    assert tools.runtime_policy_fingerprint(changed_vision_model) != fingerprint
    assert "SECRET-A" not in fingerprint
    assert "VISION-SECRET-A" not in fingerprint


def test_private_provider_requires_exact_normalized_origin_allowlist(monkeypatch):
    ctx = security.UserContext(user_id=None, role="phase1_full_access")
    settings = _egress_settings(
        llm_base_url="https://api.deepseek.com/v1",
        llm_private_base_urls="",
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)

    assert tools.tools_for(ctx) == []
    assert (
        tools.dispatch(None, "search_parts", {"query": "PN"}, ctx)["kind"]
        == "model_context_egress_denied"
    )

    settings.llm_base_url = "https://gpu0.tailnet:8443/v1"
    settings.llm_private_base_urls = "https://GPU0.TAILNET:8443/"
    assert "search_parts" in {schema["function"]["name"] for schema in tools.tools_for(ctx)}


def test_operator_asserted_private_origin_is_not_attestation_or_prod_file_release():
    settings = _egress_settings(
        environment="prod",
        agent_allow_unattested_private_for_development=True,
        llm_base_url="https://public-but-mislabelled.example/v1",
        llm_private_base_urls="https://public-but-mislabelled.example",
    )
    metadata = tools._destination_policy_metadata(
        settings,
        tools.EgressDestination.PRIMARY_MODEL,
    )

    assert metadata["private_trust_basis"] == "operator_assertion_only_v1"
    assert metadata["private_endpoint_identity_attested"] is False
    assert metadata["operator_asserted_private_origins"] == (
        "https://public-but-mislabelled.example",
    )
    # The dev/test escape hatch is ignored in prod. Exact origin alone never authorizes the
    # customer-file-classified conversation edge or Vision equivalent.
    assert tools.primary_model_call_allowed(settings) is False
    assert tools._unattested_private_development_allowed(settings) is False


def test_unattested_private_override_is_explicitly_dev_test_only():
    development = _egress_settings(
        environment="dev",
        agent_allow_unattested_private_for_development=True,
    )
    missing_override = _egress_settings(
        environment="dev",
        agent_allow_unattested_private_for_development=False,
    )

    assert tools.primary_model_call_allowed(development) is True
    assert tools.primary_model_call_allowed(missing_override) is False


@pytest.mark.parametrize(
    "base_url",
    [
        "https://user:password@gpu0.tailnet:8443/v1",
        "https://gpu0.tailnet:8443/v1?token=secret",
        "https://gpu0.tailnet:8443/v1#fragment",
        "https://10.0.0.5:8443/v1",
    ],
)
def test_private_provider_rejects_unsafe_or_unlisted_origins(monkeypatch, base_url):
    settings = _egress_settings(
        llm_base_url=base_url,
        llm_private_base_urls="https://gpu0.tailnet:8443",
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)

    assert tools.tools_for(security.UserContext(user_id=None, role="phase1_full_access")) == []


@pytest.mark.parametrize(
    "allowlist",
    [
        "https://user:password@gpu0.tailnet:8443",
        "https://gpu0.tailnet:8443?token=secret",
        "https://gpu0.tailnet:8443#fragment",
        "https://gpu0.tailnet:8443/v1",
    ],
)
def test_private_provider_rejects_unsafe_allowlist_entries(monkeypatch, allowlist):
    settings = _egress_settings(
        llm_base_url="https://gpu0.tailnet:8443/v1",
        llm_private_base_urls=allowlist,
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)

    assert tools.tools_for(security.UserContext(user_id=None, role="phase1_full_access")) == []


def test_non_loopback_http_is_rejected_even_in_private_zone(monkeypatch):
    settings = _egress_settings(
        environment="dev",
        agent_allow_loopback_http=True,
        llm_base_url="http://gpu0.tailnet:8000/v1",
        llm_private_base_urls="http://gpu0.tailnet:8000",
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)

    assert tools.tools_for(security.UserContext(user_id=None, role="phase1_full_access")) == []


def test_loopback_http_requires_explicit_dev_only_override(monkeypatch):
    settings = _egress_settings(
        environment="dev",
        agent_allow_loopback_http=False,
        llm_base_url="http://127.0.0.1:8000/v1",
        llm_private_base_urls="http://127.0.0.1:8000",
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    assert tools.tools_for(ctx) == []
    settings.agent_allow_loopback_http = True
    assert "search_parts" in {
        schema["function"]["name"] for schema in tools.tools_for(ctx)
    }
    settings.environment = "prod"
    assert tools.tools_for(ctx) == []


def test_approved_external_requires_exact_origin_allowlist(monkeypatch):
    settings = _egress_settings(
        llm_trust_zone="approved_external",
        llm_base_url="https://api.deepseek.com/v1",
        llm_private_base_urls="",
        llm_approved_external_base_urls="",
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    assert tools.tools_for(ctx) == []
    settings.llm_approved_external_base_urls = "https://api.deepseek.com"
    assert "search_parts" in {
        schema["function"]["name"] for schema in tools.tools_for(ctx)
    }


def test_vision_is_independent_and_both_vision_edges_must_pass(monkeypatch):
    settings = _egress_settings(vision_trust_zone="unknown")
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    ctx = _sys_ctx()

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    assert "read_document" in visible
    assert "read_document_with_vision" not in visible

    settings.vision_trust_zone = "private"
    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    assert "read_document_with_vision" in visible

    settings.llm_trust_zone = "approved_external"
    settings.llm_private_base_urls = ""
    settings.llm_approved_external_base_urls = "https://agent-private.test:8000"
    settings.agent_external_file_egress_enabled = False
    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    assert "read_document_with_vision" not in visible


def test_tool_arguments_are_validated_before_handler(monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    monkeypatch.setattr(
        tools.part_resolver,
        "resolve",
        lambda *_args, **_kwargs: pytest.fail("invalid arguments must not reach handler"),
    )
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    undeclared = tools.dispatch(None, "search_parts", {"query": "PN", "write": True}, ctx)
    oversized = tools.dispatch(None, "search_parts", {"query": "X" * 501}, ctx)
    over_limit = tools.dispatch(None, "search_parts", {"query": "PN", "limit": 21}, ctx)

    assert undeclared == {
        "error": "工具参数不符合安全约束",
        "kind": "validation_error",
        "code": "AGENT_TOOL_ARGS_INVALID",
        "retriable": False,
    }
    for result in (oversized, over_limit):
        assert result["kind"] == "validation_error"
        assert result["code"] == "AGENT_TOOL_BUDGET_EXCEEDED"
        assert result["retriable"] is False


@pytest.mark.parametrize(
    "args",
    [
        {"dimension": "part", "date_from": "2026-02-30"},
        {"dimension": "part", "date_from": "2026-02-01T00:00:00"},
        {
            "dimension": "part",
            "date_from": "2026-03-02",
            "date_to": "2026-03-01",
        },
    ],
)
def test_profit_ranking_invalid_dates_fail_closed_before_query(monkeypatch, args):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    monkeypatch.setattr(
        tools.profit,
        "aggregate",
        lambda *_args, **_kwargs: pytest.fail("invalid date reached profit query"),
    )

    result = tools.dispatch(
        None,
        "get_profit_ranking",
        args,
        security.UserContext(user_id=None, role="phase1_full_access"),
    )

    assert result == {
        "error": "工具参数不符合安全约束",
        "kind": "validation_error",
        "code": "AGENT_TOOL_ARGS_INVALID",
        "retriable": False,
    }


def test_profit_ranking_accepts_real_leap_day_and_preserves_typed_range(monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    captured: dict[str, object] = {}

    def aggregate(_db, dimension, date_from, date_to, include_tax, _ctx):
        captured.update(
            dimension=dimension,
            date_from=date_from,
            date_to=date_to,
            include_tax=include_tax,
        )
        return {"rows": []}

    monkeypatch.setattr(tools.profit, "aggregate", aggregate)

    result = tools.dispatch(
        None,
        "get_profit_ranking",
        {
            "dimension": "part",
            "date_from": "2024-02-29",
            "date_to": "2024-03-01",
        },
        security.UserContext(user_id=None, role="phase1_full_access"),
    )

    assert result == {"rows": []}
    assert captured == {
        "dimension": "part",
        "date_from": date(2024, 2, 29),
        "date_to": date(2024, 3, 1),
        "include_tax": False,
    }


@pytest.mark.parametrize(
    "args",
    [
        {"cells": []},
        {"cells": [{"row": 1, "col": "A", "value": ["not", "scalar"]}]},
        {"cells": [{"row": 1, "col": "A"}]},
        {"cells": [{"row": 0, "col": "A", "value": "x"}]},
        {"cells": [{"row": 1_048_577, "col": "A", "value": "x"}]},
        {"cells": [{"row": 1, "col": 0, "value": "x"}]},
        {"cells": [{"row": 1, "col": 16_385, "value": "x"}]},
        {"cells": [{"row": 1, "col": "XFE", "value": "x"}]},
        {"cells": [{"row": 1, "col": "A1", "value": "x"}]},
        {
            "base_file_id": "not-a-file-id",
            "cells": [{"row": 1, "col": "A", "value": "x"}],
        },
    ],
)
def test_excel_arguments_reject_empty_non_scalar_out_of_range_and_bad_ids(
    monkeypatch,
    args,
):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    monkeypatch.setattr(
        agent_files,
        "write_excel",
        lambda *_args, **_kwargs: pytest.fail("invalid arguments reached file service"),
    )

    result = tools.dispatch(None, "write_excel", args, _sys_ctx())

    assert result == {
        "error": "工具参数不符合安全约束",
        "kind": "validation_error",
        "code": "AGENT_TOOL_ARGS_INVALID",
        "retriable": False,
    }


def test_well_formed_uuid_file_reference_reaches_owner_gate(monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    result = tools.dispatch(
        None,
        "read_document",
        {"file_id": "123e4567-e89b-12d3-a456-426614174000"},
        _sys_ctx(),
    )

    assert result == tools._NO_ACCESS


def test_fileerror_is_stable_non_retriable_and_never_internal(monkeypatch, caplog):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    secret = "CUSTOMER-FILE-DETAIL-8472"
    monkeypatch.setattr(
        agent_files,
        "write_report",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(agent_files.FileError(secret)),
    )

    result = tools.dispatch(
        None,
        "write_report",
        {"headers": ["PN"], "rows": [["A"]]},
        _sys_ctx(),
    )

    assert result == {
        "error": "文件参数、格式或状态不符合要求",
        "kind": "file_error",
        "code": "AGENT_FILE_REJECTED",
        "retriable": False,
    }
    assert secret not in caplog.text


def test_excel_and_report_strings_are_literal_while_numbers_stay_numeric(db):
    excel = agent_files.write_excel(
        None,
        None,
        [
            {"row": 1, "col": "A", "value": "=1+1"},
            {"row": 2, "col": "A", "value": "+SUM(A1:A2)"},
            {"row": 3, "col": "A", "value": "-2+3"},
            {"row": 4, "col": "A", "value": "@cmd"},
            {"row": 5, "col": "A", "value": 42},
        ],
        "literal.xlsx",
        "alice",
    )
    excel_path, _ = agent_files.get_download(excel["file_id"])
    wb = load_workbook(excel_path, data_only=False)
    ws = wb.active
    assert [(ws.cell(row=i, column=1).value, ws.cell(row=i, column=1).data_type)
            for i in range(1, 5)] == [
        ("=1+1", "s"),
        ("+SUM(A1:A2)", "s"),
        ("-2+3", "s"),
        ("@cmd", "s"),
    ]
    assert ws["A5"].value == 42 and ws["A5"].data_type == "n"
    wb.close()

    report = agent_files.write_report(
        "=danger-title",
        ["@header", "Number"],
        [["+payload", 3.5]],
        "literal-report.xlsx",
        "alice",
    )
    report_path, _ = agent_files.get_download(report["file_id"])
    wb = load_workbook(report_path, data_only=False)
    ws = wb.active
    assert ws["A1"].value == "=danger-title" and ws["A1"].data_type == "s"
    assert ws["A2"].value == "@header" and ws["A2"].data_type == "s"
    assert ws["A3"].value == "+payload" and ws["A3"].data_type == "s"
    assert ws["B3"].value == 3.5 and ws["B3"].data_type == "n"
    wb.close()


def test_all_capabilities_reject_undeclared_arguments_before_dispatch(monkeypatch):
    ctx = _sys_ctx(role="admin")

    for spec in tools.TOOL_SPECS:
        result = tools.dispatch(None, spec.name, {"__undeclared": True}, ctx)
        assert result.get("code") == "AGENT_TOOL_ARGS_INVALID", (spec.name, result)
        assert result["retriable"] is False, spec.name


def test_public_trace_and_artifact_projections_are_hard_bounded():
    artifact_ids = [f"{index:012x}" for index in range(
        limits.MAX_ARTIFACT_IDS_PER_TRACE_ENTRY + 20
    )]
    trace = [{
        "name": "search_parts",
        "args": {
            "arg_count": 1,
            "arg_keys": ["query"],
            "string_lengths": {"query": 2},
        },
        "args_are_shape": True,
        "artifact_ids": artifact_ids,
    } for _ in range(limits.MAX_PUBLIC_TRACE_ENTRIES + 20)]

    clean = tools.sanitize_tool_trace(trace)

    assert len(clean) == limits.MAX_PUBLIC_TRACE_ENTRIES
    assert all(
        len(entry.get("artifact_ids", []))
        <= limits.MAX_ARTIFACT_IDS_PER_TRACE_ENTRY
        for entry in clean
    )


def test_nan_and_infinity_fail_before_handler_projection_or_error_sse(monkeypatch):
    monkeypatch.setattr(
        agent_files,
        "write_excel",
        lambda *_args, **_kwargs: pytest.fail("non-finite cell reached handler"),
    )

    result = tools.dispatch(
        None,
        "write_excel",
        {"cells": [{"row": 1, "col": "A", "value": float("nan")}]},
        _sys_ctx(),
    )
    public_error = runtime.project_error_event({
        "type": "error",
        "message": float("nan"),
        "code": float("inf"),
    })

    assert result["code"] == "AGENT_TOOL_ARGS_INVALID"
    assert tools.serialize_tool_result_for_model(
        "search_parts", {"value": float("nan")}
    ) is None
    assert "NaN" not in json.dumps(public_error, allow_nan=False)
    assert "Infinity" not in json.dumps(public_error, allow_nan=False)


def test_blank_multi_page_pdf_is_detected_from_real_content_only(monkeypatch, tmp_path):
    class BlankPage:
        @staticmethod
        def extract_text():
            return ""

        @staticmethod
        def extract_tables():
            return []

    class FakePdf:
        pages = [BlankPage() for _ in range(30)]

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

    import pdfplumber

    monkeypatch.setattr(pdfplumber, "open", lambda *_args, **_kwargs: FakePdf())

    text, scanned = agent_files._read_pdf(tmp_path / "blank.pdf")

    assert "[第30页]" in text
    assert scanned is True


def test_existing_tools_have_explicit_effects_sensitivity_and_egress():
    """Every model-visible tool comes from one typed, explicitly classified source."""
    expected_names = {schema["function"]["name"] for schema in tools.TOOLS}
    expected_effects = {
        "search_parts": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_part_overview": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "inspect_file": frozenset({tools.ToolEffect.FILE_READ}),
        "read_file_rows": frozenset({tools.ToolEffect.FILE_READ}),
        "lookup_prices_bulk": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "write_excel": frozenset({
            tools.ToolEffect.FILE_READ,
            tools.ToolEffect.ARTIFACT_CREATE,
        }),
        "read_document": frozenset({tools.ToolEffect.FILE_READ}),
        "read_document_with_vision": frozenset({tools.ToolEffect.FILE_READ}),
        "write_report": frozenset({tools.ToolEffect.ARTIFACT_CREATE}),
        "list_recent_purchases": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_profit_ranking": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_purchase_analysis": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_inventory": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_maintenance_board": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_maintenance_projects": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_maintenance_lines": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_cancellation_stats": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "list_skills": frozenset({tools.ToolEffect.BUSINESS_READ}),
        "get_skill": frozenset({tools.ToolEffect.BUSINESS_READ}),
    }
    customer_file_tools = {
        "inspect_file",
        "read_file_rows",
        "write_excel",
        "read_document",
        "read_document_with_vision",
    }
    internal_tools = {"list_skills", "get_skill"}
    expected_sensitivity = {
        name: (
            tools.DataSensitivity.CUSTOMER_FILE
            if name in customer_file_tools
            else tools.DataSensitivity.INTERNAL
            if name in internal_tools
            else tools.DataSensitivity.BUSINESS_CONFIDENTIAL
        )
        for name in expected_names
    }
    expected_egress = {
        name: tools._primary_egress(
            sensitivity,
            (
                tools.EgressSource.CUSTOMER_FILE
                if name in customer_file_tools
                else tools.EgressSource.TOOL_RESULT
            ),
        )
        for name, sensitivity in expected_sensitivity.items()
    }
    expected_egress["read_document_with_vision"] = tools._vision_egress()
    expected_permissions = {
        "search_parts": "page:page_parts",
        "get_part_overview": "page:page_parts",
        "inspect_file": "allow",
        "read_file_rows": "allow",
        "lookup_prices_bulk": "page:page_parts",
        "write_excel": "allow",
        "read_document": "allow",
        "read_document_with_vision": "allow",
        "write_report": "allow",
        "list_recent_purchases": "page:page_purchases",
        "get_profit_ranking": "page:page_profit:deny_scoped_sales",
        "get_purchase_analysis": "page:page_purchases",
        "get_inventory": "page:page_inventory",
        "get_maintenance_board": "page:page_maintenance:deny_scoped_sales",
        "get_maintenance_projects": "page:page_maintenance",
        "get_maintenance_lines": "page:page_maintenance",
        "get_cancellation_stats": "page:page_purchases",
        "list_skills": "allow",
        "get_skill": "allow",
    }

    assert len(expected_names) == 19
    assert {spec.name for spec in tools.TOOL_SPECS} == expected_names
    assert {spec.name: spec.effects for spec in tools.TOOL_SPECS} == expected_effects
    assert {spec.name: spec.egress for spec in tools.TOOL_SPECS} == expected_egress
    assert {spec.name: spec.sensitivity for spec in tools.TOOL_SPECS} == expected_sensitivity
    assert {spec.name: spec.permission_id for spec in tools.TOOL_SPECS} == expected_permissions
    assert set(tools._REGISTRY) == expected_names
    assert all(isinstance(spec.effects, frozenset) and spec.effects for spec in tools.TOOL_SPECS)
    assert all(
        all(effect in tools.ALLOWED_TOOL_EFFECTS for effect in spec.effects)
        for spec in tools.TOOL_SPECS
    )
    assert all(tools._valid_egress(spec.egress) for spec in tools.TOOL_SPECS)
    assert all(
        spec.sensitivity in tools.ALLOWED_DATA_SENSITIVITIES
        for spec in tools.TOOL_SPECS
    )
    assert all(callable(spec.permission) for spec in tools.TOOL_SPECS)
    assert all(isinstance(spec.enabled, bool) for spec in tools.TOOL_SPECS)
    assert all(isinstance(spec.budget, tools.ToolBudget) for spec in tools.TOOL_SPECS)
    assert all(callable(spec.validator) for spec in tools.TOOL_SPECS)
    assert all(spec.implementation_version for spec in tools.TOOL_SPECS)
    assert isinstance(tools._REGISTRY, MappingProxyType)
    assert isinstance(tools._SPEC_BY_NAME, MappingProxyType)
    for spec in tools.TOOL_SPECS:
        assert spec.name == spec.schema["function"]["name"]
        assert tools._REGISTRY[spec.name] is spec.handler

    assert not hasattr(tools._REGISTRY, "__setitem__")
    assert not hasattr(tools._SPEC_BY_NAME, "__setitem__")


def test_tool_budgets_cover_required_resource_dimensions():
    budgets = {spec.name: spec.budget for spec in tools.TOOL_SPECS}

    assert budgets["search_parts"].max_query_chars == tools._QUERY_CHARS_MAX
    assert budgets["get_part_overview"].max_pn_chars == tools._PN_CHARS_MAX
    assert budgets["search_parts"].max_limit == tools._SEARCH_LIMIT_MAX
    assert budgets["list_recent_purchases"].max_days == tools._RECENT_DAYS_MAX
    assert budgets["get_maintenance_lines"].max_page is not None
    assert budgets["read_file_rows"].max_rows == tools._READ_ROWS_MAX
    assert budgets["lookup_prices_bulk"].max_items == tools._BULK_MAX
    assert budgets["write_excel"].max_cells == agent_files._MAX_WRITE_CELLS
    assert budgets["inspect_file"].max_sheets == agent_files._MAX_INSPECT_SHEETS
    assert budgets["write_report"].max_output_name_chars is not None
    assert (
        max(budget.max_payload_bytes for budget in budgets.values())
        <= limits.MAX_TOOL_ARGUMENT_BYTES_PER_CALL
    )


def test_tools_for_returns_defensive_schema_copies():
    ctx = security.UserContext(user_id=None, role="phase1_full_access")
    first = tools.tools_for(ctx)
    original_name = first[0]["function"]["name"]
    first[0]["function"]["name"] = "attacker_mutated_name"
    first[0]["function"]["parameters"]["properties"].clear()

    second = tools.tools_for(ctx)
    by_name = {item["function"]["name"]: item for item in second}

    assert original_name in by_name
    assert by_name[original_name]["function"]["parameters"]["properties"]
    assert tools._SPEC_BY_NAME[original_name].name == original_name


def test_dispatch_executes_toolspec_handler_without_registry_lookup(monkeypatch):
    monkeypatch.setattr(tools, "_REGISTRY", None)
    monkeypatch.setattr(
        tools.part_resolver,
        "resolve",
        lambda *_args, **_kwargs: {"source": "immutable-spec"},
    )
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    assert tools.dispatch(None, "search_parts", {"query": "PN"}, ctx) == {
        "source": "immutable-spec"
    }


@pytest.mark.parametrize("tool_name", ["read_document", "write_excel", "write_report"])
def test_file_and_artifact_capabilities_require_stable_subject(monkeypatch, tool_name):
    service_name = {
        "read_document": "read_document",
        "write_excel": "write_excel",
        "write_report": "write_report",
    }[tool_name]
    monkeypatch.setattr(
        agent_files,
        service_name,
        lambda *_args, **_kwargs: pytest.fail("denied capability reached handler"),
    )
    shared = security.UserContext(
        user_id="admin",
        role="admin",
        is_authenticated=True,
        authn="shared",
    )
    visible = {schema["function"]["name"] for schema in tools.tools_for(shared)}

    result = tools.dispatch(None, tool_name, {}, shared)

    assert tool_name not in visible
    assert result["kind"] == "capability_denied"


def test_stable_subject_can_see_local_file_and_artifact_capabilities():
    visible = {schema["function"]["name"] for schema in tools.tools_for(_sys_ctx())}

    assert {"read_document", "write_excel", "write_report"} <= visible


def test_denied_capability_is_hidden_and_cannot_be_dispatched(monkeypatch):
    """Hiding a schema is UX; dispatch must independently enforce the same policy."""
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    monkeypatch.setattr(tools, "_reload_dispatch_context", lambda _db, ctx: ctx)
    ctx = security.UserContext(
        user_id="sales-1",
        role="sales",
        permissions={"page_chat": True, "page_parts": True, "page_inventory": False},
        is_authenticated=True,
        authn="sys_user",
    )
    monkeypatch.setattr(
        tools.inventory,
        "list_dynamic",
        lambda *_args, **_kwargs: pytest.fail("denied capability reached handler"),
    )

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    assert "search_parts" in visible
    assert "get_inventory" not in visible

    result = tools.dispatch(None, "get_inventory", {}, ctx)
    assert result["kind"] == "capability_denied"
    assert "无权限" in result["error"]


def test_runtime_only_sends_context_allowed_schemas_to_model(monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", True)
    monkeypatch.setattr(tools, "refresh_runtime_context", lambda _db, ctx: ctx)
    ctx = security.UserContext(
        user_id="sales-1",
        role="sales",
        permissions={"page_chat": True, "page_parts": True, "page_inventory": False},
        is_authenticated=True,
        authn="sys_user",
    )
    captured: list[dict] = []

    def fake_chat_stream(messages, schemas=None, **_kwargs):
        captured.extend(schemas or [])
        yield "result", provider.ChatResult(content="ok", tool_calls=[])

    monkeypatch.setattr(provider, "chat_stream", fake_chat_stream)

    assert runtime.run(None, [{"role": "user", "content": "hi"}], ctx)["answer"] == "ok"
    names = {schema["function"]["name"] for schema in captured}
    assert "search_parts" in names
    assert "get_inventory" not in names


@pytest.mark.parametrize(
    ("trust_zone", "model_context_enabled"),
    [("unknown", True), ("private", False), ("approved_external", False)],
)
def test_model_context_requires_known_zone_and_explicit_opt_in(
    monkeypatch,
    trust_zone,
    model_context_enabled,
):
    settings = _egress_settings(
        llm_trust_zone=trust_zone,
        agent_model_context_egress_enabled=model_context_enabled,
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    ctx = security.UserContext(user_id=None, role="phase1_full_access")
    monkeypatch.setattr(
        tools.part_resolver,
        "resolve",
        lambda *_args, **_kwargs: pytest.fail("denied capability reached handler"),
    )

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    result = tools.dispatch(None, "search_parts", {"query": "PN"}, ctx)

    assert "search_parts" not in visible
    assert result["kind"] == "model_context_egress_denied"


def test_registry_projection_cannot_be_extended_to_bypass_capability_policy():
    assert "rogue_business_write" not in tools._REGISTRY
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    result = tools.dispatch(None, "rogue_business_write", {}, ctx)
    assert result["kind"] == "capability_denied"


@pytest.mark.parametrize(
    "bad_effects",
    [None, frozenset(), frozenset({"business_write"}), {tools.ToolEffect.BUSINESS_READ}],
)
def test_unclassified_or_business_write_effect_is_rejected(bad_effects):
    original = tools._SPEC_BY_NAME["search_parts"]
    poisoned = replace(original, effects=bad_effects)
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    assert tools._allowed(poisoned, ctx) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint((poisoned,))


def test_write_excel_composite_effects_are_audited(monkeypatch):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    events: list[dict] = []
    monkeypatch.setattr(
        security,
        "record_access_log",
        lambda ctx, action, resource, filters=None: events.append(filters or {}),
    )
    monkeypatch.setattr(agent_files, "write_excel", lambda *_args: {"ok": True})
    ctx = _sys_ctx()

    assert tools.dispatch(
        None,
        "write_excel",
        {"cells": [{"row": 1, "col": "A", "value": "ok"}]},
        ctx,
    ) == {"ok": True}
    assert events[0]["effects"] == ["artifact_create", "file_read"]
    assert events[0]["sensitivity"] == "customer_file"


def test_undeclared_egress_is_rejected():
    original = tools._SPEC_BY_NAME["search_parts"]
    poisoned = replace(original, egress=None)
    ctx = _sys_ctx(role="phase1_full_access")

    assert tools._allowed(poisoned, ctx) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint((poisoned,))


def test_empty_egress_is_structurally_valid_but_never_model_visible_or_dispatchable():
    original = tools._SPEC_BY_NAME["search_parts"]
    no_result_edge = replace(original, egress=())
    ctx = _sys_ctx(role="phase1_full_access")

    assert tools._valid_egress(()) is True
    assert tools._valid_classification(no_result_edge) is False
    assert tools._allowed(no_result_edge, ctx) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint((no_result_edge,))


@pytest.mark.parametrize("bad_sensitivity", [None, "customer_file", "public"])
def test_undeclared_sensitivity_is_rejected(bad_sensitivity):
    original = tools._SPEC_BY_NAME["search_parts"]
    poisoned = replace(original, sensitivity=bad_sensitivity)
    ctx = _sys_ctx(role="phase1_full_access")

    assert tools._allowed(poisoned, ctx) is False
    with pytest.raises(ValueError):
        tools.capability_policy_fingerprint((poisoned,))


def test_approved_external_provider_needs_file_authorization_for_customer_files(monkeypatch):
    settings = _egress_settings(
        llm_trust_zone="approved_external",
        llm_base_url="https://api.deepseek.com/v1",
        llm_private_base_urls="",
        llm_approved_external_base_urls="https://api.deepseek.com",
        agent_external_file_egress_enabled=False,
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    ctx = _sys_ctx(role="phase1_full_access")
    monkeypatch.setattr(
        agent_files,
        "read_document",
        lambda *_args, **_kwargs: pytest.fail("denied capability reached handler"),
    )

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    result = tools.dispatch(None, "read_document", {"file_id": "a" * 12}, ctx)

    assert "search_parts" in visible
    assert "read_document" not in visible
    assert result["kind"] == "sensitivity_egress_denied"


def test_private_providers_can_use_customer_file_tools_without_external_file_flag(monkeypatch):
    settings = _egress_settings(agent_external_file_egress_enabled=False)
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    monkeypatch.setattr(agent_files, "read_document", lambda *_args: {"ok": True})
    monkeypatch.setattr(agent_files, "owner_of", lambda _file_id: "alice")
    ctx = _sys_ctx(role="phase1_full_access")

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}

    assert "read_document" in visible
    assert tools.dispatch(None, "read_document", {"file_id": "a" * 12}, ctx) == {"ok": True}
    assert "read_document_with_vision" in visible


@pytest.mark.parametrize("user_id", ["alice", "bob"])
def test_global_external_switch_never_counts_as_per_user_file_consent(
    monkeypatch,
    user_id,
):
    settings = _egress_settings(
        llm_trust_zone="approved_external",
        llm_base_url="https://api.deepseek.com/v1",
        llm_private_base_urls="",
        llm_approved_external_base_urls="https://api.deepseek.com",
        vision_trust_zone="approved_external",
        vision_base_url="https://vision.external.test/v1",
        vision_private_base_urls="",
        vision_approved_external_base_urls="https://vision.external.test",
        agent_external_file_egress_enabled=True,
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)
    monkeypatch.setattr(
        agent_files,
        "read_document",
        lambda *_args: pytest.fail("global switch reached customer file handler"),
    )
    monkeypatch.setattr(agent_files, "owner_of", lambda _file_id: user_id)
    ctx = _sys_ctx(user_id=user_id, role="phase1_full_access")

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    result = tools.dispatch(None, "read_document", {"file_id": "b" * 12}, ctx)

    assert tools.primary_model_call_allowed(settings) is False
    assert tools.vision_provider_call_allowed(settings) is False
    assert "read_document" not in visible
    assert "read_document_with_vision" not in visible
    assert result["kind"] == "sensitivity_egress_denied"


def test_local_read_flags_image_and_external_tool_denial_never_calls_vision(
    db,
    monkeypatch,
):
    """Local inspection is safe; the explicit Vision capability is hidden and denied."""
    upload = agent_files.save_upload(b"untrusted image bytes", "customer.png", "alice")
    ctx = _sys_ctx()
    called = False

    def forbidden_vision(images, hint):
        nonlocal called
        called = True
        return "should never run"

    monkeypatch.setattr(provider, "vision_extract", forbidden_vision)
    settings = _egress_settings(
        vision_trust_zone="approved_external",
        vision_base_url="https://vision.external.test/v1",
        vision_private_base_urls="",
        vision_approved_external_base_urls="https://vision.external.test",
        agent_external_file_egress_enabled=False,
    )
    monkeypatch.setattr(tools, "get_settings", lambda: settings)

    local = tools.dispatch(None, "read_document", {"file_id": upload["file_id"]}, ctx)
    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    denied = tools.dispatch(
        None,
        "read_document_with_vision",
        {"file_id": upload["file_id"]},
        ctx,
    )

    assert local["requires_vision"] is True
    assert local["vision_used"] is False
    assert "read_document" in visible
    assert "read_document_with_vision" not in visible
    assert denied["kind"] == "external_egress_denied"
    assert called is False


def test_external_egress_denial_does_not_block_local_document_parsing(db, monkeypatch):
    upload = agent_files.save_upload(b"local-only text", "notes.txt", "alice")
    ctx = _sys_ctx()
    monkeypatch.setattr(
        provider,
        "vision_extract",
        lambda *_args, **_kwargs: pytest.fail("local text must not call Vision"),
    )

    result = tools.dispatch(None, "read_document", {"file_id": upload["file_id"]}, ctx)

    assert result["content"] == "local-only text"
    assert result["vision_used"] is False


def test_text_pdf_stays_local_when_external_egress_is_disabled(db, monkeypatch):
    upload = agent_files.save_upload(b"fake pdf", "text.pdf", "alice")
    ctx = _sys_ctx()
    monkeypatch.setattr(agent_files, "_read_pdf", lambda path: ("local PDF text", False))
    monkeypatch.setattr(
        provider,
        "vision_extract",
        lambda *_args, **_kwargs: pytest.fail("text PDF must not call Vision"),
    )

    result = tools.dispatch(None, "read_document", {"file_id": upload["file_id"]}, ctx)

    assert result["content"] == "local PDF text"
    assert result["requires_vision"] is False


def test_authorized_external_vision_capability_is_visible_and_calls_provider(db, monkeypatch):
    upload = agent_files.save_upload(b"image bytes", "customer.png", "alice")
    ctx = _sys_ctx()
    calls: list[tuple] = []

    def recognized(images, hint, **kwargs):
        assert kwargs["_attempt_authorizer"]() is True
        calls.append((images, hint))
        return "recognized text"

    monkeypatch.setattr(provider, "vision_extract", recognized)

    visible = {schema["function"]["name"] for schema in tools.tools_for(ctx)}
    result = tools.dispatch(
        None,
        "read_document_with_vision",
        {"file_id": upload["file_id"]},
        ctx,
        _policy_lease=tools.capture_runtime_policy_lease(),
    )

    assert "read_document_with_vision" in visible
    assert result["content"] == "recognized text"
    assert result["vision_used"] is True
    assert len(calls) == 1


def test_vision_attempt_authorizer_rechecks_concrete_file_owner(db, monkeypatch):
    upload = agent_files.save_upload(b"image bytes", "customer.png", "alice")
    owner_checks = 0

    def owner_of(_file_id):
        nonlocal owner_checks
        owner_checks += 1
        # Handler gate + pre-read gate pass; the exact provider-attempt callback observes revoke.
        return "alice" if owner_checks <= 2 else "bob"

    def revoked_at_wire(_images, _hint, **kwargs):
        assert kwargs["_attempt_authorizer"]() is False
        raise provider.VisionEgressDenied("must remain private")

    monkeypatch.setattr(agent_files, "owner_of", owner_of)
    monkeypatch.setattr(provider, "vision_extract", revoked_at_wire)

    result = tools.dispatch(
        db,
        "read_document_with_vision",
        {"file_id": upload["file_id"]},
        _sys_ctx(),
        _policy_lease=tools.capture_runtime_policy_lease(),
    )

    assert result["code"] == "AGENT_VISION_EGRESS_DENIED"
    assert owner_checks >= 3


def test_vision_entry_revocation_denies_before_local_document_read(monkeypatch):
    local_reads = 0

    def forbidden_read(_file_id):
        nonlocal local_reads
        local_reads += 1
        pytest.fail("revoked Vision capability read the local document")

    monkeypatch.setattr(agent_files, "read_document", forbidden_read)

    result = agent_files.read_document_with_vision(
        "a" * 12,
        policy_lease=object(),
        attempt_authorizer=lambda: False,
    )

    assert result["code"] == "AGENT_VISION_EGRESS_DENIED"
    assert local_reads == 0


def test_authorized_vision_without_provider_config_remains_pending(db, monkeypatch):
    upload = agent_files.save_upload(b"image bytes", "customer.png", "alice")

    def unconfigured(*_args, **_kwargs):
        raise provider.VisionNotConfigured("secret provider detail")

    monkeypatch.setattr(provider, "vision_extract", unconfigured)

    result = tools.dispatch(
        None,
        "read_document_with_vision",
        {"file_id": upload["file_id"]},
        _sys_ctx(),
        _policy_lease=tools.capture_runtime_policy_lease(),
    )

    assert result["requires_vision"] is True
    assert result["vision_used"] is False
    assert "未配置视觉模型" in result["content"]
    assert "secret provider detail" not in result["content"]


def test_external_vision_capability_preserves_file_owner_acl(db, monkeypatch):
    upload = agent_files.save_upload(b"image bytes", "customer.png", "alice")
    bob = _sys_ctx("bob")
    monkeypatch.setattr(
        provider,
        "vision_extract",
        lambda *_args, **_kwargs: pytest.fail("non-owner must not call Vision"),
    )

    result = tools.dispatch(
        None,
        "read_document_with_vision",
        {"file_id": upload["file_id"]},
        bob,
        _policy_lease=tools.capture_runtime_policy_lease(),
    )

    assert result == tools._NO_ACCESS


def test_vision_policy_drift_maps_to_stable_non_retriable_error(db, monkeypatch, caplog):
    secret = "VISION-POLICY-RACE-SECRET-8472"
    upload = agent_files.save_upload(b"image bytes", "customer.png", "alice")

    def drifted_policy(*_args, **_kwargs):
        raise provider.VisionEgressDenied(secret)

    monkeypatch.setattr(provider, "vision_extract", drifted_policy)

    result = tools.dispatch(
        db,
        "read_document_with_vision",
        {"file_id": upload["file_id"]},
        _sys_ctx(),
        _policy_lease=tools.capture_runtime_policy_lease(),
    )

    assert result == {
        "error": "当前视觉模型数据出境策略未授权",
        "kind": "vision_egress_denied",
        "code": "AGENT_VISION_EGRESS_DENIED",
        "retriable": False,
    }
    assert secret not in caplog.text


def test_allowed_dispatch_audit_contains_shape_but_no_argument_values(monkeypatch, caplog):
    secret = "CUSTOMER-SECRET-NORMAL-8472"
    access_events: list[tuple] = []

    def capture_access(ctx, action, resource, filters=None):
        access_events.append((action, resource, filters))

    monkeypatch.setattr(security, "record_access_log", capture_access)
    monkeypatch.setattr(agent_files, "write_report", lambda *_args, **_kwargs: {"ok": True})
    ctx = _sys_ctx()

    result = tools.dispatch(
        None,
        "write_report",
        {"title": secret, "headers": [secret], "rows": [[secret]]},
        ctx,
    )

    assert result == {"ok": True}
    assert access_events[0][2]["arg_keys"] == ["headers", "rows", "title"]
    assert access_events[0][2]["collection_counts"] == {"headers": 1, "rows": 1}
    assert access_events[0][2]["effects"] == ["artifact_create"]
    assert access_events[0][2]["sensitivity"] == "business_confidential"
    assert secret not in str(access_events)
    assert secret not in caplog.text


def test_denied_dispatch_audit_does_not_log_untrusted_name_or_values(monkeypatch, caplog):
    secret = "https://secret.invalid/?api_key=DENIED-8472"
    access_events: list[tuple] = []

    monkeypatch.setattr(
        security,
        "record_access_log",
        lambda ctx, action, resource, filters=None: access_events.append(
            (action, resource, filters)
        ),
    )
    ctx = security.UserContext(user_id="alice", role="sales", is_authenticated=True)

    result = tools.dispatch(None, secret, {"api_key": secret}, ctx)

    assert result["kind"] == "capability_denied"
    assert access_events[0][0] == "agent_tool_denied:unknown"
    assert access_events[0][2] == {"arg_count": 1, "arg_keys": []}
    assert secret not in str(access_events)
    assert secret not in caplog.text


def test_exception_dispatch_logs_type_but_no_args_or_exception_message(monkeypatch):
    secret = "SUPPLIER-SECRET-EXCEPTION-8472"
    access_events: list[tuple] = []
    log_events: list[tuple] = []

    def explode(db, query, **kwargs):
        raise RuntimeError(f"upstream included {secret}")

    monkeypatch.setattr(
        security,
        "record_access_log",
        lambda ctx, action, resource, filters=None: access_events.append(
            (action, resource, filters)
        ),
    )
    monkeypatch.setattr(tools.part_resolver, "resolve", explode)
    monkeypatch.setattr(
        tools._log,
        "error",
        lambda message, *args, **kwargs: log_events.append((message, args, kwargs)),
    )
    ctx = security.UserContext(user_id=None, role="phase1_full_access")

    result = tools.dispatch(None, "search_parts", {"query": secret}, ctx)

    assert result["kind"] == "internal"
    assert log_events == [
        ("agent tool failed name=%s exception_type=%s", ("search_parts", "RuntimeError"), {})
    ]
    assert secret not in str(access_events)
    assert secret not in str(log_events)
