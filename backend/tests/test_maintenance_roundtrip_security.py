"""DEV-15 roundtrip security boundaries missed by the happy-path contract tests."""
from __future__ import annotations

import asyncio
import hashlib
import io
import os
import threading
import time
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Callable
from datetime import date
from decimal import Decimal
from types import SimpleNamespace

import httpx
import pytest
import starlette.formparsers as starlette_formparsers
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from starlette.formparsers import MultiPartParser as StarletteMultiPartParser

from app import permissions
from app.api import maintenance as maintenance_api
from app.auth import hash_password
from app.db import SessionLocal
from app.main import app
from app.etl import pipeline, reader
from app.etl.reader import ReaderError
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceContractWorkbookState,
    MaintenanceManualCostOverride,
    MaintenanceRoundtripOperation,
)
from app.models.system import SysAuditLog, SysImportBatch, SysUser
from app.services import maintenance_roundtrip
from tests.test_maintenance_roundtrip import (
    _edit_data_row,
    _export_to_path,
    _seed_contract,
)
from tests.test_maintenance_export_headers import _admin_client


def _expense_count(db, contract: str) -> int:
    return db.scalar(
        select(func.count(FProjectExpense.id)).where(
            FProjectExpense.linked_sales_order_no == contract
        )
    )


def _rewrite_xlsx_package(
    path,
    *,
    transforms: dict[str, Callable[[bytes], bytes]],
    extra_members: dict[str, bytes] | None = None,
) -> None:
    output = io.BytesIO()
    with (
        zipfile.ZipFile(path) as source,
        zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target,
    ):
        for member in source.infolist():
            payload = source.read(member)
            transform = transforms.get(member.filename)
            if transform is not None:
                payload = transform(payload)
            target.writestr(member, payload)
        for member_name, payload in (extra_members or {}).items():
            target.writestr(member_name, payload)
    path.write_bytes(output.getvalue())


def _rewrite_xlsx_member(path, member_name: str, transform) -> None:
    _rewrite_xlsx_package(path, transforms={member_name: transform})


def _roundtrip_worksheet_path(path, sheet_name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        return dict(reader._workbook_sheet_parts(archive))[sheet_name]


def _roundtrip_table_path(path, sheet_name: str) -> str:
    with zipfile.ZipFile(path) as archive:
        worksheet_path = dict(reader._workbook_sheet_parts(archive))[sheet_name]
        envelope = maintenance_roundtrip._roundtrip_sheet_envelopes()[sheet_name]
        table_path = maintenance_roundtrip._roundtrip_table_part(
            archive,
            worksheet_path=worksheet_path,
            sheet_name=sheet_name,
            expected_table_name=envelope.table_name,
        )
    assert table_path is not None
    return table_path


def _add_shared_strings_part(path, shared_strings: bytes) -> None:
    def add_content_type(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        ET.SubElement(
            root,
            f"{{{namespace}}}Override",
            {
                "PartName": "/xl/sharedStrings.xml",
                "ContentType": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sharedStrings+xml"
                ),
            },
        )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_relationship(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        ET.SubElement(
            root,
            f"{{{namespace}}}Relationship",
            {
                "Id": "rIdSharedStringsSecurityTest",
                "Type": (
                    "http://schemas.openxmlformats.org/officeDocument/"
                    "2006/relationships/sharedStrings"
                ),
                "Target": "sharedStrings.xml",
            },
        )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_package(
        path,
        transforms={
            "[Content_Types].xml": add_content_type,
            "xl/_rels/workbook.xml.rels": add_relationship,
        },
        extra_members={"xl/sharedStrings.xml": shared_strings},
    )


def _multipart_upload_body(size: int) -> tuple[bytes, str]:
    boundary = "itdata-roundtrip-security-boundary"
    body = (
        f"--{boundary}\r\n"
        'Content-Disposition: form-data; name="file"; '
        'filename="maintenance_roundtrip.xlsx"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n"
        "\r\n"
    ).encode() + (b"x" * size) + f"\r\n--{boundary}--\r\n".encode()
    return body, boundary


async def _asgi_roundtrip_import_with_receive_meter_async(
    body: bytes,
    boundary: str,
    *,
    token: str | None = None,
    include_content_length: bool = True,
    content_type: str | None = None,
) -> tuple[int, int, dict[bytes, bytes]]:
    consumed = 0
    sent: list[dict] = []
    delivered = False

    async def receive():
        nonlocal consumed, delivered
        if delivered:
            return {"type": "http.disconnect"}
        delivered = True
        consumed += len(body)
        return {"type": "http.request", "body": body, "more_body": False}

    async def send(message):
        sent.append(message)

    headers = [
        (
            b"content-type",
            (
                content_type
                if content_type is not None
                else f"multipart/form-data; boundary={boundary}"
            ).encode(),
        ),
    ]
    if include_content_length:
        headers.append((b"content-length", str(len(body)).encode()))
    if token is not None:
        headers.append((b"authorization", f"Bearer {token}".encode()))
    scope = {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": "/api/maintenance/roundtrip-import",
        "raw_path": b"/api/maintenance/roundtrip-import",
        "query_string": b"",
        "headers": headers,
        "client": ("127.0.0.1", 54321),
        "server": ("testserver", 80),
        "root_path": "",
    }
    await app(scope, receive, send)
    start = next(message for message in sent if message["type"] == "http.response.start")
    return start["status"], consumed, dict(start["headers"])


def _asgi_roundtrip_import_with_receive_meter(
    body: bytes,
    boundary: str,
    *,
    token: str | None = None,
    include_content_length: bool = True,
    content_type: str | None = None,
) -> tuple[int, int, dict[bytes, bytes]]:
    return asyncio.run(_asgi_roundtrip_import_with_receive_meter_async(
        body,
        boundary,
        token=token,
        include_content_length=include_content_length,
        content_type=content_type,
    ))


def test_signed_contract_scope_rejects_mixed_cross_contract_workbook_atomically(
    db,
    tmp_path,
):
    _seed_contract(db, suffix="SCOPE-P1", contract="XSDD-RT-SCOPE-P1")
    path = _export_to_path(
        db,
        tmp_path / "scope-p1.xlsx",
        contract="XSDD-RT-SCOPE-P1",
    )
    for row, contract in (
        (2, "XSDD-RT-SCOPE-P1"),
        (3, "XSDD-RT-OTHER"),
    ):
        _edit_data_row(
            path,
            "04_报销明细",
            {
                "操作": "CREATE",
                "合同号": contract,
                "报销日期": date(2026, 7, 20),
                "未税金额": Decimal("10.00"),
                "变更原因": "合同授权边界",
            },
            row=row,
        )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="合同范围",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 409
    assert _expense_count(db, "XSDD-RT-SCOPE-P1") == 0
    assert _expense_count(db, "XSDD-RT-OTHER") == 0


def test_signed_date_scope_is_closed_and_out_of_range_row_makes_zero_writes(
    db,
    tmp_path,
):
    _seed_contract(db, suffix="DATE-P1", contract="XSDD-RT-DATE-P1")
    path = _export_to_path(
        db,
        tmp_path / "date-p1.xlsx",
        contract="XSDD-RT-DATE-P1",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
    )
    for row, expense_date in (
        (2, date(2026, 7, 1)),
        (3, date(2026, 7, 31)),
        (4, date(2026, 8, 1)),
    ):
        _edit_data_row(
            path,
            "04_报销明细",
            {
                "操作": "CREATE",
                "合同号": "XSDD-RT-DATE-P1",
                "报销日期": expense_date,
                "未税金额": Decimal("10.00"),
                "变更原因": "日期授权边界",
            },
            row=row,
        )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="日期范围",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert _expense_count(db, "XSDD-RT-DATE-P1") == 0


def test_order_update_cannot_move_record_outside_signed_date_scope(db, tmp_path):
    order_id, _ = _seed_contract(
        db,
        suffix="ORDER-DATE-P1",
        contract="XSDD-RT-ORDER-DATE-P1",
    )
    path = _export_to_path(
        db,
        tmp_path / "order-date-p1.xlsx",
        contract="XSDD-RT-ORDER-DATE-P1",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
    )
    _edit_data_row(
        path,
        "02_维保订单",
        {
            "操作": "UPDATE",
            "制单日期": date(2026, 8, 1),
            "变更原因": "越界移动",
        },
    )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="日期范围",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    db.expire_all()
    assert db.get(FMaintenanceOrder, order_id).order_date == date(2026, 7, 15)


def test_logical_create_replay_survives_openpyxl_resave_and_payload_conflicts_409(
    db,
    tmp_path,
):
    _seed_contract(db, suffix="LEDGER-CREATE", contract="XSDD-RT-LEDGER-CREATE")
    path = _export_to_path(
        db,
        tmp_path / "ledger-create.xlsx",
        contract="XSDD-RT-LEDGER-CREATE",
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": "XSDD-RT-LEDGER-CREATE",
            "报销日期": date(2026, 7, 20),
            "未税金额": Decimal("100.00"),
            "变更原因": "逻辑幂等",
        },
    )
    first = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    first_hash = first["file_hash"]

    # 模拟 Excel/openpyxl 重新保存，并改一个服务端不采信的展示单元格，确保文件 hash 改变。
    _edit_data_row(
        path,
        "04_报销明细",
        {"含税金额(系统计算)": Decimal("999.00")},
    )
    assert pipeline.sha256_file(str(path)) != first_hash
    replay = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    assert replay["no_op"] is True
    assert replay["logical_replay"] is True
    assert replay["replayed_rows"] == 1
    assert _expense_count(db, "XSDD-RT-LEDGER-CREATE") == 1
    assert db.scalar(select(func.count(MaintenanceRoundtripOperation.id))) == 1

    _edit_data_row(path, "04_报销明细", {"未税金额": Decimal("101.00")})
    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="相同行键.*不同内容",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )
    assert caught.value.status_code == 409
    assert _expense_count(db, "XSDD-RT-LEDGER-CREATE") == 1


def test_logical_update_replay_skips_stale_entity_version(db, tmp_path):
    order_id, _ = _seed_contract(
        db,
        suffix="LEDGER-UPDATE",
        contract="XSDD-RT-LEDGER-UPDATE",
    )
    path = _export_to_path(
        db,
        tmp_path / "ledger-update.xlsx",
        contract="XSDD-RT-LEDGER-UPDATE",
    )
    _edit_data_row(
        path,
        "02_维保订单",
        {
            "操作": "UPDATE",
            "项目名称": "账本更新后名称",
            "变更原因": "逻辑幂等更新",
        },
    )
    maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    # 维保单号是只读展示字段，不参与 UPDATE payload；改动它只为确保 ZIP hash 改变。
    _edit_data_row(path, "02_维保订单", {"维保单号": "展示字段被重新保存"})
    replay = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    assert replay["logical_replay"] is True
    db.expire_all()
    assert db.get(FMaintenanceOrder, order_id).project_raw == "账本更新后名称"


def test_mixed_logical_replay_rejects_tampered_old_row_token_before_new_write(
    db,
    tmp_path,
):
    order_id, _ = _seed_contract(
        db,
        suffix="LEDGER-MIXED-TOKEN",
        contract="XSDD-RT-LEDGER-MIXED-TOKEN",
    )
    path = _export_to_path(
        db,
        tmp_path / "ledger-mixed-token.xlsx",
        contract="XSDD-RT-LEDGER-MIXED-TOKEN",
    )
    _edit_data_row(
        path,
        "02_维保订单",
        {
            "操作": "UPDATE",
            "项目名称": "已成功应用的旧操作",
            "变更原因": "首次应用",
        },
    )
    first = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    assert first["counts"]["update"] == 1

    # 同一本工作簿同时包含一个账本已记录的旧操作和一个待应用的新操作。
    # 旧操作 payload 未变，只有不参与 payload hash 的 HMAC 被篡改。
    _edit_data_row(
        path,
        "02_维保订单",
        {"__row_token": "0" * 64},
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": "XSDD-RT-LEDGER-MIXED-TOKEN",
            "报销日期": date(2026, 7, 20),
            "未税金额": Decimal("10.00"),
            "变更原因": "不得越过坏 token 写入",
        },
    )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="行令牌无效",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 409
    db.expire_all()
    assert db.get(FMaintenanceOrder, order_id).project_raw == "已成功应用的旧操作"
    assert _expense_count(db, "XSDD-RT-LEDGER-MIXED-TOKEN") == 0
    assert db.scalar(select(func.count(MaintenanceRoundtripOperation.id))) == 1
    assert db.scalar(
        select(func.count(SysImportBatch.id)).where(
            SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE,
            SysImportBatch.status == "success",
        )
    ) == 1


def test_tampered_keep_row_token_rejects_before_snapshot_or_recompute(
    db,
    tmp_path,
    monkeypatch,
):
    order_id, _ = _seed_contract(
        db,
        suffix="KEEP-TOKEN",
        contract="XSDD-RT-KEEP-TOKEN",
    )
    path = _export_to_path(
        db,
        tmp_path / "keep-token.xlsx",
        contract="XSDD-RT-KEEP-TOKEN",
    )
    _edit_data_row(
        path,
        "02_维保订单",
        {"__row_token": "f" * 64},
    )
    recomputed = False

    def fail_if_recomputed(_db):
        nonlocal recomputed
        recomputed = True
        raise AssertionError("tampered KEEP row reached recompute")

    monkeypatch.setattr(
        maintenance_roundtrip,
        "_recompute_in_transaction",
        fail_if_recomputed,
    )
    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="行令牌无效",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 409
    assert recomputed is False
    assert db.get(MaintenanceContractWorkbookState, "XSDD-RT-KEEP-TOKEN") is None
    assert db.scalar(
        select(func.count(SysImportBatch.id)).where(
            SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE,
        )
    ) == 0
    assert db.scalar(select(func.count(SysAuditLog.id))) == 0
    db.expire_all()
    assert db.get(FMaintenanceOrder, order_id).project_raw == "回填项目-KEEP-TOKEN"


def test_contract_revisions_use_bounded_hidden_table_instead_of_one_json_cell(
    db,
    tmp_path,
):
    _seed_contract(db, suffix="REVISIONS", contract="XSDD-RT-REVISIONS")
    path = _export_to_path(
        db,
        tmp_path / "revisions.xlsx",
        contract="XSDD-RT-REVISIONS",
    )
    workbook = load_workbook(path, data_only=False)
    try:
        assert workbook["99_合同版本"].sheet_state == "veryHidden"
        assert set(workbook["99_合同版本"].tables) == {
            "tbl_contract_revisions_v1"
        }
        metadata = {
            workbook["99_元数据"].cell(row=row, column=1).value:
            workbook["99_元数据"].cell(row=row, column=2).value
            for row in range(2, workbook["99_元数据"].max_row + 1)
        }
        assert "contract_revisions" not in metadata
        assert metadata["contract_revision_count"] == "1"
        assert len(metadata["contract_revisions_sha256"]) == 64
    finally:
        workbook.close()


def test_generic_import_rejects_roundtrip_protocol_before_hash_or_business_write(
    db,
    tmp_path,
):
    _seed_contract(db, suffix="WRONG-ENDPOINT", contract="XSDD-RT-WRONG-ENDPOINT")
    path = _export_to_path(
        db,
        tmp_path / "wrong-endpoint.xlsx",
        contract="XSDD-RT-WRONG-ENDPOINT",
    )
    file_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    generic = SysImportBatch(
        filename="legacy-wrong-success.xlsx",
        file_type="maintenance",
        file_hash=file_hash,
        status="success",
    )
    db.add(generic)
    db.commit()

    with pytest.raises(ReaderError, match="维保项目回填入口") as caught:
        pipeline.run_import(
            db,
            str(path),
            path.name,
            uploaded_by="tester",
        )
    assert caught.value.code == "roundtrip_wrong_endpoint"
    db.rollback()

    result = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    assert result["no_op"] is False
    assert result["batch_id"] != generic.id


def test_exact_hash_replay_short_circuits_before_openpyxl(db, tmp_path, monkeypatch):
    _seed_contract(db, suffix="HASH-PREFLIGHT", contract="XSDD-RT-HASH-PREFLIGHT")
    path = _export_to_path(
        db,
        tmp_path / "hash-preflight.xlsx",
        contract="XSDD-RT-HASH-PREFLIGHT",
    )
    first = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    def fail_if_openpyxl_is_reached(_path):
        raise AssertionError("exact hash replay reached openpyxl")

    monkeypatch.setattr(
        maintenance_roundtrip,
        "_load_and_parse",
        fail_if_openpyxl_is_reached,
    )
    replay = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    assert replay["no_op"] is True
    assert replay["logical_replay"] is False
    assert replay["batch_id"] == first["batch_id"]


def test_unknown_invalid_file_rolls_back_preflight_transaction(db, tmp_path):
    path = tmp_path / "invalid-roundtrip.xlsx"
    path.write_bytes(b"not-an-xlsx")

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="有效的 .xlsx|无法打开维保回填|文件无法按 .xlsx 解析",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert db.in_transaction() is False


def test_parse_failure_close_error_preserves_primary_identity_and_warns(
    db,
    tmp_path,
    monkeypatch,
):
    _seed_contract(db, suffix="PARSE-CLOSE", contract="XSDD-RT-PARSE-CLOSE")
    path = _export_to_path(
        db,
        tmp_path / "parse-close.xlsx",
        contract="XSDD-RT-PARSE-CLOSE",
    )
    primary = maintenance_roundtrip.RoundtripWorkbookError(
        "工作簿主校验失败",
        status_code=409,
    )
    events: list[str] = []
    warnings: list[tuple[str, dict]] = []

    class CloseFailingWorkbook:
        @property
        def sheetnames(self):
            events.append("validate")
            raise primary

        def close(self):
            events.append("close")
            raise OSError("工作簿关闭失败")

    monkeypatch.setattr(
        maintenance_roundtrip,
        "load_workbook",
        lambda *_args, **_kwargs: CloseFailingWorkbook(),
    )
    monkeypatch.setattr(
        maintenance_roundtrip._log,
        "warning",
        lambda message, *_args, **kwargs: warnings.append((message, kwargs)),
    )

    with pytest.raises(maintenance_roundtrip.RoundtripWorkbookError) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value is primary
    assert caught.value.status_code == 409
    assert events == ["validate", "close"]
    assert warnings == [
        ("维保工作簿解析失败后的关闭清理失败；保留原始校验错误", {"exc_info": True})
    ]
    assert db.in_transaction() is False


@pytest.mark.parametrize(
    ("counts", "expected_sheet"),
    [
        ([maintenance_roundtrip.MAX_ROWS_PER_TABLE + 1], "02_维保订单"),
        ([1, maintenance_roundtrip.MAX_ROWS_PER_TABLE + 1], "03_订单明细"),
        (
            [
                1,
                0,
                maintenance_roundtrip.MAX_ROWS_PER_TABLE
                - maintenance_roundtrip.BLANK_CREATE_ROWS
                + 1,
            ],
            "04_报销明细",
        ),
    ],
)
def test_export_row_caps_fail_before_full_orm_materialization(
    counts,
    expected_sheet,
):
    class CountOnlySession:
        def __init__(self):
            self._counts = iter(counts)
            self.execute_called = False

        def scalar(self, _statement):
            return next(self._counts)

        def scalars(self, _statement):
            # DISTINCT contract discovery is a bounded preflight query, not ORM rows.
            return SimpleNamespace(all=lambda: [])

        def execute(self, _statement):
            self.execute_called = True
            raise AssertionError("full ORM materialization happened before cap rejection")

    session = CountOnlySession()
    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match=expected_sheet,
    ) as caught:
        maintenance_roundtrip._selected_data(
            session,
            contract=None,
            date_from=None,
            date_to=None,
            blank=False,
        )

    assert caught.value.status_code == 413
    assert session.execute_called is False


def test_dynamic_text_cap_fails_before_orm_materialization_or_rendering():
    class CountAndBudgetOnlySession:
        def __init__(self):
            self._values = iter(
                [
                    1,
                    0,
                    0,
                    maintenance_roundtrip.MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK + 1,
                ]
            )
            self.execute_called = False

        def scalar(self, _statement):
            return next(self._values)

        def scalars(self, _statement):
            # DISTINCT contract discovery is bounded and does not materialize ORM rows.
            return SimpleNamespace(all=lambda: [])

        def execute(self, _statement):
            self.execute_called = True
            raise AssertionError("ORM materialization happened before byte-cap rejection")

    session = CountAndBudgetOnlySession()
    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="动态文本超过 64 MiB",
    ) as caught:
        maintenance_roundtrip._selected_data(
            session,
            contract=None,
            date_from=None,
            date_to=None,
            blank=False,
        )

    assert caught.value.status_code == 413
    assert session.execute_called is False


def test_dynamic_text_budget_counts_all_unbounded_output_and_utf8_boundary(
    db,
    monkeypatch,
):
    contract = "XSDD-RT-UTF8-BUDGET"
    _order_id, line_id = _seed_contract(
        db,
        suffix="UTF8-BUDGET",
        contract=contract,
    )
    line = db.get(FMaintenanceLine, line_id)
    line.cost_source = "none"
    line.serial_numbers = None
    expense = FProjectExpense(
        raw_line_id="EXP-RT-UTF8-BUDGET",
        bxd_no="BXD-RT-UTF8-BUDGET",
        line_no=1,
        data_status="已结束",
        expense_date=date(2026, 7, 20),
        linked_sales_order_no=contract,
        amount=Decimal("10"),
        amount_ex_tax=Decimal("10"),
        amount_inc_tax=Decimal("11.30"),
        reason=None,
        import_batch_id=line.import_batch_id,
    )
    override = MaintenanceManualCostOverride(
        line_id=line.id,
        unit_cost_ex_tax=Decimal("1"),
        unit_cost_inc_tax=Decimal("1.13"),
        tax_rate_used=Decimal("0.13"),
        reason=None,
        evidence={"note": ""},
        active=True,
        updated_by="tester",
    )
    db.add_all([expense, override])
    db.commit()

    order_filters = maintenance_roundtrip._selection_filters(
        contract=contract,
        date_from=None,
        date_to=None,
    )
    expense_filters = [FProjectExpense.linked_sales_order_no == contract]

    def measured_bytes() -> int:
        return maintenance_roundtrip._selected_dynamic_text_bytes(
            db,
            order_filters=order_filters,
            expense_filters=expense_filters,
        )

    measured = measured_bytes()
    for entity, field in (
        (line, "description"),
        (line, "serial_numbers"),
        (expense, "reason"),
        (override, "reason"),
    ):
        setattr(entity, field, f"{getattr(entity, field) or ''}中")
        db.flush()
        current = measured_bytes()
        assert current == measured + len("中".encode())
        measured = current

    override.evidence = {"note": "中"}
    db.flush()
    current = measured_bytes()
    assert current == measured + len("中".encode())
    measured = current
    db.commit()

    # 等于上限仍可构建；少 1 byte 时必须在任何 openpyxl 渲染前拒绝。
    monkeypatch.setattr(
        maintenance_roundtrip,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        measured,
    )
    output = maintenance_roundtrip.build_roundtrip_template(
        db,
        contract=contract,
        exported_by="tester",
    )
    output.close()
    db.rollback()

    rendered = []
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_instructions_sheet",
        lambda *_args, **_kwargs: rendered.append("rendered"),
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        measured - 1,
    )
    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="动态文本超过 64 MiB",
    ) as caught:
        maintenance_roundtrip.build_roundtrip_template(
            db,
            contract=contract,
            exported_by="tester",
        )

    assert caught.value.status_code == 413
    assert rendered == []


def test_roundtrip_uncompressed_package_cap_is_64_mib(tmp_path, monkeypatch):
    path = tmp_path / "oversized-expanded.xlsx"
    path.write_bytes(b"zip-placeholder")

    class FakeArchive:
        def __init__(self, _path):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def infolist(self):
            return [
                SimpleNamespace(
                    file_size=maintenance_roundtrip.MAX_ROUNDTRIP_UNCOMPRESSED_BYTES
                    + 1
                )
            ]

        def namelist(self):
            return []

    monkeypatch.setattr(
        maintenance_roundtrip.reader,
        "_check_xlsx_archive_safety",
        lambda _path: None,
    )
    monkeypatch.setattr(
        maintenance_roundtrip.reader,
        "_check_workbook_size",
        lambda _path: None,
    )
    monkeypatch.setattr(maintenance_roundtrip.zipfile, "ZipFile", FakeArchive)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="解压后超过 64 MiB",
    ) as caught:
        maintenance_roundtrip._assert_safe_workbook_package(str(path))

    assert caught.value.status_code == 413


def test_roundtrip_rejects_static_cell_outside_table_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-OUTSIDE-TABLE"
    _seed_contract(db, suffix="OUTSIDE-TABLE", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "outside-table.xlsx",
        contract=contract,
    )
    workbook = load_workbook(path, data_only=False)
    try:
        # The signed table currently ends at row 2. This is still below the generic
        # 10k-row envelope, but must not be materialized by openpyxl.
        workbook["02_维保订单"]["S5000"] = "Excel Table 范围外静态值"
        workbook.save(path)
    finally:
        workbook.close()

    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("协议范围外单元格不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="Table 范围外",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_rejects_forged_dimension_outside_table_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-FORGED-DIMENSION"
    _seed_contract(db, suffix="FORGED-DIMENSION", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "forged-dimension.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")

    def forge_dimension(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        dimension = next(
            element
            for element in root
            if element.tag.rsplit("}", 1)[-1] == "dimension"
        )
        # This stays inside the generic 5M-cell cap, but is far beyond this
        # workbook's signed A1:S2 table.
        dimension.set("ref", "A1:S10001")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, forge_dimension)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("伪造 worksheet dimension 不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="Table 范围外",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_rejects_huge_merged_range_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-HUGE-MERGE"
    _seed_contract(db, suffix="HUGE-MERGE", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "huge-merge.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "00_使用说明")

    def forge_merged_range(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        merged_cell = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "mergeCell"
        )
        merged_cell.set("ref", "A1:XFD1048576")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, forge_merged_range)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("超大合并范围不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="合并单元格.*范围",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_rejects_huge_hyperlink_range_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-HUGE-HYPERLINK"
    _seed_contract(db, suffix="HUGE-HYPERLINK", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "huge-hyperlink.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "00_使用说明")

    def add_huge_hyperlink(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        hyperlinks = ET.Element(f"{{{namespace}}}hyperlinks")
        ET.SubElement(
            hyperlinks,
            f"{{{namespace}}}hyperlink",
            {
                "ref": "A1:XFD1048576",
                "location": "'00_使用说明'!A1",
            },
        )
        root.append(hyperlinks)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, add_huge_hyperlink)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("超大超链接范围不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="超链接.*范围",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_hyperlink_must_use_hyperlink_relationship_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-HYPERLINK-REL"
    _seed_contract(db, suffix="HYPERLINK-REL", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "hyperlink-rel.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")

    def point_hyperlink_at_table_relationship(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        table_part = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "tablePart"
        )
        relation_id = next(
            value
            for name, value in table_part.attrib.items()
            if name.rsplit("}", 1)[-1] == "id"
        )
        hyperlinks = ET.Element(f"{{{namespace}}}hyperlinks")
        ET.SubElement(
            hyperlinks,
            f"{{{namespace}}}hyperlink",
            {
                "ref": "A1",
                (
                    "{http://schemas.openxmlformats.org/"
                    "officeDocument/2006/relationships}id"
                ): relation_id,
            },
        )
        root.append(hyperlinks)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(
        path,
        worksheet_path,
        point_hyperlink_at_table_relationship,
    )
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("错配的超链接关系不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="超链接关系类型",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert opened is False


def test_roundtrip_accepts_generated_instruction_merge_ranges(db, tmp_path):
    contract = "XSDD-RT-LEGAL-MERGES"
    _seed_contract(db, suffix="LEGAL-MERGES", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "legal-merges.xlsx",
        contract=contract,
    )

    workbook, _metadata, _revisions, _rows = (
        maintenance_roundtrip._load_and_parse(str(path))
    )
    try:
        merge_ranges = {
            str(cell_range)
            for cell_range in workbook["00_使用说明"].merged_cells.ranges
        }
        assert {"A1:B1", "A3:B3", "A10:B10"} <= merge_ranges
    finally:
        workbook.close()


def test_roundtrip_accepts_bounded_hyperlink_relationship(db, tmp_path):
    contract = "XSDD-RT-LEGAL-HYPERLINK"
    _seed_contract(db, suffix="LEGAL-HYPERLINK", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "legal-hyperlink.xlsx",
        contract=contract,
    )
    workbook = load_workbook(path, data_only=False)
    try:
        workbook["00_使用说明"]["A4"].hyperlink = "https://example.invalid/help"
        workbook.save(path)
    finally:
        workbook.close()

    maintenance_roundtrip._assert_safe_workbook_package(str(path))


def test_roundtrip_rejects_cumulative_merged_range_expansion_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-MERGE-EXPANSION"
    _seed_contract(db, suffix="MERGE-EXPANSION", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "merge-expansion.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "00_使用说明")

    def add_overlapping_merges(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        merge_cells = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "mergeCells"
        )
        namespace = merge_cells.tag.rsplit("}", 1)[0].lstrip("{")
        # The instruction-sheet envelope contains 38 cells. Its three legitimate
        # 2-cell ranges consume six; seventeen repeats push declared expansion to 40.
        for _index in range(17):
            ET.SubElement(
                merge_cells,
                f"{{{namespace}}}mergeCell",
                {"ref": "A1:B1"},
            )
        merge_cells.set("count", str(len(merge_cells)))
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, add_overlapping_merges)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("累计超限合并范围不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="合并单元格展开范围超过",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_rejects_cumulative_hyperlink_expansion_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-HYPERLINK-EXPANSION"
    _seed_contract(db, suffix="HYPERLINK-EXPANSION", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "hyperlink-expansion.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "00_使用说明")

    def add_overlapping_hyperlinks(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        hyperlinks = ET.Element(f"{{{namespace}}}hyperlinks")
        # Each single-cell ref is small, but openpyxl expands every declaration.
        for _index in range(39):
            ET.SubElement(
                hyperlinks,
                f"{{{namespace}}}hyperlink",
                {
                    "ref": "A1",
                    "location": "'00_使用说明'!A1",
                },
            )
        root.append(hyperlinks)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, add_overlapping_hyperlinks)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("累计超限超链接范围不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="超链接展开范围超过",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_rejects_sparse_huge_coordinate_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-SPARSE-HUGE"
    _seed_contract(db, suffix="SPARSE-HUGE", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "sparse-huge.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")

    def add_sparse_cell(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")

        def tag(local_name: str) -> str:
            return f"{{{namespace}}}{local_name}"

        dimension = next(
            element
            for element in root
            if element.tag.rsplit("}", 1)[-1] == "dimension"
        )
        # A forged small dimension must not hide the actual sparse coordinate.
        dimension.set("ref", "A1")
        first_row = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "row"
        )
        # openpyxl treats any direct row child as cell-like. A renamed <c> must
        # not bypass coordinate accounting.
        cell = ET.SubElement(
            first_row,
            tag("x"),
            {"r": "XFD1048576", "t": "inlineStr"},
        )
        inline_string = ET.SubElement(cell, tag("is"))
        ET.SubElement(inline_string, tag("t")).text = "sparse"
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, add_sparse_cell)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("稀疏超大坐标不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="Table 范围外",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_rejects_formula_nested_in_renamed_cell_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-RENAMED-FORMULA"
    _seed_contract(db, suffix="RENAMED-FORMULA", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "renamed-formula.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")

    def rename_formula_cell(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        cell = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "c"
            and element.attrib.get("r") == "A2"
        )
        namespace = cell.tag.rsplit("}", 1)[0].lstrip("{")
        cell.tag = f"{{{namespace}}}x"
        for child in list(cell):
            cell.remove(child)
        ET.SubElement(cell, f"{{{namespace}}}f").text = "1+1"
        ET.SubElement(cell, f"{{{namespace}}}v").text = "2"
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, rename_formula_cell)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("重命名单元格中的公式不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="包含公式",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert opened is False


def test_roundtrip_rejects_wide_worksheet_xml_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-WIDE-XML"
    _seed_contract(db, suffix="WIDE-XML", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "wide-xml.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")

    def widen_worksheet(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        for index in range(
            maintenance_roundtrip.MAX_ROUNDTRIP_WORKSHEET_STRUCTURAL_ELEMENTS
            + 1
        ):
            ET.SubElement(
                root,
                f"{{{namespace}}}junk",
                {"index": str(index)},
            )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_member(path, worksheet_path, widen_worksheet)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("超宽 worksheet XML 不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="结构性 XML 元素超过",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


@pytest.mark.parametrize(
    "part_kind",
    ["worksheet", "table", "relationship", "sharedStrings"],
)
def test_roundtrip_malformed_protocol_xml_is_controlled_4xx(
    db,
    tmp_path,
    part_kind,
):
    contract = f"XSDD-RT-MALFORMED-{part_kind}"
    _seed_contract(db, suffix=f"MALFORMED-{part_kind}", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / f"malformed-{part_kind}.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")
    if part_kind == "worksheet":
        target = worksheet_path
    elif part_kind == "table":
        target = _roundtrip_table_path(path, "02_维保订单")
    elif part_kind == "relationship":
        target = maintenance_roundtrip._roundtrip_relationship_part(
            worksheet_path
        )
    else:
        _add_shared_strings_part(path, b"<sst><si>")
        target = "xl/sharedStrings.xml"

    if part_kind != "sharedStrings":
        _rewrite_xlsx_member(path, target, lambda payload: payload[:-1])

    with pytest.raises(maintenance_roundtrip.RoundtripWorkbookError) as caught:
        maintenance_roundtrip._assert_safe_workbook_package(str(path))

    assert caught.value.status_code == 422
    assert caught.value.__cause__ is not None


def test_roundtrip_relationship_namespace_bomb_is_controlled_4xx(db, tmp_path):
    contract = "XSDD-RT-REL-NAMESPACE"
    _seed_contract(db, suffix="REL-NAMESPACE", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "relationship-namespace.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")
    relationship_path = maintenance_roundtrip._roundtrip_relationship_part(
        worksheet_path
    )

    def add_namespaces(payload: bytes) -> bytes:
        root_start = payload.find(b"<Relationships")
        root_end = payload.find(b">", root_start)
        assert root_start >= 0 and root_end > root_start
        declarations = b"".join(
            f' xmlns:bad{index}="urn:bad:{index}"'.encode()
            for index in range(reader._XML_NAME_LIMIT + 1)
        )
        return payload[:root_end] + declarations + payload[root_end:]

    _rewrite_xlsx_member(path, relationship_path, add_namespaces)

    with pytest.raises(maintenance_roundtrip.RoundtripWorkbookError) as caught:
        maintenance_roundtrip._assert_safe_workbook_package(str(path))

    assert caught.value.status_code == 422
    assert isinstance(caught.value.__cause__, ReaderError)


def test_roundtrip_table_part_must_match_its_table_relationship_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-TABLE-REL-MISMATCH"
    _seed_contract(db, suffix="TABLE-REL-MISMATCH", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "table-rel-mismatch.xlsx",
        contract=contract,
    )
    worksheet_path = _roundtrip_worksheet_path(path, "02_维保订单")
    relationship_path = maintenance_roundtrip._roundtrip_relationship_part(
        worksheet_path
    )

    def redirect_table_part(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        table_part = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "tablePart"
        )
        relation_attribute = next(
            name
            for name in table_part.attrib
            if name.rsplit("}", 1)[-1] == "id"
        )
        table_part.set(relation_attribute, "rIdNotATable")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def add_non_table_relationship(payload: bytes) -> bytes:
        root = ET.fromstring(payload)
        namespace = root.tag.rsplit("}", 1)[0].lstrip("{")
        ET.SubElement(
            root,
            f"{{{namespace}}}Relationship",
            {
                "Id": "rIdNotATable",
                "Type": "urn:not-a-table",
                "Target": "../tables/table1.xml",
            },
        )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    _rewrite_xlsx_package(
        path,
        transforms={
            worksheet_path: redirect_table_part,
            relationship_path: add_non_table_relationship,
        },
    )
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("错配的 tablePart 关系不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="Table|tablePart",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert opened is False


def test_roundtrip_counts_actual_shared_strings_before_openpyxl(
    db,
    tmp_path,
    monkeypatch,
):
    contract = "XSDD-RT-SHARED-STRINGS"
    _seed_contract(db, suffix="SHARED-STRINGS", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "shared-strings.xlsx",
        contract=contract,
    )

    shared_strings = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'count="1" uniqueCount="1">'
        b"<si><t>one</t></si><si><t>two</t></si>"
        b"</sst>"
    )
    _add_shared_strings_part(path, shared_strings)
    monkeypatch.setattr(maintenance_roundtrip, "MAX_ROUNDTRIP_SHARED_STRINGS", 1)
    opened = False

    def forbidden_load(*_args, **_kwargs):
        nonlocal opened
        opened = True
        raise AssertionError("超量 sharedStrings 不应进入 openpyxl")

    monkeypatch.setattr(maintenance_roundtrip, "load_workbook", forbidden_load)

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="sharedStrings.*超过",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 413
    assert opened is False


def test_roundtrip_xml_envelope_accepts_10000_row_table_boundary(db, tmp_path):
    contract = "XSDD-RT-10K-BOUNDARY"
    _seed_contract(db, suffix="10K-BOUNDARY", contract=contract)
    path = _export_to_path(
        db,
        tmp_path / "ten-thousand-row-boundary.xlsx",
        contract=contract,
    )
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook["02_维保订单"]
        worksheet.tables["tbl_orders_v1"].ref = "A1:S10001"
        worksheet["S10001"] = "10k-boundary"
        workbook.save(path)
    finally:
        workbook.close()

    # 10,000 data rows plus the header is the inclusive protocol boundary.
    maintenance_roundtrip._assert_safe_workbook_package(str(path))


def test_template_build_concurrency_is_one(db):
    first = maintenance_roundtrip.build_roundtrip_template(
        db,
        exported_by="tester",
        blank=True,
    )
    concurrent = SessionLocal()
    try:
        with pytest.raises(
            maintenance_roundtrip.RoundtripWorkbookError,
            match="正在生成",
        ) as caught:
            maintenance_roundtrip.build_roundtrip_template(
                concurrent,
                exported_by="tester",
                blank=True,
            )
        assert caught.value.status_code == 429
    finally:
        concurrent.rollback()
        concurrent.close()
        first.close()
        db.rollback()


def test_roundtrip_apply_action_defaults_fail_closed_for_non_operators():
    assert permissions.template_for("admin")[
        "action_maintenance_roundtrip_apply"
    ] is True
    assert permissions.template_for("boss")[
        "action_maintenance_roundtrip_apply"
    ] is True
    assert permissions.template_for("sales")[
        "action_maintenance_roundtrip_apply"
    ] is False
    assert permissions.template_for("purchaser")[
        "action_maintenance_roundtrip_apply"
    ] is False
    assert permissions.template_for("readonly")[
        "action_maintenance_roundtrip_apply"
    ] is False


def test_roundtrip_import_anonymous_401_does_not_read_multipart_body(db):
    body, boundary = _multipart_upload_body(2 * 1024 * 1024)

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
    )

    assert status == 401
    assert consumed == 0


def test_roundtrip_import_customer_blind_403_does_not_read_multipart_body(db):
    base = permissions.effective("boss", None)
    overrides = {"data_customer": False}
    user = SysUser(
        username="roundtrip-asgi-no-customer",
        role="boss",
        display_name="ASGI 无客户权限回填员",
        password_hash=hash_password("roundtrip-password"),
        template_code="boss",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    )
    db.add(user)
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={
            "username": "roundtrip-asgi-no-customer",
            "password": "roundtrip-password",
        },
    )
    assert login.status_code == 200
    body, boundary = _multipart_upload_body(2 * 1024 * 1024)

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
        token=login.json()["token"],
    )

    assert status == 403
    assert consumed == 0


@pytest.mark.parametrize(
    "content_type",
    [
        "application/x-www-form-urlencoded",
        "application/json",
        "text/plain",
    ],
)
def test_roundtrip_import_rejects_non_multipart_415_before_reading_body(
    db,
    content_type,
):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    body = b"file=" + (b"x" * (2 * 1024 * 1024))

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        "unused",
        token=token,
        content_type=content_type,
    )

    assert status == 415
    assert consumed == 0


def test_roundtrip_import_declared_oversize_413_before_reading_multipart_body(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    monkeypatch.setattr(maintenance_api.config, "MAX_UPLOAD_MB", 1)
    body, boundary = _multipart_upload_body(2 * 1024 * 1024)

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
        token=token,
    )

    assert status == 413
    assert consumed == 0


def test_roundtrip_import_chunked_oversize_stops_at_body_byte_limit(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    monkeypatch.setattr(maintenance_api.config, "MAX_UPLOAD_MB", 1)
    body, boundary = _multipart_upload_body(2 * 1024 * 1024)

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
        token=token,
        include_content_length=False,
    )

    assert status == 413
    assert consumed == len(body)


def test_roundtrip_import_busy_429_before_body_with_retry_after(db):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    body, boundary = _multipart_upload_body(2 * 1024 * 1024)
    assert maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.acquire(blocking=False)
    try:
        status, consumed, headers = _asgi_roundtrip_import_with_receive_meter(
            body,
            boundary,
            token=token,
        )
    finally:
        maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()

    assert status == 429
    assert consumed == 0
    assert headers[b"retry-after"] == b"5"


@pytest.mark.parametrize(
    ("content_type", "body", "expected_consumed"),
    [
        ("multipart/form-data", b"not-read-without-boundary", 0),
        ("multipart/form-data; boundary=", b"not-read-with-empty-boundary", 0),
        (
            "multipart/form-data; boundary=broken",
            b"--broken\r\nmalformed-header\r\n\r\npayload\r\n--broken--\r\n",
            None,
        ),
        (
            f"multipart/form-data; boundary={'b' * 257}",
            b"boundary-too-long-must-not-read",
            0,
        ),
    ],
)
def test_roundtrip_import_malformed_multipart_is_4xx_never_500(
    db,
    content_type,
    body,
    expected_consumed,
):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        "unused",
        token=token,
        content_type=content_type,
    )

    assert 400 <= status < 500
    if expected_consumed is not None:
        assert consumed == expected_consumed


def test_roundtrip_import_malformed_after_file_part_closes_partial_spool(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    created_spools = []
    original_spool_factory = starlette_formparsers.SpooledTemporaryFile

    def tracking_spool(*args, **kwargs):
        spool = original_spool_factory(*args, **kwargs)
        created_spools.append(spool)
        return spool

    monkeypatch.setattr(
        starlette_formparsers,
        "SpooledTemporaryFile",
        tracking_spool,
    )
    boundary = "partial-spool-boundary"
    body = (
        f"--{boundary}\r\n"
        'Content-Disposition: form-data; name="file"; filename="partial.xlsx"\r\n'
        "Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n"
        "\r\n"
        "partial-file-data\r\n"
        f"--{boundary}\r\n"
        "malformed-header\r\n"
        "\r\n"
        "payload\r\n"
        f"--{boundary}--\r\n"
    ).encode()

    status, _consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
        token=token,
    )

    assert status == 400
    assert created_spools
    assert all(spool.closed for spool in created_spools)


def test_roundtrip_import_file_limit_413_removes_owned_temp_and_closes_spool(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    created_paths: list[str] = []
    created_spools = []
    original_mkstemp = maintenance_api.tempfile.mkstemp
    original_spool_factory = starlette_formparsers.SpooledTemporaryFile

    def tracking_mkstemp(*args, **kwargs):
        fd, path = original_mkstemp(*args, **kwargs)
        created_paths.append(path)
        return fd, path

    def tracking_spool(*args, **kwargs):
        spool = original_spool_factory(*args, **kwargs)
        created_spools.append(spool)
        return spool

    monkeypatch.setattr(maintenance_api.config, "MAX_UPLOAD_MB", 0)
    monkeypatch.setattr(maintenance_api.tempfile, "mkstemp", tracking_mkstemp)
    monkeypatch.setattr(
        starlette_formparsers,
        "SpooledTemporaryFile",
        tracking_spool,
    )
    body, boundary = _multipart_upload_body(1024)

    status, _consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
        token=token,
    )

    assert status == 413
    assert created_paths and all(not os.path.exists(path) for path in created_paths)
    assert created_spools and all(spool.closed for spool in created_spools)


def test_roundtrip_import_single_large_asgi_frame_never_copies_unbounded_remainders(
    db,
    monkeypatch,
):
    class NoTailSliceBytes(bytes):
        def __getitem__(self, key):
            if isinstance(key, slice) and key.start and key.stop is None:
                raise AssertionError("unbounded remainder slice would make chunking O(n^2)")
            return super().__getitem__(key)

    client = _admin_client(db)
    token = client.headers["Authorization"].removeprefix("Bearer ")
    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        lambda *_args, **_kwargs: {"status": "success", "no_op": True},
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args, **_kwargs: None)
    raw_body, boundary = _multipart_upload_body(512 * 1024)
    body = NoTailSliceBytes(raw_body)

    status, consumed, _headers = _asgi_roundtrip_import_with_receive_meter(
        body,
        boundary,
        token=token,
    )

    assert status == 200
    assert consumed == len(body)


def test_roundtrip_import_save_value_error_is_server_error_not_multipart_400(
    db,
    monkeypatch,
):
    authenticated_client = _admin_client(db)
    client = TestClient(app, raise_server_exceptions=False)
    client.headers.update(authenticated_client.headers)

    def broken_save(_file):
        raise ValueError("server-side save bug")

    monkeypatch.setattr(maintenance_api, "_save_roundtrip_upload", broken_save)

    response = client.post(
        "/api/maintenance/roundtrip-import",
        files={
            "file": (
                "maintenance_roundtrip.xlsx",
                b"valid-multipart-envelope",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 500


def test_roundtrip_upload_cleanup_error_does_not_replace_save_primary(
    monkeypatch,
):
    original_mkstemp = maintenance_api.tempfile.mkstemp
    original_remove = os.remove
    created_paths: list[str] = []

    class BrokenUploadBody:
        def read(self, _size):
            raise RuntimeError("save primary")

    def tracking_mkstemp(*args, **kwargs):
        fd, path = original_mkstemp(*args, **kwargs)
        created_paths.append(path)
        return fd, path

    def denied_remove(path):
        if path in created_paths:
            raise PermissionError("cleanup secondary")
        return original_remove(path)

    monkeypatch.setattr(maintenance_api.tempfile, "mkstemp", tracking_mkstemp)
    monkeypatch.setattr(maintenance_api.os, "remove", denied_remove)

    try:
        with pytest.raises(RuntimeError, match="save primary"):
            maintenance_api._save_roundtrip_upload(
                SimpleNamespace(
                    filename="maintenance_roundtrip.xlsx",
                    file=BrokenUploadBody(),
                )
            )
    finally:
        for path in created_paths:
            if os.path.exists(path):
                original_remove(path)


@pytest.mark.parametrize(
    "unlink_error",
    [
        PermissionError("roundtrip temp unlink denied"),
        OSError("roundtrip temp unlink failed"),
    ],
)
def test_roundtrip_import_unlink_os_error_never_strands_process_lock(
    db,
    monkeypatch,
    tmp_path,
    unlink_error,
):
    authenticated_client = _admin_client(db)
    client = TestClient(app, raise_server_exceptions=False)
    client.headers.update(authenticated_client.headers)
    owned_path = tmp_path / "unlink-permission-error.xlsx"
    owned_path.write_bytes(b"owned-upload")
    original_remove = os.remove
    reacquired = False

    def saved_upload(_file):
        return str(owned_path), owned_path.name

    async def successful_import(*_args):
        return {"status": "success", "no_op": True}

    def denied_remove(path):
        if os.fspath(path) == os.fspath(owned_path):
            raise unlink_error
        return original_remove(path)

    monkeypatch.setattr(maintenance_api, "_save_roundtrip_upload", saved_upload)
    monkeypatch.setattr(
        maintenance_api,
        "_wait_for_roundtrip_import_worker",
        successful_import,
    )
    monkeypatch.setattr(maintenance_api.os, "remove", denied_remove)

    try:
        response = client.post(
            "/api/maintenance/roundtrip-import",
            files={
                "file": (
                    "maintenance_roundtrip.xlsx",
                    b"valid-multipart-envelope",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        reacquired = maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.acquire(
            blocking=False
        )

        assert response.status_code == 200
        assert reacquired is True
    finally:
        if reacquired:
            maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()
        elif maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.locked():
            maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()
        if owned_path.exists():
            original_remove(owned_path)


def test_roundtrip_import_repeated_cancel_after_save_waits_for_form_close_and_unlinks_owned_temp(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    authorization = client.headers["Authorization"]
    saved_paths: list[str] = []
    close_started = threading.Event()
    close_release = threading.Event()
    close_finished = threading.Event()
    original_save = maintenance_api._save_roundtrip_upload
    original_close = maintenance_api.UploadFile.close

    def tracking_save(file):
        path, filename = original_save(file)
        saved_paths.append(path)
        return path, filename

    async def delayed_close(file):
        close_started.set()
        assert await asyncio.to_thread(close_release.wait, 2)
        await original_close(file)
        close_finished.set()

    monkeypatch.setattr(maintenance_api, "_save_roundtrip_upload", tracking_save)
    monkeypatch.setattr(maintenance_api.UploadFile, "close", delayed_close)
    body, boundary = _multipart_upload_body(1024)

    async def exercise():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport,
            base_url="http://testserver",
        ) as async_client:
            upload = asyncio.create_task(
                async_client.post(
                    "/api/maintenance/roundtrip-import",
                    content=body,
                    headers={
                        "Authorization": authorization,
                        "Content-Type": f"multipart/form-data; boundary={boundary}",
                    },
                )
            )
            assert await asyncio.to_thread(close_started.wait, 1)
            upload.cancel()
            await asyncio.sleep(0.05)
            upload.cancel()
            await asyncio.sleep(0.05)
            completed_before_close = upload.done()
            close_release.set()
            with pytest.raises(asyncio.CancelledError):
                await upload
            return completed_before_close

    completed_before_close = asyncio.run(exercise())

    assert completed_before_close is False
    assert close_finished.is_set()
    assert saved_paths and all(not os.path.exists(path) for path in saved_paths)
    assert maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.acquire(blocking=False)
    maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()


def test_roundtrip_import_repeated_cancel_during_save_waits_and_unlinks_owned_temp(
    db,
    monkeypatch,
    tmp_path,
):
    client = _admin_client(db)
    authorization = client.headers["Authorization"]
    entered = threading.Event()
    release = threading.Event()
    finished = threading.Event()
    saved_paths: list[str] = []
    owned_path = tmp_path / "repeated-cancel-save.xlsx"

    def delayed_save(_file):
        entered.set()
        assert release.wait(timeout=2)
        owned_path.write_bytes(b"saved-after-cancel")
        saved_paths.append(str(owned_path))
        finished.set()
        return str(owned_path), owned_path.name

    monkeypatch.setattr(maintenance_api, "_save_roundtrip_upload", delayed_save)
    body, boundary = _multipart_upload_body(1024)

    async def exercise():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport,
            base_url="http://testserver",
        ) as async_client:
            upload = asyncio.create_task(
                async_client.post(
                    "/api/maintenance/roundtrip-import",
                    content=body,
                    headers={
                        "Authorization": authorization,
                        "Content-Type": f"multipart/form-data; boundary={boundary}",
                    },
                )
            )
            assert await asyncio.to_thread(entered.wait, 1)
            upload.cancel()
            await asyncio.sleep(0.05)
            upload.cancel()
            await asyncio.sleep(0.05)
            completed_before_save = upload.done()
            release.set()
            with pytest.raises(asyncio.CancelledError):
                await upload
            assert await asyncio.to_thread(finished.wait, 1)
            return completed_before_save

    try:
        completed_before_save = asyncio.run(exercise())

        assert completed_before_save is False
        assert saved_paths and all(not os.path.exists(path) for path in saved_paths)
        assert maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.acquire(blocking=False)
        maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()
    finally:
        release.set()
        finished.wait(timeout=2)
        for path in saved_paths:
            if os.path.exists(path):
                os.remove(path)
        if maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.locked():
            maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()


def test_roundtrip_import_cleanup_os_error_does_not_replace_cancellation(
    monkeypatch,
    tmp_path,
):
    owned_path = tmp_path / "cancel-cleanup-error.xlsx"
    owned_path.write_bytes(b"owned-upload")
    original_remove = os.remove
    reacquired = False

    async def saved_upload(_request):
        return str(owned_path), owned_path.name

    async def cancelled_import(*_args):
        raise asyncio.CancelledError

    def denied_remove(path):
        if os.fspath(path) == os.fspath(owned_path):
            raise OSError("cancel cleanup secondary")
        return original_remove(path)

    monkeypatch.setattr(
        maintenance_api,
        "_parse_and_save_roundtrip_upload",
        saved_upload,
    )
    monkeypatch.setattr(
        maintenance_api,
        "_wait_for_roundtrip_import_worker",
        cancelled_import,
    )
    monkeypatch.setattr(maintenance_api.os, "remove", denied_remove)
    ctx = SimpleNamespace(
        is_authenticated=True,
        user_id="reviewer",
        role="admin",
        permissions=None,
    )

    try:
        with pytest.raises(asyncio.CancelledError):
            asyncio.run(
                maintenance_api.roundtrip_import(
                    SimpleNamespace(),
                    ctx=ctx,
                )
            )
        reacquired = maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.acquire(
            blocking=False
        )
        assert reacquired is True
    finally:
        if reacquired:
            maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()
        elif maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.locked():
            maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()
        if owned_path.exists():
            original_remove(owned_path)


def test_roundtrip_import_lock_covers_worker_and_cleanup_busy_request_reads_zero_body(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    authorization = client.headers["Authorization"]
    token = authorization.removeprefix("Bearer ")
    entered = threading.Event()
    release = threading.Event()
    imported_path: list[str] = []

    def slow_import(_worker_db, path, *, filename, operated_by):
        imported_path.append(path)
        entered.set()
        assert release.wait(timeout=2)
        return {"status": "success", "no_op": True}

    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        slow_import,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args, **_kwargs: None)
    first_body, boundary = _multipart_upload_body(1024)
    second_body, second_boundary = _multipart_upload_body(2 * 1024 * 1024)

    async def exercise():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport,
            base_url="http://testserver",
        ) as async_client:
            first = asyncio.create_task(async_client.post(
                "/api/maintenance/roundtrip-import",
                content=first_body,
                headers={
                    "Authorization": authorization,
                    "Content-Type": f"multipart/form-data; boundary={boundary}",
                },
            ))
            assert await asyncio.to_thread(entered.wait, 1)
            second = await _asgi_roundtrip_import_with_receive_meter_async(
                second_body,
                second_boundary,
                token=token,
            )
            release.set()
            return await first, second

    first_response, (status, consumed, headers) = asyncio.run(exercise())

    assert first_response.status_code == 200
    assert status == 429
    assert consumed == 0
    assert headers[b"retry-after"] == b"5"
    assert imported_path and not os.path.exists(imported_path[0])


def test_roundtrip_import_repeated_cancel_keeps_lock_until_worker_finishes_and_cleans_temp(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    authorization = client.headers["Authorization"]
    token = authorization.removeprefix("Bearer ")
    entered = threading.Event()
    release = threading.Event()
    imported_path: list[str] = []

    def slow_import(_worker_db, path, **_kwargs):
        imported_path.append(path)
        entered.set()
        assert release.wait(timeout=2)
        return {"status": "success", "no_op": True}

    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        slow_import,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args, **_kwargs: None)
    body, boundary = _multipart_upload_body(1024)
    busy_body, busy_boundary = _multipart_upload_body(2 * 1024 * 1024)

    async def exercise():
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport,
            base_url="http://testserver",
        ) as async_client:
            upload = asyncio.create_task(
                async_client.post(
                    "/api/maintenance/roundtrip-import",
                    content=body,
                    headers={
                        "Authorization": authorization,
                        "Content-Type": f"multipart/form-data; boundary={boundary}",
                    },
                )
            )
            assert await asyncio.to_thread(entered.wait, 1)
            upload.cancel()
            await asyncio.sleep(0.05)
            upload.cancel()
            await asyncio.sleep(0.05)
            completed_before_worker = upload.done()
            busy_task = asyncio.create_task(
                _asgi_roundtrip_import_with_receive_meter_async(
                    busy_body,
                    busy_boundary,
                    token=token,
                )
            )
            await asyncio.sleep(0.05)
            release.set()
            with pytest.raises(asyncio.CancelledError):
                await upload
            return await busy_task, completed_before_worker

    (status, consumed, headers), completed_before_worker = asyncio.run(exercise())

    assert completed_before_worker is False
    assert status == 429
    assert consumed == 0
    assert headers[b"retry-after"] == b"5"
    assert imported_path and not os.path.exists(imported_path[0])
    assert maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.acquire(blocking=False)
    maintenance_api._ROUNDTRIP_IMPORT_PARSE_LOCK.release()


def test_roundtrip_import_heavy_sync_work_does_not_block_health_or_cross_session_threads(
    db,
    monkeypatch,
):
    client = _admin_client(db)
    token = client.headers["Authorization"]
    loop_thread: list[int] = []
    session_threads: list[int] = []
    import_threads: list[int] = []
    imported_paths: list[str] = []
    release = threading.Event()
    original_session_factory = maintenance_api.SessionLocal
    parser_chunk_sizes: list[int] = []
    access_log_threads: list[int] = []

    class TrackingParser(StarletteMultiPartParser):
        def __init__(self, headers, stream, **kwargs):
            async def tracked_stream():
                async for chunk in stream:
                    parser_chunk_sizes.append(len(chunk))
                    yield chunk

            super().__init__(headers, tracked_stream(), **kwargs)

    def session_factory():
        session_threads.append(threading.get_ident())
        return original_session_factory()

    def slow_import(worker_db, path, *, filename, operated_by):
        import_threads.append(threading.get_ident())
        imported_paths.append(path)
        assert filename == "maintenance_roundtrip.xlsx"
        assert operated_by
        assert worker_db.is_active
        assert release.wait(timeout=2)
        return {"status": "success", "no_op": True}

    monkeypatch.setattr(maintenance_api, "SessionLocal", session_factory)
    monkeypatch.setattr(maintenance_api, "MultiPartParser", TrackingParser)
    monkeypatch.setattr(
        maintenance_api,
        "record_access_log",
        lambda *_args, **_kwargs: access_log_threads.append(threading.get_ident()),
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        slow_import,
    )
    body, boundary = _multipart_upload_body(4 * 1024 * 1024)

    async def exercise():
        loop_thread.append(threading.get_ident())
        transport = httpx.ASGITransport(app=app)
        async with httpx.AsyncClient(
            transport=transport,
            base_url="http://testserver",
        ) as async_client:
            timer = threading.Timer(0.4, release.set)
            timer.start()
            started_at = time.perf_counter()

            async def delayed_health():
                await asyncio.sleep(0.05)
                response = await async_client.get("/health")
                return response, time.perf_counter() - started_at

            upload, heartbeat = await asyncio.gather(
                async_client.post(
                    "/api/maintenance/roundtrip-import",
                    content=body,
                    headers={
                        "Authorization": token,
                        "Content-Type": f"multipart/form-data; boundary={boundary}",
                    },
                ),
                delayed_health(),
            )
            timer.join()
            return upload, heartbeat

    upload, (health, heartbeat_elapsed) = asyncio.run(exercise())

    assert upload.status_code == 200, upload.text
    assert health.status_code == 200
    assert heartbeat_elapsed < 0.25
    assert session_threads == import_threads
    assert import_threads and import_threads[0] != loop_thread[0]
    assert access_log_threads == import_threads
    assert parser_chunk_sizes
    assert max(parser_chunk_sizes) <= maintenance_api._ROUNDTRIP_PARSE_CHUNK_BYTES
    assert imported_paths and not os.path.exists(imported_paths[0])


def test_roundtrip_import_worker_exception_rolls_back_closes_session_and_removes_temp(
    db,
    monkeypatch,
):
    authenticated_client = _admin_client(db)
    client = TestClient(app, raise_server_exceptions=False)
    client.headers.update(authenticated_client.headers)
    events: list[tuple[str, int]] = []
    imported_paths: list[str] = []

    class FakeSession:
        def rollback(self):
            events.append(("rollback", threading.get_ident()))

        def close(self):
            events.append(("close", threading.get_ident()))

    def session_factory():
        events.append(("create", threading.get_ident()))
        return FakeSession()

    def failed_import(_worker_db, path, **_kwargs):
        events.append(("use", threading.get_ident()))
        imported_paths.append(path)
        raise RuntimeError("worker failed")

    monkeypatch.setattr(maintenance_api, "SessionLocal", session_factory)
    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        failed_import,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args, **_kwargs: None)

    response = client.post(
        "/api/maintenance/roundtrip-import",
        files={
            "file": (
                "maintenance_roundtrip.xlsx",
                b"worker-failure",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 500
    assert [name for name, _thread_id in events] == [
        "create", "use", "rollback", "close",
    ]
    assert len({thread_id for _name, thread_id in events}) == 1
    assert imported_paths and not os.path.exists(imported_paths[0])


@pytest.mark.parametrize(
    ("rollback_error", "close_error"),
    [
        (LookupError("rollback secondary"), None),
        (None, OSError("close secondary")),
        (LookupError("rollback secondary"), OSError("close secondary")),
    ],
)
def test_roundtrip_import_worker_cleanup_errors_preserve_import_primary(
    monkeypatch,
    rollback_error,
    close_error,
):
    events: list[str] = []
    primary_error = RuntimeError("import primary")

    class FakeSession:
        def rollback(self):
            events.append("rollback")
            if rollback_error is not None:
                raise rollback_error

        def close(self):
            events.append("close")
            if close_error is not None:
                raise close_error

    def failed_import(*_args, **_kwargs):
        events.append("import")
        raise primary_error

    monkeypatch.setattr(maintenance_api, "SessionLocal", FakeSession)
    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        failed_import,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args, **_kwargs: None)

    with pytest.raises(RuntimeError, match="import primary") as exc_info:
        maintenance_api._import_roundtrip_in_worker(
            "/tmp/roundtrip-primary.xlsx",
            "maintenance_roundtrip.xlsx",
            "reviewer",
            SimpleNamespace(),
        )

    assert exc_info.value is primary_error
    assert events == ["import", "rollback", "close"]


def test_roundtrip_import_worker_close_error_after_success_is_warning_only(
    monkeypatch,
):
    events: list[str] = []
    warnings: list[tuple[str, dict]] = []
    expected_result = {"status": "success", "no_op": False}

    class FakeSession:
        def rollback(self):
            events.append("rollback")

        def close(self):
            events.append("close")
            raise OSError("close after committed import")

    def successful_import(*_args, **_kwargs):
        events.append("import")
        return expected_result

    monkeypatch.setattr(maintenance_api, "SessionLocal", FakeSession)
    monkeypatch.setattr(
        maintenance_roundtrip,
        "import_roundtrip_workbook",
        successful_import,
    )
    monkeypatch.setattr(maintenance_api, "record_access_log", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        maintenance_api.logger,
        "warning",
        lambda message, *_args, **kwargs: warnings.append((message, kwargs)),
    )

    result = maintenance_api._import_roundtrip_in_worker(
        "/tmp/roundtrip-committed.xlsx",
        "maintenance_roundtrip.xlsx",
        "reviewer",
        SimpleNamespace(),
    )

    assert result is expected_result
    assert events == ["import", "close"]
    assert warnings == [
        ("维保回填成功后的数据库会话关闭失败", {"exc_info": True})
    ]


def test_roundtrip_apply_action_403_happens_before_application_reads_upload(
    db,
    monkeypatch,
):
    user = SysUser(
        username="roundtrip-no-action",
        role="boss",
        display_name="无回填动作权限",
        password_hash=hash_password("roundtrip-password"),
        permissions={"action_maintenance_roundtrip_apply": False},
    )
    db.add(user)
    db.commit()
    called = False

    def fail_if_upload_is_read(_file):
        nonlocal called
        called = True
        raise AssertionError("upload body reached application save path")

    monkeypatch.setattr(
        maintenance_api,
        "_save_roundtrip_upload",
        fail_if_upload_is_read,
    )
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={
            "username": "roundtrip-no-action",
            "password": "roundtrip-password",
        },
    )
    assert login.status_code == 200
    response = client.post(
        "/api/maintenance/roundtrip-import",
        headers={"Authorization": f"Bearer {login.json()['token']}"},
        files={
            "file": (
                "maintenance_roundtrip.xlsx",
                b"must-not-reach-application-reader",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403
    assert response.json()["detail"] == "无此操作权限"
    assert called is False


def test_roundtrip_import_requires_customer_info_before_reading_upload(
    db,
    monkeypatch,
):
    base = permissions.effective("boss", None)
    overrides = {"data_customer": False}
    user = SysUser(
        username="roundtrip-no-customer-info",
        role="boss",
        display_name="无客户权限回填员",
        password_hash=hash_password("roundtrip-password"),
        template_code="boss",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    )
    db.add(user)
    db.commit()
    called = False

    def fail_if_upload_is_read(_file):
        nonlocal called
        called = True
        raise AssertionError("customer permission must fail before upload read")

    monkeypatch.setattr(
        maintenance_api,
        "_save_roundtrip_upload",
        fail_if_upload_is_read,
    )
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={
            "username": "roundtrip-no-customer-info",
            "password": "roundtrip-password",
        },
    )
    assert login.status_code == 200

    response = client.post(
        "/api/maintenance/roundtrip-import",
        headers={"Authorization": f"Bearer {login.json()['token']}"},
        files={
            "file": (
                "maintenance_roundtrip.xlsx",
                b"must-not-reach-application-reader",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 403
    assert "客户信息" in response.json()["detail"]
    assert called is False
