"""维保台账工作簿解析与 apply 测试（B2）。"""

import io
from types import SimpleNamespace
from uuid import uuid4

import pytest
from datetime import date, datetime, timezone
from decimal import Decimal

from openpyxl import Workbook
from sqlalchemy import select, text

from app.models.maintenance import FProjectExpense
from app.models.maintenance_ledger import (
    MaintenanceLedgerContractRow,
    MaintenanceLedgerExpenseRow,
    MaintenanceLedgerImportBatch,
    MaintenanceLedgerPlanRow,
)
from app.models.maintenance_manager import MaintenanceCollectionMilestone
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectAuditLog,
    MaintenanceProjectContract,
)
from app.models.sales import FSalesOrder
from app.models.maintenance_project_operations import (
    MaintenanceProjectExpenseAttribution,
    MaintenanceProjectWorkbookState,
)
from app.models.system import SysImportBatch
from app.services import maintenance_ledger as ledger
from app.services import maintenance_project_catalog as project_catalog
from app.services import maintenance_project_operations as project_operations_service


def _old_ledger_workbook_bytes(*, include_expense: bool = False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "维保项目清单"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款", "验收材料",
         "验收材料是否完成及上传附件", "巡检时间", "巡检是否完成及上传附件",
         "回款时间1", "回款金额", "回款时间2", "回款金额", "回款时间3", "回款金额"]
    )
    ws.append(
        ["XSDD-20260731-0086", "2026-07-31", "李呈辉", "整体维保",
         "阿里专有云20260608-20291205", "2026-06-08", "2029-12-05", "廖晓娟", "任鑫明",
         44756, 0, 44756, "提供设备硬件维修记录报告；服务总结报告", "否", "/", "否",
         "2026年10月", 2986.57, "2027年1月", 2986.57, "2027年4月", 2986.57]
    )
    ws.append(
        ["XSDD-20260731-0040", "2026-07-31", "李呈辉", "备件维保",
         "正大天晴20260801-20270531因", "2026-08-01", "2027-05-31", "廖晓娟", "/",
         27400, 0, 27400, "备件发货记录", "", "/", "", "2027年2月", 6850]
    )
    if include_expense:
        cost = wb.create_sheet("项目成本")
        cost.append(
            ["费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单", "项目名称",
             "销售订单", "销售人员", "费用分类", "报销金额", "备注"]
        )
        cost.append(
            ["BXD-20260425-0002", "董学晶", "维保费用", "2026年广西国税2月份第二次巡检和4月",
             "XSDD-20251028-0016", "国税总局存储20251102-20281101北京ST存储设备维保项目",
             "XSDD-20251028-0016", "余俊", "差旅费", 1068.5, ""]
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _new_ledger_workbook_bytes(*, include_expense: bool = False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "01_项目与合同"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款", "验收材料",
         "验收材料是否完成及上传附件", "验收附件", "巡检时间", "巡检是否完成及上传附件"]
    )
    ws.append(
        ["XSDD-20260731-0086", "2026-07-31", "李呈辉", "整体维保",
         "阿里专有云20260608-20291205", "2026-06-08", "2029-12-05", "廖晓娟", "任鑫明",
         44756, 0, 44756, "服务总结报告", "否", "", "2026-10", "否"]
    )
    plan = wb.create_sheet("02_回款计划")
    plan.append(["订单编号", "计划期次", "计划回款时间", "计划回款金额"])
    plan.append(["XSDD-20260731-0086", 1, "2026-10", 2986.57])
    plan.append(["XSDD-20260731-0086", 2, "2027-01", 2986.57])
    if include_expense:
        cost = wb.create_sheet("03_项目成本")
        cost.append(
            ["费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单", "项目名称",
             "销售订单", "销售人员", "费用分类", "报销金额", "备注"]
        )
        cost.append(
            ["BXD-20260425-0002", "董学晶", "维保费用", "巡检",
             "XSDD-20251028-0016", "国税总局存储20251102-20281101北京ST存储设备维保项目",
             "XSDD-20251028-0016", "余俊", "差旅费", 1068.5, ""]
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _expense_ledger_workbook_bytes(
    *,
    project_name: str = "费用联动项目20260608-20291205",
    order_no: str = "XSDD-20260801-0001",
    period_end: str = "2029-12-05",
    bxd_no: str | None = "BXD-20260801-0001",
    expense_amount: Decimal = Decimal("113.00"),
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "维保项目清单"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款",
         "验收材料", "验收材料是否完成及上传附件", "巡检时间", "巡检是否完成及上传附件"]
    )
    ws.append(
        [order_no, "2026-06-08", "李呈辉", "整体维保", project_name,
         "2026-06-08", period_end, "廖晓娟", "任鑫明", 10000, 0, 10000,
         "", "", "", ""]
    )
    if bxd_no is not None:
        cost = wb.create_sheet("项目成本")
        cost.append(
            ["费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单", "项目名称",
             "销售订单", "销售人员", "费用分类", "报销金额", "备注"]
        )
        cost.append(
            [bxd_no, "张三", "维保费用", "现场巡检", order_no, project_name,
             order_no, "李呈辉", "差旅费", expense_amount, ""]
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _formal_expense(
    db,
    *,
    raw_line_id: str = "ledger-expense-raw-1",
    bxd_no: str = "BXD-20260801-0001",
    order_no: str = "XSDD-20260801-0001",
    expense_date: date = date(2026, 8, 1),
) -> FProjectExpense:
    import_batch = SysImportBatch(
        filename=f"{raw_line_id}.xlsx",
        file_type="expense",
        file_hash=uuid4().hex * 2,
        status="success",
    )
    db.add(import_batch)
    db.flush()
    raw = FProjectExpense(
        raw_line_id=raw_line_id,
        bxd_no=bxd_no,
        line_no=1,
        data_status="已结束",
        expense_date=expense_date,
        person="张三",
        expense_type="维保费用",
        fee_category="差旅费",
        reason="现场巡检",
        linked_sales_order_no=order_no,
        amount=Decimal("113.00"),
        amount_ex_tax=Decimal("100.00"),
        amount_inc_tax=Decimal("113.00"),
        tax_basis="inc",
        tax_rate_used=Decimal("0.13"),
        import_batch_id=import_batch.id,
    )
    db.add(raw)
    db.flush()
    return raw


def _mapped_contract_expense(db):
    project_name = "费用联动项目20260608-20291205"
    order_no = "XSDD-20260801-0001"
    project = MaintenanceProject(
        project_id=str(uuid4()),
        project_code=project_name,
        display_name=project_name,
        project_manager_id="任鑫明",
        business_type="整体维保",
        cmo_name="廖晓娟",
        salesperson="李呈辉",
        period_from=date(2026, 6, 8),
        period_to=date(2029, 12, 5),
        lifecycle_status="ongoing",
        is_active=True,
        version=1,
    )
    db.add(project)
    db.flush()
    contract = MaintenanceProjectContract(
        project_contract_id=str(uuid4()),
        project_id=project.project_id,
        contract_id=order_no,
        contract_no=order_no,
        contract_amount=None,
        amount_inc_tax=Decimal("10000.00"),
        contract_status=None,
        status_mapping_state="mapped",
        status_mapping_version="project_manager_xls_v1",
        included_in_total=True,
        effective_from=date(2026, 6, 8),
        effective_to=date(2029, 12, 5),
        source="project_manager_xls_v1",
        version=1,
    )
    db.add(contract)
    raw = _formal_expense(db)
    attribution = MaintenanceProjectExpenseAttribution(
        expense_id=f"bxd:{raw.raw_line_id}",
        project_id=project.project_id,
        project_contract_id=contract.project_contract_id,
        raw_expense_line_id=raw.raw_line_id,
        expense_ref=f"{raw.bxd_no}#{raw.line_no}",
        expense_date=raw.expense_date,
        applicant=raw.person,
        category=raw.fee_category,
        expense_reason=raw.reason,
        tax_basis=raw.tax_basis,
        amount_ex_tax=raw.amount_ex_tax,
        amount_inc_tax=raw.amount_inc_tax,
        tax_rate_used=raw.tax_rate_used,
        raw_status=raw.data_status,
        status_mapping_state="mapped",
        normalized_status="approved",
        status_mapping_version="expense-loader-v1",
        ownership_mapping_state="mapped",
        ownership_mapping_version="ownership-v1",
        version=1,
    )
    db.add(attribution)
    db.commit()
    return project, contract, raw, attribution


def test_parse_expense_sheet_accepts_detail_prefix():
    wb = Workbook()
    ws = wb.active
    ws.title = "维保项目清单"
    ws.append(["订单编号"])
    ws.append(["XSDD-20260731-0086"])
    cost = wb.create_sheet("项目成本")
    cost.append(
        ["费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单", "项目名称",
         "销售订单", "销售人员", "报销明细.费用分类", "报销明细.报销金额", "备注"]
    )
    cost.append(
        ["BXD-20260425-0002", "董学晶", "维保费用", "巡检",
         "XSDD-20251028-0016", "某项目", "XSDD-20251028-0016", "余俊", "差旅费", 1068.5, ""]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    parsed = ledger.parse_ledger_workbook(buffer.getvalue(), "台账.xlsx")
    assert parsed["expense_rows"][0].values["费用分类"] == "差旅费"
    assert parsed["expense_rows"][0].values["报销金额"] == "1068.5"


def test_parse_old_ledger_structure():
    parsed = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(include_expense=True), "维保台账.xlsx"
    )
    assert parsed["source_kind"] == "project_manager_xls_v1"
    assert len(parsed["contract_rows"]) == 2
    assert len(parsed["plan_rows"]) == 4  # 3 + 1 横向回款对展开
    assert len(parsed["expense_rows"]) == 1
    # 第一个合同行的回款计划：期次 1/2/3 金额一致
    first_plans = [p for p in parsed["plan_rows"] if p.order_no_raw == "XSDD-20260731-0086"]
    assert [p.sequence for p in first_plans] == [1, 2, 3]
    assert [str(p.amount_raw) for p in first_plans] == ["2986.57"] * 3


def test_parse_new_template_structure():
    parsed = ledger.parse_ledger_workbook(
        _new_ledger_workbook_bytes(include_expense=True), "维保台账工作簿模板_v1.xlsx"
    )
    assert parsed["source_kind"] == "ledger_template_v1"
    assert len(parsed["contract_rows"]) == 1
    assert len(parsed["plan_rows"]) == 2
    assert len(parsed["expense_rows"]) == 1


def test_parse_rejects_unknown_structure():
    wb = Workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    try:
        ledger.parse_ledger_workbook(buffer.getvalue(), "x.xlsx")
        raise AssertionError("应抛出 LedgerParseError")
    except ledger.LedgerParseError:
        pass


def test_store_preview_then_apply(db):
    parsed = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(include_expense=True), "维保台账.xlsx"
    )
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key="ledger-test-key-0001")

    batch = db.get(MaintenanceLedgerImportBatch, batch_id)
    assert batch.status == "pending"
    assert batch.contract_rows == 2
    assert batch.plan_rows == 4
    assert batch.expense_rows == 1

    contract_rows = db.execute(
        select(MaintenanceLedgerContractRow).where(
            MaintenanceLedgerContractRow.batch_id == batch_id
        )
    ).scalars().all()
    row = next(r for r in contract_rows if r.order_no == "XSDD-20260731-0086")
    assert row.project_period_from == date(2026, 6, 8)
    assert row.project_period_to == date(2029, 12, 5)
    assert float(row.amount_inc_tax) == 44756.0

    plan_rows = db.execute(
        select(MaintenanceLedgerPlanRow).where(
            MaintenanceLedgerPlanRow.batch_id == batch_id
        )
    ).scalars().all()
    month_plan = next(p for p in plan_rows if p.sequence == 1)
    assert month_plan.planned_date == date(2026, 10, 1)
    assert month_plan.date_precision == "month"

    expense_rows = db.execute(
        select(MaintenanceLedgerExpenseRow).where(
            MaintenanceLedgerExpenseRow.batch_id == batch_id
        )
    ).scalars().all()
    assert expense_rows[0].bxd_no == "BXD-20260425-0002"
    assert float(expense_rows[0].amount) == 1068.5


def test_apply_syncs_canonical_tables(db):
    from app.models.system import SysImportBatch

    import_batch = SysImportBatch(
        filename="sales.xlsx", file_type="sales", file_hash="h", status="success"
    )
    db.add(import_batch)
    db.flush()
    db.add(
        FSalesOrder(
            raw_order_id="r1",
            order_no="XSDD-20260731-0086",
            order_date=date(2026, 7, 31),
            salesperson="李呈辉",
            business_type="整体维保",
            warehouse="北京成品仓",
            amount_ex_tax=39607.08,
            tax_rate=0.13,
            data_status="已生效",
            import_batch_id=import_batch.id,
        )
    )
    db.commit()
    parsed = ledger.parse_ledger_workbook(_old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key="ledger-test-key-0001")
    summary = ledger.apply_batch(db, batch_id, "合成管理员")

    assert summary["projects_created"] == 2
    assert summary["contracts_created"] == 2
    assert summary["milestones_created"] == 4
    assert summary["skipped_rows"] == 0

    project = db.execute(
        select(MaintenanceProject).where(
            MaintenanceProject.project_code == "阿里专有云20260608-20291205"
        )
    ).scalar_one()
    assert project.lifecycle_status == "ongoing"
    assert project.cmo_name == "廖晓娟"
    assert project.business_type == "整体维保"

    contract = db.execute(
        select(MaintenanceProjectContract).where(
            MaintenanceProjectContract.project_id == project.project_id
        )
    ).scalar_one()
    assert float(contract.amount_inc_tax) == 44756.0
    # 未税金额来自销售单对账：44756 / 1.13 = 39607.08
    assert float(contract.contract_amount) == 39607.08
    assert contract.included_in_total is True
    assert contract.source == "project_manager_xls_v1"

    milestones = db.execute(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id
            == contract.project_contract_id
        )
    ).scalars().all()
    assert len(milestones) == 3
    assert milestones[0].source == "project_manager_xls_v1"

    batch = db.get(MaintenanceLedgerImportBatch, batch_id)
    assert batch.status == "applied"
    assert batch.applied_by == "合成管理员"

    audit = db.execute(
        select(MaintenanceProjectAuditLog).where(
            MaintenanceProjectAuditLog.project_id == project.project_id
        )
    ).scalars().all()
    assert any(a.action == "ledger_create" for a in audit)


def test_apply_expense_e2e_uses_locked_formal_raw_and_contract(db):
    """正式 BXD + 台账合同/费用在同一事务落归因，并只 bump 一次。"""

    raw = _formal_expense(db)
    db.commit()
    parsed = ledger.parse_ledger_workbook(
        _expense_ledger_workbook_bytes(), "维保台账.xlsx"
    )
    batch_id = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-expense-e2e-1",
    )

    summary = ledger.apply_batch(db, batch_id, "合成管理员")

    project = db.scalar(select(MaintenanceProject).where(
        MaintenanceProject.project_code == "费用联动项目20260608-20291205"
    ))
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project.project_id
    ))
    attribution = db.get(
        MaintenanceProjectExpenseAttribution, f"bxd:{raw.raw_line_id}"
    )
    state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    assert summary["expenses_created"] == 1
    assert attribution.project_id == project.project_id
    assert attribution.project_contract_id == contract.project_contract_id
    assert attribution.ownership_mapping_state == "mapped"
    assert attribution.amount_ex_tax == Decimal("100.00")
    assert attribution.amount_inc_tax == Decimal("113.00")
    assert state.revision == 1


def test_apply_revalidates_full_bxd_fingerprint_after_project_locks(
    db, monkeypatch
):
    """预检后 raw 金额变化必须整批回滚，不能按旧 plan 写 canonical。"""

    raw = _formal_expense(db, raw_line_id="ledger-expense-race-1")
    db.commit()
    parsed = ledger.parse_ledger_workbook(
        _expense_ledger_workbook_bytes(), "维保台账.xlsx"
    )
    batch_id = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-expense-race-1",
    )
    original_lock = ledger._lock_target_projects

    def mutate_after_preflight(*args, **kwargs):
        result = original_lock(*args, **kwargs)
        changed = db.get(FProjectExpense, raw.id)
        changed.amount = Decimal("226.00")
        changed.amount_ex_tax = Decimal("200.00")
        changed.amount_inc_tax = Decimal("226.00")
        db.flush()
        return result

    monkeypatch.setattr(ledger, "_lock_target_projects", mutate_after_preflight)
    with pytest.raises(ledger.LedgerBatchError, match="正式 BXD 的生效集合"):
        ledger.apply_batch(db, batch_id, "合成管理员")
    db.rollback()

    assert db.scalars(select(MaintenanceProject)).all() == []
    assert db.get(FProjectExpense, raw.id).amount_inc_tax == Decimal("113.00")


def test_apply_rejects_locked_attribution_owner_outside_state_envelope(
    db, monkeypatch
):
    """锁后 owner 漂移到未预锁项目时受控冲突，绝不晚拿 state。"""

    project, _contract, raw, attribution = _mapped_contract_expense(db)
    unrelated = MaintenanceProject(
        project_id=str(uuid4()),
        project_code="未预锁费用项目",
        display_name="未预锁费用项目",
        lifecycle_status="missing",
        is_active=True,
        version=1,
    )
    db.add(unrelated)
    db.commit()
    parsed = ledger.parse_ledger_workbook(
        _expense_ledger_workbook_bytes(), "维保台账.xlsx"
    )
    batch_id = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-expense-owner-race-1",
    )
    original_lock = ledger._lock_target_projects

    def move_owner_after_preflight(*args, **kwargs):
        result = original_lock(*args, **kwargs)
        current = db.get(
            MaintenanceProjectExpenseAttribution, attribution.expense_id
        )
        current.project_id = unrelated.project_id
        current.project_contract_id = None
        current.ownership_mapping_state = "unmapped"
        db.flush()
        return result

    monkeypatch.setattr(ledger, "_lock_target_projects", move_owner_after_preflight)
    with pytest.raises(ledger.LedgerBatchError, match="报销归属项目"):
        ledger.apply_batch(db, batch_id, "合成管理员")
    db.rollback()

    preserved = db.get(
        MaintenanceProjectExpenseAttribution, attribution.expense_id
    )
    assert preserved.project_id == project.project_id
    assert db.get(MaintenanceProjectWorkbookState, unrelated.project_id) is None
    assert db.get(FProjectExpense, raw.id) is not None


def test_contract_window_change_resyncs_raw_backed_expense_without_cost_sheet(db):
    """合同窗口缩短会立刻清退既有费用映射，即使本批没有费用页。"""

    project, contract, _raw, attribution = _mapped_contract_expense(db)
    probe = ledger._prepare_contract_expense_probe(
        db, contract_order_nos={contract.contract_no}
    )
    assert probe.raw_line_ids == ("ledger-expense-raw-1",)
    assert probe.raw_identities == (
        ("ledger-expense-raw-1", "20260801-0001"),
    )
    parsed = ledger.parse_ledger_workbook(
        _expense_ledger_workbook_bytes(
            period_end="2026-07-31",
            bxd_no=None,
        ),
        "维保台账.xlsx",
    )
    batch_id = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-contract-expense-resync-1",
    )

    summary = ledger.apply_batch(db, batch_id, "合成管理员")

    db.expire_all()
    updated = db.get(
        MaintenanceProjectExpenseAttribution, attribution.expense_id
    )
    state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    assert summary["contracts_updated"] == 1
    assert summary["expenses_updated"] == 1, (
        summary,
        updated.ownership_mapping_state,
        updated.project_contract_id,
        updated.version,
        db.get(
            MaintenanceProjectContract, contract.project_contract_id
        ).effective_to,
    )
    assert db.get(
        MaintenanceProjectContract, contract.project_contract_id
    ).effective_to == date(2026, 7, 31)
    assert updated.project_id == project.project_id
    assert updated.project_contract_id is None
    assert updated.ownership_mapping_state == "unmapped"
    assert state.revision == 1


def test_sales_evidence_uses_latest_consistent_success_and_ignores_invalid_rows(
    db,
):
    """Duplicate versions are deterministic and never raise MultipleResultsFound."""

    from app.models.system import SysImportBatch

    older = SysImportBatch(
        filename="sales-older.xlsx",
        file_type="sales",
        file_hash="ledger-sales-evidence-older",
        status="success",
        uploaded_at=datetime(2026, 8, 1, tzinfo=timezone.utc),
    )
    newer = SysImportBatch(
        filename="sales-newer.xlsx",
        file_type="sales",
        file_hash="ledger-sales-evidence-newer",
        status="success",
        uploaded_at=datetime(2026, 8, 2, tzinfo=timezone.utc),
    )
    failed = SysImportBatch(
        filename="sales-failed.xlsx",
        file_type="sales",
        file_hash="ledger-sales-evidence-failed",
        status="failed",
        uploaded_at=datetime(2026, 8, 3, tzinfo=timezone.utc),
    )
    db.add_all([older, newer, failed])
    db.flush()
    order_no = "XSDD-20260731-0086"
    db.add_all([
        FSalesOrder(
            raw_order_id="ledger-sales-evidence-old",
            order_no=order_no,
            order_date=date(2026, 7, 30),
            amount_ex_tax=Decimal("39607.08"),
            tax_rate=Decimal("0.13"),
            data_status="已生效",
            import_batch_id=older.id,
        ),
        FSalesOrder(
            raw_order_id="ledger-sales-evidence-new",
            order_no=order_no,
            order_date=date(2026, 7, 31),
            amount_ex_tax=Decimal("39607.08"),
            tax_rate=Decimal("0.13"),
            data_status="已生效",
            import_batch_id=newer.id,
        ),
        # Newer failed and inactive rows must not become evidence or conflicts.
        FSalesOrder(
            raw_order_id="ledger-sales-evidence-failed",
            order_no=order_no,
            order_date=date(2026, 8, 1),
            amount_ex_tax=Decimal("1.00"),
            tax_rate=Decimal("0.13"),
            data_status="已生效",
            import_batch_id=failed.id,
        ),
        FSalesOrder(
            raw_order_id="ledger-sales-evidence-inactive",
            order_no=order_no,
            order_date=date(2026, 8, 2),
            amount_ex_tax=Decimal("2.00"),
            tax_rate=Decimal("0.13"),
            data_status="已作废",
            import_batch_id=newer.id,
        ),
        FSalesOrder(
            raw_order_id="ledger-sales-half-up",
            order_no="XSDD-LEDGER-HALF-UP",
            order_date=date(2026, 8, 2),
            amount_ex_tax=Decimal("1.00"),
            tax_rate=Decimal("0.0050"),
            data_status="已生效",
            import_batch_id=newer.id,
        ),
    ])
    db.commit()

    resolved, conflicts = ledger._load_sales_order_evidence(
        db, {order_no, "XSDD-LEDGER-HALF-UP"}
    )
    assert conflicts == set()
    assert resolved[order_no].raw_order_id == "ledger-sales-evidence-new"
    assert ledger._sales_order_inc_tax(
        resolved["XSDD-LEDGER-HALF-UP"]
    ) == Decimal("1.01")

    parsed = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-sales-evidence-consistent",
    )
    summary = ledger.apply_batch(db, batch_id, "合成管理员")
    assert summary["contracts_created"] == 2
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.contract_no == order_no
    ))
    assert contract.contract_amount == Decimal("39607.08")


def test_apply_rejects_conflicting_successful_sales_economics(db):
    """Economic ambiguity is a controlled zero-write batch rejection."""

    from app.models.system import SysImportBatch

    batches = [
        SysImportBatch(
            filename=f"sales-conflict-{index}.xlsx",
            file_type="sales",
            file_hash=f"ledger-sales-conflict-{index}",
            status="success",
            uploaded_at=datetime(2026, 8, index, tzinfo=timezone.utc),
        )
        for index in (1, 2)
    ]
    db.add_all(batches)
    db.flush()
    order_no = "XSDD-20260731-0086"
    db.add_all([
        FSalesOrder(
            raw_order_id=f"ledger-sales-conflict-{index}",
            order_no=order_no,
            order_date=date(2026, 7, 31),
            amount_ex_tax=amount,
            tax_rate=Decimal("0.13"),
            data_status="已生效",
            import_batch_id=batch.id,
        )
        for index, (batch, amount) in enumerate(zip(
            batches,
            (Decimal("39607.08"), Decimal("40000.00")),
        ), start=1)
    ])
    db.commit()

    parsed = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-sales-evidence-conflict",
    )
    with pytest.raises(ledger.LedgerBatchError, match="金额对账失败"):
        ledger.apply_batch(db, batch_id, "合成管理员")

    batch = db.get(MaintenanceLedgerImportBatch, batch_id)
    assert batch.status == "failed"
    assert "未税额/税率存在冲突" in " ".join(
        batch.report_json["reconcile_failures"]
    )
    assert db.scalars(select(MaintenanceProject)).all() == []
    assert db.scalars(select(MaintenanceProjectContract)).all() == []


def test_apply_idempotent_and_version_bump(db):
    parsed = ledger.parse_ledger_workbook(_old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key="ledger-test-key-0001")
    first = ledger.apply_batch(db, batch_id, "合成管理员")
    assert first["contracts_created"] == 2

    # 同一文件再次导入：新批次，同值不重复写（无 updated/created 增量）
    parsed2 = ledger.parse_ledger_workbook(_old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id2 = ledger.store_preview(db, parsed2, "合成管理员", idempotency_key="ledger-test-key-0002")
    second = ledger.apply_batch(db, batch_id2, "合成管理员")
    assert second["contracts_created"] == 0
    assert second["contracts_updated"] == 0
    assert second["projects_created"] == 0

    # 金额变化 → 合同版本 +1（重建一个金额不同的工作簿）
    wb2 = Workbook()
    ws = wb2.active
    ws.title = "维保项目清单"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款", "验收材料",
         "验收材料是否完成及上传附件", "巡检时间", "巡检是否完成及上传附件",
         "回款时间1", "回款金额"]
    )
    ws.append(
        ["XSDD-20260731-0086", "2026-07-31", "李呈辉", "整体维保",
         "阿里专有云20260608-20291205", "2026-06-08", "2029-12-05", "廖晓娟", "任鑫明",
         50000, 0, 50000, "", "", "", "", "2026-10", 2986.57]
    )
    buffer = io.BytesIO()
    wb2.save(buffer)
    parsed3 = ledger.parse_ledger_workbook(buffer.getvalue(), "维保台账.xlsx")
    batch_id3 = ledger.store_preview(db, parsed3, "合成管理员", idempotency_key="ledger-test-key-0003")
    third = ledger.apply_batch(db, batch_id3, "合成管理员")
    assert third["contracts_updated"] == 1

    contract = db.execute(
        select(MaintenanceProjectContract).where(
            MaintenanceProjectContract.contract_no == "XSDD-20260731-0086"
        )
    ).scalar_one()
    assert float(contract.amount_inc_tax) == 50000.0
    assert contract.version == 2


def test_target_project_lock_order_is_independent_of_uploaded_row_order(db):
    """真实 PostgreSQL advisory/state/project 锁序不得继承 Excel 行序。"""
    projects = [
        MaintenanceProject(
            project_id=str(uuid4()),
            project_code=code,
            display_name=code,
            lifecycle_status="missing",
            is_active=True,
            version=1,
        )
        for code in ("LOCK-B", "lock-a")
    ]
    db.add_all(projects)
    db.commit()
    rows = [
        SimpleNamespace(project_name="LOCK-B", project_name_raw="LOCK-B"),
        SimpleNamespace(project_name="LOCK-A", project_name_raw="LOCK-A"),
    ]

    forward_identities = ledger._target_project_identities(rows)
    reverse_identities = ledger._target_project_identities(list(reversed(rows)))
    assert forward_identities == reverse_identities == ["lock-a", "lock-b"]

    forward_locked, _, _ = ledger._lock_target_projects(db, rows)
    reverse_locked, _, _ = ledger._lock_target_projects(db, list(reversed(rows)))
    expected_project_order = sorted(project.project_id for project in projects)
    assert forward_locked == reverse_locked == expected_project_order


def test_upsert_project_reuses_case_insensitive_identity_after_prelock(db):
    existing = MaintenanceProject(
        project_id=str(uuid4()),
        project_code="CASE-LOCK",
        display_name="case-lock",
        lifecycle_status="missing",
        is_active=True,
        version=1,
    )
    db.add(existing)
    db.commit()
    row = SimpleNamespace(
        project_name="case-lock",
        project_name_raw="case-lock",
        project_period_from=None,
        project_period_to=None,
        manager=None,
        business_type=None,
        cmo=None,
        salesperson_raw=None,
    )
    ledger._lock_target_projects(db, [row])
    summary = {"projects_created": 0, "projects_updated": 0}

    resolved, changed = ledger._upsert_project(
        db,
        row,
        operated_by="合成管理员",
        summary=summary,
        today=date(2026, 8, 26),
        ledger_batch_id="unused-because-period-is-empty",
    )

    assert resolved.project_id == existing.project_id
    assert changed is False
    assert summary["projects_created"] == 0
    assert len(db.scalars(select(MaintenanceProject)).all()) == 1


def test_nonempty_ledger_salesperson_releases_manual_override_even_when_same(db):
    project = MaintenanceProject(
        project_id=str(uuid4()),
        project_code="LEDGER-SALESPERSON-OVERRIDE",
        display_name="LEDGER-SALESPERSON-OVERRIDE",
        salesperson="台账确认销售",
        salesperson_override_active=True,
        lifecycle_status="missing",
        is_active=True,
        version=3,
    )
    db.add(project)
    db.commit()
    row = SimpleNamespace(
        project_name=project.display_name,
        project_name_raw=project.display_name,
        project_period_from=None,
        project_period_to=None,
        manager=None,
        business_type=None,
        cmo=None,
        salesperson_raw="台账确认销售",
    )
    ledger._lock_target_projects(db, [row])
    summary = {"projects_created": 0, "projects_updated": 0}

    resolved, changed = ledger._upsert_project(
        db,
        row,
        operated_by="合成管理员",
        summary=summary,
        today=date(2026, 8, 31),
        ledger_batch_id="ledger-salesperson-override-reset",
    )
    db.flush()

    assert resolved.salesperson == "台账确认销售"
    assert resolved.salesperson_override_active is False
    assert changed is True
    assert resolved.version == 4
    assert summary["projects_updated"] == 1
    audit = db.scalar(
        select(MaintenanceProjectAuditLog).where(
            MaintenanceProjectAuditLog.project_id == project.project_id,
            MaintenanceProjectAuditLog.action == "ledger_update",
        )
    )
    assert audit.before_json["salesperson_override_active"] is True
    assert audit.after_json["salesperson_override_active"] is False


def test_ledger_override_reset_bumps_project_and_workbook_once(db):
    parsed = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(),
        "维保台账.xlsx",
    )
    first_batch = ledger.store_preview(
        db,
        parsed,
        "合成管理员",
        idempotency_key="ledger-salesperson-reset-first",
    )
    ledger.apply_batch(db, first_batch, "合成管理员")
    project = db.scalar(select(MaintenanceProject).order_by(MaintenanceProject.project_id))
    assert project is not None
    assert project.salesperson

    manually_confirmed = project_catalog.update_project(
        db,
        project_id=project.project_id,
        version=project.version,
        updates={"salesperson": project.salesperson},
        reason="人工确认销售人员",
        operated_by="实名管理员",
    )
    db.commit()
    assert manually_confirmed is not None
    assert manually_confirmed["salesperson_override_active"] is True
    state = db.get(MaintenanceProjectWorkbookState, project.project_id)
    assert state is not None
    version_before_ledger = manually_confirmed["version"]
    revision_before_ledger = state.revision

    parsed_again = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(),
        "维保台账.xlsx",
    )
    second_batch = ledger.store_preview(
        db,
        parsed_again,
        "合成管理员",
        idempotency_key="ledger-salesperson-reset-second",
    )
    summary = ledger.apply_batch(db, second_batch, "合成管理员")

    db.refresh(project)
    db.refresh(state)
    assert project.salesperson_override_active is False
    assert project.version == version_before_ledger + 1
    assert state.revision == revision_before_ledger + 1
    assert summary["projects_updated"] == 1


def test_ledger_reimport_preserves_workbook_contract_total_override(db):
    """项目总表明确修改后的含税额，后续同一台账导入不能静默覆盖。"""
    parsed = ledger.parse_ledger_workbook(_old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(
        db, parsed, "合成管理员", idempotency_key="ledger-preserve-manual-1")
    ledger.apply_batch(db, batch_id, "合成管理员")
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.contract_no == "XSDD-20260731-0086",
    ))
    project_operations_service.update_contract(
        db,
        project_contract_id=contract.project_contract_id,
        version=contract.version,
        updates={
            "contract_amount": Decimal("60000.00"),
            "source": "project_master_workbook",
        },
        reason="项目总表人工确认含税合同总额",
        operated_by="实名财务",
    )
    db.commit()

    parsed_again = ledger.parse_ledger_workbook(
        _old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id_again = ledger.store_preview(
        db, parsed_again, "合成管理员", idempotency_key="ledger-preserve-manual-2")
    ledger.apply_batch(db, batch_id_again, "合成管理员")
    db.expire_all()
    preserved = db.get(MaintenanceProjectContract, contract.project_contract_id)
    assert preserved.amount_inc_tax == Decimal("60000.00")
    assert preserved.source == "project_master_workbook"
    assert preserved.version == 2


def test_apply_rejects_duplicate_apply(db):
    parsed = ledger.parse_ledger_workbook(_old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key="ledger-test-key-0001")
    ledger.apply_batch(db, batch_id, "合成管理员")
    try:
        ledger.apply_batch(db, batch_id, "合成管理员")
        raise AssertionError("重复 apply 应抛出 LedgerBatchError")
    except ledger.LedgerBatchError:
        pass


def test_apply_skips_missing_period(db):
    wb = Workbook()
    ws = wb.active
    ws.title = "维保项目清单"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款", "验收材料",
         "验收材料是否完成及上传附件", "巡检时间", "巡检是否完成及上传附件"]
    )
    ws.append(
        ["XSDD-20260731-0099", "", "李呈辉", "整体维保",
         "无期限项目", "", "", "", "", 1000, 0, 1000, "", "", "", ""]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    parsed = ledger.parse_ledger_workbook(buffer.getvalue(), "维保台账.xlsx")
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key="ledger-test-key-0001")
    # 缺期限行 → 关键异常 → 整批失败关闭：project/contract/milestone 零写入
    with pytest.raises(ledger.LedgerBatchError):
        ledger.apply_batch(db, batch_id, "合成管理员")
    batch = db.get(MaintenanceLedgerImportBatch, batch_id)
    assert batch.status == "failed"
    assert db.execute(select(MaintenanceProject)).scalars().all() == []


def test_apply_reconcile_mismatch_fail_closed(db):
    """销售单存在且台账含税额 ≠ 未税×(1+税率) → 整批失败关闭、零写入。"""
    from app.models.system import SysImportBatch

    import_batch = SysImportBatch(
        filename="s.xlsx", file_type="sales", file_hash="h", status="success"
    )
    db.add(import_batch)
    db.flush()
    db.add(
        FSalesOrder(
            raw_order_id="r-rec",
            order_no="XSDD-20260731-0086",
            order_date=date(2026, 7, 31),
            salesperson="李呈辉",
            business_type="整体维保",
            warehouse="北京成品仓",
            amount_ex_tax=39607.08,
            tax_rate=0.13,
            data_status="已生效",
            import_batch_id=import_batch.id,
        )
    )
    db.commit()
    wb = Workbook()
    ws = wb.active
    ws.title = "维保项目清单"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额(含税)", "已收尾款", "待收尾款",
         "验收材料", "验收材料是否完成及上传附件", "巡检时间", "巡检是否完成及上传附件"]
    )
    # 44,756 与销售未税对不上（应为 44,756.00；此处写 50,000）
    ws.append(
        ["XSDD-20260731-0086", "2026-07-31", "李呈辉", "整体维保",
         "阿里专有云20260608-20291205", "2026-06-08", "2029-12-05", "廖晓娟", "任鑫明",
         50000, 0, 50000, "", "", "", ""]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    parsed = ledger.parse_ledger_workbook(buffer.getvalue(), "台账.xlsx")
    batch_id = ledger.store_preview(
        db, parsed, "合成管理员", idempotency_key="ledger-test-key-reconcile"
    )
    with pytest.raises(ledger.LedgerBatchError):
        ledger.apply_batch(db, batch_id, "合成管理员")
    batch = db.get(MaintenanceLedgerImportBatch, batch_id)
    assert batch.status == "failed"
    assert "reconcile_failures" in (batch.report_json or {})
    # 零 canonical 写入
    assert db.execute(select(MaintenanceProject)).scalars().all() == []
    assert db.execute(select(MaintenanceProjectContract)).scalars().all() == []


def test_apply_does_not_invent_tax_rate_when_sales_tax_is_unknown(db):
    """销售税率 NULL 时没有证明链，台账含税事实可导入且不拿 13% 猜测拒绝。"""
    from app.models.system import SysImportBatch

    import_batch = SysImportBatch(
        filename="s-null-tax.xlsx", file_type="sales", file_hash="h-null-tax",
        status="success",
    )
    db.add(import_batch)
    db.flush()
    db.add(FSalesOrder(
        raw_order_id="r-null-tax",
        order_no="XSDD-20260731-0086",
        order_date=date(2026, 7, 31),
        salesperson="李呈辉",
        business_type="整体维保",
        warehouse="北京成品仓",
        amount_ex_tax=39607.08,
        tax_rate=None,
        data_status="已生效",
        import_batch_id=import_batch.id,
    ))
    db.commit()

    parsed = ledger.parse_ledger_workbook(_old_ledger_workbook_bytes(), "维保台账.xlsx")
    batch_id = ledger.store_preview(
        db, parsed, "合成管理员", idempotency_key="ledger-null-tax-rate")
    summary = ledger.apply_batch(db, batch_id, "合成管理员")

    assert summary["contracts_created"] == 2
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.contract_no == "XSDD-20260731-0086",
    ))
    assert contract.amount_inc_tax == Decimal("44756.00")
    assert contract.contract_amount == Decimal("39607.08")


def test_apply_orphan_plan_row_fail_closed(db):
    """回款计划孤儿行（无对应合同行）→ 整批拒绝，不静默丢弃（round-4 Blocker 7）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "01_项目与合同"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款", "验收材料",
         "验收材料是否完成及上传附件", "验收附件", "巡检时间", "巡检是否完成及上传附件"]
    )
    ws.append(
        ["XSDD-20260731-0086", "2026-07-31", "李呈辉", "整体维保",
         "阿里专有云20260608-20291205", "2026-06-08", "2029-12-05", "廖晓娟", "任鑫明",
         44756, 0, 44756, "服务总结报告", "否", "", "2026-10", "否"]
    )
    plan = wb.create_sheet("02_回款计划")
    plan.append(["订单编号", "计划期次", "计划回款时间", "计划回款金额"])
    plan.append(["XSDD-20260731-0086", 1, "2026-10", 2986.57])
    plan.append(["XSDD-99999999-9999", 1, "2026-11", 5000])  # 孤儿：批次内无此合同
    buffer = io.BytesIO()
    wb.save(buffer)

    parsed = ledger.parse_ledger_workbook(buffer.getvalue(), "台账.xlsx")
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key="ledger-test-key-orphan")
    with pytest.raises(ledger.LedgerBatchError):
        ledger.apply_batch(db, batch_id, "合成管理员")
    batch = db.get(MaintenanceLedgerImportBatch, batch_id)
    assert batch.status == "failed"
    assert "孤儿" in batch.report_json["rejection_reason"]
    # 零 canonical 写入
    assert db.execute(select(MaintenanceProject)).scalars().first() is None
    assert db.execute(select(MaintenanceCollectionMilestone)).scalars().first() is None


# ---- 项目名称周期解析（REQUIREMENTS #50：周期从项目名提取，名称仅兜底源）----


@pytest.mark.parametrize(
    ("name", "expected"),
    [
        # 生产真实命名式样：客户名+8位起-8位止+服务商+业务类型
        ("上海图书馆20250610-20260609酷易信息备件外包", (date(2025, 6, 10), date(2026, 6, 9))),
        # 日期段后带空格
        ("韵达货运20240712-20251015 上海致腾 整体维保", (date(2024, 7, 12), date(2025, 10, 15))),
        # 跨多年
        ("北京华品博睿网络20221101-20280114北京天奕浩博整体维保", (date(2022, 11, 1), date(2028, 1, 14))),
        # 连字符后带空格（生产实例）
        ("北京移动网运中心20250301- 20271231云和恩墨整体外包", (date(2025, 3, 1), date(2027, 12, 31))),
        # 波浪号连接（生产实例）
        ("江西IT服务器20240301~20250228昆仑联通整体维保", (date(2024, 3, 1), date(2025, 2, 28))),
        # 6 位年月段：起月首日、止月末日（生产实例）
        ("云南电网202505-202705唐纳整体维保", (date(2025, 5, 1), date(2027, 5, 31))),
        # 无日期段（生产确有此类）
        ("广东省教育考试院Oracle数据库小型机负载均衡设备及存储系统运维", (None, None)),
        # 名称笔误不救：止日期只有 7 位
        ("黄山九章云智20260715-2070714客户直签整体维保", (None, None)),
        # 名称笔误不救：双连字符拆散日期
        ("鼎甲20251016-2026-1016服务器整体维保", (None, None)),
        # 年份段（2024-2025）不是日期段，不得误吃
        ("广州分公司2024-2025年东涌机房政务云设备维保项目", (None, None)),
        # 非法日期：13 月
        ("某客户20241301-20251231某服务商整体维保", (None, None)),
        # 起止倒置
        ("某客户20260101-20250101某服务商整体维保", (None, None)),
        ("", (None, None)),
        (None, (None, None)),
    ],
)
def test_period_from_display_name(name, expected):
    assert ledger._period_from_display_name(name) == expected


def test_resolve_lifecycle_ledger_period_wins_over_name():
    # 台账周期是权威源：即使名称写着已结束的旧周期，也按台账判 ongoing
    today = date(2026, 8, 17)
    status = ledger._resolve_lifecycle(
        date(2026, 1, 1), date(2026, 12, 31),
        "某客户20200101-20201231某服务商整体维保", today,
    )
    assert status == "ongoing"


def test_resolve_lifecycle_falls_back_to_name_when_period_missing():
    today = date(2026, 8, 17)
    assert ledger._resolve_lifecycle(
        None, None, "某客户20250610-20270609某服务商备件外包", today
    ) == "ongoing"
    assert ledger._resolve_lifecycle(
        None, None, "某客户20240101-20241231某服务商整体维保", today
    ) == "ended"
    # 未来周期：名称解析出的是**完整**起止，只是还没开始 → ongoing
    # （2026-09-03 口径修正：missing 只留给期限数据不完整的情形，
    #  未开始的项目期限是齐的，不该挂「期限缺失」假标签）
    assert ledger._resolve_lifecycle(
        None, None, "某客户20270101-20271231某服务商整体维保", today
    ) == "ongoing"
    # 名称也解析不出 → missing
    assert ledger._resolve_lifecycle(None, None, "无期限项目", today) == "missing"


# ---- 台账 writer OCC：锁序 + 工作簿版本 bump 语义（PostgreSQL 真实锁）----


def _ledger_workbook_with_rows(rows: list[list]) -> bytes:
    """旧结构工作簿：固定 16 列 + 一组横向「回款时间1/回款金额」。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "维保项目清单"
    ws.append(
        ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期",
         "维保终止日期", "CMO", "项目经理", "订单金额", "已收尾款", "待收尾款", "验收材料",
         "验收材料是否完成及上传附件", "巡检时间", "巡检是否完成及上传附件",
         "回款时间1", "回款金额"]
    )
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _occ_row(
    order_no: str,
    project_name: str,
    amount,
    *,
    manager: str = "任鑫明",
    plan_amount=1000,
) -> list:
    return [
        order_no, "2026-07-31", "李呈辉", "整体维保", project_name,
        "2026-06-08", "2029-12-05", "廖晓娟", manager, amount, 0, amount,
        "", "", "", "", "2026年10月", plan_amount,
    ]


def _apply_workbook(db, data: bytes, key: str) -> dict:
    parsed = ledger.parse_ledger_workbook(data, "维保台账.xlsx")
    batch_id = ledger.store_preview(db, parsed, "合成管理员", idempotency_key=key)
    return ledger.apply_batch(db, batch_id, "合成管理员")


def _workbook_revision(db, project_id: str) -> int:
    from app.models.maintenance_project_operations import (
        MaintenanceProjectWorkbookState,
    )

    state = db.get(MaintenanceProjectWorkbookState, project_id)
    return state.revision if state is not None else 0


def _project_by_code(db, code: str) -> MaintenanceProject:
    return db.execute(
        select(MaintenanceProject).where(MaintenanceProject.project_code == code)
    ).scalar_one()


def test_apply_bumps_workbook_revision_once_for_multi_contract_batch(db):
    """一个项目多合同同事务：首建 +1；两合同都改金额的第二批仍只 +1。"""
    project_name = "多合同项目20260608-20291205"
    first = _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-1001", project_name, 10000),
            _occ_row("XSDD-20260731-1002", project_name, 20000),
        ]),
        "occ-multi-1",
    )
    assert first["contracts_created"] == 2
    project = _project_by_code(db, project_name)
    assert _workbook_revision(db, project.project_id) == 1

    second = _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-1001", project_name, 11000),
            _occ_row("XSDD-20260731-1002", project_name, 21000),
        ]),
        "occ-multi-2",
    )
    assert second["contracts_updated"] == 2
    db.expire_all()
    assert _workbook_revision(db, project.project_id) == 2


def test_apply_bumps_workbook_revision_for_project_field_change(db):
    """仅项目字段（项目经理）真实变化也 bump；合同/节点不动。"""
    project_name = "字段变更项目20260608-20291205"
    _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-2001", project_name, 10000),
        ]),
        "occ-proj-1",
    )
    project = _project_by_code(db, project_name)
    assert _workbook_revision(db, project.project_id) == 1

    second = _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-2001", project_name, 10000, manager="新经理"),
        ]),
        "occ-proj-2",
    )
    assert second["projects_updated"] == 1
    assert second["contracts_updated"] == 0
    assert second["milestones_updated"] == 0
    db.expire_all()
    assert _workbook_revision(db, project.project_id) == 2


def test_apply_bumps_workbook_revision_for_milestone_only_change(db):
    """仅回款节点计划金额变化也 bump；项目/合同不动。"""
    project_name = "节点变更项目20260608-20291205"
    _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-2002", project_name, 10000, plan_amount=1000),
        ]),
        "occ-mile-1",
    )
    project = _project_by_code(db, project_name)
    assert _workbook_revision(db, project.project_id) == 1

    second = _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-2002", project_name, 10000, plan_amount=1500),
        ]),
        "occ-mile-2",
    )
    assert second["milestones_updated"] == 1
    assert second["contracts_updated"] == 0
    assert second["projects_updated"] == 0
    db.expire_all()
    assert _workbook_revision(db, project.project_id) == 2


def test_apply_noop_second_batch_keeps_workbook_revision(db):
    """同值重放：零语义变化 → +0，且 no-op 不再误计 milestones_updated。"""
    project_name = "幂等项目20260608-20291205"
    data = _ledger_workbook_with_rows([
        _occ_row("XSDD-20260731-2003", project_name, 10000),
    ])
    _apply_workbook(db, data, "occ-noop-1")
    project = _project_by_code(db, project_name)
    assert _workbook_revision(db, project.project_id) == 1

    second = _apply_workbook(db, data, "occ-noop-2")
    assert second["projects_created"] == 0
    assert second["projects_updated"] == 0
    assert second["contracts_created"] == 0
    assert second["contracts_updated"] == 0
    assert second["milestones_created"] == 0
    assert second["milestones_updated"] == 0
    db.expire_all()
    assert _workbook_revision(db, project.project_id) == 1


def test_new_project_gets_service_period(db):
    """新建项目同时有 MaintenanceProject.period_* 与 MaintenanceServicePeriod。"""
    from app.models.maintenance_manager import MaintenanceServicePeriod

    project_name = "服务周期项目20260608-20291205"
    _apply_workbook(
        db,
        _ledger_workbook_with_rows([
            _occ_row("XSDD-20260731-2004", project_name, 10000),
        ]),
        "occ-period-1",
    )
    project = _project_by_code(db, project_name)
    assert project.period_from == date(2026, 6, 8)
    assert project.period_to == date(2029, 12, 5)
    period = db.get(MaintenanceServicePeriod, project.project_id)
    assert period is not None
    assert period.service_start == date(2026, 6, 8)
    assert period.service_end == date(2029, 12, 5)
    assert period.completeness_state == "complete"
    assert period.ledger_batch_id is not None


def test_shared_contract_change_is_blocked_and_cards_fail_closed(db):
    """历史同号合同挂在 A/B：治理前拒绝改写，卡片继续 fail-closed。"""
    from app.services.maintenance_boss_board import _card_contracts

    project_a = MaintenanceProject(
        project_id=str(uuid4()),
        project_code="共享项目甲20260608-20291205",
        display_name="共享项目甲20260608-20291205",
        lifecycle_status="ongoing",
        is_active=True,
        version=1,
    )
    project_b = MaintenanceProject(
        project_id=str(uuid4()),
        project_code="共享项目乙20260608-20291205",
        display_name="共享项目乙20260608-20291205",
        lifecycle_status="ongoing",
        is_active=True,
        version=1,
    )
    db.add_all([project_a, project_b])
    db.flush()
    # This fixture represents duplicate historical rows that pre-date the
    # XSDD ownership trigger.  New writes are deliberately unable to create
    # this state, so disable only that user trigger while seeding the legacy
    # conflict and restore it before exercising application behavior.
    db.execute(text(
        "ALTER TABLE maintenance_project_contract DISABLE TRIGGER "
        "trg_maintenance_contract_claim_xsdd"
    ))
    try:
        for project in (project_a, project_b):
            db.add(
                MaintenanceProjectContract(
                    project_contract_id=str(uuid4()),
                    project_id=project.project_id,
                    contract_id="XSDD-20260731-3001",
                    contract_no="XSDD-20260731-3001",
                    contract_amount=None,
                    amount_inc_tax=Decimal("5000.00"),
                    contract_status=None,
                    status_mapping_state="mapped",
                    status_mapping_version="project_manager_xls_v1",
                    included_in_total=True,
                    effective_from=date(2026, 1, 1),
                    effective_to=None,
                    source="project_manager_xls_v1",
                    version=1,
                )
            )
        db.flush()
    finally:
        db.execute(text(
            "ALTER TABLE maintenance_project_contract ENABLE TRIGGER "
            "trg_maintenance_contract_claim_xsdd"
        ))
    db.commit()

    with pytest.raises(ledger.LedgerBatchError, match="历史归并预检"):
        _apply_workbook(
            db,
            _ledger_workbook_with_rows([
                _occ_row(
                    "XSDD-20260731-3001",
                    "共享项目甲20260608-20291205",
                    6000,
                ),
            ]),
            "occ-shared-1",
        )
    db.rollback()
    db.expire_all()
    assert set(db.scalars(select(MaintenanceProjectContract.amount_inc_tax).where(
        MaintenanceProjectContract.contract_no == "XSDD-20260731-3001"
    ))) == {Decimal("5000.00")}

    cards = _card_contracts(db, [project_a.project_id, project_b.project_id])
    for project_id in (project_a.project_id, project_b.project_id):
        assert cards[project_id]["contract_shared"] is True
        # 跨项目冲突 fail-closed：标 incomplete、不返回伪精确金额
        assert cards[project_id]["contract_incomplete"] is True
        assert cards[project_id]["amount_inc_tax"] is None


def test_apply_locks_all_workbook_states_before_first_project_lock(db):
    """真实 SQL 锁序：最后一个 state 锁早于第一个 project 锁（不交错）。"""
    import re

    from sqlalchemy import event

    from app.db import engine

    # 两个目标项目都已存在，预锁路径才会真实发出 state/project 行锁。
    db.add_all([
        MaintenanceProject(
            project_id=str(uuid4()),
            project_code=code,
            display_name=code,
            lifecycle_status="ongoing",
            is_active=True,
            version=1,
        )
        for code in ("阿里专有云20260608-20291205", "正大天晴20260801-20270531因")
    ])
    db.commit()

    statements: list[str] = []

    def observe(_conn, _cursor, statement, _parameters, _context, _executemany):
        statements.append(" ".join(statement.lower().split()))

    event.listen(engine, "before_cursor_execute", observe)
    try:
        _apply_workbook(db, _old_ledger_workbook_bytes(), "occ-lock-order-1")
    finally:
        event.remove(engine, "before_cursor_execute", observe)

    state_locks = [
        index for index, stmt in enumerate(statements)
        if "from maintenance_project_workbook_state" in stmt and "for update" in stmt
    ]
    project_locks = [
        index for index, stmt in enumerate(statements)
        if re.search(r"\bfrom maintenance_project\b", stmt) and "for update" in stmt
    ]
    assert state_locks, "应存在 workbook state 行锁"
    assert project_locks, "应存在 project 行锁"
    assert max(state_locks) < min(project_locks)
