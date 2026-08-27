"""项目工作簿 v3 导出测试（D1）：六 sheet + 隐藏技术 sheet + 颜色契约。"""

import io
from datetime import date
from decimal import Decimal

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app import auth
from app.api import maintenance_project_workbook_v3
from app.auth import hash_password
from app.models.dimensions import DimPart
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder, FProjectExpense
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
    MaintenanceSiteIssue,
    MaintenanceSiteIssueLine,
)
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.models.system import SysImportBatch, SysUser
from app.services import maintenance_front_stock as front_stock
from app.services import maintenance_project_workbook_v3 as workbook_v3


def _seed(db):
    project = MaintenanceProject(
        project_id="wb3-project-1",
        project_code="WB3测试项目",
        display_name="WB3维保项目",
        business_type="整体维保",
        cmo_name="廖晓娟",
        salesperson="李呈辉",
        no_return_default=False,
        lifecycle_status="ongoing",
        is_active=True,
    )
    db.add(project)
    db.flush()
    contract = MaintenanceProjectContract(
        project_contract_id="wb3-pc-1",
        project_id="wb3-project-1",
        contract_id="wb3-contract-1",
        contract_no="XSDD-20260731-0086",
        contract_amount=Decimal("39607.08"),
        amount_inc_tax=Decimal("44756.00"),
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=date(2026, 6, 9),
        effective_to=date(2029, 12, 5),
        source="synthetic-test",
    )
    db.add(contract)
    db.flush()
    import_batch = SysImportBatch(
        filename="w.xlsx", file_type="maintenance", file_hash="h-wb3", status="success"
    )
    db.add(import_batch)
    db.flush()
    order = FMaintenanceOrder(
        raw_order_id="wb3-wbdd-raw-1",
        order_no="WBDD-20260702-0014",
        order_date=date(2026, 7, 2),
        demand_type="补库供货",
        business_type="整体维保",
        project_raw="WB3维保项目",
        project_std="WB3维保项目",
        warehouse="北京成品仓",
        data_status="已生效",
        linked_sales_order_no="XSDD-20260731-0086",
        import_batch_id=import_batch.id,
    )
    db.add(order)
    db.flush()
    part = DimPart(pn_std="02311AYV", description="8G 内存")
    db.add(part)
    db.flush()
    db.add(
        FMaintenanceLine(
            raw_line_id="wb3-line-1",
            order_id=order.id,
            line_no=1,
            part_id=part.id,
            pn_std="02311AYV",
            qty=Decimal("10"),
            import_batch_id=import_batch.id,
        )
    )
    db.add(
        MaintenanceSourceOrderAssignment(
            assignment_id="wb3-assign-1",
            source_order_id="wb3-wbdd-raw-1",
            project_id="wb3-project-1",
            is_active=True,
            created_by="合成归属人",
        )
    )
    db.add(
        FProjectExpense(
            raw_line_id="wb3-fpe-1",
            bxd_no="BXD-20260721-0019",
            line_no=1,
            data_status="已结束",
            expense_date=date(2026, 7, 27),
            person="罗汇康",
            expense_type="维保费用",
            fee_category="外援费用",
            reason="北京2026年6月外援费用",
            linked_sales_order_no="XSDD-20260731-0086",
            amount=Decimal("800"),
            amount_ex_tax=Decimal("707.96"),
            amount_inc_tax=Decimal("800"),
            tax_basis="inc",
            import_batch_id=import_batch.id,
        )
    )
    db.add(
        MaintenanceProjectExpenseAttribution(
            expense_id="bxd:wb3-fpe-1",
            project_id="wb3-project-1",
            project_contract_id="wb3-pc-1",
            raw_expense_line_id="wb3-fpe-1",
            expense_ref="BXD-20260721-0019:1",
            expense_date=date(2026, 7, 27),
            applicant="罗汇康",
            category="外援费用",
            expense_reason="北京2026年6月外援费用",
            amount_ex_tax=Decimal("707.96"),
            amount_inc_tax=Decimal("799.99"),
            tax_rate_used=Decimal("0.13"),
            raw_status="已结束",
            status_mapping_state="mapped",
            normalized_status="approved",
            status_mapping_version="synthetic-v1",
            ownership_mapping_state="mapped",
            ownership_mapping_version="synthetic-v1",
        )
    )
    db.add(
        MaintenanceCollectionSnapshot(
            collection_id="wb3-snap-1",
            project_id="wb3-project-1",
            project_contract_id="wb3-pc-1",
            report_month=date(2026, 8, 1),
            cumulative_amount=Decimal("2986.57"),
            status="confirmed",
            receipt_reference="HKD-0001",
        )
    )
    # 未来月度快照不得进入导出（round-6 Blocker 7 反例）
    db.add(
        MaintenanceCollectionSnapshot(
            collection_id="wb3-snap-future",
            project_id="wb3-project-1",
            project_contract_id="wb3-pc-1",
            report_month=date(2026, 10, 1),
            cumulative_amount=Decimal("9999"),
            status="confirmed",
        )
    )
    issue = MaintenanceSiteIssue(
        issue_id="wb3-issue-1",
        project_id="wb3-project-1",
        issue_no="LY-20260710-0001",
        issue_date=date(2026, 7, 10),
        raw_status="已确认",
        status_mapping_state="mapped",
        normalized_status="confirmed",
        status_mapping_version="synthetic-map-v1",
        source="direct_api",
        version=1,
    )
    db.add(issue)
    db.flush()
    db.add(
        MaintenanceSiteIssueLine(
            issue_line_id="wb3-issue-line-1",
            issue_id="wb3-issue-1",
            line_no=1,
            part_id=part.id,
            pn="02311AYV",
            quantity=Decimal("2"),
            serial_number="S0M59S5M",
            algorithm_version="synthetic-algo-v1",
        )
    )
    front_stock.apply_movement(
        db,
        project_id="wb3-project-1",
        part_id=part.id,
        kind="shipment_in",
        source_type="f_maintenance_line",
        source_ref="WB3-SHIP-1",
        qty=Decimal("8"),
        warehouse_name="WB3测试项目",
        unit_cost_ex_tax=Decimal("100"),
        unit_cost_inc_tax=Decimal("113"),
        operated_by="合成测试员",
    )
    db.commit()


def test_build_workbook_sheets_and_data(db):
    _seed(db)
    data = workbook_v3.build_project_workbook(db, "wb3-project-1")
    workbook = load_workbook(io.BytesIO(data))
    assert workbook.sheetnames[:7] == [
        "00_使用说明", "01_项目基础信息", "02_概览数据", "03_备件订单",
        "04_报销订单", "05_项目经理回款单", "06_现场领用与返还",
    ]
    assert "98_字典" in workbook.sheetnames and "99_元数据" in workbook.sheetnames
    assert workbook["98_字典"].sheet_state == "hidden"
    assert workbook["99_元数据"].sheet_state == "hidden"

    ws01 = workbook["01_项目基础信息"]
    row = [c.value for c in ws01[2]]
    assert row[0] == "WB3测试项目"
    assert row[9] == "否"  # 项目级不返还默认
    assert row[11] == 1  # 前置库种类数
    assert row[12] == 8.0  # 前置库件数
    assert row[13] == 904.0  # 8 × 113

    ws02 = workbook["02_概览数据"]
    contract_rows = [r for r in ws02.iter_rows(min_row=3, values_only=True) if r[0]]
    assert contract_rows[0][0] == "XSDD-20260731-0086"
    assert contract_rows[0][1] == 44756.0

    ws03 = workbook["03_备件订单"]
    rows = list(ws03.iter_rows(values_only=True))
    assert rows[1][0] == "WBDD-20260702-0014"
    assert rows[1][6] == "02311AYV"
    assert rows[1][8] == 10.0
    assert rows[1][13] is None  # 未税单位成本：可回填列留空

    ws04 = workbook["04_报销订单"]
    rows = list(ws04.iter_rows(values_only=True))
    assert rows[1][0] == "BXD-20260721-0019"
    # 展示 canonical attribution 金额，不信任可漂移的 raw 报销金额。
    assert rows[1][8] == 799.99

    ws05 = workbook["05_项目经理回款单"]
    rows = list(ws05.iter_rows(values_only=True))
    assert rows[1][1] == "XSDD-20260731-0086"
    assert rows[1][2] == "2026-08"
    assert rows[1][3] == 2986.57
    # 未来 2026-10 快照被 as_of 过滤
    assert all(row[2] != "2026-10" for row in rows[1:] if row[2])

    ws06 = workbook["06_现场领用与返还"]
    rows = list(ws06.iter_rows(values_only=True))
    assert rows[1][0] == "LY-20260710-0001"
    assert rows[1][2] == "02311AYV"
    assert rows[1][4] == 2.0


def test_editable_columns_follow_header_fill_contract(db):
    _seed(db)
    data = workbook_v3.build_project_workbook(db, "wb3-project-1")
    workbook = load_workbook(io.BytesIO(data))
    editable_03 = workbook_v3.parse_editable_header_fills(workbook["03_备件订单"])
    assert editable_03 == ["未税单位成本", "变更原因"]
    editable_01 = workbook_v3.parse_editable_header_fills(workbook["01_项目基础信息"])
    assert editable_01 == []  # 全只读
    editable_06 = workbook_v3.parse_editable_header_fills(workbook["06_现场领用与返还"])
    assert editable_06 == ["是否应返还(行级)", "备注"]


def test_v3_overview_does_not_turn_missing_site_cost_into_zero(db):
    _seed(db)

    def metrics():
        data = workbook_v3.build_project_workbook(db, "wb3-project-1")
        sheet = load_workbook(io.BytesIO(data), data_only=True)["02_概览数据"]
        return {
            row[0].value: row[1].value
            for row in sheet.iter_rows(min_col=1, max_col=2)
            if row[0].value
        }

    missing = metrics()
    assert missing["项目已计成本(含税)"] is None
    assert missing["缺失成本行数"] == 1

    line = db.get(MaintenanceSiteIssueLine, "wb3-issue-line-1")
    line.cost_source = "manual"
    line.price_basis = "ex_tax"
    line.unit_cost = Decimal("0")
    line.cost_amount = Decimal("0")
    line.unit_cost_ex_tax = Decimal("0")
    line.unit_cost_inc_tax = Decimal("0")
    line.cost_amount_ex_tax = Decimal("0")
    line.cost_amount_inc_tax = Decimal("0")
    db.commit()
    complete = metrics()
    # 现场合法 0 + 已知 canonical 报销 799.99，不能因 falsy-zero 被抹成缺失。
    assert complete["项目已计成本(含税)"] == 799.99
    assert complete["缺失成本行数"] == 0


def test_v3_overview_without_any_cost_evidence_stays_blank(db):
    _seed(db)
    db.delete(db.get(MaintenanceSiteIssueLine, "wb3-issue-line-1"))
    db.delete(db.get(
        MaintenanceProjectExpenseAttribution, "bxd:wb3-fpe-1"
    ))
    db.commit()

    data = workbook_v3.build_project_workbook(db, "wb3-project-1")
    sheet = load_workbook(io.BytesIO(data), data_only=True)["02_概览数据"]
    metrics = {
        row[0].value: row[1].value
        for row in sheet.iter_rows(min_col=1, max_col=2)
        if row[0].value
    }
    assert metrics["项目已计成本(含税)"] is None
    assert metrics["成本率"] is None
    assert metrics["缺失成本行数"] == 0
    assert "暂无成本事实" in metrics["数据完整性提示"]


def test_v3_shared_current_contract_fails_closed_for_expense_rows(db):
    _seed(db)
    other = MaintenanceProject(
        project_id="wb3-project-shared",
        project_code="WB3-SHARED",
        display_name="WB3共享合同项目",
        lifecycle_status="ongoing",
        is_active=True,
    )
    db.add(other)
    db.flush()
    db.add(MaintenanceProjectContract(
        project_contract_id="wb3-pc-shared",
        project_id=other.project_id,
        contract_id="wb3-contract-1",
        contract_no="XSDD-20260731-0086",
        amount_inc_tax=Decimal("44756.00"),
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=date(2026, 6, 8),
        effective_to=date(2029, 12, 5),
        source="synthetic-test",
    ))
    line = db.get(MaintenanceSiteIssueLine, "wb3-issue-line-1")
    line.cost_source = "manual"
    line.price_basis = "ex_tax"
    line.unit_cost = Decimal("0")
    line.cost_amount = Decimal("0")
    line.unit_cost_ex_tax = Decimal("0")
    line.unit_cost_inc_tax = Decimal("0")
    line.cost_amount_ex_tax = Decimal("0")
    line.cost_amount_inc_tax = Decimal("0")
    db.commit()

    data = workbook_v3.build_project_workbook(db, "wb3-project-1")
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    expense_rows = list(workbook["04_报销订单"].iter_rows(values_only=True))
    assert all(row[0] != "BXD-20260721-0019" for row in expense_rows[1:])
    metrics = {
        row[0].value: row[1].value
        for row in workbook["02_概览数据"].iter_rows(min_col=1, max_col=2)
        if row[0].value
    }
    assert metrics["项目已计成本(含税)"] == 0


def test_v3_overview_rejects_duplicate_current_contract_relationship(db):
    _seed(db)
    db.add(MaintenanceProjectContract(
        project_contract_id="wb3-pc-duplicate",
        project_id="wb3-project-1",
        contract_id="wb3-contract-1",
        contract_no="XSDD-20260731-0086",
        contract_amount=Decimal("39607.08"),
        amount_inc_tax=Decimal("44756.00"),
        contract_status="active",
        status_mapping_state="mapped",
        status_mapping_version="synthetic-v1",
        included_in_total=True,
        effective_from=date(2026, 6, 8),
        effective_to=date(2029, 12, 5),
        source="synthetic-test",
    ))
    db.commit()

    data = workbook_v3.build_project_workbook(db, "wb3-project-1")
    sheet = load_workbook(io.BytesIO(data), data_only=True)["02_概览数据"]
    metrics = {
        row[0].value: row[1].value
        for row in sheet.iter_rows(min_col=1, max_col=2)
        if row[0].value
    }
    assert metrics["合同总额(含税)"] is None
    assert "重复" in metrics["数据完整性提示"]


def test_workbook_v3_api_export_and_404(db):
    _seed(db)
    db.add(
        SysUser(
            username="wb3_api_admin",
            role="admin",
            display_name="工作簿导出管理员",
            password_hash=hash_password("synthetic-password-123"),
        )
    )
    db.commit()
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_project_workbook_v3.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "wb3_api_admin", "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    ok = client.get("/api/maintenance/projects/stable/wb3-project-1/workbook-v3.xlsx")
    assert ok.status_code == 200, ok.text
    assert ok.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    missing = client.get(
        "/api/maintenance/projects/stable/no-such-project/workbook-v3.xlsx"
    )
    assert missing.status_code == 404


def test_workbook_v3_api_rejects_without_data_groups(db):
    """缺 data_purchase_cost 或 data_profit：导出 403（round-6 Blocker 11 负向门）。"""
    from app import permissions as _perms
    from app.auth import hash_password
    from app.models.maintenance_project import MaintenanceProjectUserAssignment
    from datetime import datetime, timezone

    _seed(db)
    graph = _perms.effective("sales", None)
    graph.update(
        {
            "page_maintenance": True,
            "data_purchase_cost": True,
            "data_profit": False,
        }
    )
    user = SysUser(
        username="wb3_no_profit",
        role="sales",
        display_name="无利润权限销售",
        password_hash=hash_password("synthetic-password-123"),
        permissions=graph,
    )
    db.add(user)
    db.flush()
    db.add(
        MaintenanceProjectUserAssignment(
            assignment_id="wb3-noprofit-assign",
            project_id="wb3-project-1",
            responsibility_type="primary_manager",
            user_id=user.id,
            assigned_at=datetime.now(timezone.utc),
            assigned_by="synthetic-admin",
            assignment_reason="合成负责人映射",
        )
    )
    db.commit()
    app = FastAPI()
    app.include_router(auth.router, prefix="/api")
    app.include_router(maintenance_project_workbook_v3.router, prefix="/api")
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "wb3_no_profit", "password": "synthetic-password-123"},
    )
    assert login.status_code == 200, login.text
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    denied = client.get(
        "/api/maintenance/projects/stable/wb3-project-1/workbook-v3.xlsx"
    )
    assert denied.status_code == 403
