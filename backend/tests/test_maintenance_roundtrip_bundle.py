"""按合同拆分的可回填工作簿 ZIP 下载契约。"""

from __future__ import annotations

import io
import zipfile

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import config, permissions
from app.api import maintenance as maintenance_api
from app.auth import hash_password
from app.main import app
from app.models.system import SysUser
from app.security import UserContext
from app.services import maintenance_roundtrip
from tests.test_maintenance_export_headers import _admin_client, _cost_blind_maintenance_client
from tests.test_maintenance_roundtrip import _seed_contract


def _metadata(payload: bytes) -> dict[str, str]:
    workbook = load_workbook(io.BytesIO(payload), data_only=False)
    try:
        sheet = workbook["99_元数据"]
        return {
            str(key): "" if value is None else str(value)
            for key, value in sheet.iter_rows(min_row=2, values_only=True)
            if key is not None
        }
    finally:
        workbook.close()


def _assert_contract_is_literal_text(payload: bytes, contract: str) -> None:
    workbook = load_workbook(io.BytesIO(payload), data_only=False)
    try:
        for sheet_name, header in (
            ("01_项目", "合同号"),
            ("02_维保订单", "合同号"),
        ):
            sheet = workbook[sheet_name]
            columns = {
                str(cell.value): cell.column
                for cell in sheet[1]
                if cell.value is not None
            }
            cell = sheet.cell(row=2, column=columns[header])
            assert cell.value == contract
            assert cell.data_type == "s"

        revision_cell = workbook["99_合同版本"].cell(row=2, column=1)
        assert revision_cell.value == contract
        assert revision_cell.data_type == "s"

        metadata_sheet = workbook["99_元数据"]
        metadata_row = next(
            row
            for row in range(2, metadata_sheet.max_row + 1)
            if metadata_sheet.cell(row=row, column=1).value == "contract_scope"
        )
        metadata_cell = metadata_sheet.cell(row=metadata_row, column=2)
        assert metadata_cell.value == contract
        assert metadata_cell.data_type == "s"
    finally:
        workbook.close()


def _customer_blind_maintenance_client(db) -> TestClient:
    base = permissions.effective("boss", None)
    overrides = {"data_customer": False}
    effective = permissions.effective_from_snapshot(base, overrides)
    db.add(SysUser(
        username="maintenance-roundtrip-customer-blind",
        role="boss",
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="boss",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=effective,
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={
            "username": "maintenance-roundtrip-customer-blind",
            "password": "pw123456",
        },
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def test_roundtrip_template_exports_require_customer_info_permission(db):
    _seed_contract(db, suffix="customer-permission", contract="XSDD-CUSTOMER-PERM")
    client = _customer_blind_maintenance_client(db)

    single = client.get(
        "/api/maintenance/roundtrip-template",
        params={"contract": "XSDD-CUSTOMER-PERM"},
    )
    bundle = client.get("/api/maintenance/roundtrip-templates")

    assert single.status_code == 403
    assert bundle.status_code == 403
    assert "客户信息" in single.json()["detail"]
    assert "客户信息" in bundle.json()["detail"]


def test_roundtrip_customer_permission_does_not_bypass_when_rbac_is_disabled(
    monkeypatch,
):
    monkeypatch.setattr(config, "ENABLE_RBAC", False)
    context = UserContext(
        user_id="legacy-user",
        role="readonly",
        permissions=None,
        is_authenticated=True,
    )

    with pytest.raises(Exception) as caught:
        maintenance_api._require_roundtrip_customer_permission(context)

    assert getattr(caught.value, "status_code", None) == 403
    assert "客户信息" in str(getattr(caught.value, "detail", ""))


def test_roundtrip_bundle_endpoint_splits_contracts_into_independently_signed_workbooks(db):
    contracts = ["XSDD/中文", r"XSDD\中文"]
    for index, contract in enumerate(contracts):
        _seed_contract(db, suffix=f"bundle-{index}", contract=contract)
    client = _admin_client(db)

    response = client.get("/api/maintenance/roundtrip-templates")

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("application/zip")
    assert "filename*=UTF-8''" in response.headers["content-disposition"]
    assert response.headers["cache-control"] == "no-store"
    assert response.headers["x-content-type-options"] == "nosniff"

    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        assert archive.testzip() is None
        names = archive.namelist()
        assert len(names) == 2
        assert len(set(names)) == 2
        assert all(name.startswith("维保回填模板/") and name.endswith(".xlsx") for name in names)
        assert all(".." not in name and "\\" not in name for name in names)
        seen_contracts = set()
        members: list[tuple[str, bytes, str]] = []
        for name in names:
            payload = archive.read(name)
            workbook = load_workbook(io.BytesIO(payload), data_only=False)
            try:
                assert workbook.sheetnames == list(maintenance_roundtrip.SHEET_NAMES)
                assert set(workbook["02_维保订单"].tables) == {"tbl_orders_v1"}
                assert set(workbook["03_订单明细"].tables) == {"tbl_order_lines_v1"}
                order_headers = {
                    str(cell.value): cell.column
                    for cell in workbook["02_维保订单"][1]
                    if cell.value is not None
                }
                line_headers = {
                    str(cell.value): cell.column
                    for cell in workbook["03_订单明细"][1]
                    if cell.value is not None
                }
                assert len(str(workbook["02_维保订单"].cell(
                    row=2,
                    column=order_headers["__row_token"],
                ).value)) == 64
                assert len(str(workbook["03_订单明细"].cell(
                    row=2,
                    column=line_headers["__row_token"],
                ).value)) == 64
                revision_rows = [
                    (str(contract), int(revision))
                    for contract, revision in workbook["99_合同版本"].iter_rows(
                        min_row=2,
                        values_only=True,
                    )
                    if contract is not None
                ]
                assert len(revision_rows) == 1
                assert revision_rows[0][1] == 0
            finally:
                workbook.close()
            metadata = _metadata(payload)
            signature = metadata.pop("metadata_hmac")
            assert signature == maintenance_roundtrip._metadata_hmac(metadata)
            seen_contracts.add(metadata["contract_scope"])
            assert revision_rows[0][0] == metadata["contract_scope"]
            members.append((name.rsplit("/", 1)[-1], payload, metadata["contract_scope"]))
        assert seen_contracts == set(contracts)

    for filename, payload, contract in members:
        imported = client.post(
            "/api/maintenance/roundtrip-import",
            files={
                "file": (
                    filename,
                    payload,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        assert imported.status_code == 200, imported.text
        assert imported.json()["contracts"] == [contract]
        assert imported.json()["changed_rows"] == 0


def test_single_roundtrip_formula_like_contract_is_literal_and_imports_unchanged(db):
    contract = '=HYPERLINK("https://127.0.0.1","x")'
    _seed_contract(db, suffix="formula-single", contract=contract)
    client = _admin_client(db)

    response = client.get(
        "/api/maintenance/roundtrip-template",
        params={"contract": contract},
    )

    assert response.status_code == 200, response.text
    _assert_contract_is_literal_text(response.content, contract)
    metadata = _metadata(response.content)
    signature = metadata.pop("metadata_hmac")
    assert signature == maintenance_roundtrip._metadata_hmac(metadata)
    imported = client.post(
        "/api/maintenance/roundtrip-import",
        files={
            "file": (
                "formula-contract.xlsx",
                response.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["contracts"] == [contract]
    assert imported.json()["changed_rows"] == 0


def test_batch_roundtrip_formula_like_contract_is_literal_and_imports_unchanged(db):
    contract = '=HYPERLINK("https://127.0.0.1","batch")'
    _seed_contract(db, suffix="formula-batch", contract=contract)
    client = _admin_client(db)

    response = client.get("/api/maintenance/roundtrip-templates")

    assert response.status_code == 200, response.text
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        assert archive.testzip() is None
        assert len(archive.namelist()) == 1
        payload = archive.read(archive.namelist()[0])
    _assert_contract_is_literal_text(payload, contract)
    metadata = _metadata(payload)
    signature = metadata.pop("metadata_hmac")
    assert signature == maintenance_roundtrip._metadata_hmac(metadata)
    imported = client.post(
        "/api/maintenance/roundtrip-import",
        files={
            "file": (
                "formula-contract-batch.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["contracts"] == [contract]
    assert imported.json()["changed_rows"] == 0


def test_roundtrip_bundle_rejects_contract_count_at_500_boundary_before_render(
    db,
    monkeypatch,
):
    for index in range(2):
        _seed_contract(db, suffix=f"count-{index}", contract=f"XSDD-COUNT-{index}")
    monkeypatch.setattr(maintenance_roundtrip, "MAX_ROUNDTRIP_BUNDLE_CONTRACTS", 2)
    rendered = False

    def forbidden_render(*_args, **_kwargs):
        nonlocal rendered
        rendered = True
        raise AssertionError("count preflight must happen before workbook rendering")

    monkeypatch.setattr(maintenance_roundtrip, "build_roundtrip_template", forbidden_render)

    with pytest.raises(maintenance_roundtrip.RoundtripWorkbookError) as caught:
        maintenance_roundtrip.build_roundtrip_template_bundle(
            db,
            exported_by="tester",
        )

    assert caught.value.status_code == 413
    assert "必须少于 2 个" in str(caught.value)
    assert rendered is False


def test_roundtrip_bundle_contract_discovery_limits_distinct_query_to_501(
    db,
    monkeypatch,
):
    _seed_contract(db, suffix="query-limit", contract="XSDD-QUERY-LIMIT")
    original_scalars = db.scalars
    captured_limits: list[int | None] = []

    def recording_scalars(statement, *args, **kwargs):
        limit_clause = getattr(statement, "_limit_clause", None)
        captured_limits.append(
            None if limit_clause is None else int(limit_clause.value)
        )
        return original_scalars(statement, *args, **kwargs)

    monkeypatch.setattr(db, "scalars", recording_scalars)

    contracts = maintenance_roundtrip._roundtrip_bundle_contracts(
        db,
        date_from=None,
        date_to=None,
    )

    assert contracts == ["XSDD-QUERY-LIMIT"]
    assert captured_limits == [
        maintenance_roundtrip.MAX_ROUNDTRIP_BUNDLE_CONTRACTS + 1
    ]


def test_roundtrip_bundle_rejects_total_uncompressed_size_at_limit_without_response(
    db,
    monkeypatch,
):
    _seed_contract(db, suffix="size", contract="XSDD-SIZE")
    monkeypatch.setattr(
        maintenance_roundtrip,
        "MAX_ROUNDTRIP_BUNDLE_UNCOMPRESSED_BYTES",
        1,
    )

    with pytest.raises(maintenance_roundtrip.RoundtripWorkbookError) as caught:
        maintenance_roundtrip.build_roundtrip_template_bundle(
            db,
            exported_by="tester",
        )

    assert caught.value.status_code == 413
    assert "总大小必须小于 512 MiB" in str(caught.value)


def test_roundtrip_bundle_inherits_per_workbook_dynamic_text_preflight(
    db,
    monkeypatch,
):
    _seed_contract(db, suffix="text-budget", contract="XSDD-TEXT-BUDGET")
    monkeypatch.setattr(
        maintenance_roundtrip,
        "MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK",
        1,
    )
    rendered = []
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_instructions_sheet",
        lambda *_args, **_kwargs: rendered.append("rendered"),
    )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="动态文本超过 64 MiB",
    ) as caught:
        maintenance_roundtrip.build_roundtrip_template_bundle(
            db,
            exported_by="tester",
        )

    assert caught.value.status_code == 413
    assert rendered == []


def test_roundtrip_bundle_zero_data_and_permissions_fail_closed(db):
    admin = _admin_client(db)

    empty = admin.get("/api/maintenance/roundtrip-templates")

    assert empty.status_code == 422
    assert empty.json()["detail"] == "所选范围内没有可导出的已关联合同维保数据"

    _seed_contract(db, suffix="permission", contract="XSDD-PERMISSION")
    cost_blind = _cost_blind_maintenance_client(db)

    forbidden = cost_blind.get("/api/maintenance/roundtrip-templates")

    assert forbidden.status_code == 403
    assert "无成本及利润查看权限" in forbidden.json()["detail"]


def test_roundtrip_bundle_requires_complete_valid_date_pair(db):
    _seed_contract(db, suffix="date", contract="XSDD-DATE")
    client = _admin_client(db)

    missing_to = client.get(
        "/api/maintenance/roundtrip-templates",
        params={"date_from": "2026-07-01"},
    )
    reversed_range = client.get(
        "/api/maintenance/roundtrip-templates",
        params={"date_from": "2026-07-31", "date_to": "2026-07-01"},
    )

    assert missing_to.status_code == 422
    assert missing_to.json()["detail"] == "date_from 与 date_to 必须同时提供"
    assert reversed_range.status_code == 422
    assert reversed_range.json()["detail"] == "date_from 不能晚于 date_to"


def test_roundtrip_template_busy_response_includes_retry_after(db, monkeypatch):
    client = _admin_client(db)

    def busy(*_args, **_kwargs):
        raise maintenance_roundtrip.RoundtripWorkbookError(
            "已有固定回填工作簿正在生成，请稍后重试",
            status_code=429,
        )

    monkeypatch.setattr(
        maintenance_roundtrip,
        "build_roundtrip_template",
        busy,
    )

    response = client.get("/api/maintenance/roundtrip-template")

    assert response.status_code == 429
    assert response.headers["retry-after"] == "5"
    assert "正在生成" in response.json()["detail"]


def test_roundtrip_bundle_busy_response_includes_retry_after(db, monkeypatch):
    client = _admin_client(db)

    def busy(*_args, **_kwargs):
        raise maintenance_roundtrip.RoundtripWorkbookError(
            "已有固定回填工作簿正在生成，请稍后重试",
            status_code=429,
        )

    monkeypatch.setattr(
        maintenance_roundtrip,
        "build_roundtrip_template_bundle",
        busy,
    )

    response = client.get("/api/maintenance/roundtrip-templates")

    assert response.status_code == 429
    assert response.headers["retry-after"] == "5"
    assert "正在生成" in response.json()["detail"]
