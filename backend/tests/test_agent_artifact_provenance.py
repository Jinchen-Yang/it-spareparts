"""Fail-closed Artifact provenance and live scope re-authorization (#220)."""

from __future__ import annotations

import copy
import hashlib
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import DBAPIError

from app import permissions, security
from app.auth import hash_password
from app.main import app
from app.models.agent_artifact import AgentArtifact, AgentArtifactAudit
from app.models.system import SysUser
from app.services import (
    agent_artifact_provenance,
    agent_artifact_reconcile,
    agent_files,
    agent_integrity,
)
from tests.artifact_test_support import force_artifact_state


def test_trusted_evidence_repr_never_discloses_opaque_provenance() -> None:
    owner = "repr-secret-owner"
    source_id = "repr-secret-source-id"
    mac = "repr-secret-mac"
    evidence = agent_artifact_provenance.TrustedEvidence(
        owner_sub=owner,
        content_fingerprint="repr-secret-renderer-fingerprint",
        source_envelopes_json=(
            f'{{"source_artifact_id":"{source_id}","mac":"{mac}"}}',
        ),
        _seal=object(),
    )

    rendered = repr(evidence)

    assert rendered == "TrustedEvidence(envelope_count=1)"
    assert owner not in rendered
    assert source_id not in rendered
    assert mac not in rendered


def _owner(db, username: str = "provenance-owner", role: str = "admin"):
    user = SysUser(
        username=username,
        role=role,
        password_hash=hash_password("pw123456"),
        permissions=permissions.effective(role, None),
    )
    db.add(user)
    db.commit()
    return agent_files.verified_artifact_owner(
        db,
        security.UserContext(
            user_id=username,
            role=role,
            permissions=permissions.runtime_safe(
                permissions.effective_for_user(user)
            ),
            is_authenticated=True,
            authn="sys_user",
            has_stable_subject=True,
            token_version=user.token_version or 0,
        ),
    )


def _workbook_bytes(
    *,
    rows: list[list] | None = None,
    comment: str | None = None,
    hidden: bool = False,
    styled: bool = False,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "申请"
    for row in rows or [["PN", "数量", "备注"]]:
        sheet.append(row)
    if comment is not None:
        from openpyxl.comments import Comment

        sheet["A1"].comment = Comment(comment, "tester")
    if hidden:
        extra = workbook.create_sheet("隐藏规则")
        extra["A1"] = "半年无采购不得补库"
        extra.sheet_state = "hidden"
    if styled:
        from openpyxl.styles import Font

        sheet["A1"].font = Font(bold=True)
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _identity_template_bytes() -> bytes:
    return agent_artifact_provenance.canonical_identity_template_bytes(
        "pn-replenishment-request", 1
    )


def _rewrite_zip(
    content: bytes,
    *,
    entry_timestamp: bool = False,
    core_timestamp: bool = False,
    member_bytes: bool = False,
) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(content)) as source, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED
    ) as target:
        for index, info in enumerate(source.infolist()):
            data = source.read(info.filename)
            if core_timestamp and info.filename == "docProps/core.xml":
                data = data.replace(
                    b"2000-01-01T00:00:00Z", b"2001-01-01T00:00:00Z"
                )
            if member_bytes and info.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(b">PN<", b">PX<", 1)
            copied = zipfile.ZipInfo(
                info.filename,
                (2002, 1, 1, 0, 0, 0) if entry_timestamp and index == 0
                else info.date_time,
            )
            copied.compress_type = info.compress_type
            copied.create_system = info.create_system
            copied.external_attr = info.external_attr
            target.writestr(copied, data)
    return output.getvalue()


def _zip_with_comment(content: bytes) -> bytes:
    output = BytesIO(content)
    with zipfile.ZipFile(output, "a") as archive:
        archive.comment = b"attacker-controlled-comment"
    return output.getvalue()


def _business_report(owner, name: str = "业务来源") -> dict:
    headers = ["PN", "采购成本"]
    rows = [["PN-1", 100]]
    evidence = agent_files._mint_report_provenance(
        owner,
        title=name,
        headers=headers,
        rows=rows,
        output_name=f"{name}.xlsx",
        money_cols=[1],
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
        required_positive_keys={"page_purchases", "data_purchase_cost"},
    )
    return agent_files.write_report(
        name,
        headers,
        rows,
        f"{name}.xlsx",
        owner,
        money_cols=[1],
        provenance=evidence,
    )


def _derived_report(owner, source_ids: list[str], name: str) -> dict:
    headers = ["摘要"]
    rows = [[name]]
    evidence = agent_files._mint_report_from_artifacts(
        owner,
        source_ids=source_ids,
        title=name,
        headers=headers,
        rows=rows,
        output_name=f"{name}.xlsx",
        money_cols=None,
    )
    return agent_files.write_report(
        name,
        headers,
        rows,
        f"{name}.xlsx",
        owner,
        provenance=evidence,
    )


def test_model_authored_report_without_server_evidence_is_rejected_before_publish(db):
    owner = _owner(db)

    with pytest.raises(
        agent_files.ProvenanceRequired,
        match="缺少可验证的数据来源",
    ):
        agent_files.write_report(
            "可能来自任意上下文",
            ["任意列"],
            [["模型生成内容"]],
            "unclassified.xlsx",
            owner,
        )

    assert db.query(AgentArtifact).count() == 0


def test_model_authored_cell_edits_without_server_evidence_are_rejected_before_publish(db):
    owner = _owner(db)

    with pytest.raises(
        agent_files.ProvenanceRequired,
        match="缺少可验证的数据来源",
    ):
        agent_files.write_excel(
            None,
            None,
            [{"row": 1, "col": "A", "value": "模型生成内容"}],
            "unclassified.xlsx",
            owner,
        )

    assert db.query(AgentArtifact).count() == 0


def test_direct_renderers_enforce_row_cell_text_and_money_column_budgets(monkeypatch):
    with monkeypatch.context() as scoped:
        scoped.setattr(agent_files, "_MAX_REPORT_ROWS", 1)
        with pytest.raises(agent_files.FileError, match="最多"):
            agent_files._validate_report_shape(
                title=None,
                headers=["value"],
                rows=[[1], [2]],
                output_name=None,
                money_cols=None,
            )

    with monkeypatch.context() as scoped:
        scoped.setattr(agent_files, "_MAX_WRITE_CELLS", 1)
        with pytest.raises(agent_files.FileError, match="单元格"):
            agent_files._validate_excel_write_shape(
                sheet=None,
                cells=[
                    {"row": 1, "col": "A", "value": 1},
                    {"row": 2, "col": "A", "value": 2},
                ],
                output_name=None,
            )

    with monkeypatch.context() as scoped:
        scoped.setattr(agent_files, "_MAX_RENDER_TEXT_BYTES", 3)
        with pytest.raises(agent_files.FileError, match="文本总量"):
            agent_files._validate_report_shape(
                title=None,
                headers=["four"],
                rows=[],
                output_name=None,
                money_cols=None,
            )

    with monkeypatch.context() as scoped:
        scoped.setattr(agent_files, "_MAX_MONEY_COLUMNS", 1)
        with pytest.raises(agent_files.FileError, match="money_cols"):
            agent_files._validate_report_shape(
                title=None,
                headers=["gross", "net"],
                rows=[],
                output_name=None,
                money_cols=[0, 1],
            )


def test_write_excel_consumes_exact_evidence_before_read_and_persists_scope(db):
    owner = _owner(db, role="boss")
    template = agent_files.save_upload(
        _identity_template_bytes(), "补库申请模板.xlsx", owner
    )
    business = _business_report(owner)
    cells = [{"row": 2, "col": "A", "value": "PN-1"}]
    evidence = agent_files._mint_excel_from_artifacts(
        owner,
        source_ids=[template["file_id"], business["file_id"]],
        base_file_id=template["file_id"],
        sheet="申请",
        cells=cells,
        output_name="已回填.xlsx",
    )

    result = agent_files.write_excel(
        template["file_id"],
        "申请",
        cells,
        "已回填.xlsx",
        owner,
        provenance=evidence,
    )

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.source_ids == [template["file_id"], business["file_id"]]
    assert artifact.access_scope["classification"] == "business_content"
    assert artifact.access_scope["contained_resources"] == ["purchases"]
    assert artifact.access_scope["contained_fields"] == ["purchase_cost"]
    download = agent_files.get_download_info(result["file_id"], owner)
    workbook = load_workbook(BytesIO(download.content), data_only=False)
    assert workbook["申请"]["A2"].value == "PN-1"
    workbook.close()


def test_write_excel_rejects_renderer_mismatch_before_loading_or_publishing(
    db, monkeypatch
):
    owner = _owner(db, role="boss")
    template = agent_files.save_upload(
        _identity_template_bytes(), "补库申请模板.xlsx", owner
    )
    business = _business_report(owner)
    signed_cells = [{"row": 2, "col": "A", "value": "PN-1"}]
    evidence = agent_files._mint_excel_from_artifacts(
        owner,
        source_ids=[template["file_id"], business["file_id"]],
        base_file_id=template["file_id"],
        sheet="申请",
        cells=signed_cells,
        output_name="已回填.xlsx",
    )

    def forbidden(*_args, **_kwargs):
        raise AssertionError("evidence must fail before workbook read or publish")

    monkeypatch.setattr(agent_files, "load_workbook", forbidden)
    monkeypatch.setattr(agent_files, "_publish_artifact", forbidden)
    with pytest.raises(agent_files.ProvenanceRequired, match="来源证明无效"):
        agent_files.write_excel(
            template["file_id"],
            "申请",
            [{"row": 2, "col": "A", "value": "模型篡改"}],
            "已回填.xlsx",
            owner,
            provenance=evidence,
        )


@pytest.mark.parametrize("drift", ["permission", "row_scope"])
def test_current_scope_drift_after_evidence_mint_blocks_before_source_read_or_publish(
    db, monkeypatch, drift
):
    owner = _owner(db, role="boss")
    template = agent_files.save_upload(
        _identity_template_bytes(), "补库申请模板.xlsx", owner
    )
    business = _business_report(owner)
    cells = [{"row": 2, "col": "A", "value": "PN-1"}]
    evidence = agent_files._mint_excel_from_artifacts(
        owner,
        source_ids=[template["file_id"], business["file_id"]],
        base_file_id=template["file_id"],
        sheet="申请",
        cells=cells,
        output_name="漂移后禁止.xlsx",
    )

    user = db.query(SysUser).filter(SysUser.username == "provenance-owner").one()
    if drift == "permission":
        user.permissions = {
            **permissions.effective("boss", None),
            "page_purchases": False,
        }
    else:
        user.permissions = {
            **permissions.effective("boss", None),
            "own_customers_only": True,
        }
        user.salesperson_name = "Alice Sales"
    db.commit()

    def forbidden(*_args, **_kwargs):
        raise AssertionError("scope drift must fail before renderer read or publish")

    monkeypatch.setattr(agent_files, "load_workbook", forbidden)
    monkeypatch.setattr(agent_files, "_publish_artifact", forbidden)
    with pytest.raises(agent_files.ProvenanceRequired, match="当前"):
        agent_files.write_excel(
            template["file_id"],
            "申请",
            cells,
            "漂移后禁止.xlsx",
            owner,
            provenance=evidence,
        )


def test_revocation_after_validating_blocks_object_store_delegate_and_fails_row(
    db, monkeypatch
):
    from app.db import SessionLocal

    owner = _owner(db, role="boss")
    source = _business_report(owner)
    title, headers, rows = "发布边界", ["摘要"], [["derived"]]
    evidence = agent_files._mint_report_from_artifacts(
        owner,
        source_ids=[source["file_id"]],
        title=title,
        headers=headers,
        rows=rows,
        output_name="发布边界.xlsx",
        money_cols=None,
    )

    delegate = agent_files.get_artifact_store()

    class NoPublishStore:
        calls = 0

        def path_for(self, storage_key):
            return delegate.path_for(storage_key)

        def inspect(self, storage_key):
            return delegate.inspect(storage_key)

        def read_bytes(self, storage_key, *, max_bytes):
            return delegate.read_bytes(storage_key, max_bytes=max_bytes)

        def publish_bytes(self, *_args, **_kwargs):
            self.calls += 1
            raise AssertionError("revoked output bytes reached the object store")

        def remove(self, storage_key):
            return delegate.remove(storage_key)

    store = NoPublishStore()
    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: store)
    real_mark_validating = agent_files._mark_artifact_validating

    def mark_then_revoke(artifact_id):
        real_mark_validating(artifact_id)
        with SessionLocal.begin() as session:
            user = session.query(SysUser).filter_by(
                username="provenance-owner"
            ).one()
            user.permissions = {
                **permissions.effective("boss", None),
                "page_purchases": False,
            }

    monkeypatch.setattr(agent_files, "_mark_artifact_validating", mark_then_revoke)

    with pytest.raises(agent_files.FileError, match="发布失败"):
        agent_files.write_report(
            title,
            headers,
            rows,
            "发布边界.xlsx",
            owner,
            provenance=evidence,
        )

    assert store.calls == 0
    db.expire_all()
    failed = db.query(AgentArtifact).filter_by(filename="发布边界.xlsx").one()
    assert failed.status == "failed"


def test_generated_source_store_outage_stays_unknown_across_publish_and_reconcile(
    db, monkeypatch
):
    from app.db import SessionLocal

    owner = _owner(db, role="boss")
    source = _business_report(owner, "Store outage source")
    title, headers, rows = "Outage derived", ["Summary"], [["derived"]]

    def evidence_for(output_name):
        return agent_files._mint_report_from_artifacts(
            owner,
            source_ids=[source["file_id"]],
            title=title,
            headers=headers,
            rows=rows,
            output_name=output_name,
            money_cols=None,
        )

    ready = agent_files.write_report(
        title,
        headers,
        rows,
        "ready-before-outage.xlsx",
        owner,
        provenance=evidence_for("ready-before-outage.xlsx"),
    )
    delegate = agent_files.get_artifact_store()
    source_row = db.get(AgentArtifact, source["file_id"])
    assert source_row is not None

    class SourceOutageStore:
        source_unavailable = False

        def read_bytes(self, storage_key, *, max_bytes):
            if self.source_unavailable and storage_key == source_row.storage_key:
                raise agent_files.ArtifactStoreUnavailable("source store unavailable")
            return delegate.read_bytes(storage_key, max_bytes=max_bytes)

        def path_for(self, storage_key):
            return delegate.path_for(storage_key)

        def inspect(self, storage_key):
            return delegate.inspect(storage_key)

        def publish_bytes(self, *args, **kwargs):
            stored = delegate.publish_bytes(*args, **kwargs)
            self.source_unavailable = True
            return stored

    store = SourceOutageStore()
    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: store)
    store.source_unavailable = True
    with pytest.raises(agent_files.ArtifactStoreUnavailable):
        agent_files.get_download_info(ready["file_id"], owner)
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "provenance-owner", "password": "pw123456"},
    )
    assert login.status_code == 200
    response = client.get(
        f"/api/agent/files/{ready['file_id']}",
        headers={"Authorization": f"Bearer {login.json()['token']}"},
    )
    assert response.status_code == 503
    db.expire_all()
    audit = db.query(AgentArtifactAudit).filter_by(
        artifact_id=ready["file_id"],
        action="http_download",
        outcome="denied",
    ).one()
    assert audit.detail["reason_code"] == "store_or_authorization_unavailable"

    store.source_unavailable = False
    with pytest.raises(agent_files.ArtifactStoreUnavailable, match="待协调"):
        agent_files.write_report(
            title,
            headers,
            rows,
            "validating-during-outage.xlsx",
            owner,
            provenance=evidence_for("validating-during-outage.xlsx"),
        )

    db.expire_all()
    pending = db.query(AgentArtifact).filter_by(
        filename="validating-during-outage.xlsx"
    ).one()
    assert pending.status == "validating"
    now = datetime.now(timezone.utc)
    pending = force_artifact_state(
        db,
        pending,
        "validating",
        created_at=now - timedelta(hours=2),
        expires_at=now + timedelta(days=1),
    )
    first = agent_artifact_reconcile.reconcile_agent_artifacts(
        apply=True,
        grace_period=timedelta(hours=1),
        session_factory=SessionLocal,
        now=now,
        artifact_root=delegate.root,
    )
    db.expire_all()
    assert first["outcome"] == "partial"
    assert first["applied"]["mark_failed"] == 0
    assert db.get(AgentArtifact, pending.id).status == "validating"

    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: delegate)
    second = agent_artifact_reconcile.reconcile_agent_artifacts(
        apply=True,
        grace_period=timedelta(hours=1),
        session_factory=SessionLocal,
        now=now,
        artifact_root=delegate.root,
    )
    db.expire_all()
    assert second["applied"]["report_nonready_without_receipt"] == 1
    assert db.get(AgentArtifact, pending.id).status == "validating"


@pytest.mark.parametrize("source_drift", ["status", "hash", "owner", "scope"])
def test_source_fact_drift_after_evidence_mint_blocks_publish(
    db, monkeypatch, source_drift
):
    owner = _owner(db, role="boss")
    source = _business_report(owner)
    title, headers, rows = "来源漂移", ["摘要"], [["derived"]]
    evidence = agent_files._mint_report_from_artifacts(
        owner,
        source_ids=[source["file_id"]],
        title=title,
        headers=headers,
        rows=rows,
        output_name="来源漂移.xlsx",
        money_cols=None,
    )

    row = db.get(AgentArtifact, source["file_id"])
    assert row is not None
    if source_drift == "status":
        row = force_artifact_state(db, row, "failed")
    elif source_drift == "hash":
        row.sha256 = "f" * 64
    elif source_drift == "owner":
        row.owner_sub = "different-owner"
    else:
        scope = copy.deepcopy(row.access_scope)
        scope["source_access_snapshots"][0]["payload"]["source_sha256"] = "e" * 64
        row.access_scope = scope
    db.commit()

    delegate = agent_files.get_artifact_store()

    class NoPublishStore:
        calls = 0

        def path_for(self, storage_key):
            return delegate.path_for(storage_key)

        def inspect(self, storage_key):
            return delegate.inspect(storage_key)

        def read_bytes(self, storage_key, *, max_bytes):
            return delegate.read_bytes(storage_key, max_bytes=max_bytes)

        def publish_bytes(self, *_args, **_kwargs):
            self.calls += 1
            raise AssertionError("drifted source bytes reached object publication")

        def remove(self, storage_key):
            return delegate.remove(storage_key)

    store = NoPublishStore()
    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: store)
    before = db.query(AgentArtifact).count()

    with pytest.raises(agent_files.ProvenanceRequired, match="当前"):
        agent_files.write_report(
            title,
            headers,
            rows,
            "来源漂移.xlsx",
            owner,
            provenance=evidence,
        )

    assert store.calls == 0
    db.expire_all()
    assert db.query(AgentArtifact).count() == before


@pytest.mark.parametrize("principal_drift", ["disabled", "token_version"])
def test_upload_principal_drift_after_validating_blocks_store_and_ready(
    db, monkeypatch, principal_drift
):
    from app.db import SessionLocal

    owner = _owner(db)
    delegate = agent_files.get_artifact_store()

    class NoPublishStore:
        calls = 0

        def path_for(self, storage_key):
            return delegate.path_for(storage_key)

        def inspect(self, storage_key):
            return delegate.inspect(storage_key)

        def read_bytes(self, storage_key, *, max_bytes):
            return delegate.read_bytes(storage_key, max_bytes=max_bytes)

        def publish_bytes(self, *_args, **_kwargs):
            self.calls += 1
            raise AssertionError("revoked upload bytes reached the object store")

        def remove(self, storage_key):
            return delegate.remove(storage_key)

    store = NoPublishStore()
    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: store)
    real_mark_validating = agent_files._mark_artifact_validating

    def mark_then_revoke(artifact_id):
        real_mark_validating(artifact_id)
        with SessionLocal.begin() as session:
            user = session.query(SysUser).filter_by(
                username="provenance-owner"
            ).one()
            if principal_drift == "disabled":
                user.is_active = False
            else:
                user.token_version = int(user.token_version or 0) + 1

    monkeypatch.setattr(agent_files, "_mark_artifact_validating", mark_then_revoke)

    with pytest.raises(agent_files.FileError, match="发布失败"):
        agent_files.save_upload(b"never-publish", "吊销上传.txt", owner)

    assert store.calls == 0
    db.expire_all()
    failed = db.query(AgentArtifact).filter_by(filename="吊销上传.txt").one()
    assert failed.status == "failed"


def test_identity_only_evidence_cannot_authorize_report_or_semantic_edit(db):
    owner = _owner(db)
    template = agent_files.save_upload(
        _identity_template_bytes(), "补库申请模板.xlsx", owner
    )
    title, headers, rows = "非法洗白", ["值"], [["业务值"]]
    report_evidence = agent_files._mint_report_from_artifacts(
        owner,
        source_ids=[template["file_id"]],
        title=title,
        headers=headers,
        rows=rows,
        output_name="非法洗白.xlsx",
        money_cols=None,
    )
    with pytest.raises(agent_files.ProvenanceRequired, match="模板身份"):
        agent_files.write_report(
            title,
            headers,
            rows,
            "非法洗白.xlsx",
            owner,
            provenance=report_evidence,
        )

    cells = [{"row": 2, "col": "A", "value": "PN-SECRET"}]
    edit_evidence = agent_files._mint_excel_from_artifacts(
        owner,
        source_ids=[template["file_id"]],
        base_file_id=template["file_id"],
        sheet="申请",
        cells=cells,
        output_name="非法洗白.xlsx",
    )
    with pytest.raises(agent_files.ProvenanceRequired, match="identity_only"):
        agent_files.write_excel(
            template["file_id"],
            "申请",
            cells,
            "非法洗白.xlsx",
            owner,
            provenance=edit_evidence,
        )


def test_opaque_server_evidence_persists_actual_containment_and_positive_keys(db):
    owner = _owner(db, role="boss")
    title = "采购成本"
    headers = ["PN", "成本"]
    rows = [["PN-1", 100]]
    evidence = agent_files._mint_report_provenance(
        owner,
        title=title,
        headers=headers,
        rows=rows,
        output_name="成本.xlsx",
        money_cols=[1],
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
        required_positive_keys={
            "page_chat", "page_purchases", "data_purchase_cost"
        },
    )

    result = agent_files.write_report(
        title,
        headers,
        rows,
        "成本.xlsx",
        owner,
        money_cols=[1],
        provenance=evidence,
    )

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    snapshots = artifact.access_scope["source_access_snapshots"]
    assert len(snapshots) == 1
    assert {key: value for key, value in artifact.access_scope.items()
            if key != "source_access_snapshots"} == {
        "schema_version": "artifact-access/v2",
        "policy": "provenance_guarded",
        "classification": "business_content",
        "proof_version": "source-union/v1",
        "required_permissions": [
            "data_purchase_cost", "page_chat", "page_purchases"
        ],
        "contained_resources": ["purchases"],
        "contained_fields": ["purchase_cost"],
        "sensitivity": "high",
        "row_subject": None,
        "predicate_version": "source-condition-set/v1",
        "condition": {"op": "all_sources"},
    }
    snapshot = snapshots[0]
    assert snapshot["header"]["purpose"] == "agent.source-snapshot"
    assert snapshot["header"]["payload_schema_version"] == "artifact-source-snapshot/v1"
    assert snapshot["payload"]["source_kind"] == "internal_test"
    assert snapshot["payload"]["owner_sub"] == "provenance-owner"
    assert snapshot["payload"]["required_positive_keys"] == [
        "data_purchase_cost", "page_chat", "page_purchases"
    ]
    assert snapshot["payload"]["condition"] == {"op": "all_rows"}
    assert snapshot["payload"]["classification"] == "business_content"
    assert artifact.sensitivity == "high"
    assert agent_files.get_download_info(result["file_id"], owner).sha256 == artifact.sha256
    revoked_chat = security.UserContext(
        user_id="provenance-owner",
        role="boss",
        permissions={
            **permissions.effective("boss", None),
            "page_chat": False,
        },
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    assert agent_files.access_allowed(result["file_id"], revoked_chat) is False


def test_multi_source_output_persists_each_artifact_snapshot_and_exact_unions(db):
    owner = _owner(db, role="boss")

    def root(name, resources, fields, required):
        headers = ["值"]
        rows = [[name]]
        evidence = agent_files._mint_report_provenance(
            owner,
            title=name,
            headers=headers,
            rows=rows,
            output_name=f"{name}.xlsx",
            money_cols=None,
            contained_resources=set(resources),
            contained_fields=set(fields),
            required_positive_keys=set(required),
        )
        return agent_files.write_report(
            name,
            headers,
            rows,
            f"{name}.xlsx",
            owner,
            provenance=evidence,
        )

    purchase = root(
        "采购",
        {"purchases"},
        {"purchase_cost"},
        {"page_purchases", "data_purchase_cost"},
    )
    profit = root(
        "利润",
        {"profit"},
        {"profit_amount"},
        {"page_profit", "data_profit"},
    )
    title = "合并"
    headers = ["摘要"]
    rows = [["服务端确定性合并"]]
    evidence = agent_files._mint_report_from_artifacts(
        owner,
        source_ids=[purchase["file_id"], profit["file_id"]],
        title=title,
        headers=headers,
        rows=rows,
        output_name="合并.xlsx",
        money_cols=None,
    )

    result = agent_files.write_report(
        title,
        headers,
        rows,
        "合并.xlsx",
        owner,
        provenance=evidence,
    )

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.source_ids == [purchase["file_id"], profit["file_id"]]
    assert artifact.access_scope["required_permissions"] == [
        "data_profit", "data_purchase_cost", "page_profit", "page_purchases"
    ]
    assert artifact.access_scope["contained_resources"] == ["profit", "purchases"]
    assert artifact.access_scope["contained_fields"] == ["profit_amount", "purchase_cost"]
    assert artifact.access_scope["sensitivity"] == artifact.sensitivity == "high"
    snapshots = artifact.access_scope["source_access_snapshots"]
    assert [item["payload"]["source_artifact_id"] for item in snapshots] == [
        purchase["file_id"], profit["file_id"]
    ]
    assert [item["payload"]["source_sha256"] for item in snapshots] == [
        db.get(AgentArtifact, purchase["file_id"]).sha256,
        db.get(AgentArtifact, profit["file_id"]).sha256,
    ]
    assert agent_files.get_download_info(result["file_id"], owner).size_bytes > 0


def test_blank_allowlisted_template_is_classified_identity_only_before_model(db):
    owner = _owner(db)
    content = _identity_template_bytes()

    assert content == _identity_template_bytes()
    assert hashlib.sha256(content).hexdigest() == (
        agent_artifact_provenance.canonical_identity_template_sha256(
            "pn-replenishment-request", 1
        )
    )

    result = agent_files.save_upload(content, "补库申请模板.xlsx", owner)

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.access_scope["policy"] == "owner_only"
    assert artifact.access_scope["classification"] == "identity_only"
    assert artifact.access_scope["proof_version"] == "identity-template-classifier/v2"
    assert artifact.access_scope["contained_resources"] == []
    assert artifact.access_scope["contained_fields"] == []
    assert artifact.access_scope["condition"] == {"op": "top"}
    assert artifact.access_scope["template_proof"] == {
        "classifier_version": "identity-template-classifier/v2",
        "profile_id": "pn-replenishment-request/v1",
        "template_id": "pn-replenishment-request",
        "template_version": 1,
        "template_sha256": artifact.sha256,
        "sheet_headers": [{"sheet": "申请", "headers": ["PN", "数量", "备注"]}],
        "safe_style_profile": "canonical-xlsx-bytes/v1",
        "pre_model": True,
    }


@pytest.mark.parametrize(
    "content",
    [
        _workbook_bytes(rows=[["PN", "数量", "备注"], ["PN-SECRET", 2, "示例值"]]),
        _workbook_bytes(comment="半年无采购不得补库"),
        _workbook_bytes(hidden=True),
        _workbook_bytes(styled=True),
        _workbook_bytes(rows=[["PN", "数量", "半年无采购不得补库"]]),
    ],
    ids=["example", "comment-rule", "hidden-business", "unsafe-style", "semantic-header"],
)
def test_template_with_business_or_non_allowlisted_content_never_becomes_identity_only(db, content):
    owner = _owner(db)

    result = agent_files.save_upload(content, "非纯身份模板.xlsx", owner)

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.access_scope["classification"] == "business_content"
    assert artifact.access_scope["containment_status"] == "unclassified"
    assert artifact.access_scope["proof_version"] == "upload-unclassified/v1"
    assert artifact.access_scope["template_proof"] is None


@pytest.mark.parametrize(
    "mutate",
    [
        lambda content: b"prefix" + content,
        lambda content: content + b"suffix",
        lambda content: _zip_with_comment(content),
        lambda content: _rewrite_zip(content, entry_timestamp=True),
        lambda content: _rewrite_zip(content, core_timestamp=True),
        lambda content: _rewrite_zip(content, member_bytes=True),
    ],
    ids=["prefix", "suffix", "zip-comment", "entry-metadata", "core-time", "member-byte"],
)
def test_any_canonical_template_byte_or_zip_metadata_drift_is_unclassified(db, mutate):
    owner = _owner(db)
    content = mutate(_identity_template_bytes())

    result = agent_files.save_upload(content, "漂移模板.xlsx", owner)

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.access_scope["classification"] == "business_content"
    assert artifact.access_scope["proof_version"] == "upload-unclassified/v1"
    assert artifact.access_scope["template_proof"] is None


def test_identity_template_contributes_empty_top_but_keeps_hash_owner_and_proof(db):
    owner = _owner(db, role="boss")
    template = agent_files.save_upload(
        _identity_template_bytes(), "补库申请模板.xlsx", owner
    )
    source_title = "采购来源"
    source_headers = ["PN", "成本"]
    source_rows = [["PN-1", 100]]
    source_evidence = agent_files._mint_report_provenance(
        owner,
        title=source_title,
        headers=source_headers,
        rows=source_rows,
        output_name="采购来源.xlsx",
        money_cols=[1],
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
        required_positive_keys={"page_purchases", "data_purchase_cost"},
    )
    business = agent_files.write_report(
        source_title,
        source_headers,
        source_rows,
        "采购来源.xlsx",
        owner,
        money_cols=[1],
        provenance=source_evidence,
    )
    title, headers, rows = "回填", ["PN", "数量"], [["PN-1", 2]]
    evidence = agent_files._mint_report_from_artifacts(
        owner,
        source_ids=[template["file_id"], business["file_id"]],
        title=title,
        headers=headers,
        rows=rows,
        output_name="回填.xlsx",
        money_cols=None,
    )

    result = agent_files.write_report(
        title,
        headers,
        rows,
        "回填.xlsx",
        owner,
        provenance=evidence,
    )

    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.access_scope["contained_resources"] == ["purchases"]
    assert artifact.access_scope["contained_fields"] == ["purchase_cost"]
    identity = artifact.access_scope["source_access_snapshots"][0]["payload"]
    assert identity["source_artifact_id"] == template["file_id"]
    assert identity["source_sha256"] == db.get(AgentArtifact, template["file_id"]).sha256
    assert identity["owner_sub"] == "provenance-owner"
    assert identity["classification"] == "identity_only"
    assert identity["proof_version"] == "identity-template-classifier/v2"
    assert identity["condition"] == {"op": "top"}
    assert identity["required_positive_keys"] == []
    assert identity["contained_resources"] == []
    assert identity["contained_fields"] == []


def test_internal_root_evidence_mint_is_unreachable_in_production(db, monkeypatch):
    owner = _owner(db, role="boss")
    from app.config import get_settings

    monkeypatch.setattr(get_settings(), "environment", "prod")
    with pytest.raises(agent_artifact_provenance.ProvenanceError, match="独立 Query Evidence"):
        agent_files._mint_report_provenance(
            owner,
            title="禁止",
            headers=["值"],
            rows=[[1]],
            output_name="禁止.xlsx",
            money_cols=None,
            contained_resources={"purchases"},
            contained_fields=set(),
            required_positive_keys={"page_purchases"},
        )


def test_identity_classifier_proof_requires_known_exact_profile(db):
    owner = _owner(db)
    template = agent_files.save_upload(
        _identity_template_bytes(), "补库申请模板.xlsx", owner
    )
    artifact = db.get(AgentArtifact, template["file_id"])
    assert artifact is not None
    forged_scope = copy.deepcopy(artifact.access_scope)
    forged_scope["template_proof"]["profile_id"] = "unknown-profile/v1"
    forged_scope["template_proof"]["sheet_headers"] = None
    artifact.access_scope = forged_scope
    with pytest.raises(DBAPIError):
        db.commit()
    db.rollback()


def _signed_snapshot(payload: dict) -> dict:
    return agent_integrity.seal(
        payload,
        purpose="agent.source-snapshot",
        payload_schema_version="artifact-source-snapshot/v1",
        keyring=agent_integrity.configured_keyring(),
    )


def _snapshot_payload() -> dict:
    return {
        "source_ref": "00000000-0000-4000-8000-000000000001",
        "source_kind": "internal_test",
        "source_artifact_id": None,
        "source_sha256": "a" * 64,
        "owner_sub": "provenance-owner",
        "required_positive_keys": ["page_purchases"],
        "contained_resources": ["purchases"],
        "contained_fields": [],
        "sensitivity": "medium",
        "row_subject": None,
        "predicate_version": "row-access/v1",
        "condition": {"op": "all_rows"},
        "classification": "business_content",
        "proof_version": "internal-test-source/v1",
    }


@pytest.mark.parametrize(
    "mutate",
    [
        lambda payload: payload.update({
            "row_subject": None,
            "condition": {"op": "row_subject_or_all", "subject": "alice"},
        }),
        lambda payload: payload.update({"condition": {"op": "narrowest"}}),
        lambda payload: payload.update({"contained_resources": [["purchases"]]}),
        lambda payload: payload.update({"required_positive_keys": [{}]}),
        lambda payload: payload.update({"required_positive_keys": []}),
        lambda payload: payload.update({
            "required_positive_keys": ["page_purchases", "unknown_permission"]
        }),
        lambda payload: payload.update({
            "required_positive_keys": ["own_customers_only", "page_purchases"]
        }),
    ],
    ids=[
        "subject-mismatch",
        "unknown-condition",
        "unhashable-containment",
        "non-string-positive-key",
        "underdeclared-containment-permission",
        "unknown-positive-key",
        "negative-permission-key",
    ],
)
def test_authenticated_but_malformed_source_snapshot_fails_closed_without_500(mutate):
    payload = _snapshot_payload()
    mutate(payload)
    envelope = _signed_snapshot(payload)

    with pytest.raises(agent_artifact_provenance.ProvenanceError):
        agent_artifact_provenance.bind_source_envelopes(
            owner_sub="provenance-owner",
            content_fingerprint_value="b" * 64,
            source_envelopes=[envelope],
        )


def test_source_snapshot_fanout_budget_fails_before_aggregation():
    envelopes = []
    for index in range(agent_artifact_provenance.MAX_SOURCE_SNAPSHOTS + 1):
        payload = _snapshot_payload()
        payload["source_ref"] = f"source-{index}"
        envelopes.append(_signed_snapshot(payload))

    with pytest.raises(agent_artifact_provenance.ProvenanceError, match="fanout"):
        agent_artifact_provenance.bind_source_envelopes(
            owner_sub="provenance-owner",
            content_fingerprint_value="b" * 64,
            source_envelopes=envelopes,
        )


@pytest.mark.parametrize("case", ["fanout", "duplicate"])
def test_artifact_source_id_budgets_fail_before_owner_or_database_authorization(
    monkeypatch, case
):
    source_ids = (
        [str(uuid.uuid4()) for _ in range(
            agent_artifact_provenance.MAX_SOURCE_SNAPSHOTS + 1
        )]
        if case == "fanout"
        else [
            "00000000-0000-4000-8000-000000000010",
            "00000000-0000-4000-8000-000000000010",
        ]
    )

    def forbidden(*_args, **_kwargs):
        raise AssertionError("source ID shape must fail before owner/DB authorization")

    monkeypatch.setattr(agent_files, "_verified_owner_context", forbidden)
    monkeypatch.setattr(agent_files, "_canonical_source_id", forbidden)
    with pytest.raises(
        agent_artifact_provenance.ProvenanceError,
        match="fanout|重复",
    ):
        agent_files._mint_report_from_artifacts(
            object(),
            source_ids=source_ids,
            title="early",
            headers=["value"],
            rows=[[1]],
            output_name="early.xlsx",
            money_cols=None,
        )


def test_duplicate_envelope_source_id_fails_before_hmac_verification(monkeypatch):
    source_id = "00000000-0000-4000-8000-000000000010"
    envelopes = [
        {"payload": {"source_artifact_id": source_id}},
        {"payload": {"source_artifact_id": source_id}},
    ]

    def forbidden(*_args, **_kwargs):
        raise AssertionError("duplicate IDs must fail before HMAC verification")

    monkeypatch.setattr(agent_artifact_provenance, "_verify_snapshot", forbidden)
    with pytest.raises(agent_artifact_provenance.ProvenanceError, match="重复"):
        agent_artifact_provenance._aggregate_envelopes(envelopes)


def test_source_snapshot_total_byte_budget_fails_closed(monkeypatch):
    envelope = _signed_snapshot(_snapshot_payload())
    monkeypatch.setattr(agent_artifact_provenance, "MAX_SOURCE_SCOPE_BYTES", 1)

    with pytest.raises(agent_artifact_provenance.ProvenanceError, match="字节预算"):
        agent_artifact_provenance.bind_source_envelopes(
            owner_sub="provenance-owner",
            content_fingerprint_value="b" * 64,
            source_envelopes=[envelope],
        )


def test_source_scope_byte_budget_stops_before_canonicalizing_later_envelopes(
    monkeypatch
):
    envelopes = [
        {"payload": {"source_artifact_id": str(uuid.uuid4())}},
        {"payload": {"source_artifact_id": str(uuid.uuid4())}},
    ]
    calls = []

    def oversized_first(value):
        calls.append(value)
        if len(calls) > 1:
            raise AssertionError("later envelope must not be canonicalized")
        return b"x" * (agent_artifact_provenance.MAX_SOURCE_SCOPE_BYTES + 1)

    monkeypatch.setattr(agent_integrity, "canonicalize", oversized_first)
    with pytest.raises(agent_artifact_provenance.ProvenanceError, match="字节预算"):
        agent_artifact_provenance._aggregate_envelopes(envelopes)
    assert calls == [envelopes[0]]


def test_server_query_snapshot_requires_exact_registered_condition_and_proof():
    payload = _snapshot_payload()
    payload.update({
        "source_kind": "server_query",
        "proof_version": "server-query/v1",
        "row_subject": "alice sales",
        "condition": {
            "op": "row_subject_or_all",
            "subject": "alice sales",
        },
    })
    valid = _signed_snapshot(payload)
    evidence = agent_artifact_provenance.bind_source_envelopes(
        owner_sub="provenance-owner",
        content_fingerprint_value="b" * 64,
        source_envelopes=[valid],
    )
    scope = agent_artifact_provenance.consume_evidence(
        evidence,
        owner_sub="provenance-owner",
        expected_fingerprint="b" * 64,
    )
    assert scope["condition"] == {"op": "all_sources"}

    for field, value in (
        ("proof_version", "model-asserted/v1"),
        ("predicate_version", "row-access/v0"),
        ("condition", {"op": "narrowest"}),
    ):
        invalid_payload = copy.deepcopy(payload)
        invalid_payload[field] = value
        with pytest.raises(agent_artifact_provenance.ProvenanceError):
            agent_artifact_provenance.bind_source_envelopes(
                owner_sub="provenance-owner",
                content_fingerprint_value="b" * 64,
                source_envelopes=[_signed_snapshot(invalid_payload)],
            )


def test_live_source_scope_cannot_be_underdeclared_by_valid_signed_snapshot(db):
    owner = _owner(db, role="boss")
    source = _business_report(owner)
    derived = _derived_report(owner, [source["file_id"]], "派生")
    row = db.get(AgentArtifact, derived["file_id"])
    assert row is not None
    payload = copy.deepcopy(row.access_scope["source_access_snapshots"][0]["payload"])
    payload.update({
        "required_positive_keys": [],
        "contained_resources": [],
        "contained_fields": [],
        "sensitivity": "high",
    })
    forged = _signed_snapshot(payload)
    forged_scope = agent_artifact_provenance._aggregate_envelopes([forged])
    row.access_scope = forged_scope
    row.sensitivity = forged_scope["sensitivity"]
    db.commit()

    current = agent_files._verified_owner_context(owner)
    assert agent_files.access_allowed(derived["file_id"], current) is False
    with pytest.raises(agent_files.ArtifactUnavailable) as exc_info:
        agent_files.get_download_info(derived["file_id"], owner)
    assert exc_info.value.reason_code == "not_found_or_forbidden"


def test_business_upload_cannot_be_laundered_as_identity_source(db):
    owner = _owner(db, role="boss")
    identity = agent_files.save_upload(
        _identity_template_bytes(), "真模板.xlsx", owner
    )
    unclassified = agent_files.save_upload(
        _workbook_bytes(
            rows=[["PN", "数量", "备注"], ["PN-SECRET", 2, "业务内容"]]
        ),
        "业务文件.xlsx",
        owner,
    )
    business = _business_report(owner)
    derived = _derived_report(
        owner, [identity["file_id"], business["file_id"]], "带模板派生"
    )
    row = db.get(AgentArtifact, derived["file_id"])
    bad_source = db.get(AgentArtifact, unclassified["file_id"])
    assert row is not None and bad_source is not None
    envelopes = copy.deepcopy(row.access_scope["source_access_snapshots"])
    identity_payload = envelopes[0]["payload"]
    identity_payload["source_artifact_id"] = unclassified["file_id"]
    identity_payload["source_sha256"] = bad_source.sha256
    envelopes[0] = _signed_snapshot(identity_payload)
    forged_scope = agent_artifact_provenance._aggregate_envelopes(envelopes)
    row.access_scope = forged_scope
    row.source_ids = [unclassified["file_id"], business["file_id"]]
    row.sensitivity = forged_scope["sensitivity"]
    db.commit()

    assert agent_files.access_allowed(
        derived["file_id"], agent_files._verified_owner_context(owner)
    ) is False


def test_each_source_reauthenticates_current_permissions_hash_and_mac(db):
    owner = _owner(db, role="boss")
    source = _business_report(owner)
    derived = _derived_report(owner, [source["file_id"]], "实时复核")
    current = agent_files._verified_owner_context(owner)
    assert agent_files.access_allowed(derived["file_id"], current) is True

    revoked = security.UserContext(
        user_id="provenance-owner",
        role="boss",
        permissions={**permissions.effective("boss", None), "page_purchases": False},
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    assert agent_files.access_allowed(derived["file_id"], revoked) is False

    source_row = db.get(AgentArtifact, source["file_id"])
    assert source_row is not None
    source_row.sha256 = "f" * 64
    db.commit()
    assert agent_files.access_allowed(derived["file_id"], current) is False

    source_row.sha256 = source["artifact"]["sha256"]
    derived_row = db.get(AgentArtifact, derived["file_id"])
    assert derived_row is not None
    tampered_scope = copy.deepcopy(derived_row.access_scope)
    tampered_scope["source_access_snapshots"][0]["payload"]["source_sha256"] = (
        "e" * 64
    )
    derived_row.access_scope = tampered_scope
    db.commit()
    assert agent_files.access_allowed(derived["file_id"], current) is False


def test_shared_ancestor_dag_uses_global_node_work_memo_and_path_budgets(
    db, monkeypatch
):
    owner = _owner(db, role="boss")
    root = _business_report(owner, "共同根")
    left = _derived_report(owner, [root["file_id"]], "左分支")
    right = _derived_report(owner, [root["file_id"]], "右分支")
    joined = _derived_report(
        owner, [left["file_id"], right["file_id"]], "合并分支"
    )
    current = agent_files._verified_owner_context(owner)
    assert agent_files.access_allowed(joined["file_id"], current) is True

    for constant, limit in (
        ("_MAX_PROVENANCE_DERIVATION_DEPTH", 1),
        ("_MAX_PROVENANCE_AUTH_NODES", 1),
        ("_MAX_PROVENANCE_AUTH_WORK", 1),
        ("_MAX_PROVENANCE_AUTH_MEMO", 1),
        ("_MAX_PROVENANCE_AUTH_PATH", 1),
    ):
        with monkeypatch.context() as scoped:
            scoped.setattr(agent_files, constant, limit)
            assert agent_files.access_allowed(joined["file_id"], current) is False
