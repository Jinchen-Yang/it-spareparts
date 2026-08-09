"""Warehouse workbook adapters use synthetic values and structural header codes only."""

from __future__ import annotations

import io
import zipfile
from xml.etree import ElementTree as ET

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import MAC_EPOCH

from app import config
from app.services.maintenance_warehouse_adapters import (
    WarehouseWorkbookError,
    parse_warehouse_workbook,
)


RETURN_PREFIX = "D107407Fd8lreq33f21ltnq5ukwjwaxb4"
SHIPMENT_PREFIX = "D107407Fvxu6voev32rlg4pkdu6nvdc83"
RECEIPT_PREFIX = "D107407Fh8tgyrcma4r2qm9qk8sgk3v92"


def _xlsx(headers: list[tuple[str, str]], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append([code for code, _label in headers])
    sheet.append([label for _code, label in headers])
    for row in rows:
        sheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _shipment_headers(*, optional_first: bool = False) -> list[tuple[str, str]]:
    required = [
        ("ObjectId", "数据ID(不可修改)"),
        ("SeqNo", "出库单号(必填)"),
        ("F0000001", "出库日期(必填)"),
        ("F0000032", "出库类别(必填)"),
        ("F0000061", "出库备件/整机(必填)"),
        ("Status", "数据状态"),
        ("F0000151", "维保需求单(备件)(必填)"),
        (f"{SHIPMENT_PREFIX}.ObjectId", "备件明细.数据ID(不可修改)"),
        (f"{SHIPMENT_PREFIX}.F0000031", "备件明细.备件PN(必填)"),
        (f"{SHIPMENT_PREFIX}.F0000044", "备件明细.备件SN号(必填)"),
        (f"{SHIPMENT_PREFIX}.F0000011", "备件明细.出库数量"),
    ]
    optional = [("SYN_OPTIONAL", "合成可选列")]
    return optional + required if optional_first else required + optional


def _shipment_row(headers: list[tuple[str, str]], *, attachment: str | None = None):
    values = {
        "ObjectId": "SYN-DOC-001",
        "SeqNo": "SYN-SHIP-001",
        "F0000001": "2026-08-01",
        "F0000032": "维保",
        "F0000061": "备件",
        "Status": "已完成",
        "F0000151": "SYN-WBDD-001",
        f"{SHIPMENT_PREFIX}.ObjectId": "SYN-LINE-001",
        f"{SHIPMENT_PREFIX}.F0000031": "SYN-PN-001",
        f"{SHIPMENT_PREFIX}.F0000044": "SYN-SN-001",
        f"{SHIPMENT_PREFIX}.F0000011": 2,
        "SYN_OPTIONAL": "synthetic optional",
        f"{SHIPMENT_PREFIX}.F0000150": attachment,
    }
    return [values.get(code) for code, _label in headers]


def test_required_signature_matches_when_optional_columns_move_and_marks_unknown_version():
    left_headers = _shipment_headers(optional_first=False)
    right_headers = _shipment_headers(optional_first=True)

    left = parse_warehouse_workbook(_xlsx(left_headers, [_shipment_row(left_headers)]))
    right = parse_warehouse_workbook(_xlsx(right_headers, [_shipment_row(right_headers)]))

    assert left.adapter_key == right.adapter_key == "shipment"
    assert left.adapter_version == right.adapter_version == "shipment_v1"
    assert left.header_signature != right.header_signature
    assert left.version_state == right.version_state == "unknown_version"
    assert {item.code for item in left.ambiguities} >= {"unknown_version"}


def test_return_wide_wins_by_required_set_specificity_not_column_count():
    headers = [
        ("F0000032", "返库类别"),
        ("F0000192", "返库类型"),
        ("F0000061", "返库备件/整机"),
        ("F0000001", "返库日期"),
        ("Status", "数据状态"),
        (f"{RETURN_PREFIX}.F0000031", "备件明细.备件PN"),
        (f"{RETURN_PREFIX}.F0000044", "备件明细.备件SN号"),
        (f"{RETURN_PREFIX}.F0000011", "备件明细.返库数量"),
        ("ObjectId", "数据ID(不可修改)"),
        ("SeqNo", "返库单号(必填)"),
        (f"{RETURN_PREFIX}.ObjectId", "备件明细.数据ID(不可修改)"),
    ]
    row = ["维保", "备件", "备件", "2026-08-01", "已完成", "SYN-PN", "SYN-SN", 1,
           "SYN-DOC", "SYN-RETURN", "SYN-LINE"]
    parsed = parse_warehouse_workbook(_xlsx(headers, [row]))
    assert parsed.adapter_key == "return"
    assert parsed.adapter_version == "return_v2"
    assert len(parsed.documents) == 1


def test_receipt_structure_has_an_independent_typed_fixture():
    headers = [
        ("ObjectId", "数据ID(不可修改)"),
        ("SeqNo", "入库单号(必填)"),
        ("F0000001", "入库日期(必填)"),
        ("F0000032", "入库类别(必填)"),
        ("F0000061", "入库备件/整机(必填)"),
        ("Status", "数据状态"),
        ("F0000142", "维保需求单(备件)(必填)"),
        (f"{RECEIPT_PREFIX}.ObjectId", "备件明细.数据ID(不可修改)"),
        (f"{RECEIPT_PREFIX}.F0000031", "备件明细.备件PN(必填)"),
        (f"{RECEIPT_PREFIX}.F0000044", "备件明细.备件SN号(必填)"),
        (f"{RECEIPT_PREFIX}.F0000011", "备件明细.入库数量"),
    ]
    row = [
        "SYN-RECEIPT-DOC", "SYN-RECEIPT-NO", "2026-08-01", "维保入库",
        "备件", "已审批", "SYN-WBDD", "SYN-RECEIPT-LINE", "SYN-PN",
        "SYN-SN", 3,
    ]

    parsed = parse_warehouse_workbook(_xlsx(headers, [row]))

    assert parsed.adapter_key == "receipt"
    assert parsed.adapter_version == "receipt_v1"
    assert parsed.documents[0].normalized_status == "confirmed"
    assert parsed.documents[0].lines[0].quantity == 3


def test_narrow_return_without_stable_ids_creates_only_ambiguity_facts():
    headers = [
        ("F0000032", "返库类别"),
        ("F0000192", "返库类型"),
        ("F0000061", "返库备件/整机"),
        ("F0000001", "返库日期"),
        ("Status", "数据状态"),
        (f"{RETURN_PREFIX}.F0000031", "备件明细.备件PN"),
        (f"{RETURN_PREFIX}.F0000044", "备件明细.备件SN号"),
        (f"{RETURN_PREFIX}.F0000011", "备件明细.返库数量"),
    ]
    parsed = parse_warehouse_workbook(
        _xlsx(headers, [["维保", "备件", "备件", "2026-08-01", "已完成", "SYN-PN", "SYN-SN", 1]])
    )
    assert parsed.adapter_version == "return_v1"
    assert parsed.documents == []
    assert {item.code for item in parsed.ambiguities} >= {
        "missing_document_id",
        "missing_line_id",
    }


def test_duplicate_internal_code_and_formula_fail_closed():
    duplicate = _shipment_headers() + [("ObjectId", "重复数据ID")]
    with pytest.raises(WarehouseWorkbookError, match="重复内部编码"):
        parse_warehouse_workbook(_xlsx(duplicate, [_shipment_row(duplicate)]))

    headers = _shipment_headers()
    row = _shipment_row(headers)
    row[0] = "=1+1"
    with pytest.raises(WarehouseWorkbookError, match="公式"):
        parse_warehouse_workbook(_xlsx(headers, [row]))


def test_external_relationship_fails_closed():
    source = _xlsx(_shipment_headers(), [_shipment_row(_shipment_headers())])
    input_zip = zipfile.ZipFile(io.BytesIO(source))
    output = io.BytesIO()
    with input_zip, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in input_zip.infolist():
            payload = input_zip.read(item.filename)
            if item.filename == "xl/_rels/workbook.xml.rels":
                root = ET.fromstring(payload)
                namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
                ET.SubElement(root, f"{{{namespace}}}Relationship", {
                    "Id": "syntheticExternal",
                    "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
                    "Target": "https://example.invalid/synthetic.xlsx",
                    "TargetMode": "External",
                })
                payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            archive.writestr(item, payload)
    with pytest.raises(WarehouseWorkbookError, match="外部链接"):
        parse_warehouse_workbook(output.getvalue())


def test_attachment_content_is_never_preserved_or_reflected():
    headers = _shipment_headers() + [
        (f"{SHIPMENT_PREFIX}.F0000150", "备件明细.图片/附件")
    ]
    secret = "SYNTHETIC-ATTACHMENT-CONTENT-MUST-NOT-SURVIVE"
    parsed = parse_warehouse_workbook(_xlsx(headers, [_shipment_row(headers, attachment=secret)]))
    serialized = repr(parsed)
    assert secret not in serialized
    assert {item.code for item in parsed.ambiguities} >= {"controlled_attachment"}


def test_attachment_redaction_uses_stable_field_code_not_mutable_label():
    headers = _shipment_headers() + [
        (f"{SHIPMENT_PREFIX}.F0000150", "Attachment payload (renamed by exporter)")
    ]
    secret = "SYNTHETIC-ENGLISH-ATTACHMENT-MUST-NOT-SURVIVE"

    parsed = parse_warehouse_workbook(
        _xlsx(headers, [_shipment_row(headers, attachment=secret)])
    )

    assert secret not in repr(parsed)
    controlled = [
        item for item in parsed.ambiguities
        if item.code == "controlled_attachment"
    ]
    assert f"{SHIPMENT_PREFIX}.F0000150" in {
        item.field_code for item in controlled
    }


def test_mac_1904_epoch_is_used_for_numeric_dates():
    headers = _shipment_headers()
    workbook = Workbook()
    workbook.epoch = MAC_EPOCH
    sheet = workbook.active
    sheet.append([code for code, _label in headers])
    sheet.append([label for _code, label in headers])
    row = _shipment_row(headers)
    row[[code for code, _label in headers].index("F0000001")] = 44773
    sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)

    parsed = parse_warehouse_workbook(output.getvalue())

    assert parsed.documents[0].document_date.isoformat() == "2026-08-01"


def test_activex_member_fails_closed():
    source = _xlsx(_shipment_headers(), [_shipment_row(_shipment_headers())])
    input_zip = zipfile.ZipFile(io.BytesIO(source))
    output = io.BytesIO()
    with input_zip, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for item in input_zip.infolist():
            archive.writestr(item, input_zip.read(item.filename))
        archive.writestr("xl/activeX/activeX1.bin", b"synthetic-active-x")

    with pytest.raises(WarehouseWorkbookError, match="嵌入对象"):
        parse_warehouse_workbook(output.getvalue())


def test_partial_wide_return_never_falls_back_to_narrow_adapter():
    headers = [
        ("F0000032", "返库类别"),
        ("F0000192", "返库类型"),
        ("F0000061", "返库备件/整机"),
        ("F0000001", "返库日期"),
        ("Status", "数据状态"),
        (f"{RETURN_PREFIX}.F0000031", "备件明细.备件PN"),
        (f"{RETURN_PREFIX}.F0000044", "备件明细.备件SN号"),
        (f"{RETURN_PREFIX}.F0000011", "备件明细.返库数量"),
        ("ObjectId", "数据ID(不可修改)"),
        ("SeqNo", "返库单号(必填)"),
    ]
    parsed = parse_warehouse_workbook(_xlsx(headers, [[
        "维保", "备件", "备件", "2026-08-01", "已完成", "SYN-PN", "SYN-SN", 1,
        "SYN-DOC", "SYN-RETURN",
    ]]))

    assert parsed.adapter_version == "return_v2"
    assert parsed.documents == []
    assert {item.code for item in parsed.ambiguities} >= {"missing_line_id"}
    assert not any(item.code == "missing_document_id" for item in parsed.ambiguities)


def test_unknown_enum_becomes_ambiguity_without_guessing():
    headers = _shipment_headers()
    row = _shipment_row(headers)
    status_index = [code for code, _label in headers].index("Status")
    row[status_index] = "SYNTHETIC-UNMAPPED-STATE"
    parsed = parse_warehouse_workbook(_xlsx(headers, [row]))
    assert parsed.documents[0].normalized_status == "unknown"
    assert {item.code for item in parsed.ambiguities} >= {"unknown_enum"}


def test_duplicate_business_label_with_distinct_internal_codes_is_preserved():
    headers = _shipment_headers() + [
        ("SYN_OPTIONAL_TWO", "合成重复业务名称"),
        ("SYN_OPTIONAL_THREE", "合成重复业务名称"),
    ]
    parsed = parse_warehouse_workbook(_xlsx(headers, [_shipment_row(headers)]))
    duplicates = [
        pair for pair in parsed.header_pairs
        if pair.business_label == "合成重复业务名称"
    ]
    assert [pair.internal_code for pair in duplicates] == [
        "SYN_OPTIONAL_TWO",
        "SYN_OPTIONAL_THREE",
    ]


@pytest.mark.parametrize(
    ("setting", "value", "message"),
    [
        ("IMPORT_XLSX_MAX_MEMBERS", 1, "成员数量"),
        ("IMPORT_XLSX_MAX_UNCOMPRESSED_BYTES", 1, "解压体积"),
        ("IMPORT_XLSX_MAX_COLUMNS", 10, "列数"),
        ("IMPORT_MAX_ROWS", 0, "行数"),
    ],
)
def test_archive_and_shape_limits_fail_closed(monkeypatch, setting, value, message):
    headers = _shipment_headers()
    content = _xlsx(headers, [_shipment_row(headers)])
    monkeypatch.setattr(config, setting, value)
    with pytest.raises(WarehouseWorkbookError, match=message):
        parse_warehouse_workbook(content)


def test_formula_in_hidden_option_sheet_also_fails_closed():
    headers = _shipment_headers()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append([code for code, _label in headers])
    sheet.append([label for _code, label in headers])
    sheet.append(_shipment_row(headers))
    hidden = workbook.create_sheet("Synthetic_Options")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "=1+1"
    output = io.BytesIO()
    workbook.save(output)
    with pytest.raises(WarehouseWorkbookError, match="公式"):
        parse_warehouse_workbook(output.getvalue())


def test_synthetic_49_and_120_column_contracts_cover_all_four_adapter_signatures(
    monkeypatch,
):
    def pad(
        headers: list[tuple[str, str]], width: int, prefix: str
    ) -> list[tuple[str, str]]:
        return headers + [
            (f"SYN_{prefix}_PAD_{index:03d}", f"合成占位列 {index:03d}")
            for index in range(1, width - len(headers) + 1)
        ]

    return_common = [
        ("F0000032", "返库类别"),
        ("F0000192", "返库类型"),
        ("F0000061", "返库备件/整机"),
        ("F0000001", "返库日期"),
        ("Status", "数据状态"),
        (f"{RETURN_PREFIX}.F0000031", "备件明细.备件PN"),
        (f"{RETURN_PREFIX}.F0000044", "备件明细.备件SN号"),
        (f"{RETURN_PREFIX}.F0000011", "备件明细.返库数量"),
    ]
    return_v2 = return_common + [
        ("ObjectId", "数据ID(不可修改)"),
        ("SeqNo", "返库单号(必填)"),
        (f"{RETURN_PREFIX}.ObjectId", "备件明细.数据ID(不可修改)"),
    ]
    receipt = [
        ("ObjectId", "数据ID(不可修改)"),
        ("SeqNo", "入库单号(必填)"),
        ("F0000001", "入库日期(必填)"),
        ("F0000032", "入库类别(必填)"),
        ("F0000061", "入库备件/整机(必填)"),
        ("Status", "数据状态"),
        ("F0000142", "维保需求单(备件)(必填)"),
        (f"{RECEIPT_PREFIX}.ObjectId", "备件明细.数据ID(不可修改)"),
        (f"{RECEIPT_PREFIX}.F0000031", "备件明细.备件PN(必填)"),
        (f"{RECEIPT_PREFIX}.F0000044", "备件明细.备件SN号(必填)"),
        (f"{RECEIPT_PREFIX}.F0000011", "备件明细.入库数量"),
    ]
    structures = {
        "return_v1": (
            pad(return_common, 49, "RETURN_NARROW"),
            ["维保", "备件", "备件", "2026-08-01", "已完成", "SYN-PN", "SYN-SN", 1],
        ),
        "return_v2": (
            pad(return_v2, 120, "RETURN_WIDE"),
            ["维保", "备件", "备件", "2026-08-01", "已完成", "SYN-PN", "SYN-SN", 1,
             "SYN-RETURN-DOC", "SYN-RETURN-NO", "SYN-RETURN-LINE"],
        ),
        "shipment_v1": (
            pad(_shipment_headers(), 49, "SHIPMENT"),
            None,
        ),
        "receipt_v1": (
            pad(receipt, 49, "RECEIPT"),
            ["SYN-RECEIPT-DOC", "SYN-RECEIPT-NO", "2026-08-01", "维保入库",
             "备件", "已审批", "SYN-WBDD", "SYN-RECEIPT-LINE", "SYN-PN", "SYN-SN", 3],
        ),
    }
    contents: dict[str, bytes] = {}
    contracts: dict[str, tuple[tuple[tuple[str, str], ...], ...]] = {}
    unknown_signatures: set[str] = set()
    for version, (headers, values) in structures.items():
        row = (
            _shipment_row(headers)
            if version == "shipment_v1"
            else [*(values or []), *([None] * (len(headers) - len(values or [])))]
        )
        content = _xlsx(headers, [row])
        contents[version] = content
        parsed = parse_warehouse_workbook(content)
        assert parsed.adapter_version == version
        assert parsed.version_state == "unknown_version"
        unknown_signatures.add(parsed.header_signature)
        contracts[version] = (tuple(
            (pair.internal_code, pair.business_label) for pair in parsed.header_pairs
        ),)

    assert len(unknown_signatures) == 4
    assert len(parse_warehouse_workbook(contents["return_v1"]).header_pairs) == 49
    assert len(parse_warehouse_workbook(contents["return_v2"]).header_pairs) == 120

    monkeypatch.setattr(
        config,
        "MAINTENANCE_WAREHOUSE_APPROVED_HEADER_CONTRACTS",
        contracts,
    )
    for version, content in contents.items():
        approved = parse_warehouse_workbook(content)
        assert approved.adapter_version == version
        assert approved.version_state == "known"
        assert approved.header_diff["state"] == "approved_exact"
