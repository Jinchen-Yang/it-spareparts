"""Pure protocol contract for the project-manager monthly workbook v3 (#206)."""

from __future__ import annotations

import io
import unittest
import zipfile
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from app.services import maintenance_manager_workbook_v3 as workbook_v3


HMAC_KEY = b"synthetic-project-manager-workbook-v3-key"


def _snapshot() -> dict:
    return {
        "owner": {
            "user_id": 17,
            "username": "synthetic-manager",
            "display_name": "合成项目经理",
        },
        "report_month": date(2026, 8, 1),
        "scope_version": "scope-synthetic-1",
        "data_version": "manager-data-synthetic-1",
        "projects": [
            {
                "project_id": "project-synthetic-1",
                "project_code": "PROJECT-SYNTHETIC-1",
                "project_name": "合成维保项目",
                "project_version": 3,
                "service_start": date(2026, 1, 1),
                "service_end": date(2026, 12, 31),
                "service_period_version": 2,
                "contracts": [
                    {
                        "project_contract_id": "project-contract-synthetic-1",
                        "contract_no": "CONTRACT-SYNTHETIC-1",
                        "contract_amount": Decimal("100000.00"),
                        "contract_version": 4,
                        "confirmed_received_amount": Decimal("30000.00"),
                        "planned_milestones": [
                            {
                                "sequence": 1,
                                "planned_date": date(2026, 9, 10),
                                "planned_amount": Decimal("20000.00"),
                                "version": 2,
                            },
                            {
                                "sequence": 2,
                                "planned_date": None,
                                "planned_amount": Decimal("15000.00"),
                                "version": 1,
                            },
                        ],
                    }
                ],
                "acceptance": {
                    "configuration_state": "pending_business_configuration",
                    "submission_status": "not_submitted",
                    "approval_status": "not_reviewed",
                    "attachment_count": 0,
                },
            }
        ],
    }


def _edit_plan(content: bytes, *, sequence: int, date_value=..., amount_value=...) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        sheet = book[workbook_v3.PLAN_SHEET]
        table = sheet.tables[workbook_v3.PLAN_TABLE]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        target = next(
            row
            for row in range(min_row + 1, max_row + 1)
            if sheet.cell(row, headers.index("项目合同关系ID") + 1).value
            == "project-contract-synthetic-1"
            and sheet.cell(row, headers.index("计划期次") + 1).value == sequence
        )
        if date_value is not ...:
            sheet.cell(target, headers.index("计划回款日期") + 1, date_value)
        if amount_value is not ...:
            sheet.cell(target, headers.index("计划回款金额（含税）") + 1, amount_value)
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _edit_overview(content: bytes, *, header: str, value) -> bytes:
    book = load_workbook(io.BytesIO(content), data_only=False)
    try:
        sheet = book[workbook_v3.OVERVIEW_SHEET]
        table = sheet.tables[workbook_v3.OVERVIEW_TABLE]
        min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
        headers = [sheet.cell(min_row, column).value for column in range(min_col, max_col + 1)]
        sheet.cell(min_row + 1, headers.index(header) + 1, value)
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _add_zip_member(content: bytes, name: str, payload: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(content), "r") as source:
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for item in source.infolist():
                target.writestr(item, source.read(item.filename))
            target.writestr(name, payload)
    return output.getvalue()


class MaintenanceManagerWorkbookV3ContractTest(unittest.TestCase):
    def test_export_is_own_scope_and_has_24_longitudinal_plan_rows(self):
        artifact = workbook_v3.build_manager_workbook(_snapshot(), hmac_key=HMAC_KEY)

        book = load_workbook(io.BytesIO(artifact.content), data_only=False)
        try:
            self.assertEqual(
                [sheet for sheet in book.sheetnames if book[sheet].sheet_state == "visible"],
                list(workbook_v3.VISIBLE_SHEETS),
            )
            sheet = book[workbook_v3.PLAN_SHEET]
            table = sheet.tables[workbook_v3.PLAN_TABLE]
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            self.assertEqual(max_row - min_row, 24)
            headers = [
                sheet.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            sequence_col = headers.index("计划期次") + 1
            self.assertEqual(
                [sheet.cell(row, sequence_col).value for row in range(min_row + 1, max_row + 1)],
                list(range(1, 25)),
            )
            overview = book[workbook_v3.OVERVIEW_SHEET]
            self.assertIn("财务确认实收（只读）", [cell.value for cell in overview[5]])
            self.assertNotIn("实收金额", headers)
        finally:
            book.close()

    def test_unchanged_roundtrip_has_zero_writes_and_preserves_partial_node(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)

        result = workbook_v3.validate_manager_workbook(
            artifact.content,
            snapshot=snapshot,
            hmac_key=HMAC_KEY,
        )

        self.assertTrue(result.can_apply)
        self.assertTrue(result.unchanged)
        self.assertEqual(result.service_period_changes, ())
        self.assertEqual(result.milestone_changes, ())
        self.assertTrue(any(issue.code == "partial_plan_node" for issue in result.warnings))

    def test_project_without_included_contracts_roundtrips_with_empty_plan(self):
        snapshot = _snapshot()
        snapshot["projects"][0]["contracts"] = []
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)

        result = workbook_v3.validate_manager_workbook(
            artifact.content,
            snapshot=snapshot,
            hmac_key=HMAC_KEY,
        )

        self.assertEqual(artifact.project_count, 1)
        self.assertEqual(artifact.milestone_row_count, 0)
        self.assertTrue(result.can_apply)
        self.assertTrue(result.unchanged)

    def test_blank_cells_do_not_delete_existing_plan_values(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)
        edited = _edit_plan(artifact.content, sequence=1, date_value=None, amount_value=None)

        result = workbook_v3.validate_manager_workbook(
            edited,
            snapshot=snapshot,
            hmac_key=HMAC_KEY,
        )

        self.assertTrue(result.can_apply)
        self.assertTrue(result.unchanged)
        self.assertEqual(result.milestone_changes, ())

    def test_date_only_new_node_is_kept_as_incomplete_warning(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)
        edited = _edit_plan(
            artifact.content,
            sequence=3,
            date_value=date(2026, 10, 12),
            amount_value=None,
        )

        result = workbook_v3.validate_manager_workbook(
            edited,
            snapshot=snapshot,
            hmac_key=HMAC_KEY,
        )

        self.assertTrue(result.can_apply)
        self.assertFalse(result.unchanged)
        self.assertEqual(len(result.milestone_changes), 1)
        change = result.milestone_changes[0]
        self.assertEqual(change.sequence, 3)
        self.assertEqual(change.planned_date, date(2026, 10, 12))
        self.assertIsNone(change.planned_amount)
        self.assertEqual(change.completeness_state, "date_only")
        self.assertTrue(any(issue.code == "partial_plan_node" for issue in result.warnings))

    def test_acceptance_due_date_is_an_editable_versioned_change(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)
        edited = _edit_overview(
            artifact.content,
            header="验收报告截止日",
            value=date(2026, 11, 20),
        )

        result = workbook_v3.validate_manager_workbook(
            edited,
            snapshot=snapshot,
            hmac_key=HMAC_KEY,
        )

        self.assertTrue(result.can_apply)
        self.assertEqual(len(result.acceptance_due_date_changes), 1)
        self.assertEqual(
            result.acceptance_due_date_changes[0].due_date,
            date(2026, 11, 20),
        )
        self.assertEqual(result.acceptance_due_date_changes[0].expected_version, 0)

    def test_manager_cannot_write_confirmed_actual_collection(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)
        book = load_workbook(io.BytesIO(artifact.content), data_only=False)
        try:
            overview = book[workbook_v3.OVERVIEW_SHEET]
            table = overview.tables[workbook_v3.OVERVIEW_TABLE]
            min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
            headers = [
                overview.cell(min_row, column).value
                for column in range(min_col, max_col + 1)
            ]
            overview.cell(
                min_row + 1,
                headers.index("财务确认实收（只读）") + 1,
                90000,
            )
            output = io.BytesIO()
            book.save(output)
        finally:
            book.close()

        result = workbook_v3.validate_manager_workbook(
            output.getvalue(),
            snapshot=snapshot,
            hmac_key=HMAC_KEY,
        )

        self.assertFalse(result.can_apply)
        self.assertTrue(any(issue.code == "readonly_actual_changed" for issue in result.errors))

    def test_metadata_tamper_fails_closed(self):
        artifact = workbook_v3.build_manager_workbook(_snapshot(), hmac_key=HMAC_KEY)
        book = load_workbook(io.BytesIO(artifact.content), data_only=False)
        try:
            metadata = book[workbook_v3.METADATA_SHEET]
            metadata.cell(2, 2, "different-owner")
            output = io.BytesIO()
            book.save(output)
        finally:
            book.close()

        with self.assertRaises(workbook_v3.ManagerWorkbookV3Error) as raised:
            workbook_v3.validate_manager_workbook(
                output.getvalue(),
                snapshot=_snapshot(),
                hmac_key=HMAC_KEY,
            )
        self.assertEqual(raised.exception.issues[0].code, "metadata_tampered")

    def test_server_signed_non_current_template_version_is_rejected(self):
        artifact = workbook_v3.build_manager_workbook(_snapshot(), hmac_key=HMAC_KEY)
        book = load_workbook(io.BytesIO(artifact.content), data_only=False)
        try:
            metadata_sheet = book[workbook_v3.METADATA_SHEET]
            table = metadata_sheet.tables[workbook_v3.METADATA_TABLE]
            _min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
            rows = {
                str(metadata_sheet.cell(row, 1).value): row
                for row in range(min_row + 1, max_row + 1)
            }
            metadata_sheet.cell(rows["template_version"], 2, "2.9.9")
            metadata = {
                key: str(metadata_sheet.cell(row, 2).value or "")
                for key, row in rows.items()
            }
            metadata_sheet.cell(
                rows["metadata_hmac"],
                2,
                workbook_v3._signature(metadata, HMAC_KEY),
            )
            output = io.BytesIO()
            book.save(output)
        finally:
            book.close()

        with self.assertRaises(workbook_v3.ManagerWorkbookV3Error) as raised:
            workbook_v3.validate_manager_workbook(
                output.getvalue(),
                snapshot=_snapshot(),
                hmac_key=HMAC_KEY,
            )

        self.assertEqual(raised.exception.issues[0].code, "template_version_mismatch")

    def test_formula_and_external_link_packages_are_rejected(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)
        formula = _edit_plan(
            artifact.content,
            sequence=3,
            amount_value="=1+1",
        )
        with self.assertRaisesRegex(
            workbook_v3.ManagerWorkbookV3Error,
            "包含公式",
        ):
            workbook_v3.validate_manager_workbook(
                formula,
                snapshot=snapshot,
                hmac_key=HMAC_KEY,
            )

        external = _add_zip_member(
            artifact.content,
            "xl/externalLinks/externalLink1.xml",
            b"<externalLink/>",
        )
        with self.assertRaisesRegex(
            workbook_v3.ManagerWorkbookV3Error,
            "外部链接",
        ):
            workbook_v3.validate_manager_workbook(
                external,
                snapshot=snapshot,
                hmac_key=HMAC_KEY,
            )

    def test_scope_or_data_drift_makes_export_stale(self):
        snapshot = _snapshot()
        artifact = workbook_v3.build_manager_workbook(snapshot, hmac_key=HMAC_KEY)
        current = _snapshot()
        current["scope_version"] = "scope-synthetic-2"

        with self.assertRaises(workbook_v3.ManagerWorkbookV3Error) as raised:
            workbook_v3.validate_manager_workbook(
                artifact.content,
                snapshot=current,
                hmac_key=HMAC_KEY,
            )

        self.assertEqual(raised.exception.status_code, 409)
        self.assertEqual(raised.exception.issues[0].code, "stale_workbook")


if __name__ == "__main__":
    unittest.main()
