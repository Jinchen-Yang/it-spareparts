"""合同详细盈亏在看板、CSV、单本和批量工作簿之间的公开契约。"""

import csv
import io
from datetime import date
from decimal import Decimal
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select

from app import permissions, security
from app.agent import tools
from app.auth import hash_password
from app.main import app
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceContractWorkbookState,
)
from app.models.sales import FSalesOrder
from app.models.system import SysUser
from app.services import maintenance_margin_evidence
from tests.test_maintenance_export_headers import (
    _admin_client,
    _cost_blind_maintenance_client,
    _profit_blind_maintenance_client,
    _readonly_client,
)
from tests.test_maintenance_margin_integration import _load_complete_contract


_WORKBOOK_LABELS = {
    "revenue_inc": "合同收入（含税）",
    "revenue_ex": "合同收入（未税）",
    "expense_inc": "报销费用（含税）",
    "expense_ex": "报销费用（未税）",
    "parts_cost_inc_tax": "备件成本（含税归一）",
    "parts_cost_ex_tax": "备件成本（未税归一）",
    "parts_gross_profit_inc": "合同级备件毛利（含税）",
    "parts_gross_profit_ex": "合同级备件毛利（未税）",
    "parts_gross_margin_inc": "合同级备件毛利率（含税）",
    "parts_gross_margin_ex": "合同级备件毛利率（未税）",
    "contribution_profit_inc": "合同级贡献毛利（含税）",
    "contribution_profit_ex": "合同级贡献毛利（未税）",
    "contribution_margin_inc": "合同级贡献毛利率（含税）",
    "contribution_margin_ex": "合同级贡献毛利率（未税）",
}
_CENT = Decimal("0.01")
_RATE = Decimal("0.0001")
_FULL_EXPENSE_COVERAGE = date.max
_PARTS_STATUS_LABELS = {
    "complete_actual": "完整：仅实际成本",
    "complete_estimated": "完整：含估算成本",
    "missing_revenue": "合同收入缺失",
    "missing_tax_rate": "合同税率缺失",
    "invalid_tax_rate": "合同税率异常",
    "ambiguous_revenue": "重复合同收入冲突",
    "incomplete_cost": "成本不完整",
    "filtered_scope": "日期筛选下暂不计算",
}
_CONTRIBUTION_STATUS_LABELS = {
    **_PARTS_STATUS_LABELS,
    "complete": "完整",
    "expense_tax_unknown": "费用税务口径缺失",
    "expense_data_unavailable": "费用数据未就绪",
}
_EXPENSE_STATUS_LABELS = {
    "complete": "完整",
    "expense_tax_unknown": "费用税务口径缺失",
    "expense_data_unavailable": "未就绪（无记录不等于0）",
}


def _summary(workbook) -> dict[str, object]:
    values: dict[str, object] = {}
    for row in workbook["项目预算"].iter_rows(min_row=3, values_only=True):
        if row[0]:
            values[str(row[0])] = row[1]
        if len(row) > 2 and row[2]:
            values[str(row[2])] = row[3]
    return values


def _csv_row(response) -> dict[str, str]:
    rows = list(csv.DictReader(io.StringIO(response.content.decode("utf-8-sig"))))
    assert len(rows) == 1
    return rows[0]


def _scoped_sales_maintenance_client(db) -> TestClient:
    base = permissions.effective("sales", None)
    overrides = {"page_maintenance": True, "own_customers_only": True}
    db.add(SysUser(
        username="maintenance_profit_export_scoped_sales",
        role="sales",
        salesperson_name="测试销售",
        password_hash=hash_password("pw123456"),
        is_active=True,
        template_code="sales",
        template_version=1,
        template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides),
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={
            "username": "maintenance_profit_export_scoped_sales",
            "password": "pw123456",
        },
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _admin_agent_ctx() -> security.UserContext:
    """Match the persisted identity created by the shared maintenance export admin helper."""
    return security.UserContext(
        user_id="maintenance_export_admin",
        role="admin",
        permissions=permissions.effective("admin", None),
        is_authenticated=True,
        authn="sys_user",
        token_version=0,
    )


def _assert_numeric_parity(
    board: dict,
    csv_row: dict[str, str],
    *workbook_summaries: dict[str, object],
) -> None:
    for field, label in _WORKBOOK_LABELS.items():
        quantum = _RATE if "margin" in field else _CENT
        board_value = _quantized(board[field], quantum)
        csv_value = _quantized(csv_row[field], quantum)
        assert csv_value == board_value
        for summary in workbook_summaries:
            workbook_value = _quantized(summary[label], quantum)
            assert workbook_value == board_value
    _assert_status_parity(board, csv_row, *workbook_summaries)


def _assert_status_parity(
    board: dict,
    csv_row: dict[str, str],
    *workbook_summaries: dict[str, object],
) -> None:
    for basis, basis_label in (("inc", "含税"), ("ex", "未税")):
        parts_status = board[f"parts_profit_status_{basis}"]
        contribution_status = board[f"contribution_status_{basis}"]
        assert csv_row[f"parts_profit_status_{basis}"] == parts_status
        assert csv_row[f"contribution_status_{basis}"] == contribution_status
        for summary in workbook_summaries:
            assert summary[f"{basis_label}备件毛利状态"] == (
                _PARTS_STATUS_LABELS[parts_status]
            )
            assert summary[f"{basis_label}贡献毛利状态"] == (
                _CONTRIBUTION_STATUS_LABELS[contribution_status]
            )

        revenue_status = (
            "available"
            if board[f"revenue_{basis}"] is not None
            else parts_status
        )
        assert csv_row[f"收入证据状态-{basis_label}"] == revenue_status

    expense_status = board["expense_evidence_status"]
    assert csv_row["费用证据状态"] == expense_status
    for summary in workbook_summaries:
        assert summary["费用证据状态"] == _EXPENSE_STATUS_LABELS[expense_status]


def _quantized(value: object, quantum: Decimal) -> Decimal | None:
    if value in (None, "", "—"):
        return None
    return Decimal(str(value)).quantize(quantum)


def _four_carriers(
    client: TestClient,
    *,
    board_params: dict[str, str] | None = None,
) -> tuple[dict, dict[str, str], dict[str, object], dict[str, object]]:
    board_params = {"lifecycle": "all", **(board_params or {})}
    date_params = {
        key: board_params[key]
        for key in ("date_from", "date_to")
        if key in board_params
    }
    board_response = client.get("/api/maintenance/board", params=board_params)
    csv_response = client.get("/api/maintenance/board/export", params=board_params)
    single_response = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XS-MARGIN", **date_params},
    )
    bundle_response = client.get(
        "/api/maintenance/export-workbooks",
        params=date_params,
    )

    assert board_response.status_code == 200, board_response.text
    assert csv_response.status_code == 200, csv_response.text
    assert single_response.status_code == 200, single_response.text
    assert bundle_response.status_code == 200, bundle_response.text
    board_rows = board_response.json()["rows"]
    assert len(board_rows) == 1
    single = load_workbook(io.BytesIO(single_response.content), data_only=False)
    try:
        single_summary = _summary(single)
    finally:
        single.close()
    with ZipFile(io.BytesIO(bundle_response.content)) as archive:
        member = next(
            name for name in archive.namelist()
            if name.startswith("项目工作簿/")
        )
        bundled = load_workbook(
            io.BytesIO(archive.read(member)),
            data_only=False,
        )
    try:
        bundled_summary = _summary(bundled)
    finally:
        bundled.close()
    return board_rows[0], _csv_row(csv_response), single_summary, bundled_summary


def test_rounding_boundary_has_decimal_parity_across_all_four_carriers(db):
    batch = _load_complete_contract(
        db,
        purchase_unit_price=Decimal("100.01"),
    )
    sales_order = db.scalar(
        select(FSalesOrder).where(FSalesOrder.order_no == "XS-MARGIN"),
    )
    assert sales_order is not None
    sales_order.amount_ex_tax = Decimal("1000.05")
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-MARGIN-PARITY",
            bxd_no="BXD-MARGIN-PARITY",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=Decimal("0.05"),
            amount_ex_tax=Decimal("0.05"),
            amount_inc_tax=Decimal("0.06"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=_FULL_EXPENSE_COVERAGE,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()
    board, csv_data, single, bundled = _four_carriers(_admin_client(db))

    assert board["parts_profit_status_inc"] == "complete_actual"
    assert board["parts_profit_status_ex"] == "complete_actual"
    assert board["contribution_status_inc"] == "complete"
    assert board["contribution_status_ex"] == "complete"
    assert csv_data["成本证据状态"] == "actual_only"
    assert csv_data["费用证据状态"] == "complete"
    assert _quantized(board["revenue_inc"], _CENT) == Decimal("1130.06")
    assert _quantized(board["revenue_ex"], _CENT) == Decimal("1000.05")
    assert _quantized(board["parts_cost_inc_tax"], _CENT) == Decimal("226.02")
    assert _quantized(board["parts_cost_ex_tax"], _CENT) == Decimal("200.02")
    assert _quantized(board["expense_inc"], _CENT) == Decimal("0.06")
    assert _quantized(board["expense_ex"], _CENT) == Decimal("0.05")
    assert _quantized(board["contribution_profit_inc"], _CENT) == Decimal(
        "903.98",
    )
    assert _quantized(board["contribution_profit_ex"], _CENT) == Decimal(
        "799.98",
    )
    _assert_numeric_parity(
        board,
        csv_data,
        single,
        bundled,
    )


def test_contract_profit_csv_matches_board_status_and_dual_cost_evidence(db):
    _load_complete_contract(db)
    client = _admin_client(db)

    matching_board = client.get(
        "/api/maintenance/board",
        params={"lifecycle": "all", "status": "expense_data_unavailable"},
    )
    matching_csv = client.get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all", "status": "expense_data_unavailable"},
    )
    excluded_board = client.get(
        "/api/maintenance/board",
        params={"lifecycle": "all", "status": "incomplete_cost"},
    )
    excluded_csv = client.get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all", "status": "incomplete_cost"},
    )

    assert matching_board.status_code == 200, matching_board.text
    assert matching_csv.status_code == 200, matching_csv.text
    assert len(matching_board.json()["rows"]) == 1
    matching_row = _csv_row(matching_csv)
    assert matching_row["成本证据状态-含税"] == "actual_only"
    assert matching_row["成本证据状态-未税"] == "actual_only"
    assert excluded_board.status_code == 200, excluded_board.text
    assert excluded_board.json()["rows"] == []
    assert excluded_csv.status_code == 422
    assert excluded_csv.json()["detail"] == "所选范围内没有可导出的合同详细盈亏数据"


def test_expense_unavailable_is_consistent_across_all_four_carriers(db):
    _load_complete_contract(db)

    board, csv_data, single, bundled = _four_carriers(_admin_client(db))

    assert board["expense_inc"] is None
    assert board["expense_ex"] is None
    assert board["contribution_status_inc"] == "expense_data_unavailable"
    assert board["contribution_status_ex"] == "expense_data_unavailable"
    assert csv_data["费用证据状态"] == "expense_data_unavailable"
    assert single["费用证据状态"] == "未就绪（无记录不等于0）"
    assert bundled["费用证据状态"] == "未就绪（无记录不等于0）"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_unknown_expense_tax_evidence_fails_closed_across_all_four_carriers(db):
    batch = _load_complete_contract(db)
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-MARGIN-TAX-UNKNOWN",
            bxd_no="BXD-MARGIN-TAX-UNKNOWN",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=None,
            amount_ex_tax=None,
            amount_inc_tax=None,
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=_FULL_EXPENSE_COVERAGE,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    board, csv_data, single, bundled = _four_carriers(_admin_client(db))

    assert board["expense_inc"] is None
    assert board["expense_ex"] is None
    assert board["contribution_status_inc"] == "expense_tax_unknown"
    assert board["contribution_status_ex"] == "expense_tax_unknown"
    assert csv_data["费用证据状态"] == "expense_tax_unknown"
    assert single["费用证据状态"] == "费用税务口径缺失"
    assert bundled["费用证据状态"] == "费用税务口径缺失"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_missing_cost_fails_closed_across_all_four_carriers(db):
    _load_complete_contract(db)
    line = db.scalar(select(FMaintenanceLine))
    assert line is not None
    line.unit_cost = None
    line.cost_amount = None
    line.unit_cost_inc_tax = None
    line.unit_cost_ex_tax = None
    line.cost_amount_inc_tax = None
    line.cost_amount_ex_tax = None
    line.cost_source = "none"
    line.cost_tax_basis = None
    db.add(MaintenanceContractWorkbookState(
        contract_no="XS-MARGIN",
        revision=1,
        expense_complete_through=_FULL_EXPENSE_COVERAGE,
        expense_snapshot_complete=True,
    ))
    db.commit()

    board, csv_data, single, bundled = _four_carriers(_admin_client(db))

    assert board["parts_profit_status_inc"] == "incomplete_cost"
    assert board["parts_profit_status_ex"] == "incomplete_cost"
    assert board["contribution_status_inc"] == "incomplete_cost"
    assert board["contribution_status_ex"] == "incomplete_cost"
    assert csv_data["成本证据状态-含税"] == "incomplete"
    assert csv_data["成本证据状态-未税"] == "incomplete"
    assert single["含税口径质量"] == "成本不完整，需补数据"
    assert single["未税口径质量"] == "成本不完整，需补数据"
    assert bundled["含税口径质量"] == "成本不完整，需补数据"
    assert bundled["未税口径质量"] == "成本不完整，需补数据"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_missing_cost_does_not_hide_independent_expense_tax_evidence(db):
    batch = _load_complete_contract(db)
    line = db.scalar(select(FMaintenanceLine))
    assert line is not None
    line.unit_cost = None
    line.cost_amount = None
    line.unit_cost_inc_tax = None
    line.unit_cost_ex_tax = None
    line.cost_amount_inc_tax = None
    line.cost_amount_ex_tax = None
    line.cost_source = "none"
    line.cost_tax_basis = None
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-MISSING-COST-TAX-UNKNOWN",
            bxd_no="BXD-MISSING-COST-TAX-UNKNOWN",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=None,
            amount_ex_tax=None,
            amount_inc_tax=None,
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=_FULL_EXPENSE_COVERAGE,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    board, csv_data, single, bundled = _four_carriers(_admin_client(db))

    assert board["parts_profit_status_inc"] == "incomplete_cost"
    assert board["parts_profit_status_ex"] == "incomplete_cost"
    assert board["contribution_status_inc"] == "incomplete_cost"
    assert board["contribution_status_ex"] == "incomplete_cost"
    assert board["expense_data_available"] is True
    assert board["expense_inc"] is None
    assert board["expense_ex"] is None
    assert board["expense_evidence_status"] == "expense_tax_unknown"
    assert csv_data["费用证据状态"] == "expense_tax_unknown"
    assert single["费用证据状态"] == "费用税务口径缺失"
    assert bundled["费用证据状态"] == "费用税务口径缺失"


def test_zero_detail_contract_stays_in_all_four_carriers_and_costs_fail_closed(db):
    batch = _load_complete_contract(db)
    line = db.scalar(select(FMaintenanceLine))
    assert line is not None
    db.delete(line)
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-MARGIN-ZERO-DETAIL",
            bxd_no="BXD-MARGIN-ZERO-DETAIL",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=Decimal("10"),
            amount_ex_tax=Decimal("10"),
            amount_inc_tax=Decimal("11.30"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=_FULL_EXPENSE_COVERAGE,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()
    client = _admin_client(db)

    board, csv_data, single, bundled = _four_carriers(client)

    assert board["contract"] == "XS-MARGIN"
    assert board["order_count"] == 1
    assert board["missing_detail_orders"] == 1
    assert board["lines"] == 0
    assert board["cost_quality"] == "incomplete"
    for field in (
        "actual_cost_inc",
        "actual_cost_ex",
        "estimated_cost_inc",
        "estimated_cost_ex",
        "known_cost_total",
        "spent_parts",
        "parts_cost_inc_tax",
        "parts_cost_ex_tax",
        "parts_gross_profit_inc",
        "parts_gross_profit_ex",
        "parts_gross_margin_inc",
        "parts_gross_margin_ex",
        "contribution_profit_inc",
        "contribution_profit_ex",
        "contribution_margin_inc",
        "contribution_margin_ex",
    ):
        assert board[field] is None
    assert board["revenue_ex"] is not None
    assert board["expense_ex"] == 10.0
    assert csv_data["成本证据状态"] == "incomplete"
    assert csv_data["order_count"] == "1"
    assert csv_data["missing_detail_orders"] == "1"
    assert single["命中维保订单"] == 1
    assert single["有明细订单"] == 0
    assert single["无明细订单"] == 1
    assert single["订单结构完整性"] == "不完整：存在无配件明细订单"
    assert bundled["命中维保订单"] == 1
    assert bundled["有明细订单"] == 0
    assert bundled["无明细订单"] == 1
    assert bundled["订单结构完整性"] == "不完整：存在无配件明细订单"
    assert single["成本完整性"] == "成本不完整，需补数据"
    assert bundled["成本完整性"] == "成本不完整，需补数据"
    _assert_numeric_parity(board, csv_data, single, bundled)

    response = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XS-MARGIN"},
    )
    workbook = load_workbook(io.BytesIO(response.content), data_only=False)
    try:
        parts = workbook["备件明细-氚云"]
        assert parts["A2"].value == "WBDD-MARGIN"
        assert parts["K2"].value == "暂无配件明细"
        assert parts["M2"].value is None
        assert parts["W2"].value == "成本缺失"
    finally:
        workbook.close()


def test_mixed_detail_contract_keeps_known_facts_but_all_margins_fail_closed(
    db,
):
    batch = _load_complete_contract(db)
    original_order = db.scalar(select(FMaintenanceOrder).where(
        FMaintenanceOrder.linked_sales_order_no == "XS-MARGIN",
    ))
    original_line = db.scalar(select(FMaintenanceLine))
    original_sale = db.scalar(select(FSalesOrder).where(
        FSalesOrder.order_no == "XS-MARGIN",
    ))
    assert original_order is not None
    assert original_line is not None
    assert original_sale is not None
    db.add_all([
        FMaintenanceLine(
            raw_line_id="ML-MARGIN-SECOND",
            order_id=original_order.id,
            line_no=2,
            part_id=original_line.part_id,
            pn_std=original_line.pn_std,
            pn_raw=original_line.pn_raw,
            description="第二条已知成本明细",
            qty=original_line.qty,
            return_qty=original_line.return_qty,
            unit_cost=original_line.unit_cost,
            cost_amount=original_line.cost_amount,
            unit_cost_inc_tax=original_line.unit_cost_inc_tax,
            unit_cost_ex_tax=original_line.unit_cost_ex_tax,
            cost_amount_inc_tax=original_line.cost_amount_inc_tax,
            cost_amount_ex_tax=original_line.cost_amount_ex_tax,
            cost_source=original_line.cost_source,
            confidence=original_line.confidence,
            cost_tax_basis=original_line.cost_tax_basis,
            anomaly_flags=original_line.anomaly_flags,
            import_batch_id=batch.id,
        ),
        FMaintenanceOrder(
            raw_order_id="M-MARGIN-ZERO",
            order_no="WBDD-MARGIN-ZERO",
            order_date=date(2026, 3, 11),
            linked_sales_order_no="XS-MARGIN",
            project_raw="双口径毛利项目",
            project_std="双口径毛利项目",
            salesperson=original_order.salesperson,
            maint_start=original_order.maint_start,
            maint_end=original_order.maint_end,
            data_status=original_order.data_status,
            import_batch_id=batch.id,
        ),
        FSalesOrder(
            raw_order_id="S-MARGIN-SECOND-EVIDENCE",
            order_no="XS-MARGIN",
            order_date=original_sale.order_date,
            salesperson=original_sale.salesperson,
            amount_ex_tax=Decimal("1000"),
            tax_rate=original_sale.tax_rate,
            data_status=original_sale.data_status,
            import_batch_id=batch.id,
        ),
        FProjectExpense(
            raw_line_id="EXP-MARGIN-MIXED-1",
            bxd_no="BXD-MARGIN-MIXED-1",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=Decimal("10"),
            amount_ex_tax=Decimal("10"),
            amount_inc_tax=Decimal("11.30"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        FProjectExpense(
            raw_line_id="EXP-MARGIN-MIXED-2",
            bxd_no="BXD-MARGIN-MIXED-2",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 13),
            amount=Decimal("20"),
            amount_ex_tax=Decimal("20"),
            amount_inc_tax=Decimal("22.60"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=_FULL_EXPENSE_COVERAGE,
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    board, csv_data, single, bundled = _four_carriers(_admin_client(db))

    assert board["contract"] == "XS-MARGIN"
    assert board["order_count"] == 2
    assert board["missing_detail_orders"] == 1
    assert board["lines"] == 2
    assert board["cost_quality"] == "incomplete"
    assert board["parts_cost_inc_tax"] == 452.0
    assert board["parts_cost_ex_tax"] == 400.0
    assert board["revenue_inc"] == 1130.0
    assert board["revenue_ex"] == 1000.0
    assert board["expense_inc"] == 33.9
    assert board["expense_ex"] == 30.0
    for field in (
        "parts_gross_profit_inc",
        "parts_gross_profit_ex",
        "parts_gross_margin_inc",
        "parts_gross_margin_ex",
        "contribution_profit_inc",
        "contribution_profit_ex",
        "contribution_margin_inc",
        "contribution_margin_ex",
    ):
        assert board[field] is None
    assert csv_data["order_count"] == "2"
    assert csv_data["missing_detail_orders"] == "1"
    assert csv_data["成本证据状态"] == "incomplete"
    for summary in (single, bundled):
        assert summary["命中维保订单"] == 2
        assert summary["有明细订单"] == 1
        assert summary["无明细订单"] == 1
        assert summary["订单结构完整性"] == "不完整：存在无配件明细订单"
        assert summary["成本完整性"] == "成本不完整，需补数据"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_project_carriers_keep_mixed_detail_orders_and_fail_cost_closed(db):
    batch = _load_complete_contract(db)
    original_order = db.scalar(select(FMaintenanceOrder).where(
        FMaintenanceOrder.linked_sales_order_no == "XS-MARGIN",
    ))
    assert original_order is not None
    db.add(FMaintenanceOrder(
        raw_order_id="M-MARGIN-PROJECT-ZERO",
        order_no="WBDD-MARGIN-PROJECT-ZERO",
        order_date=date(2026, 3, 11),
        linked_sales_order_no="XS-MARGIN",
        project_raw=original_order.project_raw,
        project_std=original_order.project_std,
        salesperson=original_order.salesperson,
        maint_start=original_order.maint_start,
        maint_end=original_order.maint_end,
        data_status=original_order.data_status,
        import_batch_id=batch.id,
    ))
    db.commit()
    client = _admin_client(db)

    projects_response = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    csv_response = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    admin_ctx = _admin_agent_ctx()
    agent_data = tools.dispatch(
        db,
        "get_maintenance_projects",
        {},
        admin_ctx,
    )

    assert projects_response.status_code == 200, projects_response.text
    assert csv_response.status_code == 200, csv_response.text
    project = projects_response.json()["rows"][0]
    assert project["project"] == "双口径毛利项目"
    assert project["order_count"] == 2
    assert project["missing_detail_orders"] == 1
    assert project["structure_complete"] is False
    assert project["lines"] == 1
    assert project["known_cost_total"] == 200.0
    assert project["cost_quality"] == "incomplete"
    assert project["parts_cost_inc_tax"] == 226.0
    assert project["parts_cost_ex_tax"] == 200.0
    assert project["parts_cost_inc_tax_complete"] is False
    assert project["parts_cost_ex_tax_complete"] is False

    csv_data = _csv_row(csv_response)
    assert csv_data["维保订单数"] == "2"
    assert csv_data["无明细订单数"] == "1"
    assert csv_data["订单结构完整性"] == "不完整"
    assert csv_data["成本完整性"] == "成本不完整，需补数据"
    assert csv_data["已知成本参考(混合原值)"] == "200.0"
    assert csv_data["合同额证据状态"] == "完整"

    agent_project = agent_data["rows"][0]
    assert agent_project["order_count"] == project["order_count"]
    assert agent_project["missing_detail_orders"] == project["missing_detail_orders"]
    assert agent_project["structure_complete"] is project["structure_complete"]
    assert agent_project["cost_quality"] == project["cost_quality"]


def test_project_carriers_keep_zero_detail_project_without_fake_zero_cost(db):
    _load_complete_contract(db)
    line = db.scalar(select(FMaintenanceLine))
    assert line is not None
    db.delete(line)
    db.commit()
    client = _admin_client(db)

    projects_response = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    csv_response = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    admin_ctx = _admin_agent_ctx()
    agent_data = tools.dispatch(
        db,
        "get_maintenance_projects",
        {},
        admin_ctx,
    )

    assert projects_response.status_code == 200, projects_response.text
    assert csv_response.status_code == 200, csv_response.text
    project = projects_response.json()["rows"][0]
    assert project["project"] == "双口径毛利项目"
    assert project["order_count"] == 1
    assert project["missing_detail_orders"] == 1
    assert project["structure_complete"] is False
    assert project["lines"] == 0
    assert project["cost_quality"] == "incomplete"
    for field in (
        "actual_cost_inc",
        "actual_cost_ex",
        "estimated_cost_inc",
        "estimated_cost_ex",
        "known_cost_total",
        "cost_inc",
        "cost_ex",
        "cost_total",
        "parts_cost_inc_tax",
        "parts_cost_ex_tax",
    ):
        assert project[field] is None

    csv_data = _csv_row(csv_response)
    assert csv_data["维保订单数"] == "1"
    assert csv_data["无明细订单数"] == "1"
    assert csv_data["订单结构完整性"] == "不完整"
    assert csv_data["已知成本参考(混合原值)"] == ""
    assert csv_data["成本完整性"] == "成本不完整，需补数据"

    agent_project = agent_data["rows"][0]
    assert agent_project["order_count"] == project["order_count"]
    assert agent_project["missing_detail_orders"] == project["missing_detail_orders"]
    assert agent_project["structure_complete"] is project["structure_complete"]
    assert agent_project["known_cost_total"] is None
    assert agent_project["cost_quality"] == "incomplete"


def test_project_carriers_exclude_blank_contract_without_dropping_known_cost(db):
    _load_complete_contract(db)
    order = db.scalar(select(FMaintenanceOrder))
    assert order is not None
    order.linked_sales_order_no = "   "
    db.commit()
    client = _admin_client(db)

    projects_response = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    csv_response = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    admin_ctx = _admin_agent_ctx()
    agent_data = tools.dispatch(
        db,
        "get_maintenance_projects",
        {},
        admin_ctx,
    )

    assert projects_response.status_code == 200, projects_response.text
    assert csv_response.status_code == 200, csv_response.text
    project = projects_response.json()["rows"][0]
    assert project["project"] == "双口径毛利项目"
    assert project["known_cost_total"] == 200.0
    assert project["cost_quality"] == "actual_only"
    assert project["sales_orders"] == []
    assert project["contract_amount"] is None
    assert project["contract_incomplete"] is False

    csv_data = _csv_row(csv_response)
    assert csv_data["已知成本参考(混合原值)"] == "200.0"
    assert csv_data["关联销售订单"] == ""
    assert csv_data["合同额(含税参考)"] == ""
    assert csv_data["合同额证据状态"] == "未关联合同"

    restricted_response = _profit_blind_maintenance_client(db).get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    assert restricted_response.status_code == 200, restricted_response.text
    restricted_csv_data = _csv_row(restricted_response)
    assert restricted_csv_data["合同额(含税参考)"] == ""
    assert restricted_csv_data["合同额证据状态"] == "未关联合同"

    agent_project = agent_data["rows"][0]
    assert agent_project["known_cost_total"] == project["known_cost_total"]
    assert agent_project["sales_orders"] == []
    assert agent_project["contract_amount"] is None


def test_project_carriers_do_not_turn_missing_contract_revenue_into_zero(db):
    _load_complete_contract(db)
    sale = db.scalar(select(FSalesOrder))
    assert sale is not None
    db.delete(sale)
    db.commit()
    client = _admin_client(db)

    projects_response = client.get(
        "/api/maintenance/projects",
        params={"lifecycle": "all"},
    )
    csv_response = client.get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    admin_ctx = _admin_agent_ctx()
    agent_data = tools.dispatch(
        db,
        "get_maintenance_projects",
        {},
        admin_ctx,
    )

    assert projects_response.status_code == 200, projects_response.text
    assert csv_response.status_code == 200, csv_response.text
    project = projects_response.json()["rows"][0]
    assert project["known_cost_total"] == 200.0
    assert project["sales_orders"] == ["XS-MARGIN"]
    assert project["contract_amount"] is None
    assert project["contract_incomplete"] is True

    csv_data = _csv_row(csv_response)
    assert csv_data["已知成本参考(混合原值)"] == "200.0"
    assert csv_data["关联销售订单"] == "XS-MARGIN"
    assert csv_data["合同额(含税参考)"] == ""
    assert csv_data["合同额证据状态"] == "不完整"

    restricted_response = _profit_blind_maintenance_client(db).get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )
    assert restricted_response.status_code == 200, restricted_response.text
    restricted_csv_data = _csv_row(restricted_response)
    assert restricted_csv_data["合同额(含税参考)"] == ""
    assert restricted_csv_data["合同额证据状态"] == "不完整"

    agent_project = agent_data["rows"][0]
    assert agent_project["known_cost_total"] == project["known_cost_total"]
    assert agent_project["sales_orders"] == ["XS-MARGIN"]
    assert agent_project["contract_amount"] is None
    assert agent_project["contract_incomplete"] is True


def test_project_summary_csv_marks_partial_contract_revenue_incomplete_without_dropping_known_amount(
    db,
):
    batch = _load_complete_contract(db)
    original_order = db.scalar(
        select(FMaintenanceOrder).where(
            FMaintenanceOrder.linked_sales_order_no == "XS-MARGIN",
        ),
    )
    assert original_order is not None
    db.add(FMaintenanceOrder(
        raw_order_id="M-MARGIN-PARTIAL-REVENUE",
        order_no="WBDD-MARGIN-PARTIAL-REVENUE",
        order_date=date(2026, 3, 11),
        linked_sales_order_no="XS-MISSING-REVENUE",
        project_raw=original_order.project_raw,
        project_std=original_order.project_std,
        salesperson=original_order.salesperson,
        maint_start=original_order.maint_start,
        maint_end=original_order.maint_end,
        data_status=original_order.data_status,
        import_batch_id=batch.id,
    ))
    db.commit()

    response = _admin_client(db).get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )

    assert response.status_code == 200, response.text
    csv_data = _csv_row(response)
    assert Decimal(csv_data["合同额(含税参考)"]) == Decimal("1130.0")
    assert csv_data["合同额证据状态"] == "不完整"


def test_project_summary_csv_marks_scoped_sales_contract_evidence_restricted(db):
    _load_complete_contract(db)

    response = _scoped_sales_maintenance_client(db).get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )

    assert response.status_code == 200, response.text
    csv_data = _csv_row(response)
    assert csv_data["关联销售订单"] == "XS-MARGIN"
    assert csv_data["合同额(含税参考)"] == ""
    assert csv_data["合同额证据状态"] == "受限"


def test_project_summary_csv_marks_hidden_contract_amount_restricted(db):
    _load_complete_contract(db)

    response = _profit_blind_maintenance_client(db).get(
        "/api/maintenance/export",
        params={"lifecycle": "all"},
    )

    assert response.status_code == 200, response.text
    csv_data = _csv_row(response)
    assert csv_data["关联销售订单"] == "XS-MARGIN"
    assert csv_data["合同额(含税参考)"] == ""
    assert csv_data["合同额证据状态"] == "受限"


def test_blank_contract_values_are_excluded_from_board_csv_single_and_zip(db):
    batch = _load_complete_contract(db)
    for index, contract in enumerate((None, "", "   "), 1):
        db.add(FMaintenanceOrder(
            raw_order_id=f"M-BLANK-CONTRACT-{index}",
            order_no=f"WBDD-BLANK-CONTRACT-{index}",
            order_date=date(2026, 3, 10 + index),
            linked_sales_order_no=contract,
            project_raw=f"空白合同项目{index}",
            project_std=f"空白合同项目{index}",
            maint_end=date(2027, 12, 31),
            data_status="已生效",
            import_batch_id=batch.id,
        ))
    db.commit()
    client = _admin_client(db)

    board_response = client.get(
        "/api/maintenance/board",
        params={"lifecycle": "all"},
    )
    csv_response = client.get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all"},
    )
    blank_single = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "   "},
    )
    bundle_response = client.get("/api/maintenance/export-workbooks")

    assert board_response.status_code == 200, board_response.text
    assert [row["contract"] for row in board_response.json()["rows"]] == [
        "XS-MARGIN",
    ]
    assert _csv_row(csv_response)["合同"] == "XS-MARGIN"
    assert blank_single.status_code == 404
    with ZipFile(io.BytesIO(bundle_response.content)) as archive:
        workbook_members = [
            name for name in archive.namelist()
            if name.startswith("项目工作簿/")
        ]
        manifest = list(csv.DictReader(io.StringIO(
            archive.read("导出清单.csv").decode("utf-8-sig"),
        )))
    assert len(workbook_members) == 1
    assert sum(row["记录类型"] == "已生成" for row in manifest) == 1
    assert sum(row["记录类型"] == "已跳过" for row in manifest) == 3


def test_date_filter_keeps_the_same_scope_across_all_four_carriers(db):
    batch = _load_complete_contract(db)
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-MARGIN-IN-RANGE",
            bxd_no="BXD-MARGIN-IN-RANGE",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=Decimal("10"),
            amount_ex_tax=Decimal("10"),
            amount_inc_tax=Decimal("11.30"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        FProjectExpense(
            raw_line_id="EXP-MARGIN-OUTSIDE-RANGE",
            bxd_no="BXD-MARGIN-OUTSIDE-RANGE",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 4, 12),
            amount=Decimal("20"),
            amount_ex_tax=Decimal("20"),
            amount_inc_tax=Decimal("22.60"),
            tax_basis="ex",
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=date(2026, 3, 31),
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    board, csv_data, single, bundled = _four_carriers(
        _admin_client(db),
        board_params={
            "date_from": "2026-03-01",
            "date_to": "2026-03-31",
        },
    )

    assert board["expense_inc"] == 11.3
    assert board["expense_ex"] == 10.0
    assert board["parts_profit_status_inc"] == "filtered_scope"
    assert board["parts_profit_status_ex"] == "filtered_scope"
    assert board["contribution_status_inc"] == "filtered_scope"
    assert board["contribution_status_ex"] == "filtered_scope"
    assert csv_data["费用证据状态"] == "complete"
    assert single["费用证据状态"] == "完整"
    assert bundled["费用证据状态"] == "完整"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_date_filter_with_null_expense_watermark_fails_closed_in_all_carriers(
    db,
):
    _load_complete_contract(db)
    db.add(MaintenanceContractWorkbookState(
        contract_no="XS-MARGIN",
        revision=1,
        expense_complete_through=None,
        expense_snapshot_complete=True,
    ))
    db.commit()

    board, csv_data, single, bundled = _four_carriers(
        _admin_client(db),
        board_params={
            "date_from": "2026-03-01",
            "date_to": "2026-03-31",
        },
    )

    assert board["expense_data_available"] is False
    assert board["parts_profit_status_inc"] == "filtered_scope"
    assert board["parts_profit_status_ex"] == "filtered_scope"
    assert board["contribution_status_inc"] == "expense_data_unavailable"
    assert board["contribution_status_ex"] == "expense_data_unavailable"
    assert csv_data["费用证据状态"] == "expense_data_unavailable"
    for summary in (single, bundled):
        assert summary["费用证据状态"] == "未就绪（无记录不等于0）"
        assert summary["含税贡献毛利状态"] == "费用数据未就绪"
        assert summary["未税贡献毛利状态"] == "费用数据未就绪"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_date_filter_after_expense_watermark_fails_closed_in_all_carriers(db):
    _load_complete_contract(db)
    db.add(MaintenanceContractWorkbookState(
        contract_no="XS-MARGIN",
        revision=1,
        expense_complete_through=date(2026, 3, 30),
        expense_snapshot_complete=True,
    ))
    db.commit()

    board, csv_data, single, bundled = _four_carriers(
        _admin_client(db),
        board_params={
            "date_from": "2026-03-01",
            "date_to": "2026-03-31",
        },
    )

    assert board["expense_data_available"] is False
    assert board["contribution_status_inc"] == "expense_data_unavailable"
    assert board["contribution_status_ex"] == "expense_data_unavailable"
    assert csv_data["费用证据状态"] == "expense_data_unavailable"
    for summary in (single, bundled):
        assert summary["费用证据状态"] == "未就绪（无记录不等于0）"
        assert summary["含税贡献毛利状态"] == "费用数据未就绪"
        assert summary["未税贡献毛利状态"] == "费用数据未就绪"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_date_filter_preserves_expense_tax_failure_across_all_four_carriers(db):
    batch = _load_complete_contract(db)
    db.add_all([
        FProjectExpense(
            raw_line_id="EXP-MARGIN-DATE-TAX-UNKNOWN",
            bxd_no="BXD-MARGIN-DATE-TAX-UNKNOWN",
            line_no=1,
            linked_sales_order_no="XS-MARGIN",
            data_status="已结束",
            expense_date=date(2026, 3, 12),
            amount=None,
            amount_ex_tax=None,
            amount_inc_tax=None,
            import_batch_id=batch.id,
        ),
        MaintenanceContractWorkbookState(
            contract_no="XS-MARGIN",
            revision=1,
            expense_complete_through=date(2026, 3, 31),
            expense_snapshot_complete=True,
        ),
    ])
    db.commit()

    board, csv_data, single, bundled = _four_carriers(
        _admin_client(db),
        board_params={
            "date_from": "2026-03-01",
            "date_to": "2026-03-31",
        },
    )

    assert board["parts_profit_status_inc"] == "filtered_scope"
    assert board["parts_profit_status_ex"] == "filtered_scope"
    assert board["contribution_status_inc"] == "expense_tax_unknown"
    assert board["contribution_status_ex"] == "expense_tax_unknown"
    assert csv_data["费用证据状态"] == "expense_tax_unknown"
    assert single["费用证据状态"] == "费用税务口径缺失"
    assert bundled["费用证据状态"] == "费用税务口径缺失"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_date_filter_preserves_missing_revenue_across_all_four_carriers(db):
    _load_complete_contract(db)
    sales_order = db.scalar(select(FSalesOrder).where(
        FSalesOrder.order_no == "XS-MARGIN",
    ))
    assert sales_order is not None
    db.delete(sales_order)
    db.add(MaintenanceContractWorkbookState(
        contract_no="XS-MARGIN",
        revision=1,
        expense_complete_through=date(2026, 3, 31),
        expense_snapshot_complete=True,
    ))
    db.commit()

    board, csv_data, single, bundled = _four_carriers(
        _admin_client(db),
        board_params={
            "date_from": "2026-03-01",
            "date_to": "2026-03-31",
        },
    )

    assert board["parts_profit_status_inc"] == "missing_revenue"
    assert board["parts_profit_status_ex"] == "missing_revenue"
    assert board["contribution_status_inc"] == "missing_revenue"
    assert board["contribution_status_ex"] == "missing_revenue"
    assert csv_data["收入证据状态-含税"] == "missing_revenue"
    assert csv_data["收入证据状态-未税"] == "missing_revenue"
    assert single["含税备件毛利状态"] == "合同收入缺失"
    assert single["未税备件毛利状态"] == "合同收入缺失"
    assert bundled["含税备件毛利状态"] == "合同收入缺失"
    assert bundled["未税备件毛利状态"] == "合同收入缺失"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_date_filter_preserves_ambiguous_revenue_across_all_four_carriers(
    db,
    monkeypatch,
):
    _load_complete_contract(db)
    db.add(MaintenanceContractWorkbookState(
        contract_no="XS-MARGIN",
        revision=1,
        expense_complete_through=date(2026, 3, 31),
        expense_snapshot_complete=True,
    ))
    db.commit()

    def ambiguous_revenue(_db, contract_nos):
        return {
            contract_no: maintenance_margin_evidence.RevenueEvidence(
                revenue_ex=Decimal("1000"),
                tax_rate=Decimal("0.13"),
                tax_rate_ambiguous=False,
                ambiguous_inc=True,
                ambiguous_ex=True,
                record_count=2,
                legacy_contract_amount_inc=Decimal("1130"),
            )
            for contract_no in contract_nos
        }

    monkeypatch.setattr(
        maintenance_margin_evidence,
        "load_contract_revenue_evidence",
        ambiguous_revenue,
    )

    board, csv_data, single, bundled = _four_carriers(
        _admin_client(db),
        board_params={
            "date_from": "2026-03-01",
            "date_to": "2026-03-31",
        },
    )

    assert board["parts_profit_status_inc"] == "ambiguous_revenue"
    assert board["parts_profit_status_ex"] == "ambiguous_revenue"
    assert board["contribution_status_inc"] == "ambiguous_revenue"
    assert board["contribution_status_ex"] == "ambiguous_revenue"
    assert csv_data["收入证据状态-含税"] == "ambiguous_revenue"
    assert csv_data["收入证据状态-未税"] == "ambiguous_revenue"
    assert single["含税备件毛利状态"] == "重复合同收入冲突"
    assert single["未税备件毛利状态"] == "重复合同收入冲突"
    assert bundled["含税备件毛利状态"] == "重复合同收入冲突"
    assert bundled["未税备件毛利状态"] == "重复合同收入冲突"
    _assert_numeric_parity(board, csv_data, single, bundled)


def test_contract_profit_csv_matches_board_lifecycle_and_project_search(db):
    _load_complete_contract(db)
    client = _admin_client(db)

    matching = {"lifecycle": "ongoing", "q": "双口径毛利"}
    matching_board = client.get("/api/maintenance/board", params=matching)
    matching_csv = client.get("/api/maintenance/board/export", params=matching)
    ended = {"lifecycle": "ended", "q": "双口径毛利"}
    ended_board = client.get("/api/maintenance/board", params=ended)
    ended_csv = client.get("/api/maintenance/board/export", params=ended)
    missing_query = {"lifecycle": "all", "q": "不存在的项目"}
    missing_board = client.get("/api/maintenance/board", params=missing_query)
    missing_csv = client.get(
        "/api/maintenance/board/export",
        params=missing_query,
    )

    assert matching_board.status_code == 200, matching_board.text
    assert matching_csv.status_code == 200, matching_csv.text
    assert [row["contract"] for row in matching_board.json()["rows"]] == [
        "XS-MARGIN",
    ]
    assert _csv_row(matching_csv)["合同"] == "XS-MARGIN"
    for board_response, csv_response in (
        (ended_board, ended_csv),
        (missing_board, missing_csv),
    ):
        assert board_response.status_code == 200, board_response.text
        assert board_response.json()["rows"] == []
        assert csv_response.status_code == 422
        assert (
            csv_response.json()["detail"]
            == "所选范围内没有可导出的合同详细盈亏数据"
        )


def test_contract_profit_csv_has_explicit_empty_and_access_semantics(db):
    admin_response = _admin_client(db).get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all"},
    )
    anonymous_response = TestClient(app).get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all"},
    )
    no_page_response = _readonly_client(db).get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all"},
    )

    assert admin_response.status_code == 422
    assert (
        admin_response.json()["detail"]
        == "所选范围内没有可导出的合同详细盈亏数据"
    )
    assert anonymous_response.status_code == 401
    assert no_page_response.status_code == 403


def test_contract_profit_carriers_mask_evidence_for_profit_blind_user(db):
    _load_complete_contract(db)
    admin_board_response = _admin_client(db).get(
        "/api/maintenance/board",
        params={"lifecycle": "all"},
    )
    client = _profit_blind_maintenance_client(db)

    board_response = client.get(
        "/api/maintenance/board",
        params={"lifecycle": "all"},
    )
    response = client.get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all"},
    )
    single_workbook_response = client.get(
        "/api/maintenance/export-workbook",
        params={"contract": "XS-MARGIN"},
    )
    bulk_workbook_response = client.get("/api/maintenance/export-workbooks")

    assert admin_board_response.status_code == 200, admin_board_response.text
    assert admin_board_response.json()["rows"][0]["expense_data_available"] is False
    assert board_response.status_code == 200, board_response.text
    board_row = board_response.json()["rows"][0]
    assert board_row["expense_data_available"] is None
    assert board_row["expense_inc"] is None
    assert board_row["expense_ex"] is None
    assert board_row["contribution_status_inc"] is None
    assert board_row["contribution_status_ex"] is None
    assert response.status_code == 200, response.text
    row = _csv_row(response)
    assert row["parts_cost_inc_tax"] == "226.0"
    assert row["parts_cost_ex_tax"] == "200.0"
    for field in set(_WORKBOOK_LABELS) - {
        "parts_cost_inc_tax",
        "parts_cost_ex_tax",
    }:
        assert row[field] == ""
    assert row["成本证据状态"] == "actual_only"
    assert row["成本证据状态-含税"] == "actual_only"
    assert row["成本证据状态-未税"] == "actual_only"
    assert row["收入证据状态-含税"] == "restricted"
    assert row["收入证据状态-未税"] == "restricted"
    assert row["费用证据状态"] == "restricted"
    assert single_workbook_response.status_code == 403
    assert bulk_workbook_response.status_code == 403


def test_contract_profit_csv_masks_cost_and_derived_profit_for_cost_blind_user(db):
    _load_complete_contract(db)

    response = _cost_blind_maintenance_client(db).get(
        "/api/maintenance/board/export",
        params={"lifecycle": "all"},
    )

    assert response.status_code == 200, response.text
    row = _csv_row(response)
    for field in _WORKBOOK_LABELS:
        assert row[field] == ""
    assert row["成本证据状态"] == ""
    assert row["成本证据状态-含税"] == ""
    assert row["成本证据状态-未税"] == ""
    assert row["收入证据状态-含税"] == "restricted"
    assert row["收入证据状态-未税"] == "restricted"
    assert row["费用证据状态"] == "restricted"
