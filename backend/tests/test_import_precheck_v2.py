import io
import time
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Callable

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook as openpyxl_load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from app import config
from app.api import imports as imports_api
from app.auth import hash_password
from app.config import get_settings
from app.etl import pipeline, reader, sheet_selection
from app.main import app
from app.models.inventory import Inventory
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder, FProjectExpense
from app.models.purchase import FPurchaseLine, FPurchaseOrder
from app.models.sales import FSalesLine, FSalesOrder
from app.models.system import SysImportBatch, SysImportJob, SysRawFile, SysUser


_XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PURCHASE_HEADER = [
    "采购单号(必填)",
    "数据ID(不可修改)",
    "明细.数据ID(不可修改)",
    "明细.产品名称(必填)",
    "明细.单价(必填)",
]
_PURCHASE_ROW = ["CGDD-1", "PO-1", "PL-1", "PN-PURCHASE", 100]
_INVENTORY_HEADER = ["产品库存ID", "产品名称(PN)", "库存数量", "仓库"]
_EXPENSE_HEADER = ["报销日期", "费用分类", "报销金额"]


@pytest.fixture()
def import_client(db):
    db.add(SysUser(
        username="precheck-admin",
        role="admin",
        display_name="预检管理员",
        password_hash=hash_password("adminpw"),
    ))
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "precheck-admin", "password": "adminpw"},
    )
    assert login.status_code == 200, login.text
    client.headers.update({"Authorization": f"Bearer {login.json()['token']}"})
    return client


def _workbook_bytes(sheets: list[tuple[str, list[list[object]]]]) -> bytes:
    workbook = Workbook()
    for index, (sheet_name, rows) in enumerate(sheets):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = sheet_name
        for row in rows:
            sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _sparse_workbook_bytes(cell_ref: str, value: object = "x") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "稀疏表"
    sheet[cell_ref] = value
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _rewrite_xlsx_members(
    payload: bytes,
    transforms: dict[str, Callable[[bytes], bytes]],
    renames: dict[str, str] | None = None,
    omit: set[str] | None = None,
) -> bytes:
    output = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(payload)) as source,
        zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target,
    ):
        for member in source.infolist():
            if omit and member.filename in omit:
                continue
            data = source.read(member)
            transform = transforms.get(member.filename)
            if transform is not None:
                data = transform(data)
            target.writestr(
                renames.get(member.filename, member.filename)
                if renames
                else member.filename,
                data,
            )
    return output.getvalue()


def _rename_worksheet_parts(payload: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        worksheet_members = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ]
    renames = {
        name: name.replace("xl/worksheets/sheet", "xl/worksheets/ws", 1)
        for name in worksheet_members
    }

    def rewrite_workbook_rels(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for relationship in root:
            target = relationship.attrib.get("Target", "")
            relationship.set(
                "Target",
                target.replace("/xl/worksheets/sheet", "/xl/worksheets/ws", 1)
                .replace("worksheets/sheet", "worksheets/ws", 1),
            )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/_rels/workbook.xml.rels": rewrite_workbook_rels},
        renames=renames,
    )


def _relocate_worksheets_to_custom_dir(payload: bytes) -> bytes:
    prefix = "xl/worksheets/sheet"
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        worksheet_members = [
            name
            for name in archive.namelist()
            if name.startswith(prefix) and name.endswith(".xml")
        ]
    renames = {
        name: f"xl/custom/ws{name.removeprefix(prefix)}"
        for name in worksheet_members
    }

    def rewrite_workbook_rels(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for relationship in root:
            target = relationship.attrib.get("Target", "")
            relationship.set(
                "Target",
                target.replace("/xl/worksheets/sheet", "/xl/custom/ws", 1)
                .replace("worksheets/sheet", "custom/ws", 1),
            )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def rewrite_content_types(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for override in root:
            part_name = override.attrib.get("PartName", "")
            if part_name.startswith("/xl/worksheets/sheet"):
                override.set(
                    "PartName",
                    part_name.replace("/xl/worksheets/sheet", "/xl/custom/ws", 1),
                )
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={
            "xl/_rels/workbook.xml.rels": rewrite_workbook_rels,
            "[Content_Types].xml": rewrite_content_types,
        },
        renames=renames,
    )


def _share_all_worksheets_with_single_custom_part(payload: bytes) -> bytes:
    prefix = "xl/worksheets/sheet"
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        worksheet_members = [
            name
            for name in archive.namelist()
            if name.startswith(prefix) and name.endswith(".xml")
        ]

    def rewrite_workbook_rels(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for relationship in root:
            if relationship.attrib.get("Type", "").endswith("/worksheet"):
                relationship.set("Target", "/xl/custom/ws1.xml")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    def rewrite_content_types(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for override in list(root):
            part_name = override.attrib.get("PartName", "")
            if part_name == "/xl/worksheets/sheet1.xml":
                override.set("PartName", "/xl/custom/ws1.xml")
            elif part_name.startswith("/xl/worksheets/sheet"):
                root.remove(override)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={
            "xl/_rels/workbook.xml.rels": rewrite_workbook_rels,
            "[Content_Types].xml": rewrite_content_types,
        },
        renames={"xl/worksheets/sheet1.xml": "xl/custom/ws1.xml"},
        omit=set(worksheet_members[1:]),
    )


def _append_malformed_workbook_tail_after_sheets(payload: bytes) -> bytes:
    def rewrite_workbook(data: bytes) -> bytes:
        return data.replace(b"</sheets>", b"<broken</sheets>", 1)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/workbook.xml": rewrite_workbook},
    )


def _rewrite_worksheet_relationship_types(payload: bytes, rel_type: str) -> bytes:
    def rewrite_workbook_rels(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for relationship in root:
            if relationship.attrib.get("Type", "").endswith("/worksheet"):
                relationship.set("Type", rel_type)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/_rels/workbook.xml.rels": rewrite_workbook_rels},
    )


def _append_duplicate_required_relationship(payload: bytes) -> bytes:
    def rewrite_workbook_rels(data: bytes) -> bytes:
        root = ET.fromstring(data)
        relationship = next(
            (
                element
                for element in root
                if element.attrib.get("Type", "").endswith("/worksheet")
            ),
            None,
        )
        assert relationship is not None
        ET.SubElement(root, relationship.tag, relationship.attrib.copy())
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/_rels/workbook.xml.rels": rewrite_workbook_rels},
    )


def _nest_junk_under_first_workbook_sheet(payload: bytes) -> bytes:
    def rewrite_workbook(data: bytes) -> bytes:
        root = ET.fromstring(data)
        sheets = next(
            (element for element in root if element.tag.rsplit("}", 1)[-1] == "sheets"),
            None,
        )
        assert sheets is not None
        assert len(sheets) > 0
        ET.SubElement(sheets[0], "junk")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/workbook.xml": rewrite_workbook},
    )


def _append_unknown_child_under_workbook_sheets(payload: bytes) -> bytes:
    def rewrite_workbook(data: bytes) -> bytes:
        root = ET.fromstring(data)
        sheets = next(
            (element for element in root if element.tag.rsplit("}", 1)[-1] == "sheets"),
            None,
        )
        assert sheets is not None
        ET.SubElement(sheets, "junk")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/workbook.xml": rewrite_workbook},
    )


def _remove_worksheet_dimension(payload: bytes) -> bytes:
    def rewrite_dimension(data: bytes) -> bytes:
        root = ET.fromstring(data)
        dimension = next(
            (
                element
                for element in root
                if element.tag.rsplit("}", 1)[-1] == "dimension"
            ),
            None,
        )
        if dimension is not None:
            root.remove(dimension)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": rewrite_dimension},
    )


def _rewrite_worksheet_dimension_ref(payload: bytes, ref: str) -> bytes:
    def rewrite_dimension(data: bytes) -> bytes:
        root = ET.fromstring(data)
        dimension = next(
            (
                element
                for element in root
                if element.tag.rsplit("}", 1)[-1] == "dimension"
            ),
            None,
        )
        assert dimension is not None
        dimension.set("ref", ref)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": rewrite_dimension},
    )


def _remove_worksheet_dimension_and_refs(payload: bytes) -> bytes:
    def rewrite_dimension_and_refs(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for element in list(root):
            if element.tag.rsplit("}", 1)[-1] == "dimension":
                root.remove(element)
        for element in root.iter():
            if element.tag.rsplit("}", 1)[-1] in {"row", "c"}:
                element.attrib.pop("r", None)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={
            "xl/worksheets/sheet1.xml": rewrite_dimension_and_refs,
        },
    )


def _rewrite_worksheet_row_ref(payload: bytes, row_ref: str) -> bytes:
    def rewrite_row_ref(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for element in list(root):
            if element.tag.rsplit("}", 1)[-1] == "dimension":
                root.remove(element)
        row = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "row"
        )
        row.set("r", row_ref)
        for element in row:
            element.attrib.pop("r", None)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": rewrite_row_ref},
    )


def _rewrite_first_row_child_ref(payload: bytes, cell_ref: str) -> bytes:
    def rewrite_cell_ref(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for element in list(root):
            if element.tag.rsplit("}", 1)[-1] == "dimension":
                root.remove(element)
        row = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "row"
        )
        row[0].set("r", cell_ref)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": rewrite_cell_ref},
    )


def _add_cross_chunk_worksheet_doctype(payload: bytes) -> bytes:
    def add_doctype(data: bytes) -> bytes:
        declaration_at = data.find(b"?>")
        declaration_end = declaration_at + 2 if declaration_at >= 0 else 0
        doctype_offset = 64 * 1024 - 4
        comment_overhead = len(b"<!--") + len(b"-->")
        padding_length = doctype_offset - declaration_end - comment_overhead
        padding = b"<!--" + b"x" * padding_length + b"-->"
        doctype = b"<!DOCTYPE worksheet [<!-- ]/> -->]>"
        rewritten = data[:declaration_end] + padding + doctype + data[declaration_end:]
        assert rewritten[doctype_offset : doctype_offset + 9] == b"<!DOCTYPE"
        return rewritten

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": add_doctype},
    )


def _add_unique_attribute_names(payload: bytes, member_name: str) -> bytes:
    def add_attributes(data: bytes) -> bytes:
        root = ET.fromstring(data)
        for index in range(reader._XML_NAME_LIMIT + 1):
            root.set(f"unique_name_{index}", "")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(payload, transforms={member_name: add_attributes})


def _add_unique_namespace_declarations(payload: bytes, member_name: str) -> bytes:
    def add_namespaces(data: bytes) -> bytes:
        declaration_end = data.find(b"?>") + 2
        root_start = data.find(b"<", declaration_end)
        root_end = data.find(b">", root_start)
        namespaces = b"".join(
            f' xmlns:unique{index}="urn:unique:{index}"'.encode()
            for index in range(reader._XML_NAME_LIMIT + 1)
        )
        return data[:root_end] + namespaces + data[root_end:]

    return _rewrite_xlsx_members(payload, transforms={member_name: add_namespaces})


def _replace_first_cell_with_inline_cdata(payload: bytes, value: str) -> bytes:
    def replace_cell(data: bytes) -> bytes:
        cell_start = data.find(b"<c ")
        cell_end = data.find(b"</c>", cell_start) + len(b"</c>")
        assert cell_start >= 0 and cell_end >= len(b"</c>")
        replacement = (
            b'<c r="A1" t="inlineStr"><is><t><![CDATA['
            + value.encode()
            + b"]]></t></is></c>"
        )
        return data[:cell_start] + replacement + data[cell_end:]

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": replace_cell},
    )


def _add_root_attribute(
    payload: bytes,
    member_name: str,
    name: str,
    value: str,
) -> bytes:
    def add_attribute(data: bytes) -> bytes:
        root = ET.fromstring(data)
        root.set(name, value)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(payload, transforms={member_name: add_attribute})


def _add_nested_xml(payload: bytes, member_name: str, depth: int) -> bytes:
    def add_nesting(data: bytes) -> bytes:
        root = ET.fromstring(data)
        parent = root
        for _ in range(depth - 1):
            parent = ET.SubElement(parent, "level")
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(payload, transforms={member_name: add_nesting})


def _replace_first_cell_with_namespaced_sibling(
    payload: bytes,
    local_name: str,
) -> bytes:
    def replace_cell(data: bytes) -> bytes:
        root = ET.fromstring(data)
        row = next(
            element
            for element in root.iter()
            if element.tag.rsplit("}", 1)[-1] == "row"
        )
        cell = row[0]
        namespace = cell.tag.rsplit("}", 1)[0].lstrip("{")
        cell.tag = f"{{{namespace}}}{local_name}"
        for element in list(root):
            if element.tag.rsplit("}", 1)[-1] == "dimension":
                root.remove(element)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)

    return _rewrite_xlsx_members(
        payload,
        transforms={"xl/worksheets/sheet1.xml": replace_cell},
    )


def _precheck(client: TestClient, payload: bytes, filename: str = "input.xlsx"):
    response = client.post(
        "/api/import/precheck",
        files=[("files", (filename, payload, _XLSX_CONTENT_TYPE))],
    )
    assert response.status_code == 200, response.text
    return response.json()


def _sheet(result: dict, name: str) -> dict:
    return next(sheet for sheet in result["sheets"] if sheet["sheet_name"] == name)


def _wait_for_job(client: TestClient, job_id: int) -> dict:
    for _ in range(200):
        response = client.get(f"/api/import/jobs/{job_id}")
        assert response.status_code == 200, response.text
        job = response.json()
        if job["status"] != "processing":
            return job
        time.sleep(0.025)
    pytest.fail("批量导入作业未在 5 秒内结束")


def _fact_counts(db) -> dict[str, int]:
    models = (
        Inventory,
        FPurchaseOrder,
        FPurchaseLine,
        FSalesOrder,
        FSalesLine,
        FMaintenanceOrder,
        FMaintenanceLine,
        FProjectExpense,
    )
    return {
        model.__tablename__: db.scalar(select(func.count()).select_from(model))
        for model in models
    }


def test_precheck_single_purchase_keeps_v1_fields_and_adds_v2_contract(db, import_client):
    payload = _workbook_bytes([("采购", [_PURCHASE_HEADER, _PURCHASE_ROW])])
    facts_before = _fact_counts(db)

    result = _precheck(import_client, payload)
    file_result = result["files"][0]

    assert {
        "filename", "file_type", "ok", "missing_price", "warning",
        "can_import", "severity", "selected_sheets", "sheets", "issues",
    } <= file_result.keys()
    assert {"any_warning", "missing_price_any", "has_errors", "can_import_all"} <= result.keys()
    assert file_result["file_type"] == "purchase"
    assert file_result["ok"] is True
    assert file_result["missing_price"] is False
    assert file_result["warning"] is None
    assert file_result["can_import"] is True
    assert file_result["severity"] == "info"
    assert file_result["selected_sheets"] == ["采购"]
    assert file_result["issues"] == []
    assert file_result["sheets"] == [{
        "sheet_name": "采购",
        "detected_type": "purchase",
        "action": "selected",
        "header_row": 1,
        "data_rows": 1,
        "duplicate_headers": [],
        "issues": [],
    }]
    assert result["any_warning"] is False
    assert result["missing_price_any"] is False
    assert result["has_errors"] is False
    assert result["can_import_all"] is True
    assert db.scalar(select(func.count()).select_from(SysImportJob)) == 0
    assert db.scalar(select(func.count()).select_from(SysImportBatch)) == 0
    assert db.scalar(select(func.count()).select_from(SysRawFile)) == 0
    assert _fact_counts(db) == facts_before


def test_precheck_recognizes_sales_with_bare_business_type(import_client):
    payload = _workbook_bytes([(
        "销售",
        [
            [
                "订单编号(必填)",
                "业务类型",
                "数据ID(不可修改)",
                "订单明细.数据ID(不可修改)",
                "订单明细.产品名称",
                "订单明细.单价",
            ],
            ["XSDD-1", "维保", "SO-1", "SL-1", "PN-SALES", 120],
        ],
    )])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["file_type"] == "sales"
    assert file_result["selected_sheets"] == ["销售"]
    assert _sheet(file_result, "销售")["detected_type"] == "sales"


def test_precheck_missing_price_is_nonblocking_warning(import_client):
    payload = _workbook_bytes([(
        "采购无价格",
        [
            _PURCHASE_HEADER[:-1],
            _PURCHASE_ROW[:-1],
        ],
    )])

    result = _precheck(import_client, payload)
    file_result = result["files"][0]
    sheet = _sheet(file_result, "采购无价格")

    assert file_result["missing_price"] is True
    assert file_result["ok"] is False
    assert file_result["severity"] == "warning"
    assert file_result["can_import"] is True
    assert result["has_errors"] is False
    assert result["can_import_all"] is True
    assert any(issue["code"] == "missing_price_columns" for issue in sheet["issues"])


def test_precheck_all_unrecognized_is_file_error(import_client):
    payload = _workbook_bytes([
        ("说明", [["说明", "备注"], ["只用于阅读", "不导入"]]),
        ("附件", [["名称", "内容"], ["附件一", "文本"]]),
    ])

    result = _precheck(import_client, payload)
    file_result = result["files"][0]

    assert file_result["file_type"] is None
    assert file_result["selected_sheets"] == []
    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["ok"] is False
    assert any(issue["code"] == "no_recognized_sheet" for issue in file_result["issues"])
    assert {sheet["action"] for sheet in file_result["sheets"]} == {"ignored_unrecognized"}
    assert all(sheet["issues"][0]["severity"] == "info" for sheet in file_result["sheets"])
    assert result["has_errors"] is True
    assert result["can_import_all"] is False


def test_precheck_selects_first_recognized_and_marks_later_sheets(import_client):
    payload = _workbook_bytes([
        ("封面", [["标题"], ["月度汇总"]]),
        ("采购", [_PURCHASE_HEADER, _PURCHASE_ROW]),
        (
            "销售",
            [
                ["订单编号(必填)", "业务类型", "订单明细.产品名称", "订单明细.单价"],
                ["XSDD-2", "销售", "PN-SALES", 130],
            ],
        ),
    ])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["file_type"] == "purchase"
    assert file_result["selected_sheets"] == ["采购"]
    assert _sheet(file_result, "封面")["action"] == "ignored_unrecognized"
    assert _sheet(file_result, "采购")["action"] == "selected"
    ignored = _sheet(file_result, "销售")
    assert ignored["action"] == "ignored_recognized"
    assert any(issue["code"] == "sheet_ignored_recognized" for issue in ignored["issues"])
    assert file_result["can_import"] is True


def test_precheck_project_workbook_selects_only_expense(import_client):
    maintenance_header = [
        "数据ID(不可修改)",
        "需求单号",
        "制单日期",
        "需求类型",
        "需求明细.数据ID(不可修改)",
        "需求明细.需供货产品",
        "需求明细.需求数量",
    ]
    payload = _workbook_bytes([
        ("备件明细", [maintenance_header, ["M-1", "WBDD-1", "2026-07-01", "报修", "ML-1", "PN-1", 1]]),
        ("报销明细", [_EXPENSE_HEADER, ["2026-07-01", "快递", 20]]),
        ("填写说明", [["说明"], ["只读"]]),
    ])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["file_type"] == "workbook"
    assert file_result["selected_sheets"] == ["报销明细"]
    assert _sheet(file_result, "备件明细")["action"] == "ignored_recognized"
    assert _sheet(file_result, "报销明细")["action"] == "selected"
    assert _sheet(file_result, "填写说明")["action"] == "ignored_unrecognized"
    assert file_result["can_import"] is True


def test_precheck_single_expense_sheet_keeps_expense_file_type(import_client):
    payload = _workbook_bytes([
        ("报销明细", [_EXPENSE_HEADER, ["2026-07-01", "快递", 20]]),
        ("填写说明", [["说明"], ["只读"]]),
    ])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["file_type"] == "expense"
    assert file_result["selected_sheets"] == ["报销明细"]
    assert _sheet(file_result, "报销明细")["action"] == "selected"
    assert _sheet(file_result, "填写说明")["action"] == "ignored_unrecognized"


def test_precheck_selects_all_expense_sheets(import_client):
    payload = _workbook_bytes([
        ("报销一", [_EXPENSE_HEADER, ["2026-07-01", "快递", 20]]),
        ("采购", [_PURCHASE_HEADER, _PURCHASE_ROW]),
        ("报销二", [_EXPENSE_HEADER, ["2026-07-02", "交通", 30]]),
    ])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["file_type"] == "workbook"
    assert file_result["selected_sheets"] == ["报销一", "报销二"]
    assert _sheet(file_result, "报销一")["action"] == "selected"
    assert _sheet(file_result, "报销二")["action"] == "selected"
    assert _sheet(file_result, "采购")["action"] == "ignored_recognized"


def test_precheck_selected_duplicate_headers_are_blocking_error(import_client):
    duplicate_header = [
        "采购单号(必填)",
        "明细.产品名称(必填)",
        "明细.产品名称(必填)",
        "明细.单价(必填)",
    ]
    payload = _workbook_bytes([
        ("重复采购", [duplicate_header, ["CGDD-1", "PN-A", "PN-B", 100]]),
    ])

    result = _precheck(import_client, payload)
    file_result = result["files"][0]
    sheet = _sheet(file_result, "重复采购")

    assert sheet["duplicate_headers"] == ["明细.产品名称(必填)"]
    assert any(
        issue["severity"] == "error" and issue["code"] == "duplicate_headers"
        for issue in sheet["issues"]
    )
    assert file_result["can_import"] is False
    assert file_result["severity"] == "error"
    assert result["has_errors"] is True


def test_precheck_ignored_duplicate_headers_do_not_block(import_client):
    duplicate_header = [
        "采购单号(必填)",
        "明细.产品名称(必填)",
        "明细.产品名称(必填)",
    ]
    payload = _workbook_bytes([
        ("库存", [_INVENTORY_HEADER, ["INV-1", "PN-A", 1, "总仓"]]),
        ("重复采购", [duplicate_header, ["CGDD-1", "PN-A", "PN-B"]]),
    ])

    result = _precheck(import_client, payload)
    file_result = result["files"][0]
    ignored = _sheet(file_result, "重复采购")

    assert ignored["action"] == "ignored_recognized"
    assert any(
        issue["severity"] == "warning" and issue["code"] == "duplicate_headers_ignored"
        for issue in ignored["issues"]
    )
    assert file_result["severity"] == "warning"
    assert file_result["can_import"] is True
    assert result["has_errors"] is False
    assert result["can_import_all"] is True


def test_precheck_legacy_warning_prefers_error_over_earlier_warning(import_client):
    payload = _workbook_bytes([
        ("采购", [_PURCHASE_HEADER, _PURCHASE_ROW]),
        (
            "报销明细",
            [["报销日期", "报销金额", "报销金额"], ["2026-07-01", 20, 30]],
        ),
    ])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert "重复非空表头" in file_result["warning"]
    assert "不会导入" not in file_result["warning"]


def test_precheck_corrupt_workbook_is_error(import_client):
    result = _precheck(import_client, b"not-an-xlsx", "broken.xlsx")
    file_result = result["files"][0]

    assert file_result["sheets"] == []
    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "invalid_workbook"


def test_precheck_oversized_row_count_is_error(import_client, monkeypatch):
    monkeypatch.setattr(config, "IMPORT_MAX_ROWS", 3)
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("超行文件不应进入 pandas 全量读取"),
    )
    payload = _workbook_bytes([(
        "采购",
        [
            _PURCHASE_HEADER,
            _PURCHASE_ROW,
            ["CGDD-2", "PO-2", "PL-2", "PN-2", 100],
            ["CGDD-3", "PO-3", "PL-3", "PN-3", 100],
        ],
    )])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "row_limit_exceeded"


def test_precheck_upload_size_limit_is_structured_error(import_client, monkeypatch):
    monkeypatch.setattr(imports_api, "MAX_UPLOAD_MB", 0)
    monkeypatch.setattr(
        imports_api.import_precheck,
        "inspect_file",
        lambda *_args, **_kwargs: pytest.fail("超大文件不应进入工作簿解析"),
    )
    payload = _workbook_bytes([("采购", [_PURCHASE_HEADER, _PURCHASE_ROW])])

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "file_too_large"


def test_single_upload_size_limit_returns_http_413(import_client, monkeypatch):
    monkeypatch.setattr(imports_api, "MAX_UPLOAD_MB", 0)
    payload = _workbook_bytes([("采购", [_PURCHASE_HEADER, _PURCHASE_ROW])])

    response = import_client.post(
        "/api/import/upload",
        files={"file": ("too-large.xlsx", payload, _XLSX_CONTENT_TYPE)},
    )

    assert response.status_code == 413


def test_batch_upload_size_limit_returns_http_413(db, import_client, monkeypatch):
    monkeypatch.setattr(imports_api, "MAX_UPLOAD_MB", 0)
    payload = _workbook_bytes([("采购", [_PURCHASE_HEADER, _PURCHASE_ROW])])

    response = import_client.post(
        "/api/import/upload-batch",
        files=[("files", ("too-large.xlsx", payload, _XLSX_CONTENT_TYPE))],
    )

    assert response.status_code == 413
    assert db.scalar(select(func.count()).select_from(SysImportJob)) == 0
    assert db.scalar(select(func.count()).select_from(SysImportBatch)) == 0
    assert db.scalar(select(func.count()).select_from(SysRawFile)) == 0


@pytest.mark.parametrize(
    ("limit_name", "limit_value", "expected_code"),
    [
        ("IMPORT_XLSX_MAX_MEMBERS", 0, "xlsx_too_many_members"),
        (
            "IMPORT_XLSX_MAX_UNCOMPRESSED_BYTES",
            0,
            "xlsx_uncompressed_size_exceeded",
        ),
        (
            "IMPORT_XLSX_MAX_COMPRESSION_RATIO",
            0.0,
            "xlsx_compression_ratio_exceeded",
        ),
    ],
)
def test_precheck_xlsx_zip_limits_run_before_openpyxl_and_pandas(
    import_client,
    monkeypatch,
    limit_name,
    limit_value,
    expected_code,
):
    payload = _workbook_bytes([("采购", [_PURCHASE_HEADER, _PURCHASE_ROW])])
    monkeypatch.setattr(config, "IMPORT_XLSX_MAX_MEMBERS", 10_000)
    monkeypatch.setattr(config, "IMPORT_XLSX_MAX_UNCOMPRESSED_BYTES", 512 * 1024 * 1024)
    monkeypatch.setattr(config, "IMPORT_XLSX_MAX_COMPRESSION_RATIO", 200.0)
    monkeypatch.setattr(config, limit_name, limit_value)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("ZIP 超限文件不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("ZIP 超限文件不应进入 pandas"),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == expected_code


@pytest.mark.parametrize(
    "member_name",
    [
        "xl/_rels/workbook.xml.rels",
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
    ],
)
def test_inspect_workbook_rejects_too_many_xml_names_before_pandas(
    monkeypatch,
    tmp_path,
    member_name,
):
    payload = _add_unique_attribute_names(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        member_name,
    )
    path = tmp_path / "too-many-xml-names.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("XML 名称超限文件不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


@pytest.mark.parametrize(
    "member_name",
    [
        "xl/_rels/workbook.xml.rels",
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
    ],
)
def test_inspect_workbook_rejects_too_many_xml_namespaces_before_excel_readers(
    monkeypatch,
    tmp_path,
    member_name,
):
    payload = _add_unique_namespace_declarations(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        member_name,
    )
    path = tmp_path / "too-many-xml-namespaces.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "XML namespace 超限文件不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("XML namespace 超限文件不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


@pytest.mark.parametrize(
    "namespace_declaration",
    [
        f'xmlns:{"p" * (reader._XML_NAME_LENGTH_LIMIT + 1)}="urn:test"',
        f'xmlns:p="{"u" * (reader._XML_NAME_LENGTH_LIMIT + 1)}"',
    ],
)
def test_xml_namespace_prefix_and_uri_have_length_limits(namespace_declaration):
    xml = f"<root {namespace_declaration} />".encode()

    with pytest.raises(reader.ReaderError) as exc_info:
        list(reader._safe_xml_iterparse(io.BytesIO(xml)))

    assert exc_info.value.code == "invalid_workbook"


@pytest.mark.parametrize(
    "xml_text",
    [
        '<?xml version="1.0" encoding="UTF-16"?><root><child /></root>',
        '<?xml version="1.0" encoding="UTF-16"?><root><![CDATA['
        + "x" * 70_000
        + "]]></root>",
        '<?xml version="1.0" encoding="UTF-16"?>'
        '<!DOCTYPE root [<!ENTITY expanded "value">]><root>&expanded;</root>',
    ],
)
def test_xml_scanner_rejects_utf16_before_elementtree_emits_events(
    monkeypatch,
    xml_text,
):
    def fail_if_elementtree_can_emit_events(source, events):
        source.read()
        pytest.fail("UTF-16 XML must be rejected before ElementTree emits events")
        yield from ()

    monkeypatch.setattr(reader.ET, "iterparse", fail_if_elementtree_can_emit_events)

    with pytest.raises(reader.ReaderError) as exc_info:
        list(reader._safe_xml_iterparse(io.BytesIO(xml_text.encode("utf-16"))))

    assert exc_info.value.code == "invalid_workbook"


def test_xml_scanner_accepts_utf8_bom():
    xml = b'\xef\xbb\xbf<?xml version="1.0" encoding="UTF-8"?><root><child /></root>'

    events = list(reader._safe_xml_iterparse(io.BytesIO(xml)))

    assert [element.tag for event, element in events if event == "start"] == [
        "root",
        "child",
    ]


@pytest.mark.parametrize(
    ("attribute_name", "attribute_value"),
    [
        ("n" * (reader._XML_NAME_LENGTH_LIMIT + 1), ""),
        (
            "oversized_markup",
            "x" * (reader._XML_MARKUP_LENGTH_LIMIT + 1),
        ),
    ],
)
def test_inspect_workbook_rejects_oversized_xml_names_and_markup(
    monkeypatch,
    tmp_path,
    attribute_name,
    attribute_value,
):
    payload = _add_root_attribute(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        "xl/workbook.xml",
        attribute_name,
        attribute_value,
    )
    path = tmp_path / "oversized-xml-token.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(config, "IMPORT_XLSX_MAX_COMPRESSION_RATIO", 10_000.0)
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("XML token 超限文件不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


def test_xml_markup_limit_does_not_treat_comment_quotes_as_start_tag_quotes():
    xml = b"<root><!-- don't --><padding>" + b"x" * 70_000 + b"</padding></root>"

    events = list(reader._safe_xml_iterparse(io.BytesIO(xml)))

    assert [event for event, _element in events].count("start") == 2


def test_xml_markup_scanner_preserves_states_across_small_chunks():
    class SmallChunkReader(io.BytesIO):
        def read(self, size=-1):
            return super().read(min(size, 2))

    xml = (
        b'<?xml version="1.0"?>'
        b'<root quoted="a>b"><!-- ]/> --><![CDATA[<not-a-tag>]]>'
        b"<child /></root>"
    )

    events = list(reader._safe_xml_iterparse(SmallChunkReader(xml)))

    assert [element.tag for event, element in events if event == "start"] == [
        "root",
        "child",
    ]


def test_xml_markup_scanner_handles_many_short_tags_before_chunk_tail_comment():
    tag_count = 10_000
    chunk_prefix = b"<root>" + b"<n/>" * tag_count
    padding = b"x" * (
        reader._XML_MARKUP_LENGTH_LIMIT - len(chunk_prefix) - len(b"<!--")
    )
    xml = chunk_prefix + padding + b"<!--tail comment--><child /></root>"

    events = list(reader._safe_xml_iterparse(io.BytesIO(xml)))
    start_tags = [element.tag for event, element in events if event == "start"]

    assert len(start_tags) == tag_count + 2
    assert start_tags[0] == "root"
    assert start_tags[-1] == "child"


@pytest.mark.parametrize(
    ("markup", "chunk_size"),
    [
        (b"<!-- text containing <!DOCTYPE worksheet -->", 64 * 1024),
        (b"<!-- text containing <!DOCTYPE worksheet -->", 2),
        (b"<![CDATA[text containing <!DOCTYPE worksheet]]>", 64 * 1024),
        (b"<![CDATA[text containing <!DOCTYPE worksheet]]>", 2),
        (b"<?target text containing <!DOCTYPE worksheet?>", 64 * 1024),
        (b"<?target text containing <!DOCTYPE worksheet?>", 2),
    ],
)
def test_xml_markup_scanner_accepts_doctype_text_inside_special_markup(
    markup, chunk_size
):
    class SmallChunkReader(io.BytesIO):
        def read(self, size=-1):
            return super().read(min(size, chunk_size))

    xml = b"<root>" + markup + b"<child /></root>"

    events = list(reader._safe_xml_iterparse(SmallChunkReader(xml)))

    assert [element.tag for event, element in events if event == "start"] == [
        "root",
        "child",
    ]


def test_long_inline_cdata_is_readable_and_not_counted_as_start_tag_markup(tmp_path):
    value = "中" * 30_000
    payload = _replace_first_cell_with_inline_cdata(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        value,
    )
    path = tmp_path / "long-cdata.xlsx"
    path.write_bytes(payload)

    workbook = openpyxl_load_workbook(path, read_only=True, data_only=True)
    try:
        assert workbook.active["A1"].value == value
    finally:
        workbook.close()

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert reader._scan_worksheet_bounds(
            archive,
            "xl/worksheets/sheet1.xml",
            sheet_name="Sheet 1",
            row_limit=config.IMPORT_MAX_ROWS,
            row_total_so_far=0,
            column_limit=config.IMPORT_XLSX_MAX_COLUMNS,
            declared_cell_limit=config.IMPORT_XLSX_MAX_DECLARED_CELLS,
            declared_cells_so_far=0,
        ) == (1, 1)

    assert (
        reader.inspect_workbook(str(path), load_data=False)[0].sheet_name == "Sheet 1"
    )


def test_inspect_workbook_rejects_cross_chunk_doctype_before_excel_readers(
    monkeypatch,
    tmp_path,
):
    payload = _add_cross_chunk_worksheet_doctype(
        _workbook_bytes([("Sheet 1", [["row-1"]])])
    )
    path = tmp_path / "doctype.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("DOCTYPE 文件不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("DOCTYPE 文件不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


@pytest.mark.parametrize("cell_ref", ["", "A1_1", "A300_001", "A1 "])
def test_inspect_workbook_rejects_invalid_explicit_cell_ref_before_excel_readers(
    monkeypatch,
    tmp_path,
    cell_ref,
):
    payload = _rewrite_first_row_child_ref(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        cell_ref,
    )
    path = tmp_path / "invalid-cell-ref.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("非法单元格坐标不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("非法单元格坐标不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


def test_workbook_sheet_parts_allows_exactly_100_sheets():
    payload = _workbook_bytes([
        (f"Sheet {index}", [[f"row-{index}"]]) for index in range(1, 101)
    ])

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert len(reader._workbook_sheet_parts(archive)) == 100


@pytest.mark.parametrize(
    "member_name",
    ["xl/workbook.xml", "xl/worksheets/sheet1.xml"],
)
def test_real_workbook_and_worksheet_xml_allow_depth_64(member_name):
    payload = _add_nested_xml(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        member_name,
        reader._XML_DEPTH_LIMIT,
    )

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        if member_name == "xl/workbook.xml":
            assert reader._workbook_sheet_refs(archive) == [
                reader._WorkbookSheetRef("Sheet 1", "rId1")
            ]
        else:
            assert reader._scan_worksheet_bounds(
                archive,
                member_name,
                sheet_name="Sheet 1",
                row_limit=config.IMPORT_MAX_ROWS,
                row_total_so_far=0,
                column_limit=config.IMPORT_XLSX_MAX_COLUMNS,
                declared_cell_limit=config.IMPORT_XLSX_MAX_DECLARED_CELLS,
                declared_cells_so_far=0,
            ) == (1, 1)


@pytest.mark.parametrize(
    "member_name",
    ["xl/workbook.xml", "xl/worksheets/sheet1.xml"],
)
def test_real_workbook_and_worksheet_xml_reject_depth_65(member_name):
    payload = _add_nested_xml(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        member_name,
        reader._XML_DEPTH_LIMIT + 1,
    )

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        with pytest.raises(reader.ReaderError) as exc_info:
            if member_name == "xl/workbook.xml":
                reader._workbook_sheet_refs(archive)
            else:
                reader._scan_worksheet_bounds(
                    archive,
                    member_name,
                    sheet_name="Sheet 1",
                    row_limit=config.IMPORT_MAX_ROWS,
                    row_total_so_far=0,
                    column_limit=config.IMPORT_XLSX_MAX_COLUMNS,
                    declared_cell_limit=config.IMPORT_XLSX_MAX_DECLARED_CELLS,
                    declared_cells_so_far=0,
                )

    assert exc_info.value.code == "invalid_workbook"


def test_precheck_rejects_too_many_worksheets_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _workbook_bytes([
        (f"Sheet {index}", [[f"row-{index}"]]) for index in range(1, 102)
    ])
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("超工作表文件不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("超工作表文件不应进入 pandas"),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "worksheet_limit_exceeded"


def test_precheck_counts_renamed_worksheet_parts_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _rename_worksheet_parts(_workbook_bytes([
        (f"Sheet {index}", [[f"row-{index}"]]) for index in range(1, 102)
    ]))
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(workbook.worksheets) == 101
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "重命名 worksheet parts 的超页工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "重命名 worksheet parts 的超页工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "worksheet_limit_exceeded"


def test_precheck_counts_custom_directory_worksheet_relationships_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _relocate_worksheets_to_custom_dir(_workbook_bytes([
        (f"Sheet {index}", [[f"row-{index}"]]) for index in range(1, 102)
    ]))
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(workbook.worksheets) == 101
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "自定义目录 worksheet parts 的超页工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "自定义目录 worksheet parts 的超页工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "worksheet_limit_exceeded"


def test_precheck_counts_logical_sheets_when_worksheets_share_one_physical_part(
    import_client, monkeypatch
):
    payload = _share_all_worksheets_with_single_custom_part(_workbook_bytes([
        (f"Sheet {index}", [[f"row-{index}"]]) for index in range(1, 102)
    ]))
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(workbook.worksheets) == 101
        assert workbook.worksheets[0]["A1"].value == "row-1"
        assert workbook.worksheets[-1]["A1"].value == "row-1"
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "共享 worksheet part 的超页工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "共享 worksheet part 的超页工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "worksheet_limit_exceeded"


def test_precheck_stops_at_logical_sheet_limit_before_malformed_workbook_tail(
    import_client, monkeypatch
):
    payload = _append_malformed_workbook_tail_after_sheets(
        _workbook_bytes([
            (f"Sheet {index}", [[f"row-{index}"]]) for index in range(1, 102)
        ])
    )
    monkeypatch.setattr(
        reader,
        "_worksheet_relationship_targets",
        lambda *_args, **_kwargs: pytest.fail(
            "第 101 个逻辑 sheet 超限时不应解析 workbook relationships"
        ),
    )
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("超工作表文件不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("超工作表文件不应进入 pandas"),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "worksheet_limit_exceeded"


def test_precheck_rejects_unknown_internal_sheet_relationships_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _rewrite_worksheet_relationship_types(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        "urn:custom:not-a-worksheet",
    )
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(workbook.worksheets) == 1
        assert workbook.worksheets[0]["A1"].value == "row-1"
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("非法关系工作簿不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("非法关系工作簿不应进入 pandas"),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "invalid_workbook"


def test_precheck_rejects_duplicate_required_sheet_relationship_ids_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _append_duplicate_required_relationship(
        _workbook_bytes([("Sheet 1", [["row-1"]])])
    )
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(workbook.worksheets) == 1
        assert workbook.worksheets[0]["A1"].value == "row-1"
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "重复 required relationship Id 的工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "重复 required relationship Id 的工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "invalid_workbook"


@pytest.mark.parametrize(
    "scenario",
    ["unknown depth-2 sibling", "nested relationship child"],
)
def test_worksheet_relationship_targets_rejects_invalid_relationship_structure_early(
    monkeypatch,
    scenario,
):
    payload = _workbook_bytes([("Sheet 1", [["row-1"]])])

    def fake_iterparse(_source, events=("end",)):
        root = ET.Element("Relationships")
        yield "start", root
        if scenario == "unknown depth-2 sibling":
            yield "start", ET.SubElement(root, "junk")
            pytest.fail("非法 relationships depth=2 sibling 后不应继续解析")

        relationship = ET.SubElement(
            root,
            "Relationship",
            {
                "Id": "rId-target",
                "Type": (
                    "http://schemas.openxmlformats.org/officeDocument/2006/"
                    "relationships/worksheet"
                ),
                "Target": "worksheets/sheet1.xml",
            },
        )
        yield "start", relationship
        yield "start", ET.SubElement(relationship, "nested-child")
        pytest.fail("Relationship 嵌套 child 后不应继续解析")

    monkeypatch.setattr(reader.ET, "iterparse", fake_iterparse)

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        with pytest.raises(reader.ReaderError, match="无法按 .xlsx 解析"):
            reader._worksheet_relationship_targets(
                archive,
                set(archive.namelist()),
                [reader._WorkbookSheetRef("Sheet 1", "rId-target")],
            )


def test_worksheet_relationship_targets_clears_processed_relationship_children(
    monkeypatch,
):
    payload = _workbook_bytes([("Sheet 1", [["row-1"]])])
    captured_root = None

    def fake_iterparse(_source, events=("end",)):
        nonlocal captured_root
        captured_root = ET.Element("Relationships")

        if "start" in events:
            yield "start", captured_root

        for index in range(1000):
            relationship = ET.SubElement(
                captured_root,
                "Relationship",
                {
                    "Id": f"unused-{index}",
                    "Type": (
                        "http://schemas.openxmlformats.org/officeDocument/2006/"
                        "relationships/worksheet"
                    ),
                    "Target": "worksheets/sheet1.xml",
                },
            )
            if "start" in events:
                yield "start", relationship
            if "end" in events:
                yield "end", relationship

        relationship = ET.SubElement(
            captured_root,
            "Relationship",
            {
                "Id": "rId-target",
                "Type": (
                    "http://schemas.openxmlformats.org/officeDocument/2006/"
                    "relationships/worksheet"
                ),
                "Target": "worksheets/sheet1.xml",
            },
        )
        if "start" in events:
            yield "start", relationship
        if "end" in events:
            yield "end", relationship
            yield "end", captured_root

    monkeypatch.setattr(reader.ET, "iterparse", fake_iterparse)

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        rel_targets = reader._worksheet_relationship_targets(
            archive,
            set(archive.namelist()),
            [reader._WorkbookSheetRef("Sheet 1", "rId-target")],
        )

    assert rel_targets == {"rId-target": "xl/worksheets/sheet1.xml"}
    assert captured_root is not None
    assert len(captured_root) == 0


def test_workbook_sheet_refs_clears_processed_workbook_children(monkeypatch):
    payload = _workbook_bytes([("Sheet 1", [["row-1"]])])
    captured_root = None

    def fake_iterparse(_source, events=("end",)):
        nonlocal captured_root
        captured_root = ET.Element("workbook")
        yield "start", captured_root

        for _ in range(1000):
            metadata = ET.SubElement(captured_root, "metadata")
            yield "start", metadata
            yield "end", metadata

        sheets = ET.SubElement(captured_root, "sheets")
        yield "start", sheets
        sheet = ET.SubElement(
            sheets,
            "sheet",
            {
                "name": "Sheet 1",
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id": "rId1",
            },
        )
        yield "start", sheet
        yield "end", sheet
        yield "end", sheets
        yield "end", captured_root

    monkeypatch.setattr(reader.ET, "iterparse", fake_iterparse)

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        sheet_refs = reader._workbook_sheet_refs(archive)

    assert sheet_refs == [reader._WorkbookSheetRef("Sheet 1", "rId1")]
    assert captured_root is not None
    assert len(captured_root) == 0


def test_precheck_rejects_nested_workbook_sheet_nodes_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _nest_junk_under_first_workbook_sheet(
        _workbook_bytes([("Sheet 1", [["row-1"]])])
    )

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "sheet 下嵌套 junk 的工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "sheet 下嵌套 junk 的工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "invalid_workbook"


def test_precheck_rejects_unknown_workbook_sheets_children_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _append_unknown_child_under_workbook_sheets(
        _workbook_bytes([("Sheet 1", [["row-1"]])])
    )

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "sheets 下出现未知 child 的工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "sheets 下出现未知 child 的工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "invalid_workbook"


def test_scan_worksheet_bounds_rejects_excessive_xml_depth_before_scanning_cells(
    monkeypatch,
):
    payload = _workbook_bytes([("Sheet 1", [["row-1"]])])

    def fake_iterparse(_source, events=("end",)):
        for depth in range(reader._XML_DEPTH_LIMIT):
            yield "start", ET.Element(f"level-{depth}")
        yield "start", ET.Element("too-deep")
        pytest.fail("超深 worksheet XML 不应继续扫描单元格")

    monkeypatch.setattr(reader.ET, "iterparse", fake_iterparse)

    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        with pytest.raises(reader.ReaderError, match="无法按 .xlsx 解析"):
            reader._scan_worksheet_bounds(
                archive,
                "xl/worksheets/sheet1.xml",
                sheet_name="Sheet 1",
                row_limit=config.IMPORT_MAX_ROWS,
                row_total_so_far=0,
                column_limit=config.IMPORT_XLSX_MAX_COLUMNS,
                declared_cell_limit=config.IMPORT_XLSX_MAX_DECLARED_CELLS,
                declared_cells_so_far=0,
            )


def test_precheck_does_not_truncate_data_rows_for_custom_directory_sheet(
    import_client,
):
    payload = _relocate_worksheets_to_custom_dir(_workbook_bytes([(
        "采购",
        [
            _PURCHASE_HEADER,
            _PURCHASE_ROW,
            ["CGDD-2", "PO-2", "PL-2", "PN-2", 100],
            ["CGDD-3", "PO-3", "PL-3", "PN-3", 100],
            ["CGDD-4", "PO-4", "PL-4", "PN-4", 100],
            ["CGDD-5", "PO-5", "PL-5", "PN-5", 100],
        ],
    )]))

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["file_type"] == "purchase"
    assert _sheet(file_result, "采购")["data_rows"] == 5


def test_precheck_rejects_declared_column_limit_before_pandas(
    import_client, monkeypatch
):
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("超列工作簿不应进入 pandas"),
    )

    file_result = _precheck(
        import_client, _sparse_workbook_bytes("XFD300000")
    )["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "column_limit_exceeded"


def test_precheck_rejects_declared_grid_limit_before_pandas(
    import_client, monkeypatch
):
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("超网格工作簿不应进入 pandas"),
    )
    cell_ref = f"{get_column_letter(512)}10000"

    file_result = _precheck(
        import_client, _sparse_workbook_bytes(cell_ref)
    )["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "declared_cell_limit_exceeded"


def test_precheck_does_not_trust_spoofed_dimension_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _rewrite_worksheet_dimension_ref(
        _workbook_bytes([("伪造维度列", [list(range(1, 514))])]),
        "A1",
    )
    workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    try:
        assert workbook.worksheets[0].max_column == 513
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "伪造 dimension 的超列工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "伪造 dimension 的超列工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "column_limit_exceeded"


@pytest.mark.parametrize("local_name", ["x", "dimension"])
def test_inspect_workbook_counts_any_direct_row_child_as_a_cell_before_pandas(
    monkeypatch,
    tmp_path,
    local_name,
):
    payload = _replace_first_cell_with_namespaced_sibling(
        _sparse_workbook_bytes("XFD1"),
        local_name,
    )
    workbook = openpyxl_load_workbook(io.BytesIO(payload), data_only=True)
    try:
        assert workbook.worksheets[0].max_column == 16_384
    finally:
        workbook.close()

    path = tmp_path / "non-c-cell.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "row 直属非 c 单元格超列文件不应进入 pandas"
        ),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "column_limit_exceeded"


def test_inspect_workbook_rejects_nested_row_elements_before_pandas(
    monkeypatch,
    tmp_path,
):
    payload = _replace_first_cell_with_namespaced_sibling(
        _workbook_bytes([("Sheet 1", [["row-1"]])]),
        "row",
    )
    path = tmp_path / "nested-row.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("嵌套 row 的工作簿不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


def test_precheck_infers_missing_column_refs_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _remove_worksheet_dimension_and_refs(_workbook_bytes([
        ("无坐标列", [list(range(1, 514))]),
    ]))
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(next(workbook.worksheets[0].iter_rows())) == 513
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "缺失 dimension 和列坐标的超列工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "缺失 dimension 和列坐标的超列工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "column_limit_exceeded"


def test_precheck_infers_missing_row_refs_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _remove_worksheet_dimension_and_refs(_workbook_bytes([
        ("无坐标行", [["r1"], ["r2"], ["r3"], ["r4"]]),
    ]))
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert len(list(workbook.worksheets[0].iter_rows())) == 4
    finally:
        workbook.close()

    monkeypatch.setattr(config, "IMPORT_MAX_ROWS", 3)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "缺失 dimension 和行坐标的超行工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "缺失 dimension 和行坐标的超行工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "row_limit_exceeded"


def test_inspect_workbook_rejects_non_decimal_row_ref_before_openpyxl_and_pandas(
    monkeypatch,
    tmp_path,
):
    payload = _rewrite_worksheet_row_ref(
        _workbook_bytes([("非法行号", [["row-1"]])]),
        "10000000.0",
    )
    path = tmp_path / "non-decimal-row-ref.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail("非法行号工作簿不应进入 openpyxl"),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail("非法行号工作簿不应进入 pandas"),
    )

    with pytest.raises(reader.ReaderError) as exc_info:
        reader.inspect_workbook(str(path), load_data=False)

    assert exc_info.value.code == "invalid_workbook"


def test_precheck_scans_sparse_sheet_without_dimension_before_openpyxl_and_pandas(
    import_client, monkeypatch
):
    payload = _remove_worksheet_dimension(_sparse_workbook_bytes("XFD10"))
    workbook = openpyxl_load_workbook(
        io.BytesIO(payload), read_only=True, data_only=True
    )
    try:
        assert workbook.worksheets[0]["XFD10"].value == "x"
    finally:
        workbook.close()

    monkeypatch.setattr(
        openpyxl,
        "load_workbook",
        lambda *_args, **_kwargs: pytest.fail(
            "缺失 dimension 的稀疏工作簿不应进入 openpyxl"
        ),
    )
    monkeypatch.setattr(
        reader.pd,
        "read_excel",
        lambda *_args, **_kwargs: pytest.fail(
            "缺失 dimension 的稀疏工作簿不应进入 pandas"
        ),
    )

    file_result = _precheck(import_client, payload)["files"][0]

    assert file_result["severity"] == "error"
    assert file_result["can_import"] is False
    assert file_result["issues"][0]["code"] == "column_limit_exceeded"


def test_batch_upload_cannot_bypass_any_fatal_precheck_error(
    db,
    import_client,
    monkeypatch,
    tmp_path,
):
    monkeypatch.setattr(config, "IMPORT_MAX_ROWS", 3)
    raw_dir = tmp_path / "raw"
    monkeypatch.setattr(get_settings(), "raw_file_dir", str(raw_dir))
    duplicate = _workbook_bytes([(
        "重复采购",
        [
            [
                "采购单号(必填)",
                "明细.产品名称(必填)",
                "明细.产品名称(必填)",
                "明细.单价(必填)",
            ],
            ["CGDD-1", "PN-A", "PN-B", 100],
        ],
    )])
    unrecognized = _workbook_bytes([
        ("说明", [["说明", "备注"], ["只用于阅读", "不导入"]]),
    ])
    over_rows = _workbook_bytes([(
        "采购",
        [
            _PURCHASE_HEADER,
            _PURCHASE_ROW,
            ["CGDD-2", "PO-2", "PL-2", "PN-2", 100],
            ["CGDD-3", "PO-3", "PL-3", "PN-3", 100],
        ],
    )])

    response = import_client.post(
        "/api/import/upload-batch?mode=upsert",
        files=[
            ("files", ("duplicate.xlsx", duplicate, _XLSX_CONTENT_TYPE)),
            ("files", ("unrecognized.xlsx", unrecognized, _XLSX_CONTENT_TYPE)),
            ("files", ("corrupt.xlsx", b"not-an-xlsx", _XLSX_CONTENT_TYPE)),
            ("files", ("over-rows.xlsx", over_rows, _XLSX_CONTENT_TYPE)),
        ],
    )

    assert response.status_code == 200, response.text
    job = _wait_for_job(import_client, response.json()["job_id"])
    assert job["status"] == "failed"
    assert job["done_files"] == 0
    assert job["error_files"] == 4
    assert len(job["batches"]) == 4
    assert all(batch["status"] == "failed" for batch in job["batches"])

    db.expire_all()
    assert db.scalar(
        select(func.count()).select_from(SysImportBatch).where(
            SysImportBatch.import_job_id == job["id"],
            SysImportBatch.status == "success",
        )
    ) == 0
    assert db.scalar(select(func.count()).select_from(SysRawFile)) == 0
    assert all(count == 0 for count in _fact_counts(db).values())
    assert not raw_dir.exists()


def test_pipeline_and_precheck_share_sheet_selection(
    db, import_client, tmp_path, monkeypatch,
):
    payload = _workbook_bytes([
        ("封面", [["标题"], ["库存导入"]]),
        ("库存一", [_INVENTORY_HEADER, ["INV-A", "PN-A", 1, "总仓"]]),
        ("库存二", [_INVENTORY_HEADER, ["INV-B", "PN-B", 2, "总仓"]]),
    ])
    path = tmp_path / "selection.xlsx"
    path.write_bytes(payload)
    monkeypatch.setattr(get_settings(), "raw_file_dir", str(tmp_path / "raw"))
    real_select = sheet_selection.select_workbook_sheets
    calls: list[list[str]] = []

    def capture_selection(sheets):
        selection = real_select(sheets)
        calls.append([sheet.sheet_name for sheet in selection.selected])
        return selection

    monkeypatch.setattr(sheet_selection, "select_workbook_sheets", capture_selection)

    precheck_result = _precheck(import_client, payload)["files"][0]
    batch = pipeline.run_import(db, str(path), path.name)
    db.commit()

    assert precheck_result["selected_sheets"] == ["库存一"]
    assert calls == [["库存一"], ["库存一"]]
    assert batch.file_type == "inventory"
    assert db.scalars(select(Inventory.raw_inventory_id).order_by(Inventory.raw_inventory_id)).all() == [
        "INV-A",
    ]


def test_precheck_does_not_log_business_cell_values(import_client, caplog):
    secret_customer = "仅用于隐私回归的客户名称"
    payload = _workbook_bytes([(
        "销售",
        [
            [
                "订单编号(必填)",
                "业务类型",
                "客户名称",
                "订单明细.产品名称",
                "订单明细.单价",
            ],
            ["XSDD-SECRET", "销售", secret_customer, "PN-SECRET", 987654.32],
        ],
    )])

    _precheck(import_client, payload)

    assert secret_customer not in caplog.text
    assert "987654.32" not in caplog.text
