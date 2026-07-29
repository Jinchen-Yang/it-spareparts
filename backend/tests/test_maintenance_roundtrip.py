"""维保项目固定工作簿协议的导出、回填与并发安全契约。"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from types import SimpleNamespace

import pytest
from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import func, select

from app import permissions
from app.api import maintenance as maintenance_api
from app.models.dimensions import DimPart
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceContractWorkbookState,
    MaintenanceManualCostOverride,
    MaintenanceRoundtripOperation,
)
from app.models.system import SysAuditLog, SysImportBatch, SysRawFile
from app.security import UserContext
from app.services import maintenance_cost
from app.services import maintenance_roundtrip


def _seed_contract(db, *, suffix: str, contract: str) -> tuple[int, int]:
    batch = SysImportBatch(
        filename=f"roundtrip-seed-{suffix}.xlsx",
        file_type="maintenance",
        file_hash=f"roundtrip-seed-{suffix}",
    )
    part = DimPart(pn_std=f"PN-RT-{suffix}")
    db.add_all([batch, part])
    db.flush()
    order = FMaintenanceOrder(
        raw_order_id=f"ORDER-RT-{suffix}",
        order_no=f"WBDD-RT-{suffix}",
        order_date=date(2026, 7, 15),
        linked_sales_order_no=contract,
        project_raw=f"回填项目-{suffix}",
        project_std=f"回填项目-{suffix}",
        data_status="已生效",
        import_batch_id=batch.id,
    )
    db.add(order)
    db.flush()
    line = FMaintenanceLine(
        raw_line_id=f"LINE-RT-{suffix}",
        order_id=order.id,
        line_no=1,
        part_id=part.id,
        pn_std=part.pn_std,
        description="原始描述",
        qty=Decimal("2"),
        anomaly_flags=[],
        import_batch_id=batch.id,
    )
    db.add(line)
    db.commit()
    return order.id, line.id


def _export_to_path(
    db,
    path,
    *,
    contract: str | None,
    date_from: date | None = None,
    date_to: date | None = None,
    blank: bool = False,
):
    output = maintenance_roundtrip.build_roundtrip_template(
        db,
        contract=contract,
        date_from=date_from,
        date_to=date_to,
        exported_by="tester",
        blank=blank,
    )
    try:
        payload = output.read()
    finally:
        output.close()
    # 导出使用事务级共享锁；模拟下载响应在流式发送前归还连接。
    db.rollback()
    path.write_bytes(payload)
    return path


def _edit_data_row(path, sheet: str, values: dict[str, object], *, row: int = 2):
    workbook = load_workbook(path, data_only=False)
    try:
        ws = workbook[sheet]
        headers = {
            str(cell.value): cell.column for cell in ws[1] if cell.value is not None
        }
        for header, value in values.items():
            ws.cell(row=row, column=headers[header], value=value)
        workbook.save(path)
    finally:
        workbook.close()


def test_blank_roundtrip_template_has_fixed_protocol_and_tables(db):
    output = maintenance_roundtrip.build_roundtrip_template(
        db,
        contract="XSDD-BLANK",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
        exported_by="tester",
        blank=True,
    )
    try:
        workbook = load_workbook(io.BytesIO(output.read()), data_only=False)
    finally:
        output.close()

    try:
        assert workbook.sheetnames == list(maintenance_roundtrip.SHEET_NAMES)
        assert workbook["98_字典"].sheet_state == "hidden"
        assert workbook["99_元数据"].sheet_state == "veryHidden"
        assert set(workbook["01_项目"].tables) == {"tbl_projects_v1"}
        assert set(workbook["02_维保订单"].tables) == {"tbl_orders_v1"}
        assert set(workbook["03_订单明细"].tables) == {"tbl_order_lines_v1"}
        assert set(workbook["04_报销明细"].tables) == {"tbl_expenses_v1"}
        assert set(workbook["05_人工成本回填"].tables) == {"tbl_manual_costs_v1"}
        assert list(workbook["01_项目"].data_validations.dataValidation) == []
        expected_validations = {
            "02_维保订单": (
                '"KEEP,UPDATE"',
                "请选择 KEEP 或 UPDATE",
            ),
            "03_订单明细": (
                '"KEEP,UPDATE"',
                "请选择 KEEP 或 UPDATE",
            ),
            "04_报销明细": (
                '"KEEP,CREATE,UPDATE,VOID"',
                "请选择 KEEP、CREATE、UPDATE 或 VOID",
            ),
            "05_人工成本回填": (
                '"KEEP,CREATE,UPDATE,VOID"',
                "请选择 KEEP、CREATE、UPDATE 或 VOID",
            ),
        }
        for sheet, (formula, error) in expected_validations.items():
            validations = list(workbook[sheet].data_validations.dataValidation)
            assert len(validations) == 1
            assert validations[0].formula1 == formula
            assert validations[0].error == error

        manual_sheet = workbook["05_人工成本回填"]
        manual_headers = {
            str(cell.value): cell.column
            for cell in manual_sheet[1]
            if cell.value is not None
        }
        required_fill = manual_sheet.cell(
            row=2,
            column=manual_headers["人工未税单位成本"],
        ).fill.fgColor.rgb
        assert (
            manual_sheet.cell(
                row=2,
                column=manual_headers["依据说明"],
            ).fill.fgColor.rgb
            == required_fill
        )
        expense_sheet = workbook["04_报销明细"]
        expense_headers = {
            str(cell.value): cell.column
            for cell in expense_sheet[1]
            if cell.value is not None
        }
        assert (
            expense_sheet.cell(
                row=2,
                column=expense_headers["变更原因"],
            ).fill.fgColor.rgb
            == required_fill
        )

        metadata = {
            workbook["99_元数据"].cell(row=row, column=1).value: str(
                workbook["99_元数据"].cell(row=row, column=2).value or ""
            )
            for row in range(2, workbook["99_元数据"].max_row + 1)
        }
        assert metadata["protocol_id"] == "ITDATA_MAINT_ROUNDTRIP"
        assert metadata["schema_version"] == "1.0"
        assert metadata["date_from"] == "2026-07-01"
        assert metadata["date_to"] == "2026-07-31"
        assert metadata["tax_rate"] == "0.13"
        assert metadata["amount_basis"] == "ex"
        assert metadata["template_mode"] == "blank"
        assert metadata["metadata_hmac"]
        scope_summary = {
            workbook["00_使用说明"].cell(row=row, column=1).value: workbook[
                "00_使用说明"
            ]
            .cell(row=row, column=2)
            .value
            for row in range(4, 9)
        }
        assert scope_summary == {
            "签名合同范围": "指定合同：XSDD-BLANK",
            "签名日期范围": ("指定日期闭区间：2026-07-01 至 2026-07-31（含边界）"),
            "模板模式": "空白",
            "数据截止日": metadata["as_of"],
            "导出时间": metadata["exported_at"],
        }
        instruction_text = " ".join(
            str(cell.value)
            for row in workbook["00_使用说明"].iter_rows()
            for cell in row
            if cell.value
        )
        assert "VOID 时，报销必须填变更原因，人工成本必须填回填原因" in (
            instruction_text
        )
        instruction_sheet = workbook["00_使用说明"]
        instruction_rows = {
            str(instruction_sheet.cell(row=row, column=1).value): row
            for row in range(1, instruction_sheet.max_row + 1)
        }
        assert (
            instruction_sheet.cell(
                row=instruction_rows["桃色单元格"],
                column=1,
            ).fill.fgColor.rgb
            == required_fill
        )
        assert "只有不带 date_from/date_to 的全量范围模板" in instruction_text
        assert (
            "带日期的范围模板和 blank=true 空白模板绝不会声明全量完整"
            in instruction_text
        )
        assert metadata["metadata_hmac"] not in instruction_text
        assert metadata["export_id"] not in instruction_text
    finally:
        workbook.close()


def test_roundtrip_template_uses_visible_content_only_for_bounded_row_heights(
    monkeypatch,
):
    contract = "XSDD-ROW-HEIGHT"

    def order(index: int):
        return SimpleNamespace(
            id=index,
            order_no=f"WBDD-{index}",
            order_date=date(2026, 7, index),
            linked_sales_order_no=contract,
            project_raw="维保项目",
            project_std="维保项目",
            maint_start=None,
            maint_end=None,
            end_customer="终端",
            demand_type="维保",
            business_type="维保",
            salesperson="销售",
            warehouse="仓库",
            data_status="已生效",
            import_batch_id=100 + index,
            version=1,
        )

    orders = [
        (order(1), "客" * 4),
        (order(2), "客" * 9),
        (order(3), "客" * 17),
        (order(4), "客" * 25),
    ]
    data = {
        "orders": orders,
        "lines": [],
        "expenses": [],
        "manual": [],
        "contracts": {contract},
    }
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_selected_data",
        lambda *args, **kwargs: data,
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_contract_state_revisions",
        lambda *args, **kwargs: {contract: 0},
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_lock_single_template_builder",
        lambda *args, **kwargs: None,
    )
    monkeypatch.setattr(
        maintenance_roundtrip,
        "_lock_shared_snapshot",
        lambda *args, **kwargs: None,
    )

    output = maintenance_roundtrip.build_roundtrip_template(
        object(),
        contract=contract,
        exported_by="tester",
    )
    try:
        workbook = load_workbook(io.BytesIO(output.read()), data_only=False)
    finally:
        output.close()

    try:
        ws = workbook["02_维保订单"]
        assert [ws.row_dimensions[row].height for row in range(2, 6)] == [
            24,
            36,
            48,
            72,
        ]
        assert ws.protection.formatRows is True

        headers = {
            str(cell.value): cell.column for cell in ws[1] if cell.value is not None
        }
        token_column = headers["__row_token"]
        assert len(str(ws.cell(row=2, column=token_column).value)) == 64
        assert (
            ws.column_dimensions[
                ws.cell(row=1, column=token_column).column_letter
            ].hidden
            is True
        )
        # 64 字符 HMAC 不得把只有短可见内容的第一行撑高。
        assert ws.row_dimensions[2].height == 24
    finally:
        workbook.close()


def test_instruction_scope_summary_shows_all_scope_snapshot_from_signed_metadata(db):
    output = maintenance_roundtrip.build_roundtrip_template(
        db,
        exported_by="tester",
    )
    try:
        workbook = load_workbook(io.BytesIO(output.read()), data_only=False)
    finally:
        output.close()

    try:
        metadata = {
            workbook["99_元数据"].cell(row=row, column=1).value: str(
                workbook["99_元数据"].cell(row=row, column=2).value or ""
            )
            for row in range(2, workbook["99_元数据"].max_row + 1)
        }
        scope_summary = {
            workbook["00_使用说明"].cell(row=row, column=1).value: workbook[
                "00_使用说明"
            ]
            .cell(row=row, column=2)
            .value
            for row in range(4, 9)
        }

        assert metadata["contract_scope"] == ""
        assert metadata["date_from"] == ""
        assert metadata["date_to"] == ""
        assert metadata["template_mode"] == "snapshot"
        assert scope_summary == {
            "签名合同范围": "全部合同",
            "签名日期范围": "全部日期",
            "模板模式": "快照",
            "数据截止日": metadata["as_of"],
            "导出时间": metadata["exported_at"],
        }
    finally:
        workbook.close()


def test_blank_template_without_dates_never_claims_full_snapshot(db, tmp_path):
    _seed_contract(db, suffix="BLANK", contract="XSDD-RT-BLANK")
    path = _export_to_path(
        db,
        tmp_path / "blank-full.xlsx",
        contract="XSDD-RT-BLANK",
        blank=True,
    )

    result = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    state = db.get(MaintenanceContractWorkbookState, "XSDD-RT-BLANK")
    assert state.expense_snapshot_complete is False
    report = db.get(SysImportBatch, result["batch_id"]).report_json
    assert report["expense_snapshot_complete"] is False
    assert report["expense_snapshot_complete_contracts"] == []


def test_unchanged_full_snapshot_only_attests_protocol_and_exact_replay_is_stable(
    db,
    tmp_path,
):
    contract = "XSDD-RT-UNCHANGED"
    order_id, line_id = _seed_contract(
        db,
        suffix="UNCHANGED",
        contract=contract,
    )
    # 先稳定派生成本字段，确保后续断言只衡量原样导入本身。
    maintenance_cost.recompute(db)
    path = _export_to_path(
        db,
        tmp_path / "unchanged-full.xlsx",
        contract=contract,
    )

    fact_models = (
        FMaintenanceOrder,
        FMaintenanceLine,
        FProjectExpense,
        MaintenanceManualCostOverride,
    )
    fact_counts_before = {
        model: db.scalar(select(func.count()).select_from(model))
        for model in fact_models
    }
    order_before = {
        column.name: getattr(db.get(FMaintenanceOrder, order_id), column.name)
        for column in FMaintenanceOrder.__table__.columns
    }
    line_before = {
        column.name: getattr(db.get(FMaintenanceLine, line_id), column.name)
        for column in FMaintenanceLine.__table__.columns
    }
    audit_count_before = db.scalar(
        select(func.count()).select_from(SysAuditLog)
    )
    roundtrip_batch_count_before = db.scalar(
        select(func.count())
        .select_from(SysImportBatch)
        .where(
            SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
        )
    )
    raw_file_count_before = db.scalar(
        select(func.count()).select_from(SysRawFile)
    )
    assert db.get(MaintenanceContractWorkbookState, contract) is None

    first = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    assert first["no_op"] is False
    assert first["changed_rows"] == 0
    assert first["counts"]["create"] == 0
    assert first["counts"]["update"] == 0
    assert first["counts"]["void"] == 0
    assert first["contracts"] == [contract]
    db.expire_all()
    assert {
        model: db.scalar(select(func.count()).select_from(model))
        for model in fact_models
    } == fact_counts_before
    assert {
        column.name: getattr(db.get(FMaintenanceOrder, order_id), column.name)
        for column in FMaintenanceOrder.__table__.columns
    } == order_before
    assert {
        column.name: getattr(db.get(FMaintenanceLine, line_id), column.name)
        for column in FMaintenanceLine.__table__.columns
    } == line_before
    assert (
        db.scalar(select(func.count()).select_from(SysAuditLog))
        == audit_count_before
    )

    state = db.get(MaintenanceContractWorkbookState, contract)
    assert state.expense_snapshot_complete is True
    assert state.revision == 1
    assert state.last_import_batch_id == first["batch_id"]
    assert (
        db.scalar(
            select(func.count())
            .select_from(SysImportBatch)
            .where(
                SysImportBatch.file_type
                == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
            )
        )
        == roundtrip_batch_count_before + 1
    )
    archived = db.scalar(
        select(SysRawFile).where(SysRawFile.batch_id == first["batch_id"])
    )
    assert archived is not None
    assert archived.file_hash == first["file_hash"]
    assert archived.storage_path
    assert (
        db.scalar(select(func.count()).select_from(SysRawFile))
        == raw_file_count_before + 1
    )

    replay = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    assert replay["no_op"] is True
    assert replay["logical_replay"] is False
    assert replay["batch_id"] == first["batch_id"]
    db.expire_all()
    replay_state = db.get(MaintenanceContractWorkbookState, contract)
    assert replay_state.expense_snapshot_complete is True
    assert replay_state.revision == 1
    assert replay_state.last_import_batch_id == first["batch_id"]
    assert (
        db.scalar(
            select(func.count())
            .select_from(SysImportBatch)
            .where(
                SysImportBatch.file_type
                == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
            )
        )
        == roundtrip_batch_count_before + 1
    )
    assert (
        db.scalar(select(func.count()).select_from(SysRawFile))
        == raw_file_count_before + 1
    )
    assert (
        db.scalar(select(func.count()).select_from(SysAuditLog))
        == audit_count_before
    )
    assert {
        model: db.scalar(select(func.count()).select_from(model))
        for model in fact_models
    } == fact_counts_before


def test_full_scope_expense_create_is_tax_normalized_audited_and_idempotent(
    db,
    tmp_path,
):
    _seed_contract(db, suffix="EXP", contract="XSDD-RT-EXP")
    path = _export_to_path(
        db,
        tmp_path / "expense-full.xlsx",
        contract="XSDD-RT-EXP",
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": "XSDD-RT-EXP",
            "报销日期": date(2026, 7, 20),
            "报销人员": "测试人员",
            "支出事由": "现场交通",
            "未税金额": Decimal("0.50"),
            "流程状态": "已结束",
            "变更原因": "补录报销",
        },
    )

    result = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    assert result["no_op"] is False
    assert result["counts"]["create"] == 1
    expense = db.scalar(
        select(FProjectExpense).where(
            FProjectExpense.linked_sales_order_no == "XSDD-RT-EXP"
        )
    )
    assert expense.amount == Decimal("0.50")
    assert expense.amount_ex_tax == Decimal("0.50")
    assert expense.amount_inc_tax == Decimal("0.57")
    assert expense.tax_basis == "ex"
    assert expense.tax_rate_used == Decimal("0.1300")
    state = db.get(MaintenanceContractWorkbookState, "XSDD-RT-EXP")
    assert state.expense_snapshot_complete is True
    batch = db.get(SysImportBatch, result["batch_id"])
    assert batch.report_json["expense_snapshot_complete"] is True
    assert batch.report_json["expense_snapshot_complete_contracts"] == ["XSDD-RT-EXP"]
    assert (
        db.scalar(
            select(func.count(SysAuditLog.id)).where(
                SysAuditLog.entity_type == "f_project_expense",
                SysAuditLog.entity_id == expense.id,
            )
        )
        == 1
    )

    second = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )
    assert second["no_op"] is True
    assert second["batch_id"] == result["batch_id"]
    assert (
        db.scalar(
            select(func.count(FProjectExpense.id)).where(
                FProjectExpense.linked_sales_order_no == "XSDD-RT-EXP"
            )
        )
        == 1
    )


def test_same_create_client_uuid_is_independent_across_exports(db, tmp_path):
    client_row_id = "11111111-1111-4111-8111-111111111111"
    for suffix in ("CREATE-EXPORT-A", "CREATE-EXPORT-B"):
        contract = f"XSDD-RT-{suffix}"
        _seed_contract(db, suffix=suffix, contract=contract)
        path = _export_to_path(
            db,
            tmp_path / f"{suffix}.xlsx",
            contract=contract,
        )
        _edit_data_row(
            path,
            "04_报销明细",
            {
                "操作": "CREATE",
                "合同号": contract,
                "报销日期": date(2026, 7, 20),
                "未税金额": Decimal("10.00"),
                "变更原因": "跨导出同 UUID",
                "__client_row_id": client_row_id,
            },
        )
        result = maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )
        assert result["counts"]["create"] == 1

    raw_ids = list(db.scalars(
        select(FProjectExpense.raw_line_id).order_by(FProjectExpense.id)
    ))
    assert len(raw_ids) == 2
    assert len(set(raw_ids)) == 2
    assert all(raw_id.startswith("RTEXP:") for raw_id in raw_ids)
    assert all(len(raw_id) <= 80 for raw_id in raw_ids)


def test_same_create_client_uuid_is_independent_across_sheets(db, tmp_path):
    client_row_id = "22222222-2222-4222-8222-222222222222"
    contract = "XSDD-RT-CREATE-CROSS-SHEET"
    _order_id, line_id = _seed_contract(
        db,
        suffix="CREATE-CROSS-SHEET",
        contract=contract,
    )
    maintenance_cost.recompute(db)
    assert db.get(FMaintenanceLine, line_id).cost_source == "none"
    path = _export_to_path(
        db,
        tmp_path / "create-cross-sheet.xlsx",
        contract=contract,
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": contract,
            "报销日期": date(2026, 7, 20),
            "未税金额": Decimal("10.00"),
            "变更原因": "跨 sheet UUID",
            "__client_row_id": client_row_id,
        },
    )
    _edit_data_row(
        path,
        "05_人工成本回填",
        {
            "操作": "CREATE",
            "人工未税单位成本": Decimal("1.00"),
            "回填原因": "跨 sheet UUID",
            "依据说明": "跨 sheet UUID 独立性测试",
            "__client_row_id": client_row_id,
        },
    )

    result = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    assert result["counts"]["create"] == 2
    assert db.scalar(select(func.count(FProjectExpense.id))) == 1
    assert db.scalar(
        select(func.count(MaintenanceManualCostOverride.id))
    ) == 1
    ledger_keys = set(db.execute(
        select(
            MaintenanceRoundtripOperation.sheet_code,
            MaintenanceRoundtripOperation.client_row_id,
        )
    ).all())
    assert ledger_keys == {
        ("04_报销明细", client_row_id),
        ("05_人工成本回填", client_row_id),
    }


def test_duplicate_create_client_uuid_in_same_sheet_rejects_entire_workbook(
    db,
    tmp_path,
):
    client_row_id = "33333333-3333-4333-8333-333333333333"
    contract = "XSDD-RT-CREATE-SAME-SHEET"
    _seed_contract(
        db,
        suffix="CREATE-SAME-SHEET",
        contract=contract,
    )
    path = _export_to_path(
        db,
        tmp_path / "create-same-sheet.xlsx",
        contract=contract,
    )
    for row in (2, 3):
        _edit_data_row(
            path,
            "04_报销明细",
            {
                "操作": "CREATE",
                "合同号": contract,
                "报销日期": date(2026, 7, 20),
                "未税金额": Decimal("10.00"),
                "变更原因": "同 sheet 重复 UUID",
                "__client_row_id": client_row_id,
            },
            row=row,
        )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="客户端行键.*重复",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert db.scalar(select(func.count(FProjectExpense.id))) == 0
    assert db.scalar(
        select(func.count(SysImportBatch.id)).where(
            SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
        )
    ) == 0


@pytest.mark.parametrize(
    ("tax_basis", "raw_amount", "amount_ex", "amount_inc"),
    [
        ("default_ex", Decimal("0.88"), Decimal("0.88"), Decimal("0.99")),
        ("ex", Decimal("0.88"), Decimal("0.88"), Decimal("0.99")),
        ("inc", Decimal("1.00"), Decimal("0.88"), Decimal("1.00")),
    ],
)
def test_expense_update_preserves_source_tax_basis_and_raw_amount(
    db,
    tmp_path,
    tax_basis,
    raw_amount,
    amount_ex,
    amount_inc,
):
    suffix = f"EXP-{tax_basis.upper()}-UPDATE"
    contract = f"XSDD-RT-EXP-{tax_basis.upper()}-UPDATE"
    order_id, _ = _seed_contract(
        db,
        suffix=suffix,
        contract=contract,
    )
    source_batch_id = db.get(FMaintenanceOrder, order_id).import_batch_id
    expense = FProjectExpense(
        raw_line_id=f"EXP-RT-{tax_basis.upper()}-UPDATE",
        linked_sales_order_no=contract,
        expense_date=date(2026, 7, 19),
        data_status="已结束",
        reason="原说明",
        amount=raw_amount,
        amount_ex_tax=amount_ex,
        amount_inc_tax=amount_inc,
        tax_basis=tax_basis,
        tax_rate_used=Decimal("0.13"),
        import_batch_id=source_batch_id,
    )
    db.add(expense)
    db.commit()
    expense_id = expense.id

    path = _export_to_path(
        db,
        tmp_path / f"expense-{tax_basis}-update.xlsx",
        contract=contract,
    )
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook["04_报销明细"]
        headers = {
            str(cell.value): cell.column
            for cell in worksheet[1]
            if cell.value is not None
        }
        assert worksheet.cell(
            row=2,
            column=headers["__tax_basis"],
        ).value == tax_basis
        assert Decimal(str(worksheet.cell(
            row=2,
            column=headers["__raw_amount"],
        ).value)) == raw_amount
        assert worksheet.column_dimensions[
            worksheet.cell(
                row=1,
                column=headers["__tax_basis"],
            ).column_letter
        ].hidden is True
        assert worksheet.column_dimensions[
            worksheet.cell(
                row=1,
                column=headers["__raw_amount"],
            ).column_letter
        ].hidden is True
    finally:
        workbook.close()

    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "UPDATE",
            "支出事由": "只修改说明",
            "变更原因": "补充说明，不改变原税务口径",
        },
    )
    maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    db.expire_all()
    updated = db.get(FProjectExpense, expense_id)
    assert updated.reason == "只修改说明"
    assert updated.tax_basis == tax_basis
    assert updated.amount == raw_amount
    assert updated.amount_ex_tax == amount_ex
    assert updated.amount_inc_tax == amount_inc
    assert updated.tax_rate_used == Decimal("0.1300")
    audit = db.scalar(
        select(SysAuditLog)
        .where(
            SysAuditLog.entity_type == "f_project_expense",
            SysAuditLog.entity_id == expense_id,
            SysAuditLog.action == "update",
        )
        .order_by(SysAuditLog.id.desc())
    )
    assert audit.before_json["tax_basis"] == tax_basis
    assert audit.after_json["tax_basis"] == tax_basis
    assert Decimal(audit.before_json["amount"]) == raw_amount
    assert Decimal(audit.after_json["amount"]) == raw_amount


def test_expense_update_can_fill_historical_null_amount_without_changing_basis(
    db,
    tmp_path,
):
    contract = "XSDD-RT-EXP-NULL-UPDATE"
    order_id, _ = _seed_contract(
        db,
        suffix="EXP-NULL-UPDATE",
        contract=contract,
    )
    source_batch_id = db.get(FMaintenanceOrder, order_id).import_batch_id
    expense = FProjectExpense(
        raw_line_id="EXP-RT-NULL-UPDATE",
        linked_sales_order_no=contract,
        expense_date=date(2026, 7, 19),
        data_status="已结束",
        reason="历史空金额",
        amount=None,
        amount_ex_tax=None,
        amount_inc_tax=None,
        tax_basis="default_ex",
        tax_rate_used=Decimal("0.13"),
        import_batch_id=source_batch_id,
    )
    db.add(expense)
    db.commit()
    expense_id = expense.id

    path = _export_to_path(
        db,
        tmp_path / "expense-null-update.xlsx",
        contract=contract,
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "UPDATE",
            "未税金额": Decimal("10.00"),
            "支出事由": "人工补齐历史空金额",
            "变更原因": "补齐金额",
        },
    )
    maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    db.expire_all()
    updated = db.get(FProjectExpense, expense_id)
    assert updated.reason == "人工补齐历史空金额"
    assert updated.tax_basis == "default_ex"
    assert updated.amount == Decimal("10.00")
    assert updated.amount_ex_tax == Decimal("10.00")
    assert updated.amount_inc_tax == Decimal("11.30")
    assert updated.tax_rate_used == Decimal("0.1300")


def test_roundtrip_expense_accepts_negative_reversal_with_half_up_tax(db, tmp_path):
    _seed_contract(db, suffix="EXP-REV", contract="XSDD-RT-EXP-REV")
    path = _export_to_path(
        db,
        tmp_path / "expense-reversal.xlsx",
        contract="XSDD-RT-EXP-REV",
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": "XSDD-RT-EXP-REV",
            "报销日期": date(2026, 7, 21),
            "支出事由": "报销冲销",
            "未税金额": Decimal("-0.50"),
            "流程状态": "已结束",
            "变更原因": "冲销重复报销",
        },
    )

    maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    expense = db.scalar(
        select(FProjectExpense).where(
            FProjectExpense.linked_sales_order_no == "XSDD-RT-EXP-REV",
        )
    )
    assert expense.amount == Decimal("-0.50")
    assert expense.amount_ex_tax == Decimal("-0.50")
    assert expense.amount_inc_tax == Decimal("-0.57")
    assert expense.tax_basis == "ex"


def test_date_scoped_import_never_marks_expenses_as_full_snapshot(db, tmp_path):
    _seed_contract(db, suffix="DATE", contract="XSDD-RT-DATE")
    path = _export_to_path(
        db,
        tmp_path / "expense-date.xlsx",
        contract="XSDD-RT-DATE",
        date_from=date(2026, 7, 1),
        date_to=date(2026, 7, 31),
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": "XSDD-RT-DATE",
            "报销日期": date(2026, 7, 20),
            "未税金额": Decimal("20.00"),
            "变更原因": "期间补录",
        },
    )

    result = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    state = db.get(MaintenanceContractWorkbookState, "XSDD-RT-DATE")
    assert state.expense_snapshot_complete is False
    assert state.expense_complete_through == date(2026, 7, 31)
    report = db.get(SysImportBatch, result["batch_id"]).report_json
    assert report["expense_snapshot_complete"] is False
    assert report["expense_snapshot_complete_contracts"] == []


def test_full_scope_rejects_unsigned_new_contract_without_any_write(db, tmp_path):
    _seed_contract(db, suffix="SCOPE", contract="XSDD-RT-SIGNED")
    path = _export_to_path(
        db,
        tmp_path / "expense-new-contract.xlsx",
        contract="XSDD-RT-SIGNED",
    )
    _edit_data_row(
        path,
        "04_报销明细",
        {
            "操作": "CREATE",
            "合同号": "XSDD-RT-UNSIGNED",
            "报销日期": date(2026, 7, 20),
            "未税金额": Decimal("30.00"),
            "变更原因": "新合同补录",
        },
    )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="签名合同范围",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 409
    assert db.get(MaintenanceContractWorkbookState, "XSDD-RT-SIGNED") is None
    assert db.get(MaintenanceContractWorkbookState, "XSDD-RT-UNSIGNED") is None
    assert (
        db.scalar(
            select(func.count(FProjectExpense.id)).where(
                FProjectExpense.linked_sales_order_no == "XSDD-RT-UNSIGNED"
            )
        )
        == 0
    )


def test_formula_anywhere_rejects_entire_workbook_without_row_writes(db, tmp_path):
    order_id, _line_id = _seed_contract(
        db,
        suffix="FORMULA",
        contract="XSDD-RT-FORMULA",
    )
    path = _export_to_path(
        db,
        tmp_path / "formula.xlsx",
        contract="XSDD-RT-FORMULA",
    )
    _edit_data_row(
        path,
        "02_维保订单",
        {
            "操作": "UPDATE",
            "项目名称": "不应写入",
            "变更原因": "测试原子回滚",
        },
    )
    workbook = load_workbook(path, data_only=False)
    try:
        workbook["00_使用说明"]["C20"] = "=1+1"
        workbook.save(path)
    finally:
        workbook.close()

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="包含公式",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    db.expire_all()
    assert db.get(FMaintenanceOrder, order_id).project_raw == "回填项目-FORMULA"
    assert (
        db.scalar(
            select(func.count(SysImportBatch.id)).where(
                SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
            )
        )
        == 0
    )


def test_metadata_tamper_is_rejected_before_any_import_batch(db, tmp_path):
    _seed_contract(db, suffix="META", contract="XSDD-RT-META")
    path = _export_to_path(
        db,
        tmp_path / "metadata-tamper.xlsx",
        contract="XSDD-RT-META",
    )
    workbook = load_workbook(path, data_only=False)
    try:
        ws = workbook["99_元数据"]
        as_of_row = next(
            row
            for row in range(2, ws.max_row + 1)
            if ws.cell(row=row, column=1).value == "as_of"
        )
        ws.cell(row=as_of_row, column=2, value="2099-12-31")
        workbook.save(path)
    finally:
        workbook.close()

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="元数据签名无效",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert (
        db.scalar(
            select(func.count(SysImportBatch.id)).where(
                SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
            )
        )
        == 0
    )


def test_stale_existing_row_version_rejects_without_overwriting(db, tmp_path):
    order_id, _line_id = _seed_contract(
        db,
        suffix="STALE",
        contract="XSDD-RT-STALE",
    )
    path = _export_to_path(
        db,
        tmp_path / "stale.xlsx",
        contract="XSDD-RT-STALE",
    )
    _edit_data_row(
        path,
        "02_维保订单",
        {
            "操作": "UPDATE",
            "项目名称": "旧工作簿不应覆盖",
            "变更原因": "旧版本测试",
        },
    )
    newer_batch = SysImportBatch(
        filename="newer-maintenance.xlsx",
        file_type="maintenance",
        file_hash="newer-maintenance-stale",
    )
    db.add(newer_batch)
    db.flush()
    db.get(FMaintenanceOrder, order_id).import_batch_id = newer_batch.id
    db.commit()

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="数据已被更新",
    ) as caught:
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert caught.value.status_code == 409
    db.expire_all()
    assert db.get(FMaintenanceOrder, order_id).project_raw == "回填项目-STALE"
    assert (
        db.scalar(
            select(func.count(SysImportBatch.id)).where(
                SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
            )
        )
        == 0
    )


def test_mid_apply_failure_rolls_back_rows_audits_batch_and_state(
    db,
    tmp_path,
    monkeypatch,
):
    _seed_contract(db, suffix="ATOMIC", contract="XSDD-RT-ATOMIC")
    path = _export_to_path(
        db,
        tmp_path / "atomic.xlsx",
        contract="XSDD-RT-ATOMIC",
    )
    for row, amount in ((2, "10.00"), (3, "20.00")):
        _edit_data_row(
            path,
            "04_报销明细",
            {
                "操作": "CREATE",
                "合同号": "XSDD-RT-ATOMIC",
                "报销日期": date(2026, 7, 20),
                "未税金额": Decimal(amount),
                "变更原因": "原子性测试",
            },
            row=row,
        )

    real_apply = maintenance_roundtrip._apply_change
    calls = 0

    def fail_after_first(*args, **kwargs):
        nonlocal calls
        calls += 1
        result = real_apply(*args, **kwargs)
        if calls == 2:
            raise RuntimeError("simulated apply failure")
        return result

    monkeypatch.setattr(maintenance_roundtrip, "_apply_change", fail_after_first)
    with pytest.raises(RuntimeError, match="simulated apply failure"):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    assert (
        db.scalar(
            select(func.count(FProjectExpense.id)).where(
                FProjectExpense.linked_sales_order_no == "XSDD-RT-ATOMIC"
            )
        )
        == 0
    )
    assert (
        db.scalar(
            select(func.count(SysAuditLog.id)).where(
                SysAuditLog.entity_type == "f_project_expense"
            )
        )
        == 0
    )
    assert (
        db.scalar(
            select(func.count(SysImportBatch.id)).where(
                SysImportBatch.file_type == maintenance_roundtrip.ROUNDTRIP_FILE_TYPE
            )
        )
        == 0
    )
    assert db.get(MaintenanceContractWorkbookState, "XSDD-RT-ATOMIC") is None


def test_existing_line_update_preserves_quantity_milliscale_and_is_audited(
    db,
    tmp_path,
):
    _order_id, line_id = _seed_contract(
        db,
        suffix="LINE",
        contract="XSDD-RT-LINE",
    )
    path = _export_to_path(
        db,
        tmp_path / "line-update.xlsx",
        contract="XSDD-RT-LINE",
    )
    _edit_data_row(
        path,
        "03_订单明细",
        {
            "操作": "UPDATE",
            "产品描述": "更新后的描述",
            "需求数量": Decimal("1.234"),
            "退货数量": Decimal("0.111"),
            "变更原因": "修正数量精度",
        },
    )

    maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    db.expire_all()
    line = db.get(FMaintenanceLine, line_id)
    assert line.description == "更新后的描述"
    assert line.qty == Decimal("1.234")
    assert line.return_qty == Decimal("0.111")
    audit = db.scalar(
        select(SysAuditLog).where(
            SysAuditLog.entity_type == "f_maintenance_line",
            SysAuditLog.entity_id == line_id,
        )
    )
    assert audit.action == "update"
    assert audit.reason == "修正数量精度"


def test_manual_cost_create_only_for_none_source_and_recompute_uses_it(db, tmp_path):
    _order_id, line_id = _seed_contract(
        db,
        suffix="MANUAL",
        contract="XSDD-RT-MANUAL",
    )
    maintenance_cost.recompute(db)
    assert db.get(FMaintenanceLine, line_id).cost_source == "none"
    path = _export_to_path(
        db,
        tmp_path / "manual.xlsx",
        contract="XSDD-RT-MANUAL",
    )
    _edit_data_row(
        path,
        "05_人工成本回填",
        {
            "操作": "CREATE",
            "人工未税单位成本": Decimal("2.50"),
            "回填原因": "三个月内无采购和销售参考",
            "依据说明": "人工询价记录",
        },
    )

    maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    override = db.scalar(
        select(MaintenanceManualCostOverride).where(
            MaintenanceManualCostOverride.line_id == line_id
        )
    )
    assert override.active is True
    assert override.unit_cost_ex_tax == Decimal("2.50")
    assert override.unit_cost_inc_tax == Decimal("2.83")
    assert override.evidence == {"note": "人工询价记录"}
    db.expire_all()
    line = db.get(FMaintenanceLine, line_id)
    assert line.cost_source == "manual"
    assert line.unit_cost_ex_tax == Decimal("2.50")
    assert line.unit_cost_inc_tax == Decimal("2.83")

    output = maintenance_roundtrip.build_roundtrip_template(
        db,
        contract="XSDD-RT-MANUAL",
        exported_by="tester",
    )
    try:
        workbook = load_workbook(io.BytesIO(output.read()), data_only=False)
    finally:
        output.close()
    try:
        headers = {
            str(cell.value): cell.column
            for cell in workbook["05_人工成本回填"][1]
            if cell.value is not None
        }
        assert (
            workbook["05_人工成本回填"]
            .cell(
                row=2,
                column=headers["状态"],
            )
            .value
            == "生效"
        )
        assert (
            workbook["05_人工成本回填"]
            .cell(
                row=2,
                column=headers["依据说明"],
            )
            .value
            == "人工询价记录"
        )
    finally:
        workbook.close()


def test_manual_cost_create_requires_evidence_and_leaves_business_data_unchanged(
    db,
    tmp_path,
):
    _order_id, line_id = _seed_contract(
        db,
        suffix="MANUAL-NO-EVIDENCE",
        contract="XSDD-RT-MANUAL-NO-EVIDENCE",
    )
    maintenance_cost.recompute(db)
    assert db.get(FMaintenanceLine, line_id).cost_source == "none"
    path = _export_to_path(
        db,
        tmp_path / "manual-no-evidence.xlsx",
        contract="XSDD-RT-MANUAL-NO-EVIDENCE",
    )
    _edit_data_row(
        path,
        "05_人工成本回填",
        {
            "操作": "CREATE",
            "人工未税单位成本": Decimal("2.50"),
            "回填原因": "三个月内无采购和销售参考",
            "依据说明": "",
        },
    )

    with pytest.raises(
        maintenance_roundtrip.RoundtripWorkbookError,
        match="缺少“依据说明”",
    ):
        maintenance_roundtrip.import_roundtrip_workbook(
            db,
            str(path),
            filename=path.name,
            operated_by="tester",
        )

    db.rollback()
    db.expire_all()
    assert db.scalar(
        select(func.count())
        .select_from(MaintenanceManualCostOverride)
        .where(MaintenanceManualCostOverride.line_id == line_id)
    ) == 0
    assert db.get(FMaintenanceLine, line_id).cost_source == "none"


def test_manual_cost_create_accepts_zero_and_recompute_applies_zero(db, tmp_path):
    _order_id, line_id = _seed_contract(
        db,
        suffix="MANUAL-ZERO",
        contract="XSDD-RT-MANUAL-ZERO",
    )
    maintenance_cost.recompute(db)
    assert db.get(FMaintenanceLine, line_id).cost_source == "none"
    path = _export_to_path(
        db,
        tmp_path / "manual-zero.xlsx",
        contract="XSDD-RT-MANUAL-ZERO",
    )
    _edit_data_row(
        path,
        "05_人工成本回填",
        {
            "操作": "CREATE",
            "人工未税单位成本": Decimal("0.00"),
            "回填原因": "确认属于免费更换",
            "依据说明": "厂家零费用确认单",
        },
    )

    result = maintenance_roundtrip.import_roundtrip_workbook(
        db,
        str(path),
        filename=path.name,
        operated_by="tester",
    )

    assert result["counts"]["create"] == 1
    assert result["recompute"]["manual"] == 1
    override = db.scalar(
        select(MaintenanceManualCostOverride).where(
            MaintenanceManualCostOverride.line_id == line_id
        )
    )
    assert override.unit_cost_ex_tax == Decimal("0.00")
    assert override.unit_cost_inc_tax == Decimal("0.00")
    db.expire_all()
    line = db.get(FMaintenanceLine, line_id)
    assert line.cost_source == "manual"
    assert line.unit_cost == Decimal("0.00")
    assert line.unit_cost_ex_tax == Decimal("0.00")
    assert line.unit_cost_inc_tax == Decimal("0.00")
    assert line.cost_amount == Decimal("0.00")
    assert line.cost_amount_ex_tax == Decimal("0.00")
    assert line.cost_amount_inc_tax == Decimal("0.00")


def test_roundtrip_import_requires_same_cost_and_profit_visibility_as_export(db):
    ctx = UserContext(
        user_id="purchaser",
        role="purchaser",
        permissions=permissions.effective("purchaser", None),
        is_authenticated=True,
    )
    upload = UploadFile(
        filename="maintenance_roundtrip.xlsx",
        file=io.BytesIO(b"must-not-be-read"),
    )

    with pytest.raises(HTTPException) as caught:
        maintenance_api.roundtrip_import(
            file=upload,
            db=db,
            _auth="purchaser",
            _page=None,
            ctx=ctx,
        )

    assert caught.value.status_code == 403
    assert "无成本及利润查看权限" in caught.value.detail
    assert upload.file.tell() == 0
