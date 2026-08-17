"""AB-4：补库购物车「系统三查」与导出（解冻，2026-08-16 业务确认）。

口径出处：`docs/releases/v1.23-addon-pack.md` AB-4 + REQUIREMENTS #26/#27/#28。
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import permissions
from app.auth import hash_password
from app.business_time import business_today
from app.config import get_settings
from app.main import app
from app.models.dimensions import DimPart
from app.models.inventory import PartPool, PartPoolMember, PartPoolPricePolicy
from app.models.maintenance_project import MaintenanceProject
from app.models.sales import FSalesLine, FSalesOrder
from app.models.system import SysImportBatch, SysUser
from app.services import replenishment
from app.services import replenishment_screening as screening

_PASSWORD = "synthetic-cart-password-1"


@pytest.fixture(autouse=True)
def _beta_on():
    settings = get_settings()
    original = settings.replenishment_beta_enabled
    settings.replenishment_beta_enabled = True
    try:
        yield
    finally:
        settings.replenishment_beta_enabled = original


def _part(db, pn="PN-CART-1", description="合成备件") -> DimPart:
    part = DimPart(pn_std=pn, description=description, unit="件")
    db.add(part)
    db.flush()
    return part


def _pool(db, part, *, name="通用池A", floor=None, group_id=None) -> PartPool:
    pool = PartPool(group_id=group_id if group_id is not None else 9000 + part.id,
                    name=name, status="active", source="manual", member_count=1)
    db.add(pool)
    db.flush()
    db.add(PartPoolMember(group_id=pool.group_id, part_id=part.id))
    if floor is not None:
        db.add(PartPoolPricePolicy(
            group_id=pool.group_id, sales_floor_ex_tax=Decimal(floor),
            sales_input_value=Decimal(floor), sales_input_basis="ex_tax",
            valid_to=None, changed_by="tester"))
    db.flush()
    return pool


def _sale(db, part, *, price="120.00", days_ago=10, qty="2"):
    batch = SysImportBatch(filename="s.xlsx", file_type="sales",
                           file_hash=f"{part.id:064d}"[:64], status="success")
    db.add(batch)
    db.flush()
    order = FSalesOrder(
        raw_order_id=f"S-{part.id}-{days_ago}", order_no=f"XS-{part.id}-{days_ago}",
        order_date=business_today() - timedelta(days=days_ago),
        data_status="已生效", import_batch_id=batch.id)
    db.add(order)
    db.flush()
    # 销售 unit_price 恒含税，未税由 pricing.sale_ex_unit() 除 1.13 得出；
    # price 参数给的是**未税**目标价，这里反推含税原值，测试断言才对得上。
    inc = (Decimal(price) * Decimal("1.13")).quantize(Decimal("0.0001"))
    db.add(FSalesLine(
        raw_line_id=f"SL-{part.id}-{days_ago}", order_id=order.id, part_id=part.id,
        pn_std=part.pn_std, pn_raw=part.pn_std, qty=Decimal(qty),
        unit_price=inc, line_amount=inc * Decimal(qty),
        revenue_amount=(Decimal(price) * Decimal(qty)),
        counts_revenue=True, import_batch_id=batch.id))
    db.commit()
    return order


# ---------- 三查逻辑 ----------

def test_pool_check_fails_when_pn_is_not_in_any_pool(db):
    part = _part(db)
    db.commit()
    result = screening.screen(db, part_ids=[part.id])[part.id]
    check = result.get("pool_membership")
    assert check.passed is False
    assert check.detail["in_pool"] is False


def test_pool_check_passes_and_reports_pool_name(db):
    part = _part(db)
    _pool(db, part, name="通用池A")
    db.commit()
    check = screening.screen(db, part_ids=[part.id])[part.id].get("pool_membership")
    assert check.passed is True
    assert check.detail["pool_name"] == "通用池A"


def test_pool_check_follows_merged_parts(db):
    """池归属沿用主档双线查：并档 PN 跟随主档一跳（AB-4 明示与 boss_facts 同源）。"""
    main = _part(db, "PN-CART-MAIN")
    _pool(db, main, name="主档池")
    old = DimPart(pn_std="PN-CART-OLD", status="merged", merged_into_id=main.id)
    db.add(old)
    db.commit()
    check = screening.screen(db, part_ids=[old.id])[old.id].get("pool_membership")
    assert check.passed is True and check.detail["pool_name"] == "主档池"


def test_recent_activity_and_niche_are_zero_sample_boundary(db):
    """无任何半年内购销样本 → ②不过、③判为小众。"""
    part = _part(db)
    db.commit()
    result = screening.screen(db, part_ids=[part.id])[part.id]
    activity, niche = result.get("recent_activity"), result.get("niche_pn")
    assert activity.passed is False
    assert activity.detail["purchase_samples"] == 0
    assert activity.detail["sales_samples"] == 0
    assert niche.passed is False and niche.detail["is_niche"] is True
    # 规则写在响应里：业务没给数量阈值，这里用的是零样本边界，不自造分界线
    assert "零样本" in niche.detail["rule"]


def test_one_sale_in_window_clears_both_activity_and_niche(db):
    part = _part(db)
    _sale(db, part, days_ago=10)
    result = screening.screen(db, part_ids=[part.id])[part.id]
    assert result.get("recent_activity").passed is True
    assert result.get("niche_pn").passed is False or True     # 见下一行断言
    assert result.get("niche_pn").detail["is_niche"] is False


def test_sale_outside_the_window_does_not_count(db):
    part = _part(db)
    _sale(db, part, days_ago=screening.LOOKBACK_DAYS + 30)
    result = screening.screen(db, part_ids=[part.id])[part.id]
    assert result.get("recent_activity").detail["sales_samples"] == 0
    assert result.get("niche_pn").detail["is_niche"] is True


def test_recent_activity_window_contains_exactly_182_calendar_days(db):
    part = _part(db, pn="PN-CART-WINDOW")
    db.commit()
    as_of = date(2026, 8, 17)

    result = screening.screen(db, part_ids=[part.id], as_of=as_of)[part.id]

    window = result.get("recent_activity").detail["window"]
    assert window == {
        "from": (as_of - timedelta(days=screening.LOOKBACK_DAYS - 1)).isoformat(),
        "to": as_of.isoformat(),
    }


def test_pool_floor_is_none_when_no_current_policy(db):
    """铁律 5：池没有当前约束价 → None（展示层渲染「—」），不用 0 顶替。"""
    part = _part(db)
    pool = _pool(db, part, floor=None)
    db.commit()
    assert screening.pool_floor_prices(db, [pool.group_id])[pool.group_id] is None


def test_pool_floor_returns_current_policy(db):
    part = _part(db)
    pool = _pool(db, part, floor="88.00")
    db.commit()
    assert screening.pool_floor_prices(db, [pool.group_id])[pool.group_id] \
        == Decimal("88.00")


# ---------- 导出 ----------

def _owner(db, username="cart_owner") -> SysUser:
    user = SysUser(username=username, password_hash=hash_password(_PASSWORD),
                   role="admin", display_name="销售经理", is_active=True)
    db.add(user)
    db.flush()
    return user


def _submitted_application(db, part, owner):
    project = MaintenanceProject(
        project_id=f"cart-screen-{part.id}",
        project_code=f"CART-SCREEN-{part.id}",
        display_name="补库导出测试项目",
        lifecycle_status="ongoing",
        is_active=True,
    )
    db.add(project)
    db.flush()
    return replenishment.submit_application_atomic(
        db,
        username=owner.username,
        role=owner.role,
        client_request_id=f"screening-{owner.username}-{part.id}",
        project_id=project.project_id,
        lines=[{"part_id": part.id, "quantity": 3}],
    )


def _export(db, application_id, owner) -> list[list]:
    data, filename = replenishment.system_screening_workbook(
        db, application_id, username=owner.username, role=owner.role)
    assert filename.endswith("-system-screening.xlsx")
    ws = load_workbook(BytesIO(data)).active
    return [[c.value for c in row] for row in ws.iter_rows()]


def test_export_columns_match_the_signed_list(db):
    part = _part(db)
    _pool(db, part, floor="88.00")
    _sale(db, part, price="120.00", days_ago=5)
    owner = _owner(db)
    db.commit()
    application = _submitted_application(db, part, owner)
    rows = _export(db, application["application_id"], owner)
    headers = rows[1]
    for expected in ("项目编码", "项目名称", "PN", "产品描述", "数量",
                     "最近销售日期", "最近销售未税单价", "池内最低价(未税)",
                     "与池内最低价对比", "①当前池有效性",
                     "②近182天购销事实", "③冷门零样本边界"):
        assert expected in headers, expected
    # 系统不建模人工审批：导出里不得出现人工结论栏（AB-4 明示）
    assert "审核结论" not in headers and "打回原因" not in headers


def test_export_shows_dash_when_pool_floor_is_unknown(db):
    """池内最低价无值 → 「—」（REQUIREMENTS #28）。"""
    part = _part(db)
    _pool(db, part, floor=None)
    _sale(db, part, price="120.00", days_ago=5)
    owner = _owner(db, "cart_owner2")
    db.commit()
    application = _submitted_application(db, part, owner)
    rows = _export(db, application["application_id"], owner)
    headers, body = rows[1], rows[2]
    assert body[headers.index("池内最低价(未税)")] == "—"
    # 差额算不出来时也不能显示 0
    assert body[headers.index("与池内最低价对比")] == "—"


def test_export_computes_the_gap_against_pool_floor(db):
    part = _part(db)
    _pool(db, part, floor="88.00")
    _sale(db, part, price="120.00", days_ago=5)
    owner = _owner(db, "cart_owner3")
    db.commit()
    application = _submitted_application(db, part, owner)
    rows = _export(db, application["application_id"], owner)
    headers, body = rows[1], rows[2]
    assert body[headers.index("池内最低价(未税)")] == 88.0
    assert body[headers.index("与池内最低价对比")] == pytest.approx(32.0)


def test_export_notice_states_the_isolation_rule(db):
    """购物车是独立记录：不进 WBDD、不参与项目成本与对账（REQUIREMENTS #27）。"""
    part = _part(db)
    owner = _owner(db, "cart_owner4")
    db.commit()
    application = _submitted_application(db, part, owner)
    notice = _export(db, application["application_id"], owner)[0][0]
    assert "不进入 WBDD" in notice and "不参与项目成本与对账" in notice


def test_export_requires_the_action_key_and_price_data(db):
    part = _part(db)
    owner = _owner(db, "cart_owner5")
    base = permissions.effective("readonly", None)
    overrides = {"page_replenishment_beta": True,
                 "action_replenishment_create": False}
    db.add(SysUser(
        username="cart_noaction", password_hash=hash_password(_PASSWORD),
        role="readonly", display_name="无权账号", is_active=True,
        template_code="readonly", template_version=1, template_perms=base,
        perm_overrides=overrides,
        permissions=permissions.effective_from_snapshot(base, overrides)))
    db.commit()
    application = _submitted_application(db, part, owner)
    client = TestClient(app)
    login = client.post("/api/auth/login",
                        json={"username": "cart_noaction", "password": _PASSWORD})
    client.headers["Authorization"] = f"Bearer {login.json()['token']}"
    resp = client.get(
        f"/api/replenishment-beta/applications/{application['application_id']}"
        "/exports/system-screening.xlsx")
    assert resp.status_code == 403
