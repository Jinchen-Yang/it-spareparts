"""R2：项目总表六 sheet / 单 sheet / 全项目备件行级表的下载与上传覆盖。

口径出处：REQUIREMENTS #38（三处下载点、在哪下载就在哪上传）、#40（一张总表改所有
数据）、表 6 六 sheet 规格。
"""
import io
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import date, timedelta
from decimal import Decimal

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import event, select, text

from app.business_time import business_today
from app.db import SessionLocal
from app.models.maintenance import FMaintenanceLine, MaintenanceManualCostOverride
from app.models.maintenance import MaintenanceDemandTombstone
from app.models.maintenance_project import MaintenanceProjectContract
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.security import UserContext
from app.services import maintenance_expense_collection_workbook as ec
from app.services import maintenance_cost
from app.services import maintenance_demands
from app.services import maintenance_project_master_workbook as master
from app.services import maintenance_project_operations as operations
from app.services import maintenance_source_assignments
from app.api import maintenance_project_master_workbook as master_api
from tests.boss_board_helpers import assign, import_wbdd, make_project
from tests.test_maintenance_expense_collection_workbook import (  # noqa: F401
    _XLSX, _client, uploader, reader,
)

_BASE = "/api/maintenance/projects/stable"
_GLOBAL = "/api/maintenance/spare-part-lines"


@pytest.fixture(autouse=True)
def _use_v1_master_workbook_by_default(monkeypatch):
    """Keep V1 contract tests independent from deployment feature flags."""
    monkeypatch.setattr(
        master_api.get_settings(), "maintenance_project_master_v2_enabled", False
    )


def test_workbook_revision_bump_is_deduplicated_per_root_transaction(
    db,
    project_with_lines,
):
    state = operations.get_or_create_workbook_state(
        db, project_id=project_with_lines.project_id
    )
    db.commit()
    initial = state.revision

    locked = operations.lock_workbook_states(
        db, project_ids=[project_with_lines.project_id]
    )[project_with_lines.project_id]
    operations.bump_locked_workbook_revision(db, state=locked)
    operations.bump_locked_workbook_revision(db, state=locked)
    db.commit()
    db.refresh(state)
    assert state.revision == initial + 1

    locked = operations.lock_workbook_states(
        db, project_ids=[project_with_lines.project_id]
    )[project_with_lines.project_id]
    operations.bump_locked_workbook_revision(db, state=locked)
    db.commit()
    db.refresh(state)
    assert state.revision == initial + 2


def test_v1_master_multi_sheet_apply_bumps_revision_once(
    db, project_with_lines,
):
    """One V1 transaction may touch many facts but advances one revision."""

    line = db.scalar(select(FMaintenanceLine).order_by(FMaintenanceLine.id))
    contract = db.scalar(select(MaintenanceProjectContract).where(
        MaintenanceProjectContract.project_id == project_with_lines.project_id
    ))
    state = operations.get_or_create_workbook_state(
        db, project_id=project_with_lines.project_id)
    db.commit()
    before = state.revision
    plan = master.MasterPlan(
        project_id=project_with_lines.project_id,
        cost_refills=(master.CostRefill(
            line_id=line.id,
            unit_cost_ex_tax=Decimal("12.00"),
            unit_cost_inc_tax=Decimal("13.56"),
            reason="revision-test",
        ),),
        inner=ec.WorkbookPlan(
            project_id=project_with_lines.project_id,
            expense_updates=(),
            collection_ops=(ec.CollectionOp(
                operation="CREATE",
                project_contract_id=contract.project_contract_id,
                contract_no=contract.contract_no,
                report_month=date(2026, 8, 1),
                cumulative_amount=Decimal("100.00"),
                receipt_reference="REV-1",
                remark=None,
            ),),
        ),
        sheets=(master.SHEET_PARTS, master.SHEET_COLLECTION),
    )

    master.apply(
        db,
        plan,
        operated_by="revision-v1",
        import_batch_id=str(uuid.uuid4()),
    )
    db.expire_all()
    state = operations.get_or_create_workbook_state(
        db, project_id=project_with_lines.project_id)
    assert state.revision == before + 1


def test_contract_total_change_requires_named_financial_project_manager(monkeypatch):
    monkeypatch.setattr(master_api.config, "ENABLE_RBAC", True)
    privileged = UserContext(
        user_id="named-finance", role="admin", is_authenticated=True,
        permissions={
            "action_maintenance_project_manage": True,
            "data_profit": True,
        },
    )
    with pytest.raises(HTTPException) as shared:
        master_api._require_contract_amount_manage(
            privileged, {"authn": "shared", "sub": "admin", "fb": True})
    assert shared.value.status_code == 403

    missing_manage = UserContext(
        user_id="named-uploader", role="readonly", is_authenticated=True,
        permissions={
            "action_maintenance_project_manage": False,
            "data_profit": True,
        },
    )
    with pytest.raises(HTTPException) as denied:
        master_api._require_contract_amount_manage(
            missing_manage,
            {"authn": "sys_user", "sub": "named-uploader", "fb": False},
        )
    assert denied.value.status_code == 403

    master_api._require_contract_amount_manage(
        privileged,
        {"authn": "sys_user", "sub": "named-finance", "fb": False},
    )


def test_cost_recompute_busy_maps_to_retryable_conflict():
    with pytest.raises(HTTPException) as raised:
        master_api._fail(ec.WorkbookError(
            "cost_recompute_busy", "成本重算正在进行"))
    assert raised.value.status_code == 409
    assert raised.value.headers["Retry-After"] == "5"


def test_stale_workbook_maps_to_conflict():
    with pytest.raises(HTTPException) as raised:
        master_api._fail(ec.WorkbookError(
            "stale_workbook", "项目数据已被其他操作更新"))
    assert raised.value.status_code == 409


@pytest.fixture()
def project_with_lines(db, tmp_path):
    from app.models.maintenance_project import MaintenanceProjectContract

    proj = make_project(db)
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()), project_id=proj.project_id,
        contract_id="C-1", contract_no="HT-001",
        amount_inc_tax=Decimal("100000.00"), included_in_total=True,
        status_mapping_state="mapped", status_mapping_version="v1",
        effective_from=date(2026, 1, 1), source="ledger", version=1))
    orders = import_wbdd(db, tmp_path, orders=1, lines_per_order=2)
    assign(db, orders[0], proj)
    db.commit()
    return proj


def _download_master(client, project, sheets=None):
    params = {"sheets": sheets} if sheets else None
    resp = client.get(f"{_BASE}/{project.project_id}/master-workbook.xlsx",
                      params=params)
    assert resp.status_code == 200, resp.text
    return resp.content


def _upload_master(client, project, content, *, action="apply"):
    return client.post(
        f"{_BASE}/{project.project_id}/master-workbook/{action}",
        files={"file": ("m.xlsx", io.BytesIO(content), _XLSX)})


def _v1_overview(db, project):
    workbook = load_workbook(io.BytesIO(master.build_project_master(
        db,
        project_id=project.project_id,
        sheets=(master.SHEET_OVERVIEW,),
    )), data_only=True)
    worksheet = workbook[master.SHEET_OVERVIEW]
    metrics = {
        row[0].value: row[1].value
        for row in worksheet.iter_rows(min_col=1, max_col=2)
        if row[0].value
    }
    return worksheet, metrics


def test_v1_overview_uses_only_current_canonical_contract_total(
    db,
    project_with_lines,
):
    """V1 fallback must not add an expired relationship to today's total."""

    today = business_today()
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project_with_lines.project_id,
        contract_id="C-HISTORICAL",
        contract_no="HT-HISTORICAL",
        amount_inc_tax=Decimal("900000.00"),
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="v1",
        effective_from=today - timedelta(days=730),
        effective_to=today - timedelta(days=1),
        source="ledger",
        version=1,
    ))
    db.commit()

    worksheet, metrics = _v1_overview(db, project_with_lines)

    assert Decimal(str(metrics["合同总额(含税)"])) == Decimal("100000.00")
    assert metrics["合同额口径"] == "完整：当前含税合同事实"
    assert "HT-HISTORICAL" not in {
        worksheet.cell(row=row_no, column=1).value
        for row_no in range(2, worksheet.max_row + 1)
    }


def test_v1_overview_masks_known_subtotal_when_current_contract_is_incomplete(
    db,
    project_with_lines,
):
    """A missing amount is unknown, not zero beside a precise known subtotal."""

    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=project_with_lines.project_id,
        contract_id="C-MISSING",
        contract_no="HT-MISSING",
        amount_inc_tax=None,
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="v1",
        effective_from=date(2026, 1, 1),
        source="ledger",
        version=1,
    ))
    db.commit()

    _worksheet, metrics = _v1_overview(db, project_with_lines)

    assert metrics["合同总额(含税)"] is None
    assert metrics["成本率(进度条口径)"] is None
    assert str(metrics["合同额口径"]).startswith("不完整：")


def test_v1_overview_masks_shared_current_contract(db, project_with_lines):
    other = make_project(db, code=f"SHARED-{uuid.uuid4().hex[:8]}")
    db.add(MaintenanceProjectContract(
        project_contract_id=str(uuid.uuid4()),
        project_id=other.project_id,
        contract_id="C-SHARED-OTHER",
        contract_no="HT-001",
        amount_inc_tax=Decimal("100000.00"),
        included_in_total=True,
        status_mapping_state="mapped",
        status_mapping_version="v1",
        effective_from=date(2026, 1, 1),
        source="ledger",
        version=1,
    ))
    db.commit()

    _worksheet, metrics = _v1_overview(db, project_with_lines)

    assert metrics["合同总额(含税)"] is None
    assert metrics["成本率(进度条口径)"] is None
    assert "共享/归属冲突" in metrics["合同额口径"]


def test_v2_validate_contract_total_uses_authenticated_identity(
    db, project_with_lines, monkeypatch,
):
    """生产开启 V2 时，合同额预检必须注入实名身份而不是 500。"""
    monkeypatch.setattr(
        master_api.get_settings(), "maintenance_project_master_v2_enabled", True)
    client = _client(
        db,
        username="named-contract-manager",
        overrides={
            "page_maintenance": True,
            "data_profit": True,
            "action_maintenance_expense_collection_upload": True,
            "action_maintenance_project_manage": True,
        },
    )
    wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
    amount_cell = next(
        row[1]
        for row in wb[master.V2_SHEET_OVERVIEW].iter_rows(min_col=1, max_col=2)
        if row[0].value == "合同总额（含税）"
    )
    amount_cell.value = Decimal("123456.78")
    buf = io.BytesIO()
    wb.save(buf)

    response = _upload_master(
        client, project_with_lines, buf.getvalue(), action="validate")

    assert response.status_code == 200, response.text
    assert response.json()["contract_updates"] == 1


def test_v2_overview_respects_explicit_empty_salesperson_override(
    db,
    project_with_lines,
    monkeypatch,
):
    """项目总表不得把人工清空覆盖成 WBDD 众数。"""
    monkeypatch.setattr(
        master_api.get_settings(), "maintenance_project_master_v2_enabled", True
    )
    project_with_lines.salesperson = None
    project_with_lines.salesperson_override_active = True
    db.commit()

    workbook = load_workbook(
        io.BytesIO(
            _download_master(
                uploader(db, "master-salesperson-override"),
                project_with_lines,
            )
        ),
        data_only=True,
    )
    overview = workbook[master.V2_SHEET_OVERVIEW]
    salesperson = next(
        row[1].value
        for row in overview.iter_rows(min_col=1, max_col=2)
        if row[0].value == "销售人员"
    )

    assert salesperson == "—"


# ---------- 项目总表：六 sheet ----------

def test_master_workbook_has_all_six_sheets(db, project_with_lines):
    wb = load_workbook(io.BytesIO(
        _download_master(uploader(db), project_with_lines)))
    visible = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"]
    assert visible == list(master.ALL_SHEETS)


def test_single_sheet_download_returns_only_that_sheet(db, project_with_lines):
    """各 tab 单 sheet 下载（#38）。"""
    wb = load_workbook(io.BytesIO(_download_master(
        uploader(db, "m-single"), project_with_lines, sheets=master.SHEET_PARTS)))
    visible = [n for n in wb.sheetnames if wb[n].sheet_state == "visible"]
    assert visible == [master.SHEET_PARTS]


def test_master_rows_api_uses_strict_cost_and_active_manual_net_qty(
    db, project_with_lines, monkeypatch,
):
    monkeypatch.setattr(
        master_api.get_settings(), "maintenance_project_master_v2_enabled", True)
    rows = master._assigned_lines(
        db, project_id=project_with_lines.project_id, window=None)
    first, first_order, _project_id = rows[0]
    second = rows[1][0]
    # 「已返」是旧件返还等流转状态，不是需求退货冲抵。即使已返=需求=2，
    # 净量仍应是 qty(2)-return_qty(0)=2，绝不能被冲成 0。
    first.qty = Decimal("2")
    first.return_qty = Decimal("0")
    first.returned_qty = Decimal("2")
    first.pending_return_qty = Decimal("1")
    first.cost_source = None
    override = MaintenanceManualCostOverride(
        line_id=first.id,
        unit_cost_ex_tax=Decimal("8"),
        unit_cost_inc_tax=Decimal("9.04"),
        active=True,
        updated_by="test",
    )
    db.add(override)
    second.cost_source = "future_source"
    second.cost_tax_basis = "inc"
    second.cost_amount = Decimal("999")
    second.cost_amount_inc_tax = Decimal("999")
    second.unit_cost_inc_tax = Decimal("999")
    db.commit()

    response = uploader(db, "m-strict-rows").get(
        f"{_BASE}/{project_with_lines.project_id}/master-workbook/rows",
        params={"sheet": master.V2_SHEET_PARTS},
    )
    assert response.status_code == 200, response.text
    by_id = {row["line_id"]: row for row in response.json()["rows"]}
    assert by_id[first.id]["cost_amount_inc_tax"] == "18.08"
    assert by_id[first.id]["cost_source"] == "manual"
    assert by_id[first.id]["qty"] == "2.000"
    assert by_id[first.id]["return_qty"] == "0.000"
    assert by_id[first.id]["returned_qty"] == "2.000"
    assert by_id[first.id]["pending_return_qty"] == "1.000"
    assert by_id[second.id]["cost_amount_inc_tax"] is None
    assert by_id[second.id]["missing_kind"] == "invalid_cost_fact"

    agent_row = maintenance_cost._serialize_project_line(
        first,
        first_order,
        override,
        hide_cost_signals=False,
    )
    assert agent_row["qty"] == 2.0
    assert agent_row["return_qty"] == 0.0
    assert agent_row["returned_qty"] == 2.0
    assert agent_row["pending_return_qty"] == 1.0
    assert agent_row["cost_amount_inc_tax"] == 18.08


def test_unknown_sheet_name_is_422(db, project_with_lines):
    resp = uploader(db, "m-badsheet").get(
        f"{_BASE}/{project_with_lines.project_id}/master-workbook.xlsx",
        params={"sheets": "99_不存在"})
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "unknown_sheet"


def test_unknown_project_fails_closed_for_scoped_account(db):
    assert uploader(db, "m-404").get(
        f"{_BASE}/{uuid.uuid4()}/master-workbook.xlsx").status_code == 403


# ---------- 03 备件订单：缺价补录 ----------

def test_cost_refill_writes_manual_override_and_computes_inc_tax(db,
                                                                 project_with_lines):
    """#40 缺价补录：未税人工填，含税系统算；落既有人工成本覆盖表（零迁移）。"""
    client = uploader(db, "m-refill")
    wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
    ws = wb[master.SHEET_PARTS]
    ws.cell(row=2, column=12, value=200)          # 未税单位成本
    ws.cell(row=2, column=14, value="补录依据：合同附件")
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload_master(client, project_with_lines, buf.getvalue())
    assert resp.status_code == 200, resp.text
    assert resp.json()["cost_refills"] == 1

    override = db.execute(select(MaintenanceManualCostOverride)).scalars().one()
    assert override.unit_cost_ex_tax == Decimal("200.00")
    assert override.unit_cost_inc_tax == Decimal("226.00")
    assert override.reason == "补录依据：合同附件"
    assert override.active is True


def test_v1_workbook_reads_active_manual_cost_in_overview_and_parts(
    db,
    project_with_lines,
    monkeypatch,
):
    monkeypatch.setattr(
        master_api.get_settings(), "maintenance_project_master_v2_enabled", False)
    rows = master._assigned_lines(
        db, project_id=project_with_lines.project_id, window=None)
    line = rows[0][0]
    line.qty = Decimal("3")
    line.return_qty = Decimal("1")
    line.cost_source = None
    line.cost_tax_basis = None
    line.unit_cost_ex_tax = None
    line.unit_cost_inc_tax = None
    line.cost_amount = None
    line.cost_amount_ex_tax = None
    line.cost_amount_inc_tax = None
    db.add(MaintenanceManualCostOverride(
        line_id=line.id,
        unit_cost_ex_tax=Decimal("8.00"),
        unit_cost_inc_tax=Decimal("9.04"),
        reason="V1 人工成本证据",
        active=True,
        updated_by="v1-test",
    ))
    db.commit()

    workbook = load_workbook(io.BytesIO(
        _download_master(uploader(db, "m-v1-manual"), project_with_lines)))
    overview = workbook[master.SHEET_OVERVIEW]
    metrics = {
        row[0].value: row[1].value
        for row in overview.iter_rows(min_col=1, max_col=2)
        if row[0].value
    }
    assert Decimal(str(metrics["项目已计成本(含税)"])) == Decimal("18.08")
    assert metrics["缺失成本行数"] == 1

    parts = workbook[master.SHEET_PARTS]
    row_by_line_id = {
        int(parts.cell(row=row_no, column=len(master._PARTS_HEADERS) + 1).value): row_no
        for row_no in range(2, parts.max_row + 1)
    }
    row_no = row_by_line_id[line.id]
    assert parts.cell(row=row_no, column=11).value == "manual"
    assert Decimal(str(parts.cell(row=row_no, column=12).value)) == Decimal("8.00")
    assert Decimal(str(parts.cell(row=row_no, column=13).value)) == Decimal("9.04")
    assert parts.cell(row=row_no, column=14).value == "V1 人工成本证据"


def test_reupload_overwrites_the_same_line(db, project_with_lines):
    client = uploader(db, "m-refill2")

    def _send(amount):
        wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
        wb[master.SHEET_PARTS].cell(row=2, column=12, value=amount)
        buf = io.BytesIO()
        wb.save(buf)
        return _upload_master(client, project_with_lines, buf.getvalue())

    assert _send(200).status_code == 200
    assert _send(300).status_code == 200
    overrides = db.execute(select(MaintenanceManualCostOverride)).scalars().all()
    assert len(overrides) == 1                     # 覆盖，不是追加
    db.expire_all()
    assert overrides[0].unit_cost_ex_tax == Decimal("300.00")


def test_blank_cost_leaves_line_untouched(db, project_with_lines):
    client = uploader(db, "m-blank")
    content = _download_master(client, project_with_lines)
    assert _upload_master(client, project_with_lines, content).json()["cost_refills"] == 0
    assert db.execute(select(MaintenanceManualCostOverride)).scalars().all() == []


def test_invented_part_row_is_rejected(db, project_with_lines):
    """需求单只能由氚云导入，本表只补价（铁律 1）。"""
    client = uploader(db, "m-invent")
    wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
    wb[master.SHEET_PARTS].append(["编的单号", "2026-07-01", "HT-001", "项目",
                                   "PN-X", "描述", 1, 0, "", "", "", 10, "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload_master(client, project_with_lines, buf.getvalue())
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "line_not_recognized"


# ---------- 04/05 复用 AB-3 ----------

def test_master_upload_also_applies_expense_and_collection(db, project_with_lines):
    """总表一次上传同时落 03/05——三类改动同一事务（#40 一张总表改所有数据）。"""
    from app.models.maintenance_project_operations import MaintenanceCollectionSnapshot

    client = uploader(db, "m-all")
    wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
    wb[master.SHEET_PARTS].cell(row=2, column=12, value=150)
    ws = wb[master.SHEET_COLLECTION]
    ws.cell(row=2, column=1, value="CREATE")
    ws.cell(row=2, column=2, value="HT-001")
    ws.cell(row=2, column=3, value="2026-07")
    ws.cell(row=2, column=4, value=50000)
    buf = io.BytesIO()
    wb.save(buf)
    body = _upload_master(client, project_with_lines, buf.getvalue()).json()
    assert body["cost_refills"] == 1 and body["collection_creates"] == 1
    assert db.execute(select(MaintenanceCollectionSnapshot)).scalars().one() \
        .cumulative_amount == Decimal("50000.00")


def test_one_bad_row_rejects_the_whole_file(db, project_with_lines):
    """整份拒绝：合法的补价也不能落库。"""
    client = uploader(db, "m-atomic")
    wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
    wb[master.SHEET_PARTS].cell(row=2, column=12, value=150)      # 合法
    ws = wb[master.SHEET_COLLECTION]
    ws.cell(row=2, column=1, value="CREATE")
    ws.cell(row=2, column=2, value="HT-001")
    ws.cell(row=2, column=3, value="不是月份")                     # 非法
    ws.cell(row=2, column=4, value=1)
    buf = io.BytesIO()
    wb.save(buf)
    assert _upload_master(client, project_with_lines, buf.getvalue()).status_code == 422
    assert db.execute(select(MaintenanceManualCostOverride)).scalars().all() == []


def test_validate_is_side_effect_free(db, project_with_lines):
    client = uploader(db, "m-validate")
    wb = load_workbook(io.BytesIO(_download_master(client, project_with_lines)))
    wb[master.SHEET_PARTS].cell(row=2, column=12, value=777)
    buf = io.BytesIO()
    wb.save(buf)
    resp = _upload_master(client, project_with_lines, buf.getvalue(),
                          action="validate")
    assert resp.status_code == 200 and resp.json()["cost_refills"] == 1
    assert db.execute(select(MaintenanceManualCostOverride)).scalars().all() == []


# ---------- 主页全局备件行级表 ----------

def test_range_presets(db):
    today = business_today()
    assert master.resolve_range("today", None, None) == (today, today)
    y = today - timedelta(days=1)
    assert master.resolve_range("yesterday", None, None) == (y, y)
    assert master.resolve_range("this_week", None, None)[0] \
        == today - timedelta(days=today.weekday())
    assert master.resolve_range("this_month", None, None)[0] == today.replace(day=1)
    assert master.resolve_range("custom", date(2026, 1, 1), date(2026, 2, 1)) \
        == (date(2026, 1, 1), date(2026, 2, 1))


def test_custom_range_requires_both_ends(db):
    with pytest.raises(ec.WorkbookError) as excinfo:
        master.resolve_range("custom", date(2026, 1, 1), None)
    assert excinfo.value.code == "invalid_range"


def test_global_download_and_refill_roundtrip(db, project_with_lines):
    client = uploader(db, "g-refill")
    resp = client.get(f"{_GLOBAL}.xlsx",
                      params={"range": "custom", "from": "2026-01-01",
                              "to": "2026-12-31"})
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb[master.GLOBAL_SHEET]
    assert ws.max_row >= 2                       # 有归属的行都在
    ws.cell(row=2, column=10, value=88)          # 未税单位成本
    buf = io.BytesIO()
    wb.save(buf)
    applied = client.post(f"{_GLOBAL}/apply",
                          files={"file": ("g.xlsx", io.BytesIO(buf.getvalue()),
                                          _XLSX)})
    assert applied.status_code == 200, applied.text
    assert applied.json()["cost_refills"] == 1
    override = db.execute(select(MaintenanceManualCostOverride)).scalars().one()
    assert override.unit_cost_ex_tax == Decimal("88.00")


def test_global_export_is_scoped_by_range(db, project_with_lines):
    """今天没有单 → 表里只有表头（不报错、不塞别的期间的行）。"""
    client = uploader(db, "g-range")
    resp = client.get(f"{_GLOBAL}.xlsx", params={"range": "today"})
    ws = load_workbook(io.BytesIO(resp.content))[master.GLOBAL_SHEET]
    assert ws.max_row == 1


def test_global_export_empty_project_scope_returns_header_only(db, project_with_lines):
    """空集合必须表示零可见项目，不能误解释成全范围。"""
    client = uploader(
        db,
        "g-empty-scope",
        assign_existing_projects=False,
    )
    response = client.get(
        f"{_GLOBAL}.xlsx",
        params={"range": "custom", "from": "2026-01-01", "to": "2026-12-31"},
    )
    assert response.status_code == 200, response.text
    ws = load_workbook(io.BytesIO(response.content))[master.GLOBAL_SHEET]
    assert ws.max_row == 1


def test_global_upload_rejects_cross_project_hidden_line_id_atomically(
    db,
    project_with_lines,
):
    """改隐藏行键也不能把无权项目的成本写进去。"""
    authorized = uploader(db, "g-scope-source")
    response = authorized.get(
        f"{_GLOBAL}.xlsx",
        params={"range": "custom", "from": "2026-01-01", "to": "2026-12-31"},
    )
    wb = load_workbook(io.BytesIO(response.content))
    wb[master.GLOBAL_SHEET].cell(row=2, column=10, value=88)
    payload = io.BytesIO()
    wb.save(payload)

    denied = uploader(
        db,
        "g-scope-denied",
        assign_existing_projects=False,
    )
    applied = denied.post(
        f"{_GLOBAL}/apply",
        files={"file": ("g.xlsx", io.BytesIO(payload.getvalue()), _XLSX)},
    )
    assert applied.status_code == 422
    assert applied.json()["detail"]["code"] == "project_scope_denied"
    assert db.scalars(select(MaintenanceManualCostOverride)).all() == []


def test_global_apply_rechecks_scope_after_validation(db, project_with_lines):
    """validate 后来源单改派到别的项目，apply 必须在任何写入前拒绝。"""
    content = master.build_global_lines(
        db,
        preset="custom",
        date_from=date(2026, 1, 1),
        date_to=date(2026, 12, 31),
        allowed_project_ids={project_with_lines.project_id},
    )
    wb = load_workbook(io.BytesIO(content))
    wb[master.GLOBAL_SHEET].cell(row=2, column=10, value=88)
    payload = io.BytesIO()
    wb.save(payload)
    plan = master.validate_global(
        db,
        data=payload.getvalue(),
        allowed_project_ids={project_with_lines.project_id},
    )

    assignment = db.scalar(select(MaintenanceSourceOrderAssignment).where(
        MaintenanceSourceOrderAssignment.is_active.is_(True)
    ))
    assert assignment is not None
    other = make_project(db, code="合成改派后项目")
    maintenance_source_assignments.assign_source_orders(
        db,
        project_id=other.project_id,
        items=[{
            "source_order_id": assignment.source_order_id,
            "expected_assignment_id": assignment.assignment_id,
            "expected_version": assignment.version,
        }],
        reason="合成 validate/apply 之间的合法改派",
        operated_by="scope-test",
        user_ctx=UserContext(
            user_id="scope-test",
            role="admin",
            is_authenticated=True,
        ),
    )
    db.commit()

    with pytest.raises(ec.WorkbookError) as excinfo:
        master.apply_global_lines(
            db,
            plan,
            operated_by="scope-test",
            import_batch_id=str(uuid.uuid4()),
            allowed_project_ids={project_with_lines.project_id},
        )
    assert excinfo.value.code == "project_scope_denied"
    assert db.scalars(select(MaintenanceManualCostOverride)).all() == []


def test_global_apply_is_idempotent_and_rejects_stale_absent_override(
    db,
    project_with_lines,
):
    """两个文件都以“无 override”为 base：同值重试无副作用，异值不能后写覆盖。"""
    state = operations.get_or_create_workbook_state(
        db, project_id=project_with_lines.project_id
    )
    db.commit()
    before_revision = state.revision
    content = master.build_global_lines(
        db,
        preset="custom",
        date_from=date(2026, 1, 1),
        date_to=date(2026, 12, 31),
        allowed_project_ids={project_with_lines.project_id},
    )

    def _plan(amount):
        workbook = load_workbook(io.BytesIO(content))
        workbook[master.GLOBAL_SHEET].cell(row=2, column=10, value=amount)
        payload = io.BytesIO()
        workbook.save(payload)
        return master.validate_global(
            db,
            data=payload.getvalue(),
            allowed_project_ids={project_with_lines.project_id},
        )

    first = _plan(88)
    stale = _plan(99)
    result = master.apply_global_lines(
        db,
        first,
        operated_by="first-writer",
        import_batch_id=str(uuid.uuid4()),
        allowed_project_ids={project_with_lines.project_id},
    )
    assert result["replayed_line_ids"] == []
    db.refresh(state)
    assert state.revision == before_revision + 1

    replay = master.apply_global_lines(
        db,
        first,
        operated_by="first-writer-retry",
        import_batch_id=str(uuid.uuid4()),
        allowed_project_ids={project_with_lines.project_id},
    )
    assert replay["replayed_line_ids"] == [first.cost_refills[0].line_id]
    db.refresh(state)
    assert state.revision == before_revision + 1
    override = db.scalar(select(MaintenanceManualCostOverride).where(
        MaintenanceManualCostOverride.line_id == first.cost_refills[0].line_id))
    assert override.version == 1
    assert override.unit_cost_ex_tax == Decimal("88.00")
    state = operations.get_or_create_workbook_state(
        db, project_id=project_with_lines.project_id
    )
    assert state.revision == before_revision + 1

    with pytest.raises(ec.WorkbookError) as raised:
        master.apply_global_lines(
            db,
            stale,
            operated_by="stale-writer",
            import_batch_id=str(uuid.uuid4()),
            allowed_project_ids={project_with_lines.project_id},
        )
    assert raised.value.code == "stale_cost_override"
    db.rollback()
    db.expire_all()
    override = db.scalar(select(MaintenanceManualCostOverride).where(
        MaintenanceManualCostOverride.line_id == first.cost_refills[0].line_id))
    assert override.version == 1
    assert override.unit_cost_ex_tax == Decimal("88.00")


def test_global_workbook_rejects_tampered_concurrency_token(db, project_with_lines):
    content = master.build_global_lines(
        db,
        preset="custom",
        date_from=date(2026, 1, 1),
        date_to=date(2026, 12, 31),
        allowed_project_ids={project_with_lines.project_id},
    )
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook[master.GLOBAL_SHEET]
    sheet.cell(row=2, column=10, value=88)
    sheet.cell(row=2, column=len(master._GLOBAL_HEADERS) + 2, value="{}")
    payload = io.BytesIO()
    workbook.save(payload)

    with pytest.raises(ec.WorkbookError) as raised:
        master.validate_global(
            db,
            data=payload.getvalue(),
            allowed_project_ids={project_with_lines.project_id},
        )
    assert raised.value.code == "invalid_concurrency_token"


def test_global_lock_order_is_independent_of_excel_row_order(db, project_with_lines):
    line_ids = [line.id for line, _order, _project_id in master._assigned_lines(
        db, project_id=project_with_lines.project_id, window=None)]
    statements: list[str] = []

    def capture(_conn, _cursor, statement, _parameters, _context, _executemany):
        normalized = " ".join(statement.lower().split())
        if "for update" in normalized:
            statements.append(normalized)

    event.listen(db.get_bind(), "before_cursor_execute", capture)
    try:
        locked_lines, _overrides, _projects = master._lock_global_refill_rows(
            db,
            line_ids=list(reversed(line_ids)),
            allowed_project_ids={project_with_lines.project_id},
        )
    finally:
        event.remove(db.get_bind(), "before_cursor_execute", capture)
    assert list(locked_lines) == sorted(line_ids)

    def lock_index(table: str) -> int:
        return next(
            index for index, statement in enumerate(statements)
            if f"from {table} " in statement
        )

    assert (
        lock_index("f_maintenance_order")
        < lock_index("maintenance_source_order_assignment")
        < lock_index("f_maintenance_line")
        < lock_index("maintenance_manual_cost_override")
    )
    db.rollback()


def test_delete_snapshot_locks_assignment_before_detail_lines(
    db,
    project_with_lines,
):
    assignment = db.scalar(select(MaintenanceSourceOrderAssignment).where(
        MaintenanceSourceOrderAssignment.is_active.is_(True)
    ))
    assert assignment is not None
    statements: list[str] = []

    def capture(_conn, _cursor, statement, _parameters, _context, _executemany):
        normalized = " ".join(statement.lower().split())
        if "for update" in normalized:
            statements.append(normalized)

    event.listen(db.get_bind(), "before_cursor_execute", capture)
    try:
        snapshots = maintenance_demands._load_snapshots(
            db,
            [assignment.source_order_id],
            lock=True,
            active_only=True,
        )
    finally:
        event.remove(db.get_bind(), "before_cursor_execute", capture)
    assert set(snapshots) == {assignment.source_order_id}

    def lock_index(table: str) -> int:
        return next(
            index for index, statement in enumerate(statements)
            if f"from {table} " in statement
        )

    assert (
        lock_index("f_maintenance_order")
        < lock_index("maintenance_source_order_assignment")
        < lock_index("f_maintenance_line")
    )
    db.rollback()


def test_global_refill_waits_for_concurrent_void_then_fails_closed(
    db,
    project_with_lines,
):
    """Real PostgreSQL interleave: void wins; refill is a controlled conflict.

    The void transaction is paused while holding the WBDD header lock.  A
    global refill for one of its lines must wait on that same parent row rather
    than jumping ahead to assignment/detail locks.  After void commits, refill
    rechecks live assignment scope and returns ``project_scope_denied`` -- no
    deadlock, lock-timeout, raw DB exception, or last-write-wins mutation.
    """

    project_id = project_with_lines.project_id
    assignment = db.scalar(select(MaintenanceSourceOrderAssignment).where(
        MaintenanceSourceOrderAssignment.is_active.is_(True)
    ))
    assert assignment is not None
    source_order_id = assignment.source_order_id
    content = master.build_global_lines(
        db,
        preset="custom",
        date_from=date(2026, 1, 1),
        date_to=date(2026, 12, 31),
        allowed_project_ids={project_id},
    )
    workbook = load_workbook(io.BytesIO(content))
    workbook[master.GLOBAL_SHEET].cell(row=2, column=10, value=88)
    payload = io.BytesIO()
    workbook.save(payload)
    plan = master.validate_global(
        db,
        data=payload.getvalue(),
        allowed_project_ids={project_id},
    )
    # Do not retain a read transaction in the fixture session while the two
    # independent writer sessions exercise the real database locks.
    db.rollback()

    void_holds_order = threading.Event()
    release_void = threading.Event()
    refill_started = threading.Event()
    refill_reached_order_lock = threading.Event()

    def run_void() -> tuple[str, dict]:
        with SessionLocal() as session:
            connection = session.connection()
            connection.execute(text("SET LOCAL lock_timeout = '5s'"))
            connection.execute(text("SET LOCAL statement_timeout = '10s'"))
            paused = False

            def pause_after_order_lock(
                _conn, _cursor, statement, _parameters, _context, _executemany,
            ):
                nonlocal paused
                normalized = " ".join(statement.lower().split())
                if (
                    not paused
                    and "from f_maintenance_order " in normalized
                    and "for update" in normalized
                ):
                    paused = True
                    void_holds_order.set()
                    assert release_void.wait(timeout=10)

            event.listen(
                connection, "after_cursor_execute", pause_after_order_lock)
            try:
                result = maintenance_demands.void_fast(
                    session,
                    source_order_ids=[source_order_id],
                    reason="并发锁序回归测试",
                    operated_by="void-lock-order-test",
                    allowed_project_ids={project_id},
                    idempotency_key=f"void-lock-order:{uuid.uuid4()}",
                )
                session.commit()
                return "success", result
            except Exception:
                session.rollback()
                raise
            finally:
                event.remove(
                    connection, "after_cursor_execute", pause_after_order_lock)

    def run_refill() -> tuple[str, str | dict]:
        refill_started.set()
        with SessionLocal() as session:
            connection = session.connection()
            connection.execute(text("SET LOCAL lock_timeout = '5s'"))
            connection.execute(text("SET LOCAL statement_timeout = '10s'"))

            def observe_order_lock(
                _conn, _cursor, statement, _parameters, _context, _executemany,
            ):
                normalized = " ".join(statement.lower().split())
                if (
                    "from f_maintenance_order " in normalized
                    and "for update" in normalized
                ):
                    refill_reached_order_lock.set()

            event.listen(connection, "before_cursor_execute", observe_order_lock)
            try:
                result = master.apply_global_lines(
                    session,
                    plan,
                    operated_by="refill-lock-order-test",
                    import_batch_id=str(uuid.uuid4()),
                    allowed_project_ids={project_id},
                )
                return "success", result
            except ec.WorkbookError as exc:
                session.rollback()
                return "conflict", exc.code
            finally:
                event.remove(
                    connection, "before_cursor_execute", observe_order_lock)

    with ThreadPoolExecutor(max_workers=2) as executor:
        void_future = executor.submit(run_void)
        assert void_holds_order.wait(timeout=10)
        refill_future = executor.submit(run_refill)
        try:
            assert refill_started.wait(timeout=10)
            # Both writers take the same transaction advisory lock first.
            # While void owns it, refill cannot have reached the order-row lock.
            assert not refill_reached_order_lock.wait(timeout=0.5)
            assert not refill_future.done()
        finally:
            release_void.set()
        void_outcome = void_future.result(timeout=20)
        refill_outcome = refill_future.result(timeout=20)

    assert void_outcome[0] == "success"
    assert refill_outcome == ("conflict", "project_scope_denied")
    db.expire_all()
    assert db.get(MaintenanceDemandTombstone, source_order_id) is not None
    assert db.scalars(select(MaintenanceManualCostOverride)).all() == []


def test_bad_range_is_422(db):
    resp = uploader(db, "g-badrange").get(f"{_GLOBAL}.xlsx",
                                          params={"range": "last_century"})
    assert resp.status_code == 422
    assert resp.json()["detail"]["code"] == "invalid_range"


# ---------- 权限 ----------

def test_download_needs_profit_upload_needs_action_key(db, project_with_lines):
    content = _download_master(uploader(db, "m-perm-up"), project_with_lines)
    # 只读账号能下载（对账），但传不了
    assert reader(db, "m-perm-read").get(
        f"{_BASE}/{project_with_lines.project_id}/master-workbook.xlsx"
    ).status_code == 200
    assert _upload_master(reader(db, "m-perm-read2"), project_with_lines,
                          content).status_code == 403
    assert reader(db, "m-perm-read3").post(
        f"{_GLOBAL}/apply",
        files={"file": ("g.xlsx", io.BytesIO(content), _XLSX)}).status_code == 403
