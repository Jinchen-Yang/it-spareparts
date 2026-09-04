"""修复模式预检与门禁/删除侧同源（2026-09-04）。

此前预检只看表头：含撞键行的报销页报 can_import=true / severity=info，前端直接给
「开始导入」按钮，而 pipeline 会整批拒绝。这里钉住：预检的门禁结论、删除侧抑制
结论都来自 expense_void（与 loader/pipeline 同一份实现），且只在修复模式下解析。
"""
import io

import pytest
from openpyxl import Workbook

from app.etl import precheck as import_precheck
from app.etl import expense_void

_COLS = ["报销日期", "报销人员", "支出事由", "报销金额", "单号", "序号", "销售订单"]


def _xlsx(tmp_path, name, rows):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"; ws.append(_COLS)
    for r in rows:
        ws.append(r)
    p = tmp_path / name; wb.save(str(p)); return str(p)


def _anchored_xlsx(tmp_path, name, rows, anchor="XSDD-A"):
    """项目工作簿报销页形态：页级锚 + 无「销售订单」列。"""
    cols = _COLS[:-1]
    wb = Workbook(); ws = wb.active; ws.title = "报销明细"
    ws.append(["销售订单", anchor]); ws.append(cols)
    for r in rows:
        ws.append(r[:-1])
    p = tmp_path / name; wb.save(str(p)); return str(p)


def _codes(result):
    return [i["code"] for i in result["issues"]]


def test_skip_mode_does_not_parse_rows(tmp_path):
    p = _xlsx(tmp_path, "dup.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"],
        ["2026-05-01", "甲", "b", 200, "B1", 1, "XSDD-A"],       # duplicate_key
    ])
    r = import_precheck.inspect_file(p, "dup.xlsx", mode="skip")
    assert r["can_import"] is True and r["severity"] == "info"
    assert _codes(r) == []


def test_upsert_blocking_errors_now_block_at_precheck(tmp_path):
    """反向漂移修复：门禁会拒绝的文件，预检不能再说可导入。"""
    p = _xlsx(tmp_path, "dup.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"],
        ["2026-05-01", "甲", "b", 200, "B1", 1, "XSDD-A"],
    ])
    r = import_precheck.inspect_file(p, "dup.xlsx", mode="upsert")
    assert r["can_import"] is False and r["severity"] == "error"
    assert "upsert_blocking_errors" in _codes(r)
    msg = next(i["message"] for i in r["issues"] if i["code"] == "upsert_blocking_errors")
    assert "duplicate_key" in msg and "整批拒绝" in msg


def test_upsert_missing_link_is_not_blocking_but_reports_suppression(tmp_path):
    p = _xlsx(tmp_path, "mix.xlsx", [
        ["2026-05-01", "甲", "项目", 100, "B1", 1, "XSDD-A"],
        ["2026-05-02", "乙", "办公用品", 50, "B2", 1, None],      # missing_link
    ])
    r = import_precheck.inspect_file(p, "mix.xlsx", mode="upsert")
    assert r["can_import"] is True
    assert _codes(r) == ["upsert_void_suppressed_dropped"]
    # 用户选了修复模式，「你要的删除不会发生」必须被确认过：warning 而非 info
    assert r["severity"] == "warning"


def test_upsert_multi_contract_reports_suppression(tmp_path):
    p = _xlsx(tmp_path, "multi.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"],
        ["2026-05-01", "甲", "b", 100, "B2", 1, "XSDD-B"],
    ])
    r = import_precheck.inspect_file(p, "multi.xlsx", mode="upsert")
    assert r["can_import"] is True
    assert _codes(r) == ["upsert_void_suppressed_multi_contract"]
    assert r["severity"] == "warning"


def test_upsert_single_contract_anchored_clean_sheet_warns_that_void_is_armed(tmp_path):
    """唯一会真正作废的形态（单合同 + 页级锚 + 无错误行）必须在确认前被标成 warning。"""
    p = _anchored_xlsx(tmp_path, "single.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, None],
    ])
    r = import_precheck.inspect_file(p, "single.xlsx", mode="upsert")
    assert r["can_import"] is True and r["severity"] == "warning"
    assert _codes(r) == ["upsert_void_armed"]
    assert "XSDD-A" in r["issues"][0]["message"]


def test_upsert_unanchored_single_contract_reports_suppression(tmp_path):
    p = _xlsx(tmp_path, "single.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"],
    ])
    r = import_precheck.inspect_file(p, "single.xlsx", mode="upsert")
    assert _codes(r) == ["upsert_void_suppressed_unanchored"]
    assert r["severity"] == "warning"


def test_armed_is_not_emitted_alongside_blocking_errors(tmp_path):
    """单合同锚定页含撞键行：只报「整批拒绝」，不能同时说「无错误行…将作废」。"""
    p = _anchored_xlsx(tmp_path, "dup.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, None],
        ["2026-05-01", "甲", "b", 200, "B1", 1, None],
    ])
    r = import_precheck.inspect_file(p, "dup.xlsx", mode="upsert")
    assert _codes(r) == ["upsert_blocking_errors"]


def test_non_reader_exception_becomes_a_file_issue_not_a_500(tmp_path, monkeypatch):
    from app.etl import pipeline

    def _boom(*a, **k):
        raise RuntimeError("boom")

    monkeypatch.setattr(pipeline, "transform_workbook", _boom)
    p = _xlsx(tmp_path, "x.xlsx", [["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"]])
    r = import_precheck.inspect_file(p, "x.xlsx", mode="upsert")
    assert _codes(r) == ["upsert_precheck_failed"]
    assert r["can_import"] is False


def test_budget_is_per_request_across_files(tmp_path):
    """预算按整个请求累计：第二个文件超出剩余预算 ⇒ skipped，而不是各自按上限放行。"""
    rows = [["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"]]
    p1 = _xlsx(tmp_path, "f1.xlsx", rows); p2 = _xlsx(tmp_path, "f2.xlsx", rows)
    budget = {"rows": 1}
    r1 = import_precheck.inspect_file(p1, "f1.xlsx", mode="upsert", budget=budget)
    r2 = import_precheck.inspect_file(p2, "f2.xlsx", mode="upsert", budget=budget)
    assert _codes(r1) == ["upsert_void_suppressed_unanchored"]
    assert _codes(r2) == ["upsert_precheck_skipped"]
    assert budget["rows"] == 0


def test_duplicate_headers_are_reported_once(tmp_path):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(["报销日期", "报销金额", "销售订单", "销售订单"])
    ws.append(["2026-05-01", 100, "XSDD-A", "XSDD-A"])
    p = tmp_path / "dup_hdr.xlsx"; wb.save(str(p))
    r = import_precheck.inspect_file(str(p), "dup_hdr.xlsx", mode="upsert")
    assert r["severity"] == "error"
    assert _codes(r) == []                                   # 文件级不再重复上报
    assert [i["code"] for i in r["sheets"][0]["issues"]] == ["duplicate_headers"]


def test_upsert_precheck_row_cap_skips_parse_and_says_so(tmp_path, monkeypatch):
    p = _xlsx(tmp_path, "big.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"],
        ["2026-05-01", "甲", "b", 200, "B1", 1, "XSDD-A"],       # 撞键，但本次不解析
    ])
    r = import_precheck.inspect_file(p, "big.xlsx", mode="upsert", budget={"rows": 1})
    assert _codes(r) == ["upsert_precheck_skipped"]
    assert r["can_import"] is True and r["severity"] == "warning"


def test_precheck_uses_expense_void_not_a_private_copy(tmp_path, monkeypatch):
    """变异：放行集合里塞进 duplicate_key，预检必须跟着不拦——证明同源。"""
    p = _xlsx(tmp_path, "dup.xlsx", [
        ["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"],
        ["2026-05-01", "甲", "b", 200, "B1", 1, "XSDD-A"],
    ])
    monkeypatch.setattr(expense_void, "NON_BLOCKING_ERROR_TYPES",
                        frozenset({"missing_link", "duplicate_key"}))
    r = import_precheck.inspect_file(p, "dup.xlsx", mode="upsert")
    assert "upsert_blocking_errors" not in _codes(r)


def test_sheet_level_shape_is_unchanged_by_upsert_issues(tmp_path):
    """文件级 issues 可加；sheet 级键集不变（既有测试对 sheets 做精确相等断言）。"""
    p = _xlsx(tmp_path, "s.xlsx", [["2026-05-01", "甲", "a", 100, "B1", 1, "XSDD-A"]])
    r = import_precheck.inspect_file(p, "s.xlsx", mode="upsert")
    assert set(r["sheets"][0].keys()) == {
        "sheet_name", "detected_type", "action", "header_row", "data_rows",
        "duplicate_headers", "issues"}
