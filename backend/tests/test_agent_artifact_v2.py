"""Trusted Agent Artifact Delivery v2 backend contracts (GitHub #218)."""

from __future__ import annotations

import asyncio
import copy
import hashlib
import json
import os
import threading
import uuid
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import unquote

import pytest
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from PIL import Image
from sqlalchemy.exc import DBAPIError

from app import permissions, security
from app.auth import _make_token, hash_password
from app.config import Settings, get_settings
from app.main import app
from app.models.agent_artifact import AgentArtifact, AgentArtifactAudit
from app.models.system import SysUser
from app.services import agent_artifact_provenance, agent_files
from tests.artifact_test_support import force_artifact_state


def _login(db, client: TestClient, username: str, role: str = "sales") -> str:
    db.add(SysUser(username=username, role=role, password_hash=hash_password("pw123456")))
    db.commit()
    response = client.post(
        "/api/auth/login",
        json={"username": username, "password": "pw123456"},
    )
    assert response.status_code == 200, response.text
    return response.json()["token"]


def _owner_from_context(db, ctx: security.UserContext) -> agent_files.VerifiedArtifactOwner:
    user = db.query(SysUser).filter(SysUser.username == ctx.user_id).one_or_none()
    if user is None:
        user = SysUser(
            username=ctx.user_id,
            role=ctx.role,
            salesperson_name=ctx.salesperson_name,
            permissions=ctx.permissions,
            password_hash=hash_password("pw123456"),
            token_version=ctx.token_version or 0,
            is_active=True,
        )
        db.add(user)
    else:
        user.role = ctx.role
        user.salesperson_name = ctx.salesperson_name
        user.permissions = ctx.permissions
        user.is_active = True
    db.commit()
    current_permissions = permissions.runtime_safe(permissions.effective_for_user(user))
    current = security.UserContext(
        user_id=user.username,
        role=user.role,
        salesperson_name=user.salesperson_name,
        permissions=current_permissions,
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
        token_version=user.token_version or 0,
    )
    return agent_files.verified_artifact_owner(db, current)


def _verified_owner(db, username: str) -> agent_files.VerifiedArtifactOwner:
    user = db.query(SysUser).filter(SysUser.username == username).one_or_none()
    if user is None:
        user = SysUser(
            username=username,
            role="admin",
            permissions=permissions.effective("admin", None),
            password_hash=hash_password("pw123456"),
        )
        db.add(user)
        db.commit()
    return _owner_from_context(db, security.UserContext(
        user_id=user.username,
        role=user.role,
        salesperson_name=user.salesperson_name,
        permissions=permissions.runtime_safe(permissions.effective_for_user(user)),
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
        token_version=user.token_version or 0,
    ))


def _artifact_object_path(db, artifact_id: str):
    """Explicit test-only store seam for mutation/publication assertions."""
    artifact = db.get(AgentArtifact, artifact_id)
    assert artifact is not None
    return agent_files.get_artifact_store().path_for(artifact.storage_key)


def _write_legacy_fixture(legacy_id: str, meta: dict) -> None:
    """Test-only construction of pre-v2 forensic sidecars; production cannot write."""
    agent_files._meta_path(legacy_id).write_text(
        json.dumps(meta, ensure_ascii=False),
        encoding="utf-8",
    )


def _marker_workbook_bytes(marker: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = marker
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


@pytest.mark.parametrize("outage", ["root", "intermediate"])
def test_store_parent_outage_is_unknown_but_final_leaf_absence_is_invalid(
    tmp_path, outage
):
    root = tmp_path / "artifact-root"
    store = agent_files.LocalArtifactStore(root)
    key = f"objects/{uuid.uuid4()}.txt"
    store.publish_bytes(key, b"durable")
    if outage == "root":
        displaced = tmp_path / "root-offline"
        root.rename(displaced)
    else:
        displaced = root / "objects-offline"
        (root / "objects").rename(displaced)

    with pytest.raises(agent_files.ArtifactStoreUnavailable):
        store.read_bytes(key, max_bytes=1024)

    if outage == "root":
        displaced.rename(root)
    else:
        displaced.rename(root / "objects")
    store.path_for(key).unlink()
    with pytest.raises(agent_files.ArtifactObjectInvalid) as caught:
        store.read_bytes(key, max_bytes=1024)
    assert caught.value.reason_code == "object_missing"


def test_store_read_during_change_is_retryable_unknown(tmp_path, monkeypatch):
    store = agent_files.LocalArtifactStore(tmp_path / "changing-root")
    key = f"objects/{uuid.uuid4()}.txt"
    store.publish_bytes(key, b"changing-object")
    real_fstat = agent_files.os.fstat
    calls = 0

    def changed_after_read(fd):
        nonlocal calls
        state = real_fstat(fd)
        calls += 1
        if calls == 2:
            return SimpleNamespace(
                st_mode=state.st_mode,
                st_dev=state.st_dev,
                st_ino=state.st_ino,
                st_size=state.st_size,
                st_mtime_ns=state.st_mtime_ns + 1,
                st_ctime_ns=state.st_ctime_ns,
            )
        return state

    monkeypatch.setattr(agent_files.os, "fstat", changed_after_read)
    with pytest.raises(agent_files.ArtifactStoreUnavailable, match="读取期间"):
        store.read_bytes(key, max_bytes=1024)


def _marker_docx_bytes(marker: str) -> bytes:
    from docx import Document

    document = Document()
    document.add_paragraph(marker)
    buffer = BytesIO()
    document.save(buffer)
    return buffer.getvalue()


def _marker_png_bytes(color: str) -> bytes:
    buffer = BytesIO()
    Image.new("RGB", (8, 8), color).save(buffer, "PNG")
    return buffer.getvalue()


def test_verified_read_value_has_no_reopenable_path_seam():
    assert set(agent_files.StoredBytes.__dataclass_fields__) == {
        "content",
        "size_bytes",
        "sha256",
    }


def _write_test_report(
    owner: agent_files.VerifiedArtifactOwner,
    *,
    title: str | None,
    headers: list[str],
    rows: list[list],
    output_name: str,
    money_cols: list[int] | None = None,
    contained_resources: set[str] | None = None,
    contained_fields: set[str] | None = None,
):
    """Mint the non-production Query Broker evidence required by the v2 renderer."""
    evidence = agent_files._mint_report_provenance(
        owner,
        title=title,
        headers=headers,
        rows=rows,
        output_name=output_name,
        money_cols=money_cols,
        contained_resources=contained_resources or set(),
        contained_fields=contained_fields or set(),
    )
    return agent_files.write_report(
        title,
        headers,
        rows,
        output_name,
        owner,
        money_cols=money_cols,
        provenance=evidence,
    )


def _write_test_excel(
    owner: agent_files.VerifiedArtifactOwner,
    *,
    source_ids: list[str],
    base_file_id: str | None,
    sheet: str | None,
    cells: list[dict],
    output_name: str,
):
    """Bind an edit to explicit, already-authorized Artifact sources in tests."""
    evidence = agent_files._mint_excel_from_artifacts(
        owner,
        source_ids=source_ids,
        base_file_id=base_file_id,
        sheet=sheet,
        cells=cells,
        output_name=output_name,
    )
    return agent_files.write_excel(
        base_file_id,
        sheet,
        cells,
        output_name,
        owner,
        provenance=evidence,
    )


def test_artifact_v2_kill_switch_defaults_off(monkeypatch):
    monkeypatch.delenv("AGENT_ARTIFACT_V2_ENABLED", raising=False)
    assert Settings(_env_file=None).agent_artifact_v2_enabled is False


def test_disabled_v2_returns_503_and_legacy_sidecars_remain_denied(db, monkeypatch):
    client = TestClient(app)
    token = _login(db, client, "switch_owner")
    headers = {"Authorization": f"Bearer {token}"}
    settings = get_settings()
    monkeypatch.setattr(settings, "agent_artifact_v2_enabled", True)
    created = client.post(
        "/api/agent/upload",
        headers=headers,
        files={"file": ("v2.txt", b"v2", "text/plain")},
    )
    assert created.status_code == 200, created.text
    artifact_id = created.json()["file_id"]

    legacy_id = "abcdef123456"
    agent_files._data_path(legacy_id, "txt").write_bytes(b"legacy")
    _write_legacy_fixture(legacy_id, {
        "filename": "legacy.txt",
        "ext": "txt",
        "kind": "upload",
        "operated_by": "switch_owner",
        "created_at": datetime.now(timezone.utc).isoformat(),
    })

    monkeypatch.setattr(settings, "agent_artifact_v2_enabled", False)
    disabled_message = "Artifact Delivery v2 已停用"
    for path in (
        f"/api/agent/files/{artifact_id}",
        f"/api/agent/files/{artifact_id}/preview",
    ):
        response = client.get(path, headers=headers)
        assert response.status_code == 503
        assert response.json()["detail"] == disabled_message
    with pytest.raises(agent_files.ArtifactV2Disabled, match=disabled_message):
        agent_files.read_document(artifact_id, _verified_owner(db, "switch_owner"))

    denied_create = client.post(
        "/api/agent/upload",
        headers=headers,
        files={"file": ("disabled.txt", b"disabled", "text/plain")},
    )
    assert denied_create.status_code == 503
    assert denied_create.json()["detail"] == disabled_message

    assert client.get(f"/api/agent/files/{legacy_id}", headers=headers).status_code == 404
    assert client.get(f"/api/agent/files/{legacy_id}/preview", headers=headers).status_code == 404


def test_upload_persists_complete_ready_artifact_metadata(db):
    content = "可信制品".encode()
    owner = _verified_owner(db, "alice")
    result = agent_files.save_upload(content, "核验记录.txt", owner)

    artifact_id = result["file_id"]
    assert str(uuid.UUID(artifact_id)) == artifact_id
    assert result["artifact"]["id"] == artifact_id
    db.expire_all()
    artifact = db.get(AgentArtifact, artifact_id)
    assert artifact is not None
    assert artifact.owner_sub == "alice"
    assert artifact.filename == "核验记录.txt"
    assert artifact.media_type == "text/plain; charset=utf-8"
    assert artifact.size_bytes == len(content)
    assert artifact.sha256 == hashlib.sha256(content).hexdigest()
    assert artifact.status == "ready"
    assert artifact.sensitivity == "critical"
    assert artifact.storage_key == f"objects/{artifact.id}.txt"
    assert artifact.source_ids == []
    assert artifact.access_scope["policy"] == "owner_only"
    assert artifact.created_at is not None
    assert artifact.expires_at > artifact.created_at
    audits = db.query(AgentArtifactAudit).filter_by(artifact_id=artifact_id).order_by(
        AgentArtifactAudit.id
    ).all()
    assert [(audit.action, audit.from_status, audit.to_status) for audit in audits] == [
        ("artifact_created", None, "prepared"),
        ("status_transition", "prepared", "validating"),
        ("status_transition", "validating", "ready"),
    ]
    assert result["artifact"]["sha256"] == artifact.sha256
    assert result["artifact"]["mime_type"] == artifact.media_type
    assert result["artifact"]["download_url"].endswith(artifact.id)
    assert not agent_files._meta_path(result["file_id"]).exists()
    assert not (agent_files._dir() / f"{result['file_id']}.txt").exists()
    assert _artifact_object_path(db, artifact.id).parent.name == "objects"


def test_publish_rejects_non_object_extra_meta_before_creating_row(db):
    owner = _verified_owner(db, "bad-extra-meta")
    before = db.query(AgentArtifact).count()

    with pytest.raises(agent_files.FileError, match="extra_meta.*JSON 对象"):
        agent_files._publish_artifact(
            b"bad-meta",
            "bad-meta.txt",
            kind="upload",
            owner=owner,
            extra_meta=[],  # type: ignore[arg-type]
        )

    assert db.query(AgentArtifact).count() == before


def test_upload_and_generated_output_are_immutable_distinct_artifacts(db):
    owner = _verified_owner(db, "alice")
    base = agent_files.save_upload(b"source", "source.txt", owner)
    original_hash = agent_files.get_download_info(base["file_id"], owner).sha256
    business_source = _write_test_report(
        owner,
        title="source",
        headers=["value"],
        rows=[["trusted"]],
        output_name="source-business.xlsx",
    )

    first = _write_test_excel(
        owner,
        source_ids=[business_source["file_id"]],
        base_file_id=None,
        sheet=None,
        cells=[{"row": 1, "col": "A", "value": "first"}],
        output_name="result.xlsx",
    )
    second = _write_test_excel(
        owner,
        source_ids=[business_source["file_id"]],
        base_file_id=None,
        sheet=None,
        cells=[{"row": 1, "col": "A", "value": "second"}],
        output_name="result.xlsx",
    )

    assert len({base["file_id"], first["file_id"], second["file_id"]}) == 3
    assert base["artifact"]["id"] == base["file_id"]
    assert first["artifact"]["id"] == first["file_id"]
    assert second["artifact"]["id"] == second["file_id"]
    assert agent_files.get_download_info(base["file_id"], owner).sha256 == original_hash


def test_generated_artifact_records_source_ids(db):
    owner = _verified_owner(db, "alice")
    base = _write_test_report(
        owner,
        title="模板",
        headers=["值"],
        rows=[["source"]],
        output_name="模板.xlsx",
    )

    result = _write_test_excel(
        owner,
        source_ids=[base["file_id"]],
        base_file_id=base["file_id"],
        sheet=None,
        cells=[{"row": 1, "col": "A", "value": "done"}],
        output_name="已回填.xlsx",
    )

    db.expire_all()
    artifact = db.get(AgentArtifact, result["artifact"]["id"])
    assert artifact is not None
    assert artifact.source_ids == [base["file_id"]]
    assert artifact.kind == "generated"


def test_service_write_rejects_cross_owner_and_revoked_source_artifacts(db):
    alice = security.UserContext(
        user_id="alice",
        role="boss",
        permissions=permissions.effective("boss", None),
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    bob = security.UserContext(
        user_id="bob",
        role="boss",
        permissions=permissions.effective("boss", None),
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    private_base = agent_files.save_upload(
        b"private", "private.txt", _owner_from_context(db, alice)
    )
    with pytest.raises(agent_files.FileError, match="无权引用来源制品"):
        agent_files.write_excel(
            private_base["file_id"], None, [{"row": 1, "col": "A", "value": 1}],
            "cross-owner.xlsx", _owner_from_context(db, bob),
        )

    generated = _write_test_report(
        _owner_from_context(db, alice),
        title="full",
        headers=["value"],
        rows=[[1]],
        output_name="full.xlsx",
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
    )
    revoked = security.UserContext(
        user_id="alice",
        role="boss",
        permissions={key: False for key in permissions.ALL_KEYS},
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    with pytest.raises(agent_files.FileError, match="无权引用来源制品"):
        agent_files.write_excel(
            generated["file_id"], None, [{"row": 1, "col": "A", "value": 1}],
            "revoked.xlsx", _owner_from_context(db, revoked),
        )


def test_failed_atomic_publish_is_not_downloadable_and_cleans_temp(db, tmp_path, monkeypatch):
    class FailAfterStageStore(agent_files.LocalArtifactStore):
        def publish_bytes(self, storage_key, content, *, validator=None):
            def fail(_path):
                raise OSError("simulated validation failure")

            return super().publish_bytes(storage_key, content, validator=fail)

    store = FailAfterStageStore(tmp_path)
    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: store)

    owner = _verified_owner(db, "failure-owner")
    with pytest.raises(agent_files.FileError, match="发布失败"):
        agent_files.save_upload(
            b"not published", "failure.txt", owner
        )

    db.expire_all()
    artifact = (
        db.query(AgentArtifact)
        .filter(AgentArtifact.owner_sub == "failure-owner")
        .order_by(AgentArtifact.created_at.desc())
        .first()
    )
    assert artifact is not None and artifact.status == "failed"
    with pytest.raises(agent_files.FileError):
        agent_files.get_download_info(artifact.id, owner)
    assert not list(tmp_path.rglob("*.part"))
    assert not list(tmp_path.glob("*.txt"))


def test_local_store_concurrent_publish_never_overwrites_existing_key(tmp_path, monkeypatch):
    store = agent_files.LocalArtifactStore(tmp_path)
    storage_key = "objects/11111111-1111-4111-8111-111111111111.txt"
    final_path = store.path_for(storage_key)
    barrier = threading.Barrier(2)
    original_exists = type(final_path).exists

    def synchronized_exists(path):
        if path == final_path:
            barrier.wait(timeout=5)
            return False
        return original_exists(path)

    monkeypatch.setattr(type(final_path), "exists", synchronized_exists)

    def publish(payload: bytes):
        try:
            return store.publish_bytes(storage_key, payload)
        except Exception as exc:  # noqa: BLE001 - result is asserted below
            return exc

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(publish, (b"first", b"second")))

    successes = [item for item in outcomes if isinstance(item, agent_files.StoredObject)]
    failures = [item for item in outcomes if isinstance(item, Exception)]
    assert len(successes) == 1
    assert len(failures) == 1
    assert isinstance(failures[0], agent_files.FileError)
    assert final_path.read_bytes() in {b"first", b"second"}


def test_non_xlsx_upload_validates_only_store_tracked_part(db, monkeypatch):
    seen_names = []
    real_validate = agent_files._validate_staged_file

    def capture(path, ext):
        seen_names.append(path.name)
        return real_validate(path, ext)

    monkeypatch.setattr(agent_files, "_validate_staged_file", capture)

    agent_files.save_upload(b"tracked", "tracked.txt", _verified_owner(db, "tracked-owner"))

    assert len(seen_names) == 1
    assert seen_names[0].startswith("artifact-")
    assert seen_names[0].endswith(".part")
    assert "upload-check" not in seen_names[0]


def test_non_ready_expired_and_integrity_mismatch_are_denied(db):
    owner = _verified_owner(db, "alice")
    failed = agent_files.save_upload(b"failed", "failed.txt", owner)
    expired = agent_files.save_upload(b"expired", "expired.txt", owner)
    corrupted = agent_files.save_upload(b"trusted", "trusted.txt", owner)

    failed_row = db.get(AgentArtifact, failed["artifact"]["id"])
    expired_row = db.get(AgentArtifact, expired["artifact"]["id"])
    assert failed_row is not None and expired_row is not None
    failed_row = force_artifact_state(db, failed_row, "failed")
    expired_row = force_artifact_state(
        db,
        expired_row,
        "ready",
        created_at=datetime.now(timezone.utc) - timedelta(seconds=2),
        expires_at=datetime.now(timezone.utc) - timedelta(seconds=1),
    )

    with pytest.raises(agent_files.FileError):
        agent_files.get_download_info(failed["file_id"], owner)
    with pytest.raises(agent_files.FileError):
        agent_files.preview(expired["file_id"], owner)

    corrupt_path = _artifact_object_path(db, corrupted["file_id"])
    corrupt_path.write_bytes(b"tampered")
    with pytest.raises(agent_files.FileError, match="完整性"):
        agent_files.get_download_info(corrupted["file_id"], owner)


def test_storage_key_tampering_cannot_escape_store_root(db, tmp_path):
    owner = _verified_owner(db, "alice")
    result = agent_files.save_upload(b"safe", "safe.txt", owner)
    artifact = db.get(AgentArtifact, result["artifact"]["id"])
    assert artifact is not None
    artifact.storage_key = "../outside-secret.txt"
    db.commit()
    (agent_files._dir().parent / "outside-secret.txt").write_text("secret", encoding="utf-8")

    with pytest.raises(agent_files.FileError):
        agent_files.get_download_info(result["file_id"], owner)


def test_download_uses_real_mime_hash_length_and_safe_unicode_filename(db):
    client = TestClient(app)
    token = _login(db, client, "artifact_alice")
    raw_name = "../报价\r\nX-Evil: yes.txt"
    result = agent_files.save_upload(
        "报价".encode(), raw_name, _verified_owner(db, "artifact_alice")
    )

    response = client.get(
        f"/api/agent/files/{result['file_id']}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "text/plain; charset=utf-8"
    assert response.headers["content-length"] == str(len(response.content))
    assert response.headers["etag"] == f'"{hashlib.sha256(response.content).hexdigest()}"'
    disposition = unquote(response.headers["content-disposition"])
    assert "报价" in disposition
    assert "../" not in disposition
    assert "\r" not in disposition and "\n" not in disposition
    assert "X-Evil:" not in disposition
    assert response.headers["x-content-type-options"] == "nosniff"


def _image_bytes(image_format: str) -> bytes:
    output = BytesIO()
    Image.new("RGB", (2, 2), color=(12, 34, 56)).save(output, format=image_format)
    return output.getvalue()


@pytest.mark.parametrize(
    ("filename", "content", "expected"),
    [
        ("proof.png", _image_bytes("PNG"), "image/png"),
        ("proof.jpg", _image_bytes("JPEG"), "image/jpeg"),
        (
            "proof.xlsx",
            agent_artifact_provenance.canonical_identity_template_bytes(
                "pn-replenishment-request", 1
            ),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    ],
)
def test_download_projects_only_server_allowlisted_real_mime_types(
    db, filename, content, expected
):
    client = TestClient(app)
    token = _login(db, client, f"mime-{filename.rsplit('.', 1)[-1]}")
    owner = _verified_owner(db, f"mime-{filename.rsplit('.', 1)[-1]}")
    result = agent_files.save_upload(content, filename, owner)

    response = client.get(
        f"/api/agent/files/{result['file_id']}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == expected
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.content == content


def test_unknown_persisted_mime_tamper_denies_then_resigned_value_degrades(db):
    client = TestClient(app)
    token = _login(db, client, "mime-unknown")
    owner = _verified_owner(db, "mime-unknown")
    result = agent_files.save_upload(b"opaque", "opaque.txt", owner)
    row = db.get(AgentArtifact, result["file_id"])
    assert row is not None
    row.media_type = "application/x-attacker-controlled"
    db.commit()

    denied = client.get(
        f"/api/agent/files/{result['file_id']}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert denied.status_code == 404

    row = force_artifact_state(
        db,
        row,
        "ready",
        media_type="application/x-attacker-controlled",
    )
    response = client.get(
        f"/api/agent/files/{row.id}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/octet-stream"
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.content == b"opaque"


@pytest.mark.parametrize("attack", ["replace", "mutate"])
def test_download_returns_same_handle_verified_bytes_despite_path_race(
    db, monkeypatch, tmp_path, attack
):
    from app.api import agent as agent_api

    client = TestClient(app)
    token = _login(db, client, f"download-race-{attack}")
    owner = _verified_owner(db, f"download-race-{attack}")
    original = b"trusted-original"
    result = agent_files.save_upload(original, "race.txt", owner)
    row = db.get(AgentArtifact, result["file_id"])
    assert row is not None
    path = agent_files.get_artifact_store().path_for(row.storage_key)

    def race(_ctx, action, outcome, **_kwargs):
        if action != "download" or outcome != "success":
            return
        if attack == "replace":
            replacement = tmp_path / "replacement"
            replacement.write_bytes(b"attacker-replace")
            os.replace(replacement, path)
        else:
            path.write_bytes(b"attacker-mutate")

    monkeypatch.setattr(agent_api, "_audit_artifact_access", race)

    response = client.get(
        f"/api/agent/files/{result['file_id']}",
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 200
    assert response.content == original
    assert response.headers["etag"] == f'"{hashlib.sha256(original).hexdigest()}"'


def test_download_rejects_same_content_symlink_and_enforces_size_budget(
    db, monkeypatch, tmp_path
):
    owner = _verified_owner(db, "download-bounds")
    result = agent_files.save_upload(b"same-content", "bounded.txt", owner)
    row = db.get(AgentArtifact, result["file_id"])
    assert row is not None
    path = agent_files.get_artifact_store().path_for(row.storage_key)
    target = tmp_path / "same-content.txt"
    target.write_bytes(b"same-content")
    path.unlink()
    path.symlink_to(target)

    with pytest.raises(agent_files.ArtifactUnavailable):
        agent_files.get_download_info(result["file_id"], owner)

    path.unlink()
    path.write_bytes(b"same-content")
    monkeypatch.setattr(agent_files, "_MAX_DOWNLOAD_BYTES", 4)
    with pytest.raises(agent_files.ArtifactUnavailable, match="大小"):
        agent_files.get_download_info(result["file_id"], owner)


@pytest.mark.parametrize(
    "consumer",
    ["inspect", "rows", "preview", "document", "write_excel"],
)
@pytest.mark.parametrize("attack", ["replace", "mutate", "symlink"])
def test_xlsx_consumers_never_reopen_the_verified_artifact_path(
    db, monkeypatch, tmp_path, consumer, attack
):
    owner = _verified_owner(db, f"xlsx-one-read-{consumer}-{attack}")
    trusted = "trusted-xlsx-marker"
    attacker = "attacker-xlsx-marker"
    base = _write_test_report(
        owner,
        title="immutable input",
        headers=["marker"],
        rows=[[trusted]],
        output_name="immutable-base.xlsx",
        contained_resources={"purchases"},
    )
    path = _artifact_object_path(db, base["file_id"])
    attacker_bytes = _marker_workbook_bytes(attacker)
    target = tmp_path / f"xlsx-race-{consumer}-{attack}.xlsx"
    target.write_bytes(attacker_bytes)
    real_load_workbook = agent_files.load_workbook
    raced = False

    def race_path_once() -> None:
        nonlocal raced
        if raced:
            return
        raced = True
        if attack == "replace":
            os.replace(target, path)
        elif attack == "mutate":
            path.write_bytes(attacker_bytes)
        else:
            path.unlink()
            path.symlink_to(target)

    def racing_load_workbook(source, *args, **kwargs):
        race_path_once()
        return real_load_workbook(source, *args, **kwargs)

    monkeypatch.setattr(agent_files, "load_workbook", racing_load_workbook)

    if consumer == "inspect":
        result = agent_files.inspect_file(base["file_id"], owner)
        rendered = repr(result)
    elif consumer == "rows":
        result = agent_files.read_rows(base["file_id"], None, 1, 10, owner)
        rendered = repr(result)
    elif consumer == "preview":
        result = agent_files.preview(base["file_id"], owner)
        rendered = repr(result)
    elif consumer == "document":
        result = agent_files.read_document(base["file_id"], owner)
        rendered = result["content"]
    else:
        cells = [{"row": 3, "col": "B", "value": "edited"}]
        evidence = agent_files._mint_excel_from_artifacts(
            owner,
            source_ids=[base["file_id"]],
            base_file_id=base["file_id"],
            sheet=None,
            cells=cells,
            output_name="immutable-result.xlsx",
        )
        captured: dict[str, bytes] = {}

        def capture_publish(content, filename, **_kwargs):
            captured["content"] = content
            return {
                "file_id": str(uuid.uuid4()),
                "filename": filename,
                "media_type": agent_files._MIME_BY_EXT["xlsx"],
                "size_bytes": len(content),
                "sha256": hashlib.sha256(content).hexdigest(),
                "status": "ready",
                "sensitivity": "high",
            }

        monkeypatch.setattr(agent_files, "_publish_artifact", capture_publish)
        result = agent_files.write_excel(
            base["file_id"],
            None,
            cells,
            "immutable-result.xlsx",
            owner,
            provenance=evidence,
        )
        workbook = load_workbook(BytesIO(captured["content"]), data_only=False)
        rendered = repr([
            workbook.worksheets[0]["A3"].value,
            workbook.worksheets[0]["B3"].value,
        ])
        workbook.close()

    assert raced is True
    assert trusted in rendered
    assert attacker not in rendered


@pytest.mark.parametrize("file_kind", ["txt", "docx", "pdf", "png"])
@pytest.mark.parametrize("attack", ["replace", "mutate", "symlink"])
def test_read_document_consumes_verified_bytes_for_every_supported_path_format(
    db, monkeypatch, tmp_path, file_kind, attack
):
    owner = _verified_owner(db, f"document-one-read-{file_kind}-{attack}")
    trusted = f"trusted-{file_kind}-marker"
    attacker = f"attacker-{file_kind}-marker"
    if file_kind == "txt":
        trusted_bytes = trusted.encode()
        attacker_bytes = attacker.encode()
    elif file_kind == "docx":
        trusted_bytes = _marker_docx_bytes(trusted)
        attacker_bytes = _marker_docx_bytes(attacker)
    elif file_kind == "pdf":
        trusted_bytes = f"%PDF-1.4\n{trusted}".encode()
        attacker_bytes = f"%PDF-1.4\n{attacker}".encode()
    else:
        trusted_bytes = _marker_png_bytes("green")
        attacker_bytes = _marker_png_bytes("red")

    created = agent_files.save_upload(
        trusted_bytes,
        f"immutable.{file_kind}",
        owner,
    )
    path = _artifact_object_path(db, created["file_id"])
    target = tmp_path / f"document-race-{file_kind}-{attack}.{file_kind}"
    target.write_bytes(attacker_bytes)
    raced = False

    def race_path_once() -> None:
        nonlocal raced
        if raced:
            return
        raced = True
        if attack == "replace":
            os.replace(target, path)
        elif attack == "mutate":
            path.write_bytes(attacker_bytes)
        else:
            path.unlink()
            path.symlink_to(target)

    delegate = agent_files.get_artifact_store()

    class RacingStore:
        def read_bytes(self, storage_key, *, max_bytes):
            stored = delegate.read_bytes(storage_key, max_bytes=max_bytes)
            race_path_once()
            return stored

    monkeypatch.setattr(agent_files, "get_artifact_store", RacingStore)

    if file_kind == "pdf":
        def read_test_pdf(source):
            data = source.read() if hasattr(source, "read") else Path(source).read_bytes()
            return data.decode("utf-8"), False

        monkeypatch.setattr(agent_files, "_read_pdf", read_test_pdf)
    elif file_kind == "png":
        def hash_vision_input(source, _hint):
            return hashlib.sha256(Path(source).read_bytes()).hexdigest()

        monkeypatch.setattr(agent_files, "_read_image_or_scanned", hash_vision_input)

    result = agent_files.read_document(created["file_id"], owner)

    assert raced is True
    if file_kind == "png":
        assert result["content"] == hashlib.sha256(trusted_bytes).hexdigest()
    else:
        assert trusted in result["content"]
        assert attacker not in result["content"]


def test_download_and_preview_are_owner_only_even_for_admin_and_boss(db):
    client = TestClient(app)
    owner_token = _login(db, client, "artifact_owner")
    admin_token = _login(db, client, "artifact_admin", "admin")
    boss_token = _login(db, client, "artifact_boss", "boss")
    upload = agent_files.save_upload(
        b"owner only", "owner.txt", _verified_owner(db, "artifact_owner")
    )
    paths = [
        f"/api/agent/files/{upload['file_id']}",
        f"/api/agent/files/{upload['artifact']['id']}",
        f"/api/agent/files/{upload['file_id']}/preview",
        f"/api/agent/files/{upload['artifact']['id']}/preview",
    ]

    for path in paths:
        assert client.get(path, headers={"Authorization": f"Bearer {owner_token}"}).status_code == 200
        admin_response = client.get(path, headers={"Authorization": f"Bearer {admin_token}"})
        boss_response = client.get(path, headers={"Authorization": f"Bearer {boss_token}"})
        assert admin_response.status_code == 404
        assert boss_response.status_code == 404
        assert admin_response.json() == boss_response.json() == {
            "detail": "文件不存在或无权访问"
        }


def test_same_owner_scope_transplant_cannot_relabel_another_output(db):
    owner = _verified_owner(db, "binding-same-owner")
    narrow = _write_test_report(
        owner,
        title="narrow",
        headers=["public"],
        rows=[["narrow"]],
        output_name="narrow.xlsx",
        contained_resources={"parts"},
    )
    sensitive = _write_test_report(
        owner,
        title="sensitive",
        headers=["cost"],
        rows=[["sensitive"]],
        output_name="sensitive.xlsx",
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
    )
    narrow_row = db.get(AgentArtifact, narrow["file_id"])
    sensitive_row = db.get(AgentArtifact, sensitive["file_id"])
    assert narrow_row is not None and sensitive_row is not None
    assert narrow_row.access_scope != sensitive_row.access_scope
    sensitive_row.access_scope = copy.deepcopy(narrow_row.access_scope)
    sensitive_row.source_ids = list(narrow_row.source_ids)
    sensitive_row.sensitivity = narrow_row.sensitivity
    sensitive_row.binding_envelope = copy.deepcopy(narrow_row.binding_envelope)
    db.commit()

    with pytest.raises(agent_files.ArtifactUnavailable):
        agent_files.get_download_info(sensitive["file_id"], owner)


def test_cross_owner_scope_and_owner_transplant_cannot_relabel_output(db):
    owner_a = _verified_owner(db, "binding-owner-a")
    owner_b = _verified_owner(db, "binding-owner-b")
    artifact_a = _write_test_report(
        owner_a,
        title="owner-a-sensitive",
        headers=["cost"],
        rows=[["owner-a"]],
        output_name="owner-a.xlsx",
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
    )
    artifact_b = _write_test_report(
        owner_b,
        title="owner-b-narrow",
        headers=["public"],
        rows=[["owner-b"]],
        output_name="owner-b.xlsx",
        contained_resources={"parts"},
    )
    row_a = db.get(AgentArtifact, artifact_a["file_id"])
    row_b = db.get(AgentArtifact, artifact_b["file_id"])
    assert row_a is not None and row_b is not None
    row_a.owner_sub = row_b.owner_sub
    row_a.access_scope = copy.deepcopy(row_b.access_scope)
    row_a.source_ids = list(row_b.source_ids)
    row_a.sensitivity = row_b.sensitivity
    row_a.binding_envelope = copy.deepcopy(row_b.binding_envelope)
    db.commit()

    with pytest.raises(agent_files.ArtifactUnavailable):
        agent_files.get_download_info(artifact_a["file_id"], owner_b)


def test_source_ids_column_must_exactly_match_verified_scope_order(db):
    owner = _verified_owner(db, "binding-source-order")
    base = _write_test_report(
        owner,
        title="source",
        headers=["cost"],
        rows=[["source"]],
        output_name="source.xlsx",
        contained_resources={"purchases"},
        contained_fields={"purchase_cost"},
    )
    derived = _write_test_excel(
        owner,
        source_ids=[base["file_id"]],
        base_file_id=base["file_id"],
        sheet=None,
        cells=[{"row": 3, "col": "B", "value": "derived"}],
        output_name="derived.xlsx",
    )
    row = db.get(AgentArtifact, derived["file_id"])
    assert row is not None and row.source_ids == [base["file_id"]]
    row.source_ids = []
    db.commit()

    with pytest.raises(agent_files.ArtifactUnavailable):
        agent_files.get_download_info(derived["file_id"], owner)


@pytest.mark.parametrize("sealed_status", ["prepared", "validating", "failed"])
def test_nonready_status_cannot_be_tampered_to_ready_without_resigning(
    db, sealed_status
):
    owner = _verified_owner(db, f"binding-status-{sealed_status}")
    created = agent_files.save_upload(
        b"status-bound",
        "status-bound.txt",
        owner,
    )
    row = db.get(AgentArtifact, created["file_id"])
    assert row is not None
    row = force_artifact_state(db, row, sealed_status)

    row.status = "ready"
    with pytest.raises(DBAPIError):
        db.commit()
    db.rollback()

    with pytest.raises(agent_files.ArtifactUnavailable):
        agent_files.get_download_info(created["file_id"], owner)


def test_lazy_ready_to_expired_transition_resigns_binding_with_audit(db):
    owner = _verified_owner(db, "binding-expiry")
    created = agent_files.save_upload(b"expires", "expires.txt", owner)
    row = db.get(AgentArtifact, created["file_id"])
    assert row is not None and row.status == "ready"
    expired_at = datetime.now(timezone.utc) - timedelta(seconds=1)
    row = force_artifact_state(
        db,
        row,
        "ready",
        created_at=expired_at - timedelta(seconds=1),
        expires_at=expired_at,
    )

    with pytest.raises(agent_files.ArtifactUnavailable) as caught:
        agent_files.get_download_info(created["file_id"], owner)
    assert caught.value.reason_code == "expired"

    db.expire_all()
    row = db.get(AgentArtifact, created["file_id"])
    assert row is not None and row.status == "expired"
    agent_artifact_provenance.verify_artifact_binding(
        row.binding_envelope,
        agent_files._binding_metadata_from_row(row),
    )
    audit = db.query(AgentArtifactAudit).filter_by(
        artifact_id=created["file_id"],
        action="status_transition",
        from_status="ready",
        to_status="expired",
    ).one()
    assert audit.outcome == "success"


@pytest.mark.parametrize("payload", ["=1+1", "+SUM(A1:A2)", "-2+3", "@cmd", "  =HYPERLINK(\"x\")"])
def test_write_excel_neutralizes_formula_injection(db, payload):
    owner = _verified_owner(db, "alice")
    source = _write_test_report(
        owner,
        title="source",
        headers=["value"],
        rows=[["trusted"]],
        output_name="formula-source.xlsx",
    )
    result = _write_test_excel(
        owner,
        source_ids=[source["file_id"]],
        base_file_id=None,
        sheet=None,
        cells=[{"row": 1, "col": "A", "value": payload}],
        output_name="safe.xlsx",
    )
    download = agent_files.get_download_info(result["file_id"], owner)
    workbook = load_workbook(BytesIO(download.content), data_only=False)
    cell = workbook.active["A1"]
    assert cell.value == f"'{payload}"
    assert cell.data_type != "f"
    workbook.close()


def test_write_report_neutralizes_formula_injection_in_all_text_fields(db):
    owner = _verified_owner(db, "alice")
    result = _write_test_report(
        owner,
        title="=TITLE",
        headers=["@HEADER"],
        rows=[["+ROW"]],
        output_name="report.xlsx",
    )
    download = agent_files.get_download_info(result["file_id"], owner)
    workbook = load_workbook(BytesIO(download.content), data_only=False)
    sheet = workbook.active
    assert sheet["A1"].value == "'=TITLE"
    assert sheet["A2"].value == "'@HEADER"
    assert sheet["A3"].value == "'+ROW"
    assert all(sheet[cell].data_type != "f" for cell in ("A1", "A2", "A3"))
    workbook.close()


def test_generated_artifact_rechecks_current_scope_after_account_downgrade(db):
    creator_permissions = permissions.effective("sales", None)
    creator = security.UserContext(
        user_id="scoped-user",
        role="sales",
        salesperson_name="Alice Sales",
        permissions=creator_permissions,
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    result = _write_test_report(
        _owner_from_context(db, creator),
        title="含成本结果",
        headers=["成本", "毛利"],
        rows=[[100, 20]],
        output_name="scoped.xlsx",
        contained_resources={"purchases"},
        contained_fields={"purchase_cost", "profit_amount"},
    )
    artifact = db.get(AgentArtifact, result["artifact"]["id"])
    assert artifact is not None
    assert artifact.access_scope["policy"] == "provenance_guarded"
    assert "data_purchase_cost" in artifact.access_scope["required_permissions"]
    assert "data_profit" in artifact.access_scope["required_permissions"]
    assert "page_purchases" in artifact.access_scope["required_permissions"]
    assert artifact.sensitivity == "high"
    assert artifact.access_scope["contained_fields"] == ["profit_amount", "purchase_cost"]
    assert len(artifact.access_scope["source_access_snapshots"]) == 1
    assert agent_files.access_allowed(result["file_id"], creator) is True

    downgraded = security.UserContext(
        user_id="scoped-user",
        role="sales",
        salesperson_name="Alice Sales",
        permissions={**creator_permissions, "data_purchase_cost": False, "data_profit": False},
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    assert agent_files.access_allowed(result["file_id"], downgraded) is False
    assert agent_files.access_allowed(result["artifact"]["id"], downgraded) is False

    user = db.query(SysUser).filter_by(username="scoped-user").one()
    user.permissions = downgraded.permissions
    db.commit()
    client = TestClient(app)
    login = client.post(
        "/api/auth/login",
        json={"username": "scoped-user", "password": "pw123456"},
    )
    assert login.status_code == 200
    response = client.get(
        f"/api/agent/files/{result['file_id']}",
        headers={"Authorization": f"Bearer {login.json()['token']}"},
    )
    assert response.status_code == 404
    db.expire_all()
    denied = db.query(AgentArtifactAudit).filter_by(
        artifact_id=result["file_id"],
        action="http_download",
        outcome="denied",
    ).one()
    assert denied.detail["reason_code"] == "not_found_or_forbidden"


def test_current_scope_must_dominate_row_scope_but_may_expand(db):
    all_scope_permissions = permissions.effective("boss", None)
    creator = security.UserContext(
        user_id="boss-user", role="boss", permissions=all_scope_permissions,
        is_authenticated=True, authn="sys_user", has_stable_subject=True
    )
    result = _write_test_report(
        _owner_from_context(db, creator),
        title="全量结果",
        headers=["值"],
        rows=[[1]],
        output_name="all.xlsx",
    )
    narrowed = security.UserContext(
        user_id="boss-user",
        role="boss",
        permissions={**all_scope_permissions, "own_customers_only": True},
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    assert agent_files.access_allowed(result["file_id"], narrowed) is False

    own_scope_permissions = permissions.effective("sales", None)
    own_creator = security.UserContext(
        user_id="sales-user", role="sales", permissions=own_scope_permissions,
        salesperson_name="Alice Sales",
        is_authenticated=True, authn="sys_user", has_stable_subject=True
    )
    own_result = _write_test_report(
        _owner_from_context(db, own_creator),
        title="本人结果",
        headers=["值"],
        rows=[[1]],
        output_name="own.xlsx",
    )
    expanded = security.UserContext(
        user_id="sales-user",
        role="sales",
        permissions={**own_scope_permissions, "own_customers_only": False},
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
    )
    assert agent_files.access_allowed(own_result["file_id"], expanded) is True


def test_own_customer_artifact_is_bound_to_canonical_salesperson_subject(db):
    own_permissions = permissions.effective("sales", None)

    def ctx(salesperson_name: str | None, *, own: bool = True):
        return security.UserContext(
            user_id="row-owner",
            role="sales",
            salesperson_name=salesperson_name,
            permissions={**own_permissions, "own_customers_only": own},
            is_authenticated=True,
            authn="sys_user",
            has_stable_subject=True,
        )

    creator_a = ctx("  Ａlice   Sales  ")
    artifact_a = _write_test_report(
        _owner_from_context(db, creator_a),
        title="本人客户",
        headers=["值"],
        rows=[[1]],
        output_name="own-a.xlsx",
    )
    assert agent_files.artifact_info(artifact_a["file_id"])["status"] == "ready"
    assert agent_files.access_allowed(artifact_a["file_id"], ctx("Alice Sales")) is True
    assert agent_files.access_allowed(artifact_a["file_id"], ctx("Bob Sales")) is False
    assert agent_files.access_allowed(artifact_a["file_id"], ctx("Bob Sales", own=False)) is True

    creator_without_subject = ctx(None)
    with pytest.raises(agent_files.FileError, match="销售主体"):
        agent_files.snapshot_access_scope(creator_without_subject)


def test_upload_is_explicit_owner_only_even_if_data_permissions_change(db):
    upload = agent_files.save_upload(
        b"private input", "input.txt", _verified_owner(db, "alice")
    )
    downgraded = security.UserContext(
        user_id="alice", role="sales", permissions={}, is_authenticated=True,
        authn="sys_user", has_stable_subject=True,
    )
    other = security.UserContext(
        user_id="bob", role="sales", permissions=permissions.effective("sales", None),
        is_authenticated=True, authn="sys_user", has_stable_subject=True,
    )
    assert agent_files.access_allowed(upload["file_id"], downgraded) is True
    assert agent_files.access_allowed(upload["file_id"], other) is False


def test_public_generated_write_is_always_server_classified(db):
    result = _write_test_report(
        _verified_owner(db, "alice"),
        title="classified",
        headers=["value"],
        rows=[[1]],
        output_name="classified.xlsx",
    )
    owner = security.UserContext(
        user_id="alice", role="admin", permissions=permissions.effective("admin", None),
        is_authenticated=True, authn="sys_user", has_stable_subject=True,
    )
    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None
    assert artifact.access_scope["policy"] == "provenance_guarded"
    assert agent_files.access_allowed(result["file_id"], owner) is True


def test_v2_creation_rejects_missing_or_unauthenticated_owner(db):
    with pytest.raises(agent_files.FileError, match="已验证身份"):
        agent_files.save_upload(b"no owner", "owner.txt", None)
    with pytest.raises(agent_files.FileError, match="已验证身份"):
        agent_files.write_report("x", ["h"], [[1]], "x.xlsx", None)
    with pytest.raises(agent_files.FileError, match="已验证身份"):
        agent_files.save_upload(b"spoof", "spoof.txt", "alice")
    with pytest.raises(TypeError):
        agent_files.write_report(
            "forged", ["h"], [[1]], "forged.xlsx", _verified_owner(db, "alice"),
            access_scope={"policy": "owner_only"},
        )
    unauthenticated = security.UserContext(
        user_id="alice", role="sales", permissions=permissions.effective("sales", None)
    )
    with pytest.raises(agent_files.FileError, match="实名系统账号"):
        agent_files.snapshot_access_scope(unauthenticated)


def test_sensitive_file_service_methods_do_not_offer_actorless_reads(db):
    from io import BytesIO

    from openpyxl import Workbook

    workbook = Workbook()
    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    owner = _verified_owner(db, "service-owner")
    artifact = agent_files.save_upload(buffer.getvalue(), "private.xlsx", owner)
    fid = artifact["file_id"]

    for call in (
        lambda: agent_files.inspect_file(fid),
        lambda: agent_files.read_rows(fid, None, 1, 10),
        lambda: agent_files.preview(fid),
        lambda: agent_files.read_document(fid),
        lambda: agent_files.get_download_info(fid),
    ):
        with pytest.raises(TypeError):
            call()


def test_verified_owner_is_derived_from_active_current_sys_user(db):
    current_permissions = permissions.effective("sales", None)
    user = SysUser(
        username="db-owner",
        role="sales",
        salesperson_name="Alice Sales",
        password_hash=hash_password("pw123456"),
        permissions=current_permissions,
        token_version=3,
        is_active=True,
    )
    db.add(user)
    db.commit()
    matching = security.UserContext(
        user_id="db-owner",
        role="sales",
        salesperson_name="Alice Sales",
        permissions=current_permissions,
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
        token_version=3,
    )
    forged = security.UserContext(
        user_id="db-owner",
        role="admin",
        permissions=permissions.effective("admin", None),
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
        token_version=3,
    )

    with pytest.raises(agent_files.FileError, match="当前账号事实"):
        agent_files.verified_artifact_owner(db, forged)

    owner = agent_files.verified_artifact_owner(db, matching)
    revoked_permissions = {key: False for key in permissions.ALL_KEYS}
    user.permissions = revoked_permissions
    user.token_version = 4
    db.commit()
    with pytest.raises(agent_files.FileError, match="当前账号事实"):
        agent_files.verified_artifact_owner(db, matching)
    with pytest.raises(agent_files.FileError, match="登录状态"):
        agent_files.save_upload(b"stale", "stale.txt", owner)

    current_after_revoke = security.UserContext(
        user_id="db-owner",
        role="sales",
        salesperson_name="Alice Sales",
        permissions=revoked_permissions,
        is_authenticated=True,
        authn="sys_user",
        has_stable_subject=True,
        token_version=4,
    )
    revoked_owner = agent_files.verified_artifact_owner(db, current_after_revoke)
    user.is_active = False
    db.commit()
    with pytest.raises(agent_files.FileError, match="登录状态"):
        agent_files.save_upload(b"revoked", "revoked.txt", revoked_owner)


def test_shared_or_unstable_subject_cannot_create_or_reopen_artifacts(db):
    upload = agent_files.save_upload(b"private", "private.txt", _verified_owner(db, "alice"))
    shared = security.UserContext(
        user_id="alice",
        role="admin",
        permissions=permissions.effective("admin", None),
        is_authenticated=True,
        authn="shared",
        has_stable_subject=False,
    )

    with pytest.raises(agent_files.FileError, match="实名系统账号"):
        agent_files.snapshot_access_scope(shared)
    assert agent_files.access_allowed(upload["file_id"], shared) is False


def test_shared_admin_token_cannot_create_or_access_v2_artifact(db):
    client = TestClient(app)
    token, _ = _make_token(
        "admin",
        "admin",
        None,
        perms=permissions.effective("admin", None),
        authn="shared",
    )
    admin = SysUser(
        username="admin", role="admin", password_hash=hash_password("pw123456"),
        permissions=permissions.effective("admin", None),
    )
    db.add(admin)
    db.commit()
    existing = agent_files.save_upload(b"private", "private.txt", _verified_owner(db, "admin"))
    headers = {"Authorization": f"Bearer {token}"}
    response = client.post(
        "/api/agent/upload",
        headers=headers,
        files={"file": ("private.txt", b"payload", "text/plain")},
    )

    assert response.status_code == 403, response.text
    denied = client.get(f"/api/agent/files/{existing['file_id']}", headers=headers)
    assert denied.status_code == 404
    assert denied.json() == {"detail": "文件不存在或无权访问"}
    assert db.query(AgentArtifact).filter(AgentArtifact.owner_sub == "admin").count() == 1


def test_upload_access_audit_never_records_raw_filename(db, monkeypatch):
    from app.api import agent as agent_api

    client = TestClient(app)
    token = _login(db, client, "audit_owner")
    other_token = _login(db, client, "audit_other")
    calls = []

    def capture(_ctx, action, resource, detail=None):
        calls.append((action, resource, detail))

    monkeypatch.setattr(agent_api, "record_access_log", capture)
    secret_name = "甲方-客户合同-秘密报价.txt"
    response = client.post(
        "/api/agent/upload",
        headers={"Authorization": f"Bearer {token}"},
        files={"file": (secret_name, b"payload", "text/plain")},
    )

    assert response.status_code == 200, response.text
    artifact_id = response.json()["file_id"]
    assert calls == [("upload", "agent_file", {
        "outcome": "success",
        "artifact_id": artifact_id,
        "size_bytes": len(b"payload"),
    })]
    assert secret_name not in repr(calls)

    calls.clear()
    denied = client.get(
        f"/api/agent/files/{artifact_id}",
        headers={"Authorization": f"Bearer {other_token}"},
    )
    assert denied.status_code == 404
    assert calls == [("download", "agent_file", {
        "outcome": "denied",
        "identifier_present": True,
        "identifier_format": "uuid",
        "reason_code": "not_found_or_forbidden",
    })]
    assert "payload" not in repr(calls)


def test_http_artifact_success_and_bound_denials_are_durable(db):
    client = TestClient(app)
    owner_token = _login(db, client, "durable-http-owner")
    other_token = _login(db, client, "durable-http-other")
    owner_headers = {"Authorization": f"Bearer {owner_token}"}
    other_headers = {"Authorization": f"Bearer {other_token}"}

    upload = client.post(
        "/api/agent/upload",
        headers=owner_headers,
        files={"file": ("durable.txt", b"durable-http", "text/plain")},
    )
    assert upload.status_code == 200, upload.text
    artifact_id = upload.json()["file_id"]
    assert client.get(
        f"/api/agent/files/{artifact_id}", headers=owner_headers
    ).status_code == 200
    assert client.get(
        f"/api/agent/files/{artifact_id}/preview", headers=owner_headers
    ).status_code == 200
    denied = client.get(
        f"/api/agent/files/{artifact_id}", headers=other_headers
    )
    assert denied.status_code == 404

    db.expire_all()
    audits = db.query(AgentArtifactAudit).filter_by(artifact_id=artifact_id).all()
    assert {(audit.action, audit.outcome) for audit in audits} >= {
        ("http_upload", "success"),
        ("http_download", "success"),
        ("http_preview", "success"),
        ("http_download", "denied"),
    }
    denied_audit = next(
        audit for audit in audits
        if audit.action == "http_download" and audit.outcome == "denied"
    )
    assert denied_audit.detail == {
        "reason_code": "not_found_or_forbidden",
        "identifier_format": "uuid",
    }

    legacy_id = "abcdef123456"
    assert client.get(
        f"/api/agent/files/{legacy_id}", headers=owner_headers
    ).status_code == 404
    db.expire_all()
    legacy = db.query(AgentArtifactAudit).filter_by(
        artifact_id=None,
        action="http_download",
        outcome="denied",
    ).one()
    assert legacy.detail["identifier_format"] == "legacy"
    assert legacy_id not in repr(legacy.detail)

    before = db.query(AgentArtifactAudit).filter_by(
        action="http_download", outcome="denied"
    ).count()
    unknown_id = str(uuid.uuid4())
    assert client.get(
        f"/api/agent/files/{unknown_id}", headers=owner_headers
    ).status_code == 404
    assert client.get(
        "/api/agent/files/not-an-artifact", headers=owner_headers
    ).status_code == 404
    db.expire_all()
    assert db.query(AgentArtifactAudit).filter_by(
        action="http_download", outcome="denied"
    ).count() == before


@pytest.mark.parametrize("delivery", ["upload", "download", "preview"])
def test_http_success_fails_closed_when_durable_audit_cannot_commit(
    db, monkeypatch, delivery
):
    client = TestClient(app)
    token = _login(db, client, f"audit-failure-{delivery}")
    headers = {"Authorization": f"Bearer {token}"}
    existing = agent_files.save_upload(
        b"must-not-be-delivered",
        f"audit-failure-{delivery}.txt",
        _verified_owner(db, f"audit-failure-{delivery}"),
    )
    real_record = agent_files.record_artifact_http_access

    def fail_success(**kwargs):
        if kwargs["outcome"] == "success":
            raise agent_files.ArtifactAuditUnavailable("audit unavailable")
        return real_record(**kwargs)

    monkeypatch.setattr(agent_files, "record_artifact_http_access", fail_success)
    if delivery == "upload":
        response = client.post(
            "/api/agent/upload",
            headers=headers,
            files={"file": ("new.txt", b"new-secret", "text/plain")},
        )
    else:
        suffix = "/preview" if delivery == "preview" else ""
        response = client.get(
            f"/api/agent/files/{existing['file_id']}{suffix}",
            headers=headers,
        )

    assert response.status_code == 503
    assert "must-not-be-delivered" not in response.text
    assert "new-secret" not in response.text


def test_denial_remains_denied_when_durable_audit_is_unavailable(db, monkeypatch):
    client = TestClient(app)
    owner_token = _login(db, client, "deny-audit-owner")
    other_token = _login(db, client, "deny-audit-other")
    created = agent_files.save_upload(
        b"private-denial",
        "private-denial.txt",
        _verified_owner(db, "deny-audit-owner"),
    )
    monkeypatch.setattr(
        agent_files,
        "record_artifact_http_access",
        lambda **_kwargs: (_ for _ in ()).throw(
            agent_files.ArtifactAuditUnavailable("audit unavailable")
        ),
    )

    response = client.get(
        f"/api/agent/files/{created['file_id']}",
        headers={"Authorization": f"Bearer {other_token}"},
    )

    assert owner_token
    assert response.status_code == 404
    assert response.json() == {"detail": "文件不存在或无权访问"}
    assert "private-denial" not in response.text


def test_durable_http_audit_wraps_transaction_failure(db, monkeypatch):
    created = agent_files.save_upload(
        b"audit transaction",
        "audit-transaction.txt",
        _verified_owner(db, "audit-transaction-owner"),
    )
    monkeypatch.setattr(
        agent_files,
        "_add_artifact_audit",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("audit insert unavailable")
        ),
    )

    with pytest.raises(agent_files.ArtifactAuditUnavailable):
        agent_files.record_artifact_http_access(
            action="download",
            outcome="success",
            actor="audit-transaction-owner",
            artifact_id=created["file_id"],
        )


def test_upload_identity_gate_denies_and_audits_before_reading_body(monkeypatch):
    from app.api import agent as agent_api

    class UnreadBody:
        filename = "customer-secret.txt"
        read_called = False

        async def read(self):
            self.read_called = True
            raise AssertionError("body must not be read")

    body = UnreadBody()
    shared = security.UserContext(
        user_id="shared-admin",
        role="admin",
        permissions=permissions.effective("admin", None),
        is_authenticated=True,
        authn="shared",
        has_stable_subject=False,
    )
    calls = []
    monkeypatch.setattr(
        agent_api,
        "record_access_log",
        lambda _ctx, action, resource, detail=None: calls.append(
            (action, resource, detail)
        ),
    )

    with pytest.raises(HTTPException) as caught:
        asyncio.run(agent_api.upload(file=body, db=None, role="admin", ctx=shared))

    assert caught.value.status_code == 403
    assert body.read_called is False
    assert calls == [("upload", "agent_file", {
        "outcome": "denied",
        "reason_code": "unstable_identity",
    })]
    assert body.filename not in repr(calls)


def test_download_audit_distinguishes_success_expired_and_missing_object(db, monkeypatch):
    from app.api import agent as agent_api

    client = TestClient(app)
    token = _login(db, client, "audit_lifecycle")
    headers = {"Authorization": f"Bearer {token}"}
    calls = []
    monkeypatch.setattr(
        agent_api,
        "record_access_log",
        lambda _ctx, action, resource, detail=None: calls.append(
            (action, resource, detail)
        ),
    )

    first = client.post(
        "/api/agent/upload", headers=headers,
        files={"file": ("first.txt", b"first-secret", "text/plain")},
    ).json()["file_id"]
    calls.clear()
    assert client.get(f"/api/agent/files/{first}", headers=headers).status_code == 200
    assert calls == [("download", "agent_file", {
        "outcome": "success", "artifact_id": first,
    })]

    row = db.get(AgentArtifact, first)
    assert row is not None
    row = force_artifact_state(db, row, "expired")
    calls.clear()
    assert client.get(f"/api/agent/files/{first}", headers=headers).status_code == 404
    assert calls == [("download", "agent_file", {
        "outcome": "denied", "identifier_present": True,
        "identifier_format": "uuid", "reason_code": "expired",
    })]
    db.expire_all()
    assert db.query(AgentArtifactAudit).filter_by(
        artifact_id=first,
        action="http_download",
        outcome="denied",
    ).order_by(AgentArtifactAudit.id.desc()).first().detail["reason_code"] == "expired"

    second = client.post(
        "/api/agent/upload", headers=headers,
        files={"file": ("second.txt", b"second-secret", "text/plain")},
    ).json()["file_id"]
    second_row = db.get(AgentArtifact, second)
    assert second_row is not None
    agent_files.get_artifact_store().path_for(second_row.storage_key).unlink()
    calls.clear()
    assert client.get(f"/api/agent/files/{second}", headers=headers).status_code == 404
    assert calls == [("download", "agent_file", {
        "outcome": "denied", "identifier_present": True,
        "identifier_format": "uuid", "reason_code": "object_missing",
    })]
    db.expire_all()
    assert db.query(AgentArtifactAudit).filter_by(
        artifact_id=second,
        action="http_download",
        outcome="denied",
    ).order_by(AgentArtifactAudit.id.desc()).first().detail["reason_code"] == "object_missing"
    assert "first-secret" not in repr(calls)
    assert "second-secret" not in repr(calls)


def test_unclassified_existing_artifact_denial_is_durable(db):
    client = TestClient(app)
    token = _login(db, client, "unclassified-http-owner")
    headers = {"Authorization": f"Bearer {token}"}
    created = agent_files.save_upload(
        b"unclassified",
        "unclassified.txt",
        _verified_owner(db, "unclassified-http-owner"),
    )
    row = db.get(AgentArtifact, created["file_id"])
    assert row is not None
    unclassified_scope = {
        "schema_version": "artifact-access/v2",
        "policy": "unclassified_deny",
        "classification": "unclassified",
        "proof_version": "legacy-generated-unproven/v1",
        "required_permissions": [],
        "contained_resources": [],
        "contained_fields": [],
        "sensitivity": "critical",
        "row_subject": None,
        "predicate_version": "unclassified/v1",
        "condition": {"op": "unknown"},
        "source_access_snapshots": [],
        "template_proof": None,
    }
    row = force_artifact_state(
        db,
        row,
        "ready",
        kind="generated",
        sensitivity="critical",
        source_ids=[],
        access_scope=unclassified_scope,
    )

    response = client.get(
        f"/api/agent/files/{row.id}",
        headers=headers,
    )

    assert response.status_code == 404
    db.expire_all()
    audit = db.query(AgentArtifactAudit).filter_by(
        artifact_id=row.id,
        action="http_download",
        outcome="denied",
    ).one()
    assert audit.detail["reason_code"] == "not_found_or_forbidden"


def test_ready_commit_fault_after_object_write_preserves_validating_marker(db, monkeypatch):
    owner = "fault-ready-commit"
    monkeypatch.setattr(
        agent_files,
        "_mark_artifact_ready",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("commit failed")),
    )

    with pytest.raises(agent_files.FileError, match="发布失败"):
        agent_files.save_upload(b"fault", "fault.txt", _verified_owner(db, owner))

    db.expire_all()
    artifact = db.query(AgentArtifact).filter(AgentArtifact.owner_sub == owner).one()
    assert artifact.status == "validating"
    assert not agent_files._meta_path(artifact.id).exists()
    assert agent_files.get_artifact_store().path_for(
        artifact.storage_key
    ).read_bytes() == b"fault"


def test_post_publish_store_read_outage_preserves_validating_marker(db, monkeypatch):
    owner = _verified_owner(db, "post-publish-store-outage")
    delegate = agent_files.get_artifact_store()

    class InspectOutageStore:
        def publish_bytes(self, *args, **kwargs):
            return delegate.publish_bytes(*args, **kwargs)

        def inspect(self, _storage_key):
            raise agent_files.ArtifactStoreUnavailable("store read unavailable")

    monkeypatch.setattr(agent_files, "get_artifact_store", InspectOutageStore)
    with pytest.raises(agent_files.ArtifactStoreUnavailable, match="待协调"):
        agent_files.save_upload(
            b"durable during outage",
            "post-publish-store-outage.txt",
            owner,
        )

    db.expire_all()
    row = db.query(AgentArtifact).filter_by(
        filename="post-publish-store-outage.txt"
    ).one()
    assert row.status == "validating"
    assert delegate.path_for(row.storage_key).read_bytes() == b"durable during outage"


def test_idempotent_ready_retry_rechecks_live_authorization(db):
    username = "ready-idempotent-revoke"
    created = agent_files.save_upload(
        b"already ready",
        "already-ready.txt",
        _verified_owner(db, username),
    )
    row = db.get(AgentArtifact, created["file_id"])
    user = db.query(SysUser).filter_by(username=username).one()
    assert row is not None and row.status == "ready"
    stored = agent_files.get_artifact_store().inspect(row.storage_key)
    user.token_version = int(user.token_version or 0) + 1
    db.commit()

    with pytest.raises(agent_files.FileError, match="授权已失效"):
        agent_files._mark_artifact_ready(row.id, stored)

    db.expire_all()
    assert db.get(AgentArtifact, row.id).status == "ready"


def test_download_and_preview_store_outage_return_503_and_durable_denial(
    db, monkeypatch
):
    client = TestClient(app)
    token = _login(db, client, "http-store-outage")
    headers = {"Authorization": f"Bearer {token}"}
    created = agent_files.save_upload(
        b"hidden during outage",
        "http-store-outage.txt",
        _verified_owner(db, "http-store-outage"),
    )
    delegate = agent_files.get_artifact_store()

    class ReadOutageStore:
        def read_bytes(self, *_args, **_kwargs):
            raise agent_files.ArtifactStoreUnavailable("store unavailable")

        def path_for(self, storage_key):
            return delegate.path_for(storage_key)

    monkeypatch.setattr(agent_files, "get_artifact_store", ReadOutageStore)
    for suffix, action in (("", "http_download"), ("/preview", "http_preview")):
        response = client.get(
            f"/api/agent/files/{created['file_id']}{suffix}",
            headers=headers,
        )
        assert response.status_code == 503
        assert "hidden during outage" not in response.text
        db.expire_all()
        audit = db.query(AgentArtifactAudit).filter_by(
            artifact_id=created["file_id"],
            action=action,
            outcome="denied",
        ).one()
        assert audit.detail["reason_code"] == "store_or_authorization_unavailable"
def test_reconciler_ready_interleaving_is_idempotent_and_never_deletes_object(
    db, tmp_path, monkeypatch
):
    """Reconciliation only reports the missing receipt; publisher still owns ready."""
    from app.db import SessionLocal
    from app.services import agent_artifact_reconcile

    now = datetime.now(timezone.utc) + timedelta(hours=2)

    class ReconcileBeforeReturnStore(agent_files.LocalArtifactStore):
        def publish_bytes(self, storage_key, content, *, validator=None):
            stored = super().publish_bytes(storage_key, content, validator=validator)
            result = agent_artifact_reconcile.reconcile_agent_artifacts(
                apply=True,
                grace_period=timedelta(hours=1),
                session_factory=SessionLocal,
                now=now,
                artifact_root=self.root,
            )
            assert result["applied"]["report_nonready_without_receipt"] == 1
            assert result["unresolved"] == 1
            return stored

    store = ReconcileBeforeReturnStore(tmp_path)
    monkeypatch.setattr(agent_files, "get_artifact_store", lambda: store)

    result = agent_files.save_upload(
        b"durable interleaving", "interleaving.txt", _verified_owner(db, "race-owner")
    )

    db.expire_all()
    artifact = db.get(AgentArtifact, result["file_id"])
    assert artifact is not None and artifact.status == "ready"
    assert store.path_for(artifact.storage_key).read_bytes() == b"durable interleaving"


def test_legacy_sidecar_same_name_sys_user_takeover_is_denied(db):
    owner = _verified_owner(db, "alice")
    legacy_id = "abcdef123456"
    meta = {
        "filename": "legacy.txt",
        "ext": "txt",
        "kind": "upload",
        "operated_by": "alice",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    agent_files._data_path(legacy_id, "txt").write_bytes(b"legacy")
    _write_legacy_fixture(legacy_id, meta)

    with pytest.raises(agent_files.ArtifactUnavailable, match="无权访问"):
        agent_files.get_download_info(legacy_id, owner)
    with pytest.raises(agent_files.ArtifactUnavailable, match="Legacy"):
        agent_files.artifact_info(legacy_id)
    with pytest.raises(agent_files.ArtifactUnavailable, match="写入已停用"):
        agent_files._save_meta(legacy_id, meta)
    assert agent_files._owner_of_unchecked(legacy_id) is None


@pytest.mark.parametrize(
    "payload",
    [
        [],
        "scalar",
        {"filename": ["legacy.txt"], "ext": "txt", "kind": "upload",
         "operated_by": "alice", "created_at": "2026-01-01T00:00:00+00:00"},
        {"filename": "legacy.txt", "ext": "txt", "kind": "upload",
         "operated_by": "alice", "created_at": "2026-01-01T00:00:00+00:00",
         "sheets": [{"name": "Sheet", "n_rows": [], "n_cols": 1}]},
    ],
    ids=["array", "scalar", "nested-filename", "nested-sheet"],
)
def test_legacy_sidecar_json_shape_fails_closed_with_stable_file_error(payload):
    legacy_id = "abcdef123456"
    agent_files._data_path(legacy_id, "txt").write_bytes(b"legacy")
    agent_files._meta_path(legacy_id).write_text(
        json.dumps(payload), encoding="utf-8"
    )

    with pytest.raises(agent_files.FileError, match="文件元数据损坏"):
        agent_files._load_meta(legacy_id)


def test_legacy_generated_without_scope_fails_closed_for_every_role(db):
    legacy_id = "fedcba654321"
    agent_files._data_path(legacy_id, "txt").write_bytes(b"legacy generated")
    _write_legacy_fixture(legacy_id, {
        "filename": "legacy-generated.txt",
        "ext": "txt",
        "kind": "generated",
        "operated_by": "alice",
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    sales = security.UserContext(
        user_id="alice", role="sales", is_authenticated=True,
        authn="sys_user", has_stable_subject=True
    )
    admin = security.UserContext(
        user_id="admin", role="admin", is_authenticated=True,
        authn="sys_user", has_stable_subject=True
    )

    assert agent_files.access_allowed(legacy_id, sales) is False
    assert agent_files.access_allowed(legacy_id, admin) is False
