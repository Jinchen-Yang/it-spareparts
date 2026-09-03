"""项目工作簿 v3 导出（D1）：按 workbook-template-design.md 的六 sheet + 隐藏技术 sheet。

颜色契约：灰底表头=系统只读列，黄底表头=可回填列（同 scripts/gen_maintenance_workbook_templates.py）。
数据同步方向：项目/合同/回款计划以台账为唯一事实源（本表重叠列只读展示）；
本表只承载台账没有的「实际 + 回填」列。缺数据照常展示，缺失仅提示（不隐藏、不按 0 计）。
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from uuid import uuid4
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, case, func, or_, select
from sqlalchemy.orm import Session

from app.business_time import business_today
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder, FProjectExpense
from app.models.maintenance_bad_return import (
    MaintenanceBadReturn,
    MaintenanceBadReturnLine,
    MaintenanceReturnObligation,
)
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
    MaintenanceSiteIssue,
    MaintenanceSiteIssueLine,
)
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.services import maintenance_front_stock as front_stock
from app.services import maintenance_cost_quality
from app.services import maintenance_periods
from app.services import maintenance_project_operations as operations
from app.services.maintenance_boss_board import _card_contracts
from app.services.maintenance_workbook_export import safe_xlsx_text

WORKBOOK_PROTOCOL_VERSION = "maintenance-project-workbook-v3.2"
GRAY = PatternFill("solid", fgColor="D9D9D9")  # 只读/系统生成
YELLOW = PatternFill("solid", fgColor="FFF2CC")  # 可编辑
TITLE_FONT = Font(bold=True, size=13)
HEAD_FONT = Font(bold=True, size=10)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

EDITABLE_HINT = "（黄底列可编辑：表尾追加新行 / 修改黄底单元格；灰底列由系统生成，改动将被忽略）"

_READONLY = "readonly"
_EDITABLE = "editable"


def _style_header(ws, headers: list[str], colors: list[str], row: int = 1) -> None:
    for idx, (header, color) in enumerate(zip(headers, colors), 1):
        cell = ws.cell(row=row, column=idx, value=header)
        cell.font = HEAD_FONT
        cell.fill = GRAY if color == _READONLY else YELLOW
        cell.border = BORDER
        cell.alignment = WRAP
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def _set_widths(ws, widths: list[float]) -> None:
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _add_blank_rows(ws, ncols: int, n: int, start: int) -> None:
    for row in range(start, start + n):
        for idx in range(1, ncols + 1):
            ws.cell(row=row, column=idx).border = BORDER


def _blank(ws, ncols: int) -> None:
    ws.append([None] * ncols)


def _contracts(db: Session, project_id: str) -> list[MaintenanceProjectContract]:
    today = business_today()
    return list(
        db.execute(
            select(MaintenanceProjectContract)
            .where(
                MaintenanceProjectContract.project_id == project_id,
                MaintenanceProjectContract.effective_from <= today,
                or_(
                    MaintenanceProjectContract.effective_to.is_(None),
                    MaintenanceProjectContract.effective_to > today,
                ),
            )
            .order_by(MaintenanceProjectContract.effective_from)
        ).scalars()
    )


def _expenses(
    db: Session,
    project_id: str,
    contracts: list[MaintenanceProjectContract],
) -> list[tuple[FProjectExpense, MaintenanceProjectExpenseAttribution]]:
    """已稳定归属本项目、且指向唯一当前合同关系的 approved 费用。"""
    if not contracts:
        return []
    today = business_today()
    candidate_identities = {contract.contract_id for contract in contracts}
    current_relations = list(db.scalars(
        select(MaintenanceProjectContract).where(
            MaintenanceProjectContract.contract_id.in_(candidate_identities),
            MaintenanceProjectContract.effective_from <= today,
            or_(
                MaintenanceProjectContract.effective_to.is_(None),
                MaintenanceProjectContract.effective_to > today,
            ),
        )
    ))
    by_identity: dict[str, list[MaintenanceProjectContract]] = {}
    for relation in current_relations:
        by_identity.setdefault(relation.contract_id, []).append(relation)
    safe_contract_ids = {
        relation.project_contract_id
        for relation in contracts
        if len(by_identity.get(relation.contract_id, ())) == 1
    }
    if not safe_contract_ids:
        return []
    return list(
        db.execute(
            select(FProjectExpense, MaintenanceProjectExpenseAttribution)
            .join(
                MaintenanceProjectExpenseAttribution,
                MaintenanceProjectExpenseAttribution.raw_expense_line_id
                == FProjectExpense.raw_line_id,
            )
            .where(
                MaintenanceProjectExpenseAttribution.project_id == project_id,
                MaintenanceProjectExpenseAttribution.project_contract_id.in_(
                    safe_contract_ids
                ),
                MaintenanceProjectExpenseAttribution.status_mapping_state == "mapped",
                MaintenanceProjectExpenseAttribution.normalized_status == "approved",
                MaintenanceProjectExpenseAttribution.ownership_mapping_state == "mapped",
            )
            .order_by(FProjectExpense.expense_date)
        ).all()
    )


def _wbdd_lines(db: Session, project_id: str) -> list[tuple[FMaintenanceOrder, FMaintenanceLine]]:
    return list(
        db.execute(
            select(FMaintenanceOrder, FMaintenanceLine)
            .join(
                MaintenanceSourceOrderAssignment,
                MaintenanceSourceOrderAssignment.source_order_id
                == FMaintenanceOrder.raw_order_id,
            )
            .join(
                FMaintenanceLine,
                FMaintenanceLine.order_id == FMaintenanceOrder.id,
            )
            .where(
                MaintenanceSourceOrderAssignment.project_id == project_id,
                MaintenanceSourceOrderAssignment.is_active.is_(True),
                # 2026-08-19：作废明细行不进工作簿 v3（#55）
                FMaintenanceLine.is_active.is_(True),
            )
            .order_by(FMaintenanceOrder.order_date, FMaintenanceLine.line_no)
        ).all()
    )


def _obligation_map(
    db: Session, project_id: str
) -> dict[str, MaintenanceReturnObligation]:
    rows = db.execute(
        select(MaintenanceReturnObligation).where(
            MaintenanceReturnObligation.project_id == project_id,
            MaintenanceReturnObligation.is_active.is_(True),
        )
    ).scalars().all()
    return {row.issue_line_id: row for row in rows}


def _bad_return_map(
    db: Session, obligation_ids: list[str]
) -> dict[str, tuple[str, str]]:
    """obligation_id → (返还单状态, 返还单号)。"""
    if not obligation_ids:
        return {}
    rows = db.execute(
        select(
            MaintenanceBadReturnLine.obligation_id,
            MaintenanceBadReturn.status,
            MaintenanceBadReturn.return_no,
        )
        .join(
            MaintenanceBadReturn,
            MaintenanceBadReturn.return_id == MaintenanceBadReturnLine.return_id,
        )
        .where(MaintenanceBadReturnLine.obligation_id.in_(obligation_ids))
    ).all()
    return {obligation_id: (status, return_no) for obligation_id, status, return_no in rows}


def _site_issue_rows(db: Session, project_id: str) -> list[dict]:
    issue_lines = db.execute(
        select(MaintenanceSiteIssue, MaintenanceSiteIssueLine)
        .join(
            MaintenanceSiteIssueLine,
            MaintenanceSiteIssueLine.issue_id == MaintenanceSiteIssue.issue_id,
        )
        .where(
            MaintenanceSiteIssue.project_id == project_id,
            MaintenanceSiteIssue.normalized_status.in_(("confirmed", "corrected")),
            # 2026-08-19：作废领用行不导出/不计入工作簿 v3（#55）
            MaintenanceSiteIssueLine.is_active.is_(True),
        )
        .order_by(MaintenanceSiteIssue.issue_date, MaintenanceSiteIssueLine.line_no)
    ).all()
    obligations = _obligation_map(db, project_id)
    bad_returns = _bad_return_map(
        db, [obligation.obligation_id for obligation in obligations.values()]
    )
    rows: list[dict] = []
    for issue, line in issue_lines:
        obligation = obligations.get(line.issue_line_id)
        status = "无需返还"
        return_no = ""
        if obligation is not None:
            if obligation.classification == "required":
                status = "待返还"
            elif obligation.classification == "pending_category":
                status = "待分类"
            linked = bad_returns.get(obligation.obligation_id)
            if linked is not None:
                status, return_no = linked
        rows.append(
            {
                "现场领用单号": issue.issue_no,
                "领用日期": issue.issue_date.isoformat(),
                "PN": line.pn,
                "备件SN": (line.serial_number or "")[:64],
                "领用数量": float(line.quantity),
                "是否应返还(行级)": (
                    "否"
                    if line.no_return is True
                    else ("是" if line.no_return is False else "")
                ),
                "应返数量(系统)": (
                    float(obligation.required_quantity)
                    if obligation is not None
                    else None
                ),
                "返还状态(系统)": status,
                "返还单号(系统)": return_no,
                "备注": "",
            }
        )
    return rows


def _append_safe(ws, values: list) -> None:
    """动态文本统一 safe_xlsx_text：公式/控制字符不外泄成 Excel 可执行内容。"""
    ws.append([safe_xlsx_text(value) if isinstance(value, str) else value for value in values])


def parse_editable_header_fills(ws) -> list[str]:
    """D3 颜色契约：按表头单元格底色识别可回填列。

    返回可编辑表头名列表；仅黄底(fgColor=FFF2CC)视为可编辑，
    其余（灰/白/无填充）一律只读——颜色即契约，不靠列名猜。
    """
    editable: list[str] = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        fill = cell.fill
        if (
            fill is not None
            and fill.fill_type == "solid"
            and (fill.fgColor.rgb or "").endswith("FFF2CC")
        ):
            editable.append(str(cell.value).strip())
    return editable


def build_project_workbook(db: Session, project_id: str) -> bytes | None:
    """按新模板 v1 构建项目工作簿（只读事实列 + 空的可回填列）。"""
    project = db.get(MaintenanceProject, project_id)
    if project is None:
        return None
    contracts = _contracts(db, project_id)
    expenses = _expenses(db, project_id, contracts)
    wbdd = _wbdd_lines(db, project_id)
    as_of = business_today()
    current_contract_ids = [contract.project_contract_id for contract in contracts]
    collections = (
        list(
            db.execute(
                select(MaintenanceCollectionSnapshot)
                .where(
                    MaintenanceCollectionSnapshot.project_id == project_id,
                    MaintenanceCollectionSnapshot.project_contract_id.in_(
                        current_contract_ids
                    ),
                    MaintenanceCollectionSnapshot.status == "confirmed",
                    # 未来月度快照不得进入当前导出（round-6 Blocker 7）
                    MaintenanceCollectionSnapshot.report_month <= as_of,
                )
                .order_by(MaintenanceCollectionSnapshot.report_month)
            ).scalars()
        )
        if current_contract_ids
        else []
    )
    # 每份合同取最新 confirmed 快照（月份升序覆盖），再跨合同求和（round-5 Blocker 6）
    latest_by_contract: dict[str, MaintenanceCollectionSnapshot] = {}
    for snapshot in collections:
        latest_by_contract[snapshot.project_contract_id] = snapshot
    site_issues = _site_issue_rows(db, project_id)
    balance = front_stock.balance_rows(db, project_id)

    wb = Workbook()
    wb.remove(wb.active)

    # 00 使用说明
    instructions = wb.create_sheet("00_使用说明", 0)
    for line in [
        "这本工作簿 = 一个维保项目的完整台账视图。月度流程：系统下载 → 在业务 Sheet 表尾追加/修改黄底列 → 整本上传。",
        "硬规则：缺行 ≠ 删除。只有「操作」列填 VOID 才会作废业务行；历史更正显式留痕。",
        EDITABLE_HINT,
        "Sheet 一览：01_项目基础信息（只读）· 02_概览数据（只读）· 03_备件订单 · 04_报销订单 · 05_项目经理回款单（月度累计，表尾追加）· 06_现场领用与返还（行级不返还标记）",
        "日期写法：完整日期写 YYYY-MM-DD；只有年月的写 YYYY-MM。",
        "数据同步方向：项目/合同/回款计划以台账工作簿为唯一来源（本表重叠列只读）；本表只回填成本、月度累计回款、凭证、领用行标记等台账没有的列。",
    ]:
        instructions.append([line])

    # 01 项目基础信息
    ws = wb.create_sheet("01_项目基础信息")
    headers = ["项目编号", "项目名称", "业务类型", "客户名称", "维保开始日期", "维保结束日期",
               "项目经理(CMO)", "维保负责人", "销售人员", "硬盘不返还默认值(项目级)", "项目状态",
               "前置库种类数", "前置库件数", "前置库金额(含税)"]
    _style_header(ws, headers, [_READONLY] * len(headers))
    primary = contracts[0] if contracts else None
    values_inc = [row["value_inc_tax"] for row in balance]
    complete_inc = all(v is not None for v in values_inc) if values_inc else True
    _append_safe(ws, [
        project.project_code,
        project.display_name,
        project.business_type or "",
        "",
        primary.effective_from.isoformat() if primary and primary.effective_from else "",
        primary.effective_to.isoformat() if primary and primary.effective_to else "",
        project.cmo_name or "",
        (project.project_manager_name if getattr(project, "project_manager_name", None) else ""),
        project.salesperson or "",
        "是" if project.no_return_default else "否",
        # 动态计算，不读存库快照（期限回填后该列会停在旧值）
        maintenance_periods.lifecycle_status(
            project.period_from, project.period_to, business_today()),
        len({row["warehouse_name"] for row in balance}),
        round(sum(row["qty"] for row in balance), 3),
        round(sum(v for v in values_inc if v is not None), 2) if complete_inc else "",
    ])
    _set_widths(ws, [14, 38, 10, 14, 13, 13, 13, 11, 10, 17, 10, 11, 10, 16])

    # 02 概览数据
    ws = wb.create_sheet("02_概览数据")
    ws["A1"] = "一、合同清单（只读）"
    ws["A1"].font = TITLE_FONT
    c_headers = ["合同编号", "合同额(含税)", "原始合同状态", "状态映射", "是否计入总额",
                 "生效日期", "失效日期", "金额完整性"]
    _style_header(ws, c_headers, [_READONLY] * len(c_headers), row=2)
    for contract in contracts:
        _append_safe(ws, [
            contract.contract_no,
            float(contract.amount_inc_tax) if contract.amount_inc_tax is not None else "",
            contract.contract_status or "",
            contract.status_mapping_state or "",
            "是" if contract.included_in_total else "否",
            contract.effective_from.isoformat() if contract.effective_from else "",
            contract.effective_to.isoformat() if contract.effective_to else "",
            "完整" if contract.amount_inc_tax is not None else "缺金额",
        ])
    _blank(ws, len(c_headers))
    first_metric_row = ws.max_row + 2
    ws.cell(row=first_metric_row, column=1, value="二、关键指标（只读）").font = TITLE_FONT
    included = [c for c in contracts if c.included_in_total]
    missing_amount_nos = [
        c.contract_no for c in included if c.amount_inc_tax is None
    ]
    contract_card = _card_contracts(db, [project_id]).get(project_id)
    # 只认当前、mapped、计入总额且不存在重复关系/跨项目共享冲突的台账事实。
    total_inc = (
        contract_card.get("amount_inc_tax")
        if contracts
        and contract_card
        and not contract_card.get("contract_incomplete")
        and not contract_card.get("contract_shared")
        else None
    )
    cumulative = (
        sum(v.cumulative_amount for v in latest_by_contract.values())
        if latest_by_contract
        else None
    )
    site_known = and_(
        MaintenanceSiteIssueLine.cost_source.in_((
            "maint_demand",
            "direct_purchase",
            "purchase_window",
            "sales_window",
            "manual",
        )),
        MaintenanceSiteIssueLine.price_basis == "ex_tax",
        maintenance_cost_quality.sql_amount_is_valid(
            MaintenanceSiteIssueLine.cost_amount_ex_tax
        ),
        maintenance_cost_quality.sql_amount_is_valid(
            MaintenanceSiteIssueLine.cost_amount_inc_tax
        ),
    )
    site_fact = db.execute(
        select(
            func.count(MaintenanceSiteIssueLine.issue_line_id),
            func.count().filter(site_known),
            func.coalesce(
                func.sum(case(
                    (site_known, MaintenanceSiteIssueLine.cost_amount_inc_tax),
                    else_=Decimal("0"),
                )),
                Decimal("0"),
            ),
        )
        .select_from(MaintenanceSiteIssueLine)
        .join(MaintenanceSiteIssue,
              MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id)
        .where(
            MaintenanceSiteIssue.project_id == project_id,
            MaintenanceSiteIssue.status_mapping_state == "mapped",
            MaintenanceSiteIssue.normalized_status.in_(("confirmed", "corrected")),
            MaintenanceSiteIssueLine.is_active.is_(True),
        )
    ).one()
    site_total_lines, site_known_lines, site_cost = site_fact
    site_cost = Decimal(site_cost or 0)
    missing_site_cost_lines = int(site_total_lines) - int(site_known_lines)
    expense_cost = Decimal("0")
    missing_expense_cost_lines = 0
    for _expense, attribution in expenses:
        tier = maintenance_cost_quality.normalized_tax_tier(
            source="direct",
            tax_basis="inc",
            legacy_amount=attribution.amount_inc_tax,
            normalized_amount=attribution.amount_inc_tax,
            normalized_basis="inc",
            anomaly_flags=(),
        )
        if tier == "missing":
            missing_expense_cost_lines += 1
        else:
            expense_cost += Decimal(attribution.amount_inc_tax)
    cost_total = site_cost + expense_cost
    missing_cost_lines = missing_site_cost_lines + missing_expense_cost_lines
    has_cost_evidence = int(site_total_lines) > 0 or bool(expenses)
    completeness_notes = []
    if missing_amount_nos:
        completeness_notes.append(f"{len(missing_amount_nos)} 份合同缺含税金额")
    if contract_card and (
        contract_card.get("contract_incomplete")
        or contract_card.get("contract_shared")
    ):
        completeness_notes.append("当前合同存在未映射、重复或跨项目共享冲突")
    if not values_inc or not complete_inc:
        completeness_notes.append("前置库存在缺成本件")
    if missing_site_cost_lines:
        completeness_notes.append(f"{missing_site_cost_lines} 行领用缺成本")
    if missing_expense_cost_lines:
        completeness_notes.append(f"{missing_expense_cost_lines} 行费用缺含税金额")
    if not has_cost_evidence:
        completeness_notes.append("暂无成本事实")
    metrics = [
        ("合同总额(含税)", round(float(total_inc), 2) if total_inc is not None else ""),
        ("累计回款(含税)", float(cumulative) if cumulative is not None else ""),
        ("回款进度", (
            f"{float(cumulative) / float(total_inc) * 100:.1f}%"
            if cumulative is not None and total_inc is not None and total_inc > 0
            else ""
        )),
        # 完整且为 0 是真实零；存在缺口时不把已知下限伪装为完整项目成本。
        ("项目已计成本(含税)", (
            round(float(cost_total), 2)
            if has_cost_evidence and missing_cost_lines == 0 else ""
        )),
        # 成本率：缺成本行或总额不完整时不发布（缺失只展示下界，round-6 Blocker 7）
        ("成本率", (
            f"{float(cost_total) / float(total_inc) * 100:.1f}%"
            if total_inc is not None and total_inc > 0
            and has_cost_evidence and missing_cost_lines == 0
            else ""
        )),
        ("缺失成本行数", missing_cost_lines),
        ("前置库存金额(含税)", round(sum(v for v in values_inc if v is not None), 2) if complete_inc else ""),
        ("超90天未领用备件行数", sum(1 for row in balance if row["stale_90d"])),
        ("数据完整性提示", "；".join(completeness_notes) or "数据完整"),
    ]
    for idx, (key, value) in enumerate(metrics, first_metric_row + 1):
        ws.cell(row=idx, column=1, value=key).font = BODY_FONT
        ws.cell(row=idx, column=2, value=value)
    _set_widths(ws, [30, 26, 14, 12, 12, 12, 12, 14])

    # 03 备件订单
    ws = wb.create_sheet("03_备件订单")
    headers = ["维保单号", "制单日期", "需求类型", "业务类型", "合同编号", "项目名称", "PN",
               "产品描述", "需求数量", "退货数量", "发货SN", "出库仓库", "成本来源(系统)",
               "未税单位成本", "含税单位成本(系统计算)", "变更原因"]
    colors = [_READONLY] * 13 + [_EDITABLE, _READONLY, _EDITABLE]
    _style_header(ws, headers, colors)
    for order, line in wbdd:
        _append_safe(ws, [
            order.order_no,
            order.order_date.isoformat() if order.order_date else "",
            order.demand_type or "",
            order.business_type or "",
            order.linked_sales_order_no or "",
            order.project_std or "",
            line.pn_std or "",
            "",
            float(line.qty) if line.qty is not None else "",
            "",
            "",
            order.warehouse or "",
            "",
            "",
            "",
            "",
        ])
    _add_blank_rows(ws, len(headers), 10, ws.max_row + 1)
    _set_widths(ws, [17, 11, 10, 9, 17, 30, 14, 22, 9, 9, 16, 11, 16, 11, 13, 12])

    # 04 报销订单
    ws = wb.create_sheet("04_报销订单")
    headers = ["报销单号", "报销日期", "报销人员", "报销类别", "费用分类", "支出事由",
               "合同编号", "未税金额", "含税金额(系统计算)", "流程状态", "备注"]
    colors = [_READONLY] * 7 + [_EDITABLE, _READONLY, _READONLY, _EDITABLE]
    _style_header(ws, headers, colors)
    for expense, attribution in expenses:
        _append_safe(ws, [
            expense.bxd_no or "",
            expense.expense_date.isoformat() if expense.expense_date else "",
            expense.person or "",
            expense.expense_type or "",
            expense.fee_category or "",
            (expense.reason or "")[:120],
            expense.linked_sales_order_no or "",
            "",
            float(attribution.amount_inc_tax),
            attribution.raw_status or expense.data_status or "",
            "",
        ])
    _add_blank_rows(ws, len(headers), 10, ws.max_row + 1)
    _set_widths(ws, [17, 11, 10, 10, 12, 26, 17, 11, 13, 10, 18])

    # 05 项目经理回款单（月度累计快照）
    ws = wb.create_sheet("05_项目经理回款单")
    headers = ["操作", "合同编号", "报告月份", "累计回款金额(含税)", "回款凭证号", "状态(系统)", "备注"]
    colors = [_EDITABLE, _READONLY, _EDITABLE, _EDITABLE, _EDITABLE, _READONLY, _EDITABLE]
    _style_header(ws, headers, colors)
    contract_no_by_id = {c.project_contract_id: c.contract_no for c in contracts}
    for snapshot in latest_by_contract.values():
        _append_safe(ws, [
            "",
            contract_no_by_id.get(snapshot.project_contract_id, ""),
            snapshot.report_month.strftime("%Y-%m"),
            float(snapshot.cumulative_amount),
            snapshot.receipt_reference or "",
            snapshot.status,
            snapshot.remark or "",
        ])
    _add_blank_rows(ws, len(headers), 8, ws.max_row + 1)
    _set_widths(ws, [9, 17, 11, 16, 15, 10, 18])

    # 06 现场领用与返还
    ws = wb.create_sheet("06_现场领用与返还")
    headers = ["现场领用单号", "领用日期", "PN", "备件SN", "领用数量", "是否应返还(行级)",
               "应返数量(系统)", "返还状态(系统)", "返还单号(系统)", "备注"]
    colors = [_READONLY] * 5 + [_EDITABLE] + [_READONLY] * 3 + [_EDITABLE]
    _style_header(ws, headers, colors)
    for row in site_issues:
        _append_safe(ws, [row[h] for h in headers])
    _add_blank_rows(ws, len(headers), 10, ws.max_row + 1)
    _set_widths(ws, [17, 11, 14, 15, 9, 15, 12, 14, 17, 16])

    # 隐藏技术 sheet
    ws_dict = wb.create_sheet("98_字典")
    for line in [
        "操作语义（05 表）：CREATE=新增月度累计快照；VOID=作废历史快照（缺行≠删除）",
        "需求类型：报修供货 / 补库供货；流程状态：草稿/已生效/已作废",
    ]:
        ws_dict.append([line])
    ws_dict.sheet_state = "hidden"
    ws_meta = wb.create_sheet("99_元数据")
    state = operations.get_or_create_workbook_state(db, project_id=project_id)
    ws_meta.append(["协议版本", WORKBOOK_PROTOCOL_VERSION])
    ws_meta.append(["导出ID", str(uuid4())])
    ws_meta.append(["项目编号", project.project_code])
    ws_meta.append(["项目ID", project_id])
    ws_meta.append(["范围", "project"])
    ws_meta.append(["业务日", business_today().isoformat()])
    ws_meta.append(["生成时间", datetime.now(timezone.utc).isoformat()])
    ws_meta.append(["基线版本", state.data_version])
    ws_meta.sheet_state = "hidden"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
