"""Sales replenishment-cart Beta domain service.

The service records application intent only.  It never writes inventory,
purchase, sales or maintenance facts.  Submitted versions and reviews are
immutable snapshots so an external reviewer can safely target an exact digest.
"""

from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.business_time import business_today
from app.config import get_settings
from app.models.dimensions import DimPart
from app.models.inventory import PartPool, PartPoolMember
from app.models.maintenance_project import MaintenanceProject
from app.models.replenishment import (
    ReplenishmentApplication,
    ReplenishmentApplicationLine,
    ReplenishmentApplicationVersion,
    ReplenishmentAuditEvent,
    ReplenishmentReview,
    ReplenishmentReviewLine,
)
from app.models.system import SysUser
from app.services import pool_price_analysis, replenishment_screening
from app.services.query_filters import col_matches_any, keyword_groups_or_substr
from app.services.part_resolver import resolve as resolve_part

MAX_LINES = 200
PRICE_WINDOW_DAYS = replenishment_screening.LOOKBACK_DAYS
MAX_EXCEL_TEXT = 32767
# 无值一律显示「—」，绝不用 0 顶替（铁律 5；AB-4 明示池内最低价无值时的展示）
_NO_VALUE = "—"
_INVALID_XML_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]")


class ReplenishmentError(ValueError):
    """Base domain error with an HTTP-friendly semantic code."""

    def __init__(self, message: str, *, code: str = "invalid", status_code: int = 400):
        super().__init__(message)
        self.code = code
        self.status_code = status_code


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _uid() -> str:
    return str(uuid.uuid4())


def _digest(value: Any) -> str:
    raw = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _clean_optional(value: str | None, *, maximum: int) -> str | None:
    cleaned = (value or "").strip()
    if not cleaned:
        return None
    if len(cleaned) > maximum:
        raise ReplenishmentError(f"文本不能超过 {maximum} 个字符")
    return cleaned


def _quantity(value: Decimal | int | float | str) -> Decimal:
    try:
        result = Decimal(str(value)).quantize(Decimal("0.001"))
    except Exception as exc:  # noqa: BLE001
        raise ReplenishmentError("数量格式不正确") from exc
    if result <= 0 or result > Decimal("999999.999"):
        raise ReplenishmentError("数量必须大于 0 且不超过 999999.999")
    return result


def _integer_quantity(value: Any) -> Decimal:
    try:
        result = Decimal(str(value))
    except Exception as exc:  # noqa: BLE001
        raise ReplenishmentError("数量格式不正确") from exc
    if result != result.to_integral_value() or result < 1 or result > Decimal("999999"):
        raise ReplenishmentError("数量必须是 1-999999 的整数", code="invalid_quantity")
    return result.quantize(Decimal("0.001"))


def _json_value(value: Any) -> Any:
    return json.loads(json.dumps(value, ensure_ascii=False, default=str))


def _actor(db: Session, username: str) -> SysUser:
    user = db.scalar(
        select(SysUser).where(SysUser.username == username, SysUser.is_active.is_(True))
    )
    if user is None:
        raise ReplenishmentError("该操作必须使用有效的独立账号", code="real_user_required", status_code=401)
    return user


def _audit(
    db: Session,
    app: ReplenishmentApplication,
    action: str,
    actor: str,
    reason: str,
    *,
    version_id: str | None = None,
    before: dict | None = None,
    after: dict | None = None,
) -> None:
    db.add(
        ReplenishmentAuditEvent(
            event_id=_uid(),
            application_id=app.application_id,
            version_id=version_id,
            action=action,
            before_json=before,
            after_json=after,
            reason=reason,
            operated_by=actor,
        )
    )


def _application_scope(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
    lock: bool = False,
) -> ReplenishmentApplication:
    stmt = select(ReplenishmentApplication).where(
        ReplenishmentApplication.application_id == application_id
    )
    if role != "admin":
        stmt = stmt.where(ReplenishmentApplication.owner_username == username)
    if lock:
        stmt = stmt.with_for_update()
    app = db.scalar(stmt)
    if app is None:
        # Owner isolation intentionally shares the same 404 as a missing object.
        raise ReplenishmentError("补库申请不存在", code="not_found", status_code=404)
    return app


def _workflow_mode(app: ReplenishmentApplication) -> str:
    return (
        "system_screening"
        if app.client_request_id is not None and app.request_digest is not None
        else "legacy_history"
    )


def _latest_version(db: Session, app: ReplenishmentApplication, *, lock: bool = False):
    stmt = select(ReplenishmentApplicationVersion).where(
        ReplenishmentApplicationVersion.application_id == app.application_id,
        ReplenishmentApplicationVersion.version_no == app.latest_version_no,
    )
    if lock:
        stmt = stmt.with_for_update()
    version = db.scalar(stmt)
    if version is None:
        raise ReplenishmentError("补库版本不存在", code="corrupt", status_code=409)
    return version


def _draft(db: Session, app: ReplenishmentApplication, *, lock: bool = True):
    version = _latest_version(db, app, lock=lock)
    if app.status != "draft" or version.status != "draft":
        raise ReplenishmentError("当前没有可编辑的购物车草稿", code="immutable", status_code=409)
    return version


def _part(db: Session, part_id: int, *, lock: bool = False) -> DimPart:
    stmt = select(DimPart).where(DimPart.id == part_id, DimPart.status == "active")
    if lock:
        stmt = stmt.with_for_update()
    part = db.scalar(stmt)
    if part is None:
        raise ReplenishmentError("PN 不存在或已合并", code="part_unavailable", status_code=404)
    if part.is_excluded:
        raise ReplenishmentError("该 PN 已被主数据治理排除，不能加入补库单", code="part_excluded", status_code=409)
    return part


def _pool_snapshots(db: Session, part_ids: list[int]) -> dict[int, dict[str, Any]]:
    if not part_ids:
        return {}
    rows = db.execute(
        select(PartPoolMember.part_id, PartPool.group_id, PartPool.name, PartPool.version)
        .select_from(PartPoolMember)
        .join(PartPool, PartPool.group_id == PartPoolMember.group_id)
        .where(PartPoolMember.part_id.in_(part_ids), PartPool.status == "active")
        .order_by(PartPoolMember.part_id, PartPool.group_id)
    )
    snapshots: dict[int, dict[str, Any]] = {}
    for row in rows:
        snapshots.setdefault(
            row.part_id,
            {"group_id": row.group_id, "name": row.name, "version": row.version},
        )
    return snapshots


def _pool_snapshot(db: Session, part_id: int) -> dict[str, Any]:
    return _pool_snapshots(db, [part_id]).get(
        part_id, {"group_id": None, "name": None, "version": None}
    )


def _auto_review_recommendations(db: Session, part: DimPart) -> list[dict]:
    resolved = resolve_part(
        db,
        f"{part.pn_std} {part.description or ''}".strip(),
        limit=10,
        log_miss=False,
        include_similar=True,
    )
    candidates = [*resolved.get("items", []), *resolved.get("similar_items", [])]
    found: list[dict] = []
    seen: set[int] = {part.id}
    for item in candidates:
        candidate_id = int(item.get("part_id") or 0)
        if not candidate_id or candidate_id in seen:
            continue
        if item.get("is_excluded") or item.get("pool_group_id") is None:
            continue
        seen.add(candidate_id)
        found.append({
            "part_id": candidate_id,
            "pn_std": item.get("pn_std"),
            "description": item.get("description"),
            "pool_group_id": item.get("pool_group_id"),
            "pool_name": item.get("pool_name"),
            "score": item.get("score"),
            "match_reason": item.get("match_reason"),
        })
        if len(found) >= 3:
            break
    return found


def _auto_review_decision(screening) -> tuple[str, str]:
    pool = screening.get("pool_membership")
    activity = screening.get("recent_activity")
    if pool.passed:
        return "approved", "pool_member"
    detail = activity.detail or {}
    if int(detail.get("purchase_samples") or 0) > 0:
        return "approved", "recent_purchase"
    if int(detail.get("sales_samples") or 0) > 0:
        return "approved", "recent_sales"
    return "rejected", "no_purchase_or_sales_in_182_days"


def _price_snapshot(db: Session, part_id: int, *, as_of: date | None = None) -> dict:
    upper = as_of or business_today()
    lower = upper - timedelta(days=PRICE_WINDOW_DAYS - 1)
    facts = pool_price_analysis.aggregate_part_price_facts(
        db, [part_id], date_from=lower, date_to=upper
    ).get(part_id, {"purchase": None, "sales": None})
    evidence = {
        "part_id": part_id,
        "window": {"date_from": lower.isoformat(), "date_to": upper.isoformat(), "days": PRICE_WINDOW_DAYS},
        "basis": "ex_tax_quantity_weighted",
        "filters": [
            "active_order",
            "positive_quantity_and_price",
            "sales_counts_revenue",
            "confirmed_source_error_excluded",
        ],
        "purchase": facts["purchase"],
        "sales": facts["sales"],
    }
    return {
        "date_from": lower,
        "date_to": upper,
        "as_of": upper,
        "purchase": facts["purchase"] or {},
        "sales": facts["sales"] or {},
        "digest": _digest(evidence),
    }


def available_projects(db: Session, *, username: str, role: str) -> list[dict]:
    user = _actor(db, username)
    predicate = [MaintenanceProject.is_active.is_(True)]
    if role == "sales":
        salesperson = user.salesperson_name
        if not salesperson or not salesperson.strip():
            return []
        predicate.append(MaintenanceProject.salesperson == salesperson)
    elif role != "admin":
        return []
    projects = db.scalars(
        select(MaintenanceProject)
        .where(*predicate)
        .order_by(MaintenanceProject.project_code, MaintenanceProject.project_id)
    )
    return [
        {
            "project_id": project.project_id,
            "project_code": project.project_code,
            "display_name": project.display_name,
        }
        for project in projects
    ]


def _authorized_project(
    db: Session, *, project_id: str, user: SysUser, role: str
) -> MaintenanceProject:
    if role not in {"admin", "sales"}:
        raise ReplenishmentError(
            "项目不存在或不可选", code="project_unavailable", status_code=404
        )
    predicate = [
        MaintenanceProject.project_id == project_id,
        MaintenanceProject.is_active.is_(True),
    ]
    if role == "sales":
        salesperson = user.salesperson_name
        if not salesperson or not salesperson.strip():
            raise ReplenishmentError(
                "项目不存在或不可选", code="project_unavailable", status_code=404
            )
        predicate.append(MaintenanceProject.salesperson == salesperson)
    project = db.scalar(
        select(MaintenanceProject).where(*predicate).with_for_update()
    )
    if project is None:
        raise ReplenishmentError(
            "项目不存在或不可选", code="project_unavailable", status_code=404
        )
    return project


def _line_snapshot(db: Session, part: DimPart, *, as_of: date | None = None) -> dict:
    pool = _pool_snapshot(db, part.id)
    price = _price_snapshot(db, part.id, as_of=as_of)
    return {"pool": pool, "price": price}


def _stats_payload(value: dict | None) -> dict | None:
    if not value:
        return None
    return {
        "weighted_avg": value.get("weighted_avg"),
        "total_qty": value.get("total_qty"),
        "order_count": value.get("order_count", 0),
        "line_count": value.get("line_count", 0),
        "latest_date": value.get("latest_date"),
    }


def catalog_search(
    db: Session,
    query: str | None,
    *,
    page: int = 1,
    page_size: int = 20,
    as_of: date | None = None,
) -> dict:
    q = (query or "").strip()
    page = max(1, page)
    page_size = min(50, max(1, page_size))
    predicate = [DimPart.status == "active", DimPart.is_excluded.is_(False)]
    if q:
        groups = keyword_groups_or_substr(q)
        predicate.extend(
            col_matches_any(DimPart.search_doc, group) for group in groups
        )
    total = int(db.scalar(select(func.count()).select_from(DimPart).where(*predicate)) or 0)
    parts = list(
        db.scalars(
            select(DimPart)
            .where(*predicate)
            .order_by(DimPart.pn_std, DimPart.id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    upper = as_of or business_today()
    lower = upper - timedelta(days=PRICE_WINDOW_DAYS - 1)
    facts = pool_price_analysis.aggregate_part_price_facts(
        db, [part.id for part in parts], date_from=lower, date_to=upper
    )
    pools = _pool_snapshots(db, [part.id for part in parts])
    items = []
    for part in parts:
        pool = pools.get(part.id, {"group_id": None, "name": None, "version": None})
        stats = facts.get(part.id, {"purchase": None, "sales": None})
        items.append(
            {
                "part_id": part.id,
                "pn_std": part.pn_std,
                "description": part.description,
                "brand": part.brand,
                "unit": part.unit,
                "needs_review": bool(part.needs_review),
                "pool": pool,
                "price_window": {
                    "date_from": lower.isoformat(),
                    "date_to": upper.isoformat(),
                    "days": PRICE_WINDOW_DAYS,
                    "basis": "未税数量加权",
                },
                "purchase": _stats_payload(stats.get("purchase")),
                "sales": _stats_payload(stats.get("sales")),
            }
        )
    return {"items": items, "total": total, "page": page, "page_size": page_size}


def create_application(
    db: Session,
    *,
    username: str,
    warehouse: str | None = None,
    request_note: str | None = None,
) -> dict:
    user = _actor(db, username)
    app_id = _uid()
    version_id = _uid()
    suffix = app_id.replace("-", "")[:10].upper()
    app = ReplenishmentApplication(
        application_id=app_id,
        application_no=f"BLK-{business_today():%Y%m%d}-{suffix}",
        owner_username=user.username,
        owner_display_name=user.display_name or user.salesperson_name or user.username,
        salesperson_name_snapshot=user.salesperson_name,
    )
    version = ReplenishmentApplicationVersion(
        version_id=version_id,
        application_id=app_id,
        version_no=1,
        status="draft",
        warehouse=_clean_optional(warehouse, maximum=64),
        request_note=_clean_optional(request_note, maximum=4000),
        created_by=username,
    )
    db.add_all([app, version])
    db.flush()
    _audit(db, app, "application_created", username, "新建补库购物车草稿", version_id=version_id)
    db.commit()
    return get_application(db, app_id, username=username, role=user.role)


def _serialize_line(line: ReplenishmentApplicationLine, review_line=None) -> dict:
    screening = line.screening_json or {}
    return {
        "line_id": line.line_id,
        "request_line_id": line.request_line_id,
        "source_line_id": line.source_line_id,
        "line_no": line.line_no,
        "part_id": line.part_id,
        "pn_std": line.pn_std,
        "description": line.description,
        "brand": line.brand,
        "unit": line.unit,
        "quantity": float(line.quantity),
        "special_note": line.special_note,
        "pool": {
            "group_id": line.pool_group_id,
            "name": line.pool_name,
            "version": line.pool_version,
        },
        "price_window": {
            "date_from": line.price_window_from.isoformat(),
            "date_to": line.price_window_to.isoformat(),
            "days": PRICE_WINDOW_DAYS,
            "basis": "未税数量加权",
        },
        "purchase": _stats_payload(line.purchase_stats_json),
        "sales": _stats_payload(line.sales_stats_json),
        "screening": screening or None,
        "latest_sales": screening.get("latest_sales"),
        "pool_floor_ex_tax": screening.get("pool_floor_ex_tax"),
        "review": (
            {"decision": review_line.decision, "reason": review_line.reason}
            if review_line is not None
            else None
        ),
    }


def _version_payload(db: Session, version: ReplenishmentApplicationVersion) -> dict:
    lines = list(
        db.scalars(
            select(ReplenishmentApplicationLine)
            .where(ReplenishmentApplicationLine.version_id == version.version_id)
            .order_by(ReplenishmentApplicationLine.line_no)
        )
    )
    review = db.scalar(select(ReplenishmentReview).where(ReplenishmentReview.version_id == version.version_id))
    review_lines: dict[str, ReplenishmentReviewLine] = {}
    if review is not None:
        review_lines = {
            item.version_line_id: item
            for item in db.scalars(
                select(ReplenishmentReviewLine).where(ReplenishmentReviewLine.review_id == review.review_id)
            )
        }
    return {
        "version_id": version.version_id,
        "version_no": version.version_no,
        "parent_version_id": version.parent_version_id,
        "status": version.status,
        "warehouse": version.warehouse,
        "request_note": version.request_note,
        "content_digest": version.content_digest,
        "submitted_by": version.submitted_by,
        "submitted_at": version.submitted_at.isoformat() if version.submitted_at else None,
        "lines": [_serialize_line(line, review_lines.get(line.line_id)) for line in lines],
        "review": (
            {
                "review_id": review.review_id,
                "external_reference": review.external_reference,
                "summary_note": review.summary_note,
                "approved_count": review.approved_count,
                "rejected_count": review.rejected_count,
                "reviewed_at": review.reviewed_at.isoformat(),
            }
            if review is not None
            else None
        ),
    }


def get_application(db: Session, application_id: str, *, username: str, role: str) -> dict:
    app = _application_scope(db, application_id, username=username, role=role)
    versions = list(
        db.scalars(
            select(ReplenishmentApplicationVersion)
            .where(ReplenishmentApplicationVersion.application_id == app.application_id)
            .order_by(ReplenishmentApplicationVersion.version_no.desc())
        )
    )
    return {
        "application_id": app.application_id,
        "application_no": app.application_no,
        "owner_username": app.owner_username,
        "owner_display_name": app.owner_display_name,
        "salesperson_name_snapshot": app.salesperson_name_snapshot,
        "is_legacy_project_unbound": app.is_legacy_project_unbound,
        "project": (
            {
                "project_id": app.project_id,
                "project_code": app.project_code_snapshot,
                "display_name": app.project_name_snapshot,
            }
            if app.project_id
            else None
        ),
        "status": app.status,
        "workflow_mode": _workflow_mode(app),
        "stage": (
            "legacy_history"
            if _workflow_mode(app) == "legacy_history"
            else "needs_revision"
            if app.status == "needs_revision"
            else "approved"
            if app.status == "approved"
            else "screening_complete"
            if app.status == "submitted"
            else app.status
        ),
        "version": app.version,
        "latest_version_no": app.latest_version_no,
        "created_at": app.created_at.isoformat(),
        "updated_at": app.updated_at.isoformat(),
        "versions": [_version_payload(db, version) for version in versions],
    }


def list_applications(db: Session, *, username: str, role: str, page: int = 1, page_size: int = 20) -> dict:
    predicate = [] if role == "admin" else [ReplenishmentApplication.owner_username == username]
    total = int(
        db.scalar(select(func.count()).select_from(ReplenishmentApplication).where(*predicate)) or 0
    )
    apps = list(
        db.scalars(
            select(ReplenishmentApplication)
            .where(*predicate)
            .order_by(ReplenishmentApplication.updated_at.desc(), ReplenishmentApplication.application_id)
            .offset((max(page, 1) - 1) * min(max(page_size, 1), 100))
            .limit(min(max(page_size, 1), 100))
        )
    )
    return {
        "items": [
            {
                "application_id": app.application_id,
                "application_no": app.application_no,
                "owner_display_name": app.owner_display_name,
                "project": (
                    {
                        "project_id": app.project_id,
                        "project_code": app.project_code_snapshot,
                        "display_name": app.project_name_snapshot,
                    }
                    if app.project_id
                    else None
                ),
                "status": app.status,
                "is_legacy_project_unbound": app.is_legacy_project_unbound,
                "workflow_mode": _workflow_mode(app),
                "stage": (
                    "legacy_history"
                    if _workflow_mode(app) == "legacy_history"
                    else "screening_complete"
                    if app.status == "submitted"
                    else app.status
                ),
                "version": app.version,
                "latest_version_no": app.latest_version_no,
                "updated_at": app.updated_at.isoformat(),
            }
            for app in apps
        ],
        "total": total,
        "page": max(page, 1),
        "page_size": min(max(page_size, 1), 100),
    }


def update_draft(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
    expected_version: int,
    warehouse: str | None,
    request_note: str | None,
) -> dict:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role, lock=True)
    if app.version != expected_version:
        raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
    version = _draft(db, app)
    before = {"warehouse": version.warehouse, "request_note": version.request_note}
    version.warehouse = _clean_optional(warehouse, maximum=64)
    version.request_note = _clean_optional(request_note, maximum=4000)
    app.version += 1
    _audit(
        db,
        app,
        "draft_updated",
        username,
        "更新补库草稿基本信息",
        version_id=version.version_id,
        before=before,
        after={"warehouse": version.warehouse, "request_note": version.request_note},
    )
    db.commit()
    return get_application(db, application_id, username=username, role=role)


def add_line(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
    expected_version: int,
    part_id: int,
    quantity: Decimal | int | float | str,
    special_note: str | None = None,
) -> dict:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role, lock=True)
    if app.version != expected_version:
        raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
    version = _draft(db, app)
    count = int(
        db.scalar(
            select(func.count()).select_from(ReplenishmentApplicationLine).where(
                ReplenishmentApplicationLine.version_id == version.version_id
            )
        )
        or 0
    )
    if count >= MAX_LINES:
        raise ReplenishmentError(f"每份补库单最多 {MAX_LINES} 条", status_code=413)
    duplicate = db.scalar(
        select(ReplenishmentApplicationLine.line_id).where(
            ReplenishmentApplicationLine.version_id == version.version_id,
            ReplenishmentApplicationLine.part_id == part_id,
        )
    )
    if duplicate is not None:
        raise ReplenishmentError("该 PN 已在购物车中，请直接修改数量", code="duplicate_part", status_code=409)
    part = _part(db, part_id)
    snapshot = _line_snapshot(db, part)
    line_id = _uid()
    price = snapshot["price"]
    pool = snapshot["pool"]
    line = ReplenishmentApplicationLine(
        line_id=line_id,
        request_line_id=_uid(),
        version_id=version.version_id,
        line_no=count + 1,
        part_id=part.id,
        pn_std=part.pn_std,
        description=part.description,
        brand=part.brand,
        unit=part.unit,
        quantity=_quantity(quantity),
        special_note=_clean_optional(special_note, maximum=4000),
        pool_group_id=pool["group_id"],
        pool_name=pool["name"],
        pool_version=pool["version"],
        price_window_from=price["date_from"],
        price_window_to=price["date_to"],
        price_as_of=price["as_of"],
        purchase_stats_json=price["purchase"],
        sales_stats_json=price["sales"],
        evidence_digest=price["digest"],
    )
    db.add(line)
    app.version += 1
    _audit(
        db,
        app,
        "line_added",
        username,
        "购物车加入 PN",
        version_id=version.version_id,
        after={"line_id": line_id, "part_id": part.id, "quantity": str(line.quantity)},
    )
    db.commit()
    return get_application(db, application_id, username=username, role=role)


def update_line(
    db: Session,
    application_id: str,
    line_id: str,
    *,
    username: str,
    role: str,
    expected_version: int,
    part_id: int,
    quantity: Decimal | int | float | str,
    special_note: str | None,
) -> dict:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role, lock=True)
    if app.version != expected_version:
        raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
    version = _draft(db, app)
    line = db.scalar(
        select(ReplenishmentApplicationLine)
        .where(
            ReplenishmentApplicationLine.line_id == line_id,
            ReplenishmentApplicationLine.version_id == version.version_id,
        )
        .with_for_update()
    )
    if line is None:
        raise ReplenishmentError("购物车条目不存在", code="not_found", status_code=404)
    duplicate = db.scalar(
        select(ReplenishmentApplicationLine.line_id).where(
            ReplenishmentApplicationLine.version_id == version.version_id,
            ReplenishmentApplicationLine.part_id == part_id,
            ReplenishmentApplicationLine.line_id != line_id,
        )
    )
    if duplicate is not None:
        raise ReplenishmentError("该 PN 已在购物车中", code="duplicate_part", status_code=409)
    part = _part(db, part_id)
    snapshot = _line_snapshot(db, part)
    before = {"part_id": line.part_id, "quantity": str(line.quantity), "special_note": line.special_note}
    line.part_id = part.id
    line.pn_std = part.pn_std
    line.description = part.description
    line.brand = part.brand
    line.unit = part.unit
    line.quantity = _quantity(quantity)
    line.special_note = _clean_optional(special_note, maximum=4000)
    pool, price = snapshot["pool"], snapshot["price"]
    line.pool_group_id, line.pool_name, line.pool_version = pool["group_id"], pool["name"], pool["version"]
    line.price_window_from, line.price_window_to, line.price_as_of = price["date_from"], price["date_to"], price["as_of"]
    line.purchase_stats_json, line.sales_stats_json = price["purchase"], price["sales"]
    line.evidence_digest = price["digest"]
    app.version += 1
    _audit(
        db,
        app,
        "line_updated",
        username,
        "更新购物车 PN、数量或特殊说明",
        version_id=version.version_id,
        before=before,
        after={"part_id": line.part_id, "quantity": str(line.quantity), "special_note": line.special_note},
    )
    db.commit()
    return get_application(db, application_id, username=username, role=role)


def remove_line(
    db: Session,
    application_id: str,
    line_id: str,
    *,
    username: str,
    role: str,
    expected_version: int,
) -> dict:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role, lock=True)
    if app.version != expected_version:
        raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
    version = _draft(db, app)
    line = db.scalar(
        select(ReplenishmentApplicationLine)
        .where(
            ReplenishmentApplicationLine.line_id == line_id,
            ReplenishmentApplicationLine.version_id == version.version_id,
        )
        .with_for_update()
    )
    if line is None:
        raise ReplenishmentError("购物车条目不存在", code="not_found", status_code=404)
    if line.source_line_id:
        raise ReplenishmentError(
            "打回项不能直接移除，请更换 PN 或填写特殊情况说明后重新提交",
            code="revision_line_required",
            status_code=409,
        )
    before = {"line_id": line.line_id, "part_id": line.part_id, "quantity": str(line.quantity)}
    db.delete(line)
    db.flush()
    # Keep presentation line numbers dense without mutating any submitted version.
    remaining = list(
        db.scalars(
            select(ReplenishmentApplicationLine)
            .where(
                ReplenishmentApplicationLine.version_id == version.version_id,
                ReplenishmentApplicationLine.line_id != line_id,
            )
            .order_by(ReplenishmentApplicationLine.line_no)
        )
    )
    # ``(version_id, line_no)`` is an immediate unique constraint. Move all
    # survivors into a disjoint range first, then compact, so deleting row 1
    # cannot collide while row 2 is being renumbered to 1.
    temporary_base = max((item.line_no for item in remaining), default=0) + MAX_LINES
    for index, item in enumerate(remaining, 1):
        item.line_no = temporary_base + index
    db.flush()
    for index, item in enumerate(remaining, 1):
        item.line_no = index
    app.version += 1
    _audit(db, app, "line_removed", username, "从购物车移除 PN", version_id=version.version_id, before=before)
    db.commit()
    return get_application(db, application_id, username=username, role=role)


def _submission_content(version: ReplenishmentApplicationVersion, lines: list[ReplenishmentApplicationLine]) -> dict:
    return {
        "version_id": version.version_id,
        "application_id": version.application_id,
        "version_no": version.version_no,
        "parent_version_id": version.parent_version_id,
        "created_by": version.created_by,
        "warehouse": version.warehouse,
        "request_note": version.request_note,
        "lines": [
            {
                "line_id": line.line_id,
                "request_line_id": line.request_line_id,
                "version_id": line.version_id,
                "source_line_id": line.source_line_id,
                "line_no": line.line_no,
                "part_id": line.part_id,
                "pn_std": line.pn_std,
                "description": line.description,
                "brand": line.brand,
                "unit": line.unit,
                "quantity": str(line.quantity),
                "special_note": line.special_note,
                "pool_group_id": line.pool_group_id,
                "pool_name": line.pool_name,
                "pool_version": line.pool_version,
                "price_window_from": line.price_window_from,
                "price_window_to": line.price_window_to,
                "price_as_of": line.price_as_of,
                "purchase": line.purchase_stats_json,
                "sales": line.sales_stats_json,
                "screening": line.screening_json,
                "evidence_digest": line.evidence_digest,
            }
            for line in lines
        ],
    }


def submit_application_atomic(
    db: Session,
    *,
    username: str,
    role: str,
    client_request_id: str,
    project_id: str,
    lines: list[dict],
    request_note: str | None = None,
    commit: bool = True,
) -> dict:
    """Validate, screen, freeze and create one immutable application transaction."""
    try:
        user = _actor(db, username)
        key = (client_request_id or "").strip()
        if not 8 <= len(key) <= 128:
            raise ReplenishmentError("client_request_id 长度必须为 8-128 个字符")
        if not 1 <= len(lines) <= MAX_LINES:
            raise ReplenishmentError(f"补库明细必须为 1-{MAX_LINES} 条")
        cleaned_items = [
            {
                "part_id": int(item["part_id"]),
                "quantity": _integer_quantity(item["quantity"]),
                "special_note": _clean_optional(item.get("special_note"), maximum=4000),
            }
            for item in lines
        ]
        part_ids = [item["part_id"] for item in cleaned_items]
        if len(part_ids) != len(set(part_ids)):
            raise ReplenishmentError(
                "同一 PN 只能出现一次", code="duplicate_part", status_code=409
            )
        note = _clean_optional(request_note, maximum=4000)
        canonical = {
            "project_id": project_id,
            "request_note": note,
            "lines": [
                {
                    "part_id": item["part_id"],
                    "quantity": str(item["quantity"]),
                    "special_note": item["special_note"],
                }
                for item in cleaned_items
            ],
        }
        request_digest = _digest(canonical)
        db.execute(
            select(
                func.pg_advisory_xact_lock(
                    func.hashtextextended(f"replenishment-submit:{username}:{key}", 0)
                )
            )
        )
        existing = db.scalar(
            select(ReplenishmentApplication).where(
                ReplenishmentApplication.owner_username == username,
                ReplenishmentApplication.client_request_id == key,
            )
        )
        if existing is not None:
            if existing.request_digest != request_digest:
                raise ReplenishmentError(
                    "相同 client_request_id 对应了不同提交内容",
                    code="idempotency_conflict",
                    status_code=409,
                )
            result = get_application(
                db, existing.application_id, username=username, role=role
            )
            result["idempotent"] = True
            if commit:
                db.commit()
            return result

        project = _authorized_project(db, project_id=project_id, user=user, role=role)
        parts = list(
            db.scalars(
                select(DimPart)
                .where(
                    DimPart.id.in_(part_ids),
                    DimPart.status == "active",
                    DimPart.is_excluded.is_(False),
                )
                .with_for_update()
            )
        )
        parts_by_id = {part.id: part for part in parts}
        if set(parts_by_id) != set(part_ids):
            raise ReplenishmentError(
                "明细包含不存在、已合并或已排除的 PN",
                code="part_unavailable",
                status_code=422,
            )

        as_of = business_today()
        lower = as_of - timedelta(days=PRICE_WINDOW_DAYS - 1)
        facts = pool_price_analysis.aggregate_part_price_facts(
            db, part_ids, date_from=lower, date_to=as_of
        )
        pools = _pool_snapshots(db, part_ids)
        screenings = replenishment_screening.screen(
            db, part_ids=part_ids, as_of=as_of, price_facts=facts
        )
        latest_sales = replenishment_screening.latest_sales_history(
            db, part_ids=part_ids, as_of=as_of
        )
        floors = replenishment_screening.pool_floor_prices(
            db, [pool["group_id"] for pool in pools.values()]
        )

        application_id = _uid()
        version_id = _uid()
        suffix = application_id.replace("-", "")[:10].upper()
        application = ReplenishmentApplication(
            application_id=application_id,
            application_no=f"BLK-{as_of:%Y%m%d}-{suffix}",
            owner_username=user.username,
            owner_display_name=user.display_name
            or user.salesperson_name
            or user.username,
            salesperson_name_snapshot=user.salesperson_name,
            project_id=project.project_id,
            project_code_snapshot=project.project_code,
            project_name_snapshot=project.display_name,
            client_request_id=key,
            request_digest=request_digest,
            status="draft",
        )
        version = ReplenishmentApplicationVersion(
            version_id=version_id,
            application_id=application_id,
            version_no=1,
            status="draft",
            request_note=note,
            created_by=username,
        )
        db.add_all([application, version])
        db.flush()
        submitted_lines: list[ReplenishmentApplicationLine] = []
        auto_decisions: list[tuple[ReplenishmentApplicationLine, str, str]] = []
        auto_review_enabled = get_settings().replenishment_auto_review_enabled
        for line_no, item in enumerate(cleaned_items, 1):
            part = parts_by_id[item["part_id"]]
            pool = pools.get(part.id, {"group_id": None, "name": None, "version": None})
            part_facts = facts.get(part.id) or {}
            result = screenings[part.id]
            checks = result.as_dict()["checks"]
            decision, reason_code = _auto_review_decision(result)
            recommendations = (
                _auto_review_recommendations(db, part)
                if auto_review_enabled and decision == "rejected"
                else []
            )
            screening_snapshot = _json_value(
                {
                    "schema_version": 2 if auto_review_enabled else 1,
                    "as_of": as_of,
                    "lookback_days": PRICE_WINDOW_DAYS,
                    "checks": checks,
                    "anomaly_count": sum(not check["passed"] for check in checks),
                    "latest_sales": latest_sales.get(part.id) or {},
                    "pool_floor_ex_tax": floors.get(pool["group_id"]),
                    **({
                        "auto_review": {
                            "decision": decision,
                            "reason_code": reason_code,
                        },
                        "recommendations": recommendations,
                    } if auto_review_enabled else {}),
                }
            )
            digest_payload = {
                "part_id": part.id,
                "window": {"from": lower, "to": as_of},
                "purchase": part_facts.get("purchase"),
                "sales": part_facts.get("sales"),
                "screening": screening_snapshot,
            }
            line = ReplenishmentApplicationLine(
                line_id=_uid(),
                request_line_id=_uid(),
                version_id=version_id,
                line_no=line_no,
                part_id=part.id,
                pn_std=part.pn_std,
                description=part.description,
                brand=part.brand,
                unit=part.unit,
                quantity=item["quantity"],
                special_note=item["special_note"],
                pool_group_id=pool["group_id"],
                pool_name=pool["name"],
                pool_version=pool["version"],
                price_window_from=lower,
                price_window_to=as_of,
                price_as_of=as_of,
                purchase_stats_json=_json_value(part_facts.get("purchase") or {}),
                sales_stats_json=_json_value(part_facts.get("sales") or {}),
                screening_json=screening_snapshot,
                evidence_digest=_digest(digest_payload),
            )
            submitted_lines.append(line)
            if auto_review_enabled:
                auto_decisions.append((line, decision, reason_code))
            db.add(line)
        db.flush()
        content_digest = _digest(
            {
                "project_id": project.project_id,
                "client_request_id": key,
                "request_digest": request_digest,
                "version": _submission_content(version, submitted_lines),
            }
        )
        version.status = "submitted"
        version.content_digest = content_digest
        version.submitted_by = username
        version.submitted_at = _now()
        if auto_review_enabled:
            approved_count = sum(decision == "approved" for _line, decision, _reason in auto_decisions)
            rejected_count = len(auto_decisions) - approved_count
            review = ReplenishmentReview(
                review_id=_uid(),
                version_id=version_id,
                idempotency_key=f"auto-review:{version_id}",
                payload_digest=_digest({
                    "version_id": version_id,
                    "decisions": [(line.line_id, decision, reason) for line, decision, reason in auto_decisions],
                }),
                summary_note="系统按池归属及近182天采购/销售事实自动裁决",
                approved_count=approved_count,
                rejected_count=rejected_count,
                reviewed_by="system:replenishment-screening",
            )
            db.add(review)
            db.flush()
            for line, decision, reason_code in auto_decisions:
                db.add(ReplenishmentReviewLine(
                    review_line_id=_uid(),
                    review_id=review.review_id,
                    version_line_id=line.line_id,
                    decision=decision,
                    reason=None if decision == "approved" else reason_code,
                ))
            application.status = "needs_revision" if rejected_count else "approved"
        else:
            application.status = "submitted"
        application.version += 1
        _audit(
            db,
            application,
            "version_submitted",
            username,
            "按项目原子提交并冻结系统三查证据",
            version_id=version_id,
            after={
                "project_id": project.project_id,
                "line_count": len(submitted_lines),
                "content_digest": content_digest,
            },
        )
        if commit:
            db.commit()
        result = get_application(db, application_id, username=username, role=role)
        result["idempotent"] = False
        return result
    except Exception:
        db.rollback()
        raise


def submit(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
    expected_version: int,
) -> dict:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role, lock=True)
    if app.version != expected_version:
        raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
    version = _draft(db, app)
    if not (version.warehouse or "").strip():
        raise ReplenishmentError("提交前必须填写出库仓库")
    lines = list(
        db.scalars(
            select(ReplenishmentApplicationLine)
            .where(ReplenishmentApplicationLine.version_id == version.version_id)
            .order_by(ReplenishmentApplicationLine.line_no)
            .with_for_update()
        )
    )
    if not lines:
        raise ReplenishmentError("购物车为空，不能提交")
    # A revision must explicitly replace a rejected PN or explain why the same PN is exceptional.
    for line in lines:
        if line.source_line_id:
            source = db.get(ReplenishmentApplicationLine, line.source_line_id)
            if source is None or (line.part_id == source.part_id and not (line.special_note or "").strip()):
                raise ReplenishmentError(
                    f"第 {line.line_no} 条打回项需更换 PN，或填写特殊情况说明"
                )
    # Refresh the half-year evidence at the submission boundary, then freeze it permanently.
    for line in lines:
        part = _part(db, line.part_id)
        snapshot = _line_snapshot(db, part)
        pool, price = snapshot["pool"], snapshot["price"]
        line.pn_std, line.description, line.brand, line.unit = part.pn_std, part.description, part.brand, part.unit
        line.pool_group_id, line.pool_name, line.pool_version = pool["group_id"], pool["name"], pool["version"]
        line.price_window_from, line.price_window_to, line.price_as_of = price["date_from"], price["date_to"], price["as_of"]
        line.purchase_stats_json, line.sales_stats_json = price["purchase"], price["sales"]
        line.evidence_digest = price["digest"]
    db.flush()
    content_digest = _digest(_submission_content(version, lines))
    version.status = "submitted"
    version.content_digest = content_digest
    version.submitted_by = username
    version.submitted_at = _now()
    app.status = "submitted"
    app.version += 1
    _audit(
        db,
        app,
        "version_submitted",
        username,
        "提交补库版本，等待审核",
        version_id=version.version_id,
        after={"version_no": version.version_no, "content_digest": content_digest, "line_count": len(lines)},
    )
    db.commit()
    return get_application(db, application_id, username=username, role=role)


def record_review(
    db: Session,
    application_id: str,
    *,
    reviewer: str,
    version_id: str,
    content_digest: str,
    idempotency_key: str,
    external_reference: str | None,
    summary_note: str | None,
    decisions: list[dict],
) -> dict:
    _actor(db, reviewer)
    if not 1 <= len(decisions) <= MAX_LINES:
        raise ReplenishmentError(f"审核结论条数必须为 1-{MAX_LINES}")
    key = (idempotency_key or "").strip()
    if len(key) < 8 or len(key) > 128:
        raise ReplenishmentError("幂等键长度必须为 8-128 个字符")
    canonical = {
        "application_id": application_id,
        "version_id": version_id,
        "content_digest": content_digest,
        "external_reference": _clean_optional(external_reference, maximum=128),
        "summary_note": _clean_optional(summary_note, maximum=4000),
        "decisions": sorted(
            [
                {
                    "line_id": str(item.get("line_id", "")),
                    "decision": str(item.get("decision", "")),
                    "reason": _clean_optional(item.get("reason"), maximum=4000),
                }
                for item in decisions
            ],
            key=lambda item: item["line_id"],
        ),
    }
    payload_digest = _digest(canonical)
    # Serialize the global idempotency namespace before reading it. A row lock on
    # the application alone cannot protect the same key used concurrently for
    # different applications, and a pre-lock read can otherwise turn a retry
    # into stale_review/unique-constraint failure instead of an idempotent replay.
    db.execute(
        select(
            func.pg_advisory_xact_lock(
                func.hashtextextended(f"replenishment-review:{key}", 0)
            )
        )
    )
    app = db.scalar(
        select(ReplenishmentApplication)
        .where(ReplenishmentApplication.application_id == application_id)
        .with_for_update()
    )
    if app is None:
        raise ReplenishmentError("补库申请不存在", code="not_found", status_code=404)
    version = db.scalar(
        select(ReplenishmentApplicationVersion)
        .where(
            ReplenishmentApplicationVersion.version_id == version_id,
            ReplenishmentApplicationVersion.application_id == application_id,
        )
        .with_for_update()
    )
    if version is None or version.status != "submitted":
        raise ReplenishmentError("该版本不在待审核状态", code="stale_review", status_code=409)
    if version.content_digest != content_digest:
        raise ReplenishmentError("提交摘要不匹配，请重新获取版本", code="digest_mismatch", status_code=409)
    if not (version.submitted_by or "").strip():
        raise ReplenishmentError("提交人信息缺失，不能审核", code="corrupt", status_code=409)
    if version.submitted_by == reviewer:
        raise ReplenishmentError(
            "提交人与审核人不能是同一账号",
            code="separation_of_duties",
            status_code=409,
        )
    existing = db.scalar(select(ReplenishmentReview).where(ReplenishmentReview.idempotency_key == key))
    if existing is not None:
        if existing.payload_digest != payload_digest:
            raise ReplenishmentError("相同幂等键对应了不同审核内容", code="idempotency_conflict", status_code=409)
        return {
            "review_id": existing.review_id,
            "idempotent": True,
            "approved_count": existing.approved_count,
            "rejected_count": existing.rejected_count,
            "application_status": app.status,
        }
    if app.status != "submitted":
        raise ReplenishmentError("该版本不在待审核状态", code="stale_review", status_code=409)
    if db.scalar(select(ReplenishmentReview.review_id).where(ReplenishmentReview.version_id == version_id)):
        raise ReplenishmentError("该版本已有审核结果", code="already_reviewed", status_code=409)
    lines = list(
        db.scalars(
            select(ReplenishmentApplicationLine).where(
                ReplenishmentApplicationLine.version_id == version_id
            )
        )
    )
    by_id = {line.line_id: line for line in lines}
    submitted_ids = set(by_id)
    decision_ids = [item["line_id"] for item in canonical["decisions"]]
    if len(decision_ids) != len(set(decision_ids)) or set(decision_ids) != submitted_ids:
        raise ReplenishmentError("审核必须对提交版本的每一条明细恰好给出一次结论")
    for item in canonical["decisions"]:
        if item["decision"] not in {"approved", "rejected"}:
            raise ReplenishmentError("审核结论仅支持 approved 或 rejected")
        if item["decision"] == "rejected" and not item["reason"]:
            raise ReplenishmentError("打回条目必须填写原因")
    approved = sum(item["decision"] == "approved" for item in canonical["decisions"])
    rejected = len(canonical["decisions"]) - approved
    review = ReplenishmentReview(
        review_id=_uid(),
        version_id=version_id,
        idempotency_key=key,
        payload_digest=payload_digest,
        external_reference=canonical["external_reference"],
        summary_note=canonical["summary_note"],
        approved_count=approved,
        rejected_count=rejected,
        reviewed_by=reviewer,
    )
    db.add(review)
    db.flush()
    for item in canonical["decisions"]:
        db.add(
            ReplenishmentReviewLine(
                review_line_id=_uid(),
                review_id=review.review_id,
                version_line_id=item["line_id"],
                decision=item["decision"],
                reason=item["reason"],
            )
        )
    app.status = "needs_revision" if rejected else "approved"
    app.version += 1
    _audit(
        db,
        app,
        "review_recorded",
        reviewer,
        "记录外部审核 Agent 对精确提交版本的反馈",
        version_id=version_id,
        after={"approved_count": approved, "rejected_count": rejected},
    )
    db.commit()
    return {"review_id": review.review_id, "idempotent": False, "approved_count": approved, "rejected_count": rejected, "application_status": app.status}


def start_revision(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
    expected_version: int,
) -> dict:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role, lock=True)
    if app.version != expected_version:
        raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
    if app.status != "needs_revision":
        raise ReplenishmentError("当前申请没有待处理的打回条目", code="invalid_state", status_code=409)
    previous = _latest_version(db, app, lock=True)
    review = db.scalar(select(ReplenishmentReview).where(ReplenishmentReview.version_id == previous.version_id))
    if review is None:
        raise ReplenishmentError("审核反馈缺失", code="corrupt", status_code=409)
    rejected_ids = list(
        db.scalars(
            select(ReplenishmentReviewLine.version_line_id).where(
                ReplenishmentReviewLine.review_id == review.review_id,
                ReplenishmentReviewLine.decision == "rejected",
            )
        )
    )
    source_lines = list(
        db.scalars(
            select(ReplenishmentApplicationLine)
            .where(ReplenishmentApplicationLine.line_id.in_(rejected_ids))
            .order_by(ReplenishmentApplicationLine.line_no)
        )
    )
    version = ReplenishmentApplicationVersion(
        version_id=_uid(),
        application_id=app.application_id,
        version_no=app.latest_version_no + 1,
        parent_version_id=previous.version_id,
        status="draft",
        warehouse=previous.warehouse,
        request_note=previous.request_note,
        created_by=username,
    )
    db.add(version)
    db.flush()
    for index, source in enumerate(source_lines, 1):
        db.add(
            ReplenishmentApplicationLine(
                line_id=_uid(),
                request_line_id=source.request_line_id,
                source_line_id=source.line_id,
                version_id=version.version_id,
                line_no=index,
                part_id=source.part_id,
                pn_std=source.pn_std,
                description=source.description,
                brand=source.brand,
                unit=source.unit,
                quantity=source.quantity,
                special_note=None,
                pool_group_id=source.pool_group_id,
                pool_name=source.pool_name,
                pool_version=source.pool_version,
                price_window_from=source.price_window_from,
                price_window_to=source.price_window_to,
                price_as_of=source.price_as_of,
                purchase_stats_json=source.purchase_stats_json,
                sales_stats_json=source.sales_stats_json,
                evidence_digest=source.evidence_digest,
            )
        )
    app.latest_version_no += 1
    app.status = "draft"
    app.version += 1
    _audit(
        db,
        app,
        "revision_started",
        username,
        "按审核反馈创建二次提交草稿，仅复制打回条目",
        version_id=version.version_id,
        after={"version_no": version.version_no, "rejected_count": len(source_lines)},
    )
    db.commit()
    return get_application(db, application_id, username=username, role=role)


def apply_revision_atomic(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
    expected_application_version: int,
    client_request_id: str,
    resolutions: list[dict],
) -> dict:
    """Resolve exactly the rejected request lines without reopening approved facts."""
    key = (client_request_id or "").strip()
    if not 8 <= len(key) <= 128:
        raise ReplenishmentError("client_request_id 长度必须为 8-128 个字符")
    try:
        db.execute(select(func.pg_advisory_xact_lock(
            func.hashtextextended(f"replenishment-revision:{application_id}:{key}", 0)
        )))
        app = _application_scope(db, application_id, username=username, role=role, lock=True)
        if app.version != expected_application_version:
            raise ReplenishmentError("申请已被其他操作更新，请刷新后重试", code="version_conflict", status_code=409)
        previous = _latest_version(db, app, lock=True)
        review = db.scalar(select(ReplenishmentReview).where(ReplenishmentReview.version_id == previous.version_id))
        if review is None or app.status != "needs_revision":
            raise ReplenishmentError("当前申请没有待处理的打回条目", code="invalid_state", status_code=409)
        rejected = {
            line.request_line_id: line
            for line in db.scalars(select(ReplenishmentApplicationLine).where(
                ReplenishmentApplicationLine.version_id == previous.version_id
            ))
        }
        rejected_ids = {
            line.request_line_id
            for line in db.scalars(select(ReplenishmentApplicationLine).where(
                ReplenishmentApplicationLine.version_id == previous.version_id,
                ReplenishmentApplicationLine.line_id.in_(select(ReplenishmentReviewLine.version_line_id).where(
                    ReplenishmentReviewLine.review_id == review.review_id,
                    ReplenishmentReviewLine.decision == "rejected",
                )),
            ))
        }
        actions = {str(item.get("request_line_id")): item for item in resolutions}
        if set(actions) != rejected_ids or len(actions) != len(resolutions):
            raise ReplenishmentError("必须逐一处理全部打回行，不能遗漏或重复", code="revision_line_required", status_code=422)
        for event in db.scalars(select(ReplenishmentAuditEvent).where(
            ReplenishmentAuditEvent.application_id == app.application_id,
            ReplenishmentAuditEvent.action == "revision_started",
        )):
            if (event.after_json or {}).get("client_request_id") == key:
                result = get_application(db, app.application_id, username=username, role=role)
                result["idempotent"] = True
                db.commit()
                return result

        replacements = [item for item in actions.values() if item.get("action") == "replace"]
        replacement_ids = [int(item.get("part_id") or 0) for item in replacements]
        if len(replacement_ids) != len(set(replacement_ids)):
            raise ReplenishmentError("替换后的 PN 不能重复", code="duplicate_part", status_code=409)
        if any(item.get("action") not in {"replace", "remove"} for item in actions.values()):
            raise ReplenishmentError("只支持 replace 或 remove", code="invalid_revision_action", status_code=422)
        parts = list(db.scalars(select(DimPart).where(
            DimPart.id.in_(replacement_ids), DimPart.status == "active", DimPart.is_excluded.is_(False)
        ))) if replacement_ids else []
        parts_by_id = {part.id: part for part in parts}
        if set(parts_by_id) != set(replacement_ids):
            raise ReplenishmentError("替换 PN 不存在、已合并或已排除", code="part_unavailable", status_code=422)

        version = ReplenishmentApplicationVersion(
            version_id=_uid(), application_id=app.application_id,
            version_no=app.latest_version_no + 1, parent_version_id=previous.version_id,
            # 先 draft 插行（guard_replenishment_line_draft_only 要求行插入时版本为 draft），
            # 全部行写入后再置 submitted（2026-08-18）
            status="draft", request_note=previous.request_note, created_by=username,
            submitted_by=None, submitted_at=None,
        )
        db.add(version)
        db.flush()
        prior_review_lines = {
            line.version_line_id: line
            for line in db.scalars(select(ReplenishmentReviewLine).where(ReplenishmentReviewLine.review_id == review.review_id))
        }
        final_lines: list[ReplenishmentApplicationLine] = []
        decisions: list[tuple[ReplenishmentApplicationLine, str, str]] = []
        for source in db.scalars(select(ReplenishmentApplicationLine).where(
            ReplenishmentApplicationLine.version_id == previous.version_id
        ).order_by(ReplenishmentApplicationLine.line_no)):
            prior = prior_review_lines.get(source.line_id)
            if prior is not None and prior.decision == "rejected":
                item = actions[source.request_line_id]
                if item.get("action") == "remove":
                    continue
                part = parts_by_id[int(item["part_id"])]
                quantity = _integer_quantity(item.get("quantity", source.quantity))
                as_of = business_today()
                lower = as_of - timedelta(days=PRICE_WINDOW_DAYS - 1)
                facts = pool_price_analysis.aggregate_part_price_facts(db, [part.id], date_from=lower, date_to=as_of)
                pools = _pool_snapshots(db, [part.id])
                screening = replenishment_screening.screen(db, part_ids=[part.id], as_of=as_of, price_facts=facts)[part.id]
                decision, reason_code = _auto_review_decision(screening)
                pool = pools.get(part.id, {"group_id": None, "name": None, "version": None})
                latest_sales = replenishment_screening.latest_sales_history(db, part_ids=[part.id], as_of=as_of)
                floors = replenishment_screening.pool_floor_prices(
                    db, [pool["group_id"]] if pool.get("group_id") else []
                )
                snapshot = _json_value({
                    "schema_version": 2,
                    "as_of": as_of,
                    "lookback_days": PRICE_WINDOW_DAYS,
                    "checks": screening.as_dict()["checks"],
                    "anomaly_count": sum(not check["passed"] for check in screening.as_dict()["checks"]),
                    "latest_sales": latest_sales.get(part.id) or {},
                    "pool_floor_ex_tax": floors.get(pool["group_id"]),
                    "auto_review": {"decision": decision, "reason_code": reason_code},
                    "recommendations": _auto_review_recommendations(db, part) if decision == "rejected" else [],
                })
                new_line = ReplenishmentApplicationLine(
                    line_id=_uid(), request_line_id=source.request_line_id,
                    source_line_id=source.line_id, version_id=version.version_id,
                    line_no=len(final_lines) + 1, part_id=part.id, pn_std=part.pn_std,
                    description=part.description, brand=part.brand, unit=part.unit,
                    quantity=quantity, special_note=_clean_optional(item.get("special_note"), maximum=4000),
                    pool_group_id=pool["group_id"], pool_name=pool["name"], pool_version=pool["version"],
                    price_window_from=lower, price_window_to=as_of, price_as_of=as_of,
                    purchase_stats_json=_json_value((facts.get(part.id) or {}).get("purchase") or {}),
                    sales_stats_json=_json_value((facts.get(part.id) or {}).get("sales") or {}),
                    screening_json=snapshot,
                    evidence_digest=_digest({"part_id": part.id, "screening": snapshot}),
                )
                final_lines.append(new_line)
                decisions.append((new_line, decision, reason_code))
                db.add(new_line)
                continue
            copied = ReplenishmentApplicationLine(
                line_id=_uid(), request_line_id=source.request_line_id,
                source_line_id=source.line_id, version_id=version.version_id,
                line_no=len(final_lines) + 1, part_id=source.part_id, pn_std=source.pn_std,
                description=source.description, brand=source.brand, unit=source.unit,
                quantity=source.quantity, special_note=source.special_note,
                pool_group_id=source.pool_group_id, pool_name=source.pool_name, pool_version=source.pool_version,
                price_window_from=source.price_window_from, price_window_to=source.price_window_to,
                price_as_of=source.price_as_of, purchase_stats_json=source.purchase_stats_json,
                sales_stats_json=source.sales_stats_json, screening_json=source.screening_json,
                evidence_digest=source.evidence_digest,
            )
            final_lines.append(copied)
            decisions.append((copied, "approved", "approved_previous_version"))
            db.add(copied)
        if not final_lines:
            raise ReplenishmentError("至少保留一条有效申请明细", code="empty_revision", status_code=422)
        db.flush()
        content_digest = _digest({"client_request_id": key, "version": _submission_content(version, final_lines)})
        version.content_digest = content_digest
        # 行已全部写入：版本从 draft 置 submitted（2026-08-18）
        version.status = "submitted"
        version.submitted_by = username
        version.submitted_at = _now()
        app.latest_version_no = version.version_no
        app.version += 1
        # status 必须在任何 flush 之前赋值：guard_replenishment_application_identity
        # 要求每次 UPDATE version = OLD+1；若 status 单独成第二条 UPDATE 会触发
        # "identity/version is immutable"（2026-08-18）
        if get_settings().replenishment_auto_review_enabled:
            approved_count = sum(decision == "approved" for _line, decision, _reason in decisions)
            rejected_count = len(decisions) - approved_count
            app.status = "needs_revision" if rejected_count else "approved"
            auto_review = ReplenishmentReview(
                review_id=_uid(), version_id=version.version_id,
                idempotency_key=f"auto-review:{version.version_id}",
                payload_digest=_digest([(line.line_id, decision, reason) for line, decision, reason in decisions]),
                summary_note="系统按池归属及近182天采购/销售事实自动裁决",
                approved_count=approved_count, rejected_count=rejected_count,
                reviewed_by="system:replenishment-screening",
            )
            db.add(auto_review)
            db.flush()
            for line, decision, reason in decisions:
                db.add(ReplenishmentReviewLine(
                    review_line_id=_uid(), review_id=auto_review.review_id,
                    version_line_id=line.line_id, decision=decision,
                    reason=None if decision == "approved" else reason,
                ))
        else:
            app.status = "submitted"
        _audit(
            db, app, "revision_started", username, "处理自动审核打回行并重新提交",
            version_id=version.version_id,
            after={"client_request_id": key, "line_count": len(final_lines)},
        )
        db.commit()
        result = get_application(db, app.application_id, username=username, role=role)
        result["idempotent"] = False
        return result
    except Exception:
        db.rollback()
        raise


def _excel_safe(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    cleaned = _INVALID_XML_CONTROL.sub("", value)
    probe = cleaned.lstrip()
    if probe[:1] in ("=", "+", "-", "@"):
        cleaned = "'" + cleaned
    if len(cleaned) > MAX_EXCEL_TEXT:
        raise ReplenishmentError("导出文本超过 Excel 单元格上限", status_code=413)
    return cleaned


def _workbook_bytes(title: str, headers: list[str], rows: list[list[Any]], *, notice: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([notice])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A1"].font = Font(bold=True)
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        ws.append([_excel_safe(value) for value in row])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(2, len(headers)).coordinate}"
    for column_index, column in enumerate(ws.columns, 1):
        letter = get_column_letter(column_index)
        ws.column_dimensions[letter].width = min(42, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
    output = BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


def manual_review_workbook(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
) -> tuple[bytes, str]:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role)
    version = _latest_version(db, app)
    if version.status != "submitted":
        raise ReplenishmentError("只有已提交版本可以导出老板审核工作簿", code="invalid_state", status_code=409)
    payload = _version_payload(db, version)
    headers = [
        "申请单号", "版本", "申请人", "出库仓库", "序号", "PN", "产品描述", "数量", "单位",
        "所属池", "半年采购未税加权均价", "半年采购数量", "半年采购单数",
        "半年销售未税加权均价", "半年销售数量", "半年销售单数", "特殊情况说明", "审核结论", "打回原因",
    ]
    rows = []
    for line in payload["lines"]:
        purchase, sales, feedback = line["purchase"] or {}, line["sales"] or {}, line["review"] or {}
        rows.append([
            app.application_no, version.version_no, app.owner_display_name, version.warehouse, line["line_no"],
            line["pn_std"], line["description"], line["quantity"], line["unit"], line["pool"]["name"] or "未加入互通池",
            purchase.get("weighted_avg") if purchase else "半年内无有效样本", purchase.get("total_qty"), purchase.get("order_count"),
            sales.get("weighted_avg") if sales else "半年内无有效样本", sales.get("total_qty"), sales.get("order_count"),
            line["special_note"], feedback.get("decision"), feedback.get("reason"),
        ])
    data = _workbook_bytes(
        "老板人工审核",
        headers,
        rows,
        notice="人工导出供线下审核；文件本身不代表系统批准，也不会改变库存。",
    )
    _audit(db, app, "manual_exported", username, "导出老板人工审核工作簿", version_id=version.version_id)
    db.commit()
    return data, f"replenishment-{app.application_no}-v{version.version_no}-manual-review.xlsx"


def system_screening_workbook(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
) -> tuple[bytes, str]:
    """Export only the immutable facts captured by atomic submission."""
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role)
    if app.is_legacy_project_unbound:
        raise ReplenishmentError(
            "历史补库申请尚未绑定真实项目，不能导出复核包",
            code="legacy_project_unbound",
            status_code=409,
        )
    if _workflow_mode(app) != "system_screening":
        raise ReplenishmentError(
            "历史补库申请没有冻结的系统三查证据",
            code="legacy_screening_unavailable",
            status_code=409,
        )
    version = _latest_version(db, app)
    if version.status != "submitted":
        raise ReplenishmentError(
            "只有已提交版本可以导出人工复核包",
            code="invalid_state",
            status_code=409,
        )
    payload = _version_payload(db, version)
    lines = payload["lines"]
    if any(line["screening"] is None for line in lines):
        raise ReplenishmentError(
            "该历史版本没有冻结的系统三查证据",
            code="frozen_evidence_unavailable",
            status_code=409,
        )

    headers = [
        "申请单号", "项目编码", "项目名称", "版本", "申请人", "序号", "PN",
        "产品描述", "数量", "单位", "最近销售日期", "最近销售未税单价",
        "池内最低价(未税)", "与池内最低价对比", "①当前池有效性",
        "②近182天购销事实", "③冷门零样本边界", "异常项数量", "异常证据", "证据日期",
    ]
    rows = []
    for line in lines:
        evidence = line["screening"]
        checks = {check["key"]: check for check in evidence["checks"]}
        pool_check = checks["pool_membership"]
        activity = checks["recent_activity"]
        niche = checks["niche_pn"]
        recent = line["latest_sales"] or {}
        floor = line["pool_floor_ex_tax"]
        recent_price = recent.get("price_ex_tax")
        if floor is None or recent_price is None:
            comparison = _NO_VALUE
        else:
            comparison = float(Decimal(str(recent_price)) - Decimal(str(floor)))
        anomalies = [
            label
            for check, label in (
                (pool_check, "无当前有效互通池证据"),
                (activity, "近182天无有效购销样本"),
                (niche, "冷门零样本"),
            )
            if not check["passed"]
        ]
        rows.append([
            app.application_no, app.project_code_snapshot, app.project_name_snapshot,
            version.version_no, app.owner_display_name, line["line_no"], line["pn_std"],
            line["description"], line["quantity"], line["unit"],
            recent.get("order_date") or _NO_VALUE,
            recent_price if recent_price is not None else _NO_VALUE,
            float(floor) if floor is not None else _NO_VALUE, comparison,
            pool_check["detail"].get("pool_name") or (
                "无法判断（PN 未标准化）"
                if pool_check["detail"].get("in_pool") is None
                else "未加入互通池"
            ),
            f"采购 {activity['detail']['purchase_samples']} 单 / "
            f"销售 {activity['detail']['sales_samples']} 单",
            "零样本" if niche["detail"]["is_niche"] else "存在样本",
            evidence["anomaly_count"],
            "；".join(anomalies) if anomalies else "未发现规则异常；仍需人工复核",
            evidence["as_of"],
        ])
    data = _workbook_bytes(
        "人工复核包",
        headers,
        rows,
        notice=("提交时冻结的系统三查事实与需注意项，供线下人工复核；系统不记录人工结论。"
                "本申请为独立记录，不进入 WBDD、不参与项目成本与对账。"),
    )
    return data, f"replenishment-{app.application_no}-v{version.version_no}-system-screening.xlsx"


def _approved_lines(db: Session, app: ReplenishmentApplication) -> list[tuple[ReplenishmentApplicationVersion, ReplenishmentApplicationLine]]:
    versions = list(
        db.scalars(
            select(ReplenishmentApplicationVersion)
            .where(
                ReplenishmentApplicationVersion.application_id == app.application_id,
                ReplenishmentApplicationVersion.status == "submitted",
            )
            .order_by(ReplenishmentApplicationVersion.version_no)
        )
    )
    approved_by_request: dict[str, tuple[ReplenishmentApplicationVersion, ReplenishmentApplicationLine]] = {}
    for version in versions:
        review = db.scalar(select(ReplenishmentReview).where(ReplenishmentReview.version_id == version.version_id))
        if review is None:
            continue
        approved_ids = list(
            db.scalars(
                select(ReplenishmentReviewLine.version_line_id).where(
                    ReplenishmentReviewLine.review_id == review.review_id,
                    ReplenishmentReviewLine.decision == "approved",
                )
            )
        )
        for line in db.scalars(
            select(ReplenishmentApplicationLine).where(
                ReplenishmentApplicationLine.line_id.in_(approved_ids),
                ReplenishmentApplicationLine.version_id == version.version_id,
            )
        ):
            approved_by_request[line.request_line_id] = (version, line)
    return sorted(approved_by_request.values(), key=lambda pair: (pair[0].version_no, pair[1].line_no))


def wbdd_subset_workbook(
    db: Session,
    application_id: str,
    *,
    username: str,
    role: str,
) -> tuple[bytes, str]:
    _actor(db, username)
    app = _application_scope(db, application_id, username=username, role=role)
    if app.status != "approved":
        raise ReplenishmentError("全部条目通过审核后才能导出 WBDD 字段子集", code="invalid_state", status_code=409)
    if not (app.salesperson_name_snapshot or "").strip():
        raise ReplenishmentError(
            "销售人员业务映射缺失，请管理员补齐账号的销售人员映射后重新建单",
            code="salesperson_mapping_missing",
            status_code=409,
        )
    approved = _approved_lines(db, app)
    headers = [
        "需求类型", "销售人员", "出库仓库(必填)", "需求明细.序号",
        "需求明细.需供货产品", "需求明细.产品描述", "需求明细.需求数量",
    ]
    rows = []
    for index, (version, line) in enumerate(approved, 1):
        rows.append([
            "补库供货", app.salesperson_name_snapshot, version.warehouse, index, line.pn_std, line.description, float(line.quantity)
        ])
    data = _workbook_bytes(
        "WBDD字段子集",
        headers,
        rows,
        notice=(
            "WBDD 字段子集（录入辅助，非直接导入）。本文件不含氚云源数据ID、明细ID、需求单号和 F 字段码，"
            "不能直接回灌；也不会改变库存。"
        ),
    )
    latest = _latest_version(db, app)
    _audit(db, app, "wbdd_draft_exported", username, "导出 WBDD 字段子集录入辅助表", version_id=latest.version_id)
    db.commit()
    return data, f"replenishment-{app.application_no}-wbdd-subset.xlsx"
