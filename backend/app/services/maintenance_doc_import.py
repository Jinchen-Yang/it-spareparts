"""氚云三单（RKD 入库 / 退货返库 / BXD 报销）导入服务（C1b）。

- parse：字段名表头识别、主明细分组、raw_json 全量保留；
- store_preview：落 raw，零 canonical 写入；
- apply：
  - return_order：维保相关返库明细按 WBDD/XSDD/项目名稳定解析项目后，对前置库
    做 return_out（未用件收回）。已消耗件不在账本 → 负结存失败关闭并跳过，
    天然只对未用件生效；坏品返还事实由 RKD（F3）统计；
  - rkd_inbound / bxd_expense：本切片只落 raw 并标记 applied，坏件返还接线（F3）
    与报销对账（C4）在对应切片实现。
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, time, timezone
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.dimensions import DimPart
from app.models.maintenance import FMaintenanceOrder
from app.models.maintenance_doc_import import (
    MaintenanceDocHeadRow,
    MaintenanceDocImportBatch,
    MaintenanceDocLineRow,
    MaintenanceRkdReturnLine,
)
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
)
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.services.date_loose import parse_amount_loose, parse_date_loose
from app.services import maintenance_front_stock as front_stock

MAX_PREVIEW_BYTES = 256 * 1024 * 1024
_RKD_RE = re.compile(r"(RKD-\d{8}-\d{4})")
_BXD_RE = re.compile(r"(BXD-\d{8}-\d{4})")
_WBDD_RE = re.compile(r"(WBDD-\d{8}-\d{4})")
_XSDD_RE = re.compile(r"(XSDD-\d{8}-\d{4})")

# doc_type → (sheet 名候选, 主表锚列, 主表列, 明细列, 归一化规则)
_SPECS: dict[str, dict] = {
    "rkd_inbound": {
        "sheets": ("Sheet1",),
        "anchor": ("入库单号", "入库日期", "入库类别"),
        "head": [
            "入库单号", "入库日期", "入库类别", "入库备件/整机", "数据状态", "仓储中心",
            "项目名称", "维保需求单", "采购订单", "供应商", "退货类型", "退返入库通知单",
            "备注",
        ],
        "line": [
            "备件明细.备件PN", "备件明细.备件自贴码", "备件明细.备件SN", "备件明细.测试结果",
            "备件明细.入库库位", "备件明细.备件描述", "备件明细.入库数量", "备件明细.销售单价",
            "备件明细.销售金额", "备件明细.采购单价", "备件明细.金额", "备件明细.数据ID(不可修改)",
            "备件明细.序号",
        ],
        "head_no": "入库单号",
        "head_date": "入库日期",
        "category": "入库类别",
        "wbdd": "维保需求单",
        "xsdd": None,
        "project": "项目名称",
        "line_key_cols": ("备件明细.数据ID(不可修改)", "备件明细.序号"),
        "pn_col": "备件明细.备件PN",
        "qty_col": "备件明细.入库数量",
        "amount_col": "备件明细.金额",
        "test_col": "备件明细.测试结果",
    },
    "return_order": {
        "sheets": ("Sheet1",),
        "anchor": ("返库类别", "返库日期", "数据ID(不可修改)"),
        "head": [
            "返库类别", "返库类型", "返库备件/整机", "返库日期", "客户名称", "项目名称",
            "维保销售订单", "维保需求单(备件)", "维保需求单", "仓储中心", "数据状态", "备注",
            "返库单号", "数据ID(不可修改)",
        ],
        "line": [
            "备件明细.备件自贴码", "备件明细.备件PN", "备件明细.备件SN号", "备件明细.备件测试结果",
            "备件明细.入库库位", "备件明细.产品描述", "备件明细.返库数量", "备件明细.数据ID(不可修改)",
            "备件明细.序号", "备件明细.入库仓库",
        ],
        "head_no": "返库单号",
        "head_date": "返库日期",
        "category": "返库类别",
        "wbdd": "维保需求单(备件)",
        "xsdd": "维保销售订单",
        "project": "项目名称",
        "line_key_cols": ("备件明细.数据ID(不可修改)", "备件明细.序号"),
        "pn_col": "备件明细.备件PN",
        "qty_col": "备件明细.返库数量",
        "amount_col": None,
        "test_col": "备件明细.备件测试结果",
    },
    "bxd_expense": {
        "sheets": ("费用报销_支付单", "Sheet1"),
        "anchor": ("费用单号", "数据ID(不可修改)"),
        "head": [
            "费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单", "销售订单",
            "客户名称", "销售人员", "报销金额", "实付金额", "付款方式", "报销日期", "备注",
            "数据状态", "数据ID(不可修改)", "流程状态",
        ],
        "line": [
            "报销明细.费用分类", "报销明细.单据数量", "报销明细.报销金额", "报销明细.备注",
            "报销明细.数据ID(不可修改)", "报销明细.序号",
        ],
        "head_no": "费用单号",
        "head_date": "报销日期",
        "category": "报销类别",
        "wbdd": None,
        "xsdd": "维保销售订单",
        "project": None,
        "line_key_cols": ("报销明细.数据ID(不可修改)", "报销明细.序号"),
        "pn_col": None,
        "qty_col": None,
        "amount_col": "报销明细.报销金额",
        "test_col": None,
    },
}


class DocParseError(RuntimeError):
    """三单文件不可解析。"""


class DocBatchError(RuntimeError):
    """三单批次状态错误。"""


class DocScopeDenied(RuntimeError):
    """批次包含操作者项目范围之外的项目：整批拒绝。"""


@dataclass
class DocHeadData:
    row_no: int
    values: dict[str, str]
    lines: list["DocLineData"] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)


@dataclass
class DocLineData:
    row_no: int
    values: dict[str, str]
    issues: list[str] = field(default_factory=list)


def _header_index(headers: list[str], name: str) -> int | None:
    variants = {name, f"{name}(必填)", f"{name}(不可修改)"}
    for idx, value in enumerate(headers, 1):
        if value in variants:
            return idx
    return None


def _cell(row: tuple, index: int | None) -> str | None:
    if index is None or index > len(row):
        return None
    value = row[index - 1]
    if value is None or isinstance(value, (bytes, bytearray)):
        return None
    text = str(value).strip()
    if len(text) > 4096:
        return None
    return text or None


def _non_empty(row: tuple) -> bool:
    return bool(row) and any(
        v is not None and not isinstance(v, (bytes, bytearray)) and str(v).strip()
        for v in row
    )


def _clean(raw: str | None, pattern: re.Pattern) -> str | None:
    if not raw:
        return None
    match = pattern.search(raw)
    return match.group(1) if match else None


def parse_doc_workbook(
    doc_type: str,
    data: bytes,
    filename: str,
    *,
    column_aliases: dict | None = None,
) -> dict:
    if doc_type not in _SPECS:
        raise DocParseError(f"未知单据类型：{doc_type}")
    spec = _SPECS[doc_type]
    if len(data) > MAX_PREVIEW_BYTES:
        raise DocParseError("单据文件超过大小上限")
    try:
        # 氚云导出的 sheet dimension 不可靠，read_only 模式会截断行——用普通模式。
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise DocParseError(f"无法读取 Excel 文件：{type(exc).__name__}") from exc
    try:
        sheet = None
        for candidate in spec["sheets"]:
            if candidate in workbook.sheetnames:
                sheet = workbook[candidate]
                break
        if sheet is None:
            sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        result = _parse_doc_rows(rows, doc_type, spec, column_aliases)
    finally:
        workbook.close()
    return {
        **result,
        "file_hash": hashlib.sha256(data).hexdigest(),
        "filename": filename,
    }


def _parse_doc_rows(rows, doc_type: str, spec: dict, column_aliases: dict | None) -> dict:
    try:
        header_rows: list[tuple] = []
        for row in rows:
            if header_rows or any(
                v is not None and str(v).startswith("F00") for v in row
            ):
                header_rows.append(row)
                if len(header_rows) >= 2:
                    break
        if len(header_rows) < 2:
            raise DocParseError("单据缺少字段码/字段名双表头")
        headers = [str(v).strip() if v is not None else "" for v in header_rows[-1]]
        if column_aliases:
            headers = [column_aliases.get(h, h) for h in headers]
        head_indexes = {name: _header_index(headers, name) for name in spec["head"]}
        line_indexes = {name: _header_index(headers, name) for name in spec["line"]}
        if all(head_indexes[a] is None for a in spec["anchor"]):
            raise DocParseError(f"单据缺少主表识别列：{spec['anchor']}")
        if spec["pn_col"] is not None and line_indexes[spec["pn_col"]] is None:
            raise DocParseError(f"单据缺少明细列：{spec['pn_col']}")

        heads: list[DocHeadData] = []
        current: DocHeadData | None = None
        row_no = 3
        for row in rows:
            if not _non_empty(row):
                continue
            anchor_hit = any(
                _cell(row, head_indexes[name]) for name in spec["anchor"]
            )
            head_values = {
                name: _cell(row, head_indexes[name]) for name in spec["head"]
            }
            line_values = {
                name: _cell(row, line_indexes[name]) for name in spec["line"]
            }
            if anchor_hit:
                current = DocHeadData(row_no=row_no, values=head_values)
                heads.append(current)
            has_line = bool(
                line_values.get(spec["pn_col"] or spec["amount_col"] or "")
                or (
                    spec["qty_col"]
                    and line_values.get(spec["qty_col"])
                )
            )
            if has_line:
                if current is None:
                    raise DocParseError("明细行出现在主表行之前")
                current.lines.append(DocLineData(row_no=row_no, values=line_values))
            row_no += 1
    except DocParseError:
        raise
    return {
        "doc_type": doc_type,
        "heads": heads,
        "line_count": sum(len(h.lines) for h in heads),
    }


def _head_no(doc_type: str, values: dict[str, str], spec: dict) -> str | None:
    col = spec["head_no"]
    raw = values.get(col) if col else None
    pattern = {"rkd_inbound": _RKD_RE, "return_order": None, "bxd_expense": _BXD_RE}[doc_type]
    if pattern is not None:
        return _clean(raw, pattern)
    # 返库单：优先返库单号，否则数据ID
    if raw:
        return raw
    return values.get("数据ID(不可修改)")


def store_preview(
    db: Session, parsed: dict, operated_by: str, *, idempotency_key: str
) -> str:
    spec = _SPECS[parsed["doc_type"]]
    existing = db.execute(
        select(MaintenanceDocImportBatch).where(
            MaintenanceDocImportBatch.uploaded_by == operated_by,
            MaintenanceDocImportBatch.idempotency_key == idempotency_key,
        )
    ).scalar_one_or_none()
    if existing is not None:
        if existing.file_hash != parsed["file_hash"]:
            raise DocBatchError("同一 Idempotency-Key 对应不同文件内容，拒绝重放")
        return existing.batch_id
    batch = MaintenanceDocImportBatch(
        batch_id=str(uuid4()),
        doc_type=parsed["doc_type"],
        file_hash=parsed["file_hash"],
        filename=parsed["filename"][:255],
        idempotency_key=idempotency_key,
        uploaded_by=operated_by,
        head_rows=len(parsed["heads"]),
        line_rows=parsed["line_count"],
        status="pending",
    )
    db.add(batch)
    db.flush()
    issue_rows = 0
    for head in parsed["heads"]:
        values = head.values
        head_no = _head_no(parsed["doc_type"], values, spec)
        head_date, _ = parse_date_loose(
            values.get(spec["head_date"]) if spec["head_date"] else None
        )
        wbdd = _clean(values.get(spec["wbdd"]), _WBDD_RE) if spec["wbdd"] else None
        xsdd = _clean(values.get(spec["xsdd"]), _XSDD_RE) if spec["xsdd"] else None
        if not head_no and not values.get("数据ID(不可修改)"):
            head.issues.append("单据编号缺失")
        if head.issues:
            issue_rows += 1
        head_row = MaintenanceDocHeadRow(
            row_id=str(uuid4()),
            batch_id=batch.batch_id,
            row_no=head.row_no,
            raw_json={k: (v or "") for k, v in values.items()},
            head_no=head_no,
            head_date=head_date,
            category=values.get(spec["category"]),
            wbdd_no=wbdd,
            xsdd_no=xsdd,
            project_name=(values.get(spec["project"]) if spec["project"] else None),
            data_status=values.get("数据状态"),
            issues=head.issues,
        )
        db.add(head_row)
        db.flush()
        for line in head.lines:
            if line.issues:
                issue_rows += 1
            line_values = line.values
            key_cols = spec["line_key_cols"]
            line_key = next(
                (line_values.get(c) for c in key_cols if line_values.get(c)), None
            )
            pn = line_values.get(spec["pn_col"]) if spec["pn_col"] else None
            qty = (
                parse_amount_loose(line_values.get(spec["qty_col"]))
                if spec["qty_col"]
                else None
            )
            amount = (
                parse_amount_loose(line_values.get(spec["amount_col"]))
                if spec["amount_col"]
                else None
            )
            db.add(
                MaintenanceDocLineRow(
                    row_id=str(uuid4()),
                    batch_id=batch.batch_id,
                    head_row_id=head_row.row_id,
                    row_no=line.row_no,
                    raw_json={k: (v or "") for k, v in line_values.items()},
                    line_key=line_key,
                    pn=(pn or None),
                    qty=qty,
                    amount=amount,
                    test_result=(
                        line_values.get(spec["test_col"]) if spec["test_col"] else None
                    ),
                    warehouse=None,
                    location=None,
                    issues=line.issues,
                )
            )
    batch.issue_rows = issue_rows
    batch.report_json = {
        "head_rows": batch.head_rows,
        "line_rows": batch.line_rows,
        "issue_rows": issue_rows,
    }
    db.commit()
    return batch.batch_id


def _resolve_project_id(db: Session, head: MaintenanceDocHeadRow) -> str | None:
    if head.wbdd_no:
        order = db.execute(
            select(FMaintenanceOrder).where(FMaintenanceOrder.order_no == head.wbdd_no)
        ).scalar_one_or_none()
        if order is not None:
            assignment = db.execute(
                select(MaintenanceSourceOrderAssignment).where(
                    MaintenanceSourceOrderAssignment.source_order_id == order.raw_order_id,
                    MaintenanceSourceOrderAssignment.is_active.is_(True),
                )
            ).scalar_one_or_none()
            if assignment is not None:
                return assignment.project_id
    if head.xsdd_no:
        contract = db.execute(
            select(MaintenanceProjectContract).where(
                MaintenanceProjectContract.contract_no == head.xsdd_no
            )
        ).scalar_one_or_none()
        if contract is not None:
            return contract.project_id
    if head.project_name:
        project = db.execute(
            select(MaintenanceProject).where(
                MaintenanceProject.project_code == head.project_name
            )
        ).scalar_one_or_none()
        if project is not None:
            return project.project_id
    return None


def _fail_batch(db: Session, batch_id: str, failures: list[str]) -> None:
    """失败关闭：回滚本批全部 canonical 写入并标记 failed。"""
    db.rollback()
    batch = db.get(MaintenanceDocImportBatch, batch_id)
    batch.status = "failed"
    batch.report_json = {
        **(batch.report_json or {}),
        "rejection_reason": "单据批次存在应用异常，整批零写入",
        "failures": failures[:20],
    }
    db.commit()


def apply_batch(
    db: Session,
    batch_id: str,
    operated_by: str,
    *,
    allowed_project_ids: set[str] | None = None,
) -> dict:
    batch = db.get(MaintenanceDocImportBatch, batch_id)
    if batch is None:
        raise DocBatchError("单据批次不存在")
    if batch.status == "applied":
        raise DocBatchError("单据批次已应用，不能重复应用")
    if batch.status == "failed":
        raise DocBatchError("单据批次已因异常被拒绝，需重新上传")
    # preview 异常行 → 整批失败关闭（raw 保留）
    bad_rows = db.execute(
        select(MaintenanceDocHeadRow).where(
            MaintenanceDocHeadRow.batch_id == batch_id,
            func.cardinality(MaintenanceDocHeadRow.issues) > 0,
        )
    ).scalars().all()
    if bad_rows:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejected_rows": len(bad_rows),
            "rejection_reason": "单据批次存在关键异常行，整批拒绝应用",
        }
        db.commit()
        raise DocBatchError(f"单据批次存在 {len(bad_rows)} 行关键异常，整批拒绝应用")
    summary = {
        "doc_type": batch.doc_type,
        "applied_lines": 0,
        "skipped_lines": 0,
        "ignored_heads": 0,
        "canonical_effect": "none",
    }
    head_rows = (
        db.execute(
            select(MaintenanceDocHeadRow)
            .where(MaintenanceDocHeadRow.batch_id == batch_id)
            .order_by(MaintenanceDocHeadRow.row_no)
        )
        .scalars()
        .all()
    )
    if batch.doc_type == "return_order":
        summary["canonical_effect"] = "front_stock_return_out"
        for head in head_rows:
            if head.data_status and head.data_status != "已生效":
                summary["ignored_heads"] += 1
                continue
            # 窄表没有稳定单号/明细键 → 待治理事实，不做 canonical 写入
            if head.head_no is None:
                summary["skipped_lines"] += 1
                continue
            project_id = _resolve_project_id(db, head)
            if project_id is None:
                summary["skipped_lines"] += 1
                continue
            if allowed_project_ids is not None and project_id not in allowed_project_ids:
                raise DocScopeDenied(
                    f"单据批次包含无权项目（{head.head_no}），整批拒绝"
                )
            project = db.get(MaintenanceProject, project_id)
            warehouse_name = project.project_code if project is not None else project_id
            lines = (
                db.execute(
                    select(MaintenanceDocLineRow)
                    .where(MaintenanceDocLineRow.head_row_id == head.row_id)
                    .order_by(MaintenanceDocLineRow.row_no)
                )
                .scalars()
                .all()
            )
            failures: list[str] = []
            for line in lines:
                # 坏品/故障件是消耗返还（F3 分子），不扣前置库账本
                if line.test_result in ("坏品", "坏件", "故障"):
                    continue
                if line.line_key is None:
                    failures.append(f"{head.head_no}: 明细缺少稳定键")
                    continue
                if line.pn is None or line.qty is None or line.qty <= 0:
                    failures.append(f"{head.head_no}: 明细缺少 PN 或数量非法")
                    continue
                part_id = db.execute(
                    select(DimPart).where(DimPart.pn_std == line.pn)
                ).scalar_one_or_none()
                if part_id is None:
                    failures.append(f"{head.head_no}: 未知 PN {line.pn}")
                    continue
                try:
                    front_stock.apply_movement(
                        db,
                        project_id=project_id,
                        part_id=part_id.id,
                        kind="return_out",
                        source_type="return_order_line",
                        source_ref=f"return:{head.head_no}:{line.line_key}",
                        qty=line.qty,
                        warehouse_name=warehouse_name,
                        occurred_at=datetime.combine(
                            head.head_date, datetime.min.time()
                        ).replace(tzinfo=timezone.utc)
                        if head.head_date
                        else None,
                        reason=f"返库单 {head.head_no} 未用件收回",
                        operated_by=operated_by,
                    )
                    summary["applied_lines"] += 1
                except front_stock.FrontStockPayloadConflict:
                    raise
                except front_stock.FrontStockError as exc:
                    failures.append(f"{head.head_no}: {exc}")
            if failures:
                # 失败关闭：回滚本批全部流水写入
                db.rollback()
                batch = db.get(MaintenanceDocImportBatch, batch_id)
                batch.status = "failed"
                batch.report_json = {
                    **(batch.report_json or {}),
                    "rejection_reason": "单据批次存在应用异常，整批零写入",
                    "failures": failures[:20],
                }
                db.commit()
                raise DocBatchError(
                    f"单据批次应用失败 {len(failures)} 行，整批拒绝（已回滚全部写入）"
                )
    elif batch.doc_type == "rkd_inbound":
        summary["canonical_effect"] = "rkd_return_facts"
        for head in head_rows:
            if head.data_status and head.data_status != "已生效":
                summary["ignored_heads"] += 1
                continue
            if head.head_no is None:
                summary["skipped_lines"] += 1
                continue
            project_id = _resolve_project_id(db, head)
            if project_id is None:
                summary["skipped_lines"] += 1
                continue
            if allowed_project_ids is not None and project_id not in allowed_project_ids:
                raise DocScopeDenied(
                    f"单据批次包含无权项目（{head.head_no}），整批拒绝"
                )
            lines = (
                db.execute(
                    select(MaintenanceDocLineRow)
                    .where(MaintenanceDocLineRow.head_row_id == head.row_id)
                    .order_by(MaintenanceDocLineRow.row_no)
                )
                .scalars()
                .all()
            )
            failures: list[str] = []
            pending_facts: list[tuple[str, MaintenanceDocLineRow]] = []
            for line in lines:
                # 只有坏品/坏件/故障是消耗返还事实（Q8：坏件返还=氚云收货入库单）
                if line.test_result not in ("坏品", "坏件", "故障"):
                    continue
                if line.qty is None or line.qty <= 0:
                    failures.append(f"{head.head_no}: 坏件明细缺少合法数量")
                    continue
                source_ref = "rkd:" + hashlib.sha1(
                    f"{head.head_no}:{line.line_key or line.row_no}".encode("utf-8")
                ).hexdigest()
                pending_facts.append((source_ref, line))
            if failures:
                _fail_batch(db, batch_id, failures)
                raise DocBatchError(
                    f"单据批次应用失败 {len(failures)} 行，整批拒绝（已回滚全部写入）"
                )
            # 幂等预检：同一 (单号, 明细键) 已入账 → 冲突失败关闭（防跨批次重复计数）
            refs = [ref for ref, _ in pending_facts]
            if refs:
                existing = db.execute(
                    select(MaintenanceRkdReturnLine.source_ref).where(
                        MaintenanceRkdReturnLine.source_ref.in_(refs)
                    )
                ).scalars().all()
                if existing:
                    _fail_batch(
                        db, batch_id, [f"入库单坏件明细已入账：{', '.join(existing[:5])}"]
                    )
                    raise DocBatchError("入库单坏件明细与已入账事实冲突，整批拒绝")
            for source_ref, line in pending_facts:
                pn = (line.pn or "").strip() or "(无PN)"
                # 未知 PN 只影响回收监控的 PN 维度，不影响返还率分子数量 → part_id 置空
                part_id = (
                    db.scalar(select(DimPart.id).where(DimPart.pn_std == line.pn))
                    if line.pn
                    else None
                )
                db.add(
                    MaintenanceRkdReturnLine(
                        rkd_line_id=str(uuid4()),
                        batch_id=batch.batch_id,
                        head_row_id=head.row_id,
                        project_id=project_id,
                        head_no=head.head_no,
                        source_ref=source_ref,
                        part_id=part_id,
                        pn=pn,
                        qty=line.qty,
                        test_result=line.test_result,
                        occurred_at=(
                            datetime.combine(
                                head.head_date, datetime.min.time()
                            ).replace(tzinfo=timezone.utc)
                            if head.head_date
                            else None
                        ),
                    )
                )
                summary["applied_lines"] += 1
    else:
        # bxd_expense：本切片只落 raw；C4 接线
        summary["canonical_effect"] = "pending_c4_expense_reconcile"
    batch.status = "applied"
    batch.applied_by = operated_by
    batch.applied_at = datetime.now(timezone.utc)
    db.commit()
    return summary
