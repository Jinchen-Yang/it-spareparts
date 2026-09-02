"""Registry-driven, all-project maintenance business-form import core.

This is deliberately separate from the existing round-trip workbooks.  Those
workbooks are signed, project-scoped editing protocols; this module accepts raw
business-system exports and projects only unambiguous rows onto canonical
maintenance facts.

Protocol::

    xlsx -> detect adapter/header -> immutable preview plan + commit token
         -> lock/recheck -> one transaction -> canonical services + audit

Preview persists evidence in ``sys_import_batch.report_json`` but performs no
domain writes.  Apply is idempotent by batch and by ``(adapter, file_hash)``.
Adapters are registered objects, so adding another form does not change the
parser, token, persistence, pagination, or apply transaction machinery.
"""

from __future__ import annotations

import hashlib
import hmac
import io
import json
import re
import secrets
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Protocol

from openpyxl import load_workbook
from sqlalchemy import and_, func, or_, select
from sqlalchemy.orm import Session

from app import config
from app.etl import mapping
from app.business_time import business_today
from app.models.maintenance import FMaintenanceOrder, MaintenanceDemandTombstone
from app.models.maintenance_project import MaintenanceProject, MaintenanceProjectContract
from app.models.maintenance_project_operations import MaintenanceCollectionSnapshot
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.models.sales import FSalesOrder
from app.models.system import SysAuditLog, SysImportBatch, SysImportError
from app.services import maintenance_project_catalog as catalog
from app.services import maintenance_project_identity
from app.services import maintenance_project_operations as operations

PROTOCOL_VERSION = "maintenance-bulk-import-v1"
MAX_PREVIEW_BYTES = 16 * 1024 * 1024
MAX_PREVIEW_ROWS = 50_000
HEADER_SCAN_ROWS = 20
MONEY_LIMIT = Decimal("1000000000000")
SUPPORTED_BATCH_TYPES = ("maint_contract", "maint_receipt")
TRANSFER_BATCH_TYPE = "maint_bulk"
TRANSFER_SCHEMA_VERSION = "maintenance-project-batch-transfer-v1"
MAX_TRANSFER_FILES = 20
MAX_TRANSFER_TOTAL_BYTES = 64 * 1024 * 1024
TRANSFER_TOKEN_TTL = timedelta(minutes=30)


class BulkImportError(ValueError):
    """Base domain error."""


class BulkImportInvalid(BulkImportError):
    def __init__(self, message: str, *, issues: list[dict] | None = None):
        super().__init__(message)
        self.issues = issues or []


class BulkImportConflict(BulkImportError):
    pass


class BulkImportNotFound(BulkImportError):
    pass


class BulkImportScopeDenied(BulkImportError):
    pass


@dataclass(frozen=True)
class DetectedSheet:
    name: str
    header_row: int
    header_rows: tuple[int, ...]
    headers: tuple[str, ...]
    system_headers: tuple[str, ...]
    field_indexes: dict[str, int]
    field_matches: dict[str, dict]
    rows: tuple[tuple[int, tuple[Any, ...]], ...]


@dataclass(frozen=True)
class PreviewArtifact:
    adapter_key: str
    file_type: str
    file_hash: str
    filename: str
    plan: dict


class FormAdapter(Protocol):
    key: str
    file_type: str
    label: str
    aliases: dict[str, tuple[str, ...]]
    system_aliases: dict[str, tuple[str, ...]]
    required_fields: frozenset[str]

    def recognize(
        self,
        headers: tuple[str, ...],
        system_headers: tuple[str, ...] | None = None,
    ) -> tuple[int, dict[str, int], dict[str, dict]] | None:
        ...

    def build_plan(self, db: Session, sheet: DetectedSheet) -> dict:
        ...

    def apply_plan(
        self,
        db: Session,
        plan: dict,
        *,
        operated_by: str,
        audit_reason: str,
    ) -> dict:
        ...


_ADAPTERS: dict[str, FormAdapter] = {}


def register_adapter(adapter: FormAdapter) -> FormAdapter:
    if adapter.key in _ADAPTERS:
        raise RuntimeError(f"duplicate maintenance bulk adapter: {adapter.key}")
    if adapter.file_type not in SUPPORTED_BATCH_TYPES:
        raise RuntimeError(f"unsupported sys_import_batch namespace: {adapter.file_type}")
    _ADAPTERS[adapter.key] = adapter
    return adapter


def registered_forms() -> list[dict]:
    return [
        {
            "form_type": adapter.key,
            "label": adapter.label,
            "required_fields": sorted(adapter.required_fields),
            "accepted_headers": {
                key: list(values) for key, values in adapter.aliases.items()
            },
            "stable_source_fields": {
                key: list(values)
                for key, values in getattr(adapter, "system_aliases", {}).items()
            },
        }
        for adapter in _ADAPTERS.values()
    ]


def _header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"\((?:必填|不可修改)\)$", "", text, flags=re.I)
    return text.casefold()


def _system_header(value: Any) -> str:
    """Normalize the machine field id from the first export header row.

    Tritium exports sometimes prefix a child-table id (``D...F0000013``).
    The final segment is stable inside that document type and remains usable
    when users reorder columns.  It is deliberately kept separate from the
    Chinese-caption namespace, preventing the historic cross-document alias
    collision on fields such as ``税率(必填)``.
    """

    value = unicodedata.normalize("NFKC", str(value or "")).strip()
    return value.rsplit(".", 1)[-1].casefold()


def _text(value: Any) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def normalize_order_no(value: Any) -> str:
    value = re.sub(r"\s+", "", _text(value)).upper()
    return value[5:] if value.startswith("XSDD-") else value


def _decimal(value: Any, *, label: str, allow_zero: bool = True) -> Decimal:
    raw = _text(value).replace(",", "").replace("￥", "").replace("¥", "")
    if raw == "":
        raise BulkImportInvalid(f"{label}为空")
    try:
        parsed = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise BulkImportInvalid(f"{label}不是合法数字：{raw!r}") from exc
    if not parsed.is_finite() or parsed < 0 or (not allow_zero and parsed == 0):
        raise BulkImportInvalid(f"{label}必须是非负有限数字")
    if parsed >= MONEY_LIMIT:
        raise BulkImportInvalid(f"{label}超出系统金额上限")
    return parsed.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _optional_decimal(value: Any, *, label: str) -> Decimal | None:
    return None if _text(value) == "" else _decimal(value, label=label)


def _tax_rate(value: Any) -> Decimal | None:
    raw = _text(value).replace("％", "%")
    if not raw:
        return None
    pct = raw.endswith("%")
    if pct:
        raw = raw[:-1]
    try:
        parsed = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise BulkImportInvalid(f"税率不是合法数字：{value!r}") from exc
    if pct or parsed > 1:
        parsed /= Decimal("100")
    if parsed < 0 or parsed > 1:
        raise BulkImportInvalid("税率必须在 0–100% 之间")
    return parsed.quantize(Decimal("0.000001"))


def _date(value: Any, *, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    raise BulkImportInvalid(f"{label}不是合法日期：{raw!r}")


def _jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(k): _jsonable(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(v) for v in value]
    return value


def _canonical_hash(value: Any) -> str:
    payload = json.dumps(
        _jsonable(value), ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _token_hash(batch_id: int, token: str) -> str:
    return hashlib.sha256(f"{batch_id}:{token}".encode("utf-8")).hexdigest()


def _operation_hash(operated_by: str, operation_key: str) -> str:
    return hashlib.sha256(
        f"{operated_by}\0{operation_key.strip()}".encode("utf-8")
    ).hexdigest()


def _advisory_lock(db: Session, key: str) -> None:
    if db.bind is not None and db.bind.dialect.name == "postgresql":
        db.execute(select(func.pg_advisory_xact_lock(func.hashtextextended(key, 0))))


class HeaderAdapter:
    aliases: dict[str, tuple[str, ...]]
    system_aliases: dict[str, tuple[str, ...]] = {}
    required_fields: frozenset[str]
    required_alternatives: tuple[frozenset[str], ...] = ()

    def recognize(
        self,
        headers: tuple[str, ...],
        system_headers: tuple[str, ...] | None = None,
    ) -> tuple[int, dict[str, int], dict[str, dict]] | None:
        normalized_headers = tuple(_header(value) for value in headers)
        normalized_system = tuple(
            _system_header(value) for value in (system_headers or ())
        )
        indexes: dict[str, int] = {}
        matches: dict[str, dict] = {}
        stable_hits = 0
        for field, aliases in self.aliases.items():
            wanted = {_header(alias) for alias in aliases}
            label_candidates = [
                idx for idx, header in enumerate(normalized_headers) if header in wanted
            ]
            system_wanted = {
                _system_header(alias)
                for alias in self.system_aliases.get(field, ())
            }
            system_candidates = [
                idx for idx, header in enumerate(normalized_system)
                if header and header in system_wanted
            ]

            # A two-row business-system export is accepted only when the
            # stable field id and its human caption agree at the same column.
            # This resolves duplicate captions without relying on column
            # letters and fails closed if the upstream schema repurposes an id.
            if len(system_candidates) == 1:
                idx = system_candidates[0]
                if idx < len(normalized_headers) and normalized_headers[idx] in wanted:
                    indexes[field] = idx
                    stable_hits += 1
                    matches[field] = {
                        "source_column": headers[idx],
                        "system_field": (system_headers or ())[idx],
                        "confidence": "exact",
                    }
                    continue
            # A duplicated caption without a stable id is ambiguous, never
            # silently resolved to the first occurrence.
            if len(label_candidates) == 1:
                idx = label_candidates[0]
                indexes[field] = idx
                matches[field] = {
                    "source_column": headers[idx],
                    "system_field": None,
                    "confidence": "alias",
                }

        alternatives = self.required_alternatives or (self.required_fields,)
        if not any(required.issubset(indexes) for required in alternatives):
            return None
        return stable_hits * 100 + len(indexes), indexes, matches


def _detect(data: bytes) -> tuple[FormAdapter, DetectedSheet]:
    try:
        # Some Tritium exports have malformed worksheet dimensions/cell order:
        # openpyxl's streaming reader then exposes only column A although the
        # workbook visibly contains fields through EX.  Upload size and row
        # limits are enforced before/while parsing, so use the normal reader to
        # obtain the real sparse cell map deterministically.
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001
        raise BulkImportInvalid(f"无法读取 .xlsx：{type(exc).__name__}") from exc

    candidates: list[
        tuple[
            int,
            str,
            int,
            FormAdapter,
            tuple[str, ...],
            tuple[str, ...],
            dict[str, int],
            dict[str, dict],
        ]
    ] = []
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible" or sheet.title.lower().startswith("hidden_"):
            continue
        header_band = [
            tuple(_text(value) for value in values)
            for values in sheet.iter_rows(
                min_row=1,
                max_row=min(HEADER_SCAN_ROWS, sheet.max_row),
                values_only=True,
            )
        ]
        for row_no, headers in enumerate(header_band, start=1):
            system_headers = header_band[row_no - 2] if row_no > 1 else ()
            for adapter in _ADAPTERS.values():
                match = adapter.recognize(headers, system_headers or None)
                if match is not None:
                    score, indexes, matches = match
                    candidates.append(
                        (
                            score,
                            sheet.title,
                            row_no,
                            adapter,
                            headers,
                            system_headers,
                            indexes,
                            matches,
                        )
                    )
    if not candidates:
        raise BulkImportInvalid(
            "无法识别表单：未找到销售合同额或收款单所需表头",
            issues=[{"code": "unknown_form", "registered_forms": list(_ADAPTERS)}],
        )
    candidates.sort(key=lambda item: (-item[0], item[1], item[2], item[3].key))
    best_score = candidates[0][0]
    best = [item for item in candidates if item[0] == best_score]
    identities = {(item[1], item[2], item[3].key) for item in best}
    if len(identities) != 1:
        raise BulkImportInvalid(
            "工作簿中存在多个同等匹配的业务表，无法自动选择",
            issues=[
                {"sheet": s, "header_row": r, "form_type": a.key}
                for _score, s, r, a, _h, _sh, _i, _m in best[:20]
            ],
        )
    (
        _score,
        sheet_name,
        header_row,
        adapter,
        headers,
        system_headers,
        indexes,
        matches,
    ) = best[0]
    sheet = workbook[sheet_name]
    rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_no, values in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        values = tuple(values)
        if all(_text(value) == "" for value in values):
            continue
        rows.append((row_no, values))
        if len(rows) > MAX_PREVIEW_ROWS:
            raise BulkImportInvalid(f"数据行超过安全上限 {MAX_PREVIEW_ROWS}")
    if not rows:
        raise BulkImportInvalid("识别到表头，但没有数据行")
    return adapter, DetectedSheet(
        name=sheet_name,
        header_row=header_row,
        header_rows=((header_row - 1, header_row) if system_headers else (header_row,)),
        headers=headers,
        system_headers=system_headers,
        field_indexes=indexes,
        field_matches=matches,
        rows=tuple(rows),
    )


def _value(sheet: DetectedSheet, values: tuple[Any, ...], field: str) -> Any:
    idx = sheet.field_indexes.get(field)
    return values[idx] if idx is not None and idx < len(values) else None


def _contract_maps(db: Session) -> tuple[
    dict[str, list[MaintenanceProjectContract]],
    dict[str, MaintenanceProjectContract],
]:
    today = business_today()
    rows = list(db.scalars(
        select(MaintenanceProjectContract)
        .join(MaintenanceProject, MaintenanceProject.project_id == MaintenanceProjectContract.project_id)
        .where(
            MaintenanceProject.is_active.is_(True),
            MaintenanceProjectContract.effective_from <= today,
            or_(
                MaintenanceProjectContract.effective_to.is_(None),
                MaintenanceProjectContract.effective_to > today,
            ),
        )
        .order_by(
            MaintenanceProjectContract.contract_no,
            MaintenanceProjectContract.project_id,
            MaintenanceProjectContract.project_contract_id,
        )
    ))
    all_current: dict[str, list[MaintenanceProjectContract]] = defaultdict(list)
    for row in rows:
        all_current[normalize_order_no(row.contract_no)].append(row)
    safe = {
        key: matches[0]
        for key, matches in all_current.items()
        if len(matches) == 1
        and matches[0].status_mapping_state == "mapped"
        and matches[0].included_in_total
    }
    return dict(all_current), safe


def _row_issue(row_no: int, code: str, message: str, *, severity: str = "error") -> dict:
    return {"row_no": row_no, "code": code, "message": message, "severity": severity}


def _all_contracts_by_order(
    db: Session,
    order_variants: set[str],
) -> dict[str, list[MaintenanceProjectContract]]:
    if not order_variants:
        return {}
    rows = list(
        db.scalars(
            select(MaintenanceProjectContract)
            .where(MaintenanceProjectContract.contract_no.in_(sorted(order_variants)))
            .order_by(
                MaintenanceProjectContract.contract_no,
                MaintenanceProjectContract.effective_from,
                MaintenanceProjectContract.project_contract_id,
            )
        )
    )
    result: dict[str, list[MaintenanceProjectContract]] = defaultdict(list)
    for row in rows:
        result[normalize_order_no(row.contract_no)].append(row)
    return dict(result)


def _assignment_evidence(
    db: Session,
    order_variants: set[str],
) -> tuple[dict[str, list[dict]], dict[str, list[dict]]]:
    """Return current-safe and historical XSDD ownership evidence.

    Current ownership is intentionally stricter than mere assignment presence:
    the WBDD, assignment and project must all still be active.  Historical rows
    remain visible to the pre-delivery auto-create guard so an old ownership
    decision can never be erased by creating a fresh project with the same
    sales order.
    """

    if not order_variants:
        return {}, {}
    rows = db.execute(
        select(
            FMaintenanceOrder.linked_sales_order_no,
            FMaintenanceOrder.raw_order_id,
            FMaintenanceOrder.data_status,
            MaintenanceSourceOrderAssignment.assignment_id,
            MaintenanceSourceOrderAssignment.project_id,
            MaintenanceSourceOrderAssignment.is_active,
            MaintenanceSourceOrderAssignment.version,
            MaintenanceProject.is_active,
            MaintenanceProject.version,
            MaintenanceProject.display_name,
        )
        .select_from(FMaintenanceOrder)
        .join(
            MaintenanceSourceOrderAssignment,
            MaintenanceSourceOrderAssignment.source_order_id
            == FMaintenanceOrder.raw_order_id,
        )
        .join(
            MaintenanceProject,
            MaintenanceProject.project_id
            == MaintenanceSourceOrderAssignment.project_id,
        )
        .where(FMaintenanceOrder.linked_sales_order_no.in_(sorted(order_variants)))
        .where(
            FMaintenanceOrder.data_status == config.ACTIVE_STATUS,
            ~select(MaintenanceDemandTombstone.source_order_id).where(
                MaintenanceDemandTombstone.source_order_id
                == FMaintenanceOrder.raw_order_id,
                MaintenanceDemandTombstone.restored_at.is_(None),
            ).exists(),
        )
        .order_by(
            FMaintenanceOrder.linked_sales_order_no,
            MaintenanceSourceOrderAssignment.assignment_id,
        )
    ).all()
    current: dict[str, list[dict]] = defaultdict(list)
    historical: dict[str, list[dict]] = defaultdict(list)
    for (
        order_no,
        source_order_id,
        source_status,
        assignment_id,
        project_id,
        assignment_active,
        assignment_version,
        project_active,
        project_version,
        project_name,
    ) in rows:
        norm = normalize_order_no(order_no)
        item = {
            "source_order_id": source_order_id,
            "source_status": source_status,
            "assignment_id": assignment_id,
            "assignment_version": assignment_version,
            "project_id": project_id,
            "project_version": project_version,
            "project_name": project_name,
            "assignment_active": bool(assignment_active),
            "project_active": bool(project_active),
        }
        historical[norm].append(item)
        if (
            assignment_active
            and project_active
            and source_status == config.ACTIVE_STATUS
        ):
            current[norm].append(item)
    return dict(current), dict(historical)


def _assignment_fingerprint(rows: list[dict]) -> list[dict]:
    return sorted(
        [
            {
                "source_order_id": row["source_order_id"],
                "assignment_id": row["assignment_id"],
                "assignment_version": row["assignment_version"],
                "project_id": row["project_id"],
                "project_version": row["project_version"],
            }
            for row in rows
        ],
        key=lambda row: (row["project_id"], row["assignment_id"]),
    )


def _sales_fingerprint(rows: list[FSalesOrder]) -> list[dict]:
    return sorted(
        [
            {
                "id": row.id,
                "raw_order_id": row.raw_order_id,
                "order_no": row.order_no,
                "data_status": row.data_status,
                "amount_ex_tax": _jsonable(row.amount_ex_tax),
                "tax_rate": _jsonable(row.tax_rate),
            }
            for row in {row.id: row for row in rows}.values()
        ],
        key=lambda row: row["id"],
    )


def _is_yes(value: Any) -> bool:
    return _text(value).casefold() in {"是", "含税", "true", "yes", "1", "y"}


def _is_maintenance_business_type(value: Any) -> bool:
    business_type = _text(value)
    return any(word in business_type for word in ("维保", "运维", "维修"))


def _maintenance_business_flag(
    sheet: DetectedSheet,
    values: tuple[Any, ...],
) -> bool | None:
    if "maintenance_business" not in sheet.field_indexes:
        return None
    raw = _text(_value(sheet, values, "maintenance_business"))
    if not raw:
        return None
    if _is_yes(raw):
        return True
    if raw.casefold() in {"否", "false", "no", "0", "n"}:
        return False
    raise BulkImportInvalid(f"维保业务标记无法识别：{raw}")


def _is_explicit_maintenance_row(
    sheet: DetectedSheet,
    values: tuple[Any, ...],
) -> bool:
    """Only explicit source facts may trigger automatic project writes."""

    flag = _maintenance_business_flag(sheet, values)
    business_type_raw = _text(_value(sheet, values, "business_type"))
    type_is_maintenance = _is_maintenance_business_type(business_type_raw)
    # Production exports use these as independent classifiers.  For example,
    # 单次维修 can legitimately carry 维保业务=否, while an explicit 是 can
    # accompany a generic sales type.  Either positive signal is sufficient.
    return flag is True or type_is_maintenance


def _ordinary_maintenance_rows(
    sheet: DetectedSheet,
) -> tuple[tuple[int, tuple[Any, ...]], ...]:
    """Explicit, active rows eligible for the ordinary sales auto path."""

    eligible: list[tuple[int, tuple[Any, ...]]] = []
    for row_no, values in sheet.rows:
        if not _is_explicit_maintenance_row(sheet, values):
            continue
        if "data_status" not in sheet.field_indexes:
            raise BulkImportInvalid("维保销售订单自动建项要求源表包含 Status 数据状态列")
        source_status = _text(_value(sheet, values, "data_status"))
        if not source_status:
            raise BulkImportInvalid("维保销售订单自动建项要求数据状态明确为已生效")
        if source_status != config.ACTIVE_STATUS:
            continue
        eligible.append((row_no, values))
    return tuple(eligible)


def _explicit_maintenance_period(
    sheet: DetectedSheet,
    values: tuple[Any, ...],
) -> tuple[date | None, date | None]:
    raw_from = _value(sheet, values, "period_from")
    raw_to = _value(sheet, values, "period_to")
    period_from = _date(raw_from, label="维保起始日期") if _text(raw_from) else None
    period_to = _date(raw_to, label="维保终止日期") if _text(raw_to) else None
    if period_from is not None and period_to is not None and period_to < period_from:
        raise BulkImportInvalid("维保终止日期不能早于起始日期")
    return period_from, period_to


def _maintenance_project_metadata(
    sheet: DetectedSheet,
    values: tuple[Any, ...],
    *,
    row_no: int,
    norm: str,
) -> dict:
    raw_name = _text(_value(sheet, values, "project_name"))
    if not raw_name:
        raise BulkImportInvalid("维保销售订单自动建项必须提供项目名称")
    if not _is_explicit_maintenance_row(sheet, values):
        raise BulkImportInvalid("自动建项只接受销售订单中明确的维保业务事实")
    # Preserve the exact sales-order name.  A pre-delivery name and a later
    # formal name for the same XSDD are peer display facts, not a hierarchy.
    display_name = raw_name.strip()

    direct_from, direct_to = _explicit_maintenance_period(sheet, values)
    period_from = direct_from
    period_to = direct_to
    if period_from is not None and period_to is not None and period_to < period_from:
        raise BulkImportInvalid("维保终止日期不能早于起始日期")
    business_type = _text(_value(sheet, values, "business_type"))
    return {
        "row_no": row_no,
        "project_code": f"XSDD-{norm}"[:64],
        "display_name": display_name[:256],
        "period_from": period_from.isoformat() if period_from else None,
        "period_to": period_to.isoformat() if period_to else None,
        "business_type": business_type or None,
    }


def _apply_sales_project_period(
    db: Session,
    *,
    item: dict,
    project_id: str,
    audit_reason: str,
    operated_by: str,
) -> None:
    raw_from = item.get("source_period_from")
    raw_to = item.get("source_period_to")
    from_present = bool(item.get("source_period_from_present"))
    to_present = bool(item.get("source_period_to_present"))
    if not from_present and not to_present:
        return
    project = db.get(MaintenanceProject, project_id)
    if project is None:
        raise BulkImportConflict("项目在同步维保期限时消失")
    desired_from = (
        date.fromisoformat(raw_from)
        if raw_from is not None
        else (None if from_present else project.period_from)
    )
    desired_to = (
        date.fromisoformat(raw_to)
        if raw_to is not None
        else (None if to_present else project.period_to)
    )
    if project.period_from == desired_from and project.period_to == desired_to:
        return
    if project.version != item.get("expected_project_version"):
        raise BulkImportConflict("预览后的项目维保期限已变化")
    updates = {"period_from": desired_from, "period_to": desired_to}
    try:
        updated = catalog.update_project(
            db,
            project_id=project_id,
            version=project.version,
            updates=updates,
            reason=f"销售订单权威维保期限同步：{audit_reason}",
            operated_by=operated_by,
        )
    except (catalog.MaintenanceProjectCatalogError, catalog.MaintenanceProjectCatalogConflict) as exc:
        raise BulkImportConflict(str(exc)) from exc
    if updated is None:
        raise BulkImportConflict("项目在同步维保期限时消失")


class SalesContractAmountAdapter(HeaderAdapter):
    key = "sales_contract_amount"
    file_type = "maint_contract"
    label = "销售订单合同含税额"
    aliases = {
        "order_no": ("订单编号(必填)", "订单编号", "销售订单", "销售单号", "合同编号"),
        "raw_order_id": ("数据ID(不可修改)", "订单数据ID", "销售订单数据ID"),
        "amount_inc_tax": ("含税金额", "合同总额(含税)", "合同总额（含税）"),
        "order_amount": ("订单金额", "合同金额", "合同总额"),
        "tax_flag": ("是否含税(必填)", "是否含税", "含税标记"),
        "tax_rate": ("税率(必填)", "税率"),
        "tax_amount": ("税金", "税额"),
        "amount_ex_tax": ("不含税金额", "未税金额", "合同金额(未税)"),
        "data_status": ("数据状态", "订单状态"),
        "maintenance_business": ("维保业务", "是否维保业务"),
        "business_type": ("业务类型#", "业务类型"),
        "project_name": ("项目名称(必填)", "项目名称"),
        "period_from": ("维保起始日期(必填)", "维保起始日期", "维保起始时间"),
        "period_to": ("维保终止日期(必填)", "维保终止日期", "维保终止时间"),
    }
    # First-row ids from the sales-order export.  Matching is by id+caption at
    # the same column; these are not physical Excel positions.
    system_aliases = {
        "order_no": ("SeqNo",),
        "raw_order_id": ("ObjectId",),
        "order_amount": ("F0000021",),
        "tax_flag": ("F0000053",),
        "tax_rate": ("F0000054",),
        "tax_amount": ("F0000055",),
        "amount_ex_tax": ("F0000056",),
        "data_status": ("Status",),
        "maintenance_business": ("F0000118",),
        # The source contains both F0000060/业务类型 and
        # F0000059/业务类型#; the latter is the project-facing field.
        "business_type": ("F0000059",),
        "project_name": ("F0000119",),
        "period_from": ("F0000131",),
        "period_to": ("F0000132",),
    }
    required_fields = frozenset(
        {
            "order_no",
            "order_amount",
            "tax_flag",
            "tax_rate",
            "tax_amount",
            "amount_ex_tax",
        }
    )
    required_alternatives = (
        required_fields,
        frozenset({"order_no", "amount_inc_tax", "tax_rate", "amount_ex_tax"}),
        frozenset({"order_no", "amount_ex_tax", "tax_amount"}),
        frozenset({"order_no", "amount_inc_tax", "amount_ex_tax"}),
        frozenset({"order_no", "order_amount", "tax_flag", "tax_rate"}),
        frozenset(
            {
                "order_no",
                "order_amount",
                "tax_flag",
                "amount_ex_tax",
                "tax_amount",
            }
        ),
    )

    @staticmethod
    def _amounts(
        sheet: DetectedSheet,
        values: tuple[Any, ...],
    ) -> tuple[Decimal, Decimal, Decimal]:
        explicit_inc = _optional_decimal(
            _value(sheet, values, "amount_inc_tax"), label="含税金额"
        )
        order_amount = _optional_decimal(
            _value(sheet, values, "order_amount"), label="订单金额"
        )
        amount_ex = _optional_decimal(
            _value(sheet, values, "amount_ex_tax"), label="不含税金额"
        )
        tax_amount = _optional_decimal(
            _value(sheet, values, "tax_amount"), label="税金"
        )
        rate = _tax_rate(_value(sheet, values, "tax_rate"))
        flag = _text(_value(sheet, values, "tax_flag"))

        # First identify an authoritative gross value.  DK is gross only when
        # DL explicitly says so; DO+DN is an independent gross proof.
        inc: Decimal | None = None
        if explicit_inc is not None:
            inc = explicit_inc
        elif order_amount is not None and flag in {"含税", "是", "含税价"}:
            # Official sales export: DK/订单金额 is the authoritative gross
            # amount when DL says 含税.  DO is the ex-tax fact and DN is tax.
            inc = order_amount
        if order_amount is not None and flag in {"不含税", "否", "未税"}:
            amount_ex = amount_ex or order_amount

        # DO+DN can prove both gross and, when DO>0, the rate even if DM is
        # absent.  That is evidence-derived 0% when DN is actually zero, not a
        # NULL-to-zero fallback.  If a gross value was independently supplied,
        # it must agree with this sum.
        if amount_ex is not None and tax_amount is not None:
            sum_inc = (amount_ex + tax_amount).quantize(Decimal("0.01"))
            if inc is not None and abs(inc - sum_inc) > Decimal("0.02"):
                raise BulkImportInvalid("订单金额与不含税金额+税金不一致")
            inc = inc or sum_inc
            if rate is None:
                if amount_ex <= 0:
                    raise BulkImportInvalid("不含税金额为 0 时无法从税金反推税率")
                derived_rate = tax_amount / amount_ex
                if derived_rate < 0 or derived_rate > 1:
                    raise BulkImportInvalid("由税金反推的税率不在 0–100% 之间")
                rate = derived_rate.quantize(Decimal("0.000001"))

        # A known gross plus ex-tax amount can also prove the omitted tax and
        # rate.  A lone ex-tax/order amount still requires explicit rate.
        if rate is None and inc is not None and amount_ex is not None:
            derived_tax = inc - amount_ex
            if amount_ex <= 0 or derived_tax < 0:
                raise BulkImportInvalid("税率为空，且含税/未税金额无法反推税率")
            derived_rate = derived_tax / amount_ex
            if derived_rate < 0 or derived_rate > 1:
                raise BulkImportInvalid("由含税/未税金额反推的税率不在 0–100% 之间")
            rate = derived_rate.quantize(Decimal("0.000001"))

        if inc is None and amount_ex is not None:
            if rate is None:
                raise BulkImportInvalid("税率为空，不能从单一未税金额推导含税合同额")
            inc = (amount_ex * (Decimal("1") + rate)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP
            )
        if inc is None:
            raise BulkImportInvalid("无法从订单金额/含税标记/税金推导含税合同额")

        if amount_ex is None:
            if tax_amount is not None:
                amount_ex = (inc - tax_amount).quantize(Decimal("0.01"))
            elif rate is not None:
                amount_ex = (inc / (Decimal("1") + rate)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP
                )
            else:
                raise BulkImportInvalid("税率为空，且没有税金/未税金额可供交叉校验")
        if amount_ex < 0:
            raise BulkImportInvalid("不含税金额不能为负")
        if rate is None:
            raise BulkImportInvalid("税率缺失且无法从完整金额证据反推")
        expected_tax = (amount_ex * rate).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if tax_amount is not None and abs(expected_tax - tax_amount) > Decimal("0.02"):
            raise BulkImportInvalid("税金与不含税金额×税率不一致")
        if tax_amount is not None and abs((amount_ex + tax_amount) - inc) > Decimal("0.02"):
            raise BulkImportInvalid("订单金额与不含税金额+税金不一致")
        return inc, amount_ex, rate

    def build_plan(self, db: Session, sheet: DetectedSheet) -> dict:
        all_contracts, safe_contracts = _contract_maps(db)
        parsed: list[dict] = []
        source_rows: list[dict] = []
        hard_issues: list[dict] = []
        seen: dict[str, tuple[Any, ...]] = {}

        raw_ids: set[str] = set()
        order_variants: set[str] = set()
        for row_no, values in sheet.rows:
            raw_order_id = _text(_value(sheet, values, "raw_order_id"))
            order_no_raw = _text(_value(sheet, values, "order_no"))
            norm = maintenance_project_identity.normalize_xsdd(order_no_raw)
            if raw_order_id:
                raw_ids.add(raw_order_id)
            if norm:
                order_variants.update({order_no_raw, norm, f"XSDD-{norm}"})

        historical_contracts = _all_contracts_by_order(db, order_variants)
        active_assignments, assignment_history = _assignment_evidence(
            db, order_variants
        )

        sales_rows = list(db.scalars(select(FSalesOrder).where(or_(
            FSalesOrder.raw_order_id.in_(sorted(raw_ids or {""})),
            FSalesOrder.order_no.in_(sorted(order_variants or {""})),
        ))))
        sales_by_raw = {row.raw_order_id: row for row in sales_rows}
        sales_by_order: dict[str, list[FSalesOrder]] = defaultdict(list)
        for row in sales_rows:
            sales_by_order[
                maintenance_project_identity.normalize_xsdd(row.order_no)
            ].append(row)

        all_projects = list(
            db.scalars(
                select(MaintenanceProject).order_by(MaintenanceProject.project_id)
            )
        )
        projects_by_id = {project.project_id: project for project in all_projects}
        projects_by_code: dict[str, list[MaintenanceProject]] = defaultdict(list)
        for project in all_projects:
            projects_by_code[project.project_code.casefold()].append(project)

        for row_no, values in sheet.rows:
            order_no_raw = _text(_value(sheet, values, "order_no"))
            norm = maintenance_project_identity.normalize_xsdd(order_no_raw)
            raw_order_id = _text(_value(sheet, values, "raw_order_id"))
            base = {
                "row_no": row_no,
                "business_key": order_no_raw,
                "normalized_order_no": norm,
                "issues": [],
            }
            if not norm:
                issue = _row_issue(
                    row_no,
                    "invalid_xsdd_order_no",
                    "维保合同事实要求有效 XSDD-YYYYMMDD-NNN/NNNN 销售订单号",
                )
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue

            source_status = _text(_value(sheet, values, "data_status"))
            if "data_status" not in sheet.field_indexes or not source_status:
                issue = _row_issue(
                    row_no,
                    "missing_source_status",
                    "维保合同事实要求源表包含明确的销售订单数据状态",
                )
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue
            if source_status and source_status != config.ACTIVE_STATUS:
                base.update(
                    action="skip",
                    issues=[
                        _row_issue(
                            row_no,
                            "inactive_sales_order",
                            f"销售订单状态为“{source_status}”，不覆盖项目合同额",
                            severity="warning",
                        )
                    ],
                )
                source_rows.append(base)
                continue
            try:
                inc, amount_ex, rate = self._amounts(sheet, values)
            except BulkImportInvalid as exc:
                issue = _row_issue(row_no, "invalid_amount", str(exc))
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue

            project_name_raw = _text(_value(sheet, values, "project_name"))
            period_from_raw = _text(_value(sheet, values, "period_from"))
            period_to_raw = _text(_value(sheet, values, "period_to"))
            source_period_from: date | None = None
            source_period_to: date | None = None
            try:
                explicit_maintenance = _is_explicit_maintenance_row(sheet, values)
            except BulkImportInvalid as exc:
                issue = _row_issue(
                    row_no, "maintenance_source_conflict", str(exc)
                )
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue
            if explicit_maintenance:
                if not project_name_raw:
                    issue = _row_issue(
                        row_no,
                        "invalid_maintenance_project",
                        "维保销售订单自动建项必须提供项目名称",
                    )
                    base.update(action="error", issues=[issue])
                    source_rows.append(base)
                    hard_issues.append(issue)
                    continue
                try:
                    source_period_from, source_period_to = _explicit_maintenance_period(
                        sheet, values
                    )
                except BulkImportInvalid as exc:
                    issue = _row_issue(
                        row_no, "invalid_maintenance_period", str(exc)
                    )
                    base.update(action="error", issues=[issue])
                    source_rows.append(base)
                    hard_issues.append(issue)
                    continue
            duplicate = seen.get(norm)
            signature = (
                inc,
                amount_ex,
                rate,
                source_status,
                project_name_raw,
                period_from_raw,
                period_to_raw,
            )
            if duplicate is not None:
                if duplicate[:-1] != signature:
                    issue = _row_issue(
                        row_no,
                        "duplicate_conflict",
                        f"同一销售订单与第 {duplicate[-1]} 行的金额或项目元数据不一致",
                    )
                    base.update(action="error", issues=[issue])
                    hard_issues.append(issue)
                else:
                    base.update(
                        action="duplicate",
                        issues=[_row_issue(
                            row_no,
                            "duplicate_same",
                            f"与第 {duplicate[-1]} 行相同，应用时只处理一次",
                            severity="warning",
                        )],
                    )
                source_rows.append(base)
                continue
            seen[norm] = (*signature, row_no)

            sales = sales_by_raw.get(raw_order_id) if raw_order_id else None
            if sales is not None and normalize_order_no(sales.order_no) != norm:
                issue = _row_issue(row_no, "sales_identity_mismatch", "数据ID与销售订单号不一致")
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue
            if sales is None:
                matches = sales_by_order.get(norm, [])
                if len(matches) == 1:
                    sales = matches[0]
                else:
                    base["issues"].append(_row_issue(
                        row_no,
                        "sales_fact_ambiguous" if matches else "sales_fact_missing",
                        (
                            "系统销售事实存在多个候选；合同仍可按源文件处理，"
                            "但本次不反写 f_sales_order"
                            if matches
                            else (
                                "系统尚无该销售事实；合同仍可按源文件处理，"
                                "且不会制造缺少明细行的 f_sales_order 头"
                            )
                        ),
                        severity="warning",
                    ))
            sync_sales = bool(
                sales is not None and sales.data_status == config.ACTIVE_STATUS
            )
            if sales is not None and not sync_sales:
                base["issues"].append(_row_issue(
                    row_no,
                    "sales_fact_inactive",
                    "系统销售事实不是已生效状态；合同仍按源文件处理，但不反写该事实",
                    severity="warning",
                ))

            sales_amount_before = sales.amount_ex_tax if sales is not None else None
            sales_rate_before = sales.tax_rate if sales is not None else None
            contract_no = sales.order_no if sales is not None else order_no_raw
            contract_id = (
                raw_order_id
                or (sales.raw_order_id if sales is not None else "")
                or f"XSDD-{norm}"
            )
            sales_candidates = list(sales_by_order.get(norm, []))
            if raw_order_id and sales_by_raw.get(raw_order_id) is not None:
                sales_candidates.append(sales_by_raw[raw_order_id])

            operation = {
                "row_no": row_no,
                "normalized_order_no": norm,
                "source_project_name": project_name_raw[:256] or None,
                "source_period_from": (
                    source_period_from.isoformat() if source_period_from else None
                ),
                "source_period_to": (
                    source_period_to.isoformat() if source_period_to else None
                ),
                "source_period_from_present": bool(
                    explicit_maintenance and "period_from" in sheet.field_indexes
                ),
                "source_period_to_present": bool(
                    explicit_maintenance and "period_to" in sheet.field_indexes
                ),
                "sales_order_id": sales.id if sync_sales else None,
                "sales_raw_order_id": sales.raw_order_id if sync_sales else None,
                "sales_match_state": (
                    "unique_active" if sync_sales else (
                        "unique_inactive" if sales is not None else (
                            "ambiguous" if sales_by_order.get(norm) else "missing"
                        )
                    )
                ),
                "expected_sales_candidates": _sales_fingerprint(
                    sales_candidates
                ),
                "contract_id": contract_id,
                "sales_order_no": contract_no,
                "contract_status": source_status or config.ACTIVE_STATUS,
                "expected_sales_amount_ex_tax": _jsonable(sales_amount_before),
                "expected_sales_tax_rate": _jsonable(sales_rate_before),
                "expected_sales_data_status": (
                    sales.data_status if sales is not None else None
                ),
                "new_contract_amount_inc_tax": _jsonable(inc),
                "new_sales_amount_ex_tax": _jsonable(amount_ex),
                "new_sales_tax_rate": _jsonable(rate),
                "contract_effective_from": min(
                    (
                        sales.order_date
                        if sales is not None and sales.order_date is not None
                        else business_today()
                    ),
                    business_today(),
                ).isoformat(),
            }
            current_relations = all_contracts.get(norm, [])
            contract = safe_contracts.get(norm)
            if contract is not None:
                contract_project = projects_by_id.get(contract.project_id)
                before = {
                    "contract_amount_inc_tax": contract.amount_inc_tax,
                    "contract_version": contract.version,
                    "sales_amount_ex_tax": sales_amount_before,
                    "sales_tax_rate": sales_rate_before,
                }
                after = {
                    "contract_amount_inc_tax": inc,
                    "sales_amount_ex_tax": amount_ex,
                    "sales_tax_rate": rate,
                }
                action = "noop" if (
                    contract.amount_inc_tax == inc
                    and (
                        not sync_sales
                        or (
                            sales_amount_before == amount_ex
                            and sales_rate_before == rate
                        )
                    )
                ) else "update_contract"
                operation.update(
                    action=action,
                    project_id=contract.project_id,
                    project_contract_id=contract.project_contract_id,
                    expected_contract_version=contract.version,
                    expected_contract_amount_inc_tax=_jsonable(
                        contract.amount_inc_tax
                    ),
                    # Production FK guarantees the project exists.  ``None``
                    # keeps planner-only fakes usable while apply still
                    # resolves/locks the real project before any write.
                    expected_project_version=(
                        contract_project.version if contract_project else None
                    ),
                )
            elif current_relations:
                issue = _row_issue(
                    row_no,
                    "contract_ambiguous",
                    "销售订单的当前项目合同关系共享、未映射或未计入总额",
                )
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue
            elif historical_contracts.get(norm):
                issue = _row_issue(
                    row_no,
                    "historical_contract_conflict",
                    "销售订单存在历史合同关系，不能自动新建当前关系",
                )
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue
            else:
                active = active_assignments.get(norm, [])
                active_project_ids = {row["project_id"] for row in active}
                historical_project_ids = {
                    row["project_id"] for row in assignment_history.get(norm, [])
                }
                try:
                    sales_owner_id = (
                        maintenance_project_identity.resolve_sales_xsdd_project(
                            db, norm
                        )
                    )
                except maintenance_project_identity.XsddProjectConflict as exc:
                    issue = _row_issue(
                        row_no,
                        "xsdd_owner_conflict",
                        str(exc),
                    )
                    base.update(action="error", issues=[issue])
                    source_rows.append(base)
                    hard_issues.append(issue)
                    continue
                if sales_owner_id is not None:
                    project = projects_by_id.get(sales_owner_id)
                    if project is None or not project.is_active:
                        issue = _row_issue(
                            row_no,
                            "xsdd_owner_project_missing",
                            "XSDD owner 项目不存在或已归档",
                        )
                        base.update(action="error", issues=[issue])
                        source_rows.append(base)
                        hard_issues.append(issue)
                        continue
                    operation.update(
                        action="create_contract",
                        project_id=sales_owner_id,
                        project_contract_id=None,
                        expected_project_version=project.version,
                    )
                    before = {
                        "contract_amount_inc_tax": None,
                        "sales_amount_ex_tax": sales_amount_before,
                        "sales_tax_rate": sales_rate_before,
                    }
                    after = {
                        "contract_amount_inc_tax": inc,
                        "sales_amount_ex_tax": amount_ex,
                        "sales_tax_rate": rate,
                    }
                    action = "create_contract"
                else:
                    # Never infer a project from customer/name text.  Only an
                    # explicit maintenance source fact may create one.
                    if not explicit_maintenance:
                        base.update(
                            action="unmatched",
                            issues=[
                                _row_issue(
                                    row_no,
                                    "project_not_found",
                                    "未命中已有维保项目，且源行未明确标记维保业务，本批跳过",
                                    severity="warning",
                                )
                            ],
                        )
                        source_rows.append(base)
                        continue
                    if active_project_ids or historical_project_ids:
                        issue = _row_issue(
                            row_no,
                            "historical_assignment_conflict",
                            "XSDD 只有 WBDD 归属、没有销售 owner，拒绝由 WBDD 反向建项",
                        )
                        base.update(action="error", issues=[issue])
                        source_rows.append(base)
                        hard_issues.append(issue)
                        continue
                    try:
                        metadata = _maintenance_project_metadata(
                            sheet, values, row_no=row_no, norm=norm
                        )
                    except BulkImportInvalid as exc:
                        issue = _row_issue(
                            row_no, "invalid_maintenance_project", str(exc)
                        )
                        base.update(action="error", issues=[issue])
                        source_rows.append(base)
                        hard_issues.append(issue)
                        continue
                    existing_codes = projects_by_code.get(
                        metadata["project_code"].casefold(), []
                    )
                    if existing_codes:
                        issue = _row_issue(
                            row_no,
                            "maintenance_project_identity_conflict",
                            "XSDD 稳定项目编号已存在，拒绝自动创建",
                        )
                        base.update(action="error", issues=[issue])
                        source_rows.append(base)
                        hard_issues.append(issue)
                        continue
                    operation.update(
                        action="create_project",
                        project_id=None,
                        project_contract_id=None,
                        new_project=metadata,
                    )
                    before = {
                        "project": None,
                        "contract_amount_inc_tax": None,
                        "sales_amount_ex_tax": sales_amount_before,
                        "sales_tax_rate": sales_rate_before,
                    }
                    after = {
                        "project": metadata,
                        "contract_amount_inc_tax": inc,
                        "sales_amount_ex_tax": amount_ex,
                        "sales_tax_rate": rate,
                    }
                    action = "create_project"

            base.update(
                action=action,
                project_id=operation.get("project_id"),
                project_contract_id=operation.get("project_contract_id"),
                sales_order_id=operation.get("sales_order_id"),
                before=_jsonable(before),
                after=_jsonable(after),
            )
            source_rows.append(base)
            parsed.append(_jsonable(operation))

        # Fail closed per sales order, not per physical row.  A conflicting
        # duplicate or any other invalid row means that another row for the
        # same order cannot be treated as a trustworthy substitute.
        operation_norms = {
            str(operation.get("normalized_order_no") or "")
            for operation in parsed
        }
        blocked_norms = {
            str(row.get("normalized_order_no") or "")
            for row in source_rows
            if row.get("normalized_order_no")
            and (
                row.get("action") == "error"
                or (
                    row.get("action") == "skip"
                    and row.get("normalized_order_no") in operation_norms
                    and any(
                        issue.get("code") == "inactive_sales_order"
                        for issue in row.get("issues") or []
                    )
                )
            )
        }
        for norm in sorted(blocked_norms):
            affected_rows = [
                int(row["row_no"])
                for row in source_rows
                if row.get("normalized_order_no") == norm
            ]
            issue = _row_issue(
                min(affected_rows),
                "order_level_fail_closed",
                f"销售订单 {norm} 存在冲突或无效行，该订单全部操作已阻断",
            )
            hard_issues.append(issue)
            for row in source_rows:
                if row.get("normalized_order_no") != norm:
                    continue
                if not any(
                    item.get("code") == "order_level_fail_closed"
                    for item in row.get("issues") or []
                ):
                    row.setdefault("issues", []).append(issue)
                row["action"] = "error"
            for operation in parsed:
                if operation.get("normalized_order_no") != norm:
                    continue
                operation["blocked_action"] = operation.get("action")
                operation["action"] = "blocked"
                operation.setdefault("issues", []).append(issue)

        return _finish_plan(
            sheet,
            source_rows,
            parsed,
            hard_issues,
            extra={
                "amount_basis": "inc_tax",
                "source_sync": "f_sales_order+maintenance_project_contract",
                "missing_project_policy": (
                    "sales_xsdd_owner=create_contract;"
                    "explicit_maintenance=create_project;otherwise_skip"
                ),
            },
        )

    def apply_plan(
        self,
        db: Session,
        plan: dict,
        *,
        operated_by: str,
        audit_reason: str,
    ) -> dict:
        # Dedicated bulk apply and ordinary upload share the same global lock
        # order: DATA_CHANGE -> XSDD identities -> per-order bulk locks.
        db.execute(select(func.pg_advisory_xact_lock(
            config.DATA_CHANGE_ADVISORY_LOCK_KEY
        )))
        maintenance_project_identity.lock_xsdd_identities(
            db,
            sorted({
                row["normalized_order_no"]
                for row in plan["operations"]
                if row.get("normalized_order_no")
            }),
        )
        writes = 0
        noops = 0
        project_ids: set[str] = set()
        for item in sorted(
            plan["operations"],
            key=lambda row: (
                row["normalized_order_no"],
                row["action"],
            ),
        ):
            norm = item["normalized_order_no"]
            _advisory_lock(db, f"maintenance-bulk-sales:{norm}")
            sales: FSalesOrder | None = None
            if item.get("sales_order_id") is not None:
                sales = db.scalar(
                    select(FSalesOrder)
                    .where(FSalesOrder.id == item["sales_order_id"])
                    .with_for_update()
                )
                if sales is None:
                    raise BulkImportConflict("预览后的销售订单已不存在")
                if (
                    sales.raw_order_id != item["sales_raw_order_id"]
                    or maintenance_project_identity.normalize_xsdd(sales.order_no) != norm
                    or _jsonable(sales.amount_ex_tax)
                    != item["expected_sales_amount_ex_tax"]
                    or _jsonable(sales.tax_rate) != item["expected_sales_tax_rate"]
                    or sales.data_status != item["expected_sales_data_status"]
                ):
                    raise BulkImportConflict("预览后的销售订单事实已变化，请重新预览")
            else:
                variants = {item["sales_order_no"], norm, f"XSDD-{norm}"}
                candidates = list(
                    db.scalars(
                        select(FSalesOrder)
                        .where(
                            or_(
                                FSalesOrder.raw_order_id == item["contract_id"],
                                FSalesOrder.order_no.in_(sorted(variants)),
                            )
                        )
                        .order_by(FSalesOrder.id)
                        .with_for_update()
                    )
                )
                if _sales_fingerprint(candidates) != item[
                    "expected_sales_candidates"
                ]:
                    raise BulkImportConflict(
                        "预览后销售事实匹配状态已变化，请重新预览"
                    )
            new_inc = Decimal(item["new_contract_amount_inc_tax"])
            new_ex = Decimal(item["new_sales_amount_ex_tax"])
            new_rate = Decimal(item["new_sales_tax_rate"])
            if item["action"] == "noop":
                contract = db.get(
                    MaintenanceProjectContract, item["project_contract_id"]
                )
                if (
                    contract is None
                    or contract.version != item["expected_contract_version"]
                    or _jsonable(contract.amount_inc_tax)
                    != item["expected_contract_amount_inc_tax"]
                ):
                    raise BulkImportConflict("预览后的合同金额已变化，请重新预览")
                _apply_sales_project_period(
                    db,
                    item=item,
                    project_id=contract.project_id,
                    audit_reason=audit_reason,
                    operated_by=operated_by,
                )
                alias_created = maintenance_project_identity.record_alias(
                    db,
                    project_id=contract.project_id,
                    alias_name=item.get("source_project_name"),
                    source=maintenance_project_identity.sales_alias_source(norm),
                )
                if alias_created:
                    operations.bump_workbook_revision(
                        db, project_id=contract.project_id
                    )
                project_ids.add(contract.project_id)
                noops += 1
                continue

            project_id: str
            if item["action"] == "update_contract":
                contract = db.get(
                    MaintenanceProjectContract, item["project_contract_id"]
                )
                if (
                    contract is None
                    or contract.version != item["expected_contract_version"]
                    or _jsonable(contract.amount_inc_tax)
                    != item["expected_contract_amount_inc_tax"]
                ):
                    raise BulkImportConflict("预览后的合同金额已变化，请重新预览")
                payload = operations.update_contract(
                    db,
                    project_contract_id=contract.project_contract_id,
                    version=contract.version,
                    updates={"contract_amount": new_inc},
                    reason=audit_reason,
                    operated_by=operated_by,
                )
                if payload is None:
                    raise BulkImportConflict("合同在应用期间消失")
                project_id = contract.project_id
            elif item["action"] == "create_contract":
                variants = {item["sales_order_no"], norm, f"XSDD-{norm}"}
                if _all_contracts_by_order(db, variants).get(norm):
                    raise BulkImportConflict("预览后该销售订单已出现合同关系")
                project_id = item["project_id"]
                try:
                    current_owner = (
                        maintenance_project_identity.resolve_sales_xsdd_project(
                            db, norm
                        )
                    )
                except maintenance_project_identity.XsddProjectConflict as exc:
                    raise BulkImportConflict(str(exc)) from exc
                if current_owner != project_id:
                    raise BulkImportConflict("预览后的 XSDD owner 已变化")
                project = db.get(MaintenanceProject, project_id)
                if (
                    project is None
                    or not project.is_active
                    or project.version != item["expected_project_version"]
                ):
                    raise BulkImportConflict("预览后的目标项目已变化")
                payload = operations.create_contract(
                    db,
                    project_id=project_id,
                    contract_id=item["contract_id"],
                    contract_no=item["sales_order_no"],
                    contract_amount=new_inc,
                    contract_status=item["contract_status"],
                    status_mapping_state="mapped",
                    status_mapping_version="sales-order-bulk-v1",
                    included_in_total=True,
                    effective_from=date.fromisoformat(
                        item["contract_effective_from"]
                    ),
                    effective_to=None,
                    source="sales_order_bulk_v1",
                    reason=audit_reason,
                    operated_by=operated_by,
                )
                if payload is None:
                    raise BulkImportConflict("项目在创建合同期间消失")
            elif item["action"] == "create_project":
                variants = {item["sales_order_no"], norm, f"XSDD-{norm}"}
                if _all_contracts_by_order(db, variants).get(norm):
                    raise BulkImportConflict("预览后该销售订单已出现合同关系")
                try:
                    current_owner = (
                        maintenance_project_identity.resolve_sales_xsdd_project(
                            db, norm
                        )
                    )
                except maintenance_project_identity.XsddProjectConflict as exc:
                    raise BulkImportConflict(str(exc)) from exc
                if current_owner is not None:
                    raise BulkImportConflict("预览后的 XSDD owner 已出现")
                current, history = _assignment_evidence(db, variants)
                if current.get(norm) or history.get(norm):
                    raise BulkImportConflict("预览后销售订单已出现项目归属")
                metadata = item["new_project"]
                for existing in db.scalars(select(MaintenanceProject)):
                    if existing.project_code.casefold() == metadata["project_code"].casefold():
                        raise BulkImportConflict("预览后 XSDD 稳定项目编号已被占用")
                created = catalog.create_project(
                    db,
                    project_code=metadata["project_code"],
                    display_name=metadata["display_name"],
                    project_manager_id=None,
                    reason=audit_reason,
                    operated_by=operated_by,
                )
                project_id = created["project_id"]
                period_updates = {
                    key: date.fromisoformat(metadata[key])
                    for key in ("period_from", "period_to")
                    if metadata.get(key)
                }
                if period_updates:
                    updated = catalog.update_project(
                        db,
                        project_id=project_id,
                        version=created["version"],
                        updates=period_updates,
                        reason=audit_reason,
                        operated_by=operated_by,
                    )
                    if updated is None:
                        raise BulkImportConflict("维保项目在创建期间消失")
                payload = operations.create_contract(
                    db,
                    project_id=project_id,
                    contract_id=item["contract_id"],
                    contract_no=item["sales_order_no"],
                    contract_amount=new_inc,
                    contract_status=item["contract_status"],
                    status_mapping_state="mapped",
                    status_mapping_version="sales-order-bulk-v1",
                    included_in_total=True,
                    effective_from=date.fromisoformat(
                        item["contract_effective_from"]
                    ),
                    effective_to=None,
                    source="sales_order_bulk_v1",
                    reason=audit_reason,
                    operated_by=operated_by,
                )
                if payload is None:
                    raise BulkImportConflict("维保项目在创建合同期间消失")
            else:
                raise BulkImportConflict("预览计划包含未知销售订单动作")

            if item["action"] != "create_project":
                _apply_sales_project_period(
                    db,
                    item=item,
                    project_id=project_id,
                    audit_reason=audit_reason,
                    operated_by=operated_by,
                )
            alias_created = maintenance_project_identity.record_alias(
                db,
                project_id=project_id,
                alias_name=item.get("source_project_name"),
                source=maintenance_project_identity.sales_alias_source(norm),
            )
            if alias_created:
                # create/update contract already bumps in this transaction;
                # the dedupe registry makes this exactly-once.  It is the sole
                # bump for an amount-noop row that contributes a new peer name.
                operations.bump_workbook_revision(db, project_id=project_id)

            if sales is not None:
                before = {
                    "amount_ex_tax": _jsonable(sales.amount_ex_tax),
                    "tax_rate": _jsonable(sales.tax_rate),
                }
                sales.amount_ex_tax = new_ex
                sales.tax_rate = new_rate
                after = {
                    "amount_ex_tax": _jsonable(new_ex),
                    "tax_rate": _jsonable(new_rate),
                }
                if before != after:
                    db.add(SysAuditLog(
                        entity_type="sales_order_contract_amount",
                        entity_id=sales.id,
                        action="overwrite",
                        before_json=before,
                        after_json=after,
                        reason=audit_reason,
                        operated_by=operated_by,
                    ))
            writes += 1
            project_ids.add(project_id)
        return {
            "written": writes,
            "noop": noops,
            "operation": "sales_contract_amount_apply",
            "project_ids": sorted(project_ids),
        }


class ReceiptCollectionAdapter(HeaderAdapter):
    key = "receipt_collection"
    file_type = "maint_receipt"
    label = "收款单累计实收"
    aliases = {
        "order_no": ("收款明细.销售订单(必填)", "收款明细.销售订单", "销售订单", "合同编号"),
        "receipt_no": ("收款单号(必填)", "收款单号", "凭证号"),
        "receipt_date": ("收款日期(必填)", "收款日期", "到账日期"),
        "gross_amount": ("收款明细.销售收款金额(必填)", "销售收款金额"),
        "discount_amount": ("收款明细.优惠金额", "优惠金额"),
        "actual_amount": ("收款明细.实收金额", "实收金额", "到账金额"),
        "remark": ("收款明细.备注", "备注"),
        "receipt_status": ("数据状态", "收款状态", "状态"),
    }
    system_aliases = {
        "order_no": ("F0000008",),
        "receipt_no": ("SeqNo",),
        "receipt_date": ("F0000001",),
        "gross_amount": ("F0000013",),
        "discount_amount": ("F0000061",),
        "actual_amount": ("F0000062",),
        "remark": ("F0000037",),
        "receipt_status": ("Status",),
    }
    required_fields = frozenset({"order_no", "receipt_no", "receipt_date", "actual_amount"})

    @staticmethod
    def _refs(values: list[tuple[date, str]]) -> str:
        refs = [ref for _day, ref in sorted(set(values))]
        joined = ",".join(refs)
        return joined if len(joined) <= 128 else f"{refs[0]}等{len(refs)}笔"

    def build_plan(self, db: Session, sheet: DetectedSheet) -> dict:
        all_contracts, safe_contracts = _contract_maps(db)
        source_rows: list[dict] = []
        hard_issues: list[dict] = []
        by_order: dict[str, list[dict]] = defaultdict(list)
        seen_keys: dict[tuple[str, str], tuple[date, Decimal, int]] = {}

        for row_no, values in sheet.rows:
            order_raw = _text(_value(sheet, values, "order_no"))
            receipt_no = _text(_value(sheet, values, "receipt_no"))
            norm = normalize_order_no(order_raw)
            base = {
                "row_no": row_no,
                "business_key": f"{receipt_no}|{order_raw}",
                "normalized_order_no": norm,
                "issues": [],
            }
            try:
                if not norm or not receipt_no:
                    raise BulkImportInvalid("销售订单号和收款单号不能为空")
                receipt_status = _text(_value(sheet, values, "receipt_status"))
                if "receipt_status" in sheet.field_indexes:
                    if receipt_status != config.ACTIVE_STATUS:
                        raise BulkImportInvalid(
                            "收款状态必须明确为已生效，不接受空值、作废或其他状态"
                        )
                receipt_date = _date(_value(sheet, values, "receipt_date"), label="收款日期")
                actual = _decimal(_value(sheet, values, "actual_amount"), label="实收金额")
                gross = _optional_decimal(
                    _value(sheet, values, "gross_amount"), label="销售收款金额"
                )
                discount = _optional_decimal(
                    _value(sheet, values, "discount_amount"), label="优惠金额"
                ) or Decimal("0.00")
                if gross is not None and abs((gross - discount) - actual) > Decimal("0.01"):
                    raise BulkImportInvalid("销售收款金额-优惠金额与实收金额不一致")
                remark = _text(_value(sheet, values, "remark"))
                risk_word = next(
                    (
                        word
                        for word in ("坏账", "红冲", "冲销", "作废")
                        if word in remark
                    ),
                    None,
                )
                if risk_word:
                    raise BulkImportInvalid(
                        f"备注命中“{risk_word}”，需要人工确认，不能自动累计"
                    )
            except BulkImportInvalid as exc:
                issue = _row_issue(row_no, "invalid_receipt", str(exc))
                base.update(action="error", issues=[issue])
                source_rows.append(base)
                hard_issues.append(issue)
                continue

            if "receipt_status" not in sheet.field_indexes:
                base["issues"].append(
                    _row_issue(
                        row_no,
                        "source_status_missing",
                        "源收款导出没有状态列；已保留警告并仅按金额/备注做保守校验",
                        severity="warning",
                    )
                )

            duplicate = seen_keys.get((norm, receipt_no))
            if duplicate is not None:
                if duplicate[:2] != (receipt_date, actual):
                    issue = _row_issue(
                        row_no,
                        "duplicate_receipt_conflict",
                        f"同一收款单+销售订单与第 {duplicate[2]} 行金额/日期不一致",
                    )
                    base.update(action="error", issues=[issue])
                    hard_issues.append(issue)
                else:
                    base.update(
                        action="duplicate",
                        issues=[_row_issue(
                            row_no,
                            "duplicate_same",
                            f"与第 {duplicate[2]} 行相同，累计时只计一次",
                            severity="warning",
                        )],
                    )
                source_rows.append(base)
                continue
            seen_keys[(norm, receipt_no)] = (receipt_date, actual, row_no)

            current_relations = all_contracts.get(norm, [])
            contract = safe_contracts.get(norm)
            if contract is None:
                if not current_relations:
                    base.update(
                        action="unmatched",
                        issues=[_row_issue(
                            row_no,
                            "project_not_found",
                            "销售订单未关联当前维保项目，本批跳过且不创建项目",
                            severity="warning",
                        )],
                    )
                else:
                    issue = _row_issue(
                        row_no,
                        "contract_ambiguous",
                        "销售订单的当前项目合同关系共享、未映射或未计入总额",
                    )
                    base.update(action="error", issues=[issue])
                    hard_issues.append(issue)
                source_rows.append(base)
                continue
            base.update(
                action="matched",
                project_id=contract.project_id,
                project_contract_id=contract.project_contract_id,
                actual_amount=_jsonable(actual),
                gross_amount=_jsonable(gross),
                discount_amount=_jsonable(discount),
                receipt_date=receipt_date.isoformat(),
            )
            source_rows.append(base)
            by_order[norm].append({
                "row_no": row_no,
                "receipt_no": receipt_no,
                "receipt_date": receipt_date,
                "actual": actual,
                "remark": remark,
                "contract": contract,
            })

        # A receipt export represents cumulative history.  If any physical row
        # for an order is invalid (risk remark, status, amount/date, or a
        # conflicting duplicate), calculating from only the remaining rows
        # would silently create a partial cumulative snapshot.  Freeze every
        # known month for that order as blocked instead.
        blocked_norms = {
            str(row.get("normalized_order_no") or "")
            for row in source_rows
            if row.get("action") == "error" and row.get("normalized_order_no")
        }
        blocked_operations: list[dict] = []
        for norm in sorted(blocked_norms):
            order_sources = [
                row
                for row in source_rows
                if row.get("normalized_order_no") == norm
            ]
            issue = _row_issue(
                min(int(row["row_no"]) for row in order_sources),
                "order_level_fail_closed",
                f"销售订单 {norm} 存在无效/风险/冲突收款行，禁止从其余行计算部分累计",
            )
            hard_issues.append(issue)
            for row in order_sources:
                if not any(
                    item.get("code") == "order_level_fail_closed"
                    for item in row.get("issues") or []
                ):
                    row.setdefault("issues", []).append(issue)
            receipts = by_order.pop(norm, [])
            monthly: dict[date, list[dict]] = defaultdict(list)
            for receipt in receipts:
                day = receipt["receipt_date"]
                monthly[date(day.year, day.month, 1)].append(receipt)
            for month, month_rows in sorted(monthly.items()):
                contract = month_rows[0]["contract"]
                blocked_operations.append({
                    "row_no": min(row["row_no"] for row in month_rows),
                    "normalized_order_no": norm,
                    "project_id": contract.project_id,
                    "project_contract_id": contract.project_contract_id,
                    "expected_contract_version": contract.version,
                    "report_month": month.isoformat(),
                    # Deliberately absent: no partial cumulative value is ever
                    # calculated from the subset that happened to parse.
                    "new_cumulative_amount": None,
                    "receipt_reference": self._refs([
                        (row["receipt_date"], row["receipt_no"])
                        for row in month_rows
                    ]),
                    "action": "conflict",
                    "issues": [issue],
                })

        contract_ids = sorted({
            rows[0]["contract"].project_contract_id for rows in by_order.values() if rows
        })
        existing_rows = list(db.scalars(
            select(MaintenanceCollectionSnapshot)
            .where(MaintenanceCollectionSnapshot.project_contract_id.in_(contract_ids or [""]))
            .order_by(
                MaintenanceCollectionSnapshot.project_contract_id,
                MaintenanceCollectionSnapshot.report_month,
            )
        ))
        existing = {
            (row.project_contract_id, row.report_month): row for row in existing_rows
        }
        operations_plan: list[dict] = blocked_operations

        for norm, receipts in sorted(by_order.items()):
            contract = receipts[0]["contract"]
            operation_start = len(operations_plan)
            monthly: dict[date, list[dict]] = defaultdict(list)
            for receipt in receipts:
                day = receipt["receipt_date"]
                monthly[date(day.year, day.month, 1)].append(receipt)
            cumulative = Decimal("0.00")
            expected_months: set[date] = set()
            for month, month_rows in sorted(monthly.items()):
                cumulative += sum((row["actual"] for row in month_rows), Decimal("0.00"))
                cumulative = cumulative.quantize(Decimal("0.01"))
                expected_months.add(month)
                refs = self._refs([
                    (row["receipt_date"], row["receipt_no"]) for row in month_rows
                ])
                current = existing.get((contract.project_contract_id, month))
                item = {
                    "row_no": min(row["row_no"] for row in month_rows),
                    "normalized_order_no": norm,
                    "project_id": contract.project_id,
                    "project_contract_id": contract.project_contract_id,
                    "expected_contract_version": contract.version,
                    "report_month": month.isoformat(),
                    "new_cumulative_amount": _jsonable(cumulative),
                    "receipt_reference": refs,
                }
                if current is None:
                    item.update(action="create", expected_collection_id=None)
                elif current.status != "confirmed" or current.cumulative_amount != cumulative:
                    issue = _row_issue(
                        item["row_no"],
                        "snapshot_conflict",
                        f"{contract.contract_no} {month:%Y-%m} 已有快照与源实收累计不一致",
                    )
                    item.update(
                        action="conflict",
                        expected_collection_id=current.collection_id,
                        expected_collection_version=current.version,
                        expected_current_amount=_jsonable(current.cumulative_amount),
                        issues=[issue],
                    )
                    hard_issues.append(issue)
                else:
                    item.update(
                        action="noop",
                        expected_collection_id=current.collection_id,
                        expected_collection_version=current.version,
                        expected_current_amount=_jsonable(current.cumulative_amount),
                        preserve_receipt_reference=current.receipt_reference,
                    )
                operations_plan.append(item)

            extras = [
                row for row in existing_rows
                if row.project_contract_id == contract.project_contract_id
                and row.report_month not in expected_months
                and row.status == "confirmed"
            ]
            if extras:
                issue = _row_issue(
                    min(row["row_no"] for row in receipts),
                    "incomplete_receipt_history",
                    f"{contract.contract_no} 的源文件缺少 {len(extras)} 个生产已有月份，不能证明是完整累计导出",
                )
                hard_issues.append(issue)
                for item in operations_plan[operation_start:]:
                    item["action"] = "conflict"
                    item.setdefault("issues", []).append(issue)
                for row in source_rows:
                    if row.get("normalized_order_no") == norm:
                        row.setdefault("issues", []).append(issue)

        return _finish_plan(
            sheet,
            source_rows,
            operations_plan,
            hard_issues,
            extra={
                "amount_basis": "actual_received_inc_tax",
                "source_amount_field": "收款明细.实收金额",
                "missing_project_policy": "skip_without_create",
                "existing_snapshot_policy": "exact_noop_else_conflict",
                "source_warnings": (
                    [
                        {
                            "code": "source_status_missing",
                            "message": "源收款工作簿没有状态列；每行均带同名 warning",
                            "severity": "warning",
                        }
                    ]
                    if "receipt_status" not in sheet.field_indexes
                    else []
                ),
            },
        )

    def apply_plan(
        self,
        db: Session,
        plan: dict,
        *,
        operated_by: str,
        audit_reason: str,
    ) -> dict:
        target_contracts = sorted({
            item["project_contract_id"] for item in plan["operations"]
            if item["action"] in {"create", "noop"}
        })
        contracts = {
            row.project_contract_id: row for row in db.scalars(
                select(MaintenanceProjectContract).where(
                    MaintenanceProjectContract.project_contract_id.in_(target_contracts or [""])
                )
            )
        }
        snapshots = {
            (row.project_contract_id, row.report_month): row for row in db.scalars(
                select(MaintenanceCollectionSnapshot).where(
                    MaintenanceCollectionSnapshot.project_contract_id.in_(target_contracts or [""])
                )
            )
        }
        for item in plan["operations"]:
            if item["action"] not in {"create", "noop"}:
                continue
            contract = contracts.get(item["project_contract_id"])
            if contract is None or contract.version != item["expected_contract_version"]:
                raise BulkImportConflict("预览后的项目合同关系已变化")
            month = date.fromisoformat(item["report_month"])
            current = snapshots.get((contract.project_contract_id, month))
            if item["action"] == "create":
                if current is not None:
                    raise BulkImportConflict("预览后同合同同月份已新增回款快照")
            else:
                if (
                    current is None
                    or current.collection_id != item["expected_collection_id"]
                    or current.version != item["expected_collection_version"]
                    or current.status != "confirmed"
                    or _jsonable(current.cumulative_amount)
                    != item["expected_current_amount"]
                ):
                    raise BulkImportConflict("预览后的既有回款快照已变化")

        writes = 0
        noops = 0
        for item in sorted(
            plan["operations"],
            key=lambda row: (
                row["project_id"], row["project_contract_id"], row["report_month"]
            ),
        ):
            if item["action"] == "noop":
                noops += 1
                continue
            if item["action"] != "create":
                continue
            payload = operations.create_collection(
                db,
                project_id=item["project_id"],
                project_contract_id=item["project_contract_id"],
                report_month=date.fromisoformat(item["report_month"]),
                cumulative_amount=Decimal(item["new_cumulative_amount"]),
                status="confirmed",
                receipt_reference=item["receipt_reference"],
                remark=None,
                reason=audit_reason,
                operated_by=operated_by,
                source="direct_api",
                import_batch_id=None,
            )
            if payload is None:
                raise BulkImportConflict("项目在应用期间消失")
            writes += 1
        return {
            "written": writes,
            "noop": noops,
            "operation": "collection_snapshot_create",
            "project_ids": sorted(
                {
                    item["project_id"]
                    for item in plan["operations"]
                    if item["action"] in {"create", "noop"}
                }
            ),
        }


def _finish_plan(
    sheet: DetectedSheet,
    rows: list[dict],
    operations_plan: list[dict],
    hard_issues: list[dict],
    *,
    extra: dict,
) -> dict:
    actions: dict[str, int] = defaultdict(int)
    for row in rows:
        actions[row.get("action", "unknown")] += 1
    op_actions: dict[str, int] = defaultdict(int)
    for row in operations_plan:
        op_actions[row.get("action", "unknown")] += 1
    return _jsonable({
        "protocol_version": PROTOCOL_VERSION,
        "sheet": sheet.name,
        "header_row": sheet.header_row,
        "header_rows": list(sheet.header_rows),
        "headers": list(sheet.headers),
        "system_headers": list(sheet.system_headers),
        "field_matches": sheet.field_matches,
        "rows": rows,
        "operations": operations_plan,
        "issues": hard_issues,
        "summary": {
            "source_rows": len(rows),
            "source_actions": dict(actions),
            "target_operations": len(operations_plan),
            "operation_actions": dict(op_actions),
            "blocking_errors": len(hard_issues),
        },
        **extra,
    })


register_adapter(SalesContractAmountAdapter())
register_adapter(ReceiptCollectionAdapter())


def transformed_sales_sheet(
    result,
    *,
    source_columns: list[str],
) -> DetectedSheet:
    """Adapt the already-bounded generic sales transform without reopening XLSX."""

    internal_by_detected = {
        "order_no": "order_no",
        "raw_order_id": "raw_order_id",
        "data_status": "data_status",
        "maintenance_business": "maintenance_business",
        "business_type": "business_type",
        "project_name": "maintenance_project_name",
        "period_from": "maintenance_period_from",
        "period_to": "maintenance_period_to",
        "order_amount": "amount_inc_tax",
        "tax_flag": "is_tax_inclusive",
        "tax_rate": "tax_rate",
        "tax_amount": "tax_amount",
        "amount_ex_tax": "amount_ex_tax",
    }
    present_internal = {
        mapping.SALES_HEAD[column]
        for column in source_columns
        if column in mapping.SALES_HEAD
    }
    fields = [
        field
        for field, internal in internal_by_detected.items()
        if internal in present_internal
    ]
    field_indexes = {field: index for index, field in enumerate(fields)}
    rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_no, order in enumerate(result.orders.values(), start=3):
        values: list[Any] = []
        for field in fields:
            value = order.get(internal_by_detected[field])
            if field == "tax_flag":
                value = "含税" if value is True else (
                    "不含税" if value is False else None
                )
            values.append(value)
        rows.append((row_no, tuple(values)))
    return DetectedSheet(
        name="generic-sales-transform",
        header_row=2,
        header_rows=(1, 2),
        headers=tuple(fields),
        system_headers=(),
        field_indexes=field_indexes,
        field_matches={},
        rows=tuple(rows),
    )


def prelock_uploaded_sales_sheet(
    db: Session,
    sheet: DetectedSheet,
    *,
    mode: str = "upsert",
    operated_by: str = "system",
) -> dict[str, list[str]]:
    """Prelock the exact WBDD backlog a maintenance sales upload may link."""

    if mode == "skip":
        incoming_raw_ids = {
            _text(_value(sheet, values, "raw_order_id"))
            for _row_no, values in sheet.rows
            if _text(_value(sheet, values, "raw_order_id"))
        }
        existing_raw_ids = set(db.scalars(
            select(FSalesOrder.raw_order_id).where(
                FSalesOrder.raw_order_id.in_(sorted(incoming_raw_ids))
            )
        )) if incoming_raw_ids else set()
        sheet = DetectedSheet(
            name=sheet.name,
            header_row=sheet.header_row,
            header_rows=sheet.header_rows,
            headers=sheet.headers,
            system_headers=sheet.system_headers,
            field_indexes=sheet.field_indexes,
            field_matches=sheet.field_matches,
            rows=tuple(
                (row_no, values)
                for row_no, values in sheet.rows
                if _text(_value(sheet, values, "raw_order_id"))
                not in existing_raw_ids
            ),
        )
    eligible_rows = _ordinary_maintenance_rows(sheet)
    incoming_amounts: dict[str, set[Decimal]] = defaultdict(set)
    for _row_no, values in eligible_rows:
        xsdd = maintenance_project_identity.normalize_xsdd(
            _text(_value(sheet, values, "order_no"))
        )
        if not xsdd:
            raise BulkImportInvalid(
                "维保销售订单号不是有效 XSDD-YYYYMMDD-NNN/NNNN"
            )
        amount_inc_tax, _amount_ex_tax, _tax_rate = (
            SalesContractAmountAdapter._amounts(sheet, values)
        )
        incoming_amounts[xsdd].add(amount_inc_tax)
    xsdds = set(incoming_amounts)
    if not xsdds:
        return {}
    ambiguous_amounts = {
        xsdd: sorted(str(amount) for amount in amounts)
        for xsdd, amounts in incoming_amounts.items()
        if len(amounts) != 1
    }
    if ambiguous_amounts:
        raise BulkImportInvalid(
            f"同一 XSDD 的销售源行含税金额不唯一：{ambiguous_amounts}"
        )
    from app.services import maintenance_source_assignments as assignments

    try:
        maintenance_project_identity.auto_merge_sales_xsdd_conflicts(
            db,
            incoming_amount_inc_tax_by_xsdd={
                xsdd: next(iter(amounts))
                for xsdd, amounts in incoming_amounts.items()
            },
            operated_by=operated_by,
        )
        return assignments.prelock_sales_xsdd_backlog(
            db, xsdd_values=xsdds
        )
    except (
        assignments.SourceAssignmentConflict,
        maintenance_project_identity.XsddProjectMergeConflict,
    ) as exc:
        raise BulkImportConflict(str(exc)) from exc


def sync_uploaded_sales_workbook(
    db: Session,
    data: bytes | None,
    filename: str,
    *,
    operated_by: str,
    import_batch_id: int,
    prelocked_xsdd_order_ids: dict[str, list[str]] | None = None,
    detected_sheet: DetectedSheet | None = None,
) -> dict:
    """Project explicit maintenance sales rows during the ordinary import.

    The ordinary ETL remains the sales fact writer.  This hook reuses the
    reviewed XSDD planner/apply path, but feeds it only rows whose source
    columns explicitly declare maintenance business.  Non-maintenance rows
    are therefore incapable of creating/updating maintenance contracts.
    """

    if detected_sheet is None:
        if data is None:
            raise BulkImportInvalid("缺少销售订单投影输入")
        try:
            adapter, sheet = _detect(data)
        except BulkImportInvalid as exc:
            return {
                "status": "not_applicable",
                "eligible_rows": 0,
                "reason": str(exc),
            }
        if adapter.key != "sales_contract_amount":
            return {
                "status": "not_applicable",
                "eligible_rows": 0,
                "reason": "未识别为销售订单合同事实表",
            }
    else:
        adapter = _ADAPTERS["sales_contract_amount"]
        sheet = detected_sheet
    written_raw_ids = set(db.scalars(select(FSalesOrder.raw_order_id).where(
        FSalesOrder.import_batch_id == import_batch_id
    )))
    written_sheet = DetectedSheet(
        name=sheet.name,
        header_row=sheet.header_row,
        header_rows=sheet.header_rows,
        headers=sheet.headers,
        system_headers=sheet.system_headers,
        field_indexes=sheet.field_indexes,
        field_matches=sheet.field_matches,
        rows=tuple(
            (row_no, values)
            for row_no, values in sheet.rows
            if _text(_value(sheet, values, "raw_order_id")) in written_raw_ids
        ),
    )
    ordinary_eligible_rows = _ordinary_maintenance_rows(written_sheet)
    if any(
        not maintenance_project_identity.normalize_xsdd(
            _text(_value(written_sheet, values, "order_no"))
        )
        for _row_no, values in ordinary_eligible_rows
    ):
        raise BulkImportInvalid("维保销售订单号不是有效 XSDD-YYYYMMDD-NNN/NNNN")
    eligible_rows = ordinary_eligible_rows
    if not eligible_rows:
        return {"status": "no_maintenance_rows", "eligible_rows": 0}
    maintenance_sheet = DetectedSheet(
        name=sheet.name,
        header_row=sheet.header_row,
        header_rows=sheet.header_rows,
        headers=sheet.headers,
        system_headers=sheet.system_headers,
        field_indexes=sheet.field_indexes,
        field_matches=sheet.field_matches,
        rows=eligible_rows,
    )
    plan = adapter.build_plan(db, maintenance_sheet)
    blocking = int((plan.get("summary") or {}).get("blocking_errors") or 0)
    if blocking:
        raise BulkImportInvalid(
            f"维保销售订单自动建项有 {blocking} 个阻断错误",
            issues=plan.get("issues") or [],
        )
    result = adapter.apply_plan(
        db,
        plan,
        operated_by=operated_by,
        audit_reason=(
            f"普通销售订单导入自动建维保项目 "
            f"batch={import_batch_id} filename={filename[:128]}"
        ),
    )
    eligible_xsdds = {
        maintenance_project_identity.normalize_xsdd(
            _text(_value(maintenance_sheet, values, "order_no"))
        )
        for _row_no, values in maintenance_sheet.rows
    }
    candidate_order_ids = {
        raw_order_id
        for xsdd in eligible_xsdds
        for raw_order_id in (prelocked_xsdd_order_ids or {}).get(xsdd, [])
    }
    linked_wbdd = None
    if candidate_order_ids:
        from app.services import maintenance_source_assignments as assignments

        try:
            linked_wbdd = assignments.auto_assign_existing_orders(
                db,
                operated_by=operated_by,
                source_order_ids=candidate_order_ids,
            )
        except assignments.SourceAssignmentConflict as exc:
            raise BulkImportConflict(str(exc)) from exc
    response = {
        "status": "applied",
        "eligible_rows": len(eligible_rows),
        "source_actions": (plan.get("summary") or {}).get("source_actions") or {},
        **result,
    }
    if linked_wbdd is not None:
        response["linked_wbdd"] = linked_wbdd
    return response


def build_preview(db: Session, data: bytes, filename: str) -> PreviewArtifact:
    if not filename.lower().endswith(".xlsx"):
        raise BulkImportInvalid("只接受 .xlsx 文件")
    file_hash = hashlib.sha256(data).hexdigest()
    adapter, sheet = _detect(data)
    plan = adapter.build_plan(db, sheet)
    plan.update({
        "form_type": adapter.key,
        "file_type": adapter.file_type,
        "filename": filename,
        "file_hash": file_hash,
    })
    return PreviewArtifact(
        adapter_key=adapter.key,
        file_type=adapter.file_type,
        file_hash=file_hash,
        filename=filename,
        plan=_jsonable(plan),
    )


def _preview_response(batch: SysImportBatch, token: str | None) -> dict:
    report = batch.report_json or {}
    plan = report.get("plan") or {}
    return {
        "batch_id": batch.id,
        "status": batch.status,
        "form_type": report.get("form_type"),
        "file_hash": batch.file_hash,
        "plan_hash": report.get("plan_hash"),
        "commit_token": token,
        "summary": plan.get("summary") or {},
        "rows": plan.get("rows") or [],
        "operations": plan.get("operations") or [],
        "issues": plan.get("issues") or [],
        "result": report.get("result"),
    }


def store_preview(
    db: Session,
    artifact: PreviewArtifact,
    *,
    operated_by: str,
    operation_key: str,
) -> dict:
    operation_hash = _operation_hash(operated_by, operation_key)
    _advisory_lock(db, f"maintenance-bulk-preview:{operation_hash}")
    existing = db.scalar(
        select(SysImportBatch)
        .where(
            SysImportBatch.uploaded_by == operated_by,
            SysImportBatch.file_type.in_(SUPPORTED_BATCH_TYPES),
            SysImportBatch.report_json["operation_key_hash"].as_string()
            == operation_hash,
        )
        .order_by(SysImportBatch.id.desc())
        .limit(1)
    )
    if existing is not None:
        if existing.file_hash != artifact.file_hash or existing.file_type != artifact.file_type:
            raise BulkImportConflict("同一 Idempotency-Key 已用于另一份文件")
        if existing.status == "success":
            return _preview_response(existing, None)
        token = secrets.token_urlsafe(32)
        report = dict(existing.report_json or {})
        report["token_hash"] = _token_hash(existing.id, token)
        report["token_rotated_at"] = datetime.now(timezone.utc).isoformat()
        existing.report_json = report
        db.commit()
        return _preview_response(existing, token)

    batch = SysImportBatch(
        filename=artifact.filename,
        file_type=artifact.file_type,
        file_hash=artifact.file_hash,
        uploaded_by=operated_by,
        rows_total=int(artifact.plan["summary"]["source_rows"]),
        rows_inserted=0,
        rows_skipped=sum(
            int(artifact.plan["summary"]["source_actions"].get(key, 0))
            for key in ("noop", "unmatched", "duplicate")
        ),
        rows_error=int(artifact.plan["summary"]["blocking_errors"]),
        status="processing",
    )
    db.add(batch)
    db.flush()
    token = secrets.token_urlsafe(32)
    plan_hash = _canonical_hash(artifact.plan)
    batch.report_json = {
        "protocol_version": PROTOCOL_VERSION,
        "form_type": artifact.adapter_key,
        "operation_key_hash": operation_hash,
        "plan_hash": plan_hash,
        "token_hash": _token_hash(batch.id, token),
        "plan": artifact.plan,
        "previewed_at": datetime.now(timezone.utc).isoformat(),
    }
    for issue in artifact.plan.get("issues") or []:
        db.add(SysImportError(
            batch_id=batch.id,
            row_no=issue.get("row_no"),
            error_type=str(issue.get("code") or "bulk_import")[:32],
            error_detail=str(issue.get("message") or "")[:4000],
            raw_row=issue,
        ))
    db.commit()
    return _preview_response(batch, token)


def preview(
    db: Session,
    data: bytes,
    filename: str,
    *,
    operated_by: str,
    operation_key: str,
) -> dict:
    return store_preview(
        db,
        build_preview(db, data, filename),
        operated_by=operated_by,
        operation_key=operation_key,
    )


def get_batch(db: Session, batch_id: int, *, operated_by: str, allow_admin: bool = False) -> dict:
    batch = db.get(SysImportBatch, batch_id)
    if batch is None or batch.file_type not in SUPPORTED_BATCH_TYPES:
        raise BulkImportNotFound("批量导入批次不存在")
    if batch.uploaded_by != operated_by and not allow_admin:
        raise BulkImportNotFound("批量导入批次不存在")
    return _preview_response(batch, None)


def apply_preview(
    db: Session,
    batch_id: int,
    *,
    commit_token: str,
    plan_hash: str,
    operated_by: str,
    allow_admin: bool = False,
) -> dict:
    batch = db.scalar(
        select(SysImportBatch)
        .where(SysImportBatch.id == batch_id)
        .with_for_update()
    )
    if batch is None or batch.file_type not in SUPPORTED_BATCH_TYPES:
        raise BulkImportNotFound("批量导入批次不存在")
    if batch.uploaded_by != operated_by and not allow_admin:
        raise BulkImportNotFound("批量导入批次不存在")
    report = dict(batch.report_json or {})
    if not hmac.compare_digest(
        str(report.get("token_hash") or ""), _token_hash(batch.id, commit_token)
    ):
        raise BulkImportConflict("提交 token 无效或已轮换")
    if not hmac.compare_digest(str(report.get("plan_hash") or ""), plan_hash):
        raise BulkImportConflict("预览计划 hash 不匹配")
    if batch.status == "success":
        return {"batch_id": batch.id, "status": "success", **(report.get("result") or {})}
    plan = report.get("plan") or {}
    if _canonical_hash(plan) != plan_hash:
        raise BulkImportConflict("服务器保存的预览计划校验失败")
    if int((plan.get("summary") or {}).get("blocking_errors") or 0) > 0:
        raise BulkImportInvalid("预览包含阻断错误，不能提交", issues=plan.get("issues") or [])

    _advisory_lock(db, f"maintenance-bulk-apply:{batch.file_type}:{batch.file_hash}")
    already = db.scalar(
        select(SysImportBatch)
        .where(
            SysImportBatch.id != batch.id,
            SysImportBatch.file_type == batch.file_type,
            SysImportBatch.file_hash == batch.file_hash,
            SysImportBatch.status == "success",
        )
        .order_by(SysImportBatch.id)
        .limit(1)
    )
    if already is not None:
        prior = (already.report_json or {}).get("result") or {}
        report["result"] = {**prior, "duplicate_of_batch_id": already.id}
        report["applied_at"] = datetime.now(timezone.utc).isoformat()
        report["applied_by"] = operated_by
        batch.report_json = report
        batch.status = "duplicate"
        db.commit()
        return {"batch_id": batch.id, "status": "duplicate", **report["result"]}

    adapter = _ADAPTERS.get(str(report.get("form_type") or ""))
    if adapter is None or adapter.file_type != batch.file_type:
        raise BulkImportConflict("批次适配器不存在或协议已变化")
    audit_reason = (
        f"全项目批量导入 batch={batch.id} form={adapter.key} "
        f"sha256={batch.file_hash}"
    )
    result = adapter.apply_plan(
        db,
        plan,
        operated_by=operated_by,
        audit_reason=audit_reason,
    )
    result = _jsonable({
        **result,
        "form_type": adapter.key,
        "file_hash": batch.file_hash,
        "plan_hash": plan_hash,
    })
    report["result"] = result
    report["applied_at"] = datetime.now(timezone.utc).isoformat()
    report["applied_by"] = operated_by
    batch.report_json = report
    batch.status = "success"
    batch.rows_inserted = int(result.get("written") or 0)
    batch.rows_skipped = int(result.get("noop") or 0) + int(
        (plan.get("summary") or {}).get("source_actions", {}).get("unmatched", 0)
    )
    batch.rows_error = 0
    db.add(SysAuditLog(
        entity_type="maintenance_bulk_import",
        entity_id=batch.id,
        action="apply",
        before_json={"status": "processing", "plan_hash": plan_hash},
        after_json={"status": "success", **result},
        reason=audit_reason,
        operated_by=operated_by,
    ))
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"batch_id": batch.id, "status": "success", **result}


# ---------------------------------------------------------------------------
# Public multi-file project-batch-transfer protocol
# ---------------------------------------------------------------------------


def _transfer_kind(adapter_key: str) -> str:
    return {
        "sales_contract_amount": "sales_contract",
        "receipt_collection": "receipt",
    }.get(adapter_key, adapter_key)


def _split_issues(issues: list[dict] | None) -> tuple[list[dict], list[dict]]:
    warnings: list[dict] = []
    errors: list[dict] = []
    for issue in issues or []:
        item = {
            "code": str(issue.get("code") or "bulk_import"),
            "message": str(issue.get("message") or ""),
            "field": issue.get("field"),
        }
        (warnings if issue.get("severity") == "warning" else errors).append(item)
    return warnings, errors


def _detected_fields(adapter: FormAdapter, sheet: DetectedSheet) -> list[dict]:
    basis = {
        "order_amount": "源订单金额；含税标记为含税时直接作为合同含税额",
        "amount_inc_tax": "含税合同额",
        "amount_ex_tax": "销售事实未税金额",
        "tax_rate": "税率；缺失不按 0%",
        "tax_amount": "仅用于与含税/未税金额交叉校验",
        "gross_amount": "销售收款金额；仅用于 F-G=H 对账",
        "discount_amount": "优惠金额；仅用于 F-G=H 对账",
        "actual_amount": "累计回款的唯一金额来源",
    }
    fields = []
    for field, match in sheet.field_matches.items():
        fields.append(
            {
                "source_column": match["source_column"],
                "canonical_field": field,
                "canonical_label": adapter.aliases[field][0],
                "confidence": match["confidence"],
                "required": field in adapter.required_fields,
                "metric_basis": basis.get(field),
            }
        )
    return fields


def _project_names_for_ids(db: Session, project_ids: set[str]) -> dict[str, str]:
    if not project_ids:
        return {}
    return {
        project_id: display_name
        for project_id, display_name in db.execute(
            select(MaintenanceProject.project_id, MaintenanceProject.display_name)
            .where(MaintenanceProject.project_id.in_(sorted(project_ids)))
        )
    }


def _public_sales_rows(
    *,
    file_id: str,
    filename: str,
    plan_index: int,
    plan: dict,
    project_names_by_id: dict[str, str],
) -> tuple[list[dict], dict[str, dict]]:
    operations_by_row = {
        int(item["row_no"]): (idx, item)
        for idx, item in enumerate(plan.get("operations") or [])
    }
    rows: list[dict] = []
    row_map: dict[str, dict] = {}
    for source in plan.get("rows") or []:
        row_no = int(source["row_no"])
        op_pair = operations_by_row.get(row_no)
        warnings, errors = _split_issues(source.get("issues"))
        op_index: int | None = None
        operation: dict | None = None
        if op_pair is not None:
            op_index, operation = op_pair
        source_action = str(source.get("action") or "error")
        if operation is not None:
            internal_action = operation["action"]
            op_warnings, op_errors = _split_issues(operation.get("issues"))
            warning_codes = {item.get("code") for item in warnings}
            error_codes = {item.get("code") for item in errors}
            warnings.extend(
                item for item in op_warnings if item.get("code") not in warning_codes
            )
            errors.extend(
                item for item in op_errors if item.get("code") not in error_codes
            )
            action = {
                "create_project": "create_project",
                "create_contract": "create_contract",
                "update_contract": "update_contract",
                "noop": "skip",
                "blocked": "block",
            }[internal_action]
            row_status = (
                "blocked"
                if internal_action == "blocked"
                else ("unchanged" if internal_action == "noop" else "ready")
            )
            match_state = "ambiguous" if internal_action == "blocked" else "matched"
            project_id = operation.get("project_id")
            new_project = operation.get("new_project") or {}
            project_name = (
                new_project.get("display_name")
                or project_names_by_id.get(project_id)
            )
            match_strategy = (
                "none"
                if internal_action in {"create_project", "blocked"}
                else (
                    "candidate"
                    if internal_action == "create_contract"
                    else "exact_contract_no"
                )
            )
            canonical = {
                "sales_order_no": operation["sales_order_no"],
                "contract_amount_inc_tax": operation[
                    "new_contract_amount_inc_tax"
                ],
                "amount_ex_tax": operation["new_sales_amount_ex_tax"],
                "tax_rate": operation["new_sales_tax_rate"],
                "project_name": project_name,
            }
            target_key = f"sales:{operation['normalized_order_no']}"
            before = source.get("before")
            after = source.get("after")
        else:
            action = "block" if errors else "skip"
            row_status = "blocked" if errors else (
                "needs_review" if source_action == "unmatched" else "unchanged"
            )
            match_state = (
                "invalid"
                if errors
                else ("unmatched" if source_action == "unmatched" else "matched")
            )
            project_id = source.get("project_id")
            project_name = project_names_by_id.get(project_id)
            match_strategy = "none"
            canonical = {
                "sales_order_no": source.get("business_key"),
                "normalized_order_no": source.get("normalized_order_no"),
            }
            target_key = None
            before = source.get("before")
            after = source.get("after")
        row_key = _canonical_hash(
            [file_id, plan_index, row_no, target_key or source_action]
        )[:32]
        public = {
            "row_key": row_key,
            "file_id": file_id,
            "filename": filename,
            "detected_sheet": plan.get("sheet"),
            "source_row": row_no,
            "canonical": canonical,
            "normalized_key": source.get("normalized_order_no"),
            "idempotency_key": row_key,
            "matched_project_id": project_id,
            "matched_project_name": project_name,
            "matched_contract_id": (
                operation.get("contract_id") if operation else None
            ),
            "match_strategy": match_strategy,
            "candidate_count": 1 if project_id else 0,
            "candidates": [],
            "match_state": match_state,
            "action": action,
            "row_status": row_status,
            "before": before,
            "after": after,
            "delta": None,
            "warnings": warnings,
            "errors": errors,
            "_target_key": target_key,
        }
        rows.append(public)
        if operation is not None and row_status == "ready" and not errors:
            row_map[row_key] = {
                "plan_index": plan_index,
                "operation_index": op_index,
            }
    return rows, row_map


def _public_receipt_rows(
    *,
    file_id: str,
    filename: str,
    plan_index: int,
    plan: dict,
    project_names_by_id: dict[str, str],
) -> tuple[list[dict], dict[str, dict]]:
    rows: list[dict] = []
    row_map: dict[str, dict] = {}
    status_warnings = [
        {
            "code": item["code"],
            "message": item["message"],
            "field": None,
        }
        for item in plan.get("source_warnings") or []
    ]
    operation_source_rows: set[int] = set()
    for op_index, operation in enumerate(plan.get("operations") or []):
        row_no = int(operation["row_no"])
        operation_source_rows.add(row_no)
        warnings, errors = _split_issues(operation.get("issues"))
        warnings = [*status_warnings, *warnings]
        internal_action = operation["action"]
        action = (
            "upsert_collection_snapshot"
            if internal_action in {"create", "noop"}
            else "block"
        )
        row_status = (
            "ready"
            if internal_action == "create"
            else ("unchanged" if internal_action == "noop" else "blocked")
        )
        match_state = "matched" if internal_action in {"create", "noop"} else "ambiguous"
        target_key = (
            f"receipt:{operation['project_contract_id']}:"
            f"{operation['report_month']}"
        )
        row_key = _canonical_hash(
            [file_id, plan_index, row_no, target_key]
        )[:32]
        public = {
            "row_key": row_key,
            "file_id": file_id,
            "filename": filename,
            "detected_sheet": plan.get("sheet"),
            "source_row": row_no,
            "canonical": {
                "sales_order_no": operation["normalized_order_no"],
                "report_month": operation["report_month"],
                "cumulative_received_inc_tax": operation[
                    "new_cumulative_amount"
                ],
                "receipt_reference": operation["receipt_reference"],
            },
            "normalized_key": operation["normalized_order_no"],
            "idempotency_key": row_key,
            "matched_project_id": operation["project_id"],
            "matched_project_name": project_names_by_id.get(
                operation["project_id"]
            ),
            "matched_contract_id": operation["project_contract_id"],
            "match_strategy": "exact_contract_no",
            "candidate_count": 1,
            "candidates": [],
            "match_state": match_state,
            "action": action,
            "row_status": row_status,
            "before": (
                {"cumulative_amount": operation.get("expected_current_amount")}
                if operation.get("expected_collection_id")
                else None
            ),
            "after": {
                "cumulative_amount": operation["new_cumulative_amount"]
            },
            "delta": None,
            "warnings": warnings,
            "errors": errors,
            "_target_key": target_key,
        }
        rows.append(public)
        if row_status == "ready" and not errors:
            row_map[row_key] = {
                "plan_index": plan_index,
                "operation_index": op_index,
            }

    # Preserve invalid/unmatched/duplicate source evidence that did not become
    # a monthly target operation.  Matched receipt source lines are represented
    # by their aggregate snapshot above, avoiding double-selection semantics.
    for source in plan.get("rows") or []:
        row_no = int(source["row_no"])
        if source.get("action") == "matched":
            continue
        warnings, errors = _split_issues(source.get("issues"))
        warnings = [*status_warnings, *warnings]
        source_action = str(source.get("action") or "error")
        match_state = (
            "invalid"
            if errors
            else ("unmatched" if source_action == "unmatched" else "matched")
        )
        row_key = _canonical_hash(
            [file_id, plan_index, row_no, source.get("business_key")]
        )[:32]
        rows.append(
            {
                "row_key": row_key,
                "file_id": file_id,
                "filename": filename,
                "detected_sheet": plan.get("sheet"),
                "source_row": row_no,
                "canonical": {
                    "receipt_key": source.get("business_key"),
                    "sales_order_no": source.get("normalized_order_no"),
                },
                "normalized_key": source.get("normalized_order_no"),
                "idempotency_key": row_key,
                "matched_project_id": source.get("project_id"),
                "matched_project_name": project_names_by_id.get(
                    source.get("project_id")
                ),
                "matched_contract_id": source.get("project_contract_id"),
                "match_strategy": "none",
                "candidate_count": 0,
                "candidates": [],
                "match_state": match_state,
                "action": "block" if errors else "skip",
                "row_status": "blocked" if errors else "unchanged",
                "before": None,
                "after": None,
                "delta": None,
                "warnings": warnings,
                "errors": errors,
                "_target_key": None,
            }
        )
    return rows, row_map


def _transfer_summary(rows: list[dict]) -> dict:
    return {
        "total": len(rows),
        "matched": sum(row["match_state"] == "matched" for row in rows),
        "ambiguous": sum(row["match_state"] == "ambiguous" for row in rows),
        "unmatched": sum(row["match_state"] == "unmatched" for row in rows),
        "invalid": sum(row["match_state"] == "invalid" for row in rows),
        "ready": sum(row["row_status"] == "ready" for row in rows),
    }


def _enforce_project_scope(
    operations: list[dict],
    allowed_project_ids: set[str] | None,
) -> None:
    """Fail closed before preview exposure or apply writes.

    ``None`` is the shared full-scope sentinel.  A scoped account cannot create
    a project because the not-yet-created id cannot belong to its current
    visible set.  Missing or out-of-scope target ids reject the whole batch.
    """

    if allowed_project_ids is None:
        return
    for operation in operations:
        if operation.get("action") == "create_project":
            raise BulkImportScopeDenied("范围账号不能通过批量导入创建新项目")
        project_id = operation.get("project_id")
        if not project_id or str(project_id) not in allowed_project_ids:
            raise BulkImportScopeDenied("批次包含当前账号无权访问的项目，整批拒绝")


def preview_transfer(
    db: Session,
    files: list[tuple[str, bytes]],
    *,
    operated_by: str,
    allowed_project_ids: set[str] | None = None,
) -> dict:
    if not files or len(files) > MAX_TRANSFER_FILES:
        raise BulkImportInvalid(f"一次必须上传 1–{MAX_TRANSFER_FILES} 个文件")
    total_bytes = sum(len(data) for _name, data in files)
    if total_bytes > MAX_TRANSFER_TOTAL_BYTES:
        raise BulkImportInvalid("批量上传总大小超过 64 MiB")
    artifacts = [build_preview(db, data, filename) for filename, data in files]
    _enforce_project_scope(
        [
            operation
            for artifact in artifacts
            for operation in artifact.plan.get("operations") or []
        ],
        allowed_project_ids,
    )
    hashes = [artifact.file_hash for artifact in artifacts]
    if len(hashes) != len(set(hashes)):
        raise BulkImportInvalid("同一批次包含内容完全相同的重复文件")
    if sum(int(artifact.plan["summary"]["source_rows"]) for artifact in artifacts) > MAX_PREVIEW_ROWS:
        raise BulkImportInvalid(f"多文件数据行合计超过安全上限 {MAX_PREVIEW_ROWS}")

    plans = []
    file_payloads = []
    project_ids = {
        str(item["project_id"])
        for artifact in artifacts
        for item in artifact.plan.get("operations") or []
        if item.get("project_id")
    }
    project_names_by_id = _project_names_for_ids(db, project_ids)
    public_rows: list[dict] = []
    row_map: dict[str, dict] = {}
    for index, artifact in enumerate(artifacts):
        adapter = _ADAPTERS[artifact.adapter_key]
        file_id = f"file-{index + 1}-{artifact.file_hash[:12]}"
        plan = artifact.plan
        plans.append(
            {
                "file_id": file_id,
                "filename": artifact.filename,
                "form_type": artifact.adapter_key,
                "plan": plan,
            }
        )
        file_payloads.append(
            {
                "file_id": file_id,
                "filename": artifact.filename,
                "import_kind": _transfer_kind(artifact.adapter_key),
                "source_sha256": artifact.file_hash,
                "detected_sheet": plan.get("sheet"),
                "header_rows": plan.get("header_rows") or [plan.get("header_row")],
                "detected_fields": _detected_fields(
                    adapter,
                    DetectedSheet(
                        name=str(plan.get("sheet") or ""),
                        header_row=int(plan.get("header_row") or 1),
                        header_rows=tuple(plan.get("header_rows") or []),
                        headers=tuple(plan.get("headers") or []),
                        system_headers=tuple(plan.get("system_headers") or []),
                        field_indexes={},
                        field_matches=plan.get("field_matches") or {},
                        rows=(),
                    ),
                ),
                "mapping_conflicts": [],
            }
        )
        if artifact.adapter_key == "sales_contract_amount":
            rows, mapping = _public_sales_rows(
                file_id=file_id,
                filename=artifact.filename,
                plan_index=index,
                plan=plan,
                project_names_by_id=project_names_by_id,
            )
        else:
            rows, mapping = _public_receipt_rows(
                file_id=file_id,
                filename=artifact.filename,
                plan_index=index,
                plan=plan,
                project_names_by_id=project_names_by_id,
            )
        public_rows.extend(rows)
        row_map.update(mapping)

    # The same canonical target appearing in two source files is never applied
    # on first-file-wins semantics.  Both rows remain visible but blocked.
    by_target: dict[str, list[dict]] = defaultdict(list)
    for row in public_rows:
        if row.get("_target_key") and row["row_status"] == "ready":
            by_target[row["_target_key"]].append(row)
    for target, duplicates in by_target.items():
        if len(duplicates) < 2:
            continue
        for row in duplicates:
            row["match_state"] = "ambiguous"
            row["action"] = "block"
            row["row_status"] = "blocked"
            row["errors"].append(
                {
                    "code": "cross_file_duplicate_target",
                    "message": f"多个文件同时修改同一目标 {target}，请只保留一份来源",
                    "field": None,
                }
            )
            row_map.pop(row["row_key"], None)

    for row in public_rows:
        row.pop("_target_key", None)
    public_rows.sort(
        key=lambda row: (row["filename"], row["source_row"], row["row_key"])
    )
    summary = _transfer_summary(public_rows)
    expires_at = datetime.now(timezone.utc) + TRANSFER_TOKEN_TTL
    data_version = _canonical_hash(
        [
            {
                "form_type": item["form_type"],
                "operations": item["plan"].get("operations") or [],
            }
            for item in plans
        ]
    )
    public_payload = {
        "schema_version": TRANSFER_SCHEMA_VERSION,
        "files": file_payloads,
        "rows": public_rows,
        "summary": summary,
        "can_apply": summary["ready"] > 0,
    }
    payload_hash = _canonical_hash(
        {
            "public": public_payload,
            "plans": plans,
            "data_version": data_version,
        }
    )
    batch = SysImportBatch(
        filename=(";".join(name for name, _data in files))[:256],
        file_type=TRANSFER_BATCH_TYPE,
        file_hash=payload_hash,
        uploaded_by=operated_by,
        rows_total=summary["total"],
        rows_inserted=0,
        rows_skipped=summary["total"] - summary["ready"],
        rows_error=summary["invalid"] + summary["ambiguous"],
        status="processing",
    )
    db.add(batch)
    db.flush()
    secret = secrets.token_urlsafe(32)
    preview_token = f"{batch.id}.{secret}"
    batch.report_json = {
        "protocol_version": TRANSFER_SCHEMA_VERSION,
        "payload_hash": payload_hash,
        "data_version": data_version,
        "token_hash": hashlib.sha256(preview_token.encode("utf-8")).hexdigest(),
        "expires_at": expires_at.isoformat(),
        "plans": plans,
        "row_map": row_map,
        "public": public_payload,
        "previewed_at": datetime.now(timezone.utc).isoformat(),
    }
    for row in public_rows:
        for issue in row["errors"]:
            db.add(
                SysImportError(
                    batch_id=batch.id,
                    row_no=row["source_row"],
                    error_type=str(issue["code"])[:32],
                    error_detail=str(issue["message"])[:4000],
                    raw_row={
                        "row_key": row["row_key"],
                        "file_id": row["file_id"],
                        "filename": row["filename"],
                        "issue": issue,
                    },
                )
            )
    db.commit()
    return {
        **public_payload,
        "preview_id": str(batch.id),
        "preview_token": preview_token,
        "payload_hash": payload_hash,
        "data_version": data_version,
        "expires_at": expires_at.isoformat(),
    }


def _batch_id_from_transfer_token(preview_token: str) -> int:
    prefix, separator, secret = preview_token.partition(".")
    if not separator or not prefix.isdigit() or len(secret) < 24:
        raise BulkImportConflict("预览 token 无效")
    return int(prefix)


def record_transfer_failure(
    db: Session,
    *,
    preview_token: str,
    operated_by: str,
    error_code: str,
    message: str,
    allow_admin: bool = False,
) -> None:
    """Persist a failed apply attempt after the domain transaction rolled back."""

    batch_id = _batch_id_from_transfer_token(preview_token)
    batch = db.scalar(
        select(SysImportBatch)
        .where(SysImportBatch.id == batch_id)
        .with_for_update()
    )
    if (
        batch is None
        or batch.file_type != TRANSFER_BATCH_TYPE
        or (batch.uploaded_by != operated_by and not allow_admin)
        or batch.status != "processing"
    ):
        return
    report = dict(batch.report_json or {})
    expected = str(report.get("token_hash") or "")
    actual = hashlib.sha256(preview_token.encode("utf-8")).hexdigest()
    if not hmac.compare_digest(expected, actual):
        return
    failure = {
        "error_code": error_code[:64],
        "message": message[:1000],
        "failed_at": datetime.now(timezone.utc).isoformat(),
        "failed_by": operated_by,
    }
    report["failure"] = failure
    batch.report_json = report
    batch.status = "failed"
    batch.rows_error = max(int(batch.rows_error or 0), 1)
    db.add(
        SysAuditLog(
            entity_type="maintenance_bulk_import",
            entity_id=batch.id,
            action="failed",
            before_json={"status": "processing"},
            after_json={"status": "failed", **failure},
            reason="批量导入应用失败，业务事务已整体回滚",
            operated_by=operated_by,
        )
    )
    db.commit()


def apply_transfer(
    db: Session,
    *,
    preview_token: str,
    payload_hash: str,
    data_version: str,
    row_keys: list[str],
    operated_by: str,
    allow_admin: bool = False,
    allowed_project_ids: set[str] | None = None,
) -> dict:
    batch_id = _batch_id_from_transfer_token(preview_token)
    batch = db.scalar(
        select(SysImportBatch)
        .where(SysImportBatch.id == batch_id)
        .with_for_update()
    )
    if batch is None or batch.file_type != TRANSFER_BATCH_TYPE:
        raise BulkImportNotFound("批量预览不存在")
    if batch.uploaded_by != operated_by and not allow_admin:
        raise BulkImportNotFound("批量预览不存在")
    report = dict(batch.report_json or {})
    expected_token = str(report.get("token_hash") or "")
    actual_token = hashlib.sha256(preview_token.encode("utf-8")).hexdigest()
    if not hmac.compare_digest(expected_token, actual_token):
        raise BulkImportConflict("预览 token 无效")
    if not hmac.compare_digest(str(report.get("payload_hash") or ""), payload_hash):
        raise BulkImportConflict("预览 payload hash 不匹配")
    if not hmac.compare_digest(str(report.get("data_version") or ""), str(data_version)):
        raise BulkImportConflict("预览数据版本不匹配")
    selected = list(dict.fromkeys(str(key) for key in row_keys))
    if not selected or len(selected) != len(row_keys):
        raise BulkImportInvalid("必须选择至少一行，且 row_keys 不能重复")
    if batch.status == "success":
        applied = [str(key) for key in report.get("selected_row_keys") or []]
        if not applied or sorted(selected) != sorted(applied):
            raise BulkImportConflict("该批次已应用，row_keys 与原应用选择不一致")
        expected_selection_hash = _canonical_hash(
            {"payload_hash": payload_hash, "row_keys": sorted(applied)}
        )
        if not hmac.compare_digest(
            str(report.get("selection_hash") or ""),
            expected_selection_hash,
        ):
            raise BulkImportConflict("已应用批次的选择证据不完整")
        result = dict(report.get("result") or {})
        _enforce_project_scope(
            [
                {
                    "action": "replay",
                    "project_id": row.get("project_id"),
                }
                for row in result.get("rows") or []
            ],
            allowed_project_ids,
        )
        return result
    if batch.status != "processing":
        raise BulkImportConflict("该预览已失败或失效，请重新上传预览")
    expires_at = datetime.fromisoformat(str(report["expires_at"]))
    if expires_at <= datetime.now(timezone.utc):
        raise BulkImportConflict("预览已过期，请重新上传预览")
    row_map = report.get("row_map") or {}
    if any(key not in row_map for key in selected):
        raise BulkImportInvalid("包含不可提交、未知或已阻断的 row_key")
    public_by_key = {
        row["row_key"]: row for row in (report.get("public") or {}).get("rows") or []
    }
    if any(
        public_by_key.get(key, {}).get("row_status") != "ready"
        for key in selected
    ):
        raise BulkImportInvalid("只能提交预览状态为 ready 的行")

    plans = report.get("plans") or []
    selected_by_plan: dict[int, list[tuple[str, int]]] = defaultdict(list)
    for row_key in selected:
        mapping = row_map[row_key]
        selected_by_plan[int(mapping["plan_index"])].append(
            (row_key, int(mapping["operation_index"]))
        )

    selected_operations = [
        plans[plan_index]["plan"]["operations"][operation_index]
        for plan_index, mappings in selected_by_plan.items()
        for _row_key, operation_index in mappings
    ]
    _enforce_project_scope(selected_operations, allowed_project_ids)

    selection_hash = _canonical_hash(
        {"payload_hash": payload_hash, "row_keys": sorted(selected)}
    )
    _advisory_lock(db, f"maintenance-transfer-apply:{selection_hash}")
    already = db.scalar(
        select(SysImportBatch)
        .where(
            SysImportBatch.id != batch.id,
            SysImportBatch.file_type == TRANSFER_BATCH_TYPE,
            SysImportBatch.file_hash == selection_hash,
            SysImportBatch.status == "success",
        )
        .order_by(SysImportBatch.id)
        .limit(1)
    )
    if already is not None:
        already_report = dict(already.report_json or {})
        already_selected = [
            str(key) for key in already_report.get("selected_row_keys") or []
        ]
        if (
            sorted(already_selected) != sorted(selected)
            or not hmac.compare_digest(
                str(already_report.get("selection_hash") or ""),
                selection_hash,
            )
        ):
            raise BulkImportConflict("幂等批次的已应用行选择证据不一致")
        return dict(already_report.get("result") or {})

    project_ids: set[str] = set()
    audit_reason = f"全项目批量传输 batch={batch.id} payload={payload_hash}"
    for plan_index in sorted(selected_by_plan):
        plan_wrapper = plans[plan_index]
        adapter = _ADAPTERS.get(plan_wrapper["form_type"])
        if adapter is None:
            raise BulkImportConflict("预览使用的表单适配器已不存在")
        source_plan = plan_wrapper["plan"]
        op_indexes = [op_index for _key, op_index in selected_by_plan[plan_index]]
        subplan = {
            **source_plan,
            "operations": [source_plan["operations"][index] for index in op_indexes],
        }
        result = adapter.apply_plan(
            db,
            subplan,
            operated_by=operated_by,
            audit_reason=f"{audit_reason} file={plan_wrapper['file_id']}",
        )
        project_ids.update(result.get("project_ids") or [])

    result_rows: list[dict] = []
    for row_key in selected:
        public = public_by_key[row_key]
        mapping = row_map[row_key]
        wrapper = plans[int(mapping["plan_index"])]
        operation = wrapper["plan"]["operations"][int(mapping["operation_index"])]
        project_id = operation.get("project_id")
        project_contract_id = operation.get("project_contract_id")
        entity_id = project_contract_id
        report_month = operation.get("report_month")
        aggregate_key = None
        if wrapper["form_type"] == "sales_contract_amount":
            relation = db.scalar(
                select(MaintenanceProjectContract)
                .where(
                    MaintenanceProjectContract.contract_id
                    == operation["contract_id"],
                    MaintenanceProjectContract.contract_no
                    == operation["sales_order_no"],
                )
                .order_by(MaintenanceProjectContract.created_at.desc())
                .limit(1)
            )
            if relation is not None:
                project_id = relation.project_id
                project_contract_id = relation.project_contract_id
                entity_id = relation.project_contract_id
        else:
            aggregate_key = f"{operation['project_contract_id']}:{operation['report_month']}"
            snapshot = db.scalar(
                select(MaintenanceCollectionSnapshot).where(
                    MaintenanceCollectionSnapshot.project_contract_id
                    == operation["project_contract_id"],
                    MaintenanceCollectionSnapshot.report_month
                    == date.fromisoformat(operation["report_month"]),
                )
            )
            entity_id = snapshot.collection_id if snapshot is not None else None
        if project_id:
            project_ids.add(project_id)
        result_rows.append(
            {
                "row_key": row_key,
                "source_file": public["filename"],
                "source_sheet": public.get("detected_sheet"),
                "source_row": public["source_row"],
                "status": "applied",
                "action": public["action"],
                "project_id": project_id,
                "contract_id": operation.get("contract_id"),
                "entity_id": entity_id,
                "message": "已按冻结预览在同一事务中应用",
                "error_code": None,
                "before_version": operation.get("expected_contract_version"),
                "after_version": None,
                "aggregate_key": aggregate_key,
                "project_contract_id": project_contract_id,
                "report_month": report_month,
            }
        )

    result = {
        "batch_id": str(batch.id),
        "status": "done",
        "applied": len(result_rows),
        "skipped": 0,
        "blocked": 0,
        "project_ids": sorted(project_ids),
        "invalidated_projects": sorted(project_ids),
        "audit_ref": f"maintenance_bulk_import:{batch.id}",
        "rows": result_rows,
    }
    report["result"] = result
    report["selected_row_keys"] = selected
    report["selection_hash"] = selection_hash
    report["applied_at"] = datetime.now(timezone.utc).isoformat()
    report["applied_by"] = operated_by
    batch.file_hash = selection_hash
    batch.report_json = report
    batch.status = "success"
    batch.rows_inserted = len(result_rows)
    batch.rows_skipped = int(batch.rows_total or 0) - len(result_rows)
    batch.rows_error = 0
    db.add(
        SysAuditLog(
            entity_type="maintenance_bulk_import",
            entity_id=batch.id,
            action="apply",
            before_json={
                "status": "processing",
                "payload_hash": payload_hash,
                "data_version": data_version,
            },
            after_json={
                "status": "success",
                "selection_hash": selection_hash,
                "applied": len(result_rows),
                "project_ids": sorted(project_ids),
            },
            reason=audit_reason,
            operated_by=operated_by,
        )
    )
    try:
        db.commit()
    except Exception:
        db.rollback()
        raise
    return result
