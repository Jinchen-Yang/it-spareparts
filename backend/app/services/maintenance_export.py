"""维保订单 XLSX 导出。"""
from contextlib import suppress
from datetime import date, datetime, timezone
import re
import threading
from tempfile import SpooledTemporaryFile
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.writer.excel import ExcelWriter
from sqlalchemy import func, select, text
from sqlalchemy.orm import Session

from app import config
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder
from app.security import UserContext, apply_field_visibility, is_field_hidden
from app.services import maintenance_cost_quality


ORDER_HEADERS = (
    "数据库ID", "原始订单ID", "维保单号", "制单日期", "销售订单", "项目名", "终端客户",
    "需求类型", "业务类型", "销售人员", "出库仓库", "维保开始日期", "维保终止日期", "流程状态",
)
LINE_HEADERS = (
    "数据库ID", "原始明细ID", "订单数据库ID", "原始订单ID", "维保单号", "制单日期", "行号",
    "PN", "原始PN", "产品描述", "需求数量", "退货数量", "发货SN", "单价", "金额",
    "含税单位成本", "未税单位成本", "含税成本金额", "未税成本金额",
    "成本事实层级", "成本来源", "含税口径", "取价月", "追溯月数", "关联采购单", "距采购天数", "置信度",
    "参考侧", "参考池ID", "参考池版本", "参考样本数", "参考起始日", "参考截止日", "最近样本日",
    "异常标记",
)
_INVALID_XML_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]")
_COST_ANOMALY_FLAGS = maintenance_cost_quality.COST_DERIVED_ANOMALY_FLAGS
MAX_DATA_ROWS_PER_SHEET = 1_048_575
MAX_DYNAMIC_TEXT_BYTES = 64 * 1024 * 1024
MAX_WORKBOOK_BYTES = 256 * 1024 * 1024
ORDER_RENDERED_TEXT_OVERHEAD_BYTES = 16
LINE_RENDERED_TEXT_OVERHEAD_BYTES = 64
ORDER_EXPORT_ADVISORY_LOCK_KEY = 0x4D4F_584C
_ORDER_EXPORT_LOCK = threading.Lock()


class ExcelCellTooLong(ValueError):
    pass


class ExcelRowLimitExceeded(ValueError):
    pass


class ExcelExportTooLarge(ValueError):
    pass


class ExcelExportEmpty(ValueError):
    pass


class ExcelExportBusy(RuntimeError):
    pass


class _SizeLimitedFile:
    """给 openpyxl/ZipFile 提供可 seek 的硬上限，禁止先写爆临时盘再拒绝。"""

    def __init__(self, raw, max_size: int):
        self._raw = raw
        self._max_size = max_size
        self._extent = raw.tell()

    def write(self, data):
        end = self._raw.tell() + len(data)
        if max(self._extent, end) > self._max_size:
            raise ExcelExportTooLarge("维保订单 XLSX 超过 256 MiB 上限")
        written = self._raw.write(data)
        self._extent = max(self._extent, self._raw.tell())
        return written

    def seek(self, *args):
        return self._raw.seek(*args)

    def tell(self):
        return self._raw.tell()

    def flush(self):
        return self._raw.flush()


def _save_workbook(workbook: Workbook, output) -> None:
    """显式持有 ZipFile，确保超限异常时也关闭 archive，不遗留析构期写入。"""
    archive = ZipFile(
        _SizeLimitedFile(output, MAX_WORKBOOK_BYTES),
        mode="w",
        compression=ZIP_DEFLATED,
        allowZip64=True,
    )
    workbook.properties.modified = datetime.now(timezone.utc).replace(tzinfo=None)
    try:
        ExcelWriter(workbook, archive).save()
    finally:
        if archive.fp is not None:
            archive.close()


def _safe_row(values: tuple) -> tuple:
    safe = []
    for value in values:
        if isinstance(value, str):
            value = _INVALID_XML_CONTROL.sub("", value)
            if value[:1] in ("=", "+", "-", "@", "\t", "\r", "\n"):
                value = "'" + value
            if len(value) > 32767:
                raise ExcelCellTooLong("文本超过 Excel 单元格上限 32767 个字符")
        safe.append(value)
    return tuple(safe)


def _formatted_row(worksheet, values: tuple, formats: dict[int, str]) -> tuple:
    cells = []
    for index, value in enumerate(_safe_row(values)):
        cell = WriteOnlyCell(worksheet, value=value)
        if index in formats:
            cell.number_format = formats[index]
        cells.append(cell)
    return tuple(cells)


def _cleanup_workbook(workbook: Workbook) -> None:
    for worksheet in workbook.worksheets:
        with suppress(Exception):
            if not worksheet.closed:
                worksheet.close()
        writer = getattr(worksheet, "_writer", None)
        if writer is None or getattr(writer, "out", None) is None:
            continue
        with suppress(Exception):
            writer.close()
        with suppress(Exception):
            writer.cleanup()


def _date_filters(date_from: date | None, date_to: date | None) -> tuple:
    filters = (FMaintenanceOrder.data_status == config.ACTIVE_STATUS,)
    if date_from is not None and date_to is not None:
        filters += (
            FMaintenanceOrder.order_date >= date_from,
            FMaintenanceOrder.order_date <= date_to,
        )
    return filters


def _octet_length(value):
    """在 PostgreSQL 端统计即将进入 XLSX 的动态文本 UTF-8 字节。"""
    return func.octet_length(func.coalesce(value, ""))


def _preflight_row_limits(
    db: Session,
    date_from: date | None,
    date_to: date | None,
) -> None:
    filters = _date_filters(date_from, date_to)
    rendered_project = func.coalesce(
        func.nullif(FMaintenanceOrder.project_raw, ""),
        FMaintenanceOrder.project_std,
        "",
    )
    order_text_bytes = (
        _octet_length(FMaintenanceOrder.raw_order_id)
        + _octet_length(FMaintenanceOrder.order_no)
        + _octet_length(FMaintenanceOrder.linked_sales_order_no)
        + _octet_length(rendered_project)
        + _octet_length(FMaintenanceOrder.end_customer)
        + _octet_length(FMaintenanceOrder.demand_type)
        + _octet_length(FMaintenanceOrder.business_type)
        + _octet_length(FMaintenanceOrder.salesperson)
        + _octet_length(FMaintenanceOrder.warehouse)
        + _octet_length(FMaintenanceOrder.data_status)
        + ORDER_RENDERED_TEXT_OVERHEAD_BYTES
    )
    order_count, order_dynamic_text_bytes = db.execute(
        select(
            func.count(FMaintenanceOrder.id),
            func.coalesce(func.sum(order_text_bytes), 0),
        ).where(*filters)
    ).one()
    if order_count == 0:
        raise ExcelExportEmpty("所选范围内没有可导出的维保订单")
    if order_count > MAX_DATA_ROWS_PER_SHEET:
        raise ExcelRowLimitExceeded("维保订单超过 Excel 单 Sheet 数据行上限 1048575")

    line_text_bytes = (
        _octet_length(FMaintenanceLine.raw_line_id)
        + _octet_length(FMaintenanceOrder.raw_order_id)
        + _octet_length(FMaintenanceOrder.order_no)
        + _octet_length(FMaintenanceLine.pn_std)
        + _octet_length(FMaintenanceLine.pn_raw)
        + _octet_length(FMaintenanceLine.description)
        + _octet_length(FMaintenanceLine.serial_numbers)
        + _octet_length(FMaintenanceLine.cost_source)
        + _octet_length(FMaintenanceLine.cost_tax_basis)
        + _octet_length(FMaintenanceLine.price_month)
        + _octet_length(FMaintenanceLine.linked_purchase_order_no)
        + _octet_length(FMaintenanceLine.confidence)
        + _octet_length(FMaintenanceLine.reference_side)
        + _octet_length(func.array_to_string(FMaintenanceLine.anomaly_flags, "、"))
        + LINE_RENDERED_TEXT_OVERHEAD_BYTES
    )
    line_count, line_dynamic_text_bytes = db.execute(
        select(
            func.count(FMaintenanceLine.id),
            func.coalesce(func.sum(line_text_bytes), 0),
        )
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .where(*filters)
    ).one()
    if line_count > MAX_DATA_ROWS_PER_SHEET:
        raise ExcelRowLimitExceeded("订单明细超过 Excel 单 Sheet 数据行上限 1048575")
    if (
        int(order_dynamic_text_bytes) + int(line_dynamic_text_bytes)
        > MAX_DYNAMIC_TEXT_BYTES
    ):
        raise ExcelExportTooLarge("维保订单 XLSX 动态文本超过 64 MiB 上限")


def _build_workbook(
    db: Session,
    user_ctx: UserContext,
    date_from: date | None = None,
    date_to: date | None = None,
) -> SpooledTemporaryFile:
    _preflight_row_limits(db, date_from, date_to)
    workbook = Workbook(write_only=True)
    output = None
    completed = False
    try:
        orders = workbook.create_sheet("维保订单")
        lines = workbook.create_sheet("订单明细")
        orders.append(ORDER_HEADERS)
        lines.append(LINE_HEADERS)

        statement = (
            select(FMaintenanceOrder, FMaintenanceLine)
            .outerjoin(FMaintenanceLine, FMaintenanceLine.order_id == FMaintenanceOrder.id)
            .where(*_date_filters(date_from, date_to))
            .order_by(FMaintenanceOrder.id, FMaintenanceLine.id)
            .execution_options(stream_results=True, yield_per=1000)
        )
        previous_order_id = None
        order_count = 0
        line_count = 0
        for order, line in db.execute(statement):
            if order.id != previous_order_id:
                order_count += 1
                if order_count > MAX_DATA_ROWS_PER_SHEET:
                    raise ExcelRowLimitExceeded("维保订单超过 Excel 单 Sheet 数据行上限 1048575")
                visible_order = apply_field_visibility({
                    "end_customer": order.end_customer,
                }, user_ctx)
                orders.append(_formatted_row(orders, (
                    order.id, order.raw_order_id, order.order_no, order.order_date,
                    order.linked_sales_order_no, order.project_raw or order.project_std,
                    visible_order["end_customer"], order.demand_type, order.business_type, order.salesperson,
                    order.warehouse, order.maint_start, order.maint_end, order.data_status,
                ), {3: "yyyy-mm-dd", 11: "yyyy-mm-dd", 12: "yyyy-mm-dd"}))
                previous_order_id = order.id
            if line is not None:
                line_count += 1
                if line_count > MAX_DATA_ROWS_PER_SHEET:
                    raise ExcelRowLimitExceeded("订单明细超过 Excel 单 Sheet 数据行上限 1048575")
                cost_tier = maintenance_cost_quality.source_tier(
                    line.cost_source,
                    line.cost_tax_basis,
                    line.cost_amount,
                )
                has_known_cost = cost_tier != "missing"
                cost = apply_field_visibility({
                    "unit_cost": line.unit_cost if has_known_cost else None,
                    "cost_amount": line.cost_amount if has_known_cost else None,
                    "unit_cost_inc_tax": line.unit_cost_inc_tax if has_known_cost else None,
                    "unit_cost_ex_tax": line.unit_cost_ex_tax if has_known_cost else None,
                    "cost_amount_inc_tax": line.cost_amount_inc_tax if has_known_cost else None,
                    "cost_amount_ex_tax": line.cost_amount_ex_tax if has_known_cost else None,
                    "cost_tier": cost_tier,
                    "cost_source": line.cost_source,
                    "cost_tax_basis": line.cost_tax_basis,
                    "price_month": line.price_month,
                    "trace_months": line.trace_months,
                    "linked_purchase_order_no": line.linked_purchase_order_no,
                    "price_distance_days": line.price_distance_days,
                    "confidence": line.confidence,
                    "reference_side": line.reference_side,
                    "reference_pool_group_id": line.reference_pool_group_id,
                    "reference_pool_version": line.reference_pool_version,
                    "reference_sample_count": line.reference_sample_count,
                    "reference_from_date": line.reference_from_date,
                    "reference_to_date": line.reference_to_date,
                    "reference_latest_date": line.reference_latest_date,
                }, user_ctx)
                anomaly_flags = line.anomaly_flags or []
                if is_field_hidden(user_ctx, "unit_cost"):
                    anomaly_flags = [flag for flag in anomaly_flags if flag not in _COST_ANOMALY_FLAGS]
                lines.append(_formatted_row(lines, (
                    line.id, line.raw_line_id, order.id, order.raw_order_id, order.order_no,
                    order.order_date, line.line_no, line.pn_std, line.pn_raw, line.description,
                    line.qty, line.return_qty, line.serial_numbers, cost["unit_cost"], cost["cost_amount"],
                    cost["unit_cost_inc_tax"], cost["unit_cost_ex_tax"],
                    cost["cost_amount_inc_tax"], cost["cost_amount_ex_tax"],
                    {"actual": "实际采购参考", "estimated": "估算参考", "missing": "成本缺失"}
                    .get(cost["cost_tier"], cost["cost_tier"]),
                    cost["cost_source"], cost["cost_tax_basis"], cost["price_month"], cost["trace_months"],
                    cost["linked_purchase_order_no"], cost["price_distance_days"], cost["confidence"],
                    cost["reference_side"], cost["reference_pool_group_id"],
                    cost["reference_pool_version"], cost["reference_sample_count"],
                    cost["reference_from_date"], cost["reference_to_date"],
                    cost["reference_latest_date"],
                    "、".join(anomaly_flags),
                ), {
                    5: "yyyy-mm-dd", 10: "0.00", 11: "0.00",
                    13: "#,##0.00", 14: "#,##0.00",
                    15: "#,##0.00", 16: "#,##0.00",
                    17: "#,##0.00", 18: "#,##0.00",
                    31: "yyyy-mm-dd", 32: "yyyy-mm-dd", 33: "yyyy-mm-dd",
                }))

        output = SpooledTemporaryFile(max_size=5 * 1024 * 1024, mode="w+b")
        _save_workbook(workbook, output)
        output.seek(0)
        completed = True
        return output
    finally:
        if not completed:
            if output is not None:
                output.close()
            _cleanup_workbook(workbook)
        workbook.close()


def _acquire_shared_source_lock(db: Session) -> None:
    """冻结导出事实快照，覆盖资源预检到工作簿完整物化的整个事务。"""
    db.execute(
        text("SELECT pg_advisory_xact_lock_shared(:k)"),
        {"k": config.DATA_CHANGE_ADVISORY_LOCK_KEY},
    )


def build_workbook(
    db: Session,
    user_ctx: UserContext,
    date_from: date | None = None,
    date_to: date | None = None,
) -> SpooledTemporaryFile:
    """跨进程非阻塞构建逐单工作簿，避免并发任务叠加临时磁盘和内存。"""
    if not _ORDER_EXPORT_LOCK.acquire(blocking=False):
        raise ExcelExportBusy("已有逐单维保导出正在执行，请稍后重试")
    try:
        acquired = db.scalar(
            text("SELECT pg_try_advisory_xact_lock(:k)"),
            {"k": ORDER_EXPORT_ADVISORY_LOCK_KEY},
        )
        if not acquired:
            raise ExcelExportBusy("已有逐单维保导出正在执行，请稍后重试")
        _acquire_shared_source_lock(db)
        return _build_workbook(db, user_ctx, date_from, date_to)
    finally:
        _ORDER_EXPORT_LOCK.release()
