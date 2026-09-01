"""项目成本工作簿渲染器。

本模块只负责把已经查询好的合同工作簿数据渲染成 openpyxl Workbook。
查询、权限、批量选择、临时文件与 ZIP 生命周期由上层导出服务负责。
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app import config


SOURCE_LABELS = {
    "direct": "实际·专属采购",
    "window": "实际·±7天最近价",
    "month_avg": "实际·当月均价",
    "trace_avg": "预估·追溯均价",
    "sales_ref": "预估·销售参考",
    "pool_purchase": "预估·互通池采购均价",
    "pool_sales": "预估·互通池销售均价",
    "purchase_history": "预估·本PN采购历史",
    "sales_history": "预估·本PN销售历史",
    "manual": "实际·人工回填",
    "none": "成本缺失",
}
CONFIDENCE_LABELS = {"high": "高", "medium": "中", "low": "低"}
FEE_CATEGORY_HEADER_PREFIX = "费用分类："
_HDR_FILL = PatternFill("solid", fgColor="35506B")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=15)
_SUB_FONT = Font(color="8C8C8C", size=10)
_KV_FILL = PatternFill("solid", fgColor="EFEBE3")
_ALT_FILL = PatternFill("solid", fgColor="F7F4EE")
_DOC_FILL = PatternFill("solid", fgColor="FDF3D7")
_TOTAL_FILL = PatternFill("solid", fgColor="E8E2D6")
_THIN = Side(style="thin", color="D8D2C6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0.00"
_PERCENT = "0.00%"
_CENTER = Alignment(horizontal="center", vertical="center")
_STATUS_STYLE = {
    "incomplete_cost": ("成本不完整，需补数据", "8C8C8C"),
    "expense_data_unavailable": ("费用数据未就绪", "8C8C8C"),
    "filtered_scope": ("日期筛选下不计算合同预算余量/红黄绿", "8C8C8C"),
    "red": ("预算已用完或超预算", "C0524A"),
    "yellow": ("预算余量≤20%", "B8860B"),
    "green": ("预算余量>20%", "3F7A45"),
    "no_budget": ("无预算", "8C8C8C"),
}
_QUALITY_LABELS = {
    "actual_only": "完整：仅实际采购参考",
    "contains_estimate": "完整：含估算参考",
    "incomplete": "成本不完整，需补数据",
}
_TIER_LABELS = {
    "actual": "实际采购参考",
    "estimated": "估算参考",
    "missing": "成本缺失",
}
_PARTS_PROFIT_STATUS_LABELS = {
    "complete_actual": "完整：仅实际成本",
    "complete_estimated": "完整：含估算成本",
    "missing_revenue": "合同收入缺失",
    "missing_tax_rate": "合同税率缺失",
    "invalid_tax_rate": "合同税率异常",
    "ambiguous_revenue": "重复合同收入冲突",
    "incomplete_cost": "成本不完整",
    "filtered_scope": "日期筛选下暂不计算",
}
_CONTRIBUTION_STATUS_LABELS = {
    **_PARTS_PROFIT_STATUS_LABELS,
    "complete": "完整",
    "expense_tax_unknown": "费用税务口径缺失",
    "expense_data_unavailable": "费用数据未就绪",
}


def _hdr_row(ws, row: int, ncols: int) -> None:
    for column in range(1, ncols + 1):
        cell = ws.cell(row=row, column=column)
        cell.fill, cell.font = _HDR_FILL, _HDR_FONT
        cell.border, cell.alignment = _BORDER, _CENTER


def _col_widths(ws, widths: list[float]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def render_contract_workbook(
    contract: str,
    data: dict,
    safe_text,
    *,
    workbook: Workbook | None = None,
) -> Workbook:
    """把合同工作簿数据渲染成四个 Sheet。"""
    owns_workbook = workbook is None
    workbook = workbook or Workbook()
    try:
        return _populate_contract_workbook(contract, data, safe_text, workbook)
    except BaseException:
        if owns_workbook:
            workbook.close()
        raise


def _populate_contract_workbook(
    contract: str,
    data: dict,
    safe_text,
    workbook: Workbook,
) -> Workbook:
    safe_contract = safe_text(contract)
    budget = data["budget"]
    cost_summary = data["cost_summary"]
    dual_cost_summary = data["dual_cost_summary"]
    margin = data.get("margin")
    decision = data["decision"]
    date_filtered = data.get("date_filtered") is True
    expense_data_available = data.get("expense_data_available") is True
    order_count = int(data.get("order_count", len(data.get("orders", []))))
    orders_with_details = int(data.get(
        "orders_with_details",
        len({order.id for _line, order in data["lines"]}),
    ))
    missing_detail_orders = int(data.get(
        "missing_detail_orders",
        max(order_count - orders_with_details, 0),
    ))
    structure_complete = (
        data.get("structure_complete", missing_detail_orders == 0) is True
        and missing_detail_orders == 0
    )
    has_parts_lines = bool(data["lines"])
    spent_parts = (
        float(cost_summary["known_cost_total"])
        if has_parts_lines else None
    )
    expense_inc = data.get("expense_inc")
    expense_ex = data.get("expense_ex")
    expense_evidence_status = data.get(
        "expense_evidence_status",
        "expense_data_unavailable",
    )
    expense_status_label = {
        "complete": "完整",
        "expense_tax_unknown": "费用税务口径缺失",
        "expense_data_unavailable": "未就绪（无记录不等于0）",
    }.get(expense_evidence_status, "未就绪（无记录不等于0）")
    spent = (
        float(decision["known_spend_total"])
        if has_parts_lines and expense_data_available else None
    )
    remaining = (
        float(decision["remaining"])
        if decision["remaining"] is not None else None
    )
    status = decision["decision_status"]
    status_label, status_color = _STATUS_STYLE[status]
    quality_label = _QUALITY_LABELS[cost_summary["cost_quality"]]
    tax_status = data.get("contract_tax_status", "missing")
    if tax_status == "available" and data.get("contract_tax_rate") is not None:
        tax_rate_value = float(data["contract_tax_rate"])
        tax_rate_format = _PERCENT
    else:
        tax_rate_value = {
            "ambiguous": "冲突",
            "invalid": "异常",
        }.get(tax_status, "—")
        tax_rate_format = None

    budget_sheet = workbook.active
    budget_sheet.title = "项目预算"
    budget_sheet.merge_cells("A1:F1")
    budget_sheet["A1"] = safe_text(f"维保合同成本工作簿 · {contract}")
    budget_sheet["A1"].font = _TITLE_FONT
    budget_sheet.row_dimensions[1].height = 26
    budget_sheet.merge_cells("A2:F2")
    scope_note = (
        "所选期间成本/报销事实仅供参考 · "
        "日期筛选下不计算合同预算余量/红黄绿 · "
        if date_filtered
        else "成本或费用数据未就绪时不计算预算余量或红黄绿 · "
    )
    budget_sheet["A2"] = (
        "实际、估算、缺失分层同源 · 已知金额为含税/不含税混合原值参考 · "
        f"{scope_note}导出自 IT 备件智能管理系统"
    )
    budget_sheet["A2"].font = _SUB_FONT

    key_values = [
        ("合同（销售订单）", safe_contract, None,
         "成本完整性", quality_label, None),
        ("命中维保订单", order_count, None,
         "有明细订单", orders_with_details, None),
        ("无明细订单", missing_detail_orders, None,
         "订单结构完整性",
         (
             "完整"
             if structure_complete
             else "不完整：存在无配件明细订单"
         ),
         None),
        ("合同金额（含税参考）",
         float(budget) if budget is not None else "（销售表未找到）",
         _MONEY if budget is not None else None,
         "预算消耗参考状态", status_label, None),
        ("实际采购参考（含税）",
         float(cost_summary["actual_cost_inc"]) if has_parts_lines else None,
         _MONEY if has_parts_lines else None,
         "实际参考行", cost_summary["actual_lines"], None),
        ("实际采购参考（不含税）",
         float(cost_summary["actual_cost_ex"]) if has_parts_lines else None,
         _MONEY if has_parts_lines else None,
         "税率", tax_rate_value, tax_rate_format),
        ("估算参考（含税）",
         float(cost_summary["estimated_cost_inc"]) if has_parts_lines else None,
         _MONEY if has_parts_lines else None,
         "估算参考行", cost_summary["estimated_lines"], None),
        ("估算参考（不含税）",
         float(cost_summary["estimated_cost_ex"]) if has_parts_lines else None,
         _MONEY if has_parts_lines else None,
         "缺失成本行", cost_summary["missing_cost_lines"], None),
        ("备件成本（含税归一）",
         (
             float(dual_cost_summary["parts_cost_inc_tax"])
             if dual_cost_summary["parts_cost_inc_tax"] is not None else None
         ),
         _MONEY if dual_cost_summary["parts_cost_inc_tax"] is not None else None,
         "含税口径质量",
         _QUALITY_LABELS[dual_cost_summary["parts_cost_inc_tax_quality"]], None),
        ("备件成本（未税归一）",
         (
             float(dual_cost_summary["parts_cost_ex_tax"])
             if dual_cost_summary["parts_cost_ex_tax"] is not None else None
         ),
         _MONEY if dual_cost_summary["parts_cost_ex_tax"] is not None else None,
         "未税口径质量",
         _QUALITY_LABELS[dual_cost_summary["parts_cost_ex_tax_quality"]], None),
        ("已知备件成本参考（混合原值）", spent_parts,
         _MONEY if spent_parts is not None else None,
         "费用证据状态", expense_status_label, None),
        (
         "报销费用（含税）",
         float(expense_inc) if expense_inc is not None else "—",
         _MONEY if expense_inc is not None else None,
         "报销费用（未税）",
         float(expense_ex) if expense_ex is not None else "—",
         _MONEY if expense_ex is not None else None,
        ),
        (
         "所选期间支出参考（备件+报销）"
         if date_filtered else "完整项目支出参考（备件+报销）",
         spent if spent is not None else "—",
         _MONEY if spent is not None else None,
         "剩余预算", remaining if remaining is not None else "—",
         _MONEY if remaining is not None else None),
    ]
    if margin is not None:
        def margin_value(
            field: str,
            number_format: str | None = _MONEY,
            *,
            allowed: bool = True,
        ):
            value = margin.get(field) if allowed else None
            return (
                float(value) if value is not None else "—",
                number_format if value is not None else None,
            )

        complete_parts_statuses = {"complete_actual", "complete_estimated"}
        parts_status_inc = margin.get("parts_profit_status_inc")
        parts_status_ex = margin.get("parts_profit_status_ex")
        parts_complete_inc = parts_status_inc in complete_parts_statuses
        parts_complete_ex = parts_status_ex in complete_parts_statuses
        revenue_inc, revenue_inc_format = margin_value(
            "revenue_inc",
            allowed=parts_complete_inc
            or parts_status_inc in {"incomplete_cost", "filtered_scope"},
        )
        revenue_ex, revenue_ex_format = margin_value(
            "revenue_ex",
            allowed=parts_complete_ex
            or parts_status_ex in {"incomplete_cost", "filtered_scope"},
        )
        parts_profit_inc, parts_profit_inc_format = margin_value(
            "parts_gross_profit_inc",
            allowed=parts_complete_inc,
        )
        parts_profit_ex, parts_profit_ex_format = margin_value(
            "parts_gross_profit_ex",
            allowed=parts_complete_ex,
        )
        parts_margin_inc, parts_margin_inc_format = margin_value(
            "parts_gross_margin_inc",
            _PERCENT,
            allowed=parts_complete_inc,
        )
        parts_margin_ex, parts_margin_ex_format = margin_value(
            "parts_gross_margin_ex",
            _PERCENT,
            allowed=parts_complete_ex,
        )
        contribution_inc, contribution_inc_format = margin_value(
            "contribution_profit_inc",
            allowed=(
                parts_complete_inc
                and margin.get("contribution_status_inc") == "complete"
            ),
        )
        contribution_ex, contribution_ex_format = margin_value(
            "contribution_profit_ex",
            allowed=(
                parts_complete_ex
                and margin.get("contribution_status_ex") == "complete"
            ),
        )
        contribution_margin_inc, contribution_margin_inc_format = margin_value(
            "contribution_margin_inc",
            _PERCENT,
            allowed=(
                parts_complete_inc
                and margin.get("contribution_status_inc") == "complete"
            ),
        )
        contribution_margin_ex, contribution_margin_ex_format = margin_value(
            "contribution_margin_ex",
            _PERCENT,
            allowed=(
                parts_complete_ex
                and margin.get("contribution_status_ex") == "complete"
            ),
        )
        key_values.extend([
            (
                "合同收入（含税）",
                revenue_inc,
                revenue_inc_format,
                "合同收入（未税）",
                revenue_ex,
                revenue_ex_format,
            ),
            (
                "合同级备件毛利（含税）",
                parts_profit_inc,
                parts_profit_inc_format,
                "合同级备件毛利（未税）",
                parts_profit_ex,
                parts_profit_ex_format,
            ),
            (
                "合同级备件毛利率（含税）",
                parts_margin_inc,
                parts_margin_inc_format,
                "合同级备件毛利率（未税）",
                parts_margin_ex,
                parts_margin_ex_format,
            ),
            (
                "合同级贡献毛利（含税）",
                contribution_inc,
                contribution_inc_format,
                "合同级贡献毛利（未税）",
                contribution_ex,
                contribution_ex_format,
            ),
            (
                "合同级贡献毛利率（含税）",
                contribution_margin_inc,
                contribution_margin_inc_format,
                "合同级贡献毛利率（未税）",
                contribution_margin_ex,
                contribution_margin_ex_format,
            ),
            (
                "含税备件毛利状态",
                _PARTS_PROFIT_STATUS_LABELS.get(
                    margin.get("parts_profit_status_inc"),
                    margin.get("parts_profit_status_inc") or "—",
                ),
                None,
                "未税备件毛利状态",
                _PARTS_PROFIT_STATUS_LABELS.get(
                    margin.get("parts_profit_status_ex"),
                    margin.get("parts_profit_status_ex") or "—",
                ),
                None,
            ),
            (
                "含税贡献毛利状态",
                _CONTRIBUTION_STATUS_LABELS.get(
                    margin.get("contribution_status_inc"),
                    margin.get("contribution_status_inc") or "—",
                ),
                None,
                "未税贡献毛利状态",
                _CONTRIBUTION_STATUS_LABELS.get(
                    margin.get("contribution_status_ex"),
                    margin.get("contribution_status_ex") or "—",
                ),
                None,
            ),
        ])
    row = 4
    for left_label, left_value, left_format, right_label, right_value, right_format in key_values:
        for label, value, number_format, label_column in (
            (left_label, left_value, left_format, 1),
            (right_label, right_value, right_format, 3),
        ):
            label_cell = budget_sheet.cell(row=row, column=label_column, value=label)
            value_cell = budget_sheet.cell(row=row, column=label_column + 1, value=value)
            label_cell.fill, label_cell.font, label_cell.border = (
                _KV_FILL,
                Font(bold=True),
                _BORDER,
            )
            value_cell.border = _BORDER
            if number_format:
                value_cell.number_format = number_format
            if label == "剩余预算" and remaining is not None:
                value_cell.font = Font(bold=True, color=status_color)
            if label == "预算消耗参考状态":
                value_cell.fill = PatternFill("solid", fgColor=status_color)
                value_cell.font = Font(bold=True, color="FFFFFF")
                value_cell.alignment = _CENTER
        row += 1

    row += 1
    categories = sorted({
        category
        for month in data["monthly_expenses"].values()
        for category in month
    })
    expense_columns = [
        (
            f"{FEE_CATEGORY_HEADER_PREFIX}{category}"
            if expense_data_available
            else f"当前已导入报销（非全量）·{category}"
        )
        for category in categories
    ]
    columns = [
        "月份",
        "已知备件成本参考（混合原值·兼容）",
        *expense_columns,
        (
            "当月合计"
            if expense_data_available
            else "当月已知合计（费用非全量）"
        ),
    ]
    for index, heading in enumerate(columns, 1):
        budget_sheet.cell(row=row, column=index, value=safe_text(heading))
    _hdr_row(budget_sheet, row, len(columns))
    band = False
    totals = [0.0] * (len(columns) - 2)
    year_months = sorted(
        set(data["monthly_parts"]) | set(data["monthly_expenses"]),
    )
    for year_month in year_months:
        row += 1
        band = not band
        month_expenses = data["monthly_expenses"].get(year_month, {})
        values = [
            (
                float(data["monthly_parts"].get(year_month, 0))
                if has_parts_lines else None
            )
        ]
        values.extend(
            float(month_expenses.get(category, 0))
            for category in categories
        )
        for index, value in enumerate(values):
            if value is not None:
                totals[index] += value
        row_values = [
            safe_text(year_month),
            *values,
            round(sum(value or 0 for value in values), 2),
        ]
        for index, value in enumerate(row_values, 1):
            cell = budget_sheet.cell(row=row, column=index, value=value)
            cell.border = _BORDER
            if band:
                cell.fill = _ALT_FILL
            if index > 1:
                cell.number_format = _MONEY
    row += 1
    total_row = [
        (
            "合计"
            if expense_data_available
            else "当前已知合计（费用非全量）"
        ),
        *[
            None if index == 0 and not has_parts_lines else round(total, 2)
            for index, total in enumerate(totals)
        ],
        round(sum(totals), 2),
    ]
    for index, value in enumerate(total_row, 1):
        cell = budget_sheet.cell(row=row, column=index, value=value)
        cell.font, cell.fill, cell.border = Font(bold=True), _TOTAL_FILL, _BORDER
        if index > 1:
            cell.number_format = _MONEY
    _col_widths(
        budget_sheet,
        [24, 16] + [14] * (len(categories) + 1),
    )

    parts_sheet = workbook.create_sheet("备件明细-氚云")
    part_headers = [
        "数据标题(WBDD单号)", "制单日期", "销售订单", "项目名", "需求类型",
        "出库仓库", "销售人员", "业务类型", "序号", "需供货产品", "产品描述",
        "需求数量", "已知成本参考", "单价", "合计", "发货SN", "行成本单价",
        "行成本金额", "含税单位成本", "未税单位成本", "含税成本金额", "未税成本金额",
        "成本来源", "置信度", "取价月", "距采购天数", "含税口径",
        "成本事实层级", "参考侧", "参考池ID", "参考池版本", "参考样本数",
        "参考起始日", "参考截止日", "最近样本日",
    ]
    parts_sheet.append(part_headers)
    _hdr_row(parts_sheet, 1, len(part_headers))
    parts_sheet.freeze_panes = "A2"
    previous_order, band = None, False
    rendered_order_ids: set[int] = set()
    for line, order in data["lines"]:
        rendered_order_ids.add(order.id)
        first = order.order_no != previous_order
        if first:
            band = not band
        previous_order = order.order_no
        document_cost = data["doc_total"].get(order.order_no)
        display = data.get("line_cost_display", {}).get(line.id)
        tier = (
            display["tier"]
            if display is not None
            else data["line_cost_tiers"][line.id]
        )
        known_line = tier != "missing"
        manual_fallback = bool(display and display.get("manual_fallback"))
        source = display["source"] if display is not None else line.cost_source
        confidence = (
            display["confidence"] if display is not None else line.confidence
        )
        tax_basis = (
            display["tax_basis"] if display is not None else line.cost_tax_basis
        )

        def displayed_cost(field: str):
            if not known_line:
                return None
            return (
                display[field]
                if display is not None
                else getattr(line, field)
            )

        parts_sheet.append([
            safe_text(order.order_no),
            order.order_date.isoformat() if order.order_date else None,
            safe_text(order.linked_sales_order_no),
            safe_text(order.project_raw or order.project_std),
            safe_text(order.demand_type),
            safe_text(order.warehouse),
            safe_text(order.salesperson),
            safe_text(order.business_type),
            line.line_no,
            safe_text(line.pn_std),
            safe_text(line.description),
            float(line.qty) if line.qty is not None else None,
            float(document_cost) if first and document_cost is not None else None,
            None,
            None,
            safe_text(line.serial_numbers),
            float(displayed_cost("unit_cost"))
            if displayed_cost("unit_cost") is not None else None,
            float(displayed_cost("cost_amount"))
            if displayed_cost("cost_amount") is not None else None,
            float(displayed_cost("unit_cost_inc_tax"))
            if displayed_cost("unit_cost_inc_tax") is not None else None,
            float(displayed_cost("unit_cost_ex_tax"))
            if displayed_cost("unit_cost_ex_tax") is not None else None,
            float(displayed_cost("cost_amount_inc_tax"))
            if displayed_cost("cost_amount_inc_tax") is not None else None,
            float(displayed_cost("cost_amount_ex_tax"))
            if displayed_cost("cost_amount_ex_tax") is not None else None,
            safe_text(SOURCE_LABELS.get(source, source)),
            safe_text(CONFIDENCE_LABELS.get(confidence, confidence or "")),
            safe_text(None if manual_fallback else line.price_month),
            None if manual_fallback else line.price_distance_days,
            safe_text(tax_basis),
            safe_text(_TIER_LABELS[tier]),
            safe_text(None if manual_fallback else line.reference_side),
            None if manual_fallback else line.reference_pool_group_id,
            None if manual_fallback else line.reference_pool_version,
            None if manual_fallback else line.reference_sample_count,
            line.reference_from_date.isoformat()
            if not manual_fallback and line.reference_from_date else None,
            line.reference_to_date.isoformat()
            if not manual_fallback and line.reference_to_date else None,
            line.reference_latest_date.isoformat()
            if not manual_fallback and line.reference_latest_date else None,
        ])
        rendered_row = parts_sheet.max_row
        for column in range(1, len(part_headers) + 1):
            cell = parts_sheet.cell(row=rendered_row, column=column)
            cell.border = _BORDER
            if band:
                cell.fill = _ALT_FILL
        for column in (13, 14, 15, 17, 18, 19, 20, 21, 22):
            parts_sheet.cell(row=rendered_row, column=column).number_format = _MONEY
        if parts_sheet.cell(row=rendered_row, column=13).value is not None:
            document_cost_cell = parts_sheet.cell(row=rendered_row, column=13)
            document_cost_cell.fill = _DOC_FILL
            document_cost_cell.font = Font(bold=True)
        if confidence == "low":
            parts_sheet.cell(row=rendered_row, column=24).font = Font(
                color="B8860B",
                bold=True,
            )
    for order in data.get("orders", []):
        if order.id in rendered_order_ids:
            continue
        band = not band
        parts_sheet.append([
            safe_text(order.order_no),
            order.order_date.isoformat() if order.order_date else None,
            safe_text(order.linked_sales_order_no),
            safe_text(order.project_raw or order.project_std),
            safe_text(order.demand_type),
            safe_text(order.warehouse),
            safe_text(order.salesperson),
            safe_text(order.business_type),
            None,
            None,
            "暂无配件明细",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "成本缺失",
            None,
            None,
            None,
            None,
            "成本缺失",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ])
        rendered_row = parts_sheet.max_row
        for column in range(1, len(part_headers) + 1):
            cell = parts_sheet.cell(row=rendered_row, column=column)
            cell.border = _BORDER
            if band:
                cell.fill = _ALT_FILL
    parts_sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(part_headers))}{parts_sheet.max_row}"
    )
    _col_widths(
        parts_sheet,
        [
            20, 11, 16, 26, 10, 12, 9, 10, 6, 20, 36, 9, 13, 9, 9, 18, 11,
            12, 12, 12, 12, 12, 16, 8, 9, 11, 9, 12, 9, 10, 10, 10, 11, 11, 11,
        ],
    )

    expense_sheet = workbook.create_sheet("报销明细")
    expense_sheet.cell(row=1, column=1, value="销售订单").font = Font(bold=True)
    expense_sheet.cell(row=1, column=1).fill = _KV_FILL
    expense_sheet.cell(row=1, column=2, value=safe_contract).font = Font(bold=True)
    for column in (1, 2):
        expense_sheet.cell(row=1, column=column).border = _BORDER
    expense_headers = [
        "报销日期", "报销人员", "报销类别", "费用分类", "支出事由",
        "报销金额", "报销金额（含税）", "报销金额（未税）",
        "费用证据状态", "流程状态", "单号", "序号",
    ]
    amount_column = 6
    for column, heading in enumerate(expense_headers, 1):
        expense_sheet.cell(row=2, column=column, value=heading)
    _hdr_row(expense_sheet, 2, len(expense_headers))
    expense_sheet.freeze_panes = "A3"
    for index, expense in enumerate(data["expenses"]):
        expense_sheet.append([
            expense.expense_date.isoformat() if expense.expense_date else None,
            safe_text(expense.person),
            safe_text(expense.expense_type),
            safe_text(expense.fee_category),
            safe_text(expense.reason),
            float(expense.amount) if expense.amount is not None else None,
            (
                float(expense.amount_inc_tax)
                if expense_data_available
                and expense.amount_inc_tax is not None else None
            ),
            (
                float(expense.amount_ex_tax)
                if expense_data_available
                and expense.amount_ex_tax is not None else None
            ),
            (
                "费用快照未就绪"
                if not expense_data_available
                else "双口径已确认"
                if (
                    expense.amount_inc_tax is not None
                    and expense.amount_ex_tax is not None
                )
                else "费用税务口径缺失"
            ),
            safe_text(expense.data_status),
            safe_text(expense.bxd_no),
            expense.line_no,
        ])
        rendered_row = expense_sheet.max_row
        inactive = expense.data_status != config.MAINT_EXPENSE_ACTIVE_STATUS
        for column in range(1, len(expense_headers) + 1):
            cell = expense_sheet.cell(row=rendered_row, column=column)
            cell.border = _BORDER
            if index % 2:
                cell.fill = _ALT_FILL
            if inactive:
                cell.font = Font(color="A0A0A0")
        for column in (amount_column, amount_column + 1, amount_column + 2):
            expense_sheet.cell(
                row=rendered_row,
                column=column,
            ).number_format = _MONEY
    if data["expenses"]:
        expense_sheet.auto_filter.ref = f"A2:L{expense_sheet.max_row}"
        total_row_index = expense_sheet.max_row + 1
        expense_sheet.cell(
            row=total_row_index,
            column=amount_column - 1,
            value="双口径合计（仅已结束）",
        ).font = Font(bold=True)
        total_values = (
            None,
            float(expense_inc) if expense_inc is not None else None,
            float(expense_ex) if expense_ex is not None else None,
        )
        for offset, value in enumerate(total_values):
            total_cell = expense_sheet.cell(
                row=total_row_index,
                column=amount_column + offset,
                value=value,
            )
            total_cell.font, total_cell.fill = Font(bold=True), _TOTAL_FILL
            total_cell.number_format = _MONEY
        expense_sheet.cell(
            row=total_row_index,
            column=amount_column + 3,
            value=expense_status_label,
        ).font = Font(bold=True)
    _col_widths(
        expense_sheet,
        [12, 10, 12, 14, 42, 13, 13, 13, 16, 10, 18, 6],
    )

    instructions_sheet = workbook.create_sheet("填写说明")
    instructions_sheet.sheet_view.showGridLines = False
    notes = [
        (
            "这本工作簿怎么用",
            "这是系统导出的「项目追踪工作簿」：报销明细页由你续填，其余页由系统生成。"
            "填好后把整本工作簿拖回系统「数据导入」页——系统只吃报销明细页，其它页自动跳过。",
        ),
        (
            "报销明细页",
            "必填仅两列：报销日期、报销金额，且**每行都要填日期**（不允许留空表示"
            "「同上」——有金额没日期的行会报错打回）。行内没有销售订单列时，按第 1 行锚"
            "（销售订单=本合同）归集；流程状态留空视为「已结束」（计入项目已花）；"
            "单号/序号选填，有则参与防重（同一文件内单号+序号不能重复）。",
        ),
        (
            "导入模式",
            "「跳过」=增量，只进新行；「修复」=以本表为准——本合同在系统里的报销行"
            "会被本表整体替换（改了金额/删了行都以表为准），因此修复模式要求本页没有任何"
            "错误行，有错先修再导。",
        ),
        (
            "备件明细页",
            "系统按取价瀑布自动回填（产品成本=单据级总额填首行，行级取价在附加列），"
            "此页导入时忽略——备件出库数据请一直用氚云「维保需求单」导出上传，样式不变。",
        ),
        (
            "空白表单",
            "新项目可直接导出本工作簿当作空白表单分发：报销页只有表头和锚行，填完传回即可。",
        ),
    ]
    instructions_sheet.column_dimensions["A"].width = 16
    instructions_sheet.column_dimensions["B"].width = 96
    title = instructions_sheet.cell(row=1, column=1, value="项目追踪工作簿 · 填写说明")
    title.font = _TITLE_FONT
    for row, (heading, text) in enumerate(notes, 3):
        heading_cell = instructions_sheet.cell(row=row, column=1, value=heading)
        heading_cell.font = Font(bold=True)
        heading_cell.alignment = Alignment(vertical="top")
        text_cell = instructions_sheet.cell(row=row, column=2, value=text)
        text_cell.alignment = Alignment(wrap_text=True, vertical="top")
        instructions_sheet.row_dimensions[row].height = 42
    return workbook
