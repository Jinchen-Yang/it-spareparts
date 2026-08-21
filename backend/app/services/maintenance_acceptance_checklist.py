"""验收需求清单 Excel 导入服务（2026-08-21 客户反馈）。

两阶段（复刻 maintenance_doc_import 范式）：
1. ``parse_checklist_workbook`` + ``store_preview``——openpyxl 解析「验收需求 /
   是否完成」两列，raw 全量落库，零 canonical 写入；
2. ``apply_batch``——无问题行才生效，**整表替换**当前清单（当前 = 该项目
   applied_at 最新的批次；旧批次留档为历史，replaced_batch_id 串链）。

行级校验（fail-closed）：需求为空、是否完成无法识别 → issue 行，apply 整批拒绝。
"""

from __future__ import annotations

import hashlib
import io
import uuid
from datetime import UTC, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.maintenance_acceptance_checklist import (
    MaintenanceAcceptanceChecklistBatch,
    MaintenanceAcceptanceChecklistItem,
)
from app.models.maintenance_project import MaintenanceProject

MAX_CHECKLIST_BYTES = 8 * 1024 * 1024
MAX_CHECKLIST_ROWS = 2000

SHEET_CANDIDATES = ("验收清单", "验收需求")
HEADER_SCAN_ROWS = 10
COL_REQUIREMENT = "验收需求"
COL_DONE = "是否完成"

_DONE_TRUE = {"是", "已完成", "完成", "已完", "yes", "y", "true", "1"}
_DONE_FALSE = {"否", "未完成", "未完", "no", "n", "false", "0"}
EXAMPLE_PREFIX = "（示例）"


class ChecklistParseError(Exception):
    """文件不是可识别的验收清单。"""


class ChecklistBatchError(Exception):
    """批次状态/幂等冲突。"""


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_checklist_workbook(data: bytes, filename: str) -> dict:
    """解析清单：定位表头 → 逐行归一化。返回 preview 载荷（不含 batch_id）。"""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ChecklistParseError(f"无法读取 Excel 文件：{type(exc).__name__}") from exc
    try:
        ws = None
        for name in SHEET_CANDIDATES:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.worksheets[0] if wb.worksheets else None
        if ws is None:
            raise ChecklistParseError("工作簿没有任何工作表")

        rows = list(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS + MAX_CHECKLIST_ROWS,
                                 max_col=8, values_only=True))
        # 表头定位：前 HEADER_SCAN_ROWS 行内找「验收需求」列
        header_idx = None
        req_col = done_col = None
        for i, row in enumerate(rows[:HEADER_SCAN_ROWS]):
            texts = {_cell_text(v): j for j, v in enumerate(row) if _cell_text(v)}
            if COL_REQUIREMENT in texts:
                header_idx = i
                req_col = texts[COL_REQUIREMENT]
                done_col = texts.get(COL_DONE)
                break
        if header_idx is None or req_col is None:
            raise ChecklistParseError(
                f"未找到表头：需要「{COL_REQUIREMENT}」列（可选「{COL_DONE}」列）")

        items: list[dict] = []
        data_rows = rows[header_idx + 1:]
        for offset, row in enumerate(data_rows):
            row_no = header_idx + offset + 2  # Excel 实际行号（1 基）
            raw_req = _cell_text(row[req_col]) if req_col < len(row) else ""
            raw_done = (_cell_text(row[done_col]) if done_col is not None
                        and done_col < len(row) else "")
            if not raw_req and not raw_done:
                continue  # 整行空白跳过
            if raw_req.startswith(EXAMPLE_PREFIX):
                continue  # 模板灰底示例行——防呆跳过（同总表 V2.3.0 示例行约定）
            done: bool | None = None
            issues: list[str] = []
            if not raw_req:
                issues.append("验收需求为空")
            if done_col is None:
                issues.append("缺少「是否完成」列")
            elif raw_done:
                lowered = raw_done.lower()
                if lowered in _DONE_TRUE:
                    done = True
                elif lowered in _DONE_FALSE:
                    done = False
                else:
                    issues.append(f"无法识别「是否完成」：{raw_done[:32]}")
            elif not issues:
                # 需求有值但完成为空：视为待验收（未完成），不拦导入
                done = False
            items.append({
                "row_no": row_no,
                "requirement": raw_req,
                "done": done,
                "raw": {"row_no": row_no, "requirement": raw_req, "done": raw_done},
                "issues": issues,
            })
            if len(items) >= MAX_CHECKLIST_ROWS:
                break
        if not items:
            raise ChecklistParseError("清单没有任何数据行")
        return {
            "file_hash": hashlib.sha256(data).hexdigest(),
            "filename": filename,
            "items": items,
            "item_rows": len(items),
            "issue_rows": sum(1 for it in items if it["issues"]),
        }
    finally:
        wb.close()


def store_preview(
    db: Session,
    parsed: dict,
    *,
    project_id: str,
    uploaded_by: str,
    idempotency_key: str,
) -> str:
    """落 pending 批次 + raw 行。幂等重放：同 (uploaded_by, key) 且同 hash 返回原批次。"""
    existing = db.scalar(
        select(MaintenanceAcceptanceChecklistBatch).where(
            MaintenanceAcceptanceChecklistBatch.uploaded_by == uploaded_by,
            MaintenanceAcceptanceChecklistBatch.idempotency_key == idempotency_key,
        )
    )
    if existing is not None:
        if existing.file_hash != parsed["file_hash"]:
            raise ChecklistBatchError("同一幂等键已用于不同文件，请换一个上传重试")
        db.rollback()
        return existing.batch_id
    project = db.get(MaintenanceProject, project_id)
    if project is None:
        raise ChecklistBatchError("维保项目不存在")
    batch = MaintenanceAcceptanceChecklistBatch(
        batch_id=str(uuid.uuid4()),
        project_id=project_id,
        file_hash=parsed["file_hash"],
        filename=parsed["filename"][:255],
        idempotency_key=idempotency_key,
        uploaded_by=uploaded_by,
        item_rows=parsed["item_rows"],
        issue_rows=parsed["issue_rows"],
        status="pending",
        report_json={
            "item_rows": parsed["item_rows"],
            "issue_rows": parsed["issue_rows"],
            "done_rows": sum(1 for it in parsed["items"] if it["done"] is True),
            "todo_rows": sum(1 for it in parsed["items"] if it["done"] is False),
        },
    )
    db.add(batch)
    db.flush()
    for it in parsed["items"]:
        db.add(MaintenanceAcceptanceChecklistItem(
            item_id=str(uuid.uuid4()),
            batch_id=batch.batch_id,
            row_no=it["row_no"],
            raw_json=it["raw"],
            requirement=it["requirement"],
            done=it["done"],
            issues=it["issues"],
        ))
    db.flush()
    return batch.batch_id


def _current_batch(db: Session, project_id: str
                   ) -> MaintenanceAcceptanceChecklistBatch | None:
    return db.scalar(
        select(MaintenanceAcceptanceChecklistBatch)
        .where(
            MaintenanceAcceptanceChecklistBatch.project_id == project_id,
            MaintenanceAcceptanceChecklistBatch.status == "applied",
        )
        .order_by(MaintenanceAcceptanceChecklistBatch.applied_at.desc())
        .limit(1)
    )


def apply_batch(db: Session, batch_id: str, *, operated_by: str) -> dict:
    """应用批次：整表替换当前清单。有 issue 行 → 整批拒绝（fail-closed）。"""
    batch = db.get(MaintenanceAcceptanceChecklistBatch, batch_id)
    if batch is None:
        raise ChecklistBatchError("清单批次不存在")
    if batch.uploaded_by != operated_by:
        raise ChecklistBatchError("只能应用本人上传的清单批次")
    if batch.status == "applied":
        db.rollback()
        return _batch_summary(db, batch)
    if batch.status == "failed":
        raise ChecklistBatchError("批次已失败，请重新上传")

    items = list(db.scalars(
        select(MaintenanceAcceptanceChecklistItem)
        .where(MaintenanceAcceptanceChecklistItem.batch_id == batch_id)
        .order_by(MaintenanceAcceptanceChecklistItem.row_no)))
    issues = [(it.row_no, msg) for it in items for msg in (it.issues or [])]
    if issues:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "failures": [f"第 {row} 行：{msg}" for row, msg in issues[:20]],
        }
        db.flush()
        raise ChecklistBatchError(
            f"清单有 {len(issues)} 个问题行，整批未生效："
            + "；".join(f"第 {row} 行 {msg}" for row, msg in issues[:5]))

    previous = _current_batch(db, batch.project_id)
    batch.status = "applied"
    batch.applied_by = operated_by
    batch.applied_at = datetime.now(UTC)
    batch.replaced_batch_id = previous.batch_id if previous else None
    db.flush()
    return {
        "batch_id": batch.batch_id,
        "item_rows": batch.item_rows,
        "done_rows": sum(1 for it in items if it.done is True),
        "todo_rows": sum(1 for it in items if it.done is False),
        "replaced_batch_id": batch.replaced_batch_id,
    }


def _batch_summary(db: Session, batch: MaintenanceAcceptanceChecklistBatch) -> dict:
    items = list(db.scalars(
        select(MaintenanceAcceptanceChecklistItem)
        .where(MaintenanceAcceptanceChecklistItem.batch_id == batch.batch_id)
        .order_by(MaintenanceAcceptanceChecklistItem.row_no)))
    return {
        "batch_id": batch.batch_id,
        "item_rows": batch.item_rows,
        "done_rows": sum(1 for it in items if it.done is True),
        "todo_rows": sum(1 for it in items if it.done is False),
        "replaced_batch_id": batch.replaced_batch_id,
    }


def project_checklist(db: Session, project_id: str) -> dict:
    """当前生效清单 + 历史批次列表（面板「验收」页签数据）。"""
    current = _current_batch(db, project_id)
    history = list(db.scalars(
        select(MaintenanceAcceptanceChecklistBatch)
        .where(
            MaintenanceAcceptanceChecklistBatch.project_id == project_id,
            MaintenanceAcceptanceChecklistBatch.status == "applied",
        )
        .order_by(MaintenanceAcceptanceChecklistBatch.applied_at.desc())
        .limit(20)))
    current_payload = None
    if current is not None:
        items = list(db.scalars(
            select(MaintenanceAcceptanceChecklistItem)
            .where(MaintenanceAcceptanceChecklistItem.batch_id == current.batch_id)
            .order_by(MaintenanceAcceptanceChecklistItem.row_no)))
        current_payload = {
            "batch_id": current.batch_id,
            "filename": current.filename,
            "uploaded_by": current.uploaded_by,
            "applied_by": current.applied_by,
            "applied_at": current.applied_at.isoformat() if current.applied_at else None,
            "item_rows": current.item_rows,
            "done_rows": sum(1 for it in items if it.done is True),
            "todo_rows": sum(1 for it in items if it.done is False),
            "items": [
                {
                    "item_id": it.item_id,
                    "row_no": it.row_no,
                    "requirement": it.requirement,
                    "done": it.done,
                }
                for it in items
            ],
        }
    return {
        "current": current_payload,
        "history": [
            {
                "batch_id": b.batch_id,
                "filename": b.filename,
                "applied_by": b.applied_by,
                "applied_at": b.applied_at.isoformat() if b.applied_at else None,
                "item_rows": b.item_rows,
            }
            for b in history
        ],
    }


def build_template() -> bytes:
    """标准模板：验收需求 / 是否完成 两列 + 两行灰底示例行。

    示例行以「（示例）」开头，解析器自动跳过（防呆：模板原样上传不会产生垃圾行）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_CANDIDATES[0]
    ws.append([COL_REQUIREMENT, COL_DONE])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF3CD")
    ws.append([f"{EXAMPLE_PREFIX}设备巡检报告已归档（本行为示例，导入时自动忽略）", "是"])
    ws.append([f"{EXAMPLE_PREFIX}备件损耗清单双方签字确认（本行为示例）", "否"])
    for cell in ws[2] + ws[3]:
        cell.font = Font(color="999999", italic=True)
    ws.column_dimensions["A"].width = 64
    ws.column_dimensions["B"].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
