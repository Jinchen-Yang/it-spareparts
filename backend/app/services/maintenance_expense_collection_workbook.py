"""报销/回款往返工作簿（增补包 AB-3，F3 改判并入 v1）。

一个 .xlsx、两张 sheet，**下载 → 本地改 → 上传覆盖**：

- ``04_报销订单``：黄底可编辑「未税金额」；正式金额列 = ``amount_inc_tax``
  （REQUIREMENTS #8），由未税 × (1+税率) 系统算出，不接受人工直填含税。
- ``05_项目经理回款单``：**月度累计快照**（REQUIREMENTS #30），表尾追加；
  黄底可编辑「操作 / 报告月份 / 累计回款金额 / 回款凭证号 / 备注」。
  操作 CREATE=新增或覆盖该合同该月快照；VOID=作废历史快照（缺行 ≠ 删除）。

v1 明确不做（AB-3 裁剪）：氚云 BXD 逐条对账、凭证附件、进老板看板三槽。
回款**计划**不在本表——唯一事实源是台账 ``02_回款计划``（REQUIREMENTS #31），
本表只记已收到的累计额，不重复录计划。

与冻结的「项目工作簿 v3 导出」的关系：两张 sheet 的列定义原样取自 v3 模板
（`maintenance_project_workbook_v3.py`），但本模块是**独立的两表工作簿**，
不导出 v3 的其余四表，因此不解冻 v3。
"""
from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app import config
from app.business_time import business_today
from app.models.maintenance import FProjectExpense
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
    MaintenanceProjectWorkbookState,
)
from app.models.system import SysImportBatch

PROTOCOL_VERSION = "expense-collection-v1"
SHEET_EXPENSE = "04_报销订单"
SHEET_COLLECTION = "05_项目经理回款单"
_META_SHEET = "99_元数据"

# 与 f_project_expense 的 CHECK 约束一致（税率固定 13%，双口径保留原值）
TAX_RATE = Decimal("0.13")

_EXPENSE_HEADERS = ["报销单号", "报销日期", "报销人员", "报销类别", "费用分类",
                    "支出事由", "合同编号", "未税金额", "含税金额(系统计算)",
                    "流程状态", "备注"]
_COLLECTION_HEADERS = ["操作", "合同编号", "报告月份", "累计回款金额(含税)",
                       "回款凭证号", "状态(系统)", "备注"]

_EDITABLE = PatternFill("solid", fgColor="FFF3C4")   # 黄底=可编辑
_READONLY = PatternFill("solid", fgColor="EFEFEF")


class WorkbookError(ValueError):
    """回传给 API 层映射 422 的用户可读校验错误。"""

    def __init__(self, code: str, message: str, issues: list[dict] | None = None):
        super().__init__(message)
        self.code = code
        self.message = message
        self.issues = issues or []


@dataclass(frozen=True)
class ExpenseUpdate:
    raw_line_id: str
    is_create: bool = False
    bxd_no: str | None = None
    line_no: int | None = None
    # 2026-08-17 全面放开：所有事实列可改后回传覆盖
    expense_date: date | None = None
    person: str | None = None
    expense_type: str | None = None
    fee_category: str | None = None
    reason: str | None = None
    contract_no: str | None = None
    amount_ex_tax: Decimal | None = None
    amount_inc_tax: Decimal | None = None
    # Source amount basis is read-only in workbooks but must be preserved when
    # another field or the displayed ex-tax amount is updated.
    tax_basis: str | None = None
    data_status: str | None = None
    remark: str | None = None


@dataclass(frozen=True)
class CollectionOp:
    operation: str                       # CREATE | VOID | UPDATE
    project_contract_id: str
    contract_no: str
    report_month: date
    cumulative_amount: Decimal | None    # VOID 时为 None
    receipt_reference: str | None
    remark: str | None
    # 2026-08-17 全面放开：非 VOID 时可覆盖状态
    collection_status: str | None = None


@dataclass(frozen=True)
class WorkbookPlan:
    project_id: str
    expense_updates: tuple[ExpenseUpdate, ...]
    collection_ops: tuple[CollectionOp, ...]

    @property
    def summary(self) -> dict:
        creates = sum(1 for op in self.collection_ops if op.operation == "CREATE")
        updates = sum(1 for op in self.collection_ops if op.operation == "UPDATE")
        voids = sum(1 for op in self.collection_ops if op.operation == "VOID")
        return {
            "expense_creates": sum(item.is_create for item in self.expense_updates),
            "expense_updates": sum(not item.is_create for item in self.expense_updates),
            "collection_creates": creates,
            "collection_updates": updates,
            "collection_voids": voids,
        }


# ------------------------------------------------------------------ 导出

def _contracts(db: Session, project_id: str) -> list[MaintenanceProjectContract]:
    return list(db.execute(
        select(MaintenanceProjectContract)
        .where(MaintenanceProjectContract.project_id == project_id)
        .order_by(MaintenanceProjectContract.effective_from)
    ).scalars())


def _current_unique_contracts(
    db: Session,
    project_id: str,
    *,
    contract_nos: set[str] | None = None,
    lock: bool = False,
) -> dict[str, MaintenanceProjectContract]:
    """Return current contract identities owned by exactly one live relation.

    ``linked_sales_order_no`` is not a foreign key.  Treating it as project
    ownership leaked the same BXD row through every project that happened to
    share a contract number.  Expense mutation is therefore allowed only when
    the URL project owns one current relation and no other current relation in
    any project uses that contract number.

    During apply, all matching live relationship rows are locked in one global
    order before uniqueness is recomputed.  This prevents two workbook applies
    from observing different subsets or taking cross-project rows in reverse
    order.
    """

    today = business_today()
    current_condition = (
        (MaintenanceProjectContract.effective_from <= today)
        & (
            MaintenanceProjectContract.effective_to.is_(None)
            | (MaintenanceProjectContract.effective_to > today)
        )
    )
    candidates = set(db.scalars(
        select(MaintenanceProjectContract.contract_no)
        .where(
            MaintenanceProjectContract.project_id == project_id,
            current_condition,
        )
        .order_by(
            MaintenanceProjectContract.contract_no,
            MaintenanceProjectContract.project_contract_id,
        )
    ).all())
    if contract_nos is not None:
        candidates &= {str(value) for value in contract_nos if value}
    if not candidates:
        return {}

    statement = (
        select(MaintenanceProjectContract)
        .where(
            MaintenanceProjectContract.contract_no.in_(sorted(candidates)),
            current_condition,
        )
        .order_by(
            MaintenanceProjectContract.contract_no,
            MaintenanceProjectContract.project_id,
            MaintenanceProjectContract.project_contract_id,
        )
    )
    if lock:
        statement = statement.with_for_update()
    rows = list(db.scalars(statement).all())
    by_no: dict[str, list[MaintenanceProjectContract]] = {}
    for row in rows:
        by_no.setdefault(row.contract_no, []).append(row)
    return {
        contract_no: matches[0]
        for contract_no, matches in by_no.items()
        if len(matches) == 1 and matches[0].project_id == project_id
    }


def project_expenses(db: Session, project_id: str) -> list[FProjectExpense]:
    """Project-scoped editable expenses from canonical attribution.

    作废行（data_status='已作废'）彻底不导出（#264/#267 读侧修复 3：
    用户拍板 2026-08-19）——导出口径即回传口径，缺行=作废的对账语义
    要求导出侧从源头剔除已作废行，否则作废行永远循环出现在往返文件里。
    """
    writable_contracts = _current_unique_contracts(db, project_id)
    contract_nos = set(writable_contracts)
    if not contract_nos:
        return []
    rows = list(db.execute(
        select(FProjectExpense, MaintenanceProjectExpenseAttribution)
        .join(
            MaintenanceProjectExpenseAttribution,
            MaintenanceProjectExpenseAttribution.raw_expense_line_id
            == FProjectExpense.raw_line_id,
        )
        .where(
            MaintenanceProjectExpenseAttribution.project_id == project_id,
            MaintenanceProjectExpenseAttribution.project_contract_id.in_([
                contract.project_contract_id
                for contract in writable_contracts.values()
            ]),
            FProjectExpense.linked_sales_order_no.in_(contract_nos),
            or_(FProjectExpense.data_status.is_(None),
                FProjectExpense.data_status != "已作废"),
        )
        .order_by(FProjectExpense.expense_date, FProjectExpense.raw_line_id)
    ).all())
    return [
        expense
        for expense, attribution in rows
        if (
            expense.linked_sales_order_no in writable_contracts
            and attribution.project_contract_id
            == writable_contracts[expense.linked_sales_order_no].project_contract_id
        )
    ]


def _expense_in_project(
    db: Session,
    *,
    project_id: str,
    raw_line_id: str,
    writable_contracts: dict[str, MaintenanceProjectContract] | None = None,
) -> FProjectExpense | None:
    """Resolve one raw BXD only through its live canonical project edge."""

    writable_contracts = (
        writable_contracts
        if writable_contracts is not None
        else _current_unique_contracts(db, project_id)
    )
    if not writable_contracts:
        return None
    row = db.execute(
        select(FProjectExpense, MaintenanceProjectExpenseAttribution)
        .join(
            MaintenanceProjectExpenseAttribution,
            MaintenanceProjectExpenseAttribution.raw_expense_line_id
            == FProjectExpense.raw_line_id,
        )
        .where(
            FProjectExpense.raw_line_id == raw_line_id,
            MaintenanceProjectExpenseAttribution.project_id == project_id,
            MaintenanceProjectExpenseAttribution.project_contract_id.in_([
                contract.project_contract_id
                for contract in writable_contracts.values()
            ]),
            FProjectExpense.linked_sales_order_no.in_(writable_contracts),
        )
    ).one_or_none()
    if row is None:
        return None
    expense, attribution = row
    contract = writable_contracts.get(expense.linked_sales_order_no or "")
    if (
        contract is None
        or attribution.project_contract_id != contract.project_contract_id
    ):
        return None
    return expense


def _latest_snapshots(db: Session, project_id: str
                      ) -> dict[str, MaintenanceCollectionSnapshot]:
    """每份合同取最新 confirmed 快照；未来月份不进导出（与 v3 同口径）。"""
    rows = db.execute(
        select(MaintenanceCollectionSnapshot)
        .where(MaintenanceCollectionSnapshot.project_id == project_id,
               MaintenanceCollectionSnapshot.status == "confirmed",
               MaintenanceCollectionSnapshot.report_month <= business_today())
        .order_by(MaintenanceCollectionSnapshot.report_month)
    ).scalars()
    latest: dict[str, MaintenanceCollectionSnapshot] = {}
    for snapshot in rows:
        latest[snapshot.project_contract_id] = snapshot
    return latest


def _style_header(ws, headers: list[str], fills: list[PatternFill]) -> None:
    ws.append(headers)
    for idx, fill in enumerate(fills, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = fill
        cell.font = Font(bold=True)


def build_workbook(db: Session, *, project_id: str) -> bytes | None:
    """导出两表工作簿；项目不存在返回 None（由 API 映射 404）。"""
    project = db.get(MaintenanceProject, project_id)
    if project is None:
        return None
    contracts = _contracts(db, project_id)

    wb = Workbook()
    wb.remove(wb.active)

    _build_expense_sheet(wb, db, project_id=project_id)
    _build_collection_sheet(wb, db, project_id, contracts)

    meta = wb.create_sheet(_META_SHEET)
    meta.append(["协议版本", PROTOCOL_VERSION])
    meta.append(["项目ID", project_id])
    meta.append(["导出ID", str(uuid4())])
    meta.append(["导出时间", datetime.now(timezone.utc).isoformat()])
    meta.append(["口径", "正式金额=含税(系统按未税×1.13 计算)；回款=月度累计快照"])
    meta.sheet_state = "hidden"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_expense_sheet(wb, db: Session, *, project_id: str) -> None:
    """04_报销订单。抽出供项目总表（六 sheet）复用，口径只此一份。"""
    expenses = project_expenses(db, project_id)
    ws = wb.create_sheet(SHEET_EXPENSE)
    # 2026-08-17 全面放开：除含税金额(系统计算)外所有列黄底可改
    _style_header(ws, _EXPENSE_HEADERS,
                  [_EDITABLE] * 7 + [_EDITABLE, _READONLY, _EDITABLE, _EDITABLE])
    for expense in expenses:
        ws.append([
            expense.bxd_no or "",
            expense.expense_date.isoformat() if expense.expense_date else "",
            expense.person or "",
            expense.expense_type or "",
            expense.fee_category or "",
            (expense.reason or "")[:120],
            expense.linked_sales_order_no or "",
            float(expense.amount_ex_tax) if expense.amount_ex_tax is not None else "",
            float(expense.amount_inc_tax) if expense.amount_inc_tax is not None else "",
            expense.data_status or "",
            expense.remark or "",
        ])
        # 行标识写进隐藏元数据列，避免用可编辑列做主键（改了单号就对不上行）
        ws.cell(row=ws.max_row, column=len(_EXPENSE_HEADERS) + 1,
                value=expense.raw_line_id)
    ws.column_dimensions[ws.cell(row=1, column=len(_EXPENSE_HEADERS) + 1)
                         .column_letter].hidden = True

def _build_collection_sheet(wb, db: Session, project_id: str, contracts) -> None:
    """05_项目经理回款单（月度累计快照）。同样抽出供总表复用。"""
    # Contract numbers are also user-editable here.  Keep the legacy workbook
    # fail-closed when one live identity is shared/duplicated across projects.
    contracts = list(_current_unique_contracts(db, project_id).values())
    contract_no_by_id = {c.project_contract_id: c.contract_no for c in contracts}
    latest = _latest_snapshots(db, project_id)
    ws = wb.create_sheet(SHEET_COLLECTION)
    # 2026-08-17 全面放开：所有列黄底可改（含合同编号/状态）
    _style_header(ws, _COLLECTION_HEADERS, [_EDITABLE] * 7)
    for snapshot in latest.values():
        if snapshot.project_contract_id not in contract_no_by_id:
            continue
        ws.append([
            "",                                    # 操作留空=本行不动
            contract_no_by_id.get(snapshot.project_contract_id, ""),
            snapshot.report_month.strftime("%Y-%m"),
            float(snapshot.cumulative_amount),
            snapshot.receipt_reference or "",
            snapshot.status,
            snapshot.remark or "",
        ])
    # 表尾留空行供追加新月份（月度累计快照按月追加，不是改历史行）
    for _ in range(8):
        ws.append([""] * len(_COLLECTION_HEADERS))


# ------------------------------------------------------------------ 解析

def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _decimal(
    value, *, label: str, row_no: int, allow_negative: bool = False
) -> Decimal:
    try:
        parsed = Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, AttributeError, ValueError):
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}不是合法数字")
    if not parsed.is_finite() or abs(parsed) >= Decimal("1000000000000"):
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}超出允许范围")
    if parsed < 0 and not allow_negative:
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}不能为负")
    return parsed.quantize(Decimal("0.01"))


def _month(value, *, row_no: int) -> date:
    raw = _text(value)
    if isinstance(value, datetime):
        return date(value.year, value.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    for fmt in ("%Y-%m", "%Y/%m", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(raw, fmt)
        except ValueError:
            continue
        return date(parsed.year, parsed.month, 1)
    raise WorkbookError("invalid_month",
                        f"第 {row_no} 行报告月份必须是 YYYY-MM（收到 {raw!r}）")


def validate(db: Session, *, project_id: str, data: bytes) -> WorkbookPlan:
    """解析工作簿并产出**无副作用**的应用计划；任何一行不合法即整份拒绝。

    整份拒绝而非跳过：这是一份人工编辑的表，静默丢掉一行意味着操作者以为改了、
    实际没改——比报错难发现得多。
    """
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:                                    # noqa: BLE001
        raise WorkbookError("invalid_file",
                            f"无法读取 .xlsx：{type(exc).__name__}") from exc
    missing = [name for name in (SHEET_EXPENSE, SHEET_COLLECTION)
               if name not in wb.sheetnames]
    if missing:
        raise WorkbookError("missing_sheet", f"缺少工作表：{'、'.join(missing)}")

    expense_updates = _parse_expenses(db, wb[SHEET_EXPENSE], project_id=project_id)
    collection_ops = _parse_collections(db, wb[SHEET_COLLECTION],
                                        project_id=project_id)
    return WorkbookPlan(project_id=project_id,
                        expense_updates=tuple(expense_updates),
                        collection_ops=tuple(collection_ops))


def validate_partial(db: Session, *, project_id: str, workbook) -> WorkbookPlan:
    """只解析 workbook 里实际存在的 04/05，供项目总表与单 sheet 上传复用。"""
    expense_updates = (_parse_expenses(db, workbook[SHEET_EXPENSE],
                                       project_id=project_id)
                       if SHEET_EXPENSE in workbook.sheetnames else [])
    collection_ops = (_parse_collections(db, workbook[SHEET_COLLECTION],
                                         project_id=project_id)
                      if SHEET_COLLECTION in workbook.sheetnames else [])
    return WorkbookPlan(project_id=project_id,
                        expense_updates=tuple(expense_updates),
                        collection_ops=tuple(collection_ops))


def _parse_expenses(db: Session, ws, *, project_id: str) -> list[ExpenseUpdate]:
    id_col = len(_EXPENSE_HEADERS) + 1
    writable_contracts = _current_unique_contracts(db, project_id)
    updates: list[ExpenseUpdate] = []
    seen_ids: set[str] = set()
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(_text(v) == "" for v in row):
            continue
        def _col(idx: int) -> str:
            return _text(row[idx]) if len(row) > idx else ""

        bxd_no = _col(0) or None
        raw_date = _col(1)
        dt = None
        if raw_date:
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
                try:
                    dt = datetime.strptime(raw_date, fmt).date()
                    break
                except (ValueError, TypeError):
                    continue
            if dt is None:
                raise WorkbookError(
                    "invalid_expense_date",
                    f"第 {row_no} 行报销日期 {raw_date!r} 格式无效（需 YYYY-MM-DD）")

        person = _col(2) or None
        expense_type = _col(3) or None
        fee_category = _col(4) or None
        reason = _col(5) or None
        contract_no = _col(6) or None
        data_status = _col(9) or None

        raw_amount = row[7] if len(row) > 7 else None
        remark_col = _EXPENSE_HEADERS.index("备注")
        raw_remark = row[remark_col] if len(row) > remark_col else None
        remark = _text(raw_remark) or None

        ex_tax = inc_tax = None
        if _text(raw_amount) != "":
            ex_tax = _decimal(
                raw_amount,
                label="未税金额",
                row_no=row_no,
                allow_negative=True,
            )
            inc_tax = (ex_tax * (Decimal("1") + TAX_RATE)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP)
        raw_line_id = _text(row[id_col - 1]) if len(row) >= id_col else ""
        is_create = False
        expense = None
        if contract_no and contract_no not in writable_contracts:
            raise WorkbookError(
                "contract_not_found",
                f"第 {row_no} 行合同编号 {contract_no!r} 已失效、存在共享或不属于本项目",
            )
        if not raw_line_id:
            if not contract_no:
                raise WorkbookError(
                    "contract_not_found",
                    f"第 {row_no} 行合同编号不能为空")
            if dt is None:
                raise WorkbookError("missing_expense_date",
                                    f"第 {row_no} 行手工新增报销必须填写报销日期")
            if ex_tax is None:
                raise WorkbookError("missing_amount",
                                    f"第 {row_no} 行手工新增报销必须填写未税金额")
            basis = "|".join([
                contract_no, bxd_no or "", dt.isoformat(), str(ex_tax),
                reason or "", person or "",
            ])
            raw_line_id = f"EXP:{hashlib.sha1(basis.encode('utf-8')).hexdigest()[:36]}#0"
            existing = db.scalar(select(FProjectExpense).where(
                FProjectExpense.raw_line_id == raw_line_id))
            if existing is None:
                is_create = True
            else:
                expense = _expense_in_project(
                    db,
                    project_id=project_id,
                    raw_line_id=raw_line_id,
                    writable_contracts=writable_contracts,
                )
                if expense is None:
                    raise WorkbookError(
                        "expense_not_in_project",
                        f"第 {row_no} 行生成的报销标识已被其他项目占用，请调整内容后重试",
                    )
        else:
            expense = _expense_in_project(
                db,
                project_id=project_id,
                raw_line_id=raw_line_id,
                writable_contracts=writable_contracts,
            )
            if expense is None:
                raise WorkbookError(
                    "expense_not_in_project",
                    f"第 {row_no} 行的报销行已不存在、归因变化或不属于本项目，请重新下载",
                )

        if raw_line_id in seen_ids:
            raise WorkbookError(
                "duplicate_expense",
                f"第 {row_no} 行与前面重复引用同一报销事实，请删除重复行",
            )
        seen_ids.add(raw_line_id)

        if not is_create and expense.amount_ex_tax is not None and ex_tax == expense.amount_ex_tax:
            ex_tax = inc_tax = None

        # 2026-08-17 全面放开：判断任一字段是否有变化
        changed = is_create or (
            ex_tax is not None
            or remark != (expense.remark or None)
            or (dt is not None and dt != expense.expense_date)
            or (person is not None and person != (expense.person or ""))
            or (expense_type is not None and expense_type != (expense.expense_type or ""))
            or (fee_category is not None and fee_category != (expense.fee_category or ""))
            or (reason is not None and reason != (expense.reason or ""))
            or (contract_no is not None and contract_no != (expense.linked_sales_order_no or ""))
            or (data_status is not None and data_status != (expense.data_status or ""))
        )
        if not changed:
            continue

        updates.append(ExpenseUpdate(
            raw_line_id=raw_line_id, is_create=is_create,
            bxd_no=bxd_no, line_no=None,
            expense_date=dt,
            person=person,
            expense_type=expense_type,
            fee_category=fee_category,
            reason=reason,
            contract_no=contract_no,
            amount_ex_tax=ex_tax,
            amount_inc_tax=inc_tax,
            tax_basis=(expense.tax_basis if expense is not None else "ex"),
            data_status=data_status,
            remark=remark,
        ))
    return updates


def _parse_collections(db: Session, ws, *, project_id: str) -> list[CollectionOp]:
    contracts = _current_unique_contracts(db, project_id)
    ops: list[CollectionOp] = []
    seen: set[tuple[str, date]] = set()
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(_text(v) == "" for v in row):
            continue
        operation = _text(row[0]).upper()
        if operation == "":
            continue
        if operation not in ("CREATE", "VOID", "UPDATE"):
            raise WorkbookError("invalid_operation",
                                f"第 {row_no} 行操作必须是 CREATE、UPDATE 或 VOID")
        contract_no = _text(row[1])
        contract = contracts.get(contract_no)
        if contract is None:
            raise WorkbookError(
                "contract_not_found",
                f"第 {row_no} 行合同编号 {contract_no!r} 不属于本项目")
        report_month = _month(row[2] if len(row) > 2 else None, row_no=row_no)
        key = (contract.project_contract_id, report_month)
        if key in seen:
            raise WorkbookError(
                "duplicate_month",
                f"第 {row_no} 行与前面重复：同一合同同一月份只能有一条累计快照")
        seen.add(key)
        amount = None
        if operation in ("CREATE", "UPDATE"):
            if _text(row[3] if len(row) > 3 else None) == "":
                raise WorkbookError("missing_amount",
                                    f"第 {row_no} 行 {operation} 必须填累计回款金额")
            amount = _decimal(row[3], label="累计回款金额", row_no=row_no)
        # 2026-08-17 全面放开：状态列也可编辑
        status_raw = _text(row[5] if len(row) > 5 else None) or None
        ops.append(CollectionOp(
            operation=operation,
            project_contract_id=contract.project_contract_id,
            contract_no=contract_no,
            report_month=report_month,
            cumulative_amount=amount,
            receipt_reference=_text(row[4] if len(row) > 4 else None) or None,
            remark=_text(row[6] if len(row) > 6 else None) or None,
            collection_status=status_raw,
        ))
    return ops


# ------------------------------------------------------------------ 应用

def _sync_expense_attributions(
    db: Session,
    *,
    project_id: str,
    updates: tuple[ExpenseUpdate, ...],
    contracts_by_no: dict[str, MaintenanceProjectContract],
) -> bool:
    """Keep the canonical attribution fact in the same transaction as raw BXD."""

    from app.services.maintenance_expense_integrity import (
        ExpenseIntegrityError,
        OwnershipConflictError,
        sync_attribution_from_raw,
    )

    changed_any = False
    db.flush()
    for update in updates:
        expense = db.scalar(select(FProjectExpense).where(
            FProjectExpense.raw_line_id == update.raw_line_id))
        if expense is None:
            raise WorkbookError("expense_not_found", "报销行已不存在，请重新下载")
        contract = contracts_by_no.get(expense.linked_sales_order_no or "")
        if contract is None:
            raise WorkbookError(
                "contract_not_found", "报销归集合同已失效、存在共享或不属于本项目，请重新下载")
        expense_id = f"bxd:{expense.raw_line_id}"
        attribution = db.get(MaintenanceProjectExpenseAttribution, expense_id)
        if attribution is not None and attribution.project_id != project_id:
            raise WorkbookError("expense_not_in_project", "报销归因已变化，请重新下载")
        try:
            result = sync_attribution_from_raw(
                db,
                raw=expense,
                project_id=project_id,
                status_mapping_version=PROTOCOL_VERSION,
            )
        except OwnershipConflictError as exc:
            raise WorkbookError(
                "expense_not_in_project", "报销历史合同归属已变化，请重新下载"
            ) from exc
        except ExpenseIntegrityError as exc:
            raise WorkbookError("expense_incomplete", str(exc)) from exc
        synced = db.get(MaintenanceProjectExpenseAttribution, result.expense_id)
        if (
            synced is None
            or synced.ownership_mapping_state != "mapped"
            or synced.project_contract_id != contract.project_contract_id
        ):
            raise WorkbookError(
                "contract_not_found",
                "报销发生日无法唯一匹配当前项目合同，整份工作簿未写入",
            )
        changed_any = result.changed or changed_any
    return changed_any


def _lock_project_apply_context(
    db: Session,
    *,
    project_id: str,
    workbook_state: MaintenanceProjectWorkbookState | None = None,
) -> MaintenanceProjectWorkbookState:
    """Take the canonical project-workbook lock prefix: state -> project.

    The unlocked existence probe is only there to avoid trying to insert a
    state row whose project FK cannot exist.  It grants no write authority;
    the active project row is re-read under ``FOR UPDATE`` after the state lock.
    Nested master-workbook callers may pass the state they already locked so
    the complete transaction keeps one lock hierarchy and one revision bump.
    """

    from app.services import maintenance_project_operations as operations

    project_exists = db.scalar(
        select(MaintenanceProject.project_id).where(
            MaintenanceProject.project_id == project_id
        )
    )
    if project_exists is None:
        raise WorkbookError("project_not_editable", "项目已不存在，请重新下载")
    if workbook_state is None:
        workbook_state = operations.get_or_create_workbook_state(
            db, project_id=project_id, lock=True
        )
    elif workbook_state.project_id != project_id:
        raise WorkbookError("invalid_plan", "工作簿并发状态与项目不匹配")

    project = db.scalar(
        select(MaintenanceProject)
        .where(MaintenanceProject.project_id == project_id)
        .with_for_update()
    )
    if project is None or not project.is_active:
        raise WorkbookError("project_not_editable", "项目已不存在或归档，请重新下载")
    return workbook_state


def _lock_and_recheck_apply_scope(
    db: Session,
    plan: WorkbookPlan,
    *,
    workbook_state: MaintenanceProjectWorkbookState | None = None,
) -> tuple[
    dict[str, MaintenanceProjectContract],
    MaintenanceProjectWorkbookState,
]:
    """Lock and rebuild every live ownership edge before the first write."""

    workbook_state = _lock_project_apply_context(
        db,
        project_id=plan.project_id,
        workbook_state=workbook_state,
    )

    existing_ids = sorted({
        update.raw_line_id for update in plan.expense_updates
        if not update.is_create
    })
    prelock_contract_nos = {
        str(value)
        for value in db.scalars(
            select(FProjectExpense.linked_sales_order_no).where(
                FProjectExpense.raw_line_id.in_(existing_ids),
                FProjectExpense.linked_sales_order_no.is_not(None),
            )
        ).all()
        if value
    } if existing_ids else set()
    requested_contract_nos = (
        {update.contract_no for update in plan.expense_updates if update.contract_no}
        | {op.contract_no for op in plan.collection_ops}
        | prelock_contract_nos
    )
    contracts_by_no = _current_unique_contracts(
        db,
        plan.project_id,
        contract_nos=requested_contract_nos,
        lock=True,
    )
    explicit_contract_nos = (
        {update.contract_no for update in plan.expense_updates if update.contract_no}
        | {op.contract_no for op in plan.collection_ops}
    )
    if not explicit_contract_nos.issubset(contracts_by_no):
        raise WorkbookError(
            "contract_not_found",
            "工作簿包含已失效、存在共享或不属于本项目的合同，请重新下载",
        )

    raw_ids = sorted({update.raw_line_id for update in plan.expense_updates})
    # A row lock cannot protect an absent deterministic manual-create key.
    # Serialize the key first so concurrent first creates become one writer plus
    # an idempotent scoped re-read instead of a unique-constraint 500.
    for raw_line_id in raw_ids:
        db.execute(select(func.pg_advisory_xact_lock(
            func.hashtextextended(f"maintenance-expense-row:{raw_line_id}", 0)
        )))
    expense_keys = [f"bxd:{raw_line_id}" for raw_line_id in raw_ids]
    locked_attributions = list(db.scalars(
        select(MaintenanceProjectExpenseAttribution)
        .where(MaintenanceProjectExpenseAttribution.expense_id.in_(expense_keys))
        .order_by(MaintenanceProjectExpenseAttribution.expense_id)
        .with_for_update()
    ).all()) if expense_keys else []
    attribution_by_id = {
        attribution.expense_id: attribution for attribution in locked_attributions
    }
    locked_expenses = list(db.scalars(
        select(FProjectExpense)
        .where(FProjectExpense.raw_line_id.in_(raw_ids))
        .order_by(FProjectExpense.raw_line_id)
        .with_for_update()
    ).all()) if raw_ids else []
    expenses_by_id = {expense.raw_line_id: expense for expense in locked_expenses}

    for update in plan.expense_updates:
        expense = expenses_by_id.get(update.raw_line_id)
        attribution = attribution_by_id.get(f"bxd:{update.raw_line_id}")
        if expense is None:
            if not update.is_create or attribution is not None:
                raise WorkbookError(
                    "expense_not_in_project",
                    "报销行已不存在、归因异常或不属于本项目，请重新下载",
                )
            continue
        if attribution is None or attribution.project_id != plan.project_id:
            raise WorkbookError(
                "expense_not_in_project",
                "报销行的稳定项目归因已变化，请重新下载",
            )
        contract = contracts_by_no.get(expense.linked_sales_order_no or "")
        if (
            contract is None
            or attribution.project_contract_id != contract.project_contract_id
        ):
            raise WorkbookError(
                "expense_not_in_project",
                "报销行的归集合同已失效、存在共享或不属于本项目，请重新下载",
            )

    for op in plan.collection_ops:
        contract = contracts_by_no.get(op.contract_no)
        if contract is None or contract.project_contract_id != op.project_contract_id:
            raise WorkbookError(
                "contract_not_found", "回款合同关系已变化，请重新下载")
    # Contract locks serialize absent monthly snapshots.  Lock existing rows in
    # the same contract/month order before mutating any of them.
    for op in sorted(plan.collection_ops,
                     key=lambda item: (item.project_contract_id, item.report_month)):
        db.scalar(
            select(MaintenanceCollectionSnapshot)
            .where(
                MaintenanceCollectionSnapshot.project_id == plan.project_id,
                MaintenanceCollectionSnapshot.project_contract_id
                == op.project_contract_id,
                MaintenanceCollectionSnapshot.report_month == op.report_month,
            )
            .with_for_update()
        )
    return contracts_by_no, workbook_state

def apply(db: Session, plan: WorkbookPlan, *, operated_by: str,
          import_batch_id: str, commit: bool = True,
          workbook_state: MaintenanceProjectWorkbookState | None = None,
          bump_revision: bool = True,
          track_change: bool = False) -> dict:
    """整份事务应用。上传即覆盖——同合同同月份的 CREATE 覆盖既有累计额。"""
    from app.services import maintenance_project_operations as operations

    contracts_by_no, workbook_state = _lock_and_recheck_apply_scope(
        db, plan, workbook_state=workbook_state)
    operating_fact_changed = False
    manual_batch = None
    missing_create_count = sum(
        1
        for update in plan.expense_updates
        if update.is_create
        and db.scalar(select(FProjectExpense.raw_line_id).where(
            FProjectExpense.raw_line_id == update.raw_line_id
        )) is None
    )
    if missing_create_count:
        manual_batch = SysImportBatch(
            filename="manual-expense-workbook.xlsx",
            file_type="maintenance",
            file_hash=hashlib.sha256(
                f"manual-expense:{import_batch_id}".encode("utf-8")
            ).hexdigest(),
            uploaded_by=operated_by,
            rows_total=missing_create_count,
            rows_inserted=missing_create_count,
            status="success",
            report_json={"source": "workbook_manual_create", "project_id": plan.project_id},
        )
        db.add(manual_batch)
        db.flush()
    for update in sorted(plan.expense_updates, key=lambda item: item.raw_line_id):
        normalized_status = (
            config.MAINT_EXPENSE_ACTIVE_STATUS
            if update.data_status == "已生效"
            else update.data_status
        )
        expense = db.scalar(select(FProjectExpense).where(
            FProjectExpense.raw_line_id == update.raw_line_id))
        if expense is None and update.is_create:
            tax_basis = update.tax_basis or "ex"
            if tax_basis not in {"default_ex", "ex", "inc"}:
                raise WorkbookError("invalid_tax_basis", "报销金额口径无效")
            expense = FProjectExpense(
                raw_line_id=update.raw_line_id,
                bxd_no=update.bxd_no,
                line_no=update.line_no,
                data_status=normalized_status or config.MAINT_EXPENSE_ACTIVE_STATUS,
                expense_date=update.expense_date,
                person=update.person,
                expense_type=update.expense_type,
                fee_category=update.fee_category,
                reason=update.reason,
                linked_sales_order_no=update.contract_no,
                amount=(
                    update.amount_inc_tax
                    if tax_basis == "inc"
                    else update.amount_ex_tax
                ),
                amount_ex_tax=update.amount_ex_tax,
                amount_inc_tax=update.amount_inc_tax,
                tax_basis=tax_basis,
                tax_rate_used=TAX_RATE,
                remark=update.remark,
                import_batch_id=manual_batch.id,
            )
            db.add(expense)
            operating_fact_changed = True
        elif expense is None:
            raise WorkbookError("expense_not_found", "报销行已不存在，请重新下载")
        else:
            # 2026-08-17 全面放开：所有事实列可回传覆盖
            if update.expense_date is not None and update.expense_date != expense.expense_date:
                expense.expense_date = update.expense_date
                operating_fact_changed = True
            if update.person is not None and update.person != (expense.person or ""):
                expense.person = update.person
                operating_fact_changed = True
            if update.expense_type is not None and update.expense_type != (expense.expense_type or ""):
                expense.expense_type = update.expense_type
                operating_fact_changed = True
            if update.fee_category is not None and update.fee_category != (expense.fee_category or ""):
                expense.fee_category = update.fee_category
                operating_fact_changed = True
            if update.reason is not None and update.reason != (expense.reason or ""):
                expense.reason = update.reason
                operating_fact_changed = True
            if update.contract_no is not None and update.contract_no != (expense.linked_sales_order_no or ""):
                expense.linked_sales_order_no = update.contract_no
                operating_fact_changed = True
            if update.amount_ex_tax is not None:
                tax_basis = update.tax_basis or expense.tax_basis or "ex"
                if tax_basis not in {"default_ex", "ex", "inc"}:
                    raise WorkbookError("invalid_tax_basis", "报销金额口径无效")
                if (
                    expense.amount != (
                        update.amount_inc_tax
                        if tax_basis == "inc"
                        else update.amount_ex_tax
                    )
                    or expense.amount_ex_tax != update.amount_ex_tax
                    or expense.amount_inc_tax != update.amount_inc_tax
                    or expense.tax_basis != tax_basis
                ):
                    operating_fact_changed = True
                expense.amount = (
                    update.amount_inc_tax
                    if tax_basis == "inc"
                    else update.amount_ex_tax
                )
                expense.amount_ex_tax = update.amount_ex_tax
                expense.amount_inc_tax = update.amount_inc_tax
                expense.tax_basis = tax_basis
            if normalized_status is not None and normalized_status != (expense.data_status or ""):
                expense.data_status = normalized_status
                operating_fact_changed = True
            if expense.remark != update.remark:
                expense.remark = update.remark
                operating_fact_changed = True

    operating_fact_changed = _sync_expense_attributions(
        db,
        project_id=plan.project_id,
        updates=plan.expense_updates,
        contracts_by_no=contracts_by_no,
    ) or operating_fact_changed

    for op in sorted(plan.collection_ops,
                     key=lambda item: (item.project_contract_id, item.report_month)):
        existing = db.execute(
            select(MaintenanceCollectionSnapshot)
            .where(MaintenanceCollectionSnapshot.project_contract_id
                   == op.project_contract_id,
                   MaintenanceCollectionSnapshot.project_id == plan.project_id,
                   MaintenanceCollectionSnapshot.report_month == op.report_month)
        ).scalar_one_or_none()
        if op.operation == "VOID":
            if existing is None:
                raise WorkbookError(
                    "void_target_missing",
                    f"{op.contract_no} {op.report_month:%Y-%m} 没有可作废的快照")
            existing.status = "void"
            existing.version += 1
            operating_fact_changed = True
            continue
        if op.operation == "UPDATE":
            if existing is None:
                raise WorkbookError(
                    "update_target_missing",
                    f"{op.contract_no} {op.report_month:%Y-%m} 没有可更新的快照")
            existing.cumulative_amount = op.cumulative_amount
            existing.receipt_reference = op.receipt_reference
            existing.remark = op.remark
            if op.collection_status is not None:
                existing.status = op.collection_status
            existing.source = "workbook"
            existing.import_batch_id = import_batch_id
            existing.version += 1
            operating_fact_changed = True
            continue
        if existing is None:
            db.add(MaintenanceCollectionSnapshot(
                collection_id=str(uuid4()),
                project_id=plan.project_id,
                project_contract_id=op.project_contract_id,
                report_month=op.report_month,
                cumulative_amount=op.cumulative_amount,
                status=op.collection_status or "confirmed",
                receipt_reference=op.receipt_reference,
                remark=op.remark,
                source="workbook",
                import_batch_id=import_batch_id,
                version=1,
            ))
            operating_fact_changed = True
        else:
            existing.cumulative_amount = op.cumulative_amount
            existing.status = op.collection_status or "confirmed"
            existing.receipt_reference = op.receipt_reference
            existing.remark = op.remark
            existing.source = "workbook"
            existing.import_batch_id = import_batch_id
            existing.version += 1
            operating_fact_changed = True
    if bump_revision and operating_fact_changed:
        operations.bump_locked_workbook_revision(db, state=workbook_state)
    if commit:
        db.commit()
    result = {"applied_by": operated_by, "import_batch_id": import_batch_id,
              **plan.summary}
    if track_change:
        result["_operating_fact_changed"] = operating_fact_changed
    return result
