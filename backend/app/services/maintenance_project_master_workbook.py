"""项目总表（表 6 六 sheet）与全项目备件行级表（页面重设计 R2）。

页面定稿的核心原则（REQUIREMENTS #40）：**一张项目总表改所有数据**——缺价补录、
报销、回款全部走「下载 Excel → 改 → 上传覆盖」，系统内不再手工散改。
下载/上传共 3 处，**在哪下载就在哪上传**（#38）：

1. 主页全局：全项目备件行级表（本模块 `build_global_lines` / `apply_global_lines`）；
2. 项目面板：本项目总表六 sheet（`build_project_master` / `apply_project_master`）；
3. 各 tab：单 sheet（同两个函数，传 `sheets=` 子集）。

本模块**扩展** AB-3 的 `maintenance_expense_collection_workbook`，不另起炉灶：
`04_报销订单` / `05_项目经理回款单` 的解析与写库直接复用那边已验收的实现，
本模块只新增 `01/02`（只读）、`03_备件订单`（缺价补录）、`06_现场领用与返还`
（行级不返还标记）。

零迁移：`03` 的未税单位成本落 `maintenance_manual_cost_override`（该表的既有用途
就是「自动取价瀑布仍缺失时的人工成本证据」，line_id 唯一，天然可覆盖）；
`06` 的不返还标记落 `maintenance_site_issue_line.no_return`。
"""
from __future__ import annotations

import io
import hashlib
import hmac
import json
import re
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timedelta, timezone
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import and_, func, select
from sqlalchemy.orm import Session

from app import config
from app.business_time import business_today
from app.config import get_settings
from app.models.dimensions import DimPart, PartAlias
from app.models.maintenance import (
    FMaintenanceLine,
    FMaintenanceOrder,
    FProjectExpense,
    MaintenanceManualCostOverride,
)
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectContract,
)
from app.models.maintenance_project_operations import (
    MaintenanceCollectionSnapshot,
    MaintenanceProjectExpenseAttribution,
    MaintenanceProjectOperationAudit,
    MaintenanceProjectWorkbookOperation,
    MaintenanceSiteIssue,
    MaintenanceSiteIssueLine,
)
from app.models.system import SysImportBatch
from app.models.maintenance_manager import MaintenanceCollectionMilestone
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.models.sales import FSalesOrder
from app.security import FULL_SCOPE_ROLES, UserContext
from app.services import maintenance_cost_quality, maintenance_project_identity
from app.services import maintenance_expense_collection_workbook as ec
from app.services.maintenance_boss_board import _card_contracts
from app.services.maintenance_collection_milestones import write_collection_milestone
from app.services.maintenance_expense_collection_workbook import (
    TAX_RATE,
    WorkbookError,
)

PROTOCOL_VERSION = "project-master-v1"
V2_PROTOCOL_ID = "ITDATA_MAINT_PROJECT_MASTER/2.0"
# 2026-08-19（#264/#267）：03 全字段可编辑 + 04 作废/缺行=作废——新增「操作」列、
# 放开行级数据列、只读哈希收窄。旧 2.0.0 工作簿因哈希列集合变更一律拒绝并
# 提示重新下载。
V2_TEMPLATE_VERSION = "2.7.1"
V2_META_SIGNATURE_ALGORITHM = "HMAC-SHA256"
_V2_META_SIGNATURE_DOMAIN = b"ITDATA_MAINT_PROJECT_MASTER_META_V1\x00"

SHEET_BASICS = "01_项目基础信息"
SHEET_OVERVIEW = "02_概览数据"
SHEET_PARTS = "03_备件订单"
SHEET_EXPENSE = ec.SHEET_EXPENSE          # 04_报销订单（AB-3 已实现）
SHEET_COLLECTION = ec.SHEET_COLLECTION    # 05_项目经理回款单（AB-3 已实现）
SHEET_SITE = "06_现场领用与返还"
_META_SHEET = "99_元数据"
_GLOBAL_COST_TOKEN_DOMAIN = b"ITDATA_MAINT_GLOBAL_COST_BASE_V1\x00"

ALL_SHEETS = (SHEET_BASICS, SHEET_OVERVIEW, SHEET_PARTS, SHEET_EXPENSE,
              SHEET_COLLECTION, SHEET_SITE)
# 可回填的四张：01/02 是只读呈现（表 6 规格明写「全只读」）
EDITABLE_SHEETS = (SHEET_PARTS, SHEET_EXPENSE, SHEET_COLLECTION, SHEET_SITE)

_EDITABLE = PatternFill("solid", fgColor="FFF3C4")
_READONLY = PatternFill("solid", fgColor="EFEFEF")

# 全项目备件行级表（主页全局下载）
GLOBAL_SHEET = "备件行级表"
_RANGE_PRESETS = ("today", "yesterday", "this_week", "this_month", "custom")

_PARTS_HEADERS = ["维保单号", "制单日期", "合同编号", "项目名称", "PN", "产品描述",
                  "需求数量", "退货数量", "发货SN", "出库仓库", "成本来源(系统)",
                  "未税单位成本", "含税单位成本(系统计算)", "变更原因"]
_SITE_HEADERS = ["现场领用单号", "领用日期", "PN", "备件SN", "领用数量",
                 "是否应返还(行级)", "备注"]
_GLOBAL_HEADERS = ["项目名称", "合同编号", "维保单号", "制单日期", "PN", "产品描述",
                   "需求数量", "退货数量", "成本来源(系统)", "未税单位成本",
                   "含税单位成本(系统计算)", "变更原因"]

V2_SHEET_USAGE = "00_使用说明"
V2_SHEET_OVERVIEW = "01_项目概览"
V2_SHEET_PLAN = "02_回款计划"
V2_SHEET_PARTS = "03_备件明细"
V2_SHEET_EXPENSE = "04_费用报销"
V2_SHEET_RECEIPTS = "05_实收回款"
V2_SHEET_SITE = "06_领用返还"
V2_SHEET_DICTIONARY = "98_字段说明"
V2_SHEET_META = "99_元数据"
V2_ALL_SHEETS = (V2_SHEET_OVERVIEW, V2_SHEET_PLAN, V2_SHEET_PARTS,
                 V2_SHEET_EXPENSE, V2_SHEET_RECEIPTS, V2_SHEET_SITE,
                 V2_SHEET_DICTIONARY, V2_SHEET_USAGE, V2_SHEET_META)
V2_EDITABLE = PatternFill("solid", fgColor="FFE699")
V2_READONLY = PatternFill("solid", fgColor="D9E2F3")
V2_HEADER = PatternFill("solid", fgColor="1F4E78")
V2_TITLE = PatternFill("solid", fgColor="17365D")
V2_PART_HEADERS = [
    "操作", "维保单号", "制单日期", "XSDD", "需求类型", "仓库", "销售人员", "业务类型",
    "PN", "描述", "需求数量", "SN", "退货数量", "成本来源", "置信度",
    "系统未税单位成本", "系统含税单位成本", "人工未税单位成本", "人工成本原因",
    "成本缺失类型", "可补价", "实体ID", "备件主键", "只读哈希",
    "备注", "来源", "基线令牌",
]
# 2026-08-19 全字段可编辑（1-based 列号）：操作列 + 行级数据列（PN/描述/数量/SN/
# 退货数量）+ 人工成本两列 + 备注。维保单号/制单日期/XSDD/头级字段/系统成本/来源仍锁。
V2_PART_EDITABLE = {1, 9, 10, 11, 12, 13, 18, 19, 25}
# 参与只读哈希的列（按表头名定位，避免插列后错位）：行级身份与系统事实。
# 可编辑列一律不进哈希，改黄底列不再触发 readonly_cell_modified。
V2_PART_HASH_COLUMNS = [
    "维保单号", "制单日期", "XSDD", "需求类型", "仓库", "销售人员", "业务类型",
    "成本来源", "置信度", "系统未税单位成本", "系统含税单位成本", "来源", "实体ID",
]
V2_PLAN_HEADERS = ["操作", "合同编号", "期次", "计划回款日期", "日期精度", "计划回款金额（含税）",
                   "累计计划金额", "最新累计实收", "到款状态", "提醒状态", "负责人", "备注", "实体ID", "基础版本"]
V2_EXPENSE_HEADERS = ["操作", "费用单号", "明细序号", "报销日期", "报销人员", "报销类别", "费用分类",
                      "支出事由", "维保销售订单（归集键）", "项目名称", "销售人员", "流程状态",
                      "原始报销金额", "金额口径", "未税金额", "含税金额（系统）", "备注", "实体ID",
                      "基线令牌"]
V2_RECEIPT_HEADERS = ["合同编号", "报告月份", "累计实收金额（含税）", "状态", "回款凭证号", "备注", "实体ID",
                      "基线令牌"]
V2_SITE_HEADERS = ["领用单号", "领用日期", "PN", "SN", "领用数量", "是否应返还", "备注",
                   "应返数量", "返还状态", "返还单号", "实体ID", "基线令牌"]
# 行级基线令牌域（2.7.0）：把每行可编辑字段的导出值签进隐藏列，上传侧据此做
# 三路合并（用户值 vs 导出基线 vs 服务端现值），取代整本 revision 作废。
_V2_ROW_BASE_DOMAIN = b"ITDATA_MAINT_MASTER_ROW_BASE_V1\x00"
V2_BASE_COLUMN = "基线令牌"
V2_PART_BASE_FIELDS = ("PN", "描述", "需求数量", "SN", "退货数量",
                       "人工未税单位成本", "人工成本原因", "备注")
V2_EXPENSE_BASE_FIELDS = ("报销日期", "报销人员", "报销类别", "费用分类", "支出事由",
                          "维保销售订单（归集键）", "原始报销金额", "未税金额", "流程状态", "备注")
V2_RECEIPT_BASE_FIELDS = ("累计实收金额（含税）", "状态", "回款凭证号", "备注")
V2_SITE_BASE_FIELDS = ("领用单号", "领用日期", "PN", "SN", "领用数量", "是否应返还", "备注")


# ------------------------------------------------------------------ 计划

@dataclass(frozen=True)
class CostRefill:
    line_id: int | None
    unit_cost_ex_tax: Decimal | None
    unit_cost_inc_tax: Decimal | None
    reason: str | None
    note: str | None = None
    # 2026-08-19 全字段可编辑（03，#264/#267）：UPDATE/VOID 作用于既有行；
    # CREATE 新增行时 line_id 为 None，order_no+pn+qty 必填。
    operation: str = "UPDATE"
    pn: str | None = None
    description: str | None = None
    qty: Decimal | None = None
    return_qty: Decimal | None = None
    serial_numbers: str | None = None
    order_no: str | None = None
    is_create: bool = False
    order_id: int | None = None
    part_id: int | None = None
    # Global workbook optimistic-concurrency snapshot.  ``None`` version means
    # no override existed at export; the signed token distinguishes that from
    # older workbooks which carried no snapshot at all.
    base_override_version: int | None = None
    base_override_active: bool | None = None
    base_unit_cost_ex_tax: Decimal | None = None
    base_unit_cost_inc_tax: Decimal | None = None
    base_reason: str | None = None


@dataclass(frozen=True)
class SourceOrderAssignmentChange:
    """项目内人工回传确认的一张 WBDD 归属更正。

    ``expected_*`` 把 validate 时看到的活跃挂靠版本带到 apply，避免校验后有人
    已经改挂而本次上传又把新结果覆盖掉。旧挂靠不会删除；正式 apply 复用来源单
    归属服务，生成不可变历史和双边审计。
    """

    source_order_id: str
    order_no: str
    expected_assignment_id: str | None
    expected_version: int | None
    previous_project_id: str | None
    previous_project_name: str | None


@dataclass(frozen=True)
class SiteReturnFlag:
    issue_line_id: str
    no_return: bool | None
    is_create: bool = False
    # 2026-08-23：06 缺行=作废（与 03/04 同语义）——删除导出中存在的领用行
    is_void: bool = False
    issue_id: str | None = None
    line_no: int | None = None
    part_id: int | None = None
    # 2026-08-17 全面放开：可回传覆盖的领用事实
    issue_no: str | None = None
    issue_date: date | None = None
    pn: str | None = None
    serial_number: str | None = None
    quantity: Decimal | None = None
    remark: str | None = None


@dataclass(frozen=True)
class MasterPlan:
    project_id: str | None
    cost_refills: tuple[CostRefill, ...] = ()
    site_flags: tuple[SiteReturnFlag, ...] = ()
    inner: ec.WorkbookPlan | None = None      # 04/05 复用 AB-3 的计划
    sheets: tuple[str, ...] = field(default=ALL_SHEETS)

    @property
    def summary(self) -> dict:
        out = {"cost_refills": len(self.cost_refills),
               "site_return_flags": len(self.site_flags),
               "expense_updates": 0, "collection_creates": 0,
               "collection_voids": 0}
        if self.inner is not None:
            out.update(self.inner.summary)
        return out


# ------------------------------------------------------------------ 取数

def project_sales_order_nos(db: Session, project_id: str) -> list[str]:
    """项目可用于费用归集的 XSDD：合同台账优先，并兼容已挂靠 WBDD 的销售订单。"""
    values = set(db.scalars(select(MaintenanceProjectContract.contract_no).where(
        MaintenanceProjectContract.project_id == project_id,
    )).all())
    values.update(db.scalars(
        select(FMaintenanceOrder.linked_sales_order_no)
        .join(
            MaintenanceSourceOrderAssignment,
            MaintenanceSourceOrderAssignment.source_order_id
            == FMaintenanceOrder.raw_order_id,
        )
        .where(
            MaintenanceSourceOrderAssignment.project_id == project_id,
            MaintenanceSourceOrderAssignment.is_active.is_(True),
            FMaintenanceOrder.linked_sales_order_no.is_not(None),
        )
    ).all())
    return sorted(value for value in values if value)


def _current_contracts(
    db: Session,
    project_id: str,
    *,
    lock: bool = False,
) -> list[MaintenanceProjectContract]:
    """Only relationships effective at the current business date."""

    today = business_today()
    statement = (
        select(MaintenanceProjectContract)
        .where(
            MaintenanceProjectContract.project_id == project_id,
            MaintenanceProjectContract.effective_from <= today,
            (
                MaintenanceProjectContract.effective_to.is_(None)
                | (MaintenanceProjectContract.effective_to > today)
            ),
        )
        .order_by(
            MaintenanceProjectContract.contract_no,
            MaintenanceProjectContract.project_contract_id,
        )
    )
    if lock:
        statement = statement.with_for_update()
    return list(db.scalars(statement).all())


def _canonical_contract_read_model(db: Session, project_id: str) -> dict:
    """Workbook-safe current contract summary.

    ``_card_contracts`` is the canonical read model shared with the project
    cards.  It deliberately retains a known subtotal alongside
    ``contract_incomplete`` for diagnostic consumers.  A workbook cell labelled
    "contract total", however, must not expose that subtotal as if it were the
    complete project cap.  Mask the amount whenever ownership or completeness
    is unresolved while preserving a real, complete zero.
    """

    card = _card_contracts(db, [project_id]).get(project_id)
    if card is None:
        return {
            "contract_nos": [],
            "amount_inc_tax": None,
            "contract_shared": False,
            "contract_incomplete": True,
        }
    shared = bool(card.get("contract_shared"))
    amount = card.get("amount_inc_tax")
    incomplete = bool(card.get("contract_incomplete")) or amount is None
    return {
        **card,
        "amount_inc_tax": None if shared or incomplete else amount,
        "contract_shared": shared,
        "contract_incomplete": incomplete,
    }


def _current_contract_by_no(
    db: Session,
    project_id: str,
    contract_no: str,
    *,
    lock: bool = False,
) -> MaintenanceProjectContract | None:
    matches = [
        contract for contract in _current_contracts(db, project_id, lock=lock)
        if contract.contract_no == contract_no
    ]
    if len(matches) > 1:
        raise WorkbookError(
            "contract_ambiguous",
            f"合同 {contract_no} 当前存在多条生效关系，请先消歧后重新下载",
        )
    return matches[0] if matches else None


def _assigned_xsdd_nos(db: Session, project_id: str) -> set[str]:
    return {
        str(value)
        for value in db.scalars(
            select(FMaintenanceOrder.linked_sales_order_no)
            .join(
                MaintenanceSourceOrderAssignment,
                (MaintenanceSourceOrderAssignment.source_order_id
                 == FMaintenanceOrder.raw_order_id)
                & MaintenanceSourceOrderAssignment.is_active.is_(True),
            )
            .where(
                MaintenanceSourceOrderAssignment.project_id == project_id,
                FMaintenanceOrder.linked_sales_order_no.is_not(None),
                FMaintenanceOrder.linked_sales_order_no != "",
            )
            .group_by(FMaintenanceOrder.linked_sales_order_no)
        ).all()
        if value
    }


def _writable_contract_nos(db: Session, project_id: str) -> set[str]:
    """Current, uniquely owned XSDD identities allowed for expense writes."""

    current = _current_contracts(db, project_id)
    local_counts: dict[str, int] = {}
    for contract in current:
        local_counts[contract.contract_no] = (
            local_counts.get(contract.contract_no, 0) + 1
        )
    candidate_nos = {
        contract_no for contract_no, count in local_counts.items()
        if count == 1
    }
    if not candidate_nos:
        return set()
    today = business_today()
    owner_counts = dict(db.execute(
        select(
            MaintenanceProjectContract.contract_no,
            func.count(func.distinct(MaintenanceProjectContract.project_id)),
        )
        .where(
            MaintenanceProjectContract.contract_no.in_(candidate_nos),
            MaintenanceProjectContract.effective_from <= today,
            (
                MaintenanceProjectContract.effective_to.is_(None)
                | (MaintenanceProjectContract.effective_to > today)
            ),
        )
        .group_by(MaintenanceProjectContract.contract_no)
    ).all())
    return {contract_no for contract_no in candidate_nos
            if int(owner_counts.get(contract_no, 0)) == 1}


def _project_expenses(db: Session, project_id: str) -> list[FProjectExpense]:
    """Editable raw expenses with stable project attribution and current XSDD."""
    return ec.project_expenses(db, project_id)


def _exact_part_for_pn(db: Session, pn: str) -> DimPart | None:
    normalized = pn.strip().upper()
    part = db.scalar(select(DimPart).where(
        DimPart.pn_std == normalized,
        DimPart.status != "merged",
    ))
    if part is not None:
        return part
    return db.scalar(
        select(DimPart)
        .join(PartAlias, PartAlias.part_id == DimPart.id)
        .where(
            PartAlias.pn_raw == pn.strip(),
            PartAlias.status == "active",
            DimPart.status != "merged",
        )
    )

def resolve_range(preset: str, date_from: date | None,
                  date_to: date | None) -> tuple[date, date]:
    """今天/昨天/本周/本月/自定义（#38）。周一为周首。"""
    if preset not in _RANGE_PRESETS:
        raise WorkbookError("invalid_range", f"时间预设必须是 {'/'.join(_RANGE_PRESETS)}")
    today = business_today()
    if preset == "today":
        return today, today
    if preset == "yesterday":
        return today - timedelta(days=1), today - timedelta(days=1)
    if preset == "this_week":
        return today - timedelta(days=today.weekday()), today
    if preset == "this_month":
        return today.replace(day=1), today
    if date_from is None or date_to is None:
        raise WorkbookError("invalid_range", "自定义区间必须同时给 from 与 to")
    if date_from > date_to:
        raise WorkbookError("invalid_range", "起始日期不能晚于结束日期")
    return date_from, date_to


def _assigned_lines(
    db: Session,
    *,
    project_id: str | None,
    window: tuple[date, date] | None,
    allowed_project_ids: set[str] | None = None,
):
    """需求单明细 + 其项目归属（未归属单不进这两张表——它们还没有项目口径）。

    2026-08-19（#267 读侧修复 1）：墓碑整单作废与行级软作废都在此过滤——
    项目总表 03、02 概览、主页全局行级表三处共用本入口。
    2026-08-21（客户反馈）：order_date 倒序——最新日期排前面；行键/行级哈希
    与行序解耦，Excel 回传不受行序影响。
    """
    from app.services import maintenance_demands

    stmt = (
        select(FMaintenanceLine, FMaintenanceOrder,
               MaintenanceSourceOrderAssignment.project_id)
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .join(MaintenanceSourceOrderAssignment,
              (MaintenanceSourceOrderAssignment.source_order_id
               == FMaintenanceOrder.raw_order_id)
              & MaintenanceSourceOrderAssignment.is_active.is_(True))
        .where(FMaintenanceLine.is_active.is_(True),
               maintenance_demands.active_demand_condition())
        .order_by(FMaintenanceOrder.order_date.desc(), FMaintenanceOrder.order_no,
                  FMaintenanceLine.line_no)
    )
    if project_id is not None:
        stmt = stmt.where(MaintenanceSourceOrderAssignment.project_id == project_id)
    if allowed_project_ids is not None:
        # Scope helper contract: None = full scope; empty set = no visible rows.
        stmt = stmt.where(
            MaintenanceSourceOrderAssignment.project_id.in_(allowed_project_ids)
        )
    if window is not None:
        stmt = stmt.where(FMaintenanceOrder.order_date >= window[0],
                          FMaintenanceOrder.order_date <= window[1])
    return db.execute(stmt).all()


def _style(ws, headers: list[str], fills: list[PatternFill]) -> None:
    ws.append(headers)
    for idx, fill in enumerate(fills, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = fill
        cell.font = Font(bold=True)


def _hide_key_column(ws, headers: list[str]) -> None:
    """行标识写进隐藏列：可编辑列不能当主键（改了单号就对不上行）。"""
    ws.column_dimensions[
        ws.cell(row=1, column=len(headers) + 1).column_letter].hidden = True


def _num(value) -> float | str:
    return float(value) if value is not None else ""


# ------------------------------------------------------------------ 导出

def _sheet_basics(wb, db: Session, project: MaintenanceProject,
                  contracts: list[MaintenanceProjectContract]) -> None:
    ws = wb.create_sheet(SHEET_BASICS)
    _style(ws, ["字段", "值"], [_READONLY, _READONLY])
    for label, value in (
        ("项目编号", project.project_code),
        ("项目名称", project.display_name),
        ("业务类型", project.business_type or ""),
        ("项目经理(CMO)", project.cmo_name or ""),
        ("销售人员", project.salesperson or ""),
        ("项目状态", project.lifecycle_status),
        ("硬盘不返还默认值(项目级)", "是" if project.no_return_default else "否"),
        ("合同编号(XSDD)", "、".join(c.contract_no for c in contracts)),
        ("前置库种类数 / 件数 / 金额(含税)", "尚未接入"),
    ):
        ws.append([label, value])


def _sheet_overview(wb, db: Session, project: MaintenanceProject,
                    contracts: list[MaintenanceProjectContract],
                    lines, *, overrides=None) -> None:
    overrides = overrides or {}
    ws = wb.create_sheet(SHEET_OVERVIEW)
    _style(ws, ["合同编号", "合同额(含税)", "原始合同状态", "是否计入总额",
                "生效日期", "失效日期"],
           [_READONLY] * 6)
    for c in contracts:
        ws.append([c.contract_no, _num(c.amount_inc_tax), c.contract_status or "",
                   "是" if c.included_in_total else "否",
                   c.effective_from.isoformat() if c.effective_from else "",
                   c.effective_to.isoformat() if c.effective_to else ""])
    ws.append([])
    ws.append(["关键指标", "值"])
    contract = _canonical_contract_read_model(db, project.project_id)
    total = contract["amount_inc_tax"]
    cost_facts = [
        _line_cost_evidence(ln, overrides.get(ln.id), basis="inc")
        for ln, _o, _p in lines
    ]
    known = sum(
        (fact["amount"] or Decimal(0))
        for fact in cost_facts
        if fact["tier"] != "missing"
    )
    missing = sum(1 for fact in cost_facts if fact["tier"] == "missing")
    collected = db.execute(
        select(func.coalesce(func.sum(MaintenanceCollectionSnapshot.cumulative_amount), 0))
        .where(MaintenanceCollectionSnapshot.project_id == project.project_id,
               MaintenanceCollectionSnapshot.status == "confirmed")
    ).scalar_one()
    # 缺数据的行照常展示，缺失只提示、不按 0 计（表 6 §02 口径）
    ws.append(["合同总额(含税)", _num(total)])
    ws.append(["累计回款(含税)", _num(collected)])
    ws.append(["项目已计成本(含税)", _num(known)])
    ws.append(["成本率(进度条口径)",
               f"{known / total * 100:.1f}%"
               if total is not None and total > 0 else ""])
    if contract["contract_shared"]:
        contract_basis = "不完整：当前合同存在跨项目共享/归属冲突，总额与成本率不展示"
    elif contract["contract_incomplete"]:
        contract_basis = "不完整：当前合同金额、映射或唯一性事实缺失，总额与成本率不展示"
    else:
        contract_basis = "完整：当前含税合同事实"
    ws.append(["合同额口径", contract_basis])
    ws.append(["缺失成本行数", missing])


def _sheet_parts(wb, db: Session, lines, *, project_name_by_id=None,
                 overrides=None) -> None:
    overrides = overrides or {}
    ws = wb.create_sheet(SHEET_PARTS)
    # 2026-08-17 全面放开：未税单价+含税单价+变更原因 三列黄底可改
    _style(ws, _PARTS_HEADERS,
           [_READONLY] * 11 + [_EDITABLE, _EDITABLE, _EDITABLE])
    for ln, order, _pid in lines:
        override = overrides.get(ln.id)
        ex_fact = _line_cost_evidence(ln, override, basis="ex")
        inc_fact = _line_cost_evidence(ln, override, basis="inc")
        resolved_source = (
            inc_fact["source"]
            if inc_fact["tier"] != "missing"
            else (ln.cost_source or "none")
        )
        ws.append([
            order.order_no, order.order_date.isoformat() if order.order_date else "",
            order.linked_sales_order_no or "", order.project_raw or "",
            ln.pn_std or ln.pn_raw or "", ln.description or "",
            _num(ln.qty), _num(ln.return_qty), ln.serial_numbers or "",
            order.warehouse or "", resolved_source,
            _num(ex_fact["unit_cost"]), _num(inc_fact["unit_cost"]),
            override.reason if resolved_source == "manual" and override else "",
        ])
        ws.cell(row=ws.max_row, column=len(_PARTS_HEADERS) + 1, value=ln.id)
    _hide_key_column(ws, _PARTS_HEADERS)


def _sheet_site(wb, db: Session, project_id: str) -> None:
    ws = wb.create_sheet(SHEET_SITE)
    # 2026-08-17 全面放开：所有列黄底可改
    _style(ws, _SITE_HEADERS, [_EDITABLE] * 7)
    rows = db.execute(
        select(MaintenanceSiteIssueLine, MaintenanceSiteIssue)
        .join(MaintenanceSiteIssue,
              MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id)
        .where(MaintenanceSiteIssue.project_id == project_id)
        .order_by(MaintenanceSiteIssue.issue_date, MaintenanceSiteIssueLine.line_no)
    ).all()
    for line, issue in rows:
        ws.append([
            issue.issue_no, issue.issue_date.isoformat() if issue.issue_date else "",
            line.pn, line.serial_number or "", _num(line.quantity),
            "" if line.no_return is None else ("否" if line.no_return else "是"),
            "",
        ])
        ws.cell(row=ws.max_row, column=len(_SITE_HEADERS) + 1, value=line.issue_line_id)
    _hide_key_column(ws, _SITE_HEADERS)


def build_project_master(db: Session, *, project_id: str,
                         sheets: tuple[str, ...] = ALL_SHEETS) -> bytes | None:
    """项目总表：默认六 sheet；传 sheets 子集即「各 tab 单 sheet 下载」。"""
    project = db.get(MaintenanceProject, project_id)
    if project is None:
        return None
    unknown = [name for name in sheets if name not in ALL_SHEETS]
    if unknown:
        raise WorkbookError("unknown_sheet", f"未知 sheet：{'、'.join(unknown)}")
    contracts = ec._contracts(db, project_id)
    current_contracts = _current_contracts(db, project_id)
    lines = _assigned_lines(db, project_id=project_id, window=None)
    line_ids = [line.id for line, _order, _project_id in lines]
    active_overrides = {
        override.line_id: override
        for override in db.scalars(
            select(MaintenanceManualCostOverride)
            .where(
                MaintenanceManualCostOverride.line_id.in_(line_ids),
                MaintenanceManualCostOverride.active.is_(True),
            )
            .order_by(MaintenanceManualCostOverride.line_id)
        ).all()
    } if line_ids else {}

    wb = Workbook()
    wb.remove(wb.active)
    for name in sheets:
        if name == SHEET_BASICS:
            _sheet_basics(wb, db, project, current_contracts)
        elif name == SHEET_OVERVIEW:
            _sheet_overview(
                wb, db, project, current_contracts, lines,
                overrides=active_overrides)
        elif name == SHEET_PARTS:
            _sheet_parts(wb, db, lines, overrides=active_overrides)
        elif name == SHEET_EXPENSE:
            ec._build_expense_sheet(wb, db, project_id=project_id)
        elif name == SHEET_COLLECTION:
            ec._build_collection_sheet(wb, db, project_id, contracts)
        elif name == SHEET_SITE:
            _sheet_site(wb, db, project_id)

    meta = wb.create_sheet(_META_SHEET)
    meta.append(["协议版本", PROTOCOL_VERSION])
    meta.append(["项目ID", project_id])
    meta.append(["导出ID", str(uuid4())])
    meta.append(["导出时间", datetime.now(timezone.utc).isoformat()])
    meta.append(["包含 sheet", "、".join(sheets)])
    meta.sheet_state = "hidden"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _global_cost_decimal(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(Decimal(str(value)).quantize(Decimal("0.01")), "f")


def _global_cost_base_token(
    *,
    line: FMaintenanceLine,
    override: MaintenanceManualCostOverride | None,
    display_unit_cost_ex_tax: Decimal | None,
    display_unit_cost_inc_tax: Decimal | None,
    visible_reason: str | None,
) -> str:
    """Signed base snapshot used to reject stale concurrent global uploads."""

    payload = {
        "line_id": int(line.id),
        "override_version": override.version if override is not None else None,
        "override_active": override.active if override is not None else None,
        "unit_cost_ex_tax": _global_cost_decimal(display_unit_cost_ex_tax),
        "unit_cost_inc_tax": _global_cost_decimal(display_unit_cost_inc_tax),
        "reason": visible_reason or None,
    }
    payload_bytes = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    key_id, signing_key = get_settings().maintenance_manifest_signing_material()
    signature = hmac.new(
        signing_key,
        _GLOBAL_COST_TOKEN_DOMAIN + payload_bytes,
        hashlib.sha256,
    ).hexdigest()
    return json.dumps(
        {"key_id": key_id, "payload": payload, "signature": signature},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )


def _parse_global_cost_base_token(value, *, line_id: int, row_no: int) -> dict:
    try:
        envelope = json.loads(str(value))
        key_id = str(envelope["key_id"])
        payload = envelope["payload"]
        signature = str(envelope["signature"])
        if not isinstance(payload, dict) or int(payload["line_id"]) != line_id:
            raise ValueError("line mismatch")
        payload_bytes = json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ).encode("utf-8")
        verification_key = (
            get_settings().maintenance_manifest_verification_keys().get(key_id)
        )
        if verification_key is None:
            raise ValueError("unknown key")
        expected = hmac.new(
            verification_key,
            _GLOBAL_COST_TOKEN_DOMAIN + payload_bytes,
            hashlib.sha256,
        ).hexdigest()
        if not hmac.compare_digest(expected, signature):
            raise ValueError("bad signature")
        raw_version = payload.get("override_version")
        version = None if raw_version is None else int(raw_version)
        if version is not None and version < 1:
            raise ValueError("bad version")
        raw_active = payload.get("override_active")
        if raw_active is not None and not isinstance(raw_active, bool):
            raise ValueError("bad active state")

        def _money(key: str) -> Decimal | None:
            raw = payload.get(key)
            return None if raw is None else Decimal(str(raw)).quantize(Decimal("0.01"))

        return {
            "override_version": version,
            "override_active": raw_active,
            "unit_cost_ex_tax": _money("unit_cost_ex_tax"),
            "unit_cost_inc_tax": _money("unit_cost_inc_tax"),
            "reason": str(payload["reason"]) if payload.get("reason") is not None else None,
        }
    except (KeyError, TypeError, ValueError, InvalidOperation, json.JSONDecodeError) as exc:
        raise WorkbookError(
            "invalid_concurrency_token",
            f"第 {row_no} 行并发校验信息无效，请重新下载全局备件表",
        ) from exc


def build_global_lines(
    db: Session,
    *,
    preset: str,
    date_from: date | None = None,
    date_to: date | None = None,
    allowed_project_ids: set[str] | None = None,
) -> bytes:
    """主页全局下载：当前账号可见项目的备件行级表（系统回填价之后）。"""
    window = resolve_range(preset, date_from, date_to)
    lines = _assigned_lines(
        db,
        project_id=None,
        window=window,
        allowed_project_ids=allowed_project_ids,
    )
    project_ids = {project_id for _line, _order, project_id in lines}
    names = dict(db.execute(
        select(MaintenanceProject.project_id, MaintenanceProject.display_name)
        .where(MaintenanceProject.project_id.in_(project_ids))
    ).all()) if project_ids else {}
    line_ids = [line.id for line, _order, _project_id in lines]
    overrides = {
        override.line_id: override
        for override in db.scalars(
            select(MaintenanceManualCostOverride)
            .where(MaintenanceManualCostOverride.line_id.in_(line_ids))
            .order_by(MaintenanceManualCostOverride.line_id)
        ).all()
    } if line_ids else {}

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(GLOBAL_SHEET)
    _style(ws, _GLOBAL_HEADERS, [_READONLY] * 9 + [_EDITABLE, _READONLY, _EDITABLE])
    for ln, order, pid in lines:
        override = overrides.get(ln.id)
        ex_fact = _line_cost_evidence(ln, override, basis="ex")
        inc_fact = _line_cost_evidence(ln, override, basis="inc")
        resolved_source = (
            inc_fact["source"]
            if inc_fact["tier"] != "missing"
            else (ln.cost_source or "none")
        )
        visible_reason = (
            override.reason
            if resolved_source == "manual" and override is not None else None
        )
        ws.append([
            names.get(pid, ""), order.linked_sales_order_no or "",
            order.order_no, order.order_date.isoformat() if order.order_date else "",
            ln.pn_std or ln.pn_raw or "", ln.description or "",
            _num(ln.qty), _num(ln.return_qty), resolved_source,
            _num(ex_fact["unit_cost"]), _num(inc_fact["unit_cost"]),
            visible_reason or "",
        ])
        ws.cell(row=ws.max_row, column=len(_GLOBAL_HEADERS) + 1, value=ln.id)
        ws.cell(
            row=ws.max_row,
            column=len(_GLOBAL_HEADERS) + 2,
            value=_global_cost_base_token(
                line=ln,
                override=override,
                display_unit_cost_ex_tax=ex_fact["unit_cost"],
                display_unit_cost_inc_tax=inc_fact["unit_cost"],
                visible_reason=visible_reason,
            ),
        )
    _hide_key_column(ws, _GLOBAL_HEADERS)
    ws.column_dimensions[
        ws.cell(row=1, column=len(_GLOBAL_HEADERS) + 2).column_letter
    ].hidden = True

    meta = wb.create_sheet(_META_SHEET)
    meta.append(["协议版本", PROTOCOL_VERSION])
    meta.append(["时间预设", preset])
    meta.append(["区间", f"{window[0].isoformat()}~{window[1].isoformat()}"])
    meta.append(["导出ID", str(uuid4())])
    meta.sheet_state = "hidden"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------ 解析

def _decimal(value, *, label: str, row_no: int) -> Decimal:
    try:
        parsed = Decimal(str(value).strip().replace(",", ""))
    except (InvalidOperation, AttributeError, ValueError):
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}不是合法数字")
    if parsed < 0:
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}不能为负")
    return parsed.quantize(Decimal("0.01"))


def _global_line_in_scope(
    db: Session,
    *,
    line_id: int,
    allowed_project_ids: set[str] | None,
    lock_assignment: bool = False,
) -> FMaintenanceLine | None:
    """Return an active, assigned demand line only when its project is visible."""
    from app.services import maintenance_demands

    stmt = (
        select(FMaintenanceLine)
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .join(
            MaintenanceSourceOrderAssignment,
            (MaintenanceSourceOrderAssignment.source_order_id
             == FMaintenanceOrder.raw_order_id)
            & MaintenanceSourceOrderAssignment.is_active.is_(True),
        )
        .where(
            FMaintenanceLine.id == line_id,
            FMaintenanceLine.is_active.is_(True),
            maintenance_demands.active_demand_condition(),
        )
    )
    if allowed_project_ids is not None:
        stmt = stmt.where(
            MaintenanceSourceOrderAssignment.project_id.in_(allowed_project_ids)
        )
    if lock_assignment:
        # Keep the active project edge stable between scope validation and the
        # override write. Reassignment uses the same row lock and therefore
        # either completes first (this query sees the new scope) or waits.
        stmt = stmt.with_for_update(of=MaintenanceSourceOrderAssignment)
    return db.scalar(stmt)


def _parse_cost_refills(
    db: Session,
    ws,
    *,
    headers: list[str],
    cost_col: int,
    reason_col: int,
    require_active_assignment: bool = False,
    allowed_project_ids: set[str] | None = None,
    require_concurrency_token: bool = False,
) -> list[CostRefill]:
    """缺价补录：只认隐藏列里的 line_id，未填金额的行原样不动。
    2026-08-17 全面放开：含税单价也可直接填写。"""
    key_col = len(headers) + 1
    token_col = key_col + 1
    out: list[CostRefill] = []
    seen_line_ids: set[int] = set()
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(ec._text(v) == "" for v in row):
            continue
        raw_key = ec._text(row[key_col - 1]) if len(row) >= key_col else ""
        if not raw_key:
            raise WorkbookError(
                "line_not_recognized",
                f"第 {row_no} 行不是导出的备件行——需求单只能由氚云导入，本表只补价")
        try:
            line_id = int(raw_key)
        except (TypeError, ValueError):
            raise WorkbookError(
                "line_not_recognized",
                f"第 {row_no} 行的备件行标识无效，请重新下载",
            )
        if line_id in seen_line_ids:
            raise WorkbookError(
                "duplicate_line",
                f"第 {row_no} 行与前面重复引用同一备件行，请删除重复行",
            )
        seen_line_ids.add(line_id)

        base = None
        if require_concurrency_token:
            raw_token = row[token_col - 1] if len(row) >= token_col else None
            if raw_token in (None, ""):
                raise WorkbookError(
                    "missing_concurrency_token",
                    f"第 {row_no} 行缺少并发校验信息，请重新下载全局备件表",
                )
            base = _parse_global_cost_base_token(
                raw_token, line_id=line_id, row_no=row_no)

        raw_ex = row[cost_col - 1] if len(row) >= cost_col else None
        # 含税单价列（紧跟未税之后）
        inc_col = cost_col + 1
        raw_inc = row[inc_col - 1] if len(row) >= inc_col else None
        reason = (
            ec._text(row[reason_col - 1]) or None
            if len(row) >= reason_col else None
        )
        has_ex = ec._text(raw_ex) != ""
        has_inc = ec._text(raw_inc) != ""
        line = (
            _global_line_in_scope(
                db,
                line_id=line_id,
                allowed_project_ids=allowed_project_ids,
            )
            if require_active_assignment
            else db.get(FMaintenanceLine, line_id)
        )
        if line is None:
            code = "project_scope_denied" if require_active_assignment else "line_not_found"
            raise WorkbookError(
                code,
                f"第 {row_no} 行的备件行已不存在或不在当前可见项目范围，请重新下载"
                if require_active_assignment
                else f"第 {row_no} 行的备件行已不存在，请重新下载",
            )
        if has_ex:
            ex_tax = _decimal(raw_ex, label="未税单位成本", row_no=row_no)
            inc_tax = (ex_tax * (Decimal("1") + TAX_RATE)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP)
        elif has_inc:
            # 只填含税 → 推算未税
            inc_val = _decimal(raw_inc, label="含税单位成本", row_no=row_no)
            inc_tax = inc_val
            ex_tax = (inc_val / (Decimal("1") + TAX_RATE)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            ex_tax = inc_tax = None

        if base is not None:
            if (
                ex_tax == base["unit_cost_ex_tax"]
                and inc_tax == base["unit_cost_inc_tax"]
                and reason == base["reason"]
            ):
                continue
            if ex_tax is None and base["override_version"] is None:
                raise WorkbookError(
                    "missing_amount",
                    f"第 {row_no} 行尚无人工成本，不能只填写变更原因",
                )
        elif not has_ex and not has_inc:
            continue
        out.append(CostRefill(
            line_id=line.id,
            unit_cost_ex_tax=ex_tax,
            unit_cost_inc_tax=inc_tax,
            reason=reason,
            base_override_version=(base["override_version"] if base else None),
            base_override_active=(base["override_active"] if base else None),
            base_unit_cost_ex_tax=(base["unit_cost_ex_tax"] if base else None),
            base_unit_cost_inc_tax=(base["unit_cost_inc_tax"] if base else None),
            base_reason=(base["reason"] if base else None),
        ))
    return out


def _parse_site_flags(
    db: Session,
    ws,
    *,
    project_id: str | None = None,
) -> list[SiteReturnFlag]:
    """2026-08-17 全面放开：06 所有列可改后回传覆盖。"""
    key_col = len(_SITE_HEADERS) + 1
    out: list[SiteReturnFlag] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(ec._text(v) == "" for v in row):
            continue
        raw_key = ec._text(row[key_col - 1]) if len(row) >= key_col else ""
        if not raw_key:
            raise WorkbookError("line_not_recognized",
                                f"第 {row_no} 行不是导出的领用行")
        existing_line = db.get(MaintenanceSiteIssueLine, raw_key)
        existing_issue = (
            db.get(MaintenanceSiteIssue, existing_line.issue_id)
            if existing_line is not None
            else None
        )
        if project_id is not None:
            if (
                existing_line is None
                or not existing_line.is_active
                or existing_issue is None
                or existing_issue.project_id != project_id
            ):
                raise WorkbookError(
                    "project_scope_denied",
                    f"第 {row_no} 行领用事实不属于当前项目，请重新下载",
                )

        issue_no = ec._text(row[0]) if len(row) > 0 and ec._text(row[0]) else None
        raw_date = ec._text(row[1]) if len(row) > 1 else ""
        parsed_date = None
        if raw_date:
            from datetime import datetime as dt_cls
            try:
                parsed_date = dt_cls.strptime(raw_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                try:
                    parsed_date = dt_cls.strptime(raw_date, "%Y/%m/%d").date()
                except (ValueError, TypeError):
                    parsed_date = None
        pn = ec._text(row[2]) if len(row) > 2 and ec._text(row[2]) else None
        sn = ec._text(row[3]) if len(row) > 3 and ec._text(row[3]) else None
        raw_qty = row[4] if len(row) > 4 else None
        qty = _decimal(raw_qty, label="领用数量", row_no=row_no) if ec._text(raw_qty) != "" else None
        raw_flag = ec._text(row[5]) if len(row) > 5 else ""
        no_return: bool | None = None
        if raw_flag != "":
            if raw_flag not in ("是", "否"):
                raise WorkbookError("invalid_flag",
                                    f"第 {row_no} 行「是否应返还」只能填 是 / 否")
            no_return = raw_flag == "否"
        remark = ec._text(row[6]) if len(row) > 6 and ec._text(row[6]) else None
        # The unsigned V1 template has no row/header OCC envelope.  Identity,
        # date, PN and quantity edits could otherwise detach PN from part_id or
        # overwrite a shared issue header in row order.  Current V2 supports
        # those edits safely; reject them here rather than silently corrupting
        # cost provenance.
        if existing_line is None or existing_issue is None:
            raise WorkbookError("line_not_found", f"第 {row_no} 行领用事实已不存在")
        requested_no = issue_no or existing_issue.issue_no
        requested_date = parsed_date or existing_issue.issue_date
        requested_pn = pn or existing_line.pn
        requested_qty = qty if qty is not None else existing_line.quantity
        if (
            requested_no != existing_issue.issue_no
            or requested_date != existing_issue.issue_date
            or requested_pn != existing_line.pn
            or requested_qty != existing_line.quantity
        ):
            raise WorkbookError(
                "legacy_site_identity_readonly",
                f"第 {row_no} 行使用旧版项目总表，领用单号/日期/PN/数量不可修改；"
                "请重新下载当前 V2 项目总表",
            )
        issue_no = parsed_date = pn = qty = None
        # 至少有一个字段变化才记录
        if no_return is None and remark is None and parsed_date is None and pn is None and sn is None and qty is None and issue_no is None:
            continue
        out.append(SiteReturnFlag(
            issue_line_id=raw_key,
            no_return=no_return,
            issue_no=issue_no,
            issue_date=parsed_date,
            pn=pn,
            serial_number=sn,
            quantity=qty,
            remark=remark,
        ))
    return out


def validate(db: Session, *, project_id: str | None, data: bytes) -> MasterPlan:
    """解析并产出无副作用计划；任何一行不合法即整份拒绝（与 AB-3 同语义）。"""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:                                    # noqa: BLE001
        raise WorkbookError("invalid_file",
                            f"无法读取 .xlsx：{type(exc).__name__}") from exc
    present = [name for name in ALL_SHEETS if name in wb.sheetnames]
    if not present:
        raise WorkbookError(
            "missing_sheet",
            f"文件里没有任何可识别的工作表（期望其一：{'、'.join(ALL_SHEETS)}）")

    refills: list[CostRefill] = []
    flags: list[SiteReturnFlag] = []
    if SHEET_PARTS in wb.sheetnames:
        refills = _parse_cost_refills(
            db,
            wb[SHEET_PARTS],
            headers=_PARTS_HEADERS,
            cost_col=12,
            reason_col=14,
            require_active_assignment=project_id is not None,
            allowed_project_ids=({project_id} if project_id is not None else None),
        )
    if SHEET_SITE in wb.sheetnames:
        flags = _parse_site_flags(
            db, wb[SHEET_SITE], project_id=project_id)

    inner = None
    if project_id is not None and (SHEET_EXPENSE in wb.sheetnames
                                   or SHEET_COLLECTION in wb.sheetnames):
        # 04/05 直接复用 AB-3 已验收的解析；缺哪张就跳过哪张
        inner = ec.validate_partial(db, project_id=project_id, workbook=wb)
    return MasterPlan(project_id=project_id, cost_refills=tuple(refills),
                      site_flags=tuple(flags), inner=inner,
                      sheets=tuple(present))


def validate_global(
    db: Session,
    *,
    data: bytes,
    allowed_project_ids: set[str] | None = None,
) -> MasterPlan:
    """主页全局备件行级表回传：只有补价一件事。"""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:                                    # noqa: BLE001
        raise WorkbookError("invalid_file",
                            f"无法读取 .xlsx：{type(exc).__name__}") from exc
    if GLOBAL_SHEET not in wb.sheetnames:
        raise WorkbookError("missing_sheet", f"缺少工作表：{GLOBAL_SHEET}")
    refills = _parse_cost_refills(db, wb[GLOBAL_SHEET], headers=_GLOBAL_HEADERS,
                                  cost_col=10, reason_col=12,
                                  require_active_assignment=True,
                                  allowed_project_ids=allowed_project_ids,
                                  require_concurrency_token=True)
    return MasterPlan(project_id=None, cost_refills=tuple(refills),
                      sheets=(GLOBAL_SHEET,))


# ------------------------------------------------------------------ 应用

def apply(
    db: Session,
    plan: MasterPlan,
    *,
    operated_by: str,
    import_batch_id: str,
    _prelocked_workbook_states: dict | None = None,
    _line_project_ids: dict[int, str] | None = None,
) -> dict:
    """整份事务应用；上传即覆盖。"""
    workbook_state = None
    operating_fact_changed = False
    changed_project_ids: set[str] = set()
    if plan.project_id is not None:
        # Standalone V1 and nested 04/05 uploads share the same global lock
        # prefix as V2: workbook state -> project.  The inner service receives
        # this already-locked state and therefore neither reverses the order nor
        # advances the revision a second time.
        workbook_state = ec._lock_project_apply_context(
            db, project_id=plan.project_id)
        refill_line_ids = [
            int(refill.line_id)
            for refill in plan.cost_refills
            if refill.line_id is not None
        ]
        if len(refill_line_ids) != len(plan.cost_refills):
            raise WorkbookError(
                "invalid_plan", "V1 项目总表包含无效备件行标识")
        _lock_global_refill_rows(
            db,
            line_ids=refill_line_ids,
            allowed_project_ids={plan.project_id},
            prelocked_project_ids={plan.project_id},
        )
        _lock_v1_site_rows(
            db,
            project_id=plan.project_id,
            issue_line_ids=[flag.issue_line_id for flag in plan.site_flags],
        )
    # Stable lock order prevents two uploads with different Excel row orders
    # from taking assignment locks in opposite order.
    for refill in sorted(plan.cost_refills, key=lambda item: item.line_id or -1):
        refill_changed = False
        # 备注与成本独立——只改备注/原因不写覆盖表
        if refill.unit_cost_ex_tax is not None or refill.reason is not None:
            refill_changed = _merge_manual_cost_to_line(
                db, refill, operated_by=operated_by
            )
            operating_fact_changed = refill_changed or operating_fact_changed
        if refill.note is not None:
            line = db.get(FMaintenanceLine, refill.line_id)
            if line is not None and line.line_note != refill.note:
                line.line_note = refill.note
                refill_changed = True
                operating_fact_changed = True
        if refill_changed and _line_project_ids is not None:
            project_id = _line_project_ids.get(int(refill.line_id))
            if project_id is None:
                raise WorkbookError(
                    "project_scope_denied",
                    "工作簿包含归属已变化的备件行，请重新下载",
                )
            changed_project_ids.add(project_id)

    for flag in plan.site_flags:
        line = db.get(MaintenanceSiteIssueLine, flag.issue_line_id)
        if line is None:
            raise WorkbookError("line_not_found",
                                f"领用行 {flag.issue_line_id} 已不存在，请重新下载")
        if line.no_return != flag.no_return:
            line.no_return = flag.no_return
            operating_fact_changed = True
        # 2026-08-17 全面放开：领用行级字段覆盖
        if flag.pn is not None and flag.pn != line.pn:
            line.pn = flag.pn
            operating_fact_changed = True
        if (flag.serial_number is not None
                and flag.serial_number != line.serial_number):
            line.serial_number = flag.serial_number
            operating_fact_changed = True
        if flag.quantity is not None and flag.quantity != line.quantity:
            line.quantity = flag.quantity
            operating_fact_changed = True
        if flag.remark is not None and flag.remark != line.remark:
            line.remark = flag.remark
            operating_fact_changed = True
        # 现场领用单头字段（issue_no/issue_date）需要通过关联的单头更新
        if flag.issue_no is not None or flag.issue_date is not None:
            issue = db.get(MaintenanceSiteIssue, line.issue_id)
            if issue is not None:
                if flag.issue_no is not None and flag.issue_no != issue.issue_no:
                    issue.issue_no = flag.issue_no
                    operating_fact_changed = True
                if (flag.issue_date is not None
                        and flag.issue_date != issue.issue_date):
                    issue.issue_date = flag.issue_date
                    operating_fact_changed = True

    if plan.inner is not None:
        inner_result = ec.apply(
            db,
            plan.inner,
            operated_by=operated_by,
            import_batch_id=import_batch_id,
            commit=False,
            workbook_state=workbook_state,
            bump_revision=False,
            track_change=True,
        )
        operating_fact_changed = (
            bool(inner_result.pop("_operating_fact_changed", False))
            or operating_fact_changed
        )
    if workbook_state is not None and operating_fact_changed:
        from app.services import maintenance_project_operations as operations

        operations.bump_locked_workbook_revision(db, state=workbook_state)
    elif changed_project_ids:
        from app.services import maintenance_project_operations as operations

        if _prelocked_workbook_states is None:
            raise WorkbookError(
                "invalid_plan", "全局备件工作簿缺少项目并发上下文"
            )
        for project_id in sorted(changed_project_ids):
            state = _prelocked_workbook_states.get(project_id)
            if state is None:
                raise WorkbookError(
                    "project_scope_denied",
                    "工作簿包含未预锁的项目备件行，请重新下载",
                )
            operations.bump_locked_workbook_revision(db, state=state)
    db.commit()
    return {"applied_by": operated_by, "import_batch_id": import_batch_id,
            "sheets": list(plan.sheets), **plan.summary}


def _lock_global_refill_rows(
    db: Session,
    *,
    line_ids: list[int],
    allowed_project_ids: set[str] | None,
    prelocked_project_ids: set[str] | None = None,
) -> tuple[
    dict[int, FMaintenanceLine],
    dict[int, MaintenanceManualCostOverride],
    dict[int, str],
]:
    """Canonical lock order: orders -> assignments -> lines -> overrides.

    Order and assignment rows use source-order order (their global key); detail
    facts use ascending ``line_id``.  The plan's Excel row order never
    influences lock acquisition.  This also matches demand void/delete and
    source-assignment writers, so those paths cannot form an order/assignment
    inversion with a global refill.
    """

    from app.services import maintenance_demands

    requested = sorted(set(line_ids))
    if len(requested) != len(line_ids):
        raise WorkbookError("duplicate_line", "全局备件表重复引用同一备件行")
    if not requested:
        return {}, {}, {}
    edges = list(db.execute(
        select(
            FMaintenanceLine.id,
            FMaintenanceLine.order_id,
            FMaintenanceOrder.raw_order_id,
        )
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .where(FMaintenanceLine.id.in_(requested))
        .order_by(FMaintenanceLine.id)
    ).all())
    if {line_id for line_id, _order_id, _source_id in edges} != set(requested):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已不存在或不在当前可见项目范围的备件行，请重新下载",
        )

    expected_source_by_order = {
        order_id: source_id for _line_id, order_id, source_id in edges
    }
    locked_orders = list(db.scalars(
        select(FMaintenanceOrder)
        .where(FMaintenanceOrder.id.in_(expected_source_by_order))
        .order_by(FMaintenanceOrder.raw_order_id, FMaintenanceOrder.id)
        .with_for_update()
    ).all())
    if (
        {order.id for order in locked_orders} != set(expected_source_by_order)
        or any(
            order.raw_order_id != expected_source_by_order[order.id]
            for order in locked_orders
        )
    ):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已不存在或已变化的维保需求单，请重新下载",
        )

    source_ids = sorted(expected_source_by_order.values())
    assignments = list(db.scalars(
        select(MaintenanceSourceOrderAssignment)
        .where(
            MaintenanceSourceOrderAssignment.source_order_id.in_(source_ids),
            MaintenanceSourceOrderAssignment.is_active.is_(True),
        )
        .order_by(MaintenanceSourceOrderAssignment.source_order_id)
        .with_for_update()
    ).all())
    assignment_by_source = {
        assignment.source_order_id: assignment for assignment in assignments
    }
    if (
        set(assignment_by_source) != set(source_ids)
        or (
            allowed_project_ids is not None
            and any(
                assignment.project_id not in allowed_project_ids
                for assignment in assignments
            )
        )
        or (
            prelocked_project_ids is not None
            and any(
                assignment.project_id not in prelocked_project_ids
                for assignment in assignments
            )
        )
    ):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已改派或不在当前可见项目范围的备件行，请重新下载",
        )

    lines = list(db.scalars(
        select(FMaintenanceLine)
        .where(
            FMaintenanceLine.id.in_(requested),
            FMaintenanceLine.is_active.is_(True),
        )
        .order_by(FMaintenanceLine.id)
        .with_for_update()
    ).all())
    lines_by_id = {line.id: line for line in lines}
    expected_order_by_line = {
        line_id: order_id for line_id, order_id, _source_id in edges
    }
    if (
        set(lines_by_id) != set(requested)
        or any(
            line.order_id != expected_order_by_line[line_id]
            for line_id, line in lines_by_id.items()
        )
    ):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已不存在或已作废的备件行，请重新下载",
        )

    overrides = list(db.scalars(
        select(MaintenanceManualCostOverride)
        .where(MaintenanceManualCostOverride.line_id.in_(requested))
        .order_by(MaintenanceManualCostOverride.line_id)
        .with_for_update()
    ).all())
    overrides_by_line = {override.line_id: override for override in overrides}

    # Re-read the complete join after all three lock layers.  This catches an
    # assignment/tombstone change that committed between validation and apply.
    scoped_statement = (
        select(FMaintenanceLine.id)
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .join(
            MaintenanceSourceOrderAssignment,
            (MaintenanceSourceOrderAssignment.source_order_id
             == FMaintenanceOrder.raw_order_id)
            & MaintenanceSourceOrderAssignment.is_active.is_(True),
        )
        .where(
            FMaintenanceLine.id.in_(requested),
            FMaintenanceLine.is_active.is_(True),
            maintenance_demands.active_demand_condition(),
        )
    )
    if allowed_project_ids is not None:
        scoped_statement = scoped_statement.where(
            MaintenanceSourceOrderAssignment.project_id.in_(allowed_project_ids)
        )
    scoped_ids = set(db.scalars(scoped_statement).all())
    if scoped_ids != set(requested):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已改派、已删除或不在当前可见项目范围的备件行，请重新下载",
        )
    source_by_line = {
        line_id: source_id for line_id, _order_id, source_id in edges
    }
    project_by_line = {
        line_id: assignment_by_source[source_id].project_id
        for line_id, source_id in source_by_line.items()
    }
    return lines_by_id, overrides_by_line, project_by_line


def _lock_v1_site_rows(
    db: Session,
    *,
    project_id: str,
    issue_line_ids: list[str],
) -> dict[str, MaintenanceSiteIssueLine]:
    """Lock V1 06 identities and prove their live URL-project ownership."""

    requested = sorted(set(issue_line_ids))
    if len(requested) != len(issue_line_ids):
        raise WorkbookError("duplicate_line", "项目总表重复引用同一领用行")
    if not requested:
        return {}

    edges = list(db.execute(
        select(
            MaintenanceSiteIssueLine.issue_line_id,
            MaintenanceSiteIssueLine.issue_id,
        )
        .where(MaintenanceSiteIssueLine.issue_line_id.in_(requested))
        .order_by(MaintenanceSiteIssueLine.issue_line_id)
    ).all())
    if {line_id for line_id, _issue_id in edges} != set(requested):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已不存在或不属于当前项目的领用行，请重新下载",
        )
    expected_issue_by_line = dict(edges)
    issue_ids = sorted(set(expected_issue_by_line.values()))
    issues = list(db.scalars(
        select(MaintenanceSiteIssue)
        .where(MaintenanceSiteIssue.issue_id.in_(issue_ids))
        .order_by(MaintenanceSiteIssue.issue_id)
        .with_for_update()
    ).all())
    if (
        {issue.issue_id for issue in issues} != set(issue_ids)
        or any(issue.project_id != project_id for issue in issues)
    ):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含不属于当前项目的领用事实，请重新下载",
        )

    lines = list(db.scalars(
        select(MaintenanceSiteIssueLine)
        .where(
            MaintenanceSiteIssueLine.issue_line_id.in_(requested),
            MaintenanceSiteIssueLine.is_active.is_(True),
        )
        .order_by(MaintenanceSiteIssueLine.issue_line_id)
        .with_for_update()
    ).all())
    lines_by_id = {line.issue_line_id: line for line in lines}
    if (
        set(lines_by_id) != set(requested)
        or any(
            line.issue_id != expected_issue_by_line[line.issue_line_id]
            for line in lines
        )
    ):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已作废或归属已变化的领用事实，请重新下载",
        )
    return lines_by_id


def _global_refill_matches_override(
    refill: CostRefill,
    override: MaintenanceManualCostOverride | None,
) -> bool:
    if override is None or not override.active:
        return False
    if refill.unit_cost_ex_tax is None:
        return override.reason == refill.reason
    return (
        override.unit_cost_ex_tax == refill.unit_cost_ex_tax
        and override.unit_cost_inc_tax == refill.unit_cost_inc_tax
        and override.reason == refill.reason
    )


def apply_global_lines(
    db: Session,
    plan: MasterPlan,
    *,
    operated_by: str,
    import_batch_id: str,
    allowed_project_ids: set[str] | None = None,
) -> dict:
    """Recheck global-workbook row scope immediately before any write."""
    if (
        plan.project_id is not None
        or plan.site_flags
        or plan.inner is not None
        or plan.sheets != (GLOBAL_SHEET,)
    ):
        raise WorkbookError("invalid_plan", "全局备件工作簿计划无效，请重新上传")
    line_ids = [
        int(refill.line_id)
        for refill in plan.cost_refills
        if refill.line_id is not None
    ]
    if len(line_ids) != len(plan.cost_refills):
        raise WorkbookError("invalid_plan", "全局备件工作簿包含无效行，请重新上传")

    # The global sheet can touch several projects in one upload.  Freeze the
    # complete project set before taking any order/assignment/detail lock so a
    # concurrent reassignment cannot create state -> fact / fact -> state
    # inversion or leave one affected workbook token unchanged.
    db.execute(
        select(func.pg_advisory_xact_lock(config.DATA_CHANGE_ADVISORY_LOCK_KEY))
    )
    probed_rows = list(db.execute(
        select(
            FMaintenanceLine.id,
            MaintenanceSourceOrderAssignment.project_id,
        )
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .join(
            MaintenanceSourceOrderAssignment,
            and_(
                MaintenanceSourceOrderAssignment.source_order_id
                == FMaintenanceOrder.raw_order_id,
                MaintenanceSourceOrderAssignment.is_active.is_(True),
            ),
        )
        .where(FMaintenanceLine.id.in_(sorted(set(line_ids))))
        .order_by(FMaintenanceLine.id)
    ).all())
    probed_project_ids = {project_id for _line_id, project_id in probed_rows}
    if (
        {line_id for line_id, _project_id in probed_rows} != set(line_ids)
        or (
            allowed_project_ids is not None
            and not probed_project_ids.issubset(allowed_project_ids)
        )
    ):
        raise WorkbookError(
            "project_scope_denied",
            "工作簿包含已改派或不在当前可见项目范围的备件行，请重新下载",
        )
    from app.services import maintenance_project_operations as operations

    workbook_states = operations.lock_workbook_states(
        db,
        project_ids=probed_project_ids,
    )
    if probed_project_ids:
        locked_projects = set(db.scalars(
            select(MaintenanceProject.project_id)
            .where(
                MaintenanceProject.project_id.in_(sorted(probed_project_ids)),
                MaintenanceProject.is_active.is_(True),
            )
            .order_by(MaintenanceProject.project_id)
            .with_for_update()
        ).all())
        if locked_projects != probed_project_ids:
            raise WorkbookError(
                "project_scope_denied",
                "工作簿包含已归档或已不存在的项目，请重新下载",
            )

    lines_by_id, overrides_by_line, line_project_ids = _lock_global_refill_rows(
        db,
        line_ids=line_ids,
        allowed_project_ids=allowed_project_ids,
        prelocked_project_ids=set(workbook_states),
    )
    effective_refills: list[CostRefill] = []
    replayed_line_ids: list[int] = []
    for refill in sorted(plan.cost_refills, key=lambda item: item.line_id or -1):
        line = lines_by_id.get(int(refill.line_id))
        if line is None:
            raise WorkbookError(
                "project_scope_denied",
                "工作簿包含已不存在或不在当前可见项目范围的备件行，请重新下载",
            )
        override = overrides_by_line.get(int(refill.line_id))
        if _global_refill_matches_override(refill, override):
            replayed_line_ids.append(int(refill.line_id))
            continue
        expected_version = refill.base_override_version
        current_version = override.version if override is not None else None
        if current_version != expected_version:
            raise WorkbookError(
                "stale_cost_override",
                f"备件行 {refill.line_id} 的人工成本已被他人更新，请重新下载后再改",
            )
        if (
            override is not None
            and override.active != refill.base_override_active
        ):
            raise WorkbookError(
                "stale_cost_override",
                f"备件行 {refill.line_id} 的人工成本状态已变化，请重新下载后再改",
            )
        current_ex = _line_cost_evidence(line, override, basis="ex")["unit_cost"]
        current_inc = _line_cost_evidence(line, override, basis="inc")["unit_cost"]
        if (
            current_ex != refill.base_unit_cost_ex_tax
            or current_inc != refill.base_unit_cost_inc_tax
        ):
            raise WorkbookError(
                "stale_cost_fact",
                f"备件行 {refill.line_id} 的系统成本已变化，请重新下载后再改",
            )
        effective_refills.append(refill)

    effective_plan = MasterPlan(
        project_id=None,
        cost_refills=tuple(effective_refills),
        sheets=plan.sheets,
    )
    result = apply(
        db,
        effective_plan,
        operated_by=operated_by,
        import_batch_id=import_batch_id,
        _prelocked_workbook_states=workbook_states,
        _line_project_ids=line_project_ids,
    )
    # An ACK-loss replay reports the original requested count while making no
    # second mutation/version bump.
    result["cost_refills"] = len(plan.cost_refills)
    result["replayed_line_ids"] = replayed_line_ids
    return result


# ====================================================================== V2

@dataclass(frozen=True)
class V2MilestoneChange:
    operation: str
    contract_no: str
    sequence: int
    planned_date: date | None
    date_precision: str
    planned_amount: Decimal | None
    entity_id: str | None
    base_version: int | None


@dataclass(frozen=True)
class ContractAmountChange:
    """01 项目概览中唯一生效合同的含税金额变更。"""

    project_contract_id: str
    base_version: int
    before_amount_inc_tax: Decimal | None
    amount_inc_tax: Decimal


@dataclass(frozen=True)
class MasterV2Plan:
    project_id: str
    sheets: tuple[str, ...]
    export_id: str = ""
    file_sha256: str = ""
    expected_workbook_revision: int | None = None
    expected_workbook_data_version: str = ""
    cost_refills: tuple[CostRefill, ...] = ()
    site_flags: tuple[SiteReturnFlag, ...] = ()
    expense_updates: tuple[ec.ExpenseUpdate, ...] = ()
    receipt_ops: tuple[ec.CollectionOp, ...] = ()
    milestone_changes: tuple[V2MilestoneChange, ...] = ()
    assignment_changes: tuple[SourceOrderAssignmentChange, ...] = ()
    contract_amount_change: ContractAmountChange | None = None
    # 04 报销作废（#264/#267 契约）：显式 VOID 操作列 + 缺行=作废。
    expense_voids: tuple[str, ...] = ()
    will_void_rows: tuple[dict, ...] = ()
    # 2.7.0 行级三路合并产物：字段级变更/冲突/接管明细。
    force_takeover: bool = False
    field_changes: tuple[dict, ...] = ()
    conflicts: tuple[dict, ...] = ()
    overridden: tuple[dict, ...] = ()
    warnings: tuple[str, ...] = ()

    @property
    def summary(self) -> dict:
        updates = [x for x in self.cost_refills if not x.is_create and x.operation != "VOID"]
        return {
            # E2E #5：改数量不再被误报为 cost_overrides——行更新与成本覆盖分开计
            "line_updates": len(updates),
            "qty_updates": sum(x.qty is not None or x.return_qty is not None for x in updates),
            "cost_overrides": sum(
                x.unit_cost_ex_tax is not None or x.reason is not None for x in updates),
            "line_creates": sum(x.is_create for x in self.cost_refills),
            "line_voids": sum(x.operation == "VOID" for x in self.cost_refills),
            "order_reassignments": len(self.assignment_changes),
            "contract_updates": 1 if self.contract_amount_change is not None else 0,
            "expense_creates": sum(getattr(x, "is_create", False) for x in self.expense_updates),
            "expense_updates": sum(not getattr(x, "is_create", False) for x in self.expense_updates),
            "expense_voids": len(self.expense_voids),
            "plan_creates": sum(x.operation == "CREATE" for x in self.milestone_changes),
            "plan_updates": sum(x.operation == "UPDATE" for x in self.milestone_changes),
            "plan_voids": sum(x.operation == "VOID" for x in self.milestone_changes),
            "collection_updates": sum(
                operation.operation != "VOID" for operation in self.receipt_ops
            ),
            "collection_voids": sum(
                operation.operation == "VOID" for operation in self.receipt_ops
            ),
            "site_creates": sum(x.is_create for x in self.site_flags),
            "site_voids": sum(x.is_void for x in self.site_flags),
            "site_updates": sum(not x.is_create and not x.is_void
                                for x in self.site_flags),
        }


def _v2_hash_value(value) -> str:
    """哈希输入归一化：Excel 往返后 Decimal 变 float、date 变 datetime，
    直接 json 序列化会导致导出侧与解析侧哈希不一致（checkpoint 潜伏 bug）。"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float, Decimal)):
        return format(Decimal(str(value)).normalize(), "f")
    return str(value)


def _v2_hash(values: list[object]) -> str:
    return hashlib.sha256(json.dumps(
        [_v2_hash_value(v) for v in values],
        ensure_ascii=False, separators=(",", ":")).encode()).hexdigest()


def _v2_row_base_token(
    *, sheet: str, entity_id, values: dict[str, object]
) -> str:
    """行级签名基线（2.7.0）：该行可编辑字段的导出值快照 + HMAC。

    与全局表 ``_global_cost_base_token`` 同一模式（域分隔 + 规范 JSON），
    供上传侧三路合并区分「用户改的」与「文件里带出来的导出旧值」。
    """
    payload = {
        "sheet": sheet,
        "entity": str(entity_id),
        "base": {
            str(name): _v2_hash_value(values.get(name)) for name in sorted(values)
        },
    }
    payload_bytes = json.dumps(
        payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    key_id, signing_key = get_settings().maintenance_manifest_signing_material()
    signature = hmac.new(
        signing_key, _V2_ROW_BASE_DOMAIN + payload_bytes, hashlib.sha256
    ).hexdigest()
    return json.dumps(
        {"key_id": key_id, "payload": payload, "signature": signature},
        ensure_ascii=False, sort_keys=True, separators=(",", ":"),
    )


def _parse_v2_row_base_token(
    value, *, sheet: str, entity_id, row_no: int
) -> dict[str, str]:
    try:
        envelope = json.loads(str(value))
        key_id = str(envelope["key_id"])
        payload = envelope["payload"]
        signature = str(envelope["signature"])
        if (
            not isinstance(payload, dict)
            or payload.get("sheet") != sheet
            or str(payload.get("entity")) != str(entity_id)
        ):
            raise ValueError("row mismatch")
        payload_bytes = json.dumps(
            payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ).encode("utf-8")
        verification_key = (
            get_settings().maintenance_manifest_verification_keys().get(key_id)
        )
        if verification_key is None:
            raise ValueError("unknown key")
        expected = hmac.new(
            verification_key, _V2_ROW_BASE_DOMAIN + payload_bytes, hashlib.sha256
        ).hexdigest()
        if not hmac.compare_digest(expected, signature):
            raise ValueError("bad signature")
        base = payload.get("base")
        if not isinstance(base, dict):
            raise ValueError("bad base")
        return {str(k): ("" if v is None else str(v)) for k, v in base.items()}
    except (KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise WorkbookError(
            "invalid_concurrency_token",
            f"第 {row_no} 行基线令牌无效或与实体不匹配，请重新下载当前项目总表（2.7.0）",
        ) from exc


def _v2_row_base_hash(values: dict[str, object], fields: tuple[str, ...]) -> str:
    """行级基线哈希（进 99_元数据）：缺行=作废时判定「服务端该行是否已变」。"""
    return _v2_hash([_v2_hash_value(values.get(name)) for name in fields])


class _V2MergeContext:
    """validate 期间收集三路合并结果（冲突/变更/接管/PN 解析明细）。"""

    def __init__(self, *, force_takeover: bool) -> None:
        self.force_takeover = force_takeover
        self.changes: list[dict] = []
        self.conflicts: list[dict] = []
        self.overridden: list[dict] = []
        # 2026-09-03：手工新增行的 PN 柔性解析——成功规范化记入 warnings，
        # 失败批量收集（不再首错即断），validate 汇总一次性报全量。
        self.warnings: list[str] = []
        self.unresolved_pns: dict[str, list[int]] = {}

    def note_unresolved_pn(self, pn: str, row_no: int) -> None:
        self.unresolved_pns.setdefault(pn, []).append(row_no)

    def record(
        self, *, sheet: str, row_label: str, entity_id, field: str,
        old: str, new: str, conflict: bool, base: str = "",
    ) -> None:
        entry = {
            "sheet": sheet, "row": row_label, "entity_id": str(entity_id),
            "field": field, "old": old or "", "new": new or "",
        }
        if conflict:
            # 三值对照：base=导出基线 / old=服务端现值 / new=本次上传值。
            # force_takeover 时冲突当场解决：不进 conflicts（放行 apply），
            # 落 overridden 留痕并计入 changes。
            if self.force_takeover:
                enriched = {**entry, "base": base or ""}
                self.overridden.append(enriched)
                self.changes.append({**enriched, "overridden": True})
            else:
                self.conflicts.append({
                    **entry, "base": base or "",
                    "reason": "server_changed_since_export"})
        else:
            self.changes.append({**entry, "overridden": False})

    def record_row_conflict(
        self, *, sheet: str, row_label: str, entity_id, action: str,
    ) -> bool:
        """整行级冲突（删除行/作废行但服务端该行已变）。返回是否放行。"""
        entry = {
            "sheet": sheet, "row": row_label, "entity_id": str(entity_id),
            "field": "（整行）", "old": "", "new": action,
        }
        if self.force_takeover:
            self.overridden.append(dict(entry))
            return True
        self.conflicts.append({
            **entry, "reason": "server_changed_since_export"})
        return False


def _resolve_part_flexible(
    db: Session, pn_raw: str, *, row_no: int, sheet: str,
    merge: "_V2MergeContext",
) -> DimPart:
    """手工新增行的 PN 柔性解析（2026-09-03，中国电信云 1028 行实录教训）。

    依次尝试：精确（含别名）→ 空格归一 → 纯数字补前导零（华为/中兴货号
    6200288→06200288 实测）→ PN+WWN/固件号粘连取首段（首段须 ≥6 位且含
    数字，防描述性文本误中）。命中规范化路径记 warning；全部失败则记入
    批量收集器，validate 结束一次性报全量（不再首错即断——客户源数据
    一次几十个杂 PN，逐个报错要循环上传几十次）。
    """
    pn = str(pn_raw or "").strip()
    part = _exact_part_for_pn(db, pn)
    if part is not None:
        return part

    collapsed = " ".join(pn.split())
    if collapsed != pn:
        part = _exact_part_for_pn(db, collapsed)
        if part is not None:
            merge.warnings.append(
                f"{sheet}第 {row_no} 行 PN 空格已归一：{pn!r} → {part.pn_std}")
            return part

    if pn.isdigit():
        part = _exact_part_for_pn(db, "0" + pn)
        if part is not None:
            merge.warnings.append(
                f"{sheet}第 {row_no} 行货号已补前导零：{pn} → {part.pn_std}")
            return part

    first_token = collapsed.split(" ")[0] if " " in collapsed else ""
    if (len(first_token) >= 6 and any(ch.isdigit() for ch in first_token)
            and first_token.upper() == first_token):
        part = _exact_part_for_pn(db, first_token)
        if part is not None:
            merge.warnings.append(
                f"{sheet}第 {row_no} 行 PN 含粘连后缀已取标准段：{pn!r} → {part.pn_std}")
            return part

    merge.note_unresolved_pn(pn, row_no)
    return _exact_part_for_pn(db, pn)  # None；真正报错在 validate 汇总


def _unmatched_pn_error(
    db: Session, merge: "_V2MergeContext",
) -> "WorkbookError | None":
    if not merge.unresolved_pns:
        return None
    lines = []
    for pn, rows in sorted(merge.unresolved_pns.items()):
        suggestions = ", ".join(
            value for value in db.scalars(
                select(DimPart.pn_std).where(
                    DimPart.pn_std.ilike(f"%{pn[:18]}%"),
                    DimPart.status != "merged",
                ).limit(3)
            ).all()
        ) if len(pn) >= 4 else ""
        suffix = f"（相近：{suggestions}）" if suggestions else ""
        lines.append(f"{pn}（第 {'、'.join(str(r) for r in rows[:6])} 行等 {len(rows)} 行）{suffix}")
    return WorkbookError(
        "part_not_found",
        f"共 {len(merge.unresolved_pns)} 个 PN 未匹配备件主数据，请修正后重传："
        + "；".join(lines[:20])
        + ("……" if len(lines) > 20 else ""),
    )


def _v2_merge_row(
    *,
    sheet: str, row_label: str, entity_id, row, index: dict[str, int],
    base_fields: tuple[str, ...], server_values: dict[str, object],
    baseline: dict[str, str], ctx: _V2MergeContext,
) -> bool:
    """三路合并一行：返回该行是否被用户触碰。

    - 未触碰（用户值==导出基线）：返回 False，调用方跳过该行（服务端已变
      也不回写——自动 rebase，取代整本 stale 作废）。
    - 触碰且服务端未变：登记变更，走既有应用路径。
    - 触碰且服务端已变且与用户值不同：冲突；force_takeover 时登记接管放行。
    """
    user_values = {
        name: _v2_hash_value(_cell(row, index, name)) for name in base_fields
    }
    touched = {
        name: value for name, value in user_values.items()
        if value != baseline.get(name, "")
    }
    if not touched:
        return False
    for name, user_value in touched.items():
        server_value = _v2_hash_value(server_values.get(name))
        if user_value == server_value:
            continue
        ctx.record(
            sheet=sheet, row_label=row_label, entity_id=entity_id, field=name,
            old=server_value, new=user_value,
            conflict=server_value != baseline.get(name, ""),
            base=baseline.get(name, ""),
        )
    return True


def _v2_server_row_changed(
    *, server_values: dict[str, object], baseline: dict[str, str],
    fields: tuple[str, ...],
) -> bool:
    return any(
        _v2_hash_value(server_values.get(name)) != baseline.get(name, "")
        for name in fields
    )


def _v2_metadata_signature(metadata: dict[str, str], hmac_key: bytes) -> str:
    """Bind the complete export identity set to a server-only key.

    ``parts_row_ids`` / ``site_row_ids`` / ``expense_row_ids`` /
    ``receipt_row_ids`` drive the
    destructive "missing row = VOID" reconciliation.  Merely hiding those
    cells is not an integrity boundary, so every metadata field (including
    project and included sheets) is signed as one canonical envelope.
    """

    payload = json.dumps(
        {
            key: str(metadata[key])
            for key in sorted(metadata)
            if key != "metadata_hmac"
        },
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hmac.new(
        hmac_key,
        _V2_META_SIGNATURE_DOMAIN + payload,
        hashlib.sha256,
    ).hexdigest()


_META_CELL_SAFE_LENGTH = 30000
_META_CHUNK_SEP = "@"


def _chunk_meta_value(key: str, value: str) -> list[tuple[str, str]]:
    """Excel 单元格硬上限 32767：超长值（大项目 parts_base_hashes 实测 40KB+）
    会被静默截断，签名必炸（2.7.0 生产回归）。按 30000 分块为 key@2/@3...，
    读取端 _v2_meta 合并后再参与验签/使用。首块保留原名，兼容短值零分块。"""
    text = str(value)
    if len(text) <= _META_CELL_SAFE_LENGTH:
        return [(key, text)]
    chunks = [text[i:i + _META_CELL_SAFE_LENGTH]
              for i in range(0, len(text), _META_CELL_SAFE_LENGTH)]
    return [(key, chunks[0])] + [
        (f"{key}{_META_CHUNK_SEP}{n + 2}", chunk)
        for n, chunk in enumerate(chunks[1:])
    ]


def _v2_sign_meta_rows(rows: list[tuple[str, str]]) -> list[tuple[str, str]]:
    key_id, signing_key = get_settings().maintenance_manifest_signing_material()
    flat_rows: list[tuple[str, str]] = []
    for key, value in rows:
        flat_rows.extend(_chunk_meta_value(str(key), str(value)))
    signed_rows = [
        *flat_rows,
        ("metadata_hmac_algorithm", V2_META_SIGNATURE_ALGORITHM),
        ("metadata_hmac_key_id", key_id),
    ]
    # 签名覆盖「逻辑值（未分块）+ 算法/键ID」，与读取端合并后的全键一致。
    # 教训：3cf4f53 曾误签 dict(rows)（缺 algorithm/key_id 两键）导致全量必炸，
    # 由 test_v271 回归测试守住。
    metadata = dict(rows) | {
        "metadata_hmac_algorithm": V2_META_SIGNATURE_ALGORITHM,
        "metadata_hmac_key_id": key_id,
    }
    signed_rows.append(("metadata_hmac", _v2_metadata_signature(metadata, signing_key)))
    return signed_rows


def _v2_finalize(ws, headers: list[str], *, hidden_from: int | None = None,
                 editable: set[int] | None = None,
                 operation_col: int | None = None,
                 operation_choices: tuple[str, ...] = ("UPDATE", "VOID", "CREATE")) -> None:
    """收尾样式：黄底=可编辑（表头+数据区整列），操作列做成下拉只能选不能乱填。"""
    editable = editable or set()
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row or 1, len(headers)).coordinate}"
    for index, header in enumerate(headers, 1):
        cell = ws.cell(1, index)
        if index in editable:
            cell.fill = V2_EDITABLE
            cell.font = Font(bold=True, color="4F3B00")
        else:
            cell.fill = V2_HEADER
            cell.font = Font(bold=True, color="FFFFFF")
        width = min(60, max(12, len(str(header)) * 2 + 2))
        ws.column_dimensions[cell.column_letter].width = width
    if hidden_from is not None:
        for index in range(hidden_from, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, index).column_letter].hidden = True
    # 数据区可编辑列整列黄底（含空白新增行，向下多刷 20 行备用）
    if editable:
        last_row = (ws.max_row or 1) + 20
        for index in editable:
            letter = ws.cell(1, index).column_letter
            for r in range(2, last_row + 1):
                ws[f"{letter}{r}"].fill = V2_EDITABLE
    # 操作列：Excel 数据验证下拉（允许留空），杜绝自由输入拼错
    if operation_col is not None:
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(operation_choices) + '"',
            allow_blank=True,
            showDropDown=False,  # False=显示下拉箭头（openpyxl 语义反转）
        )
        dv.error = "操作列只能从下拉选择：{}".format("/".join(operation_choices))
        dv.errorTitle = "操作列取值无效"
        ws.add_data_validation(dv)
        letter = ws.cell(1, operation_col).column_letter
        dv.add(f"{letter}2:{letter}{(ws.max_row or 1) + 20}")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = min(72, max(18, 18 + 12 * sum("\n" in str(c.value or "") for c in row)))


def _v2_header(ws, headers: list[str], editable: set[int] = set()) -> None:
    ws.append(headers)
    for idx in range(1, len(headers) + 1):
        cell = ws.cell(1, idx)
        cell.fill = V2_EDITABLE if idx in editable else V2_HEADER
        cell.font = Font(bold=True, color="FFFFFF" if idx not in editable else "4F3B00")


def _project_order_salesperson(db: Session, project_id: str) -> str | None:
    """项目销售人员（2026-08-20 用户口径）：本项目挂靠需求单上的销售众数。

    调用方仅可在 canonical salesperson 未人工覆盖时使用这个遗留兜底值。
    """
    from app.services import maintenance_demands

    rows = db.execute(
        select(FMaintenanceOrder.salesperson, func.count())
        .join(MaintenanceSourceOrderAssignment,
              (MaintenanceSourceOrderAssignment.source_order_id
               == FMaintenanceOrder.raw_order_id)
              & MaintenanceSourceOrderAssignment.is_active.is_(True))
        .where(
            MaintenanceSourceOrderAssignment.project_id == project_id,
            FMaintenanceOrder.salesperson.is_not(None),
            FMaintenanceOrder.salesperson != "",
            maintenance_demands.active_demand_condition(),
        )
        .group_by(FMaintenanceOrder.salesperson)
        .order_by(func.count().desc())
        .limit(1)
    ).all()
    return rows[0][0] if rows else None


def _account_display_name(db: Session, username: str | None) -> str | None:
    """账号 → 显示人名（找不到回退账号本身）。"""
    if not username:
        return None
    from app.models.system import SysUser

    row = db.execute(
        select(SysUser.display_name).where(SysUser.username == username)
    ).scalar_one_or_none()
    return row or username


_EXAMPLE_FONT = Font(color="999999", italic=True)
_EXAMPLE_FILL = PatternFill("solid", fgColor="F5F5F5")


def _v2_append_example_row(ws, headers: list[str], values: dict[str, object]) -> None:
    """每个数据 sheet 底部一行灰色「示例」：告诉用户怎么填，上传时系统忽略。

    标记约定：操作列填「示例」（02/03/04）；无操作列的 sheet（05/06）在备注列
    填「【示例】…」。解析侧 _is_example_row 统一识别跳过。
    """
    row = [values.get(name, "") for name in headers]
    ws.append(row)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(r, c)
        cell.font = _EXAMPLE_FONT
        cell.fill = _EXAMPLE_FILL


def _is_example_row(row) -> bool:
    """上传侧识别示例行：任一单元格为「示例」或以「【示例】」开头。"""
    for value in row or ():
        text = str(value or "").strip()
        if text == "示例" or text.startswith("【示例】"):
            return True
    return False


def _line_cost_evidence(line, override, *, basis: str) -> dict:
    """Canonical normalized line cost for workbook and row-API consumers."""
    normalized_amount = (
        line.cost_amount_inc_tax if basis == "inc" else line.cost_amount_ex_tax
    )
    manual_unit = None
    if override is not None:
        manual_unit = (
            override.unit_cost_inc_tax
            if basis == "inc"
            else override.unit_cost_ex_tax
        )
    fact = maintenance_cost_quality.normalized_line_cost(
        source=line.cost_source,
        tax_basis=line.cost_tax_basis,
        legacy_amount=line.cost_amount,
        normalized_amount=normalized_amount,
        normalized_basis=basis,
        anomaly_flags=line.anomaly_flags,
        qty=line.qty,
        return_qty=line.return_qty,
        manual_unit_cost=manual_unit,
        manual_active=override is not None and override.active is True,
    )
    if fact["tier"] == "missing":
        unit = None
    elif fact["source"] == "manual" and override is not None:
        unit = manual_unit
    else:
        unit = line.unit_cost_inc_tax if basis == "inc" else line.unit_cost_ex_tax
    return {**fact, "unit_cost": unit}


def _v2_build_overview(wb, project, contracts, db, lines) -> None:
    ws = wb.create_sheet(V2_SHEET_OVERVIEW)
    ws.append(["项目总览", None])
    ws.merge_cells("A1:B1")
    ws["A1"].fill = V2_TITLE
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    card = _canonical_contract_read_model(db, project.project_id)
    # 展示与卡片共用同一总额函数；不在工作簿里再发明第三套税口径。
    # canonical read model 可能携带诊断用的已知小计；总表必须在不完整/共享时
    # fail closed，不能把小计标成“合同总额”或据此计算成本率。
    total_contract = card["amount_inc_tax"]
    contract_shared = card["contract_shared"]
    contract_incomplete = card["contract_incomplete"]
    editable_contracts = _v2_editable_contracts(db, project.project_id, contracts)
    today = business_today()
    current_included = [
        contract for contract in contracts
        if contract.included_in_total
        and contract.effective_from <= today
        and (contract.effective_to is None or contract.effective_to > today)
    ]
    if len(editable_contracts) == 1:
        contract_edit_state = (
            "可编辑：写回唯一当前含税合同事实"
            + ("（注意：该合同跨项目共享，修改将影响所有计入项目）"
               if contract_shared else "")
        )
    elif contract_shared:
        contract_edit_state = "只读：多份当前计入合同且存在共享，须先收敛合同关系"
    elif len(current_included) > 1:
        contract_edit_state = "只读：当前有多份合同，须逐合同维护，不能把项目总额写入其中一份"
    elif not current_included:
        contract_edit_state = "只读：仅有 XSDD 参考或尚无合同台账，请先建立唯一合同关系"
    else:
        contract_edit_state = "只读：当前合同事实不完整，请先完成合同治理"
    # 2026-08-19：备件成本合并人工覆盖——主表无成本但有 override 的行按
    # override 含税金额×数量计入（与看板/面板口径一致）
    line_ids = [line.id for line, _order, _pid in lines]
    override_map = {
        item.line_id: item for item in db.scalars(
            select(MaintenanceManualCostOverride).where(
                MaintenanceManualCostOverride.line_id.in_(line_ids),
                MaintenanceManualCostOverride.active.is_(True),
            )
        )
    } if line_ids else {}

    cost_facts = [
        _line_cost_evidence(line, override_map.get(line.id), basis="inc")
        for line, _order, _pid in lines
    ]
    missing_cost_lines = sum(
        fact["tier"] == "missing" for fact in cost_facts
    )
    cost = sum(
        (fact["amount"] for fact in cost_facts if fact["amount"] is not None),
        Decimal("0"),
    )
    cost_complete = bool(cost_facts) and missing_cost_lines == 0
    values = [
        ("项目编号", project.project_code), ("项目名称", project.display_name),
        ("生命周期", project.lifecycle_status), ("服务期", f"{project.period_from or '—'} ~ {project.period_to or '—'}"),
        # 负责人＝项目经理（显示人名）；销售人员＝canonical 优先，未人工覆盖的
        # 遗留空值才回落到挂靠需求单众数；CMO＝台账来源（无台账则缺）。
        ("项目经理（负责人）",
         _account_display_name(db, project.project_manager_id) or "未关联账号"),
        (
            "销售人员",
            (
                project.salesperson
                if project.salesperson_override_active
                else (
                    project.salesperson
                    or _project_order_salesperson(db, project.project_id)
                )
            )
            or "—",
        ),
        ("CMO", project.cmo_name or "—"),
        ("合同编号", "、".join(c.contract_no for c in contracts) or "—"),
        ("合同总额（含税）", str(total_contract) if total_contract is not None else "—"),
        ("合同总额编辑状态", contract_edit_state),
        ("合同额口径", (
            "合同事实不完整/存在共享冲突"
            if contract_shared or contract_incomplete
            else "含税合同事实"
        )),
        # 完整且为 0 是真实零；有任一缺口时已知小计不是完整项目成本，不能冒充。
        ("备件成本（含税）", str(cost) if cost_complete else "—"),
        ("成本率", (
            f"{(cost / total_contract * 100).quantize(Decimal('0.1'))}%"
            if cost_complete and total_contract and total_contract > 0 else "—"
        )),
        ("缺成本行数", missing_cost_lines),
    ]
    for item in values:
        ws.append(list(item))
        if item[0] == "合同总额（含税）":
            # 只有“唯一生效且计入总额”的合同，项目总额才能无歧义地回写到一条
            # 合同事实。多合同不自动分摊；无合同不凭空创建身份。
            if len(editable_contracts) == 1:
                ws.cell(ws.max_row, 2).fill = V2_EDITABLE
            else:
                ws.cell(ws.max_row, 2).fill = V2_READONLY
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 54
    ws.freeze_panes = "A2"


def _v2_editable_contracts(
    db: Session,
    project_id: str,
    contracts,
) -> list[MaintenanceProjectContract]:
    today = business_today()
    included_current = [
        contract for contract in contracts
        if contract.included_in_total
        and contract.effective_from <= today
        and (contract.effective_to is None or contract.effective_to > today)
    ]
    # 2026-09-02 拍板：合同额对项目负责人/销售与管理员全量放开。
    # 仍要求「当前计入的合同恰好一份」，保证项目总额能一一对应回写到唯一
    # 合同事实；mapped/共享不再硬拒——共享合同可编辑，apply 侧以
    # contract base_version CAS + 审计兜底，共享状态在导出侧提示。
    if len(included_current) != 1:
        return []
    only = included_current[0]
    if only.project_id != project_id:
        return []
    return [only]


def _v2_build_plan(wb, db, project_id: str, contracts) -> None:
    ws = wb.create_sheet(V2_SHEET_PLAN)
    _v2_header(ws, V2_PLAN_HEADERS, editable={1, 2, 3, 4, 5, 6, 12})
    contract_by_id = {c.project_contract_id: c.contract_no for c in contracts}
    milestones = list(db.scalars(select(MaintenanceCollectionMilestone).where(
        MaintenanceCollectionMilestone.project_id == project_id,
        MaintenanceCollectionMilestone.is_active.is_(True),
        MaintenanceCollectionMilestone.project_contract_id.in_(contract_by_id),
    ).order_by(MaintenanceCollectionMilestone.project_contract_id, MaintenanceCollectionMilestone.sequence)))
    cumulative: dict[str, Decimal] = {}
    for milestone in milestones:
        contract_no = contract_by_id.get(milestone.project_contract_id, "")
        cumulative[contract_no] = cumulative.get(contract_no, Decimal(0)) + (milestone.planned_amount or Decimal(0))
        ws.append([
            "", contract_no, milestone.sequence, milestone.planned_date,
            milestone.date_precision, milestone.planned_amount,
            cumulative[contract_no], "—", "未上报", milestone.follow_up_status,
            project_id, milestone.follow_up_note or "", milestone.milestone_id, milestone.version,
        ])
    for _ in range(4):
        ws.append([""] * len(V2_PLAN_HEADERS))
    sample_contract = ((contracts[0].contract_no if contracts else None)
                       or next(iter(project_sales_order_nos(db, project_id)), ""))
    _v2_append_example_row(ws, V2_PLAN_HEADERS, {
        "操作": "示例",
        "合同编号": sample_contract or "（先在需求单挂靠后自动获得 XSDD）",
        "期次": 1,
        "计划回款日期": "2026-09-30",
        "日期精度": "day",
        "计划回款金额（含税）": 50000,
        "备注": "新增一条计划：操作选 CREATE，按合同节点填期次/日期/金额",
    })
    _v2_finalize(ws, V2_PLAN_HEADERS, hidden_from=13,
                 editable={1, 2, 3, 4, 5, 6, 12}, operation_col=1)


def _v2_part_row_values(line, order, override) -> dict[str, object]:
    """03 一行的全部展示值（按表头名）。build/parse 共用，保证哈希口径一致。"""
    has_cost = line.unit_cost_ex_tax is not None
    return {
        "操作": "",
        "维保单号": order.order_no or "",
        "制单日期": order.order_date,
        "XSDD": order.linked_sales_order_no or "",
        "需求类型": order.demand_type or "",
        "仓库": order.warehouse or "",
        "销售人员": order.salesperson or "",
        "业务类型": order.business_type or "",
        "PN": line.pn_std or line.pn_raw or "",
        "描述": line.description or "",
        "需求数量": line.qty,
        "SN": line.serial_numbers or "",
        "退货数量": line.return_qty,
        "成本来源": line.cost_source or "",
        "置信度": line.confidence or "none",
        "系统未税单位成本": line.unit_cost_ex_tax,
        "系统含税单位成本": line.unit_cost_inc_tax,
        "人工未税单位成本": override.unit_cost_ex_tax if override else None,
        "人工成本原因": override.reason if override else None,
        "成本缺失类型": (
            "out_of_scope" if line.cost_source is None and not has_cost
            else ("none" if not has_cost else "")
        ),
        "可补价": "是" if line.cost_source in (None, "none") and not has_cost else "否",
        "实体ID": line.id,
        "备件主键": line.part_id,
        "只读哈希": "",  # 占位，下方按锁定列回填
        "备注": line.line_note or "",
        "来源": "工作簿" if line.edited_source == "workbook_manual" else "WBDD",
    }


def _v2_build_parts(wb, db, project_id: str, lines) -> None:
    ws = wb.create_sheet(V2_SHEET_PARTS)
    _v2_header(ws, V2_PART_HEADERS, editable=V2_PART_EDITABLE)
    line_ids = [line.id for line, _order, _pid in lines]
    overrides = {
        item.line_id: item for item in db.scalars(select(MaintenanceManualCostOverride).where(
            MaintenanceManualCostOverride.line_id.in_(line_ids),
            MaintenanceManualCostOverride.active.is_(True),
        ))
    } if line_ids else {}
    base_hashes: dict[str, str] = {}
    for line, order, _pid in lines:
        values = _v2_part_row_values(line, order, overrides.get(line.id))
        digest = _v2_hash([values[name] for name in V2_PART_HASH_COLUMNS])
        values["只读哈希"] = digest
        values[V2_BASE_COLUMN] = _v2_row_base_token(
            sheet="03_备件明细", entity_id=line.id,
            values={name: values.get(name) for name in V2_PART_BASE_FIELDS})
        base_hashes[str(line.id)] = _v2_row_base_hash(values, V2_PART_BASE_FIELDS)
        ws.append([values[name] for name in V2_PART_HEADERS])
    # 空白新增行（无实体ID）：用户填操作=CREATE 的新明细
    for _ in range(5):
        ws.append([""] * len(V2_PART_HEADERS))
    sample_line = lines[0] if lines else None
    _v2_append_example_row(ws, V2_PART_HEADERS, {
        "操作": "示例",
        "维保单号": (sample_line[1].order_no if sample_line else "（本项目已有需求单号）"),
        "PN": ((sample_line[0].pn_std or sample_line[0].pn_raw or "") if sample_line
               else "（标准PN）"),
        "描述": "新增一行备件",
        "需求数量": 2,
        "退货数量": 0,
        "人工未税单位成本": 88.5,
        "人工成本原因": "采购价依据",
        "备注": "新增示例：操作选 CREATE，必须挂本项目已有单号",
    })
    _v2_finalize(ws, V2_PART_HEADERS,
                 editable=V2_PART_EDITABLE, operation_col=1)
    # 仅隐藏技术列：实体ID(22)/备件主键(23)/只读哈希(24)/基线令牌(27)。
    # 备注(25)/来源(26) 可见。
    for col in (22, 23, 24, len(V2_PART_HEADERS)):
        ws.column_dimensions[ws.cell(1, col).column_letter].hidden = True
    return base_hashes


def _v2_expense_row_values(expense, project_display: str) -> dict[str, object]:
    """04 一行的展示值（按表头名）。build/parse 共用，保证基线口径一致。"""
    return {
        "操作": "",
        "费用单号": expense.bxd_no or "",
        "明细序号": expense.line_no,
        "报销日期": expense.expense_date,
        "报销人员": expense.person or "",
        "报销类别": expense.expense_type or "",
        "费用分类": expense.fee_category or "",
        "支出事由": expense.reason or "",
        "维保销售订单（归集键）": expense.linked_sales_order_no or "",
        "项目名称": project_display,
        "销售人员": "—",
        "流程状态": expense.data_status or "",
        "原始报销金额": expense.amount,
        "金额口径": expense.tax_basis or "ex",
        "未税金额": expense.amount_ex_tax,
        "含税金额（系统）": expense.amount_inc_tax,
        "备注": expense.remark or "",
        "实体ID": expense.raw_line_id,
    }


def _v2_build_expense(wb, db, project_id: str, contracts, project=None) -> dict[str, str]:
    ws = wb.create_sheet(V2_SHEET_EXPENSE)
    # 金额口径 is source provenance and cannot be edited independently.  The
    # user edits the explicit ex-tax column; apply preserves the source basis.
    _v2_header(ws, V2_EXPENSE_HEADERS, editable={1, 4, 5, 6, 7, 8, 9, 12, 15, 17})
    expenses = _project_expenses(db, project_id)
    base_hashes: dict[str, str] = {}
    project_display = (
        project.display_name if getattr(project, "display_name", None) else project_id
    )
    for expense in expenses:
        values = _v2_expense_row_values(expense, project_display)
        values[V2_BASE_COLUMN] = _v2_row_base_token(
            sheet="04_费用报销", entity_id=expense.raw_line_id,
            values={name: values.get(name) for name in V2_EXPENSE_BASE_FIELDS})
        base_hashes[str(expense.raw_line_id)] = _v2_row_base_hash(
            values, V2_EXPENSE_BASE_FIELDS)
        ws.append([values[name] for name in V2_EXPENSE_HEADERS])
    _v2_append_example_row(ws, V2_EXPENSE_HEADERS, {
        "操作": "示例",
        "费用单号": "BXD-20260901-0001",
        "明细序号": 1,
        "报销日期": "2026-09-01",
        "报销人员": "张三",
        "报销类别": "维保费用",
        "费用分类": "交通费",
        "支出事由": "现场交通",
        "维保销售订单（归集键）": "（填本项目XSDD合同号）",
        "原始报销金额": 226,
        "未税金额": 200,
        "备注": "手工新增示例：实体ID留空，费用单号+明细序号必填",
    })
    _v2_finalize(ws, V2_EXPENSE_HEADERS, hidden_from=18,
                 editable={1, 4, 5, 6, 7, 8, 9, 12, 15, 17}, operation_col=1)
    return base_hashes


def _v2_receipt_row_values(snapshot, contract_no: str) -> dict[str, object]:
    """05 一行的展示值（按表头名）。build/parse 共用。"""
    month = snapshot.report_month
    month_text = (
        month.isoformat()[:7] if hasattr(month, "isoformat") else str(month or "")
    )
    return {
        "合同编号": contract_no,
        "报告月份": month_text,
        "累计实收金额（含税）": snapshot.cumulative_amount,
        "状态": snapshot.status,
        "回款凭证号": snapshot.receipt_reference or "",
        "备注": snapshot.remark or "",
        "实体ID": snapshot.collection_id,
    }


def _v2_build_receipts(wb, db, project_id: str, contracts) -> dict[str, str]:
    ws = wb.create_sheet(V2_SHEET_RECEIPTS)
    _v2_header(ws, V2_RECEIPT_HEADERS)
    contract_by_id = {c.project_contract_id: c.contract_no for c in contracts}
    rows = list(db.scalars(select(MaintenanceCollectionSnapshot).where(
        MaintenanceCollectionSnapshot.project_id == project_id,
        MaintenanceCollectionSnapshot.status == "confirmed",
    ).order_by(MaintenanceCollectionSnapshot.report_month)))
    base_hashes: dict[str, str] = {}
    for row in rows:
        contract_no = contract_by_id.get(row.project_contract_id, "")
        values = _v2_receipt_row_values(row, contract_no)
        values[V2_BASE_COLUMN] = _v2_row_base_token(
            sheet="05_实收回款", entity_id=row.collection_id,
            values={name: values.get(name) for name in V2_RECEIPT_BASE_FIELDS})
        base_hashes[str(row.collection_id)] = _v2_row_base_hash(
            values, V2_RECEIPT_BASE_FIELDS)
        ws.append([values[name] for name in V2_RECEIPT_HEADERS])
    _v2_append_example_row(ws, V2_RECEIPT_HEADERS, {
        "合同编号": next(iter(contract_by_id.values()), "") or "（本项目合同号）",
        "报告月份": "2026-09",
        "累计实收金额（含税）": 50000,
        "状态": "confirmed",
        "回款凭证号": "PJ-202609-001",
        "备注": "【示例】实收=每月累计快照：同一合同同月重复上传即覆盖更新，凭证号选填",
    })
    _v2_finalize(ws, V2_RECEIPT_HEADERS, hidden_from=7)
    return base_hashes


def _v2_site_row_values(line, issue) -> dict[str, object]:
    """06 一行的展示值（按表头名）。build/parse 共用。"""
    return {
        "领用单号": issue.issue_no,
        "领用日期": issue.issue_date,
        "PN": line.pn,
        "SN": line.serial_number or "",
        "领用数量": line.quantity,
        "是否应返还": (
            "" if line.no_return is None else ("否" if line.no_return else "是")
        ),
        "备注": line.remark or "",
        "应返数量": "—",
        "返还状态": "待确认品类",
        "返还单号": "—",
        "实体ID": line.issue_line_id,
    }


def _v2_build_site(wb, db, project_id: str) -> dict[str, str]:
    ws = wb.create_sheet(V2_SHEET_SITE)
    _v2_header(ws, V2_SITE_HEADERS, editable={1, 2, 3, 4, 5, 6, 7})
    rows = db.execute(select(MaintenanceSiteIssueLine, MaintenanceSiteIssue).join(
        MaintenanceSiteIssue, MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id
    ).where(MaintenanceSiteIssue.project_id == project_id,
            MaintenanceSiteIssueLine.is_active.is_(True)
    ).order_by(MaintenanceSiteIssue.issue_date, MaintenanceSiteIssueLine.line_no)).all()
    base_hashes: dict[str, str] = {}
    for line, issue in rows:
        values = _v2_site_row_values(line, issue)
        values[V2_BASE_COLUMN] = _v2_row_base_token(
            sheet="06_领用返还", entity_id=line.issue_line_id,
            values={name: values.get(name) for name in V2_SITE_BASE_FIELDS})
        base_hashes[str(line.issue_line_id)] = _v2_row_base_hash(
            values, V2_SITE_BASE_FIELDS)
        ws.append([values[name] for name in V2_SITE_HEADERS])
    _v2_append_example_row(ws, V2_SITE_HEADERS, {
        "领用单号": "CKD-20260901-0001",
        "领用日期": "2026-09-01",
        "PN": (rows[0][0].pn if rows else "（标准PN）"),
        "SN": "SN-001",
        "领用数量": 1,
        "是否应返还": "是",
        "备注": "【示例】手工新增领用：实体ID留空，单号/日期/PN/数量必填",
    })
    _v2_finalize(ws, V2_SITE_HEADERS, hidden_from=11,
                 editable={1, 2, 3, 4, 5, 6, 7})
    return base_hashes


def _v2_build_usage(wb) -> None:
    """00_使用说明：怎么改、怎么删、哪些能改（黄底）、防呆规则。"""
    ws = wb.create_sheet(V2_SHEET_USAGE)
    rows = [
        ("如何使用本项目总表", ""),
        ("", ""),
        ("【总原则】", "黄底单元格 = 可以修改；其他颜色/无底色 = 系统只读，改了会被拒绝并要求重新下载。"),
        ("", "每个数据行的第 1 列是「操作」下拉框：留空=更新该行 / UPDATE=更新 / VOID=作废该行 / CREATE=新增行（只用于表尾空白行）。"),
        ("", ""),
        ("【03 备件明细】", ""),
        ("改数量/退货数量", "直接改黄底单元格，回传后系统按 数量-退货数量×单价 重算金额。"),
        ("改 PN", "填新的标准 PN（必须能匹配备件主数据），回传后行会整体换绑新备件。"),
        ("新增一行", "在表尾空白行：操作列选 CREATE，填 维保单号（必须是本项目已有需求单）+ PN + 数量。"),
        ("删除一行", "两种等价方式：① 操作列选 VOID；② 直接把该行整行删掉。回传后该行作废，不再进入任何统计与导出。"),
        ("整单删除", "把该单的所有行都删掉/标 VOID，回传后整张需求单自动作废（可由管理员在系统内恢复）。"),
        ("【04 费用报销】", ""),
        ("删除一行", "直接删行或操作列选 VOID。作废的行从此不再导出、不计金额。"),
        ("新增一行", "表尾填 费用单号+明细序号+报销日期+金额+归集键（XSDD），操作列留空或 CREATE。"),
        ("【02 回款计划】", "计划=打算什么时候收多少钱：一行一个合同期次。操作选 CREATE 新增，填合同号、期次（第几期）、计划回款日期、金额；改已有行选 UPDATE（带基础版本防冲突）；作废选 VOID。"),
        ("【05 实收回款】", "实收=每月实际收到的累计数：同一合同同一月份只保留一行，报告月份填 YYYY-MM，金额填「截至该月累计实收」（不是当月增量）。同月重复上传=覆盖更新，凭证号选填。"),
        ("【06 领用返还】", "按黄底提示编辑；手工新增领用行填单号/日期/PN/数量，实体ID留空。"),
        ("删除领用行", "2026-08-23 起：直接把该行整行删掉再上传 = 该领用行作废（退出成本与返还计算）；一张单的行全删 = 整单作废。"),
        ("【灰色示例行】", "每个数据页最后一行灰色斜体是填写示例，系统上传时自动忽略，不会入库——照着它的格式填，填完可保留或删除该行。"),
        ("", ""),
        ("【安全规则】", ""),
        ("行数防呆", "上传的数据行少于导出时的一半 → 整本拒绝（防止筛选后误传删掉大片数据）。"),
        ("缺行判定", "「删行=作废」只对下载时存在的行生效；下载之后系统里新导入的行不受影响。"),
        ("原样回传", "基于「本次下载」的文件什么都不改直接传回去 = 零变更（幂等）。注意：已经应用过一次的旧文件再传，其中已作废的行会被拒绝（提示重新下载），这是正常的保护。"),
        ("改错了怎么办", "行级作废需管理员在系统内恢复；整单作废同。恢复入口：维保需求单页面 → 搜索（勾选「含已作废」）→ 恢复。"),
    ]
    for k, v in rows:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 100
    ws["A1"].fill = V2_TITLE
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws.merge_cells("A1:B1")
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1)
        if a.value and str(a.value).startswith("【"):
            a.font = Font(bold=True)
        ws.cell(r, 2).alignment = ws.cell(r, 2).alignment.copy(wrap_text=True, vertical="top")


def _v2_build_dictionary(wb) -> None:
    ws = wb.create_sheet(V2_SHEET_DICTIONARY)
    headers = ["字段", "业务含义", "来源/表字段", "编辑性", "格式/允许值", "空值意义", "覆盖影响"]
    _v2_header(ws, headers)
    rows = [
        ("part_id", "备件主键", "f_maintenance_line.part_id → dim_part.id", "只读", "整数", "异常需治理", "关系身份不随 PN 文本改变"),
        ("人工未税单位成本", "缺成本时的人工补价", "maintenance_manual_cost_override.unit_cost_ex_tax", "可编辑", "非负金额", "不创建 override", "覆盖该行成本人工证据"),
        ("计划回款金额", "合同期次计划金额", "maintenance_collection_milestone.planned_amount", "可编辑", "正数", "节点 incomplete", "更新计划节点并触发提醒复核"),
        ("最新累计实收", "已确认回款快照累计额", "maintenance_collection_snapshot.cumulative_amount", "只读", "金额", "未上报，不等于0", "仅影响展示和到款状态"),
        ("实体ID", "稳定写回身份", "各事实表主键", "只读", "UUID/字符串", "非法文件", "缺失或跨项目拒绝"),
    ]
    for row in rows:
        ws.append(row)
    _v2_finalize(ws, headers)


def build_project_master_v2(
    db: Session,
    *,
    project_id: str,
    sheets: tuple[str, ...] | None = None,
) -> bytes | None:
    if db.scalar(
        select(MaintenanceProject.project_id).where(
            MaintenanceProject.project_id == project_id
        )
    ) is None:
        return None
    # Freeze one coherent project snapshot for the complete multi-query export.
    # Every V2-visible writer takes this state row before project/fact rows, so
    # holding it through workbook serialization prevents a mixed-sheet download
    # instead of merely making that freshly downloaded file immediately stale.
    from app.services import maintenance_project_operations as operations

    workbook_state = operations.lock_workbook_states(
        db,
        project_ids={project_id},
    )[project_id]
    project = db.scalar(
        select(MaintenanceProject)
        .where(MaintenanceProject.project_id == project_id)
        .with_for_update()
    )
    if project is None:
        return None
    exported_revision = workbook_state.revision
    exported_data_version = workbook_state.data_version
    contracts = ec._contracts(db, project_id)
    current_contracts = _current_contracts(db, project_id)
    wanted = tuple(sheets or V2_ALL_SHEETS[:-2])
    unknown = [name for name in wanted if name not in V2_ALL_SHEETS[:-2]]
    if unknown:
        raise WorkbookError("unsupported_sheet", f"不支持的 V2 工作表：{'、'.join(unknown)}")
    lines = (_assigned_lines(db, project_id=project_id, window=None)
             if V2_SHEET_OVERVIEW in wanted or V2_SHEET_PARTS in wanted else [])
    wb = Workbook()
    wb.remove(wb.active)
    base_hash_maps: dict[str, dict[str, str]] = {}
    if V2_SHEET_OVERVIEW in wanted:
        _v2_build_overview(wb, project, contracts, db, lines)
    if V2_SHEET_PLAN in wanted:
        _v2_build_plan(wb, db, project_id, current_contracts)
    if V2_SHEET_PARTS in wanted:
        base_hash_maps["parts"] = _v2_build_parts(wb, db, project_id, lines)
    if V2_SHEET_EXPENSE in wanted:
        base_hash_maps["expense"] = _v2_build_expense(
            wb, db, project_id, contracts, project=project)
    if V2_SHEET_RECEIPTS in wanted:
        base_hash_maps["receipt"] = _v2_build_receipts(
            wb, db, project_id, current_contracts)
    if V2_SHEET_SITE in wanted:
        base_hash_maps["site"] = _v2_build_site(wb, db, project_id)
    _v2_build_dictionary(wb)
    _v2_build_usage(wb)
    meta_rows = [
        ("protocol_id", V2_PROTOCOL_ID), ("template_version", V2_TEMPLATE_VERSION),
        ("project_id", project_id), ("export_id", str(uuid4())),
        ("exported_at", datetime.now(timezone.utc).isoformat()),
        ("included_sheets", ",".join(wanted)),
        ("workbook_revision", str(exported_revision)),
        ("workbook_data_version", exported_data_version),
    ]
    if V2_SHEET_OVERVIEW in wanted:
        editable_contracts = _v2_editable_contracts(db, project_id, contracts)
        card = _canonical_contract_read_model(db, project_id)
        exported_total = card["amount_inc_tax"]
        meta_rows.extend([
            ("contract_total_exported",
             "" if exported_total is None else str(exported_total)),
            ("contract_editable",
             "true" if len(editable_contracts) == 1 else "false"),
        ])
        if len(editable_contracts) == 1:
            editable_contract = editable_contracts[0]
            meta_rows.extend([
                ("contract_edit_id", editable_contract.project_contract_id),
                ("contract_edit_version", str(editable_contract.version)),
            ])
    if V2_SHEET_PLAN in wanted:
        from app.models.maintenance_manager import MaintenanceCollectionMilestone

        plan_ids = db.scalars(select(MaintenanceCollectionMilestone.milestone_id).where(
            MaintenanceCollectionMilestone.project_id == project_id,
            MaintenanceCollectionMilestone.is_active.is_(True),
            MaintenanceCollectionMilestone.project_contract_id.in_(
                [contract.project_contract_id for contract in current_contracts]
            ),
        )).all()
        meta_rows.append(("plan_row_ids", _encode_row_ids(plan_ids)))
    if V2_SHEET_PARTS in wanted:
        meta_rows.append(("parts_row_ids",
                          _encode_row_ids([ln.id for ln, _o, _p in lines])))
        meta_rows.append(("parts_base_hashes",
                          _encode_base_hashes(base_hash_maps.get("parts", {}))))
    if V2_SHEET_EXPENSE in wanted:
        meta_rows.append(("expense_row_ids",
                          _encode_row_ids(_expected_expense_ids(db, project_id))))
        meta_rows.append(("expense_base_hashes",
                          _encode_base_hashes(base_hash_maps.get("expense", {}))))
    if V2_SHEET_RECEIPTS in wanted:
        receipt_ids = db.scalars(
            select(MaintenanceCollectionSnapshot.collection_id).where(
                MaintenanceCollectionSnapshot.project_id == project_id,
                MaintenanceCollectionSnapshot.status == "confirmed",
            )
        ).all()
        meta_rows.append(("receipt_row_ids", _encode_row_ids(receipt_ids)))
        meta_rows.append(("receipt_base_hashes",
                          _encode_base_hashes(base_hash_maps.get("receipt", {}))))
    if V2_SHEET_SITE in wanted:
        site_line_ids = db.scalars(
            select(MaintenanceSiteIssueLine.issue_line_id)
            .join(MaintenanceSiteIssue,
                  MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id)
            .where(MaintenanceSiteIssue.project_id == project_id,
                   MaintenanceSiteIssueLine.is_active.is_(True))
            .order_by(MaintenanceSiteIssue.issue_date, MaintenanceSiteIssueLine.line_no)
        ).all()
        meta_rows.append(("site_row_ids", _encode_row_ids(site_line_ids)))
        meta_rows.append(("site_base_hashes",
                          _encode_base_hashes(base_hash_maps.get("site", {}))))
    meta = wb.create_sheet(V2_SHEET_META)
    for key, value in _v2_sign_meta_rows(meta_rows):
        meta.append([key, value])
    meta.sheet_state = "hidden"
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _v2_date(value, *, row_no: int, label: str) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S", "%Y-%m"):
        try:
            parsed = datetime.strptime(raw, fmt).date()
            return parsed.replace(day=1) if fmt == "%Y-%m" else parsed
        except ValueError:
            continue
    raise WorkbookError("invalid_date", f"第 {row_no} 行{label}日期格式无效")


def _v2_decimal(
    value,
    *,
    row_no: int,
    label: str,
    required: bool = False,
    allow_negative: bool = False,
) -> Decimal | None:
    if value in (None, ""):
        if required:
            raise WorkbookError("missing_amount", f"第 {row_no} 行{label}不能为空")
        return None
    try:
        parsed = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}不是合法数字")
    normalized = parsed.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if (
        not normalized.is_finite()
        or abs(normalized) >= Decimal("1000000000000")
        or (normalized < 0 and not allow_negative)
    ):
        raise WorkbookError("invalid_amount", f"第 {row_no} 行{label}超出允许范围")
    return normalized


def _cell(row, index: dict[str, int], name: str):
    return row[index[name]] if name in index and index[name] < len(row) else None


def _exact_part_for_pn(db: Session, pn: str) -> DimPart | None:
    normalized = pn.strip().upper()
    part = db.scalar(select(DimPart).where(
        DimPart.pn_std == normalized,
        DimPart.status != "merged",
    ))
    if part is not None:
        return part
    return db.scalar(
        select(DimPart)
        .join(PartAlias, PartAlias.part_id == DimPart.id)
        .where(
            PartAlias.pn_raw == pn.strip(),
            PartAlias.status == "active",
            DimPart.status != "merged",
        )
    )


def _merge_manual_cost_to_line(
    db: Session, refill: "CostRefill", *, operated_by: str,
) -> bool:
    """人工成本合并到主表——写 override 表（审计/回滚）并同步 f_maintenance_line：
    unit_cost_ex_tax/inc_tax + cost_source='manual' + cost_amount 字段。
    面板/看板/概览读主表即得人工值，无需各处 merge。

    仅当主表当前无自动成本（cost_source IN (NULL,'none')）时才合并到主表——已有
    direct/window 等自动证据的行不被人工值覆盖；'manual' 行允许修正人工价。
    override 表始终写入审计。
    """
    if refill.line_id is None:
        raise WorkbookError("line_not_found", "备件行标识无效，请重新下载")
    # Every caller uses the same row hierarchy.  Locking the parent line first
    # also serializes the otherwise-unlockable "no override row yet" state, so
    # concurrent first fills cannot both INSERT the unique line_id.
    line = db.scalar(
        select(FMaintenanceLine)
        .where(FMaintenanceLine.id == refill.line_id)
        .with_for_update()
    )
    if line is None:
        raise WorkbookError("line_not_found", "备件行已不存在，请重新下载")
    existing = db.scalar(
        select(MaintenanceManualCostOverride)
        .where(MaintenanceManualCostOverride.line_id == refill.line_id)
        .with_for_update()
    )
    changed = False
    if refill.unit_cost_ex_tax is not None:
        if existing is None:
            db.add(MaintenanceManualCostOverride(
                line_id=refill.line_id,
                unit_cost_ex_tax=refill.unit_cost_ex_tax,
                unit_cost_inc_tax=refill.unit_cost_inc_tax,
                reason=refill.reason, active=True, version=1,
                updated_by=operated_by))
            changed = True
        else:
            override_changed = (
                existing.unit_cost_ex_tax != refill.unit_cost_ex_tax
                or existing.unit_cost_inc_tax != refill.unit_cost_inc_tax
                or existing.reason != refill.reason
                or not existing.active
            )
            if override_changed:
                existing.unit_cost_ex_tax = refill.unit_cost_ex_tax
                existing.unit_cost_inc_tax = refill.unit_cost_inc_tax
                existing.reason = refill.reason
                existing.active = True
                existing.version += 1
                existing.updated_by = operated_by
                changed = True
        if line.cost_source in (None, "none", "manual"):
            line_changed = (
                line.unit_cost_ex_tax != refill.unit_cost_ex_tax
                or line.unit_cost_inc_tax != refill.unit_cost_inc_tax
                or line.cost_source != "manual"
            )
            _recompute_line_amounts(line, unit_cost_ex_tax=refill.unit_cost_ex_tax,
                                    unit_cost_inc_tax=refill.unit_cost_inc_tax)
            line.cost_source = "manual"
            changed = line_changed or changed
    elif refill.reason is not None:
        if existing is not None and existing.reason != refill.reason:
            existing.reason = refill.reason
            existing.version += 1
            existing.updated_by = operated_by
            changed = True
    return changed


def _recompute_line_amounts(line: FMaintenanceLine, *,
                            unit_cost_ex_tax: Decimal | None = None,
                            unit_cost_inc_tax: Decimal | None = None) -> None:
    """数量/退货数量/单价变更后重算两套成本金额（有效数量×单价）。
    无成本的行（none/NULL）保持 NULL——不知道≠0（铁律 5）。"""
    if unit_cost_ex_tax is not None:
        line.unit_cost_ex_tax = unit_cost_ex_tax
    if unit_cost_inc_tax is not None:
        line.unit_cost_inc_tax = unit_cost_inc_tax
    effective_qty = max((line.qty or Decimal(0)) - (line.return_qty or Decimal(0)),
                        Decimal(0))
    if line.unit_cost_ex_tax is not None:
        line.cost_amount_ex_tax = (line.unit_cost_ex_tax * effective_qty).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP)
    if line.unit_cost_inc_tax is not None:
        line.cost_amount_inc_tax = (line.unit_cost_inc_tax * effective_qty).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP)


def _write_audit(db: Session, *, project_id: str, entity_type: str,
                 entity_id: str, action: str, operated_by: str,
                 reason: str, before: dict | None = None,
                 after: dict | None = None) -> None:
    """行级审计（行级即可，不做字段级全量 diff）。"""
    db.add(MaintenanceProjectOperationAudit(
        project_id=project_id, entity_type=entity_type, entity_id=str(entity_id),
        action=action, before_json=before, after_json=after,
        reason=reason or "工作簿编辑", operated_by=operated_by,
    ))


def _cascade_void_site_lines(
    db: Session,
    line: FMaintenanceLine,
    *,
    project_id: str,
) -> int:
    """03 行作废时，按 source_line_id 文本匹配级联作废 06 领用返还行
    （06↔03 无 FK，best-effort，匹配不到不报错）。返回级联条数。"""
    if not line.raw_line_id:
        return 0
    site_lines = db.scalars(
        select(MaintenanceSiteIssueLine)
        .join(MaintenanceSiteIssue,
              MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id)
        .where(MaintenanceSiteIssue.project_id == project_id,
               MaintenanceSiteIssueLine.source_line_id == line.raw_line_id,
               MaintenanceSiteIssueLine.is_active.is_(True))
    ).all()
    for sl in site_lines:
        sl.is_active = False
    return len(site_lines)


def _v2_meta(wb) -> dict[str, str]:
    if V2_SHEET_META not in wb.sheetnames:
        raise WorkbookError("template_version_mismatch", "工作簿缺少 V2 元数据，请重新下载当前项目总表")
    raw: dict[str, str] = {}
    for row in wb[V2_SHEET_META].iter_rows(min_col=1, max_col=2):
        if not row[0].value:
            continue
        key = str(row[0].value).strip()
        raw[key] = str(row[1].value or "").strip()
    # 合并 2.7.1 分块（key@2/key@3... 按数字序拼接——字典序会让 @10 排在 @2 前）；
    # 签名在发射端按逻辑值计算，因此必须在任何使用/验签之前完成合并。
    merged: dict[str, str] = {}
    chunks: dict[str, list[tuple[int, str]]] = {}
    for key, value in raw.items():
        if (_META_CHUNK_SEP in key
                and key.rsplit(_META_CHUNK_SEP, 1)[-1].isdigit()):
            base, suffix = key.rsplit(_META_CHUNK_SEP, 1)
            chunks.setdefault(base, []).append((int(suffix), value))
        else:
            merged[key] = value
    for base, items in chunks.items():
        if base not in merged:
            merged[base] = ""
        merged[base] += "".join(value for _n, value in sorted(items))
    return merged


def _v2_verify_meta(db: Session, wb, project_id: str) -> dict[str, str]:
    meta = _v2_meta(wb)
    if meta.get("protocol_id") != V2_PROTOCOL_ID or meta.get("template_version") != V2_TEMPLATE_VERSION:
        raise WorkbookError("template_version_mismatch", "工作簿版本已更新，请重新下载当前项目总表后再上传。")
    key_id = meta.get("metadata_hmac_key_id", "")
    verification_key = get_settings().maintenance_manifest_verification_keys().get(key_id)
    supplied_signature = meta.get("metadata_hmac", "")
    if (
        meta.get("metadata_hmac_algorithm") != V2_META_SIGNATURE_ALGORITHM
        or verification_key is None
        or not supplied_signature
        or not hmac.compare_digest(
            supplied_signature,
            _v2_metadata_signature(meta, verification_key),
        )
    ):
        raise WorkbookError(
            "template_signature_invalid",
            "工作簿元数据签名无效或已过期，请重新下载当前项目总表后再上传。",
        )
    if meta.get("project_id") != project_id:
        # 2026-08-23：带上工作簿真正所属的项目名——批量处理多项目时极易
        # 拿错文件，只说「不一致」用户不知道这表是谁的
        from app.models.maintenance_project import MaintenanceProject

        other = db.get(MaintenanceProject, meta.get("project_id"))
        other_name = other.display_name if other is not None else "未知项目"
        raise WorkbookError(
            "project_mismatch",
            f"这份工作簿属于项目「{other_name}」，请到该项目的面板上传；"
            f"或重新从当前项目下载后再改。")
    included = tuple(
        name.strip() for name in meta.get("included_sheets", "").split(",")
        if name.strip()
    )
    if not included:
        raise WorkbookError("template_version_mismatch", "V2 工作簿缺少 included_sheets 元数据")
    unknown = [name for name in included if name not in V2_ALL_SHEETS[:-1]]
    if unknown:
        raise WorkbookError("template_version_mismatch", "V2 工作簿 included_sheets 无效")
    missing = [name for name in included if name not in wb.sheetnames]
    if missing:
        raise WorkbookError("missing_sheet", f"V2 工作簿缺少工作表：{'、'.join(missing)}")
    try:
        workbook_revision = int(meta.get("workbook_revision", ""))
    except ValueError as exc:
        raise WorkbookError(
            "template_version_mismatch",
            "工作簿缺少有效的项目并发版本，请重新下载当前项目总表。",
        ) from exc
    workbook_data_version = meta.get("workbook_data_version", "")
    if (
        workbook_revision < 0
        or re.fullmatch(r"[0-9a-f]{64}", workbook_data_version) is None
    ):
        raise WorkbookError(
            "template_version_mismatch",
            "工作簿项目并发版本无效，请重新下载当前项目总表。",
        )
    return meta


def _v2_refill_for_existing_line(
    db: Session,
    *,
    row,
    index: dict[str, int],
    row_no: int,
    line: FMaintenanceLine,
) -> CostRefill | None:
    """把一行人工回传与既有事实做 diff；原样行返回 ``None``。

    带实体 ID 的正常回传和「人工认证后、尚未出现在本项目导出中的全局 WBDD
    行」共用这一套字段语义，避免认领归属时顺手制造一份重复明细。
    """

    pn = None
    part_id = None
    pn_cell = str(_cell(row, index, "PN") or "").strip()
    if pn_cell and pn_cell != (line.pn_std or line.pn_raw or ""):
        part = _exact_part_for_pn(db, pn_cell)
        if part is None:
            raise WorkbookError("part_not_found",
                                f"第 {row_no} 行 PN {pn_cell!r} 未匹配备件主数据")
        pn = part.pn_std
        part_id = part.id
    desc_cell = str(_cell(row, index, "描述") or "").strip()
    description = desc_cell if desc_cell != (line.description or "") else None
    qty_cell = _cell(row, index, "需求数量")
    qty_parsed = (
        _v2_decimal(qty_cell, row_no=row_no, label="需求数量")
        if qty_cell not in (None, "") else None
    )
    if qty_parsed is not None and qty_parsed <= 0:
        raise WorkbookError("invalid_amount", f"第 {row_no} 行需求数量必须大于 0")
    qty = (
        qty_parsed
        if qty_parsed is not None and qty_parsed != (line.qty or Decimal(0))
        else None
    )
    raw_return = _cell(row, index, "退货数量")
    return_parsed = (
        _v2_decimal(raw_return, row_no=row_no, label="退货数量")
        if raw_return not in (None, "") else None
    )
    return_qty = (
        return_parsed
        if return_parsed is not None
        and return_parsed != (line.return_qty or Decimal(0))
        else None
    )
    effective_qty = qty if qty is not None else (line.qty or Decimal(0))
    effective_return = (
        return_qty if return_qty is not None else (line.return_qty or Decimal(0))
    )
    if effective_return > effective_qty:
        raise WorkbookError("invalid_amount",
                            f"第 {row_no} 行退货数量不能大于需求数量")
    sn_cell = str(_cell(row, index, "SN") or "").strip()
    serial_numbers = sn_cell if sn_cell != (line.serial_numbers or "") else None
    note_cell = str(_cell(row, index, "备注") or "").strip()
    note = note_cell if note_cell != (line.line_note or "") else None

    override = db.scalar(
        select(MaintenanceManualCostOverride).where(
            MaintenanceManualCostOverride.line_id == line.id,
            MaintenanceManualCostOverride.active.is_(True),
        ))
    raw_manual = _cell(row, index, "人工未税单位成本")
    amount = inc = None
    if raw_manual not in (None, ""):
        parsed_amount = _v2_decimal(raw_manual, row_no=row_no,
                                    label="人工未税单位成本", required=True)
        if override is None or parsed_amount != (override.unit_cost_ex_tax or Decimal(0)):
            amount = parsed_amount
            inc = (amount * (Decimal("1") + TAX_RATE)).quantize(
                Decimal("0.01"), rounding=ROUND_HALF_UP)
    reason_cell = str(_cell(row, index, "人工成本原因") or "").strip() or None
    reason = reason_cell if reason_cell != (override.reason if override else None) else None

    if (pn is None and description is None and qty is None
            and return_qty is None and serial_numbers is None and note is None
            and amount is None and reason is None):
        return None
    return CostRefill(
        line_id=line.id, operation="UPDATE",
        pn=pn, part_id=part_id,
        description=description, qty=qty, return_qty=return_qty,
        serial_numbers=serial_numbers,
        unit_cost_ex_tax=amount, unit_cost_inc_tax=inc,
        reason=reason, note=note,
    )


def _v2_parse_parts(
    db: Session,
    project_id: str,
    ws,
    merge: _V2MergeContext,
) -> tuple[
    list[CostRefill],
    int,
    set[int],
    tuple[SourceOrderAssignmentChange, ...],
]:
    headers = [str(cell.value or "") for cell in ws[1]]
    index = {name: i for i, name in enumerate(headers)}
    required = {"操作", "维保单号", "PN", "需求数量", "人工未税单位成本",
                "人工成本原因", "实体ID", "只读哈希", "备注", V2_BASE_COLUMN}
    if not required.issubset(index):
        raise WorkbookError("template_version_mismatch",
                            "03_备件明细列定义不是当前 V2.7 版本，请重新下载当前项目总表")
    out: list[CostRefill] = []
    uploaded_entity_rows = 0
    uploaded_entity_ids: set[int] = set()
    claimed_existing_line_ids: set[int] = set()
    assignment_changes: dict[str, SourceOrderAssignmentChange] = {}
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(value in (None, "") for value in row):
            continue
        if _is_example_row(row):
            continue
        raw_id = _cell(row, index, "实体ID")
        operation = str(_cell(row, index, "操作") or "").strip().upper()
        has_entity = raw_id not in (None, "")
        if has_entity:
            uploaded_entity_rows += 1
            uploaded_entity_ids.add(int(raw_id))

        if not has_entity:
            # 空实体行有两种明确语义：
            # - 操作留空：优先唯一识别全局 WBDD 的既有行（人工认证/幂等回传）；
            #   无匹配才新增。
            # - 显式 CREATE：强制新增。
            # 一旦该 WBDD 当前挂在别的项目，本项目内人工回传就是改挂确认；apply
            # 阶段仍经来源单归属服务做权限、乐观锁、历史和审计。
            order_no = str(_cell(row, index, "维保单号") or "").strip()
            pn_raw = str(_cell(row, index, "PN") or "").strip()
            qty_raw = _cell(row, index, "需求数量")
            if not order_no and not pn_raw and qty_raw in (None, ""):
                continue
            if operation and operation != "CREATE":
                raise WorkbookError("invalid_operation",
                                    f"第 {row_no} 行无实体ID，只能填 CREATE 或留空")
            if not order_no:
                raise WorkbookError("missing_field", f"第 {row_no} 行新增明细必须填写维保单号")
            from app.services.query_filters import active_beta_maintenance_orders

            order_stmt = (
                select(FMaintenanceOrder)
                .where(FMaintenanceOrder.order_no == order_no)
                .order_by(FMaintenanceOrder.raw_order_id)
            )
            orders = list(db.scalars(
                active_beta_maintenance_orders(order_stmt, FMaintenanceOrder)
            ).all())
            if not orders:
                raise WorkbookError(
                    "order_not_found",
                    f"第 {row_no} 行维保单号 {order_no!r} 尚未导入统一 WBDD，"
                    "请先导入原始 WBDD，再回传项目总表")
            if len(orders) > 1:
                raise WorkbookError(
                    "order_no_ambiguous",
                    f"第 {row_no} 行维保单号 {order_no!r} 对应多张活动来源单，"
                    "无法安全确认归属，请管理员核查来源 ID")
            order = orders[0]
            if not pn_raw:
                raise WorkbookError("missing_field", f"第 {row_no} 行新增明细必须填写 PN")
            part = _resolve_part_flexible(
                db, pn_raw, row_no=row_no, sheet="03_备件明细", merge=merge)
            if part is None:
                continue
            qty = _v2_decimal(qty_raw, row_no=row_no, label="需求数量", required=True)
            if qty is None or qty <= 0:
                raise WorkbookError("invalid_amount",
                                    f"第 {row_no} 行需求数量必须大于 0")
            return_qty = _v2_decimal(_cell(row, index, "退货数量"),
                                     row_no=row_no, label="退货数量") or Decimal(0)
            if return_qty > qty:
                raise WorkbookError("invalid_amount",
                                    f"第 {row_no} 行退货数量不能大于需求数量")
            current_assignment = db.scalar(
                select(MaintenanceSourceOrderAssignment).where(
                    MaintenanceSourceOrderAssignment.source_order_id == order.raw_order_id,
                    MaintenanceSourceOrderAssignment.is_active.is_(True),
                )
            )
            if current_assignment is None or current_assignment.project_id != project_id:
                previous_project = (
                    db.get(MaintenanceProject, current_assignment.project_id)
                    if current_assignment is not None else None
                )
                change = SourceOrderAssignmentChange(
                    source_order_id=order.raw_order_id,
                    order_no=order_no,
                    expected_assignment_id=(
                        current_assignment.assignment_id
                        if current_assignment is not None else None
                    ),
                    expected_version=(
                        current_assignment.version
                        if current_assignment is not None else None
                    ),
                    previous_project_id=(
                        current_assignment.project_id
                        if current_assignment is not None else None
                    ),
                    previous_project_name=(
                        previous_project.display_name if previous_project is not None else None
                    ),
                )
                prior = assignment_changes.get(order.raw_order_id)
                if prior is not None and prior != change:
                    raise WorkbookError(
                        "stale_assignment",
                        f"第 {row_no} 行维保单号 {order_no!r} 的项目归属已变化，"
                        "请重新预检")
                assignment_changes[order.raw_order_id] = change

            # 操作留空时先认领已存在的全局明细。PN+数量+退货数量是基础身份；
            # SN/描述只在出现多个候选时用于收窄，仍多义则整本拒绝。
            existing_line: FMaintenanceLine | None = None
            if not operation:
                candidates = [
                    candidate
                    for candidate in db.scalars(
                        select(FMaintenanceLine)
                        .where(
                            FMaintenanceLine.order_id == order.id,
                            FMaintenanceLine.part_id == part.id,
                            FMaintenanceLine.is_active.is_(True),
                        )
                        .order_by(FMaintenanceLine.id)
                    ).all()
                    if candidate.id not in claimed_existing_line_ids
                    and (candidate.qty or Decimal(0)) == qty
                    and (candidate.return_qty or Decimal(0)) == return_qty
                ]
                sn_value = str(_cell(row, index, "SN") or "").strip()
                if len(candidates) > 1 and sn_value:
                    candidates = [
                        candidate for candidate in candidates
                        if (candidate.serial_numbers or "").strip() == sn_value
                    ]
                description_value = str(_cell(row, index, "描述") or "").strip()
                if len(candidates) > 1 and description_value:
                    narrowed = [
                        candidate for candidate in candidates
                        if (candidate.description or "").strip() == description_value
                    ]
                    if narrowed:
                        candidates = narrowed
                if len(candidates) > 1:
                    raise WorkbookError(
                        "manual_line_ambiguous",
                        f"第 {row_no} 行 {order_no}/{part.pn_std} 在统一 WBDD 中有"
                        "多条相同数量明细，无法唯一认领；请保留实体ID或明确填 CREATE")
                if candidates:
                    existing_line = candidates[0]

            if existing_line is not None:
                claimed_existing_line_ids.add(existing_line.id)
                uploaded_entity_rows += 1
                uploaded_entity_ids.add(existing_line.id)
                refill = _v2_refill_for_existing_line(
                    db, row=row, index=index, row_no=row_no, line=existing_line)
                if refill is not None:
                    out.append(refill)
                continue

            raw_manual = _cell(row, index, "人工未税单位成本")
            amount = inc = None
            if raw_manual not in (None, ""):
                amount = _v2_decimal(raw_manual, row_no=row_no,
                                     label="人工未税单位成本", required=True)
                inc = (amount * (Decimal("1") + TAX_RATE)).quantize(
                    Decimal("0.01"), rounding=ROUND_HALF_UP)
            out.append(CostRefill(
                line_id=None, operation="CREATE", is_create=True,
                order_no=order_no, order_id=order.id, part_id=part.id, pn=part.pn_std,
                description=str(_cell(row, index, "描述") or "").strip() or None,
                qty=qty, return_qty=return_qty,
                serial_numbers=str(_cell(row, index, "SN") or "").strip() or None,
                unit_cost_ex_tax=amount, unit_cost_inc_tax=inc,
                reason=(str(_cell(row, index, "人工成本原因") or "").strip() or None),
                note=(str(_cell(row, index, "备注") or "").strip() or None),
            ))
            continue

        line = db.get(FMaintenanceLine, int(raw_id))
        if line is None or not line.is_active:
            raise WorkbookError("line_not_found",
                                f"第 {row_no} 行备件事实已不存在或已作废，请重新下载")
        # 行必须归属本项目（P1，Codex review #272）：防止拿别项目工作簿的
        # 实体ID（或构造行）越权 UPDATE/VOID 本项目之外的事实。
        owner_project = db.scalar(
            select(MaintenanceSourceOrderAssignment.project_id)
            .join(FMaintenanceOrder,
                  FMaintenanceOrder.id == line.order_id)
            .where(
                MaintenanceSourceOrderAssignment.source_order_id
                == FMaintenanceOrder.raw_order_id,
                MaintenanceSourceOrderAssignment.is_active.is_(True),
            )
            .limit(1)
        )
        if owner_project != project_id:
            raise WorkbookError("line_not_in_project",
                                f"第 {row_no} 行备件事实不属于本项目，请重新下载")
        # 只读事实哈希校验（只覆盖锁定列；可编辑列改动不触发）
        readonly_values = [_cell(row, index, name) for name in V2_PART_HASH_COLUMNS]
        if _v2_hash(readonly_values) != str(_cell(row, index, "只读哈希") or ""):
            raise WorkbookError(
                "readonly_cell_modified",
                f"第 {row_no} 行只读事实（单号/日期/XSDD/系统成本等）已修改或已过期，"
                "请重新下载；如需改数量/PN/备注请改黄底列")
        # 行级基线三路合并（2.7.0）：未触碰行直接跳过（服务端已变也不回写），
        # 触碰行按字段判定冲突；VOID 行在服务端已变时登记整行冲突。
        baseline = _parse_v2_row_base_token(
            _cell(row, index, V2_BASE_COLUMN), sheet="03_备件明细",
            entity_id=line.id, row_no=row_no)
        baseline_order = db.get(FMaintenanceOrder, line.order_id)
        baseline_override = db.scalar(
            select(MaintenanceManualCostOverride).where(
                MaintenanceManualCostOverride.line_id == line.id,
                MaintenanceManualCostOverride.active.is_(True),
            ))
        server_values = _v2_part_row_values(
            line, baseline_order, baseline_override)
        row_label = str(server_values.get("维保单号") or line.id)
        if operation == "VOID":
            if _v2_server_row_changed(
                    server_values=server_values, baseline=baseline,
                    fields=V2_PART_BASE_FIELDS):
                if merge.record_row_conflict(
                        sheet="03_备件明细", row_label=row_label,
                        entity_id=line.id, action="VOID（整行作废）"):
                    out.append(CostRefill(line_id=line.id, operation="VOID",
                                          unit_cost_ex_tax=None, unit_cost_inc_tax=None,
                                          reason=None))
                continue
            out.append(CostRefill(line_id=line.id, operation="VOID",
                                  unit_cost_ex_tax=None, unit_cost_inc_tax=None,
                                  reason=None))
            continue
        if operation and operation != "UPDATE":
            raise WorkbookError("invalid_operation",
                                f"第 {row_no} 行操作只能是 UPDATE、VOID 或留空")
        if not _v2_merge_row(
                sheet="03_备件明细", row_label=row_label, entity_id=line.id,
                row=row, index=index, base_fields=V2_PART_BASE_FIELDS,
                server_values=server_values, baseline=baseline, ctx=merge):
            continue

        refill = _v2_refill_for_existing_line(
            db, row=row, index=index, row_no=row_no, line=line)
        if refill is not None:
            out.append(refill)
    return (
        out,
        uploaded_entity_rows,
        uploaded_entity_ids,
        tuple(assignment_changes.values()),
    )


def _v2_parse_site(
    db: Session, project_id: str, ws, merge: _V2MergeContext,
) -> tuple[list[SiteReturnFlag], set[str]]:
    headers = [str(cell.value or "") for cell in ws[1]]
    index = {name: i for i, name in enumerate(headers)}
    if V2_BASE_COLUMN not in index:
        raise WorkbookError("template_version_mismatch",
                            "06_领用返还列定义不是当前 V2.7 版本，请重新下载当前项目总表")
    out: list[SiteReturnFlag] = []
    present_ids: set[str] = set()
    new_counts: dict[str, int] = {}
    # One issue header is repeated on every line in Excel.  Track changes
    # relative to the stored header so changing a single row propagates once;
    # two different new values fail closed instead of "last row wins".
    header_changes: dict[str, dict[str, object]] = {}
    new_issue_dates: dict[str, date] = {}
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(value in (None, "") for value in row):
            continue
        if _is_example_row(row):
            continue
        raw_id = _cell(row, index, "实体ID")
        if raw_id not in (None, ""):
            present_ids.add(str(raw_id))
        issue_no = str(_cell(row, index, "领用单号") or "").strip() or None
        issue_date = (_v2_date(_cell(row, index, "领用日期"), row_no=row_no, label="领用")
                      if _cell(row, index, "领用日期") not in (None, "") else None)
        pn = str(_cell(row, index, "PN") or "").strip() or None
        serial_number = str(_cell(row, index, "SN") or "").strip() or None
        quantity = (_v2_decimal(_cell(row, index, "领用数量"), row_no=row_no, label="领用数量")
                    if _cell(row, index, "领用数量") not in (None, "") else None)
        flag = str(_cell(row, index, "是否应返还") or "").strip()
        if flag and flag not in {"是", "否"}:
            raise WorkbookError("invalid_flag", f"第 {row_no} 行是否应返还只能填 是 / 否")
        is_create = raw_id in (None, "")
        issue_id = None
        line_no = None
        part_id = None
        if is_create:
            if not issue_no or issue_date is None or not pn or quantity is None:
                raise WorkbookError(
                    "missing_site_identity",
                    f"06_领用返还第 {row_no} 行手工新增必须填写领用单号、日期、PN 和数量",
                )
            if quantity <= 0:
                raise WorkbookError("invalid_amount", f"第 {row_no} 行领用数量必须大于 0")
            part = _resolve_part_flexible(
                db, pn, row_no=row_no, sheet="06_领用返还", merge=merge)
            if part is None:
                continue
            part_id = part.id
            pn = part.pn_std
            document_date = re.search(r"(?:^|-)20\d{6}(?:-|$)", issue_no)
            if document_date:
                raw_document_date = document_date.group(0).strip("-")
                issue_date = datetime.strptime(raw_document_date, "%Y%m%d").date()
            identity = "|".join([project_id, issue_no, pn, serial_number or ""])
            raw_id = f"manual-site:{hashlib.sha1(identity.encode('utf-8')).hexdigest()}"
            existing_line = db.get(MaintenanceSiteIssueLine, raw_id)
            if existing_line is not None:
                is_create = False
                issue_id = existing_line.issue_id
                line_no = existing_line.line_no
                part_id = existing_line.part_id
            else:
                base = db.scalar(
                    select(func.max(MaintenanceSiteIssueLine.line_no))
                    .join(MaintenanceSiteIssue,
                          MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id)
                    .where(
                        MaintenanceSiteIssue.project_id == project_id,
                        MaintenanceSiteIssue.issue_no == issue_no,
                    )
                ) or 0
                new_counts[issue_no] = new_counts.get(issue_no, 0) + 1
                line_no = int(base) + new_counts[issue_no]
                previous_date = new_issue_dates.setdefault(issue_no, issue_date)
                if previous_date != issue_date:
                    raise WorkbookError(
                        "conflicting_issue_header",
                        f"06_领用返还中同一新增领用单 {issue_no!r} 填了不同日期",
                    )
        else:
            raw_id = str(raw_id)
            existing_line = db.get(MaintenanceSiteIssueLine, raw_id)
            if existing_line is None:
                raise WorkbookError("line_not_found", f"第 {row_no} 行领用事实已不存在，请重新下载")
            issue = db.get(MaintenanceSiteIssue, existing_line.issue_id)
            if issue is None or issue.project_id != project_id:
                raise WorkbookError("project_mismatch", f"第 {row_no} 行领用事实不属于本项目")
            issue_id = existing_line.issue_id
            line_no = existing_line.line_no
            baseline = _parse_v2_row_base_token(
                _cell(row, index, V2_BASE_COLUMN), sheet="06_领用返还",
                entity_id=raw_id, row_no=row_no)
            server_values = _v2_site_row_values(existing_line, issue)
            if not _v2_merge_row(
                    sheet="06_领用返还", row_label=(
                        str(server_values.get("领用单号") or raw_id)),
                    entity_id=raw_id,
                    row=row, index=index, base_fields=V2_SITE_BASE_FIELDS,
                    server_values=server_values, baseline=baseline, ctx=merge):
                continue
            if pn is not None and pn != existing_line.pn:
                part = _exact_part_for_pn(db, pn)
                if part is None:
                    raise WorkbookError(
                        "part_not_found",
                        f"06_领用返还第 {row_no} 行 PN {pn!r} 未匹配备件主数据",
                    )
                part_id = part.id
                pn = part.pn_std
            else:
                part_id = existing_line.part_id
                pn = existing_line.pn
        if issue_id is not None:
            existing_issue = db.get(MaintenanceSiteIssue, issue_id)
            if existing_issue is None or existing_issue.project_id != project_id:
                raise WorkbookError(
                    "project_mismatch", f"第 {row_no} 行领用事实不属于本项目"
                )
            requested_no = issue_no or existing_issue.issue_no
            requested_date = issue_date or existing_issue.issue_date
            header = header_changes.setdefault(
                issue_id,
                {
                    "original_no": existing_issue.issue_no,
                    "original_date": existing_issue.issue_date,
                    "new_nos": set(),
                    "new_dates": set(),
                },
            )
            if requested_no != header["original_no"]:
                header["new_nos"].add(requested_no)
            if requested_date != header["original_date"]:
                header["new_dates"].add(requested_date)
        out.append(SiteReturnFlag(
            issue_line_id=str(raw_id), is_create=is_create,
            issue_id=issue_id, line_no=line_no, part_id=part_id,
            no_return=(flag == "否") if flag else None,
            issue_no=issue_no,
            issue_date=issue_date,
            pn=pn,
            serial_number=serial_number,
            quantity=quantity,
            remark=str(_cell(row, index, "备注") or "").strip() or None,
        ))
    present_ids |= {f.issue_line_id for f in out if not f.is_create}
    chosen_headers: dict[str, tuple[str, date]] = {}
    for issue_id, header in header_changes.items():
        new_nos = header["new_nos"]
        new_dates = header["new_dates"]
        if len(new_nos) > 1 or len(new_dates) > 1:
            raise WorkbookError(
                "conflicting_issue_header",
                "06_领用返还同一领用单的多行填写了不同单号或日期，请统一后重试",
            )
        chosen_no = next(iter(new_nos), header["original_no"])
        chosen_date = next(iter(new_dates), header["original_date"])
        duplicate_issue = db.scalar(
            select(MaintenanceSiteIssue.issue_id).where(
                MaintenanceSiteIssue.project_id == project_id,
                MaintenanceSiteIssue.issue_no == chosen_no,
                MaintenanceSiteIssue.issue_id != issue_id,
            )
        )
        if duplicate_issue is not None:
            raise WorkbookError(
                "duplicate_issue_no",
                f"领用单号 {chosen_no!r} 已被本项目另一张领用单使用",
            )
        chosen_headers[issue_id] = (chosen_no, chosen_date)
    merged_out = [
        replace(
            flag,
            issue_no=chosen_headers[flag.issue_id][0],
            issue_date=chosen_headers[flag.issue_id][1],
        )
        if flag.issue_id in chosen_headers
        else flag
        for flag in out
    ]
    return merged_out, present_ids


def _v2_parse_plan(db: Session, project_id: str, ws) -> list[V2MilestoneChange]:
    headers = [str(cell.value or "") for cell in ws[1]]
    index = {name: i for i, name in enumerate(headers)}
    out: list[V2MilestoneChange] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(value in (None, "") for value in row):
            continue
        if _is_example_row(row):
            continue
        operation = str(row[index["操作"]] or "").strip().upper()
        raw_entity = str(row[index["实体ID"]] or "").strip() or None
        if not operation and not raw_entity:
            continue
        if operation and operation not in {"CREATE", "UPDATE", "VOID"}:
            raise WorkbookError("invalid_operation", f"第 {row_no} 行操作必须是 CREATE、UPDATE 或 VOID")
        milestone = None
        if raw_entity:
            milestone = db.get(MaintenanceCollectionMilestone, raw_entity)
            if milestone is None or not milestone.is_active or milestone.project_id != project_id:
                raise WorkbookError(
                    "milestone_not_found",
                    f"第 {row_no} 行回款计划已不存在或已作废，请重新下载",
                )
        # VOID is identity-bound: editable contract/sequence cells never choose
        # the target.  This also makes a forged A-id/B-coordinate row harmless.
        if operation == "VOID":
            if milestone is None:
                raise WorkbookError("invalid_operation", f"第 {row_no} 行无实体ID不能 VOID")
            original_contract = db.get(
                MaintenanceProjectContract, milestone.project_contract_id)
            if original_contract is None or original_contract.project_id != project_id:
                raise WorkbookError(
                    "milestone_not_found",
                    f"第 {row_no} 行回款计划关联合同已不存在，请重新下载",
                )
            out.append(V2MilestoneChange(
                operation="VOID",
                contract_no=original_contract.contract_no,
                sequence=milestone.sequence,
                planned_date=milestone.planned_date,
                date_precision=milestone.date_precision,
                planned_amount=milestone.planned_amount,
                entity_id=raw_entity,
                base_version=(
                    int(row[index["基础版本"]])
                    if row[index["基础版本"]] not in (None, "") else None
                ),
            ))
            continue
        # 留空操作 + 已有实体行 = 与 03/04 同语义：改了才更新（diff 幂等），
        # 原样回传零变更（2026-08-20 用户三连问：改/删都要同步）。
        contract_no = str(row[index["合同编号"]] or "").strip()
        contract = _resolve_plan_contract(db, project_id, contract_no, row_no)
        sequence = int(row[index["期次"]])
        if not 1 <= sequence <= 24:
            raise WorkbookError("invalid_sequence", f"第 {row_no} 行期次必须为 1-24")
        precision = str(row[index["日期精度"]] or "day").strip()
        if precision not in {"day", "month"}:
            raise WorkbookError("invalid_date_precision", f"第 {row_no} 行日期精度只能是 day/month")
        amount = _v2_decimal(row[index["计划回款金额（含税）"]], row_no=row_no, label="计划回款金额", required=operation == "CREATE")
        planned_date = _v2_date(row[index["计划回款日期"]], row_no=row_no, label="计划回款")
        if operation not in ("VOID",) and not raw_entity and planned_date is None and amount is None:
            raise WorkbookError("incomplete_milestone", f"第 {row_no} 行计划日期和金额不能同时为空")
        note = str(row[index["备注"]] or "").strip() if "备注" in index else None
        if raw_entity:
            if (
                contract is None
                or contract.project_contract_id != milestone.project_contract_id
                or sequence != milestone.sequence
            ):
                raise WorkbookError(
                    "milestone_identity_mismatch",
                    f"第 {row_no} 行合同编号或期次不能修改；请新增目标节点后再作废原节点",
                )
            # 留空或 UPDATE：空单元格=保留现值；与现值 diff，未变化不下发
            # （原样回传零变更；writer 对 None 会当清空，故传生效值）。
            eff_date = planned_date if planned_date is not None else milestone.planned_date
            eff_amount = amount if amount is not None else milestone.planned_amount
            changed = (
                eff_date != milestone.planned_date
                or eff_amount != milestone.planned_amount
                or precision != milestone.date_precision
                or (note is not None and note != (milestone.follow_up_note or ""))
            )
            if not changed:
                continue
            out.append(V2MilestoneChange(
                operation="UPDATE", contract_no=contract_no, sequence=sequence,
                planned_date=eff_date, date_precision=precision,
                planned_amount=eff_amount,
                entity_id=raw_entity,
                base_version=int(row[index["基础版本"]]) if row[index["基础版本"]] not in (None, "") else None,
            ))
            continue
        out.append(V2MilestoneChange(
            operation="CREATE", contract_no=contract_no, sequence=sequence,
            planned_date=planned_date, date_precision=precision, planned_amount=amount,
            entity_id=None,
            base_version=int(row[index["基础版本"]]) if row[index["基础版本"]] not in (None, "") else None,
        ))
    return out


def _v2_parse_expenses(
    db: Session, project_id: str, ws, merge: _V2MergeContext,
) -> tuple[list[ec.ExpenseUpdate], list[str], set[str]]:
    """04 解析（V2.7）：操作列（空/CREATE/UPDATE/VOID）+ 空白实体ID 手工新增
    （27c95fa 既有语义，实体ID 空 + 费用单号/明细序号必填 → CREATE）。

    返回 (updates, void_raw_line_ids, present_ids)；present_ids 是文件中出现
    实体ID 的全部行（含未触碰行），供缺行=作废的对账使用。
    """
    headers = [str(cell.value or "") for cell in ws[1]]
    index = {name: i for i, name in enumerate(headers)}
    required = {"操作", "实体ID", V2_BASE_COLUMN}
    if not required.issubset(index):
        raise WorkbookError("template_version_mismatch",
                            "04_费用报销列定义不是当前 V2.7 版本，请重新下载当前项目总表")
    # Expense mutation requires a current, unshared contract identity.  Stable
    # row ownership itself is checked against expense attribution below.
    contract_nos = _writable_contract_nos(db, project_id)
    expected_expense_ids = set(_expected_expense_ids(db, project_id))
    out: list[ec.ExpenseUpdate] = []
    voids: list[str] = []
    present_ids: set[str] = set()
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(value in (None, "") for value in row):
            continue
        if _is_example_row(row):
            continue
        operation = str(_cell(row, index, "操作") or "").strip().upper()
        if operation and operation not in {"VOID", "UPDATE", "CREATE"}:
            raise WorkbookError("invalid_operation",
                                f"第 {row_no} 行操作只能是 VOID、UPDATE、CREATE 或留空")
        raw_id = str(_cell(row, index, "实体ID") or "").strip()
        if raw_id:
            present_ids.add(raw_id)
        bxd_no = str(_cell(row, index, "费用单号") or "").strip() or None
        raw_line_no = _cell(row, index, "明细序号")
        line_no = None
        if raw_line_no not in (None, ""):
            try:
                line_no = int(raw_line_no)
            except (TypeError, ValueError):
                raise WorkbookError("invalid_line_no", f"第 {row_no} 行明细序号必须是整数")
            if line_no < 1:
                raise WorkbookError("invalid_line_no", f"第 {row_no} 行明细序号必须大于 0")
        contract_no = str(_cell(row, index, "维保销售订单（归集键）") or "").strip() or None
        if contract_no not in contract_nos:
            raise WorkbookError(
                "contract_not_found",
                f"第 {row_no} 行维保销售订单 {contract_no or ''!r} 不属于本项目",
            )
        amount = _v2_decimal(
            _cell(row, index, "未税金额"),
            row_no=row_no,
            label="未税金额",
            allow_negative=True,
        )
        inc = (amount * (Decimal("1") + TAX_RATE)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if amount is not None else None
        exported_basis = str(_cell(row, index, "金额口径") or "").strip()
        expense_date = (_v2_date(_cell(row, index, "报销日期"), row_no=row_no, label="报销")
                        if _cell(row, index, "报销日期") not in (None, "") else None)
        is_create = False
        expense = None
        if raw_id:
            if raw_id not in expected_expense_ids:
                # 跨项目/伪造实体先于基线解码拒绝（与 03/06 同序）。
                raise WorkbookError(
                    "expense_not_in_project",
                    f"第 {row_no} 行报销事实不属于本项目，请重新下载")
            expense = db.scalar(select(FProjectExpense).where(FProjectExpense.raw_line_id == raw_id))
            if expense is None:
                raise WorkbookError("expense_not_found", f"第 {row_no} 行报销事实已不存在，请重新下载")
            baseline = _parse_v2_row_base_token(
                _cell(row, index, V2_BASE_COLUMN), sheet="04_费用报销",
                entity_id=raw_id, row_no=row_no)
            server_values = _v2_expense_row_values(expense, "")
            row_label = str(server_values.get("费用单号") or raw_id)
            if operation == "VOID":
                if _v2_server_row_changed(
                        server_values=server_values, baseline=baseline,
                        fields=V2_EXPENSE_BASE_FIELDS):
                    if merge.record_row_conflict(
                            sheet="04_费用报销", row_label=row_label,
                            entity_id=raw_id, action="VOID（整行作废）"):
                        voids.append(raw_id)
                    continue
                voids.append(raw_id)
                continue
            if not _v2_merge_row(
                    sheet="04_费用报销", row_label=row_label, entity_id=raw_id,
                    row=row, index=index, base_fields=V2_EXPENSE_BASE_FIELDS,
                    server_values=server_values, baseline=baseline, ctx=merge):
                continue
        else:
            if not bxd_no or line_no is None:
                raise WorkbookError(
                    "missing_expense_identity",
                    f"第 {row_no} 行手工新增报销必须填写费用单号和明细序号",
                )
            if expense_date is None:
                raise WorkbookError("missing_expense_date", f"第 {row_no} 行手工新增报销必须填写报销日期")
            if amount is None:
                raise WorkbookError("missing_amount", f"第 {row_no} 行手工新增报销必须填写未税金额")
            scope = hashlib.sha1(contract_no.encode("utf-8")).hexdigest()[:8]
            raw_id = f"{bxd_no[:40]}#{line_no}@{scope}"
            expense = db.scalar(select(FProjectExpense).where(FProjectExpense.raw_line_id == raw_id))
            is_create = expense is None

        if not is_create and expense.amount_ex_tax is not None and amount == expense.amount_ex_tax:
            amount = inc = None
        if is_create:
            if exported_basis not in {"", "default_ex", "ex"}:
                raise WorkbookError(
                    "invalid_tax_basis",
                    f"第 {row_no} 行新增报销仅支持未税金额口径",
                )
            tax_basis = "ex"
        else:
            tax_basis = expense.tax_basis or "ex"
            if exported_basis and exported_basis != tax_basis:
                raise WorkbookError(
                    "readonly_tax_basis_changed",
                    f"第 {row_no} 行金额口径是来源只读字段，请重新下载后仅修改未税金额",
                )
        out.append(ec.ExpenseUpdate(
            raw_line_id=raw_id, is_create=is_create,
            bxd_no=bxd_no, line_no=line_no,
            expense_date=expense_date,
            person=str(_cell(row, index, "报销人员") or "").strip() or None,
            expense_type=str(_cell(row, index, "报销类别") or "").strip() or None,
            fee_category=str(_cell(row, index, "费用分类") or "").strip() or None,
            reason=str(_cell(row, index, "支出事由") or "").strip() or None,
            contract_no=contract_no,
            amount_ex_tax=amount,
            amount_inc_tax=inc,
            tax_basis=tax_basis,
            data_status=str(_cell(row, index, "流程状态") or "").strip() or None,
            remark=str(_cell(row, index, "备注") or "").strip() or None,
        ))
    return out, voids, present_ids


def _ensure_contract_for_xsdd_apply(
    db: Session,
    *,
    project: MaintenanceProject,
    contract_no: str,
    operated_by: str,
    assigned_xsdd_nos: set[str],
) -> tuple[MaintenanceProjectContract, bool]:
    """Create one current fallback relation inside the locked apply transaction."""

    contract_id = f"xsdd-{contract_no}"
    try:
        maintenance_project_identity.claim_xsdd_project(
            db,
            value=contract_no,
            project_id=project.project_id,
            source="project_master_workbook",
        )
    except maintenance_project_identity.XsddProjectConflict as exc:
        raise WorkbookError("contract_shared", str(exc)) from exc
    # Serialize this stable identity across projects; project row locks alone do
    # not protect two different projects from concurrently claiming one XSDD.
    db.execute(select(func.pg_advisory_xact_lock(
        func.hashtextextended(f"maintenance-contract:{contract_id}", 0)
    )))
    current = _current_contract_by_no(
        db, project.project_id, contract_no, lock=True)
    today = business_today()
    shared = db.scalar(
        select(MaintenanceProjectContract.project_contract_id)
        .where(
            (
                (MaintenanceProjectContract.contract_no == contract_no)
                | (
                    MaintenanceProjectContract.contract_id
                    == (current.contract_id if current is not None else contract_id)
                )
            ),
            MaintenanceProjectContract.project_id != project.project_id,
            MaintenanceProjectContract.effective_from <= today,
            (
                MaintenanceProjectContract.effective_to.is_(None)
                | (MaintenanceProjectContract.effective_to > today)
            ),
        )
        .limit(1)
        .with_for_update()
    )
    if shared is not None:
        raise WorkbookError(
            "contract_shared",
            f"合同 {contract_no} 已被其他项目当前使用，请先完成合同归属消歧",
        )
    if current is not None:
        return current, False
    # The caller took order -> active-assignment locks for every referenced
    # XSDD before any detail row or contract lock.  Use that locked snapshot;
    # do not reopen a TOCTOU window with an unlocked ownership query here.
    if contract_no not in assigned_xsdd_nos:
        raise WorkbookError(
            "contract_not_found",
            f"合同 {contract_no} 已不再由本项目当前 WBDD 挂靠，请重新下载",
        )
    if project.period_from is None:
        raise WorkbookError(
            "contract_effective_date_missing",
            "项目缺少业务开始日期，无法建立 XSDD 合同关系；请先补齐项目日期",
        )
    sales_candidates = list(db.execute(
        select(
            FSalesOrder.id,
            FSalesOrder.amount_ex_tax,
            FSalesOrder.tax_rate,
        )
        .join(SysImportBatch, SysImportBatch.id == FSalesOrder.import_batch_id)
        .where(
            FSalesOrder.order_no == contract_no,
            FSalesOrder.data_status == config.ACTIVE_STATUS,
            FSalesOrder.amount_ex_tax.is_not(None),
            SysImportBatch.file_type == "sales",
            SysImportBatch.status == "success",
        )
        .order_by(
            SysImportBatch.uploaded_at.desc().nullslast(),
            FSalesOrder.created_at.desc().nullslast(),
            FSalesOrder.import_batch_id.desc(),
            FSalesOrder.id.desc(),
        )
        .with_for_update(of=FSalesOrder)
    ).all())
    economic_candidates = {
        (
            Decimal(str(row.amount_ex_tax)),
            Decimal(str(row.tax_rate)) if row.tax_rate is not None else None,
        )
        for row in sales_candidates
    }
    if len(economic_candidates) > 1:
        raise WorkbookError(
            "sales_order_ambiguous",
            f"销售订单 {contract_no} 存在多条有效且金额/税率冲突的成功导入记录，"
            "无法自动建立合同关系，请先完成销售订单消歧。",
        )
    sale = sales_candidates[0] if sales_candidates else None
    amount_ex_tax = (
        Decimal(str(sale.amount_ex_tax)) if sale is not None else None
    )
    tax_rate = (
        Decimal(str(sale.tax_rate))
        if sale is not None and sale.tax_rate is not None else None
    )
    inc_tax = (
        amount_ex_tax * (Decimal("1") + tax_rate)
    ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) \
        if amount_ex_tax is not None and tax_rate is not None else None
    contract = MaintenanceProjectContract(
        project_contract_id=str(uuid4()),
        project_id=project.project_id,
        contract_id=contract_id,
        contract_no=contract_no,
        contract_amount=amount_ex_tax,
        amount_inc_tax=inc_tax,
        contract_status="正常",
        status_mapping_state="mapped",
        status_mapping_version="workbook-v2-xsdd",
        included_in_total=True,
        effective_from=max(project.period_from, today),
        source="sales_fallback",
        version=1,
    )
    db.add(contract)
    db.flush()
    _write_audit(
        db,
        project_id=project.project_id,
        entity_type="project_contract",
        entity_id=contract.project_contract_id,
        action="CREATE",
        operated_by=operated_by,
        reason="项目总表应用时按当前 XSDD 建立合同关系",
        after={"contract_no": contract_no, "source": "sales_fallback"},
    )
    return contract, True


def _xsdd_contract_for_project(
    db: Session, project_id: str, row_no: int
) -> tuple[str, MaintenanceProjectContract | None]:
    """合同编号留空 → 按项目唯一挂靠 XSDD 自动回填（多义/缺失整本拒绝）。"""
    xsdd_rows = db.scalars(
        select(FMaintenanceOrder.linked_sales_order_no)
        .join(MaintenanceSourceOrderAssignment,
              (MaintenanceSourceOrderAssignment.source_order_id
               == FMaintenanceOrder.raw_order_id)
              & MaintenanceSourceOrderAssignment.is_active.is_(True))
        .where(
            MaintenanceSourceOrderAssignment.project_id == project_id,
            FMaintenanceOrder.linked_sales_order_no.is_not(None),
        )
        .group_by(FMaintenanceOrder.linked_sales_order_no)
    ).all()
    xsdd_nos = sorted({str(o) for o in xsdd_rows if o})
    if not xsdd_nos:
        raise WorkbookError(
            "contract_not_found",
            f"第 {row_no} 行合同编号留空，但该项目未挂靠任何 XSDD 单据，无法自动回填")
    if len(xsdd_nos) > 1:
        raise WorkbookError(
            "contract_not_found",
            f"第 {row_no} 行合同编号留空，但该项目挂靠多个 XSDD（{'、'.join(xsdd_nos)}），请明确填写")
    contract_no = xsdd_nos[0]
    return contract_no, _current_contract_by_no(db, project_id, contract_no)


def _resolve_plan_contract(db: Session, project_id: str, contract_no: str, row_no: int):
    """Read-only validate: current ledger or deferred current WBDD XSDD."""
    current_contracts = _current_contracts(db, project_id)
    current = _current_contract_by_no(db, project_id, contract_no)
    if current is not None:
        return current
    xsdd = _assigned_xsdd_nos(db, project_id)
    if contract_no in xsdd:
        return None
    available = sorted({contract.contract_no for contract in current_contracts} | xsdd)
    raise WorkbookError(
        "contract_not_found",
        f"第 {row_no} 行合同编号 {contract_no or ''!r} 不属于本项目"
        + (f"（可用：{'、'.join(available[:8])}）" if available else "（本项目还没有合同，请先填挂靠单据上的 XSDD 号）"))


def _v2_parse_receipts(
    db: Session, project_id: str, ws, merge: _V2MergeContext,
) -> list[ec.CollectionOp]:
    headers = [str(cell.value or "") for cell in ws[1]]
    index = {name: i for i, name in enumerate(headers)}
    if V2_BASE_COLUMN not in index:
        raise WorkbookError("template_version_mismatch",
                            "05_实收回款列定义不是当前 V2.7 版本，请重新下载当前项目总表")
    out: list[ec.CollectionOp] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(value in (None, "") for value in row):
            continue
        if _is_example_row(row):
            continue
        contract_no = str(row[index["合同编号"]] or "").strip()
        contract = (
            _current_contract_by_no(db, project_id, contract_no)
            if contract_no else None
        )
        if contract is None:
            if contract_no:
                # Validate availability only; contract creation is deferred to
                # apply after workbook-state/project locks.
                contract = _resolve_plan_contract(
                    db, project_id, contract_no, row_no)
            else:
                contract_no, contract = _xsdd_contract_for_project(
                    db, project_id, row_no)
        month = _v2_date(row[index["报告月份"]], row_no=row_no, label="报告月份")
        if month is None:
            raise WorkbookError("invalid_month", f"第 {row_no} 行报告月份不能为空")
        # Excel 日期单元格通常保存为某月任意一天；数据库快照的身份键是
        # “月份”，并由约束强制 day=1。解析边界统一归一，避免 validate 通过、
        # apply 到 commit 才 500（生产 2026-08-26 已出现）。
        month = month.replace(day=1)
        amount = _v2_decimal(row[index["累计实收金额（含税）"]], row_no=row_no, label="累计实收金额", required=True)
        existing = (
            db.scalar(select(MaintenanceCollectionSnapshot).where(
                MaintenanceCollectionSnapshot.project_contract_id
                == contract.project_contract_id,
                MaintenanceCollectionSnapshot.report_month == month,
            ))
            if contract is not None else None
        )
        # 行级基线三路合并：带实体ID的既有行未触碰则跳过（05 的缺行=作废
        # 对账以文件实体ID列为准，与本 ops 列表无关）。
        raw_receipt_id = str(_cell(row, index, "实体ID") or "").strip()
        if existing is not None and raw_receipt_id:
            baseline = _parse_v2_row_base_token(
                _cell(row, index, V2_BASE_COLUMN), sheet="05_实收回款",
                entity_id=raw_receipt_id, row_no=row_no)
            server_values = _v2_receipt_row_values(
                existing, contract_no or "")
            if not _v2_merge_row(
                    sheet="05_实收回款", row_label=(
                        f"{server_values.get('合同编号') or ''}"
                        f"/{server_values.get('报告月份') or ''}"),
                    entity_id=raw_receipt_id,
                    row=row, index=index, base_fields=V2_RECEIPT_BASE_FIELDS,
                    server_values=server_values, baseline=baseline, ctx=merge):
                continue
        out.append(ec.CollectionOp(
            operation="UPDATE" if existing is not None else "CREATE",
            project_contract_id=(contract.project_contract_id if contract else ""),
            contract_no=contract_no,
            report_month=month, cumulative_amount=amount,
            receipt_reference=str(row[index["回款凭证号"]] or "").strip() or None,
            remark=str(row[index["备注"]] or "").strip() or None,
            collection_status=str(row[index["状态"]] or "confirmed").strip() or "confirmed",
        ))
    return out


def _encode_base_hashes(hashes: dict[str, str]) -> str:
    """行基线哈希图（entity_id -> sha256），用于缺行=作废的行级冲突判定。"""
    return ",".join(f"{eid}:{digest}" for eid, digest in sorted(hashes.items()))


def _decode_base_hashes(value: str | None) -> dict[str, str]:
    out: dict[str, str] = {}
    for chunk in str(value or "").split(","):
        chunk = chunk.strip()
        if not chunk or ":" not in chunk:
            continue
        eid, digest = chunk.rsplit(":", 1)
        out[eid] = digest
    return out


def _encode_row_ids(ids) -> str:
    """导出时行ID全集（P1，Codex review #272）：缺行=作废只允许命中
    「导出时存在、上传时消失」的行；导出之后新导入的行天然豁免，
    应用后原样重传（幂等重试）也不受影响。"""
    return ",".join(str(i) for i in sorted(ids))


def _decode_row_ids(value: str | None) -> set[str]:
    if not value:
        return set()
    return {part for part in value.split(",") if part}


def _expected_expense_ids(db: Session, project_id: str) -> list[str]:
    # 与 _v2_build_expense 同口径：稳定 attribution + current/unshared XSDD。
    return [expense.raw_line_id for expense in _project_expenses(db, project_id)]


def _expected_part_ids(db: Session, project_id: str) -> set[int]:
    """Active 03 entities whose current assignment belongs to this project."""

    from app.services import maintenance_demands

    return set(db.scalars(
        select(FMaintenanceLine.id)
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .join(
            MaintenanceSourceOrderAssignment,
            (MaintenanceSourceOrderAssignment.source_order_id
             == FMaintenanceOrder.raw_order_id)
            & MaintenanceSourceOrderAssignment.is_active.is_(True),
        )
        .where(
            MaintenanceSourceOrderAssignment.project_id == project_id,
            FMaintenanceLine.is_active.is_(True),
            maintenance_demands.active_demand_condition(),
        )
    ).all())


def _expected_site_ids(db: Session, project_id: str) -> set[str]:
    """Active 06 entities whose owning issue belongs to this project."""

    return set(db.scalars(
        select(MaintenanceSiteIssueLine.issue_line_id)
        .join(
            MaintenanceSiteIssue,
            MaintenanceSiteIssue.issue_id == MaintenanceSiteIssueLine.issue_id,
        )
        .where(
            MaintenanceSiteIssue.project_id == project_id,
            MaintenanceSiteIssueLine.is_active.is_(True),
        )
    ).all())


def _v2_assert_entity_scope(
    db: Session,
    *,
    project_id: str,
    part_ids: set[int] | None = None,
    site_ids: set[str] | None = None,
    expense_ids: set[str] | None = None,
    linked_sales_order_nos: set[str] | None = None,
    lock_rows: bool = False,
) -> set[str]:
    """Fail the complete workbook before a foreign fact can be mutated.

    Signed export metadata proves which rows the server exported, but it is not
    an authorization cache: ownership can change between export, validation,
    and apply.  Each boundary therefore rebuilds the three live project sets.
    """

    requested_parts = {int(value) for value in (part_ids or set())}
    requested_sites = {str(value) for value in (site_ids or set())}
    requested_expenses = {str(value) for value in (expense_ids or set())}
    requested_contract_nos = {
        str(value) for value in (linked_sales_order_nos or set()) if value
    }
    locked_project_xsdd_nos: set[str] = set()

    if lock_rows and (requested_parts or requested_contract_nos):
        # Canonical reassignment order: source order -> active assignment ->
        # detail line.  Include every WBDD referenced by a deferred XSDD
        # contract before locking *any* detail line, otherwise a plan-only
        # apply can invert the assignment writer's project->order lock order.
        line_edges = list(db.execute(
            select(FMaintenanceLine.id, FMaintenanceLine.order_id)
            .where(FMaintenanceLine.id.in_(requested_parts))
            .order_by(FMaintenanceLine.id)
        ).all()) if requested_parts else []
        order_ids = {int(order_id) for _line_id, order_id in line_edges}
        if requested_contract_nos:
            order_ids.update(db.scalars(
                select(FMaintenanceOrder.id).where(
                    FMaintenanceOrder.linked_sales_order_no.in_(
                        requested_contract_nos
                    )
                )
            ).all())
        locked_orders = list(db.scalars(
            select(FMaintenanceOrder)
            .where(FMaintenanceOrder.id.in_(sorted(order_ids)))
            .order_by(FMaintenanceOrder.raw_order_id)
            .with_for_update()
        ).all()) if order_ids else []
        source_ids = [order.raw_order_id for order in locked_orders]
        locked_assignments: list[MaintenanceSourceOrderAssignment] = []
        if source_ids:
            locked_assignments = list(db.scalars(
                select(MaintenanceSourceOrderAssignment)
                .where(
                    MaintenanceSourceOrderAssignment.source_order_id.in_(source_ids),
                    MaintenanceSourceOrderAssignment.is_active.is_(True),
                )
                .order_by(MaintenanceSourceOrderAssignment.source_order_id)
                .with_for_update()
            ).all())
        project_sources = {
            assignment.source_order_id
            for assignment in locked_assignments
            if assignment.project_id == project_id
        }
        locked_project_xsdd_nos = {
            str(order.linked_sales_order_no)
            for order in locked_orders
            if (
                order.raw_order_id in project_sources
                and order.linked_sales_order_no
            )
        }
        if requested_parts:
            locked_lines = list(db.scalars(
                select(FMaintenanceLine)
                .where(FMaintenanceLine.id.in_(requested_parts))
                .order_by(FMaintenanceLine.id)
                .with_for_update()
            ).all())
            if {line.id for line in locked_lines} != requested_parts:
                raise WorkbookError(
                    "line_not_in_project",
                    "工作簿包含已不存在、已作废或不属于本项目的备件事实，请重新下载",
                )

    foreign_parts = requested_parts - _expected_part_ids(db, project_id)
    if foreign_parts:
        raise WorkbookError(
            "line_not_in_project",
            "工作簿包含已不存在、已作废或不属于本项目的备件事实，请重新下载",
        )
    if lock_rows and requested_sites:
        # Site writers use parent issue -> child line.  Resolve the edge first,
        # lock parents in stable order, then lock children and re-read the live
        # join; reversing that order deadlocks with an issue-level writer.
        site_edges = list(db.execute(
            select(
                MaintenanceSiteIssueLine.issue_line_id,
                MaintenanceSiteIssueLine.issue_id,
            )
            .where(MaintenanceSiteIssueLine.issue_line_id.in_(requested_sites))
            .order_by(MaintenanceSiteIssueLine.issue_line_id)
        ).all())
        if {line_id for line_id, _issue_id in site_edges} != requested_sites:
            raise WorkbookError(
                "site_line_not_in_project",
                "工作簿包含已不存在、已作废或不属于本项目的领用事实，请重新下载",
            )
        parent_issue_ids = sorted({issue_id for _line_id, issue_id in site_edges})
        list(db.scalars(
            select(MaintenanceSiteIssue)
            .where(MaintenanceSiteIssue.issue_id.in_(parent_issue_ids))
            .order_by(MaintenanceSiteIssue.issue_id)
            .with_for_update()
        ).all())
        locked_site_lines = list(db.scalars(
            select(MaintenanceSiteIssueLine)
            .where(MaintenanceSiteIssueLine.issue_line_id.in_(requested_sites))
            .order_by(MaintenanceSiteIssueLine.issue_line_id)
            .with_for_update()
        ).all())
        if (
            {line.issue_line_id for line in locked_site_lines} != requested_sites
            or {line.issue_id for line in locked_site_lines} != set(parent_issue_ids)
        ):
            raise WorkbookError(
                "site_line_not_in_project",
                "领用事实归属已变化，请重新下载当前项目总表",
            )

    foreign_sites = requested_sites - _expected_site_ids(db, project_id)
    if foreign_sites:
        raise WorkbookError(
            "site_line_not_in_project",
            "工作簿包含已不存在、已作废或不属于本项目的领用事实，请重新下载",
        )
    if lock_rows and requested_expenses:
        expense_keys = {f"bxd:{raw_line_id}" for raw_line_id in requested_expenses}
        locked_attributions = list(db.scalars(
            select(MaintenanceProjectExpenseAttribution)
            .where(
                MaintenanceProjectExpenseAttribution.expense_id.in_(expense_keys)
            )
            .order_by(MaintenanceProjectExpenseAttribution.expense_id)
            .with_for_update()
        ).all())
        if (
            {attribution.expense_id for attribution in locked_attributions}
            != expense_keys
            or any(
                attribution.project_id != project_id
                for attribution in locked_attributions
            )
        ):
            raise WorkbookError(
                "expense_not_in_project",
                "工作簿包含已不存在、已作废或不属于本项目的报销归因，请重新下载",
            )
        locked_expenses = list(db.scalars(
            select(FProjectExpense)
            .where(FProjectExpense.raw_line_id.in_(requested_expenses))
            .order_by(FProjectExpense.raw_line_id)
            .with_for_update()
        ).all())
        if {expense.raw_line_id for expense in locked_expenses} != requested_expenses:
            raise WorkbookError(
                "expense_not_in_project",
                "工作簿包含已不存在、已作废或不属于本项目的报销事实，请重新下载",
            )
    expected_expenses = set(_expected_expense_ids(db, project_id))
    foreign_expenses = requested_expenses - expected_expenses
    if foreign_expenses:
        raise WorkbookError(
            "expense_not_in_project",
            "工作簿包含已不存在、已作废或不属于本项目的报销事实，请重新下载",
        )
    return locked_project_xsdd_nos

def _v2_parse_contract_amount(
    db: Session,
    *,
    project_id: str,
    ws,
    meta: dict[str, str],
) -> ContractAmountChange | None:
    """解析 01 中的项目合同总额；只有唯一生效合同行可无歧义写回。"""

    raw_value = None
    found = False
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if str(row[0] or "").strip() == "合同总额（含税）":
            raw_value = row[1]
            found = True
            break
    if not found:
        raise WorkbookError("missing_contract_total", "01_项目概览缺少合同总额（含税）行，请重新下载")

    exported_text = str(meta.get("contract_total_exported", "")).strip()
    raw_text = "" if raw_value is None else str(raw_value).strip()
    if raw_text in {"", "—"} and exported_text == "":
        return None
    if raw_text in {"", "—"}:
        raise WorkbookError("missing_contract_total", "合同总额（含税）不能清空")
    amount = _v2_decimal(raw_value, row_no=1, label="合同总额（含税）", required=True)
    exported = (
        _v2_decimal(exported_text, row_no=1, label="导出合同总额（含税）", required=True)
        if exported_text else None
    )
    if amount == exported:
        return None

    if meta.get("contract_editable") != "true":
        raise WorkbookError(
            "contract_total_ambiguous",
            "当前项目没有唯一的生效计入合同：不能把项目总额暗中分摊或写入未知合同。"
            "请先在项目合同中完成挂靠；多合同项目请逐合同修改。",
        )
    contract_id = meta.get("contract_edit_id", "")
    try:
        base_version = int(meta.get("contract_edit_version", ""))
    except ValueError as exc:
        raise WorkbookError("template_version_mismatch", "合同版本元数据无效，请重新下载") from exc
    contracts = list(db.scalars(
        select(MaintenanceProjectContract)
        .where(MaintenanceProjectContract.project_id == project_id)
        .order_by(MaintenanceProjectContract.project_contract_id)
    ))
    editable_contracts = _v2_editable_contracts(db, project_id, contracts)
    if len(editable_contracts) != 1:
        raise WorkbookError(
            "contract_total_ambiguous",
            "当前项目已不再只有一条生效计入合同，请重新下载后按最新合同关系处理",
        )
    contract = editable_contracts[0]
    if contract.project_contract_id != contract_id:
        raise WorkbookError("contract_not_found", "合同关系已不存在或不属于当前项目，请重新下载")
    if contract.version != base_version:
        raise WorkbookError(
            "stale_contract",
            f"合同金额已被他人更新（当前版本 {contract.version}），请重新下载后再改",
        )
    return ContractAmountChange(
        project_contract_id=contract.project_contract_id,
        base_version=base_version,
        before_amount_inc_tax=contract.amount_inc_tax,
        amount_inc_tax=amount,
    )


# 上传数据行（带实体ID）低于导出行数的该比例 → 疑似筛选/复制粘贴事故，整本拒绝。


def validate_project_master_v2(
    db: Session,
    *,
    project_id: str,
    data: bytes,
    user_ctx: UserContext | None = None,
    force_takeover: bool = False,
) -> MasterV2Plan:
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise WorkbookError("invalid_file", f"无法读取 .xlsx：{type(exc).__name__}") from exc
    meta = _v2_verify_meta(db, wb, project_id)
    merge = _V2MergeContext(force_takeover=force_takeover)
    included_meta = tuple(
        name.strip() for name in meta["included_sheets"].split(",") if name.strip()
    )
    included = tuple(name for name in included_meta if name in V2_ALL_SHEETS[:-2])

    cost_refills: tuple[CostRefill, ...] = ()
    site_flags: tuple[SiteReturnFlag, ...] = ()
    expense_updates: list[ec.ExpenseUpdate] = []
    expense_voids: list[str] = []
    receipt_ops: tuple[ec.CollectionOp, ...] = ()
    milestone_changes: tuple[V2MilestoneChange, ...] = ()
    will_void_rows: list[dict] = []
    uploaded_line_rows = 0
    uploaded_line_ids: set[int] = set()
    assignment_changes: tuple[SourceOrderAssignmentChange, ...] = ()
    contract_amount_change = None
    if V2_SHEET_OVERVIEW in included:
        contract_amount_change = _v2_parse_contract_amount(
            db,
            project_id=project_id,
            ws=wb[V2_SHEET_OVERVIEW],
            meta=meta,
        )
    if V2_SHEET_PARTS in included:
        (parsed_refills, uploaded_line_rows, uploaded_line_ids,
         assignment_changes) = _v2_parse_parts(
            db, project_id, wb[V2_SHEET_PARTS], merge)
        cost_refills = tuple(parsed_refills)
        if (assignment_changes and user_ctx is not None
                and user_ctx.role not in FULL_SCOPE_ROLES):
            raise WorkbookError(
                "assignment_permission_denied",
                "这份项目总表会更正统一 WBDD 的项目归属，仅管理员或全量项目账号可确认")
    if V2_SHEET_SITE in included:
        site_flags, uploaded_site_ids = _v2_parse_site(
            db, project_id, wb[V2_SHEET_SITE], merge)
        site_flags = list(site_flags)
        # 2026-08-23：06 缺行=作废（用户口径：Excel 删行覆盖上传，没有的默认作废）
        export_site_ids = _decode_row_ids(meta.get("site_row_ids"))
        export_site_hashes = _decode_base_hashes(meta.get("site_base_hashes"))
        missing_site_ids = [sid for sid in export_site_ids
                            if sid not in uploaded_site_ids]
        for sid in missing_site_ids:
            line_row = db.get(MaintenanceSiteIssueLine, sid)
            if line_row is not None and line_row.is_active:
                issue_row = db.get(MaintenanceSiteIssue, line_row.issue_id)
                values = _v2_site_row_values(line_row, issue_row) if issue_row else {}
                exported_hash = export_site_hashes.get(sid)
                if (exported_hash is not None
                        and _v2_row_base_hash(values, V2_SITE_BASE_FIELDS)
                        != exported_hash):
                    if merge.record_row_conflict(
                            sheet="06_领用返还",
                            row_label=str(values.get("领用单号") or sid),
                            entity_id=sid, action="删行=作废（服务端该行已变）"):
                        site_flags.append(SiteReturnFlag(
                            issue_line_id=sid, no_return=None, is_void=True))
                    continue
                site_flags.append(SiteReturnFlag(
                    issue_line_id=sid, no_return=None, is_void=True))
        site_flags = tuple(site_flags)
    if V2_SHEET_EXPENSE in included:
        (expense_updates, expense_voids,
         uploaded_expense_present) = _v2_parse_expenses(
            db, project_id, wb[V2_SHEET_EXPENSE], merge)
    if V2_SHEET_RECEIPTS in included:
        receipt_ws = wb[V2_SHEET_RECEIPTS]
        receipt_ops = tuple(_v2_parse_receipts(
            db, project_id, receipt_ws, merge))
        # 05 缺行=作废：只允许命中导出时由 HMAC 签名的稳定 collection_id。
        # 导出之后新产生的回款不在 envelope 中，不能被旧文件误删。
        export_receipt_ids = _decode_row_ids(meta.get("receipt_row_ids"))
        receipt_headers = [str(cell.value or "") for cell in receipt_ws[1]]
        try:
            entity_index = receipt_headers.index("实体ID")
        except ValueError as exc:
            raise WorkbookError(
                "template_column_missing",
                "05_实收回款缺少隐藏实体ID列，请重新下载项目总表",
            ) from exc
        uploaded_receipt_ids = {
            str(row[entity_index]).strip()
            for row in receipt_ws.iter_rows(min_row=2, values_only=True)
            if row and row[entity_index] not in (None, "")
        }
        missing_receipt_ids = sorted(export_receipt_ids - uploaded_receipt_ids)
        if missing_receipt_ids:
            contracts_by_id = {
                contract.project_contract_id: contract
                for contract in _current_contracts(db, project_id)
            }
            parsed_receipt_keys = {
                (operation.project_contract_id, operation.report_month)
                for operation in receipt_ops
                if operation.project_contract_id
            }
            implicit_voids: list[ec.CollectionOp] = []
            for collection_id in missing_receipt_ids:
                snapshot = db.get(MaintenanceCollectionSnapshot, collection_id)
                if (
                    snapshot is None
                    or snapshot.project_id != project_id
                    or snapshot.status != "confirmed"
                ):
                    raise WorkbookError(
                        "collection_not_found",
                        "导出后的实收回款已不存在、已作废或不属于本项目，请重新下载",
                    )
                contract = contracts_by_id.get(snapshot.project_contract_id)
                if contract is None:
                    raise WorkbookError(
                        "collection_contract_not_current",
                        "导出后的实收回款关联合同已失效，请重新下载并先处理合同关系",
                    )
                if (
                    snapshot.project_contract_id,
                    snapshot.report_month,
                ) in parsed_receipt_keys:
                    raise WorkbookError(
                        "receipt_identity_lost",
                        "05_实收回款中仍有相同合同和月份的行，但隐藏实体ID已丢失；"
                        "为避免把保留的回款误作废，请重新下载项目总表后修改",
                    )
                # 2.7.0 行级冲突：删行=作废时若服务端该快照已变，未接管则跳过
                exported_hash = _decode_base_hashes(
                    meta.get("receipt_base_hashes")).get(collection_id)
                if exported_hash is not None and _v2_row_base_hash(
                        _v2_receipt_row_values(snapshot, contract.contract_no),
                        V2_RECEIPT_BASE_FIELDS) != exported_hash:
                    if not merge.record_row_conflict(
                            sheet=V2_SHEET_RECEIPTS,
                            row_label=(
                                f"{contract.contract_no} "
                                f"{snapshot.report_month:%Y-%m}"),
                            entity_id=collection_id,
                            action="删行=作废（服务端该行已变）"):
                        continue
                implicit_voids.append(ec.CollectionOp(
                    operation="VOID",
                    project_contract_id=contract.project_contract_id,
                    contract_no=contract.contract_no,
                    report_month=snapshot.report_month,
                    cumulative_amount=None,
                    receipt_reference=None,
                    remark=None,
                    collection_status=None,
                ))
                will_void_rows.append({
                    "sheet": V2_SHEET_RECEIPTS,
                    "entity_id": collection_id,
                    "label": (
                        f"{contract.contract_no} "
                        f"{snapshot.report_month:%Y-%m}"
                    ),
                    "reason": "上传文件缺行",
                })
            receipt_ops = receipt_ops + tuple(implicit_voids)
    if V2_SHEET_PLAN in included:
        plan_ws = wb[V2_SHEET_PLAN]
        milestone_changes = tuple(_v2_parse_plan(db, project_id, plan_ws))
        # 02 缺行=作废（2026-08-20 用户三连问）：只命中导出时存在的里程碑
        export_plan_ids = _decode_row_ids(meta.get("plan_row_ids"))
        plan_headers = [str(cell.value or "") for cell in plan_ws[1]]
        entity_index = plan_headers.index("实体ID")
        uploaded_plan_ids = {
            str(row[entity_index]).strip()
            for row in plan_ws.iter_rows(min_row=2, values_only=True)
            if row and row[entity_index] not in (None, "")
        }
        # Parser intentionally elides unchanged rows.  Missing-row detection
        # must compare the physical uploaded identity set, not the mutation
        # list, otherwise every unchanged milestone is silently voided.
        extra_voids = sorted(export_plan_ids - uploaded_plan_ids)
        if extra_voids:
            from app.models.maintenance_manager import MaintenanceCollectionMilestone

            for mid in extra_voids:
                ms = db.get(MaintenanceCollectionMilestone, mid)
                if ms is None or not ms.is_active or ms.project_id != project_id:
                    raise WorkbookError(
                        "milestone_not_found",
                        "导出后的回款计划已不存在、已作废或不属于本项目，请重新下载",
                    )
                contract_no = next(
                    (c.contract_no for c in _current_contracts(db, project_id)
                     if c.project_contract_id == ms.project_contract_id), "")
                if not contract_no:
                    raise WorkbookError(
                        "milestone_contract_not_current",
                        "导出后的回款计划关联合同已失效，请重新下载并先处理合同关系",
                    )
                milestone_changes = milestone_changes + (V2MilestoneChange(
                    operation="VOID", contract_no=contract_no, sequence=ms.sequence,
                    planned_date=None, date_precision="day", planned_amount=None,
                    entity_id=mid, base_version=None),)
                will_void_rows.append({"sheet": "02_回款计划", "entity_id": mid,
                                       "label": f"第{ms.sequence}期", "reason": "上传文件缺行"})

    # ---- 缺行=作废（04，#264 契约）+ 行数骤减防呆（03/04 共用一道） ----
    for refill in cost_refills:
        if refill.operation == "VOID":
            will_void_rows.append({"sheet": "03_备件明细", "entity_id": str(refill.line_id),
                                   "label": f"{refill.order_no or ''}"})

    # 03 缺行=作废（用户 2026-08-20 拍板，与 04 同语义）：只命中「导出时存在、
    # 上传时消失」的行；导出后新导入的行不在导出全集里，天然豁免。
    if V2_SHEET_PARTS in included:
        export_parts_ids = _decode_row_ids(meta.get("parts_row_ids"))
        export_parts_hashes = _decode_base_hashes(meta.get("parts_base_hashes"))
        voided_line_ids = {r.line_id for r in cost_refills if r.operation == "VOID"}
        for line_id_str in export_parts_ids - {str(i) for i in uploaded_line_ids}:
            line_id = int(line_id_str)
            if line_id in voided_line_ids:
                continue
            missing_line = db.get(FMaintenanceLine, line_id)
            exported_hash = export_parts_hashes.get(line_id_str)
            if (missing_line is not None and missing_line.is_active
                    and exported_hash is not None):
                missing_order = db.get(FMaintenanceOrder, missing_line.order_id)
                missing_override = db.scalar(
                    select(MaintenanceManualCostOverride).where(
                        MaintenanceManualCostOverride.line_id == line_id,
                        MaintenanceManualCostOverride.active.is_(True),
                    ))
                values = _v2_part_row_values(
                    missing_line, missing_order, missing_override)
                if _v2_row_base_hash(values, V2_PART_BASE_FIELDS) != exported_hash:
                    if merge.record_row_conflict(
                            sheet="03_备件明细",
                            row_label=str(values.get("维保单号") or line_id_str),
                            entity_id=line_id_str,
                            action="删行=作废（服务端该行已变）"):
                        cost_refills = cost_refills + (CostRefill(
                            line_id=line_id, operation="VOID",
                            unit_cost_ex_tax=None, unit_cost_inc_tax=None,
                            reason=None),)
                        will_void_rows.append(
                            {"sheet": "03_备件明细", "entity_id": line_id_str,
                             "label": "", "reason": "上传文件缺行（接管覆盖）"})
                    continue
            cost_refills = cost_refills + (CostRefill(
                line_id=line_id, operation="VOID",
                unit_cost_ex_tax=None, unit_cost_inc_tax=None, reason=None),)
            will_void_rows.append({"sheet": "03_备件明细", "entity_id": line_id_str,
                                   "label": "", "reason": "上传文件缺行"})

    if V2_SHEET_EXPENSE in included:
        # 缺行=作废只针对导出时存在的行（P1，Codex review #272）：
        # 导出后新导入的行不在导出全集里，天然豁免误杀。
        export_expense_ids = _decode_row_ids(meta.get("expense_row_ids"))
        export_expense_hashes = _decode_base_hashes(meta.get("expense_base_hashes"))
        missing_expense_ids = [rid for rid in export_expense_ids
                               if rid not in uploaded_expense_present]
        for rid in missing_expense_ids:
            expense_row = db.scalar(select(FProjectExpense).where(
                FProjectExpense.raw_line_id == rid))
            exported_hash = export_expense_hashes.get(rid)
            if (expense_row is not None and exported_hash is not None
                    and _v2_row_base_hash(
                        _v2_expense_row_values(expense_row, ""),
                        V2_EXPENSE_BASE_FIELDS) != exported_hash):
                if merge.record_row_conflict(
                        sheet="04_费用报销",
                        row_label=str(expense_row.bxd_no or rid),
                        entity_id=rid, action="删行=作废（服务端该行已变）"):
                    expense_voids.append(rid)
                    will_void_rows.append(
                        {"sheet": "04_费用报销", "entity_id": rid, "label": ""})
                continue
            expense_voids.append(rid)
            will_void_rows.append({"sheet": "04_费用报销", "entity_id": rid, "label": ""})
        # 2026-08-22 用户拍板：撤销行损失防呆（原 50% 批量损失拦截）——
        # 线上进入大批量作废期，需要能全量增删改。缺行=作废的语义与审计不变；
        # 误传风险由前端的 will_void_rows 确认弹窗兜底。

    if V2_SHEET_PARTS in included:
        export_parts_ids = _decode_row_ids(meta.get("parts_row_ids"))

    unmatched_error = _unmatched_pn_error(db, merge)
    if unmatched_error is not None:
        raise unmatched_error

    # Metadata HMAC proves the exported identity envelope was not forged, while
    # this live check proves every explicit/implicit target still belongs to
    # the URL project.  Do not treat signed metadata as an authorization cache.
    _v2_assert_entity_scope(
        db,
        project_id=project_id,
        part_ids={
            int(refill.line_id)
            for refill in cost_refills
            if not refill.is_create and refill.line_id is not None
        },
        site_ids={
            flag.issue_line_id for flag in site_flags if not flag.is_create
        },
        expense_ids=(
            {
                update.raw_line_id
                for update in expense_updates
                if not getattr(update, "is_create", False)
            }
            | set(expense_voids)
        ),
    )

    return MasterV2Plan(
        project_id=project_id,
        sheets=included,
        export_id=meta["export_id"],
        file_sha256=hashlib.sha256(data).hexdigest(),
        expected_workbook_revision=int(meta["workbook_revision"]),
        expected_workbook_data_version=meta["workbook_data_version"],
        cost_refills=cost_refills,
        site_flags=site_flags,
        expense_updates=tuple(expense_updates),
        receipt_ops=receipt_ops,
        milestone_changes=milestone_changes,
        assignment_changes=assignment_changes,
        contract_amount_change=contract_amount_change,
        expense_voids=tuple(expense_voids),
        will_void_rows=tuple(will_void_rows),
        force_takeover=merge.force_takeover,
        field_changes=tuple(merge.changes),
        conflicts=tuple(merge.conflicts),
        overridden=tuple(merge.overridden),
        warnings=tuple(merge.warnings),
    )


def _v2_apply_operation_key(plan: MasterV2Plan, operated_by: str) -> str | None:
    if not plan.export_id or not plan.file_sha256:
        return None
    identity = "|".join([
        plan.project_id, plan.export_id, plan.file_sha256, operated_by,
    ])
    return "master-v2:" + hashlib.sha256(identity.encode("utf-8")).hexdigest()


def _v2_apply_result(
    plan: MasterV2Plan,
    *,
    operated_by: str,
    import_batch_id: str,
    replayed: bool,
    revision_drift: bool = False,
) -> dict:
    return {
        "applied_by": operated_by,
        "import_batch_id": import_batch_id,
        "protocol_id": V2_PROTOCOL_ID,
        "template_version": V2_TEMPLATE_VERSION,
        "project_id": plan.project_id,
        "sheets": list(plan.sheets),
        **plan.summary,
        "will_void_rows": [dict(row) for row in plan.will_void_rows],
        "changes": [dict(item) for item in plan.field_changes],
        "conflicts": [dict(item) for item in plan.conflicts],
        "overridden": [dict(item) for item in plan.overridden],
        "force_takeover": plan.force_takeover,
        "revision_drift": revision_drift,
        "warnings": list(plan.warnings),
        "replayed": replayed,
    }


def apply_project_master_v2(
    db: Session,
    plan: MasterV2Plan,
    *,
    operated_by: str,
    import_batch_id: str,
    user_ctx: UserContext | None = None,
) -> dict:
    # All mutations deliberately happen on this one Session transaction.
    apply_operation_key = _v2_apply_operation_key(plan, operated_by)
    if apply_operation_key is not None:
        # The caller creates a fresh request UUID for every HTTP attempt.  A
        # deterministic server batch identity makes an ACK-loss replay return
        # the same observable result and keeps downstream import references
        # stable as well.
        import_batch_id = (
            "v2-" + hashlib.sha256(
                apply_operation_key.encode("utf-8")
            ).hexdigest()[:36]
        )
    audit_reason = f"项目总表应用 {import_batch_id[:8]}"
    from app.services import maintenance_project_operations as operations
    operating_fact_changed = False

    # Global write lock order for the complete workbook transaction:
    # data-change advisory -> workbook states(sorted) -> project -> source
    # order -> assignment -> detail facts.  Import/recompute already take the
    # advisory first; doing the same here prevents state->global inversion when
    # this apply later cascades a demand tombstone or recomputes line costs.
    db.execute(
        select(func.pg_advisory_xact_lock(config.DATA_CHANGE_ADVISORY_LOCK_KEY))
    )
    project_exists = db.scalar(
        select(MaintenanceProject.project_id).where(
            MaintenanceProject.project_id == plan.project_id)
    )
    if project_exists is None:
        raise WorkbookError("project_not_editable", "项目已不存在，请重新下载")
    state_project_ids = {plan.project_id}
    assignment_source_ids = {
        change.source_order_id for change in plan.assignment_changes
    }
    prestate_contract_nos = {
        change.contract_no for change in plan.milestone_changes
    } | {
        operation.contract_no for operation in plan.receipt_ops
    } | {
        update.contract_no for update in plan.expense_updates
        if update.contract_no
    }
    prestate_assignment_xsdds = list(db.scalars(
        select(FMaintenanceOrder.linked_sales_order_no).where(
            FMaintenanceOrder.raw_order_id.in_(assignment_source_ids or {""}),
            FMaintenanceOrder.linked_sales_order_no.is_not(None),
        )
    ))
    # VOID is identity-bound, so its contract comes from the raw fact, never
    # from an editable workbook cell.  This unlocked pre-read only defines the
    # advisory-lock envelope; the value is re-read after the expense row lock.
    prelock_void_contract_nos = {
        str(value)
        for value in db.scalars(
            select(FProjectExpense.linked_sales_order_no).where(
                FProjectExpense.raw_line_id.in_(plan.expense_voids),
                FProjectExpense.linked_sales_order_no.is_not(None),
            )
        ).all()
        if value
    } if plan.expense_voids else set()
    # XSDD advisory identities are absent-row locks and therefore precede every
    # workbook/project row lock.  Later contract/assignment triggers reacquire
    # the same transaction locks without creating a state -> identity inversion.
    maintenance_project_identity.lock_xsdd_identities(
        db,
        prestate_contract_nos
        | set(prestate_assignment_xsdds)
        | prelock_void_contract_nos,
    )
    if assignment_source_ids:
        state_project_ids.update(
            value
            for value in db.scalars(
                select(MaintenanceSourceOrderAssignment.project_id).where(
                    MaintenanceSourceOrderAssignment.source_order_id.in_(
                        assignment_source_ids
                    ),
                    MaintenanceSourceOrderAssignment.is_active.is_(True),
                )
            ).all()
            if value
        )
    void_line_ids = {
        int(refill.line_id)
        for refill in plan.cost_refills
        if (
            refill.operation == "VOID"
            and refill.line_id is not None
            and not refill.is_create
        )
    }
    if void_line_ids:
        # A line-level VOID may cascade into an order tombstone and deactivate
        # its active assignment.  Include every current owner in the same
        # initial sorted state-lock pass; the cascade service only reuses this
        # map and fails closed if ownership changes afterwards.
        state_project_ids.update(
            value
            for value in db.scalars(
                select(MaintenanceSourceOrderAssignment.project_id)
                .join(
                    FMaintenanceOrder,
                    FMaintenanceOrder.raw_order_id
                    == MaintenanceSourceOrderAssignment.source_order_id,
                )
                .join(
                    FMaintenanceLine,
                    FMaintenanceLine.order_id == FMaintenanceOrder.id,
                )
                .where(
                    FMaintenanceLine.id.in_(sorted(void_line_ids)),
                    MaintenanceSourceOrderAssignment.is_active.is_(True),
                )
            ).all()
            if value
        )
    workbook_states = operations.lock_workbook_states(
        db,
        project_ids=state_project_ids,
    )
    workbook_state = workbook_states[plan.project_id]
    project = db.scalar(
        select(MaintenanceProject)
        .where(MaintenanceProject.project_id == plan.project_id)
        .with_for_update()
    )
    if project is None or not project.is_active:
        raise WorkbookError("project_not_editable", "项目已不存在或归档，请重新下载")
    if apply_operation_key is not None:
        # SELECT ... FOR UPDATE cannot lock an absent unique key.  Serialize
        # the deterministic receipt identity first so concurrent first applies
        # become writer + replay instead of two writers racing at commit.
        db.execute(select(func.pg_advisory_xact_lock(
            func.hashtextextended(
                f"maintenance-master-v2-apply:{apply_operation_key}", 0
            )
        )))
        receipt = db.scalar(
            select(MaintenanceProjectWorkbookOperation)
            .where(
                MaintenanceProjectWorkbookOperation.operation_key
                == apply_operation_key
            )
            .with_for_update()
        )
        if receipt is not None:
            if (
                receipt.project_id != plan.project_id
                or receipt.export_id != plan.export_id
                or receipt.file_sha256 != plan.file_sha256
                or receipt.payload_hash != plan.file_sha256
                or receipt.operated_by != operated_by
                or receipt.operation_type != "file_apply"
            ):
                raise WorkbookError(
                    "idempotency_conflict",
                    "工作簿应用回执与当前请求不一致，请重新下载",
                )
            db.commit()
            return _v2_apply_result(
                plan,
                operated_by=operated_by,
                import_batch_id=import_batch_id,
                replayed=True,
            )  # replay：以原回执语义返回，冲突字段为空
    carries_export_identity = bool(plan.export_id or plan.file_sha256)
    if carries_export_identity and (
        plan.expected_workbook_revision is None
        or not plan.expected_workbook_data_version
    ):
        raise WorkbookError(
            "invalid_concurrency_token",
            "工作簿缺少项目并发版本，请重新下载当前项目总表。",
        )
    # 2.7.0：整本 revision 不再硬拒（行级基线三路合并已区分用户改动与
    # 导出旧值）；漂移仅记录，供回执与审计观测。
    revision_drift = bool(
        plan.expected_workbook_revision is not None
        and (
            workbook_state.revision != plan.expected_workbook_revision
            or workbook_state.data_version != plan.expected_workbook_data_version
        )
    )
    if plan.conflicts and not plan.force_takeover:
        raise WorkbookError(
            "row_conflicts",
            "部分行已被他人更新，本次上传未写入任何数据；请处理后重试或选择强制接管。",
            issues=[dict(item) for item in plan.conflicts],
        )
    plan_contract_nos = prestate_contract_nos
    prelock_contract_nos = plan_contract_nos | prelock_void_contract_nos
    # The plan may have been validated minutes earlier.  Rebuild all live
    # ownership sets and lock the target facts before *any* contract, demand,
    # site, or expense write, so a stale/crafted plan fails as one zero-write
    # transaction.
    assigned_xsdd_nos = _v2_assert_entity_scope(
        db,
        project_id=plan.project_id,
        part_ids={
            int(refill.line_id)
            for refill in plan.cost_refills
            if not refill.is_create and refill.line_id is not None
        },
        site_ids={
            flag.issue_line_id for flag in plan.site_flags if not flag.is_create
        },
        expense_ids=(
            {
                update.raw_line_id
                for update in plan.expense_updates
                if not getattr(update, "is_create", False)
            }
            | set(plan.expense_voids)
        ),
        linked_sales_order_nos=prelock_contract_nos,
        lock_rows=True,
    )
    locked_void_contract_nos = {
        str(value)
        for value in db.scalars(
            select(FProjectExpense.linked_sales_order_no).where(
                FProjectExpense.raw_line_id.in_(plan.expense_voids),
                FProjectExpense.linked_sales_order_no.is_not(None),
            )
        ).all()
        if value
    } if plan.expense_voids else set()
    if not locked_void_contract_nos.issubset(prelock_void_contract_nos):
        raise WorkbookError(
            "expense_not_in_project",
            "报销归集合同在应用前已变化，请重新下载",
        )
    needed_contract_nos = sorted(plan_contract_nos | locked_void_contract_nos)
    current_contract_by_no: dict[str, MaintenanceProjectContract] = {}
    for contract_no in needed_contract_nos:
        contract, contract_created = _ensure_contract_for_xsdd_apply(
            db,
            project=project,
            contract_no=contract_no,
            operated_by=operated_by,
            assigned_xsdd_nos=assigned_xsdd_nos,
        )
        current_contract_by_no[contract_no] = contract
        operating_fact_changed = contract_created or operating_fact_changed
    receipt_ops = tuple(
        ec.CollectionOp(
            operation=operation.operation,
            project_contract_id=current_contract_by_no[
                operation.contract_no
            ].project_contract_id,
            contract_no=operation.contract_no,
            report_month=operation.report_month,
            cumulative_amount=operation.cumulative_amount,
            receipt_reference=operation.receipt_reference,
            remark=operation.remark,
            collection_status=operation.collection_status,
        )
        for operation in plan.receipt_ops
    )
    if plan.contract_amount_change is not None:
        change = plan.contract_amount_change
        # Direct contract create/update also takes the state and project locks.
        # Therefore no second current contract can appear after this recheck.
        contracts = list(db.scalars(
            select(MaintenanceProjectContract)
            .where(MaintenanceProjectContract.project_id == plan.project_id)
            .order_by(MaintenanceProjectContract.project_contract_id)
            .with_for_update()
        ))
        editable_contracts = _v2_editable_contracts(db, plan.project_id, contracts)
        if (len(editable_contracts) != 1
                or editable_contracts[0].project_contract_id
                != change.project_contract_id):
            raise WorkbookError(
                "contract_total_ambiguous",
                "合同关系已变化，当前项目不再只有一条生效计入合同，请重新下载",
            )
        contract = editable_contracts[0]
        if contract.version != change.base_version:
            raise WorkbookError(
                "stale_contract",
                f"合同金额已被他人更新（当前版本 {contract.version}），请重新下载后再改",
            )
        before_amount = contract.amount_inc_tax
        if before_amount != change.amount_inc_tax:
            before = operations.contract_dict(contract)
            contract.amount_inc_tax = change.amount_inc_tax
            contract.source = "project_master_workbook"
            contract.version += 1
            operations._audit_contract(
                db,
                contract,
                action="workbook_update_amount",
                operated_by=operated_by,
                reason=audit_reason,
                before=before,
                after=operations.contract_dict(contract),
            )
            operating_fact_changed = True
    if plan.assignment_changes:
        if user_ctx is None:
            raise WorkbookError(
                "assignment_permission_denied",
                "本次回传包含 WBDD 项目归属更正，缺少可审计的登录身份")
        from app.services import maintenance_source_assignments as source_assignments

        assignment_changed_project_ids: set[str] = set()
        try:
            source_assignments.assign_source_orders(
                db,
                project_id=plan.project_id,
                items=[{
                    "source_order_id": change.source_order_id,
                    "expected_assignment_id": change.expected_assignment_id,
                    "expected_version": change.expected_version,
                } for change in plan.assignment_changes],
                reason=f"项目总表人工认证更正 WBDD 归属（{import_batch_id[:8]}）",
                operated_by=operated_by,
                user_ctx=user_ctx,
                _prelocked_states=workbook_states,
                _changed_project_ids=assignment_changed_project_ids,
            )
            operating_fact_changed = (
                bool(assignment_changed_project_ids)
                or operating_fact_changed
            )
        except source_assignments.SourceAssignmentConflict as exc:
            raise WorkbookError("stale_assignment", str(exc)) from exc
        except source_assignments.SourceAssignmentPermissionError as exc:
            raise WorkbookError("assignment_permission_denied", str(exc)) from exc
        except source_assignments.SourceAssignmentError as exc:
            raise WorkbookError("assignment_invalid", str(exc)) from exc
    # 手工新增行需要一个 import_batch（NOT NULL FK）
    manual_batch: SysImportBatch | None = None
    if any(r.is_create for r in plan.cost_refills):
        manual_batch = SysImportBatch(
            filename="manual-maintenance-line-workbook.xlsx",
            file_type="maintenance",
            file_hash=hashlib.sha256(
                f"manual-line:{apply_operation_key or import_batch_id}".encode(
                    "utf-8"
                )
            ).hexdigest(),
            uploaded_by=operated_by,
            rows_total=sum(r.is_create for r in plan.cost_refills),
            rows_inserted=sum(r.is_create for r in plan.cost_refills),
            status="success",
            report_json={"source": "workbook_manual_create",
                         "project_id": plan.project_id},
        )
        db.add(manual_batch)
        db.flush()

    demand_reprice_line_ids: set[int] = set()
    for refill_index, refill in enumerate(plan.cost_refills):
        if refill.is_create:
            order = db.get(FMaintenanceOrder, refill.order_id)
            base_line_no = db.scalar(
                select(func.coalesce(func.max(FMaintenanceLine.line_no), 0))
                .where(FMaintenanceLine.order_id == refill.order_id)) or 0
            stable_create_identity = "|".join([
                plan.export_id or import_batch_id,
                plan.file_sha256 or import_batch_id,
                "parts",
                str(refill_index),
                str(refill.order_id),
                str(refill.part_id),
            ])
            new_line = FMaintenanceLine(
                raw_line_id=(
                    "manual-line:"
                    + hashlib.sha256(
                        stable_create_identity.encode("utf-8")
                    ).hexdigest()[:48]
                ),
                order_id=refill.order_id,
                line_no=int(base_line_no) + 1,
                part_id=refill.part_id,
                pn_std=refill.pn, pn_raw=refill.pn,
                description=refill.description,
                qty=refill.qty, return_qty=refill.return_qty or Decimal(0),
                serial_numbers=refill.serial_numbers,
                line_note=refill.note,
                edited_source="workbook_manual",
                is_active=True,
                import_batch_id=manual_batch.id,
            )
            if refill.unit_cost_ex_tax is not None:
                new_line.cost_source = "manual"
                new_line.cost_tax_basis = "ex"
                _recompute_line_amounts(new_line, unit_cost_ex_tax=refill.unit_cost_ex_tax,
                                        unit_cost_inc_tax=refill.unit_cost_inc_tax)
            db.add(new_line)
            db.flush()
            if refill.unit_cost_ex_tax is not None:
                db.add(MaintenanceManualCostOverride(
                    line_id=new_line.id,
                    unit_cost_ex_tax=refill.unit_cost_ex_tax,
                    unit_cost_inc_tax=refill.unit_cost_inc_tax,
                    reason=refill.reason, active=True, version=1,
                    updated_by=operated_by))
            # 新行无论是否同时填了人工价都进入权威取价瀑布：自动采购/销售
            # 证据优先，人工 override 只在自动证据仍缺失时兜底。
            demand_reprice_line_ids.add(new_line.id)
            _write_audit(db, project_id=plan.project_id,
                         entity_type="maintenance_line", entity_id=new_line.id,
                         action="CREATE", operated_by=operated_by, reason=audit_reason,
                         after={"pn": refill.pn, "qty": str(refill.qty),
                                "order_no": refill.order_no})
            operating_fact_changed = True
            continue

        line = db.get(FMaintenanceLine, refill.line_id)
        if line is None:
            raise WorkbookError("line_not_found",
                                f"备件行 {refill.line_id} 已不存在，请重新下载")
        if refill.operation == "VOID":
            if line.is_active:
                line.is_active = False
                line.voided_at = datetime.now(timezone.utc)
                line.voided_by = operated_by
                line.void_reason = audit_reason
                cascaded = _cascade_void_site_lines(
                    db, line, project_id=plan.project_id)
                _write_audit(db, project_id=plan.project_id,
                             entity_type="maintenance_line", entity_id=line.id,
                             action="VOID", operated_by=operated_by, reason=audit_reason,
                             after={"cascaded_site_lines": cascaded})
                operating_fact_changed = True
            continue

        # UPDATE：行级数据字段
        before: dict[str, object] = {}
        if refill.pn is not None:
            before["pn"] = line.pn_std or line.pn_raw
            line.pn_std = refill.pn
            line.pn_raw = refill.pn
            # 跨表身份同步换 part_id（P1，Codex review #272）：库存/成本/池
            # 等 join 都走 part_id，只改文本不改主键会腐化下游聚合。
            if refill.part_id is not None:
                before["part_id"] = line.part_id
                line.part_id = refill.part_id
            # 旧 PN 的人工证据不能跟着 line_id 偷渡到新 PN。若用户本行同时填了
            # 新人工价，下面 _merge_manual_cost_to_line 会以新证据重新激活。
            stale_override = db.scalar(
                select(MaintenanceManualCostOverride).where(
                    MaintenanceManualCostOverride.line_id == line.id,
                    MaintenanceManualCostOverride.active.is_(True),
                )
            )
            if stale_override is not None:
                stale_override.active = False
                stale_override.version += 1
                stale_override.updated_by = operated_by
            demand_reprice_line_ids.add(line.id)
        if refill.description is not None:
            before["description"] = line.description
            line.description = refill.description
        if refill.qty is not None:
            before["qty"] = str(line.qty) if line.qty is not None else None
            line.qty = refill.qty
        if refill.return_qty is not None:
            before["return_qty"] = str(line.return_qty) if line.return_qty is not None else None
            line.return_qty = refill.return_qty
        if refill.serial_numbers is not None:
            before["sn"] = line.serial_numbers
            line.serial_numbers = refill.serial_numbers
        if refill.note is not None:
            before["note"] = line.line_note
            line.line_note = refill.note
        # 数量类字段有变化 → 重算金额（无成本行保持 NULL）
        if "qty" in before or "return_qty" in before:
            _recompute_line_amounts(line)
        # 人工成本（写 override + 合并主表，沿用既有语义）
        cost_changed = False
        if refill.unit_cost_ex_tax is not None or refill.reason is not None:
            cost_changed = _merge_manual_cost_to_line(
                db, refill, operated_by=operated_by)
        line.edited_source = "workbook_manual"
        if before or cost_changed:
            after = {"pn": line.pn_std, "qty": str(line.qty),
                     "return_qty": str(line.return_qty)}
            if refill.unit_cost_ex_tax is not None:
                after["cost"] = str(refill.unit_cost_ex_tax)
            if refill.reason is not None:
                after["cost_reason"] = refill.reason
            _write_audit(db, project_id=plan.project_id,
                         entity_type="maintenance_line", entity_id=line.id,
                         action="UPDATE", operated_by=operated_by, reason=audit_reason,
                         before=before or None, after=after)
            operating_fact_changed = True
    # 行级作废的整单级联：被 VOID 的行所属需求单若活动行归零 → 整单墓碑
    # （需求单搜索消失、氚云重传不复活；恢复走 demands 页面既有入口）。
    db.flush()  # Session autoflush=False：先把 is_active=False 刷库，级联计数才准
    voided_order_ids = {
        line.order_id
        for line in (
            db.get(FMaintenanceLine, r.line_id)
            for r in plan.cost_refills if r.operation == "VOID"
        )
        if line is not None
    }
    if voided_order_ids:
        from app.services import maintenance_demands as _demands

        raw_ids = db.scalars(
            select(FMaintenanceOrder.raw_order_id).where(
                FMaintenanceOrder.id.in_(voided_order_ids))
        ).all()
        try:
            cascaded = _demands.cascade_tombstone_orders(
                db, source_order_ids=list(raw_ids),
                operated_by=operated_by,
                reason=f"{audit_reason}（行全部作废级联）",
                _prelocked_states=workbook_states,
            )
        except _demands.DeleteIntentConflict as exc:
            raise WorkbookError(
                "stale_workbook",
                "需求单状态或项目归属已变化，请重新下载项目总表后再操作",
            ) from exc
        if cascaded:
            # entity_id 是 varchar(64)，不能把多张 source_order_id 逗号拼接到一行。
            # 每张单独立审计也让后续检索/追责无需再拆字符串。
            for source_order_id in cascaded:
                _write_audit(
                    db,
                    project_id=plan.project_id,
                    entity_type="maintenance_order",
                    entity_id=source_order_id,
                    action="VOID",
                    operated_by=operated_by,
                    reason=audit_reason,
                    after={
                        "cascaded_order": source_order_id,
                        "cascade_count": len(cascaded),
                    },
                )
    if demand_reprice_line_ids:
        # 与 WBDD/采购导入共用同一权威价格瀑布；commit=False 保持整本工作簿
        # “全成或全败”。这一步也会清除 PN 变化前的旧成本 provenance。
        from app.services import maintenance_cost as _maintenance_cost

        db.flush()
        try:
            _maintenance_cost.recompute(
                db,
                commit=False,
                line_ids=demand_reprice_line_ids,
            )
        except _maintenance_cost.MaintenanceCostRecomputeBusy as exc:
            raise WorkbookError("cost_recompute_busy", str(exc)) from exc
    # 2026-08-24：工作簿建行/改量后立即取价——此前建行不调取价，新行成本
    # 为空要等下一次全局回填才恢复（8-24 两个项目新增 38 行无价的根因）。
    # Apply each repeated issue header once.  Parser normalization guarantees
    # all rows of the same issue carry one chosen value, but recheck the plan so
    # a forged in-process plan cannot restore row-order-dependent behavior.
    requested_headers: dict[str, tuple[str, date]] = {}
    for flag in plan.site_flags:
        if flag.is_void or flag.issue_id is None:
            continue
        requested = (flag.issue_no, flag.issue_date)
        if requested[0] is None or requested[1] is None:
            continue
        previous = requested_headers.setdefault(flag.issue_id, requested)
        if previous != requested:
            raise WorkbookError(
                "conflicting_issue_header",
                "同一领用单的多行包含不同单号或日期，本次未写入",
            )
    header_changed_issue_ids: set[str] = set()
    for issue_id, (issue_no, issue_date) in requested_headers.items():
        issue = db.get(MaintenanceSiteIssue, issue_id)
        if issue is None or issue.project_id != plan.project_id:
            raise WorkbookError("project_mismatch", "领用单归属已变化，请重新下载")
        if issue.issue_no != issue_no or issue.issue_date != issue_date:
            before_header = {
                "issue_no": issue.issue_no,
                "issue_date": issue.issue_date.isoformat(),
            }
            issue.issue_no = issue_no
            issue.issue_date = issue_date
            issue.version += 1
            header_changed_issue_ids.add(issue_id)
            _write_audit(
                db,
                project_id=plan.project_id,
                entity_type="site_issue",
                entity_id=issue_id,
                action="UPDATE",
                operated_by=operated_by,
                reason=audit_reason,
                before=before_header,
                after={
                    "issue_no": issue_no,
                    "issue_date": issue_date.isoformat(),
                },
            )
            operating_fact_changed = True

    pricing_entries: dict[str, tuple[date, MaintenanceSiteIssueLine]] = {}
    for flag in plan.site_flags:
        line = db.get(MaintenanceSiteIssueLine, flag.issue_line_id)
        # 2026-08-23：缺行=作废——领用行软作废，退出成本与返还义务计算；
        # 一张单全部行都作废时，单据状态同步置 void（与系统内作废同观感）
        if flag.is_void:
            if line is not None and line.is_active:
                line.is_active = False
                db.flush()
                issue_doc = db.get(MaintenanceSiteIssue, line.issue_id)
                if issue_doc is not None:
                    remaining = db.scalar(
                        select(func.count()).select_from(MaintenanceSiteIssueLine).where(
                            MaintenanceSiteIssueLine.issue_id == issue_doc.issue_id,
                            MaintenanceSiteIssueLine.is_active.is_(True)))
                    if not remaining:
                        issue_doc.normalized_status = "void"
                        issue_doc.version += 1
                        db.flush()
                _write_audit(db, project_id=plan.project_id,
                             entity_type="site_issue_line", entity_id=line.issue_line_id,
                             action="VOID", operated_by=operated_by, reason=audit_reason,
                             after={"issue_no": (db.get(MaintenanceSiteIssue, line.issue_id).issue_no
                                                 if db.get(MaintenanceSiteIssue, line.issue_id) else None)})
                operating_fact_changed = True
            continue
        if line is None and flag.is_create:
            issue = db.scalar(select(MaintenanceSiteIssue).where(
                MaintenanceSiteIssue.project_id == plan.project_id,
                MaintenanceSiteIssue.issue_no == flag.issue_no,
            ))
            if issue is None:
                issue = MaintenanceSiteIssue(
                    issue_id=str(uuid4()),
                    project_id=plan.project_id,
                    issue_no=flag.issue_no,
                    issue_date=flag.issue_date,
                    raw_status="已确认",
                    status_mapping_state="mapped",
                    normalized_status="confirmed",
                    status_mapping_version="workbook-manual-v1",
                    source="workbook",
                    import_batch_id=import_batch_id,
                    created_by=operated_by,
                    version=1,
                )
                db.add(issue)
                db.flush()
            line = MaintenanceSiteIssueLine(
                issue_line_id=flag.issue_line_id,
                issue_id=issue.issue_id,
                line_no=flag.line_no,
                part_id=flag.part_id,
                pn=flag.pn,
                quantity=flag.quantity,
                serial_number=flag.serial_number,
                remark=flag.remark,
                no_return=flag.no_return,
                is_active=True,
                algorithm_version="workbook-manual-v1",
            )
            db.add(line)
            pricing_entries[line.issue_line_id] = (issue.issue_date, line)
            operating_fact_changed = True
            continue
        if line is None:
            raise WorkbookError("line_not_found", f"领用行 {flag.issue_line_id} 已不存在，请重新下载")
        if line.no_return != flag.no_return:
            line.no_return = flag.no_return
            operating_fact_changed = True
        if flag.pn is not None and (
            flag.pn != line.pn
            or (flag.part_id is not None and flag.part_id != line.part_id)
        ):
            line.pn = flag.pn
            if flag.part_id is None:
                raise WorkbookError(
                    "part_not_found", "PN 已变化但未解析到备件主数据"
                )
            line.part_id = flag.part_id
            # A line-specific manual override is evidence for the old part and
            # cannot follow a PN identity change.
            line.manual_unit_cost = None
            line.manual_unit_cost_inc_tax = None
            line.manual_evidence = None
            issue_for_price = db.get(MaintenanceSiteIssue, line.issue_id)
            if issue_for_price is not None:
                pricing_entries[line.issue_line_id] = (
                    issue_for_price.issue_date,
                    line,
                )
            operating_fact_changed = True
        if (flag.serial_number is not None
                and flag.serial_number != line.serial_number):
            line.serial_number = flag.serial_number
            operating_fact_changed = True
        if flag.quantity is not None and flag.quantity != line.quantity:
            # 先把金额同步到新数量再等批量重取价——赋值后任何 autoflush 都会
            # 把 UPDATE 刷进库，若金额停留在旧数量口径会撞
            # ck_maintenance_site_issue_line_dual_tax_amounts（金额=数量×单价）。
            line.quantity = flag.quantity
            if line.unit_cost_ex_tax is not None:
                line.cost_amount_ex_tax = (
                    Decimal(line.quantity) * line.unit_cost_ex_tax
                ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                line.cost_amount_inc_tax = (
                    Decimal(line.quantity) * line.unit_cost_inc_tax
                ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                line.cost_amount = line.cost_amount_ex_tax
            qty_issue = db.get(MaintenanceSiteIssue, line.issue_id)
            if qty_issue is not None:
                pricing_entries[line.issue_line_id] = (qty_issue.issue_date, line)
            operating_fact_changed = True
        if flag.remark is not None and flag.remark != line.remark:
            line.remark = flag.remark
            operating_fact_changed = True
    if header_changed_issue_ids:
        for issue, line in db.execute(
            select(MaintenanceSiteIssue, MaintenanceSiteIssueLine)
            .join(
                MaintenanceSiteIssueLine,
                MaintenanceSiteIssueLine.issue_id == MaintenanceSiteIssue.issue_id,
            )
            .where(
                MaintenanceSiteIssue.issue_id.in_(
                    sorted(header_changed_issue_ids)
                ),
                MaintenanceSiteIssueLine.is_active.is_(True),
            )
            .order_by(
                MaintenanceSiteIssue.issue_id,
                MaintenanceSiteIssueLine.line_no,
            )
        ):
            pricing_entries[line.issue_line_id] = (issue.issue_date, line)
    if pricing_entries:
        from app.services import maintenance_consumption_cost as _consumption_cost
        _consumption_cost.resolve_lines(db, lines=list(pricing_entries.values()))
    if plan.expense_updates or receipt_ops:
        inner_result = ec.apply(
            db,
            ec.WorkbookPlan(plan.project_id, plan.expense_updates, receipt_ops),
            operated_by=operated_by,
            import_batch_id=import_batch_id,
            commit=False,
            workbook_state=workbook_state,
            bump_revision=False,
            track_change=True,
        )
        operating_fact_changed = (
            bool(inner_result.pop("_operating_fact_changed", False))
            or operating_fact_changed
        )
    # 04 作废（显式 VOID + 缺行=作废）：软删标记，读侧从此不导出（#264 契约）。
    for raw_line_id in plan.expense_voids:
        expense = db.scalar(select(FProjectExpense).where(FProjectExpense.raw_line_id == raw_line_id))
        if expense is None or expense.data_status == "已作废":
            continue
        expense.data_status = "已作废"
        attribution = db.get(
            MaintenanceProjectExpenseAttribution, f"bxd:{raw_line_id}")
        if attribution is None or attribution.project_id != plan.project_id:
            raise WorkbookError(
                "expense_not_in_project", "报销归因已变化，请重新下载")
        attribution.raw_status = "已作废"
        attribution.status_mapping_state = "mapped"
        attribution.normalized_status = "void"
        attribution.status_mapping_version = ec.PROTOCOL_VERSION
        attribution.version += 1
        _write_audit(db, project_id=plan.project_id,
                     entity_type="project_expense", entity_id=raw_line_id,
                     action="VOID", operated_by=operated_by, reason=audit_reason,
                     before={"data_status": None})
        operating_fact_changed = True
    milestone_entity_ids = {
        change.entity_id
        for change in plan.milestone_changes
        if change.entity_id is not None
    }
    locked_milestones = list(db.scalars(
        select(MaintenanceCollectionMilestone)
        .where(
            MaintenanceCollectionMilestone.milestone_id.in_(
                milestone_entity_ids
            )
        )
        .order_by(MaintenanceCollectionMilestone.milestone_id)
        .with_for_update()
    ).all()) if milestone_entity_ids else []
    if (
        {milestone.milestone_id for milestone in locked_milestones}
        != milestone_entity_ids
        or any(
            milestone.project_id != plan.project_id
            or not milestone.is_active
            for milestone in locked_milestones
        )
    ):
        raise WorkbookError(
            "milestone_not_found",
            "回款计划已不存在、已作废或不属于本项目，请重新下载",
        )
    milestones_by_id = {
        milestone.milestone_id: milestone for milestone in locked_milestones
    }
    create_coordinates: set[tuple[str, int]] = set()
    for change in plan.milestone_changes:
        contract = current_contract_by_no.get(change.contract_no)
        if contract is None:
            raise WorkbookError(
                "contract_not_found",
                f"合同 {change.contract_no} 已不是本项目当前合同，请重新下载",
            )
        existing = (
            milestones_by_id.get(change.entity_id)
            if change.entity_id is not None else None
        )
        if existing is not None:
            # The hidden ID is the identity.  Editable contract/sequence cells
            # are attributes only and may never redirect UPDATE/VOID to a
            # different row.
            if (
                existing.project_contract_id != contract.project_contract_id
                or existing.sequence != change.sequence
            ):
                raise WorkbookError(
                    "milestone_identity_mismatch",
                    "回款计划的合同或期次已变化，请重新下载",
                )
            if (
                change.base_version is not None
                and existing.version != change.base_version
            ):
                raise WorkbookError(
                    "stale_row",
                    f"合同 {change.contract_no} 第 {change.sequence} 期已被更新",
                )
        if change.operation == "VOID":
            if existing is None:
                raise WorkbookError(
                    "void_target_missing",
                    "作废行缺少可验证的回款计划实体ID",
                )
            existing.is_active = False
            existing.version += 1
            _write_audit(db, project_id=plan.project_id,
                         entity_type="collection_milestone",
                         entity_id=existing.milestone_id, action="VOID",
                         operated_by=operated_by, reason=audit_reason,
                         after={"contract_no": change.contract_no,
                                "sequence": change.sequence})
            operating_fact_changed = True
            continue
        if change.operation == "CREATE":
            coordinate = (contract.project_contract_id, change.sequence)
            if coordinate in create_coordinates:
                raise WorkbookError(
                    "milestone_target_conflict",
                    f"合同 {change.contract_no} 第 {change.sequence} 期在工作簿中重复新建",
                )
            create_coordinates.add(coordinate)
            coordinate_owner = db.scalar(
                select(MaintenanceCollectionMilestone)
                .where(
                    MaintenanceCollectionMilestone.project_contract_id
                    == contract.project_contract_id,
                    MaintenanceCollectionMilestone.sequence == change.sequence,
                )
                .with_for_update()
            )
            if coordinate_owner is not None:
                raise WorkbookError(
                    "milestone_target_conflict",
                    f"合同 {change.contract_no} 第 {change.sequence} 期已有回款计划，请重新下载",
                )
        elif existing is None:
            raise WorkbookError(
                "milestone_not_found",
                "更新行缺少可验证的回款计划实体ID",
            )
        milestone = write_collection_milestone(
            db, project_id=plan.project_id, project_contract_id=contract.project_contract_id,
            sequence=change.sequence, planned_date=change.planned_date,
            planned_amount=change.planned_amount, completeness_state=("complete" if change.planned_date and change.planned_amount else "date_only" if change.planned_date else "amount_only"),
            source="project_master_v2", date_precision=change.date_precision, operator=operated_by,
        )
        milestone.is_active = True
        operating_fact_changed = True
    if operating_fact_changed:
        # One successful workbook transaction invalidates every older V3/data
        # version exactly once, regardless of how many sheets it changed.
        operations.bump_locked_workbook_revision(db, state=workbook_state)
    if apply_operation_key is not None:
        db.add(MaintenanceProjectWorkbookOperation(
            project_id=plan.project_id,
            export_id=plan.export_id,
            file_sha256=plan.file_sha256,
            operation_key=apply_operation_key,
            payload_hash=plan.file_sha256,
            operation_type="file_apply",
            operated_by=operated_by,
        ))
    db.commit()
    return _v2_apply_result(
        plan,
        operated_by=operated_by,
        import_batch_id=import_batch_id,
        replayed=False,
        revision_drift=revision_drift,
    )
