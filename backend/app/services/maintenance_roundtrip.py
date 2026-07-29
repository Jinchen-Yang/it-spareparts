"""维保项目可编辑工作簿协议。

协议的公开边界只有两个：

* :func:`build_roundtrip_template` 导出一份带版本、行令牌和合同 revision 的固定 XLSX；
* :func:`import_roundtrip_workbook` 校验整本工作簿并在同一事务中应用显式操作。

工作簿不是数据库备份。缺行永远不表示删除；只有 ``VOID`` 才作废业务行。Excel
公式、隐藏技术列、客户端提供的含税值都不可信，服务端会重新校验和计算。
"""

from __future__ import annotations

import hashlib
import hmac
import json
import logging
import os
import posixpath
import re
import unicodedata
import uuid
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from tempfile import SpooledTemporaryFile
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy import Text, cast, distinct, func, or_, select, text
from sqlalchemy.orm import Session

from app import config, tax_policy
from app.business_time import business_today
from app.config import get_settings
from app.etl import cleaner, pipeline, reader
from app.models import maintenance as maintenance_models
from app.models.dimensions import DimCustomer
from app.models.maintenance import FMaintenanceLine, FMaintenanceOrder, FProjectExpense
from app.models.maintenance import MaintenanceRoundtripOperation
from app.models.system import SysAuditLog, SysImportBatch, SysRawFile
from app.services import maintenance_cost
from app.services.maintenance_workbook_export import (
    MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK as STANDARD_MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK,
)

_log = logging.getLogger(__name__)

PROTOCOL_ID = reader.ROUNDTRIP_PROTOCOL_ID
SCHEMA_VERSION = "1.0"
APP_VERSION = "1.20.0"
ROUNDTRIP_FILE_TYPE = "maint_roundtrip"
TAX_RATE = tax_policy.TAX_RATE
SHEET_NAMES = (
    "00_使用说明",
    "01_项目",
    "02_维保订单",
    "03_订单明细",
    "04_报销明细",
    "05_人工成本回填",
    "98_字典",
    "99_合同版本",
    "99_元数据",
)

MAX_ROWS_PER_TABLE = 10_000
MAX_CONTRACT_REVISIONS = 100_000
MAX_OPERATIONS_PER_WORKBOOK = 100_000
MAX_METADATA_ROWS = 128
MAX_CELL_CHARS = 32_767
MAX_WORKBOOK_BYTES = 256 * 1024 * 1024
MAX_ROUNDTRIP_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK = STANDARD_MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK
MAX_ROUNDTRIP_BUNDLE_CONTRACTS = 500
MAX_ROUNDTRIP_BUNDLE_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_ROUNDTRIP_BUNDLE_MEMBER_BYTES = 240
BLANK_CREATE_ROWS = 50
ROUNDTRIP_EXPORT_ADVISORY_LOCK_KEY = 0x5254_584C  # "RTXL"
_ORDER_RENDERED_TEXT_OVERHEAD_BYTES = 128
_LINE_RENDERED_TEXT_OVERHEAD_BYTES = 128
_EXPENSE_RENDERED_TEXT_OVERHEAD_BYTES = 128
_MANUAL_RENDERED_TEXT_OVERHEAD_BYTES = 128
_BLANK_RENDERED_TEXT_OVERHEAD_BYTES = BLANK_CREATE_ROWS * 64 * 2

_INVALID_BUNDLE_MEMBER_CHARS = re.compile(r'[\x00-\x1f\x7f/\\:*?"<>|]+')
_INVALID_XLSX_TEXT_CONTROLS = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]"
)

_TECH_HEADERS = (
    "操作",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
_PROJECT_HEADERS = (
    "操作",
    "合同号",
    "项目名称",
    "维保开始",
    "维保结束",
    "维保订单数",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
_ORDER_HEADERS = (
    "操作",
    "维保单号",
    "制单日期",
    "合同号",
    "项目名称",
    "维保开始",
    "维保结束",
    "客户名称",
    "最终客户",
    "需求类型",
    "业务类型",
    "销售人员",
    "出库仓库",
    "数据状态",
    "变更原因",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
_LINE_HEADERS = (
    "操作",
    "维保单号",
    "序号",
    "PN",
    "产品描述",
    "需求数量",
    "退货数量",
    "发货SN",
    "成本来源",
    "未税单位成本",
    "含税单位成本",
    "变更原因",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
_EXPENSE_HEADERS = (
    "操作",
    "合同号",
    "报销日期",
    "报销人员",
    "报销类别",
    "费用分类",
    "支出事由",
    "未税金额",
    "含税金额(系统计算)",
    "流程状态",
    "单号",
    "序号",
    "变更原因",
    "__tax_basis",
    "__raw_amount",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
_MANUAL_HEADERS = (
    "操作",
    "维保明细ID",
    "维保单号",
    "PN",
    "需求数量",
    "当前成本来源",
    "人工未税单位成本",
    "含税单位成本(系统计算)",
    "税率",
    "回填原因",
    "依据说明",
    "状态",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
_TABLE_SPECS = {
    "01_项目": ("tbl_projects_v1", _PROJECT_HEADERS),
    "02_维保订单": ("tbl_orders_v1", _ORDER_HEADERS),
    "03_订单明细": ("tbl_order_lines_v1", _LINE_HEADERS),
    "04_报销明细": ("tbl_expenses_v1", _EXPENSE_HEADERS),
    "05_人工成本回填": ("tbl_manual_costs_v1", _MANUAL_HEADERS),
}
MAX_ROUNDTRIP_ROW_ELEMENTS = (
    19
    + len(_TABLE_SPECS) * (MAX_ROWS_PER_TABLE + 1)
    + 7
    + (MAX_CONTRACT_REVISIONS + 1)
    + (MAX_METADATA_ROWS + 1)
)
MAX_ROUNDTRIP_CELL_ELEMENTS = (
    2 * 19
    + sum(
        len(headers) * (MAX_ROWS_PER_TABLE + 1)
        for _table_name, headers in _TABLE_SPECS.values()
    )
    + 3 * 7
    + 2 * (MAX_CONTRACT_REVISIONS + 1)
    + 2 * (MAX_METADATA_ROWS + 1)
)
MAX_ROUNDTRIP_SHARED_STRINGS = MAX_ROUNDTRIP_CELL_ELEMENTS
MAX_ROUNDTRIP_XML_ELEMENTS_PER_PART = 4 * MAX_ROUNDTRIP_SHARED_STRINGS
MAX_ROUNDTRIP_CONTENT_TYPE_ELEMENTS = config.IMPORT_XLSX_MAX_MEMBERS + 32
MAX_ROUNDTRIP_RELATIONSHIP_ELEMENTS = 256
MAX_ROUNDTRIP_TABLE_ELEMENTS = 512
MAX_ROUNDTRIP_WORKSHEET_STRUCTURAL_ELEMENTS = 10_000
MAX_ROUNDTRIP_CELL_XML_ELEMENTS = 128
MAX_ROUNDTRIP_SHARED_STRING_XML_ELEMENTS = 128

_HEADER_FILL = PatternFill("solid", fgColor="35506B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SYSTEM_FILL = PatternFill("solid", fgColor="E8EEF3")
_EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
_REQUIRED_FILL = PatternFill("solid", fgColor="FCE4D6")
_ERROR_FILL = PatternFill("solid", fgColor="F4CCCC")
_TITLE_FILL = PatternFill("solid", fgColor="DDEBF7")
_THIN = Side(style="thin", color="D9E1F2")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(vertical="top", wrap_text=True)
_MONEY = "#,##0.00"
_DATE = "yyyy-mm-dd"
_ROW_HEIGHTS = (24, 36, 48, 72)

_EDITABLE_COLUMNS = {
    "02_维保订单": {
        "操作",
        "制单日期",
        "合同号",
        "项目名称",
        "维保开始",
        "维保结束",
        "客户名称",
        "最终客户",
        "需求类型",
        "业务类型",
        "销售人员",
        "出库仓库",
        "数据状态",
        "变更原因",
    },
    "03_订单明细": {
        "操作",
        "产品描述",
        "需求数量",
        "退货数量",
        "发货SN",
        "变更原因",
    },
    "04_报销明细": {
        "操作",
        "合同号",
        "报销日期",
        "报销人员",
        "报销类别",
        "费用分类",
        "支出事由",
        "未税金额",
        "流程状态",
        "单号",
        "序号",
        "变更原因",
    },
    "05_人工成本回填": {
        "操作",
        "人工未税单位成本",
        "回填原因",
        "依据说明",
    },
}
_REQUIRED_COLUMNS = {
    "02_维保订单": {"操作", "变更原因"},
    "03_订单明细": {"操作", "变更原因"},
    "04_报销明细": {
        "操作",
        "合同号",
        "报销日期",
        "未税金额",
        "变更原因",
    },
    "05_人工成本回填": {
        "操作",
        "人工未税单位成本",
        "回填原因",
        "依据说明",
    },
}
_ALLOWED_OPERATIONS = {"", "KEEP", "CREATE", "UPDATE", "VOID"}
_OPERATION_CHOICES_BY_SHEET = {
    "02_维保订单": ("KEEP", "UPDATE"),
    "03_订单明细": ("KEEP", "UPDATE"),
    "04_报销明细": ("KEEP", "CREATE", "UPDATE", "VOID"),
    "05_人工成本回填": ("KEEP", "CREATE", "UPDATE", "VOID"),
}
_CONTRACT_REVISION_TABLE = "tbl_contract_revisions_v1"
_CONTRACT_REVISION_HEADERS = ("合同号", "revision")

_OPERATION_PAYLOAD_HEADERS: dict[tuple[str, str], tuple[str, ...]] = {
    ("02_维保订单", "UPDATE"): (
        "制单日期",
        "合同号",
        "项目名称",
        "维保开始",
        "维保结束",
        "客户名称",
        "最终客户",
        "需求类型",
        "业务类型",
        "销售人员",
        "出库仓库",
        "数据状态",
        "变更原因",
    ),
    ("03_订单明细", "UPDATE"): (
        "产品描述",
        "需求数量",
        "退货数量",
        "发货SN",
        "变更原因",
    ),
    ("04_报销明细", "CREATE"): (
        "合同号",
        "报销日期",
        "报销人员",
        "报销类别",
        "费用分类",
        "支出事由",
        "未税金额",
        "流程状态",
        "单号",
        "序号",
        "变更原因",
        "__tax_basis",
        "__raw_amount",
    ),
    ("04_报销明细", "UPDATE"): (
        "合同号",
        "报销日期",
        "报销人员",
        "报销类别",
        "费用分类",
        "支出事由",
        "未税金额",
        "流程状态",
        "单号",
        "序号",
        "变更原因",
        "__tax_basis",
        "__raw_amount",
    ),
    ("04_报销明细", "VOID"): ("变更原因",),
    ("05_人工成本回填", "CREATE"): (
        "维保明细ID",
        "人工未税单位成本",
        "回填原因",
        "依据说明",
    ),
    ("05_人工成本回填", "UPDATE"): (
        "维保明细ID",
        "人工未税单位成本",
        "回填原因",
        "依据说明",
    ),
    ("05_人工成本回填", "VOID"): ("回填原因",),
}


class RoundtripWorkbookError(ValueError):
    """可安全返回给客户端的整本工作簿拒绝原因。"""

    def __init__(self, detail: str, *, status_code: int = 422):
        self.status_code = status_code
        super().__init__(detail)


@dataclass(frozen=True, slots=True)
class _ParsedRow:
    sheet: str
    excel_row: int
    values: dict[str, Any]

    @property
    def operation(self) -> str:
        return str(self.values.get("操作") or "").strip().upper()


@dataclass(frozen=True, slots=True)
class _Change:
    kind: str
    operation: str
    row: _ParsedRow
    entity: Any | None
    values: dict[str, Any]
    contracts: frozenset[str]
    scope_dates: tuple[date | None, ...] = ()


@dataclass(frozen=True, slots=True)
class _OperationIntent:
    row: _ParsedRow
    client_row_id: str
    payload_hash: str


@dataclass(frozen=True, slots=True)
class _AppliedChange:
    """Prepared mutation whose audit can be emitted after one batch flush.

    CREATE entities do not have primary keys until SQLAlchemy flushes them. Keeping
    audit metadata beside the entity lets the importer flush every business entity
    once, then batch-insert audits and operation-ledger rows without per-row SQL.
    """

    change: _Change
    entity: Any
    entity_type: str
    audit_action: str
    audit_fields: tuple[str, ...]
    before: dict[str, Any] | None
    reason: str | None


def _manual_model():
    model = getattr(maintenance_models, "MaintenanceManualCostOverride", None)
    if model is None:
        raise RuntimeError("MaintenanceManualCostOverride model is not installed")
    return model


def _state_model():
    model = getattr(maintenance_models, "MaintenanceContractWorkbookState", None)
    if model is None:
        raise RuntimeError("MaintenanceContractWorkbookState model is not installed")
    return model


def _safe(value: Any) -> Any:
    if isinstance(value, str):
        if _INVALID_XLSX_TEXT_CONTROLS.search(value):
            raise RoundtripWorkbookError("文本包含 XLSX 不允许的控制字符")
        if len(value) > MAX_CELL_CHARS:
            raise RoundtripWorkbookError(
                f"文本超过 Excel 单元格上限 {MAX_CELL_CHARS} 个字符",
                status_code=413,
            )
    return value


def _append_literal_row(ws, values: Iterable[Any]) -> None:
    """Append data while forcing every string to OpenXML text, never a formula."""
    ws.append([_safe(value) for value in values])
    row = ws.max_row
    for column in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=column)
        if isinstance(cell.value, str):
            cell.data_type = "s"


def _iso(value: date | datetime | None) -> str:
    return value.isoformat() if value is not None else ""


def _decimal(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _evidence_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dict):
        note = value.get("note")
        if set(value) == {"note"} and isinstance(note, str):
            return note
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def _base_version(entity: Any) -> int:
    value = getattr(entity, "version", None)
    if value is None:
        value = getattr(entity, "import_batch_id", None)
    return int(value or 0)


def _signing_secret() -> bytes:
    return get_settings().secret_key.encode("utf-8")


def _row_token(export_id: str, kind: str, entity_id: int, base_version: int) -> str:
    payload = (
        f"{PROTOCOL_ID}|{SCHEMA_VERSION}|{export_id}|{kind}|{entity_id}|{base_version}"
    ).encode("utf-8")
    return hmac.new(_signing_secret(), payload, hashlib.sha256).hexdigest()


def _export_client_row_id(
    export_id: str,
    sheet: str,
    kind: str,
    entity_id: int,
    base_version: int,
) -> str:
    """Stable operation id for an exported existing row.

    UUID5 avoids random-state persistence while remaining stable across
    Excel/openpyxl saves of this export.
    """
    return str(
        uuid.uuid5(
            uuid.NAMESPACE_URL,
            (
                f"{PROTOCOL_ID}|{SCHEMA_VERSION}|{export_id}|{sheet}|{kind}|"
                f"{entity_id}|{base_version}"
            ),
        )
    )


def _expense_create_raw_line_id(export_id: str, client_row_id: str) -> str:
    """Stable business key aligned with the operation-ledger identity."""
    identity = uuid.uuid5(
        uuid.NAMESPACE_URL,
        (
            f"{PROTOCOL_ID}|{SCHEMA_VERSION}|{export_id}|"
            f"04_报销明细|{client_row_id}"
        ),
    )
    return f"RTEXP:{identity}"


def _contract_revisions_payload(revisions: dict[str, int]) -> bytes:
    return json.dumps(
        revisions,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _contract_revisions_digest(revisions: dict[str, int]) -> str:
    return hashlib.sha256(_contract_revisions_payload(revisions)).hexdigest()


def _metadata_payload(metadata: dict[str, str]) -> bytes:
    unsigned = {
        key: metadata[key] for key in sorted(metadata) if key != "metadata_hmac"
    }
    return json.dumps(
        unsigned,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")


def _metadata_hmac(metadata: dict[str, str]) -> str:
    return hmac.new(
        _signing_secret(),
        _metadata_payload(metadata),
        hashlib.sha256,
    ).hexdigest()


def _contract_state_revisions(db: Session, contracts: Iterable[str]) -> dict[str, int]:
    values = sorted(
        {contract.strip() for contract in contracts if contract and contract.strip()}
    )
    if not values:
        return {}
    model = getattr(maintenance_models, "MaintenanceContractWorkbookState", None)
    if model is None:
        return {contract: 0 for contract in values}
    rows = db.execute(
        select(model.contract_no, model.revision).where(model.contract_no.in_(values))
    ).all()
    revisions = {contract: int(revision) for contract, revision in rows}
    return {contract: revisions.get(contract, 0) for contract in values}


def _lock_shared_snapshot(db: Session) -> None:
    db.execute(
        text("SELECT pg_advisory_xact_lock_shared(:key)"),
        {"key": config.DATA_CHANGE_ADVISORY_LOCK_KEY},
    )


def _lock_single_template_builder(db: Session) -> None:
    acquired = db.scalar(
        text("SELECT pg_try_advisory_xact_lock(:key)"),
        {"key": ROUNDTRIP_EXPORT_ADVISORY_LOCK_KEY},
    )
    if acquired is not True:
        raise RoundtripWorkbookError(
            "已有维保回填模板正在生成，请稍后重试",
            status_code=429,
        )


def _selection_filters(
    *,
    contract: str | None,
    date_from: date | None,
    date_to: date | None,
) -> list:
    filters = []
    if contract:
        filters.append(FMaintenanceOrder.linked_sales_order_no == contract.strip())
    if date_from is not None:
        filters.append(FMaintenanceOrder.order_date >= date_from)
    if date_to is not None:
        filters.append(FMaintenanceOrder.order_date <= date_to)
    return filters


def _validate_date_pair(date_from: date | None, date_to: date | None) -> None:
    if (date_from is None) != (date_to is None):
        raise RoundtripWorkbookError("date_from 与 date_to 必须同时提供")
    if date_from is not None and date_to is not None and date_from > date_to:
        raise RoundtripWorkbookError("date_from 不能晚于 date_to")


def _octet_length(value):
    """在 PostgreSQL 端统计固定模板将输出的 UTF-8 文本字节。"""
    return func.octet_length(func.coalesce(value, ""))


def _selected_dynamic_text_bytes(
    db: Session,
    *,
    order_filters: list,
    expense_filters: list,
) -> int:
    """在 ORM 行物化前，以一个标量查询预估当前工作簿的动态文本总量。

    项目页按订单行保守重复计入合同号和项目名，避免为精确去重再增加一次数据库
    往返。其余表达式与固定模板真实输出列保持一致；行令牌、客户端行 ID 和预留
    新增行使用固定开销计入。
    """
    rendered_project = func.coalesce(
        func.nullif(FMaintenanceOrder.project_raw, ""),
        FMaintenanceOrder.project_std,
        "",
    )
    order_text_bytes = (
        _octet_length(FMaintenanceOrder.order_no)
        # 订单页各一次，项目页的合同号/项目名及 entity id 各再保守计两次。
        + _octet_length(FMaintenanceOrder.linked_sales_order_no) * 3
        + _octet_length(rendered_project) * 3
        + _octet_length(DimCustomer.name_raw)
        + _octet_length(FMaintenanceOrder.end_customer)
        + _octet_length(FMaintenanceOrder.demand_type)
        + _octet_length(FMaintenanceOrder.business_type)
        + _octet_length(FMaintenanceOrder.salesperson)
        + _octet_length(FMaintenanceOrder.warehouse)
        + _octet_length(FMaintenanceOrder.data_status)
        + _ORDER_RENDERED_TEXT_OVERHEAD_BYTES
    )
    order_total = (
        select(func.coalesce(func.sum(order_text_bytes), 0))
        .select_from(FMaintenanceOrder)
        .outerjoin(DimCustomer, DimCustomer.id == FMaintenanceOrder.customer_id)
        .where(*order_filters)
        .scalar_subquery()
    )

    line_text_bytes = (
        _octet_length(FMaintenanceOrder.order_no)
        + _octet_length(FMaintenanceLine.pn_std)
        + _octet_length(FMaintenanceLine.description)
        + _octet_length(FMaintenanceLine.serial_numbers)
        + _octet_length(FMaintenanceLine.cost_source)
        + _LINE_RENDERED_TEXT_OVERHEAD_BYTES
    )
    line_total = (
        select(func.coalesce(func.sum(line_text_bytes), 0))
        .select_from(FMaintenanceLine)
        .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
        .where(*order_filters)
        .scalar_subquery()
    )

    expense_text_bytes = (
        _octet_length(FProjectExpense.linked_sales_order_no)
        + _octet_length(FProjectExpense.person)
        + _octet_length(FProjectExpense.expense_type)
        + _octet_length(FProjectExpense.fee_category)
        + _octet_length(FProjectExpense.reason)
        + _octet_length(FProjectExpense.data_status)
        + _octet_length(FProjectExpense.bxd_no)
        + _octet_length(FProjectExpense.tax_basis)
        + _EXPENSE_RENDERED_TEXT_OVERHEAD_BYTES
    )
    expense_total = (
        select(func.coalesce(func.sum(expense_text_bytes), 0))
        .select_from(FProjectExpense)
        .where(*expense_filters)
        .scalar_subquery()
    )

    model = getattr(maintenance_models, "MaintenanceManualCostOverride", None)
    if model is None:
        manual_text_bytes = (
            _octet_length(FMaintenanceOrder.order_no)
            + _octet_length(FMaintenanceLine.pn_std)
            + _octet_length(FMaintenanceLine.cost_source)
            + _MANUAL_RENDERED_TEXT_OVERHEAD_BYTES
        )
        manual_total = (
            select(func.coalesce(func.sum(manual_text_bytes), 0))
            .select_from(FMaintenanceLine)
            .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
            .where(*order_filters, FMaintenanceLine.cost_source == "none")
            .scalar_subquery()
        )
    else:
        manual_text_bytes = (
            _octet_length(FMaintenanceOrder.order_no)
            + _octet_length(FMaintenanceLine.pn_std)
            + _octet_length(FMaintenanceLine.cost_source)
            + _octet_length(model.reason)
            # JSONB::text is equal to or more conservative than the displayed
            # ensure_ascii=False JSON/note representation.
            + _octet_length(cast(model.evidence, Text))
            + _MANUAL_RENDERED_TEXT_OVERHEAD_BYTES
        )
        manual_total = (
            select(func.coalesce(func.sum(manual_text_bytes), 0))
            .select_from(FMaintenanceLine)
            .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
            .outerjoin(model, model.line_id == FMaintenanceLine.id)
            .where(
                *order_filters,
                or_(model.id.is_not(None), FMaintenanceLine.cost_source == "none"),
            )
            .scalar_subquery()
        )

    total = db.scalar(
        select(
            order_total
            + line_total
            + expense_total
            + manual_total
            + _BLANK_RENDERED_TEXT_OVERHEAD_BYTES
        )
    )
    return int(total or 0)


def _selected_data(
    db: Session,
    *,
    contract: str | None,
    date_from: date | None,
    date_to: date | None,
    blank: bool,
) -> dict[str, Any]:
    if blank:
        contracts = {contract.strip()} if contract and contract.strip() else set()
        return {
            "orders": [],
            "lines": [],
            "expenses": [],
            "manual": [],
            "contracts": contracts,
        }

    filters = _selection_filters(
        contract=contract,
        date_from=date_from,
        date_to=date_to,
    )
    if contract and contract.strip():
        contract_value = contract.strip()
        contract_exists = bool(
            db.scalar(
                select(func.count(FMaintenanceOrder.id)).where(
                    FMaintenanceOrder.linked_sales_order_no == contract_value,
                )
            )
        )
        if not contract_exists:
            raise RoundtripWorkbookError(
                f"合同不存在：{contract_value}",
                status_code=404,
            )
    order_count = int(
        db.scalar(select(func.count(FMaintenanceOrder.id)).where(*filters)) or 0
    )
    if order_count == 0:
        raise RoundtripWorkbookError(
            (
                "合同存在，但所选范围内没有可导出的维保数据"
                if contract and contract.strip()
                else "所选范围内没有可导出的维保数据"
            ),
        )
    if order_count > MAX_ROWS_PER_TABLE:
        raise RoundtripWorkbookError(
            f"02_维保订单命中 {order_count} 行，超过 {MAX_ROWS_PER_TABLE} 行导出上限；"
            "请按合同或日期缩小范围",
            status_code=413,
        )

    line_count = int(
        db.scalar(
            select(func.count(FMaintenanceLine.id))
            .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
            .where(*filters)
        )
        or 0
    )
    if line_count > MAX_ROWS_PER_TABLE:
        raise RoundtripWorkbookError(
            f"03_订单明细命中 {line_count} 行，超过 {MAX_ROWS_PER_TABLE} 行导出上限；"
            "请使用单合同“回填模板”或缩小日期范围",
            status_code=413,
        )

    contracts = {
        str(value).strip()
        for value in db.scalars(
            select(distinct(FMaintenanceOrder.linked_sales_order_no)).where(
                *filters,
                FMaintenanceOrder.linked_sales_order_no.is_not(None),
            )
        ).all()
        if value and str(value).strip()
    }
    if contract and contract.strip():
        contracts.add(contract.strip())

    expense_filters = []
    if contracts:
        expense_filters.append(FProjectExpense.linked_sales_order_no.in_(contracts))
    else:
        expense_filters.append(text("false"))
    if date_from is not None:
        expense_filters.append(FProjectExpense.expense_date >= date_from)
    if date_to is not None:
        expense_filters.append(FProjectExpense.expense_date <= date_to)
    expense_count = int(
        db.scalar(select(func.count(FProjectExpense.id)).where(*expense_filters)) or 0
    )
    expense_row_cap = MAX_ROWS_PER_TABLE - BLANK_CREATE_ROWS
    if expense_count > expense_row_cap:
        raise RoundtripWorkbookError(
            f"04_报销明细命中 {expense_count} 行，连同预留新增行将超过 "
            f"{MAX_ROWS_PER_TABLE} 行导出上限；请按合同或日期缩小范围",
            status_code=413,
        )

    dynamic_text_bytes = _selected_dynamic_text_bytes(
        db,
        order_filters=filters,
        expense_filters=expense_filters,
    )
    if dynamic_text_bytes > MAX_DYNAMIC_TEXT_BYTES_PER_WORKBOOK:
        raise RoundtripWorkbookError(
            "单个维保回填工作簿动态文本超过 64 MiB 安全上限；"
            "请按合同或日期缩小范围",
            status_code=413,
        )

    # 上述 COUNT、轻量 DISTINCT 和字节聚合都在任何 ORM 行全量物化/openpyxl 构建前完成。
    orders = db.execute(
        select(FMaintenanceOrder, DimCustomer.name_raw)
        .outerjoin(DimCustomer, DimCustomer.id == FMaintenanceOrder.customer_id)
        .where(*filters)
        .order_by(
            FMaintenanceOrder.order_date.asc().nullslast(),
            FMaintenanceOrder.order_no,
            FMaintenanceOrder.id,
        )
    ).all()
    order_ids = [order.id for order, _customer in orders]
    lines = []
    if order_ids:
        lines = db.execute(
            select(FMaintenanceLine, FMaintenanceOrder)
            .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
            .where(FMaintenanceLine.order_id.in_(order_ids))
            .order_by(
                FMaintenanceOrder.order_date.asc().nullslast(),
                FMaintenanceOrder.order_no,
                FMaintenanceLine.line_no.asc().nullslast(),
                FMaintenanceLine.id,
            )
        ).all()

    expenses = db.scalars(
        select(FProjectExpense)
        .where(*expense_filters)
        .order_by(
            FProjectExpense.linked_sales_order_no,
            FProjectExpense.expense_date.asc().nullslast(),
            FProjectExpense.id,
        )
    ).all()

    manual_rows = []
    model = getattr(maintenance_models, "MaintenanceManualCostOverride", None)
    if model is not None and order_ids:
        manual_rows = db.execute(
            select(model, FMaintenanceLine, FMaintenanceOrder)
            .join(FMaintenanceLine, FMaintenanceLine.id == model.line_id)
            .join(FMaintenanceOrder, FMaintenanceOrder.id == FMaintenanceLine.order_id)
            .where(FMaintenanceLine.order_id.in_(order_ids))
            .order_by(FMaintenanceOrder.order_no, FMaintenanceLine.line_no, model.id)
        ).all()

    return {
        "orders": orders,
        "lines": lines,
        "expenses": expenses,
        "manual": manual_rows,
        "contracts": contracts,
    }


def _table_style(table: Table) -> None:
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )


def _configure_print_layout(
    ws,
    *,
    landscape: bool,
    repeat_header: bool = False,
) -> None:
    """Keep wide operational sheets readable when staff print or save as PDF."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.4,
        bottom=0.4,
        header=0.2,
        footer=0.2,
    )
    if repeat_header:
        ws.print_title_rows = "1:1"


def _display_width(value: Any) -> int:
    width = 0
    for character in str(value):
        if character == "\t":
            width += 4
        elif unicodedata.combining(character):
            continue
        elif unicodedata.east_asian_width(character) in {"F", "W"}:
            width += 2
        else:
            width += 1
    return width


def _wrapped_line_count(value: Any, column_width: float | None) -> int:
    if value is None:
        return 1
    capacity = max(1, int(column_width or 16))
    return max(
        1,
        sum(
            max(1, (_display_width(part) + capacity - 1) // capacity)
            for part in str(value).replace("\r\n", "\n").replace("\r", "\n").split("\n")
        ),
    )


def _visible_row_height(ws, row: int, headers: tuple[str, ...]) -> int:
    wrapped_lines = 1
    for column, header in enumerate(headers, 1):
        column_letter = get_column_letter(column)
        if header.startswith("__") or ws.column_dimensions[column_letter].hidden:
            continue
        wrapped_lines = max(
            wrapped_lines,
            _wrapped_line_count(
                ws.cell(row=row, column=column).value,
                ws.column_dimensions[column_letter].width,
            ),
        )
    return _ROW_HEIGHTS[min(wrapped_lines, len(_ROW_HEIGHTS)) - 1]


def _append_table(
    ws,
    *,
    table_name: str,
    headers: tuple[str, ...],
    rows: list[list[Any]],
    editable: set[str] | None = None,
    required: set[str] | None = None,
) -> None:
    editable = editable or set()
    required = required or set()
    if len(rows) > MAX_ROWS_PER_TABLE:
        raise RoundtripWorkbookError(
            f"{ws.title} 超过 {MAX_ROWS_PER_TABLE} 行导出安全上限",
            status_code=413,
        )
    _append_literal_row(ws, headers)
    if not rows:
        rows = [[None] * len(headers)]
    for values in rows:
        _append_literal_row(ws, values)

    table = Table(
        displayName=table_name,
        ref=f"A1:{get_column_letter(len(headers))}{len(rows) + 1}",
    )
    _table_style(table)
    ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.auto_filter.ref = table.ref
    ws.row_dimensions[1].height = 38
    _configure_print_layout(ws, landscape=True, repeat_header=True)
    ws.print_area = table.ref

    for column, header in enumerate(headers, 1):
        header_cell = ws.cell(row=1, column=column)
        header_cell.fill = _HEADER_FILL
        header_cell.font = _HEADER_FONT
        header_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        header_cell.border = _BORDER
        width = 16
        if header in {
            "项目名称",
            "产品描述",
            "支出事由",
            "依据说明",
            "回填原因",
            "变更原因",
        }:
            width = 28
        elif "单位成本" in header or "系统计算" in header:
            width = 20
        elif header.startswith("__"):
            width = 4
            ws.column_dimensions[get_column_letter(column)].hidden = True
        elif header in {"操作", "序号", "税率", "状态"}:
            width = 11
        ws.column_dimensions[get_column_letter(column)].width = width

        for row in range(2, len(rows) + 2):
            cell = ws.cell(row=row, column=column)
            cell.border = _BORDER
            cell.alignment = _CENTER if header.startswith("__") else _WRAP
            if header in editable:
                cell.fill = _REQUIRED_FILL if header in required else _EDIT_FILL
                cell.protection = Protection(locked=False)
            else:
                cell.fill = _SYSTEM_FILL
                cell.protection = Protection(locked=True)
            if "日期" in header or header in {"维保开始", "维保结束"}:
                cell.number_format = _DATE
            if "金额" in header or "单位成本" in header:
                cell.number_format = _MONEY

    for row in range(2, len(rows) + 2):
        # Visible business text may wrap to several lines. Hidden HMAC/UUID
        # columns are deliberately excluded so they cannot inflate every row.
        ws.row_dimensions[row].height = _visible_row_height(ws, row, headers)

    operation_choices = _OPERATION_CHOICES_BY_SHEET.get(ws.title)
    if "操作" in headers and operation_choices:
        operation_column = get_column_letter(headers.index("操作") + 1)
        choice_text = ",".join(operation_choices)
        choice_prompt = (
            f"{'、'.join(operation_choices[:-1])} 或 {operation_choices[-1]}"
        )
        validation = DataValidation(
            type="list",
            formula1=f'"{choice_text}"',
            allow_blank=True,
        )
        validation.error = f"请选择 {choice_prompt}"
        validation.errorTitle = "操作无效"
        validation.showErrorMessage = True
        ws.add_data_validation(validation)
        validation.add(f"{operation_column}2:{operation_column}{len(rows) + 1}")

    ws.conditional_formatting.add(
        f"A2:{get_column_letter(len(headers))}{len(rows) + 1}",
        FormulaRule(formula=['$A2="VOID"'], fill=_ERROR_FILL),
    )
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.insertRows = False


def _instructions_sheet(workbook: Workbook, metadata: dict[str, str]) -> None:
    ws = workbook.active
    ws.title = "00_使用说明"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:B1")
    ws["A1"] = "维保项目 · 固定回填工作簿"
    ws["A1"].font = Font(size=18, bold=True, color="1F4E78")
    ws["A1"].fill = _TITLE_FILL
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    contract_scope = metadata["contract_scope"]
    date_from = metadata["date_from"]
    date_to = metadata["date_to"]
    scope_items = [
        (
            "签名合同范围",
            f"指定合同：{contract_scope}" if contract_scope else "全部合同",
        ),
        (
            "签名日期范围",
            (
                f"指定日期闭区间：{date_from} 至 {date_to}（含边界）"
                if date_from and date_to
                else "全部日期"
            ),
        ),
        (
            "模板模式",
            "空白" if metadata["template_mode"] == "blank" else "快照",
        ),
        ("数据截止日", metadata["as_of"]),
        ("导出时间", metadata["exported_at"]),
    ]
    notes = [
        ("黄色单元格", "允许填写；蓝灰单元格由系统维护，请勿修改隐藏技术列。"),
        (
            "桃色单元格",
            "按所选操作必填：CREATE / UPDATE 填对应业务字段；"
            "VOID 时，报销必须填变更原因，人工成本必须填回填原因。",
        ),
        ("KEEP / 留空", "不写数据库；即使删除整行，也不会被解释为删除。"),
        ("CREATE", "新增报销或人工成本；订单、订单明细在 1.0 版暂不允许新增。"),
        ("UPDATE", "更新现有行；系统会校验导出时版本，旧工作簿不会覆盖新数据。"),
        ("VOID", "软作废并保留审计；不物理删除业务事实。"),
        ("税务口径", "报销和人工成本只填写未税金额，系统固定按 13% 重算含税值。"),
        (
            "完整快照",
            "只有不带 date_from/date_to 的全量范围模板导入成功后，才会把合同报销标记为"
            "完整快照；带日期的范围模板和 blank=true 空白模板绝不会声明全量完整。",
        ),
        ("错误处理", "任一行校验失败或版本冲突，整本工作簿零写入。"),
    ]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 88
    ws.merge_cells("A3:B3")
    ws["A3"] = "本工作簿签名范围"
    ws["A3"].font = Font(bold=True, color="1F4E78")
    ws["A3"].fill = _TITLE_FILL
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 28
    for row, (title, content) in enumerate(scope_items, 4):
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=content)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = _SYSTEM_FILL
        ws.cell(row=row, column=2).fill = _TITLE_FILL
        for column in (1, 2):
            ws.cell(row=row, column=column).border = _BORDER
            ws.cell(row=row, column=column).alignment = _WRAP
        ws.row_dimensions[row].height = 28

    rules_heading_row = len(scope_items) + 5
    ws.merge_cells(
        start_row=rules_heading_row,
        start_column=1,
        end_row=rules_heading_row,
        end_column=2,
    )
    ws.cell(row=rules_heading_row, column=1, value="填写与回填规则")
    ws.cell(row=rules_heading_row, column=1).font = Font(bold=True, color="1F4E78")
    ws.cell(row=rules_heading_row, column=1).fill = _TITLE_FILL
    ws.cell(row=rules_heading_row, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[rules_heading_row].height = 28
    notes_start_row = rules_heading_row + 1
    for row, (title, content) in enumerate(notes, notes_start_row):
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=content)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = {
            "黄色单元格": _EDIT_FILL,
            "桃色单元格": _REQUIRED_FILL,
        }.get(title, _SYSTEM_FILL)
        for column in (1, 2):
            ws.cell(row=row, column=column).border = _BORDER
            ws.cell(row=row, column=column).alignment = _WRAP
        ws.row_dimensions[row].height = 34
    _configure_print_layout(ws, landscape=True)
    ws.print_area = f"A1:B{notes_start_row + len(notes) - 1}"


def _project_rows(data: dict[str, Any], export_id: str) -> list[list[Any]]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for order, _customer in data["orders"]:
        key = (
            order.linked_sales_order_no or "",
            order.project_raw or order.project_std or "",
        )
        slot = grouped.setdefault(
            key,
            {"start": None, "end": None, "count": 0},
        )
        slot["count"] += 1
        if order.maint_start is not None:
            slot["start"] = (
                order.maint_start
                if slot["start"] is None
                else min(slot["start"], order.maint_start)
            )
        if order.maint_end is not None:
            slot["end"] = (
                order.maint_end
                if slot["end"] is None
                else max(slot["end"], order.maint_end)
            )
    rows = []
    for index, ((contract, project), values) in enumerate(sorted(grouped.items()), 1):
        rows.append(
            [
                "KEEP",
                contract,
                project,
                values["start"],
                values["end"],
                values["count"],
                f"{contract}|{project}" if contract or project else index,
                0,
                "",
                "",
            ]
        )
    return rows


def _order_rows(data: dict[str, Any], export_id: str) -> list[list[Any]]:
    rows = []
    for order, customer in data["orders"]:
        base = _base_version(order)
        rows.append(
            [
                "KEEP",
                order.order_no,
                order.order_date,
                order.linked_sales_order_no,
                order.project_raw or order.project_std,
                order.maint_start,
                order.maint_end,
                customer,
                order.end_customer,
                order.demand_type,
                order.business_type,
                order.salesperson,
                order.warehouse,
                order.data_status,
                "",
                order.id,
                base,
                _row_token(export_id, "order", order.id, base),
                _export_client_row_id(
                    export_id, "02_维保订单", "order", order.id, base
                ),
            ]
        )
    return rows


def _line_rows(data: dict[str, Any], export_id: str) -> list[list[Any]]:
    rows = []
    for line, order in data["lines"]:
        base = _base_version(line)
        rows.append(
            [
                "KEEP",
                order.order_no,
                line.line_no,
                line.pn_std,
                line.description,
                _decimal(line.qty),
                _decimal(line.return_qty),
                line.serial_numbers,
                line.cost_source,
                _decimal(line.unit_cost_ex_tax),
                _decimal(line.unit_cost_inc_tax),
                "",
                line.id,
                base,
                _row_token(export_id, "line", line.id, base),
                _export_client_row_id(export_id, "03_订单明细", "line", line.id, base),
            ]
        )
    return rows


def _expense_rows(data: dict[str, Any], export_id: str) -> list[list[Any]]:
    rows = []
    for expense in data["expenses"]:
        base = _base_version(expense)
        amount_ex = getattr(expense, "amount_ex_tax", None)
        if amount_ex is None:
            amount_ex = expense.amount
        amount_inc = getattr(expense, "amount_inc_tax", None)
        if amount_inc is None and amount_ex is not None:
            amount_inc = tax_policy.inc_from_ex(amount_ex)
        rows.append(
            [
                "KEEP",
                expense.linked_sales_order_no,
                expense.expense_date,
                expense.person,
                expense.expense_type,
                expense.fee_category,
                expense.reason,
                _decimal(amount_ex),
                _decimal(amount_inc),
                expense.data_status,
                expense.bxd_no,
                expense.line_no,
                "",
                expense.tax_basis,
                _decimal(expense.amount),
                expense.id,
                base,
                _row_token(export_id, "expense", expense.id, base),
                _export_client_row_id(
                    export_id, "04_报销明细", "expense", expense.id, base
                ),
            ]
        )
    for _ in range(BLANK_CREATE_ROWS):
        rows.append(
            [
                "",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                config.MAINT_EXPENSE_ACTIVE_STATUS,
                None,
                None,
                "",
                None,
                None,
                None,
                None,
                None,
                str(uuid.uuid4()),
            ]
        )
    return rows


def _manual_rows(data: dict[str, Any], export_id: str) -> list[list[Any]]:
    rows = []
    override_by_line: dict[int, tuple[Any, Any, Any]] = {
        line.id: (override, line, order) for override, line, order in data["manual"]
    }
    for line, order in data["lines"]:
        existing = override_by_line.get(line.id)
        if existing is not None:
            override, _line, _order = existing
            base = _base_version(override)
            rows.append(
                [
                    "KEEP",
                    line.id,
                    order.order_no,
                    line.pn_std,
                    _decimal(line.qty),
                    line.cost_source,
                    _decimal(override.unit_cost_ex_tax),
                    _decimal(override.unit_cost_inc_tax),
                    _decimal(override.tax_rate_used),
                    override.reason,
                    _evidence_text(override.evidence),
                    "生效" if override.active else "已作废",
                    override.id,
                    base,
                    _row_token(export_id, "manual", override.id, base),
                    _export_client_row_id(
                        export_id,
                        "05_人工成本回填",
                        "manual",
                        override.id,
                        base,
                    ),
                ]
            )
        elif line.cost_source == "none":
            base = _base_version(line)
            rows.append(
                [
                    "",
                    line.id,
                    order.order_no,
                    line.pn_std,
                    _decimal(line.qty),
                    line.cost_source,
                    None,
                    None,
                    float(TAX_RATE),
                    "",
                    "",
                    "待回填",
                    None,
                    base,
                    _row_token(export_id, "manual-create", line.id, base),
                    str(uuid.uuid4()),
                ]
            )
    if not rows:
        for _ in range(BLANK_CREATE_ROWS):
            rows.append(
                [
                    "",
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    float(TAX_RATE),
                    "",
                    "",
                    "待回填",
                    None,
                    None,
                    None,
                    str(uuid.uuid4()),
                ]
            )
    return rows


def _dictionary_sheet(workbook: Workbook) -> None:
    ws = workbook.create_sheet("98_字典")
    ws.append(["字典", "值", "说明"])
    values = [
        ("操作", "KEEP", "保持，不写数据库"),
        ("操作", "CREATE", "新增"),
        ("操作", "UPDATE", "更新"),
        ("操作", "VOID", "软作废"),
        ("流程状态", config.MAINT_EXPENSE_ACTIVE_STATUS, "计入报销"),
        ("税率", "0.13", "统一业务税率"),
    ]
    for value in values:
        ws.append(value)
    table = Table(displayName="tbl_dictionary_v1", ref=f"A1:C{len(values) + 1}")
    _table_style(table)
    ws.add_table(table)
    ws.sheet_state = "hidden"


def _metadata_sheet(workbook: Workbook, metadata: dict[str, str]) -> None:
    ws = workbook.create_sheet("99_元数据")
    _append_literal_row(ws, ["key", "value"])
    for key in sorted(metadata):
        _append_literal_row(ws, [key, metadata[key]])
    table = Table(displayName="tbl_metadata_v1", ref=f"A1:B{len(metadata) + 1}")
    _table_style(table)
    ws.add_table(table)
    ws.sheet_state = "veryHidden"


def _contract_revisions_sheet(
    workbook: Workbook,
    revisions: dict[str, int],
) -> None:
    if len(revisions) > MAX_CONTRACT_REVISIONS:
        raise RoundtripWorkbookError(
            f"合同 revision 超过 {MAX_CONTRACT_REVISIONS} 条安全上限",
            status_code=413,
        )
    ws = workbook.create_sheet("99_合同版本")
    _append_literal_row(ws, _CONTRACT_REVISION_HEADERS)
    for contract, revision in sorted(revisions.items()):
        if not contract or len(contract) > 64:
            raise RoundtripWorkbookError("合同号为空或超过 64 字符，不能生成回填模板")
        if revision < 0 or revision > 2_147_483_647:
            raise RoundtripWorkbookError("合同 revision 超出允许范围")
        _append_literal_row(ws, [contract, revision])
    # Excel Table 至少保留一行；空集合用全空哨兵，导入解析时忽略。
    if not revisions:
        ws.append([None, None])
    table = Table(
        displayName=_CONTRACT_REVISION_TABLE,
        ref=f"A1:B{len(revisions) + 1 if revisions else 2}",
    )
    _table_style(table)
    ws.add_table(table)
    ws.sheet_state = "veryHidden"


def build_roundtrip_template(
    db: Session,
    *,
    contract: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    exported_by: str | None,
    blank: bool = False,
) -> SpooledTemporaryFile:
    """构建固定、可回填的维保工作簿。

    日期闭区间同时作用于维保订单、订单明细和报销明细；``blank=True`` 只生成空表及
    指定合同的 revision，不读取业务行。
    """
    _validate_date_pair(date_from, date_to)
    _lock_single_template_builder(db)
    _lock_shared_snapshot(db)
    data = _selected_data(
        db,
        contract=contract,
        date_from=date_from,
        date_to=date_to,
        blank=blank,
    )
    export_id = str(uuid.uuid4())
    revisions = _contract_state_revisions(db, data["contracts"])
    as_of = date_to or business_today()
    metadata = {
        "protocol_id": PROTOCOL_ID,
        "schema_version": SCHEMA_VERSION,
        "export_id": export_id,
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_by": exported_by or "",
        "app_version": APP_VERSION,
        "date_from": _iso(date_from),
        "date_to": _iso(date_to),
        "as_of": as_of.isoformat(),
        "contract_scope": contract.strip() if contract and contract.strip() else "",
        "template_mode": "blank" if blank else "snapshot",
        "contract_revision_count": str(len(revisions)),
        "contract_revisions_sha256": _contract_revisions_digest(revisions),
        "tax_rate": str(TAX_RATE),
        "amount_basis": "ex",
        "table_map": json.dumps(
            {sheet: spec[0] for sheet, spec in _TABLE_SPECS.items()},
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
        ),
    }
    metadata["metadata_hmac"] = _metadata_hmac(metadata)

    workbook = Workbook()
    output = SpooledTemporaryFile(max_size=8 * 1024 * 1024, mode="w+b")
    try:
        _instructions_sheet(workbook, metadata)
        rows_by_sheet = {
            "01_项目": _project_rows(data, export_id),
            "02_维保订单": _order_rows(data, export_id),
            "03_订单明细": _line_rows(data, export_id),
            "04_报销明细": _expense_rows(data, export_id),
            "05_人工成本回填": _manual_rows(data, export_id),
        }
        for sheet, (table_name, headers) in _TABLE_SPECS.items():
            ws = workbook.create_sheet(sheet)
            _append_table(
                ws,
                table_name=table_name,
                headers=headers,
                rows=rows_by_sheet[sheet],
                editable=_EDITABLE_COLUMNS.get(sheet),
                required=_REQUIRED_COLUMNS.get(sheet),
            )
        _dictionary_sheet(workbook)
        _contract_revisions_sheet(workbook, revisions)
        _metadata_sheet(workbook, metadata)
        workbook.calculation.fullCalcOnLoad = False
        workbook.calculation.forceFullCalc = False
        workbook.save(output)
        output.seek(0, os.SEEK_END)
        if output.tell() > MAX_WORKBOOK_BYTES:
            raise RoundtripWorkbookError(
                "维保回填工作簿超过 256 MiB 安全上限",
                status_code=413,
            )
        output.seek(0)
        return output
    except BaseException:
        output.close()
        raise
    finally:
        workbook.close()


def _roundtrip_bundle_contracts(
    db: Session,
    *,
    date_from: date | None,
    date_to: date | None,
) -> list[str]:
    filters = _selection_filters(
        contract=None,
        date_from=date_from,
        date_to=date_to,
    )
    values = db.scalars(
        select(FMaintenanceOrder.linked_sales_order_no)
        .distinct()
        .where(
            *filters,
            FMaintenanceOrder.linked_sales_order_no.is_not(None),
            func.btrim(FMaintenanceOrder.linked_sales_order_no) != "",
        )
        .order_by(FMaintenanceOrder.linked_sales_order_no)
        .limit(MAX_ROUNDTRIP_BUNDLE_CONTRACTS + 1)
    ).all()
    return [str(value).strip() for value in values if value and str(value).strip()]


def _bundle_member_collision_key(name: str) -> str:
    return unicodedata.normalize("NFKC", name).casefold().rstrip(" .")


def _utf8_prefix(value: str, max_bytes: int) -> str:
    encoded = value.encode("utf-8")
    if len(encoded) <= max_bytes:
        return value
    return encoded[:max_bytes].decode("utf-8", errors="ignore")


def _roundtrip_bundle_member_name(contract: str, used_names: set[str]) -> str:
    clean = unicodedata.normalize("NFKC", contract)
    clean = "".join(
        "_" if unicodedata.category(character) in {"Cc", "Cf"} else character
        for character in clean
    )
    clean = _INVALID_BUNDLE_MEMBER_CHARS.sub("_", clean)
    while ".." in clean:
        clean = clean.replace("..", "_")
    clean = clean.strip(" ._") or "contract"

    def leaf(suffix: str = "") -> str:
        prefix = "maintenance_roundtrip_"
        extension = ".xlsx"
        fixed_bytes = len((prefix + suffix + extension).encode("utf-8"))
        stem = _utf8_prefix(
            clean,
            max(MAX_ROUNDTRIP_BUNDLE_MEMBER_BYTES - fixed_bytes, 0),
        ).rstrip(" ._") or "contract"
        return f"{prefix}{stem}{suffix}{extension}"

    name = leaf()
    collision_key = _bundle_member_collision_key(name)
    if collision_key in used_names:
        digest = hashlib.sha256(contract.encode("utf-8")).hexdigest()[:10]
        name = leaf(f"_{digest}")
        collision_key = _bundle_member_collision_key(name)
        suffix = 2
        while collision_key in used_names:
            name = leaf(f"_{digest}_{suffix}")
            collision_key = _bundle_member_collision_key(name)
            suffix += 1
    used_names.add(collision_key)
    return f"维保回填模板/{name}"


def build_roundtrip_template_bundle(
    db: Session,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    exported_by: str | None,
) -> SpooledTemporaryFile:
    """按合同拆分生成可独立校验、仍需逐本导入的固定回填工作簿 ZIP。"""
    _validate_date_pair(date_from, date_to)
    _lock_single_template_builder(db)
    _lock_shared_snapshot(db)
    contracts = _roundtrip_bundle_contracts(
        db,
        date_from=date_from,
        date_to=date_to,
    )
    if not contracts:
        raise RoundtripWorkbookError("所选范围内没有可导出的已关联合同维保数据")
    if len(contracts) >= MAX_ROUNDTRIP_BUNDLE_CONTRACTS:
        raise RoundtripWorkbookError(
            f"命中合同至少 {MAX_ROUNDTRIP_BUNDLE_CONTRACTS} 个，必须少于 "
            f"{MAX_ROUNDTRIP_BUNDLE_CONTRACTS} 个；请缩小日期范围",
            status_code=413,
        )

    output = SpooledTemporaryFile(max_size=16 * 1024 * 1024, mode="w+b")
    used_names: set[str] = set()
    total_uncompressed = 0
    try:
        with zipfile.ZipFile(
            output,
            mode="w",
            compression=zipfile.ZIP_STORED,
            allowZip64=True,
        ) as archive:
            for contract in contracts:
                workbook = build_roundtrip_template(
                    db,
                    contract=contract,
                    date_from=date_from,
                    date_to=date_to,
                    exported_by=exported_by,
                )
                try:
                    workbook.seek(0, os.SEEK_END)
                    workbook_bytes = workbook.tell()
                    total_uncompressed += workbook_bytes
                    if total_uncompressed >= MAX_ROUNDTRIP_BUNDLE_UNCOMPRESSED_BYTES:
                        raise RoundtripWorkbookError(
                            "批量可回填工作簿解压后总大小必须小于 512 MiB；"
                            "请缩小日期范围",
                            status_code=413,
                        )
                    workbook.seek(0)
                    member_name = _roundtrip_bundle_member_name(contract, used_names)
                    with archive.open(member_name, mode="w") as member:
                        while chunk := workbook.read(1024 * 1024):
                            member.write(chunk)
                finally:
                    workbook.close()
        output.seek(0)
        return output
    except BaseException:
        output.close()
        raise


@dataclass(frozen=True)
class _RoundtripSheetEnvelope:
    table_name: str | None
    max_column: int
    max_row: int


def _roundtrip_sheet_envelopes() -> dict[str, _RoundtripSheetEnvelope]:
    envelopes = {
        "00_使用说明": _RoundtripSheetEnvelope(None, 2, 19),
        **{
            sheet: _RoundtripSheetEnvelope(
                table_name,
                len(headers),
                MAX_ROWS_PER_TABLE + 1,
            )
            for sheet, (table_name, headers) in _TABLE_SPECS.items()
        },
        "98_字典": _RoundtripSheetEnvelope("tbl_dictionary_v1", 3, 7),
        "99_合同版本": _RoundtripSheetEnvelope(
            _CONTRACT_REVISION_TABLE,
            2,
            MAX_CONTRACT_REVISIONS + 1,
        ),
        "99_元数据": _RoundtripSheetEnvelope(
            "tbl_metadata_v1",
            2,
            MAX_METADATA_ROWS + 1,
        ),
    }
    return envelopes


def _roundtrip_relationship_part(source_path: str) -> str:
    directory, filename = posixpath.split(source_path)
    return posixpath.join(directory, "_rels", f"{filename}.rels")


def _enforce_roundtrip_xml_element_limit(
    count: int,
    *,
    part: str,
    limit: int | None = None,
) -> None:
    effective_limit = (
        MAX_ROUNDTRIP_XML_ELEMENTS_PER_PART if limit is None else limit
    )
    if count > effective_limit:
        raise RoundtripWorkbookError(
            f"{part} XML 元素总量超过协议安全上限",
            status_code=413,
        )


def _roundtrip_range_bounds(ref: str) -> tuple[int, int, int, int] | None:
    refs = ref.split(":")
    if len(refs) == 1:
        refs *= 2
    if len(refs) != 2:
        return None
    start = reader._cell_reference_bounds(refs[0])
    end = reader._cell_reference_bounds(refs[1])
    if start is None or end is None:
        return None
    min_row, min_column = start
    max_row, max_column = end
    if (
        min_row < 1
        or min_column < 1
        or max_row < min_row
        or max_column < min_column
    ):
        return None
    return min_column, min_row, max_column, max_row


def _roundtrip_table_relation_id(
    archive: zipfile.ZipFile,
    *,
    worksheet_path: str,
    sheet_name: str,
    expected_table_name: str | None,
) -> str | None:
    declared_count: str | None = None
    relation_ids: list[str] = []
    saw_table_parts = False
    open_tags: list[str] = []
    open_elements: list[ET.Element] = []
    element_count = 0
    with archive.open(worksheet_path) as worksheet_xml:
        for event, element in reader._safe_xml_iterparse(worksheet_xml):
            tag = reader._xml_local_name(element.tag)
            if event == "start":
                element_count += 1
                _enforce_roundtrip_xml_element_limit(
                    element_count,
                    part=worksheet_path,
                )
                depth = len(open_tags) + 1
                if depth > reader._XML_DEPTH_LIMIT:
                    raise RoundtripWorkbookError("工作簿 XML 层级超过安全上限")
                parent = open_tags[-1] if open_tags else None
                if tag == "tableParts":
                    if saw_table_parts:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 tableParts 结构无效"
                        )
                    saw_table_parts = True
                    declared_count = element.attrib.get("count")
                elif tag == "tablePart":
                    if parent != "tableParts":
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 tableParts 结构无效"
                        )
                    relation_id = next(
                        (
                            value
                            for name, value in element.attrib.items()
                            if reader._xml_local_name(name) == "id"
                        ),
                        None,
                    )
                    if not relation_id or relation_ids:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 tableParts 数量或关系无效"
                        )
                    relation_ids.append(relation_id)
                open_tags.append(tag)
                open_elements.append(element)
                continue

            if not open_tags or open_tags[-1] != tag:
                raise RoundtripWorkbookError(
                    f"{sheet_name} 的 worksheet XML 结构无效"
                )
            parent_element = (
                open_elements[-2] if len(open_elements) > 1 else None
            )
            open_tags.pop()
            open_elements.pop()
            element.clear()
            if parent_element is not None:
                parent_element.remove(element)

    if expected_table_name is None:
        if relation_ids or declared_count not in {None, "0"}:
            raise RoundtripWorkbookError(f"{sheet_name} 不允许附加 Excel Table")
        return None
    if declared_count != "1" or len(relation_ids) != 1:
        raise RoundtripWorkbookError(
            f"{sheet_name} 的 tableParts 数量或关系不符合协议"
        )
    return relation_ids[0]


def _roundtrip_table_part(
    archive: zipfile.ZipFile,
    *,
    worksheet_path: str,
    sheet_name: str,
    expected_table_name: str | None,
) -> str | None:
    expected_relation_id = _roundtrip_table_relation_id(
        archive,
        worksheet_path=worksheet_path,
        sheet_name=sheet_name,
        expected_table_name=expected_table_name,
    )
    relationship_path = _roundtrip_relationship_part(worksheet_path)
    if relationship_path not in archive.namelist():
        if expected_relation_id is None:
            return None
        raise RoundtripWorkbookError(f"{sheet_name} 缺少协议 Excel Table")

    table_targets: list[str] = []
    with archive.open(relationship_path) as relationships:
        open_tags: list[str] = []
        open_elements: list[ET.Element] = []
        saw_root = False
        element_count = 0
        for event, element in reader._safe_xml_iterparse(relationships):
            tag = reader._xml_local_name(element.tag)
            if event == "start":
                element_count += 1
                _enforce_roundtrip_xml_element_limit(
                    element_count,
                    part=relationship_path,
                    limit=MAX_ROUNDTRIP_RELATIONSHIP_ELEMENTS,
                )
                depth = len(open_tags) + 1
                if depth > reader._XML_DEPTH_LIMIT:
                    raise RoundtripWorkbookError("工作簿 XML 层级超过安全上限")
                if depth == 1:
                    if tag != "Relationships":
                        raise RoundtripWorkbookError("工作簿关系文件结构无效")
                    saw_root = True
                elif depth == 2:
                    if tag != "Relationship":
                        raise RoundtripWorkbookError("工作簿关系文件结构无效")
                    relation_id = element.attrib.get("Id")
                    is_table = element.attrib.get("Type", "").endswith("/table")
                    if is_table:
                        if (
                            expected_relation_id is None
                            or relation_id != expected_relation_id
                            or table_targets
                        ):
                            raise RoundtripWorkbookError(
                                f"{sheet_name} 的 Excel Table 名称或数量不符合协议"
                            )
                        if (
                            element.attrib.get("TargetMode", "").casefold()
                            == "external"
                        ):
                            raise RoundtripWorkbookError(
                                f"{sheet_name} 的 Excel Table 不能使用外部关系"
                            )
                        target = reader._normalize_internal_target(
                            source_path=worksheet_path,
                            target=element.attrib.get("Target", ""),
                        )
                        if target not in archive.namelist():
                            raise RoundtripWorkbookError(
                                f"{sheet_name} 的 Excel Table 部件不存在"
                            )
                        table_targets.append(target)
                    elif relation_id == expected_relation_id:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 tablePart 关系类型不符合协议"
                        )
                else:
                    raise RoundtripWorkbookError("工作簿关系文件结构无效")
                open_tags.append(tag)
                open_elements.append(element)
                continue

            if not open_tags or open_tags[-1] != tag:
                raise RoundtripWorkbookError("工作簿关系文件结构无效")
            parent = open_elements[-2] if len(open_elements) > 1 else None
            open_tags.pop()
            open_elements.pop()
            element.clear()
            if parent is not None:
                parent.remove(element)
    if not saw_root:
        raise RoundtripWorkbookError("工作簿关系文件结构无效")
    if expected_relation_id is None:
        return None
    if len(table_targets) != 1:
        raise RoundtripWorkbookError(
            f"{sheet_name} 的 Excel Table 名称或数量不符合协议"
        )
    return table_targets[0]


def _roundtrip_table_bounds(
    archive: zipfile.ZipFile,
    *,
    table_path: str,
    sheet_name: str,
    envelope: _RoundtripSheetEnvelope,
) -> tuple[int, int]:
    table_ref: str | None = None
    table_name: str | None = None
    declared_columns: str | None = None
    actual_columns = 0
    saw_root = False
    open_tags: list[str] = []
    open_elements: list[ET.Element] = []
    element_count = 0
    with archive.open(table_path) as table_xml:
        for event, element in reader._safe_xml_iterparse(table_xml):
            tag = reader._xml_local_name(element.tag)
            if event == "start":
                element_count += 1
                _enforce_roundtrip_xml_element_limit(
                    element_count,
                    part=table_path,
                    limit=MAX_ROUNDTRIP_TABLE_ELEMENTS,
                )
                depth = len(open_tags) + 1
                if depth > reader._XML_DEPTH_LIMIT:
                    raise RoundtripWorkbookError("工作簿 XML 层级超过安全上限")
                if depth == 1:
                    if tag != "table":
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 Excel Table 结构无效"
                        )
                    saw_root = True
                    table_ref = element.attrib.get("ref")
                    table_name = element.attrib.get("displayName")
                    internal_name = element.attrib.get("name")
                    if internal_name not in {None, envelope.table_name}:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 Excel Table 名称不符合协议"
                        )
                elif tag == "tableColumns":
                    declared_columns = element.attrib.get("count")
                elif tag == "tableColumn":
                    actual_columns += 1
                    if actual_columns > envelope.max_column:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 Excel Table 列数超过协议上限",
                            status_code=413,
                        )
                open_tags.append(tag)
                open_elements.append(element)
                continue

            if not open_tags or open_tags[-1] != tag:
                raise RoundtripWorkbookError(
                    f"{sheet_name} 的 Excel Table 结构无效"
                )
            parent = open_elements[-2] if len(open_elements) > 1 else None
            open_tags.pop()
            open_elements.pop()
            element.clear()
            if parent is not None:
                parent.remove(element)

    if not saw_root or table_name != envelope.table_name or table_ref is None:
        raise RoundtripWorkbookError(
            f"{sheet_name} 的 Excel Table 名称或结构不符合协议"
        )
    if (
        declared_columns is None
        or re.fullmatch(r"[0-9]{1,3}", declared_columns) is None
        or int(declared_columns) != envelope.max_column
        or actual_columns != envelope.max_column
    ):
        raise RoundtripWorkbookError(
            f"{sheet_name} 的 Excel Table 列数不符合协议"
        )
    bounds = _roundtrip_range_bounds(table_ref)
    if bounds is None:
        raise RoundtripWorkbookError(f"{sheet_name} 的 Excel Table 范围无效")
    min_column, min_row, max_column, max_row = bounds
    if (min_column, min_row, max_column) != (1, 1, envelope.max_column):
        raise RoundtripWorkbookError(f"{sheet_name} 的 Excel Table 范围不符合协议")
    if max_row < 2 or max_row > envelope.max_row:
        raise RoundtripWorkbookError(
            f"{sheet_name} 的 Excel Table 超过 {envelope.max_row - 1} 行安全上限",
            status_code=413,
        )
    return max_row, max_column


def _scan_roundtrip_worksheet(
    archive: zipfile.ZipFile,
    *,
    worksheet_path: str,
    sheet_name: str,
    max_row: int,
    max_column: int,
    table_scoped: bool,
) -> tuple[int, int]:
    current_row = 0
    current_column = 0
    row_elements = 0
    cell_elements = 0
    open_tags: list[str] = []
    open_elements: list[ET.Element] = []
    element_count = 0
    structural_elements = 0
    cell_root_depth: int | None = None
    cell_xml_elements = 0
    scope_name = "Table 范围外" if table_scoped else "协议范围外"
    outside_coordinate: str | None = None

    def note_outside(coordinate: str) -> None:
        nonlocal outside_coordinate
        if outside_coordinate is None:
            outside_coordinate = coordinate

    def reject_outside(coordinate: str) -> None:
        raise RoundtripWorkbookError(
            f"{sheet_name} 存在 {scope_name}单元格或维度：{coordinate}",
            status_code=413,
        )

    with archive.open(worksheet_path) as worksheet_xml:
        for event, element in reader._safe_xml_iterparse(worksheet_xml):
            tag = reader._xml_local_name(element.tag)
            if event == "start":
                element_count += 1
                _enforce_roundtrip_xml_element_limit(
                    element_count,
                    part=worksheet_path,
                )
                depth = len(open_tags) + 1
                if depth > reader._XML_DEPTH_LIMIT:
                    raise RoundtripWorkbookError("工作簿 XML 层级超过安全上限")
                parent = open_tags[-1] if open_tags else None
                if cell_root_depth is not None:
                    cell_xml_elements += 1
                    if cell_xml_elements > MAX_ROUNDTRIP_CELL_XML_ELEMENTS:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 单个 cell 的 XML 元素超过协议安全上限",
                            status_code=413,
                        )
                elif tag != "row" and parent != "row":
                    structural_elements += 1
                    if (
                        structural_elements
                        > MAX_ROUNDTRIP_WORKSHEET_STRUCTURAL_ELEMENTS
                    ):
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 结构性 XML 元素超过协议安全上限",
                            status_code=413,
                        )
                if tag == "dimension":
                    ref = element.attrib.get("ref")
                    bounds = _roundtrip_range_bounds(ref or "")
                    if bounds is None:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 worksheet dimension 无效"
                        )
                    min_col, min_row, dimension_max_col, dimension_max_row = bounds
                    if (
                        min_col < 1
                        or min_row < 1
                        or dimension_max_col > max_column
                        or dimension_max_row > max_row
                    ):
                        note_outside(ref or "")
                elif tag == "row":
                    if parent == "row":
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 worksheet row 结构无效"
                        )
                    row_elements += 1
                    if row_elements > max_row:
                        reject_outside(f"row-count={row_elements}")
                    row_ref = element.attrib.get("r")
                    if row_ref is None:
                        current_row += 1
                    elif re.fullmatch(r"[0-9]{1,7}", row_ref) is None:
                        raise RoundtripWorkbookError(
                            f"{sheet_name} 的 worksheet row 坐标无效"
                        )
                    else:
                        current_row = reader._row_index(row_ref)
                    current_column = 0
                    if current_row < 1 or current_row > max_row:
                        note_outside(f"row={current_row}")
                elif parent == "row":
                    # openpyxl treats every direct row child as cell-like, even
                    # when an attacker renames <c>; mirror that behavior here.
                    cell_root_depth = depth
                    cell_xml_elements = 1
                    cell_elements += 1
                    if cell_elements > max_row * max_column:
                        reject_outside(f"cell-count={cell_elements}")
                    ref = element.attrib.get("r")
                    if ref is None:
                        current_column += 1
                        cell_row, cell_column = current_row, current_column
                        coordinate = f"row={cell_row},column={cell_column}"
                    else:
                        bounds = reader._cell_reference_bounds(ref)
                        if bounds is None:
                            raise RoundtripWorkbookError(
                                f"{sheet_name} 的 worksheet cell 坐标无效"
                            )
                        cell_row, cell_column = bounds
                        current_column = cell_column
                        coordinate = ref
                    if (
                        cell_row != current_row
                        or cell_row < 1
                        or cell_row > max_row
                        or cell_column < 1
                        or cell_column > max_column
                    ):
                        note_outside(coordinate)
                elif tag == "c":
                    raise RoundtripWorkbookError(
                        f"{sheet_name} 的 worksheet cell 结构无效"
                    )
                elif (
                    tag == "f"
                    and len(open_tags) >= 2
                    and open_tags[-2] == "row"
                ):
                    raise RoundtripWorkbookError(
                        f"{sheet_name} 包含公式，固定回填工作簿只允许静态值"
                    )
                open_tags.append(tag)
                open_elements.append(element)
                continue

            if not open_tags or open_tags[-1] != tag:
                raise RoundtripWorkbookError(
                    f"{sheet_name} 的 worksheet XML 结构无效"
                )
            parent_element = (
                open_elements[-2] if len(open_elements) > 1 else None
            )
            closing_cell = (
                cell_root_depth is not None
                and len(open_tags) == cell_root_depth
            )
            open_tags.pop()
            open_elements.pop()
            if tag == "row":
                current_column = 0
            element.clear()
            if parent_element is not None:
                parent_element.remove(element)
            if closing_cell:
                cell_root_depth = None
                cell_xml_elements = 0
    if outside_coordinate is not None:
        reject_outside(outside_coordinate)
    return row_elements, cell_elements


def _roundtrip_shared_strings_part(archive: zipfile.ZipFile) -> str | None:
    content_types_path = "[Content_Types].xml"
    if content_types_path not in archive.namelist():
        raise RoundtripWorkbookError("工作簿缺少 Content Types")
    shared_string_type = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sharedStrings+xml"
    )
    targets: list[str] = []
    open_tags: list[str] = []
    open_elements: list[ET.Element] = []
    saw_root = False
    element_count = 0
    with archive.open(content_types_path) as content_types:
        for event, element in reader._safe_xml_iterparse(content_types):
            tag = reader._xml_local_name(element.tag)
            if event == "start":
                element_count += 1
                _enforce_roundtrip_xml_element_limit(
                    element_count,
                    part=content_types_path,
                    limit=MAX_ROUNDTRIP_CONTENT_TYPE_ELEMENTS,
                )
                depth = len(open_tags) + 1
                if depth > reader._XML_DEPTH_LIMIT:
                    raise RoundtripWorkbookError("工作簿 XML 层级超过安全上限")
                if depth == 1:
                    if tag != "Types":
                        raise RoundtripWorkbookError("Content Types 结构无效")
                    saw_root = True
                elif depth == 2 and tag == "Override":
                    if element.attrib.get("ContentType") == shared_string_type:
                        target = reader._normalize_internal_target(
                            source_path=content_types_path,
                            target=element.attrib.get("PartName", ""),
                        )
                        if target not in archive.namelist():
                            raise RoundtripWorkbookError(
                                "sharedStrings 部件不存在"
                            )
                        targets.append(target)
                open_tags.append(tag)
                open_elements.append(element)
                continue

            if not open_tags or open_tags[-1] != tag:
                raise RoundtripWorkbookError("Content Types 结构无效")
            parent = open_elements[-2] if len(open_elements) > 1 else None
            open_tags.pop()
            open_elements.pop()
            element.clear()
            if parent is not None:
                parent.remove(element)
    if not saw_root:
        raise RoundtripWorkbookError("Content Types 结构无效")
    if len(targets) > 1:
        raise RoundtripWorkbookError("工作簿包含多个 sharedStrings 部件")
    return targets[0] if targets else None


def _scan_roundtrip_shared_strings(
    archive: zipfile.ZipFile,
    shared_strings_path: str | None,
) -> int:
    if shared_strings_path is None:
        return 0
    item_count = 0
    current_chars = 0
    current_item_elements = 0
    open_tags: list[str] = []
    open_elements: list[ET.Element] = []
    saw_root = False
    element_count = 0
    with archive.open(shared_strings_path) as shared_strings:
        for event, element in reader._safe_xml_iterparse(shared_strings):
            tag = reader._xml_local_name(element.tag)
            if event == "start":
                element_count += 1
                _enforce_roundtrip_xml_element_limit(
                    element_count,
                    part=shared_strings_path,
                )
                depth = len(open_tags) + 1
                if depth > reader._XML_DEPTH_LIMIT:
                    raise RoundtripWorkbookError("工作簿 XML 层级超过安全上限")
                if depth == 1:
                    if tag != "sst":
                        raise RoundtripWorkbookError("sharedStrings 结构无效")
                    saw_root = True
                elif tag == "si":
                    if len(open_tags) != 1 or open_tags[-1] != "sst":
                        raise RoundtripWorkbookError("sharedStrings 结构无效")
                    item_count += 1
                    current_chars = 0
                    current_item_elements = 1
                    if item_count > MAX_ROUNDTRIP_SHARED_STRINGS:
                        raise RoundtripWorkbookError(
                            "sharedStrings 实际条目数超过协议安全上限",
                            status_code=413,
                        )
                elif "si" in open_tags:
                    current_item_elements += 1
                    if (
                        current_item_elements
                        > MAX_ROUNDTRIP_SHARED_STRING_XML_ELEMENTS
                    ):
                        raise RoundtripWorkbookError(
                            "sharedStrings 单条 XML 元素超过协议安全上限",
                            status_code=413,
                        )
                open_tags.append(tag)
                open_elements.append(element)
                continue

            if not open_tags or open_tags[-1] != tag:
                raise RoundtripWorkbookError("sharedStrings 结构无效")
            if tag == "t" and "si" in open_tags:
                current_chars += len(element.text or "")
                if current_chars > MAX_CELL_CHARS:
                    raise RoundtripWorkbookError(
                        "sharedStrings 单条文本超过 Excel 文本上限",
                        status_code=413,
                    )
            parent = open_elements[-2] if len(open_elements) > 1 else None
            open_tags.pop()
            open_elements.pop()
            element.clear()
            if parent is not None:
                parent.remove(element)
    if not saw_root:
        raise RoundtripWorkbookError("sharedStrings 结构无效")
    return item_count


def _assert_roundtrip_xml_envelope(archive: zipfile.ZipFile) -> None:
    envelopes = _roundtrip_sheet_envelopes()
    sheet_parts = reader._workbook_sheet_parts(archive)
    if [name for name, _path in sheet_parts] != list(SHEET_NAMES):
        raise RoundtripWorkbookError(
            "工作表名称、顺序或数量被修改，请重新导出模板"
        )
    if len({path for _name, path in sheet_parts}) != len(sheet_parts):
        raise RoundtripWorkbookError("多个工作表不能共享同一个 worksheet 部件")

    _scan_roundtrip_shared_strings(
        archive,
        _roundtrip_shared_strings_part(archive),
    )
    total_rows = 0
    total_cells = 0
    for sheet_name, worksheet_path in sheet_parts:
        envelope = envelopes[sheet_name]
        table_path = _roundtrip_table_part(
            archive,
            worksheet_path=worksheet_path,
            sheet_name=sheet_name,
            expected_table_name=envelope.table_name,
        )
        if table_path is None:
            max_row, max_column = envelope.max_row, envelope.max_column
        else:
            max_row, max_column = _roundtrip_table_bounds(
                archive,
                table_path=table_path,
                sheet_name=sheet_name,
                envelope=envelope,
            )
        rows, cells = _scan_roundtrip_worksheet(
            archive,
            worksheet_path=worksheet_path,
            sheet_name=sheet_name,
            max_row=max_row,
            max_column=max_column,
            table_scoped=table_path is not None,
        )
        total_rows += rows
        total_cells += cells
        if total_rows > MAX_ROUNDTRIP_ROW_ELEMENTS:
            raise RoundtripWorkbookError(
                "维保回填工作簿 row 元素总量超过协议安全上限",
                status_code=413,
            )
        if total_cells > MAX_ROUNDTRIP_CELL_ELEMENTS:
            raise RoundtripWorkbookError(
                "维保回填工作簿 cell 元素总量超过协议安全上限",
                status_code=413,
            )


def _assert_safe_workbook_package(path: str) -> None:
    try:
        size = os.path.getsize(path)
    except OSError as exc:
        raise RoundtripWorkbookError("无法读取上传的工作簿") from exc
    if size > MAX_WORKBOOK_BYTES:
        raise RoundtripWorkbookError(
            "维保回填工作簿超过 256 MiB 安全上限",
            status_code=413,
        )
    try:
        reader._check_xlsx_archive_safety(path)
        with zipfile.ZipFile(path) as archive:
            total_uncompressed = sum(member.file_size for member in archive.infolist())
            if total_uncompressed > MAX_ROUNDTRIP_UNCOMPRESSED_BYTES:
                raise RoundtripWorkbookError(
                    "维保回填工作簿解压后超过 64 MiB 安全上限",
                    status_code=413,
                )
            unsafe = [
                name
                for name in archive.namelist()
                if (
                    "vbaproject.bin" in name.casefold()
                    or name.casefold().startswith("xl/externallinks/")
                    or name.casefold().startswith("xl/embeddings/")
                )
            ]
            if unsafe:
                raise RoundtripWorkbookError(
                    "工作簿包含宏、外部链接或嵌入对象，拒绝导入"
                )
            _assert_roundtrip_xml_envelope(archive)
    except RoundtripWorkbookError:
        raise
    except reader.ReaderError as exc:
        raise RoundtripWorkbookError(str(exc)) from exc
    except Exception as exc:
        raise RoundtripWorkbookError("文件不是有效的 .xlsx 工作簿") from exc


def _metadata_from_workbook(workbook) -> dict[str, str]:
    ws = workbook["99_元数据"]
    table = ws.tables.get("tbl_metadata_v1")
    if table is None:
        raise RoundtripWorkbookError("99_元数据缺少 tbl_metadata_v1")
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if (min_col, min_row, max_col) != (1, 1, 2):
        raise RoundtripWorkbookError("99_元数据表结构无效")
    if max_row - 1 > MAX_METADATA_ROWS:
        raise RoundtripWorkbookError(
            f"99_元数据超过 {MAX_METADATA_ROWS} 行安全上限",
            status_code=413,
        )
    metadata: dict[str, str] = {}
    for row in range(2, max_row + 1):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if not key:
            continue
        key = str(key).strip()
        if len(key) > 128:
            raise RoundtripWorkbookError("99_元数据键名超过 128 字符")
        if key in metadata:
            raise RoundtripWorkbookError(f"99_元数据存在重复键：{key}")
        normalized = "" if value is None else str(value)
        if len(normalized) > MAX_CELL_CHARS:
            raise RoundtripWorkbookError("99_元数据值超过 Excel 文本上限")
        metadata[key] = normalized

    required = {
        "protocol_id",
        "schema_version",
        "export_id",
        "exported_at",
        "app_version",
        "date_from",
        "date_to",
        "as_of",
        "contract_scope",
        "contract_revision_count",
        "contract_revisions_sha256",
        "tax_rate",
        "amount_basis",
        "table_map",
        "template_mode",
        "metadata_hmac",
    }
    missing = sorted(required - metadata.keys())
    if missing:
        raise RoundtripWorkbookError(f"99_元数据缺少字段：{', '.join(missing)}")
    if metadata["protocol_id"] != PROTOCOL_ID:
        raise RoundtripWorkbookError("未知工作簿协议，必须重新从系统导出模板")
    if metadata["schema_version"] != SCHEMA_VERSION:
        raise RoundtripWorkbookError("不兼容的工作簿版本，必须重新从系统导出模板")
    if metadata["tax_rate"] != str(TAX_RATE) or metadata["amount_basis"] != "ex":
        raise RoundtripWorkbookError("工作簿税率或金额权威口径被修改")
    if metadata["template_mode"] not in {"snapshot", "blank"}:
        raise RoundtripWorkbookError("工作簿模板模式无效")
    expected_hmac = _metadata_hmac(metadata)
    if not hmac.compare_digest(metadata["metadata_hmac"], expected_hmac):
        raise RoundtripWorkbookError("工作簿元数据签名无效，请重新导出")
    try:
        uuid.UUID(metadata["export_id"])
        table_map = json.loads(metadata["table_map"])
        date.fromisoformat(metadata["as_of"])
        revision_count = int(metadata["contract_revision_count"])
    except (ValueError, TypeError, json.JSONDecodeError) as exc:
        raise RoundtripWorkbookError("工作簿元数据格式无效") from exc
    if not 0 <= revision_count <= MAX_CONTRACT_REVISIONS:
        raise RoundtripWorkbookError("合同 revision 数量元数据超出安全上限")
    if re.fullmatch(r"[0-9a-f]{64}", metadata["contract_revisions_sha256"]) is None:
        raise RoundtripWorkbookError("合同 revision 摘要格式无效")
    scope = metadata["contract_scope"]
    if len(scope) > 64:
        raise RoundtripWorkbookError("合同范围超过 64 字符")
    expected_table_map = {sheet: spec[0] for sheet, spec in _TABLE_SPECS.items()}
    if table_map != expected_table_map:
        raise RoundtripWorkbookError("工作簿数据表映射被修改")
    if bool(metadata["date_from"]) != bool(metadata["date_to"]):
        raise RoundtripWorkbookError("工作簿日期范围元数据不完整")
    if metadata["date_from"]:
        try:
            start = date.fromisoformat(metadata["date_from"])
            end = date.fromisoformat(metadata["date_to"])
        except ValueError as exc:
            raise RoundtripWorkbookError("工作簿日期范围元数据无效") from exc
        if start > end:
            raise RoundtripWorkbookError("工作簿日期范围元数据无效")
    return metadata


def _contract_revisions_from_workbook(
    workbook,
    metadata: dict[str, str],
) -> dict[str, int]:
    ws = workbook["99_合同版本"]
    if set(ws.tables) != {_CONTRACT_REVISION_TABLE}:
        raise RoundtripWorkbookError("99_合同版本的数据表名称或数量不符合协议")
    table = ws.tables[_CONTRACT_REVISION_TABLE]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if (min_col, min_row, max_col) != (1, 1, 2):
        raise RoundtripWorkbookError("99_合同版本表结构无效")
    if max_row - 1 > MAX_CONTRACT_REVISIONS:
        raise RoundtripWorkbookError(
            f"合同 revision 超过 {MAX_CONTRACT_REVISIONS} 条安全上限",
            status_code=413,
        )
    headers = tuple(
        str(ws.cell(row=1, column=column).value or "") for column in range(1, 3)
    )
    if headers != _CONTRACT_REVISION_HEADERS:
        raise RoundtripWorkbookError("99_合同版本表头被修改")

    revisions: dict[str, int] = {}
    for row_number in range(2, max_row + 1):
        contract_value = ws.cell(row=row_number, column=1).value
        revision_value = ws.cell(row=row_number, column=2).value
        if contract_value in {None, ""} and revision_value in {None, ""}:
            continue
        contract = str(contract_value or "").strip()
        if not contract or len(contract) > 64:
            raise RoundtripWorkbookError(
                f"99_合同版本第 {row_number} 行合同号为空或超过 64 字符"
            )
        if contract in revisions:
            raise RoundtripWorkbookError(f"99_合同版本存在重复合同号：{contract}")
        if isinstance(revision_value, bool):
            raise RoundtripWorkbookError(
                f"99_合同版本第 {row_number} 行 revision 必须是整数"
            )
        try:
            revision_decimal = Decimal(str(revision_value))
        except InvalidOperation as exc:
            raise RoundtripWorkbookError(
                f"99_合同版本第 {row_number} 行 revision 必须是整数"
            ) from exc
        if (
            not revision_decimal.is_finite()
            or revision_decimal != revision_decimal.to_integral_value()
            or revision_decimal < 0
            or revision_decimal > 2_147_483_647
        ):
            raise RoundtripWorkbookError(
                f"99_合同版本第 {row_number} 行 revision 超出允许范围"
            )
        revisions[contract] = int(revision_decimal)

    expected_count = int(metadata["contract_revision_count"])
    if len(revisions) != expected_count:
        raise RoundtripWorkbookError("合同 revision 数量与签名元数据不一致")
    if not hmac.compare_digest(
        _contract_revisions_digest(revisions),
        metadata["contract_revisions_sha256"],
    ):
        raise RoundtripWorkbookError("合同 revision 内容与签名元数据不一致")
    scope = metadata["contract_scope"]
    if scope and scope not in revisions:
        raise RoundtripWorkbookError("签名合同范围缺少对应 revision")
    return revisions


def _parse_table(workbook, sheet: str) -> list[_ParsedRow]:
    table_name, expected_headers = _TABLE_SPECS[sheet]
    ws = workbook[sheet]
    if set(ws.tables) != {table_name}:
        raise RoundtripWorkbookError(f"{sheet} 的 Excel Table 名称或数量不符合协议")
    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if min_row != 1 or min_col != 1 or max_col != len(expected_headers):
        raise RoundtripWorkbookError(f"{sheet} 的数据表范围不符合协议")
    if max_row - 1 > MAX_ROWS_PER_TABLE:
        raise RoundtripWorkbookError(
            f"{sheet} 超过 {MAX_ROWS_PER_TABLE} 行导入安全上限",
            status_code=413,
        )
    headers = tuple(
        str(ws.cell(row=1, column=column).value or "")
        for column in range(1, max_col + 1)
    )
    if headers != expected_headers:
        raise RoundtripWorkbookError(f"{sheet} 表头被修改，请重新导出模板")

    rows: list[_ParsedRow] = []
    for row_number in range(2, max_row + 1):
        values: dict[str, Any] = {}
        for column, header in enumerate(headers, 1):
            cell: Cell = ws.cell(row=row_number, column=column)
            if cell.data_type == "f":
                raise RoundtripWorkbookError(
                    f"{sheet} 第 {row_number} 行包含公式，导入字段只允许静态值"
                )
            value = cell.value
            if isinstance(value, str):
                if len(value) > MAX_CELL_CHARS:
                    raise RoundtripWorkbookError(
                        f"{sheet} 第 {row_number} 行“{header}”超过 Excel 文本上限"
                    )
            values[header] = value
        operation = str(values.get("操作") or "").strip().upper()
        if operation not in _ALLOWED_OPERATIONS:
            raise RoundtripWorkbookError(
                f"{sheet} 第 {row_number} 行操作无效：{operation}"
            )
        rows.append(_ParsedRow(sheet, row_number, values))
    return rows


def _load_and_parse(path: str):
    _assert_safe_workbook_package(path)
    try:
        workbook = load_workbook(path, data_only=False, keep_links=False)
    except Exception as exc:
        raise RoundtripWorkbookError("无法打开维保回填工作簿") from exc
    try:
        if workbook.sheetnames != list(SHEET_NAMES):
            raise RoundtripWorkbookError("工作表名称、顺序或数量被修改，请重新导出模板")
        if (
            workbook["98_字典"].sheet_state != "hidden"
            or workbook["99_合同版本"].sheet_state != "veryHidden"
            or workbook["99_元数据"].sheet_state != "veryHidden"
        ):
            raise RoundtripWorkbookError("系统字典或元数据工作表状态被修改")
        if getattr(workbook, "_external_links", None):
            raise RoundtripWorkbookError("工作簿包含外部链接，拒绝导入")
        for worksheet in workbook.worksheets:
            # 只遍历实际存储的单元格，避免恶意放大的 worksheet dimension 触发全区扫描。
            for cell in worksheet._cells.values():
                if cell.data_type == "f":
                    raise RoundtripWorkbookError(
                        f"{worksheet.title} 包含公式，固定回填工作簿只允许静态值"
                    )
        metadata = _metadata_from_workbook(workbook)
        revisions = _contract_revisions_from_workbook(workbook, metadata)
        rows = {sheet: _parse_table(workbook, sheet) for sheet in _TABLE_SPECS}
        return workbook, metadata, revisions, rows
    except BaseException:
        try:
            workbook.close()
        except BaseException:
            try:
                _log.warning(
                    "维保工作簿解析失败后的关闭清理失败；保留原始校验错误",
                    exc_info=True,
                )
            except BaseException:
                # 关闭与日志都只是清理动作，绝不能覆盖原始校验异常。
                pass
        raise


def _text_value(
    value: Any, *, field: str, row: _ParsedRow, limit: int, required: bool = False
) -> str | None:
    if value is None:
        result = None
    else:
        result = str(value).strip()
        if not result:
            result = None
    if required and result is None:
        raise RoundtripWorkbookError(f"{row.sheet} 第 {row.excel_row} 行缺少“{field}”")
    if result is not None and len(result) > limit:
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”超过 {limit} 字符"
        )
    return result


def _int_value(
    value: Any, *, field: str, row: _ParsedRow, required: bool = False
) -> int | None:
    if value is None or value == "":
        if required:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行缺少“{field}”"
            )
        return None
    if isinstance(value, bool):
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”必须是整数"
        )
    try:
        number = Decimal(str(value))
    except InvalidOperation as exc:
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”必须是整数"
        ) from exc
    if not number.is_finite() or number != number.to_integral_value():
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”必须是整数"
        )
    return int(number)


def _date_value(
    value: Any, *, field: str, row: _ParsedRow, required: bool = False
) -> date | None:
    if value is None or value == "":
        if required:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行缺少“{field}”"
            )
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        parsed = cleaner.parse_date(value)
    except ValueError as exc:
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”日期无效"
        ) from exc
    if parsed is None and required:
        raise RoundtripWorkbookError(f"{row.sheet} 第 {row.excel_row} 行缺少“{field}”")
    return parsed


def _decimal_value(
    value: Any,
    *,
    field: str,
    row: _ParsedRow,
    required: bool = False,
    allow_zero: bool = True,
    allow_negative: bool = False,
    quantum: Decimal = Decimal("0.01"),
) -> Decimal | None:
    if value is None or value == "":
        if required:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行缺少“{field}”"
            )
        return None
    try:
        number = Decimal(str(value))
    except InvalidOperation as exc:
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”不是有效数字"
        ) from exc
    if (
        not number.is_finite()
        or (not allow_negative and number < 0)
        or (not allow_zero and number == 0)
        or abs(number) >= Decimal("1000000000000")
    ):
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行“{field}”超出允许范围"
        )
    if quantum == tax_policy.MONEY_QUANTUM:
        return tax_policy.round_money(number)
    return number.quantize(quantum)


def _client_id(row: _ParsedRow, *, required: bool) -> str | None:
    value = _text_value(
        row.values.get("__client_row_id"),
        field="__client_row_id",
        row=row,
        limit=64,
        required=required,
    )
    if value is not None:
        try:
            value = str(uuid.UUID(value))
        except ValueError as exc:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行客户端行键不是 UUID"
            ) from exc
    return value


def _verify_existing_row(
    row: _ParsedRow,
    *,
    export_id: str,
    kind: str,
    entity: Any | None,
) -> Any:
    entity_id = _int_value(
        row.values.get("__entity_id"),
        field="__entity_id",
        row=row,
        required=True,
    )
    base_version = _int_value(
        row.values.get("__base_version"),
        field="__base_version",
        row=row,
        required=True,
    )
    token = _text_value(
        row.values.get("__row_token"),
        field="__row_token",
        row=row,
        limit=128,
        required=True,
    )
    if entity is None or entity.id != entity_id:
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行目标记录不存在",
            status_code=409,
        )
    expected = _row_token(export_id, kind, entity_id, base_version)
    if not hmac.compare_digest(token, expected):
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行令牌无效，请重新导出",
            status_code=409,
        )
    if _base_version(entity) != base_version:
        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行数据已被更新，请重新导出后合并",
            status_code=409,
        )
    return entity


def _validate_exported_row_identities(
    rows: dict[str, list[_ParsedRow]],
    *,
    export_id: str,
) -> None:
    """Verify every signed exported identity, including KEEP/no-op rows.

    This intentionally authenticates the exported entity/base version without
    comparing it to the current live version. Pending writes perform the normal
    optimistic-lock comparison later, while a ledger-backed logical replay must
    remain idempotent after its first application advanced the live version.
    """

    def verify(row: _ParsedRow, *, kind: str, entity_value: Any) -> None:
        entity_id = _int_value(
            entity_value,
            field="__entity_id",
            row=row,
            required=True,
        )
        base_version = _int_value(
            row.values.get("__base_version"),
            field="__base_version",
            row=row,
            required=True,
        )
        token = _text_value(
            row.values.get("__row_token"),
            field="__row_token",
            row=row,
            limit=128,
            required=True,
        )
        expected = _row_token(
            export_id,
            kind,
            entity_id,
            base_version,
        )
        if not hmac.compare_digest(token, expected):
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行令牌无效，请重新导出",
                status_code=409,
            )

    for sheet, kind in (
        ("02_维保订单", "order"),
        ("03_订单明细", "line"),
        ("04_报销明细", "expense"),
    ):
        for row in rows[sheet]:
            identity_values = (
                row.values.get("__entity_id"),
                row.values.get("__base_version"),
                row.values.get("__row_token"),
            )
            if all(value in {None, ""} for value in identity_values):
                continue
            verify(
                row,
                kind=kind,
                entity_value=row.values.get("__entity_id"),
            )

    for row in rows["05_人工成本回填"]:
        entity_id = row.values.get("__entity_id")
        base_version = row.values.get("__base_version")
        token = row.values.get("__row_token")
        line_id = row.values.get("维保明细ID")
        if entity_id not in {None, ""}:
            verify(row, kind="manual", entity_value=entity_id)
        elif any(value not in {None, ""} for value in (base_version, token, line_id)):
            verify(row, kind="manual-create", entity_value=line_id)


def _project_std(value: str | None) -> str | None:
    return re.sub(r"^预交付-", "", value).strip() if value else value


def _change_contracts(*values: str | None) -> frozenset[str]:
    return frozenset(value.strip() for value in values if value and value.strip())


def _validate_project_rows(rows: list[_ParsedRow]) -> None:
    for row in rows:
        if row.operation not in {"", "KEEP"}:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行项目摘要是只读区"
            )


def _validate_order_changes(
    db: Session,
    rows: list[_ParsedRow],
    *,
    export_id: str,
) -> list[_Change]:
    changes = []
    target_ids = [
        _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        for row in rows
        if row.operation not in {"", "KEEP"}
    ]
    entities = (
        {
            entity.id: entity
            for entity in db.scalars(
                select(FMaintenanceOrder).where(FMaintenanceOrder.id.in_(target_ids))
            ).all()
        }
        if target_ids
        else {}
    )
    for row in rows:
        if row.operation in {"", "KEEP"}:
            continue
        if row.operation != "UPDATE":
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行：1.0 版维保订单只支持 UPDATE"
            )
        entity_id = _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        entity = entities.get(entity_id)
        _verify_existing_row(row, export_id=export_id, kind="order", entity=entity)
        reason = _text_value(
            row.values.get("变更原因"),
            field="变更原因",
            row=row,
            limit=1000,
            required=True,
        )
        contract = _text_value(
            row.values.get("合同号"), field="合同号", row=row, limit=64
        )
        project = _text_value(
            row.values.get("项目名称"), field="项目名称", row=row, limit=256
        )
        values = {
            "order_date": _date_value(
                row.values.get("制单日期"), field="制单日期", row=row
            ),
            "linked_sales_order_no": contract,
            "project_raw": project,
            "project_std": _project_std(project),
            "maint_start": _date_value(
                row.values.get("维保开始"), field="维保开始", row=row
            ),
            "maint_end": _date_value(
                row.values.get("维保结束"), field="维保结束", row=row
            ),
            "customer_name": _text_value(
                row.values.get("客户名称"), field="客户名称", row=row, limit=256
            ),
            "end_customer": _text_value(
                row.values.get("最终客户"), field="最终客户", row=row, limit=256
            ),
            "demand_type": _text_value(
                row.values.get("需求类型"), field="需求类型", row=row, limit=16
            ),
            "business_type": _text_value(
                row.values.get("业务类型"), field="业务类型", row=row, limit=16
            ),
            "salesperson": _text_value(
                row.values.get("销售人员"), field="销售人员", row=row, limit=64
            ),
            "warehouse": _text_value(
                row.values.get("出库仓库"), field="出库仓库", row=row, limit=64
            ),
            "data_status": _text_value(
                row.values.get("数据状态"), field="数据状态", row=row, limit=16
            ),
            "reason": reason,
        }
        if (
            values["maint_start"] is not None
            and values["maint_end"] is not None
            and values["maint_start"] > values["maint_end"]
        ):
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行维保开始不能晚于维保结束"
            )
        changes.append(
            _Change(
                "order",
                row.operation,
                row,
                entity,
                values,
                _change_contracts(entity.linked_sales_order_no, contract),
                (entity.order_date, values["order_date"]),
            )
        )
    return changes


def _validate_line_changes(
    db: Session,
    rows: list[_ParsedRow],
    *,
    export_id: str,
) -> list[_Change]:
    changes = []
    target_ids = [
        _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        for row in rows
        if row.operation not in {"", "KEEP"}
    ]
    entities = (
        {
            entity.id: entity
            for entity in db.scalars(
                select(FMaintenanceLine).where(FMaintenanceLine.id.in_(target_ids))
            ).all()
        }
        if target_ids
        else {}
    )
    order_ids = {entity.order_id for entity in entities.values()}
    orders = (
        {
            order.id: order
            for order in db.scalars(
                select(FMaintenanceOrder).where(FMaintenanceOrder.id.in_(order_ids))
            ).all()
        }
        if order_ids
        else {}
    )
    for row in rows:
        if row.operation in {"", "KEEP"}:
            continue
        if row.operation != "UPDATE":
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行：1.0 版订单明细只支持 UPDATE"
            )
        entity_id = _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        entity = entities.get(entity_id)
        _verify_existing_row(row, export_id=export_id, kind="line", entity=entity)
        qty = _decimal_value(
            row.values.get("需求数量"),
            field="需求数量",
            row=row,
            quantum=Decimal("0.001"),
        )
        return_qty = _decimal_value(
            row.values.get("退货数量"),
            field="退货数量",
            row=row,
            quantum=Decimal("0.001"),
        )
        if qty is not None and return_qty is not None and return_qty > qty:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行退货数量不能大于需求数量"
            )
        reason = _text_value(
            row.values.get("变更原因"),
            field="变更原因",
            row=row,
            limit=1000,
            required=True,
        )
        order = orders.get(entity.order_id)
        changes.append(
            _Change(
                "line",
                row.operation,
                row,
                entity,
                {
                    "description": _text_value(
                        row.values.get("产品描述"),
                        field="产品描述",
                        row=row,
                        limit=32767,
                    ),
                    "qty": qty,
                    "return_qty": return_qty,
                    "serial_numbers": _text_value(
                        row.values.get("发货SN"), field="发货SN", row=row, limit=32767
                    ),
                    "reason": reason,
                },
                _change_contracts(order.linked_sales_order_no if order else None),
                (order.order_date if order else None,),
            )
        )
    return changes


def _expense_values(
    row: _ParsedRow,
    *,
    entity: FProjectExpense | None = None,
) -> dict[str, Any]:
    amount_ex = _decimal_value(
        row.values.get("未税金额"),
        field="未税金额",
        row=row,
        required=True,
        allow_negative=True,
    )
    if entity is None:
        amount = amount_ex
        amount_inc = tax_policy.inc_from_ex(amount_ex)
        tax_basis = "ex"
    else:
        tax_basis = _text_value(
            row.values.get("__tax_basis"),
            field="__tax_basis",
            row=row,
            limit=16,
            required=True,
        )
        raw_amount = _decimal_value(
            row.values.get("__raw_amount"),
            field="__raw_amount",
            row=row,
            allow_negative=True,
        )
        if (
            tax_basis not in {"default_ex", "ex", "inc"}
            or tax_basis != entity.tax_basis
            or raw_amount != entity.amount
        ):
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行报销税务审计字段与导出记录不一致，"
                "请重新导出",
                status_code=409,
            )
        # An explicitly inc-basis source can lose a cent when converted
        # inc -> ex -> inc. If the editable ex-tax amount is unchanged, retain
        # the exact audited source amount and both persisted tax values.
        if amount_ex == entity.amount_ex_tax:
            amount = entity.amount
            amount_ex = entity.amount_ex_tax
            amount_inc = entity.amount_inc_tax
        else:
            amount_inc = tax_policy.inc_from_ex(amount_ex)
            amount = amount_inc if tax_basis == "inc" else amount_ex
    return {
        "linked_sales_order_no": _text_value(
            row.values.get("合同号"),
            field="合同号",
            row=row,
            limit=64,
            required=True,
        ),
        "expense_date": _date_value(
            row.values.get("报销日期"),
            field="报销日期",
            row=row,
            required=True,
        ),
        "person": _text_value(
            row.values.get("报销人员"), field="报销人员", row=row, limit=64
        ),
        "expense_type": _text_value(
            row.values.get("报销类别"), field="报销类别", row=row, limit=64
        ),
        "fee_category": _text_value(
            row.values.get("费用分类"), field="费用分类", row=row, limit=64
        ),
        "reason": _text_value(
            row.values.get("支出事由"), field="支出事由", row=row, limit=32767
        ),
        "amount": amount,
        "amount_ex_tax": amount_ex,
        "amount_inc_tax": amount_inc,
        "tax_basis": tax_basis,
        "tax_rate_used": TAX_RATE,
        "data_status": _text_value(
            row.values.get("流程状态"), field="流程状态", row=row, limit=16
        )
        or config.MAINT_EXPENSE_ACTIVE_STATUS,
        "bxd_no": _text_value(row.values.get("单号"), field="单号", row=row, limit=64),
        "line_no": _int_value(row.values.get("序号"), field="序号", row=row),
        "change_reason": _text_value(
            row.values.get("变更原因"), field="变更原因", row=row, limit=1000
        ),
    }


def _validate_expense_changes(
    db: Session,
    rows: list[_ParsedRow],
    *,
    export_id: str,
) -> list[_Change]:
    changes = []
    seen_clients: set[str] = set()
    existing_ids = [
        _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        for row in rows
        if row.operation in {"UPDATE", "VOID"}
    ]
    entities = (
        {
            entity.id: entity
            for entity in db.scalars(
                select(FProjectExpense).where(FProjectExpense.id.in_(existing_ids))
            ).all()
        }
        if existing_ids
        else {}
    )
    create_raw_ids = {
        _expense_create_raw_line_id(
            export_id,
            _client_id(row, required=True),
        )
        for row in rows
        if row.operation == "CREATE"
    }
    existing_raw_ids = (
        set(
            db.scalars(
                select(FProjectExpense.raw_line_id).where(
                    FProjectExpense.raw_line_id.in_(create_raw_ids)
                )
            ).all()
        )
        if create_raw_ids
        else set()
    )
    for row in rows:
        if row.operation in {"", "KEEP"}:
            continue
        if row.operation == "CREATE":
            if (
                _int_value(
                    row.values.get("__entity_id"),
                    field="__entity_id",
                    row=row,
                )
                is not None
            ):
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行 CREATE 不能带实体 ID"
                )
            if any(
                row.values.get(field) not in {None, ""}
                for field in ("__tax_basis", "__raw_amount")
            ):
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行 CREATE 不能带来源税务审计字段"
                )
            client_id = _client_id(row, required=True)
            if client_id in seen_clients:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行客户端行键重复"
                )
            seen_clients.add(client_id)
            raw_line_id = _expense_create_raw_line_id(
                export_id,
                client_id,
            )
            if raw_line_id in existing_raw_ids:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行 CREATE 已经应用，请重新导出",
                    status_code=409,
                )
            values = _expense_values(row)
            values["raw_line_id"] = raw_line_id
            changes.append(
                _Change(
                    "expense",
                    "CREATE",
                    row,
                    None,
                    values,
                    _change_contracts(values["linked_sales_order_no"]),
                    (values["expense_date"],),
                )
            )
            continue

        if row.operation not in {"UPDATE", "VOID"}:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行报销操作无效"
            )
        entity_id = _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        entity = entities.get(entity_id)
        _verify_existing_row(row, export_id=export_id, kind="expense", entity=entity)
        if row.operation == "VOID":
            values = {
                "data_status": "已作废",
                "change_reason": _text_value(
                    row.values.get("变更原因"),
                    field="变更原因",
                    row=row,
                    limit=1000,
                    required=True,
                ),
            }
            target_contract = entity.linked_sales_order_no
        else:
            values = _expense_values(row, entity=entity)
            target_contract = values["linked_sales_order_no"]
        changes.append(
            _Change(
                "expense",
                row.operation,
                row,
                entity,
                values,
                _change_contracts(entity.linked_sales_order_no, target_contract),
                (
                    entity.expense_date,
                    values.get("expense_date", entity.expense_date),
                ),
            )
        )
    return changes


def _manual_values(row: _ParsedRow) -> dict[str, Any]:
    unit_ex = _decimal_value(
        row.values.get("人工未税单位成本"),
        field="人工未税单位成本",
        row=row,
        required=True,
    )
    reason = _text_value(
        row.values.get("回填原因"),
        field="回填原因",
        row=row,
        limit=4000,
        required=True,
    )
    evidence_text = _text_value(
        row.values.get("依据说明"),
        field="依据说明",
        row=row,
        limit=32767,
        required=True,
    )
    assert evidence_text is not None
    try:
        decoded_evidence = json.loads(evidence_text)
    except json.JSONDecodeError:
        evidence = {"note": evidence_text}
    else:
        evidence = (
            decoded_evidence
            if isinstance(decoded_evidence, dict)
            else {"note": evidence_text}
        )
    return {
        "unit_cost_ex_tax": unit_ex,
        "unit_cost_inc_tax": tax_policy.inc_from_ex(unit_ex),
        "tax_rate_used": TAX_RATE,
        "reason": reason,
        "evidence": evidence,
    }


def _validate_manual_changes(
    db: Session,
    rows: list[_ParsedRow],
    *,
    export_id: str,
) -> list[_Change]:
    model = _manual_model()
    changes = []
    seen_clients: set[str] = set()
    existing_ids = [
        _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        for row in rows
        if row.operation in {"UPDATE", "VOID"}
    ]
    entities = (
        {
            entity.id: entity
            for entity in db.scalars(
                select(model).where(model.id.in_(existing_ids))
            ).all()
        }
        if existing_ids
        else {}
    )
    create_line_ids = [
        _int_value(
            row.values.get("维保明细ID"),
            field="维保明细ID",
            row=row,
            required=True,
        )
        for row in rows
        if row.operation == "CREATE"
    ]
    line_ids = {
        *create_line_ids,
        *(entity.line_id for entity in entities.values()),
    }
    lines = (
        {
            line.id: line
            for line in db.scalars(
                select(FMaintenanceLine).where(FMaintenanceLine.id.in_(line_ids))
            ).all()
        }
        if line_ids
        else {}
    )
    existing_override_line_ids = (
        set(
            db.scalars(
                select(model.line_id).where(model.line_id.in_(create_line_ids))
            ).all()
        )
        if create_line_ids
        else set()
    )
    order_ids = {line.order_id for line in lines.values()}
    orders = (
        {
            order.id: order
            for order in db.scalars(
                select(FMaintenanceOrder).where(FMaintenanceOrder.id.in_(order_ids))
            ).all()
        }
        if order_ids
        else {}
    )
    for row in rows:
        if row.operation in {"", "KEEP"}:
            continue
        if row.operation == "CREATE":
            if (
                _int_value(
                    row.values.get("__entity_id"),
                    field="__entity_id",
                    row=row,
                )
                is not None
            ):
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行 CREATE 不能带实体 ID"
                )
            client_id = _client_id(row, required=True)
            if client_id in seen_clients:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行客户端行键重复"
                )
            seen_clients.add(client_id)
            line_id = _int_value(
                row.values.get("维保明细ID"),
                field="维保明细ID",
                row=row,
                required=True,
            )
            line = lines.get(line_id)
            if line is None:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行维保明细不存在",
                    status_code=409,
                )
            base = _int_value(
                row.values.get("__base_version"),
                field="__base_version",
                row=row,
                required=True,
            )
            token = _text_value(
                row.values.get("__row_token"),
                field="__row_token",
                row=row,
                limit=128,
                required=True,
            )
            expected = _row_token(export_id, "manual-create", line.id, base)
            if not hmac.compare_digest(token, expected) or _base_version(line) != base:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行维保明细已变化，请重新导出",
                    status_code=409,
                )
            if line.cost_source != "none":
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行已有自动成本，不能人工回填",
                    status_code=409,
                )
            if line.id in existing_override_line_ids:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行已有人工成本记录，请重新导出",
                    status_code=409,
                )
            order = orders.get(line.order_id)
            values = _manual_values(row)
            values.update(
                {
                    "line_id": line.id,
                    "client_id": client_id,
                }
            )
            changes.append(
                _Change(
                    "manual",
                    "CREATE",
                    row,
                    None,
                    values,
                    _change_contracts(order.linked_sales_order_no if order else None),
                    (order.order_date if order else None,),
                )
            )
            continue

        if row.operation not in {"UPDATE", "VOID"}:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行人工成本操作无效"
            )
        entity_id = _int_value(
            row.values.get("__entity_id"),
            field="__entity_id",
            row=row,
            required=True,
        )
        entity = entities.get(entity_id)
        _verify_existing_row(row, export_id=export_id, kind="manual", entity=entity)
        line = lines.get(entity.line_id) if entity is not None else None
        if line is None:
            raise RoundtripWorkbookError(
                f"{row.sheet} 第 {row.excel_row} 行维保明细不存在",
                status_code=409,
            )
        if row.operation == "UPDATE":
            if line.cost_source not in {"none", "manual"}:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行已有自动成本，不能人工覆盖",
                    status_code=409,
                )
            values = _manual_values(row)
        else:
            values = {
                "reason": _text_value(
                    row.values.get("回填原因"),
                    field="回填原因",
                    row=row,
                    limit=4000,
                    required=True,
                )
            }
        order = orders.get(line.order_id)
        changes.append(
            _Change(
                "manual",
                row.operation,
                row,
                entity,
                values,
                _change_contracts(order.linked_sales_order_no if order else None),
                (order.order_date if order else None,),
            )
        )
    return changes


def _validate_unique_targets(changes: list[_Change]) -> None:
    seen: set[tuple[str, int]] = set()
    for change in changes:
        if change.entity is None:
            continue
        key = (change.kind, int(change.entity.id))
        if key in seen:
            raise RoundtripWorkbookError(
                f"{change.row.sheet} 第 {change.row.excel_row} 行重复操作同一记录"
            )
        seen.add(key)


def _validate_unique_client_ids(changes: list[_Change]) -> None:
    seen: set[tuple[str, str]] = set()
    for change in changes:
        if change.operation != "CREATE":
            continue
        client_id = _client_id(change.row, required=True)
        key = (change.row.sheet, client_id)
        if key in seen:
            raise RoundtripWorkbookError(
                f"{change.row.sheet} 第 {change.row.excel_row} 行客户端行键在工作簿中重复"
            )
        seen.add(key)


def _canonical_operation_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (Decimal, int, float)):
        try:
            number = Decimal(str(value))
        except InvalidOperation:
            return str(value)
        if not number.is_finite():
            return str(value)
        if number == 0:
            return "0"
        return format(number.normalize(), "f")
    if isinstance(value, str):
        normalized = value.strip()
        return normalized or None
    return str(value).strip() or None


def _operation_payload_hash(row: _ParsedRow) -> str:
    headers = _OPERATION_PAYLOAD_HEADERS.get((row.sheet, row.operation), ())
    payload = {
        "protocol_id": PROTOCOL_ID,
        "schema_version": SCHEMA_VERSION,
        "sheet_code": row.sheet,
        "operation": row.operation,
        "entity_id": _canonical_operation_value(row.values.get("__entity_id")),
        "base_version": _canonical_operation_value(row.values.get("__base_version")),
        "values": {
            header: _canonical_operation_value(row.values.get(header))
            for header in headers
        },
    }
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _operation_intents(
    rows: dict[str, list[_ParsedRow]],
) -> list[_OperationIntent]:
    intents: list[_OperationIntent] = []
    seen: set[tuple[str, str]] = set()
    for sheet in (
        "02_维保订单",
        "03_订单明细",
        "04_报销明细",
        "05_人工成本回填",
    ):
        for row in rows[sheet]:
            if row.operation in {"", "KEEP"}:
                continue
            client_row_id = _client_id(row, required=True)
            key = (sheet, client_row_id)
            if key in seen:
                raise RoundtripWorkbookError(
                    f"{sheet} 第 {row.excel_row} 行客户端行键在工作表中重复"
                )
            seen.add(key)
            intents.append(
                _OperationIntent(
                    row=row,
                    client_row_id=client_row_id,
                    payload_hash=_operation_payload_hash(row),
                )
            )
    if len(intents) > MAX_OPERATIONS_PER_WORKBOOK:
        raise RoundtripWorkbookError(
            f"显式写操作超过 {MAX_OPERATIONS_PER_WORKBOOK} 条安全上限",
            status_code=413,
        )
    return intents


def _partition_operation_replays(
    db: Session,
    *,
    export_id: str,
    intents: list[_OperationIntent],
) -> tuple[list[_OperationIntent], list[_OperationIntent]]:
    if not intents:
        return [], []
    ledgers = {
        (entry.sheet_code, entry.client_row_id): entry
        for entry in db.scalars(
            select(MaintenanceRoundtripOperation).where(
                MaintenanceRoundtripOperation.export_id == export_id,
            )
        ).all()
    }
    replayed: list[_OperationIntent] = []
    pending: list[_OperationIntent] = []
    for intent in intents:
        ledger = ledgers.get((intent.row.sheet, intent.client_row_id))
        if ledger is None:
            pending.append(intent)
            continue
        if ledger.operation != intent.row.operation or not hmac.compare_digest(
            ledger.payload_hash,
            intent.payload_hash,
        ):
            raise RoundtripWorkbookError(
                f"{intent.row.sheet} 第 {intent.row.excel_row} 行相同行键已用"
                "不同内容成功应用，整本工作簿拒绝；请重新导出",
                status_code=409,
            )
        replayed.append(intent)
    return pending, replayed


def _validate_replayed_intents(
    *,
    export_id: str,
    replayed: list[_OperationIntent],
) -> None:
    """Authenticate replay rows without comparing against post-apply versions.

    A successful UPDATE advances the live entity version, so normal optimistic
    locking would reject its own retry. The matched operation-ledger payload
    fixes the original entity/base version; the row HMAC then proves that exact
    version came from this signed export. No replay row may bypass these checks.
    """
    existing_kinds = {
        ("02_维保订单", "UPDATE"): "order",
        ("03_订单明细", "UPDATE"): "line",
        ("04_报销明细", "UPDATE"): "expense",
        ("04_报销明细", "VOID"): "expense",
        ("05_人工成本回填", "UPDATE"): "manual",
        ("05_人工成本回填", "VOID"): "manual",
    }
    for intent in replayed:
        row = intent.row
        kind = existing_kinds.get((row.sheet, row.operation))
        if kind is not None:
            entity_id = _int_value(
                row.values.get("__entity_id"),
                field="__entity_id",
                row=row,
                required=True,
            )
            base_version = _int_value(
                row.values.get("__base_version"),
                field="__base_version",
                row=row,
                required=True,
            )
            token = _text_value(
                row.values.get("__row_token"),
                field="__row_token",
                row=row,
                limit=128,
                required=True,
            )
            expected = _row_token(
                export_id,
                kind,
                entity_id,
                base_version,
            )
            if not hmac.compare_digest(token, expected):
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行令牌无效，请重新导出",
                    status_code=409,
                )
            continue

        if (row.sheet, row.operation) == ("04_报销明细", "CREATE"):
            if any(
                row.values.get(field) not in {None, ""}
                for field in ("__entity_id", "__base_version", "__row_token")
            ):
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行 CREATE 技术字段结构无效",
                    status_code=409,
                )
            continue

        if (row.sheet, row.operation) == ("05_人工成本回填", "CREATE"):
            if _int_value(
                row.values.get("__entity_id"),
                field="__entity_id",
                row=row,
            ) is not None:
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行 CREATE 不能带实体 ID"
                )
            line_id = _int_value(
                row.values.get("维保明细ID"),
                field="维保明细ID",
                row=row,
                required=True,
            )
            base_version = _int_value(
                row.values.get("__base_version"),
                field="__base_version",
                row=row,
                required=True,
            )
            token = _text_value(
                row.values.get("__row_token"),
                field="__row_token",
                row=row,
                limit=128,
                required=True,
            )
            expected = _row_token(
                export_id,
                "manual-create",
                line_id,
                base_version,
            )
            if not hmac.compare_digest(token, expected):
                raise RoundtripWorkbookError(
                    f"{row.sheet} 第 {row.excel_row} 行令牌无效，请重新导出",
                    status_code=409,
                )
            continue

        raise RoundtripWorkbookError(
            f"{row.sheet} 第 {row.excel_row} 行重放操作结构无效",
            status_code=409,
        )


def _rows_without_replays(
    rows: dict[str, list[_ParsedRow]],
    replayed: list[_OperationIntent],
) -> dict[str, list[_ParsedRow]]:
    replay_keys = {(intent.row.sheet, intent.row.excel_row) for intent in replayed}
    return {
        sheet: [
            row for row in sheet_rows if (row.sheet, row.excel_row) not in replay_keys
        ]
        for sheet, sheet_rows in rows.items()
    }


def _validate_signed_scope(
    metadata: dict[str, str],
    revisions: dict[str, int],
    changes: list[_Change],
) -> None:
    explicit_contract = metadata["contract_scope"].strip()
    allowed_contracts = {explicit_contract} if explicit_contract else set(revisions)
    for change in changes:
        if not change.contracts or not change.contracts.issubset(allowed_contracts):
            raise RoundtripWorkbookError(
                f"{change.row.sheet} 第 {change.row.excel_row} 行超出签名合同范围，"
                "整本工作簿未写入",
                status_code=409,
            )

    if not metadata["date_from"]:
        return
    start = date.fromisoformat(metadata["date_from"])
    end = date.fromisoformat(metadata["date_to"])
    for change in changes:
        if not change.scope_dates or any(
            value is None or value < start or value > end
            for value in change.scope_dates
        ):
            raise RoundtripWorkbookError(
                f"{change.row.sheet} 第 {change.row.excel_row} 行超出签名日期范围"
                f" {start.isoformat()} 至 {end.isoformat()}（含边界），整本工作簿未写入",
                status_code=409,
            )


def _current_states(db: Session, contracts: set[str]) -> dict[str, Any]:
    if not contracts:
        return {}
    model = _state_model()
    return {
        state.contract_no: state
        for state in db.scalars(
            select(model).where(model.contract_no.in_(contracts))
        ).all()
    }


def _verify_contract_revisions(
    db: Session,
    exported: dict[str, int],
    changes: list[_Change],
) -> tuple[set[str], dict[str, Any]]:
    contracts = set(exported)
    for change in changes:
        contracts.update(change.contracts)
    states = _current_states(db, contracts)
    for contract in sorted(contracts):
        expected = int(exported.get(contract, 0))
        current = int(states[contract].revision) if contract in states else 0
        if current != expected:
            raise RoundtripWorkbookError(
                f"合同 {contract} 已由其他工作簿更新（当前 revision={current}，"
                f"导出 revision={expected}），请重新导出后合并",
                status_code=409,
            )
    return contracts, states


def _jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _jsonable(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_jsonable(item) for item in value]
    return value


def _snapshot(entity: Any, fields: Iterable[str]) -> dict[str, Any]:
    return {field: _jsonable(getattr(entity, field, None)) for field in fields}


def _audit(
    db: Session,
    *,
    entity_type: str,
    entity: Any,
    action: str,
    before: dict[str, Any] | None,
    after: dict[str, Any] | None,
    reason: str | None,
    operated_by: str,
) -> None:
    db.add(
        SysAuditLog(
            entity_type=entity_type,
            entity_id=int(entity.id),
            action=action,
            before_json=_jsonable(before),
            after_json=_jsonable(after),
            reason=reason,
            operated_by=operated_by,
        )
    )


def _customer_ids_for_changes(
    db: Session,
    changes: list[_Change],
) -> dict[str, int]:
    """Resolve every order customer with one SELECT and at most one batch INSERT."""
    names = sorted(
        {
            str(change.values["customer_name"])
            for change in changes
            if (change.kind == "order" and change.values.get("customer_name"))
        }
    )
    if not names:
        return {}
    customers = {
        customer.name_raw: customer
        for customer in db.scalars(
            select(DimCustomer).where(DimCustomer.name_raw.in_(names))
        ).all()
    }
    missing = [
        DimCustomer(name_raw=name, name_normalized=name)
        for name in names
        if name not in customers
    ]
    if missing:
        db.add_all(missing)
        # One flush assigns all generated customer IDs. It also prevents a later
        # order UPDATE from referring to an unresolved transient scalar FK.
        db.flush()
        customers.update({customer.name_raw: customer for customer in missing})
    return {name: customer.id for name, customer in customers.items()}


_ORDER_AUDIT_FIELDS = (
    "order_date",
    "linked_sales_order_no",
    "project_raw",
    "project_std",
    "customer_id",
    "end_customer",
    "demand_type",
    "business_type",
    "salesperson",
    "warehouse",
    "maint_start",
    "maint_end",
    "data_status",
    "import_batch_id",
)
_LINE_AUDIT_FIELDS = (
    "description",
    "qty",
    "return_qty",
    "serial_numbers",
    "import_batch_id",
)
_EXPENSE_AUDIT_FIELDS = (
    "raw_line_id",
    "bxd_no",
    "line_no",
    "data_status",
    "expense_date",
    "person",
    "expense_type",
    "fee_category",
    "reason",
    "linked_sales_order_no",
    "amount",
    "amount_ex_tax",
    "amount_inc_tax",
    "tax_basis",
    "tax_rate_used",
    "import_batch_id",
)
_MANUAL_AUDIT_FIELDS = (
    "line_id",
    "unit_cost_ex_tax",
    "unit_cost_inc_tax",
    "tax_rate_used",
    "reason",
    "evidence",
    "version",
    "active",
    "updated_by",
    "updated_at",
)


def _apply_change(
    db: Session,
    change: _Change,
    *,
    batch: SysImportBatch,
    operated_by: str,
    customer_ids: dict[str, int],
) -> _AppliedChange:
    """Prepare one business mutation without issuing per-row SQL.

    The caller flushes the full entity set once and only then creates audits and
    logical-operation ledgers, so generated IDs remain available without N+1
    ``flush()`` calls.
    """
    if change.kind == "order":
        entity = change.entity
        before = _snapshot(entity, _ORDER_AUDIT_FIELDS)
        values = dict(change.values)
        reason = values.pop("reason")
        customer_name = values.pop("customer_name")
        values["customer_id"] = customer_ids.get(customer_name)
        values["import_batch_id"] = batch.id
        for field, value in values.items():
            setattr(entity, field, value)
        return _AppliedChange(
            change=change,
            entity=entity,
            entity_type="f_maintenance_order",
            audit_action="update",
            audit_fields=_ORDER_AUDIT_FIELDS,
            before=before,
            reason=reason,
        )

    if change.kind == "line":
        entity = change.entity
        before = _snapshot(entity, _LINE_AUDIT_FIELDS)
        values = dict(change.values)
        reason = values.pop("reason")
        values["import_batch_id"] = batch.id
        for field, value in values.items():
            setattr(entity, field, value)
        return _AppliedChange(
            change=change,
            entity=entity,
            entity_type="f_maintenance_line",
            audit_action="update",
            audit_fields=_LINE_AUDIT_FIELDS,
            before=before,
            reason=reason,
        )

    if change.kind == "expense":
        values = dict(change.values)
        reason = values.pop("change_reason", None)
        if change.operation == "CREATE":
            entity = FProjectExpense(
                **values,
                import_batch_id=batch.id,
            )
            db.add(entity)
            return _AppliedChange(
                change=change,
                entity=entity,
                entity_type="f_project_expense",
                audit_action="create",
                audit_fields=_EXPENSE_AUDIT_FIELDS,
                before=None,
                reason=reason,
            )
        entity = change.entity
        before = _snapshot(entity, _EXPENSE_AUDIT_FIELDS)
        values["import_batch_id"] = batch.id
        for field, value in values.items():
            setattr(entity, field, value)
        return _AppliedChange(
            change=change,
            entity=entity,
            entity_type="f_project_expense",
            audit_action="delete" if change.operation == "VOID" else "update",
            audit_fields=_EXPENSE_AUDIT_FIELDS,
            before=before,
            reason=reason,
        )

    model = _manual_model()
    values = dict(change.values)
    client_id = values.pop("client_id", None)
    _ = client_id
    reason = values.get("reason")
    now = datetime.now(timezone.utc)
    if change.operation == "CREATE":
        entity = model(
            **values,
            version=1,
            active=True,
            updated_by=operated_by,
            updated_at=now,
        )
        db.add(entity)
        return _AppliedChange(
            change=change,
            entity=entity,
            entity_type="maintenance_manual_cost_override",
            audit_action="create",
            audit_fields=_MANUAL_AUDIT_FIELDS,
            before=None,
            reason=reason,
        )
    entity = change.entity
    before = _snapshot(entity, _MANUAL_AUDIT_FIELDS)
    if change.operation == "VOID":
        entity.active = False
        entity.reason = reason
    else:
        for field, value in values.items():
            setattr(entity, field, value)
        entity.active = True
    entity.version += 1
    entity.updated_by = operated_by
    entity.updated_at = now
    return _AppliedChange(
        change=change,
        entity=entity,
        entity_type="maintenance_manual_cost_override",
        audit_action="delete" if change.operation == "VOID" else "update",
        audit_fields=_MANUAL_AUDIT_FIELDS,
        before=before,
        reason=reason,
    )


def _record_applied_audit(
    db: Session,
    applied: _AppliedChange,
    *,
    operated_by: str,
) -> None:
    _audit(
        db,
        entity_type=applied.entity_type,
        entity=applied.entity,
        action=applied.audit_action,
        before=applied.before,
        after=_snapshot(applied.entity, applied.audit_fields),
        reason=applied.reason,
        operated_by=operated_by,
    )


def _record_operation_result(
    db: Session,
    *,
    export_id: str,
    intent: _OperationIntent,
    change: _Change,
    entity: Any,
    batch: SysImportBatch,
    operated_by: str,
) -> None:
    result = {
        "status": "applied",
        "kind": change.kind,
        "operation": change.operation,
        "entity_id": int(entity.id),
        "import_batch_id": int(batch.id),
    }
    db.add(
        MaintenanceRoundtripOperation(
            export_id=export_id,
            sheet_code=intent.row.sheet,
            client_row_id=intent.client_row_id,
            operation=intent.row.operation,
            payload_hash=intent.payload_hash,
            result_json=result,
            import_batch_id=batch.id,
            applied_by=operated_by,
        )
    )


def _update_contract_states(
    db: Session,
    *,
    contracts: set[str],
    states: dict[str, Any],
    metadata: dict[str, str],
    signed_contracts: set[str],
    batch: SysImportBatch,
    operated_by: str,
) -> None:
    model = _state_model()
    as_of = date.fromisoformat(metadata["as_of"])
    full_scope = (
        not metadata["date_from"]
        and not metadata["date_to"]
        and metadata["template_mode"] == "snapshot"
    )
    for contract in sorted(contracts):
        # 新增报销或修改订单可能把一个未出现在导出快照中的合同带入变更集。
        # 即使模板本身无日期范围，也不能据此宣称该新合同的全部报销已被快照覆盖。
        snapshot_complete = full_scope and contract in signed_contracts
        state = states.get(contract)
        if state is None:
            state = model(
                contract_no=contract,
                revision=1,
                expense_complete_through=as_of,
                expense_snapshot_complete=snapshot_complete,
                last_export_id=metadata["export_id"],
                last_import_batch_id=batch.id,
                updated_by=operated_by,
                updated_at=datetime.now(timezone.utc),
            )
            db.add(state)
        else:
            state.revision += 1
            state.expense_complete_through = as_of
            # 期间模板或未被签名导出快照覆盖的新合同都不具备全生命周期完整性。
            state.expense_snapshot_complete = snapshot_complete
            state.last_export_id = metadata["export_id"]
            state.last_import_batch_id = batch.id
            state.updated_by = operated_by
            state.updated_at = datetime.now(timezone.utc)


def _recompute_in_transaction(db: Session) -> dict:
    try:
        return maintenance_cost.recompute(db, commit=False)
    except TypeError as exc:
        if "commit" in str(exc):
            raise RuntimeError(
                "maintenance_cost.recompute 尚未提供 commit=False，不能保证回填原子性"
            ) from exc
        raise


def _log_import_cleanup_failure(operation: str) -> None:
    try:
        _log.exception(
            "维保工作簿回填的%s清理失败；保留原始处理结果",
            operation,
        )
    except BaseException:
        # Logging is diagnostic only and must never replace the import outcome.
        pass


def import_roundtrip_workbook(
    db: Session,
    path: str,
    *,
    filename: str,
    operated_by: str,
) -> dict[str, Any]:
    """完整预检后原子应用一份固定协议工作簿。

    该函数自行提交成功事务；任何异常都会回滚。返回 ``no_op=True`` 表示同一文件 hash
    已成功应用过，未创建新批次且未重算。
    """
    if not operated_by or not operated_by.strip():
        raise RoundtripWorkbookError("维保回填必须使用实名账号", status_code=401)
    workbook = None
    try:
        # Hash/lock/idempotency are deliberately ahead of openpyxl. Exact byte-for-byte
        # retries must not spend memory parsing a 10k-row workbook, and the transaction
        # lock serializes the duplicate decision with the eventual successful batch.
        file_hash = pipeline.sha256_file(path)
        db.execute(
            text("SELECT pg_advisory_xact_lock(:key)"),
            {"key": config.DATA_CHANGE_ADVISORY_LOCK_KEY},
        )
        duplicate = pipeline.successful_batch_ids_by_hash(
            db,
            {file_hash},
            file_type=ROUNDTRIP_FILE_TYPE,
        ).get(file_hash)
        if duplicate is not None:
            db.rollback()
            return {
                "status": "success",
                "no_op": True,
                "logical_replay": False,
                "batch_id": duplicate,
                "file_hash": file_hash,
            }

        workbook, metadata, revisions, rows = _load_and_parse(path)
        _validate_project_rows(rows["01_项目"])
        _validate_exported_row_identities(
            rows,
            export_id=metadata["export_id"],
        )
        intents = _operation_intents(rows)
        pending_intents, replayed_intents = _partition_operation_replays(
            db,
            export_id=metadata["export_id"],
            intents=intents,
        )
        _validate_replayed_intents(
            export_id=metadata["export_id"],
            replayed=replayed_intents,
        )
        if intents and not pending_intents:
            original_batch_id = db.scalar(
                select(MaintenanceRoundtripOperation.import_batch_id)
                .where(
                    MaintenanceRoundtripOperation.export_id == metadata["export_id"],
                )
                .order_by(MaintenanceRoundtripOperation.id)
                .limit(1)
            )
            db.rollback()
            return {
                "status": "success",
                "no_op": True,
                "logical_replay": True,
                "replayed_rows": len(replayed_intents),
                "batch_id": original_batch_id,
                "file_hash": file_hash,
            }
        pending_rows = _rows_without_replays(rows, replayed_intents)
        changes = [
            *_validate_order_changes(
                db,
                pending_rows["02_维保订单"],
                export_id=metadata["export_id"],
            ),
            *_validate_line_changes(
                db,
                pending_rows["03_订单明细"],
                export_id=metadata["export_id"],
            ),
            *_validate_expense_changes(
                db,
                pending_rows["04_报销明细"],
                export_id=metadata["export_id"],
            ),
            *_validate_manual_changes(
                db,
                pending_rows["05_人工成本回填"],
                export_id=metadata["export_id"],
            ),
        ]
        _validate_unique_targets(changes)
        _validate_unique_client_ids(changes)
        _validate_signed_scope(metadata, revisions, changes)
        contracts, states = _verify_contract_revisions(db, revisions, changes)

        batch = SysImportBatch(
            filename=(filename or "maintenance_roundtrip.xlsx")[:256],
            file_type=ROUNDTRIP_FILE_TYPE,
            file_hash=file_hash,
            uploaded_by=operated_by,
            status="processing",
        )
        db.add(batch)
        db.flush()

        counts = {
            "create": 0,
            "update": 0,
            "void": 0,
            "replay": len(replayed_intents),
            "keep": sum(
                1
                for sheet_rows in rows.values()
                for row in sheet_rows
                if row.operation in {"", "KEEP"}
            ),
        }
        pending_by_row = {
            (intent.row.sheet, intent.row.excel_row): intent
            for intent in pending_intents
        }
        customer_ids = _customer_ids_for_changes(db, changes)
        applied_changes: list[tuple[_AppliedChange, _OperationIntent]] = []
        for change in changes:
            applied = _apply_change(
                db,
                change,
                batch=batch,
                operated_by=operated_by,
                customer_ids=customer_ids,
            )
            intent = pending_by_row[(change.row.sheet, change.row.excel_row)]
            applied_changes.append((applied, intent))
            counts[change.operation.casefold()] += 1

        # One entity flush for the entire workbook: UPDATEs use executemany and
        # CREATEs use PostgreSQL insertmanyvalues/RETURNING to assign every ID.
        db.flush()
        for applied, intent in applied_changes:
            _record_applied_audit(
                db,
                applied,
                operated_by=operated_by,
            )
            _record_operation_result(
                db,
                export_id=metadata["export_id"],
                intent=intent,
                change=applied.change,
                entity=applied.entity,
                batch=batch,
                operated_by=operated_by,
            )
        # Audits and operation ledgers are also emitted as two batched INSERTs.
        db.flush()

        recompute_stats = _recompute_in_transaction(db)
        _update_contract_states(
            db,
            contracts=contracts,
            states=states,
            metadata=metadata,
            signed_contracts=set(revisions),
            batch=batch,
            operated_by=operated_by,
        )

        storage_path = pipeline._archive(path, file_hash)
        db.add(
            SysRawFile(
                batch_id=batch.id,
                filename=(filename or "maintenance_roundtrip.xlsx")[:256],
                file_hash=file_hash,
                storage_path=storage_path,
            )
        )
        changed = len(changes)
        batch.rows_total = sum(len(sheet_rows) for sheet_rows in rows.values())
        batch.rows_inserted = counts["create"]
        batch.rows_skipped = counts["keep"]
        batch.rows_error = 0
        batch.rows_inactive = counts["void"]
        signed_contracts = set(revisions)
        complete_contracts = sorted(
            contract
            for contract in contracts
            if not metadata["date_from"]
            and not metadata["date_to"]
            and metadata["template_mode"] == "snapshot"
            and contract in signed_contracts
        )
        batch.report_json = {
            "protocol_id": PROTOCOL_ID,
            "schema_version": SCHEMA_VERSION,
            "export_id": metadata["export_id"],
            "counts": counts,
            "contracts": sorted(contracts),
            "expense_snapshot_complete": (
                bool(contracts) and len(complete_contracts) == len(contracts)
            ),
            "expense_snapshot_complete_contracts": complete_contracts,
            "expense_complete_through": metadata["as_of"],
            "recompute": recompute_stats,
        }
        batch.status = "success"
        db.commit()
        return {
            "status": "success",
            "no_op": False,
            "logical_replay": False,
            "batch_id": batch.id,
            "file_hash": file_hash,
            "changed_rows": changed,
            "counts": counts,
            "contracts": sorted(contracts),
            "recompute": recompute_stats,
        }
    except BaseException:
        try:
            db.rollback()
        except BaseException:
            _log_import_cleanup_failure("事务回滚")
        raise
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except BaseException:
                _log_import_cleanup_failure("工作簿关闭")
