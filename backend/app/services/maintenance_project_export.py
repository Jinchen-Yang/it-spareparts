"""维保项目卡墙可选字段 XLSX 导出。

导出只接受 ``EXPORT_FIELDS`` 中声明的字段。客户端提交的是稳定 key，永远不会被
解释成 SQL、ORM 属性名或 Excel 公式；项目范围、筛选和财务口径复用老板看板读模型。
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import re
from typing import Callable, Literal

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.maintenance_project import MaintenanceProject
from app.security import UserContext, is_field_hidden
from app.services import maintenance_boss_board as board


MAX_EXPORT_ROWS = 5_000
MAX_DYNAMIC_TEXT_BYTES = 16 * 1024 * 1024
MAX_WORKBOOK_BYTES = 32 * 1024 * 1024
_INVALID_XML_CONTROL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]")

DEFAULT_FIELDS = (
    "project_name",
    "period_from",
    "period_to",
    "contract_nos",
    "contract_amount_inc_tax",
    "collection_received_inc_tax",
)

PermissionGroup = Literal["profit", "cost", "cost_profit"]


class ProjectExportError(Exception):
    """维保项目导出错误基类。"""


class UnknownProjectExportField(ProjectExportError):
    def __init__(self, fields: list[str]):
        self.fields = fields
        super().__init__("存在不支持的导出字段")


class ForbiddenProjectExportField(ProjectExportError):
    def __init__(self, fields: list[str]):
        self.fields = fields
        super().__init__("所选导出字段超出当前账号的数据权限")


class ProjectExportTooLarge(ProjectExportError):
    pass


class ProjectExportFilterNotPermitted(ProjectExportError):
    pass


@dataclass(frozen=True)
class ExportField:
    key: str
    label: str
    group: str
    accessor: Callable[[dict], object]
    permission: PermissionGroup | None = None
    number_format: str | None = None
    width: int = 16


def _value(row: dict, key: str):
    return row.get(key)


def _stat_value(row: dict, key: str, *, nested_key: str | None = None):
    envelope = row.get(key)
    if not isinstance(envelope, dict):
        return None
    value = envelope.get("value")
    if nested_key is not None:
        return value.get(nested_key) if isinstance(value, dict) else None
    return value


def _stat_state(row: dict, key: str):
    envelope = row.get(key)
    if not isinstance(envelope, dict):
        return None
    return {
        "ready": "完整",
        "partial": "部分数据",
        "stale": "数据较旧",
        "not_imported": "未导入",
        "restricted": "无权限",
        "error": "异常",
    }.get(envelope.get("state"), envelope.get("state"))


def _contract_amount_state(row: dict) -> str:
    envelope = row.get("contract_amount_inc_tax")
    if not isinstance(envelope, dict):
        return "无有效合同"
    state = envelope.get("state")
    value = envelope.get("value")
    if state == "restricted":
        return "无权限"
    if state == "not_imported":
        return "未导入"
    if state == "error":
        return "异常"
    if row.get("contract_incomplete"):
        return (
            "合同事实不完整（暂无已知小计）"
            if value is None
            else "合同事实不完整（已知小计）"
        )
    if value is None:
        return "无有效合同"
    return "数据较旧" if state == "stale" else "完整"


def _collection_state(row: dict) -> str:
    envelope = row.get("collection_preview_inc_tax")
    if not isinstance(envelope, dict):
        return "尚未上报"
    state = envelope.get("state")
    value = envelope.get("value")
    if state == "restricted":
        return "无权限"
    if state == "not_imported":
        return "未导入"
    if state == "error":
        return "异常"
    # 看板以 ready(None) 表达“没有 confirmed 累计快照”；导出状态必须把
    # 它写成人话，不能把空值标成“完整”，更不能落成 0。
    if value is None:
        return "尚未上报"
    if state == "partial":
        return "部分数据"
    return "数据较旧" if state == "stale" else "已上报"


def _iso_date(row: dict, key: str):
    raw = row.get(key)
    if not raw:
        return None
    if isinstance(raw, date):
        return raw
    try:
        return date.fromisoformat(str(raw))
    except ValueError:
        return str(raw)


def _period_text(row: dict) -> str:
    start = row.get("period_from")
    end = row.get("period_to")
    if start and end:
        return f"{start} 至 {end}"
    if start:
        return f"自 {start}（终止时间缺失）"
    if end:
        return f"截至 {end}（起始时间缺失）"
    return "期限缺失"


def _joined(row: dict, key: str) -> str | None:
    values = row.get(key)
    if not values:
        return None
    return "、".join(str(value) for value in values)


def _yes_no(row: dict, key: str) -> str:
    return "是" if bool(row.get(key)) else "否"


def _lifecycle(row: dict) -> str:
    return {
        "ongoing": "进行中",
        "ended": "已结束",
        "missing": "期限缺失",
    }.get(row.get("lifecycle"), str(row.get("lifecycle") or ""))


def _card_status(row: dict) -> str:
    return {
        "normal": "正常",
        "warning": "提醒",
        "alert": "报警",
        None: "数据不足",
    }.get(row.get("card_status"), str(row.get("card_status") or ""))


# 明确白名单：key、中文标题、权限组和取值函数全部由服务端固定。
# ``_master_*`` 来自 maintenance_project 主表；其余来自既有项目卡聚合结果。
EXPORT_FIELDS: tuple[ExportField, ...] = (
    ExportField("project_name", "项目名称", "项目基础", lambda r: _value(r, "display_name"), width=30),
    ExportField("project_code", "项目编号", "项目基础", lambda r: _value(r, "project_code"), width=20),
    ExportField("project_id", "项目数据库ID", "项目基础", lambda r: _value(r, "project_id"), width=38),
    ExportField("business_type", "业务类型", "项目基础", lambda r: _value(r, "_master_business_type"), width=18),
    ExportField("cmo_name", "CMO名称", "项目基础", lambda r: _value(r, "_master_cmo_name"), width=20),
    ExportField("project_manager", "维保负责人", "项目基础", lambda r: _value(r, "project_manager"), width=16),
    ExportField("project_manager_id", "维保负责人账号", "项目基础", lambda r: _value(r, "_master_project_manager_id"), width=20),
    ExportField("salesperson", "销售人员", "项目基础", lambda r: _value(r, "salesperson"), width=16),
    ExportField("no_return_default", "默认不返还", "项目基础", lambda r: _yes_no(r, "_master_no_return_default"), width=14),
    ExportField("is_active", "主档有效", "项目基础", lambda r: _yes_no(r, "_master_is_active"), width=12),
    ExportField("is_archived", "已归档", "项目基础", lambda r: _yes_no(r, "is_archived"), width=12),
    ExportField("version", "主档版本", "项目基础", lambda r: _value(r, "_master_version"), number_format="0", width=12),
    ExportField("created_at", "创建时间", "项目基础", lambda r: _value(r, "_master_created_at"), width=22),
    ExportField("updated_at", "更新时间", "项目基础", lambda r: _value(r, "_master_updated_at"), width=22),
    ExportField("period_from", "维保起始时间", "期限", lambda r: _iso_date(r, "period_from"), number_format="yyyy-mm-dd", width=16),
    ExportField("period_to", "维保终止时间", "期限", lambda r: _iso_date(r, "period_to"), number_format="yyyy-mm-dd", width=16),
    ExportField("maintenance_period", "维保期限", "期限", _period_text, width=28),
    ExportField("lifecycle", "期限状态", "期限", _lifecycle, width=14),
    ExportField("contract_nos", "销售单号（合同号）", "合同与回款", lambda r: _joined(r, "contract_nos"), width=28),
    ExportField("contract_amount_inc_tax", "合同总额（含税）", "合同与回款", lambda r: _stat_value(r, "contract_amount_inc_tax"), permission="profit", number_format='#,##0.00', width=18),
    ExportField("contract_amount_state", "合同总额数据状态", "合同与回款", _contract_amount_state, permission="profit", width=30),
    ExportField("contract_shared", "销售单跨项目共用", "合同与回款", lambda r: _yes_no(r, "contract_shared"), permission="profit", width=18),
    ExportField("contract_incomplete", "合同数据不完整", "合同与回款", lambda r: _yes_no(r, "contract_incomplete"), permission="profit", width=18),
    ExportField("collection_received_inc_tax", "累计已回款（含税）", "合同与回款", lambda r: _stat_value(r, "collection_preview_inc_tax"), permission="profit", number_format='#,##0.00', width=20),
    ExportField("collection_state", "累计回款数据状态", "合同与回款", _collection_state, permission="profit", width=18),
    ExportField("has_activity", "有维保单据", "业务统计", lambda r: _yes_no(r, "has_activity_in_window"), width=14),
    ExportField("pre_delivery_order_count", "预交付单数", "业务统计", lambda r: _value(r, "pre_delivery_order_count"), number_format="0", width=14),
    ExportField("order_count", "维保订单数", "业务统计", lambda r: _stat_value(r, "orders_ytd"), number_format="0", width=14),
    ExportField("line_count", "维保明细行数", "业务统计", lambda r: _stat_value(r, "lines_ytd"), number_format="0", width=16),
    ExportField("procured_qty", "维保备件采购数", "业务统计", lambda r: _stat_value(r, "procured_qty"), number_format='#,##0.00', width=18),
    ExportField("shipped_qty", "已发货数量", "业务统计", lambda r: _stat_value(r, "shipped_qty"), number_format='#,##0.00', width=16),
    ExportField("returned_good_qty", "已返良品数量", "业务统计", lambda r: _stat_value(r, "returned_good_qty"), number_format='#,##0.00', width=16),
    ExportField("returned_bad_qty", "已返不良品数量", "业务统计", lambda r: _stat_value(r, "returned_bad_qty"), number_format='#,##0.00', width=18),
    ExportField("known_apply_cost_inc_tax", "已知申请成本（含税）", "成本", lambda r: _stat_value(r, "known_apply_cost_inc_tax", nested_key="known_amount"), permission="cost", number_format='#,##0.00', width=20),
    ExportField("known_apply_cost_ex_tax", "已知申请成本（未税）", "成本", lambda r: _stat_value(r, "known_apply_cost_ex_tax", nested_key="known_amount"), permission="cost", number_format='#,##0.00', width=20),
    ExportField("expense_cost_inc_tax", "报销成本（含税）", "成本", lambda r: _stat_value(r, "expense_cost_inc_tax"), permission="cost", number_format='#,##0.00', width=18),
    ExportField("requisition_cost_inc_tax", "已领用成本（含税）", "成本", lambda r: _stat_value(r, "requisition_cost_inc_tax"), permission="cost", number_format='#,##0.00', width=20),
    ExportField("cost_ratio_pct", "成本率（%）", "成本", lambda r: _stat_value(r, "cost_ratio_pct"), permission="cost_profit", number_format='0.0', width=14),
    ExportField("card_status", "项目状态", "成本", _card_status, permission="cost_profit", width=14),
)

_FIELDS_BY_KEY = {field.key: field for field in EXPORT_FIELDS}


def _can_view_contract(user_ctx: UserContext) -> bool:
    """合同额/回款沿用现有字段脱敏契约，禁止用成本权限代替利润权限。"""
    return not is_field_hidden(user_ctx, "contract_amount")


def _permission_allowed(field: ExportField, user_ctx: UserContext) -> bool:
    if field.permission == "profit":
        return _can_view_contract(user_ctx)
    if field.permission == "cost":
        return board.can_view_cost(user_ctx)
    if field.permission == "cost_profit":
        return board.can_view_cost(user_ctx) and _can_view_contract(user_ctx)
    return True


def export_options(user_ctx: UserContext) -> dict:
    available = [
        field for field in EXPORT_FIELDS if _permission_allowed(field, user_ctx)
    ]
    keys = {field.key for field in available}
    defaults = [key for key in DEFAULT_FIELDS if key in keys]
    default_set = set(defaults)
    return {
        "fields": [
            {
                "key": field.key,
                "label": field.label,
                "group": field.group,
                "default_selected": field.key in default_set,
            }
            for field in available
        ],
        "default_fields": defaults,
    }


def resolve_fields(keys: list[str], user_ctx: UserContext) -> list[ExportField]:
    unknown = list(dict.fromkeys(key for key in keys if key not in _FIELDS_BY_KEY))
    if unknown:
        raise UnknownProjectExportField(unknown)
    forbidden = list(dict.fromkeys(
        key for key in keys if not _permission_allowed(_FIELDS_BY_KEY[key], user_ctx)
    ))
    if forbidden:
        raise ForbiddenProjectExportField(forbidden)
    return [_FIELDS_BY_KEY[key] for key in keys]


def _merge_master_facts(db: Session, rows: list[dict]) -> None:
    project_ids = [row["project_id"] for row in rows]
    if not project_ids:
        return
    projects = {
        project.project_id: project
        for project in db.scalars(
            select(MaintenanceProject).where(
                MaintenanceProject.project_id.in_(project_ids)
            )
        )
    }
    for row in rows:
        project = projects.get(row["project_id"])
        if project is None:
            continue
        row.update({
            "_master_business_type": project.business_type,
            "_master_cmo_name": project.cmo_name,
            "_master_project_manager_id": project.project_manager_id,
            "_master_no_return_default": project.no_return_default,
            "_master_is_active": project.is_active,
            "_master_version": project.version,
            "_master_created_at": project.created_at,
            "_master_updated_at": project.updated_at,
        })


def _safe_excel_value(value: object) -> object:
    if value is None or isinstance(value, (bool, int, float, Decimal, date)):
        if isinstance(value, datetime) and value.tzinfo is not None:
            return value.replace(tzinfo=None)
        return value
    text = _INVALID_XML_CONTROL.sub("", str(value))
    if len(text) > 32_767:
        raise ProjectExportTooLarge("导出内容超过 Excel 单元格长度上限")
    if text[:1] in ("=", "+", "-", "@", "\t", "\r", "\n"):
        text = "'" + text
    return text


def _build_xlsx(rows: list[dict], fields: list[ExportField]) -> bytes:
    workbook = Workbook(write_only=True)
    try:
        worksheet = workbook.create_sheet("维保项目清单")
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = (
            f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
        )
        for index, field in enumerate(fields, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = field.width

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header = []
        for field in fields:
            cell = WriteOnlyCell(worksheet, value=field.label)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            header.append(cell)
        worksheet.append(header)

        text_bytes = 0
        for row in rows:
            cells = []
            for field in fields:
                value = _safe_excel_value(field.accessor(row))
                if isinstance(value, str):
                    text_bytes += len(value.encode("utf-8"))
                    if text_bytes > MAX_DYNAMIC_TEXT_BYTES:
                        raise ProjectExportTooLarge(
                            "维保项目导出文本超过 16 MiB 上限"
                        )
                cell = WriteOnlyCell(worksheet, value=value)
                if field.number_format:
                    cell.number_format = field.number_format
                cells.append(cell)
            worksheet.append(cells)

        output = BytesIO()
        try:
            workbook.save(output)
            payload = output.getvalue()
        finally:
            output.close()
    finally:
        workbook.close()
    if len(payload) > MAX_WORKBOOK_BYTES:
        raise ProjectExportTooLarge("维保项目 XLSX 超过 32 MiB 上限")
    return payload


def build_project_export(
    db: Session,
    *,
    user_ctx: UserContext,
    field_keys: list[str],
    q_text: str | None,
    lifecycle: str,
    card_status: str | None,
    sort: str,
    allowed_project_ids: set[str] | None,
) -> tuple[bytes, int]:
    fields = resolve_fields(field_keys, user_ctx)
    if (card_status is not None or sort == "cost_ratio") and not (
        board.can_view_cost(user_ctx) and _can_view_contract(user_ctx)
    ):
        raise ProjectExportFilterNotPermitted()
    result = board.projects(
        db,
        user_ctx=user_ctx,
        page=1,
        page_size=MAX_EXPORT_ROWS + 1,
        lifecycle=lifecycle,
        sort=sort,
        q_text=q_text.strip() if q_text and q_text.strip() else None,
        card_status_filter=card_status,
        allowed_project_ids=allowed_project_ids,
    )
    if result["total"] > MAX_EXPORT_ROWS:
        raise ProjectExportTooLarge(
            f"符合条件的项目超过 {MAX_EXPORT_ROWS} 条，请缩小筛选范围"
        )
    # 未归属桶是看板守恒用的伪项目，不属于维保项目主表，项目清单不导出。
    rows = [
        dict(row)
        for row in result["rows"]
        if row.get("project_id") != board.UNASSIGNED_BUCKET
    ]
    _merge_master_facts(db, rows)
    return _build_xlsx(rows, fields), len(rows)
