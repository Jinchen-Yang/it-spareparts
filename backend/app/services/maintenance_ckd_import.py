"""氚云发货单（CKD）导入服务（C1a/F1）。

- parse：识别 166 列双表头（字段码行 + 字段名行），主表+明细分组展开；
- store_preview：落 raw 行，零业务写入；
- apply：仅「维保供货」行按 WBDD→稳定项目关联，把明细数量入前置库账本
  （front_stock kind=shipment_in，source_type=f_maintenance_line，幂等）；
  销售出库/采购退货不计入前置库；无法稳定关联的行进异常清单，不按名称猜。
"""
from __future__ import annotations

import hashlib
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.dimensions import DimPart
from app.models.maintenance import FMaintenanceOrder
from app.models.maintenance_project import MaintenanceProject
from app.models.maintenance_ckd_import import (
    MaintenanceCkdHeadRow,
    MaintenanceCkdImportBatch,
    MaintenanceCkdLineRow,
)
from app.models.maintenance_source_assignment import MaintenanceSourceOrderAssignment
from app.services.date_loose import parse_amount_loose, parse_date_loose
from app.services import maintenance_front_stock as front_stock

MAX_PREVIEW_BYTES = 256 * 1024 * 1024  # 真实文件含内嵌图片，允许大文件
_CKD_RE = re.compile(r"(CKD-\d{8}-\d{4})")
_WBDD_RE = re.compile(r"(WBDD-\d{8}-\d{4})")
_XSDD_RE = re.compile(r"(XSDD-\d{8}-\d{3,4})(?!\d)")

_HEAD_COLUMNS = [
    "出库单号", "出库日期", "出库类别", "出库备件/整机", "出库仓库", "仓储中心",
    "维保需求单(备件)", "维保需求单", "销售订单(备件)", "销售订单", "销售人员",
    "项目经理", "维保需求人", "备注", "数据状态",
]
_LINE_COLUMNS = [
    "备件明细.数据ID(不可修改)", "备件明细.序号", "备件明细.数据标题", "备件明细.产品名称", "备件明细.备件自贴码",
    "备件明细.备件PN", "备件明细.备件SN号", "备件明细.备件描述", "备件明细.所在仓库",
    "备件明细.所在库位", "备件明细.产品大类", "备件明细.产品小类", "备件明细.品牌",
    "备件明细.单位", "备件明细.出库数量", "备件明细.成本单价", "备件明细.成本金额",
    "备件明细.备件测试合格",
]


class CkdParseError(RuntimeError):
    """发货单文件不可解析。"""


class CkdBatchError(RuntimeError):
    """发货单批次状态错误。"""


class CkdScopeDenied(RuntimeError):
    """批次包含操作者项目范围之外的项目：整批拒绝。"""


@dataclass
class CkdHeadData:
    row_no: int
    values: dict[str, str]
    lines: list["CkdLineData"] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)


@dataclass
class CkdLineData:
    row_no: int
    values: dict[str, str]
    issues: list[str] = field(default_factory=list)


def _header_index(headers: list[str], name: str) -> int | None:
    variants = {name, f"{name}(必填)", f"{name}(不可修改)"}
    matches = [idx for idx, value in enumerate(headers, 1) if value in variants]
    if len(matches) > 1:
        raise CkdParseError(f"列「{name}」存在重复映射，拒绝解析")
    return matches[0] if matches else None


def _cell(row: tuple, index: int | None) -> str | None:
    if index is None or index > len(row):
        return None
    value = row[index - 1]
    if value is None:
        return None
    if isinstance(value, (bytes, bytearray)):
        return None
    text = str(value).strip()
    if len(text) > 4096:
        return None  # 内嵌图片 base64：不落库
    return text or None


def _non_empty(row: tuple) -> bool:
    return bool(row) and any(
        v is not None and not isinstance(v, (bytes, bytearray)) and str(v).strip()
        for v in row
    )


def parse_ckd_workbook(
    data: bytes, filename: str, *, column_aliases: dict | None = None
) -> dict:
    """解析发货单工作簿。返回 {source_kind, file_hash, heads: [CkdHeadData], line_count}。"""
    if len(data) > MAX_PREVIEW_BYTES:
        raise CkdParseError("发货单文件超过大小上限")
    from app.services import import_safety

    if len(data) > import_safety.STREAM_THRESHOLD_BYTES:
        # 大文件：流式 XML 解析，剥离内嵌图片单元格（54MB 文件普通模式需 7.5 分钟）
        rows = iter(import_safety.stream_first_sheet_rows(data))
        try:
            result = _parse_ckd_rows(rows, column_aliases)
        except Exception as exc:  # noqa: BLE001
            raise CkdParseError(f"流式解析失败：{type(exc).__name__}") from exc
        return _with_file_meta(result, data, filename)
    try:
        # 氚云导出的 sheet dimension 不可靠，read_only 模式会截断行——用普通模式。
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise CkdParseError(f"无法读取 Excel 文件：{type(exc).__name__}") from exc
    try:
        sheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        result = _parse_ckd_rows(rows, column_aliases)
    finally:
        workbook.close()
    return _with_file_meta(result, data, filename)


def _with_file_meta(result: dict, data: bytes, filename: str) -> dict:
    return {
        **result,
        "file_hash": hashlib.sha256(data).hexdigest(),
        "filename": filename,
    }


def _parse_ckd_rows(rows, column_aliases: dict | None) -> dict:
    """在行元组迭代器上执行表头识别与主明细分组（openpyxl 与流式共用）。"""
    try:
        header_rows: list[tuple] = []
        for row in rows:
            if header_rows or any(
                v is not None and str(v).startswith("F00") for v in row
            ):
                header_rows.append(row)
                if len(header_rows) >= 2:
                    break
        if len(header_rows) < 2:
            raise CkdParseError("发货单缺少字段码/字段名双表头")
        headers = [
            str(v).strip() if v is not None else "" for v in header_rows[-1]
        ]
        try:
            from app.services import import_safety as _safety

            headers = _safety.apply_column_aliases(headers, column_aliases)
        except _safety.UploadSafetyError as exc:
            raise CkdParseError(str(exc)) from exc
        head_indexes = {name: _header_index(headers, name) for name in _HEAD_COLUMNS}
        line_indexes = {name: _header_index(headers, name) for name in _LINE_COLUMNS}
        if head_indexes["出库单号"] is None:
            raise CkdParseError("发货单缺少「出库单号」列")
        if line_indexes["备件明细.备件PN"] is None:
            raise CkdParseError("发货单缺少「备件明细.备件PN」列")

        heads: list[CkdHeadData] = []
        current: CkdHeadData | None = None
        row_no = 3
        for row in rows:
            if not _non_empty(row):
                continue
            head_anchor = _cell(row, head_indexes["出库单号"])
            if head_anchor:
                current = CkdHeadData(
                    row_no=row_no,
                    values={
                        name: _cell(row, head_indexes[name]) for name in _HEAD_COLUMNS
                    },
                )
                heads.append(current)
            line_values = {
                name: _cell(row, line_indexes[name]) for name in _LINE_COLUMNS
            }
            if line_values["备件明细.备件PN"] or line_values["备件明细.出库数量"]:
                line = CkdLineData(row_no=row_no, values=line_values)
                if current is None:
                    raise CkdParseError("明细行出现在主表行之前")
                current.lines.append(line)
            row_no += 1
    except CkdParseError:
        raise
    return {
        "source_kind": "ckd_shipment_v1",
        "heads": heads,
        "line_count": sum(len(h.lines) for h in heads),
    }


def store_preview(
    db: Session, parsed: dict, operated_by: str, *, idempotency_key: str, commit: bool = True
) -> str:
    """落 raw 行并返回 batch_id。零业务写入；Idempotency-Key 重放收敛。

    commit=False 供 AI accept 使用：把 batch 落库与 proposal 状态更新放进
    同一个事务（round-5 Blocker 8），避免「batch 已提交、proposal 未更新」
    的孤儿窗口。
    """
    existing = db.execute(
        select(MaintenanceCkdImportBatch).where(
            MaintenanceCkdImportBatch.uploaded_by == operated_by,
            MaintenanceCkdImportBatch.idempotency_key == idempotency_key,
        )
    ).scalar_one_or_none()
    if existing is not None:
        if existing.file_hash != parsed["file_hash"]:
            raise CkdBatchError("同一 Idempotency-Key 对应不同文件内容，拒绝重放")
        return existing.batch_id
    batch = MaintenanceCkdImportBatch(
        batch_id=str(uuid4()),
        file_hash=parsed["file_hash"],
        filename=parsed["filename"][:255],
        idempotency_key=idempotency_key,
        uploaded_by=operated_by,
        head_rows=len(parsed["heads"]),
        line_rows=parsed["line_count"],
        status="pending",
    )
    db.add(batch)
    db.flush()
    issue_rows = 0
    for head in parsed["heads"]:
        values = head.values
        order_no = _clean(values["出库单号"], _CKD_RE)
        if not order_no:
            head.issues.append("出库单号缺失或格式异常")
        order_date, _ = parse_date_loose(values["出库日期"])
        wbdd = _clean(values["维保需求单(备件)"] or values["维保需求单"], _WBDD_RE)
        if values["出库类别"] == "维保供货" and not wbdd:
            head.issues.append("维保供货缺少维保需求单关联")
        if values["出库日期"] and order_date is None:
            head.issues.append("出库日期无法解析")
        if not values["出库日期"] and order_date is None:
            # 空日期 fail-closed：不允许回退当前时间冒充业务发生时间（round-4 Blocker 7）
            head.issues.append("出库日期缺失")
        if values["出库类别"] == "维保供货" and not values["数据状态"]:
            head.issues.append("维保供货缺少数据状态")
        if head.issues:
            issue_rows += 1
        head_row = MaintenanceCkdHeadRow(
            row_id=str(uuid4()),
            batch_id=batch.batch_id,
            row_no=head.row_no,
            order_no_raw=values["出库单号"],
            order_date_raw=values["出库日期"],
            category_raw=values["出库类别"],
            machine_or_part_raw=values["出库备件/整机"],
            warehouse_raw=values["出库仓库"],
            wh_center_raw=values["仓储中心"],
            wbdd_raw=values["维保需求单(备件)"] or values["维保需求单"],
            sales_order_raw=values["销售订单(备件)"] or values["销售订单"],
            salesperson_raw=values["销售人员"],
            project_manager_raw=values["项目经理"],
            maintainer_raw=values["维保需求人"],
            data_status_raw=values["数据状态"],
            remark_raw=values["备注"],
            order_no=order_no,
            order_date=order_date,
            category=values["出库类别"] or None,
            wbdd_no=wbdd,
            sales_order_no=_clean(
                values["销售订单(备件)"] or values["销售订单"], _XSDD_RE
            ),
            issues=head.issues,
        )
        db.add(head_row)
        db.flush()
        for line in head.lines:
            line_values = line.values
            if line.issues:
                issue_rows += 1
            pn = (line_values["备件明细.备件PN"] or "").strip() or None
            qty = parse_amount_loose(line_values["备件明细.出库数量"])
            if pn and qty is None:
                line.issues.append("出库数量无法解析")
            if not line_values["备件明细.数据ID(不可修改)"]:
                line.issues.append("明细缺少稳定数据ID，无法安全入账")
            db.add(
                MaintenanceCkdLineRow(
                    row_id=str(uuid4()),
                    batch_id=batch.batch_id,
                    head_row_id=head_row.row_id,
                    row_no=line.row_no,
                    data_id_raw=line_values["备件明细.数据ID(不可修改)"],
                    seq_raw=line_values["备件明细.序号"],
                    title_raw=line_values["备件明细.数据标题"],
                    part_name_raw=line_values["备件明细.产品名称"],
                    self_code_raw=line_values["备件明细.备件自贴码"],
                    pn_raw=line_values["备件明细.备件PN"],
                    sn_raw=line_values["备件明细.备件SN号"],
                    desc_raw=line_values["备件明细.备件描述"],
                    warehouse_raw=line_values["备件明细.所在仓库"],
                    location_raw=line_values["备件明细.所在库位"],
                    brand_raw=line_values["备件明细.品牌"],
                    category_major_raw=line_values["备件明细.产品大类"],
                    category_minor_raw=line_values["备件明细.产品小类"],
                    unit_raw=line_values["备件明细.单位"],
                    out_qty_raw=line_values["备件明细.出库数量"],
                    unit_cost_raw=line_values["备件明细.成本单价"],
                    cost_amount_raw=line_values["备件明细.成本金额"],
                    test_result_raw=line_values["备件明细.备件测试合格"],
                    pn=pn,
                    out_qty=qty,
                    unit_cost=parse_amount_loose(line_values["备件明细.成本单价"]),
                    cost_amount=parse_amount_loose(line_values["备件明细.成本金额"]),
                    issues=line.issues,
                )
            )
    batch.issue_rows = issue_rows
    batch.report_json = {
        "head_rows": batch.head_rows,
        "line_rows": batch.line_rows,
        "issue_rows": issue_rows,
    }
    if commit:
        db.commit()
    return batch.batch_id


def _clean(raw: str | None, pattern: re.Pattern) -> str | None:
    if not raw:
        return None
    match = pattern.search(raw)
    return match.group(1) if match else None


def _resolve_project_id(db: Session, wbdd_no: str) -> str | None:
    order = db.execute(
        select(FMaintenanceOrder).where(FMaintenanceOrder.order_no == wbdd_no)
    ).scalar_one_or_none()
    if order is None:
        return None
    assignment = db.execute(
        select(MaintenanceSourceOrderAssignment).where(
            MaintenanceSourceOrderAssignment.source_order_id == order.raw_order_id,
            MaintenanceSourceOrderAssignment.is_active.is_(True),
        )
    ).scalar_one_or_none()
    return assignment.project_id if assignment is not None else None


def _project_warehouse(db: Session, project_id: str) -> str:
    project = db.get(MaintenanceProject, project_id)
    return project.project_code if project is not None else project_id


def _resolve_part_id(db: Session, pn: str) -> int | None:
    part = db.execute(
        select(DimPart).where(DimPart.pn_std == pn)
    ).scalar_one_or_none()
    return part.id if part is not None else None


def apply_batch(
    db: Session,
    batch_id: str,
    operated_by: str,
    *,
    allowed_project_ids: set[str] | None = None,
) -> dict:
    """把「维保供货」明细入前置库账本。失败关闭：任何关键异常整批零写。

    allowed_project_ids=None 表示全范围（admin/boss）；非 None 时批次解析出的
    任何项目超出范围即整批拒绝（CkdScopeDenied）。
    """
    batch = db.get(MaintenanceCkdImportBatch, batch_id)
    if batch is None:
        raise CkdBatchError("发货单批次不存在")
    if batch.status == "applied":
        raise CkdBatchError("发货单批次已应用，不能重复应用")
    if batch.status == "failed":
        raise CkdBatchError("发货单批次已因异常被拒绝，需重新上传")
    bad_head_rows = db.execute(
        select(MaintenanceCkdHeadRow).where(
            MaintenanceCkdHeadRow.batch_id == batch_id,
            func.cardinality(MaintenanceCkdHeadRow.issues) > 0,
        )
    ).scalars().all()
    bad_line_rows = db.execute(
        select(MaintenanceCkdLineRow).where(
            MaintenanceCkdLineRow.batch_id == batch_id,
            func.cardinality(MaintenanceCkdLineRow.issues) > 0,
        )
    ).scalars().all()
    if bad_head_rows or bad_line_rows:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejected_rows": len(bad_head_rows) + len(bad_line_rows),
            "rejection_reason": "发货单批次存在关键异常行，整批拒绝应用",
        }
        db.commit()
        raise CkdBatchError(
            f"发货单批次存在 {len(bad_head_rows) + len(bad_line_rows)} 行关键异常，"
            "整批拒绝应用（raw 已保留）"
        )
    summary = {
        "maintenance_heads": 0,
        "applied_lines": 0,
        "skipped_lines": 0,
        "ignored_heads": 0,
    }
    head_rows = (
        db.execute(
            select(MaintenanceCkdHeadRow)
            .where(MaintenanceCkdHeadRow.batch_id == batch_id)
            .order_by(MaintenanceCkdHeadRow.row_no)
        )
        .scalars()
        .all()
    )
    # 跨批完整性：line.batch_id 与 head.batch_id 必须一致（防跨批行绕过筛查）
    cross_batch = db.execute(
        select(MaintenanceCkdLineRow)
        .join(
            MaintenanceCkdHeadRow,
            MaintenanceCkdHeadRow.row_id == MaintenanceCkdLineRow.head_row_id,
        )
        .where(
            MaintenanceCkdHeadRow.batch_id == batch_id,
            MaintenanceCkdLineRow.batch_id != batch_id,
        )
        .limit(1)
    ).first()
    if cross_batch is not None:
        raise CkdBatchError("发货单批次存在跨批明细，拒绝应用")
    failures: list[str] = []
    for head in head_rows:
        if head.category != "维保供货":
            summary["ignored_heads"] += 1
            continue
        if head.data_status_raw and head.data_status_raw != "已生效":
            # 作废/草稿不入账；未知状态视为异常（失败关闭）
            if head.data_status_raw not in ("已取消", "草稿", "作废"):
                failures.append(f"{head.order_no}: 未知数据状态 {head.data_status_raw}")
            else:
                summary["ignored_heads"] += 1
            continue
        summary["maintenance_heads"] += 1
        if head.wbdd_no is None:
            failures.append(f"{head.order_no}: 维保供货缺少维保需求单关联")
            continue
        project_id = _resolve_project_id(db, head.wbdd_no)
        if project_id is None:
            failures.append(f"{head.wbdd_no}: 无法解析到稳定项目归属")
            continue
        if allowed_project_ids is not None and project_id not in allowed_project_ids:
            raise CkdScopeDenied(
                f"发货单批次包含无权项目（{head.wbdd_no}），整批拒绝"
            )
        warehouse_name = _project_warehouse(db, project_id)
        lines = (
            db.execute(
                select(MaintenanceCkdLineRow)
                .where(MaintenanceCkdLineRow.head_row_id == head.row_id)
                .order_by(MaintenanceCkdLineRow.row_no)
            )
            .scalars()
            .all()
        )
        for line in lines:
            if line.pn is None or line.out_qty is None or line.out_qty <= 0:
                failures.append(f"{head.order_no}: 明细缺少 PN 或数量非法")
                continue
            part_id = _resolve_part_id(db, line.pn)
            if part_id is None:
                failures.append(f"{head.order_no}: 未知 PN {line.pn}")
                continue
            try:
                front_stock.apply_movement(
                    db,
                    project_id=project_id,
                    part_id=part_id,
                    kind="shipment_in",
                    source_type="ckd_shipment_line",
                    source_ref=f"ckd:{head.order_no}:{line.data_id_raw}",
                    qty=line.out_qty,
                    warehouse_name=warehouse_name,
                    occurred_at=datetime.combine(
                        head.order_date, datetime.min.time()
                    ).replace(tzinfo=timezone.utc)
                    if head.order_date
                    else None,
                    reason=f"发货单 {head.order_no} 维保供货入前置库",
                    operated_by=operated_by,
                )
                summary["applied_lines"] += 1
            except front_stock.FrontStockPayloadConflict:
                raise
            except front_stock.FrontStockError as exc:
                failures.append(f"{head.order_no}: {exc}")
    if failures:
        # 失败关闭：回滚全部已写入流水，批次标记 failed
        db.rollback()
        batch = db.get(MaintenanceCkdImportBatch, batch_id)
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejection_reason": "发货单批次存在应用异常，整批零写入",
            "failures": failures[:20],
        }
        db.commit()
        raise CkdBatchError(
            f"发货单批次应用失败 {len(failures)} 行，整批拒绝（已回滚全部写入）"
        )
    batch.status = "applied"
    batch.applied_by = operated_by
    batch.applied_at = datetime.now(timezone.utc)
    db.commit()
    return summary
