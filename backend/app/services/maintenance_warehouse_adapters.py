"""Fail-closed adapters for warehouse exports used by maintenance reconciliation.

The first worksheet has a two-row contract: row 1 is the system's immutable
internal field code and row 2 is the human business label.  Adapter selection
uses required internal-code sets.  Ordered dual-header hashes are version
evidence only; an added/moved optional column must never silently select a
different adapter.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
import io
import json
import posixpath
import re
from typing import Iterator
import zipfile
from xml.etree import ElementTree as ET

from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, from_excel

from app import config


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_DOC_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS = {"m": _MAIN_NS, "r": _DOC_REL_NS, "p": _PKG_REL_NS}
_CELL_REF = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")
_MAX_CELL_CHARS = 32_767
_MAX_REL_XML_BYTES = 4 * 1024 * 1024
_MAX_TOTAL_TEXT_BYTES = 64 * 1024 * 1024

_RETURN_PREFIX = "D107407Fd8lreq33f21ltnq5ukwjwaxb4"
_SHIPMENT_PREFIX = "D107407Fvxu6voev32rlg4pkdu6nvdc83"
_RECEIPT_PREFIX = "D107407Fh8tgyrcma4r2qm9qk8sgk3v92"


class WarehouseWorkbookError(ValueError):
    """A generic client-safe workbook rejection."""

    def __init__(self, message: str, *, code: str = "invalid_workbook") -> None:
        super().__init__(message)
        self.code = code


@dataclass(frozen=True)
class HeaderPair:
    position: int
    internal_code: str
    business_label: str


@dataclass(frozen=True)
class WarehouseAmbiguityFact:
    code: str
    field_code: str | None = None
    source_row: int | None = None
    document_source_id: str | None = None
    line_source_id: str | None = None
    value_hash: str | None = None
    candidate_refs: tuple[dict, ...] = ()


@dataclass
class WarehouseLineFact:
    source_line_id: str
    source_row: int
    line_no: int | None
    pn: str | None
    sn: str | None
    self_code: str | None
    quantity: Decimal | None
    raw_fields: dict
    raw_fingerprint: str


@dataclass
class WarehouseDocumentFact:
    document_type: str
    source_document_id: str
    document_no: str | None
    document_date: date | None
    raw_status: str | None
    normalized_status: str
    stable_refs: dict[str, str]
    raw_fields: dict
    raw_fingerprint: str
    lines: list[WarehouseLineFact] = field(default_factory=list)


@dataclass
class ParsedWarehouseWorkbook:
    adapter_key: str
    adapter_version: str
    version_state: str
    source_file_hash: str
    header_signature: str
    header_pairs: tuple[HeaderPair, ...]
    header_diff: dict
    documents: list[WarehouseDocumentFact]
    ambiguities: list[WarehouseAmbiguityFact]
    data_row_count: int


@dataclass(frozen=True)
class AdapterSpec:
    key: str
    version: str
    document_type: str
    line_prefix: str
    required_headers: frozenset[str]
    document_id_code: str | None
    document_no_code: str | None
    document_date_code: str
    status_code: str
    line_id_code: str | None
    pn_code: str
    sn_code: str
    self_code: str
    quantity_code: str
    maintenance_order_codes: tuple[str, ...]
    upstream_document_codes: tuple[str, ...] = ()
    controlled_field_codes: frozenset[str] = frozenset()


_RETURN_COMMON = frozenset({
    "F0000032", "F0000192", "F0000061", "F0000001", "Status",
    f"{_RETURN_PREFIX}.F0000031",
    f"{_RETURN_PREFIX}.F0000044",
    f"{_RETURN_PREFIX}.F0000011",
})
_SHIPMENT_REQUIRED = frozenset({
    "ObjectId", "SeqNo", "F0000001", "F0000032", "F0000061", "Status",
    "F0000151",
    f"{_SHIPMENT_PREFIX}.ObjectId",
    f"{_SHIPMENT_PREFIX}.F0000031",
    f"{_SHIPMENT_PREFIX}.F0000044",
    f"{_SHIPMENT_PREFIX}.F0000011",
})
_RECEIPT_REQUIRED = frozenset({
    "ObjectId", "SeqNo", "F0000001", "F0000032", "F0000061", "Status",
    "F0000142",
    f"{_RECEIPT_PREFIX}.ObjectId",
    f"{_RECEIPT_PREFIX}.F0000031",
    f"{_RECEIPT_PREFIX}.F0000044",
    f"{_RECEIPT_PREFIX}.F0000011",
})

ADAPTERS: tuple[AdapterSpec, ...] = (
    AdapterSpec(
        key="return", version="return_v2", document_type="return",
        line_prefix=_RETURN_PREFIX,
        required_headers=_RETURN_COMMON | {"ObjectId", "SeqNo", f"{_RETURN_PREFIX}.ObjectId"},
        document_id_code="ObjectId", document_no_code="SeqNo",
        document_date_code="F0000001", status_code="Status",
        line_id_code=f"{_RETURN_PREFIX}.ObjectId",
        pn_code=f"{_RETURN_PREFIX}.F0000031",
        sn_code=f"{_RETURN_PREFIX}.F0000044",
        self_code=f"{_RETURN_PREFIX}.F0000043",
        quantity_code=f"{_RETURN_PREFIX}.F0000011",
        maintenance_order_codes=("F0000139", "F0000156"),
        upstream_document_codes=("F0000166", "F0000165"),
        controlled_field_codes=frozenset({f"{_RETURN_PREFIX}.F0000150"}),
    ),
    AdapterSpec(
        key="return", version="return_v1", document_type="return",
        line_prefix=_RETURN_PREFIX, required_headers=_RETURN_COMMON,
        document_id_code=None, document_no_code=None,
        document_date_code="F0000001", status_code="Status", line_id_code=None,
        pn_code=f"{_RETURN_PREFIX}.F0000031",
        sn_code=f"{_RETURN_PREFIX}.F0000044",
        self_code=f"{_RETURN_PREFIX}.F0000043",
        quantity_code=f"{_RETURN_PREFIX}.F0000011",
        maintenance_order_codes=("F0000139", "F0000156"),
        upstream_document_codes=("F0000166", "F0000165"),
        controlled_field_codes=frozenset({f"{_RETURN_PREFIX}.F0000150"}),
    ),
    AdapterSpec(
        key="shipment", version="shipment_v1", document_type="shipment",
        line_prefix=_SHIPMENT_PREFIX, required_headers=_SHIPMENT_REQUIRED,
        document_id_code="ObjectId", document_no_code="SeqNo",
        document_date_code="F0000001", status_code="Status",
        line_id_code=f"{_SHIPMENT_PREFIX}.ObjectId",
        pn_code=f"{_SHIPMENT_PREFIX}.F0000031",
        sn_code=f"{_SHIPMENT_PREFIX}.F0000044",
        self_code=f"{_SHIPMENT_PREFIX}.F0000043",
        quantity_code=f"{_SHIPMENT_PREFIX}.F0000011",
        maintenance_order_codes=("F0000151", "F0000192"),
        upstream_document_codes=("F0000147", f"{_SHIPMENT_PREFIX}.F0000148"),
        controlled_field_codes=frozenset({f"{_SHIPMENT_PREFIX}.F0000150"}),
    ),
    AdapterSpec(
        key="receipt", version="receipt_v1", document_type="receipt",
        line_prefix=_RECEIPT_PREFIX, required_headers=_RECEIPT_REQUIRED,
        document_id_code="ObjectId", document_no_code="SeqNo",
        document_date_code="F0000001", status_code="Status",
        line_id_code=f"{_RECEIPT_PREFIX}.ObjectId",
        pn_code=f"{_RECEIPT_PREFIX}.F0000031",
        sn_code=f"{_RECEIPT_PREFIX}.F0000044",
        self_code=f"{_RECEIPT_PREFIX}.F0000043",
        quantity_code=f"{_RECEIPT_PREFIX}.F0000011",
        maintenance_order_codes=("F0000142",),
        upstream_document_codes=("F0000179", "F0000178", "F0000147"),
        controlled_field_codes=frozenset({f"{_RECEIPT_PREFIX}.F0000150"}),
    ),
)

_KNOWN_STATUS = {
    "已完成": "confirmed", "完成": "confirmed", "已审批": "confirmed",
    "已生效": "confirmed", "有效": "confirmed", "正常": "confirmed",
    "已结束": "confirmed", "进行中": "pending", "审批中": "pending",
    "草稿": "pending", "待审批": "pending", "处理中": "pending",
    "作废": "void", "已作废": "void", "已取消": "void", "取消": "void",
}
_KNOWN_DOCUMENT_CATEGORY = {
    "维保", "维保出库", "维保入库", "销售", "销售出库", "采购", "采购入库",
    "退货", "退货入库", "退返入库", "返库", "换货", "维修", "其他",
}
_KNOWN_ASSET_TYPE = {"备件", "整机", "备件/整机", "备件及整机"}
_KNOWN_RETURN_TYPE = {
    "备件", "整机", "好件", "坏件", "好件返库", "坏件返库", "未使用", "现场备件",
}


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _canonical_hash(value: object) -> str:
    encoded = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _approved_header_contracts(
    adapter_version: str,
) -> tuple[tuple[tuple[str, str], ...], ...]:
    """Return code-reviewed full header contracts or fail on bad server config."""

    raw_contracts = config.MAINTENANCE_WAREHOUSE_APPROVED_HEADER_CONTRACTS.get(
        adapter_version, ()
    )
    contracts: list[tuple[tuple[str, str], ...]] = []
    try:
        for raw_contract in raw_contracts:
            contract = tuple(
                (str(code), str(label)) for code, label in raw_contract
            )
            if (
                not contract
                or any(not code or len(code) > 256 or len(label) > 512 for code, label in contract)
                or len({code for code, _label in contract}) != len(contract)
            ):
                raise ValueError
            contracts.append(contract)
    except (TypeError, ValueError) as exc:
        raise WarehouseWorkbookError(
            "服务端仓库模板批准合同配置无效",
            code="invalid_server_contract",
        ) from exc
    return tuple(contracts)


def _header_diff(
    current: tuple[tuple[str, str], ...],
    contracts: tuple[tuple[tuple[str, str], ...], ...],
) -> dict:
    """Explain the nearest approved full-header contract without guessing."""

    current_signature = _canonical_hash([list(pair) for pair in current])
    if current in contracts:
        return {
            "state": "approved_exact",
            "baseline_signature": current_signature,
            "added": [],
            "removed": [],
            "moved": [],
            "label_changed": [],
        }
    if not contracts:
        return {
            "state": "approved_baseline_unavailable",
            "baseline_signature": None,
            "added": [
                {"position": index, "internal_code": code}
                for index, (code, _label) in enumerate(current, start=1)
            ],
            "removed": [],
            "moved": [],
            "label_changed": [],
        }

    def describe(contract: tuple[tuple[str, str], ...]) -> dict:
        baseline_by_code = {
            code: (index, label)
            for index, (code, label) in enumerate(contract, start=1)
        }
        current_by_code = {
            code: (index, label)
            for index, (code, label) in enumerate(current, start=1)
        }
        added = [
            {"position": index, "internal_code": code}
            for code, (index, _label) in current_by_code.items()
            if code not in baseline_by_code
        ]
        removed = [
            {"position": index, "internal_code": code}
            for code, (index, _label) in baseline_by_code.items()
            if code not in current_by_code
        ]
        moved = [
            {
                "internal_code": code,
                "from_position": baseline_by_code[code][0],
                "to_position": current_by_code[code][0],
            }
            for code in current_by_code.keys() & baseline_by_code.keys()
            if current_by_code[code][0] != baseline_by_code[code][0]
        ]
        label_changed = [
            {
                "internal_code": code,
                "position": current_by_code[code][0],
                "approved_label_hash": _canonical_hash(baseline_by_code[code][1]),
                "current_label_hash": _canonical_hash(current_by_code[code][1]),
            }
            for code in current_by_code.keys() & baseline_by_code.keys()
            if current_by_code[code][1] != baseline_by_code[code][1]
        ]
        return {
            "state": "unapproved_difference",
            "baseline_signature": _canonical_hash([list(pair) for pair in contract]),
            "added": sorted(added, key=lambda item: item["position"]),
            "removed": sorted(removed, key=lambda item: item["position"]),
            "moved": sorted(moved, key=lambda item: item["internal_code"]),
            "label_changed": sorted(
                label_changed, key=lambda item: item["internal_code"]
            ),
        }

    candidates = [describe(contract) for contract in contracts]
    return min(
        candidates,
        key=lambda item: (
            len(item["added"])
            + len(item["removed"])
            + len(item["moved"])
            + len(item["label_changed"]),
            item["baseline_signature"],
        ),
    )


def _clean_text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _bounded_text(value: object | None, limit: int, field: str) -> str | None:
    text = _clean_text(value)
    if text is not None and len(text) > limit:
        raise WarehouseWorkbookError(f"工作簿{field}超过安全长度上限", code="cell_limit")
    return text


def _safe_archive(content: bytes) -> tuple[zipfile.ZipFile, list[str]]:
    if len(content) > config.MAX_UPLOAD_MB * 1024 * 1024:
        raise WarehouseWorkbookError("工作簿超过上传安全上限", code="upload_limit")
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
        infos = archive.infolist()
    except (zipfile.BadZipFile, OSError) as exc:
        raise WarehouseWorkbookError("工作簿不是有效的 XLSX 文件") from exc
    if not infos or len(infos) > config.IMPORT_XLSX_MAX_MEMBERS:
        archive.close()
        raise WarehouseWorkbookError("工作簿 ZIP 成员数量超过安全上限", code="zip_member_limit")
    uncompressed = 0
    compressed = 0
    names: list[str] = []
    for info in infos:
        name = info.filename
        parts = name.replace("\\", "/").split("/")
        if name.startswith(("/", "\\")) or ".." in parts or "\\" in name:
            archive.close()
            raise WarehouseWorkbookError("工作簿 ZIP 路径无效", code="zip_path")
        if info.flag_bits & 0x1:
            archive.close()
            raise WarehouseWorkbookError("不接受加密工作簿", code="encrypted_archive")
        if info.compress_size and info.file_size / info.compress_size > config.IMPORT_XLSX_MAX_COMPRESSION_RATIO:
            archive.close()
            raise WarehouseWorkbookError("工作簿压缩比超过安全上限", code="zip_bomb")
        uncompressed += info.file_size
        compressed += info.compress_size
        names.append(name)
    if uncompressed > config.IMPORT_XLSX_MAX_UNCOMPRESSED_BYTES:
        archive.close()
        raise WarehouseWorkbookError("工作簿解压体积超过安全上限", code="zip_bomb")
    if compressed and uncompressed / compressed > config.IMPORT_XLSX_MAX_COMPRESSION_RATIO:
        archive.close()
        raise WarehouseWorkbookError("工作簿压缩比超过安全上限", code="zip_bomb")
    lowered = {name.lower() for name in names}
    if any(
        name.startswith("xl/externallinks/")
        or name.startswith("xl/embeddings/")
        or name.startswith("xl/oleobjects/")
        or name.startswith("xl/activex/")
        or name.endswith("vbaproject.bin")
        for name in lowered
    ):
        archive.close()
        raise WarehouseWorkbookError("工作簿含外部链接或嵌入对象", code="external_link")
    for name in names:
        if not name.lower().endswith(".rels"):
            continue
        info = archive.getinfo(name)
        if info.file_size > _MAX_REL_XML_BYTES:
            archive.close()
            raise WarehouseWorkbookError("工作簿关系文件超过安全上限", code="relationship_limit")
        payload = archive.read(name)
        if b"<!DOCTYPE" in payload[:4096].upper() or b"<!ENTITY" in payload[:4096].upper():
            archive.close()
            raise WarehouseWorkbookError("工作簿 XML 声明无效", code="unsafe_xml")
        try:
            root = ET.fromstring(payload)
        except ET.ParseError as exc:
            archive.close()
            raise WarehouseWorkbookError("工作簿关系文件无效") from exc
        for relation in root:
            if relation.attrib.get("TargetMode", "").lower() == "external":
                archive.close()
                raise WarehouseWorkbookError("工作簿含外部链接", code="external_link")
    return archive, names


def _reject_unsafe_xml_prefix(archive: zipfile.ZipFile, name: str) -> None:
    with archive.open(name) as stream:
        prefix = stream.read(4096).upper()
    if b"<!DOCTYPE" in prefix or b"<!ENTITY" in prefix:
        raise WarehouseWorkbookError("工作簿 XML 声明无效", code="unsafe_xml")


def _workbook_sheets(archive: zipfile.ZipFile) -> tuple[list[str], datetime]:
    required = {"xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
    if not required.issubset(set(archive.namelist())):
        raise WarehouseWorkbookError("工作簿缺少必要结构")
    for name in required:
        _reject_unsafe_xml_prefix(archive, name)
    try:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relations = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    except ET.ParseError as exc:
        raise WarehouseWorkbookError("工作簿结构无效") from exc
    workbook_properties = workbook.find("m:workbookPr", _NS)
    date_1904 = (
        workbook_properties.attrib.get("date1904", "0").strip().lower()
        if workbook_properties is not None
        else "0"
    )
    if date_1904 not in {"0", "1", "false", "true"}:
        raise WarehouseWorkbookError("工作簿日期系统无效")
    epoch = MAC_EPOCH if date_1904 in {"1", "true"} else WINDOWS_EPOCH
    targets = {
        rel.attrib.get("Id"): rel.attrib.get("Target", "")
        for rel in relations
    }
    sheet_nodes = workbook.findall("m:sheets/m:sheet", _NS)
    if not sheet_nodes or len(sheet_nodes) > config.IMPORT_XLSX_MAX_WORKSHEETS:
        raise WarehouseWorkbookError("工作表数量超过安全上限", code="worksheet_limit")
    paths: list[str] = []
    for sheet in sheet_nodes:
        relation_id = sheet.attrib.get(f"{{{_DOC_REL_NS}}}id")
        target = targets.get(relation_id, "")
        if not target:
            raise WarehouseWorkbookError("工作表关系无效")
        path = target.lstrip("/")
        if not path.startswith("xl/"):
            path = posixpath.normpath(posixpath.join("xl", path))
        if path not in archive.namelist() or not path.startswith("xl/worksheets/"):
            raise WarehouseWorkbookError("工作表关系无效")
        paths.append(path)
    return paths, epoch


def _shared_strings(archive: zipfile.ZipFile) -> list[str]:
    name = "xl/sharedStrings.xml"
    if name not in archive.namelist():
        return []
    _reject_unsafe_xml_prefix(archive, name)
    values: list[str] = []
    total = 0
    try:
        with archive.open(name) as stream:
            for _event, elem in ET.iterparse(stream, events=("end",)):
                if _local_name(elem.tag) != "si":
                    continue
                value = "".join(node.text or "" for node in elem.iter() if _local_name(node.tag) == "t")
                if len(value) > _MAX_CELL_CHARS:
                    raise WarehouseWorkbookError("工作簿单元格文本超过安全上限", code="cell_limit")
                total += len(value.encode("utf-8"))
                if total > _MAX_TOTAL_TEXT_BYTES:
                    raise WarehouseWorkbookError("工作簿动态文本超过安全上限", code="text_limit")
                values.append(value)
                elem.clear()
    except ET.ParseError as exc:
        raise WarehouseWorkbookError("共享字符串结构无效") from exc
    return values


def _column_index(reference: str) -> int:
    match = _CELL_REF.fullmatch(reference)
    if match is None:
        raise WarehouseWorkbookError("工作簿单元格坐标无效")
    value = 0
    for char in match.group(1):
        value = value * 26 + ord(char) - ord("A") + 1
    if value > config.IMPORT_XLSX_MAX_COLUMNS:
        raise WarehouseWorkbookError("工作簿列数超过安全上限", code="column_limit")
    return value - 1


def _cell_value(cell: ET.Element, shared: list[str]) -> object | None:
    if any(_local_name(child.tag) == "f" for child in cell):
        raise WarehouseWorkbookError("工作簿含公式，拒绝导入", code="formula")
    kind = cell.attrib.get("t", "")
    if kind == "inlineStr":
        value: object | None = "".join(
            node.text or "" for node in cell.iter() if _local_name(node.tag) == "t"
        )
    else:
        value_node = next((child for child in cell if _local_name(child.tag) == "v"), None)
        raw = value_node.text if value_node is not None else None
        if raw is None:
            return None
        if kind == "s":
            try:
                value = shared[int(raw)]
            except (ValueError, IndexError) as exc:
                raise WarehouseWorkbookError("共享字符串索引无效") from exc
        elif kind == "b":
            if raw not in {"0", "1"}:
                raise WarehouseWorkbookError("布尔单元格值无效")
            value = raw == "1"
        elif kind == "e":
            raise WarehouseWorkbookError("工作簿含错误单元格", code="cell_error")
        else:
            value = raw
    if isinstance(value, str) and len(value) > _MAX_CELL_CHARS:
        raise WarehouseWorkbookError("工作簿单元格文本超过安全上限", code="cell_limit")
    return value


def _iter_rows(
    archive: zipfile.ZipFile,
    path: str,
    shared: list[str],
    *,
    yield_values: bool,
) -> Iterator[tuple[int, dict[int, object | None]]]:
    _reject_unsafe_xml_prefix(archive, path)
    cell_count = 0
    previous_row = 0
    try:
        with archive.open(path) as stream:
            for _event, elem in ET.iterparse(stream, events=("end",)):
                if _local_name(elem.tag) != "row":
                    continue
                row_number = int(elem.attrib.get("r", previous_row + 1))
                if row_number <= previous_row:
                    raise WarehouseWorkbookError("工作簿行号重复或逆序")
                previous_row = row_number
                if row_number > config.IMPORT_MAX_ROWS + 2:
                    raise WarehouseWorkbookError("工作簿行数超过安全上限", code="row_limit")
                values: dict[int, object | None] = {}
                for cell in elem:
                    if _local_name(cell.tag) != "c":
                        continue
                    index = _column_index(cell.attrib.get("r", ""))
                    if index in values:
                        raise WarehouseWorkbookError("工作簿同一行存在重复列")
                    value = _cell_value(cell, shared)
                    if yield_values:
                        values[index] = value
                    cell_count += 1
                    if cell_count > config.IMPORT_XLSX_MAX_DECLARED_CELLS:
                        raise WarehouseWorkbookError("工作簿单元格数超过安全上限", code="cell_limit")
                if yield_values:
                    yield row_number, values
                elem.clear()
    except ET.ParseError as exc:
        raise WarehouseWorkbookError("工作表 XML 无效") from exc


def _ordered_row(values: dict[int, object | None], width: int) -> list[object | None]:
    if values and max(values) >= width:
        raise WarehouseWorkbookError("数据列没有对应的双表头")
    return [values.get(index) for index in range(width)]


def _select_adapter(internal_codes: set[str]) -> AdapterSpec:
    # A return export that contains any wide identity column belongs to the
    # wide protocol even when another identity column is missing.  Falling
    # back to the narrow adapter would discard a real document ID and replace
    # a precise missing-line blocker with two misleading missing-ID blockers.
    return_wide_identity = {"ObjectId", "SeqNo", f"{_RETURN_PREFIX}.ObjectId"}
    if _RETURN_COMMON <= internal_codes and return_wide_identity & internal_codes:
        return next(adapter for adapter in ADAPTERS if adapter.version == "return_v2")
    matches = [adapter for adapter in ADAPTERS if adapter.required_headers <= internal_codes]
    if not matches:
        raise WarehouseWorkbookError("未识别的仓库单据表头协议", code="unknown_adapter")
    largest = max(len(adapter.required_headers) for adapter in matches)
    finalists = [adapter for adapter in matches if len(adapter.required_headers) == largest]
    if len(finalists) != 1:
        raise WarehouseWorkbookError("仓库单据表头匹配存在歧义", code="ambiguous_adapter")
    return finalists[0]


def _typed_date(value: object | None, *, epoch: datetime = WINDOWS_EPOCH) -> date | None:
    text = _clean_text(value)
    if text is None:
        return None
    try:
        number = float(text)
    except ValueError:
        number = None
    if number is not None:
        try:
            converted = from_excel(number, epoch=epoch)
            return converted.date() if isinstance(converted, datetime) else converted
        except (OverflowError, ValueError):
            return None
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None


def _typed_decimal(value: object | None) -> Decimal | None:
    text = _clean_text(value)
    if text is None:
        return None
    try:
        result = Decimal(text)
    except InvalidOperation:
        return None
    if not result.is_finite() or result < 0 or result >= Decimal("1000000000000"):
        return None
    return result


def _raw_json_value(value: object | None) -> object | None:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _first_stable_ref(
    row: dict[str, object | None],
    codes: tuple[str, ...],
    *,
    source_row: int,
    document_source_id: str,
    ambiguities: list[WarehouseAmbiguityFact],
) -> str | None:
    values = {_bounded_text(row.get(code), 128, "稳定关联键") for code in codes}
    values.discard(None)
    if len(values) > 1:
        ambiguities.append(WarehouseAmbiguityFact(
            code="field_conflict", field_code="|".join(codes), source_row=source_row,
            document_source_id=document_source_id,
        ))
        return None
    return next(iter(values), None)


def _declared_fact_codes(adapter: AdapterSpec) -> set[str]:
    """Codes whose scalar values may enter typed facts for this adapter.

    Unknown full-header versions fail closed for every optional code: its
    value is replaced by a controlled marker until a reviewed contract names
    the field.  This protects attachment payloads even when a human label is
    translated, renamed, blank, or duplicated.
    """

    return {
        "F0000032",
        "F0000061",
        "F0000192",
        *adapter.required_headers,
        *adapter.maintenance_order_codes,
        *adapter.upstream_document_codes,
        *filter(None, (
            adapter.document_id_code,
            adapter.document_no_code,
            adapter.document_date_code,
            adapter.status_code,
            adapter.line_id_code,
            adapter.pn_code,
            adapter.sn_code,
            adapter.self_code,
            adapter.quantity_code,
        )),
    }


def parse_warehouse_workbook(content: bytes) -> ParsedWarehouseWorkbook:
    """Parse one XLSX without writing database state or retaining attachment values."""

    source_hash = hashlib.sha256(content).hexdigest()
    archive, _names = _safe_archive(content)
    try:
        sheet_paths, epoch = _workbook_sheets(archive)
        shared = _shared_strings(archive)
        rows = _iter_rows(archive, sheet_paths[0], shared, yield_values=True)
        try:
            _row1, internal_values = next(rows)
            _row2, business_values = next(rows)
        except StopIteration as exc:
            raise WarehouseWorkbookError("工作簿缺少两行双表头") from exc
        if (_row1, _row2) != (1, 2):
            raise WarehouseWorkbookError("工作簿双表头必须位于前两行")
        width = max(
            max(internal_values, default=-1), max(business_values, default=-1)
        ) + 1
        if width <= 0 or width > config.IMPORT_XLSX_MAX_COLUMNS:
            raise WarehouseWorkbookError("工作簿双表头为空或超过列上限")
        internal = [_clean_text(value) or "" for value in _ordered_row(internal_values, width)]
        business = [_clean_text(value) or "" for value in _ordered_row(business_values, width)]
        if any(not code for code in internal):
            raise WarehouseWorkbookError("工作簿内部编码表头不能为空")
        if any(len(code) > 256 for code in internal) or any(len(label) > 512 for label in business):
            raise WarehouseWorkbookError("工作簿双表头超过安全长度上限", code="cell_limit")
        duplicates = sorted({code for code in internal if internal.count(code) > 1})
        if duplicates:
            raise WarehouseWorkbookError("工作簿存在重复内部编码", code="duplicate_header")
        pairs = tuple(
            HeaderPair(index + 1, code, business[index])
            for index, code in enumerate(internal)
        )
        if len({(item.internal_code, item.business_label) for item in pairs}) != len(pairs):
            raise WarehouseWorkbookError("工作簿存在重复双表头", code="duplicate_header")
        signature = _canonical_hash(
            [[item.internal_code, item.business_label] for item in pairs]
        )
        adapter = _select_adapter(set(internal))
        current_contract = tuple(
            (item.internal_code, item.business_label) for item in pairs
        )
        approved_contracts = _approved_header_contracts(adapter.version)
        version_state = (
            "known" if current_contract in approved_contracts else "unknown_version"
        )
        header_diff = _header_diff(current_contract, approved_contracts)
        ambiguities: list[WarehouseAmbiguityFact] = []
        if version_state == "unknown_version":
            ambiguities.append(WarehouseAmbiguityFact(
                code="unknown_version", value_hash=signature
            ))
        controlled_codes = set(adapter.controlled_field_codes)
        if version_state == "unknown_version":
            declared_codes = _declared_fact_codes(adapter)
            controlled_codes.update(
                pair.internal_code for pair in pairs
                if pair.internal_code not in declared_codes
            )
        by_document: dict[str, WarehouseDocumentFact] = {}
        line_index: dict[tuple[str, str], WarehouseLineFact] = {}
        data_row_count = 0
        for row_number, sparse in rows:
            ordered = _ordered_row(sparse, width)
            if not any(_clean_text(value) is not None for value in ordered):
                continue
            data_row_count += 1
            row = dict(zip(internal, ordered, strict=True))
            document_source_id = (
                _bounded_text(row.get(adapter.document_id_code), 128, "单据稳定 ID")
                if adapter.document_id_code else None
            )
            line_source_id = (
                _bounded_text(row.get(adapter.line_id_code), 128, "明细稳定 ID")
                if adapter.line_id_code else None
            )
            if document_source_id is None:
                ambiguities.append(WarehouseAmbiguityFact(
                    code="missing_document_id", source_row=row_number,
                    line_source_id=line_source_id,
                ))
            if line_source_id is None:
                ambiguities.append(WarehouseAmbiguityFact(
                    code="missing_line_id", source_row=row_number,
                    document_source_id=document_source_id,
                ))

            redacted: dict[str, object | None] = {}
            for code, value in row.items():
                if _clean_text(value) is None:
                    continue
                if code in controlled_codes:
                    redacted[code] = {"controlled": True}
                    ambiguities.append(WarehouseAmbiguityFact(
                        code="controlled_attachment", field_code=code,
                        source_row=row_number, document_source_id=document_source_id,
                        line_source_id=line_source_id,
                    ))
                else:
                    redacted[code] = _raw_json_value(value)
            if document_source_id is None or line_source_id is None:
                continue

            raw_status = _bounded_text(row.get(adapter.status_code), 128, "状态")
            normalized_status = _KNOWN_STATUS.get(raw_status or "", "unknown")
            if raw_status is None or normalized_status == "unknown":
                ambiguities.append(WarehouseAmbiguityFact(
                    code="unknown_enum", field_code=adapter.status_code,
                    source_row=row_number, document_source_id=document_source_id,
                    line_source_id=line_source_id,
                    value_hash=_canonical_hash(raw_status or ""),
                ))
            enum_rules = [
                ("F0000032", _KNOWN_DOCUMENT_CATEGORY),
                ("F0000061", _KNOWN_ASSET_TYPE),
            ]
            if adapter.key == "return":
                enum_rules.append(("F0000192", _KNOWN_RETURN_TYPE))
            for enum_code, allowed_values in enum_rules:
                raw_enum = _clean_text(row.get(enum_code))
                if raw_enum is not None and raw_enum not in allowed_values:
                    ambiguities.append(WarehouseAmbiguityFact(
                        code="unknown_enum", field_code=enum_code,
                        source_row=row_number, document_source_id=document_source_id,
                        line_source_id=line_source_id,
                        value_hash=_canonical_hash(raw_enum),
                    ))
            document_date = _typed_date(
                row.get(adapter.document_date_code), epoch=epoch
            )
            if _clean_text(row.get(adapter.document_date_code)) and document_date is None:
                ambiguities.append(WarehouseAmbiguityFact(
                    code="field_conflict", field_code=adapter.document_date_code,
                    source_row=row_number, document_source_id=document_source_id,
                ))
            quantity = _typed_decimal(row.get(adapter.quantity_code))
            if _clean_text(row.get(adapter.quantity_code)) and quantity is None:
                ambiguities.append(WarehouseAmbiguityFact(
                    code="field_conflict", field_code=adapter.quantity_code,
                    source_row=row_number, document_source_id=document_source_id,
                    line_source_id=line_source_id,
                ))
            raw_header = {
                code: value for code, value in redacted.items()
                if not code.startswith(f"{adapter.line_prefix}.")
            }
            raw_line = {
                code: value for code, value in redacted.items()
                if code.startswith(f"{adapter.line_prefix}.")
            }
            maintenance_order = _first_stable_ref(
                row, adapter.maintenance_order_codes, source_row=row_number,
                document_source_id=document_source_id, ambiguities=ambiguities,
            )
            upstream = _first_stable_ref(
                row, adapter.upstream_document_codes, source_row=row_number,
                document_source_id=document_source_id, ambiguities=ambiguities,
            )
            stable_refs = {}
            if maintenance_order:
                stable_refs["maintenance_order"] = maintenance_order
            if upstream:
                stable_refs["upstream_document"] = upstream
            header_fingerprint = _canonical_hash(raw_header)
            existing = by_document.get(document_source_id)
            if existing is None:
                existing = WarehouseDocumentFact(
                    document_type=adapter.document_type,
                    source_document_id=document_source_id,
                    document_no=(
                        _bounded_text(row.get(adapter.document_no_code), 128, "单号")
                        if adapter.document_no_code else None
                    ),
                    document_date=document_date,
                    raw_status=raw_status,
                    normalized_status=normalized_status,
                    stable_refs=stable_refs,
                    raw_fields=raw_header,
                    raw_fingerprint=header_fingerprint,
                )
                by_document[document_source_id] = existing
            elif existing.raw_fingerprint != header_fingerprint:
                ambiguities.append(WarehouseAmbiguityFact(
                    code="field_conflict", field_code="document_header",
                    source_row=row_number, document_source_id=document_source_id,
                    line_source_id=line_source_id,
                ))
            line = WarehouseLineFact(
                source_line_id=line_source_id,
                source_row=row_number,
                line_no=row_number - 2,
                pn=_bounded_text(row.get(adapter.pn_code), 256, "PN"),
                sn=_bounded_text(row.get(adapter.sn_code), 256, "SN"),
                self_code=_bounded_text(row.get(adapter.self_code), 256, "自贴码"),
                quantity=quantity,
                raw_fields=raw_line,
                raw_fingerprint=_canonical_hash(raw_line),
            )
            key = (document_source_id, line_source_id)
            prior_line = line_index.get(key)
            if prior_line is None:
                line_index[key] = line
                existing.lines.append(line)
            elif prior_line.raw_fingerprint != line.raw_fingerprint:
                ambiguities.append(WarehouseAmbiguityFact(
                    code="field_conflict", field_code="document_line",
                    source_row=row_number, document_source_id=document_source_id,
                    line_source_id=line_source_id,
                ))

        # Formula/external content in hidden option sheets is rejected too.
        for path in sheet_paths[1:]:
            for _ in _iter_rows(archive, path, shared, yield_values=False):
                pass
        return ParsedWarehouseWorkbook(
            adapter_key=adapter.key,
            adapter_version=adapter.version,
            version_state=version_state,
            source_file_hash=source_hash,
            header_signature=signature,
            header_pairs=pairs,
            header_diff=header_diff,
            documents=list(by_document.values()),
            ambiguities=ambiguities,
            data_row_count=data_row_count,
        )
    finally:
        archive.close()
