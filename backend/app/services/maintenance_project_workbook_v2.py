"""项目维保四表工作簿 v2 的无数据库业务内核。

调用方先组装项目级 ``workspace`` 快照，再使用本模块导出、校验和生成应用计划。
数据库事务、权限与审计由 API 适配器负责；本模块不会导入 ORM，也不会读取默认密钥。
"""

from __future__ import annotations

import hashlib
import hmac
import io
import json
import math
import re
import uuid
import zipfile
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Mapping, Protocol, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PROTOCOL_ID = "ITDATA_MAINT_PROJECT_WORKBOOK/2.0"
SCHEMA_VERSION = "2.0"
ENDPOINT_CONTRACT = {
    "export": "GET /api/maintenance/projects/stable/{project_id}/workbook",
    "validate": "POST /api/maintenance/projects/stable/{project_id}/workbook/validate",
    "apply": "POST /api/maintenance/projects/stable/{project_id}/workbook/apply",
    "errors": "GET /api/maintenance/workbook-validations/{validation_id}/errors.xlsx",
}
VISIBLE_SHEETS = (
    "01_总览",
    "02_备件消耗",
    "03_报销单",
    "04_项目经理追踪与提醒",
)
HIDDEN_SHEETS = ("98_字典", "99_实体版本", "99_元数据")
SHEET_NAMES = VISIBLE_SHEETS + HIDDEN_SHEETS
BLANK_COLLECTION_ROWS = 50
MAX_WORKBOOK_BYTES = 64 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_ZIP_MEMBERS = 256
MAX_COMPRESSION_RATIO = 200
MAX_ROWS_PER_TABLE = 20_000
MAX_WORKSHEET_ROWS = MAX_ROWS_PER_TABLE + 200
MAX_WORKSHEET_COLUMNS = 64
MAX_DECLARED_CELLS = 500_000
MAX_CELL_CHARS = 32_767

CONTRACT_TABLE = "tbl_project_contracts_v2"
COLLECTION_TABLE = "tbl_collections_v2"
CONSUMPTION_TABLE = "tbl_consumptions_v2"
EXPENSE_TABLE = "tbl_expenses_v2"
TASK_TABLE = "tbl_tasks_v2"
ENTITY_VERSION_TABLE = "tbl_entity_versions_v2"
METADATA_TABLE = "tbl_metadata_v2"

CONTRACT_HEADERS = (
    "项目合同关系ID",
    "合同编号",
    "合同额（全部合同）",
    "原始合同状态",
    "状态映射",
    "是否计入全部合同额",
    "当前是否生效",
    "金额完整性",
    "生效日期",
    "失效日期",
    "__base_version",
)
COLLECTION_HEADERS = (
    "操作",
    "项目合同关系ID",
    "合同编号",
    "报告月份",
    "累计回款金额",
    "回款凭证号",
    "状态",
    "备注",
    "__entity_id",
    "__base_version",
    "__row_token",
    "__client_row_id",
)
CONSUMPTION_HEADERS = (
    "现场领用明细ID",
    "现场领用单号",
    "领用日期",
    "PN",
    "备件名称",
    "实际领用数量",
    "未税单位成本",
    "实际消耗成本",
    "成本完整性",
    "成本来源",
)
EXPENSE_HEADERS = (
    "报销明细ID",
    "报销单号",
    "报销日期",
    "报销人",
    "费用分类",
    "已审批金额",
    "审批状态",
    "备注",
)
TASK_HEADERS = (
    "追踪项ID",
    "类型",
    "标题",
    "截止日期",
    "状态",
    "负责人",
    "详细说明",
)

_TABLE_MAP = {
    "01_总览": (CONTRACT_TABLE, COLLECTION_TABLE),
    "02_备件消耗": (CONSUMPTION_TABLE,),
    "03_报销单": (EXPENSE_TABLE,),
    "04_项目经理追踪与提醒": (TASK_TABLE,),
}
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MONEY_FORMAT = "#,##0.00"
_DATE_FORMAT = "yyyy-mm-dd"
_HEADER_FILL = PatternFill("solid", fgColor="35506B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUBHEADER_FILL = PatternFill("solid", fgColor="E8EEF3")
_EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
_WRAP = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class WorkbookIssue:
    code: str
    message: str
    sheet: str | None = None
    row: int | None = None
    column: str | None = None


class ProjectWorkbookV2Error(ValueError):
    """A controlled workbook error suitable for an HTTP response."""

    def __init__(
        self,
        message: str,
        *,
        status_code: int = 422,
        issues: Sequence[WorkbookIssue] = (),
    ) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.issues = tuple(issues) or (WorkbookIssue("workbook_invalid", message),)


@dataclass(frozen=True)
class WorkbookExportArtifact:
    content: bytes
    preview: dict[str, int]
    export_id: str
    filename: str


@dataclass(frozen=True)
class CollectionCreate:
    operation_key: str
    client_row_id: str
    project_contract_id: str
    contract_no: str
    report_month: date
    cumulative_amount: Decimal
    voucher_no: str | None
    status: str
    remark: str | None
    payload_hash: str


@dataclass(frozen=True)
class WorkbookValidation:
    validation_id: str
    project_id: str
    export_id: str
    expected_revision: int
    file_sha256: str
    creates: tuple[CollectionCreate, ...]
    unchanged: bool
    metadata: Mapping[str, str] = field(repr=False)


@dataclass(frozen=True)
class WorkbookApplyResult:
    status: str
    created: int
    replayed: int
    validation_id: str


@dataclass(frozen=True)
class WorkbookValidationAttempt:
    valid: bool
    validation_id: str
    validation: WorkbookValidation | None
    issues: tuple[WorkbookIssue, ...]
    can_apply: bool


class ProjectWorkbookApplyRepository(Protocol):
    """API/database adapter contract.

    ``apply_collections_atomically`` must check ``expected_revision`` again and
    persist business rows, operation-ledger entries, file hash and audit log in
    one database transaction. Any exception must roll the whole transaction back.
    """

    def current_revision(self, project_id: str) -> int: ...

    def applied_file(self, file_sha256: str) -> bool: ...

    def applied_operation(self, operation_key: str) -> str | None: ...

    def apply_collections_atomically(
        self,
        validation: WorkbookValidation,
        creates: Sequence[CollectionCreate],
    ) -> None: ...


class ProjectWorkbookEndpointAdapter(ProjectWorkbookApplyRepository, Protocol):
    """Thin API adapter boundary for :data:`ENDPOINT_CONTRACT`.

    The HTTP layer must inject the HMAC key from application configuration,
    load an all-time project workspace, and persist validation plans server-side
    (never trust a client-posted plan). Export uses the existing maintenance
    export permission; validate/apply use ``action_maintenance_roundtrip_apply``
    plus the existing maintenance page/data dependencies. No v2 permission key
    is introduced here.
    """

    def load_workspace(self, project_id: str) -> Mapping[str, Any]: ...

    def save_validation(self, validation: WorkbookValidation) -> None: ...

    def load_validation(self, validation_id: str) -> WorkbookValidation | None: ...

    def save_validation_error(
        self,
        validation_id: str,
        project_id: str,
        issues: Sequence[WorkbookIssue],
        error_workbook: bytes,
    ) -> None: ...

    def load_validation_error(self, validation_id: str) -> bytes | None: ...


def _require_hmac_key(hmac_key: bytes) -> bytes:
    if not isinstance(hmac_key, bytes) or len(hmac_key) < 16:
        raise ValueError("hmac_key must be injected as at least 16 bytes")
    return hmac_key


def _safe_text(value: Any) -> Any:
    if value is None or not isinstance(value, str):
        return value
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]", "", value)
    if len(value) > MAX_CELL_CHARS:
        raise ProjectWorkbookV2Error("单元格文本超过 Excel 安全上限")
    if value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ProjectWorkbookV2Error(f"金额或数量不是有效数字：{value}") from exc
    if not number.is_finite():
        raise ProjectWorkbookV2Error("金额或数量不能是 NaN 或 Infinity")
    try:
        rendered = float(number)
    except (OverflowError, ValueError) as exc:
        raise ProjectWorkbookV2Error("金额或数量超过工作簿安全范围") from exc
    if not math.isfinite(rendered):
        raise ProjectWorkbookV2Error("金额或数量超过工作簿安全范围")
    return rendered


def _iso(value: Any) -> str:
    if isinstance(value, datetime):
        if value.hour == value.minute == value.second == value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (Decimal, float, int)) and not isinstance(value, bool):
        number = Decimal(str(value))
        if not number.is_finite():
            raise ProjectWorkbookV2Error("工作簿包含非有限数值")
        normalized = format(number.normalize(), "f")
        return "0" if normalized in {"-0", "-0.0"} else normalized
    if value is None:
        return ""
    return str(value)


def _canonical(rows: Iterable[Sequence[Any]]) -> str:
    return json.dumps(
        [[_iso(_safe_text(value)) for value in row] for row in rows],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _digest(rows_by_table: Mapping[str, Sequence[Sequence[Any]]]) -> str:
    payload = {
        table: [
            [_iso(_safe_text(value)) for value in row]
            for row in rows_by_table[table]
        ]
        for table in sorted(rows_by_table)
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _metadata_signature(metadata: Mapping[str, str], hmac_key: bytes) -> str:
    payload = json.dumps(
        {key: metadata[key] for key in sorted(metadata) if key != "metadata_hmac"},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hmac.new(hmac_key, payload, hashlib.sha256).hexdigest()


def _safe_metadata(metadata: Mapping[str, Any]) -> dict[str, str]:
    """Return exactly the strings that will be written into the metadata table."""

    return {
        str(key): str(_safe_text(str(value)))
        for key, value in metadata.items()
    }


def preview_project_workbook(workspace: Mapping[str, Any]) -> dict[str, int]:
    consumptions = list(workspace.get("consumptions") or [])
    collections = [
        row
        for row in workspace.get("collections") or []
        if str(row.get("status") or "") == "已确认"
    ]
    expenses = [
        row
        for row in workspace.get("expenses") or []
        if str(row.get("approval_status") or "") == "已审批"
    ]
    return {
        "contracts": len(workspace.get("contracts") or []),
        "collections": len(collections),
        "consumptions": len(consumptions),
        "expenses": len(expenses),
        "tasks": len(workspace.get("tasks") or []),
        "missing_cost_rows": sum(
            1
            for row in consumptions
            if row.get("unit_cost") is None or row.get("cost_amount") is None
        ),
    }


def compute_project_summary(workspace: Mapping[str, Any]) -> dict[str, Any]:
    """Compute project KPIs without double-counting monthly cumulative snapshots."""

    all_contracts = list(workspace.get("contracts") or [])
    contracts = [
        row
        for row in all_contracts
        if (
            bool(row.get("is_effective"))
            if "is_effective" in row
            else bool(row.get("included_in_total", True))
        )
    ]
    known_contract_amount = sum(
        (
            Decimal(str(row["contract_amount"]))
            for row in contracts
            if row.get("contract_amount") is not None
        ),
        Decimal("0"),
    )
    contract_amount_complete = bool(contracts) and all(
        row.get("contract_amount") is not None for row in contracts
    )
    total_contract_amount = known_contract_amount if contract_amount_complete else None
    as_of = workspace.get("as_of")
    as_of_month = _month_text(as_of) or "9999-12"
    latest: dict[str, tuple[str, Decimal]] = {}
    for row in workspace.get("collections") or []:
        if str(row.get("status") or "") != "已确认":
            continue
        relation_id = str(row.get("project_contract_id") or "")
        report_month = _month_text(row.get("report_month"))
        if not relation_id or not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", report_month):
            continue
        if report_month > as_of_month:
            continue
        amount = Decimal(str(row.get("cumulative_amount") or 0))
        previous = latest.get(relation_id)
        if previous is None or report_month > previous[0]:
            latest[relation_id] = (report_month, amount)
    confirmed = sum((item[1] for item in latest.values()), Decimal("0"))

    known_cost = Decimal("0")
    missing_cost_rows = 0
    for row in workspace.get("consumptions") or []:
        if row.get("unit_cost") is None or row.get("cost_amount") is None:
            missing_cost_rows += 1
        if row.get("cost_amount") is not None:
            known_cost += Decimal(str(row["cost_amount"]))
    approved_expense = sum(
        (
            Decimal(str(row["amount"]))
            for row in workspace.get("expenses") or []
            if str(row.get("approval_status") or "") == "已审批"
            and row.get("amount") is not None
        ),
        Decimal("0"),
    )
    known_project_cost = known_cost + approved_expense
    collection_rate = (
        confirmed / total_contract_amount
        if total_contract_amount is not None and total_contract_amount > 0
        else None
    )
    cost_rate = (
        known_project_cost / total_contract_amount
        if total_contract_amount is not None and total_contract_amount > 0
        else None
    )
    if cost_rate is None:
        cost_alert = "no_contract_amount"
    elif cost_rate > Decimal("1"):
        cost_alert = "red"
    elif cost_rate >= Decimal("0.8"):
        cost_alert = "yellow"
    elif missing_cost_rows:
        cost_alert = "incomplete"
    else:
        cost_alert = "green"
    return {
        "total_contract_amount": total_contract_amount,
        "known_contract_amount": known_contract_amount,
        "contract_amount_complete": contract_amount_complete,
        "confirmed_cumulative_collection_amount": confirmed,
        "collection_rate": collection_rate,
        "known_consumption_cost": known_cost,
        "approved_expense": approved_expense,
        "known_project_cost": known_project_cost,
        "missing_cost_rows": missing_cost_rows,
        "cost_rate_lower_bound": cost_rate,
        "cost_alert": cost_alert,
    }


def _table(
    sheet,
    *,
    name: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    start_row: int,
) -> tuple[int, int]:
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(start_row, col, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
    rendered_rows = rows or ([None] * len(headers),)
    for row_offset, row in enumerate(rendered_rows, 1):
        if len(row) != len(headers):
            raise ValueError(f"{name} row width does not match headers")
        for col, value in enumerate(row, 1):
            sheet.cell(start_row + row_offset, col, _safe_text(value))
    end_row = start_row + len(rendered_rows)
    ref = f"A{start_row}:{get_column_letter(len(headers))}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    return start_row, end_row


def _month_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    return str(value or "").strip()


def _project_rows(workspace: Mapping[str, Any], hmac_key: bytes, export_id: str):
    contracts = [
        (
            row.get("project_contract_id"),
            row.get("contract_no"),
            _number(row.get("contract_amount")),
            row.get("contract_status"),
            (
                "已映射"
                if row.get("status_mapping_state") == "mapped"
                else "待映射"
            ),
            "是" if row.get("included_in_total", True) else "否",
            (
                "是"
                if row.get("is_effective", row.get("included_in_total", True))
                else "否"
            ),
            "金额完整" if row.get("contract_amount") is not None else "缺少合同额",
            row.get("effective_from"),
            row.get("effective_to"),
            row.get("version"),
        )
        for row in workspace.get("contracts") or []
    ]
    collections = []
    for row in workspace.get("collections") or []:
        if str(row.get("status") or "") != "已确认":
            continue
        entity_id = str(row.get("collection_id") or "")
        base_version = str(row.get("version") or "")
        token_payload = f"{export_id}|collection|{entity_id}|{base_version}".encode()
        row_token = hmac.new(hmac_key, token_payload, hashlib.sha256).hexdigest()
        collections.append((
            "KEEP",
            row.get("project_contract_id"),
            row.get("contract_no"),
            _month_text(row.get("report_month")),
            _number(row.get("cumulative_amount")),
            row.get("voucher_no"),
            row.get("status"),
            row.get("remark"),
            entity_id,
            base_version,
            row_token,
            "",
        ))
    for _ in range(max(BLANK_COLLECTION_ROWS, len(contracts))):
        collections.append((None,) * 11 + (str(uuid.uuid4()),))
    return contracts, collections


def _consumption_rows(workspace: Mapping[str, Any]):
    return [(
        row.get("consumption_id"),
        row.get("issue_no"),
        row.get("issue_date"),
        row.get("part_no"),
        row.get("part_name"),
        _number(row.get("quantity")),
        _number(row.get("unit_cost")),
        _number(row.get("cost_amount")),
        row.get("cost_status") or (
            "缺少价格成本"
            if row.get("unit_cost") is None or row.get("cost_amount") is None
            else "成本完整"
        ),
        row.get("cost_source"),
    ) for row in workspace.get("consumptions") or []]


def _expense_rows(workspace: Mapping[str, Any]):
    return [(
        row.get("expense_id"),
        row.get("expense_no"),
        row.get("expense_date"),
        row.get("applicant"),
        row.get("category"),
        _number(row.get("amount")),
        row.get("approval_status"),
        row.get("remark"),
    ) for row in workspace.get("expenses") or []
       if str(row.get("approval_status") or "") == "已审批"]


def _task_rows(workspace: Mapping[str, Any]):
    return [(
        row.get("task_id"),
        row.get("task_type"),
        row.get("title"),
        row.get("due_date"),
        row.get("status"),
        row.get("owner"),
        row.get("detail"),
    ) for row in workspace.get("tasks") or []]


def _summary_rows(workspace: Mapping[str, Any]) -> tuple[tuple[str, Any], ...]:
    project = workspace["project"]
    metrics = compute_project_summary(workspace)
    return (
        ("项目编号", project.get("project_code")),
        ("项目名称", project.get("project_name")),
        ("项目经理", project.get("manager_name")),
        ("数据截至", workspace.get("as_of")),
        (
            "口径",
            "合同额为该项目全部合同额；回款为每合同每月累计快照；实际消耗为现场领用单",
        ),
        ("全部合同额（当前计入口径）", _number(metrics["total_contract_amount"])),
        ("已确认累计回款", _number(metrics["confirmed_cumulative_collection_amount"])),
        ("回款进度", _number(metrics["collection_rate"])),
        ("已知实际消耗", _number(metrics["known_consumption_cost"])),
        ("已审批报销", _number(metrics["approved_expense"])),
        ("项目实际成本（已知）", _number(metrics["known_project_cost"])),
        ("缺少价格成本行", metrics["missing_cost_rows"]),
        ("项目成本/合同额（已知下界）", _number(metrics["cost_rate_lower_bound"])),
        ("成本预警", metrics["cost_alert"]),
    )


def _summary_sheet(sheet, workspace, contracts, collections):
    summary = _summary_rows(workspace)
    for row_index, (label, value) in enumerate(summary, 1):
        sheet.cell(row_index, 1, label).fill = _SUBHEADER_FILL
        sheet.cell(row_index, 1).font = Font(bold=True)
        sheet.cell(row_index, 2, _safe_text(value))
    for row_index in (6, 7, 9, 10, 11):
        sheet.cell(row_index, 2).number_format = _MONEY_FORMAT
    for row_index in (8, 13):
        sheet.cell(row_index, 2).number_format = "0.00%"
    contract_start = len(summary) + 3
    _table(
        sheet,
        name=CONTRACT_TABLE,
        headers=CONTRACT_HEADERS,
        rows=contracts,
        start_row=contract_start,
    )
    collection_start = contract_start + max(len(contracts), 1) + 3
    sheet.cell(collection_start - 1, 1, "回款明细：既有行只保留 KEEP；请在黄色尾部行追加 CREATE")
    _, collection_end = _table(
        sheet,
        name=COLLECTION_TABLE,
        headers=COLLECTION_HEADERS,
        rows=collections,
        start_row=collection_start,
    )
    for row in range(collection_start + 1, collection_end + 1):
        is_new_row = sheet.cell(row, 9).value in (None, "")
        for col in range(1, 9):
            if is_new_row:
                sheet.cell(row, col).fill = _EDIT_FILL
                sheet.cell(row, col).protection = Protection(locked=False)
    validation = DataValidation(type="list", formula1='"KEEP,CREATE"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"A{collection_start + 1}:A{collection_end}")
    for col in range(9, 13):
        sheet.column_dimensions[get_column_letter(col)].hidden = True
    sheet.freeze_panes = f"A{collection_start + 1}"
    sheet.protection.sheet = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = True


def _entity_versions(workspace: Mapping[str, Any]) -> list[tuple[str, str, str]]:
    rows = [(
        "project",
        str(workspace["project"]["project_id"]),
        str(workspace["project"].get("version") or ""),
    )]
    rows.extend(
        (
            "project_contract",
            str(item.get("project_contract_id") or ""),
            str(item.get("version") or ""),
        )
        for item in workspace.get("contracts") or []
    )
    rows.extend(
        ("collection", str(item.get("collection_id") or ""), str(item.get("version") or ""))
        for item in workspace.get("collections") or []
        if str(item.get("status") or "") == "已确认"
    )
    return rows


def _metadata_sheet(book: Workbook, metadata: Mapping[str, str]) -> None:
    sheet = book.create_sheet("99_元数据")
    rows = [(key, metadata[key]) for key in sorted(metadata)]
    _table(
        sheet,
        name=METADATA_TABLE,
        headers=("key", "value"),
        rows=rows,
        start_row=1,
    )
    sheet.sheet_state = "veryHidden"


def build_project_workbook(
    workspace: Mapping[str, Any],
    *,
    hmac_key: bytes,
    exported_by: str,
    export_id: str | None = None,
    exported_at: datetime | None = None,
) -> WorkbookExportArtifact:
    """Render a signed four-visible-sheet project workbook."""

    key = _require_hmac_key(hmac_key)
    project = workspace.get("project") or {}
    project_id = str(project.get("project_id") or "").strip()
    if not project_id:
        raise ProjectWorkbookV2Error("缺少稳定项目 ID")
    try:
        revision = int(workspace["workbook_revision"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ProjectWorkbookV2Error("缺少有效的项目工作簿 revision") from exc
    export_id = export_id or str(uuid.uuid4())
    exported_at = exported_at or datetime.now(timezone.utc)
    contracts, collection_rows = _project_rows(workspace, key, export_id)
    consumptions = _consumption_rows(workspace)
    expenses = _expense_rows(workspace)
    tasks = _task_rows(workspace)
    existing_count = sum(
        1
        for row in workspace.get("collections") or []
        if str(row.get("status") or "") == "已确认"
    )
    existing_collections = collection_rows[:existing_count]
    collection_client_ids = [
        str(row[11]) for row in collection_rows if row[8] in (None, "")
    ]
    snapshot_rows = {
        CONTRACT_TABLE: contracts,
        COLLECTION_TABLE: existing_collections,
        CONSUMPTION_TABLE: consumptions,
        EXPENSE_TABLE: expenses,
        TASK_TABLE: tasks,
    }
    summary_rows = _summary_rows(workspace)
    entity_versions = _entity_versions(workspace)
    metadata = _safe_metadata({
        "protocol_id": PROTOCOL_ID,
        "schema_version": SCHEMA_VERSION,
        "error_report": "false",
        "project_id": project_id,
        "project_version": str(project.get("version") or ""),
        "workbook_revision": str(revision),
        "export_id": export_id,
        "exported_at": exported_at.isoformat(),
        "exported_by": exported_by,
        "as_of": _iso(workspace.get("as_of")),
        "data_version": str(workspace.get("data_version") or ""),
        "table_map": json.dumps(_TABLE_MAP, ensure_ascii=False, sort_keys=True),
        "snapshot_digest": _digest(snapshot_rows),
        "summary_digest": _digest({"summary": summary_rows}),
        "entity_versions_digest": hashlib.sha256(
            _canonical(entity_versions).encode("utf-8")
        ).hexdigest(),
        "collection_existing_count": str(len(existing_collections)),
        "collection_client_ids_digest": hashlib.sha256(
            _canonical([(value,) for value in collection_client_ids]).encode("utf-8")
        ).hexdigest(),
    })
    metadata["metadata_hmac"] = _metadata_signature(metadata, key)

    book = Workbook()
    try:
        overview = book.active
        overview.title = "01_总览"
        _summary_sheet(overview, workspace, contracts, collection_rows)

        consumption_sheet = book.create_sheet("02_备件消耗")
        _table(
            consumption_sheet,
            name=CONSUMPTION_TABLE,
            headers=CONSUMPTION_HEADERS,
            rows=consumptions,
            start_row=1,
        )
        expense_sheet = book.create_sheet("03_报销单")
        _table(
            expense_sheet,
            name=EXPENSE_TABLE,
            headers=EXPENSE_HEADERS,
            rows=expenses,
            start_row=1,
        )
        task_sheet = book.create_sheet("04_项目经理追踪与提醒")
        _table(
            task_sheet,
            name=TASK_TABLE,
            headers=TASK_HEADERS,
            rows=tasks,
            start_row=1,
        )
        dictionary = book.create_sheet("98_字典")
        dictionary.append(("字段", "允许值"))
        dictionary.append(("回款操作", "KEEP / CREATE"))
        dictionary.append(("回款状态", "已确认"))
        dictionary.sheet_state = "hidden"
        versions = book.create_sheet("99_实体版本")
        _table(
            versions,
            name=ENTITY_VERSION_TABLE,
            headers=("entity_type", "entity_id", "base_version"),
            rows=entity_versions,
            start_row=1,
        )
        versions.sheet_state = "veryHidden"
        _metadata_sheet(book, metadata)

        for sheet in book.worksheets:
            if sheet.title in VISIBLE_SHEETS:
                sheet.sheet_view.showGridLines = False
                for cell in sheet[1]:
                    cell.alignment = _WRAP
                for col in range(1, min(sheet.max_column, 16) + 1):
                    sheet.column_dimensions[get_column_letter(col)].width = 18
        for sheet_name in ("02_备件消耗", "03_报销单", "04_项目经理追踪与提醒"):
            book[sheet_name].protection.sheet = True
        for sheet_name, date_headers in {
            "02_备件消耗": {"领用日期"},
            "03_报销单": {"报销日期"},
            "04_项目经理追踪与提醒": {"截止日期"},
        }.items():
            sheet = book[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            for header in date_headers:
                col = headers.index(header) + 1
                for row in range(2, sheet.max_row + 1):
                    sheet.cell(row, col).number_format = _DATE_FORMAT
        for sheet_name, money_headers in {
            "01_总览": {"合同额（全部合同）", "累计回款金额"},
            "02_备件消耗": {"未税单位成本", "实际消耗成本"},
            "03_报销单": {"已审批金额"},
        }.items():
            sheet = book[sheet_name]
            for table in sheet.tables.values():
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                headers = [sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)]
                for header in money_headers.intersection(headers):
                    col = min_col + headers.index(header)
                    for row in range(min_row + 1, max_row + 1):
                        sheet.cell(row, col).number_format = _MONEY_FORMAT

        output = io.BytesIO()
        book.save(output)
        content = output.getvalue()
        if len(content) > MAX_WORKBOOK_BYTES:
            raise ProjectWorkbookV2Error("导出工作簿超过安全大小上限")
    finally:
        book.close()

    safe_code = re.sub(r"[^0-9A-Za-z._-]+", "_", str(project.get("project_code") or project_id))
    return WorkbookExportArtifact(
        content=content,
        preview=preview_project_workbook(workspace),
        export_id=export_id,
        filename=f"maintenance_project_{safe_code}.xlsx",
    )


def build_error_workbook(
    issues: Sequence[WorkbookIssue],
    *,
    hmac_key: bytes,
    project_id: str,
    source_sha256: str,
) -> bytes:
    """Build a signed, human-readable and deliberately non-importable report."""

    key = _require_hmac_key(hmac_key)
    issue_rows = [(
        issue.code,
        issue.message,
        issue.sheet,
        issue.row,
        issue.column,
    ) for issue in issues]
    if not issue_rows:
        issue_rows = [("workbook_invalid", "工作簿校验失败", None, None, None)]
    empty_digest = hashlib.sha256(b"[]").hexdigest()
    metadata = _safe_metadata({
        "protocol_id": PROTOCOL_ID,
        "schema_version": SCHEMA_VERSION,
        "error_report": "true",
        "project_id": project_id,
        "project_version": "",
        "workbook_revision": "0",
        "export_id": str(uuid.uuid4()),
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_by": "system-error-report",
        "as_of": "",
        "data_version": "error-report",
        "table_map": "{}",
        "snapshot_digest": empty_digest,
        "summary_digest": empty_digest,
        "entity_versions_digest": empty_digest,
        "collection_existing_count": "0",
        "collection_client_ids_digest": empty_digest,
        "source_sha256": source_sha256,
        "error_count": str(len(issue_rows)),
    })
    metadata["metadata_hmac"] = _metadata_signature(metadata, key)
    book = Workbook()
    try:
        sheet = book.active
        sheet.title = "00_错误清单"
        _table(
            sheet,
            name="tbl_workbook_errors_v2",
            headers=("错误代码", "说明", "工作表", "行号", "字段"),
            rows=issue_rows,
            start_row=1,
        )
        sheet.freeze_panes = "A2"
        for width, column in zip((28, 48, 28, 12, 28), range(1, 6)):
            sheet.column_dimensions[get_column_letter(column)].width = width
        _metadata_sheet(book, metadata)
        output = io.BytesIO()
        book.save(output)
        return output.getvalue()
    finally:
        book.close()


def _source_bytes(source: bytes | bytearray | memoryview | io.BufferedIOBase) -> bytes:
    if isinstance(source, (bytes, bytearray, memoryview)):
        content = bytes(source)
    elif hasattr(source, "read"):
        current = source.tell() if hasattr(source, "tell") else None
        content = source.read(MAX_WORKBOOK_BYTES + 1)
        if current is not None and hasattr(source, "seek"):
            source.seek(current)
    else:
        raise TypeError("workbook source must be bytes or a binary file")
    if not content or len(content) > MAX_WORKBOOK_BYTES:
        raise ProjectWorkbookV2Error("上传工作簿为空或超过安全大小上限")
    return content


def _assert_safe_package(content: bytes) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except (zipfile.BadZipFile, OSError) as exc:
        raise ProjectWorkbookV2Error("上传文件不是有效的 XLSX 工作簿") from exc
    with archive:
        infos = archive.infolist()
        if len(infos) > MAX_ZIP_MEMBERS:
            raise ProjectWorkbookV2Error("XLSX 内部文件数量超过安全上限")
        total = 0
        names = {item.filename.casefold() for item in infos}
        if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
            raise ProjectWorkbookV2Error("上传文件缺少 XLSX 核心结构")
        for info in infos:
            name = info.filename.replace("\\", "/")
            folded = name.casefold()
            if name.startswith("/") or ".." in name.split("/"):
                raise ProjectWorkbookV2Error("XLSX 包含非法内部路径")
            total += info.file_size
            if total > MAX_UNCOMPRESSED_BYTES:
                raise ProjectWorkbookV2Error("XLSX 解压后大小超过安全上限")
            if info.file_size > 1_000_000:
                ratio = info.file_size / max(info.compress_size, 1)
                if ratio > MAX_COMPRESSION_RATIO:
                    raise ProjectWorkbookV2Error("XLSX 压缩比异常，疑似 ZIP bomb")
            if (
                "vbaproject.bin" in folded
                or "macrosheet" in folded
                or folded.startswith("xl/externallinks/")
            ):
                raise ProjectWorkbookV2Error("不允许宏或外部链接工作簿")
            if folded.endswith((".xml", ".rels")):
                raw = archive.read(info)
                lower = raw.lower()
                if b"macroenabled" in lower or re.search(
                    rb"targetmode\s*=\s*['\"]\s*external\s*['\"]",
                    lower,
                ):
                    raise ProjectWorkbookV2Error("不允许宏或外部链接工作簿")


def _assert_safe_dimensions(book) -> None:
    """Reject sparse far-away cells before any rectangular worksheet iteration."""

    for sheet in book.worksheets:
        max_row = int(sheet.max_row or 0)
        max_column = int(sheet.max_column or 0)
        if (
            max_row > MAX_WORKSHEET_ROWS
            or max_column > MAX_WORKSHEET_COLUMNS
            or max_row * max_column > MAX_DECLARED_CELLS
        ):
            raise ProjectWorkbookV2Error(
                f"{sheet.title} 的声明范围超过工作簿安全上限"
            )


def _read_table(book, sheet_name: str, table_name: str, headers: Sequence[str]):
    sheet = book[sheet_name]
    table = sheet.tables.get(table_name)
    if table is None:
        raise ProjectWorkbookV2Error(f"{sheet_name} 缺少 {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if min_col != 1 or max_col != len(headers) or max_row - min_row > MAX_ROWS_PER_TABLE:
        raise ProjectWorkbookV2Error(f"{sheet_name} 的 {table_name} 范围不符合协议")
    actual_headers = tuple(
        sheet.cell(min_row, column).value
        for column in range(min_col, max_col + 1)
    )
    if actual_headers != tuple(headers):
        raise ProjectWorkbookV2Error(f"{sheet_name} 的 {table_name} 表头被修改")
    return [
        tuple(sheet.cell(row, column).value for column in range(min_col, max_col + 1))
        for row in range(min_row + 1, max_row + 1)
    ]


def _read_metadata(book) -> dict[str, str]:
    rows = _read_table(book, "99_元数据", METADATA_TABLE, ("key", "value"))
    metadata: dict[str, str] = {}
    for key, value in rows:
        key_text = str(key or "")
        if not key_text or key_text in metadata:
            raise ProjectWorkbookV2Error("99_元数据存在空键或重复键")
        metadata[key_text] = str(value if value is not None else "")
    required = {
        "protocol_id",
        "schema_version",
        "error_report",
        "project_id",
        "workbook_revision",
        "export_id",
        "table_map",
        "snapshot_digest",
        "summary_digest",
        "entity_versions_digest",
        "collection_existing_count",
        "collection_client_ids_digest",
        "metadata_hmac",
    }
    missing = sorted(required - metadata.keys())
    if missing:
        raise ProjectWorkbookV2Error("99_元数据缺少字段：" + "、".join(missing))
    return metadata


def _raise_issue(code: str, message: str, *, sheet=None, row=None, column=None):
    raise ProjectWorkbookV2Error(
        message,
        issues=(WorkbookIssue(code, message, sheet, row, column),),
    )


def _parse_month(value: Any, *, row: int) -> date:
    text = _month_text(value)
    if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", text):
        return date.fromisoformat(text + "-01")
    _raise_issue(
        "invalid_report_month",
        "报告月份必须为 YYYY-MM",
        sheet="01_总览",
        row=row,
        column="报告月份",
    )


def _parse_positive_decimal(value: Any, *, row: int) -> Decimal:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, ValueError):
        amount = Decimal("NaN")
    if (
        not amount.is_finite()
        or amount <= 0
        or amount >= Decimal("1000000000000")
    ):
        _raise_issue(
            "invalid_cumulative_amount",
            "累计回款金额必须大于 0",
            sheet="01_总览",
            row=row,
            column="累计回款金额",
        )
    try:
        return amount.quantize(Decimal("0.01"))
    except InvalidOperation:
        _raise_issue(
            "invalid_cumulative_amount",
            "累计回款金额超过安全范围",
            sheet="01_总览",
            row=row,
            column="累计回款金额",
        )


def validate_project_workbook(
    source: bytes | bytearray | memoryview | io.BufferedIOBase,
    *,
    workspace: Mapping[str, Any],
    hmac_key: bytes,
) -> WorkbookValidation:
    """Fully validate a workbook and return a side-effect-free apply plan."""

    key = _require_hmac_key(hmac_key)
    content = _source_bytes(source)
    _assert_safe_package(content)
    try:
        book = load_workbook(io.BytesIO(content), data_only=False, keep_links=False)
    except Exception as exc:
        raise ProjectWorkbookV2Error("无法安全读取上传的 XLSX 工作簿") from exc
    try:
        if "99_元数据" not in book.sheetnames:
            raise ProjectWorkbookV2Error("未知工作簿协议；请从项目面板重新导出 v2 工作簿")
        metadata = _read_metadata(book)
        if metadata["protocol_id"] != PROTOCOL_ID:
            raise ProjectWorkbookV2Error(
                f"工作簿协议 {metadata['protocol_id'] or 'unknown'} 与 {PROTOCOL_ID} 不兼容；请勿跨版本上传"
            )
        if metadata["schema_version"] != SCHEMA_VERSION:
            raise ProjectWorkbookV2Error("工作簿 schema 版本不兼容；请重新导出")
        if metadata["error_report"].casefold() == "true":
            raise ProjectWorkbookV2Error("错误报告工作簿不可再次导入")
        expected_signature = _metadata_signature(metadata, key)
        if not hmac.compare_digest(metadata["metadata_hmac"], expected_signature):
            raise ProjectWorkbookV2Error("工作簿元数据签名无效；请重新导出")
        if tuple(book.sheetnames) != SHEET_NAMES:
            raise ProjectWorkbookV2Error("工作表名称、顺序或数量不符合 v2 协议")
        _assert_safe_dimensions(book)
        visible = tuple(sheet.title for sheet in book.worksheets if sheet.sheet_state == "visible")
        if visible != VISIBLE_SHEETS:
            raise ProjectWorkbookV2Error("可见工作表必须严格为项目四表")
        if book["98_字典"].sheet_state != "hidden" or any(
            book[name].sheet_state != "veryHidden"
            for name in ("99_实体版本", "99_元数据")
        ):
            raise ProjectWorkbookV2Error("协议隐藏工作表状态被修改")
        for sheet in book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        raise ProjectWorkbookV2Error(f"{sheet.title} 包含公式，工作簿不可导入")
        if set(book["01_总览"].tables) != {CONTRACT_TABLE, COLLECTION_TABLE}:
            raise ProjectWorkbookV2Error("01_总览 的 Excel Table 名称或数量不符合协议")
        for sheet_name, expected in {
            "02_备件消耗": {CONSUMPTION_TABLE},
            "03_报销单": {EXPENSE_TABLE},
            "04_项目经理追踪与提醒": {TASK_TABLE},
            "99_实体版本": {ENTITY_VERSION_TABLE},
            "99_元数据": {METADATA_TABLE},
        }.items():
            if set(book[sheet_name].tables) != expected:
                raise ProjectWorkbookV2Error(f"{sheet_name} 的 Excel Table 名称或数量不符合协议")
        if json.loads(metadata["table_map"]) != {key: list(value) for key, value in _TABLE_MAP.items()}:
            raise ProjectWorkbookV2Error("工作簿数据表映射被修改")

        expected_summary_rows = len(_summary_rows(workspace))
        summary_rows = [
            (
                book["01_总览"].cell(row, 1).value,
                book["01_总览"].cell(row, 2).value,
            )
            for row in range(1, expected_summary_rows + 1)
        ]
        if not hmac.compare_digest(
            _digest({"summary": summary_rows}),
            metadata["summary_digest"],
        ):
            raise ProjectWorkbookV2Error("01_总览项目摘要被修改；请重新导出")

        current_project_id = str((workspace.get("project") or {}).get("project_id") or "")
        if metadata["project_id"] != current_project_id:
            raise ProjectWorkbookV2Error("工作簿项目与当前项目不一致", status_code=409)
        if not bool((workspace.get("project") or {}).get("is_active", True)):
            raise ProjectWorkbookV2Error(
                "项目已归档，不能校验或应用回款工作簿", status_code=409
            )
        try:
            expected_revision = int(metadata["workbook_revision"])
            current_revision = int(workspace["workbook_revision"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ProjectWorkbookV2Error("工作簿 revision 无效") from exc
        if expected_revision != current_revision:
            raise ProjectWorkbookV2Error(
                "项目数据已更新，当前工作簿已过期；请重新导出后回填",
                status_code=409,
                issues=(WorkbookIssue("stale_workbook", "项目工作簿 revision 已变化"),),
            )

        contracts = [row for row in _read_table(book, "01_总览", CONTRACT_TABLE, CONTRACT_HEADERS) if any(value not in (None, "") for value in row)]
        collection_rows = _read_table(book, "01_总览", COLLECTION_TABLE, COLLECTION_HEADERS)
        consumptions = [row for row in _read_table(book, "02_备件消耗", CONSUMPTION_TABLE, CONSUMPTION_HEADERS) if any(value not in (None, "") for value in row)]
        expenses = [row for row in _read_table(book, "03_报销单", EXPENSE_TABLE, EXPENSE_HEADERS) if any(value not in (None, "") for value in row)]
        tasks = [row for row in _read_table(book, "04_项目经理追踪与提醒", TASK_TABLE, TASK_HEADERS) if any(value not in (None, "") for value in row)]
        try:
            existing_count = int(metadata["collection_existing_count"])
        except ValueError as exc:
            raise ProjectWorkbookV2Error("既有回款数量元数据无效") from exc
        existing = [row for row in collection_rows if row[8] not in (None, "")]
        if len(existing) != existing_count:
            raise ProjectWorkbookV2Error("既有回款行被删除或新增；请保留全部 KEEP 行")
        if any(str(row[0] or "").strip() != "KEEP" for row in existing):
            raise ProjectWorkbookV2Error("既有回款只能保留 KEEP，不能修改或删除")
        if any(
            collection_rows[index][8] in (None, "")
            for index in range(existing_count)
        ) or any(
            row[8] not in (None, "")
            for row in collection_rows[existing_count:]
        ):
            raise ProjectWorkbookV2Error(
                "既有 KEEP 行必须完整保留在回款表前部，新增只能写在尾部"
            )
        client_ids = [
            str(row[11] or "")
            for row in collection_rows
            if row[8] in (None, "")
        ]
        if any(not value for value in client_ids):
            raise ProjectWorkbookV2Error("回款尾部行缺少系统生成的 client_row_id")
        client_digest = hashlib.sha256(
            _canonical([(value,) for value in client_ids]).encode("utf-8")
        ).hexdigest()
        if not hmac.compare_digest(
            client_digest,
            metadata["collection_client_ids_digest"],
        ):
            raise ProjectWorkbookV2Error(
                "系统生成的 client_row_id 列表被修改；请重新导出"
            )
        for row in existing:
            token_payload = f"{metadata['export_id']}|collection|{row[8]}|{row[9]}".encode()
            expected_token = hmac.new(key, token_payload, hashlib.sha256).hexdigest()
            if not hmac.compare_digest(str(row[10] or ""), expected_token):
                raise ProjectWorkbookV2Error("既有回款行令牌无效；请重新导出")
        snapshot = {
            CONTRACT_TABLE: contracts,
            COLLECTION_TABLE: existing,
            CONSUMPTION_TABLE: consumptions,
            EXPENSE_TABLE: expenses,
            TASK_TABLE: tasks,
        }
        if not hmac.compare_digest(_digest(snapshot), metadata["snapshot_digest"]):
            raise ProjectWorkbookV2Error("只读业务数据或既有回款被修改；仅允许在尾部新增回款")

        version_rows = [
            row for row in _read_table(
                book,
                "99_实体版本",
                ENTITY_VERSION_TABLE,
                ("entity_type", "entity_id", "base_version"),
            )
            if any(value not in (None, "") for value in row)
        ]
        version_digest = hashlib.sha256(_canonical(version_rows).encode("utf-8")).hexdigest()
        if not hmac.compare_digest(version_digest, metadata["entity_versions_digest"]):
            raise ProjectWorkbookV2Error("实体版本快照被修改；请重新导出")

        project_contract_nos = {str(row[0]): str(row[1] or "") for row in contracts}
        project_contract_ids = set(project_contract_nos)
        contract_facts = {
            str(row.get("project_contract_id") or ""): row
            for row in workspace.get("contracts") or []
        }
        as_of_text = _iso(workspace.get("as_of"))[:10]
        try:
            latest_report_month = date.fromisoformat(as_of_text).replace(day=1)
        except ValueError as exc:
            raise ProjectWorkbookV2Error("项目数据截至日期无效") from exc
        occupied_contract_months = {
            (
                str(row.get("project_contract_id") or ""),
                _month_text(row.get("report_month")),
            )
            for row in workspace.get("collections") or []
            if row.get("project_contract_id")
            and re.fullmatch(
                r"\d{4}-(0[1-9]|1[0-2])",
                _month_text(row.get("report_month")),
            )
        }
        snapshot_amounts: dict[tuple[str, str], Decimal] = {}
        for row in existing:
            key_tuple = (str(row[1]), _month_text(row[3]))
            if key_tuple in snapshot_amounts:
                raise ProjectWorkbookV2Error(
                    "导出快照内存在重复的项目合同关系/月"
                )
            try:
                snapshot_amounts[key_tuple] = Decimal(str(row[4]))
            except (InvalidOperation, ValueError):
                raise ProjectWorkbookV2Error(
                    "既有累计回款金额无效；请先修复系统数据"
                ) from None
        creates: list[CollectionCreate] = []
        seen_client_ids: set[str] = set()
        for index, row in enumerate(collection_rows, 1):
            if row[8] not in (None, ""):
                continue
            business_values = row[:8]
            if all(value in (None, "") for value in business_values):
                continue
            excel_row = index + range_boundaries(book["01_总览"].tables[COLLECTION_TABLE].ref)[1]
            if str(row[0] or "").strip() != "CREATE":
                _raise_issue("create_operation_required", "新增回款必须选择 CREATE", sheet="01_总览", row=excel_row, column="操作")
            project_contract_id = str(row[1] or "").strip()
            if project_contract_id not in project_contract_ids:
                _raise_issue(
                    "unknown_project_contract",
                    "新增回款必须关联本项目内的合同关系",
                    sheet="01_总览",
                    row=excel_row,
                    column="项目合同关系ID",
                )
            contract_no = str(row[2] or "").strip()
            if contract_no != project_contract_nos[project_contract_id]:
                _raise_issue(
                    "project_contract_number_mismatch",
                    "合同编号与项目合同关系不一致",
                    sheet="01_总览",
                    row=excel_row,
                    column="合同编号",
                )
            client_row_id = str(row[11] or "").strip()
            try:
                uuid.UUID(client_row_id)
            except (ValueError, AttributeError):
                _raise_issue("invalid_client_row_id", "新增回款行缺少系统生成的 client_row_id", sheet="01_总览", row=excel_row)
            if client_row_id in seen_client_ids:
                _raise_issue("duplicate_client_row_id", "新增回款 client_row_id 重复", sheet="01_总览", row=excel_row)
            seen_client_ids.add(client_row_id)
            report_month = _parse_month(row[3], row=excel_row)
            if report_month > latest_report_month:
                _raise_issue(
                    "future_report_month",
                    "回款报告月份不能晚于项目数据截至月份",
                    sheet="01_总览",
                    row=excel_row,
                    column="报告月份",
                )
            contract_fact = contract_facts.get(project_contract_id) or {}
            if contract_fact.get("status_mapping_state") not in (None, "mapped"):
                _raise_issue(
                    "unmapped_project_contract",
                    "新增回款关联合同状态尚未映射",
                    sheet="01_总览",
                    row=excel_row,
                    column="项目合同关系ID",
                )
            month_end = report_month.replace(
                day=monthrange(report_month.year, report_month.month)[1]
            )
            effective_from = contract_fact.get("effective_from")
            effective_to = contract_fact.get("effective_to")
            try:
                starts = date.fromisoformat(str(effective_from)) if effective_from else None
                ends = date.fromisoformat(str(effective_to)) if effective_to else None
            except ValueError as exc:
                raise ProjectWorkbookV2Error("项目合同生效日期无效") from exc
            if (starts is not None and starts > month_end) or (
                ends is not None and ends <= report_month
            ):
                _raise_issue(
                    "contract_inactive_for_report_month",
                    "新增回款关联合同在该报告月份无效",
                    sheet="01_总览",
                    row=excel_row,
                    column="报告月份",
                )
            amount = _parse_positive_decimal(row[4], row=excel_row)
            snapshot_key = (project_contract_id, report_month.strftime("%Y-%m"))
            if (
                snapshot_key in snapshot_amounts
                or snapshot_key in occupied_contract_months
            ):
                _raise_issue(
                    "duplicate_contract_month",
                    "同一项目合同关系的同一报告月份只能有一条累计回款快照",
                    sheet="01_总览",
                    row=excel_row,
                    column="报告月份",
                )
            earlier = [
                existing_amount
                for (relation, month), existing_amount in snapshot_amounts.items()
                if relation == project_contract_id and month < snapshot_key[1]
            ]
            later = [
                existing_amount
                for (relation, month), existing_amount in snapshot_amounts.items()
                if relation == project_contract_id and month > snapshot_key[1]
            ]
            if earlier and amount < max(earlier):
                _raise_issue(
                    "cumulative_decrease",
                    "累计回款金额不能低于该合同关系更早月份的累计值",
                    sheet="01_总览",
                    row=excel_row,
                    column="累计回款金额",
                )
            if later and amount > min(later):
                _raise_issue(
                    "cumulative_exceeds_later_month",
                    "累计回款金额不能高于该合同关系更晚月份的累计值",
                    sheet="01_总览",
                    row=excel_row,
                    column="累计回款金额",
                )
            snapshot_amounts[snapshot_key] = amount
            status = str(row[6] or "已确认").strip()
            if status != "已确认":
                _raise_issue("unconfirmed_collection", "只允许导入已确认的回款", sheet="01_总览", row=excel_row, column="状态")
            payload = {
                "project_contract_id": project_contract_id,
                "contract_no": contract_no,
                "report_month": report_month.strftime("%Y-%m"),
                "cumulative_amount": format(amount, "f"),
                "voucher_no": str(row[5]).strip() if row[5] not in (None, "") else None,
                "status": status,
                "remark": str(row[7]).strip() if row[7] not in (None, "") else None,
            }
            payload_hash = hashlib.sha256(
                json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
            ).hexdigest()
            creates.append(CollectionCreate(
                operation_key=f"{metadata['export_id']}|01_总览|{client_row_id}",
                client_row_id=client_row_id,
                project_contract_id=project_contract_id,
                contract_no=payload["contract_no"],
                report_month=report_month,
                cumulative_amount=amount,
                voucher_no=payload["voucher_no"],
                status=status,
                remark=payload["remark"],
                payload_hash=payload_hash,
            ))
        return WorkbookValidation(
            validation_id=str(uuid.uuid4()),
            project_id=metadata["project_id"],
            export_id=metadata["export_id"],
            expected_revision=expected_revision,
            file_sha256=hashlib.sha256(content).hexdigest(),
            creates=tuple(creates),
            unchanged=not creates,
            metadata=metadata,
        )
    finally:
        book.close()


def apply_project_workbook(
    validation: WorkbookValidation,
    *,
    repository: ProjectWorkbookApplyRepository,
) -> WorkbookApplyResult:
    """Apply a fully validated plan through one transactional repository call."""

    if repository.applied_file(validation.file_sha256):
        return WorkbookApplyResult(
            status="file_replay",
            created=0,
            replayed=len(validation.creates),
            validation_id=validation.validation_id,
        )

    pending: list[CollectionCreate] = []
    replayed = 0
    for create in validation.creates:
        applied_hash = repository.applied_operation(create.operation_key)
        if applied_hash is None:
            pending.append(create)
            continue
        if not hmac.compare_digest(applied_hash, create.payload_hash):
            raise ProjectWorkbookV2Error(
                "同一导出行已用不同内容应用，拒绝覆盖",
                status_code=409,
                issues=(WorkbookIssue(
                    "operation_payload_conflict",
                    "export_id/sheet/client_row_id 已存在但 payload 不同",
                ),),
            )
        replayed += 1

    if not pending:
        return WorkbookApplyResult(
            status="logical_replay" if replayed else "unchanged",
            created=0,
            replayed=replayed,
            validation_id=validation.validation_id,
        )

    if repository.current_revision(validation.project_id) != validation.expected_revision:
        raise ProjectWorkbookV2Error(
            "项目数据已更新，应用计划已过期；请重新导出后回填",
            status_code=409,
            issues=(WorkbookIssue("stale_workbook", "项目工作簿 revision 已变化"),),
        )
    repository.apply_collections_atomically(validation, tuple(pending))
    return WorkbookApplyResult(
        status="applied",
        created=len(pending),
        replayed=replayed,
        validation_id=validation.validation_id,
    )


def validate_and_store_project_workbook(
    project_id: str,
    source: bytes | bytearray | memoryview | io.BufferedIOBase,
    *,
    adapter: ProjectWorkbookEndpointAdapter,
    hmac_key: bytes,
) -> WorkbookValidationAttempt:
    """Endpoint orchestration: store a server-owned plan or error workbook."""

    try:
        content = _source_bytes(source)
        validation = validate_project_workbook(
            content,
            workspace=adapter.load_workspace(project_id),
            hmac_key=hmac_key,
        )
        if validation.project_id != project_id:
            raise ProjectWorkbookV2Error(
                "工作簿项目与 URL 项目不一致",
                status_code=409,
            )
    except ProjectWorkbookV2Error as exc:
        if exc.status_code == 409:
            raise
        validation_id = str(uuid.uuid4())
        try:
            source_hash = hashlib.sha256(content).hexdigest()
        except UnboundLocalError:
            source_hash = hashlib.sha256(b"").hexdigest()
        report = build_error_workbook(
            exc.issues,
            hmac_key=hmac_key,
            project_id=project_id,
            source_sha256=source_hash,
        )
        adapter.save_validation_error(
            validation_id,
            project_id,
            exc.issues,
            report,
        )
        return WorkbookValidationAttempt(
            valid=False,
            validation_id=validation_id,
            validation=None,
            issues=exc.issues,
            can_apply=False,
        )
    adapter.save_validation(validation)
    return WorkbookValidationAttempt(
        valid=True,
        validation_id=validation.validation_id,
        validation=validation,
        issues=(),
        can_apply=not validation.unchanged,
    )


def apply_stored_project_workbook(
    project_id: str,
    validation_id: str,
    *,
    adapter: ProjectWorkbookEndpointAdapter,
) -> WorkbookApplyResult:
    """Apply only the server-side validation plan identified by the token."""

    validation = adapter.load_validation(validation_id)
    if validation is None:
        raise ProjectWorkbookV2Error("validation_id 不存在或已过期", status_code=404)
    if validation.project_id != project_id:
        raise ProjectWorkbookV2Error("validation_id 不属于当前项目", status_code=409)
    return apply_project_workbook(validation, repository=adapter)


def load_validation_error_workbook(
    validation_id: str,
    *,
    adapter: ProjectWorkbookEndpointAdapter,
) -> bytes:
    content = adapter.load_validation_error(validation_id)
    if content is None:
        raise ProjectWorkbookV2Error("错误工作簿不存在或已过期", status_code=404)
    return content
