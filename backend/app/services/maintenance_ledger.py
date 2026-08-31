"""维保台账工作簿：解析（preview）与 apply（B2）。

台账 = 商务线唯一事实源。支持两种结构：
- 旧结构（业务现行文件）：`维保项目清单`（16 固定列 + 24 组横向「回款时间N/回款金额」对）
  + `项目成本`；
- 新模板 v1（docs/maintenance/templates/）：`01_项目与合同` + `02_回款计划` + `03_项目成本`。

preview 只解析落 raw 行（零 canonical 写入）；apply 才同步 project / contract /
milestone，并在台账金额与正式 BXD 含税事实一致后同步 expense attribution。
"""
from __future__ import annotations

import hashlib
import io
import re
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import func, or_, select, text
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app import config, tax_policy
from app.business_time import business_today
from app.models.maintenance_ledger import (
    LEDGER_SOURCE,
    LEDGER_TEMPLATE_SOURCE,
    MaintenanceLedgerContractRow,
    MaintenanceLedgerExpenseRow,
    MaintenanceLedgerImportBatch,
    MaintenanceLedgerPlanRow,
)
from app.models.maintenance_manager import MaintenanceCollectionMilestone
from app.models.maintenance_project import (
    MaintenanceProject,
    MaintenanceProjectAuditLog,
    MaintenanceProjectContract,
)
from app.models.sales import FSalesOrder
from app.models.system import SysImportBatch
from app.services.date_loose import (
    parse_amount_loose,
    parse_date_loose,
    parse_project_name,
)
from app.services.maintenance_periods import (
    MaintenancePeriodError,
    SOURCE_LEDGER,
    apply_canonical_period_locked,
    lifecycle_status as canonical_lifecycle_status,
)

MAX_PREVIEW_BYTES = 16 * 1024 * 1024
MAX_PROJECT_CODE_LEN = 64

_CONTRACT_HEADERS = [
    "订单编号", "订单日期", "销售人员", "业务类型", "项目名称",
    "维保起始日期", "维保终止日期", "CMO", "项目经理", "订单金额",
    "已收尾款", "待收尾款", "验收材料", "验收材料是否完成及上传附件",
    "验收附件", "巡检时间", "巡检是否完成及上传附件",
]
_EXPENSE_HEADERS = [
    "费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单",
    "项目名称", "销售订单", "销售人员", "费用分类", "报销金额", "备注",
]
_PLAN_HEADERS = ["订单编号", "计划期次", "计划回款时间", "计划回款金额"]

_XSDD_RE = re.compile(r"(XSDD-\d{8}-\d{4})")
_BXD_RE = re.compile(r"(BXD-\d{8}-\d{4})")


class LedgerParseError(RuntimeError):
    """台账文件不可解析（结构/大小/加密等）。"""


class LedgerBatchError(RuntimeError):
    """台账批次状态错误（重复 apply / 不存在等）。"""


@dataclass
class ContractRowData:
    row_no: int
    values: dict[str, str]
    issues: list[str] = field(default_factory=list)


@dataclass
class PlanRowData:
    row_no: int
    order_no_raw: str
    sequence: int
    time_raw: str | None
    amount_raw: str | None
    issues: list[str] = field(default_factory=list)


@dataclass
class ExpenseRowData:
    row_no: int
    values: dict[str, str]
    issues: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class LedgerExpenseSyncPlan:
    bxd_no: str
    target_order_no: str
    raw_line_ids: tuple[str, ...]
    observed_raw_line_ids: tuple[str, ...]
    expected_raw_fingerprint: str


@dataclass(frozen=True)
class LedgerContractExpenseProbe:
    """Unlocked snapshot used only to define the later lock envelope."""

    raw_line_ids: tuple[str, ...]
    raw_identities: tuple[tuple[str, str], ...]
    expected_raw_fingerprint: str
    owner_project_ids: frozenset[str]


def _aliased_headers(headers: list, column_aliases: dict | None) -> list:
    from app.services import import_safety

    try:
        return import_safety.apply_column_aliases(headers, column_aliases)
    except import_safety.UploadSafetyError as exc:
        raise LedgerParseError(str(exc)) from exc


def _header_index(headers: list, name: str) -> int | None:
    variants = {
        name,
        f"{name}(必填)",
        f"{name}(不可修改)",
        f"{name}(含税)",
    }
    for idx, value in enumerate(headers, 1):
        if value in variants:
            return idx
    return None


def _cell(row: tuple, index: int | None) -> str | None:
    if index is None or index > len(row):
        return None
    value = row[index - 1]
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _clean_order_no(raw: str | None) -> str | None:
    if not raw:
        return None
    match = _XSDD_RE.search(raw)
    return match.group(1) if match else None


def _clean_bxd_no(raw: str | None) -> str | None:
    if not raw:
        return None
    match = _BXD_RE.search(raw)
    return match.group(1) if match else None


def _non_empty_row(row: tuple) -> bool:
    return bool(row) and any(v is not None and str(v).strip() != "" for v in row)


def _parse_old_ledger(workbook, column_aliases: dict | None = None):
    try:
        contract_sheet = workbook["维保项目清单"]
    except KeyError as exc:
        raise LedgerParseError("缺少「维保项目清单」Sheet") from exc

    rows = list(contract_sheet.iter_rows(values_only=True))
    if not rows:
        raise LedgerParseError("「维保项目清单」为空")
    headers = _aliased_headers(
        [str(v).strip() if v is not None else "" for v in rows[0]], column_aliases
    )
    indexes = {name: _header_index(headers, name) for name in _CONTRACT_HEADERS}
    if indexes["订单编号"] is None:
        raise LedgerParseError("「维保项目清单」缺少「订单编号」列")

    time_cols: list[tuple[int, int]] = []
    amount_cols: list[int] = []
    for idx, name in enumerate(headers, 1):
        match = re.fullmatch(r"回款时间(\d{1,2})", name)
        if match:
            time_cols.append((int(match.group(1)), idx))
        elif name == "回款金额":
            amount_cols.append(idx)
    time_cols.sort()
    pairs = list(zip(time_cols, amount_cols))[:24]

    contract_rows: list[ContractRowData] = []
    plan_rows: list[PlanRowData] = []
    for row_no, row in enumerate(rows[1:], 2):
        if not _non_empty_row(row):
            continue
        values = {name: _cell(row, indexes[name]) for name in _CONTRACT_HEADERS}
        issues: list[str] = []
        order_no = _clean_order_no(values["订单编号"])
        if not order_no:
            issues.append("订单编号缺失或格式异常")
        for (seq, time_idx), amount_idx in pairs:
            time_raw = _cell(row, time_idx)
            amount_raw = _cell(row, amount_idx)
            if not time_raw and not amount_raw:
                continue
            plan_rows.append(
                PlanRowData(
                    row_no=row_no,
                    order_no_raw=order_no or "",
                    sequence=seq,
                    time_raw=time_raw,
                    amount_raw=amount_raw,
                    issues=list(issues),
                )
            )
        contract_rows.append(ContractRowData(row_no=row_no, values=values, issues=issues))

    expense_rows: list[ExpenseRowData] = []
    if "项目成本" in workbook.sheetnames:
        expense_rows.extend(_parse_expense_sheet(workbook["项目成本"], column_aliases))
    return contract_rows, plan_rows, expense_rows


def _parse_expense_sheet(sheet, column_aliases: dict | None = None) -> list[ExpenseRowData]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = _aliased_headers(
        [str(v).strip() if v is not None else "" for v in rows[0]], column_aliases
    )
    indexes: dict[str, int | None] = {}
    for name in _EXPENSE_HEADERS:
        index = _header_index(headers, name)
        if index is None:
            # 兼容「报销明细.报销金额」「报销明细.费用分类」前缀变体
            index = _header_index(headers, f"报销明细.{name}")
        indexes[name] = index
    result: list[ExpenseRowData] = []
    for row_no, row in enumerate(rows[1:], 2):
        if not _non_empty_row(row):
            continue
        values = {name: _cell(row, indexes[name]) for name in _EXPENSE_HEADERS}
        issues: list[str] = []
        bxd_no = _clean_bxd_no(values["费用单号"])
        primary_order = _clean_order_no(values["维保销售订单"])
        duplicate_order = _clean_order_no(values["销售订单"])
        if not bxd_no:
            issues.append("费用单号缺失或格式异常")
        if primary_order and duplicate_order and primary_order != duplicate_order:
            issues.append("维保销售订单与销售订单不一致")
        if values["报销金额"] and parse_amount_loose(values["报销金额"]) is None:
            issues.append("报销金额无法解析")
        result.append(ExpenseRowData(row_no=row_no, values=values, issues=issues))
    return result


def _parse_new_ledger(workbook, column_aliases: dict | None = None):
    try:
        contract_sheet = workbook["01_项目与合同"]
    except KeyError as exc:
        raise LedgerParseError("缺少「01_项目与合同」Sheet") from exc

    rows = list(contract_sheet.iter_rows(values_only=True))
    if not rows:
        raise LedgerParseError("「01_项目与合同」为空")
    headers = _aliased_headers(
        [str(v).strip() if v is not None else "" for v in rows[0]], column_aliases
    )
    indexes = {name: _header_index(headers, name) for name in _CONTRACT_HEADERS}
    if indexes["订单编号"] is None:
        raise LedgerParseError("「01_项目与合同」缺少「订单编号」列")

    contract_rows: list[ContractRowData] = []
    for row_no, row in enumerate(rows[1:], 2):
        if not _non_empty_row(row):
            continue
        values = {name: _cell(row, indexes[name]) for name in _CONTRACT_HEADERS}
        issues: list[str] = []
        if not _clean_order_no(values["订单编号"]):
            issues.append("订单编号缺失或格式异常")
        contract_rows.append(ContractRowData(row_no=row_no, values=values, issues=issues))

    plan_rows: list[PlanRowData] = []
    if "02_回款计划" in workbook.sheetnames:
        plan_data = list(workbook["02_回款计划"].iter_rows(values_only=True))
        if plan_data:
            plan_headers = _aliased_headers(
                [str(v).strip() if v is not None else "" for v in plan_data[0]],
                column_aliases,
            )
            plan_indexes = {name: _header_index(plan_headers, name) for name in _PLAN_HEADERS}
            for row_no, row in enumerate(plan_data[1:], 2):
                if not _non_empty_row(row):
                    continue
                order_no_raw = _cell(row, plan_indexes["订单编号"]) or ""
                seq_raw = _cell(row, plan_indexes["计划期次"])
                time_raw = _cell(row, plan_indexes["计划回款时间"])
                amount_raw = _cell(row, plan_indexes["计划回款金额"])
                try:
                    sequence = int(str(seq_raw).strip()) if seq_raw else 0
                except ValueError:
                    sequence = 0
                issues: list[str] = []
                if not _clean_order_no(order_no_raw):
                    issues.append("订单编号缺失或格式异常")
                if not 1 <= sequence <= 24:
                    issues.append("计划期次无效（应为 1-24）")
                plan_rows.append(
                    PlanRowData(
                        row_no=row_no,
                        order_no_raw=order_no_raw,
                        sequence=sequence,
                        time_raw=time_raw,
                        amount_raw=amount_raw,
                        issues=issues,
                    )
                )

    expense_rows: list[ExpenseRowData] = []
    if "03_项目成本" in workbook.sheetnames:
        expense_rows.extend(_parse_expense_sheet(workbook["03_项目成本"], column_aliases))
    return contract_rows, plan_rows, expense_rows


def parse_ledger_workbook(
    data: bytes, filename: str, *, column_aliases: dict | None = None
) -> dict:
    """解析台账工作簿字节流。返回 {source_kind, file_hash, contract_rows, plan_rows, expense_rows}。"""
    if len(data) > MAX_PREVIEW_BYTES:
        raise LedgerParseError("台账文件超过大小上限")
    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - 文件损坏时统一业务错误
        raise LedgerParseError(f"无法读取 Excel 文件：{type(exc).__name__}") from exc
    try:
        sheet_names = set(workbook.sheetnames)
        if "维保项目清单" in sheet_names:
            source_kind = LEDGER_SOURCE
            contract_rows, plan_rows, expense_rows = _parse_old_ledger(
                workbook, column_aliases
            )
        elif "01_项目与合同" in sheet_names:
            source_kind = LEDGER_TEMPLATE_SOURCE
            contract_rows, plan_rows, expense_rows = _parse_new_ledger(
                workbook, column_aliases
            )
        else:
            raise LedgerParseError(
                "无法识别台账结构：需要「维保项目清单」或「01_项目与合同」Sheet"
            )
    finally:
        workbook.close()
    return {
        "source_kind": source_kind,
        "file_hash": hashlib.sha256(data).hexdigest(),
        "filename": filename,
        "contract_rows": contract_rows,
        "plan_rows": plan_rows,
        "expense_rows": expense_rows,
    }


def store_preview(
    db: Session,
    parsed: dict,
    operated_by: str,
    *,
    idempotency_key: str,
    commit: bool = True,
) -> str:
    """落 raw 行并返回 batch_id。零 canonical 写入。

    Idempotency-Key 生效：(uploaded_by, idempotency_key) 唯一；
    同 key 同 hash 返回既有批次（重放），同 key 异 hash 拒绝。
    """
    existing = db.execute(
        select(MaintenanceLedgerImportBatch).where(
            MaintenanceLedgerImportBatch.uploaded_by == operated_by,
            MaintenanceLedgerImportBatch.idempotency_key == idempotency_key,
        )
    ).scalar_one_or_none()
    if existing is not None:
        if existing.file_hash != parsed["file_hash"]:
            raise LedgerBatchError("同一 Idempotency-Key 对应不同文件内容，拒绝重放")
        return existing.batch_id
    batch = MaintenanceLedgerImportBatch(
        batch_id=str(uuid4()),
        file_hash=parsed["file_hash"],
        filename=parsed["filename"][:255],
        idempotency_key=idempotency_key,
        source_kind=parsed["source_kind"],
        uploaded_by=operated_by,
        contract_rows=len(parsed["contract_rows"]),
        plan_rows=len(parsed["plan_rows"]),
        expense_rows=len(parsed["expense_rows"]),
        status="pending",
    )
    db.add(batch)
    db.flush()
    issue_rows = 0
    for data in parsed["contract_rows"]:
        order_date, _ = parse_date_loose(data.values["订单日期"])
        start, _ = parse_date_loose(data.values["维保起始日期"])
        end, _ = parse_date_loose(data.values["维保终止日期"])
        parsed_project = parse_project_name(data.values["项目名称"] or "")
        if start is None:
            start = parsed_project["period_from"]
        if end is None:
            end = parsed_project["period_to"]
        if not data.values["项目名称"]:
            data.issues.append("项目名称缺失")
        if data.values["维保起始日期"] and start is None:
            data.issues.append("维保起始日期无法解析")
        if data.values["维保终止日期"] and end is None:
            data.issues.append("维保终止日期无法解析")
        if data.values["订单日期"] and order_date is None:
            data.issues.append("订单日期无法解析")
        amount = parse_amount_loose(data.values["订单金额"])
        if data.values["订单金额"] and amount is None:
            data.issues.append("订单金额无法解析")
        elif not data.values["订单金额"]:
            data.issues.append("订单金额缺失")
        if start is None and end is None:
            data.issues.append("维保期限缺失（台账日期与项目名称内均未找到）")
        if data.issues:
            issue_rows += 1
        db.add(
            MaintenanceLedgerContractRow(
                row_id=str(uuid4()),
                batch_id=batch.batch_id,
                row_no=data.row_no,
                order_no_raw=data.values["订单编号"],
                order_date_raw=data.values["订单日期"],
                salesperson_raw=data.values["销售人员"],
                business_type_raw=data.values["业务类型"],
                project_name_raw=data.values["项目名称"],
                maint_start_raw=data.values["维保起始日期"],
                maint_end_raw=data.values["维保终止日期"],
                cmo_raw=data.values["CMO"],
                manager_raw=data.values["项目经理"],
                amount_raw=data.values["订单金额"],
                collected_raw=data.values["已收尾款"],
                receivable_raw=data.values["待收尾款"],
                acceptance_material_raw=data.values["验收材料"],
                acceptance_done_raw=data.values["验收材料是否完成及上传附件"],
                acceptance_attachment_raw=data.values["验收附件"],
                inspection_time_raw=data.values["巡检时间"],
                inspection_done_raw=data.values["巡检是否完成及上传附件"],
                order_no=_clean_order_no(data.values["订单编号"]),
                order_date=order_date,
                business_type=data.values["业务类型"] or None,
                project_name=parsed_project["identity"] or None,
                project_period_from=start,
                project_period_to=end,
                cmo=data.values["CMO"] or None,
                manager=data.values["项目经理"] or None,
                amount_inc_tax=parse_amount_loose(data.values["订单金额"]),
                collected_amount=parse_amount_loose(data.values["已收尾款"]),
                receivable_amount=parse_amount_loose(data.values["待收尾款"]),
                issues=data.issues,
            )
        )
    for data in parsed["plan_rows"]:
        planned_date, precision = parse_date_loose(data.time_raw)
        if data.time_raw and planned_date is None:
            data.issues.append("计划回款时间无法解析")
        if data.amount_raw and parse_amount_loose(data.amount_raw) is None:
            data.issues.append("计划回款金额无法解析")
        if data.issues:
            issue_rows += 1
        db.add(
            MaintenanceLedgerPlanRow(
                row_id=str(uuid4()),
                batch_id=batch.batch_id,
                row_no=data.row_no,
                order_no_raw=data.order_no_raw,
                sequence=data.sequence if 1 <= data.sequence <= 24 else 0,
                time_raw=data.time_raw,
                amount_raw=data.amount_raw,
                order_no=_clean_order_no(data.order_no_raw),
                planned_date=planned_date,
                date_precision=precision,
                planned_amount=parse_amount_loose(data.amount_raw),
                issues=data.issues,
            )
        )
    for data in parsed["expense_rows"]:
        if data.issues:
            issue_rows += 1
        db.add(
            MaintenanceLedgerExpenseRow(
                row_id=str(uuid4()),
                batch_id=batch.batch_id,
                row_no=data.row_no,
                bxd_no_raw=data.values["费用单号"],
                person_raw=data.values["报销人员"],
                expense_type_raw=data.values["报销类别"],
                reason_raw=data.values["支出事由"],
                sales_order_raw=data.values["维保销售订单"],
                project_name_raw=data.values["项目名称"],
                sales_order_dup_raw=data.values["销售订单"],
                salesperson_raw=data.values["销售人员"],
                fee_category_raw=data.values["费用分类"],
                amount_raw=data.values["报销金额"],
                remark_raw=data.values["备注"],
                bxd_no=_clean_bxd_no(data.values["费用单号"]),
                sales_order=(
                    _clean_order_no(data.values["维保销售订单"])
                    or _clean_order_no(data.values["销售订单"])
                ),
                amount=parse_amount_loose(data.values["报销金额"]),
                issues=data.issues,
            )
        )
    batch.issue_rows = issue_rows
    batch.report_json = {
        "contract_rows": batch.contract_rows,
        "plan_rows": batch.plan_rows,
        "expense_rows": batch.expense_rows,
        "issue_rows": issue_rows,
    }
    if commit:
        db.commit()
    return batch.batch_id


def _lifecycle_status(period_from: date | None, period_to: date | None, today: date) -> str:
    return canonical_lifecycle_status(period_from, period_to, today)


# 项目名称里嵌着服务周期：`客户名YYYYMMDD-YYYYMMDD服务商 业务类型`（REQUIREMENTS #50，
# 业务 2026-08-17 指示：周期从项目名提取）。按生产 415 个名称实测定的两档式样：
# 8 位日起止为主（89%），连接符有 `-`/`~`/空格混排；另有 6 位年月段（YYYYMM）个例。
# 前后负向断言防止从更长的数字串中间截取；名称笔误（7 位日期、双连字符）不救——
# 名称只是兜底源，宁缺毋错。
_NAME_PERIOD_RE = re.compile(r"(?<!\d)(\d{8})\s*[-—~至]\s*(\d{8})(?!\d)")
_NAME_PERIOD_YM_RE = re.compile(r"(?<!\d)(\d{6})\s*[-—~至]\s*(\d{6})(?!\d)")


def _period_from_display_name(name: str | None) -> tuple[date | None, date | None]:
    """从项目名称提取服务周期；解析不出（无日期段/非法日期/起止倒置）一律返回
    (None, None)。年月段取起月首日、止月末日。

    迁移 e8b2c6f4d1a7 内嵌同款逻辑做存量回填（迁移必须自包含、不得 import 本模块，
    否则 CI 重放整链会绑死 app 包的历史形状）；改这里记得同步那份副本。
    """
    if not name:
        return (None, None)
    start = end = None
    match = _NAME_PERIOD_RE.search(name)
    if match is not None:
        try:
            start = datetime.strptime(match.group(1), "%Y%m%d").date()
            end = datetime.strptime(match.group(2), "%Y%m%d").date()
        except ValueError:
            return (None, None)
    else:
        match = _NAME_PERIOD_YM_RE.search(name)
        if match is None:
            return (None, None)
        try:
            start = datetime.strptime(match.group(1), "%Y%m").date()
            end_month = datetime.strptime(match.group(2), "%Y%m").date()
        except ValueError:
            return (None, None)
        end = end_month.replace(day=monthrange(end_month.year, end_month.month)[1])
    if start > end:
        return (None, None)
    return (start, end)


def _resolve_lifecycle(
    period_from: date | None,
    period_to: date | None,
    display_name: str | None,
    today: date,
) -> str:
    """台账周期是权威源；台账没给（missing）时才回退到名称解析（#50）。

    注意这是**快照语义**：与台账导入一致，lifecycle 反映计算当日的状态，
    ongoing→ended 的翻转靠下一次台账导入（或下一次名称回填）刷新。
    """
    status = _lifecycle_status(period_from, period_to, today)
    if status != "missing":
        return status
    name_from, name_to = _period_from_display_name(display_name)
    return _lifecycle_status(name_from, name_to, today)


def _resolve_period(
    ledger_from: date | None,
    ledger_to: date | None,
    display_name: str | None,
) -> tuple[date | None, date | None]:
    """期限主数据的写入值（#51）：台账优先，缺位回退名称解析，都没有返回 (None, None)。

    返回 (None, None) 时调用方**不得清空**项目上已有的 period——那可能是迁移
    f3b5d7c9e2a4 按 WBDD 挂靠聚合回填的值，比「没有」更接近事实。
    """
    if ledger_from is not None or ledger_to is not None:
        return (ledger_from, ledger_to)
    return _period_from_display_name(display_name)


def _project_code_for_row(row: MaintenanceLedgerContractRow) -> str:
    return (row.project_name or row.project_name_raw or "未命名项目")[
        :MAX_PROJECT_CODE_LEN
    ]


def _target_project_identities(
    rows: list[MaintenanceLedgerContractRow],
) -> list[str]:
    """Return the canonical lock identities, independent of workbook row order."""
    return sorted({_project_code_for_row(row).lower() for row in rows})


def _lock_contract_evidence_identities(
    db: Session,
    contract_nos: list[str] | set[str] | tuple[str, ...],
) -> list[str]:
    """Serialize a ledger apply with readers that will later lock projects.

    ``apply_batch`` already owns its import-batch row before reaching this
    point.  The remediation tool first locks every existing matching batch,
    then takes these identities, and only then enters the canonical
    workbook-state -> project -> contract protocol.  Keeping this shared
    advisory step between those two phases also covers a newly committed batch
    that did not exist when remediation took its batch-row envelope.
    """

    from app.services.maintenance_expense_integrity import normalize_contract_no

    original = {
        str(contract_no).strip()
        for contract_no in contract_nos
        if contract_no is not None and str(contract_no).strip()
    }
    normalized = {
        normalize_contract_no(contract_no)
        for contract_no in original
        if normalize_contract_no(contract_no)
    }
    # Contract writers elsewhere take the original, bare normalized and
    # XSDD-prefixed normalized identities.  Ledger apply must use the same
    # identity set and order, otherwise historical bare-number rows can race a
    # prefixed ledger row even though ownership treats them as the same key.
    identities = sorted(
        original
        | normalized
        | {f"XSDD-{identity}" for identity in normalized}
    )
    for contract_no in identities:
        db.execute(
            text(
                "SELECT pg_advisory_xact_lock("
                "hashtextextended(:identity, 0))"
            ),
            {"identity": f"maintenance-ledger-contract:{contract_no}"},
        )
    return identities


def _normalized_contract_sql(column):
    """PostgreSQL expression matching expense-integrity contract identity."""

    return func.regexp_replace(
        func.upper(func.regexp_replace(func.btrim(column), r"\s+", "", "g")),
        "^XSDD-",
        "",
    )


def _related_project_ids_by_contract(
    db: Session,
    contract_nos: set[str],
) -> dict[str, set[str]]:
    """既有项目里与本批合同同号（contract_id/contract_no）的关联项目。

    ``_card_contracts`` 按 contract_id/contract_no 全局判定 shared/incomplete，
    同号合同的语义变化（create/amount/effective_to/version）可能让其他项目
    从确定翻成 incomplete 或反向，因此这些项目必须与目标项目同批锁定、
    同批失效导出工作簿。
    """
    from app.services.maintenance_expense_integrity import normalize_contract_no

    related: dict[str, set[str]] = {}
    if not contract_nos:
        return related
    originals_by_identity: dict[str, set[str]] = {}
    for contract_no in contract_nos:
        normalized = normalize_contract_no(contract_no)
        if normalized:
            originals_by_identity.setdefault(normalized, set()).add(contract_no)
    if not originals_by_identity:
        return related
    for contract_id, contract_no, project_id in db.execute(
        select(
            MaintenanceProjectContract.contract_id,
            MaintenanceProjectContract.contract_no,
            MaintenanceProjectContract.project_id,
        ).where(
            or_(
                _normalized_contract_sql(
                    MaintenanceProjectContract.contract_id
                ).in_(sorted(originals_by_identity)),
                _normalized_contract_sql(
                    MaintenanceProjectContract.contract_no
                ).in_(sorted(originals_by_identity)),
            )
        )
    ):
        for key in (contract_id, contract_no):
            normalized = normalize_contract_no(key)
            for original in originals_by_identity.get(normalized, ()):
                related.setdefault(original, set()).add(project_id)
    return related


def _lock_target_projects(
    db: Session,
    rows: list[MaintenanceLedgerContractRow],
    contract_nos: set[str] | None = None,
    extra_project_ids: set[str] | None = None,
) -> tuple[
    list[str],
    dict[str, object],
    dict[str, set[str]],
]:
    """Pre-lock every target and contract-related project before any mutation.

    Advisory identities cover absent rows (``SELECT FOR UPDATE`` cannot), so
    concurrent ledger batches cannot race on the case-insensitive project-code
    unique index.  Existing target projects plus every existing project sharing
    a batch contract_id/contract_no then follow the system-wide
    workbook-state -> project order: **all** workbook states first (one sorted
    pass), then **all** project rows (sorted by stable project id), never an
    interleaved state A -> project A -> state B.  Row order in the uploaded
    workbook is deliberately not part of the database lock order.

    Returns ``(locked_project_ids, workbook_states, related_by_contract)`` so
    the caller can bump revisions without re-locking.
    """
    identities = _target_project_identities(rows)
    for identity in identities:
        db.execute(
            text(
                "SELECT pg_advisory_xact_lock("
                "hashtextextended(:identity, 0))"
            ),
            {"identity": f"maintenance-ledger-project:{identity}"},
        )
    related = _related_project_ids_by_contract(db, contract_nos or set())

    from app.services import maintenance_project_operations as operations

    existing: list[MaintenanceProject] = []
    if identities:
        existing = list(db.scalars(
            select(MaintenanceProject)
            .where(func.lower(MaintenanceProject.project_code).in_(identities))
            .order_by(MaintenanceProject.project_id)
        ))
    all_project_ids = sorted(
        {project.project_id for project in existing}
        | {pid for pids in related.values() for pid in pids}
        | set(extra_project_ids or ())
    )
    workbook_states = operations.lock_workbook_states(
        db, project_ids=all_project_ids
    )
    locked_project_ids: list[str] = []
    for project_id in all_project_ids:
        locked = db.scalar(
            select(MaintenanceProject.project_id)
            .where(MaintenanceProject.project_id == project_id)
            .with_for_update()
        )
        if locked is None:
            raise LedgerBatchError(
                "项目在应用前已不存在，整批未应用，请重试"
            )
        locked_project_ids.append(str(locked))
    return locked_project_ids, workbook_states, related


def _lock_related_contracts(
    db: Session,
    *,
    project_ids: list[str],
    contract_nos: set[str],
) -> list[MaintenanceProjectContract]:
    """Lock matching contract versions after all state/project rows are locked."""

    from app.services.maintenance_expense_integrity import normalize_contract_no

    identities = sorted({
        normalize_contract_no(contract_no)
        for contract_no in contract_nos
        if normalize_contract_no(contract_no)
    })
    if not project_ids or not identities:
        return []
    return list(db.scalars(
        select(MaintenanceProjectContract)
        .where(
            MaintenanceProjectContract.project_id.in_(sorted(project_ids)),
            or_(
                _normalized_contract_sql(
                    MaintenanceProjectContract.contract_id
                ).in_(identities),
                _normalized_contract_sql(
                    MaintenanceProjectContract.contract_no
                ).in_(identities),
            ),
        )
        .order_by(MaintenanceProjectContract.project_contract_id)
        .with_for_update()
        .execution_options(populate_existing=True)
    ))


def _upsert_project(
    db: Session,
    row: MaintenanceLedgerContractRow,
    operated_by: str,
    summary: dict,
    today: date,
    ledger_batch_id: str,
) -> tuple[MaintenanceProject, bool]:
    """Upsert 项目主数据，返回 ``(project, changed)``。

    ``changed`` 只表达既有项目的字段级真实变化（新建不算 field change），
    供 apply_batch 汇总决定是否 bump 工作簿版本；本子函数不做任何 bump。
    """
    from app.services import maintenance_project_identity

    code = _project_code_for_row(row)
    project = db.execute(
        select(MaintenanceProject).where(
            func.lower(MaintenanceProject.project_code) == code.lower()
        )
    ).scalar_one_or_none()
    display_name = row.project_name_raw or code
    # 期限主数据（#51）：台账优先、名称解析兜底；两者都没有时保留项目上已有的
    # 回填值（WBDD 挂靠聚合），并以「实际生效的期限」计算 lifecycle——否则台账
    # 导入会把已回填的期限/状态打回 missing（update 分支对 lifecycle 无条件同步）。
    new_from, new_to = _resolve_period(
        row.project_period_from, row.project_period_to, display_name
    )
    eff_from = new_from if (new_from or new_to) else (
        project.period_from if project is not None else None)
    eff_to = new_to if (new_from or new_to) else (
        project.period_to if project is not None else None)
    lifecycle = _lifecycle_status(eff_from, eff_to, today)
    if project is None:
        project = MaintenanceProject(
            project_id=str(uuid4()),
            project_code=code,
            display_name=display_name,
            project_manager_id=(row.manager or "")[:64] or None,
            business_type=row.business_type or None,
            cmo_name=(row.cmo or "")[:128] or None,
            salesperson=(row.salesperson_raw or "")[:64] or None,
            period_from=new_from,
            period_to=new_to,
            lifecycle_status=lifecycle,
            is_active=True,
            version=1,
        )
        db.add(project)
        try:
            db.flush()
        except IntegrityError as exc:
            # A non-ledger project creator may not participate in the ledger
            # advisory protocol.  Surface the remaining unique race as a
            # controlled whole-batch conflict; the API rolls this transaction
            # back and the caller can safely retry against the now-visible row.
            raise LedgerBatchError(
                f"项目编号「{code}」被并发创建，整批未应用，请重试"
            ) from exc
        maintenance_project_identity.record_alias(
            db,
            project_id=project.project_id,
            alias_name=display_name,
            source=LEDGER_SOURCE,
        )
        db.add(
            MaintenanceProjectAuditLog(
                project_id=project.project_id,
                entity_type="project",
                entity_id=project.project_id,
                action="ledger_create",
                after_json={"display_name": display_name, "code": code},
                reason="台账导入创建项目",
                operated_by=operated_by,
            )
        )
        summary["projects_created"] += 1
        # 台账确实提供了期限时才声明 ledger provenance；空输入不制造空投影。
        if new_from is not None or new_to is not None:
            try:
                apply_canonical_period_locked(
                    db,
                    project=project,
                    period_from=new_from,
                    period_to=new_to,
                    source=SOURCE_LEDGER,
                    ledger_batch_id=ledger_batch_id,
                    as_of=today,
                    operated_by=operated_by,
                    reason="台账导入创建项目期限",
                )
            except MaintenancePeriodError as exc:
                raise LedgerBatchError(str(exc)) from exc
        return project, False
    changed = False
    before = {
        "display_name": project.display_name,
        "lifecycle_status": project.lifecycle_status,
        "period_from": project.period_from.isoformat() if project.period_from else None,
        "period_to": project.period_to.isoformat() if project.period_to else None,
        "business_type": project.business_type,
        "cmo_name": project.cmo_name,
        "salesperson": project.salesperson,
        "salesperson_override_active": project.salesperson_override_active,
        "project_manager_id": project.project_manager_id,
    }
    after = dict(before)
    if project.display_name != display_name:
        maintenance_project_identity.record_alias(
            db,
            project_id=project.project_id,
            alias_name=project.display_name,
            source=LEDGER_SOURCE,
        )
        project.display_name = display_name
        maintenance_project_identity.record_alias(
            db,
            project_id=project.project_id,
            alias_name=display_name,
            source=LEDGER_SOURCE,
        )
        after["display_name"] = display_name
        changed = True
    # 台账/名称给出了期限才覆盖（#51）；项目列是唯一业务事实，service-period
    # 行仅作为 manager workbook 的 OCC/provenance 投影。
    if new_from is not None or new_to is not None:
        try:
            period_result = apply_canonical_period_locked(
                db,
                project=project,
                period_from=new_from,
                period_to=new_to,
                source=SOURCE_LEDGER,
                ledger_batch_id=ledger_batch_id,
                as_of=today,
                operated_by=operated_by,
                reason="台账导入更新项目期限",
            )
        except MaintenancePeriodError as exc:
            raise LedgerBatchError(str(exc)) from exc
        if period_result["project_changed"]:
            after["period_from"] = new_from.isoformat() if new_from else None
            after["period_to"] = new_to.isoformat() if new_to else None
            after["lifecycle_status"] = project.lifecycle_status
            changed = True
    elif project.lifecycle_status != lifecycle:
        # 空期限输入不接管 provenance，但生命周期仍按既有 canonical 日期刷新。
        project.lifecycle_status = lifecycle
        after["lifecycle_status"] = lifecycle
        changed = True
    if row.manager and project.project_manager_id != row.manager[:64]:
        project.project_manager_id = row.manager[:64]
        after["project_manager_id"] = row.manager[:64]
        changed = True
    if row.business_type and project.business_type != row.business_type:
        project.business_type = row.business_type
        after["business_type"] = row.business_type
        changed = True
    if row.cmo and project.cmo_name != row.cmo[:128]:
        project.cmo_name = row.cmo[:128]
        after["cmo_name"] = row.cmo[:128]
        changed = True
    if row.salesperson_raw:
        ledger_salesperson = row.salesperson_raw[:64]
        if (
            project.salesperson != ledger_salesperson
            or project.salesperson_override_active
        ):
            project.salesperson = ledger_salesperson
            project.salesperson_override_active = False
            after["salesperson"] = ledger_salesperson
            after["salesperson_override_active"] = False
            changed = True
    if changed:
        project.version += 1
        db.add(
            MaintenanceProjectAuditLog(
                project_id=project.project_id,
                entity_type="project",
                entity_id=project.project_id,
                action="ledger_update",
                before_json=before,
                after_json=after,
                reason="台账导入更新项目",
                operated_by=operated_by,
            )
        )
        summary["projects_updated"] += 1
    return project, changed


def _sales_order_inc_tax(sales_order: FSalesOrder | None) -> Decimal | None:
    """Derive explicit-rate inc-tax money with the shared HALF_UP policy."""

    if (
        sales_order is None
        or sales_order.amount_ex_tax is None
        or sales_order.tax_rate is None
    ):
        return None
    amount_ex_tax = Decimal(str(sales_order.amount_ex_tax))
    tax_rate = Decimal(str(sales_order.tax_rate))
    return tax_policy.round_money(
        amount_ex_tax * (Decimal("1") + tax_rate)
    )


def _load_sales_order_evidence(
    db: Session,
    order_nos: set[str],
) -> tuple[dict[str, FSalesOrder], set[str]]:
    """Resolve one deterministic successful sales fact per XSDD.

    Only active rows from successful ``sales`` batches are evidence.  The
    latest row is selected deterministically when every surviving version has
    the same economic tuple.  Conflicting amount/tax/inc-tax versions are
    returned separately and deliberately omitted from the usable map.
    """

    if not order_nos:
        return {}, set()
    candidates = list(db.scalars(
        select(FSalesOrder)
        .join(
            SysImportBatch,
            SysImportBatch.id == FSalesOrder.import_batch_id,
        )
        .where(
            FSalesOrder.order_no.in_(sorted(order_nos)),
            FSalesOrder.data_status == config.ACTIVE_STATUS,
            SysImportBatch.file_type == "sales",
            SysImportBatch.status == "success",
        )
        .order_by(
            FSalesOrder.order_no,
            SysImportBatch.uploaded_at.desc().nullslast(),
            FSalesOrder.created_at.desc().nullslast(),
            SysImportBatch.id.desc(),
            FSalesOrder.id.desc(),
        )
    ).all())
    resolved: dict[str, FSalesOrder] = {}
    economics: dict[
        str, set[tuple[Decimal | None, Decimal | None, Decimal | None]]
    ] = {}
    for sales_order in candidates:
        resolved.setdefault(sales_order.order_no, sales_order)
        amount_ex_tax = (
            Decimal(str(sales_order.amount_ex_tax))
            if sales_order.amount_ex_tax is not None else None
        )
        tax_rate = (
            Decimal(str(sales_order.tax_rate))
            if sales_order.tax_rate is not None else None
        )
        economics.setdefault(sales_order.order_no, set()).add((
            amount_ex_tax,
            tax_rate,
            _sales_order_inc_tax(sales_order),
        ))
    conflicts = {
        order_no
        for order_no, values in economics.items()
        if len(values) > 1
    }
    for order_no in conflicts:
        resolved.pop(order_no, None)
    return resolved, conflicts


def _upsert_contract(
    db: Session,
    row: MaintenanceLedgerContractRow,
    project: MaintenanceProject,
    operated_by: str,
    summary: dict,
    *,
    sales_order: FSalesOrder | None,
) -> tuple[MaintenanceProjectContract | None, bool, bool]:
    """Return ``(contract, changed, ownership_changed)`` without bumping.

    workbook-state -> project 锁已在 apply_batch 开头一次性取齐；语义变化
    由 apply_batch 汇总后统一 bump 目标与同号关联项目。只有合同身份首次创建
    或历史有效窗变化会触发 raw-backed attribution 重算；金额变化不会制造费用
    版本噪声。no-op / 纯 provenance 变化不改变 ``changed``。
    """
    from app.services import maintenance_project_operations as operations

    sales_amount_inc_tax = _sales_order_inc_tax(sales_order)
    effective_from = row.project_period_from or row.order_date
    if effective_from is None and sales_order is not None:
        effective_from = sales_order.order_date
    if effective_from is None:
        summary["skipped_rows"] += 1
        return None, False, False
    contract = db.execute(
        select(MaintenanceProjectContract).where(
            MaintenanceProjectContract.project_id == project.project_id,
            MaintenanceProjectContract.contract_no == row.order_no,
        ).with_for_update().execution_options(populate_existing=True)
    ).scalar_one_or_none()
    if contract is None:
        # 台账未给含税额时，只有销售单税率明确才可换算；未知税率保持 NULL。
        inc_from_sales = (
            sales_amount_inc_tax
            if row.amount_inc_tax is None else None
        )
        contract = MaintenanceProjectContract(
            project_contract_id=str(uuid4()),
            project_id=project.project_id,
            contract_id=row.order_no,
            contract_no=row.order_no,
            contract_amount=(sales_order.amount_ex_tax if sales_order is not None else None),
            amount_inc_tax=row.amount_inc_tax if row.amount_inc_tax is not None else inc_from_sales,
            contract_status=None,
            status_mapping_state="mapped",
            status_mapping_version=LEDGER_SOURCE,
            included_in_total=True,
            effective_from=effective_from,
            effective_to=row.project_period_to,
            source=LEDGER_SOURCE,
            version=1,
        )
        db.add(contract)
        db.flush()
        db.add(
            MaintenanceProjectAuditLog(
                project_id=project.project_id,
                entity_type="project_contract",
                entity_id=contract.contract_no,
                action="ledger_create",
                after_json={
                    "contract_no": row.order_no,
                    "amount_inc_tax": (
                        str(row.amount_inc_tax) if row.amount_inc_tax is not None else None
                    ),
                    "effective_from": str(effective_from),
                },
                reason="台账导入创建合同",
                operated_by=operated_by,
            )
        )
        summary["contracts_created"] += 1
        return contract, True, True
    before = operations.contract_dict(contract)
    changed = False
    ownership_changed = False
    amount_changed_from_ledger = False
    if (contract.source != "project_master_workbook"
            and row.amount_inc_tax is not None
            and contract.amount_inc_tax != row.amount_inc_tax):
        contract.amount_inc_tax = row.amount_inc_tax
        changed = True
        amount_changed_from_ledger = True
    elif (contract.source != "project_master_workbook"
          and contract.amount_inc_tax is None
          and sales_amount_inc_tax is not None):
        contract.amount_inc_tax = sales_amount_inc_tax
        changed = True
        amount_changed_from_ledger = True
    if sales_order is not None and contract.contract_amount != sales_order.amount_ex_tax:
        contract.contract_amount = sales_order.amount_ex_tax
        changed = True
    if contract.effective_from != effective_from:
        contract.effective_from = effective_from
        changed = True
        ownership_changed = True
    if contract.effective_to != row.project_period_to:
        contract.effective_to = row.project_period_to
        changed = True
        ownership_changed = True
    if changed:
        if amount_changed_from_ledger:
            contract.source = LEDGER_SOURCE
        contract.version += 1
        after = operations.contract_dict(contract)
        db.add(
            MaintenanceProjectAuditLog(
                project_id=project.project_id,
                entity_type="project_contract",
                entity_id=contract.project_contract_id,
                action="ledger_update",
                before_json=before,
                after_json=after,
                reason="台账导入更新合同",
                operated_by=operated_by,
            )
        )
        summary["contracts_updated"] += 1
    return contract, changed, ownership_changed


def _upsert_milestone(
    db: Session,
    row: MaintenanceLedgerPlanRow,
    contract: MaintenanceProjectContract,
    ledger_batch_id: str,
    summary: dict,
) -> bool:
    """回款计划节点走唯一受控 helper（保留已跟进待复核语义）。

    返回是否有计划事实（日期/金额/完整度）级真实变化。金额可空，比较必须
    先判 None，不能 ``Decimal(None)``；来源/批次只是引用关系，纯来源或批次
    变化不算 semantic update（受控 helper 同样只为计划事实 bump version）。
    """
    from app.services import maintenance_collection_milestones as milestone_svc

    if row.planned_date is None and row.planned_amount is None:
        return False
    completeness = "complete"
    if row.planned_date is None:
        completeness = "amount_only"
    elif row.planned_amount is None:
        completeness = "date_only"
    existing = db.execute(
        select(MaintenanceCollectionMilestone).where(
            MaintenanceCollectionMilestone.project_contract_id
            == contract.project_contract_id,
            MaintenanceCollectionMilestone.sequence == row.sequence,
        )
    ).scalar_one_or_none()
    if existing is None:
        semantic_changed = True
    else:
        old_amount = existing.planned_amount
        new_amount = row.planned_amount
        amount_changed = (
            (old_amount is None) != (new_amount is None)
            or (
                old_amount is not None
                and new_amount is not None
                and old_amount != new_amount
            )
        )
        semantic_changed = (
            existing.planned_date != row.planned_date
            or amount_changed
            or existing.completeness_state != completeness
        )
    # 台账来源默认精度为 month；由 helper 按来源派生并校验。
    milestone_svc.write_collection_milestone(
        db,
        project_id=contract.project_id,
        project_contract_id=contract.project_contract_id,
        sequence=row.sequence,
        planned_date=row.planned_date,
        planned_amount=row.planned_amount,
        completeness_state=completeness,
        source=LEDGER_SOURCE,
        ledger_batch_id=ledger_batch_id,
        date_precision=None,
    )
    if existing is None:
        summary["milestones_created"] += 1
    elif semantic_changed:
        summary["milestones_updated"] += 1
    return semantic_changed


_EXPENSE_RAW_FINGERPRINT_FIELDS = (
    "raw_line_id",
    "bxd_no",
    "line_no",
    "data_status",
    "expense_date",
    "person",
    "expense_type",
    "fee_category",
    "reason",
    "linked_sales_order_no",
    "amount",
    "amount_ex_tax",
    "amount_inc_tax",
    "tax_basis",
    "tax_rate_used",
    "remark",
    "import_batch_id",
)


def _expense_raw_fingerprint(raw_rows) -> str:
    """Stable hash of every raw fact that can affect reconciliation or sync."""

    def stable(value) -> str:
        if value is None:
            return "<NULL>"
        if isinstance(value, (date, datetime)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return format(value, "f")
        return str(value)

    snapshot = tuple(
        tuple(stable(getattr(raw, field_name)) for field_name in
              _EXPENSE_RAW_FINGERPRINT_FIELDS)
        for raw in sorted(raw_rows, key=lambda item: item.raw_line_id)
    )
    return hashlib.sha256(repr(snapshot).encode("utf-8")).hexdigest()


def _load_formal_expense_rows(
    db: Session,
    *,
    bxd_nos: list[str],
    lock: bool = False,
):
    """Load every successful formal row for the requested BXD identities."""

    from app.models.maintenance import FProjectExpense

    if not bxd_nos:
        return []
    statement = (
        select(FProjectExpense)
        .join(SysImportBatch, SysImportBatch.id == FProjectExpense.import_batch_id)
        .where(
            FProjectExpense.bxd_no.in_(bxd_nos),
            SysImportBatch.status == "success",
        )
        .order_by(FProjectExpense.bxd_no, FProjectExpense.raw_line_id)
    )
    if lock:
        # Do not late-lock SysImportBatch rows after project/contract rows.
        statement = statement.with_for_update(of=FProjectExpense)
    return list(db.scalars(
        statement.execution_options(populate_existing=lock)
    ))


def _build_expense_sync_plans(
    *,
    expense_rows: list[MaintenanceLedgerExpenseRow],
    contract_order_nos: set[str],
    raw_rows,
) -> tuple[list[LedgerExpenseSyncPlan], list[str]]:
    """Pure reconciliation used before and after the canonical lock envelope."""

    from collections import defaultdict

    from app.services.maintenance_expense_integrity import normalize_contract_no

    failures: list[str] = []
    by_bxd: dict[str, list[MaintenanceLedgerExpenseRow]] = defaultdict(list)
    for row in expense_rows:
        if row.bxd_no:
            by_bxd[row.bxd_no].append(row)
    raw_by_bxd: dict[str, list] = defaultdict(list)
    for raw in raw_rows:
        if raw.bxd_no:
            raw_by_bxd[str(raw.bxd_no)].append(raw)

    contract_by_normalized: dict[str, str] = {}
    duplicate_contract_identities: set[str] = set()
    for order_no in sorted(contract_order_nos):
        normalized = normalize_contract_no(order_no)
        if not normalized:
            continue
        if normalized in contract_by_normalized:
            duplicate_contract_identities.add(normalized)
        contract_by_normalized[normalized] = order_no

    plans: list[LedgerExpenseSyncPlan] = []
    for bxd_no, rows in sorted(by_bxd.items()):
        if any(row.amount is None for row in rows):
            failures.append(f"{bxd_no}: 台账含税报销金额缺失")
            continue
        ledger_amount = sum((Decimal(row.amount) for row in rows), Decimal("0.00"))
        declared_orders = {
            normalize_contract_no(row.sales_order)
            for row in rows
            if normalize_contract_no(row.sales_order)
        }
        if len(declared_orders) > 1:
            failures.append(f"{bxd_no}: 台账多行销售订单不一致")
            continue
        observed_raw = sorted(
            raw_by_bxd.get(bxd_no, ()), key=lambda item: item.raw_line_id
        )
        active_raw = [
            raw for raw in observed_raw
            if raw.data_status == config.MAINT_EXPENSE_ACTIVE_STATUS
        ]
        if not active_raw:
            failures.append(f"{bxd_no}: 未找到生效的正式 BXD 报销行")
            continue
        if any(
            raw.expense_date is None
            or raw.amount_ex_tax is None
            or raw.amount_inc_tax is None
            for raw in active_raw
        ):
            failures.append(f"{bxd_no}: 正式 BXD 缺少日期或双税金额")
            continue
        normalized_raw_orders = [
            normalize_contract_no(raw.linked_sales_order_no)
            for raw in active_raw
        ]
        if (
            any(not order_no for order_no in normalized_raw_orders)
            or len(set(normalized_raw_orders)) != 1
        ):
            failures.append(f"{bxd_no}: 正式 BXD 销售订单缺失或不唯一")
            continue
        normalized_order = normalized_raw_orders[0]
        if declared_orders and declared_orders != {normalized_order}:
            failures.append(f"{bxd_no}: 台账销售订单与正式 BXD 不一致")
            continue
        if normalized_order in duplicate_contract_identities:
            failures.append(f"{bxd_no}: 台账合同身份重复，无法唯一归集")
            continue
        target_order_no = contract_by_normalized.get(normalized_order)
        if target_order_no is None:
            failures.append(f"{bxd_no}: 正式 BXD 销售订单不在本批合同清单")
            continue
        formal_amount = sum(
            (Decimal(raw.amount_inc_tax) for raw in active_raw), Decimal("0.00")
        )
        if ledger_amount != formal_amount:
            failures.append(
                f"{bxd_no}: 台账含税 {ledger_amount} ≠ 正式 BXD 含税 {formal_amount}"
            )
            continue
        plans.append(LedgerExpenseSyncPlan(
            bxd_no=bxd_no,
            target_order_no=target_order_no,
            raw_line_ids=tuple(raw.raw_line_id for raw in active_raw),
            observed_raw_line_ids=tuple(raw.raw_line_id for raw in observed_raw),
            expected_raw_fingerprint=_expense_raw_fingerprint(observed_raw),
        ))
    return plans, failures


def _expense_attribution_owners(db: Session, raw_ids: list[str]) -> set[str]:
    from app.models.maintenance_project_operations import (
        MaintenanceProjectExpenseAttribution,
    )
    from app.services.maintenance_expense_integrity import expense_id_for

    if not raw_ids:
        return set()
    expense_ids = [expense_id_for(raw_id) for raw_id in raw_ids]
    return set(db.scalars(
        select(MaintenanceProjectExpenseAttribution.project_id).where(
            or_(
                MaintenanceProjectExpenseAttribution.raw_expense_line_id.in_(raw_ids),
                MaintenanceProjectExpenseAttribution.expense_id.in_(expense_ids),
            )
        )
    ))


def _prepare_contract_expense_probe(
    db: Session,
    *,
    contract_order_nos: set[str],
) -> LedgerContractExpenseProbe:
    """Snapshot existing raw-backed attributions affected by contract windows."""

    from app.models.maintenance import FProjectExpense
    from app.models.maintenance_project_operations import (
        MaintenanceProjectExpenseAttribution,
    )
    from app.services.maintenance_expense_integrity import (
        expense_id_for,
        normalize_contract_no,
    )

    identities = sorted({
        normalize_contract_no(contract_no)
        for contract_no in contract_order_nos
        if normalize_contract_no(contract_no)
    })
    if not identities:
        return LedgerContractExpenseProbe(
            (), (), _expense_raw_fingerprint(()), frozenset()
        )
    candidates = list(db.scalars(
        select(FProjectExpense)
        .where(
            FProjectExpense.linked_sales_order_no.is_not(None),
            _normalized_contract_sql(
                FProjectExpense.linked_sales_order_no
            ).in_(identities),
        )
        .order_by(FProjectExpense.raw_line_id)
    ))
    if not candidates:
        return LedgerContractExpenseProbe(
            (), (), _expense_raw_fingerprint(()), frozenset()
        )
    candidate_by_id = {raw.raw_line_id: raw for raw in candidates}
    expense_to_raw = {
        expense_id_for(raw_id): raw_id for raw_id in candidate_by_id
    }
    attributions = list(db.scalars(
        select(MaintenanceProjectExpenseAttribution).where(
            or_(
                MaintenanceProjectExpenseAttribution.raw_expense_line_id.in_(
                    sorted(candidate_by_id)
                ),
                MaintenanceProjectExpenseAttribution.expense_id.in_(
                    sorted(expense_to_raw)
                ),
            )
        )
    ))
    impacted_ids: set[str] = set()
    owner_project_ids: set[str] = set()
    for attribution in attributions:
        raw_id = attribution.raw_expense_line_id
        if raw_id not in candidate_by_id:
            raw_id = expense_to_raw.get(attribution.expense_id)
        if raw_id is None:
            continue
        impacted_ids.add(raw_id)
        owner_project_ids.add(attribution.project_id)
    impacted_raw = [candidate_by_id[raw_id] for raw_id in sorted(impacted_ids)]
    return LedgerContractExpenseProbe(
        raw_line_ids=tuple(raw.raw_line_id for raw in impacted_raw),
        raw_identities=tuple(
            (raw.raw_line_id, normalize_contract_no(raw.linked_sales_order_no))
            for raw in impacted_raw
        ),
        expected_raw_fingerprint=_expense_raw_fingerprint(impacted_raw),
        owner_project_ids=frozenset(owner_project_ids),
    )


def _prepare_expense_sync_plans(
    db: Session,
    *,
    expense_rows: list[MaintenanceLedgerExpenseRow],
    contract_order_nos: set[str],
) -> tuple[list[LedgerExpenseSyncPlan], set[str], list[str]]:
    """Reconcile ledger totals and snapshot every raw expectation before locks."""

    if not expense_rows:
        return [], set(), []
    bxd_nos = sorted({row.bxd_no for row in expense_rows if row.bxd_no})
    raw_rows = _load_formal_expense_rows(db, bxd_nos=bxd_nos)
    plans, failures = _build_expense_sync_plans(
        expense_rows=expense_rows,
        contract_order_nos=contract_order_nos,
        raw_rows=raw_rows,
    )
    observed_ids = sorted({
        raw_id
        for plan in plans
        for raw_id in plan.observed_raw_line_ids
    })
    return plans, _expense_attribution_owners(db, observed_ids), failures


def apply_batch(db: Session, batch_id: str, operated_by: str) -> dict:
    """把 raw 行同步进 canonical 表。失败关闭：任何关键异常行都整批零写。"""
    # Serialize the same batch before reading its status.  ``populate_existing``
    # matters when the API ownership probe already placed a stale ORM instance
    # in this Session while another request was finishing the apply.
    batch = db.scalar(
        select(MaintenanceLedgerImportBatch)
        .where(MaintenanceLedgerImportBatch.batch_id == batch_id)
        .with_for_update()
        .execution_options(populate_existing=True)
    )
    if batch is None:
        raise LedgerBatchError("台账批次不存在")
    if batch.status == "applied":
        raise LedgerBatchError("台账批次已应用，不能重复应用")
    if batch.status == "failed":
        raise LedgerBatchError("台账批次已因异常被拒绝，需重新上传")
    issue_rows = (
        db.execute(
            select(MaintenanceLedgerContractRow).where(
                MaintenanceLedgerContractRow.batch_id == batch_id,
                func.cardinality(MaintenanceLedgerContractRow.issues) > 0,
            )
        ).scalars().all()
    )
    issue_plan_rows = (
        db.execute(
            select(MaintenanceLedgerPlanRow).where(
                MaintenanceLedgerPlanRow.batch_id == batch_id,
                func.cardinality(MaintenanceLedgerPlanRow.issues) > 0,
            )
        ).scalars().all()
    )
    contract_rows = (
        db.execute(
            select(MaintenanceLedgerContractRow)
            .where(MaintenanceLedgerContractRow.batch_id == batch_id)
            .order_by(MaintenanceLedgerContractRow.row_no)
        )
        .scalars()
        .all()
    )
    contract_order_nos = {
        row.order_no for row in contract_rows if row.order_no is not None
    }
    sales_orders, sales_conflicts = _load_sales_order_evidence(
        db, contract_order_nos
    )
    # 对账预检：销售单存在时，台账含税额必须与未税额×(1+税率) 一致，否则整批拒绝。
    reconcile_failures: list[str] = []
    for row in contract_rows:
        if row.order_no is None:
            continue
        if row.order_no in sales_conflicts:
            reconcile_failures.append(
                f"{row.order_no}: 多条生效销售成功批次的未税额/税率存在冲突，"
                "无法确定合同金额"
            )
            continue
        if row.amount_inc_tax is None:
            continue
        sales_order = sales_orders.get(row.order_no)
        expected_inc = _sales_order_inc_tax(sales_order)
        # 税率未知时无法从未税额证明台账含税额对错；保留台账权威值，不能
        # 再拿全局 13% 假设把整个批次拒绝。
        if expected_inc is None:
            continue
        if row.amount_inc_tax != expected_inc:
            reconcile_failures.append(
                f"{row.order_no}: 台账含税 {row.amount_inc_tax} ≠ "
                f"销售未税 {sales_order.amount_ex_tax}×"
                f"(1+{sales_order.tax_rate}) = {expected_inc}"
            )
    if reconcile_failures:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejection_reason": "合同金额对账失败",
            "reconcile_failures": reconcile_failures[:20],
        }
        db.commit()
        raise LedgerBatchError(
            f"台账批次金额对账失败 {len(reconcile_failures)} 行，整批拒绝应用"
        )
    expense_rows = list(db.scalars(
        select(MaintenanceLedgerExpenseRow)
        .where(MaintenanceLedgerExpenseRow.batch_id == batch_id)
        .order_by(MaintenanceLedgerExpenseRow.row_no)
    ))
    issue_expense_rows = [row for row in expense_rows if row.issues]
    if issue_expense_rows:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejected_rows": len(issue_expense_rows),
            "rejection_reason": "台账批次报销归集存在关键异常行，整批拒绝应用",
        }
        db.commit()
        raise LedgerBatchError(
            f"台账批次存在 {len(issue_expense_rows)} 行报销归集异常，整批拒绝应用"
        )
    if issue_rows or issue_plan_rows:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejected_rows": len(issue_rows) + len(issue_plan_rows),
            "rejection_reason": "台账批次存在关键异常行，整批拒绝应用",
        }
        db.commit()
        raise LedgerBatchError(
            f"台账批次存在 {len(issue_rows) + len(issue_plan_rows)} 行关键异常，"
            "整批拒绝应用（raw 已保留）"
        )
    # 孤儿回款计划行（无对应合同行）fail-closed：不静默丢弃（round-4 Blocker 7）
    orphan_plan_rows = [
        row
        for row in db.execute(
            select(MaintenanceLedgerPlanRow).where(
                MaintenanceLedgerPlanRow.batch_id == batch_id
            )
        ).scalars()
        if row.order_no not in contract_order_nos
    ]
    if orphan_plan_rows:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejected_rows": len(orphan_plan_rows),
            "rejection_reason": "台账批次回款计划存在无对应合同的孤儿行，整批拒绝应用",
        }
        db.commit()
        raise LedgerBatchError(
            f"台账批次存在 {len(orphan_plan_rows)} 行无对应合同的回款计划，"
            "整批拒绝应用（raw 已保留）"
        )
    expense_sync_plans, expense_old_project_ids, expense_failures = (
        _prepare_expense_sync_plans(
            db,
            expense_rows=expense_rows,
            contract_order_nos=contract_order_nos,
        )
    )
    if expense_failures:
        batch.status = "failed"
        batch.report_json = {
            **(batch.report_json or {}),
            "rejection_reason": "报销台账与正式 BXD 对账失败",
            "expense_reconcile_failures": expense_failures[:20],
        }
        db.commit()
        raise LedgerBatchError(
            f"台账批次报销对账失败 {len(expense_failures)} 项，整批拒绝应用"
        )
    contract_expense_probe = _prepare_contract_expense_probe(
        db, contract_order_nos=contract_order_nos
    )
    today = business_today()
    summary = {
        "projects_created": 0,
        "projects_updated": 0,
        "contracts_created": 0,
        "contracts_updated": 0,
        "milestones_created": 0,
        "milestones_updated": 0,
        "expenses_created": 0,
        "expenses_updated": 0,
        "skipped_rows": 0,
    }
    plan_rows = (
        db.execute(
            select(MaintenanceLedgerPlanRow)
            .where(MaintenanceLedgerPlanRow.batch_id == batch_id)
            .order_by(MaintenanceLedgerPlanRow.row_no)
        )
        .scalars()
        .all()
    )
    # Global order for contract evidence writers/readers:
    # import batch -> XSDD advisory -> contract advisory -> project-code
    # advisory -> workbook states (one sorted pass) -> projects (sorted) ->
    # contracts/milestones.  Direct contract writes use the same XSDD ->
    # contract-advisory order, avoiding an AB-BA deadlock on the two identities.
    # The remediation utility uses the same advisory identities after it has
    # locked the matching batch/row evidence envelope, so neither side can hold
    # workbook state while waiting for the other's batch lock.
    from app.services import maintenance_project_identity

    # XSDD identity must be locked before workbook/project rows.  Otherwise
    # the database trigger would first acquire it during contract INSERT while
    # this transaction already holds project locks, inverting the lock order
    # used by manual assignment and master-workbook apply.
    maintenance_project_identity.lock_xsdd_identities(db, contract_order_nos)
    _lock_contract_evidence_identities(db, contract_order_nos)
    locked_project_ids, workbook_states, related_by_contract = _lock_target_projects(
        db,
        contract_rows,
        contract_nos=contract_order_nos,
        extra_project_ids=(
            expense_old_project_ids
            | set(contract_expense_probe.owner_project_ids)
        ),
    )
    _lock_related_contracts(
        db,
        project_ids=locked_project_ids,
        contract_nos=contract_order_nos,
    )
    from app.services import maintenance_project_operations as operations

    # 真实变化集中在这里汇总：project 字段、合同语义（create/amount/
    # effective_to/version）、milestone 计划事实。no-op / 纯 provenance
    # 变化不进集合；一个项目多合同/多节点同事务只 bump 一次。
    changed_project_ids: set[str] = set()
    new_project_ids: set[str] = set()
    ownership_changed_identities: set[str] = set()
    contract_by_order: dict[str, MaintenanceProjectContract] = {}
    from app.services.maintenance_expense_integrity import normalize_contract_no

    for row in contract_rows:
        if row.order_no is None:
            summary["skipped_rows"] += 1
            continue
        project, project_changed = _upsert_project(
            db, row, operated_by, summary, today, ledger_batch_id=batch_id
        )
        if project.project_id not in locked_project_ids:
            new_project_ids.add(project.project_id)
        if project_changed:
            changed_project_ids.add(project.project_id)
        try:
            maintenance_project_identity.claim_xsdd_project(
                db,
                value=row.order_no,
                project_id=project.project_id,
                source=LEDGER_SOURCE,
            )
        except maintenance_project_identity.XsddProjectConflict as exc:
            raise LedgerBatchError(str(exc)) from exc
        contract, contract_changed, ownership_changed = _upsert_contract(
            db,
            row,
            project,
            operated_by,
            summary,
            sales_order=sales_orders.get(row.order_no),
        )
        if contract is not None:
            contract_by_order[row.order_no] = contract
        if ownership_changed:
            normalized = normalize_contract_no(row.order_no)
            if normalized:
                ownership_changed_identities.add(normalized)
        if contract_changed:
            # 同号合同语义变化会翻转其他项目卡片上的 shared/incomplete 判定
            # （_card_contracts 按 contract_id/contract_no 全局判断），目标
            # 项目与相关项目一起失效导出工作簿。
            changed_project_ids.add(project.project_id)
            changed_project_ids.update(
                related_by_contract.get(row.order_no, ())
            )
    for row in plan_rows:
        contract = contract_by_order.get(row.order_no)
        if contract is None:
            summary["skipped_rows"] += 1
            continue
        if not 1 <= row.sequence <= 24:
            summary["skipped_rows"] += 1
            continue
        if _upsert_milestone(db, row, contract, batch_id, summary):
            changed_project_ids.add(contract.project_id)
    contract_resync_raw_ids: set[str] = set()
    contract_probe_lock_ids: set[str] = set()
    if ownership_changed_identities:
        # The test/session factory deliberately disables autoflush.  Ownership
        # resolution below executes fresh SQL, so make the just-written
        # contract identity/window visible to that query before re-probing and
        # syncing the raw-backed attributions.
        db.flush()
        # Re-probe after contract locks/writes but before attribution/raw locks.
        # Any concurrent membership/raw change invalidates the prelocked owner
        # envelope; retrying is safer than taking a late workbook-state lock.
        current_probe = _prepare_contract_expense_probe(
            db, contract_order_nos=contract_order_nos
        )
        if current_probe != contract_expense_probe:
            raise LedgerBatchError(
                "合同关联报销在应用前已变化，整批未写入，请重试"
            )
        contract_probe_lock_ids.update(contract_expense_probe.raw_line_ids)
        # The prelocked probe is already scoped to identities carried by this
        # batch.  Re-resolving the full envelope is deliberate: overlapping
        # normalized identities can change ambiguity even when only one row's
        # effective window changed.  Preserved mapping-version fields keep true
        # no-ops from bumping attribution versions.
        contract_resync_raw_ids.update(contract_expense_probe.raw_line_ids)

    if expense_sync_plans or contract_resync_raw_ids:
        from app.models.maintenance import FProjectExpense
        from app.models.maintenance_project_operations import (
            MaintenanceProjectExpenseAttribution,
        )
        from app.services.maintenance_expense_integrity import (
            ExpenseIntegrityError,
            OWNERSHIP_MAPPING_VERSION,
            OwnershipConflictError,
            expense_id_for,
            expense_ref_for,
            find_ownership_candidates,
            sync_attribution_from_raw,
        )

        observed_expense_raw_ids = {
            raw_id
            for expense_plan in expense_sync_plans
            for raw_id in expense_plan.observed_raw_line_ids
        }
        raw_ids = sorted(observed_expense_raw_ids | contract_probe_lock_ids)
        for raw_id in raw_ids:
            db.execute(select(func.pg_advisory_xact_lock(
                func.hashtextextended(f"maintenance-expense-row:{raw_id}", 0)
            )))
        expense_ids = [expense_id_for(raw_id) for raw_id in raw_ids]
        locked_attributions = list(db.scalars(
            select(MaintenanceProjectExpenseAttribution)
            .where(or_(
                MaintenanceProjectExpenseAttribution.raw_expense_line_id.in_(raw_ids),
                MaintenanceProjectExpenseAttribution.expense_id.in_(expense_ids),
            ))
            .order_by(MaintenanceProjectExpenseAttribution.expense_id)
            .with_for_update()
            .execution_options(populate_existing=True)
        ))
        # This check belongs immediately after attribution locking.  An owner
        # outside the prelocked state envelope is a controlled OCC conflict;
        # never repair it by taking a state lock after attribution/raw locks.
        unexpected_owner_ids = sorted({
            attribution.project_id
            for attribution in locked_attributions
            if attribution.project_id not in workbook_states
        })
        if unexpected_owner_ids:
            raise LedgerBatchError(
                "报销归属项目在应用前已变化，整批未写入，请重试"
            )
        raw_by_expense_id = dict(zip(expense_ids, raw_ids))
        attribution_by_raw_id: dict[
            str, MaintenanceProjectExpenseAttribution
        ] = {}
        for attribution in locked_attributions:
            raw_id = attribution.raw_expense_line_id
            if raw_id not in raw_ids:
                raw_id = raw_by_expense_id.get(attribution.expense_id)
            if raw_id is None:
                continue
            prior = attribution_by_raw_id.get(raw_id)
            if prior is not None and prior.expense_id != attribution.expense_id:
                raise LedgerBatchError(
                    "同一正式报销行存在多条归因，整批未写入，请先治理重复数据"
                )
            attribution_by_raw_id[raw_id] = attribution
        raw_by_id = {
            raw.raw_line_id: raw
            for raw in db.scalars(
                select(FProjectExpense)
                .where(FProjectExpense.raw_line_id.in_(raw_ids))
                .order_by(FProjectExpense.raw_line_id)
                .with_for_update()
                .execution_options(populate_existing=True)
            )
        }
        if set(raw_by_id) != set(raw_ids):
            raise LedgerBatchError(
                "正式 BXD 在应用前已变化，整批未写入，请重试"
            )

        if expense_sync_plans:
            bxd_nos = sorted({plan.bxd_no for plan in expense_sync_plans})
            locked_formal_rows = _load_formal_expense_rows(
                db, bxd_nos=bxd_nos, lock=True
            )
            locked_plans, locked_failures = _build_expense_sync_plans(
                expense_rows=expense_rows,
                contract_order_nos=contract_order_nos,
                raw_rows=locked_formal_rows,
            )
            if locked_failures or locked_plans != expense_sync_plans:
                raise LedgerBatchError(
                    "正式 BXD 的生效集合、状态、日期、金额或订单在应用前已变化，"
                    "整批未写入，请重试"
                )
            raw_by_id.update({
                raw.raw_line_id: raw for raw in locked_formal_rows
            })

        locked_contract_raw = [
            raw_by_id[raw_id]
            for raw_id in contract_expense_probe.raw_line_ids
            if raw_id in raw_by_id
        ]
        if contract_resync_raw_ids and (
            len(locked_contract_raw) != len(contract_expense_probe.raw_line_ids)
            or _expense_raw_fingerprint(locked_contract_raw)
            != contract_expense_probe.expected_raw_fingerprint
        ):
            raise LedgerBatchError(
                "合同关联报销原始事实在应用前已变化，整批未写入，请重试"
            )

        explicit_plan_by_raw_id = {
            raw_id: expense_plan
            for expense_plan in expense_sync_plans
            for raw_id in expense_plan.raw_line_ids
        }
        sync_raw_ids = sorted(
            set(explicit_plan_by_raw_id) | contract_resync_raw_ids
        )
        expense_audit_counts: dict[str, dict[str, int]] = {}
        for raw_id in sync_raw_ids:
            raw = raw_by_id[raw_id]
            expense_plan = explicit_plan_by_raw_id.get(raw_id)
            expected_contract = None
            existing_attribution = attribution_by_raw_id.get(raw_id)
            if expense_plan is not None:
                expected_contract = contract_by_order.get(
                    expense_plan.target_order_no
                )
                if expected_contract is None:
                    raise LedgerBatchError(
                        f"{expense_plan.bxd_no}: 合同应用后不存在，整批未写入"
                    )
                target_project_id = expected_contract.project_id
                status_mapping_version = LEDGER_SOURCE
                ownership_mapping_version = OWNERSHIP_MAPPING_VERSION
            else:
                if existing_attribution is None:
                    raise LedgerBatchError(
                        "合同关联报销归因在应用前已变化，整批未写入，请重试"
                    )
                if raw.expense_date is None:
                    raise LedgerBatchError(
                        "合同变更命中缺少报销日期的既有费用，整批未写入"
                    )
                candidates = find_ownership_candidates(
                    db,
                    linked_sales_order_no=raw.linked_sales_order_no,
                    expense_date=raw.expense_date,
                )
                candidate_project_ids = {
                    candidate.project_id for candidate in candidates
                }
                target_project_id = (
                    next(iter(candidate_project_ids))
                    if len(candidate_project_ids) == 1
                    else existing_attribution.project_id
                )
                status_mapping_version = (
                    existing_attribution.status_mapping_version
                )
                ownership_mapping_version = (
                    existing_attribution.ownership_mapping_version
                    or OWNERSHIP_MAPPING_VERSION
                )
            if (
                target_project_id not in workbook_states
                and target_project_id not in new_project_ids
            ):
                raise LedgerBatchError(
                    "合同变更产生了预锁范围外的报销归属，整批未写入，请重试"
                )
            duplicate = db.scalar(
                select(MaintenanceProjectExpenseAttribution.expense_id).where(
                    MaintenanceProjectExpenseAttribution.project_id
                    == target_project_id,
                    MaintenanceProjectExpenseAttribution.expense_ref
                    == expense_ref_for(raw),
                    MaintenanceProjectExpenseAttribution.expense_id
                    != expense_id_for(raw_id),
                )
            )
            if duplicate is not None:
                raise LedgerBatchError(
                    "合同变更命中同项目重复报销单号与明细序号，整批未写入"
                )
            try:
                result = sync_attribution_from_raw(
                    db,
                    raw=raw,
                    project_id=target_project_id,
                    status_mapping_version=status_mapping_version,
                    ownership_mapping_version=ownership_mapping_version,
                )
            except (ExpenseIntegrityError, OwnershipConflictError) as exc:
                label = expense_plan.bxd_no if expense_plan is not None else raw.bxd_no
                raise LedgerBatchError(
                    f"{label or raw_id}: 报销归属无法唯一确认，整批未写入"
                ) from exc
            attribution = db.get(
                MaintenanceProjectExpenseAttribution, result.expense_id
            )
            if expected_contract is not None and (
                attribution is None
                or attribution.ownership_mapping_state != "mapped"
                or attribution.project_contract_id
                != expected_contract.project_contract_id
            ):
                raise LedgerBatchError(
                    f"{expense_plan.bxd_no}: 发生日未命中本批唯一合同，整批未写入"
                )
            if not result.changed:
                continue
            summary[
                "expenses_created" if result.created else "expenses_updated"
            ] += 1
            changed_project_ids.update(result.affected_project_ids)
            for project_id in result.affected_project_ids:
                counts = expense_audit_counts.setdefault(
                    project_id, {"created": 0, "updated": 0}
                )
                counts["created" if result.created else "updated"] += 1
        for project_id, counts in expense_audit_counts.items():
            operations._fact_audit(
                db,
                project_id=project_id,
                entity_type="expense",
                entity_id=f"ledger:{batch_id[:32]}",
                action="bulk_sync",
                before=None,
                after=counts,
                reason="台账合同/报销与正式 BXD 对账后同步",
                operated_by=operated_by,
            )
    for project_id in sorted(changed_project_ids):
        state = workbook_states.get(project_id)
        if state is None:
            if project_id not in new_project_ids:
                raise LedgerBatchError(
                    "费用或合同变更触及预锁范围外的既有项目，整批未写入，请重试"
                )
            # 本批新建项目不在预锁集合内：其 state 行此刻才创建，并发事务在
            # 本项目提交前看不到该项目，不存在锁序竞争。
            state = operations.get_or_create_workbook_state(
                db, project_id=project_id, lock=True
            )
        operations.bump_locked_workbook_revision(db, state=state)
    batch.status = "applied"
    batch.applied_by = operated_by
    batch.applied_at = datetime.now(timezone.utc)
    db.commit()
    return summary
