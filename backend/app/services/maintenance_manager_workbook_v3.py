"""Pure project-manager monthly workbook v3 protocol (#206).

The workbook is deliberately independent from the frozen project workbook v2.
It contains only the authenticated manager's explicit project scope, stores
planned collection milestones longitudinally, and never exposes a write path to
financially confirmed collections.  Database transactions live in the adapter;
this module only builds a signed workbook and returns a side-effect-free plan.
"""

from __future__ import annotations

import hashlib
import hmac
import io
import json
import math
import re
import uuid
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PROTOCOL_ID = "ITDATA_MAINT_MANAGER_WORKBOOK/3.0"
SCHEMA_VERSION = "3.0"
# 2026-08-25 拍板：验收无截止日概念，总览移除「验收报告截止日」列，模板版本随之升位。
TEMPLATE_VERSION = "3.2.0"

OVERVIEW_SHEET = "01_本人项目总览"
PLAN_SHEET = "02_计划回款节点"
ACCEPTANCE_SHEET = "03_验收材料"
INSTRUCTIONS_SHEET = "04_使用说明与校验"
DICTIONARY_SHEET = "98_字典"
ENTITY_VERSION_SHEET = "99_实体版本"
METADATA_SHEET = "99_元数据"
VISIBLE_SHEETS = (
    OVERVIEW_SHEET,
    PLAN_SHEET,
    ACCEPTANCE_SHEET,
    INSTRUCTIONS_SHEET,
)
HIDDEN_SHEETS = (DICTIONARY_SHEET, ENTITY_VERSION_SHEET, METADATA_SHEET)
SHEET_NAMES = VISIBLE_SHEETS + HIDDEN_SHEETS

OVERVIEW_TABLE = "tbl_manager_projects_v3"
PLAN_TABLE = "tbl_manager_collection_plans_v3"
ACCEPTANCE_TABLE = "tbl_manager_acceptance_v3"
RULE_TABLE = "tbl_manager_rules_v3"
DICTIONARY_TABLE = "tbl_manager_dictionary_v3"
ENTITY_VERSION_TABLE = "tbl_manager_entity_versions_v3"
METADATA_TABLE = "tbl_manager_metadata_v3"

OVERVIEW_HEADERS = (
    "项目ID",
    "项目编号",
    "项目名称",
    "维保开始日期",
    "维保结束日期",
    "维保期限完整性（系统生成）",
    "全部合同额（含税）",
    "合同额完整性（系统生成）",
    "财务确认实收（只读）",
    "实收/合同额（只读）",
    "计划回款合计（只读）",
    "计划/合同额（只读）",
    "验收提交状态（只读）",
    "验收审批状态（只读）",
    "业务配置状态（只读）",
    "__project_version",
    "__service_period_version",
    "__acceptance_version",
    "__row_token",
)
PLAN_HEADERS = (
    "项目ID",
    "项目编号",
    "项目合同关系ID",
    "合同编号",
    "合同额（含税，只读）",
    "计划期次",
    "计划回款日期",
    "计划回款金额（含税）",
    "完整性状态（系统生成）",
    "__project_version",
    "__contract_version",
    "__milestone_version",
    "__row_token",
)
ACCEPTANCE_HEADERS = (
    "项目ID",
    "项目编号",
    "项目名称",
    "提交状态（只读）",
    "提交时间（只读）",
    "附件数量（只读）",
    "审批状态（只读）",
    "审批时间（只读）",
    "审批人（只读）",
    "业务配置状态（只读）",
    "说明",
)
RULE_HEADERS = ("主题", "规则", "结果")
DICTIONARY_HEADERS = ("字段", "允许值或语义")
ENTITY_VERSION_HEADERS = ("entity_type", "entity_id", "base_version")
METADATA_HEADERS = ("key", "value")

MAX_WORKBOOK_BYTES = 64 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_ZIP_MEMBERS = 256
MAX_COMPRESSION_RATIO = 200
MAX_ROWS_PER_TABLE = 20_000
MAX_WORKSHEET_ROWS = MAX_ROWS_PER_TABLE + 200
MAX_WORKSHEET_COLUMNS = 64
MAX_DECLARED_CELLS = 500_000
MAX_CELL_CHARS = 32_767
MILESTONES_PER_CONTRACT = 24
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MONEY_FORMAT = "#,##0.00"
_PERCENT_FORMAT = "0.00%"
_DATE_FORMAT = "yyyy-mm-dd"
_HEADER_FILL = PatternFill("solid", fgColor="35506B")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_EDIT_FILL = PatternFill("solid", fgColor="FFF2CC")
_READONLY_FILL = PatternFill("solid", fgColor="F2F4F7")
_WARNING_FILL = PatternFill("solid", fgColor="FFF4CC")
_WRAP = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class WorkbookIssue:
    code: str
    message: str
    sheet: str | None = None
    row: int | None = None
    column: str | None = None
    severity: str = "error"


class ManagerWorkbookV3Error(ValueError):
    """Controlled structural/security error for the v3 workbook."""

    def __init__(
        self,
        message: str,
        *,
        status_code: int = 422,
        issues: Sequence[WorkbookIssue] = (),
    ) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.issues = tuple(issues) or (WorkbookIssue("workbook_invalid", message),)


@dataclass(frozen=True)
class ManagerWorkbookExportArtifact:
    content: bytes
    export_id: str
    filename: str
    project_count: int
    milestone_row_count: int
    file_sha256: str


@dataclass(frozen=True)
class ServicePeriodChange:
    project_id: str
    project_version: int
    expected_version: int
    service_start: date | None
    service_end: date | None
    completeness_state: str


@dataclass(frozen=True)
class MilestoneChange:
    project_id: str
    project_contract_id: str
    sequence: int
    project_version: int
    contract_version: int
    expected_version: int
    planned_date: date | None
    planned_amount: Decimal | None
    completeness_state: str


@dataclass(frozen=True)
class ManagerWorkbookValidation:
    validation_id: str
    export_id: str
    owner_user_id: int
    report_month: date
    scope_version: str
    data_version: str
    file_sha256: str
    service_period_changes: tuple[ServicePeriodChange, ...]
    milestone_changes: tuple[MilestoneChange, ...]
    warnings: tuple[WorkbookIssue, ...] = field(default_factory=tuple)
    errors: tuple[WorkbookIssue, ...] = field(default_factory=tuple)

    @property
    def unchanged(self) -> bool:
        return not self.service_period_changes and not self.milestone_changes

    @property
    def can_apply(self) -> bool:
        return not self.errors


def _require_hmac_key(hmac_key: bytes) -> bytes:
    if not isinstance(hmac_key, bytes) or len(hmac_key) < 16:
        raise ValueError("hmac_key must be injected as at least 16 bytes")
    return hmac_key


def _safe_text(value: Any) -> Any:
    if value is None or not isinstance(value, str):
        return value
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ufffe\uffff]", "", value)
    if len(value) > MAX_CELL_CHARS:
        raise ManagerWorkbookV3Error("单元格文本超过 Excel 安全上限")
    return "'" + value if value.startswith(_FORMULA_PREFIXES) else value


def _iso(value: Any) -> str:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        if not value.is_finite():
            raise ManagerWorkbookV3Error("工作簿包含非有限数值")
        normalized = format(value.normalize(), "f")
        return "0" if normalized in {"-0", "-0.0"} else normalized
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ManagerWorkbookV3Error("工作簿包含非有限数值")
        return _iso(Decimal(str(value)))
    if value is None:
        return ""
    return str(value)


def _canonical(rows: Sequence[Sequence[Any]]) -> str:
    return json.dumps(
        [[_iso(_safe_text(value)) for value in row] for row in rows],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _digest(rows: Sequence[Sequence[Any]]) -> str:
    return hashlib.sha256(_canonical(rows).encode("utf-8")).hexdigest()


def _signature(metadata: Mapping[str, str], hmac_key: bytes) -> str:
    payload = json.dumps(
        {key: metadata[key] for key in sorted(metadata) if key != "metadata_hmac"},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hmac.new(hmac_key, payload, hashlib.sha256).hexdigest()


def _row_token(export_id: str, kind: str, identity: str, versions: str, key: bytes) -> str:
    payload = f"{export_id}|{kind}|{identity}|{versions}".encode("utf-8")
    return hmac.new(key, payload, hashlib.sha256).hexdigest()


def _as_month(value: Any) -> date:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.replace(day=1)
    text = str(value or "").strip()
    if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", text):
        return date.fromisoformat(text + "-01")
    if re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])-01", text):
        return date.fromisoformat(text)
    raise ManagerWorkbookV3Error("报告月份必须为 YYYY-MM")


def _date_or_none(value: Any, *, sheet: str, row: int, column: str) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        if value.time().isoformat() != "00:00:00":
            raise ManagerWorkbookV3Error(
                f"{column}只能填写日期，不能包含时间",
                issues=(WorkbookIssue("invalid_date", f"{column}只能填写日期", sheet, row, column),),
            )
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise ManagerWorkbookV3Error(
            f"{column}必须为 YYYY-MM-DD",
            issues=(WorkbookIssue("invalid_date", f"{column}必须为 YYYY-MM-DD", sheet, row, column),),
        ) from exc


def _amount_or_none(value: Any, *, row: int) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        amount = Decimal(str(value))
        amount = amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ManagerWorkbookV3Error(
            "计划回款金额必须为有效数字",
            issues=(WorkbookIssue("invalid_planned_amount", "计划回款金额必须为有效数字", PLAN_SHEET, row, "计划回款金额（含税）"),),
        ) from exc
    if not amount.is_finite() or amount <= 0 or amount >= Decimal("1000000000000"):
        raise ManagerWorkbookV3Error(
            "计划回款金额必须大于 0 且不超过安全范围",
            issues=(WorkbookIssue("invalid_planned_amount", "计划回款金额必须大于 0 且不超过安全范围", PLAN_SHEET, row, "计划回款金额（含税）"),),
        )
    return amount


def _money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ManagerWorkbookV3Error("系统金额不是有效数字") from exc
    if not number.is_finite() or abs(number) >= Decimal("1000000000000"):
        raise ManagerWorkbookV3Error("系统金额超过安全范围")
    return float(number)


def _completion(planned_date: date | None, amount: Decimal | None) -> str:
    if planned_date is not None and amount is not None:
        return "complete"
    if planned_date is not None:
        return "date_only"
    if amount is not None:
        return "amount_only"
    return "empty"


def _service_completion(start: date | None, end: date | None) -> str:
    if start and end:
        return "complete" if end >= start else "invalid_range"
    if start:
        return "start_only"
    if end:
        return "end_only"
    return "empty"


def _table(sheet, name: str, headers: Sequence[str], rows: Sequence[Sequence[Any]], *, start_row: int = 1) -> None:
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(start_row, column, header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _WRAP
    materialized = list(rows)
    if len(materialized) > MAX_ROWS_PER_TABLE:
        raise ManagerWorkbookV3Error(f"{sheet.title} 超过 v3 单表行数上限")
    materialized = materialized or [tuple(None for _ in headers)]
    for row_index, values in enumerate(materialized, start_row + 1):
        for column, value in enumerate(values, 1):
            sheet.cell(row_index, column, _safe_text(value)).alignment = _WRAP
    ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(materialized)}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = f"A{start_row + 1}"
    sheet.auto_filter.ref = ref


def _snapshot_maps(snapshot: Mapping[str, Any]) -> tuple[dict[str, Mapping[str, Any]], dict[str, tuple[Mapping[str, Any], Mapping[str, Any]]]]:
    projects: dict[str, Mapping[str, Any]] = {}
    contracts: dict[str, tuple[Mapping[str, Any], Mapping[str, Any]]] = {}
    for project in snapshot.get("projects") or []:
        project_id = str(project.get("project_id") or "")
        if not project_id or project_id in projects:
            raise ManagerWorkbookV3Error("项目范围包含空或重复的项目 ID")
        projects[project_id] = project
        for contract in project.get("contracts") or []:
            relation_id = str(contract.get("project_contract_id") or "")
            if not relation_id or relation_id in contracts:
                raise ManagerWorkbookV3Error("项目范围包含空或重复的合同关系 ID")
            contracts[relation_id] = (project, contract)
    return projects, contracts


def _scope_rows(snapshot: Mapping[str, Any]) -> list[tuple[Any, ...]]:
    projects, _ = _snapshot_maps(snapshot)
    return sorted(
        (
            project_id,
            int(project.get("project_version") or 0),
            str(project.get("assignment_id") or ""),
            int(project.get("assignment_version") or 0),
        )
        for project_id, project in projects.items()
    )


def _entity_version_rows(snapshot: Mapping[str, Any]) -> list[tuple[str, str, int]]:
    projects, contracts = _snapshot_maps(snapshot)
    rows: list[tuple[str, str, int]] = []
    for project_id, project in projects.items():
        rows.append(("project", project_id, int(project.get("project_version") or 0)))
        rows.append(("service_period", project_id, int(project.get("service_period_version") or 0)))
        acceptance = project.get("acceptance") or {}
        rows.append(("acceptance", project_id, int(acceptance.get("version") or 0)))
    for relation_id, (_project, contract) in contracts.items():
        rows.append(("project_contract", relation_id, int(contract.get("contract_version") or 0)))
        by_sequence = {
            int(row.get("sequence") or 0): row
            for row in contract.get("planned_milestones") or []
        }
        for sequence in range(1, MILESTONES_PER_CONTRACT + 1):
            rows.append((
                "collection_milestone",
                f"{relation_id}:{sequence}",
                int((by_sequence.get(sequence) or {}).get("version") or 0),
            ))
    return sorted(rows)


def _overview_rows(snapshot: Mapping[str, Any], *, export_id: str, key: bytes) -> list[tuple[Any, ...]]:
    projects, _ = _snapshot_maps(snapshot)
    rows: list[tuple[Any, ...]] = []
    for project_id, project in sorted(projects.items()):
        contracts = list(project.get("contracts") or [])
        known_amounts = [Decimal(str(row["contract_amount"])) for row in contracts if row.get("contract_amount") is not None]
        contract_complete = bool(contracts) and len(known_amounts) == len(contracts)
        total_contract = sum(known_amounts, Decimal("0")) if contract_complete else None
        actual = sum(
            (Decimal(str(row.get("confirmed_received_amount") or 0)) for row in contracts),
            Decimal("0"),
        )
        planned = sum(
            (
                Decimal(str(node.get("planned_amount") or 0))
                for contract in contracts
                for node in contract.get("planned_milestones") or []
            ),
            Decimal("0"),
        )
        start = project.get("service_start")
        end = project.get("service_end")
        acceptance = project.get("acceptance") or {}
        project_version = int(project.get("project_version") or 0)
        service_version = int(project.get("service_period_version") or 0)
        acceptance_version = int(acceptance.get("version") or 0)
        rows.append((
            project_id,
            str(project.get("project_code") or ""),
            str(project.get("project_name") or project.get("display_name") or ""),
            start,
            end,
            _service_completion(start, end),
            _money(total_contract),
            "complete" if contract_complete else "missing",
            _money(actual),
            float(actual / total_contract) if total_contract else None,
            _money(planned),
            float(planned / total_contract) if total_contract else None,
            str(acceptance.get("submission_status") or "not_submitted"),
            str(acceptance.get("approval_status") or "not_reviewed"),
            str(acceptance.get("configuration_state") or "pending_business_configuration"),
            project_version,
            service_version,
            acceptance_version,
            _row_token(
                export_id,
                "project",
                project_id,
                f"{project_version}:{service_version}:{acceptance_version}",
                key,
            ),
        ))
    return rows


def _plan_rows(snapshot: Mapping[str, Any], *, export_id: str, key: bytes) -> list[tuple[Any, ...]]:
    _projects, contracts = _snapshot_maps(snapshot)
    rows: list[tuple[Any, ...]] = []
    for relation_id, (project, contract) in sorted(contracts.items()):
        project_id = str(project.get("project_id") or "")
        project_version = int(project.get("project_version") or 0)
        contract_version = int(contract.get("contract_version") or 0)
        by_sequence = {
            int(node.get("sequence") or 0): node
            for node in contract.get("planned_milestones") or []
        }
        if any(sequence < 1 or sequence > MILESTONES_PER_CONTRACT for sequence in by_sequence):
            raise ManagerWorkbookV3Error("计划回款期次必须在 1 到 24 之间")
        for sequence in range(1, MILESTONES_PER_CONTRACT + 1):
            node = by_sequence.get(sequence) or {}
            milestone_version = int(node.get("version") or 0)
            planned_date = node.get("planned_date")
            planned_amount = (
                Decimal(str(node["planned_amount"]))
                if node.get("planned_amount") is not None
                else None
            )
            rows.append((
                project_id,
                str(project.get("project_code") or ""),
                relation_id,
                str(contract.get("contract_no") or ""),
                _money(contract.get("contract_amount")),
                sequence,
                planned_date,
                _money(planned_amount),
                _completion(planned_date, planned_amount),
                project_version,
                contract_version,
                milestone_version,
                _row_token(
                    export_id,
                    "milestone",
                    f"{relation_id}:{sequence}",
                    f"{project_version}:{contract_version}:{milestone_version}",
                    key,
                ),
            ))
    return rows


def _acceptance_rows(snapshot: Mapping[str, Any]) -> list[tuple[Any, ...]]:
    projects, _ = _snapshot_maps(snapshot)
    rows: list[tuple[Any, ...]] = []
    for project_id, project in sorted(projects.items()):
        acceptance = project.get("acceptance") or {}
        configuration = str(acceptance.get("configuration_state") or "pending_business_configuration")
        rows.append((
            project_id,
            str(project.get("project_code") or ""),
            str(project.get("project_name") or project.get("display_name") or ""),
            str(acceptance.get("submission_status") or "not_submitted"),
            acceptance.get("submitted_at"),
            int(acceptance.get("attachment_count") or 0),
            str(acceptance.get("approval_status") or "not_reviewed"),
            acceptance.get("approved_at"),
            str(acceptance.get("approved_by") or ""),
            configuration,
            "附件载体待业务配置；外部链接不作为附件证据" if configuration != "configured" else "提交即生效，无需独立审批（2026-08-24 起）",
        ))
    return rows


def _readonly_overview_projection(rows: Sequence[Sequence[Any]]) -> list[tuple[Any, ...]]:
    # Columns 4-5 are the only editable cells in this sheet.
    return [
        tuple(value for index, value in enumerate(row) if index not in (3, 4))
        for row in rows
    ]


def _readonly_plan_projection(rows: Sequence[Sequence[Any]]) -> list[tuple[Any, ...]]:
    # Columns 7-8 are the only editable cells in this sheet.
    return [tuple(value for index, value in enumerate(row) if index not in (6, 7)) for row in rows]


def _metadata_rows(metadata: Mapping[str, str]) -> list[tuple[str, str]]:
    return [(key, metadata[key]) for key in sorted(metadata)]


def build_manager_workbook(
    snapshot: Mapping[str, Any],
    *,
    hmac_key: bytes,
    export_id: str | None = None,
    exported_at: datetime | None = None,
) -> ManagerWorkbookExportArtifact:
    """Build a signed, full-scope workbook without persisting anything."""

    key = _require_hmac_key(hmac_key)
    owner = snapshot.get("owner") or {}
    try:
        owner_user_id = int(owner["user_id"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ManagerWorkbookV3Error("项目经理范围缺少稳定账号 ID") from exc
    owner_username = str(owner.get("username") or "").strip()
    if not owner_username:
        raise ManagerWorkbookV3Error("项目经理范围缺少账号名")
    report_month = _as_month(snapshot.get("report_month"))
    export_id = export_id or str(uuid.uuid4())
    exported_at = exported_at or datetime.now(timezone.utc)
    overview_rows = _overview_rows(snapshot, export_id=export_id, key=key)
    plan_rows = _plan_rows(snapshot, export_id=export_id, key=key)
    acceptance_rows = _acceptance_rows(snapshot)
    version_rows = _entity_version_rows(snapshot)
    metadata = {
        "protocol_id": PROTOCOL_ID,
        "schema_version": SCHEMA_VERSION,
        "template_version": TEMPLATE_VERSION,
        "owner_user_id": str(owner_user_id),
        "owner_username": owner_username,
        "report_month": report_month.strftime("%Y-%m"),
        "scope_version": str(snapshot.get("scope_version") or ""),
        "data_version": str(snapshot.get("data_version") or ""),
        "export_id": export_id,
        "exported_at": exported_at.isoformat(),
        "project_count": str(len(overview_rows)),
        "milestone_row_count": str(len(plan_rows)),
        "scope_digest": _digest(_scope_rows(snapshot)),
        "readonly_overview_digest": _digest(_readonly_overview_projection(overview_rows)),
        "readonly_plan_digest": _digest(_readonly_plan_projection(plan_rows)),
        "acceptance_digest": _digest(acceptance_rows),
        "entity_versions_digest": _digest(version_rows),
    }
    metadata["metadata_hmac"] = _signature(metadata, key)

    book = Workbook()
    book.remove(book.active)
    overview = book.create_sheet(OVERVIEW_SHEET)
    plan = book.create_sheet(PLAN_SHEET)
    acceptance = book.create_sheet(ACCEPTANCE_SHEET)
    instructions = book.create_sheet(INSTRUCTIONS_SHEET)
    dictionary = book.create_sheet(DICTIONARY_SHEET)
    versions = book.create_sheet(ENTITY_VERSION_SHEET)
    metadata_sheet = book.create_sheet(METADATA_SHEET)

    overview["A1"] = "项目经理月度全量工作簿 v3"
    overview["A1"].font = Font(size=16, bold=True, color="35506B")
    overview["A2"] = f"范围：{owner.get('display_name') or owner_username} 本人负责项目；月份：{report_month:%Y-%m}"
    overview["A3"] = "黄色单元格可回填。计划回款与财务确认实收严格分开，上传不会改写实收。"
    _table(overview, OVERVIEW_TABLE, OVERVIEW_HEADERS, overview_rows, start_row=5)
    _table(plan, PLAN_TABLE, PLAN_HEADERS, plan_rows, start_row=5)
    plan["A1"] = "24 期计划回款节点（纵向）"
    plan["A1"].font = Font(size=16, bold=True, color="35506B")
    plan["A2"] = "每个项目合同关系固定 24 行；日期或金额只填一项也会保留并标注不完整。"
    plan["A3"] = "留空表示保留现有值，不执行删除；实际回款由财务确认链路维护。"
    acceptance["A1"] = "验收材料状态（只读）"
    acceptance["A1"].font = Font(size=16, bold=True, color="35506B")
    acceptance["A2"] = "提交即生效（2026-08-24 起取消独立审批）；提交人即生效人，全部实名审计。"
    acceptance["A3"] = "附件载体未配置时系统关闭写入口；外部链接不视为附件。"
    _table(acceptance, ACCEPTANCE_TABLE, ACCEPTANCE_HEADERS, acceptance_rows, start_row=5)

    rules = (
        ("范围", "只包含当前登录项目经理被显式分配且仍有效的项目", "越权项目整批拒绝"),
        ("计划与实收", "计划节点仅用于提醒；财务确认实收只读", "两条数据链不互相覆盖"),
        ("空白", "空白保留已有值；新行日期和金额都空则忽略", "不删除已存在事实"),
        ("不完整节点", "仅日期或仅金额仍保留", "预览显示黄色警告"),
        ("验收", "提交即生效，可重新提交新版本", "生效后补附件走受控上传"),
        ("附件", "只接受受控文件对象；外部链接不是附件", "未配置载体时关闭操作"),
        ("应用", "先校验预览，再一次性确认应用", "任一冲突则零写入"),
        ("版本", "v3 独立；不得上传至 v2 项目工作簿入口", "跨版本整批拒绝"),
    )
    instructions["A1"] = "使用说明与校验规则"
    instructions["A1"].font = Font(size=16, bold=True, color="35506B")
    _table(instructions, RULE_TABLE, RULE_HEADERS, rules, start_row=3)
    _table(dictionary, DICTIONARY_TABLE, DICTIONARY_HEADERS, (
        ("计划节点完整性", "complete / date_only / amount_only / empty"),
        ("验收提交状态", "not_submitted / submitted"),
        ("验收生效状态", "not_reviewed / approved / rejected（提交即 approved，其余为历史存量）"),
        ("业务配置状态", "configured / pending_business_configuration"),
    ))
    _table(versions, ENTITY_VERSION_TABLE, ENTITY_VERSION_HEADERS, version_rows)
    _table(metadata_sheet, METADATA_TABLE, METADATA_HEADERS, _metadata_rows(metadata))

    for sheet in book.worksheets:
        sheet.sheet_view.showGridLines = False
        for column in range(1, min(sheet.max_column, MAX_WORKSHEET_COLUMNS) + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18
    overview.column_dimensions["C"].width = 30
    acceptance.column_dimensions["C"].width = 30
    acceptance.column_dimensions["K"].width = 48
    instructions.column_dimensions["A"].width = 20
    instructions.column_dimensions["B"].width = 56
    instructions.column_dimensions["C"].width = 28

    for sheet, editable_columns in ((overview, (4, 5)), (plan, (7, 8))):
        for row in range(6, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, column)
                cell.protection = Protection(locked=column not in editable_columns)
                if column in editable_columns:
                    cell.fill = _EDIT_FILL
                elif column >= sheet.max_column - 2:
                    cell.fill = _READONLY_FILL
        sheet.protection.sheet = True
        sheet.protection.password = "itdata-v3"
        sheet.protection.selectLockedCells = False
        sheet.protection.selectUnlockedCells = True

    for sheet in (acceptance, instructions):
        sheet.protection.sheet = True
        sheet.protection.password = "itdata-v3"
    for hidden in HIDDEN_SHEETS:
        book[hidden].sheet_state = "hidden" if hidden == DICTIONARY_SHEET else "veryHidden"

    for sheet in (overview, plan):
        for row in range(6, sheet.max_row + 1):
            sheet.cell(row, 4 if sheet is overview else 7).number_format = _DATE_FORMAT
            if sheet is overview:
                sheet.cell(row, 5).number_format = _DATE_FORMAT
    for row in range(6, overview.max_row + 1):
        for column in (7, 9, 11):
            overview.cell(row, column).number_format = _MONEY_FORMAT
        for column in (10, 12):
            overview.cell(row, column).number_format = _PERCENT_FORMAT
    for row in range(6, plan.max_row + 1):
        for column in (5, 8):
            plan.cell(row, column).number_format = _MONEY_FORMAT
    overview_date_validation = DataValidation(type="date", operator="between", formula1="DATE(2000,1,1)", formula2="DATE(2199,12,31)", allow_blank=True)
    overview.add_data_validation(overview_date_validation)
    if overview.max_row >= 6:
        overview_date_validation.add(f"D6:E{overview.max_row}")
    date_validation = DataValidation(type="date", operator="between", formula1="DATE(2000,1,1)", formula2="DATE(2199,12,31)", allow_blank=True)
    amount_validation = DataValidation(type="decimal", operator="between", formula1="0.01", formula2="999999999999.99", allow_blank=True)
    plan.add_data_validation(date_validation)
    plan.add_data_validation(amount_validation)
    if plan.max_row >= 6:
        date_validation.add(f"G6:G{plan.max_row}")
        amount_validation.add(f"H6:H{plan.max_row}")

    output = io.BytesIO()
    try:
        book.save(output)
    finally:
        book.close()
    content = output.getvalue()
    filename = f"维保项目经理月度全量工作簿_{report_month:%Y-%m}.xlsx"
    return ManagerWorkbookExportArtifact(
        content=content,
        export_id=export_id,
        filename=filename,
        project_count=len(overview_rows),
        milestone_row_count=len(plan_rows),
        file_sha256=hashlib.sha256(content).hexdigest(),
    )


def _source_bytes(source: bytes | bytearray | memoryview | io.BufferedIOBase) -> bytes:
    if isinstance(source, (bytes, bytearray, memoryview)):
        content = bytes(source)
    elif hasattr(source, "read"):
        position = source.tell() if hasattr(source, "tell") else None
        content = source.read(MAX_WORKBOOK_BYTES + 1)
        if position is not None and hasattr(source, "seek"):
            source.seek(position)
    else:
        raise TypeError("workbook source must be bytes or a binary file")
    if not content or len(content) > MAX_WORKBOOK_BYTES:
        raise ManagerWorkbookV3Error("上传工作簿为空或超过安全大小上限")
    return content


def _assert_safe_package(content: bytes) -> None:
    try:
        archive = zipfile.ZipFile(io.BytesIO(content))
    except (zipfile.BadZipFile, OSError) as exc:
        raise ManagerWorkbookV3Error("上传文件不是有效的 XLSX 工作簿") from exc
    with archive:
        infos = archive.infolist()
        if len(infos) > MAX_ZIP_MEMBERS:
            raise ManagerWorkbookV3Error("XLSX 内部文件数量超过安全上限")
        names = {item.filename.casefold() for item in infos}
        if "[content_types].xml" not in names or "xl/workbook.xml" not in names:
            raise ManagerWorkbookV3Error("上传文件缺少 XLSX 核心结构")
        total = 0
        for info in infos:
            normalized = info.filename.replace("\\", "/")
            folded = normalized.casefold()
            if normalized.startswith("/") or ".." in normalized.split("/"):
                raise ManagerWorkbookV3Error("XLSX 包含非法内部路径")
            total += info.file_size
            if total > MAX_UNCOMPRESSED_BYTES:
                raise ManagerWorkbookV3Error("XLSX 解压后大小超过安全上限")
            if info.file_size > 1_000_000 and info.file_size / max(info.compress_size, 1) > MAX_COMPRESSION_RATIO:
                raise ManagerWorkbookV3Error("XLSX 压缩比异常，疑似 ZIP bomb")
            if "vbaproject.bin" in folded or "macrosheet" in folded or folded.startswith("xl/externallinks/"):
                raise ManagerWorkbookV3Error("不允许宏或外部链接工作簿")
            if folded.endswith((".xml", ".rels")):
                raw = archive.read(info).lower()
                if b"macroenabled" in raw or re.search(rb"targetmode\s*=\s*['\"]\s*external\s*['\"]", raw):
                    raise ManagerWorkbookV3Error("不允许宏或外部链接工作簿")


def _read_table(book, sheet_name: str, table_name: str, headers: Sequence[str]) -> tuple[list[tuple[Any, ...]], int]:
    sheet = book[sheet_name]
    table = sheet.tables.get(table_name)
    if table is None:
        raise ManagerWorkbookV3Error(f"{sheet_name} 缺少 {table_name}")
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if min_col != 1 or max_col != len(headers) or max_row - min_row > MAX_ROWS_PER_TABLE:
        raise ManagerWorkbookV3Error(f"{sheet_name} 的表格范围不符合 v3 协议")
    actual_headers = tuple(sheet.cell(min_row, column).value for column in range(min_col, max_col + 1))
    if actual_headers != tuple(headers):
        raise ManagerWorkbookV3Error(f"{sheet_name} 的表头被修改")
    rows = [
        tuple(sheet.cell(row, column).value for column in range(min_col, max_col + 1))
        for row in range(min_row + 1, max_row + 1)
    ]
    return rows, min_row


def _read_metadata(book) -> dict[str, str]:
    rows, _ = _read_table(book, METADATA_SHEET, METADATA_TABLE, METADATA_HEADERS)
    metadata: dict[str, str] = {}
    for key, value in rows:
        key_text = str(key or "")
        if not key_text or key_text in metadata:
            raise ManagerWorkbookV3Error("99_元数据存在空键或重复键")
        metadata[key_text] = str(value if value is not None else "")
    required = {
        "protocol_id", "schema_version", "template_version", "owner_user_id",
        "owner_username", "report_month", "scope_version", "data_version",
        "export_id", "scope_digest", "readonly_overview_digest",
        "readonly_plan_digest", "acceptance_digest", "entity_versions_digest",
        "metadata_hmac",
    }
    missing = sorted(required - metadata.keys())
    if missing:
        raise ManagerWorkbookV3Error("99_元数据缺少字段：" + "、".join(missing))
    return metadata


def validate_manager_workbook(
    source: bytes | bytearray | memoryview | io.BufferedIOBase,
    *,
    snapshot: Mapping[str, Any],
    hmac_key: bytes,
) -> ManagerWorkbookValidation:
    """Validate the whole workbook and return an atomic, server-owned plan."""

    key = _require_hmac_key(hmac_key)
    content = _source_bytes(source)
    _assert_safe_package(content)
    try:
        book = load_workbook(io.BytesIO(content), data_only=False, keep_links=False)
    except Exception as exc:
        raise ManagerWorkbookV3Error("无法安全读取上传的 XLSX 工作簿") from exc
    try:
        if METADATA_SHEET not in book.sheetnames:
            raise ManagerWorkbookV3Error("未知工作簿协议；请重新下载 v3 模板")
        metadata = _read_metadata(book)
        if metadata["protocol_id"] != PROTOCOL_ID or metadata["schema_version"] != SCHEMA_VERSION:
            raise ManagerWorkbookV3Error("工作簿不是兼容的 v3 模板；v2 与 v3 不可混用")
        if not hmac.compare_digest(metadata["metadata_hmac"], _signature(metadata, key)):
            raise ManagerWorkbookV3Error(
                "工作簿元数据签名无效；请重新下载",
                issues=(WorkbookIssue("metadata_tampered", "工作簿元数据签名无效"),),
            )
        if metadata["template_version"] != TEMPLATE_VERSION:
            raise ManagerWorkbookV3Error(
                "工作簿模板版本已过期；请重新下载当前模板",
                status_code=409,
                issues=(
                    WorkbookIssue(
                        "template_version_mismatch",
                        "工作簿模板版本不是当前版本；请重新下载",
                        METADATA_SHEET,
                    ),
                ),
            )
        if tuple(book.sheetnames) != SHEET_NAMES:
            raise ManagerWorkbookV3Error("工作表名称、顺序或数量不符合 v3 协议")
        visible = tuple(sheet.title for sheet in book.worksheets if sheet.sheet_state == "visible")
        if visible != VISIBLE_SHEETS:
            raise ManagerWorkbookV3Error("可见工作表不符合 v3 协议")
        if book[DICTIONARY_SHEET].sheet_state != "hidden" or any(
            book[name].sheet_state != "veryHidden" for name in (ENTITY_VERSION_SHEET, METADATA_SHEET)
        ):
            raise ManagerWorkbookV3Error("协议隐藏工作表状态被修改")
        for sheet in book.worksheets:
            if int(sheet.max_row or 0) > MAX_WORKSHEET_ROWS or int(sheet.max_column or 0) > MAX_WORKSHEET_COLUMNS or int(sheet.max_row or 0) * int(sheet.max_column or 0) > MAX_DECLARED_CELLS:
                raise ManagerWorkbookV3Error(f"{sheet.title} 的声明范围超过安全上限")
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        raise ManagerWorkbookV3Error(f"{sheet.title} 包含公式，工作簿不可导入")
        expected_tables = {
            OVERVIEW_SHEET: {OVERVIEW_TABLE},
            PLAN_SHEET: {PLAN_TABLE},
            ACCEPTANCE_SHEET: {ACCEPTANCE_TABLE},
            INSTRUCTIONS_SHEET: {RULE_TABLE},
            DICTIONARY_SHEET: {DICTIONARY_TABLE},
            ENTITY_VERSION_SHEET: {ENTITY_VERSION_TABLE},
            METADATA_SHEET: {METADATA_TABLE},
        }
        for sheet_name, names in expected_tables.items():
            if set(book[sheet_name].tables) != names:
                raise ManagerWorkbookV3Error(f"{sheet_name} 的 Excel Table 名称或数量不符合协议")

        owner = snapshot.get("owner") or {}
        report_month = _as_month(snapshot.get("report_month"))
        try:
            owner_user_id = int(owner.get("user_id"))
        except (TypeError, ValueError) as exc:
            raise ManagerWorkbookV3Error("当前项目经理账号无稳定 ID") from exc
        if metadata["owner_user_id"] != str(owner_user_id) or metadata["owner_username"] != str(owner.get("username") or ""):
            raise ManagerWorkbookV3Error("工作簿不属于当前账号", status_code=403, issues=(WorkbookIssue("owner_mismatch", "工作簿不属于当前账号"),))
        if metadata["report_month"] != report_month.strftime("%Y-%m"):
            raise ManagerWorkbookV3Error("工作簿报告月份与当前任务不一致", status_code=409)
        for key_name in ("scope_version", "data_version"):
            if metadata[key_name] != str(snapshot.get(key_name) or ""):
                raise ManagerWorkbookV3Error(
                    "项目范围或数据版本已变化；请重新下载后回填",
                    status_code=409,
                    issues=(WorkbookIssue("stale_workbook", "项目范围或数据版本已变化"),),
                )
        if not hmac.compare_digest(metadata["scope_digest"], _digest(_scope_rows(snapshot))):
            raise ManagerWorkbookV3Error(
                "本人负责项目范围已变化；请重新下载",
                status_code=409,
                issues=(WorkbookIssue("scope_conflict", "本人负责项目范围已变化"),),
            )

        overview_rows, overview_header_row = _read_table(book, OVERVIEW_SHEET, OVERVIEW_TABLE, OVERVIEW_HEADERS)
        plan_rows, plan_header_row = _read_table(book, PLAN_SHEET, PLAN_TABLE, PLAN_HEADERS)
        # Excel tables require one physical data row. A manager project with no
        # included contracts therefore exports one all-blank placeholder; at
        # the protocol level that is an empty plan, not an unknown row.
        if len(plan_rows) == 1 and all(value in (None, "") for value in plan_rows[0]):
            plan_rows = []
        acceptance_rows, _ = _read_table(book, ACCEPTANCE_SHEET, ACCEPTANCE_TABLE, ACCEPTANCE_HEADERS)
        version_rows, _ = _read_table(book, ENTITY_VERSION_SHEET, ENTITY_VERSION_TABLE, ENTITY_VERSION_HEADERS)
        if not hmac.compare_digest(metadata["readonly_overview_digest"], _digest(_readonly_overview_projection(overview_rows))):
            return ManagerWorkbookValidation(
                validation_id=str(uuid.uuid4()), export_id=metadata["export_id"], owner_user_id=owner_user_id,
                report_month=report_month, scope_version=metadata["scope_version"], data_version=metadata["data_version"],
                file_sha256=hashlib.sha256(content).hexdigest(), service_period_changes=(), milestone_changes=(),
                errors=(WorkbookIssue("readonly_actual_changed", "项目只读字段（含财务确认实收）被修改", OVERVIEW_SHEET),),
            )
        if not hmac.compare_digest(metadata["readonly_plan_digest"], _digest(_readonly_plan_projection(plan_rows))):
            return ManagerWorkbookValidation(
                validation_id=str(uuid.uuid4()), export_id=metadata["export_id"], owner_user_id=owner_user_id,
                report_month=report_month, scope_version=metadata["scope_version"], data_version=metadata["data_version"],
                file_sha256=hashlib.sha256(content).hexdigest(), service_period_changes=(), milestone_changes=(),
                errors=(WorkbookIssue("readonly_plan_changed", "计划表的只读身份或版本字段被修改", PLAN_SHEET),),
            )
        if not hmac.compare_digest(metadata["acceptance_digest"], _digest(acceptance_rows)):
            return ManagerWorkbookValidation(
                validation_id=str(uuid.uuid4()), export_id=metadata["export_id"], owner_user_id=owner_user_id,
                report_month=report_month, scope_version=metadata["scope_version"], data_version=metadata["data_version"],
                file_sha256=hashlib.sha256(content).hexdigest(), service_period_changes=(), milestone_changes=(),
                errors=(WorkbookIssue("readonly_acceptance_changed", "验收提交、附件或审批状态被修改", ACCEPTANCE_SHEET),),
            )
        normalized_versions = [tuple(row) for row in version_rows if any(value not in (None, "") for value in row)]
        if not hmac.compare_digest(metadata["entity_versions_digest"], _digest(normalized_versions)):
            raise ManagerWorkbookV3Error("实体版本快照被修改；请重新下载")

        projects, contracts = _snapshot_maps(snapshot)
        if len(overview_rows) != len(projects):
            raise ManagerWorkbookV3Error("项目总览行数与本人范围不一致")
        if len(plan_rows) != len(contracts) * MILESTONES_PER_CONTRACT:
            raise ManagerWorkbookV3Error("每个项目合同关系必须完整保留 24 个计划期次")

        service_changes: list[ServicePeriodChange] = []
        warnings: list[WorkbookIssue] = []
        errors: list[WorkbookIssue] = []
        seen_projects: set[str] = set()
        for offset, row in enumerate(overview_rows, 1):
            excel_row = overview_header_row + offset
            project_id = str(row[0] or "")
            project = projects.get(project_id)
            if project is None or project_id in seen_projects:
                raise ManagerWorkbookV3Error("项目总览包含未知或重复项目")
            seen_projects.add(project_id)
            expected_start = project.get("service_start")
            expected_end = project.get("service_end")
            entered_start = _date_or_none(row[3], sheet=OVERVIEW_SHEET, row=excel_row, column="维保开始日期")
            entered_end = _date_or_none(row[4], sheet=OVERVIEW_SHEET, row=excel_row, column="维保结束日期")
            # Blank means preserve, never erase an existing fact.
            proposed_start = entered_start if entered_start is not None else expected_start
            proposed_end = entered_end if entered_end is not None else expected_end
            state = _service_completion(proposed_start, proposed_end)
            if state == "invalid_range":
                errors.append(WorkbookIssue("invalid_service_period", "维保结束日期不能早于开始日期", OVERVIEW_SHEET, excel_row))
                continue
            if proposed_start != expected_start or proposed_end != expected_end:
                service_changes.append(ServicePeriodChange(
                    project_id=project_id,
                    project_version=int(project.get("project_version") or 0),
                    expected_version=int(project.get("service_period_version") or 0),
                    service_start=proposed_start,
                    service_end=proposed_end,
                    completeness_state=state,
                ))
            if state in {"start_only", "end_only"}:
                warnings.append(WorkbookIssue("partial_service_period", "维保期限只填写了一端，系统会保留并标注不完整", OVERVIEW_SHEET, excel_row, severity="warning"))

        milestone_changes: list[MilestoneChange] = []
        seen_plan_keys: set[tuple[str, int]] = set()
        for offset, row in enumerate(plan_rows, 1):
            excel_row = plan_header_row + offset
            relation_id = str(row[2] or "")
            pair = contracts.get(relation_id)
            try:
                sequence = int(row[5])
            except (TypeError, ValueError) as exc:
                raise ManagerWorkbookV3Error("计划期次无效") from exc
            key_tuple = (relation_id, sequence)
            if pair is None or not 1 <= sequence <= MILESTONES_PER_CONTRACT or key_tuple in seen_plan_keys:
                raise ManagerWorkbookV3Error("计划表包含未知、越界或重复的合同期次")
            seen_plan_keys.add(key_tuple)
            project, contract = pair
            existing = next(
                (node for node in contract.get("planned_milestones") or [] if int(node.get("sequence") or 0) == sequence),
                {},
            )
            expected_date = existing.get("planned_date")
            expected_amount = Decimal(str(existing["planned_amount"])) if existing.get("planned_amount") is not None else None
            entered_date = _date_or_none(row[6], sheet=PLAN_SHEET, row=excel_row, column="计划回款日期")
            entered_amount = _amount_or_none(row[7], row=excel_row)
            proposed_date = entered_date if entered_date is not None else expected_date
            proposed_amount = entered_amount if entered_amount is not None else expected_amount
            state = _completion(proposed_date, proposed_amount)
            if state in {"date_only", "amount_only"}:
                warnings.append(WorkbookIssue("partial_plan_node", f"第 {sequence} 期计划仅填写了日期或金额，系统会保留并标注不完整", PLAN_SHEET, excel_row, severity="warning"))
            if proposed_date != expected_date or proposed_amount != expected_amount:
                milestone_changes.append(MilestoneChange(
                    project_id=str(project.get("project_id") or ""),
                    project_contract_id=relation_id,
                    sequence=sequence,
                    project_version=int(project.get("project_version") or 0),
                    contract_version=int(contract.get("contract_version") or 0),
                    expected_version=int(existing.get("version") or 0),
                    planned_date=proposed_date,
                    planned_amount=proposed_amount,
                    completeness_state=state,
                ))
        if len(seen_plan_keys) != len(contracts) * MILESTONES_PER_CONTRACT:
            raise ManagerWorkbookV3Error("计划表未完整保留全部 24 期行")

        return ManagerWorkbookValidation(
            validation_id=str(uuid.uuid4()),
            export_id=metadata["export_id"],
            owner_user_id=owner_user_id,
            report_month=report_month,
            scope_version=metadata["scope_version"],
            data_version=metadata["data_version"],
            file_sha256=hashlib.sha256(content).hexdigest(),
            service_period_changes=tuple(service_changes),
            milestone_changes=tuple(milestone_changes),
            warnings=tuple(warnings),
            errors=tuple(errors),
        )
    finally:
        book.close()
