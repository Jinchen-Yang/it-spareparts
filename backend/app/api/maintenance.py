"""维保项目成本 API（docs/维保出库成本核算-开发方案.md §5）。

鉴权双层：current_role 硬鉴权（缺/失效凭证 → 401，与 purchases/parts 一致，不依赖 RBAC 开关）
+ require_page('page_maintenance') 页面准入（admin/boss/purchaser 模板默认开）。
成本金额字段随 data_purchase_cost 脱敏（成本 7 键 + 聚合派生键已登记 FIELD_GROUPS）。
"""
import csv
import io
import re
from datetime import date
from decimal import Decimal
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from starlette.background import BackgroundTask

from app.auth import current_role, require_admin
from app.business_time import business_today
from app.db import get_db
from app.security import (
    UserContext, apply_field_visibility, get_current_user_context, record_access_log,
    require_page,
)
from app import config
from app.services import maintenance_cost, maintenance_export

router = APIRouter(prefix="/maintenance", tags=["maintenance"])


@router.post("/recompute")
def recompute(db: Session = Depends(get_db),
              _auth: str = Depends(require_admin),   # 全表重算(~1min 写库)：限管理员，与导入触发方口径一致
              ctx: UserContext = Depends(get_current_user_context)) -> dict:
    record_access_log(ctx, "recompute", "maintenance")
    return maintenance_cost.recompute(db)


@router.get("/projects")
def projects(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    q: str | None = Query(None, max_length=128),
    lifecycle: str = Query("ongoing", pattern=r"^(ongoing|ended|missing|all)$"),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),   # 硬鉴权：缺/失效凭证 → 401，不依赖全局 RBAC 开关
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> dict:
    record_access_log(ctx, "projects", "maintenance")
    data = maintenance_cost.projects_aggregate(
        db, date_from, date_to, q, user_ctx=ctx,
        lifecycle=lifecycle, as_of=business_today(),
    )
    return apply_field_visibility(data, ctx)


@router.get("/lines")
def lines(
    project: str = Query(..., max_length=256),
    month: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),   # 硬鉴权：缺/失效凭证 → 401
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> dict:
    record_access_log(ctx, "lines", "maintenance", {"project": project})
    data = maintenance_cost.project_lines(db, project, month, date_from, date_to, page, page_size)
    return apply_field_visibility(data, ctx)


_SOURCE_LABEL = {"direct": "实际·专属采购", "window": "实际·±7天最近价",
                 "month_avg": "实际·当月均价",
                 "trace_avg": "预估·追溯均价", "sales_ref": "没有采购有销售", "none": "无成本"}
_CONF_LABEL = {"high": "高", "medium": "中", "low": "低"}


def _safe(v):
    # 防 CSV 公式注入：以 = + - @ 制表/回车开头的文本前置单引号
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + v
    return v


_INVALID_DOWNLOAD_NAME = re.compile(r'[\x00-\x1f\x7f/\\:*?"<>|]+')
_INVALID_ASCII_FALLBACK = re.compile(r"[^A-Za-z0-9._-]+")
_RFC5987_SAFE = "!#$&+-.^_`|~"


def _content_disposition(filename: str, ascii_fallback: str | None = None) -> str:
    """生成纯 ASCII 下载响应头，同时保留 RFC 5987 UTF-8 文件名。"""
    clean_name = _INVALID_DOWNLOAD_NAME.sub("_", filename).strip(" .") or "download"
    fallback_source = ascii_fallback or clean_name
    fallback = _INVALID_DOWNLOAD_NAME.sub("_", fallback_source)
    fallback = fallback.encode("ascii", "ignore").decode()
    fallback = _INVALID_ASCII_FALLBACK.sub("_", fallback).strip(" ._") or "download"
    encoded_name = quote(clean_name, safe=_RFC5987_SAFE)
    return (
        f'attachment; filename="{fallback}"; '
        f"filename*=UTF-8''{encoded_name}"
    )


def _csv_stream(
    header: list,
    rows: list,
    filename: str,
    ascii_fallback: str | None = None,
) -> StreamingResponse:
    buf = io.StringIO()
    buf.write("﻿")  # BOM，Excel 正确识别 UTF-8
    w = csv.writer(buf)
    w.writerow(header)
    for r in rows:
        w.writerow(r)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]), media_type="text/csv",
        headers={"Content-Disposition": _content_disposition(filename, ascii_fallback)},
    )


@router.get("/export")
def export(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    q: str | None = Query(None, max_length=128),
    lifecycle: str = Query("ongoing", pattern=r"^(ongoing|ended|missing|all)$"),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),   # 硬鉴权：缺/失效凭证 → 401
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> StreamingResponse:
    record_access_log(ctx, "export", "maintenance")
    data = maintenance_cost.projects_aggregate(
        db, date_from, date_to, q, user_ctx=ctx,
        lifecycle=lifecycle, as_of=business_today(),
    )
    data = apply_field_visibility(data, ctx)   # 导出同样过脱敏层（§8.5）
    header = ["项目", "期限状态", "维保终止日期",
              "出库行数", "出库数量", "备件成本-含税小计", "备件成本-不含税小计",
              "成本合计(混合口径参考)", "覆盖率%",
              *(_SOURCE_LABEL[s] + "(行)" for s in ("direct", "window", "month_avg",
                                                    "trace_avg", "sales_ref", "none")),
              "月份数", "关联销售订单", "合同额(含税参考)", "合同被多项目共用"]
    rows = []
    for r in data["rows"]:
        bs = r["by_source"]
        lifecycle_label = {"ongoing": "进行中", "ended": "已结束", "missing": "期限缺失"}
        rows.append([_safe(r["project"]), lifecycle_label[r["lifecycle_status"]], r["maint_end"],
                     r["lines"], r["qty"],
                     r["cost_inc"], r["cost_ex"], r["cost_total"], r["coverage_pct"],
                     bs.get("direct", 0), bs.get("window", 0), bs.get("month_avg", 0),
                     bs.get("trace_avg", 0), bs.get("sales_ref", 0), bs.get("none", 0),
                     r["months"], _safe("、".join(r["sales_orders"])),
                     r["contract_amount"], "是" if r["contract_shared"] else ""])
    return _csv_stream(header, rows, "maintenance_projects.csv")


@router.get("/orders/export")
def orders_export(
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> StreamingResponse:
    if (date_from is None) != (date_to is None):
        raise HTTPException(status_code=422, detail="date_from 与 date_to 必须同时提供")
    if date_from is not None and date_to is not None and date_from > date_to:
        raise HTTPException(status_code=422, detail="date_from 不能晚于 date_to")
    audit_scope = (
        {"date_from": date_from.isoformat(), "date_to": date_to.isoformat()}
        if date_from is not None and date_to is not None
        else {"scope": "all"}
    )
    record_access_log(ctx, "orders_export", "maintenance", audit_scope)
    try:
        output = maintenance_export.build_workbook(db, ctx, date_from, date_to)
    except (maintenance_export.ExcelCellTooLong, maintenance_export.ExcelRowLimitExceeded) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    scope = f"{date_from.isoformat()}_{date_to.isoformat()}" if date_from and date_to else "all"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(
            f"maintenance_orders_{scope}.xlsx",
        )},
        background=BackgroundTask(output.close),
    )


@router.get("/lines/export")
def lines_export(
    project: str = Query(..., max_length=256),
    month: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),   # 硬鉴权：缺/失效凭证 → 401
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> StreamingResponse:
    """单项目 SKU 明细导出（财务逐行复核入账用）——全量、含成本来源/税口径。"""
    record_access_log(ctx, "lines_export", "maintenance", {"project": project})
    data = maintenance_cost.project_lines(db, project, month, date_from, date_to,
                                          page=1, page_size=1_000_000)
    data = apply_field_visibility(data, ctx)
    header = ["日期", "维保单号", "需求类型", "业务类型", "出库仓库", "PN", "描述",
              "数量", "退货", "单价", "金额", "成本来源", "置信度", "含税口径", "取价月",
              "追溯月数", "距采购天数", "关联采购单", "异常标记"]
    rows = []
    for r in data["rows"]:
        rows.append([r["order_date"], _safe(r["order_no"]), r["demand_type"], r["business_type"],
                     r["warehouse"], _safe(r["pn_std"]), _safe(r["description"]),
                     r["qty"], r["return_qty"], r["unit_cost"], r["cost_amount"],
                     _SOURCE_LABEL.get(r["cost_source"], r["cost_source"]),
                     _CONF_LABEL.get(r["confidence"], r["confidence"] or ""),
                     r["cost_tax_basis"], r["price_month"], r["trace_months"],
                     r["price_distance_days"],
                     _safe(r["linked_purchase_order_no"]), "、".join(r["anomaly_flags"] or [])])
    return _csv_stream(
        header,
        rows,
        f"maintenance_lines_{project[:40]}.csv",
        ascii_fallback="maintenance_lines.csv",
    )


@router.get("/board")
def board(
    status: str | None = Query(None, pattern=r"^(red|yellow|green|no_budget)$"),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    q: str | None = Query(None, max_length=128),
    lifecycle: str = Query("ongoing", pattern=r"^(ongoing|ended|missing|all)$"),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),   # 硬鉴权：缺/失效凭证 → 401
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> dict:
    """盈亏看板（§16.2）：合同(XSDD)级 红/黄/绿 状态灯，黄红置顶。"""
    record_access_log(ctx, "board", "maintenance")
    data = maintenance_cost.board(
        db, date_from, date_to, status, user_ctx=ctx, q_text=q,
        lifecycle=lifecycle, as_of=business_today(),
    )
    return apply_field_visibility(data, ctx)


# ─────────── §16.4 工作簿模板样式（财务件规范：深色表头/千分位/斑马纹/冻结/筛选）───────────
from openpyxl import Workbook                      # noqa: E402  （openpyxl 为既有依赖）
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter       # noqa: E402

_HDR_FILL = PatternFill("solid", fgColor="35506B")
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=15)
_SUB_FONT = Font(color="8C8C8C", size=10)
_KV_FILL = PatternFill("solid", fgColor="EFEBE3")
_ALT_FILL = PatternFill("solid", fgColor="F7F4EE")
_DOC_FILL = PatternFill("solid", fgColor="FDF3D7")   # 产品成本（单据级回填）高亮：财务第一眼找它
_TOTAL_FILL = PatternFill("solid", fgColor="E8E2D6")
_THIN = Side(style="thin", color="D8D2C6")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_MONEY = "#,##0.00"
_CENTER = Alignment(horizontal="center", vertical="center")
_STATUS_STYLE = {"red": ("超支/亏损", "C0524A"), "yellow": ("预警·剩余≤20%", "B8860B"),
                 "green": ("健康", "3F7A45"), "no_budget": ("无预算", "8C8C8C")}


def _hdr_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = _HDR_FILL, _HDR_FONT
        cell.border, cell.alignment = _BORDER, _CENTER


def _col_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _money_cell(cell, value):
    cell.value = value
    cell.number_format = _MONEY


def _build_workbook(contract: str, data: dict) -> Workbook:
    """三 sheet 财务模板：项目预算（抬头+盈亏+月度汇总）/ 备件明细（按单分组斑马纹）/ 报销明细。"""
    wb = Workbook()
    budget = data["budget"]
    so = data["sales_order"]
    spent_parts = float(sum(data["doc_total"].values(), Decimal(0)))
    spent_exp = float(sum((e.amount for e in data["expenses"]
                           if e.data_status == config.MAINT_EXPENSE_ACTIVE_STATUS
                           and e.amount is not None), Decimal(0)))
    spent = round(spent_parts + spent_exp, 2)
    if budget:
        b = float(budget)
        remaining = round(b - spent, 2)
        status = ("red" if spent >= b
                  else "yellow" if remaining <= b * float(config.MAINT_BUDGET_WARN_PCT)
                  else "green")
    else:
        remaining, status = None, "no_budget"
    st_label, st_color = _STATUS_STYLE[status]

    # ── Sheet1 项目预算 ──
    ws = wb.active
    ws.title = "项目预算"
    ws.merge_cells("A1:F1")
    ws["A1"] = f"维保项目成本工作簿 · {contract}"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 26
    ws.merge_cells("A2:F2")
    ws["A2"] = "系统按取价瀑布自动核算 · 金额为含税/不含税原值混合参考口径 · 导出自 IT 备件智能管理系统"
    ws["A2"].font = _SUB_FONT

    kv = [
        ("合同（销售订单）", contract, None),
        ("合同金额（含税参考）", float(budget) if budget is not None else "（销售表未找到）", _MONEY if budget is not None else None),
        ("税率", float(so.tax_rate) if so is not None and so.tax_rate is not None else "—", None),
        ("已花合计（备件+报销）", spent, _MONEY),
        ("　├ 备件成本", spent_parts, _MONEY),
        ("　└ 报销费用（已结束）", spent_exp, _MONEY),
        ("剩余预算", remaining if remaining is not None else "—", _MONEY if remaining is not None else None),
        ("状态", st_label, None),
    ]
    r = 4
    for label, value, fmt in kv:
        lc, vc = ws.cell(row=r, column=1, value=label), ws.cell(row=r, column=2, value=value)
        lc.fill, lc.font, lc.border = _KV_FILL, Font(bold=True), _BORDER
        vc.border = _BORDER
        if fmt:
            vc.number_format = fmt
        if label == "剩余预算" and remaining is not None:
            vc.font = Font(bold=True, color=st_color)
        if label == "状态":
            vc.fill = PatternFill("solid", fgColor=st_color)
            vc.font = Font(bold=True, color="FFFFFF")
            vc.alignment = _CENTER
        r += 1

    r += 1
    cats = sorted({c for m in data["monthly"].values() for c in m if c != "备件消耗"})
    cols = ["月份", "备件消耗", *cats, "当月合计"]
    for i, h in enumerate(cols, 1):
        ws.cell(row=r, column=i, value=h)
    _hdr_row(ws, r, len(cols))
    band = False
    totals = [0.0] * (len(cols) - 2)
    for ym in sorted(data["monthly"]):
        r += 1
        band = not band
        m = data["monthly"][ym]
        vals = [float(m.get("备件消耗", 0))] + [float(m.get(c, 0)) for c in cats]
        for i, v in enumerate(vals):
            totals[i] += v
        row_vals = [ym, *vals, round(sum(vals), 2)]
        for i, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.border = _BORDER
            if band:
                cell.fill = _ALT_FILL
            if i > 1:
                cell.number_format = _MONEY
    r += 1
    total_row = ["合计", *[round(t, 2) for t in totals], round(sum(totals), 2)]
    for i, v in enumerate(total_row, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font, cell.fill, cell.border = Font(bold=True), _TOTAL_FILL, _BORDER
        if i > 1:
            cell.number_format = _MONEY
    _col_widths(ws, [24, 16] + [14] * (len(cats) + 1))

    # ── Sheet2 备件明细-氚云（原列 + 回填/附加列；按 WBDD 单分组斑马纹）──
    ws2 = wb.create_sheet("备件明细-氚云")
    hdr2 = ["数据标题(WBDD单号)", "制单日期", "销售订单", "项目名", "需求类型", "出库仓库",
            "销售人员", "业务类型", "序号", "需供货产品", "产品描述", "需求数量",
            "产品成本", "单价", "合计", "发货SN",
            "行成本单价", "行成本金额", "成本来源", "置信度", "取价月", "距采购天数", "含税口径"]
    ws2.append(hdr2)
    _hdr_row(ws2, 1, len(hdr2))
    ws2.freeze_panes = "A2"
    prev_order, band = None, False
    for ln, o in data["lines"]:
        first = o.order_no != prev_order
        if first:
            band = not band
        prev_order = o.order_no
        doc_cost = data["doc_total"].get(o.order_no)
        ws2.append([
            o.order_no, o.order_date.isoformat() if o.order_date else None,
            o.linked_sales_order_no, o.project_raw or o.project_std, o.demand_type,
            o.warehouse, o.salesperson, o.business_type,
            ln.line_no, ln.pn_std, ln.description,
            float(ln.qty) if ln.qty is not None else None,
            # 财务习惯：单据级总成本恒填首行（§16.4 实证 1407/1407 张多行单据如此）
            float(doc_cost) if (first and doc_cost is not None) else None,
            None, None, ln.serial_numbers,
            float(ln.unit_cost) if ln.unit_cost is not None else None,
            float(ln.cost_amount) if ln.cost_amount is not None else None,
            _SOURCE_LABEL.get(ln.cost_source, ln.cost_source),
            _CONF_LABEL.get(ln.confidence, ln.confidence or ""),
            ln.price_month, ln.price_distance_days, ln.cost_tax_basis,
        ])
        rr = ws2.max_row
        for c in range(1, len(hdr2) + 1):
            cell = ws2.cell(row=rr, column=c)
            cell.border = _BORDER
            if band:
                cell.fill = _ALT_FILL
        for c in (13, 14, 15, 17, 18):
            ws2.cell(row=rr, column=c).number_format = _MONEY
        if ws2.cell(row=rr, column=13).value is not None:      # 产品成本（单据级）高亮
            dc = ws2.cell(row=rr, column=13)
            dc.fill, dc.font = _DOC_FILL, Font(bold=True)
        if ln.confidence == "low":
            ws2.cell(row=rr, column=20).font = Font(color="B8860B", bold=True)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(hdr2))}{ws2.max_row}"
    _col_widths(ws2, [20, 11, 16, 26, 10, 12, 9, 10, 6, 20, 36, 9,
                      13, 9, 9, 18, 11, 12, 16, 8, 9, 11, 9])

    # ── Sheet3 报销明细（§17.3 canonical：人填区，导入=导出同格式）──
    # 第 1 行=归集锚（销售订单|XSDD），第 2 行=表头；员工在下方续填后整本上传即可导入。
    ws3 = wb.create_sheet("报销明细")
    ws3.cell(row=1, column=1, value="销售订单").font = Font(bold=True)
    ws3.cell(row=1, column=1).fill = _KV_FILL
    ws3.cell(row=1, column=2, value=contract).font = Font(bold=True)
    for c in (1, 2):
        ws3.cell(row=1, column=c).border = _BORDER
    hdr3 = ["报销日期", "报销人员", "报销类别", "费用分类", "支出事由",
            "报销金额", "流程状态", "单号", "序号"]
    _AMT_COL = 6
    for c, h in enumerate(hdr3, 1):
        ws3.cell(row=2, column=c, value=h)
    _hdr_row(ws3, 2, len(hdr3))
    ws3.freeze_panes = "A3"
    for i, e in enumerate(data["expenses"]):
        ws3.append([e.expense_date.isoformat() if e.expense_date else None,
                    e.person, e.expense_type, e.fee_category, e.reason,
                    float(e.amount) if e.amount is not None else None,
                    e.data_status, e.bxd_no, e.line_no])
        rr = ws3.max_row
        inactive = e.data_status != config.MAINT_EXPENSE_ACTIVE_STATUS
        for c in range(1, len(hdr3) + 1):
            cell = ws3.cell(row=rr, column=c)
            cell.border = _BORDER
            if i % 2:
                cell.fill = _ALT_FILL
            if inactive:                                       # 未生效：置灰（不计入已花）
                cell.font = Font(color="A0A0A0")
        ws3.cell(row=rr, column=_AMT_COL).number_format = _MONEY
    if data["expenses"]:
        ws3.auto_filter.ref = f"A2:I{ws3.max_row}"
        rr = ws3.max_row + 1
        # 合计行不填日期 → 再导入时按「缺日期」自然跳过（§17.3）
        ws3.cell(row=rr, column=_AMT_COL - 1, value="合计（仅已结束）").font = Font(bold=True)
        tc = ws3.cell(row=rr, column=_AMT_COL, value=spent_exp)
        tc.font, tc.fill = Font(bold=True), _TOTAL_FILL
        tc.number_format = _MONEY
    _col_widths(ws3, [12, 10, 12, 14, 42, 13, 10, 18, 6])

    # ── Sheet4 填写说明（系统区，导入时忽略）──
    ws4 = wb.create_sheet("填写说明")
    ws4.sheet_view.showGridLines = False
    notes = [
        ("这本工作簿怎么用", "这是系统导出的「项目追踪工作簿」：报销明细页由你续填，其余页由系统生成。"
         "填好后把整本工作簿拖回系统「数据导入」页——系统只吃报销明细页，其它页自动跳过。"),
        ("报销明细页", "必填仅两列：报销日期、报销金额，且**每行都要填日期**（不允许留空表示"
         "「同上」——有金额没日期的行会报错打回）。行内没有销售订单列时，按第 1 行锚"
         "（销售订单=本合同）归集；流程状态留空视为「已结束」（计入项目已花）；"
         "单号/序号选填，有则参与防重（同一文件内单号+序号不能重复）。"),
        ("导入模式", "「跳过」=增量，只进新行；「修复」=以本表为准——本合同在系统里的报销行"
         "会被本表整体替换（改了金额/删了行都以表为准），因此修复模式要求本页没有任何"
         "错误行，有错先修再导。"),
        ("备件明细页", "系统按取价瀑布自动回填（产品成本=单据级总额填首行，行级取价在附加列），"
         "此页导入时忽略——备件出库数据请一直用氚云「维保需求单」导出上传，样式不变。"),
        ("空白表单", "新项目可直接导出本工作簿当作空白表单分发：报销页只有表头和锚行，填完传回即可。"),
    ]
    ws4.column_dimensions["A"].width = 16
    ws4.column_dimensions["B"].width = 96
    tt = ws4.cell(row=1, column=1, value="项目追踪工作簿 · 填写说明")
    tt.font = _TITLE_FONT
    for i, (k, v) in enumerate(notes, 3):
        kc = ws4.cell(row=i, column=1, value=k)
        kc.font = Font(bold=True)
        kc.alignment = Alignment(vertical="top")
        vc = ws4.cell(row=i, column=2, value=v)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws4.row_dimensions[i].height = 42
    return wb


@router.get("/export-workbook")
def export_workbook(
    contract: str = Query(..., max_length=64),
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),   # 硬鉴权：缺/失效凭证 → 401
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> StreamingResponse:
    """§16.4 工作簿形态回填导出：与财务现用「项目预算工作簿」同构（三 sheet）。

    xlsx 无法走字段级脱敏 → 显式门禁：本质是成本导出，无 data_purchase_cost 直接 403。
    「产品成本」按财务习惯填单据级总成本于每张 WBDD 首行；行级取价明细作附加列（增强不破坏）。
    """
    from fastapi import HTTPException

    if apply_field_visibility({"unit_cost": 1}, ctx).get("unit_cost") is None:
        raise HTTPException(status_code=403, detail="无成本查看权限，不能导出项目成本工作簿")
    record_access_log(ctx, "export_workbook", "maintenance", {"contract": contract})
    data = maintenance_cost.contract_workbook_data(db, contract)
    wb = _build_workbook(contract, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(
            f"project_workbook_{contract[:40]}.xlsx",
            ascii_fallback="project_workbook.xlsx",
        )},
    )
