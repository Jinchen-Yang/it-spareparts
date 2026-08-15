"""AI 兜底列映射 API（C3）：propose → accept → 提案列表。

AI 只提案、人工确认后走标准解析/预览/应用；未配置 LLM 时优雅降级 503。
"""
from fastapi import APIRouter, Depends, HTTPException, Path, Query, Request, Response, status
from sqlalchemy import select
from sqlalchemy.orm import Session
from starlette.datastructures import UploadFile

from app import permissions as _perm
from app.auth import current_identity, current_role
from app.db import get_db
from app.models.maintenance_ai_fallback import MaintenanceAiMappingProposal
from app.models.system import SysUser
from app.security import (
    UserContext,
    get_current_user_context,
    require_action,
    require_page,
)
from app.services import import_safety
from app.services import maintenance_ai_fallback as ai

router = APIRouter(prefix="/maintenance", tags=["maintenance"])
MAX_PREVIEW_BYTES = 16 * 1024 * 1024
_ACTION_KEYS = {
    "ckd_shipment": "action_maintenance_doc_import",
    "rkd_inbound": "action_maintenance_doc_import",
    "return_order": "action_maintenance_doc_import",
    "bxd_expense": "action_maintenance_doc_import",
    "ledger": "action_maintenance_ledger_import",
}


def _action_key(doc_type: str) -> str:
    key = _ACTION_KEYS.get(doc_type)
    if key is None:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_doc_type", "message": "未知单据类型"},
        )
    return key


def _enforce_doc_action(ctx: UserContext, action_key: str) -> None:
    """动作门 + 数据依赖（能改必须能看）；admin 恒放行（与 require_action 同口径）。"""
    from app import config as _config
    from app import permissions as _perm

    if not _config.ENABLE_RBAC or ctx.role == "admin":
        return
    if not ctx.is_authenticated:
        raise HTTPException(status.HTTP_401_UNAUTHORIZED, "请先登录")
    perms = ctx.permissions or _perm.effective(ctx.role, None)
    if not perms.get(action_key, False):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "无此操作权限")
    require_data = (
        "data_purchase_cost"
        if action_key == "action_maintenance_doc_import"
        else "data_profit"
    )
    if require_data and not perms.get(require_data, False):
        raise HTTPException(
            status.HTTP_403_FORBIDDEN,
            "该操作需要同时具备对应数据的查看权限",
        )


def _real_operator(ident: dict) -> str:
    return str(ident.get("username") or ident.get("sub") or "unknown")


def _read_headers_and_samples(data: bytes, doc_type: str) -> tuple[list[str], list[list]]:
    import io

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    try:
        sheet = workbook.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))
        # 与各解析器一致：字段名行是第二个表头行（氚云双表头），字段码行跳过
        header_row_index = 1 if len(rows) >= 2 else 0
        headers = [
            str(v).strip() if v is not None else ""
            for v in (rows[header_row_index] if rows else [])
        ]
        samples: list[list] = []
        for row in rows[header_row_index + 1 : header_row_index + 6]:
            samples.append([str(v)[:40] if v is not None else "" for v in row])
        return headers, samples
    finally:
        workbook.close()


@router.post("/ai-fallback/propose")
async def propose_ai_mapping(
    doc_type: str = Query(...),
    request: Request = None,
    response: Response = None,
    db: Session = Depends(get_db),
    ident: dict = Depends(current_identity),
    _auth: str = Depends(current_role),
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> dict:
    response.headers["Cache-Control"] = "no-store"
    action_key = _action_key(doc_type)
    _enforce_doc_action(ctx, action_key)

    try:
        form = await request.form(max_files=2, max_fields=8)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_request", "message": f"multipart 表单解析失败：{type(exc).__name__}"},
        )
    file = form.get("file")
    if not isinstance(file, UploadFile):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_request", "message": "缺少 file 上传字段"},
        )
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            {"code": "unsupported_media_type", "message": "只接受 .xlsx 文件"},
        )
    try:
        data = await import_safety.read_limited(file, MAX_PREVIEW_BYTES)
        import_safety.validate_xlsx_zip(data, max_bytes=MAX_PREVIEW_BYTES)
    except import_safety.UploadSafetyError as exc:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_file", "message": str(exc)},
        )
    try:
        headers, samples = await import_safety.parse_in_threadpool(
            _read_headers_and_samples, data, doc_type
        )
        result = ai.propose(
            db,
            doc_type=doc_type,
            data=data,
            filename=filename,
            headers=headers,
            samples=samples,
            operated_by=_real_operator(ident),
        )
    except ai.AIUnavailable as exc:
        raise HTTPException(
            status.HTTP_503_SERVICE_UNAVAILABLE,
            {"code": "ai_not_configured", "message": str(exc)},
        )
    except ai.AIProposalInvalid as exc:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_proposal", "message": str(exc)},
        )
    return result


@router.post("/ai-fallback/{proposal_id}/accept")
async def accept_ai_mapping(
    proposal_id: str = Path(..., min_length=36, max_length=36),
    request: Request = None,
    response: Response = None,
    db: Session = Depends(get_db),
    ident: dict = Depends(current_identity),
    _auth: str = Depends(current_role),
    _page: None = Depends(require_page("page_maintenance")),
    ctx: UserContext = Depends(get_current_user_context),
) -> dict:
    response.headers["Cache-Control"] = "no-store"
    row = db.get(MaintenanceAiMappingProposal, proposal_id)
    if row is None:
        raise HTTPException(
            status.HTTP_404_NOT_FOUND,
            {"code": "not_found", "message": "提案不存在"},
        )
    action_key = _action_key(row.doc_type)
    _enforce_doc_action(ctx, action_key)

    idempotency_key = request.headers.get("idempotency-key", "")
    if not (8 <= len(idempotency_key) <= 128):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_request", "message": "Idempotency-Key 必填且长度必须在 8–128 字符之间"},
        )
    try:
        form = await request.form(max_files=2, max_fields=8)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_request", "message": f"multipart 表单解析失败：{type(exc).__name__}"},
        )
    file = form.get("file")
    if not isinstance(file, UploadFile):
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_request", "message": "缺少 file 上传字段"},
        )
    try:
        data = await import_safety.read_limited(file, MAX_PREVIEW_BYTES)
        import_safety.validate_xlsx_zip(data, max_bytes=MAX_PREVIEW_BYTES)
    except import_safety.UploadSafetyError as exc:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            {"code": "invalid_file", "message": str(exc)},
        )
    try:
        batch_id = ai.accept_proposal(
            db,
            proposal_id=proposal_id,
            data=data,
            filename=file.filename or "upload.xlsx",
            operated_by=_real_operator(ident),
            idempotency_key=idempotency_key,
        )
    except ai.AIProposalError as exc:
        db.rollback()
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            {"code": "proposal_conflict", "message": str(exc)},
        )
    return {"proposal_id": proposal_id, "batch_id": batch_id}


@router.get("/ai-fallback/proposals")
def list_ai_proposals(
    doc_type: str | None = Query(None),
    response: Response = None,
    db: Session = Depends(get_db),
    _auth: str = Depends(current_role),
    _page: None = Depends(require_page("page_maintenance")),
) -> dict:
    response.headers["Cache-Control"] = "no-store"
    return {"rows": ai.list_proposals(db, doc_type=doc_type)}
