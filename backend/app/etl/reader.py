"""文件解析：双表头探测 + 列名规整 + 重复列校验 + 文件识别 + ffill（§6.2）；
多 sheet 工作簿逐页探测（§17.5）。"""
import re
from collections import Counter
from collections.abc import Iterable
from itertools import repeat
from typing import NamedTuple

import pandas as pd
from openpyxl import load_workbook

from app import config
from app.etl import mapping, sheet_selection

_FCODE = re.compile(r"^F\d{7}$")
# 空表头单元格归一：pandas 以 dtype=object 读空单元格为 float('nan')→str 后是小写 'nan'，
# openpyxl 流式读则是 None；统一折叠为 ""，避免被误判为「重复列名」（审计 2026-06-28 I-1）。
_BLANK_HEADER = ("", "nan", "none", "<na>")


class ReaderError(Exception):
    """无法解析（识别失败 / 重复列名 / 超限等），导致整批 failed。"""

    def __init__(self, message: str, *, code: str = "reader_error"):
        self.code = code
        super().__init__(message)


def _check_workbook_size(path: str) -> dict[str, int] | None:
    """流式探测行规模（全部 sheet 合计），超 IMPORT_MAX_ROWS 直接拒绝，
    避免 pandas 全量物化 OOM（I-2；§17.5 起多 sheet 均会被解析，按合计控）。"""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None  # 损坏/非 xlsx 交给下方 pd.read_excel 给出统一「无法解析」提示
    try:
        limit = config.IMPORT_MAX_ROWS
        total = 0
        row_counts: dict[str, int] = {}
        for ws in wb.worksheets:
            max_row = ws.max_row  # read_only 下通常来自 <dimension> 标签，O(1)；缺失时为 None
            if max_row is None:
                max_row = 0  # 维度未知 → 流式计数封顶（仍是低内存逐行迭代）
                for _ in ws.iter_rows(values_only=True):
                    max_row += 1
                    if total + max_row > limit:
                        break
            row_counts[ws.title] = max_row
            total += max_row
            if total > limit:
                raise ReaderError(
                    f"文件行数（全部工作表合计）超过 {limit} 行上限，请拆分后再导入"
                    "（避免占用过多内存影响其他用户）。",
                    code="row_limit_exceeded",
                )
        return row_counts
    finally:
        wb.close()


def detect_header(raw: pd.DataFrame) -> int:
    """第 0 行命中 F\\d{7} ≥3 个 → 双表头(表头在第1行)，否则单表头(第0行)。"""
    first = raw.iloc[0].astype(str)
    return 1 if first.str.match(_FCODE).sum() >= 3 else 0


def _norm_cols(values) -> list[str]:
    out: list[str] = []
    for c in values:
        s = "" if c is None else str(c).strip()
        if s.lower() in _BLANK_HEADER:
            s = ""
        out.append(s)
    return out


class SheetData(NamedTuple):
    """单个可识别 sheet 的解析结果。anchor：报销页归集锚（§17.3，如 XSDD 单号），其余类型 None。

    dup_cols：表头重复列名清单。此处不抛错——只有真正要入库的页才值得整批失败，
    被跳过的页（如粘贴的副本页）带着重复列名不该拖垮整个文件；由调用方按取用情况裁决。
    """

    sheet_name: str
    df: pd.DataFrame
    file_type: str
    anchor: str | None
    dup_cols: list
    header_row: int
    data_rows: int
    columns: list[str]


class SheetInspection(NamedTuple):
    sheet_name: str
    file_type: str | None
    header_row: int | None
    data_rows: int
    columns: list[str]
    dup_cols: list[str]
    parsed: SheetData | None


_ANCHOR_LABELS = {"销售订单", "维保销售订单"}
# 表头探测扫描的行数上限：锚行(≤1) + F码行(≤1) + 表头行
_HEADER_SCAN_ROWS = 3


def _scan_anchor(raw: pd.DataFrame, upto: int) -> str | None:
    """在表头行之前的行里找归集锚：形如「销售订单 | XSDD-…」的相邻单元格对（§17.3）。"""
    for i in range(min(upto, len(raw))):
        cells = _norm_cols(raw.iloc[i].tolist())
        for j, c in enumerate(cells[:-1]):
            if mapping._strip_opt(c) in _ANCHOR_LABELS and cells[j + 1]:
                return cells[j + 1]
    return None


def _coalesce_value_aliases(df: pd.DataFrame, file_type: str) -> pd.DataFrame:
    for target, aliases in mapping.VALUE_ALIASES.get(file_type, {}).items():
        present_aliases = [alias for alias in aliases if alias in df.columns]
        if target not in df.columns and not present_aliases:
            continue
        values = (df[target].replace("", pd.NA) if target in df.columns
                  else pd.Series(pd.NA, index=df.index, dtype=object))
        for alias in present_aliases:
            values = values.combine_first(df[alias].replace("", pd.NA))
        df[target] = values
        if present_aliases:
            df = df.drop(columns=present_aliases)
    return df


def _identity_value(value: object) -> object | None:
    if value is None or pd.isna(value) or value == "":
        return None
    return value


def _order_group_ids(
    raw_ids: Iterable[object],
    order_nos: Iterable[object],
) -> list[int]:
    current_raw = None
    current_order = None
    group_id = 0
    groups: list[int] = []

    for raw_value, order_value in zip(raw_ids, order_nos, strict=True):
        raw_id = _identity_value(raw_value)
        order_no = _identity_value(order_value)
        has_identity = raw_id is not None or order_no is not None
        raw_matches = (
            raw_id is not None
            and current_raw is not None
            and raw_id == current_raw
        )
        order_matches = (
            order_no is not None
            and current_order is not None
            and order_no == current_order
        )
        raw_conflicts = (
            raw_id is not None
            and current_raw is not None
            and raw_id != current_raw
        )
        order_conflicts = (
            order_no is not None
            and current_order is not None
            and order_no != current_order
        )
        starts_new_group = has_identity and (
            group_id == 0
            or raw_conflicts
            or order_conflicts
            or not (raw_matches or order_matches)
        )

        if starts_new_group:
            group_id += 1
            current_raw = raw_id
            current_order = order_no
        else:
            if raw_id is not None and current_raw is None:
                current_raw = raw_id
            if order_no is not None and current_order is None:
                current_order = order_no

        groups.append(group_id)

    return groups


def _ffill_head_columns(df: pd.DataFrame, file_type: str) -> pd.DataFrame:
    ffill_cols = [c for c in mapping.FFILL_COLS[file_type] if c in df.columns]
    if not ffill_cols:
        return df
    identity_cols = {
        internal: source for source, internal in mapping.MAPPINGS[file_type]["head"].items()
        if internal in ("raw_order_id", "order_no") and source in df.columns
    }
    if not identity_cols:
        return df
    raw_ids: Iterable[object] = (
        df[identity_cols["raw_order_id"]].array
        if "raw_order_id" in identity_cols else repeat(None, len(df))
    )
    order_nos: Iterable[object] = (
        df[identity_cols["order_no"]].array
        if "order_no" in identity_cols else repeat(None, len(df))
    )
    order_groups = pd.Series(
        _order_group_ids(raw_ids, order_nos),
        index=df.index,
    )
    df[ffill_cols] = (
        df[ffill_cols].replace("", pd.NA).groupby(order_groups, sort=False).ffill()
    )
    return df


def _inspect_frame(
    raw: pd.DataFrame,
    sheet_name: str,
    *,
    load_data: bool,
    row_count: int | None,
) -> SheetInspection:
    """单 sheet：前 _HEADER_SCAN_ROWS 行内找可识别表头，并保留未识别页的元数据。

    兼容既有双表头（第 0 行 F 码、第 1 行真表头：F 码行识别不出类型，自然落到第 1 行）
    与报销页锚行（第 1 行锚、第 2 行表头，同理）。重复列校验只记录不抛错（见 SheetData）。
    """
    if raw.empty:
        return SheetInspection(sheet_name, None, None, 0, [], [], None)
    for h in range(min(_HEADER_SCAN_ROWS, len(raw))):
        cols = mapping.canonicalize_columns(_norm_cols(raw.iloc[h].tolist()))
        file_type = mapping.detect_file_type(cols)
        if file_type is None:
            continue
        dup = [c for c, n in Counter(cols).items() if n > 1 and c != ""]
        data_rows = (
            len(raw) - h - 1
            if load_data or row_count is None
            else max(row_count - h - 1, 0)
        )
        anchor = _scan_anchor(raw, h) if file_type == mapping.EXPENSE else None
        parsed = None
        if load_data:
            df = raw.iloc[h + 1:].reset_index(drop=True)
            df.columns = cols
            if not dup:
                df = _coalesce_value_aliases(df, file_type)
                df = _ffill_head_columns(df, file_type)
            parsed = SheetData(
                sheet_name,
                df,
                file_type,
                anchor,
                dup,
                h + 1,
                data_rows,
                cols,
            )
        return SheetInspection(
            sheet_name,
            file_type,
            h + 1,
            data_rows,
            cols,
            dup,
            parsed,
        )
    first_cols = mapping.canonicalize_columns(_norm_cols(raw.iloc[0].tolist()))
    data_rows = len(raw) if load_data or row_count is None else row_count
    return SheetInspection(sheet_name, None, None, data_rows, first_cols, [], None)


def require_clean_columns(sheet: SheetData) -> None:
    """入库前置校验：将被取用的页若有重复列名 → 整批拒绝（数据质量问题必须响）。"""
    if sheet.dup_cols:
        raise ReaderError(
            f"工作表「{sheet.sheet_name}」表头存在重复列名：{sheet.dup_cols}，请确认导出模版")


def inspect_workbook(path: str, *, load_data: bool = True) -> list[SheetInspection]:
    """统一探测全部工作表；预检仅读表头，正式导入同时构造完整 DataFrame。"""
    row_counts = _check_workbook_size(path)
    try:
        sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=object,
                               engine="openpyxl",
                               nrows=None if load_data else _HEADER_SCAN_ROWS)
    except Exception as exc:  # BadZipFile / openpyxl 格式错误 / 损坏
        raise ReaderError(
            "文件无法按 .xlsx 解析：可能是旧版 .xls 格式、非 Excel 文件或文件已损坏。"
            "请在 Excel 中打开后「另存为 → Excel 工作簿 (.xlsx)」再上传。",
            code="invalid_workbook",
        ) from exc
    if not sheets or all(raw.empty for raw in sheets.values()):
        raise ReaderError("文件为空", code="empty_workbook")
    return [
        _inspect_frame(
            raw,
            name,
            load_data=load_data,
            row_count=row_counts.get(name) if row_counts is not None else None,
        )
        for name, raw in sheets.items()
    ]


def read_workbook(path: str) -> list[SheetData]:
    """兼容入口：返回全部可识别页，未识别页仍不进入正式导入。"""
    return [
        sheet.parsed
        for sheet in inspect_workbook(path)
        if sheet.parsed is not None
    ]


def read_excel(path: str) -> tuple[pd.DataFrame, str]:
    """单表读取（兼容入口）：取第一个可识别 sheet → (DataFrame, file_type)。"""
    parsed = read_workbook(path)
    if not parsed:
        raise ReaderError(
            "无法识别文件类型，请确认是采购/销售/库存/维保出库/报销明细导出文件",
            code="no_recognized_sheet",
        )
    require_clean_columns(parsed[0])
    return parsed[0].df, parsed[0].file_type


def peek_columns(path: str) -> tuple[list[str], str | None]:
    """兼容入口：复用正式 reader 与共享选表规则，返回首个选中页列名与文件类型。"""
    sheets = inspect_workbook(path, load_data=False)
    selection = sheet_selection.select_workbook_sheets(sheets)
    if selection.selected:
        return selection.selected[0].columns, selection.file_type
    return sheets[0].columns, None
