#!/usr/bin/env python3
"""生成维保业务两套 Excel 模板（非氚云机打单部分）。

产物：
  1. 项目工作簿模板（每项目一本；系统下载 → 表尾回填 → 上传）
  2. 台账工作簿模板（业务手工维护的项目/合同/回款/报销归集台账）

设计口径与 docs/maintenance/workbook-template-design.md 保持一致。
颜色约定：
  灰底表头 = 系统生成只读列；黄底表头 = 可回填/可追加列；白底 = 普通事实列。
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = "docs/maintenance/templates"

# ---------------------------------------------------------------- 样式
GRAY = PatternFill("solid", fgColor="D9D9D9")   # 只读/系统生成
YELLOW = PatternFill("solid", fgColor="FFF2CC")  # 可编辑
WHITE = PatternFill("solid", fgColor="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
HEAD_FONT = Font(bold=True, size=10)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

EDITABLE_HINT = "（黄底列可编辑：表尾追加新行 / 修改黄底单元格；灰底列由系统生成，改动将被忽略）"

# ---------------------------------------------------------------- 通用
def style_header(ws, headers, colors, row=1):
    for idx, (h, color) in enumerate(zip(headers, colors), 1):
        c = ws.cell(row=row, column=idx, value=h)
        c.font = HEAD_FONT
        c.fill = GRAY if color == "readonly" else YELLOW if color == "editable" else WHITE
        c.border = BORDER
        c.alignment = WRAP
    ws.freeze_panes = f"A{row + 1}"
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(headers))}{row}"


def set_widths(ws, widths):
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def add_blank_rows(ws, headers, n, start):
    for r in range(start, start + n):
        for idx in range(1, len(headers) + 1):
            ws.cell(row=r, column=idx).border = BORDER


def add_hidden_tech_sheets(wb):
    ws_dict = wb.create_sheet("98_字典")
    ws_dict.append(("字段", "允许值或语义"))
    ws_dict.append(("操作", "留空=不动；CREATE=新增；UPDATE=修改；VOID=作废（缺行永远不等于删除）"))
    ws_dict.append(("需求类型", "报修供货 / 补库供货 / 已供货补流程"))
    ws_dict.append(("流程状态", "已生效 / 进行中 / 已取消 / 草稿"))
    ws_dict.sheet_state = "hidden"
    ws_meta = wb.create_sheet("99_元数据")
    ws_meta.append(("key", "value"))
    ws_meta.append(("协议版本", "项目工作簿模板 v1"))
    ws_meta.append(("生成时间", str(date.today())))
    ws_meta.append(("生成方", "IT 备件智能管理系统"))
    ws_meta.sheet_state = "hidden"


def add_instructions_sheet(wb, lines):
    ws = wb.create_sheet("00_使用说明", 0)
    ws["A1"] = "使用说明"
    ws["A1"].font = TITLE_FONT
    for i, line in enumerate(lines, 2):
        ws.cell(row=i, column=1, value=line).alignment = WRAP
    ws.column_dimensions["A"].width = 100
    return ws


# ---------------------------------------------------------------- 1. 项目工作簿
def build_project_workbook(path):
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, [
        "这本工作簿 = 一个维保项目的完整台账视图。月度流程：系统下载 → 在业务 Sheet 表尾追加/修改黄底列 → 整本上传。",
        "硬规则：缺行 ≠ 删除。只有「操作」列填 VOID 才会作废业务行；历史更正显式留痕。",
        EDITABLE_HINT,
        "Sheet 一览：01_项目基础信息（只读）· 02_概览数据（只读）· 03_备件订单 · 04_报销订单 · 05_项目经理回款单（月度累计，表尾追加）· 06_现场领用与返还（行级不返还标记）",
        "日期写法：完整日期写 YYYY-MM-DD（如 2026-10-01）；只有年月的写 YYYY-MM（如 2026-10）。中文写法可被系统识别，但建议按规范写，避免歧义。",
        "数据同步方向：项目/合同/回款计划以台账工作簿为唯一来源（本表重叠列只读）；本表只回填成本、月度累计回款、凭证、领用行标记等台账没有的列。",
    ])

    # 01 项目基础信息（只读参考）
    ws = wb.create_sheet("01_项目基础信息")
    headers = ["项目编号", "项目名称", "业务类型", "客户名称", "维保开始日期", "维保结束日期",
               "项目经理(CMO)", "维保负责人", "销售人员", "硬盘不返还默认值(项目级)", "项目状态", "前置库种类数", "前置库件数", "前置库金额(含税)"]
    colors = ["readonly"] * len(headers)
    ws.append(headers)
    ws.append(["示例：WX-001", "示例：大疆20260201-20261231新华三整体维保", "整体维保", "新华三集团",
               "2026-02-01", "2026-12-31", "廖晓娟", "李冰冰", "李呈辉", "是", "服务中",
               "尚未接入", "尚未接入", "尚未接入"])
    set_widths(ws, [14, 38, 10, 14, 13, 13, 13, 11, 10, 17, 10, 11, 10, 16])

    # 02 概览数据（只读）
    ws = wb.create_sheet("02_概览数据")
    ws["A1"] = "一、合同清单（只读）"
    ws["A1"].font = TITLE_FONT
    c_headers = ["合同编号", "合同额(含税)", "原始合同状态", "状态映射", "是否计入总额", "生效日期", "失效日期", "金额完整性"]
    for idx, h in enumerate(c_headers, 1):
        ws.cell(row=2, column=idx, value=h)
    style_header(ws, c_headers, ["readonly"] * len(c_headers), row=2)
    ws.append([])
    first_metric_row = ws.max_row + 2
    ws.cell(row=first_metric_row, column=1, value="二、关键指标（只读）").font = TITLE_FONT
    metrics = [("合同总额(含税)", ""), ("累计回款(含税)", ""), ("回款进度", ""),
               ("项目已计成本(含税)", ""), ("成本率", ""), ("缺失成本行数", ""),
               ("前置库存金额(含税)", ""), ("超90天未领用备件行数", ""), ("数据完整性提示", "")]
    for i, (k, v) in enumerate(metrics, first_metric_row + 1):
        ws.cell(row=i, column=1, value=k).font = BODY_FONT
        ws.cell(row=i, column=2, value=v)
    set_widths(ws, [30, 26, 14, 12, 12, 12, 12, 14])

    # 03 备件订单（回填：成本来源/未税单位成本/变更原因）
    ws = wb.create_sheet("03_备件订单")
    headers = ["维保单号", "制单日期", "需求类型", "业务类型", "合同编号", "项目名称", "PN", "产品描述",
               "需求数量", "退货数量", "发货SN", "出库仓库", "成本来源(系统)", "未税单位成本", "含税单位成本(系统计算)", "变更原因"]
    colors = ["readonly"] * 13 + ["editable", "readonly", "editable"]
    ws.append(headers)
    style_header(ws, headers, colors)
    ws.append(["WBDD-20260702-0014", "2026-07-02", "补库供货", "整体维保", "XSDD-20250731-0035",
               "示例项目", "02311AYV", "8G 内存", 10, 0, "", "上海成品仓", "采购±7天加权", 120.5, 136.17, ""])
    add_blank_rows(ws, headers, 10, ws.max_row + 1)
    set_widths(ws, [17, 11, 10, 9, 17, 30, 14, 22, 9, 9, 16, 11, 16, 11, 13, 12])

    # 04 报销订单（回填：未税金额/备注）
    ws = wb.create_sheet("04_报销订单")
    headers = ["报销单号", "报销日期", "报销人员", "报销类别", "费用分类", "支出事由", "合同编号",
               "未税金额", "含税金额(系统计算)", "流程状态", "备注"]
    colors = ["readonly"] * 7 + ["editable", "readonly", "readonly", "editable"]
    ws.append(headers)
    style_header(ws, headers, colors)
    ws.append(["BXD-20260721-0019", "2026-07-27", "罗汇康", "维保费用", "外援费用", "北京2026年6月外援费用",
               "XSDD-20260203-0029", 707.96, 800, "已结束", ""])
    add_blank_rows(ws, headers, 10, ws.max_row + 1)
    set_widths(ws, [17, 11, 10, 10, 12, 26, 17, 11, 13, 10, 18])

    # 05 项目经理回款单（月度累计快照，表尾追加）
    ws = wb.create_sheet("05_项目经理回款单")
    headers = ["操作", "合同编号", "报告月份", "累计回款金额(含税)", "回款凭证号", "状态(系统)", "备注"]
    colors = ["editable", "readonly", "editable", "editable", "editable", "readonly", "editable"]
    ws.append(headers)
    style_header(ws, headers, colors)
    ws.append(["", "XSDD-20260731-0086", "2026-10", 2986.57, "", "已确认", ""])
    add_blank_rows(ws, headers, 8, ws.max_row + 1)
    dv = DataValidation(type="list", formula1='"CREATE,VOID"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"A3:A{ws.max_row}")
    set_widths(ws, [9, 17, 11, 16, 15, 10, 18])

    # 06 现场领用与返还（行级不返还标记；默认值继承 01 表项目级设置）
    ws = wb.create_sheet("06_现场领用与返还")
    headers = ["现场领用单号", "领用日期", "PN", "备件SN", "领用数量", "是否应返还(行级)", "应返数量(系统)", "返还状态(系统)", "返还单号(系统)", "备注"]
    colors = ["readonly"] * 5 + ["editable"] + ["readonly"] * 3 + ["editable"]
    ws.append(headers)
    style_header(ws, headers, colors)
    ws.append(["LY-20260710-0001", "2026-07-10", "02311AYV", "S0M59S5M", 2, "是", 2, "待返还", "", ""])
    add_blank_rows(ws, headers, 10, ws.max_row + 1)
    dv_yn = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"F3:F{ws.max_row}")
    set_widths(ws, [17, 11, 14, 15, 9, 15, 12, 14, 17, 16])

    add_hidden_tech_sheets(wb)
    wb.save(path)
    print("written:", path)


# ---------------------------------------------------------------- 2. 台账工作簿
def build_ledger_workbook(path):
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb, [
        "这本工作簿 = 维保项目/合同/回款/报销归集的业务台账（手工维护，商务线唯一事实源）。",
        "月度流程：业务负责人在表尾追加新行 / 修改黄底列后整本上传。",
        EDITABLE_HINT,
        "Sheet 一览：01_项目与合同 · 02_回款计划 · 03_项目成本(报销归集)",
        "回款计划请逐行填写（一期限一行），不再使用横向 N 组「回款时间/回款金额」列。",
        "日期写法：完整日期写 YYYY-MM-DD（如 2026-10-01）；只有年月的写 YYYY-MM（如 2026-10）。中文写法可被系统识别，但建议按规范写，避免歧义。",
        "数据同步方向：本表是项目/合同/回款计划的唯一事实源；项目工作簿里的同名列由系统生成（只读），不要在两处都维护。",
    ])

    # 01 项目与合同
    ws = wb.create_sheet("01_项目与合同")
    headers = ["订单编号", "订单日期", "销售人员", "业务类型", "项目名称", "维保起始日期", "维保终止日期",
               "CMO", "项目经理", "订单金额(含税)", "已收尾款", "待收尾款", "验收材料", "验收完成标记",
               "验收附件", "巡检时间", "巡检完成标记"]
    colors = ["editable"] * 17
    ws.append(headers)
    style_header(ws, headers, colors)
    ws.append(["XSDD-20260731-0086", "2026-07-31", "李呈辉", "整体维保", "阿里专有云20260608-20291205",
               "2026-06-08", "2029-12-05", "廖晓娟", "任鑫明", 44756, 0, 44756,
               "提供设备硬件维修记录报告；服务总结报告", "否", "", "", "否"])
    add_blank_rows(ws, headers, 10, ws.max_row + 1)
    dv_yn = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
    ws.add_data_validation(dv_yn)
    for col, letter in ((14, "N"), (17, "Q")):
        dv_yn.add(f"{letter}3:{letter}{ws.max_row}")
    dv_type = DataValidation(type="list", formula1='"整体维保,备件维保,算力运维,单次维修,整机销售"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add(f"D3:D{ws.max_row}")
    set_widths(ws, [18, 12, 10, 10, 34, 13, 13, 10, 10, 13, 10, 10, 30, 12, 12, 12, 12])

    # 02 回款计划（纵向化）
    ws = wb.create_sheet("02_回款计划")
    headers = ["订单编号", "计划期次", "计划回款时间", "计划回款金额"]
    colors = ["editable"] * 4
    ws.append(headers)
    style_header(ws, headers, colors)
    for seq, (t, amt) in enumerate([
        ("2026-10", 2986.57), ("2027-01", 2986.57), ("2027-04", 2986.57), ("2027-07", 2986.57),
    ], 1):
        ws.append(["XSDD-20260731-0086", seq, t, amt])
    add_blank_rows(ws, headers, 12, ws.max_row + 1)
    set_widths(ws, [18, 10, 14, 14])

    # 03 项目成本（报销归集）
    ws = wb.create_sheet("03_项目成本")
    headers = ["费用单号", "报销人员", "报销类别", "支出事由", "维保销售订单", "项目名称", "销售订单", "销售人员", "费用分类", "报销金额", "备注"]
    colors = ["editable"] * 11
    ws.append(headers)
    style_header(ws, headers, colors)
    ws.append(["BXD-20260425-0002", "董学晶", "维保费用", "2026年广西国税2月份第二次巡检和4月",
               "XSDD-20251028-0016", "国税总局存储20251102-20281101北京ST存储设备维保项目",
               "XSDD-20251028-0016", "余俊", "差旅费", 1068.5, ""])
    add_blank_rows(ws, headers, 10, ws.max_row + 1)
    set_widths(ws, [18, 10, 10, 30, 18, 36, 18, 10, 14, 11, 16])

    add_hidden_tech_sheets(wb)
    wb.save(path)
    print("written:", path)


if __name__ == "__main__":
    import os
    os.makedirs(OUT_DIR, exist_ok=True)
    build_project_workbook(os.path.join(OUT_DIR, "维保项目工作簿模板_v1.xlsx"))
    build_ledger_workbook(os.path.join(OUT_DIR, "维保台账工作簿模板_v1.xlsx"))
